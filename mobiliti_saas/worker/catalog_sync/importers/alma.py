from __future__ import annotations

import hashlib
import json
import math
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Literal, Mapping

from openpyxl.utils.cell import column_index_from_string, get_column_letter

from mobiliti_saas.quote_engine.supplier_catalog import load_supplier_catalog_data
from mobiliti_saas.worker.catalog_sync.kundesign_links import (
    load_kundesign_link_index,
    resolve_kundesign_link,
)
from mobiliti_saas.worker.catalog_sync.mondecasa_links import (
    load_mondecasa_link_index,
    resolve_mondecasa_link,
)
from mobiliti_saas.worker.catalog_sync.mondecasa_images import (
    load_mondecasa_image_index,
    resolve_mondecasa_image,
)

from .common import (
    CellRef,
    ImageAsset,
    extract_xlsx_image_galleries_from_bytes,
    extract_xlsx_images_from_bytes,
    neutralize_spreadsheet_text,
    open_xlsx_data_only_from_bytes,
    _normalize_image,
    read_validated_source,
    source_ref,
)


@dataclass(frozen=True)
class AlmaAssetBinding:
    internal_id: str
    asset_sha256: str
    object_name: str
    image_kind: Literal["official"]
    match_status: Literal[
        "exact_xlsx", "merged_xlsx", "family_xlsx", "exact_web", "model_web"
    ]
    source_references: tuple[dict, ...]


@dataclass(frozen=True)
class AlmaSnapshotBuild:
    snapshot: dict
    assets_by_sha256: Mapping[str, ImageAsset]
    bindings: tuple[AlmaAssetBinding, ...]


@dataclass(frozen=True)
class _AlmaImageCandidate:
    identity: str
    asset: ImageAsset
    match_status: Literal["exact_xlsx", "merged_xlsx", "family_xlsx"]
    source_reference: dict
    selection_reason: str = ""


@dataclass(frozen=True)
class _AlmaResolvedImage:
    asset: ImageAsset
    match_status: Literal[
        "exact_xlsx", "merged_xlsx", "family_xlsx", "exact_web", "model_web"
    ]
    source_references: tuple[dict, ...]
    selection_reason: str = ""


_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_KUN_PATH = "SPEC Guide-Alma-KUN.xlsx"
_KUN_PRICE_PATH = "SPEC GUIDES 2026/ALMA/Spec guide-Alma-KUN Design.xlsx"
_MONDECASA_PATH = "SPEC Guide-Alma-Mondecasa.xlsx"
_ALMA_ADAPTER_VERSION = "alma-configurable-pricing-v5"
_FAMILY_ASSET_OVERRIDES = {
    "kun\0bagel\0side table": (
        _KUN_PRICE_PATH,
        "SPEC Alma",
        "A153",
        "057ddd131acf2648f2d8f2cb64113af7312a75f7761b7334f16f04590d481a16",
    ),
    "kun\0flow\0led light strip for flow swing": (
        _KUN_PRICE_PATH,
        "SPEC Alma",
        "A262",
        "36d762c2fe4a92fb532745c54fa22575c4d8637d34eba0f0b63c1ab3d1ca190e",
    ),
    "kun\0bergen\0round dining table": (
        _KUN_PRICE_PATH,
        "SPEC Alma",
        "A287",
        "9b281582a9b891e323cf3af77c1c19a06e309f27076ec37ae3201a58cd865595",
    ),
    "kun\0lotus\0led light strip for lotus round dining table": (
        _KUN_PRICE_PATH,
        "SPEC Alma",
        "A318",
        "fc299f9a024161f3058f7b778a0fba266151047e4120c4b2416feccc35c03444",
    ),
    "kun\0decorative pillows\0lumbar pillow": (
        _KUN_PATH,
        "KUN DESIGN",
        "B68",
        "d14738794c065dca5cb246bc71ed7ccfd51196e3687410fbfa54bd2942b073f1",
    ),
    "kun\0decorative pillows\0decorative pillow": (
        _KUN_PATH,
        "KUN DESIGN",
        "B69",
        "36daf830457b940f559f2e34e1d89ad6e96425f23cfe13d8fbb388ce34fb72be",
    ),
}
_MULTI_IMAGE_IDENTITY_CURATIONS = {
    "kun\0chic\0deluxe folding armchair c": (
        "f436bc94d403e0efc4f5a443729ff045254d5912ef097308ebcfd8631c99917b",
        "official_front_view",
    ),
    "kun\0shell\0dining side chair with pvc rope": (
        "00d8a71122ef4f9a6ad3d328def562fae55c16afd3d7650f02f6d0c5f6b63725",
        "official_same_product_highest_resolution",
    ),
    "kun\0shell\0lounge side chair with pvc rope": (
        "5e23d7b5df7550bd5c5b4a42748ec9985ddfa43225eda251bd9b0c14c825e042",
        "official_same_product_highest_resolution",
    ),
    "kun\0shell\0barstool": (
        "632c085e34f7eeaba21504b0ee52df2f875d939a00db5902fe813a36484346f3",
        "official_same_product_highest_resolution",
    ),
    "kun\0shell\0barstool counter height": (
        "632c085e34f7eeaba21504b0ee52df2f875d939a00db5902fe813a36484346f3",
        "official_same_product_highest_resolution",
    ),
    "kun\0lotus planter\0xxl": (
        "17c1f97b5574ce29eb8c0767e8ef78fc3101849a40e65c8c9c9059e16fdf0955",
        "official_dimensions_65x45",
    ),
    "kun\0lotus planter\0xxxl": (
        "6c51076edec7df85ab59c179bd8a9bda9e9149a659f94c685074cde29fe02ae5",
        "official_dimensions_65x83",
    ),
}
_EXPECTED = {
    _KUN_PATH: {
        "kind": "spec_guide",
        "brand": "KUN",
    },
    _KUN_PRICE_PATH: {
        "kind": "spec_guide",
        "brand": "KUN",
    },
    _MONDECASA_PATH: {
        "kind": "spec_guide",
        "brand": "Mondecasa",
    },
}
_KUN_SHEETS = {"KUN DESIGN", "PAVILION "}
_KUN_PRICE_SHEETS = {"SPEC Alma", "Costo Alma"}
_MONDECASA_SHEETS = {"MONDECASA", "PAVILIONS"}
_KUN_DESIGN_COUNT = 307
_KUN_PAVILION_COUNT = 3
_KUN_MIN_IDENTITY_MATCHES = 299
_KUN_MAX_RECONCILED_GROUPS = 8
_KUN_MAX_DERIVED_GROUPS = 2
_KUN_HEADERS = (
    "Código",
    "Imagen",
    "Descripción",
    "Dimensiones",
    "Sin cojín Aluminio: Recubrimiento en polvo",
    "Sin cojín Aluminio: Aspecto de teca",
    "Solo Cojín Calidad: Tela A Espuma Normal Ceramica A",
    "Solo Cojín Calidad: Tela A+ Espuma Normal Ceramica A+",
    "Solo Cojín Calidad: Tela A++ Espuma Normal Ceramica A++",
)
_MONEY_LIMIT = Decimal("1000000000")
_SIX_PLACES = Decimal("0.000001")
_PRICE_TOLERANCE = Decimal("0.000005")
_DASHES = {"", "-", "—", "–", "/"}


def _plain(value) -> str:
    return re.sub(r"\s+", " ", neutralize_spreadsheet_text(value)).strip()


def _fold(value) -> str:
    text = unicodedata.normalize("NFKD", _plain(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.findall(r"[^\W_]+", text.casefold(), re.UNICODE))


def _slug(value) -> str:
    return "-".join(re.findall(r"[^\W_]+", _fold(value), re.UNICODE))[:120] or "unknown"


def _image_family_key(record: dict) -> str:
    brand = _fold(record.get("brand"))
    collection = _fold(record.get("collection"))
    lines = str(record.get("link_description") or record.get("description") or "").splitlines()
    product_type = _fold(lines[0] if lines else "")
    prefix = f"{collection} "
    if product_type.startswith(prefix):
        product_type = product_type[len(prefix):]
    return "\0".join((brand, collection, product_type))


def _resolve_kun_asset_candidates(
    records: list[dict], candidates: list[_AlmaImageCandidate]
) -> dict[str, _AlmaResolvedImage]:
    records_by_identity = {record["identity"]: record for record in records}
    if len(records_by_identity) != len(records):
        raise ValueError("ALMA_IMAGE_COVERAGE")
    by_identity: dict[str, list[_AlmaImageCandidate]] = {}
    for candidate in candidates:
        if candidate.identity not in records_by_identity:
            raise ValueError("ALMA_IMAGE_COVERAGE")
        by_identity.setdefault(candidate.identity, []).append(candidate)

    resolved = {}
    direct_by_family: dict[str, list[_AlmaResolvedImage]] = {}
    for identity, record in records_by_identity.items():
        options = by_identity.get(identity, [])
        if not options:
            continue
        chosen = min(
            options,
            key=lambda row: (
                {"exact_xlsx": 0, "merged_xlsx": 1, "family_xlsx": 2}[
                    row.match_status
                ],
                json.dumps(row.source_reference, sort_keys=True, separators=(",", ":")),
                row.asset.sha256,
            ),
        )
        match = _AlmaResolvedImage(
            chosen.asset,
            chosen.match_status,
            (chosen.source_reference,),
            chosen.selection_reason,
        )
        resolved[identity] = match
        if match.match_status != "family_xlsx":
            direct_by_family.setdefault(_image_family_key(record), []).append(match)

    for identity, record in records_by_identity.items():
        if identity in resolved:
            continue
        donors = direct_by_family.get(_image_family_key(record), [])
        assets = {donor.asset.sha256: donor.asset for donor in donors}
        if len(assets) != 1:
            raise ValueError("ALMA_IMAGE_COVERAGE")
        references = _unique_refs(
            [reference for donor in donors for reference in donor.source_references]
        )
        asset = next(iter(assets.values()))
        resolved[identity] = _AlmaResolvedImage(
            asset, "family_xlsx", tuple(references), "closed_family_unique_sha"
        )
    return resolved


def _money(value) -> Decimal | None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    if isinstance(value, float) and not math.isfinite(value):
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return None
    if not number.is_finite() or number <= 0 or number > _MONEY_LIMIT:
        return None
    rounded = number.quantize(_SIX_PLACES, rounding=ROUND_HALF_UP)
    return rounded if rounded > 0 else None


def _money_text(value: Decimal) -> str:
    return f"{value:.6f}"


def _blank_or_dash(value) -> bool:
    if value is None:
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    return not text or text == "/" or all(character in "-\u2013\u2014" for character in text)


def _source_hash(
    files,
    link_index,
    mondecasa_link_index=None,
    mondecasa_image_index=None,
) -> str:
    mondecasa_link_index = mondecasa_link_index or load_mondecasa_link_index()
    mondecasa_image_index = mondecasa_image_index or load_mondecasa_image_index()
    material = {
        "adapter_version": _ALMA_ADAPTER_VERSION,
        "kundesign_resource_sha256": link_index.resource_fingerprint,
        "mondecasa_resource_sha256": mondecasa_link_index.resource_fingerprint,
        "mondecasa_image_resource_sha256": mondecasa_image_index.resource_fingerprint,
        "sources": [
            {
                "path": row.path,
                "kind": row.kind,
                "brand": row.brand,
                "sha256": row.sha256,
            }
            for row in sorted(files, key=lambda value: value.path)
        ],
        "family_asset_overrides": [
            {
                "family_key": key,
                "path": value[0],
                "sheet": value[1],
                "cell": value[2],
                "sha256": value[3],
            }
            for key, value in sorted(_FAMILY_ASSET_OVERRIDES.items())
        ],
        "multi_image_identity_curations": [
            {"family_key": key, "sha256": value[0], "reason": value[1]}
            for key, value in sorted(_MULTI_IMAGE_IDENTITY_CURATIONS.items())
        ],
    }
    canonical = json.dumps(
        material, sort_keys=True, separators=(",", ":"), ensure_ascii=True
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _validated_bundle(files):
    rows = tuple(files)
    if len(rows) != len(_EXPECTED):
        raise ValueError("ALMA_BUNDLE")
    bundle = {}
    source_data = {}
    for row in rows:
        logical_path = getattr(row, "path", None)
        expected = _EXPECTED.get(logical_path)
        local_path = getattr(row, "local_path", None)
        declared_hash = getattr(row, "sha256", None)
        if (
            expected is None
            or logical_path in bundle
            or getattr(row, "kind", None) != expected["kind"]
            or getattr(row, "brand", None) != expected["brand"]
            or getattr(row, "mime_type", None) != _MIME
            or not isinstance(declared_hash, str)
            or re.fullmatch(r"[0-9a-f]{64}", declared_hash) is None
            or not isinstance(local_path, Path)
            or local_path.suffix.casefold() != ".xlsx"
        ):
            raise ValueError("ALMA_BUNDLE")
        validated, data = read_validated_source(local_path, ".xlsx")
        if validated.sha256 != declared_hash:
            raise ValueError("ALMA_HASH")
        bundle[logical_path] = row
        source_data[logical_path] = data
    if set(bundle) != set(_EXPECTED):
        raise ValueError("ALMA_BUNDLE")
    return bundle, source_data


def _require_sheets(workbook, expected) -> None:
    if set(workbook.sheetnames) != expected or len(workbook.sheetnames) != len(expected):
        raise ValueError("ALMA_SHEETS")


def _require_kun_headers(workbook) -> None:
    design = workbook["KUN DESIGN"]
    if any(_fold(design.cell(6, column).value) != _fold(value) for column, value in enumerate(_KUN_HEADERS, 1)):
        raise ValueError("ALMA_HEADER")
    pavilion = workbook["PAVILION "]
    expected = ("Picture", "Item no.", "Description", "Dimensions", "FOB Price")
    if any(_fold(pavilion.cell(1, column).value) != _fold(value) for column, value in enumerate(expected, 1)):
        raise ValueError("ALMA_HEADER")


def _require_kun_price_headers(workbook) -> None:
    _require_sheets(workbook, _KUN_PRICE_SHEETS)
    sales = workbook["SPEC Alma"]
    costs = workbook["Costo Alma"]
    sales_headers = (
        "Imagen.", "Cod.", "Descripcion.",
        "Precio Venta", "Precio Venta", "Precio Venta", "Precio Venta", "Precio Venta",
        "Moneda.",
    )
    cost_headers = (
        "Imagen.", "Cod.", "Descripcion.",
        "P. Unitario.", "P. Unitario.", "P. Unitario.", "P. Unitario.", "P. Unitario.",
        "Tipo de Cambio", "LAB Cedis",
    )
    if any(
        _fold(sales.cell(8, column).value) != _fold(value)
        for column, value in enumerate(sales_headers, 1)
    ) or any(
        _fold(costs.cell(8, column).value) != _fold(value)
        for column, value in enumerate(cost_headers, 1)
    ):
        raise ValueError("ALMA_HEADER")


def _mondecasa_header(sheet, row: int) -> bool:
    return (
        _fold(sheet.cell(row, 1).value) in {"cod", "codigo"}
        and "descripcion" in _fold(sheet.cell(row, 3).value)
        and "dimension" in _fold(sheet.cell(row, 4).value)
        and bool(_plain(sheet.cell(row, 5).value))
    )


def _pavilions_header(sheet, row: int) -> bool:
    return _fold(sheet.cell(row, 1).value) == "item no" and "description" in _fold(
        sheet.cell(row, 3).value
    )


def _require_mondecasa_headers(workbook) -> None:
    products = workbook["MONDECASA"]
    if not any(_mondecasa_header(products, row) for row in range(1, products.max_row + 1)):
        raise ValueError("ALMA_HEADER")
    pavilions = workbook["PAVILIONS"]
    if not {5, 10, 17}.issubset(
        {row for row in range(1, pavilions.max_row + 1) if _pavilions_header(pavilions, row)}
    ):
        raise ValueError("ALMA_HEADER")


def _merge_index(sheet) -> dict[tuple[int, int], tuple[int, int, int, int, int, int]]:
    index = {}
    for merged in sheet.merged_cells.ranges:
        anchor = (merged.min_row, merged.min_col)
        record = (*anchor, merged.min_row, merged.max_row, merged.min_col, merged.max_col)
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                index[(row, column)] = record
    return index


def _anchor(index, row: int, column: int) -> tuple[int, int]:
    merged = index.get((row, column))
    return (merged[0], merged[1]) if merged else (row, column)


def _value(sheet, index, row: int, column: int):
    anchor_row, anchor_column = _anchor(index, row, column)
    return sheet.cell(anchor_row, anchor_column).value


def _vertical_product_block(index, row: int) -> tuple[int, int] | None:
    ranges = []
    for column in range(1, 7):
        merged = index.get((row, column))
        if merged is None or merged[4] != merged[5] or merged[4] != column:
            return None
        ranges.append((merged[2], merged[3]))
    return ranges[0] if len(set(ranges)) == 1 else None


def _vertical_code_block(index, row: int) -> tuple[int, int] | None:
    merged = index.get((row, 1))
    if merged is None or merged[4:] != (1, 1) or merged[2] == merged[3]:
        return None
    return merged[2], merged[3]


def _mondecasa_product_block(index, row: int) -> tuple[int, int] | None:
    """Return the row span that represents one Mondecasa product.

    Some Mondecasa products merge the code cell vertically, while others leave
    the continuation code blank and merge only the image or dimensions.  Those
    continuation rows are configurations of the same product, not standalone
    products.
    """
    code_block = _vertical_code_block(index, row)
    if code_block is not None:
        return code_block

    candidates = []
    for column in (2, 4):
        merged = index.get((row, column))
        if (
            merged is not None
            and merged[4:] == (column, column)
            and merged[2] == row
            and merged[3] > row
        ):
            candidates.append((merged[2], merged[3]))
    return max(candidates, key=lambda block: block[1]) if candidates else None


def _category(sheet, index, row: int) -> str | None:
    merged = index.get((row, 1))
    if merged and merged[2] == row and merged[4] == 1 and merged[5] >= 9:
        return _plain(sheet.cell(merged[0], merged[1]).value) or None
    return None


def _cell_ref(file_id: str, sheet, index, row: int, column: int) -> dict:
    anchor_row, anchor_column = _anchor(index, row, column)
    return source_ref(file_id, sheet.title, sheet.cell(anchor_row, anchor_column).coordinate)


def _unique_refs(refs: list[dict]) -> list[dict]:
    by_key = {json.dumps(row, sort_keys=True, separators=(",", ":")): row for row in refs}
    return [by_key[key] for key in sorted(by_key)]


def _summary_price_refs(refs: list[dict]) -> list[dict]:
    if len(refs) <= 3:
        return refs
    coordinates = []
    for ref in refs:
        match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", ref["cell_or_bbox"])
        if match is None:
            return refs
        coordinates.append((column_index_from_string(match[1]), int(match[2])))
    first = refs[0]
    if any(
        ref["file_id"] != first["file_id"] or ref["sheet_or_page"] != first["sheet_or_page"]
        for ref in refs[1:]
    ):
        return refs
    columns, rows = zip(*coordinates)
    bbox = f"{get_column_letter(min(columns))}{min(rows)}:{get_column_letter(max(columns))}{max(rows)}"
    return [source_ref(first["file_id"], first["sheet_or_page"], bbox)]


def _aggregate_price_ref(refs: list[dict]) -> dict:
    if len(refs) == 1:
        return refs[0]
    first = refs[0]
    coordinates = []
    for ref in refs:
        match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", ref["cell_or_bbox"])
        if (
            match is None
            or ref["file_id"] != first["file_id"]
            or ref["sheet_or_page"] != first["sheet_or_page"]
        ):
            raise ValueError("ALMA_PRICE_EVIDENCE")
        coordinates.append((column_index_from_string(match[1]), int(match[2])))
    columns, rows = zip(*coordinates)
    bbox = (
        f"{get_column_letter(min(columns))}{min(rows)}:"
        f"{get_column_letter(max(columns))}{max(rows)}"
    )
    return source_ref(first["file_id"], first["sheet_or_page"], bbox)


def _source_images(images, file_id: str, sheet, index, first_row: int, last_row: int, column: int):
    result = {}
    for row in range(first_row, last_row + 1):
        anchor_row, anchor_column = _anchor(index, row, column)
        coordinate = sheet.cell(anchor_row, anchor_column).coordinate
        reference = CellRef(sheet.title, coordinate)
        image = images.get(reference)
        if image is None:
            continue
        source = source_ref(file_id, sheet.title, coordinate)
        result[(image.sha256, coordinate)] = {
            "sha256": image.sha256,
            "width": image.width,
            "height": image.height,
            "source": source,
        }
    return [result[key] for key in sorted(result)]


def _price_evidence(kind: str, label: str, price: Decimal, available: bool, source: dict) -> dict:
    return {
        "kind": kind,
        "label": label,
        "price_net": _money_text(price),
        "available": available,
        "source": source,
    }


def _base_option(option_id: str, name: str, price: Decimal) -> dict:
    return {"id": option_id, "name": name, "price_net": _money_text(price), "available": True}


def _kun_configured_base_label(header: str, column: int) -> str:
    clean = re.sub(r"\bsolo\s+coj[ií]n\b", "", header, flags=re.IGNORECASE)
    clean = re.sub(r"[-–—]{3,}", " · ", clean)
    clean = re.sub(
        r"\s+(?=(?:espuma|cer[aá]mica)\b)",
        " · ",
        clean,
        flags=re.IGNORECASE,
    )
    clean = re.sub(r"(?:\s*·\s*)+", " · ", clean)
    clean = re.sub(r"\s+", " ", clean).strip(" ·-")
    return clean or f"Configuración {column - 6}"


def _add_on(
    option_id: str,
    name: str,
    family: str,
    price: Decimal,
    available: bool,
    compatible_base_option_ids: list[str] | None = None,
) -> dict:
    option = {
        "id": option_id,
        "name": name,
        "family": family,
        "price_net": _money_text(price),
        "available": available,
    }
    if compatible_base_option_ids:
        option["compatible_base_option_ids"] = compatible_base_option_ids
    return option


def _record(
    *,
    source,
    sheet,
    index,
    row: int,
    brand: str,
    collection: str,
    code,
    description,
    dimensions,
    direct_price: Decimal | None,
    base_options: list[dict],
    add_on_options: list[dict],
    price_evidence: list[dict],
    source_images: list[dict],
    warnings: list[str],
    extra_refs: list[dict] | None = None,
    extra_attributes: dict | None = None,
) -> dict:
    source_code = _plain(code)
    clean_description = _plain(description)
    clean_dimensions = _plain(dimensions)
    name = clean_description.split(" - ", 1)[0].strip() or source_code or "Producto ALMA por verificar"
    refs = []
    if source_code:
        refs.append(_cell_ref(source.sha256, sheet, index, row, 1 if sheet.title != "PAVILION " else 2))
    if clean_description:
        refs.append(_cell_ref(source.sha256, sheet, index, row, 3))
    if clean_dimensions:
        refs.append(_cell_ref(source.sha256, sheet, index, row, 4))
    refs.extend(_summary_price_refs([entry["source"] for entry in price_evidence]))
    refs.extend(entry["source"] for entry in source_images)
    refs.extend(extra_refs or [])
    if direct_price is None and not base_options:
        warnings.append("Precio base no disponible; verificar antes de cotizar.")
    if brand == "KUN":
        identity = "\0".join(
            (
                brand.casefold(),
                collection.casefold(),
                source_code.casefold(),
                clean_description.casefold(),
                clean_dimensions.casefold(),
            )
        )
    else:
        identity = "\0".join(
            (
                source.path,
                sheet.title,
                str(row),
                source_code.casefold(),
                clean_description.casefold(),
                clean_dimensions.casefold(),
            )
        )
    return {
        "identity": identity,
        "source_code": source_code,
        "brand": brand,
        "collection": collection,
        "name": name,
        "description": clean_description,
        "link_description": description,
        "dimensions": clean_dimensions,
        "direct_price": direct_price,
        "base_options": base_options,
        "add_on_options": add_on_options,
        "price_evidence": price_evidence,
        "source_images": source_images,
        "extra_attributes": dict(extra_attributes or {}),
        "warnings": list(dict.fromkeys(warnings)),
        "refs": _unique_refs(refs),
    }


def _invalid_price(value, reference: CellRef, formulas: set[CellRef]) -> bool:
    return reference in formulas or (not _blank_or_dash(value) and _money(value) is None)


def _parse_kun_design(source, sheet, images, formulas) -> list[dict]:
    index = _merge_index(sheet)
    headers = {column: _plain(sheet.cell(6, column).value) for column in range(5, 10)}
    records = []
    collection = ""
    row = 7
    while row <= sheet.max_row:
        category = _category(sheet, index, row)
        if category:
            collection = category
            row += 1
            continue
        block = _vertical_product_block(index, row)
        if block and block[0] < row:
            row += 1
            continue
        last_row = block[1] if block and block[0] == row else row
        code = _value(sheet, index, row, 1)
        description = _value(sheet, index, row, 3)
        dimensions = _value(sheet, index, row, 4)
        if not (_plain(code) or _plain(description)):
            row = last_row + 1
            continue

        warnings = []
        evidence = []
        rejected_refs = []
        base_options = []
        base_prices = {}
        for column in (5, 6):
            value = sheet.cell(row, column).value
            price = _money(value)
            reference = source_ref(source.sha256, sheet.title, sheet.cell(row, column).coordinate)
            if price is not None:
                option_id = f"base-{sheet.cell(6, column).column_letter.casefold()}"
                base_options.append(_base_option(option_id, headers[column], price))
                evidence.append(_price_evidence("base", headers[column], price, True, reference))
                base_prices[column] = price
            elif _invalid_price(value, CellRef(sheet.title, sheet.cell(row, column).coordinate), formulas):
                warnings.append("Se ignoró un precio base no numérico, negativo o fuera de rango.")
                rejected_refs.append(reference)
        if 5 in base_prices and 6 in base_prices and base_prices[6] < base_prices[5] / 2:
            warnings.append("La alternativa de materialidad F es materialmente menor que E; verificar precio.")

        add_ons = []
        levels = {7: "A", 8: "A+", 9: "A++"}
        for price_row in range(row, last_row + 1):
            for column in (7, 8, 9):
                value = sheet.cell(price_row, column).value
                price = _money(value)
                coordinate = sheet.cell(price_row, column).coordinate
                reference = source_ref(source.sha256, sheet.title, coordinate)
                if price is not None:
                    name = f"Agregado por verificar {levels[column]}"
                    add_ons.append(
                        _add_on(
                            f"review-r{price_row}-c{column}",
                            name,
                            f"review-kun-design-r{price_row}",
                            price,
                            False,
                        )
                    )
                    evidence.append(_price_evidence("add_on", name, price, False, reference))
                elif _invalid_price(value, CellRef(sheet.title, coordinate), formulas):
                    warnings.append("Se ignoró un agregado no numérico, negativo o fuera de rango.")
                    rejected_refs.append(reference)
        records.append(
            _record(
                source=source,
                sheet=sheet,
                index=index,
                row=row,
                brand="KUN",
                collection=collection,
                code=code,
                description=description,
                dimensions=dimensions,
                direct_price=None,
                base_options=base_options,
                add_on_options=add_ons,
                price_evidence=evidence,
                source_images=_source_images(images, source.sha256, sheet, index, row, last_row, 2),
                warnings=warnings,
                extra_refs=rejected_refs,
            )
        )
        row = last_row + 1
    return records


def _parse_kun_pavilion(source, sheet, images, formulas) -> list[dict]:
    index = _merge_index(sheet)
    records = []
    for row in range(3, sheet.max_row + 1):
        code = sheet.cell(row, 2).value
        description = sheet.cell(row, 3).value
        if not (_plain(code) or _plain(description)):
            continue
        value = sheet.cell(row, 5).value
        price = _money(value)
        coordinate = sheet.cell(row, 5).coordinate
        warnings = []
        evidence = []
        rejected_refs = []
        if price is not None:
            evidence.append(
                _price_evidence(
                    "direct",
                    "FOB Price",
                    price,
                    True,
                    source_ref(source.sha256, sheet.title, coordinate),
                )
            )
        elif _invalid_price(value, CellRef(sheet.title, coordinate), formulas):
            warnings.append("Se ignoró un precio FOB no numérico, negativo o fuera de rango.")
            rejected_refs.append(source_ref(source.sha256, sheet.title, coordinate))
        records.append(
            _record(
                source=source,
                sheet=sheet,
                index=index,
                row=row,
                brand="KUN",
                collection="PAVILION",
                code=code,
                description=description,
                dimensions=sheet.cell(row, 4).value,
                direct_price=price,
                base_options=[],
                add_on_options=[],
                price_evidence=evidence,
                source_images=_source_images(images, source.sha256, sheet, index, row, row, 1),
                warnings=warnings,
                extra_refs=rejected_refs,
                extra_attributes={
                    "price_reconciliation": {
                        "method": "identity_direct",
                        "source_path": source.path,
                    }
                },
            )
        )
    return records


def _kun_groups(sheet, *, code_column: int, description_column: int, price_columns, first_row: int):
    index = _merge_index(sheet)
    groups = {}
    order = []
    collection = ""
    for row in range(first_row, sheet.max_row + 1):
        category = _category(sheet, index, row) if code_column == 1 else None
        if category:
            collection = category
            continue
        code = _value(sheet, index, row, code_column)
        prices = [_money(sheet.cell(row, column).value) for column in price_columns]
        if not _plain(code) or not any(prices):
            continue
        key = (_anchor(index, row, code_column), _anchor(index, row, description_column))
        if key not in groups:
            groups[key] = {
                "row": row,
                "rows": [],
                "code": code,
                "description": _value(sheet, index, row, description_column),
                "collection": collection,
                "index": index,
            }
            order.append(key)
        groups[key]["rows"].append(row)
    return [groups[key] for key in order]


def _kun_price_matrix(sheet, group, columns, formulas) -> list[list[Decimal | None]]:
    matrix = []
    for row in group["rows"]:
        values = []
        for column in columns:
            raw = sheet.cell(row, column).value
            reference = CellRef(sheet.title, sheet.cell(row, column).coordinate)
            numeric_zero = (
                isinstance(raw, (int, float, Decimal))
                and not isinstance(raw, bool)
                and Decimal(str(raw)) == 0
            )
            if not numeric_zero and _invalid_price(raw, reference, formulas):
                raise ValueError("ALMA_PRICE")
            values.append(_money(raw))
        matrix.append(values)
    return matrix


def _same_price_matrix(left, right) -> bool:
    return left == right


def _derived_sale(cost: Decimal) -> Decimal:
    return (cost / Decimal("0.3") / Decimal("0.5")).quantize(
        _SIX_PLACES, rounding=ROUND_HALF_UP
    )


def _validate_current_kun_prices(source, workbook):
    _require_kun_price_headers(workbook)
    sales = workbook["SPEC Alma"]
    costs = workbook["Costo Alma"]
    sales_groups = _kun_groups(
        sales, code_column=2, description_column=3, price_columns=range(4, 9), first_row=9
    )
    cost_groups = _kun_groups(
        costs, code_column=2, description_column=3, price_columns=range(4, 9), first_row=9
    )
    if (
        len(sales_groups) != _KUN_DESIGN_COUNT
        or len(cost_groups) != _KUN_DESIGN_COUNT
    ):
        raise ValueError("ALMA_KUN_COUNT")
    sales_formulas = set(workbook.formula_cells)
    cost_formulas = set(workbook.formula_cells)
    saw_currency = False
    for sale_group, cost_group in zip(sales_groups, cost_groups):
        if (
            _fold(sale_group["code"]) != _fold(cost_group["code"])
            or _fold(sale_group["description"]) != _fold(cost_group["description"])
            or len(sale_group["rows"]) != len(cost_group["rows"])
        ):
            raise ValueError("ALMA_KUN_ALIGNMENT")
        sale_matrix = _kun_price_matrix(sales, sale_group, range(4, 9), sales_formulas)
        cost_matrix = _kun_price_matrix(costs, cost_group, range(4, 9), cost_formulas)
        currencies = {
            _fold(_value(sales, sale_group["index"], row, 9))
            for row in sale_group["rows"]
            if _plain(_value(sales, sale_group["index"], row, 9))
        }
        if currencies - {"usd"}:
            raise ValueError("ALMA_CURRENCY")
        saw_currency = saw_currency or bool(currencies)
        for sale_row, cost_row, sale_number, cost_number in zip(
            sale_group["rows"], cost_group["rows"], sale_matrix, cost_matrix
        ):
            factors = (_money(costs.cell(cost_row, 9).value), _money(costs.cell(cost_row, 10).value))
            if factors != (Decimal("0.300000"), Decimal("0.500000")):
                raise ValueError("ALMA_FACTORS")
            for sale, cost in zip(sale_number, cost_number):
                if (sale is None) != (cost is None) or (
                    sale is not None and abs(sale - _derived_sale(cost)) > _PRICE_TOLERANCE
                ):
                    raise ValueError("ALMA_PRICE_ALIGNMENT")
        sale_group["matrix"] = sale_matrix
        sale_group["source"] = source
        cost_group["matrix"] = cost_matrix
        cost_group["source"] = source
    if not saw_currency:
        raise ValueError("ALMA_CURRENCY")
    return sales_groups, cost_groups


def _parse_kun_design_2026(identity_source, identity_sheet, images, formulas, sales_groups, cost_groups):
    identity_groups = _kun_groups(
        identity_sheet,
        code_column=1,
        description_column=3,
        price_columns=range(5, 10),
        first_row=7,
    )
    if (
        len(identity_groups) != _KUN_DESIGN_COUNT
        or len(sales_groups) != _KUN_DESIGN_COUNT
        or len(cost_groups) != _KUN_DESIGN_COUNT
    ):
        raise ValueError("ALMA_KUN_COUNT")
    headers = {column: _plain(identity_sheet.cell(6, column).value) for column in range(5, 10)}
    records = []
    identity_matches = 0
    reconciled_groups = 0
    derived_groups = 0
    for identity, sales, costs in zip(identity_groups, sales_groups, cost_groups):
        if not (
            len(identity["rows"]) == len(sales["rows"]) == len(costs["rows"])
        ):
            raise ValueError("ALMA_KUN_ALIGNMENT")
        identity_matrix = _kun_price_matrix(identity_sheet, identity, range(5, 10), formulas)
        same_cost = _same_price_matrix(identity_matrix, costs["matrix"])
        same_identity = (
            _fold(identity["code"]) == _fold(costs["code"])
            or _fold(identity["description"]) == _fold(costs["description"])
        )
        derived = not same_cost and not same_identity
        identity_matches += int(same_identity)
        reconciled_groups += int(not same_cost)
        derived_groups += int(derived)
        price_matrix = identity_matrix if derived else costs["matrix"]
        warnings = []
        method = "official_cost"
        if derived:
            warnings.append(
                "Costo oficial vigente conservado pese a diferencia con la conciliación histórica."
            )
        elif not same_cost:
            warnings.append("Costo oficial vigente conciliado por identidad.")

        base_options = []
        cushion_prices: dict[int, Decimal] = {}
        cushion_price_refs: dict[int, list[dict]] = {}
        evidence = []
        cost_refs = []
        for row_index, (identity_row, _sale_row, cost_row) in enumerate(
            zip(identity["rows"], sales["rows"], costs["rows"])
        ):
            for column_index, price in enumerate(price_matrix[row_index]):
                if price is None:
                    continue
                identity_column = column_index + 5
                cost_column = column_index + 4
                label = headers[identity_column] or f"Opción {identity_column}"
                price_source = (
                    source_ref(
                        identity_source.sha256,
                        identity_sheet.title,
                        identity_sheet.cell(identity_row, identity_column).coordinate,
                    )
                    if derived
                    else source_ref(
                        costs["source"].sha256,
                        "Costo Alma",
                        f"{get_column_letter(cost_column)}{cost_row}",
                    )
                )
                if identity_column <= 6:
                    dimension = _plain(
                        _value(identity_sheet, identity["index"], identity_row, 4)
                    )
                    option_label = (
                        f"{label} · {dimension}"
                        if len(identity["rows"]) > 1 and dimension
                        else label
                    )
                    base_options.append(
                        _base_option(
                            f"base-v{row_index + 1}-c{identity_column}",
                            option_label,
                            price,
                        )
                    )
                    evidence.append(
                        _price_evidence("base", option_label, price, True, price_source)
                    )
                else:
                    cushion_prices[identity_column] = (
                        cushion_prices.get(identity_column, Decimal(0)) + price
                    )
                    cushion_price_refs.setdefault(identity_column, []).append(price_source)
                cost_refs.append(price_source)

        add_ons = []
        promote_cushion_prices = not base_options
        for identity_column in range(7, 10):
            price = cushion_prices.get(identity_column)
            if price is None:
                continue
            refs = cushion_price_refs[identity_column]
            source = _aggregate_price_ref(refs)
            label = headers[identity_column] or f"Cojín {identity_column}"
            if promote_cushion_prices:
                configured_label = _kun_configured_base_label(label, identity_column)
                base_options.append(
                    _base_option(
                        f"base-c{identity_column}",
                        configured_label,
                        price,
                    )
                )
                evidence.append(
                    _price_evidence("base", configured_label, price, True, source)
                )
                continue
            add_ons.append(
                _add_on(
                    f"cushion-c{identity_column}",
                    label,
                    "cushion",
                    price,
                    True,
                )
            )
            evidence.append(
                _price_evidence(
                    "add_on",
                    label,
                    price,
                    True,
                    source,
                )
            )

        dimension_values = []
        for identity_row in identity["rows"]:
            dimension = _plain(
                _value(identity_sheet, identity["index"], identity_row, 4)
            )
            if dimension and dimension not in dimension_values:
                dimension_values.append(dimension)
        dimensions = " / ".join(dimension_values) or None
        records.append(
            _record(
                source=identity_source,
                sheet=identity_sheet,
                index=identity["index"],
                row=identity["row"],
                brand="KUN",
                collection=identity["collection"],
                code=identity["code"],
                description=identity["description"],
                dimensions=dimensions,
                direct_price=None,
                base_options=base_options,
                add_on_options=add_ons,
                price_evidence=evidence,
                source_images=_source_images(
                    images,
                    identity_source.sha256,
                    identity_sheet,
                    identity["index"],
                    min(identity["rows"]),
                    max(identity["rows"]),
                    2,
                ),
                warnings=warnings,
                extra_refs=_summary_price_refs(cost_refs) + [
                    source_ref(
                        costs["source"].sha256,
                        "Costo Alma",
                        f"I{min(costs['rows'])}:J{max(costs['rows'])}",
                    )
                ],
                extra_attributes={
                    "price_reconciliation": {
                        "method": method,
                        "identity_path": identity_source.path,
                        "price_path": sales["source"].path,
                        "cost_factor": "0.3",
                        "sale_factor": "0.5",
                    }
                },
            )
        )
    if (
        identity_matches < _KUN_MIN_IDENTITY_MATCHES
        or reconciled_groups > _KUN_MAX_RECONCILED_GROUPS
        or derived_groups > _KUN_MAX_DERIVED_GROUPS
    ):
        raise ValueError("ALMA_KUN_RECONCILIATION")
    return records


def _parse_kun(
    identity_source,
    identity_data: bytes,
    price_source,
    price_data: bytes,
    *,
    identity_images=None,
) -> list[dict]:
    images = identity_images if identity_images is not None else extract_xlsx_images_from_bytes(identity_data)
    identity_workbook = open_xlsx_data_only_from_bytes(identity_data)
    price_workbook = open_xlsx_data_only_from_bytes(price_data)
    try:
        _require_sheets(identity_workbook, _KUN_SHEETS)
        _require_kun_headers(identity_workbook)
        sales_groups, cost_groups = _validate_current_kun_prices(price_source, price_workbook)
        formulas = set(identity_workbook.formula_cells)
        records = _parse_kun_design_2026(
            identity_source,
            identity_workbook["KUN DESIGN"],
            images,
            formulas,
            sales_groups,
            cost_groups,
        )
        pavilion_records = _parse_kun_pavilion(
            identity_source, identity_workbook["PAVILION "], images, formulas
        )
        if (
            len(pavilion_records) != _KUN_PAVILION_COUNT
            or any(record["direct_price"] is None for record in pavilion_records)
        ):
            raise ValueError("ALMA_KUN_COUNT")
        records.extend(pavilion_records)
        return records
    finally:
        identity_workbook.close()
        price_workbook.close()


def _group_image_candidates(record, source, sheet, group, images, image_column: int):
    identity = record["identity"]
    candidates = {}
    product_rows = set(group["rows"])
    for row in group["rows"]:
        for column in range(1, 5):
            merged = group["index"].get((row, column))
            if merged is not None and merged[2] < merged[3]:
                product_rows.update(range(merged[2], merged[3] + 1))
    status = "merged_xlsx" if len(product_rows) > 1 else "exact_xlsx"
    for row in sorted(product_rows):
        coordinate = f"{get_column_letter(image_column)}{row}"
        raw_assets = images.get(CellRef(sheet.title, coordinate), ())
        assets = _curated_gallery_assets(
            record,
            raw_assets,
        )
        curation = _MULTI_IMAGE_IDENTITY_CURATIONS.get(_image_family_key(record))
        reason = curation[1] if len(raw_assets) > 1 and curation is not None else ""
        for asset in assets:
            reference = source_ref(source.sha256, sheet.title, coordinate)
            key = (status, json.dumps(reference, sort_keys=True), asset.sha256)
            candidates[key] = _AlmaImageCandidate(
                identity, asset, status, reference, reason
            )
    return [candidates[key] for key in sorted(candidates)]


def _curated_gallery_assets(record, assets):
    if len(assets) <= 1:
        return assets
    curation = _MULTI_IMAGE_IDENTITY_CURATIONS.get(_image_family_key(record))
    if curation is None:
        return ()
    expected_sha = curation[0]
    matches = tuple(asset for asset in assets if asset.sha256 == expected_sha)
    if len(matches) != 1:
        raise ValueError("ALMA_IMAGE_COVERAGE")
    return matches


def _require_family_override_asset(assets, expected_sha: str) -> ImageAsset:
    if len(assets) != 1 or assets[0].sha256 != expected_sha:
        raise ValueError("ALMA_IMAGE_COVERAGE")
    return assets[0]


def _kun_asset_candidates(
    identity_source,
    identity_data: bytes,
    price_source,
    price_data: bytes,
    records: list[dict],
    identity_images,
    price_images,
) -> list[_AlmaImageCandidate]:
    identity_workbook = open_xlsx_data_only_from_bytes(identity_data)
    price_workbook = open_xlsx_data_only_from_bytes(price_data)
    try:
        identity_sheet = identity_workbook["KUN DESIGN"]
        identity_groups = _kun_groups(
            identity_sheet,
            code_column=1,
            description_column=3,
            price_columns=range(5, 10),
            first_row=7,
        )
        sales_sheet = price_workbook["SPEC Alma"]
        sales_groups = _kun_groups(
            sales_sheet,
            code_column=2,
            description_column=3,
            price_columns=range(4, 9),
            first_row=9,
        )
        costs_sheet = price_workbook["Costo Alma"]
        costs_groups = _kun_groups(
            costs_sheet,
            code_column=2,
            description_column=3,
            price_columns=range(4, 9),
            first_row=9,
        )
        design_records = records[:_KUN_DESIGN_COUNT]
        if not (
            len(design_records)
            == len(identity_groups)
            == len(sales_groups)
            == len(costs_groups)
            == _KUN_DESIGN_COUNT
        ):
            raise ValueError("ALMA_IMAGE_COVERAGE")
        candidates = []
        for record, identity_group, sales_group, costs_group in zip(
            design_records, identity_groups, sales_groups, costs_groups
        ):
            candidates.extend(
                _group_image_candidates(
                    record,
                    identity_source,
                    identity_sheet,
                    identity_group,
                    identity_images,
                    2,
                )
            )
            candidates.extend(
                _group_image_candidates(
                    record,
                    price_source,
                    sales_sheet,
                    sales_group,
                    price_images,
                    1,
                )
            )
            candidates.extend(
                _group_image_candidates(
                    record,
                    price_source,
                    costs_sheet,
                    costs_group,
                    price_images,
                    1,
                )
            )

        pavilion_sheet = identity_workbook["PAVILION "]
        pavilion_rows = [
            row
            for row in range(3, pavilion_sheet.max_row + 1)
            if _plain(pavilion_sheet.cell(row, 2).value)
            or _plain(pavilion_sheet.cell(row, 3).value)
        ]
        pavilion_records = records[_KUN_DESIGN_COUNT:]
        if not len(pavilion_rows) == len(pavilion_records) == _KUN_PAVILION_COUNT:
            raise ValueError("ALMA_IMAGE_COVERAGE")
        for record, row in zip(pavilion_records, pavilion_rows):
            coordinate = pavilion_sheet.cell(row, 1).coordinate
            assets = _curated_gallery_assets(
                record,
                identity_images.get(CellRef(pavilion_sheet.title, coordinate), ()),
            )
            if not assets:
                raise ValueError("ALMA_IMAGE_COVERAGE")
            reference = source_ref(
                identity_source.sha256, pavilion_sheet.title, coordinate
            )
            candidates.extend(
                _AlmaImageCandidate(
                    record["identity"], asset, "exact_xlsx", reference
                )
                for asset in assets
            )

        sources = {_KUN_PATH: identity_source, _KUN_PRICE_PATH: price_source}
        galleries = {_KUN_PATH: identity_images, _KUN_PRICE_PATH: price_images}
        for record in records:
            override = _FAMILY_ASSET_OVERRIDES.get(_image_family_key(record))
            if override is None:
                continue
            path, sheet_name, coordinate, expected_sha = override
            asset = _require_family_override_asset(
                galleries[path].get(CellRef(sheet_name, coordinate), ()),
                expected_sha,
            )
            reference = source_ref(sources[path].sha256, sheet_name, coordinate)
            candidates.append(
                _AlmaImageCandidate(
                    record["identity"],
                    asset,
                    "family_xlsx",
                    reference,
                    "curated_family_override",
                )
            )
        return candidates
    finally:
        identity_workbook.close()
        price_workbook.close()


def _configuration_labels(sheet, index, first_row: int, last_row: int, extra_columns=()) -> dict[int, str]:
    labels = {}
    for row in range(first_row, last_row + 1):
        parts = []
        for column in (3, 4, *extra_columns):
            value = _value(sheet, index, row, column)
            if column in extra_columns and _money(value) is not None:
                continue
            text = _plain(value)
            if text and text not in parts and not _blank_or_dash(value):
                parts.append(text)
        labels[row] = " | ".join(parts)[:1000] or f"Configuración fila {row}"
    counts = Counter(labels.values())
    return {
        row: label if counts[label] == 1 else f"{label[:970]} (fila {row})"
        for row, label in labels.items()
    }


def _parse_mondecasa_products(source, sheet, images, formulas) -> list[dict]:
    index = _merge_index(sheet)
    records = []
    collection = ""
    headers = {}
    row = 1
    while row <= sheet.max_row:
        if _mondecasa_header(sheet, row):
            collection = _plain(sheet.cell(row, 2).value)
            headers = {column: _plain(sheet.cell(row, column).value) for column in range(5, 9)}
            row += 1
            continue
        if not headers:
            row += 1
            continue
        block = _mondecasa_product_block(index, row)
        if block and block[0] < row:
            row += 1
            continue
        last_row = block[1] if block and block[0] == row else row
        code = _value(sheet, index, row, 1)
        description = next(
            (_value(sheet, index, candidate, 3) for candidate in range(row, last_row + 1) if _plain(_value(sheet, index, candidate, 3))),
            None,
        )
        dimensions = next(
            (_value(sheet, index, candidate, 4) for candidate in range(row, last_row + 1) if _plain(_value(sheet, index, candidate, 4))),
            None,
        )
        if not (_plain(code) or _plain(description)):
            row = last_row + 1
            continue

        grouped = last_row > row
        labels = _configuration_labels(sheet, index, row, last_row, (6,))
        warnings = []
        evidence = []
        rejected_refs = []
        base_options = []
        add_ons = []
        direct_price = None
        for price_row in range(row, last_row + 1):
            direct_value = sheet.cell(price_row, 5).value
            price = _money(direct_value)
            coordinate = sheet.cell(price_row, 5).coordinate
            reference = source_ref(source.sha256, sheet.title, coordinate)
            base_id = None
            if price is not None:
                if grouped:
                    base_id = f"base-r{price_row}-c5"
                    base_options.append(_base_option(base_id, labels[price_row], price))
                    evidence.append(_price_evidence("base", labels[price_row], price, True, reference))
                else:
                    direct_price = price
                    evidence.append(_price_evidence("direct", headers[5], price, True, reference))
            elif _invalid_price(direct_value, CellRef(sheet.title, coordinate), formulas):
                warnings.append("Se ignoró un precio base no numérico, negativo o fuera de rango.")
                rejected_refs.append(reference)

            for column in (6, 7, 8):
                value = sheet.cell(price_row, column).value
                option_price = _money(value)
                option_coordinate = sheet.cell(price_row, column).coordinate
                option_reference = source_ref(source.sha256, sheet.title, option_coordinate)
                if option_price is not None:
                    name = headers[column] or f"Cojín opción {sheet.cell(price_row, column).column_letter}"
                    available = not grouped or base_id is not None
                    option_id = f"cushion-r{price_row}-c{column}" if grouped else f"cushion-c{column}"
                    add_ons.append(
                        _add_on(
                            option_id,
                            name,
                            "cushion",
                            option_price,
                            available,
                            [base_id] if grouped and base_id else None,
                        )
                    )
                    evidence.append(_price_evidence("add_on", name, option_price, available, option_reference))
                    if not available:
                        warnings.append("Agregado sin precio base compatible; verificar antes de cotizar.")
                elif _invalid_price(value, CellRef(sheet.title, option_coordinate), formulas):
                    warnings.append("Se ignoró un precio de cojín no numérico, negativo o fuera de rango.")
                    rejected_refs.append(option_reference)

        config_refs = (
            [source_ref(source.sha256, sheet.title, f"C{row}:F{last_row}")]
            if grouped
            else []
        )
        records.append(
            _record(
                source=source,
                sheet=sheet,
                index=index,
                row=row,
                brand="Mondecasa",
                collection=collection,
                code=code,
                description=description,
                dimensions=dimensions,
                direct_price=direct_price,
                base_options=base_options,
                add_on_options=add_ons,
                price_evidence=evidence,
                source_images=_source_images(images, source.sha256, sheet, index, row, last_row, 2),
                warnings=warnings,
                extra_refs=config_refs + rejected_refs,
            )
        )
        row = last_row + 1
    return records


def _add_on_family(header: str, *, header_row: int, column: int) -> tuple[str | None, bool]:
    folded = _fold(header)
    if not folded or "qty" in folded or "loading" in folded or "carton" in folded:
        return None, False
    if "roof" in folded or "shade" in folded:
        return "roof", True
    if "floor" in folded:
        return "floor", True
    if "panel" in folded and ("side" in folded or "back" in folded):
        return f"panel-{_slug(header)}", True
    if ("curtain" in folded or "blind" in folded) and "side" in folded:
        return "side-curtain", True
    if "cushion" in folded or "foam" in folded:
        return "cushion", True
    return f"review-pavilions-r{header_row}-c{column}", False


def _parse_pavilions(source, sheet, images, formulas) -> list[dict]:
    index = _merge_index(sheet)
    records = []
    headers = {}
    header_row = 0
    row = 1
    while row <= sheet.max_row:
        if _pavilions_header(sheet, row):
            header_row = row if _plain(sheet.cell(row, 5).value) else row + 1
            headers = {
                column: _plain(sheet.cell(header_row, column).value)
                for column in range(5, sheet.max_column + 1)
            }
            row = header_row + 1
            continue
        if not headers:
            row += 1
            continue
        block = _vertical_code_block(index, row)
        if block and block[0] < row:
            row += 1
            continue
        last_row = block[1] if block and block[0] == row else row
        code = _value(sheet, index, row, 1)
        description = next(
            (_value(sheet, index, candidate, 3) for candidate in range(row, last_row + 1) if _plain(_value(sheet, index, candidate, 3))),
            None,
        )
        dimensions = next(
            (_value(sheet, index, candidate, 4) for candidate in range(row, last_row + 1) if _plain(_value(sheet, index, candidate, 4))),
            None,
        )
        if not (_plain(code) or _plain(description)):
            row = last_row + 1
            continue
        grouped = last_row > row
        labels = _configuration_labels(sheet, index, row, last_row)
        warnings = []
        evidence = []
        rejected_refs = []
        base_options = []
        add_ons = []
        direct_price = None
        for price_row in range(row, last_row + 1):
            direct_value = sheet.cell(price_row, 5).value
            price = _money(direct_value)
            direct_coordinate = sheet.cell(price_row, 5).coordinate
            direct_reference = source_ref(source.sha256, sheet.title, direct_coordinate)
            base_id = None
            if price is not None:
                if grouped:
                    base_id = f"base-r{price_row}-c5"
                    base_options.append(_base_option(base_id, labels[price_row], price))
                    evidence.append(_price_evidence("base", labels[price_row], price, True, direct_reference))
                else:
                    direct_price = price
                    evidence.append(
                        _price_evidence("direct", headers.get(5) or "Precio base", price, True, direct_reference)
                    )
            elif _invalid_price(direct_value, CellRef(sheet.title, direct_coordinate), formulas):
                warnings.append("Se ignoró un precio base no numérico, negativo o fuera de rango.")
                rejected_refs.append(direct_reference)

            for column in range(6, sheet.max_column + 1):
                header = headers.get(column, "")
                family, header_available = _add_on_family(header, header_row=header_row, column=column)
                if family is None:
                    continue
                value = sheet.cell(price_row, column).value
                option_price = _money(value)
                coordinate = sheet.cell(price_row, column).coordinate
                option_reference = source_ref(source.sha256, sheet.title, coordinate)
                available = header_available and (not grouped or base_id is not None)
                if option_price is not None:
                    name = header if header_available else "Agregado por verificar"
                    option_id = (
                        f"addon-r{price_row}-c{column}"
                        if grouped
                        else f"addon-r{header_row}-c{column}"
                    )
                    add_ons.append(
                        _add_on(
                            option_id,
                            name,
                            family,
                            option_price,
                            available,
                            [base_id] if grouped and base_id else None,
                        )
                    )
                    evidence.append(
                        _price_evidence("add_on", name, option_price, available, option_reference)
                    )
                    if not available:
                        warnings.append("Agregado conservado como no disponible hasta clasificación Admin.")
                elif _invalid_price(value, CellRef(sheet.title, coordinate), formulas):
                    warnings.append("Se ignoró un agregado no numérico, negativo o fuera de rango.")
                    rejected_refs.append(option_reference)

        config_refs = (
            [source_ref(source.sha256, sheet.title, f"C{row}:E{last_row}")]
            if grouped
            else []
        )
        records.append(
            _record(
                source=source,
                sheet=sheet,
                index=index,
                row=row,
                brand="Mondecasa",
                collection="PAVILIONS",
                code=code,
                description=description,
                dimensions=dimensions,
                direct_price=direct_price,
                base_options=base_options,
                add_on_options=add_ons,
                price_evidence=evidence,
                source_images=_source_images(images, source.sha256, sheet, index, row, last_row, 2),
                warnings=warnings,
                extra_refs=config_refs + rejected_refs,
            )
        )
        row = last_row + 1
    return records


def _parse_mondecasa(source, data: bytes, *, images=None) -> list[dict]:
    images = images if images is not None else extract_xlsx_images_from_bytes(data)
    workbook = open_xlsx_data_only_from_bytes(data)
    try:
        _require_sheets(workbook, _MONDECASA_SHEETS)
        _require_mondecasa_headers(workbook)
        formulas = set(workbook.formula_cells)
        return _parse_mondecasa_products(
            source, workbook["MONDECASA"], images, formulas
        ) + _parse_pavilions(source, workbook["PAVILIONS"], images, formulas)
    finally:
        workbook.close()


def _blocked_options(options: list[dict]) -> list[dict]:
    return [{**option, "price_net": "0.000000", "available": False} for option in options]


def _kun_internal_id(record: dict) -> str:
    identity_hash = hashlib.sha256(record["identity"].encode()).hexdigest()
    return f"alma:{_slug(record['brand'])}:variant:{identity_hash[:20]}"


def _embedded_record_assets(records: list[dict], images) -> dict[str, _AlmaResolvedImage]:
    assets_by_sha256 = {asset.sha256: asset for asset in images.values()}
    resolved = {}
    for record in records:
        source_images = record.get("source_images") or []
        source_shas = {row.get("sha256") for row in source_images if row.get("sha256")}
        if len(source_shas) != 1:
            continue
        asset = assets_by_sha256.get(next(iter(source_shas)))
        if asset is None:
            continue
        references = _unique_refs(
            [row["source"] for row in source_images if isinstance(row.get("source"), dict)]
        )
        resolved[record["identity"]] = _AlmaResolvedImage(
            asset,
            "exact_xlsx",
            tuple(references),
            "exact_embedded_raster",
        )
    return resolved


def _mondecasa_web_record_assets(
    records: list[dict],
    existing_assets: Mapping[str, _AlmaResolvedImage],
    link_index,
    image_index,
) -> dict[str, _AlmaResolvedImage]:
    normalized_assets: dict[str, ImageAsset] = {}
    resolved = {}
    for record in records:
        identity = record["identity"]
        if record["brand"] != "Mondecasa" or identity in existing_assets:
            continue
        link = resolve_mondecasa_link(
            record["source_code"], record["collection"], link_index
        )
        if link.status != "exact_index":
            continue
        references = link.evidence.get("references")
        if not isinstance(references, list) or not references:
            continue
        matches = [
            resolve_mondecasa_image(reference, image_index) for reference in references
        ]
        if any(match is None for match in matches):
            continue
        signatures = {
            (
                match.page_url,
                match.image_url,
                match.original_sha256,
                match.match_status,
            )
            for match in matches
        }
        match = matches[0]
        if (
            len(signatures) != 1
            or match.page_url != link.product_url
            or any(reference not in match.reference_numbers for reference in references)
        ):
            raise ValueError("ALMA_MONDECASA_WEB_IMAGE")
        asset = normalized_assets.get(match.original_sha256)
        if asset is None:
            asset = _normalize_image(match.asset_bytes)
            normalized_assets[match.original_sha256] = asset
        reason = (
            "official_mondecasa_product_page_gallery"
            if match.match_status == "exact_web"
            else "official_mondecasa_model_page_gallery"
        )
        resolved[identity] = _AlmaResolvedImage(
            asset,
            match.match_status,
            (
                {
                    "file_id": match.original_sha256,
                    "sheet_or_page": match.page_url,
                    "cell_or_bbox": match.image_url,
                },
            ),
            reason,
        )
    return resolved


def _finalize(
    records: list[dict],
    resolved_assets=None,
    link_index=None,
    mondecasa_link_index=None,
) -> list[dict]:
    kun_identities = Counter(
        record["identity"] for record in records if record["brand"] == "KUN"
    )
    if any(count > 1 for count in kun_identities.values()):
        raise ValueError("ALMA_IDENTITY_COLLISION")
    items = []
    for record in records:
        source_code = record["source_code"]
        has_base_price = record["direct_price"] is not None or any(
            option["available"] and Decimal(option["price_net"]) > 0
            for option in record["base_options"]
        )
        verified = bool(source_code and has_base_price)
        warnings = list(record["warnings"])
        if not source_code:
            warnings.append("Código ausente; producto publicado solo para revisión.")
        identity_hash = hashlib.sha256(record["identity"].encode()).hexdigest()
        if verified:
            internal_id = _kun_internal_id(record)
            product_key = f"{_slug(record['brand'])}:{_slug(source_code)}"
            sku = f"{_slug(record['brand'])}:{_slug(source_code)}:{identity_hash[:16]}"
            price = record["direct_price"]
            base_options = record["base_options"]
            add_on_options = record["add_on_options"]
        else:
            internal_id = _kun_internal_id(record)
            product_key = f"review:{identity_hash[:24]}"
            sku = ""
            price = None
            base_options = _blocked_options(record["base_options"])
            add_on_options = _blocked_options(record["add_on_options"])
        attributes = {
            "source_code": source_code,
            "dimensions": record["dimensions"],
            "price_evidence": record["price_evidence"],
            "source_images": record["source_images"],
        }
        attributes.update(record["extra_attributes"])
        product_url = ""
        image_kind = "placeholder"
        if record["brand"] == "KUN" and link_index is not None:
            link = resolve_kundesign_link(
                record["collection"], record["link_description"], link_index
            )
            product_url = link.product_url
            attributes["product_url_match"] = link.metadata
        elif record["brand"] == "Mondecasa" and mondecasa_link_index is not None:
            link = resolve_mondecasa_link(
                source_code,
                record["collection"],
                mondecasa_link_index,
            )
            product_url = link.product_url
            attributes["product_url_match"] = link.metadata
        asset_match = (resolved_assets or {}).get(record["identity"])
        if record["brand"] == "KUN" and resolved_assets is not None:
            if asset_match is None:
                raise ValueError("ALMA_IMAGE_COVERAGE")
        if asset_match is not None:
            image_kind = "official"
            if asset_match.match_status == "model_web":
                warnings.append(
                    "Imagen oficial del modelo Mondecasa; el acabado mostrado puede variar."
                )
            attributes["image_match"] = {
                "status": asset_match.match_status,
                "asset_sha256": asset_match.asset.sha256,
                "source_references": list(asset_match.source_references),
            }
            if asset_match.selection_reason:
                attributes["image_match"]["selection_reason"] = (
                    asset_match.selection_reason
                )
            attributes["approved_asset"] = {
                "bucket": "catalog-assets",
                "path": f"{asset_match.asset.sha256}.png",
                "image_kind": "official",
                "label": (
                    "Imagen oficial de la ficha Mondecasa"
                    if asset_match.match_status == "exact_web"
                    else "Imagen oficial del modelo Mondecasa"
                    if asset_match.match_status == "model_web"
                    else "Imagen oficial del XLSX ALMA"
                ),
                "approved": True,
            }
        items.append(
            {
                "internal_id": internal_id,
                "supplier": "alma",
                "product_key": product_key,
                "sku": sku,
                "code_status": "verified" if verified else "needs_review",
                "brand": record["brand"],
                "collection": record["collection"],
                "name": record["name"],
                "description": record["description"],
                "unit": "PZA",
                "availability_type": "made_to_order",
                "stock": None,
                "lead_time": "Sobre pedido",
                "base_price_options": base_options,
                "add_on_options": add_on_options,
                "base_currency": "USD",
                "price_net": _money_text(price) if price is not None else "0.000000",
                "tax_rate": "0.160000",
                "attributes": attributes,
                "image_url": "",
                "image_kind": image_kind,
                "product_url": product_url,
                "warnings": list(dict.fromkeys(warnings)),
                "source_reference": json.dumps(
                    record["refs"], sort_keys=True, separators=(",", ":"), ensure_ascii=True
                ),
            }
        )
    items.sort(key=lambda item: item["internal_id"])
    return items


def _build_alma(files, *, include_assets: bool):
    bundle, source_data = _validated_bundle(files)
    link_index = load_kundesign_link_index()
    mondecasa_link_index = load_mondecasa_link_index()
    mondecasa_image_index = load_mondecasa_image_index()
    identity_galleries = (
        extract_xlsx_image_galleries_from_bytes(source_data[_KUN_PATH])
        if include_assets
        else None
    )
    price_galleries = (
        extract_xlsx_image_galleries_from_bytes(source_data[_KUN_PRICE_PATH])
        if include_assets
        else None
    )
    identity_images = (
        {
            reference: assets[0]
            for reference, assets in identity_galleries.items()
            if len(assets) == 1
        }
        if include_assets
        else None
    )
    records = _parse_kun(
        bundle[_KUN_PATH],
        source_data[_KUN_PATH],
        bundle[_KUN_PRICE_PATH],
        source_data[_KUN_PRICE_PATH],
        identity_images=identity_images,
    )
    kun_records = list(records)
    resolved_assets = None
    if include_assets:
        candidates = _kun_asset_candidates(
            bundle[_KUN_PATH],
            source_data[_KUN_PATH],
            bundle[_KUN_PRICE_PATH],
            source_data[_KUN_PRICE_PATH],
            kun_records,
            identity_galleries,
            price_galleries,
        )
        resolved_assets = _resolve_kun_asset_candidates(kun_records, candidates)
    mondecasa_images = (
        extract_xlsx_images_from_bytes(source_data[_MONDECASA_PATH])
        if include_assets
        else None
    )
    mondecasa_records = _parse_mondecasa(
        bundle[_MONDECASA_PATH],
        source_data[_MONDECASA_PATH],
        images=mondecasa_images,
    )
    records.extend(mondecasa_records)
    if include_assets:
        resolved_assets.update(_embedded_record_assets(mondecasa_records, mondecasa_images))
        resolved_assets.update(
            _mondecasa_web_record_assets(
                mondecasa_records,
                resolved_assets,
                mondecasa_link_index,
                mondecasa_image_index,
            )
        )
    if not records:
        raise ValueError("ALMA_EMPTY")
    snapshot = {
        "supplier": "alma",
        "source_hash": _source_hash(
            tuple(bundle.values()),
            link_index,
            mondecasa_link_index,
            mondecasa_image_index,
        ),
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "items": _finalize(
            records,
            resolved_assets,
            link_index,
            mondecasa_link_index,
        ),
    }
    load_supplier_catalog_data(snapshot, expected_supplier="alma")
    if not include_assets:
        return snapshot
    bindings = tuple(
        sorted(
            (
                AlmaAssetBinding(
                    _kun_internal_id(record),
                    resolved_assets[record["identity"]].asset.sha256,
                    f"{resolved_assets[record['identity']].asset.sha256}.png",
                    "official",
                    resolved_assets[record["identity"]].match_status,
                    resolved_assets[record["identity"]].source_references,
                )
                for record in records
                if record["identity"] in resolved_assets
            ),
            key=lambda binding: binding.internal_id,
        )
    )
    assets = {
        match.asset.sha256: match.asset for match in resolved_assets.values()
    }
    return AlmaSnapshotBuild(snapshot, assets, bindings)


def build_alma_snapshot(files) -> dict:
    return _build_alma(files, include_assets=False)


def build_alma_snapshot_with_assets(files) -> AlmaSnapshotBuild:
    return _build_alma(files, include_assets=True)
