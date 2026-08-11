"""Parser determinista de evidencia textual de los catálogos IDÉLIKA.

La salida de este módulo todavía no es un snapshot publicable.  Son filas de
evidencia inmutables que conservan el PDF y la página de origen para que la
etapa posterior pueda construir y validar el SPEC Guide.

``extract_idelika_rows`` acepta documentos del sincronizador (con
``local_path``) y fixtures de texto (con ``pages``).  En ambos casos cada
página se representa por ``PdfPage`` o por un objeto estructural equivalente
con ``page_number`` y ``text``.
"""

from __future__ import annotations

import hashlib
import json
import re
import tempfile
import unicodedata
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any

import fitz
from openpyxl import load_workbook

from mobiliti_saas.quote_engine.supplier_catalog import (
    PUBLIC_ITEM_FIELDS,
    load_supplier_catalog_data,
)

from .common import CatalogSnapshotBuild, PdfPage


@dataclass(frozen=True)
class IdelikaEvidenceRow:
    subcatalog: str
    source_file: str
    source_page: int
    stable_key: str
    sku: str | None
    product: str
    family: str | None
    variant: str | None
    material: str | None
    dimensions: str | None
    description: str
    unit: str
    cost_mxn: Decimal | None
    reference_price_mxn: Decimal | None
    original_price_text: str | None
    price_status: str
    quotable: bool
    minimum_order: Decimal | None
    source_url: str
    identity_hash: str
    notes: str | None


@dataclass(frozen=True)
class _DocumentEvidence:
    subcatalog: str
    source_file: str
    source_url: str
    pages: tuple[PdfPage, ...]


@dataclass(frozen=True)
class _PriceEvidence:
    cost: Decimal
    reference: Decimal | None
    original: str
    variant: str | None
    dimensions: str | None


@dataclass(frozen=True)
class _ParsedProduct:
    product: str
    sku: str | None
    family: str | None
    variant: str | None
    material: str | None
    dimensions: str | None
    description: str
    unit: str
    minimum_order: Decimal | None
    prices: tuple[_PriceEvidence, ...]
    pending_price_text: str | None


@dataclass(frozen=True)
class _ImplicitAnchor:
    start: int
    marker: int
    title_indices: tuple[int, ...]
    titles: tuple[str, ...]


_SUBCATALOGS = {
    "fabricacion": "Fabricacion",
    "fabrication": "Fabricacion",
    "stock": "Stock",
    "inventory": "Stock",
    "school": "School Series",
    "school-series": "School Series",
    "school_series": "School Series",
    "school series": "School Series",
}
_SUBCATALOG_ORDER = {"Fabricacion": 0, "Stock": 1, "School Series": 2}
_FIELD_ALIASES = {
    "producto": "product",
    "product": "product",
    "sku": "sku",
    "codigo": "sku",
    "clave": "sku",
    "familia": "family",
    "coleccion": "family",
    "variante": "variant",
    "material": "material",
    "materiales": "material",
    "medida": "dimensions",
    "medidas": "dimensions",
    "dimensiones": "dimensions",
    "descripcion": "description",
    "unidad": "unit",
    "pedido minimo": "minimum_order",
    "minimo de compra": "minimum_order",
}
_PRODUCT_FIELD = re.compile(r"^\s*(?:producto|product)\s*:\s*(.+?)\s*$", re.IGNORECASE)
_PRICE_FIELD = re.compile(r"^\s*precios?\s*:\s*(.+?)\s*$", re.IGNORECASE)
_FIELD = re.compile(r"^\s*([^:]{2,40})\s*:\s*(.*?)\s*$")
_MONEY = re.compile(
    r"\$\s*(?P<amount>(?:\d{1,3}(?:[,.\s]\d{3})+|\d+)(?:\.\d{1,2})?)"
)
_DIMENSIONS = re.compile(
    r"(?i)(?:\d+(?:[.,]\d+)?\s*(?:\*|x|×)\s*)+\d+(?:[.,]\d+)?"
    r"|\b\d+(?:[.,]\d+)?\s*(?:cm|mm|m|metros?|di[aá]metro)\b"
)
_MINIMUM_ORDER = re.compile(
    r"(?i)(?:pedido\s+)?m[ií]nimo(?:\s+de\s+compra)?\s*:?[\s]*(\d+(?:[.,]\d+)?)"
    r"(?:\s*(?:piezas?|pzs?\.?|sets?))?"
)
_PRODUCTION_MARKER = re.compile(r"(?i)^(?:fabricaci[oó]n|importaci[oó]n)(?:\b|\s)")
_PENDING_PRICE = re.compile(
    r"(?i)(?:\bpor\s+confirmar\b|\bconsultar\s+(?:el\s+)?precio\b|"
    r"\bprecios?\s+disponibles?\b|\bprecio\s+por\s+proyecto\b|\bcot[ií]z)"
)
_AMBIGUOUS_PRICE_CONCEPT = re.compile(
    r"(?i)\b(?:flete|env[ií]o|iva|impuesto|subtotal|usd|d[oó]lares?)\b"
)
_PAIR_SEPARATOR = re.compile(r"(?i)^(?:[-–—/]|a)$")
_MATERIAL_WORDS = re.compile(
    r"(?i)\b(?:acero|aluminio|bejuco|cemento|cristal|cuerda|eucalipto|fibra|"
    r"laminado|loneta|madera|m[aá]rmol|melamina|mimbre|parota|piel|pino|"
    r"policarbonato|polipropileno|polycemento|polywood|rat[aá]n|rattan|tela|"
    r"teca|terrazo|textilene|tzalam|vinipiel|whicker|yute)\b"
)
_TITLE_NOUN = re.compile(
    r"(?i)\b(?:alfombra|banca|banco|base|bur[oó]|cama|camastro|colecci[oó]n|"
    r"colch[oó]n|columpio|comedor|cubierta|escritorio|mesa|pantalla|perchero|"
    r"p[eé]rgola|pintarr[oó]n|pupitre|rec[aá]mara|sala|seccional|set|silla|"
    r"sill[oó]n|sof[aá]|sombrilla|taburete|tapete|touch\s+all)\b"
)
_NON_TITLE_PREFIX = re.compile(
    r"(?i)^(?:acabado|acero|aluminio|beige|calidad|coj[ií]n|color(?:es)?|"
    r"con\b|cubierta\b|derecha\b|dise[nñ]o\b|estructura\b|fabricado\b|"
    r"electroest[aá]tica\b|madera|material(?:es)?|medidas?|negra?|no incluye|"
    r"pedido\b|producto\b|"
    r"resistente\b|uso\b|varios\b)"
)
_DECORATIVE = re.compile(
    r"(?i)^(?:id[eé]lika(?:\s*[|—-].*)?|cat[aá]logo\s+2026|p[aá]gina\s+\d+|"
    r"www\.|ventas@|zapopan showroom|playa del carmen showroom|"
    r"entra a www\.|reg[ií]strate en www\.)"
)


def _value(row: object, *names: str) -> Any:
    if isinstance(row, Mapping):
        for name in names:
            if name in row:
                return row[name]
        return None
    for name in names:
        if hasattr(row, name):
            return getattr(row, name)
    return None


def _clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\x00", " ")).strip()


def _fold(value: object) -> str:
    text = unicodedata.normalize("NFKD", _clean(value))
    return "".join(character for character in text if not unicodedata.combining(character)).casefold()


def _normalize_text(value: object) -> str:
    """Normaliza texto de contrato conservando palabras y eliminando acentos."""

    return _fold(value)


def _canonical(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _fold(value)).strip()


def _slug(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "-", _fold(value)).strip("-") or "sin-dato"


def _filename(value: object) -> str:
    text = _clean(value).replace("\\", "/")
    if not text:
        return ""
    # PurePath de la plataforma no reconoce necesariamente rutas del otro SO.
    return PurePosixPath(text).name


def _subcatalog(document: object, source_file: str) -> str:
    declared = _clean(_value(document, "subcatalog", "subcatalogo", "kind"))
    normalized = _fold(declared).replace("_", "-")
    if normalized in _SUBCATALOGS:
        return _SUBCATALOGS[normalized]
    filename = _fold(source_file)
    if "school" in filename:
        return "School Series"
    if "fabricacion" in filename:
        return "Fabricacion"
    if "stock" in filename:
        return "Stock"
    raise ValueError("IDELIKA_SUBCATALOG")


def _page(row: object, fallback: int) -> PdfPage:
    if isinstance(row, str):
        return PdfPage(fallback, row)
    number = _value(row, "page_number", "number", "source_page", "page")
    text = _value(row, "text", "page_text")
    if type(number) is not int or number < 1 or not isinstance(text, str):
        raise ValueError("IDELIKA_PAGE")
    return PdfPage(number, text, int(_value(row, "image_count") or 0))


def _read_pdf_pages(path: Path) -> tuple[PdfPage, ...]:
    """Extrae texto con pypdf sin materializar archivos intermedios."""

    try:
        from pypdf import PdfReader

        reader = PdfReader(path)
        if reader.is_encrypted:
            raise ValueError("IDELIKA_PDF")
        return tuple(
            PdfPage(page_number=index, text=page.extract_text() or "")
            for index, page in enumerate(reader.pages, start=1)
        )
    except ValueError:
        raise
    except Exception:
        raise ValueError("IDELIKA_PDF") from None


def _document(document: object) -> _DocumentEvidence:
    path_value = _value(document, "source_file", "path", "filename", "name")
    local_path = _value(document, "local_path")
    source_file = _filename(path_value) or _filename(local_path)
    if not source_file or Path(source_file).suffix.casefold() != ".pdf":
        raise ValueError("IDELIKA_DOCUMENT")
    subcatalog = _subcatalog(document, source_file)
    source_url = _clean(_value(document, "source_url", "web_url", "url"))

    declared_pages = _value(document, "pages")
    if declared_pages is not None:
        if isinstance(declared_pages, (str, bytes)):
            raise ValueError("IDELIKA_PAGES")
        try:
            pages = tuple(
                _page(candidate, index)
                for index, candidate in enumerate(declared_pages, start=1)
            )
        except TypeError:
            raise ValueError("IDELIKA_PAGES") from None
    else:
        if not isinstance(local_path, Path) or local_path.suffix.casefold() != ".pdf":
            raise ValueError("IDELIKA_DOCUMENT")
        pages = _read_pdf_pages(local_path)

    page_numbers = [page.page_number for page in pages]
    if len(page_numbers) != len(set(page_numbers)) or any(number < 1 for number in page_numbers):
        raise ValueError("IDELIKA_PAGES")
    return _DocumentEvidence(
        subcatalog=subcatalog,
        source_file=source_file,
        source_url=source_url,
        pages=tuple(sorted(pages, key=lambda row: row.page_number)),
    )


def _decimal(value: object) -> Decimal | None:
    text = _clean(value)
    if not text:
        return None
    normalized = text.replace(" ", "").replace(",", "")
    try:
        result = Decimal(normalized)
    except InvalidOperation:
        return None
    if not result.is_finite() or result <= 0:
        return None
    return result


def _amounts(text: str) -> tuple[Decimal, ...]:
    amounts = []
    for match in _MONEY.finditer(text):
        amount = _decimal(match.group("amount"))
        if amount is not None:
            amounts.append(amount)
    return tuple(amounts)


def _variant_after_price(text: str) -> str | None:
    matches = tuple(_MONEY.finditer(text))
    if not matches:
        return None
    suffix = _clean(text[matches[-1].end() :])
    suffix = re.sub(r"^[\s|/,:;()—–-]+", "", suffix).strip()
    suffix = re.sub(r"(?i)\bMXN\b", "", suffix).strip(" |/,:;()—–-")
    return suffix or None


def _is_pending_price(value: object) -> bool:
    return _PENDING_PRICE.search(_clean(value)) is not None


def _pending_price_text(lines: Sequence[str], *, explicit: bool) -> str | None:
    for index, line in enumerate(lines):
        text = _clean(line)
        match = _PRICE_FIELD.match(text) if explicit else None
        evidence = _clean(match.group(1)) if match else text
        if not _is_pending_price(evidence):
            continue
        if (
            _fold(evidence).endswith("precios disponibles en")
            and index + 1 < len(lines)
            and _fold(lines[index + 1]).startswith("www.")
        ):
            evidence = _clean(f"{evidence} {lines[index + 1]}")
        return evidence
    return None


def _price_context_is_unambiguous(text: str, amount_count: int) -> bool:
    original = _clean(text)
    if _AMBIGUOUS_PRICE_CONCEPT.search(original):
        return False
    if amount_count == 1:
        return True
    if amount_count != 2:
        return False
    matches = tuple(_MONEY.finditer(original))
    if len(matches) != 2:
        return False
    between = _clean(original[matches[0].end() : matches[1].start()])
    between = re.sub(r"(?i)\b(?:MXN|pesos?)\b", "", between).strip()
    return _PAIR_SEPARATOR.fullmatch(between) is not None


def _price_evidence(
    text: str,
    *,
    variant: str | None = None,
    dimensions: str | None = None,
) -> _PriceEvidence | None:
    original = _clean(text)
    values = _amounts(original)
    if len(values) not in {1, 2} or not _price_context_is_unambiguous(original, len(values)):
        return None
    cost = min(values)
    reference = max(values) if len(values) == 2 else None
    return _PriceEvidence(
        cost=cost,
        reference=reference,
        original=original,
        variant=_clean(variant) or _variant_after_price(original),
        dimensions=_clean(dimensions) or None,
    )


def _price_lines(lines: Sequence[str], *, explicit: bool) -> tuple[tuple[_PriceEvidence, ...], bool]:
    prices: list[_PriceEvidence] = []
    ambiguous = False
    last_dimensions: str | None = None
    index = 0
    while index < len(lines):
        line = _clean(lines[index])
        dimension_match = _DIMENSIONS.search(line)
        first_money = _MONEY.search(line)
        if dimension_match:
            if first_money is None:
                last_dimensions = line
            elif dimension_match.start() < first_money.start():
                last_dimensions = _clean(line[: first_money.start()])
        match = _PRICE_FIELD.match(line) if explicit else None
        candidate = match.group(1) if match else line
        if (explicit and match is None) or (not explicit and "$" not in candidate):
            index += 1
            continue

        if match is not None and _is_pending_price(candidate):
            index += 1
            continue

        original = _clean(candidate)
        values = _amounts(original)
        if len(values) == 1 and index + 1 < len(lines):
            next_line = _clean(lines[index + 1])
            if next_line.startswith("$") and len(_amounts(next_line)) == 1:
                original = _clean(f"{original} {next_line}")
                values = _amounts(original)
                index += 1
        if len(values) > 2 or not values:
            ambiguous = True
            index += 1
            continue
        price = _price_evidence(original, dimensions=last_dimensions)
        if price is None:
            ambiguous = True
        else:
            prices.append(price)
            last_dimensions = None
        index += 1
    return tuple(prices), ambiguous


def _prices_have_proven_variants(prices: Sequence[_PriceEvidence]) -> bool:
    if len(prices) <= 1:
        return True
    identities = []
    for price in prices:
        identity = (_canonical(price.variant), _canonical(price.dimensions))
        if identity == ("", "") or identity in identities:
            return False
        identities.append(identity)
    return True


def _field_name(label: str) -> str | None:
    return _FIELD_ALIASES.get(_canonical(label))


def _explicit_blocks(lines: Sequence[str]) -> tuple[tuple[str, ...], ...]:
    starts = [index for index, line in enumerate(lines) if _PRODUCT_FIELD.match(line)]
    return tuple(
        tuple(lines[start : starts[position + 1] if position + 1 < len(starts) else len(lines)])
        for position, start in enumerate(starts)
    )


def _minimum_order(text: object) -> Decimal | None:
    match = _MINIMUM_ORDER.search(_clean(text))
    return _decimal(match.group(1).replace(",", ".")) if match else None


def _parse_explicit(block: Sequence[str]) -> _ParsedProduct | None:
    fields: dict[str, str] = {}
    free_text: list[str] = []
    for line in block:
        match = _FIELD.match(line)
        if match:
            name = _field_name(match.group(1))
            if name is not None and name not in fields:
                fields[name] = _clean(match.group(2))
                continue
            if _PRICE_FIELD.match(line):
                continue
        if not _DECORATIVE.match(_clean(line)):
            free_text.append(_clean(line))

    product = fields.get("product", "")
    if not product:
        return None
    prices, ambiguous = _price_lines(block, explicit=True)
    pending_price_text = _pending_price_text(block, explicit=True)
    if (
        ambiguous
        or (prices and pending_price_text is not None)
        or not _prices_have_proven_variants(prices)
    ):
        return None
    description = fields.get("description") or _clean(" ".join(free_text)) or product
    minimum_order = _decimal(fields.get("minimum_order"))
    if minimum_order is None:
        minimum_order = _minimum_order(" ".join(block))
    return _ParsedProduct(
        product=product,
        sku=fields.get("sku") or None,
        family=fields.get("family") or None,
        variant=fields.get("variant") or None,
        material=fields.get("material") or None,
        dimensions=fields.get("dimensions") or None,
        description=description,
        unit=fields.get("unit") or "pieza",
        minimum_order=minimum_order,
        prices=prices,
        pending_price_text=pending_price_text,
    )


def _looks_like_title(line: str) -> bool:
    text = _clean(line)
    if not text or len(text) > 90:
        return False
    if "$" in text or _DIMENSIONS.search(text) or _DECORATIVE.match(text):
        return False
    if _NON_TITLE_PREFIX.match(text) or _MATERIAL_WORDS.search(text):
        return False
    if text.endswith((".", ",", ";", ":")):
        return False
    return len(text.split()) <= 9


def _is_production_marker(line: str) -> bool:
    text = _clean(line)
    if not _PRODUCTION_MARKER.match(text) or len(text) > 80 or text.endswith("."):
        return False
    markers = re.findall(r"(?i)\b(?:fabricaci[oó]n|importaci[oó]n)\b", text)
    return len(markers) == 1


def _title_before_marker(lines: Sequence[str], marker: int) -> _ImplicitAnchor | None:
    lower = max(0, marker - 8)
    for index in range(marker - 1, lower - 1, -1):
        if "$" in lines[index] or _is_production_marker(lines[index]) or _DECORATIVE.match(lines[index]):
            lower = index + 1
            break
    candidates = [index for index in range(lower, marker) if _looks_like_title(lines[index])]
    positives = [index for index in candidates if _TITLE_NOUN.search(lines[index])]
    if not positives:
        return None

    first_positive = positives[0]
    first_words = set(_canonical(lines[first_positive]).split())
    if len(positives) > 1 and "set" not in first_words:
        # Dos rótulos completos consecutivos ("Mesa..." / "Silla...") son
        # productos separados; solo se aceptarán si los precios los nombran.
        title_indices = tuple(positives)
        titles = tuple(_clean(lines[index]) for index in positives)
    else:
        # Un nombre puede continuar en renglones cortos ("Pupitre" / "Mentor")
        # siempre después de un rótulo positivo, nunca por simple proximidad
        # antes de él (evita absorber domicilios y texto decorativo).
        start = positives[0]
        title_indices_list = [start]
        index = start + 1
        while index < marker and _looks_like_title(lines[index]):
            title_indices_list.append(index)
            index += 1
        title_indices = tuple(title_indices_list)
        titles = (_clean(" ".join(lines[index] for index in title_indices)),)
    return _ImplicitAnchor(
        start=min(title_indices),
        marker=marker,
        title_indices=title_indices,
        titles=titles,
    )


def _page_family(lines: Sequence[str]) -> str | None:
    for index, line in enumerate(lines):
        match = re.match(r"(?i)^colecci[oó]n\s*(.*)$", line)
        if not match:
            continue
        value = _clean(match.group(1))
        if not value and index + 1 < len(lines) and _looks_like_title(lines[index + 1]):
            value = _clean(lines[index + 1])
        if value and len(value) <= 80:
            return value
    return None


def _material(lines: Sequence[str]) -> str | None:
    matches = [_clean(line) for line in lines if _MATERIAL_WORDS.search(line) and "$" not in line]
    return _clean(" ".join(matches[:4])) or None


def _dimensions(lines: Sequence[str]) -> str | None:
    matches = [_clean(line) for line in lines if _DIMENSIONS.search(line) and "$" not in line]
    return _clean(" | ".join(matches)) or None


def _description(lines: Sequence[str], product: str) -> str:
    useful = []
    for line in lines:
        text = _clean(line)
        if (
            not text
            or _is_production_marker(text)
            or _DECORATIVE.match(text)
            or "$" in text
            or _fold(text).startswith("coleccion")
        ):
            continue
        useful.append(text)
    return _clean(" ".join(useful)) or product


def _bounded_implicit_end(lines: Sequence[str], marker: int, limit: int) -> int:
    saw_price = False
    for index in range(marker + 1, limit):
        line = lines[index]
        if "$" in line:
            saw_price = True
            continue
        if (
            saw_price
            and _looks_like_title(line)
            and _TITLE_NOUN.search(line)
            and not _fold(line).startswith("coleccion")
        ):
            return index
    return limit


def _price_matches_title(price: _PriceEvidence, title: str) -> bool:
    variant_words = set(_canonical(price.variant).split())
    title_words = set(_canonical(title).split())
    return bool(variant_words & title_words & set(_TITLE_NOUN.findall(_canonical(title))))


def _without_mapping_variant(price: _PriceEvidence) -> _PriceEvidence:
    return _PriceEvidence(
        cost=price.cost,
        reference=price.reference,
        original=price.original,
        variant=None,
        dimensions=price.dimensions,
    )


def _products_from_anchor(
    anchor: _ImplicitAnchor,
    content: Sequence[str],
    family: str | None,
) -> tuple[_ParsedProduct, ...]:
    prices, ambiguous = _price_lines(content, explicit=False)
    pending_price_text = _pending_price_text(content, explicit=False)
    if ambiguous or (prices and pending_price_text is not None):
        return ()

    material = _material(content)
    minimum_order = _minimum_order(" ".join(content))
    if len(anchor.titles) > 1:
        if len(prices) != len(anchor.titles) or pending_price_text is not None:
            return ()
        assignments: dict[int, _PriceEvidence] = {}
        for price in prices:
            matches = [
                index
                for index, title in enumerate(anchor.titles)
                if _price_matches_title(price, title)
            ]
            if len(matches) != 1 or matches[0] in assignments:
                return ()
            assignments[matches[0]] = _without_mapping_variant(price)
        if len(assignments) != len(anchor.titles):
            return ()
        return tuple(
            _ParsedProduct(
                product=title,
                sku=None,
                family=family,
                variant=None,
                material=material,
                dimensions=None,
                description=_description(content, title),
                unit="pieza",
                minimum_order=minimum_order,
                prices=(assignments[index],),
                pending_price_text=None,
            )
            for index, title in enumerate(anchor.titles)
        )

    if not _prices_have_proven_variants(prices):
        return ()
    product = anchor.titles[0]
    return (
        _ParsedProduct(
            product=product,
            sku=None,
            family=family,
            variant=None,
            material=material,
            dimensions=_dimensions(content),
            description=_description(content, product),
            unit="pieza",
            minimum_order=minimum_order,
            prices=prices,
            pending_price_text=pending_price_text,
        ),
    )


def _parse_implicit(lines: Sequence[str]) -> tuple[_ParsedProduct, ...]:
    markers = [index for index, line in enumerate(lines) if _is_production_marker(line)]
    anchors: list[_ImplicitAnchor] = []
    for marker in markers:
        anchor = _title_before_marker(lines, marker)
        if anchor is None or any(row.start == anchor.start for row in anchors):
            continue
        anchors.append(anchor)
    anchors.sort(key=lambda row: (row.start, row.marker))

    family = _page_family(lines)
    products: list[_ParsedProduct] = []
    for position, anchor in enumerate(anchors):
        limit = anchors[position + 1].start if position + 1 < len(anchors) else len(lines)
        end = _bounded_implicit_end(lines, anchor.marker, limit)
        content = tuple(
            lines[index]
            for index in range(anchor.start, end)
            if index not in anchor.title_indices
        )
        products.extend(_products_from_anchor(anchor, content, family))
    return tuple(products)


def _page_products(text: str) -> tuple[_ParsedProduct, ...]:
    lines = tuple(_clean(line) for line in text.replace("\r", "\n").split("\n") if _clean(line))
    explicit = _explicit_blocks(lines)
    if explicit:
        return tuple(product for block in explicit if (product := _parse_explicit(block)) is not None)
    return _parse_implicit(lines)


def _identity_material(
    document: _DocumentEvidence,
    page_number: int,
    product: _ParsedProduct,
    *,
    variant: str | None,
    dimensions: str | None,
) -> dict[str, object]:
    material: dict[str, object] = {
        "supplier": "idelika",
        "subcatalog": _canonical(document.subcatalog),
        "source_file": _canonical(document.source_file),
        "source_page": page_number,
        "sku": _canonical(product.sku) or None,
        "product": _canonical(product.product),
        "family": _canonical(product.family) or None,
        "variant": _canonical(variant) or None,
        "material": _canonical(product.material) or None,
        "dimensions": _canonical(dimensions) or None,
    }
    if product.sku is None:
        material["description"] = _canonical(product.description)
    return material


def _row(
    document: _DocumentEvidence,
    page_number: int,
    product: _ParsedProduct,
    price: _PriceEvidence | None,
) -> IdelikaEvidenceRow:
    variant = product.variant or (price.variant if price else None)
    # Una medida localizada junto a un precio pertenece a esa variante. No se
    # reemplaza con el agregado del bloque, que colapsaría opciones distintas.
    dimensions = price.dimensions if price else product.dimensions
    identity_data = _identity_material(
        document,
        page_number,
        product,
        variant=variant,
        dimensions=dimensions,
    )
    serialized = json.dumps(identity_data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    identity_hash = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
    sku_part = _slug(product.sku) if product.sku else "sin-sku"
    stable_key = (
        f"idelika:{_slug(document.subcatalog)}:{sku_part}:{identity_hash[:20]}"
    )

    notes = []
    if product.sku is None:
        notes.append("SKU no publicado; se conserva una identidad técnica estable.")
    if price is None:
        notes.append("Precio por confirmar.")
    return IdelikaEvidenceRow(
        subcatalog=document.subcatalog,
        source_file=document.source_file,
        source_page=page_number,
        stable_key=stable_key,
        sku=product.sku,
        product=product.product,
        family=product.family,
        variant=variant,
        material=product.material,
        dimensions=dimensions,
        description=product.description,
        unit=product.unit,
        cost_mxn=price.cost if price else None,
        reference_price_mxn=price.reference if price else None,
        original_price_text=price.original if price else product.pending_price_text,
        price_status="confirmado" if price else "precio_por_confirmar",
        quotable=True,
        minimum_order=product.minimum_order,
        source_url=document.source_url,
        identity_hash=identity_hash,
        notes=" ".join(notes) or None,
    )


def _unique_public_rows(rows: Sequence[IdelikaEvidenceRow]) -> tuple[IdelikaEvidenceRow, ...]:
    by_identity: dict[str, list[IdelikaEvidenceRow]] = {}
    by_key: dict[str, list[IdelikaEvidenceRow]] = {}
    for row in rows:
        by_identity.setdefault(row.identity_hash, []).append(row)
        by_key.setdefault(row.stable_key, []).append(row)
    conflicts = {
        row.identity_hash
        for row in rows
        if any(candidate != row for candidate in by_identity[row.identity_hash])
        or any(candidate.identity_hash != row.identity_hash for candidate in by_key[row.stable_key])
    }
    result = []
    seen_hashes = set()
    seen_keys = set()
    for row in rows:
        if (
            row.identity_hash in conflicts
            or row.identity_hash in seen_hashes
            or row.stable_key in seen_keys
        ):
            continue
        seen_hashes.add(row.identity_hash)
        seen_keys.add(row.stable_key)
        result.append(row)
    return tuple(result)


def extract_idelika_rows(documents: Iterable[object]) -> tuple[IdelikaEvidenceRow, ...]:
    """Convierte documentos PDF o páginas textuales en evidencia auditable.

    La función no fusiona productos por similitud. Cada bloque textual genera
    su propia identidad y las variantes solo se expanden cuando aparecen como
    precios separados dentro de ese mismo bloque.
    """

    if isinstance(documents, (str, bytes, Mapping)):
        raise ValueError("IDELIKA_BUNDLE")
    try:
        evidence = tuple(_document(document) for document in documents)
    except TypeError:
        raise ValueError("IDELIKA_BUNDLE") from None
    if not evidence:
        return ()

    ordered = sorted(
        evidence,
        key=lambda row: (_SUBCATALOG_ORDER[row.subcatalog], _fold(row.source_file)),
    )
    rows: list[IdelikaEvidenceRow] = []
    for document in ordered:
        for page in document.pages:
            for product in _page_products(page.text):
                if product.prices:
                    rows.extend(
                        _row(document, page.page_number, product, price)
                        for price in product.prices
                    )
                else:
                    rows.append(_row(document, page.page_number, product, None))
    return _unique_public_rows(rows)


class IdelikaSpecValidationError(ValueError):
    """Indica que el SPEC validado de IDÉLIKA fue alterado o quedó incompleto."""


_SPEC_ERROR_PREFIX = "IDELIKA_SPEC_VALIDATION_FAILED:"
_PDF_MIME = "application/pdf"
_IDELIKA_PDF_BUNDLE = {
    "IDELIKA/1 CATALOGO FABRICACION 2026B.pdf": ("catalog", "fabricacion"),
    "IDELIKA/2 CATALOGO STOCK 2026.pdf": ("inventory", "stock"),
    "IDELIKA/4 SCHOOL SERIES 2026.pdf": ("catalog", "school"),
}
_EXPECTED_SPEC_SHEETS = (
    "Consolidado",
    "Fabricacion",
    "Stock",
    "School Series",
    "Fuentes_Reglas",
)
_EXPECTED_SPEC_COLUMNS = (
    "Proveedor",
    "Subcatalogo",
    "Archivo_origen",
    "Pagina_origen",
    "Clave_estable",
    "SKU",
    "Estado_codigo",
    "Producto",
    "Familia",
    "Variante",
    "Material",
    "Medidas",
    "Descripcion",
    "Unidad",
    "Costo_MXN",
    "Precio_referencia_MXN",
    "Precio_original",
    "Estado_precio",
    "Cotizable",
    "Minimo_compra",
    "Imagen_referencia",
    "URL_fuente",
    "Identidad_hash",
    "Notas",
)
_IDELIKA_TRUST_MANIFEST_PATH = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "idelika_spec_contract.v1.json"
)
_IDELIKA_BUNDLED_SPEC_PATH = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "idelika_spec_guide.v1.xlsx"
)
_TRUST_MANIFEST_SCHEMA = "mobiliti.idelika_spec_contract"
_TRUST_MANIFEST_VERSION = 1
_TRUST_MANIFEST_FIELDS = frozenset({
    "schema",
    "version",
    "supplier",
    "normalized_spec_sha256",
    "source_hashes",
    "total_rows",
    "counts",
})


def _spec_fail(reason: str) -> None:
    raise IdelikaSpecValidationError(f"{_SPEC_ERROR_PREFIX}{reason}")


def _validated_source_bundle(documents: Sequence[object]) -> dict[str, Path]:
    """Verifica el contrato de los tres PDF oficiales antes de generar el SPEC."""

    if len(documents) != len(_IDELIKA_PDF_BUNDLE):
        _spec_fail("invalid_source_bundle")
    expected_hashes = _load_idelika_trust_manifest()["source_hashes"]
    sources: dict[str, Path] = {}
    for document in documents:
        path = getattr(document, "path", None)
        local_path = getattr(document, "local_path", None)
        expected = _IDELIKA_PDF_BUNDLE.get(path)
        if (
            expected is None
            or path in sources
            or getattr(document, "kind", None) != expected[0]
            or getattr(document, "mime_type", None) != _PDF_MIME
            or not isinstance(local_path, Path)
            or local_path.suffix.casefold() != ".pdf"
            or not local_path.is_file()
        ):
            _spec_fail("invalid_source_bundle")
        source_name = PurePosixPath(path).name
        if hashlib.sha256(local_path.read_bytes()).hexdigest() != expected_hashes[source_name]:
            _spec_fail(f"source_hash_mismatch:{_slug(source_name)}")
        sources[path] = local_path
    if set(sources) != set(_IDELIKA_PDF_BUNDLE):
        _spec_fail("invalid_source_bundle")
    return {
        logical_name: sources[logical_name]
        for logical_name in _IDELIKA_PDF_BUNDLE
    }


def build_idelika_spec_artifact(documents: Sequence[object]) -> Path:
    """Genera un SPEC/sidecar persistente desde el bundle PDF oficial.

    La carpeta única queda retenida en el temporal del sistema como evidencia
    auditable. No se copian los PDF descargados ni se ejecuta limpieza alguna;
    la retención puede acumular artefactos y requiere mantenimiento manual
    autorizado por separado.
    """

    sources = _validated_source_bundle(tuple(documents))
    try:
        from mobiliti_saas.worker.catalog_sync.tools.build_idelika_spec_guide import (
            build_spec_guide,
        )
        working_directory = Path(tempfile.mkdtemp(prefix="mobiliti-idelika-spec-"))
        generated = working_directory / "Spec guide-IDELIKA-2026.xlsx"
        build_spec_guide(
            fabricacion=sources["IDELIKA/1 CATALOGO FABRICACION 2026B.pdf"],
            stock=sources["IDELIKA/2 CATALOGO STOCK 2026.pdf"],
            school=sources["IDELIKA/4 SCHOOL SERIES 2026.pdf"],
            output=generated,
        )
        _load_validated_idelika_contract(generated)
        return generated
    except IdelikaSpecValidationError:
        raise
    except Exception as error:
        _spec_fail(f"spec_build_failed:{type(error).__name__}")
    raise AssertionError("unreachable")


def _rows_and_digest_from_source(
    spec_path: str | Path | Sequence[object],
) -> tuple[tuple[IdelikaEvidenceRow, ...], str]:
    """Carga un SPEC explícito o construye y valida uno desde los tres PDF."""

    if isinstance(spec_path, (str, Path)):
        return _load_validated_idelika_contract(spec_path)
    try:
        bundle = tuple(spec_path)
    except TypeError:
        _spec_fail("invalid_spec_bundle")
    if len(bundle) == 1:
        document = bundle[0]
        local_path = getattr(document, "local_path", None)
        if (
            getattr(document, "kind", None) != "spec_guide"
            or getattr(document, "mime_type", None)
            != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or not isinstance(local_path, Path)
            or local_path.suffix.casefold() != ".xlsx"
        ):
            _spec_fail("invalid_spec_bundle")
        return _load_validated_idelika_contract(local_path)

    generated = build_idelika_spec_artifact(bundle)
    return _load_validated_idelika_contract(generated)


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_decimal(value: Any, *, field: str, row_number: int) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        _spec_fail(f"invalid_decimal:{field}:row={row_number}")
    return None


def _cell_bool(value: Any, *, field: str, row_number: int) -> bool:
    if isinstance(value, bool):
        return value
    normalized = _normalize_text(str(value or ""))
    if normalized in {"si", "true", "1", "yes"}:
        return True
    if normalized in {"no", "false", "0"}:
        return False
    _spec_fail(f"invalid_boolean:{field}:row={row_number}")
    return False


def _normalized_spec_digest(path: Path) -> str:
    # Importación diferida: el constructor del SPEC reutiliza este importador.
    from mobiliti_saas.worker.catalog_sync.tools.build_idelika_spec_guide import (
        _Package,
        _normalized_digest,
    )

    package = _Package(path)
    try:
        return _normalized_digest(package)
    finally:
        close = getattr(package, "close", None)
        if callable(close):
            close()


def _load_idelika_trust_manifest() -> dict[str, Any]:
    """Carga la raíz de confianza versionada desde una ruta fija del paquete.

    El manifiesto obtiene su confianza de revisión de código separada del
    workbook/sidecar generado. No es una firma y no protege frente a una
    modificación autorizada del propio repositorio.
    """

    path = _IDELIKA_TRUST_MANIFEST_PATH
    if not path.is_file():
        _spec_fail("trust_manifest_missing")
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        _spec_fail("trust_manifest_invalid")
    if not isinstance(manifest, dict) or set(manifest) != _TRUST_MANIFEST_FIELDS:
        _spec_fail("trust_manifest_invalid")
    if (
        manifest.get("schema") != _TRUST_MANIFEST_SCHEMA
        or type(manifest.get("version")) is not int
        or manifest.get("version") != _TRUST_MANIFEST_VERSION
        or manifest.get("supplier") != "idelika"
        or not isinstance(manifest.get("normalized_spec_sha256"), str)
        or re.fullmatch(
            r"[0-9a-f]{64}", manifest["normalized_spec_sha256"]
        )
        is None
    ):
        _spec_fail("trust_manifest_invalid")

    source_hashes = manifest.get("source_hashes")
    expected_source_names = {
        PurePosixPath(logical_name).name
        for logical_name in _IDELIKA_PDF_BUNDLE
    }
    if (
        not isinstance(source_hashes, dict)
        or set(source_hashes) != expected_source_names
        or any(
            not isinstance(value, str)
            or re.fullmatch(r"[0-9a-f]{64}", value) is None
            for value in source_hashes.values()
        )
    ):
        _spec_fail("trust_manifest_invalid")

    counts = manifest.get("counts")
    total_rows = manifest.get("total_rows")
    if (
        not isinstance(counts, dict)
        or set(counts) != set(_SUBCATALOG_ORDER)
        or any(type(value) is not int or value < 0 for value in counts.values())
        or type(total_rows) is not int
        or total_rows <= 0
        or sum(counts.values()) != total_rows
    ):
        _spec_fail("trust_manifest_invalid")
    return manifest


def _load_validated_idelika_contract(
    spec_path: str | Path,
) -> tuple[tuple[IdelikaEvidenceRow, ...], str]:
    path = Path(spec_path).resolve()
    sidecar_path = path.with_suffix(".validation.json")
    if not path.is_file():
        _spec_fail("workbook_missing")
    if not sidecar_path.is_file():
        _spec_fail("sidecar_missing")

    trust_manifest = _load_idelika_trust_manifest()

    try:
        sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        _spec_fail("sidecar_invalid_json")
    if not isinstance(sidecar, dict):
        _spec_fail("sidecar_not_object")
    package_validation = sidecar.get("package_validation")
    determinism = sidecar.get("determinism")
    if (
        not isinstance(package_validation, dict)
        or package_validation.get("valid") is not True
    ):
        _spec_fail("package_not_validated")
    if not isinstance(determinism, dict) or determinism.get("passed") is not True:
        _spec_fail("determinism_not_validated")

    digest = _normalized_spec_digest(path)
    trust_digest = trust_manifest["normalized_spec_sha256"]
    if digest.lower() != trust_digest:
        _spec_fail("workbook_digest_mismatch")
    if (
        determinism.get("first_normalized_sha256") != trust_digest
        or determinism.get("second_normalized_sha256") != trust_digest
    ):
        _spec_fail("sidecar_digest_mismatch")
    if sidecar.get("source_hashes") != trust_manifest["source_hashes"]:
        _spec_fail("sidecar_source_hashes_mismatch")
    if sidecar.get("total_rows") != trust_manifest["total_rows"]:
        _spec_fail("sidecar_total_mismatch")
    if sidecar.get("counts") != trust_manifest["counts"]:
        _spec_fail("sidecar_counts_mismatch")

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if tuple(workbook.sheetnames) != _EXPECTED_SPEC_SHEETS:
            _spec_fail("sheet_contract_mismatch")
        sheet = workbook["Consolidado"]
        header = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        if header != _EXPECTED_SPEC_COLUMNS:
            _spec_fail("column_contract_mismatch")

        rows: list[IdelikaEvidenceRow] = []
        stable_keys: set[str] = set()
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(value not in (None, "") for value in values):
                continue
            record = dict(zip(_EXPECTED_SPEC_COLUMNS, values, strict=True))
            provider = _cell_text(record["Proveedor"])
            if _normalize_text(provider or "") != "idelika":
                _spec_fail(f"provider_mismatch:row={row_number}")
            stable_key = _cell_text(record["Clave_estable"])
            product = _cell_text(record["Producto"])
            source_file = _cell_text(record["Archivo_origen"])
            source_url = _cell_text(record["URL_fuente"])
            identity_hash = (_cell_text(record["Identidad_hash"]) or "").lower()
            if not stable_key or not product or not source_file or not source_url:
                _spec_fail(f"required_value_missing:row={row_number}")
            if stable_key in stable_keys:
                _spec_fail(f"duplicate_stable_key:{stable_key}")
            stable_keys.add(stable_key)
            if not re.fullmatch(r"[0-9a-f]{64}", identity_hash):
                _spec_fail(f"identity_hash_invalid:row={row_number}")

            try:
                source_page = int(record["Pagina_origen"])
            except (TypeError, ValueError):
                _spec_fail(f"source_page_invalid:row={row_number}")
            rows.append(
                IdelikaEvidenceRow(
                    subcatalog=_cell_text(record["Subcatalogo"]) or "",
                    source_file=source_file,
                    source_page=source_page,
                    stable_key=stable_key,
                    sku=_cell_text(record["SKU"]),
                    product=product,
                    family=_cell_text(record["Familia"]),
                    variant=_cell_text(record["Variante"]),
                    material=_cell_text(record["Material"]),
                    dimensions=_cell_text(record["Medidas"]),
                    description=_cell_text(record["Descripcion"]) or product,
                    unit=_cell_text(record["Unidad"]) or "PZA",
                    cost_mxn=_cell_decimal(
                        record["Costo_MXN"], field="Costo_MXN", row_number=row_number
                    ),
                    reference_price_mxn=_cell_decimal(
                        record["Precio_referencia_MXN"],
                        field="Precio_referencia_MXN",
                        row_number=row_number,
                    ),
                    original_price_text=_cell_text(record["Precio_original"]),
                    price_status=_cell_text(record["Estado_precio"]) or "pending",
                    quotable=_cell_bool(
                        record["Cotizable"], field="Cotizable", row_number=row_number
                    ),
                    minimum_order=_cell_decimal(
                        record["Minimo_compra"],
                        field="Minimo_compra",
                        row_number=row_number,
                    ),
                    source_url=source_url,
                    identity_hash=identity_hash,
                    notes=_cell_text(record["Notas"]),
                )
            )
    except IdelikaSpecValidationError:
        raise
    except Exception as exc:
        _spec_fail(f"workbook_read_failed:{type(exc).__name__}")
    finally:
        if workbook is not None:
            workbook.close()

    if trust_manifest["total_rows"] != len(rows):
        _spec_fail("workbook_total_mismatch")
    actual_counts: dict[str, int] = {}
    for row in rows:
        actual_counts[row.subcatalog] = actual_counts.get(row.subcatalog, 0) + 1
    if trust_manifest["counts"] != actual_counts:
        _spec_fail("workbook_counts_mismatch")
    return tuple(rows), digest


def load_validated_idelika_spec(
    spec_path: str | Path,
) -> tuple[IdelikaEvidenceRow, ...]:
    """Carga exclusivamente un SPEC IDÉLIKA validado y no alterado."""

    rows, _digest = _load_validated_idelika_contract(spec_path)
    return rows


def _money_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value.quantize(Decimal("0.000001")), "f")


_PDF_RENDER_SCALE = 1.25
_PDF_MIN_IMAGE_AREA_RATIO = 0.005


@dataclass(frozen=True)
class _PdfAssetMatch:
    asset: ImageAsset
    source_reference: dict[str, Any]


def _rect_payload(rect: fitz.Rect) -> list[float]:
    return [round(float(value), 3) for value in (rect.x0, rect.y0, rect.x1, rect.y1)]


def _pdf_title_rect(page: fitz.Page, product: str) -> fitz.Rect | None:
    matches = page.search_for(product)
    if matches:
        return fitz.Rect(min(matches, key=lambda rect: (rect.y0, rect.x0)))

    target = _canonical(product)
    if not target:
        return None
    candidates: list[tuple[int, float, float, fitz.Rect]] = []
    text_lines: list[tuple[str, fitz.Rect]] = []
    page_dict = page.get_text("dict")
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            line_text = "".join(str(span.get("text", "")) for span in spans)
            canonical_line = _canonical(line_text)
            if not canonical_line or not spans:
                continue
            rect = fitz.Rect(spans[0]["bbox"])
            for span in spans[1:]:
                rect.include_rect(fitz.Rect(span["bbox"]))
            text_lines.append((canonical_line, rect))
            if target in canonical_line:
                candidates.append(
                    (len(canonical_line) - len(target), rect.y0, rect.x0, rect)
                )
    if not candidates:
        compact_target = target.replace(" ", "")
        for start in range(len(text_lines)):
            for size in (2, 3):
                window = text_lines[start : start + size]
                if len(window) != size:
                    continue
                compact_text = "".join(text.replace(" ", "") for text, _rect in window)
                if compact_target not in compact_text:
                    continue
                rect = fitz.Rect(window[0][1])
                for _text, line_rect in window[1:]:
                    rect.include_rect(line_rect)
                candidates.append(
                    (len(compact_text) - len(compact_target), rect.y0, rect.x0, rect)
                )
    if not candidates:
        return None
    return fitz.Rect(min(candidates, key=lambda row: row[:3])[3])


def _pdf_product_images(page: fitz.Page) -> list[fitz.Rect]:
    page_area = page.rect.get_area()
    minimum_area = page_area * _PDF_MIN_IMAGE_AREA_RATIO
    found: dict[tuple[float, float, float, float], fitz.Rect] = {}
    for image in page.get_images(full=True):
        xref = image[0]
        for raw_rect in page.get_image_rects(xref):
            rect = fitz.Rect(raw_rect)
            rect.intersect(page.rect)
            if rect.is_empty or rect.is_infinite or rect.get_area() < minimum_area:
                continue
            aspect = rect.width / rect.height if rect.height else float("inf")
            if aspect > 3.5 and rect.height < page.rect.height * 0.2:
                continue
            key = tuple(round(float(value), 3) for value in rect)
            found[key] = rect
    return list(found.values())


def _expand_pdf_crop(
    page_rect: fitz.Rect,
    title_rect: fitz.Rect,
    image_rect: fitz.Rect,
) -> fitz.Rect:
    crop = fitz.Rect(
        min(title_rect.x0, image_rect.x0) - 48,
        min(title_rect.y0, image_rect.y0) - 12,
        max(title_rect.x1, image_rect.x1) + 48,
        max(title_rect.y1, image_rect.y1) + 12,
    )
    crop.intersect(page_rect)

    def bounded_interval(
        center: float,
        requested_size: float,
        lower: float,
        upper: float,
    ) -> tuple[float, float]:
        size = min(requested_size, upper - lower)
        start = max(lower, min(center - size / 2, upper - size))
        return start, start + size

    if crop.height and crop.width / crop.height > 2:
        crop.y0, crop.y1 = bounded_interval(
            (crop.y0 + crop.y1) / 2,
            crop.width / 2,
            page_rect.y0,
            page_rect.y1,
        )
    if crop.width and crop.height / crop.width > 1.6:
        crop.x0, crop.x1 = bounded_interval(
            (crop.x0 + crop.x1) / 2,
            crop.height / 1.6,
            page_rect.x0,
            page_rect.x1,
        )
    return crop


def _pdf_asset_match(
    page: fitz.Page,
    row: IdelikaEvidenceRow,
    source_sha256: str,
) -> _PdfAssetMatch:
    from .common import _normalize_image, source_ref

    title_rect = _pdf_title_rect(page, row.product)
    if title_rect is None:
        _spec_fail(f"pdf_title_not_found:{_slug(row.stable_key)}")

    images = _pdf_product_images(page)
    if not images:
        _spec_fail(f"pdf_image_not_found:{_slug(row.stable_key)}")

    def horizontal_gap(image_rect: fitz.Rect) -> float:
        if image_rect.x1 < title_rect.x0:
            return title_rect.x0 - image_rect.x1
        if title_rect.x1 < image_rect.x0:
            return image_rect.x0 - title_rect.x1
        return 0

    image_rect = min(
        images,
        key=lambda rect: (
            abs((title_rect.y0 + title_rect.y1) - (rect.y0 + rect.y1))
            + horizontal_gap(rect),
            rect.get_area(),
        ),
    )
    crop_rect = _expand_pdf_crop(page.rect, title_rect, image_rect)
    pixmap = page.get_pixmap(
        matrix=fitz.Matrix(_PDF_RENDER_SCALE, _PDF_RENDER_SCALE),
        clip=crop_rect,
        alpha=False,
    )
    asset = _normalize_image(pixmap.tobytes("png"))
    reference = source_ref(source_sha256, row.source_page, _rect_payload(crop_rect))
    reference.update(
        {
            "file": row.source_file,
            "product": row.product,
            "identity_hash": row.identity_hash,
            "url": row.source_url,
            "title_match": "exact_text",
            "title_bbox": _rect_payload(title_rect),
            "image_bbox": _rect_payload(image_rect),
            "crop_bbox": _rect_payload(crop_rect),
        }
    )
    return _PdfAssetMatch(asset=asset, source_reference=reference)


def _build_idelika_pdf_assets(
    rows: Sequence[IdelikaEvidenceRow],
    sources: Mapping[str, Path],
) -> tuple[dict[str, ImageAsset], dict[str, _PdfAssetMatch]]:
    source_paths = {_filename(name): path for name, path in sources.items()}
    documents: dict[str, tuple[fitz.Document, str]] = {}
    try:
        for source_file, path in source_paths.items():
            data = path.read_bytes()
            source_sha256 = hashlib.sha256(data).hexdigest()
            documents[source_file] = (
                fitz.open(stream=data, filetype="pdf"),
                source_sha256,
            )

        assets: dict[str, ImageAsset] = {}
        matches: dict[str, _PdfAssetMatch] = {}
        for row in rows:
            source = documents.get(row.source_file)
            if source is None:
                _spec_fail(f"pdf_source_missing:{_slug(row.source_file)}")
            document, source_sha256 = source
            if not 1 <= row.source_page <= document.page_count:
                _spec_fail(f"pdf_page_invalid:{_slug(row.stable_key)}")
            match = _pdf_asset_match(
                document[row.source_page - 1],
                row,
                source_sha256,
            )
            assets.setdefault(match.asset.sha256, match.asset)
            matches[row.stable_key] = match
        return assets, matches
    except IdelikaSpecValidationError:
        raise
    except Exception as error:
        _spec_fail(f"pdf_asset_failed:{type(error).__name__}")
    finally:
        for document, _source_sha256 in documents.values():
            document.close()
    raise AssertionError("unreachable")


def _public_idelika_items(
    rows: Sequence[IdelikaEvidenceRow],
) -> list[dict[str, Any]]:
    public_items: list[dict[str, Any]] = []
    for row in rows:
        warnings: list[str] = []
        code_status = "verified"
        if not row.sku:
            warnings.append("missing_code")
            code_status = "needs_review"
        if row.cost_mxn is None:
            warnings.append("price_pending")
        if not row.quotable:
            warnings.append("not_quotable")

        source_reference = {
            "file": row.source_file,
            "page": row.source_page,
            "url": row.source_url,
            "identity_hash": row.identity_hash,
        }
        item = {
            "internal_id": f"idelika:{row.stable_key}",
            "supplier": "idelika",
            "product_key": row.stable_key,
            "sku": row.sku or "",
            "code_status": code_status,
            "brand": "IDÉLIKA",
            "collection": row.subcatalog,
            "name": row.product,
            "description": row.description,
            "unit": row.unit,
            "availability_type": "unknown",
            "stock": None,
            "lead_time": "Por confirmar",
            "base_price_options": [],
            "add_on_options": [],
            "base_currency": "MXN",
            "price_net": _money_text(row.cost_mxn),
            "tax_rate": "0.160000",
            "attributes": {
                "family": row.family or "",
                "variant": row.variant or "",
                "material": row.material or "",
                "dimensions": row.dimensions or "",
                "reference_price_mxn": _money_text(row.reference_price_mxn),
                "original_price_text": row.original_price_text or "",
                "price_status": row.price_status,
                "quotable": row.quotable,
                "minimum_order": _money_text(row.minimum_order),
                "provenance": source_reference,
                "identity_hash": row.identity_hash,
            },
            "image_url": "",
            "image_kind": "placeholder",
            "product_url": row.source_url,
            "warnings": warnings,
            "source_reference": json.dumps(
                [source_reference],
                sort_keys=True,
                separators=(",", ":"),
                ensure_ascii=True,
            ),
        }
        if tuple(item) != tuple(PUBLIC_ITEM_FIELDS):
            _spec_fail("public_item_field_contract_mismatch")
        public_items.append(item)
    public_items.sort(key=lambda item: item["internal_id"])
    return public_items


def build_idelika_snapshot_with_assets(
    spec_path: str | Path | Sequence[object],
    *,
    synced_at: datetime | None = None,
) -> CatalogSnapshotBuild:
    """Compone el snapshot público únicamente desde el SPEC validado."""

    source: str | Path | tuple[object, ...] = spec_path
    pdf_sources: Mapping[str, Path] | None = None
    if not isinstance(spec_path, (str, Path)):
        source = tuple(spec_path)
        if len(source) == len(_IDELIKA_PDF_BUNDLE):
            pdf_sources = _validated_source_bundle(source)
            source = _IDELIKA_BUNDLED_SPEC_PATH

    rows, digest = _rows_and_digest_from_source(source)
    generated_at = synced_at or datetime.now(timezone.utc)
    if generated_at.tzinfo is None:
        generated_at = generated_at.replace(tzinfo=timezone.utc)
    items = _public_idelika_items(rows)
    assets: dict[str, ImageAsset] = {}
    bindings: list[CatalogAssetBinding] = []
    if pdf_sources is not None:
        from .common import CatalogAssetBinding

        assets, matches = _build_idelika_pdf_assets(rows, pdf_sources)
        for item in items:
            match = matches.get(item["product_key"])
            if match is None:
                _spec_fail(f"pdf_asset_binding_missing:{_slug(item['product_key'])}")
            asset = match.asset
            object_name = f"{asset.sha256}.png"
            item["image_kind"] = "official"
            item["attributes"]["image_sha256"] = asset.sha256
            item["attributes"]["image_width"] = asset.width
            item["attributes"]["image_height"] = asset.height
            item["attributes"]["image_match"] = {
                "status": "exact_pdf",
                "asset_sha256": asset.sha256,
                "source_references": [match.source_reference],
            }
            item["attributes"]["approved_asset"] = {
                "bucket": "catalog-assets",
                "path": object_name,
                "image_kind": "official",
                "label": "Recorte oficial de ficha PDF IDÉLIKA",
                "approved": True,
            }
            bindings.append(
                CatalogAssetBinding(
                    item["internal_id"],
                    asset.sha256,
                    object_name,
                    "official",
                    "exact_pdf",
                    (match.source_reference,),
                )
            )
    snapshot = {
        "supplier": "idelika",
        "source_hash": digest,
        "generated_at": generated_at.astimezone(timezone.utc).isoformat(),
        "items": items,
    }
    load_supplier_catalog_data(snapshot, expected_supplier="idelika")
    return CatalogSnapshotBuild(
        snapshot=snapshot,
        assets_by_sha256=assets,
        bindings=tuple(sorted(bindings, key=lambda binding: binding.internal_id)),
    )


def build_idelika_snapshot(
    spec_path: str | Path | Sequence[object],
    *,
    synced_at: datetime | None = None,
) -> dict[str, Any]:
    """Atajo compatible para consumidores que solo requieren el JSON."""

    return build_idelika_snapshot_with_assets(spec_path, synced_at=synced_at).snapshot


__all__ = (
    "IdelikaEvidenceRow",
    "IdelikaSpecValidationError",
    "build_idelika_spec_artifact",
    "build_idelika_snapshot",
    "build_idelika_snapshot_with_assets",
    "extract_idelika_rows",
    "load_validated_idelika_spec",
)
