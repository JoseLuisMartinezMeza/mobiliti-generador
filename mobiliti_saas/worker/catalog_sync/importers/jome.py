"""Importador determinista de los dos catálogos de costo oficiales JOME.

La interfaz de documentos es deliberadamente estructural: cada documento debe
aportar ``path``, ``kind``, ``sha256``, ``mime_type`` y ``local_path``. Es la
misma forma que entrega hoy el sincronizador, hasta que ``SourceDocument`` se
publique como tipo compartido.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl.utils.cell import coordinate_to_tuple

from .common import (
    CatalogAssetBinding,
    CatalogSnapshotBuild,
    CellRef,
    ImageAsset,
    extract_xlsx_images,
    neutralize_spreadsheet_text,
    open_xlsx_data_only,
    source_ref,
    validate_source_file,
)
from mobiliti_saas.quote_engine.supplier_catalog import load_supplier_catalog_data


_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_SHEETS = {
    "estructuras": "COSTO ESTRUCTURAS 2026",
    "laminado": "COSTO LAMINADO 2026",
}


def _plain(value: object) -> str:
    return re.sub(r"\s+", " ", neutralize_spreadsheet_text(value).strip())


def _slug(value: object) -> str:
    text = unicodedata.normalize("NFKD", _plain(value))
    text = "".join(char for char in text if not unicodedata.combining(char)).casefold()
    return re.sub(r"[^a-z0-9]+", "-", text).strip("-") or "sin-dato"


def _currency(value: object) -> str:
    return _plain(value).upper()


def _decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        result = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        result = Decimal(str(value))
    elif isinstance(value, str):
        text = value.strip().replace("$", "").replace(",", "")
        try:
            result = Decimal(text)
        except InvalidOperation:
            raise ValueError("JOME_COST") from None
    else:
        raise ValueError("JOME_COST")
    if not result.is_finite() or result <= 0:
        raise ValueError("JOME_COST")
    return result


def _source_hash(documents: tuple[object, ...]) -> str:
    material = "\n".join(
        f"{document.path}\0{document.sha256}"
        for document in sorted(documents, key=lambda row: row.path)
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def _validated_documents(documents: object) -> dict[str, object]:
    try:
        rows = tuple(documents)
    except TypeError:
        raise ValueError("JOME_BUNDLE") from None
    if len(rows) != len(_SHEETS):
        raise ValueError("JOME_BUNDLE")
    by_kind = {}
    for document in rows:
        source_kind = getattr(document, "kind", None)
        brand = getattr(document, "brand", None)
        kind = (
            brand
            if source_kind == "spec_guide" and brand in _SHEETS
            else source_kind
        )
        local_path = getattr(document, "local_path", None)
        declared_hash = getattr(document, "sha256", None)
        path = getattr(document, "path", None)
        if (
            kind not in _SHEETS
            or kind in by_kind
            or not isinstance(path, str)
            or Path(path).suffix.lower() != ".xlsx"
            or not isinstance(local_path, Path)
            or local_path.suffix.lower() != ".xlsx"
            or getattr(document, "mime_type", None) != _MIME
            or not isinstance(declared_hash, str)
            or re.fullmatch(r"[0-9a-f]{64}", declared_hash) is None
        ):
            raise ValueError("JOME_BUNDLE")
        validated = validate_source_file(local_path, ".xlsx")
        if validated.sha256 != declared_hash:
            raise ValueError("JOME_HASH")
        by_kind[kind] = document
    return by_kind


def _heading_value(row: tuple[object, ...], label: str) -> str | None:
    for value in row:
        text = _plain(value)
        match = re.match(rf"(?i)^{label}\s*:\s*(.+)$", text)
        if match:
            return match.group(1).strip()
    return None


def _is_block_heading(sheet, row_number: int) -> bool:
    code = _plain(sheet.cell(row_number, 2).value)
    description = _plain(sheet.cell(row_number, 3).value)
    if (
        code
        or not description
        or sheet.cell(row_number, 5).value is not None
        or "descrip" in _slug(description)
    ):
        return False
    for candidate in range(row_number + 1, min(sheet.max_row, row_number + 3) + 1):
        candidate_code = _plain(sheet.cell(candidate, 2).value)
        candidate_cost = sheet.cell(candidate, 5).value
        if candidate_code and candidate_cost is not None:
            return True
    return False


def _images_by_row(images: dict[CellRef, object], sheet_name: str):
    rows = []
    for reference, asset in images.items():
        if reference.sheet != sheet_name:
            continue
        row, column = coordinate_to_tuple(reference.cell)
        rows.append((row, column, reference, asset))
    return sorted(rows)


def _image_payload(asset, document, reference: CellRef) -> dict:
    return {
        "sha256": asset.sha256,
        "media_type": asset.media_type,
        "width": asset.width,
        "height": asset.height,
        "source_reference": source_ref(document.sha256, reference.sheet, reference.cell),
    }


def _parse_document(kind: str, document: object) -> list[dict]:
    expected_sheet = _SHEETS[kind]
    images = extract_xlsx_images(document.local_path)
    workbook = open_xlsx_data_only(document.local_path)
    try:
        if expected_sheet not in workbook.sheetnames:
            raise ValueError("JOME_SHEET")
        sheet = workbook[expected_sheet]
        image_rows = _images_by_row(images, expected_sheet)
        image_index = 0
        block_image = None
        system = "sin-sistema"
        block = "sin-bloque"
        inherited_code = ""
        records = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            found_system = _heading_value(row, "sistema")
            found_block = _heading_value(row, "bloque")
            found_family = _heading_value(row, "fami(?:lia|la)")
            if found_system:
                system = found_system
                inherited_code = ""
            if found_block:
                block = found_block
                inherited_code = ""
                block_image = None
            elif found_family:
                block = found_family
                inherited_code = ""
                block_image = None
            while image_index < len(image_rows) and image_rows[image_index][0] <= row_number:
                _, _, reference, asset = image_rows[image_index]
                block_image = _image_payload(asset, document, reference)
                image_index += 1

            code = _plain(sheet.cell(row_number, 2).value)
            description = _plain(sheet.cell(row_number, 3).value)
            dimensions = _plain(sheet.cell(row_number, 4).value)
            cost_value = sheet.cell(row_number, 5).value
            if _slug(code) in {"codigo", "cod", "clave"} and "descrip" in _slug(description):
                continue
            if _is_block_heading(sheet, row_number):
                block = description
                inherited_code = ""
                block_image = None
                continue
            if code:
                inherited_code = code
                description = description or code
            elif description and dimensions and cost_value is not None and block != "sin-bloque":
                code = inherited_code
            if not code or not description or cost_value is None:
                continue

            raw_cost = _decimal(cost_value)
            declared_currency = _currency(sheet.cell(row_number, 8).value)
            identity = ":".join(
                (
                    kind,
                    _slug(system),
                    _slug(block),
                    _slug(code),
                    _slug(dimensions),
                    str(row_number),
                )
            )
            provenance = {
                "file": document.path,
                "file_hash": document.sha256,
                "sheet": expected_sheet,
                "row": row_number,
                "code_cell": f"B{row_number}",
                "cost_cell": f"E{row_number}",
                "declared_currency_cell": f"H{row_number}",
                "declared_currency": declared_currency,
                "currency_normalization": (
                    "human_source_error_to_mxn" if declared_currency != "MXN" else None
                ),
            }
            record = {
                "identity": identity,
                "subcatalog": kind,
                "system": system,
                "block": block,
                "code": code,
                "name": description,
                "dimensions": dimensions,
                "raw_cost": raw_cost,
                "base_currency": "MXN",
                "provenance": provenance,
            }
            if block_image is not None:
                record["image"] = block_image
            records.append(record)
        return records
    finally:
        workbook.close()


def import_jome_catalog(documents, *, synced_at: datetime) -> dict:
    """Lee costos JOME de E, ignora I y publica sus costos canónicos en MXN."""
    if not isinstance(synced_at, datetime) or synced_at.tzinfo is None:
        raise ValueError("JOME_SYNCED_AT")
    bundle = _validated_documents(documents)
    rows = []
    for kind in sorted(_SHEETS):
        rows.extend(_parse_document(kind, bundle[kind]))
    if not rows:
        raise ValueError("JOME_EMPTY")
    rows.sort(key=lambda row: row["identity"])
    return {
        "supplier": "jome",
        "source_hash": _source_hash(tuple(bundle.values())),
        "generated_at": synced_at.isoformat().replace("+00:00", "Z"),
        "items": rows,
    }


def _money_text(value: Decimal) -> str:
    return f"{value:.6f}"


def _available_assets(documents: tuple[object, ...]) -> dict[str, ImageAsset]:
    assets: dict[str, ImageAsset] = {}
    for document in documents:
        for asset in extract_xlsx_images(document.local_path).values():
            assets[asset.sha256] = asset
    return assets


def _public_items(raw_items: list[dict], assets: dict[str, ImageAsset] | None):
    code_counts = Counter(item["code"].casefold() for item in raw_items)
    source_images = {
        (
            row["subcatalog"],
            row["system"].casefold(),
            row["code"].casefold(),
            row["dimensions"].casefold(),
        ): row["image"]
        for row in raw_items
        if row.get("image") is not None
    }
    items = []
    bindings = []
    selected_assets: dict[str, ImageAsset] = {}
    for row in raw_items:
        code = row["code"]
        duplicate_code = code_counts[code.casefold()] > 1
        source_reference = source_ref(
            row["provenance"]["file_hash"],
            row["provenance"]["sheet"],
            row["provenance"]["cost_cell"],
        )
        attributes = {
            "source_code": code,
            "subcatalog": row["subcatalog"],
            "system": row["system"],
            "block": row["block"],
            "dimensions": row["dimensions"],
            "raw_cost": _money_text(row["raw_cost"]),
            "provenance": row["provenance"],
        }
        warnings = []
        image_kind = "placeholder"
        image = row.get("image")
        if image is None and code.casefold().endswith("-28m"):
            image = source_images.get(
                (
                    row["subcatalog"],
                    row["system"].casefold(),
                    code[:-4].casefold(),
                    row["dimensions"].casefold(),
                )
            )
        if image is not None and assets is not None:
            asset = assets.get(image["sha256"])
            if asset is not None:
                image_kind = "official"
                object_name = f"{asset.sha256}.png"
                image_reference = image["source_reference"]
                attributes["image_match"] = {
                    "status": "family_xlsx",
                    "asset_sha256": asset.sha256,
                    "source_references": [image_reference],
                }
                attributes["approved_asset"] = {
                    "bucket": "catalog-assets",
                    "path": object_name,
                    "image_kind": "official",
                    "label": "Imagen oficial del bloque XLSX JOME",
                    "approved": True,
                }
                selected_assets[asset.sha256] = asset
                bindings.append(
                    CatalogAssetBinding(
                        row["identity"], asset.sha256, object_name, "official", "family_xlsx",
                        (image_reference,),
                    )
                )
        if duplicate_code:
            warnings.append("Codigo duplicado entre sistemas; SKU requiere revision.")
        items.append(
            {
                "internal_id": row["identity"],
                "supplier": "jome",
                "product_key": row["identity"],
                "sku": "" if duplicate_code else code,
                "code_status": "needs_review" if duplicate_code else "verified",
                "brand": "Jome",
                "collection": row["block"],
                "name": row["name"],
                "description": row["name"],
                "unit": "PZA",
                "availability_type": "made_to_order",
                "stock": None,
                "lead_time": "",
                "base_price_options": [],
                "add_on_options": [],
                "base_currency": "MXN",
                "price_net": _money_text(row["raw_cost"]),
                "tax_rate": "0.160000",
                "attributes": attributes,
                "image_url": "",
                "image_kind": image_kind,
                "product_url": "",
                "warnings": warnings,
                "source_reference": json.dumps(
                    [source_reference], sort_keys=True, separators=(",", ":"), ensure_ascii=True
                ),
            }
        )
    items.sort(key=lambda item: item["internal_id"])
    bindings.sort(key=lambda binding: binding.internal_id)
    return items, selected_assets, tuple(bindings)


def _build_jome(documents, *, synced_at: datetime | None, include_assets: bool):
    timestamp = synced_at or datetime.now(timezone.utc)
    raw = import_jome_catalog(documents, synced_at=timestamp)
    bundle = _validated_documents(documents)
    source_documents = tuple(bundle.values())
    assets = _available_assets(source_documents) if include_assets else None
    items, selected_assets, bindings = _public_items(raw["items"], assets)
    snapshot = {
        "supplier": "jome",
        "source_hash": raw["source_hash"],
        "generated_at": raw["generated_at"],
        "items": items,
    }
    load_supplier_catalog_data(snapshot, expected_supplier="jome")
    if include_assets:
        return CatalogSnapshotBuild(snapshot, selected_assets, bindings)
    return snapshot


def build_jome_snapshot(documents, *, synced_at: datetime | None = None) -> dict:
    """Convierte el registro JOME crudo al contrato público del catálogo."""
    return _build_jome(documents, synced_at=synced_at, include_assets=False)


def build_jome_snapshot_with_assets(
    documents, *, synced_at: datetime | None = None
) -> CatalogSnapshotBuild:
    """Construye el snapshot JOME publicable con sus activos OOXML aprobados."""
    return _build_jome(documents, synced_at=synced_at, include_assets=True)
