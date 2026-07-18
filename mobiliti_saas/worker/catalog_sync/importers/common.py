import hashlib
import io
import math
import multiprocessing
import os
import posixpath
import re
import stat
import time
import warnings
import zipfile
import zlib
from collections import Counter
from dataclasses import dataclass
from pathlib import Path, PurePosixPath, PureWindowsPath
from typing import Iterator, Literal, Mapping, Protocol, runtime_checkable
from urllib.parse import urlsplit
from xml.etree import ElementTree

import fitz
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.workbook.workbook import Workbook
from PIL import Image, ImageOps


MAX_FILE_BYTES = 67_108_864
MAX_ZIP_ENTRIES = 10_000
MAX_ZIP_EXPANDED_BYTES = 268_435_456
MAX_ZIP_RATIO = 200
MAX_WORKBOOK_SHEETS = 128
MAX_WORKBOOK_CELLS = 1_100_000
MAX_WORKBOOK_FORMULAS = 100_000
MAX_WORKBOOK_MERGED_RANGES = 10_000
MAX_MERGED_RANGE_CELLS = 100_000
MAX_WORKBOOK_MERGED_CELLS = 1_100_000
MAX_WORKBOOK_IMAGES = 10_000
MAX_WORKBOOK_IMAGES_PER_CELL = 16
MAX_WORKBOOK_UNIQUE_IMAGES = 2_000
MAX_WORKBOOK_IMAGE_PIXELS = 400_000_000
MAX_WORKBOOK_NORMALIZED_IMAGE_BYTES = 128 * 1024 * 1024
MAX_IMAGE_BYTES = 8 * 1024 * 1024
MAX_IMAGE_DIMENSION = 16_384
MAX_IMAGE_PIXELS = 40_000_000
MAX_PDF_PAGES = 2_000
MAX_PDF_XREFS = 100_000
MAX_PDF_STREAMS = 50_000
MAX_PDF_STREAM_RAW_BYTES = 8 * 1024 * 1024
MAX_PDF_STREAM_DECODED_BYTES = 64 * 1024 * 1024
MAX_PDF_STREAM_EXPANDED_BYTES = 192 * 1024 * 1024
MAX_PDF_STREAM_RATIO = 200
MAX_PDF_TEXT_BYTES = 64 * 1024 * 1024
MAX_PDF_TEXT_WORKER_BYTES = 1024 * 1024 * 1024
MAX_PDF_TEXT_SECONDS = 30
MAX_PDF_TEXT_CHUNK_CHARS = 16 * 1024
MAX_PDF_WORKER_SHUTDOWN_SECONDS = 1
MAX_PDF_IMAGES = 10_000
MAX_PDF_IMAGE_BYTES = 128 * 1024 * 1024
MAX_TEXT_LENGTH = 32_768

_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPE_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_RELATIONSHIP_ATTR_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_SUPPORTED_COMPRESSION = {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}
_SUPPORTED_IMAGE_FORMATS = {"BMP", "GIF", "JPEG", "PNG", "TIFF", "WEBP"}
_ALLOWED_CONTENT_TYPES = {
    "application/xml",
    "text/xml",
    "application/vnd.openxmlformats-package.relationships+xml",
    "application/vnd.openxmlformats-package.core-properties+xml",
    "application/vnd.openxmlformats-officedocument.extended-properties+xml",
    "application/vnd.openxmlformats-officedocument.custom-properties+xml",
    "application/vnd.openxmlformats-officedocument.customxmlproperties+xml",
    "application/vnd.openxmlformats-officedocument.theme+xml",
    "application/vnd.openxmlformats-officedocument.drawing+xml",
    "application/vnd.openxmlformats-officedocument.vmldrawing",
    "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
    "application/vnd.openxmlformats-officedocument.drawingml.chartshapes+xml",
    "application/vnd.openxmlformats-officedocument.drawingml.diagramcolors+xml",
    "application/vnd.openxmlformats-officedocument.drawingml.diagramdata+xml",
    "application/vnd.openxmlformats-officedocument.drawingml.diagramlayout+xml",
    "application/vnd.openxmlformats-officedocument.drawingml.diagramstyle+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.chartsheet+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedstrings+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.calcchain+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivottable+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotcachedefinition+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotcacherecords+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.querytable+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.printersettings",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheetmetadata+xml",
    "application/vnd.ms-office.chartstyle+xml",
    "application/vnd.ms-office.chartcolorstyle+xml",
    "application/vnd.ms-excel.threadedcomments+xml",
    "application/vnd.ms-excel.person+xml",
    "application/vnd.ms-excel.slicer+xml",
    "application/vnd.ms-excel.slicercache+xml",
    "application/vnd.ms-excel.timeline+xml",
    "application/vnd.ms-excel.timelinecache+xml",
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/bmp",
    "image/tiff",
    "image/webp",
    "image/x-emf",
}
_ALLOWED_RELATIONSHIPS = {
    "http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/custom-properties",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chartsheet",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chartUserShapes",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/printerSettings",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/queryTable",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXmlProps",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/metadata",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramColors",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramData",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramLayout",
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramQuickStyle",
    "http://schemas.microsoft.com/office/2007/relationships/slicer",
    "http://schemas.microsoft.com/office/2007/relationships/slicerCache",
    "http://schemas.microsoft.com/office/2011/relationships/slicer",
    "http://schemas.microsoft.com/office/2011/relationships/slicerCache",
    "http://schemas.microsoft.com/office/2011/relationships/timeline",
    "http://schemas.microsoft.com/office/2011/relationships/timelineCache",
    "http://schemas.microsoft.com/office/2017/10/relationships/threadedComment",
    "http://schemas.microsoft.com/office/2017/10/relationships/person",
}
_PASSIVE_EXTERNAL_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
)
_NULL_EXTERNAL_IMAGE_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)
_PDF_IMAGE_FILTERS = {
    "ASCII85Decode",
    "ASCIIHexDecode",
    "CCITTFaxDecode",
    "DCTDecode",
    "FlateDecode",
    "JBIG2Decode",
    "JPXDecode",
}
_ACTIVE_PDF_TOKEN = re.compile(
    rb"/(?:JavaScript|JS|OpenAction|AA|Launch|EmbeddedFiles|RichMedia|Filespec|GoToR|SubmitForm|ImportData|URI)\b"
)
_DDE_FORMULA = re.compile(r"(?:\bDDE\s*\(|\|[^!]{0,1024}!)", re.IGNORECASE)
_EXTERNAL_FORMULA = re.compile(r"\[[^\]]{1,1024}\][^!]{0,1024}!")
_FILE_ID = re.compile(r"[A-Za-z0-9_.:-]{1,128}\Z")
_CELL = re.compile(r"[A-Z]{1,3}[1-9][0-9]{0,6}(?::[A-Z]{1,3}[1-9][0-9]{0,6})?\Z")
_MAX_XLSX_ROW = 1_048_576
_MAX_XLSX_COLUMN = 16_384


class SourceSafetyError(ValueError):
    pass


@dataclass(frozen=True, order=True)
class CellRef:
    sheet: str
    cell: str

    @property
    def coordinate(self) -> str:
        return self.cell


@dataclass(frozen=True)
class ValidatedSource:
    extension: str
    size: int
    sha256: str


@dataclass(frozen=True)
class PdfPage:
    page_number: int
    text: str
    image_count: int = 0

    @property
    def number(self) -> int:
        return self.page_number


@dataclass(frozen=True)
class ImageAsset:
    data: bytes
    media_type: str
    width: int
    height: int
    sha256: str

    @property
    def content_type(self) -> str:
        return self.media_type


CatalogAssetMatchStatus = Literal[
    "exact_xlsx",
    "merged_xlsx",
    "family_xlsx",
    "exact_pdf",
    "exact_web",
    "model_web",
]


@dataclass(frozen=True)
class CatalogAssetBinding:
    internal_id: str
    asset_sha256: str
    object_name: str
    image_kind: Literal["official"]
    match_status: CatalogAssetMatchStatus
    source_references: tuple[dict, ...]


@dataclass(frozen=True)
class CatalogSnapshotBuild:
    snapshot: dict
    assets_by_sha256: Mapping[str, ImageAsset]
    bindings: tuple[CatalogAssetBinding, ...]


@runtime_checkable
class CatalogSnapshotBuildLike(Protocol):
    snapshot: dict
    assets_by_sha256: Mapping[str, ImageAsset]
    bindings: tuple[CatalogAssetBinding, ...]


def _fail(code: str):
    raise SourceSafetyError(code) from None


def _extension(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    extension = value.lower()
    if not extension.startswith("."):
        extension = "." + extension
    return extension if extension in {".xlsx", ".pdf", ".png", ".jpg", ".jpeg"} else None


def _same_file(left, right) -> bool:
    return all(
        getattr(left, name, None) == getattr(right, name, None)
        for name in ("st_dev", "st_ino", "st_size", "st_mtime_ns")
    ) and stat.S_ISREG(left.st_mode) and stat.S_ISREG(right.st_mode)


def _read_source(path: Path, expected_extension: str, max_bytes: int) -> tuple[ValidatedSource, bytes]:
    extension = _extension(expected_extension)
    if extension is None or type(max_bytes) is not int or not 0 < max_bytes <= MAX_FILE_BYTES:
        _fail("SOURCE_ARGUMENT")
    try:
        source = Path(path)
        suffix = source.suffix.lower()
    except Exception:
        _fail("SOURCE_ARGUMENT")
    if suffix != extension:
        _fail("SOURCE_TYPE")

    descriptor = None
    failed = None
    data = b""
    try:
        before = os.lstat(source)
        if stat.S_ISLNK(before.st_mode) or not stat.S_ISREG(before.st_mode):
            failed = "SOURCE_FILE"
        else:
            flags = os.O_RDONLY | getattr(os, "O_BINARY", 0) | getattr(os, "O_NOFOLLOW", 0)
            descriptor = os.open(source, flags)
            opened = os.fstat(descriptor)
            if not _same_file(before, opened):
                failed = "SOURCE_CHANGED"
            elif not 0 < opened.st_size <= max_bytes:
                failed = "SOURCE_SIZE"
            else:
                remaining = max_bytes + 1
                chunks = []
                while remaining:
                    chunk = os.read(descriptor, min(64 * 1024, remaining))
                    if not chunk:
                        break
                    chunks.append(chunk)
                    remaining -= len(chunk)
                data = b"".join(chunks)
                after = os.fstat(descriptor)
                if not _same_file(opened, after) or len(data) != opened.st_size:
                    failed = "SOURCE_CHANGED"
                elif len(data) > max_bytes:
                    failed = "SOURCE_SIZE"
    except Exception:
        failed = failed or "SOURCE_FILE"
    finally:
        if descriptor is not None:
            try:
                os.close(descriptor)
            except Exception:
                failed = "SOURCE_FILE"
    if failed:
        _fail(failed)
    magic_ok = {
        ".xlsx": data.startswith(b"PK\x03\x04"),
        ".pdf": data.startswith(b"%PDF-"),
        ".png": data.startswith(b"\x89PNG\r\n\x1a\n"),
        ".jpg": data.startswith(b"\xff\xd8\xff"),
        ".jpeg": data.startswith(b"\xff\xd8\xff"),
    }[extension]
    if not magic_ok:
        _fail("SOURCE_TYPE")
    return ValidatedSource(extension, len(data), hashlib.sha256(data).hexdigest()), data


def validate_source_file(
    path: Path, expected_extension: str, max_bytes: int = MAX_FILE_BYTES
) -> ValidatedSource:
    source, _ = read_validated_source(path, expected_extension, max_bytes)
    return source


def read_validated_source(
    path: Path, expected_extension: str, max_bytes: int = MAX_FILE_BYTES
) -> tuple[ValidatedSource, bytes]:
    source, data = _read_source(path, expected_extension, max_bytes)
    if source.extension == ".xlsx":
        _validate_xlsx(data)
    elif source.extension == ".pdf":
        _pdf_pages(data)
    else:
        _normalize_image(data)
    return source, data


def _safe_member(name: str) -> bool:
    windows = PureWindowsPath(name)
    path = PurePosixPath(name)
    return bool(name) and not (
        "\\" in name
        or path.is_absolute()
        or windows.drive
        or windows.root
        or any(part in {"", ".", ".."} for part in path.parts)
    )


def _part_target(source_part: str, target: str) -> str | None:
    if not target or "\\" in target or "#" in target or "?" in target:
        return None
    if re.match(r"^[A-Za-z][A-Za-z0-9+.-]*:", target):
        return None
    if target.startswith("/"):
        resolved = posixpath.normpath(target.lstrip("/"))
    else:
        resolved = posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))
    return resolved if _safe_member(resolved) else None


def _relationship_source(name: str) -> str | None:
    if name == "_rels/.rels":
        return ""
    marker = "/_rels/"
    if marker not in name or not name.endswith(".rels"):
        return None
    directory, leaf = name.split(marker, 1)
    return f"{directory}/{leaf[:-5]}"


def _validate_content_types(parts: dict[str, bytes]) -> dict[str, str]:
    try:
        root = ElementTree.fromstring(parts["[Content_Types].xml"])
    except Exception:
        _fail("XLSX_INVALID")
    if root.tag != f"{{{_CONTENT_TYPE_NS}}}Types" or root.attrib:
        _fail("XLSX_INVALID")
    defaults = {}
    overrides = {}
    for entry in root:
        if entry.tag == f"{{{_CONTENT_TYPE_NS}}}Default" and set(entry.attrib) == {
            "Extension",
            "ContentType",
        }:
            extension = entry.get("Extension")
            content_type = entry.get("ContentType", "").lower()
            if (
                not extension
                or re.fullmatch(r"[A-Za-z0-9]{1,20}", extension) is None
                or extension.lower() in defaults
            ):
                _fail("XLSX_INVALID")
            defaults[extension.lower()] = content_type
        elif entry.tag == f"{{{_CONTENT_TYPE_NS}}}Override" and set(entry.attrib) == {
            "PartName",
            "ContentType",
        }:
            part_name = entry.get("PartName", "")
            content_type = entry.get("ContentType", "").lower()
            part = part_name[1:] if part_name.startswith("/") else ""
            if not part or not _safe_member(part) or part not in parts or part.casefold() in overrides:
                _fail("XLSX_INVALID")
            overrides[part.casefold()] = content_type
        else:
            _fail("XLSX_INVALID")
        if content_type not in _ALLOWED_CONTENT_TYPES:
            _fail("XLSX_UNSAFE")

    content_types = {}
    for part in parts:
        if part == "[Content_Types].xml" or part.endswith("/"):
            continue
        content_type = overrides.get(part.casefold())
        if content_type is None:
            extension = part.rsplit(".", 1)[-1].lower() if "." in part else ""
            content_type = defaults.get(extension)
        if content_type not in _ALLOWED_CONTENT_TYPES:
            _fail("XLSX_UNSAFE")
        content_types[part.casefold()] = content_type
    return content_types


def _relationships(parts: dict[str, bytes], name: str) -> dict[str, tuple[str, str]]:
    if name not in parts:
        return {}
    source = _relationship_source(name)
    if source is None:
        _fail("XLSX_INVALID")
    try:
        root = ElementTree.fromstring(parts[name])
    except Exception:
        _fail("XLSX_INVALID")
    if root.tag != f"{{{_REL_NS}}}Relationships":
        _fail("XLSX_INVALID")
    result = {}
    for relationship in root:
        if relationship.tag != f"{{{_REL_NS}}}Relationship" or set(relationship.attrib) - {
            "Id",
            "Type",
            "Target",
            "TargetMode",
        }:
            _fail("XLSX_INVALID")
        relationship_id = relationship.get("Id")
        relation_type = relationship.get("Type")
        target = relationship.get("Target")
        mode = relationship.get("TargetMode", "Internal")
        if (
            not relationship_id
            or relationship_id in result
            or not relation_type
            or len(relation_type) > 512
            or mode not in {"Internal", "External"}
        ):
            _fail("XLSX_INVALID")
        if mode == "External":
            if relation_type == _NULL_EXTERNAL_IMAGE_RELATIONSHIP and target == "NULL":
                continue
            try:
                link = urlsplit(target or "")
            except ValueError:
                _fail("XLSX_UNSAFE")
            if (
                relation_type != _PASSIVE_EXTERNAL_RELATIONSHIP
                or link.scheme.lower() not in {"http", "https"}
                or not link.hostname
                or link.username is not None
                or link.password is not None
                or len(target or "") > 2048
            ):
                _fail("XLSX_UNSAFE")
            continue
        if relation_type not in _ALLOWED_RELATIONSHIPS:
            _fail("XLSX_UNSAFE")
        resolved = _part_target(source, target or "")
        if resolved is None or resolved not in parts:
            _fail("XLSX_INVALID")
        result[relationship_id] = (relation_type, resolved)
    return result


def _worksheet_parts(parts: dict[str, bytes]) -> dict[str, str]:
    try:
        workbook = ElementTree.fromstring(parts["xl/workbook.xml"])
    except Exception:
        _fail("XLSX_INVALID")
    for defined_name in workbook.findall(f".//{{{_SHEET_NS}}}definedName"):
        _validate_formula_text(defined_name.text or "")
    relations = _relationships(parts, "xl/_rels/workbook.xml.rels")
    sheets = {}
    for sheet in workbook.findall(f".//{{{_SHEET_NS}}}sheet"):
        name = sheet.get("name")
        relation_id = sheet.get(f"{{{_DOC_REL_NS}}}id")
        relation = relations.get(relation_id)
        if (
            not name
            or len(name) > 31
            or name in sheets
            or relation is None
            or relation[1] not in parts
        ):
            _fail("XLSX_INVALID")
        sheets[name] = relation[1]
    if not sheets or len(sheets) > MAX_WORKBOOK_SHEETS:
        _fail("XLSX_LIMIT")
    return sheets


def _validate_formula_text(text: str):
    if len(text) > MAX_TEXT_LENGTH or _DDE_FORMULA.search(text) or _EXTERNAL_FORMULA.search(text):
        _fail("XLSX_UNSAFE")


def _formulas(parts: dict[str, bytes], sheets: dict[str, str]) -> tuple[CellRef, ...]:
    cells = 0
    formulas = []
    merged_ranges = 0
    merged_cells = 0
    for sheet_name, part in sheets.items():
        try:
            root = ElementTree.fromstring(parts[part])
        except Exception:
            _fail("XLSX_INVALID")
        for cell in root.iter(f"{{{_SHEET_NS}}}c"):
            cells += 1
            coordinate = cell.get("r")
            formula = cell.find(f"{{{_SHEET_NS}}}f")
            if formula is None:
                continue
            if not coordinate:
                _fail("XLSX_INVALID")
            try:
                coordinate_to_tuple(coordinate)
            except Exception:
                _fail("XLSX_INVALID")
            text = formula.text or ""
            _validate_formula_text(text)
            formulas.append(CellRef(sheet_name, coordinate.upper()))
        for merged in root.iter(f"{{{_SHEET_NS}}}mergeCell"):
            merged_ranges += 1
            if merged_ranges > MAX_WORKBOOK_MERGED_RANGES:
                _fail("XLSX_LIMIT")
            reference = merged.get("ref")
            if not reference or _CELL.fullmatch(reference.upper()) is None:
                _fail("XLSX_INVALID")
            start, separator, end = reference.upper().partition(":")
            if not separator:
                end = start
            try:
                start_row, start_column = coordinate_to_tuple(start)
                end_row, end_column = coordinate_to_tuple(end)
            except Exception:
                _fail("XLSX_INVALID")
            if (
                start_row > _MAX_XLSX_ROW
                or end_row > _MAX_XLSX_ROW
                or start_column > _MAX_XLSX_COLUMN
                or end_column > _MAX_XLSX_COLUMN
                or end_row < start_row
                or end_column < start_column
            ):
                _fail("XLSX_INVALID")
            area = (end_row - start_row + 1) * (end_column - start_column + 1)
            merged_cells += area
            if area > MAX_MERGED_RANGE_CELLS or merged_cells > MAX_WORKBOOK_MERGED_CELLS:
                _fail("XLSX_LIMIT")
    if cells > MAX_WORKBOOK_CELLS or len(formulas) > MAX_WORKBOOK_FORMULAS:
        _fail("XLSX_LIMIT")
    return tuple(sorted(formulas))


def _validate_xlsx(
    data: bytes,
) -> tuple[dict[str, bytes], dict[str, str], tuple[CellRef, ...], dict[str, str]]:
    invalid = None
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            if not infos or len(infos) > MAX_ZIP_ENTRIES:
                _fail("XLSX_LIMIT")
            names = [info.filename for info in infos]
            if len({name.casefold() for name in names}) != len(names):
                _fail("XLSX_UNSAFE")
            expanded = 0
            for info in infos:
                local_flags = int.from_bytes(data[info.header_offset + 6 : info.header_offset + 8], "little")
                if (
                    not _safe_member(info.filename)
                    or info.flag_bits & 1
                    or local_flags & 1
                    or info.compress_type not in _SUPPORTED_COMPRESSION
                ):
                    _fail("XLSX_UNSAFE")
                expanded += info.file_size
                if info.file_size and (
                    info.compress_size == 0 or info.file_size / info.compress_size > MAX_ZIP_RATIO
                ):
                    _fail("XLSX_LIMIT")
            if expanded > MAX_ZIP_EXPANDED_BYTES:
                _fail("XLSX_LIMIT")
            lowered = {name.casefold() for name in names}
            if any(
                "vbaproject" in name
                or name.startswith(("xl/embeddings/", "xl/activex/"))
                or name.startswith("xl/externallinks/")
                for name in lowered
            ):
                _fail("XLSX_UNSAFE")
            parts = {info.filename: archive.read(info) for info in infos}
    except SourceSafetyError:
        raise
    except Exception:
        invalid = "XLSX_INVALID"
    if invalid:
        _fail(invalid)
    if "[Content_Types].xml" not in parts or "xl/workbook.xml" not in parts:
        _fail("XLSX_INVALID")
    content_types = _validate_content_types(parts)
    for name in sorted(parts):
        if name.endswith(".rels"):
            _relationships(parts, name)
    sheets = _worksheet_parts(parts)
    formulas = _formulas(parts, sheets)
    return parts, sheets, formulas, content_types


def _without_unsupported_images(parts: dict[str, bytes], content_types: dict[str, str]) -> bytes | None:
    cleaned = dict(parts)
    changed = False
    for name, raw in tuple(cleaned.items()):
        if not name.endswith(".rels"):
            continue
        try:
            root = ElementTree.fromstring(raw)
        except Exception:
            _fail("XLSX_INVALID")
        source = _relationship_source(name)
        unsupported_ids = set()
        for relationship in root:
            if relationship.get("Type") != _NULL_EXTERNAL_IMAGE_RELATIONSHIP:
                continue
            mode = relationship.get("TargetMode", "Internal")
            target = relationship.get("Target", "")
            unsupported = target == "NULL" and mode == "External"
            if mode == "Internal" and source is not None:
                resolved = _part_target(source, target)
                unsupported = (
                    resolved is not None
                    and content_types.get(resolved.casefold()) == "image/x-emf"
                )
            if unsupported:
                unsupported_ids.add(relationship.get("Id"))
        unsupported_ids.discard(None)
        if not unsupported_ids:
            continue
        changed = True
        root[:] = [relationship for relationship in root if relationship.get("Id") not in unsupported_ids]
        cleaned[name] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

        if source not in cleaned or not source.startswith("xl/drawings/"):
            continue
        try:
            drawing = ElementTree.fromstring(cleaned[source])
        except Exception:
            _fail("XLSX_INVALID")
        drawing[:] = [
            anchor
            for anchor in drawing
            if not any(
                blip.get(f"{{{_RELATIONSHIP_ATTR_NS}}}embed") in unsupported_ids
                or blip.get(f"{{{_RELATIONSHIP_ATTR_NS}}}link") in unsupported_ids
                for blip in anchor.findall(f".//{{{_DRAWING_MAIN_NS}}}blip")
            )
        ]
        cleaned[source] = ElementTree.tostring(drawing, encoding="utf-8", xml_declaration=True)
    if not changed:
        return None
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name in sorted(cleaned):
            archive.writestr(name, cleaned[name])
    return output.getvalue()


def open_xlsx_data_only_from_bytes(data: bytes) -> Workbook:
    if type(data) is not bytes or not 0 < len(data) <= MAX_FILE_BYTES or not data.startswith(b"PK\x03\x04"):
        _fail("SOURCE_TYPE")
    parts, _, formulas, content_types = _validate_xlsx(data)
    safe_data = _without_unsupported_images(parts, content_types) or data
    failed = False
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(safe_data), data_only=True, read_only=False, keep_links=False)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink is not None:
                        cell.hyperlink = None
        for reference in formulas:
            workbook[reference.sheet][reference.cell].value = None
        workbook.formula_cells = formulas
    except Exception:
        failed = True
    if failed:
        _fail("XLSX_INVALID")
    return workbook


def open_xlsx_data_only(path: Path) -> Workbook:
    _, data = _read_source(path, ".xlsx", MAX_FILE_BYTES)
    return open_xlsx_data_only_from_bytes(data)


def _pdf_integer(document, xref: int, key: str) -> int | None:
    kind, value = document.xref_get_key(xref, key)
    if kind == "int" and re.fullmatch(r"[0-9]+", value):
        return int(value)
    if kind == "xref" and re.fullmatch(r"[1-9][0-9]* 0 R", value):
        referenced = int(value.split(" ", 1)[0])
        raw = document.xref_object(referenced, compressed=False).strip()
        return int(raw) if re.fullmatch(r"[0-9]+", raw) else None
    return None


def _pdf_flate_size(raw: bytes) -> int:
    decoder = zlib.decompressobj()
    total = 0
    invalid = False
    try:
        for offset in range(0, len(raw), 64 * 1024):
            pending = raw[offset : offset + 64 * 1024]
            while pending:
                output = decoder.decompress(
                    pending,
                    min(64 * 1024, MAX_PDF_STREAM_DECODED_BYTES - total + 1),
                )
                total += len(output)
                if total > MAX_PDF_STREAM_DECODED_BYTES:
                    _fail("PDF_LIMIT")
                remaining = decoder.unconsumed_tail
                if remaining == pending and not output:
                    invalid = True
                    break
                pending = remaining
            if invalid:
                break
        if not invalid and not decoder.eof:
            output = decoder.flush(MAX_PDF_STREAM_DECODED_BYTES - total + 1)
            total += len(output)
    except SourceSafetyError:
        raise
    except Exception:
        invalid = True
    if invalid or not decoder.eof or decoder.unused_data:
        _fail("PDF_INVALID")
    if total > MAX_PDF_STREAM_DECODED_BYTES or (
        total and (not raw or total / len(raw) > MAX_PDF_STREAM_RATIO)
    ):
        _fail("PDF_LIMIT")
    return total


def _pdf_filters(document, xref: int) -> tuple[str, ...] | None:
    kind, value = document.xref_get_key(xref, "Filter")
    if kind == "null":
        return ()
    if kind not in {"name", "array"}:
        return None
    names = tuple(re.findall(r"/([A-Za-z0-9]+)", value))
    return names or None


def _pdf_image_components(document, xref: int) -> int | None:
    kind, value = document.xref_get_key(xref, "ColorSpace")
    direct = {"/DeviceGray": 1, "/DeviceRGB": 3, "/DeviceCMYK": 4}
    if kind == "name":
        return direct.get(value)
    raw_color_space = value
    if kind == "xref" and re.fullmatch(r"[1-9][0-9]* 0 R", value):
        color_xref = int(value.split(" ", 1)[0])
        raw_color_space = document.xref_object(color_xref, compressed=False).strip()
    indexed = re.fullmatch(
        r"\[\s*/Indexed\s+(/DeviceGray|/DeviceRGB|/DeviceCMYK)\s+([0-9]+)\s+([1-9][0-9]*)\s+0\s+R\s*\]",
        raw_color_space,
    )
    if indexed is not None:
        base_components = direct[indexed.group(1)]
        highest_index = int(indexed.group(2))
        palette_xref = int(indexed.group(3))
        if not 0 < palette_xref < document.xref_length():
            return None
        expected_palette_bytes = (highest_index + 1) * base_components
        declared_palette_bytes = _pdf_integer(document, palette_xref, "Length")
        palette_filters = _pdf_filters(document, palette_xref)
        palette_raw = (
            document.xref_stream_raw(palette_xref)
            if document.xref_is_stream(palette_xref)
            else None
        )
        decoded_palette_bytes = None
        if isinstance(palette_raw, bytes) and declared_palette_bytes == len(palette_raw):
            if palette_filters == ():
                decoded_palette_bytes = len(palette_raw)
            elif palette_filters == ("FlateDecode",):
                decoded_palette_bytes = _pdf_flate_size(palette_raw)
        if (
            highest_index <= 255
            and document.xref_is_stream(palette_xref)
            and declared_palette_bytes is not None
            and declared_palette_bytes <= expected_palette_bytes + 64
            and decoded_palette_bytes == expected_palette_bytes
        ):
            return 1
        return None
    profile = None
    if kind == "array":
        match = re.fullmatch(r"\[\s*/ICCBased\s+([1-9][0-9]*)\s+0\s+R\s*\]", value)
        profile = int(match.group(1)) if match else None
    elif kind == "xref" and re.fullmatch(r"[1-9][0-9]* 0 R", value):
        match = re.fullmatch(r"\[\s*/ICCBased\s+([1-9][0-9]*)\s+0\s+R\s*\]", raw_color_space)
        profile = int(match.group(1)) if match else None
    components = _pdf_integer(document, profile, "N") if profile is not None else None
    return components if components in {1, 2, 3, 4} else None


def _pdf_image_decoded_size(document, xref: int) -> int:
    width = _pdf_integer(document, xref, "Width")
    height = _pdf_integer(document, xref, "Height")
    bits = _pdf_integer(document, xref, "BitsPerComponent")
    components = _pdf_image_components(document, xref)
    if width is None or height is None or bits is None or components is None:
        _fail("PDF_INVALID")
    if (
        width <= 0
        or height <= 0
        or width > MAX_IMAGE_DIMENSION
        or height > MAX_IMAGE_DIMENSION
        or width * height > MAX_IMAGE_PIXELS
        or not 1 <= bits <= 16
    ):
        _fail("PDF_LIMIT")
    decoded = ((width * components * bits + 7) // 8) * height
    if decoded > MAX_PDF_STREAM_DECODED_BYTES:
        _fail("PDF_LIMIT")
    return decoded


def _passive_http_uri(value: object) -> bool:
    if not isinstance(value, str) or len(value) > 2048:
        return False
    try:
        parsed = urlsplit(value)
    except ValueError:
        return False
    return (
        parsed.scheme.lower() in {"http", "https"}
        and bool(parsed.hostname)
        and parsed.username is None
        and parsed.password is None
    )


def _pdf_passive_uri_action(document, xref: int, prefix: str = "") -> bool:
    action_kind, action_value = document.xref_get_key(xref, f"{prefix}S")
    uri_kind, uri_value = document.xref_get_key(xref, f"{prefix}URI")
    return (
        action_kind == "name"
        and action_value == "/URI"
        and uri_kind == "string"
        and _passive_http_uri(uri_value)
    )


def _pdf_passive_internal_open_action(document, xref: int) -> bool:
    action_kind, action_value = document.xref_get_key(xref, "OpenAction")
    if action_kind != "xref":
        return False
    action_match = re.fullmatch(r"(\d+) 0 R", action_value)
    if action_match is None:
        return False
    action_xref = int(action_match.group(1))
    if not 0 < action_xref < document.xref_length():
        return False
    action_keys = set(document.xref_get_keys(action_xref))
    action_type, action_name = document.xref_get_key(action_xref, "S")
    destination_type, destination = document.xref_get_key(action_xref, "D")
    destination_match = re.fullmatch(r"\[\s*(\d+)\s+0\s+R\s*/Fit\s*\]", destination)
    if (
        action_keys & {"AA", "Next"}
        or action_type != "name"
        or action_name != "/GoTo"
        or destination_type != "array"
        or destination_match is None
    ):
        return False
    page_xref = int(destination_match.group(1))
    return (
        0 < page_xref < document.xref_length()
        and document.xref_get_key(page_xref, "Type") == ("name", "/Page")
    )


def _pdf_preflight(document) -> dict[int, int]:
    xref_count = document.xref_length()
    if not 1 < xref_count <= MAX_PDF_XREFS:
        _fail("PDF_LIMIT")
    stream_count = 0
    total_raw = 0
    total_expanded = 0
    raw_sizes = {}
    for xref in range(1, xref_count):
        keys = set(document.xref_get_keys(xref))
        passive_open_action = "OpenAction" in keys and _pdf_passive_internal_open_action(document, xref)
        if "AA" in keys or ("OpenAction" in keys and not passive_open_action):
            _fail("PDF_UNSAFE")
        type_value = document.xref_get_key(xref, "Type")[1]
        action_value = document.xref_get_key(xref, "S")[1]
        passive_uri = _pdf_passive_uri_action(document, xref)
        passive_link = _pdf_passive_uri_action(document, xref, "A/")
        chained_action = (
            document.xref_get_key(xref, "Next")[0] != "null"
            and type_value == "/Action"
        ) or document.xref_get_key(xref, "A/Next")[0] != "null"
        if chained_action:
            _fail("PDF_UNSAFE")
        if (type_value == "/Action" and not passive_uri) or (
            action_value in {
            "/GoToR",
            "/ImportData",
            "/JavaScript",
            "/Launch",
            "/Named",
            "/SubmitForm",
            "/URI",
            }
            and not passive_uri
        ):
            _fail("PDF_UNSAFE")
        raw_object = document.xref_object(xref, compressed=False).encode("latin-1", "ignore")
        active_tokens = {match.group(0) for match in _ACTIVE_PDF_TOKEN.finditer(raw_object)}
        allowed_active_tokens = (
            active_tokens == {b"/URI"} and (passive_uri or passive_link)
        ) or (active_tokens == {b"/OpenAction"} and passive_open_action)
        if active_tokens and not allowed_active_tokens:
            _fail("PDF_UNSAFE")
        if not document.xref_is_stream(xref):
            continue

        stream_count += 1
        declared = _pdf_integer(document, xref, "Length")
        if declared is None or declared > MAX_PDF_STREAM_RAW_BYTES:
            _fail("PDF_LIMIT" if declared is not None else "PDF_INVALID")
        total_raw += declared
        if stream_count > MAX_PDF_STREAMS or total_raw > MAX_FILE_BYTES:
            _fail("PDF_LIMIT")
        raw = document.xref_stream_raw(xref)
        if not isinstance(raw, bytes) or len(raw) != declared:
            _fail("PDF_INVALID")
        raw_sizes[xref] = len(raw)
        filters = _pdf_filters(document, xref)
        subtype = document.xref_get_key(xref, "Subtype")[1]
        if subtype == "/Image" and filters and set(filters) <= _PDF_IMAGE_FILTERS:
            decoded = _pdf_image_decoded_size(document, xref)
        elif filters == ():
            decoded = len(raw)
        elif filters == ("FlateDecode",):
            decoded = _pdf_flate_size(raw)
        else:
            _fail("PDF_UNSAFE")
        total_expanded += decoded
        if total_expanded > MAX_PDF_STREAM_EXPANDED_BYTES:
            _fail("PDF_LIMIT")
    return raw_sizes


def _pdf_page_preflight(document, page, raw_sizes: dict[int, int]) -> tuple[int, int]:
    annotation_xrefs = {item[0] for item in page.annot_xrefs()}
    widgets = page.widgets()
    if widgets is not None:
        annotation_xrefs.update(widget.xref for widget in widgets)
    for xref in annotation_xrefs:
        keys = set(document.xref_get_keys(xref))
        if "AA" in keys or ("A" in keys and not _pdf_passive_uri_action(document, xref, "A/")):
            _fail("PDF_UNSAFE")
    for link in page.get_links():
        if link.get("kind") != fitz.LINK_URI or not _passive_http_uri(link.get("uri")):
            _fail("PDF_UNSAFE")

    images = page.get_images(full=True)
    image_bytes = 0
    for image in images:
        xref, width, height = image[0], image[2], image[3]
        if (
            width <= 0
            or height <= 0
            or width > MAX_IMAGE_DIMENSION
            or height > MAX_IMAGE_DIMENSION
            or width * height > MAX_IMAGE_PIXELS
        ):
            _fail("PDF_LIMIT")
        image_bytes += raw_sizes.get(xref, 0)
    return len(images), image_bytes


def _pdf_text_worker(control, output, data: bytes, maximum: int):
    document = None
    try:
        if os.name != "nt":
            import resource

            current = resource.getrlimit(resource.RLIMIT_AS)[1]
            hard = MAX_PDF_TEXT_WORKER_BYTES if current == resource.RLIM_INFINITY else min(
                current, MAX_PDF_TEXT_WORKER_BYTES
            )
            resource.setrlimit(resource.RLIMIT_AS, (hard, hard))
        if control.recv_bytes(1) != b"G":
            return
        document = fitz.open(stream=data, filetype="pdf")
        total = 0
        for index in range(document.page_count):
            text = document.load_page(index).get_text("text")
            if not isinstance(text, str):
                output.send_bytes(b"X")
                return
            output.send_bytes(b"S" + index.to_bytes(4, "big"))
            for offset in range(0, len(text), MAX_PDF_TEXT_CHUNK_CHARS):
                chunk = text[offset : offset + MAX_PDF_TEXT_CHUNK_CHARS].encode("utf-8")
                if total + len(chunk) > maximum:
                    output.send_bytes(b"L")
                    return
                output.send_bytes(b"C" + chunk)
                total += len(chunk)
            output.send_bytes(b"E")
        output.send_bytes(b"D")
    except Exception:
        try:
            output.send_bytes(b"X")
        except Exception:
            pass
    finally:
        if document is not None:
            try:
                document.close()
            except Exception:
                pass
        control.close()
        output.close()


def _windows_text_job(process) -> int | None:
    if os.name != "nt":
        return 0
    import ctypes
    from ctypes import wintypes

    class IoCounters(ctypes.Structure):
        _fields_ = [(name, ctypes.c_ulonglong) for name in (
            "ReadOperationCount",
            "WriteOperationCount",
            "OtherOperationCount",
            "ReadTransferCount",
            "WriteTransferCount",
            "OtherTransferCount",
        )]

    class BasicLimits(ctypes.Structure):
        _fields_ = [
            ("PerProcessUserTimeLimit", ctypes.c_longlong),
            ("PerJobUserTimeLimit", ctypes.c_longlong),
            ("LimitFlags", wintypes.DWORD),
            ("MinimumWorkingSetSize", ctypes.c_size_t),
            ("MaximumWorkingSetSize", ctypes.c_size_t),
            ("ActiveProcessLimit", wintypes.DWORD),
            ("Affinity", ctypes.c_size_t),
            ("PriorityClass", wintypes.DWORD),
            ("SchedulingClass", wintypes.DWORD),
        ]

    class ExtendedLimits(ctypes.Structure):
        _fields_ = [
            ("BasicLimitInformation", BasicLimits),
            ("IoInfo", IoCounters),
            ("ProcessMemoryLimit", ctypes.c_size_t),
            ("JobMemoryLimit", ctypes.c_size_t),
            ("PeakProcessMemoryUsed", ctypes.c_size_t),
            ("PeakJobMemoryUsed", ctypes.c_size_t),
        ]

    kernel = ctypes.WinDLL("kernel32", use_last_error=True)
    kernel.CreateJobObjectW.restype = wintypes.HANDLE
    kernel.OpenProcess.restype = wintypes.HANDLE
    job = kernel.CreateJobObjectW(None, None)
    process_handle = kernel.OpenProcess(0x0001 | 0x0100 | 0x1000, False, process.pid)
    if not job or not process_handle:
        if job:
            kernel.CloseHandle(job)
        if process_handle:
            kernel.CloseHandle(process_handle)
        return None
    limits = ExtendedLimits()
    limits.BasicLimitInformation.LimitFlags = 0x0100 | 0x2000
    limits.ProcessMemoryLimit = MAX_PDF_TEXT_WORKER_BYTES
    configured = kernel.SetInformationJobObject(job, 9, ctypes.byref(limits), ctypes.sizeof(limits))
    assigned = configured and kernel.AssignProcessToJobObject(job, process_handle)
    kernel.CloseHandle(process_handle)
    if not assigned:
        kernel.CloseHandle(job)
        return None
    return int(job)


def _close_windows_handle(handle: int | None):
    if os.name == "nt" and handle:
        import ctypes

        ctypes.WinDLL("kernel32", use_last_error=True).CloseHandle(handle)


def _pdf_worker_alive(process) -> bool:
    try:
        return process.pid is not None and process.is_alive()
    except Exception:
        return True


def _pdf_worker_join(process):
    try:
        process.join(MAX_PDF_WORKER_SHUTDOWN_SECONDS)
    except Exception:
        pass


def _shutdown_pdf_worker(process, job: int | None, code: str | None) -> tuple[str | None, None]:
    alive = _pdf_worker_alive(process)
    if code is None:
        _pdf_worker_join(process)
        alive = _pdf_worker_alive(process)
    if alive:
        try:
            process.terminate()
        except Exception:
            pass
        code = code or "PDF_LIMIT"
        _pdf_worker_join(process)
        alive = _pdf_worker_alive(process)
    if alive:
        killer = getattr(process, "kill", None)
        if callable(killer):
            try:
                killer()
            except Exception:
                pass
    try:
        _close_windows_handle(job)
    except Exception:
        code = code or "PDF_LIMIT"
    job = None
    _pdf_worker_join(process)
    alive = _pdf_worker_alive(process)
    if alive:
        code = code or "PDF_LIMIT"
        popen = getattr(process, "_popen", None)
        closer = getattr(popen, "close", None)
        if callable(closer):
            try:
                closer()
            except Exception:
                pass
    else:
        try:
            process.close()
        except Exception:
            code = code or "PDF_LIMIT"
    return code, job


def _close_pdf_connections(*connections) -> bool:
    failed = False
    for connection in connections:
        try:
            connection.close()
        except Exception:
            failed = True
    return failed


def _pdf_text_isolated(data: bytes, page_count: int) -> tuple[str, ...]:
    context = multiprocessing.get_context("spawn")
    control_read, control_write = context.Pipe(duplex=False)
    output_read, output_write = context.Pipe(duplex=False)
    process = context.Process(
        target=_pdf_text_worker,
        args=(control_read, output_write, data, MAX_PDF_TEXT_BYTES),
    )
    process.daemon = True
    code = None
    job = None
    texts = []
    current = None
    total = 0
    deadline = time.monotonic() + MAX_PDF_TEXT_SECONDS
    try:
        process.start()
        control_read.close()
        output_write.close()
        job = _windows_text_job(process)
        if job is None:
            code = "PDF_LIMIT"
        else:
            control_write.send_bytes(b"G")
        while code is None:
            remaining = deadline - time.monotonic()
            if remaining <= 0 or not output_read.poll(remaining):
                code = "PDF_LIMIT"
                break
            try:
                message = output_read.recv_bytes(4 * MAX_PDF_TEXT_CHUNK_CHARS + 1)
            except EOFError:
                code = "PDF_LIMIT"
                break
            if not message:
                code = "PDF_INVALID"
                break
            tag, payload = message[:1], message[1:]
            if tag == b"S" and len(payload) == 4 and current is None:
                if int.from_bytes(payload, "big") != len(texts):
                    code = "PDF_INVALID"
                else:
                    current = []
            elif tag == b"C" and current is not None:
                total += len(payload)
                if total > MAX_PDF_TEXT_BYTES:
                    code = "PDF_LIMIT"
                else:
                    try:
                        current.append(payload.decode("utf-8"))
                    except UnicodeDecodeError:
                        code = "PDF_INVALID"
            elif tag == b"E" and not payload and current is not None:
                texts.append("".join(current))
                current = None
            elif tag == b"D" and not payload and current is None:
                break
            elif tag == b"L":
                code = "PDF_LIMIT"
            elif tag == b"X":
                code = "PDF_INVALID"
            else:
                code = "PDF_INVALID"
        if code is None and len(texts) != page_count:
            code = "PDF_INVALID"
    except Exception:
        code = code or "PDF_INVALID"
    finally:
        code, job = _shutdown_pdf_worker(process, job, code)
        if _close_pdf_connections(control_read, control_write, output_read, output_write):
            code = code or "PDF_INVALID"
    if code:
        _fail(code)
    return tuple(texts)


def _pdf_pages(data: bytes) -> tuple[PdfPage, ...]:
    failed = None
    pages = []
    document = None
    try:
        document = fitz.open(stream=data, filetype="pdf")
        if not document.is_pdf or document.needs_pass or document.is_encrypted:
            failed = "PDF_UNSAFE"
        elif not 0 < document.page_count <= MAX_PDF_PAGES:
            failed = "PDF_LIMIT"
        elif document.embfile_names():
            failed = "PDF_UNSAFE"
        else:
            raw_sizes = _pdf_preflight(document)
        total_images = 0
        total_image_bytes = 0
        page_image_counts = []
        if failed is None:
            for index in range(document.page_count):
                page = document.load_page(index)
                image_count, image_bytes = _pdf_page_preflight(document, page, raw_sizes)
                page_image_counts.append(image_count)
                total_images += image_count
                total_image_bytes += image_bytes
                if total_images > MAX_PDF_IMAGES or total_image_bytes > MAX_PDF_IMAGE_BYTES:
                    _fail("PDF_LIMIT")
            texts = _pdf_text_isolated(data, document.page_count)
            pages = [
                PdfPage(index + 1, text, page_image_counts[index])
                for index, text in enumerate(texts)
            ]
    except SourceSafetyError:
        raise
    except Exception:
        failed = failed or "PDF_INVALID"
    finally:
        if document is not None:
            try:
                document.close()
            except Exception:
                failed = failed or "PDF_INVALID"
    if failed:
        _fail(failed)
    return tuple(pages)


def iter_pdf_pages(path: Path) -> Iterator[PdfPage]:
    _, data = _read_source(path, ".pdf", MAX_FILE_BYTES)
    return iter(_pdf_pages(data))


def _probe_image_dimensions(data: bytes) -> tuple[int, int]:
    if not 0 < len(data) <= MAX_IMAGE_BYTES:
        _fail("IMAGE_LIMIT")
    failed = None
    width = height = 0
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(io.BytesIO(data)) as probe:
                if (
                    probe.format not in _SUPPORTED_IMAGE_FORMATS
                    or getattr(probe, "is_animated", False)
                    or probe.width <= 0
                    or probe.height <= 0
                    or probe.width > MAX_IMAGE_DIMENSION
                    or probe.height > MAX_IMAGE_DIMENSION
                    or probe.width * probe.height > MAX_IMAGE_PIXELS
                ):
                    failed = "IMAGE_UNSAFE"
                else:
                    width, height = probe.size
    except Exception:
        failed = failed or "IMAGE_UNSAFE"
    if failed:
        _fail(failed)
    return width, height


def _normalize_image(data: bytes) -> ImageAsset:
    width, height = _probe_image_dimensions(data)
    failed = None
    normalized = b""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(io.BytesIO(data)) as probe:
                probe.verify()
            if failed is None:
                with Image.open(io.BytesIO(data)) as source:
                    source.load()
                    image = ImageOps.exif_transpose(source)
                    if image.mode not in {"1", "L", "LA", "P", "RGB", "RGBA"}:
                        image = image.convert("RGBA" if "A" in image.getbands() else "RGB")
                    width, height = image.size
                    output = io.BytesIO()
                    image.save(output, format="PNG", optimize=False, compress_level=6)
                    normalized = output.getvalue()
    except Exception:
        failed = failed or "IMAGE_UNSAFE"
    if failed:
        _fail(failed)
    if len(normalized) > MAX_IMAGE_BYTES:
        _fail("IMAGE_LIMIT")
    return ImageAsset(normalized, "image/png", width, height, hashlib.sha256(normalized).hexdigest())


def _sheet_relationship_name(sheet_part: str) -> str:
    directory, leaf = posixpath.split(sheet_part)
    return f"{directory}/_rels/{leaf}.rels"


def _drawing_images(
    parts: dict[str, bytes],
    sheets: dict[str, str],
    content_types: dict[str, str],
    *,
    keep_ambiguous: bool = False,
) -> list[tuple[CellRef, str]]:
    found: list[tuple[CellRef, str]] = []
    anchors_per_cell: Counter[CellRef] = Counter()
    for sheet_name, sheet_part in sheets.items():
        sheet_relations = _relationships(parts, _sheet_relationship_name(sheet_part))
        drawing_parts = sorted(
            {
                target
                for relation_type, target in sheet_relations.values()
                if relation_type.endswith("/drawing")
            }
        )
        for drawing_part in drawing_parts:
            if drawing_part not in parts:
                _fail("IMAGE_UNSAFE")
            drawing_rels_name = _sheet_relationship_name(drawing_part)
            drawing_relations = _relationships(parts, drawing_rels_name)
            try:
                root = ElementTree.fromstring(parts[drawing_part])
            except Exception:
                _fail("IMAGE_UNSAFE")
            for anchor in root:
                start = anchor.find(f"{{{_DRAWING_NS}}}from")
                blip = anchor.find(f".//{{{_DRAWING_MAIN_NS}}}blip")
                relation_id = blip.get(f"{{{_RELATIONSHIP_ATTR_NS}}}embed") if blip is not None else None
                relation = drawing_relations.get(relation_id)
                if start is None or relation is None or not relation[0].endswith("/image"):
                    continue
                row_node = start.find(f"{{{_DRAWING_NS}}}row")
                column_node = start.find(f"{{{_DRAWING_NS}}}col")
                try:
                    row = int(row_node.text) + 1
                    column = int(column_node.text) + 1
                except Exception:
                    _fail("IMAGE_UNSAFE")
                if not (1 <= row <= _MAX_XLSX_ROW and 1 <= column <= _MAX_XLSX_COLUMN):
                    _fail("IMAGE_UNSAFE")
                target = relation[1]
                if target not in parts or not target.lower().startswith("xl/media/"):
                    _fail("IMAGE_UNSAFE")
                if content_types.get(target.casefold()) == "image/x-emf":
                    continue
                reference = CellRef(sheet_name, f"{get_column_letter(column)}{row}")
                anchors_per_cell[reference] += 1
                if anchors_per_cell[reference] > MAX_WORKBOOK_IMAGES_PER_CELL:
                    _fail("XLSX_LIMIT")
                found.append((reference, target))
                if len(found) > MAX_WORKBOOK_IMAGES:
                    _fail("XLSX_LIMIT")
    found.sort(key=lambda item: item[0])
    if keep_ambiguous:
        return found
    ambiguous = {reference for reference, count in Counter(row[0] for row in found).items() if count > 1}
    return [row for row in found if row[0] not in ambiguous]


def _normalized_xlsx_images(
    parts: dict[str, bytes], found: list[tuple[CellRef, str]]
) -> list[tuple[CellRef, ImageAsset]]:
    raw_by_part = {target: parts[target] for _, target in found}
    unique_raw = tuple(dict.fromkeys(raw_by_part.values()))
    if len(unique_raw) > MAX_WORKBOOK_UNIQUE_IMAGES:
        _fail("XLSX_LIMIT")

    asset_by_raw: dict[bytes, ImageAsset] = {}
    asset_by_normalized: dict[bytes, ImageAsset] = {}
    total_pixels = 0
    total_bytes = 0
    for raw in unique_raw:
        width, height = _probe_image_dimensions(raw)
        total_pixels += width * height
        if total_pixels > MAX_WORKBOOK_IMAGE_PIXELS:
            _fail("XLSX_LIMIT")
        asset = _normalize_image(raw)
        total_bytes += len(asset.data)
        if total_bytes > MAX_WORKBOOK_NORMALIZED_IMAGE_BYTES:
            _fail("XLSX_LIMIT")
        canonical = asset_by_normalized.get(asset.data)
        if canonical is None:
            canonical = asset
            asset_by_normalized[asset.data] = canonical
        asset_by_raw[raw] = canonical

    asset_by_part = {
        target: asset_by_raw[raw]
        for target, raw in raw_by_part.items()
    }
    return [(reference, asset_by_part[target]) for reference, target in found]


def extract_xlsx_images_from_bytes(data: bytes) -> dict[CellRef, ImageAsset]:
    if type(data) is not bytes or not 0 < len(data) <= MAX_FILE_BYTES or not data.startswith(b"PK\x03\x04"):
        _fail("SOURCE_TYPE")
    parts, sheets, _, content_types = _validate_xlsx(data)
    found = _drawing_images(parts, sheets, content_types)
    return dict(_normalized_xlsx_images(parts, found))


def extract_xlsx_image_galleries_from_bytes(
    data: bytes,
) -> dict[CellRef, tuple[ImageAsset, ...]]:
    if type(data) is not bytes or not 0 < len(data) <= MAX_FILE_BYTES or not data.startswith(b"PK\x03\x04"):
        _fail("SOURCE_TYPE")
    parts, sheets, _, content_types = _validate_xlsx(data)
    galleries: dict[CellRef, list[ImageAsset]] = {}
    found = _drawing_images(parts, sheets, content_types, keep_ambiguous=True)
    for reference, asset in _normalized_xlsx_images(parts, found):
        galleries.setdefault(reference, []).append(asset)
    return {
        reference: tuple(
            sorted(
                {asset.sha256: asset for asset in assets}.values(),
                key=lambda asset: asset.sha256,
            )
        )
        for reference, assets in galleries.items()
    }


def extract_xlsx_images(path: Path) -> dict[CellRef, ImageAsset]:
    _, data = _read_source(path, ".xlsx", MAX_FILE_BYTES)
    return extract_xlsx_images_from_bytes(data)


def _bounded_text(value: object, maximum: int) -> str | None:
    if not isinstance(value, str) or not value or len(value) > maximum:
        return None
    if any(ord(character) < 32 or ord(character) == 127 for character in value):
        return None
    return value


def source_ref(file_id, sheet_or_page, cell_or_bbox) -> dict:
    if not isinstance(file_id, str) or _FILE_ID.fullmatch(file_id) is None:
        _fail("SOURCE_REF")
    if isinstance(sheet_or_page, str):
        if _bounded_text(sheet_or_page, 128) is None:
            _fail("SOURCE_REF")
        location = sheet_or_page
    elif type(sheet_or_page) is int and 1 <= sheet_or_page <= MAX_PDF_PAGES:
        location = sheet_or_page
    else:
        _fail("SOURCE_REF")
    if isinstance(cell_or_bbox, str):
        if _CELL.fullmatch(cell_or_bbox) is None:
            _fail("SOURCE_REF")
        position = cell_or_bbox
    elif isinstance(cell_or_bbox, (tuple, list)) and len(cell_or_bbox) == 4:
        if any(type(value) not in {int, float} or not math.isfinite(value) or abs(value) > 1_000_000 for value in cell_or_bbox):
            _fail("SOURCE_REF")
        position = list(cell_or_bbox)
        if position[0] > position[2] or position[1] > position[3]:
            _fail("SOURCE_REF")
    else:
        _fail("SOURCE_REF")
    return {"file_id": file_id, "sheet_or_page": location, "cell_or_bbox": position}


def neutralize_spreadsheet_text(value: object) -> str:
    if value is None:
        return ""
    try:
        text = value if isinstance(value, str) else str(value)
    except Exception:
        return ""
    text = text[:MAX_TEXT_LENGTH]
    if text.lstrip().startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text[:MAX_TEXT_LENGTH]
