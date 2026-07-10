from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import ipaddress
import mimetypes
import re
import socket
import tempfile
import urllib.error
import urllib.request
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


MAX_IMAGE_BYTES = 8 * 1024 * 1024
WARNING_FILL = "FFF2CC"
OFFICIAL_IMAGE_HOSTS = {
    "offiho_cart": frozenset(
        {
            "offiho.com",
            "www.offiho.com",
            "econosillas.com",
            "www.econosillas.com",
            "offihoblack.com",
            "www.offihoblack.com",
        }
    ),
    "tarkett_cart": frozenset(
        {
            "tarkett.com.mx",
            "www.tarkett.com.mx",
            "tarkett.com.ar",
            "www.tarkett.com.ar",
            "profesional.tarkett.es",
            "media.tarkett-image.com",
        }
    ),
}


def create_catalog_quotation_workbook(
    payload: dict[str, Any],
    output_path: str | Path,
    *,
    source_type: str,
    category_label: str,
    image_dir: str | Path | None = None,
) -> Path:
    if payload.get("source_type") != source_type:
        raise ValueError(f"Payload {category_label} invalido")
    items = list(payload.get("items") or [])
    if not items:
        raise ValueError(f"Payload {category_label} sin productos")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    tmp_context = None
    if image_dir is None:
        tmp_context = tempfile.TemporaryDirectory(prefix="catalog_images_")
        images_root = Path(tmp_context.name)
    else:
        images_root = Path(image_dir)
        images_root.mkdir(parents=True, exist_ok=True)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        _write_headers(ws)
        ws.cell(8, 1).value = f"- {category_label}"
        ws.cell(8, 1).font = Font(bold=True)

        for index, item in enumerate(items, start=1):
            row = index + 8
            code = str(item.get("code", "")).strip()
            name = str(item.get("name", "")).strip()
            unit = str(item.get("unit", "")).strip()
            url = str(item.get("product_url", "") or "").strip()
            description, warning = _description_for_item(item, code, url)
            ws.cell(row, 1).value = index
            ws.cell(row, 2).value = name
            ws.cell(row, 4).value = description
            ws.cell(row, 5).value = unit
            ws.cell(row, 7).value = float(_decimal(item.get("quantity", 0)))
            ws.cell(row, 10).value = _excel_number(_decimal(item.get("unit_price", 0)))
            ws.cell(row, 11).value = url
            if warning:
                ws.cell(row, 4).fill = PatternFill("solid", fgColor=WARNING_FILL)
            ws.row_dimensions[row].height = 72
            _add_catalog_image(ws, row, item.get("image_url"), images_root, code, source_type)

        _set_column_widths(ws)
        wb.save(output)
        wb.close()
    finally:
        if tmp_context is not None:
            tmp_context.cleanup()
    return output


def _write_headers(ws) -> None:
    headers = {
        1: "No.",
        2: "Item",
        3: "Image",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
        11: "URL",
    }
    for col, title in headers.items():
        cell = ws.cell(7, col)
        cell.value = title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2F6B")


def _description_for_item(item: dict[str, Any], code: str, url: str) -> tuple[str, str]:
    parts = [f"Clave: {code}" if code else ""]
    variant = str(item.get("variant", "") or "").strip()
    if variant:
        parts.append(f"Variante: {variant}")
    if url:
        parts.append(f"URL: {url}")
    warning = _stock_warning(item)
    if warning:
        parts.append(warning)
    return " | ".join(part for part in parts if part), warning


def _stock_warning(item: dict[str, Any]) -> str:
    status = str(item.get("stock_status", "") or "").strip()
    if status == "out_of_stock":
        return "ADVERTENCIA: PRODUCTO AGOTADO"
    if status == "insufficient_stock":
        quantity = _excel_number(_decimal(item.get("quantity", 0)))
        available = _excel_number(_decimal(item.get("available_quantity", 0)))
        return f"ADVERTENCIA: EXISTENCIA INSUFICIENTE (solicitado: {quantity}; disponible: {available})"
    return ""


def _set_column_widths(ws) -> None:
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["K"].width = 42


def _add_catalog_image(ws, row: int, image_url: Any, image_dir: Path, code: str, source_type: str) -> None:
    image_path = _download_catalog_image(image_url, image_dir, code, source_type)
    if not image_path:
        return
    try:
        image = XlsxImage(str(image_path))
        if image.width and image.height:
            scale = min(90 / image.width, 66 / image.height)
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
        image.anchor = f"C{row}"
        ws.add_image(image)
    except Exception:
        return


class _OfficialRedirectHandler(urllib.request.HTTPRedirectHandler):
    def __init__(self, allowed_hosts: frozenset[str]):
        super().__init__()
        self._allowed_hosts = allowed_hosts

    def redirect_request(self, req, fp, code, msg, headers, newurl):
        _validate_official_https_url(newurl, self._allowed_hosts)
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def _download_catalog_image(url: Any, image_dir: Path, code: str, source_type: str) -> Path | None:
    clean_url = str(url or "").strip()
    allowed_hosts = OFFICIAL_IMAGE_HOSTS.get(source_type, frozenset())
    if not allowed_hosts:
        return None
    try:
        _validate_official_https_url(clean_url, allowed_hosts)
        request = urllib.request.Request(clean_url, headers={"User-Agent": "Mobiliti Official Catalog/1.0"})
        opener = urllib.request.build_opener(_OfficialRedirectHandler(allowed_hosts))
        with opener.open(request, timeout=18) as response:
            content_type = response.headers.get("content-type", "").split(";", 1)[0].strip().lower()
            content_length = response.headers.get("content-length")
            if not content_type.startswith("image/"):
                return None
            if content_length and int(content_length) > MAX_IMAGE_BYTES:
                return None
            data = response.read(MAX_IMAGE_BYTES + 1)
    except (OSError, ValueError, urllib.error.URLError, urllib.error.HTTPError, socket.gaierror):
        return None
    if not data or len(data) > MAX_IMAGE_BYTES:
        return None
    suffix = mimetypes.guess_extension(content_type) or Path(urlsplit(clean_url).path).suffix or ".jpg"
    safe_code = re.sub(r"[^A-Za-z0-9_-]+", "_", code or "producto")
    destination = image_dir / f"{safe_code}{suffix}"
    destination.write_bytes(data)
    return destination


def _validate_official_https_url(url: str, allowed_hosts: frozenset[str]) -> None:
    parsed = urlsplit(url)
    host = (parsed.hostname or "").lower().rstrip(".")
    if (
        parsed.scheme.lower() != "https"
        or not host
        or parsed.username
        or parsed.password
        or parsed.port not in (None, 443)
        or host not in allowed_hosts
    ):
        raise ValueError("URL de imagen no es HTTPS oficial")
    _resolve_public_host(host)


def _resolve_public_host(host: str) -> None:
    addresses = {record[4][0] for record in socket.getaddrinfo(host, 443, type=socket.SOCK_STREAM)}
    if not addresses or any(not ipaddress.ip_address(address).is_global for address in addresses):
        raise ValueError("Host de imagen no resuelve a una direccion publica")


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def _excel_number(value: Decimal) -> int | float:
    if value == value.to_integral():
        return int(value)
    return float(value)
