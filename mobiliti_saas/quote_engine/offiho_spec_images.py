from __future__ import annotations

import hashlib
import itertools
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from PIL import Image


EMU_PER_PIXEL = 9525

VARIANT_CANONICAL = {
    "AMARILLA": "AMARILLO",
    "AZABACHE": "NEGRO",
    "BLANCA": "BLANCO",
    "CEREZA": "CEREZO",
    "CROMADA": "CROMO",
    "CROMADO": "CROMO",
    "GRISVERDE": "GRIS VERDE",
    "NARANANJA": "NARANJA",
    "NEGRA": "NEGRO",
    "OBSCURO": "OSCURO",
    "OCENAO": "OCEANO",
    "ROJA": "ROJO",
}

FEATURE_ALIASES = {
    "ARO": ("ARO",),
    "BRAZO": ("BRAZO", "BRAZOS"),
    "BRAZOS": ("BRAZO", "BRAZOS"),
    "CABECERA": ("CABECERA", "CAB"),
    "CONECTOR": ("CONECTOR",),
    "CUBIERTA": ("CUBIERTA",),
    "ESTRUCTURA": ("ESTRUCTURA",),
    "KIT": ("KIT",),
    "PEDESTAL": ("PEDESTAL",),
    "PISTON": ("PISTON",),
}


def extract_offiho_spec_images(
    workbook_paths: Iterable[str | Path],
    inventory_items: Iterable[Mapping[str, Any]],
    *,
    assets_dir: str | Path,
    base_url: str,
    source_urls: Mapping[str, str] | None = None,
) -> dict[str, dict[str, Any]]:
    """Materializa imágenes con evidencia exacta de código y variante."""
    items = [dict(item) for item in inventory_items]
    destination = Path(assets_dir)
    destination.mkdir(parents=True, exist_ok=True)
    candidate_groups: dict[str, dict[tuple[str, str], dict[str, Any]]] = {}

    paths = sorted(
        (Path(raw_path) for raw_path in workbook_paths),
        key=lambda path: (path.name.casefold(), str(path.resolve()).casefold()),
    )
    for path in paths:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
        try:
            for sheet in workbook.worksheets:
                code_column, description_column = _product_columns(sheet)
                if code_column is None or description_column is None:
                    continue
                for picture in sheet._images:
                    anchor = getattr(picture, "anchor", None)
                    marker = getattr(anchor, "_from", None)
                    if marker is None:
                        continue
                    row = _picture_product_row(sheet, anchor, code_column, description_column)
                    if row is None:
                        continue
                    anchor_row = int(marker.row) + 1
                    code = _text(sheet.cell(row, code_column).value)
                    description = _text(sheet.cell(row, description_column).value)
                    if not code or not description:
                        continue
                    exact_items = [
                        item for item in items if _is_exact_match(code, description, item)
                    ]
                    if not exact_items:
                        continue
                    image_bytes = picture._data()
                    sha256 = hashlib.sha256(image_bytes).hexdigest()
                    content_type, extension, width, height = _image_metadata(image_bytes)
                    product_url = _source_url(path, source_urls)
                    reference = {
                        "workbook": path.name,
                        "sheet": sheet.title,
                        "row": row,
                        "code": code,
                        "anchor_row_delta": row - anchor_row,
                        "product_url": product_url,
                    }
                    representative = {
                        "product_url": product_url,
                        "description": description,
                        "source_workbook": path.name,
                        "source_sheet": sheet.title,
                        "source_row": row,
                        "source_code": code,
                        "anchor_row_delta": row - anchor_row,
                    }
                    rank = _source_rank(path, sheet.title, row, code)
                    for item in exact_items:
                        inventory_key = _text(item.get("inventory_key"))
                        if not inventory_key:
                            continue
                        groups = candidate_groups.setdefault(inventory_key, {})
                        group_key = (_compact(code), sha256)
                        group = groups.get(group_key)
                        if group is None:
                            group = {
                                "image_bytes": image_bytes,
                                "image_sha256": sha256,
                                "image_content_type": content_type,
                                "image_extension": extension,
                                "image_width": width,
                                "image_height": height,
                                "representative": representative,
                                "representative_rank": rank,
                                "source_references": [],
                            }
                            groups[group_key] = group
                        if rank < group["representative_rank"]:
                            group["representative"] = representative
                            group["representative_rank"] = rank
                        group["source_references"].append(reference)
        finally:
            workbook.close()

    matches: dict[str, dict[str, Any]] = {}
    for inventory_key in sorted(candidate_groups):
        groups = list(candidate_groups[inventory_key].values())
        selected = min(
            groups,
            key=lambda group: (group["representative_rank"], group["image_sha256"]),
        )
        filename = (
            f"{_slug(inventory_key)}-{selected['image_sha256'][:16]}."
            f"{selected['image_extension']}"
        )
        _materialize_exact_image(destination / filename, selected["image_bytes"])
        references = sorted(
            selected["source_references"],
            key=lambda ref: (
                not str(ref["sheet"]).upper().startswith("SPEC"),
                str(ref["workbook"]).casefold(),
                str(ref["sheet"]).casefold(),
                int(ref["row"]),
            ),
        )
        representative = selected["representative"]
        matches[inventory_key] = {
            "inventory_key": inventory_key,
            "image_url": _asset_url(base_url, filename),
            "description_source": "spec_guide",
            "match_status": "spec_guide_exact",
            "image_sha256": selected["image_sha256"],
            "image_content_type": selected["image_content_type"],
            "image_content_length": len(selected["image_bytes"]),
            "image_width": selected["image_width"],
            "image_height": selected["image_height"],
            "source_references": references,
            "source_reference_count": len(references),
            "exact_image_candidate_count": len(groups),
            **representative,
        }
    return matches


def _source_rank(path: Path, sheet: str, row: int, code: str) -> tuple[Any, ...]:
    return (
        not sheet.upper().startswith("SPEC"),
        path.name.casefold(),
        str(path.resolve()).casefold(),
        sheet.casefold(),
        row,
        _compact(code),
    )


def _materialize_exact_image(path: Path, payload: bytes) -> None:
    if path.exists():
        existing_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        payload_hash = hashlib.sha256(payload).hexdigest()
        if existing_hash != payload_hash:
            raise RuntimeError(f"Colision de asset Offiho en {path}")
        return
    path.write_bytes(payload)


def _product_columns(sheet: Any) -> tuple[int | None, int | None]:
    for row in sheet.iter_rows():
        normalized = {
            cell.column: _compact(cell.value)
            for cell in row
            if cell.value not in (None, "")
        }
        code = next(
            (
                column
                for column, value in normalized.items()
                if value in {"COD", "CODIGO"}
            ),
            None,
        )
        description = next(
            (
                column
                for column, value in normalized.items()
                if value in {"DESCRIPCION", "DESCRIPTION"}
            ),
            None,
        )
        if code is not None and description is not None:
            return code, description
    return None, None


def _picture_product_row(
    sheet: Any,
    anchor: Any,
    code_column: int,
    description_column: int,
) -> int | None:
    start = getattr(anchor, "_from", None)
    if start is None:
        return None
    start_row = int(start.row) + 1
    end_marker = getattr(anchor, "to", None)
    required_rows = max(sheet.max_row, start_row + 1)
    if end_marker is not None:
        required_rows = max(required_rows, int(end_marker.row) + 2)
    row_starts, row_heights = _row_geometry(sheet, required_rows)
    top = row_starts[start_row] + int(start.rowOff or 0) / EMU_PER_PIXEL
    if end_marker is not None:
        end_row = int(end_marker.row) + 1
        bottom = row_starts[end_row] + int(end_marker.rowOff or 0) / EMU_PER_PIXEL
    else:
        extent = getattr(anchor, "ext", None)
        bottom = top + int(getattr(extent, "cy", 0) or 0) / EMU_PER_PIXEL
    if bottom <= top:
        return None

    overlaps: list[tuple[float, int]] = []
    for row in range(1, sheet.max_row + 1):
        if not _text(sheet.cell(row, code_column).value) or not _text(
            sheet.cell(row, description_column).value
        ):
            continue
        overlap = max(
            0.0,
            min(bottom, row_starts[row] + row_heights[row]) - max(top, row_starts[row]),
        )
        if overlap > 0:
            overlaps.append((overlap, row))
    return max(overlaps, default=(0.0, 0))[1] or None


def _row_geometry(sheet: Any, max_row: int) -> tuple[dict[int, float], dict[int, float]]:
    default_points = float(sheet.sheet_format.defaultRowHeight or 15.0)
    starts = {1: 0.0}
    heights: dict[int, float] = {}
    for row in range(1, max_row + 2):
        explicit = sheet.row_dimensions[row].height
        points = default_points if explicit is None else float(explicit)
        heights[row] = max(0.0, points * 96.0 / 72.0)
        starts[row + 1] = starts[row] + heights[row]
    return starts, heights


def _is_exact_match(
    row_code: str,
    description: str,
    item: Mapping[str, Any],
) -> bool:
    base = _compact(item.get("code"))
    code = _compact(row_code)
    if not base or not code.startswith(base):
        return False
    suffix = code[len(base) :]
    if base[-1:].isdigit() and suffix[:1].isdigit():
        return False
    signatures = _variant_signatures(item.get("variant"))
    if not signatures:
        return _deduplicated_compact(row_code) == _deduplicated_compact(
            item.get("inventory_key")
        ) and _required_features_supported(row_code, description, suffix, item)
    inventory_key = _compact(item.get("inventory_key"))
    configuration = ""
    for signature in signatures:
        if not signature or signature not in suffix:
            continue
        remainder = suffix.replace(signature, "", 1)
        if not remainder or remainder in inventory_key:
            configuration = remainder
            break
    return bool(configuration or suffix in signatures) and _required_features_supported(
        row_code,
        description,
        configuration,
        item,
    )


def _required_features_supported(
    row_code: str,
    description: str,
    configuration: str,
    item: Mapping[str, Any],
) -> bool:
    required = [token for token in _tokens(item.get("name")) if token in FEATURE_ALIASES]
    if not required:
        return True
    evidence_tokens = set(_tokens(description)) | set(_tokens(row_code))
    base = _compact(item.get("code"))
    return all(
        any(
            alias in evidence_tokens
            or _compact(alias) == base
            or _compact(alias) in configuration
            for alias in FEATURE_ALIASES[feature]
        )
        for feature in required
    )


def _variant_signatures(value: Any) -> set[str]:
    variant_tokens = [VARIANT_CANONICAL.get(token, token) for token in _tokens(value)]
    if not variant_tokens:
        return set()
    options: list[tuple[str, ...]] = []
    for token in variant_tokens:
        aliases = {token}
        if token == "CROMO":
            aliases.add("CR")
        options.append(tuple(sorted(aliases)))
    return {
        "".join(ordering)
        for chosen in itertools.product(*options)
        for ordering in set(itertools.permutations(chosen))
    }


def _image_metadata(payload: bytes) -> tuple[str, str, int, int]:
    with Image.open(BytesIO(payload)) as image:
        image_format = str(image.format or "").upper()
        if image_format == "PNG":
            return "image/png", "png", int(image.width), int(image.height)
        if image_format in {"JPEG", "JPG"}:
            return "image/jpeg", "jpg", int(image.width), int(image.height)
        raise ValueError(f"Formato de imagen Offiho no soportado: {image_format or 'desconocido'}")


def _source_url(path: Path, source_urls: Mapping[str, str] | None) -> str:
    if source_urls:
        for key in (str(path), str(path.resolve()), path.name):
            value = _text(source_urls.get(key))
            if value:
                return value
    return path.resolve().as_uri()


def _asset_url(base_url: str, filename: str) -> str:
    prefix = str(base_url or "").rstrip("/")
    return f"{prefix}/{filename}" if prefix else filename


def _slug(value: Any) -> str:
    normalized = _ascii_upper(value).lower()
    return re.sub(r"[^a-z0-9]+", "-", normalized).strip("-") or "offiho-spec"


def _compact(value: Any) -> str:
    return "".join(_tokens(value))


def _tokens(value: Any) -> list[str]:
    return re.findall(r"[A-Z0-9]+", _ascii_upper(value))


def _deduplicated_compact(value: Any) -> str:
    result: list[str] = []
    for token in _tokens(value):
        if not result or token != result[-1]:
            result.append(token)
    return "".join(result)


def _ascii_upper(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _text(value).upper())
    return "".join(character for character in normalized if not unicodedata.combining(character))


def _text(value: Any) -> str:
    return " ".join(str("" if value is None else value).split())
