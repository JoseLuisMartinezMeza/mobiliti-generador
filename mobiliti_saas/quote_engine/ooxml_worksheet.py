"""Mutación OOXML de la hoja oficial ``Mobiliti`` sin guardar el libro.

El editor conserva el árbol original y reemplaza únicamente las filas de la
tabla dinámica. Las fórmulas, estilos, propiedades de fila, combinaciones y
reglas se clonan del XML oficial.
"""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from decimal import Decimal
import re
from typing import Literal, Sequence
import xml.etree.ElementTree as ET

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from .mobiliti_layout import MobilitiRowMap, SectionNeed, plan_mobiliti_layout
from .ooxml_formula import translate_formula


MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XM = "http://schemas.microsoft.com/office/excel/2006/main"
XML = "http://www.w3.org/XML/1998/namespace"
TABLE_LAST_COLUMN = 34  # AH
CANONICAL_TOTAL_ROW = 573
CANONICAL_AUXILIARY_START = 574
CANONICAL_AUXILIARY_END = 610
INPUT_COLUMNS = frozenset((4, 5, 6, 8, 10, 11, 16))
CANONICAL_SUBTOTAL_ROWS = tuple(range(47, 573, 35))
XLSX_MAX_ROW = 1_048_576
XLSX_MAX_COLUMN = 16_384
MAX_EXCEL_CELL_TEXT_LENGTH = 32_767
_CELL_REFERENCE = re.compile(r"(?P<column>\$?[A-Z]{1,3})(?P<row_abs>\$?)(?P<row>[1-9][0-9]*)$")
_RANGE_REFERENCE = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[^'!]+)!)?"
    r"(?P<first>\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
    r"(?::(?P<last>\$?[A-Z]{1,3}\$?[1-9][0-9]*))?$"
)

for prefix, namespace in (
    ("", MAIN),
    ("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships"),
    ("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006"),
    ("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"),
    ("xr", "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"),
    ("xr2", "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"),
    ("xr3", "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"),
    ("x14", X14),
    ("xm", XM),
):
    ET.register_namespace(prefix, namespace)


@dataclass(frozen=True)
class MobilitiCellWrite:
    coordinate: str
    kind: Literal["number", "text", "boolean"]
    value: Decimal | str | bool


@dataclass(frozen=True)
class MobilitiSheetMutation:
    xml: bytes
    row_map: MobilitiRowMap


@dataclass(frozen=True)
class OfficialMobilitiBlock:
    first_section_header: ET.Element
    first_product_row: ET.Element
    first_subtotal_row: ET.Element
    section_header: ET.Element
    product_row: ET.Element
    subtotal_row: ET.Element
    total_row: ET.Element
    auxiliary_rows: tuple[ET.Element, ...]
    merges: tuple[ET.Element, ...]


@dataclass(frozen=True)
class SharedFormulaMaster:
    coordinate: str
    text: str


class FormulaIndex:
    """Resuelve seguidores shared desde los masters del XML oficial."""

    def __init__(self, sheet_data: ET.Element):
        masters: dict[str, SharedFormulaMaster] = {}
        followers: list[tuple[str, str]] = []
        for cell in sheet_data.findall(f".//{{{MAIN}}}c"):
            formula = cell.find(f"{{{MAIN}}}f")
            if formula is None or formula.attrib.get("t") != "shared":
                continue
            coordinate = cell.attrib["r"]
            shared_index = formula.attrib.get("si")
            if shared_index is None:
                raise ValueError(
                    f"La estructura de Mobiliti contiene shared formula sin si: {coordinate}"
                )
            if formula.text:
                if shared_index in masters:
                    raise ValueError(
                        f"La estructura de Mobiliti repite master shared si={shared_index}"
                    )
                masters[shared_index] = SharedFormulaMaster(coordinate, formula.text)
            else:
                followers.append((coordinate, shared_index))
        missing = sorted({index for _coordinate, index in followers if index not in masters})
        if missing:
            raise ValueError(
                f"La estructura de Mobiliti contiene followers shared sin master: {missing}"
            )
        self.masters = masters

    def source_text(self, formula: ET.Element, coordinate: str) -> str:
        formula_type = formula.attrib.get("t")
        if formula_type != "shared":
            if not formula.text:
                raise ValueError(f"Fórmula oficial vacía: {coordinate}")
            return formula.text
        if formula.text:
            return formula.text
        master = self.masters[formula.attrib["si"]]
        return _translate_formula_ranges(
            "=" + master.text,
            origin=master.coordinate,
            target=coordinate,
        )[1:]


class WorksheetEditor:
    """Editor estrecho sobre un ``worksheet`` SpreadsheetML existente."""

    def __init__(self, root: ET.Element):
        self.root = root
        self.row_map: MobilitiRowMap | None = None
        self.sheet_data = root.find(f"{{{MAIN}}}sheetData")
        if self.sheet_data is None:
            raise ValueError("Mobiliti no contiene sheetData")
        _validate_worksheet_structure(self.sheet_data)
        self.formulas = FormulaIndex(self.sheet_data)

    @classmethod
    def from_xml(cls, payload: bytes) -> "WorksheetEditor":
        try:
            root = ET.fromstring(payload)
        except (ET.ParseError, TypeError) as error:
            raise ValueError("El XML de Mobiliti no es válido") from error
        if root.tag != f"{{{MAIN}}}worksheet":
            raise ValueError("La parte indicada no es un worksheet SpreadsheetML")
        return cls(root)

    def to_xml(self) -> bytes:
        return ET.tostring(self.root, encoding="utf-8", xml_declaration=True)

    def row(self, number: int) -> ET.Element | None:
        return self.sheet_data.find(f"{{{MAIN}}}row[@r='{number}']")

    def require_row(self, number: int) -> ET.Element:
        row = self.row(number)
        if row is None:
            raise ValueError(f"La plantilla oficial no contiene la fila Mobiliti {number}")
        return row

    def replace_table_row(
        self,
        target_row: int,
        source: ET.Element,
        source_row: int,
        row_map: MobilitiRowMap,
        *,
        last_column: int = TABLE_LAST_COLUMN,
        product_range: tuple[int, int] | None = None,
        clear_input_formulas: bool = False,
    ) -> ET.Element:
        existing = self.row(target_row)
        sidecar = []
        if existing is not None:
            sidecar = [
                deepcopy(cell)
                for cell in existing.findall(f"{{{MAIN}}}c")
                if _cell_column(cell) > last_column
            ]
        clone = deepcopy(source)
        clone.set("r", str(target_row))
        for cell in list(clone.findall(f"{{{MAIN}}}c")):
            if _cell_column(cell) > last_column:
                clone.remove(cell)
                continue
            _move_cell(
                cell,
                source_row=source_row,
                target_row=target_row,
                row_map=row_map,
                product_range=product_range,
                clear_input_formulas=clear_input_formulas,
                formula_index=self.formulas,
            )
        for cell in sidecar:
            clone.append(cell)
        if existing is not None:
            self.sheet_data.remove(existing)
        self.sheet_data.append(clone)
        return clone

    def replace_whole_row(
        self,
        target_row: int,
        source: ET.Element,
        source_row: int,
        row_map: MobilitiRowMap,
    ) -> ET.Element:
        existing = self.row(target_row)
        clone = deepcopy(source)
        clone.set("r", str(target_row))
        for cell in clone.findall(f"{{{MAIN}}}c"):
            _move_cell(
                cell,
                source_row=source_row,
                target_row=target_row,
                row_map=row_map,
                formula_index=self.formulas,
            )
        if existing is not None:
            self.sheet_data.remove(existing)
        self.sheet_data.append(clone)
        return clone

    def sort_rows(self) -> None:
        children = list(self.sheet_data)
        children.sort(key=lambda row: int(row.attrib["r"]))
        self.sheet_data[:] = children

    def set_boolean(self, coordinate: str, value: bool) -> None:
        """Escribe un booleano en una fila existente con coordenada A1 estricta."""

        self.set_typed_values((MobilitiCellWrite(coordinate, "boolean", value),))

    def set_inline_string(self, coordinate: str, value: str) -> None:
        """Escribe un ``inlineStr`` XML 1.0 acotado en una fila existente."""

        self.set_typed_values((MobilitiCellWrite(coordinate, "text", value),))

    def set_typed_values(self, writes: Sequence[MobilitiCellWrite]) -> None:
        """Prevalida y aplica atómicamente escrituras tipadas a celdas existentes."""

        if isinstance(writes, (str, bytes, bytearray)) or not isinstance(
            writes, Sequence
        ):
            raise TypeError("Las escrituras tipadas Mobiliti deben ser una secuencia")
        prepared: list[tuple[ET.Element, MobilitiCellWrite]] = []
        seen: set[str] = set()
        for write in writes:
            if not isinstance(write, MobilitiCellWrite):
                raise TypeError("Escritura tipada Mobiliti inválida")
            row, normalized = self._typed_write_target(write.coordinate)
            if normalized in seen:
                raise ValueError(f"Escritura tipada Mobiliti duplicada: {normalized}")
            seen.add(normalized)
            _validate_write_value(write.kind, write.value)
            prepared.append((row, write))
        for row, write in prepared:
            _set_cell_value(
                row,
                write.coordinate,
                write.kind,
                write.value,
            )

    def _typed_write_target(self, coordinate: str) -> tuple[ET.Element, str]:
        if not isinstance(coordinate, str) or re.fullmatch(
            r"[A-Z]{1,3}[1-9][0-9]*", coordinate
        ) is None:
            raise ValueError(f"Coordenada Mobiliti inválida: {coordinate!r}")
        match = _CELL_REFERENCE.fullmatch(coordinate)
        assert match is not None
        column = column_index_from_string(match.group("column"))
        row_number = int(match.group("row"))
        if column > XLSX_MAX_COLUMN or row_number > XLSX_MAX_ROW:
            raise ValueError(f"Coordenada Mobiliti fuera de XLSX: {coordinate}")
        row = self.require_row(row_number)
        if _find_cell(row, column) is None:
            raise ValueError(f"Celda destino Mobiliti ausente: {coordinate}")
        return row, coordinate


def capture_official_mobiliti_block(
    editor: WorksheetEditor,
    first_section_row: int = 13,
    second_section_row: int = 48,
    total_row: int = CANONICAL_TOTAL_ROW,
) -> OfficialMobilitiBlock:
    """Captura las siete filas canónicas y el bloque auxiliar oficial."""

    merge_cells = editor.root.find(f"{{{MAIN}}}mergeCells")
    merges = tuple(deepcopy(node) for node in (() if merge_cells is None else merge_cells))
    return OfficialMobilitiBlock(
        first_section_header=deepcopy(editor.require_row(first_section_row)),
        first_product_row=deepcopy(editor.require_row(first_section_row + 1)),
        first_subtotal_row=deepcopy(editor.require_row(second_section_row - 1)),
        section_header=deepcopy(editor.require_row(second_section_row)),
        product_row=deepcopy(editor.require_row(second_section_row + 1)),
        subtotal_row=deepcopy(editor.require_row(second_section_row + 34)),
        total_row=deepcopy(editor.require_row(total_row)),
        auxiliary_rows=tuple(
            deepcopy(editor.require_row(row))
            for row in range(total_row + 1, CANONICAL_AUXILIARY_END + 1)
        ),
        merges=merges,
    )


def apply_mobiliti_layout(editor: WorksheetEditor, row_map: MobilitiRowMap) -> None:
    """Limpia solo la tabla oficial que será reconstruida desde sus clones."""

    editor.row_map = row_map
    if row_map.total_row != CANONICAL_TOTAL_ROW:
        canonical_total = editor.require_row(CANONICAL_TOTAL_ROW)
        for cell in list(canonical_total.findall(f"{{{MAIN}}}c")):
            if _cell_column(cell) <= 36:  # El bloque total oficial llega hasta AJ.
                canonical_total.remove(cell)
    target_rows = {
        row
        for section in row_map.sections
        for row in range(section.section_row, section.subtotal_row + 1)
    } | {row_map.total_row}
    for row in list(editor.sheet_data):
        number = int(row.attrib["r"])
        if 13 <= number <= CANONICAL_TOTAL_ROW and number not in target_rows:
            for cell in list(row.findall(f"{{{MAIN}}}c")):
                if _cell_column(cell) <= TABLE_LAST_COLUMN:
                    row.remove(cell)
        if CANONICAL_AUXILIARY_START <= number <= CANONICAL_AUXILIARY_END:
            editor.sheet_data.remove(row)


def _translate_static_structural_formulas(
    editor: WorksheetEditor, row_map: MobilitiRowMap
) -> None:
    """Actualiza referencias al total en cabecera y panel lateral preservados."""

    for row in editor.sheet_data:
        row_number = int(row.attrib["r"])
        if row_number > CANONICAL_TOTAL_ROW:
            continue
        for cell in row.findall(f"{{{MAIN}}}c"):
            if row_number >= 13 and _cell_column(cell) <= TABLE_LAST_COLUMN:
                continue
            formula = cell.find(f"{{{MAIN}}}f")
            if formula is None:
                continue
            coordinate = cell.attrib["r"]
            if formula.attrib.get("t") in {"array", "dataTable"}:
                if formula.attrib.get("ref") == coordinate:
                    _translate_special_formula(
                        cell,
                        formula,
                        origin=coordinate,
                        target=coordinate,
                        row_map=row_map,
                    )
                continue
            source_text = editor.formulas.source_text(formula, coordinate)
            translated_text = _translate_official_formula(
                "=" + source_text,
                origin=coordinate,
                target=coordinate,
                row_map=row_map,
            )[1:]
            formula.text = translated_text
            if translated_text != source_text:
                cached_value = cell.find(f"{{{MAIN}}}v")
                if cached_value is not None:
                    cell.remove(cached_value)
            if formula.attrib.get("t") == "shared":
                for shared_attribute in ("t", "ref", "si"):
                    formula.attrib.pop(shared_attribute, None)


def relocate_official_auxiliary_rows(
    editor: WorksheetEditor,
    row_map: MobilitiRowMap,
    canonical: OfficialMobilitiBlock,
) -> None:
    """Mueve las filas oficiales 574:610 inmediatamente después del total."""

    for offset, source in enumerate(canonical.auxiliary_rows, start=1):
        editor.replace_whole_row(
            row_map.total_row + offset,
            source,
            CANONICAL_TOTAL_ROW + offset,
            row_map,
        )
    _update_dimension(editor, row_map.total_row + len(canonical.auxiliary_rows))


def _preflight_cloneable_formulas(canonical: OfficialMobilitiBlock) -> None:
    rows = (
        canonical.first_section_header,
        canonical.first_product_row,
        canonical.first_subtotal_row,
        canonical.section_header,
        canonical.product_row,
        canonical.subtotal_row,
        canonical.total_row,
        *canonical.auxiliary_rows,
    )
    for row in rows:
        for cell in row.findall(f"{{{MAIN}}}c"):
            formula = cell.find(f"{{{MAIN}}}f")
            if formula is None:
                continue
            formula_type = formula.attrib.get("t")
            if formula_type in {"array", "dataTable"}:
                coordinate = cell.attrib.get("r", "?")
                if formula.attrib.get("ref") != coordinate or (
                    formula_type == "array" and not formula.text
                ):
                    raise ValueError(
                        f"La fórmula {formula_type} de {coordinate} no se puede "
                        "clonar de forma segura"
                    )


def _preflight_static_special_formulas(
    editor: WorksheetEditor, row_map: MobilitiRowMap
) -> None:
    if row_map.total_row == CANONICAL_TOTAL_ROW:
        return
    for cell in editor.sheet_data.findall(f".//{{{MAIN}}}c"):
        formula = cell.find(f"{{{MAIN}}}f")
        if formula is None or formula.attrib.get("t") not in {"array", "dataTable"}:
            continue
        coordinate = cell.attrib["r"]
        for attribute in ("ref", "r1", "r2"):
            if attribute in formula.attrib and _RANGE_REFERENCE.fullmatch(
                formula.attrib[attribute]
            ) is None:
                raise ValueError(
                    f"La fórmula {formula.attrib['t']} de {coordinate} tiene "
                    f"{attribute} inválido en preflight"
                )
        if formula.attrib.get("ref") == coordinate:
            continue
        if _special_formula_has_moved_structural_reference(formula):
            raise ValueError(
                f"La fórmula {formula.attrib['t']} de {coordinate} requiere "
                "preflight antes de mover el total"
            )


def _special_formula_has_moved_structural_reference(formula: ET.Element) -> bool:
    values = list(formula.attrib.values())
    if formula.text:
        values.extend(
            token.value
            for token in Tokenizer("=" + formula.text).items
            if token.type == "OPERAND" and token.subtype == "RANGE"
        )
    for value in values:
        match = _RANGE_REFERENCE.fullmatch(value)
        if match is None or not _is_local_mobiliti_reference(match.group("sheet")):
            continue
        endpoints = [match.group("first")]
        if match.group("last"):
            endpoints.append(match.group("last"))
        parsed = [_CELL_REFERENCE.fullmatch(endpoint) for endpoint in endpoints]
        if all(item is not None for item in parsed) and any(
            CANONICAL_TOTAL_ROW <= int(item.group("row")) <= CANONICAL_AUXILIARY_END
            for item in parsed
            if item is not None
        ):
            return True
    return False


def clone_section_header(
    editor: WorksheetEditor,
    canonical_row: ET.Element,
    target_row: int,
    title: str,
    row_map: MobilitiRowMap | None = None,
    *,
    source_row: int = 48,
) -> None:
    effective_row_map = row_map or editor.row_map
    if effective_row_map is None:
        raise ValueError("apply_mobiliti_layout debe ejecutarse antes de clonar secciones")
    clone = editor.replace_table_row(
        target_row, canonical_row, source_row, effective_row_map
    )
    title_column = 4 if source_row == 13 else 1
    _set_cell_value(clone, f"{get_column_letter(title_column)}{target_row}", "text", title)


def clone_formula_row(
    editor: WorksheetEditor,
    canonical_row: ET.Element,
    target_row: int,
    row_map: MobilitiRowMap,
    *,
    source_row: int = 49,
) -> None:
    editor.replace_table_row(
        target_row,
        canonical_row,
        source_row,
        row_map,
        clear_input_formulas=True,
    )


def clone_subtotal_row(
    editor: WorksheetEditor,
    canonical_row: ET.Element,
    section,
    row_map: MobilitiRowMap,
    *,
    source_row: int = 82,
) -> None:
    product_range = (
        (section.product_start, section.product_start + section.item_count - 1)
        if section.item_count
        else (0, 0)
    )
    clone = editor.replace_table_row(
        section.subtotal_row,
        canonical_row,
        source_row,
        row_map,
        product_range=product_range,
    )
    section_number = row_map.sections.index(section) + 1
    _set_cell_value(
        clone,
        f"A{section.subtotal_row}",
        "text",
        f"Subtotales Sección {section_number}",
    )


def clone_total_row(
    editor: WorksheetEditor,
    canonical_row: ET.Element,
    target_row: int,
    row_map: MobilitiRowMap,
) -> None:
    clone = editor.replace_table_row(
        target_row,
        canonical_row,
        CANONICAL_TOTAL_ROW,
        row_map,
        last_column=36,
    )
    for cell in clone.findall(f"{{{MAIN}}}c"):
        formula = cell.find(f"{{{MAIN}}}f")
        if formula is None or not formula.text:
            continue
        # replace_table_row ya tradujo la fórmula; se reconstruye de nuevo desde
        # el original para incluir cualquier sección adicional a las 16 oficiales.
        source_cell = _find_cell(canonical_row, _cell_column(cell))
        source_formula = None if source_cell is None else source_cell.find(f"{{{MAIN}}}f")
        if source_formula is not None:
            source_text = editor.formulas.source_text(
                source_formula, source_cell.attrib["r"]
            )
            formula.text = _translate_total_formula(
                "=" + source_text,
                origin=source_cell.attrib["r"],
                target=cell.attrib["r"],
                row_map=row_map,
            )[1:]


def clear_mobiliti_input_cells(editor: WorksheetEditor, row_map: MobilitiRowMap) -> None:
    """Elimina valores contaminados exclusivamente de las columnas de entrada."""

    for section in row_map.sections:
        for row_number in range(section.product_start, section.product_start + section.capacity):
            row = editor.require_row(row_number)
            for column in INPUT_COLUMNS:
                cell = _find_cell(row, column)
                if cell is not None:
                    _clear_cell(cell)


def apply_mobiliti_cell_writes(
    editor: WorksheetEditor,
    cell_writes: Sequence[MobilitiCellWrite],
    row_map: MobilitiRowMap,
) -> None:
    product_rows = set(row_map.item_rows)
    validated: list[tuple[MobilitiCellWrite, int]] = []
    for write in cell_writes:
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", write.coordinate.upper())
        if match is None:
            raise ValueError(f"Coordenada Mobiliti inválida: {write.coordinate!r}")
        column = column_index_from_string(match.group(1))
        row_number = int(match.group(2))
        if row_number not in product_rows or column not in INPUT_COLUMNS:
            raise ValueError(f"Escritura fuera de inputs Mobiliti: {write.coordinate}")
        _validate_write_value(write.kind, write.value)
        validated.append((write, row_number))
    for write, row_number in validated:
        row = editor.require_row(row_number)
        _set_cell_value(row, write.coordinate.upper(), write.kind, write.value)


def translate_mobiliti_validations(editor: WorksheetEditor, row_map: MobilitiRowMap) -> None:
    """Traduce validaciones clásicas y x14, conservando sus extensiones."""

    validations = editor.root.find(f"{{{MAIN}}}dataValidations")
    if validations is not None:
        for validation in validations:
            if "sqref" in validation.attrib:
                validation.set("sqref", _translate_sqref(validation.attrib["sqref"], row_map))
    for sqref in editor.root.findall(f".//{{{XM}}}sqref"):
        if sqref.text:
            sqref.text = _translate_sqref(sqref.text, row_map)


def translate_mobiliti_conditional_formatting(
    editor: WorksheetEditor, row_map: MobilitiRowMap
) -> None:
    """Mueve y amplía reglas oficiales de formato condicional por sección."""

    original = [
        node
        for node in list(editor.root)
        if node.tag == f"{{{MAIN}}}conditionalFormatting"
    ]
    dynamic: list[tuple[int, ET.Element]] = []
    fixed: list[ET.Element] = []
    for node in original:
        section_indices = _sqref_dynamic_indices(node.attrib.get("sqref", ""))
        if not section_indices:
            fixed.append(node)
        else:
            dynamic.extend((section_index, node) for section_index in section_indices)
        editor.root.remove(node)

    translated: list[ET.Element] = []
    seen_nodes: set[int] = set()
    for section_index, node in dynamic:
        if section_index < len(row_map.sections):
            translated.append(
                _translate_cf_node(
                    node,
                    section_index,
                    row_map,
                    source_section_index=section_index,
                    include_fixed=id(node) not in seen_nodes,
                )
            )
            seen_nodes.add(id(node))
    second_templates = [node for index, node in dynamic if index == 1]
    for section_index in range(len(CANONICAL_SUBTOTAL_ROWS), len(row_map.sections)):
        translated.extend(
            _translate_cf_node(node, section_index, row_map, source_section_index=1)
            for node in second_templates
        )

    insertion_index = next(
        (
            index
            for index, node in enumerate(editor.root)
            if node.tag in {f"{{{MAIN}}}dataValidations", f"{{{MAIN}}}pageMargins"}
        ),
        len(editor.root),
    )
    for node in [*fixed, *translated]:
        editor.root.insert(insertion_index, node)
        insertion_index += 1


def _translate_cf_node(
    source: ET.Element,
    target_section_index: int,
    row_map: MobilitiRowMap,
    *,
    source_section_index: int | None = None,
    include_fixed: bool = False,
) -> ET.Element:
    clone = deepcopy(source)
    source_index = (
        _sqref_section_index(source.attrib.get("sqref", ""))
        if source_section_index is None
        else source_section_index
    )
    if source_index is None:
        return clone
    source_start = 14 + source_index * 35
    target_start = row_map.sections[target_section_index].product_start
    sqref_tokens: list[str] = []
    for token in source.attrib["sqref"].split():
        token_index = _sqref_token_section_index(token)
        if token_index is None:
            if include_fixed and token not in sqref_tokens:
                sqref_tokens.append(token)
        elif token_index == source_index:
            translated_token = _translate_sqref_token(
                token,
                source_start,
                target_start,
                row_map.sections[target_section_index].capacity,
            )
            if translated_token not in sqref_tokens:
                sqref_tokens.append(translated_token)
    if not sqref_tokens:
        raise ValueError("Formato condicional Mobiliti sin sqref después de traducir")
    clone.set("sqref", " ".join(sqref_tokens))
    for formula in clone.findall(f".//{{{MAIN}}}formula"):
        if formula.text:
            formula.text = _translate_official_formula(
                "=" + formula.text,
                origin=f"A{source_start}",
                target=f"A{target_start}",
                row_map=row_map,
            )[1:]
    return clone


def _validate_worksheet_structure(sheet_data: ET.Element) -> None:
    previous_row = 0
    seen_rows: set[int] = set()
    for row in sheet_data:
        if row.tag != f"{{{MAIN}}}row":
            raise ValueError("La estructura de Mobiliti contiene un hijo inválido en sheetData")
        raw_row = row.attrib.get("r")
        if raw_row is None or not raw_row.isdecimal():
            raise ValueError("La estructura de Mobiliti contiene row@r inválido")
        row_number = int(raw_row)
        if (
            row_number < 1
            or row_number > XLSX_MAX_ROW
            or row_number in seen_rows
            or row_number <= previous_row
        ):
            raise ValueError("La estructura de Mobiliti contiene filas duplicadas o desordenadas")
        seen_rows.add(row_number)
        previous_row = row_number

        previous_column = 0
        seen_cells: set[str] = set()
        for cell in row.findall(f"{{{MAIN}}}c"):
            coordinate = cell.attrib.get("r", "")
            match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", coordinate)
            if match is None:
                raise ValueError("La estructura de Mobiliti contiene c@r inválido")
            column = column_index_from_string(match.group(1))
            cell_row = int(match.group(2))
            if (
                coordinate in seen_cells
                or cell_row != row_number
                or column > XLSX_MAX_COLUMN
                or column <= previous_column
            ):
                raise ValueError(
                    "La estructura de Mobiliti contiene celdas duplicadas, fuera de fila o desordenadas"
                )
            seen_cells.add(coordinate)
            previous_column = column


def _cell_column(cell: ET.Element) -> int:
    match = re.match(r"([A-Z]{1,3})", cell.attrib.get("r", ""))
    if match is None:
        raise ValueError("Celda oficial sin coordenada válida")
    return column_index_from_string(match.group(1))


def _find_cell(row: ET.Element, column: int) -> ET.Element | None:
    for cell in row.findall(f"{{{MAIN}}}c"):
        if _cell_column(cell) == column:
            return cell
    return None


def _move_cell(
    cell: ET.Element,
    *,
    source_row: int,
    target_row: int,
    row_map: MobilitiRowMap,
    product_range: tuple[int, int] | None = None,
    clear_input_formulas: bool = False,
    formula_index: FormulaIndex,
) -> None:
    column = get_column_letter(_cell_column(cell))
    origin = cell.attrib.get("r", f"{column}{source_row}")
    target = f"{column}{target_row}"
    cell.set("r", target)
    formula = cell.find(f"{{{MAIN}}}f")
    if formula is None:
        return
    if clear_input_formulas and _cell_column(cell) in INPUT_COLUMNS:
        _clear_cell(cell)
        return
    formula_type = formula.attrib.get("t")
    if formula_type in {"array", "dataTable"}:
        if formula.attrib.get("ref") != origin or (
            formula_type == "array" and not formula.text
        ):
            raise ValueError(
                f"La fórmula {formula_type} de {origin} no se puede clonar de forma segura"
            )
        _translate_special_formula(
            cell,
            formula,
            origin=origin,
            target=target,
            row_map=row_map,
            product_range=product_range,
        )
        return
    source_text = formula_index.source_text(formula, origin)
    formula.text = _translate_official_formula(
        "=" + source_text,
        origin=origin,
        target=target,
        row_map=row_map,
        product_range=product_range,
    )[1:]
    if formula_type == "shared":
        for shared_attribute in ("t", "ref", "si"):
            formula.attrib.pop(shared_attribute, None)
    value = cell.find(f"{{{MAIN}}}v")
    if value is not None:
        cell.remove(value)


def _translate_special_formula(
    cell: ET.Element,
    formula: ET.Element,
    *,
    origin: str,
    target: str,
    row_map: MobilitiRowMap,
    product_range: tuple[int, int] | None = None,
) -> None:
    formula_type = formula.attrib.get("t")
    if formula_type not in {"array", "dataTable"}:
        raise ValueError(f"Fórmula especial inválida en {origin}")
    if formula_type == "array" and not formula.text:
        raise ValueError(f"La fórmula array de {origin} no tiene texto")

    changed = False
    if formula.text:
        translated_text = _translate_official_formula(
            "=" + formula.text,
            origin=origin,
            target=target,
            row_map=row_map,
            product_range=product_range,
        )[1:]
        changed = translated_text != formula.text
        formula.text = translated_text

    for attribute, value in list(formula.attrib.items()):
        if attribute == "t" or _RANGE_REFERENCE.fullmatch(value) is None:
            continue
        translated_value = _translate_official_formula(
            "=" + value,
            origin=origin,
            target=target,
            row_map=row_map,
            product_range=product_range,
        )[1:]
        changed = changed or translated_value != value
        formula.set(attribute, translated_value)

    if changed:
        cached_value = cell.find(f"{{{MAIN}}}v")
        if cached_value is not None:
            cell.remove(cached_value)


def _translate_official_formula(
    formula: str,
    *,
    origin: str,
    target: str,
    row_map: MobilitiRowMap,
    product_range: tuple[int, int] | None = None,
) -> str:
    translated = _translate_formula_ranges(formula, origin=origin, target=target)
    result = ["="]
    original_tokens = Tokenizer(formula).items
    translated_tokens = Tokenizer(translated).items
    if len(original_tokens) != len(translated_tokens):
        raise ValueError(f"La traducción oficial cambió la estructura de {origin}")
    for original_token, translated_token in zip(
        original_tokens, translated_tokens, strict=True
    ):
        value = translated_token.value
        if (
            original_token.type == "OPERAND"
            and original_token.subtype == "RANGE"
            and _RANGE_REFERENCE.fullmatch(original_token.value)
        ):
            value = _structural_reference(
                original_token.value,
                value,
                row_map,
                product_range=product_range,
            )
        result.append(value)
    return "".join(result)


def _translate_formula_ranges(formula: str, *, origin: str, target: str) -> str:
    result = ["="]
    for token in Tokenizer(formula).items:
        value = token.value
        if token.type == "OPERAND" and token.subtype == "RANGE" and _RANGE_REFERENCE.fullmatch(value):
            value = translate_formula(
                "=" + value,
                origin=origin,
                target=target,
                sheet="Mobiliti",
            )[1:]
        result.append(value)
    return "".join(result)


def _structural_reference(
    original: str,
    translated: str,
    row_map: MobilitiRowMap,
    *,
    product_range: tuple[int, int] | None,
) -> str:
    match = _RANGE_REFERENCE.fullmatch(original)
    if match is None:
        return translated
    if not _is_local_mobiliti_reference(match.group("sheet")):
        return translated
    endpoints = [match.group("first")]
    if match.group("last"):
        endpoints.append(match.group("last"))
    parsed = [_CELL_REFERENCE.fullmatch(endpoint) for endpoint in endpoints]
    if any(item is None for item in parsed):
        return translated
    rows = [int(item.group("row")) for item in parsed if item is not None]
    if product_range is not None and (
        rows == [14, 46] or rows == [49, 81]
    ):
        if product_range == (0, 0):
            return "0"
        first, last = product_range
        return _reference_with_rows(original, (first, last))
    if all(CANONICAL_TOTAL_ROW <= row <= CANONICAL_AUXILIARY_END for row in rows):
        return _reference_with_rows(
            original,
            tuple(row_map.total_row + row - CANONICAL_TOTAL_ROW for row in rows),
        )
    return translated


def _is_local_mobiliti_reference(sheet: str | None) -> bool:
    if sheet is None:
        return True
    normalized = sheet
    if normalized.startswith("'") and normalized.endswith("'"):
        normalized = normalized[1:-1].replace("''", "'")
    return normalized.casefold() == "mobiliti"


def _reference_with_rows(reference: str, rows: tuple[int, ...]) -> str:
    match = _RANGE_REFERENCE.fullmatch(reference)
    if match is None:
        return reference
    endpoints = [match.group("first")]
    if match.group("last"):
        endpoints.append(match.group("last"))
    rewritten = []
    for endpoint, row in zip(endpoints, rows, strict=True):
        cell = _CELL_REFERENCE.fullmatch(endpoint)
        if cell is None:
            return reference
        rewritten.append(f"{cell.group('column')}{cell.group('row_abs')}{row}")
    prefix = f"{match.group('sheet')}!" if match.group("sheet") else ""
    return prefix + ":".join(rewritten)


def _translate_total_formula(
    formula: str,
    *,
    origin: str,
    target: str,
    row_map: MobilitiRowMap,
) -> str:
    translated = _translate_official_formula(
        formula,
        origin=origin,
        target=target,
        row_map=row_map,
    )
    tokens = Tokenizer(formula).items
    translated_tokens = Tokenizer(translated).items
    if len(tokens) != len(translated_tokens):
        raise ValueError(f"La traducción del total cambió su estructura: {origin}")
    reference_indices: list[int] = []
    source_rows: list[int] = []
    source_references: list[str] = []
    reference_signature: tuple[str, str, str] | None = None
    for index, token in enumerate(tokens):
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        match = _RANGE_REFERENCE.fullmatch(token.value)
        if (
            match is None
            or match.group("last") is not None
            or not _is_local_mobiliti_reference(match.group("sheet"))
        ):
            continue
        cell = _CELL_REFERENCE.fullmatch(match.group("first"))
        if cell is None or int(cell.group("row")) not in CANONICAL_SUBTOTAL_ROWS:
            continue
        signature = (
            (match.group("sheet") or "").casefold(),
            cell.group("column"),
            cell.group("row_abs"),
        )
        if reference_signature is None:
            reference_signature = signature
        elif signature != reference_signature:
            raise ValueError(
                f"El total oficial requiere 16 subtotales únicos en orden canónico: {origin}"
            )
        reference_indices.append(index)
        source_rows.append(int(cell.group("row")))
        source_references.append(token.value)
    ascending = list(CANONICAL_SUBTOTAL_ROWS)
    descending = list(reversed(CANONICAL_SUBTOTAL_ROWS))
    if len(reference_indices) != len(CANONICAL_SUBTOTAL_ROWS) or source_rows not in (
        ascending,
        descending,
    ):
        raise ValueError(
            f"El total oficial requiere 16 subtotales únicos en orden canónico: {origin}"
        )
    first, second, last = reference_indices[0], reference_indices[1], reference_indices[-1]
    separator = "".join(token.value for token in translated_tokens[first + 1 : second])
    for left, right in zip(reference_indices, reference_indices[1:]):
        candidate = "".join(token.value for token in translated_tokens[left + 1 : right])
        if candidate != separator:
            raise ValueError(
                f"El total oficial requiere 16 subtotales únicos en orden canónico: {origin}"
            )
    rows = (
        list(reversed(row_map.subtotal_rows))
        if source_rows == descending
        else list(row_map.subtotal_rows)
    )
    references = separator.join(
        _reference_with_rows(source_references[0], (row,)) for row in rows
    )
    prefix = "".join(token.value for token in translated_tokens[:first])
    suffix = "".join(token.value for token in translated_tokens[last + 1 :])
    return "=" + prefix + references + suffix


def _clear_cell(cell: ET.Element) -> None:
    for child in list(cell):
        if child.tag in {
            f"{{{MAIN}}}f",
            f"{{{MAIN}}}v",
            f"{{{MAIN}}}is",
        }:
            cell.remove(child)
    cell.attrib.pop("t", None)


def _set_cell_value(
    row: ET.Element,
    coordinate: str,
    kind: Literal["number", "text", "boolean"],
    value: Decimal | str | bool,
) -> None:
    _validate_write_value(kind, value)
    column = column_index_from_string(re.match(r"[A-Z]{1,3}", coordinate).group())
    cell = _find_cell(row, column)
    if cell is None:
        cell = ET.Element(f"{{{MAIN}}}c", {"r": coordinate})
        row.append(cell)
    _clear_cell(cell)
    cell.set("r", coordinate)
    if kind == "text":
        if not isinstance(value, str):
            raise TypeError("Una escritura text requiere str")
        cell.set("t", "inlineStr")
        inline = ET.SubElement(cell, f"{{{MAIN}}}is")
        text = ET.SubElement(inline, f"{{{MAIN}}}t")
        if _has_significant_whitespace(value):
            text.set(f"{{{XML}}}space", "preserve")
        text.text = value
    elif kind == "boolean":
        if type(value) is not bool:
            raise TypeError("Una escritura boolean requiere bool")
        cell.set("t", "b")
        ET.SubElement(cell, f"{{{MAIN}}}v").text = "1" if value else "0"
    elif kind == "number":
        if isinstance(value, bool) or not isinstance(value, Decimal):
            raise TypeError("Una escritura number requiere Decimal")
        ET.SubElement(cell, f"{{{MAIN}}}v").text = format(value, "f")
    else:
        raise ValueError(f"Tipo de escritura Mobiliti inválido: {kind!r}")
    _sort_cells(row)


def _validate_write_value(
    kind: Literal["number", "text", "boolean"], value: Decimal | str | bool
) -> None:
    if kind == "text":
        if not isinstance(value, str):
            raise TypeError("Una escritura text requiere str")
        if len(value) > MAX_EXCEL_CELL_TEXT_LENGTH:
            raise ValueError("Una escritura text excede 32767 caracteres")
        if any(not _is_xml_10_character(ord(character)) for character in value):
            raise ValueError("El texto Mobiliti contiene caracteres inválidos para XML 1.0")
    elif kind == "boolean":
        if type(value) is not bool:
            raise TypeError("Una escritura boolean requiere bool")
    elif kind == "number":
        if isinstance(value, bool) or not isinstance(value, Decimal):
            raise TypeError("Una escritura number requiere Decimal")
        if not value.is_finite():
            raise ValueError("Una escritura number requiere un Decimal finito")
    else:
        raise ValueError(f"Tipo de escritura Mobiliti inválido: {kind!r}")


def _is_xml_10_character(codepoint: int) -> bool:
    return (
        codepoint in {0x9, 0xA, 0xD}
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def _has_significant_whitespace(value: str) -> bool:
    return (
        value != value.strip()
        or any(character in value for character in "\t\r\n")
        or "  " in value
    )


def _sort_cells(row: ET.Element) -> None:
    cells = list(row.findall(f"{{{MAIN}}}c"))
    others = [node for node in row if node.tag != f"{{{MAIN}}}c"]
    cells.sort(key=_cell_column)
    row[:] = [*cells, *others]


def _translate_sqref(value: str, row_map: MobilitiRowMap) -> str:
    result = []
    seen = set()
    canonical_templates: list[str] = []
    for token in value.split():
        match = _RANGE_REFERENCE.fullmatch(token)
        section_index = _sqref_token_section_index(token)
        if match is None or section_index is None:
            translated = token
        else:
            section = row_map.sections[section_index]
            translated = _translate_sqref_token(
                token, 14 + section_index * 35, section.product_start, section.capacity
            )
            if section_index == 1 and token not in canonical_templates:
                canonical_templates.append(token)
        if translated not in seen:
            result.append(translated)
            seen.add(translated)
    for section in row_map.sections[len(CANONICAL_SUBTOTAL_ROWS) :]:
        for template in canonical_templates:
            token = _translate_sqref_token(
                template,
                49,
                section.product_start,
                section.capacity,
            )
            if token not in seen:
                result.append(token)
                seen.add(token)
    return " ".join(result)


def _translate_sqref_token(
    token: str, source_start: int, target_start: int, capacity: int
) -> str:
    match = _RANGE_REFERENCE.fullmatch(token)
    if match is None:
        return token
    first = _CELL_REFERENCE.fullmatch(match.group("first"))
    last = _CELL_REFERENCE.fullmatch(match.group("last"))
    if first is None or last is None:
        return token
    prefix = f"{match.group('sheet')}!" if match.group("sheet") else ""
    first_delta = int(first.group("row")) - source_start
    last_delta = int(last.group("row")) - (source_start + 32)
    return (
        f"{prefix}{first.group('column')}{first.group('row_abs')}{target_start + first_delta}:"
        f"{last.group('column')}{last.group('row_abs')}"
        f"{target_start + capacity - 1 + last_delta}"
    )


def _sqref_token_section_index(token: str) -> int | None:
    match = _RANGE_REFERENCE.fullmatch(token)
    if match is None or match.group("last") is None:
        return None
    first = _CELL_REFERENCE.fullmatch(match.group("first"))
    last = _CELL_REFERENCE.fullmatch(match.group("last"))
    if first is None or last is None:
        return None
    first_row, last_row = int(first.group("row")), int(last.group("row"))
    if first_row < 14 or (first_row - 14) % 35 or last_row != first_row + 32:
        return None
    index = (first_row - 14) // 35
    return index if index < len(CANONICAL_SUBTOTAL_ROWS) else None


def _sqref_section_index(value: str) -> int | None:
    indices = _sqref_dynamic_indices(value)
    return indices[0] if len(indices) == 1 else None


def _sqref_dynamic_indices(value: str) -> tuple[int, ...]:
    result: list[int] = []
    for token in value.split():
        index = _sqref_token_section_index(token)
        if index is not None and index not in result:
            result.append(index)
    return tuple(result)


def _update_dimension(editor: WorksheetEditor, last_row: int) -> None:
    dimension = editor.root.find(f"{{{MAIN}}}dimension")
    if dimension is None:
        return
    reference = dimension.attrib.get("ref", "A1:AV610")
    if ":" not in reference:
        return
    start, end = reference.split(":", 1)
    column = re.match(r"\$?[A-Z]{1,3}", end)
    if column:
        dimension.set("ref", f"{start}:{column.group()}{last_row}")


def _replace_merges(
    editor: WorksheetEditor,
    canonical: OfficialMobilitiBlock,
    row_map: MobilitiRowMap,
) -> None:
    container = editor.root.find(f"{{{MAIN}}}mergeCells")
    if container is None:
        container = ET.Element(f"{{{MAIN}}}mergeCells")
        sheet_data_index = list(editor.root).index(editor.sheet_data)
        editor.root.insert(sheet_data_index + 1, container)
    preserved = []
    for merge in canonical.merges:
        bounds = _merge_bounds(merge.attrib["ref"])
        if bounds is None:
            preserved.append(deepcopy(merge))
            continue
        min_col, min_row, max_col, max_row = bounds
        cross_boundary = min_col <= TABLE_LAST_COLUMN < max_col
        canonical_total = (
            min_row == max_row == CANONICAL_TOTAL_ROW and max_col <= 36
        )
        dynamic_table = (
            min_row >= 13
            and max_row <= CANONICAL_TOTAL_ROW
            and min_col <= TABLE_LAST_COLUMN
            and not cross_boundary
        )
        auxiliary = min_row >= CANONICAL_AUXILIARY_START and max_row <= CANONICAL_AUXILIARY_END
        if not dynamic_table and not auxiliary and not canonical_total:
            preserved.append(deepcopy(merge))
    generated = []
    for index, section in enumerate(row_map.sections):
        templates = (
            ((13, section.section_row), (47, section.subtotal_row))
            if index == 0
            else ((48, section.section_row), (82, section.subtotal_row))
        )
        for source_row, target_row in templates:
            generated.extend(_row_merges(canonical.merges, source_row, target_row, TABLE_LAST_COLUMN))
        product_source = 14 if index == 0 else 49
        for target_row in range(
            section.product_start, section.product_start + section.capacity
        ):
            generated.extend(
                _row_merges(
                    canonical.merges,
                    product_source,
                    target_row,
                    TABLE_LAST_COLUMN,
                )
            )
    generated.extend(
        _row_merges(
            canonical.merges,
            CANONICAL_TOTAL_ROW,
            row_map.total_row,
            36,
        )
    )
    for merge in canonical.merges:
        bounds = _merge_bounds(merge.attrib["ref"])
        if bounds and bounds[1] >= CANONICAL_AUXILIARY_START and bounds[3] <= CANONICAL_AUXILIARY_END:
            generated.append(_shift_merge(merge, row_map.total_row - CANONICAL_TOTAL_ROW))
    by_reference: dict[str, ET.Element] = {}
    for merge in (*preserved, *generated):
        by_reference.setdefault(merge.attrib["ref"], merge)
    ordered = sorted(by_reference.values(), key=_merge_element_sort_key)
    container[:] = ordered
    container.set("count", str(len(container)))


def _row_merges(
    merges: Sequence[ET.Element], source_row: int, target_row: int, max_column: int
) -> list[ET.Element]:
    result = []
    for merge in merges:
        bounds = _merge_bounds(merge.attrib["ref"])
        if bounds and bounds[1] == bounds[3] == source_row and bounds[2] <= max_column:
            result.append(_shift_merge(merge, target_row - source_row))
    return result


def _merge_bounds(reference: str) -> tuple[int, int, int, int] | None:
    match = _RANGE_REFERENCE.fullmatch(reference)
    if match is None or match.group("last") is None:
        return None
    first = _CELL_REFERENCE.fullmatch(match.group("first"))
    last = _CELL_REFERENCE.fullmatch(match.group("last"))
    if first is None or last is None:
        return None
    return (
        column_index_from_string(first.group("column").replace("$", "")),
        int(first.group("row")),
        column_index_from_string(last.group("column").replace("$", "")),
        int(last.group("row")),
    )


def _shift_merge(merge: ET.Element, row_delta: int) -> ET.Element:
    clone = deepcopy(merge)
    match = _RANGE_REFERENCE.fullmatch(clone.attrib["ref"])
    rows = []
    for endpoint in (match.group("first"), match.group("last")):
        cell = _CELL_REFERENCE.fullmatch(endpoint)
        rows.append(int(cell.group("row")) + row_delta)
    clone.set("ref", _reference_with_rows(clone.attrib["ref"], tuple(rows)))
    return clone


def _merge_element_sort_key(merge: ET.Element) -> tuple[int, int, int, int, str]:
    reference = merge.attrib.get("ref", "")
    bounds = _merge_bounds(reference)
    if bounds is None:
        return (XLSX_MAX_ROW + 1, XLSX_MAX_COLUMN + 1, 0, 0, reference)
    min_col, min_row, max_col, max_row = bounds
    return (min_row, min_col, max_row, max_col, reference)


def build_mobiliti_sheet(
    official_sheet_xml: bytes,
    needs: list[SectionNeed],
    cell_writes: Sequence[MobilitiCellWrite],
) -> MobilitiSheetMutation:
    """Construye la mutación de ``Mobiliti`` a partir del XML oficial."""

    row_map = plan_mobiliti_layout(needs)
    editor = WorksheetEditor.from_xml(official_sheet_xml)
    canonical = capture_official_mobiliti_block(editor)
    _preflight_cloneable_formulas(canonical)
    _preflight_static_special_formulas(editor, row_map)
    _translate_static_structural_formulas(editor, row_map)
    apply_mobiliti_layout(editor, row_map)

    for index, section in enumerate(row_map.sections):
        first = index == 0
        header = canonical.first_section_header if first else canonical.section_header
        # La fila 49 del archivo oficial no contiene la superficie completa de
        # formulas (L/N estan vacias), aunque secciones oficiales posteriores
        # si las contienen y calcChain las declara. La fila 14 es la plantilla
        # oficial completa y el traductor estructural la adapta a cada bloque.
        product = canonical.first_product_row
        subtotal = canonical.first_subtotal_row if first else canonical.subtotal_row
        header_source_row = 13 if first else 48
        product_source_row = 14
        subtotal_source_row = 47 if first else 82
        clone_section_header(
            editor,
            header,
            section.section_row,
            section.title,
            row_map,
            source_row=header_source_row,
        )
        for target_row in range(section.product_start, section.product_start + section.capacity):
            clone_formula_row(
                editor,
                product,
                target_row,
                row_map,
                source_row=product_source_row,
            )
        clone_subtotal_row(
            editor,
            subtotal,
            section,
            row_map,
            source_row=subtotal_source_row,
        )

    clone_total_row(editor, canonical.total_row, row_map.total_row, row_map)
    relocate_official_auxiliary_rows(editor, row_map, canonical)
    clear_mobiliti_input_cells(editor, row_map)
    apply_mobiliti_cell_writes(editor, cell_writes, row_map)
    _replace_merges(editor, canonical, row_map)
    translate_mobiliti_validations(editor, row_map)
    translate_mobiliti_conditional_formatting(editor, row_map)
    editor.sort_rows()
    return MobilitiSheetMutation(xml=editor.to_xml(), row_map=row_map)
