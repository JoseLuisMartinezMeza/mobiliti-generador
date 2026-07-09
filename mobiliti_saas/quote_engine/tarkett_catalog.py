from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import json
import mimetypes
import re
import tempfile
import urllib.error
import urllib.request

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


CATALOG_PATH = Path(__file__).resolve().parent / "data" / "tarkett_catalog.json"
TARKETT_CART_SOURCE_TYPE = "tarkett_cart"


@dataclass(frozen=True)
class TarkettCatalogItem:
    code: str
    name: str
    unit: str
    available_quantity: Decimal
    product_url: str = ""
    image_url: str = ""
    match_status: str = "unmatched"

    @classmethod
    def from_dict(cls, raw: dict[str, Any]) -> "TarkettCatalogItem":
        return cls(
            code=str(raw.get("code", "")).strip(),
            name=str(raw.get("name", "")).strip(),
            unit=str(raw.get("unit", "")).strip(),
            available_quantity=_decimal(raw.get("available_quantity", 0)),
            product_url=str(raw.get("product_url", "") or "").strip(),
            image_url=str(raw.get("image_url", "") or "").strip(),
            match_status=str(raw.get("match_status", "unmatched") or "unmatched").strip(),
        )

    def to_public_dict(self, reserved_quantity: Decimal = Decimal("0"), reserved_by_others: bool = False) -> dict[str, Any]:
        return {
            "code": self.code,
            "name": self.name,
            "unit": self.unit,
            "available_quantity": _json_number(self.available_quantity),
            "reserved_quantity": _json_number(reserved_quantity),
            "reserved_by_others": bool(reserved_by_others),
            "product_url": self.product_url,
            "image_url": self.image_url,
            "match_status": self.match_status,
        }


def load_tarkett_catalog(path: str | Path | None = None) -> dict[str, Any]:
    catalog_path = Path(path or CATALOG_PATH)
    raw = json.loads(catalog_path.read_text(encoding="utf-8"))
    items = [TarkettCatalogItem.from_dict(item) for item in raw.get("items", [])]
    return {
        "source_hash": str(raw.get("source_hash", "")),
        "generated_at": str(raw.get("generated_at", "")),
        "source_file": str(raw.get("source_file", "")),
        "items": items,
        "by_code": {item.code: item for item in items},
    }


def build_tarkett_cart_payload(
    raw_items: list[dict[str, Any]],
    *,
    catalog: dict[str, Any] | None = None,
) -> dict[str, Any]:
    loaded_catalog = catalog or load_tarkett_catalog()
    by_code: dict[str, TarkettCatalogItem] = loaded_catalog["by_code"]
    if not raw_items:
        raise ValueError("El carrito Tarkett esta vacio")

    lines: list[dict[str, Any]] = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            raise ValueError("Cada producto Tarkett debe ser un objeto")
        code = str(raw.get("code", raw.get("clave", ""))).strip()
        if not code:
            raise ValueError("Cada producto Tarkett requiere clave")
        item = by_code.get(code)
        if item is None:
            raise ValueError(f"Producto Tarkett no encontrado: {code}")
        quantity = _decimal(raw.get("quantity", raw.get("cantidad", 0)))
        if quantity <= 0:
            raise ValueError(f"Cantidad invalida para {code}")
        if quantity > item.available_quantity:
            raise ValueError(f"Cantidad mayor a existencia para {code}")
        lines.append(
            {
                "code": item.code,
                "name": item.name,
                "unit": item.unit,
                "quantity": _json_number(quantity),
                "unit_price": 0,
                "available_quantity": _json_number(item.available_quantity),
                "product_url": item.product_url,
                "image_url": item.image_url,
            }
        )

    return {
        "source_type": TARKETT_CART_SOURCE_TYPE,
        "catalog_source_hash": loaded_catalog["source_hash"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "items": lines,
    }


def create_tarkett_quotation_workbook(
    cart_payload: dict[str, Any],
    output_path: str | Path,
    *,
    image_dir: str | Path | None = None,
) -> Path:
    if cart_payload.get("source_type") != TARKETT_CART_SOURCE_TYPE:
        raise ValueError("Payload Tarkett invalido")
    items = list(cart_payload.get("items") or [])
    if not items:
        raise ValueError("Payload Tarkett sin productos")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    tmp_context = None
    if image_dir is None:
        tmp_context = tempfile.TemporaryDirectory(prefix="tarkett_images_")
        images_root = Path(tmp_context.name)
    else:
        images_root = Path(image_dir)
        images_root.mkdir(parents=True, exist_ok=True)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        headers = {
            1: "No.",
            2: "Item",
            3: "Image",
            4: "Description",
            5: "Dimension",
            7: "Qty",
            10: "List Price",
            11: "URL",
        }
        for col, title in headers.items():
            cell = ws.cell(7, col)
            cell.value = title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B2F6B")
        ws.cell(8, 1).value = "- Tarkett"
        ws.cell(8, 1).font = Font(bold=True)

        for index, item in enumerate(items, start=1):
            row = index + 8
            code = str(item.get("code", "")).strip()
            name = str(item.get("name", "")).strip()
            unit = str(item.get("unit", "")).strip()
            url = str(item.get("product_url", "") or "").strip()
            ws.cell(row, 1).value = index
            ws.cell(row, 2).value = name
            ws.cell(row, 4).value = f"Clave: {code}" + (f" | URL: {url}" if url else "")
            ws.cell(row, 5).value = unit
            ws.cell(row, 7).value = float(_decimal(item.get("quantity", 0)))
            ws.cell(row, 10).value = 0
            ws.cell(row, 11).value = url
            ws.row_dimensions[row].height = 72
            image_path = _download_image(item.get("image_url"), images_root, code)
            if image_path:
                try:
                    img = XlsxImage(str(image_path))
                    if img.width and img.height:
                        scale = min(90 / img.width, 66 / img.height)
                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)
                    img.anchor = f"C{row}"
                    ws.add_image(img)
                except Exception:
                    pass

        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["D"].width = 46
        ws.column_dimensions["K"].width = 42
        wb.save(output)
        wb.close()
    finally:
        if tmp_context is not None:
            tmp_context.cleanup()
    return output


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def _json_number(value: Any) -> int | float:
    if not isinstance(value, Decimal):
        value = _decimal(value)
    if value == value.to_integral():
        return int(value)
    return float(value)


def _download_image(url: Any, image_dir: Path, code: str) -> Path | None:
    clean_url = str(url or "").strip()
    if not re.match(r"^https?://", clean_url, flags=re.I):
        return None
    try:
        req = urllib.request.Request(clean_url, headers={"User-Agent": "Mobiliti Tarkett Catalog/1.0"})
        with urllib.request.urlopen(req, timeout=18) as response:
            content_type = response.headers.get("content-type", "").split(";")[0].strip()
            data = response.read(8 * 1024 * 1024)
    except (OSError, urllib.error.URLError, TimeoutError):
        return None
    if not data or not content_type.startswith("image/"):
        return None
    suffix = mimetypes.guess_extension(content_type) or Path(clean_url.split("?", 1)[0]).suffix or ".jpg"
    safe_code = re.sub(r"[^A-Za-z0-9_-]+", "_", code or "producto")
    dest = image_dir / f"{safe_code}{suffix}"
    dest.write_bytes(data)
    return dest
