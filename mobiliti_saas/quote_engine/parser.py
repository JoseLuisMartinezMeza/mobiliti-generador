from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


Q_HEADER_ROW = 7


@dataclass
class QuoteItem:
    tipo: str
    row: int
    nombre: Any = ""
    descripcion: Any = ""
    dimension: Any = ""
    m3: Any = None
    cantidad: Any = None
    precio: Any = None
    categoria: str = ""
    proveedor: Any = ""
    descuento: Any = None
    moneda_original: Any = ""
    precio_original: Any = None
    tipo_cambio_congelado: Any = None
    referencia_fuente: Any = ""
    modo_precio: Any = ""
    electrificacion_automatica: Any = None
    canonical_key: Any = ""
    source_hash: Any = ""
    source_row: Any = None
    upstream_row_hash: Any = ""


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).lower().strip()
    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _find_header_column(ws, terms: list[str], row: int = Q_HEADER_ROW) -> str | None:
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=row, column=col).value)
        if not header:
            continue
        tokens = set(header.split())
        if any(term in header or header in term or term in tokens for term in terms):
            return get_column_letter(col)
    return None


def _find_exact_header_column(ws, terms: list[str], row: int = Q_HEADER_ROW) -> str | None:
    expected = {normalize_header(term) for term in terms}
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=row, column=col).value) in expected:
            return get_column_letter(col)
    return None


def detect_columns(ws) -> dict[str, str]:
    keywords = {
        "cantidad": ["qty", "quantity", "cantidad"],
        "unit_price": ["unit price", "unitprice", "price unit", "precio unitario"],
        "total_price": ["tot price", "total price", "totprice", "amount", "precio total"],
        "list_price": ["list price", "listprice", "price list"],
        "descripcion": ["description", "desc", "descripcion"],
        "dimension": ["dimension", "dimensions", "size", "medida"],
        "proveedor": ["supplier", "provider", "proveedor"],
        "descuento": ["discount percent", "discount", "descuento"],
        "moneda_original": ["original currency", "base currency", "moneda original"],
        "precio_original": ["original unit price", "base unit price", "precio original"],
        "tipo_cambio_congelado": ["frozen exchange rate", "exchange rate", "tipo de cambio"],
        "referencia_fuente": ["source reference", "referencia fuente"],
        "modo_precio": ["price mode", "modo precio"],
        "electrificacion_automatica": ["auto electrification", "electrificacion automatica"],
    }
    technical_keywords = {
        "canonical_key": ["canonical key", "clave canonica"],
        "source_hash": ["source hash", "hash fuente"],
        "source_row": ["original source row", "fila fuente original"],
        "upstream_row_hash": ["upstream row hash", "hash fila upstream"],
    }
    column_map: dict[str, str] = {}

    vol_col = _find_header_column(ws, ["vol", "volumen", "volume"])
    if vol_col:
        column_map["m3"] = vol_col

    for key, terms in keywords.items():
        found = _find_header_column(ws, terms)
        if found:
            column_map[key] = found
    for key, terms in technical_keywords.items():
        found = _find_exact_header_column(ws, terms)
        if found:
            column_map[key] = found

    if "m3" not in column_map:
        column_map["m3"] = column_map.get("dimension", "E")
    if "unit_price" not in column_map and "list_price" not in column_map:
        for col in range(1, ws.max_column + 1):
            if "price" in normalize_header(ws.cell(row=Q_HEADER_ROW, column=col).value):
                column_map["unit_price"] = get_column_letter(col)
                break
    return column_map


def col_index(column_map: dict[str, str], key: str, fallback: str) -> int:
    return column_index_from_string(column_map.get(key, fallback))


def read_items(source_path: str | Path) -> tuple[list[QuoteItem], dict[str, str]]:
    wb = load_workbook(source_path, data_only=True)
    if "Quotation" not in wb.sheetnames:
        wb.close()
        raise ValueError("El archivo no contiene hoja Quotation")

    ws = wb["Quotation"]
    column_map = detect_columns(ws)
    desc_col = col_index(column_map, "descripcion", "D")
    dim_col = col_index(column_map, "dimension", "E")
    qty_col = col_index(column_map, "cantidad", "G")
    price_col = col_index(column_map, "list_price", column_map.get("unit_price", "J"))

    def optional_value(row: int, key: str, default: Any) -> Any:
        column = column_map.get(key)
        if column is None:
            return default
        return ws.cell(row=row, column=column_index_from_string(column)).value

    last_row = ws.max_row
    for row in range(last_row, 0, -1):
        if ws.cell(row=row, column=1).value is not None:
            last_row = row
            break

    current_category = ""
    items: list[QuoteItem] = []
    for row in range(Q_HEADER_ROW + 1, last_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        descripcion = ws.cell(row=row, column=desc_col).value
        dimension = ws.cell(row=row, column=dim_col).value

        if isinstance(no_val, str) and no_val.startswith("-"):
            current_category = no_val.strip("- ").strip()
            items.append(QuoteItem(tipo="categoria", row=row, nombre=current_category))
            continue
        if (item_name is None or item_name == "") and (no_val is None or no_val == ""):
            continue
        if isinstance(no_val, (int, float)):
            items.append(
                QuoteItem(
                    tipo="producto",
                    row=row,
                    nombre=item_name,
                    descripcion=descripcion,
                    dimension=dimension,
                    m3=optional_value(row, "m3", None),
                    cantidad=ws.cell(row=row, column=qty_col).value,
                    precio=ws.cell(row=row, column=price_col).value,
                    categoria=current_category,
                    proveedor=optional_value(row, "proveedor", ""),
                    descuento=optional_value(row, "descuento", None),
                    moneda_original=optional_value(row, "moneda_original", ""),
                    precio_original=optional_value(row, "precio_original", None),
                    tipo_cambio_congelado=optional_value(row, "tipo_cambio_congelado", None),
                    referencia_fuente=optional_value(row, "referencia_fuente", ""),
                    modo_precio=optional_value(row, "modo_precio", ""),
                    electrificacion_automatica=optional_value(
                        row, "electrificacion_automatica", None
                    ),
                    canonical_key=optional_value(row, "canonical_key", ""),
                    source_hash=optional_value(row, "source_hash", ""),
                    source_row=optional_value(row, "source_row", None),
                    upstream_row_hash=optional_value(row, "upstream_row_hash", ""),
                )
            )
        elif (no_val is None or no_val == "") and item_name:
            current_category = str(item_name).strip()
            items.append(QuoteItem(tipo="categoria", row=row, nombre=current_category))

    wb.close()
    return items, column_map
