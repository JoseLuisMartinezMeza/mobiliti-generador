from __future__ import annotations

from collections import OrderedDict
from copy import copy, deepcopy
from dataclasses import dataclass, replace
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
import hashlib
import json
import math
import os
import posixpath
from pathlib import Path
import re
import shutil
import time
from typing import Any, Mapping, Sequence
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from PIL import Image as PILImage

from .catalog_cart import WARNING_FILL
from .classification import classify_product_name, load_category_dictionary
from .descriptions import build_product_description, normalize_description_language
from .ai_image_provider import (
    dezgo_config_from_env,
    generate_with_dezgo,
    image_provider_failure_is_fatal,
    normalize_image_provider,
)
from .image_processing import improve_image_map, improve_product_image_bytes
from .images import center_image_in_cell, extract_images, fit_image_to_cell, image_scale_for_category
from .mobiliti_layout import (
    BASE_FIRST_SECTION_ROW,
    BASE_PRODUCT_CAPACITY,
    BASE_SECTION_COUNT,
    SectionNeed,
    plan_mobiliti_layout,
)
from .mobiliti_pricing import (
    PricingRowBinding,
    build_mobiliti_pricing_writes,
    lumbro_frozen_cost,
    write_official_currency_selector,
)
from .ooxml_package import (
    PackageMutation,
    XlsxPackage,
    relationship_part_name,
    relationship_type_uris,
    resolve_internal_target,
)
from .ooxml_worksheet import (
    MobilitiCellWrite,
    MobilitiSheetMutation,
    WorksheetEditor,
    build_mobiliti_sheet,
)
from .official_composer import (
    CDMX_LUMBRO_SHEET,
    ComposeRequest,
    CotizacionImageSource,
    CotizacionMetadata,
    CotizacionPriceTerm,
    CotizacionProduct,
    CotizacionSection,
    CotizacionSheetEditor,
    FLETE_ROUTES,
    _clone_row_region,
    _next_picture_id,
    _next_relationship_id,
    _product_image_anchor,
    _set_formula,
    _set_inline_string,
    _set_number,
    _stable_image_relationship_type,
    _validate_compose_paths,
    compose_official_quote,
)
from .official_template import load_template_contract
from .parser import QuoteItem, col_index, read_items
from .quotation_sheets import (
    QuotationDataRow,
    _parse_worksheet_document,
    _strip_external_image_links_with_embedded_fallback,
    _with_canonical_hash,
    _xml_bytes,
    build_quotation_data_sheet,
    official_provider_name,
    official_region_name,
    transplant_quotation,
)
from .supplier_catalog import safe_excel_text
from .sunon_image_provider import (
    extract_product_code,
    fetch_sunon_catalog_product_image,
    fetch_sunon_product_image,
    find_sunon_catalog_match,
    normalize_sunon_code,
    sunon_lookup_enabled,
)
from .template_profiles import available_template_profiles


_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)

MONEY_FORMAT = '$#,##0.00;[Red]-$#,##0.00;"-"'
MIXED_MONEY_FORMATS = {
    "MXN": '"MXN" $#,##0.00;[Red]-"MXN" $#,##0.00;"-"',
    "USD": '"USD" $#,##0.00;[Red]-"USD" $#,##0.00;"-"',
    "EUR": '"EUR" €#,##0.00;[Red]-"EUR" €#,##0.00;"-"',
}
MIXED_CATALOG_ORDER = (
    "tarkett", "offiho", "cr-global", "sonara", "sunon", "alma", "lumbro",
    "jome", "lauco", "idelika", "conceptos", "labenze", "requiez",
)
MIXED_CATALOG_LABELS = {
    "tarkett": "Tarkett",
    "offiho": "Offiho",
    "cr-global": "CR Global",
    "sonara": "Sonara",
    "sunon": "Sunon",
    "alma": "ALMA",
    "lumbro": "Lumbro",
    "jome": "JOME",
    "lauco": "Lauco",
    "idelika": "IDÉLIKA",
    "conceptos": "Conceptos",
    "labenze": "Labenze",
    "requiez": "Requiez",
}
MIXED_CATALOG_BASE_CURRENCIES = {
    "tarkett": "MXN",
    "offiho": "MXN",
    "cr-global": "MXN",
    "sonara": "MXN",
    "sunon": "USD",
    "alma": "USD",
    "lumbro": "MXN",
    "jome": "MXN",
    "lauco": "MXN",
    "idelika": "MXN",
    "conceptos": "MXN",
    "labenze": "MXN",
    "requiez": "MXN",
}
MIXED_RATE_FIELDS = {
    "catalog",
    "base_currency",
    "quote_currency",
    "exchange_rate",
    "rate_source",
    "rate_effective_date",
    "rate_retrieved_at",
}
MIXED_AUTO_RATE_FIELDS = MIXED_RATE_FIELDS - {"catalog"}
MIXED_MONEY_QUANTUM = Decimal("0.01")
MIXED_MOBILITI_MONEY_COLS = (
    10,
    13,
    14,
    *range(17, 26),
    *range(28, 31),
    32,
    33,
)
PERCENT_FORMAT = "0%"
MOBILITI_FIRST_SECTION_ROW = BASE_FIRST_SECTION_ROW
MOBILITI_SECTION_BLOCK_HEIGHT = BASE_PRODUCT_CAPACITY + 2
SECTION_CATS = [
    MOBILITI_FIRST_SECTION_ROW + index * MOBILITI_SECTION_BLOCK_HEIGHT
    for index in range(BASE_SECTION_COUNT)
]
SECTION_PROD_STARTS = [row + 1 for row in SECTION_CATS]
SECTION_SUBTOTAL_ROWS = [
    row + BASE_PRODUCT_CAPACITY + 1 for row in SECTION_CATS
]
MOBILITI_TOTAL_ROW = SECTION_SUBTOTAL_ROWS[-1] + 1
DEFAULT_EXCHANGE_RATE = 20.0
DEFAULT_DELIVERY_PLACE = "Guadalajara"
FRANKFURTER_USD_MXN_URL = "https://api.frankfurter.dev/v2/rate/USD/MXN"
EXCHANGE_RATE_CACHE_SECONDS = 60 * 60
EXCHANGE_RATE_CACHE_PATH = Path(os.environ.get("TEMP", "/tmp")) / "mobiliti_usd_mxn_rate.json"
DEFAULT_DISCOUNT_PERCENT = 40.0
DEFAULT_MOBILITI_REGION = "Centro"
DEFAULT_SUNON_LOOKUP_LIMIT = 40
DEFAULT_SUNON_LOOKUP_TIMEOUT_SECONDS = 6
DEFAULT_SUNON_LOOKUP_BUDGET_SECONDS = 180
MOBILITI_REGION_COL = 16
MOBILITI_PROVIDER_COL = 6
MOBILITI_PRODUCT_CATEGORY_COL = 5
MOBILITI_UNIT_PRICE_COL = 23
MOBILITI_MIN_UNIT_PRICE_COL = 24
MOBILITI_LIST_TOTAL_COL = 25
MOBILITI_MAX_DISCOUNT_COL = 26
MOBILITI_COVER_DISCOUNT_COL = 27
MOBILITI_DISCOUNT_AMOUNT_COL = 28
MOBILITI_FINAL_PRICE_COL = 29
MOBILITI_COMMERCIAL_TOTAL_COL = 30
MOBILITI_CLIENT_DISCOUNT_COL = 31
MOBILITI_CLIENT_PRICE_COL = 32
MOBILITI_PROJECT_TOTAL_COL = 33
MOBILITI_GP_COL = 34
MOBILITI_STATUS_COL = 35
MOBILITI_CLEAR_COLS = tuple(range(4, MOBILITI_STATUS_COL + 1))
MOBILITI_AUX_START_COL = 44
MOBILITI_AUX_END_COL = 49
MOBILITI_AUX_PRESERVE_MAX_ROW = 18
MOBILITI_PROVIDER_LIST_NAME = "Lista_Proveedores_Mobiliti"
MOBILITI_SUBTOTAL_FILL_RGB = "FF404040"
MOBILITI_SECTION_FILL_RGB = "FF3E2500"
MOBILITI_SECTION_TRAILING_FILL_RGB = "FF262626"
MOBILITI_SECTION_TEXT_RGB = "FFFFFFFF"
LUMBRO_PRICE_ROWS = {
    "MULT-LIDO-INT": 348,
    "LIDO.OP-INT": 380,
    "JUMP-1.5M": 396,
    "CAJA-FUS": 406,
}
LUMBRO_CATEGORY = "Multicontactos"
LUMBRO_PROVIDER = "Lumbro CH"
LUMBRO_LEGACY_PROVIDER = "Lumbro"
LUMBRO_PROVIDER_CODE = "P00720"
LUMBRO_PROVIDER_MARGIN = 0.85
LUMBRO_PROVIDER_EXTRA_MARGIN = 0
LUMBRO_PROVIDER_CATEGORY = "Importado"
LUMBRO_PROVIDER_MAX_DISCOUNT = 0.5
LUMBRO_ACCESSORY_IMAGE = Path(__file__).resolve().parent / "assets" / "lumbro_multicontacto_blanco.png"
LUMBRO_WORKSTATION_IMAGE = Path(__file__).resolve().parent / "assets" / "lumbro_workstation_multiusuario.png"
CUB_HIVE_REFERENCE_IMAGE = Path(__file__).resolve().parent / "assets" / "cub_hive_reference.png"
OFFICIAL_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 Oficial.xlsx"
)
OFFICIAL_TEMPLATE_CONTRACT_PATH = (
    Path(__file__).resolve().parents[1]
    / "worker"
    / "templates"
    / "formato-cotizacion-2026-oficial.contract.json"
)


def _contract_path_for_template(template_path: str | Path) -> Path:
    """Empareja activos registrados con su contrato sin aceptar rutas cliente."""

    resolved_template = Path(template_path).resolve(strict=True)
    template_hash = _official_file_hash(resolved_template)
    for profile in available_template_profiles():
        try:
            registered_template = profile.template_path.resolve(strict=True)
            registered_contract = profile.contract_path.resolve(strict=True)
        except FileNotFoundError:
            continue
        if resolved_template == registered_template:
            return registered_contract
        if load_template_contract(registered_contract).sha256 == template_hash:
            return registered_contract
    return OFFICIAL_TEMPLATE_CONTRACT_PATH.resolve(strict=True)


_ARGUMENT_OMITTED = object()
_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
@dataclass(frozen=True)
class MobilitiSectionLayout:
    section_row: int
    product_start: int
    capacity: int
    subtotal_row: int


@dataclass(frozen=True)
class LumbroPriceRef:
    row: int
    price_mxn: float


@dataclass(frozen=True)
class _OfficialPresentationLine:
    item_key: str
    section_id: str
    section_title: str
    item: QuoteItem | None
    name: str
    description: str
    dimensions: str
    m3: Decimal
    quantity: Decimal
    category: str
    provider: str
    region: str
    original_currency: str
    original_cost: Decimal | None
    frozen_rate: Decimal | None
    converted_cost: Decimal | None
    origin: str
    source_row: int | None
    upstream_row_hash: str
    parent_item_key: str | None = None
    image_content: bytes | None = None
    image_content_type: str | None = None


def _sheet_name(name: str) -> str:
    return "'{}'".format(name.replace("'", "''")) if any(ch in name for ch in " :!'-*[]?/\\") else name


def _formula(sheet: str, cell: str) -> str:
    return f"={_sheet_name(sheet)}!{cell}"


def _num(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return default


def _positive_num(value: Any) -> float | None:
    number = _num(value, 0)
    return number if number > 0 else None


def _extract_usd_mxn_rate(payload: bytes | str) -> float | None:
    try:
        data = json.loads(payload.decode("utf-8") if isinstance(payload, bytes) else payload)
    except (TypeError, ValueError, UnicodeDecodeError):
        return None
    if not isinstance(data, dict):
        return None
    direct_rate = _positive_num(data.get("rate"))
    if direct_rate:
        return direct_rate
    rates = data.get("rates")
    return _positive_num(rates.get("MXN")) if isinstance(rates, dict) else None


def _fetch_latest_usd_mxn_row() -> dict[str, str] | None:
    """Obtiene la referencia diaria USD/MXN más reciente disponible."""

    reference_day = date.today()
    request = Request(
        f"{FRANKFURTER_USD_MXN_URL}?date={reference_day.isoformat()}&providers=BANXICO",
        headers={"User-Agent": "mobiliti-quote-engine/1.0"},
    )
    try:
        with urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (OSError, TypeError, ValueError, UnicodeDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    rate = _positive_num(payload.get("rate"))
    effective_date = str(payload.get("date") or "").strip()
    try:
        effective_day = date.fromisoformat(effective_date)
    except ValueError:
        return None
    if effective_day > reference_day:
        return None
    if (
        not rate
        or str(payload.get("base") or "").upper() != "USD"
        or str(payload.get("quote") or "").upper() != "MXN"
    ):
        return None
    return {
        "currency": "USD",
        "effective_date": effective_date,
        "mxn_per_unit": f"{rate:.6f}",
        "retrieved_at": datetime.now(timezone.utc).isoformat(),
    }


def _read_cached_usd_mxn_rate(now: float | None = None) -> float | None:
    try:
        data = json.loads(EXCHANGE_RATE_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    timestamp = _num(data.get("timestamp"), 0) if isinstance(data, dict) else 0
    rate = _positive_num(data.get("rate")) if isinstance(data, dict) else None
    if not rate:
        return None
    if (now or time.time()) - timestamp > EXCHANGE_RATE_CACHE_SECONDS:
        return None
    return rate


def _write_cached_usd_mxn_rate(rate: float) -> None:
    try:
        EXCHANGE_RATE_CACHE_PATH.write_text(
            json.dumps({"timestamp": time.time(), "rate": rate}),
            encoding="utf-8",
        )
    except OSError:
        pass


def _fetch_usd_mxn_exchange_rate() -> float | None:
    row = _fetch_latest_usd_mxn_row()
    return _positive_num(row.get("mxn_per_unit")) if row is not None else None


def _exchange_rate(metadata: dict[str, Any]) -> float:
    explicit_rate = _positive_num(metadata.get("tipo_cambio", metadata.get("exchange_rate")))
    if explicit_rate:
        return explicit_rate
    cached_rate = _read_cached_usd_mxn_rate()
    if cached_rate:
        return cached_rate
    fetched_rate = _fetch_usd_mxn_exchange_rate()
    if fetched_rate:
        _write_cached_usd_mxn_rate(fetched_rate)
        return fetched_rate
    return DEFAULT_EXCHANGE_RATE


def _discount_rate(metadata: dict[str, Any]) -> float:
    raw = metadata.get("descuento", metadata.get("discount_percent", DEFAULT_DISCOUNT_PERCENT))
    value = _num(raw, DEFAULT_DISCOUNT_PERCENT)
    if 0 <= value <= 1:
        return value
    return max(0.0, min(value, 100.0)) / 100.0


def _excel_decimal(value: float) -> str:
    text = f"{float(value):.6f}".rstrip("0").rstrip(".")
    return text or "0"


def _normalized_text(value: Any) -> str:
    text = "" if value is None else str(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _detect_user_count(item: QuoteItem) -> int | None:
    text = _normalized_text(
        f"{item.nombre or ''} {item.descripcion or ''} {item.dimension or ''} {item.categoria or ''}"
    )
    patterns = [
        r"(\d+)\s*(?:pax|px)\b",
        r"(\d+)\s*(?:usuarios?|personas?|users?)\b",
        r"(?:pax|px|capacidad|usuarios?|personas?|users?)\s*(?:de|para)?\s*(\d+)\b",
    ]
    matches: list[int] = []
    for pattern in patterns:
        matches.extend(int(match) for match in re.findall(pattern, text))
    return max(matches) if matches else None


def _item_quantity(item: QuoteItem) -> int:
    return max(1, int(math.ceil(_num(item.cantidad, 1))))


def _lumbro_accessories_for_item(item: QuoteItem, category: str) -> list[tuple[str, int]]:
    quantity = _item_quantity(item)
    users_per_item = _detect_user_count(item)

    if category == "Escritorios-WorkStation":
        if users_per_item:
            total_users = users_per_item * quantity
            fuse_count = math.ceil(total_users / 8) * 2
            return [
                ("LIDO.OP-INT", total_users),
                ("JUMP-1.5M", total_users),
                ("CAJA-FUS", fuse_count),
            ]
        return [("MULT-LIDO-INT", quantity)]

    if category == "Mesas de Juntas":
        total_users = (users_per_item or 4) * quantity
        multicontacts = max(1, math.ceil(total_users / 4))
        return [
            ("MULT-LIDO-INT", multicontacts),
            ("JUMP-1.5M", multicontacts + 1),
        ]

    return []


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)


def _copy_external_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.protection = copy(src.protection)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int = 10) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def _find_mobiliti_total_row(ws) -> int | None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 6).value or "").strip().upper() == "TOTAL PIEZAS":
            return row
    return None


def _find_provider_discount_start(ws) -> int | None:
    for row in range(1, ws.max_row + 1):
        for col in (36, 35):
            if str(ws.cell(row, col).value or "").strip() == "Sunon Inc":
                return row
    return None


def _official_mobiliti_provider(value: Any) -> str:
    return official_provider_name(value)


def _set_defined_name_range(wb: Workbook, name: str, attr_text: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=attr_text))


def _ensure_lumbro_ch_provider_contract(ws) -> None:
    wb = ws.parent
    if "Proveedores" not in wb.sheetnames:
        return

    provider_ws = wb["Proveedores"]
    provider_rows = [
        row
        for row in range(2, provider_ws.max_row + 1)
        if str(provider_ws.cell(row, 1).value or "").strip()
        and str(provider_ws.cell(row, 5).value or "").strip().casefold()
        in {"importado", "nacional"}
    ]
    if not provider_rows:
        return

    target_row = next(
        (
            row
            for row in provider_rows
            if _official_mobiliti_provider(provider_ws.cell(row, 1).value)
            == LUMBRO_PROVIDER
            and str(provider_ws.cell(row, 1).value or "").strip().casefold()
            == LUMBRO_PROVIDER.casefold()
        ),
        provider_rows[-1] + 1,
    )
    if target_row not in provider_rows:
        style_row = next(
            (
                row
                for row in provider_rows
                if str(provider_ws.cell(row, 5).value or "").strip().casefold()
                == LUMBRO_PROVIDER_CATEGORY.casefold()
            ),
            provider_rows[0],
        )
        provider_ws.row_dimensions[target_row].height = provider_ws.row_dimensions[style_row].height
        for col in range(1, 6):
            _copy_external_cell_style(
                provider_ws.cell(style_row, col),
                provider_ws.cell(target_row, col),
            )

    values = (
        LUMBRO_PROVIDER,
        LUMBRO_PROVIDER_CODE,
        LUMBRO_PROVIDER_MARGIN,
        LUMBRO_PROVIDER_EXTRA_MARGIN,
        LUMBRO_PROVIDER_CATEGORY,
    )
    for col, value in enumerate(values, start=1):
        provider_ws.cell(target_row, col).value = value

    last_provider_row = max([*provider_rows, target_row])
    _set_defined_name_range(wb, "Proveedores", f"Proveedores!$A$2:$D${last_provider_row}")
    _set_defined_name_range(wb, "ProveedoreS_TC", f"Proveedores!$A$2:$E${last_provider_row}")
    _set_defined_name_range(
        wb,
        "Tabla_Proveedores_1",
        f"Proveedores!$A$2:$B${last_provider_row}",
    )


def _ensure_lumbro_ch_discount_contract(ws) -> None:
    discount_start = _find_provider_discount_start(ws)
    if discount_start is None:
        return

    name_col = 36 if str(ws.cell(discount_start, 36).value or "").strip() else 35
    value_col = name_col + 1
    discount_end = discount_start
    while (
        discount_end + 1 <= ws.max_row
        and str(ws.cell(discount_end + 1, name_col).value or "").strip()
    ):
        discount_end += 1

    existing_row = next(
        (
            row
            for row in range(discount_start, discount_end + 1)
            if str(ws.cell(row, name_col).value or "").strip().casefold()
            == LUMBRO_PROVIDER.casefold()
        ),
        None,
    )
    if existing_row is not None:
        ws.cell(existing_row, value_col).value = LUMBRO_PROVIDER_MAX_DISCOUNT
        return

    target_row = next(
        (
            row
            for row in range(discount_start, discount_end + 1)
            if str(ws.cell(row, name_col).value or "").strip() == "Tarkett MX"
        ),
        discount_end + 1,
    )
    for row in range(discount_end, target_row - 1, -1):
        for col in (name_col, value_col):
            source = ws.cell(row, col)
            target = ws.cell(row + 1, col)
            target.value = source.value
            _copy_external_cell_style(source, target)
    ws.cell(target_row, name_col).value = LUMBRO_PROVIDER
    ws.cell(target_row, value_col).value = LUMBRO_PROVIDER_MAX_DISCOUNT


def _snapshot_mobiliti_value(value: Any) -> Any:
    if isinstance(value, ArrayFormula):
        return value.text
    return value


def _snapshot_mobiliti_row(ws, row: int, max_col: int) -> dict[str, Any]:
    return {
        "height": ws.row_dimensions[row].height,
        "cells": [
            {
                "value": _snapshot_mobiliti_value(ws.cell(row, col).value),
                "style": copy(ws.cell(row, col)._style),
                "number_format": ws.cell(row, col).number_format,
                "alignment": copy(ws.cell(row, col).alignment),
            }
            for col in range(1, max_col + 1)
        ],
        "merges": [
            str(merged)
            for merged in ws.merged_cells.ranges
            if merged.min_row == row and merged.max_row == row
        ],
    }


def _snapshot_mobiliti_auxiliary_area(ws) -> dict[str, Any]:
    cells: dict[tuple[int, int], dict[str, Any]] = {}
    for row in range(1, MOBILITI_AUX_PRESERVE_MAX_ROW + 1):
        for col in range(MOBILITI_AUX_START_COL, MOBILITI_AUX_END_COL + 1):
            cell = ws.cell(row, col)
            cells[(row, col)] = {
                "value": _snapshot_mobiliti_value(cell.value),
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
            }
    merges = [
        str(merged)
        for merged in ws.merged_cells.ranges
        if (
            merged.min_row <= MOBILITI_AUX_PRESERVE_MAX_ROW
            and merged.max_row >= 1
            and merged.min_col <= MOBILITI_AUX_END_COL
            and merged.max_col >= MOBILITI_AUX_START_COL
        )
    ]
    return {"cells": cells, "merges": merges}


def _restore_mobiliti_auxiliary_area(
    ws,
    snapshot: dict[str, Any],
    total_row: int,
) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col <= MOBILITI_AUX_END_COL and merged.max_col >= MOBILITI_AUX_START_COL:
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(merged)
                except KeyError:
                    pass

    for row in range(MOBILITI_AUX_PRESERVE_MAX_ROW + 1, ws.max_row + 1):
        for col in range(MOBILITI_AUX_START_COL, MOBILITI_AUX_END_COL + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for (row, col), data in snapshot["cells"].items():
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = data["value"]
        cell._style = copy(data["style"])
        cell.number_format = data["number_format"]
        cell.alignment = copy(data["alignment"])

    ws["AR14"] = "TABLA DE ESQUEMA COMISION INTERMEDIARIO"
    ws["AS15"] = "Precio de Venta Base"
    ws["AS16"] = "Precio de Venta Final"
    ws["AS17"] = "Monto Comisi\u00f3n"
    ws["AS18"] = "% Resultante"
    ws["AT15"] = f"=AD{total_row}"
    ws["AT16"] = f"=AG{total_row}"
    ws["AT17"] = "=AT16-AT15"
    ws["AT18"] = "=(AT16/AT15)-1"

    for merge in snapshot["merges"]:
        try:
            ws.merge_cells(merge)
        except ValueError:
            pass


def _translate_formula_value(value: Any, origin: str, target: str) -> Any:
    if isinstance(value, str) and value.startswith("="):
        try:
            return Translator(value, origin=origin).translate_formula(target)
        except Exception:
            return value
    return value


def _copy_mobiliti_row_from_snapshot(
    ws,
    snapshot: dict[str, Any],
    target_row: int,
    *,
    source_row: int,
    max_col: int,
    translate_formulas: bool = True,
) -> None:
    _unmerge_row(ws, target_row)
    for key, cell in list(ws._cells.items()):
        if key[0] == target_row and isinstance(cell, MergedCell):
            del ws._cells[key]
    ws.row_dimensions[target_row].height = snapshot["height"]
    for col, data in enumerate(snapshot["cells"], start=1):
        cell = ws.cell(target_row, col)
        cell.value = None
        cell._style = copy(data["style"])
        cell.number_format = data["number_format"]
        cell.alignment = copy(data["alignment"])
        value = data["value"]
        if translate_formulas:
            value = _translate_formula_value(
                value,
                f"{get_column_letter(col)}{source_row}",
                f"{get_column_letter(col)}{target_row}",
            )
        cell.value = value

    row_offset = target_row - source_row
    for merge in snapshot["merges"]:
        shifted = CellRange(merge)
        shifted.shift(row_shift=row_offset, col_shift=0)
        try:
            ws.merge_cells(str(shifted))
        except ValueError:
            pass


def _mobiliti_landed_cost_formula(row: int, total_row: int, discount_start: int | None) -> str:
    threshold_base = discount_start or 577
    small_m3_row = threshold_base + 1
    medium_m3_row = threshold_base + 2
    large_m3_row = threshold_base + 3
    excluded = [
        "Sunon Inc",
        "Alma - Exterior",
        "Yabo - Hoteler\u00eda",
        "OSJ - Medical",
        "Seezo - Home",
        "A&D -Home",
        "Encore Alfombras Hoteleria Asia",
        "Zhong Xian - Leds",
        "Tarkett Europa",
        "2tec2",
        "Balsan Europa",
        "Armstrong Pisos USA",
        "Tarkett USA",
        "Tarkett Brasil",
        LUMBRO_PROVIDER,
    ]
    provider_checks = ",".join(f'F{row}<>"{provider}"' for provider in excluded)
    return (
        f'=IF(F{row}="Offiho", J{row}*0.55, '
        f'IF(AND({provider_checks}),J{row},'
        f'IFERROR(IF(K{total_row}<=$AP${small_m3_row},J{row}*2.14,'
        f'IF(K{total_row}<=$AO${medium_m3_row},J{row}*1.8,'
        f'IF(K{total_row}>$AO${large_m3_row},J{row}*1.5))),J{row}))*1.1)'
    )


def _mobiliti_provider_discount_formula(row: int, discount_start: int | None) -> str:
    if not discount_start:
        return "=0.5"
    discount_end = discount_start + 31
    return f'=IFERROR(VLOOKUP(F{row},$AJ${discount_start}:$AK${discount_end},2,FALSE),0.5)'


def _mobiliti_blank_product_guard(row: int) -> str:
    return f"COUNTA($D{row},$E{row},$F{row},$H{row},$J{row},$K{row})=0"


def _blank_safe_mobiliti_formula(row: int, formula: str) -> str:
    return f'=IF({_mobiliti_blank_product_guard(row)},"",{formula[1:]})'


def _set_mobiliti_row_fill(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    fill_rgb: str,
    *,
    font_rgb: str | None = None,
) -> None:
    row_merges = [
        str(merged)
        for merged in ws.merged_cells.ranges
        if merged.min_row == row and merged.max_row == row
    ]
    for merge in row_merges:
        try:
            ws.unmerge_cells(merge)
        except KeyError:
            pass

    fill = PatternFill(fill_type="solid", fgColor=fill_rgb)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        cell.fill = copy(fill)
        if font_rgb:
            font = copy(cell.font)
            font.color = font_rgb
            cell.font = font

    for merge in row_merges:
        try:
            ws.merge_cells(merge)
        except ValueError:
            pass


def _apply_mobiliti_subtotal_row_visual_style(ws, row: int) -> None:
    _set_mobiliti_row_fill(
        ws,
        row,
        1,
        MOBILITI_STATUS_COL - 1,
        MOBILITI_SUBTOTAL_FILL_RGB,
        font_rgb=MOBILITI_SECTION_TEXT_RGB,
    )


def _apply_mobiliti_section_row_visual_style(ws, row: int) -> None:
    _set_mobiliti_row_fill(
        ws,
        row,
        1,
        10,
        MOBILITI_SECTION_FILL_RGB,
        font_rgb=MOBILITI_SECTION_TEXT_RGB,
    )
    _set_mobiliti_row_fill(
        ws,
        row,
        11,
        MOBILITI_STATUS_COL - 1,
        MOBILITI_SECTION_TRAILING_FILL_RGB,
        font_rgb=MOBILITI_SECTION_TEXT_RGB,
    )


def _conditional_range_piece(cell_range: CellRange, start_row: int, end_row: int) -> str:
    start_col = get_column_letter(cell_range.min_col)
    end_col = get_column_letter(cell_range.max_col)
    return f"{start_col}{start_row}:{end_col}{end_row}"


def _exclude_mobiliti_separator_rows_from_conditional_formatting(
    ws,
    layouts: list[MobilitiSectionLayout],
) -> None:
    cf_rules = getattr(ws.conditional_formatting, "_cf_rules", None)
    if not cf_rules:
        return

    excluded_rows = {layout.section_row for layout in layouts}
    excluded_rows.update(layout.subtotal_row for layout in layouts)
    updated_rules = OrderedDict()

    for conditional_formatting, rules in cf_rules.items():
        range_pieces: list[str] = []
        for cell_range in conditional_formatting.sqref.ranges:
            rows_to_exclude = sorted(
                row
                for row in excluded_rows
                if cell_range.min_row <= row <= cell_range.max_row
            )
            if not rows_to_exclude:
                range_pieces.append(str(cell_range))
                continue

            start_row = cell_range.min_row
            for excluded_row in rows_to_exclude:
                if start_row <= excluded_row - 1:
                    range_pieces.append(
                        _conditional_range_piece(cell_range, start_row, excluded_row - 1)
                    )
                start_row = excluded_row + 1
            if start_row <= cell_range.max_row:
                range_pieces.append(_conditional_range_piece(cell_range, start_row, cell_range.max_row))

        if range_pieces:
            updated_rules[ConditionalFormatting(sqref=" ".join(range_pieces))] = rules

    cf_rules.clear()
    cf_rules.update(updated_rules)


def _set_mobiliti_total_formulas(
    ws,
    row: int,
    subtotal_rows: list[int] | None = None,
) -> None:
    subtotal_rows = subtotal_rows or SECTION_SUBTOTAL_ROWS
    ws.cell(row, 8).value = "=" + "+".join(f"H{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, 11).value = "=" + "+".join(f"L{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, 13).value = "=" + "+".join(f"N{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, MOBILITI_LIST_TOTAL_COL).value = "=" + "+".join(f"Y{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, MOBILITI_COMMERCIAL_TOTAL_COL).value = "=" + "+".join(f"AD{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, MOBILITI_PROJECT_TOTAL_COL).value = "=" + "+".join(f"AG{subtotal}" for subtotal in subtotal_rows)
    ws.cell(row, MOBILITI_GP_COL).value = f"=AVERAGE({','.join(f'AH{subtotal}' for subtotal in subtotal_rows)})"


def _cell_has_conditional_format(ws, coordinate: str) -> bool:
    for cf in getattr(ws.conditional_formatting, "_cf_rules", {}):
        for cell_range in str(cf.sqref).split():
            if coordinate in CellRange(cell_range):
                return True
    return False


def _mobiliti_status_conditional_rules(ws) -> list[Any]:
    for cf, rules in getattr(ws.conditional_formatting, "_cf_rules", {}).items():
        for cell_range in str(cf.sqref).split():
            if "AI49" in CellRange(cell_range) or "AH49" in CellRange(cell_range):
                return [deepcopy(rule) for rule in rules]
    return []


def _remove_mobiliti_status_conditional_formatting(ws) -> None:
    cf_rules = getattr(ws.conditional_formatting, "_cf_rules", None)
    if not cf_rules:
        return

    old_status_col = MOBILITI_STATUS_COL - 1
    status_cols = {old_status_col, MOBILITI_STATUS_COL}
    updated_rules = OrderedDict()
    for conditional_formatting, rules in cf_rules.items():
        kept_ranges = [
            str(cell_range)
            for cell_range in conditional_formatting.sqref.ranges
            if not any(cell_range.min_col <= col <= cell_range.max_col for col in status_cols)
        ]
        if kept_ranges:
            updated_rules[ConditionalFormatting(sqref=" ".join(kept_ranges))] = rules

    cf_rules.clear()
    cf_rules.update(updated_rules)


def _status_rule_for_range(rule: Any, status_letter: str, start_row: int) -> Any:
    copied = deepcopy(rule)
    formulas = getattr(copied, "formula", None)
    if formulas:
        copied.formula = [
            re.sub(r"\b(?:AH|AI)\$?\d+\b", f"{status_letter}{start_row}", formula)
            for formula in formulas
        ]
    return copied


def _apply_mobiliti_status_conditional_formatting(ws) -> None:
    rules = _mobiliti_status_conditional_rules(ws)
    if not rules:
        return

    _remove_mobiliti_status_conditional_formatting(ws)
    status_letter = get_column_letter(MOBILITI_STATUS_COL)
    for start_row, capacity in _mobiliti_product_ranges(ws):
        end_row = start_row + capacity - 1
        target_range = f"{status_letter}{start_row}:{status_letter}{end_row}"
        for rule in rules:
            ws.conditional_formatting.add(target_range, _status_rule_for_range(rule, status_letter, start_row))


def _clear_mobiliti_row_values(ws, row: int) -> None:
    for col in MOBILITI_CLEAR_COLS:
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None
    provider_cell = ws.cell(row, MOBILITI_PROVIDER_COL)
    if not isinstance(provider_cell, MergedCell):
        product_cell = ws.cell(row, MOBILITI_PRODUCT_CATEGORY_COL)
        provider_cell.fill = copy(product_cell.fill)
        provider_cell.border = copy(product_cell.border)


def _mobiliti_product_ranges(ws) -> list[tuple[int, int]]:
    ranges = getattr(ws, "_mobiliti_product_ranges", None)
    if ranges:
        return list(ranges)
    return [
        (start_row, BASE_PRODUCT_CAPACITY)
        for start_row in SECTION_PROD_STARTS
    ]


def _set_mobiliti_auxiliary_total_references(ws, total_row: int) -> None:
    ws["P9"] = f"=P8/H{total_row}"


def _write_mobiliti_section_title(
    ws,
    layout: MobilitiSectionLayout,
    section_number: int,
    title: str,
) -> None:
    anchor_col = _merged_anchor_column(ws, layout.section_row, 4)
    ws.cell(layout.section_row, anchor_col).value = f"Secci\u00f3n {section_number} - {title}"


def _unmerge_row(ws, row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row:
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(merged)
                except KeyError:
                    pass


def _clear_row(ws, row: int, max_col: int = 10) -> None:
    _unmerge_row(ws, row)
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def _merged_anchor_column(ws, row: int, default_col: int) -> int:
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row == row
            and merged.max_row == row
            and merged.min_col <= default_col <= merged.max_col
        ):
            return merged.min_col
    if isinstance(ws.cell(row, default_col), MergedCell):
        for col in range(default_col, 0, -1):
            if not isinstance(ws.cell(row, col), MergedCell):
                return col
    return default_col


def _find_terms_row(ws) -> int:
    for row in range(16, min(ws.max_row, 140) + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and "CONDICIONES" in value.upper():
            return row
    return 32


def _find_totals_row(ws, before_row: int) -> int:
    for row in range(16, max(16, before_row - 4) + 1):
        first_label = str(ws.cell(row, 4).value or "").strip().upper()
        second_label = str(ws.cell(row + 1, 4).value or "").strip().upper()
        total_label = str(ws.cell(row + 4, 4).value or "").strip().upper()
        if first_label == "SUBTOTAL:" and "FLETE" in second_label and total_label == "TOTAL:":
            return row
    return 21


def _snapshot_rows(ws, start_row: int, end_row: int, max_col: int = 10) -> dict[str, Any]:
    rows = []
    for row in range(start_row, end_row + 1):
        cells = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cells.append(
                {
                    "value": cell.value,
                    "style": copy(cell._style),
                    "number_format": cell.number_format,
                    "alignment": copy(cell.alignment),
                }
            )
        rows.append(cells)
    merges = [
        str(merged)
        for merged in ws.merged_cells.ranges
        if start_row <= merged.min_row and merged.max_row <= end_row
    ]
    heights = {row: ws.row_dimensions[row].height for row in range(start_row, end_row + 1)}
    return {"rows": rows, "merges": merges, "heights": heights}


def _restore_rows(ws, start_row: int, snapshot: dict[str, Any]) -> int:
    row = start_row
    source_start = min(snapshot["heights"].keys()) if snapshot["heights"] else start_row
    row_offset = start_row - source_start
    for cells in snapshot["rows"]:
        _clear_row(ws, row)
        for col, data in enumerate(cells, start=1):
            cell = ws.cell(row, col)
            cell.value = data["value"]
            cell._style = copy(data["style"])
            cell.number_format = data["number_format"]
            cell.alignment = copy(data["alignment"])
        source_row = row - row_offset
        ws.row_dimensions[row].height = snapshot["heights"].get(source_row)
        row += 1
    for merged in snapshot["merges"]:
        shifted = CellRange(merged)
        shifted.shift(row_shift=row_offset, col_shift=0)
        try:
            ws.merge_cells(str(shifted))
        except ValueError:
            pass
    return row - 1


def _set_cotizacion_image_column_px(
    ws: Worksheet, pixel_width: float = 1100, column: str = "B"
) -> None:
    ws.column_dimensions[column].width = max(1.0, (float(pixel_width) - 5) / 7)


def _load_lumbro_prices(template_path: str | Path | None) -> dict[str, LumbroPriceRef]:
    if not template_path:
        return {}
    path = Path(template_path)
    if not path.exists():
        return {}

    wb = load_workbook(path, data_only=True, keep_links=False)
    try:
        if "SPEC-GUIDE-LUMBRO" not in wb.sheetnames:
            return {}
        ws = wb["SPEC-GUIDE-LUMBRO"]
        prices: dict[str, LumbroPriceRef] = {}
        for code, row in LUMBRO_PRICE_ROWS.items():
            prices[code] = LumbroPriceRef(row=row, price_mxn=_num(ws.cell(row, 5).value, 0))
        return prices
    finally:
        wb.close()


def _first_product_row(items: list[QuoteItem]) -> int:
    return next((item.row for item in items if item.tipo == "producto"), 9)


def _uses_catalog_list_prices(metadata: dict[str, Any] | None) -> bool:
    return str((metadata or {}).get("catalog_price_mode") or "").strip() == "list_price_net"


def _uses_mixed_catalog_prices(metadata: dict[str, Any] | None) -> bool:
    return str((metadata or {}).get("catalog_price_mode") or "").strip() == "mixed_catalog_converted"


def _uses_converted_catalog_prices(metadata: dict[str, Any] | None) -> bool:
    return _uses_catalog_list_prices(metadata) or _uses_mixed_catalog_prices(metadata)


def _item_discount_rate(item: QuoteItem, metadata: dict[str, Any]) -> float:
    if _uses_mixed_catalog_prices(metadata):
        return float(_mixed_item_discount_fraction(item))
    return _discount_rate(metadata)


def _mixed_decimal(value: Any, message: str, *, positive: bool = False) -> Decimal:
    if isinstance(value, bool):
        raise ValueError(message)
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(message) from exc
    if not number.is_finite() or number < 0 or (positive and number <= 0):
        raise ValueError(message)
    return number


def _mixed_item_discount_fraction(item: QuoteItem) -> Decimal:
    mode = str(item.modo_precio or "").strip().lower()
    provider = str(item.proveedor or "").strip()
    if mode not in {"list", "net", "imported", "pending"}:
        raise ValueError("Modo de precio mixto invalido")
    value = _mixed_decimal(item.descuento, "Descuento mixto por linea invalido")
    if value > Decimal("100") or max(-value.as_tuple().exponent, 0) > 6:
        raise ValueError("Descuento mixto por linea invalido")
    if mode in {"net", "pending"} and value != 0:
        raise ValueError("Precio neto mixto no admite descuento")
    if mode == "list" and provider not in {"Tarkett", "Offiho"}:
        raise ValueError("Precio de lista mixto solo admite Tarkett u Offiho")
    return value / Decimal("100")


def _mixed_metadata_discount_fraction(metadata: dict[str, Any]) -> Decimal:
    raw = metadata.get(
        "descuento",
        metadata.get("discount_percent", DEFAULT_DISCOUNT_PERCENT),
    )
    value = _mixed_decimal(raw, "Descuento mixto general invalido")
    if value > Decimal("100"):
        raise ValueError("Descuento mixto general invalido")
    return value if value <= Decimal("1") else value / Decimal("100")


def _mixed_decimal_literal(value: Decimal) -> str:
    text = format(value, "f").rstrip("0").rstrip(".")
    return text or "0"


def _item_discount_literal(item: QuoteItem, metadata: dict[str, Any]) -> str:
    if _uses_mixed_catalog_prices(metadata):
        return _mixed_decimal_literal(_mixed_metadata_discount_fraction(metadata))
    return _excel_decimal(_discount_rate(metadata))


def _item_auto_electrification(item: QuoteItem, metadata: dict[str, Any]) -> bool:
    if _uses_mixed_catalog_prices(metadata):
        if not isinstance(item.electrificacion_automatica, bool):
            raise ValueError("Politica de electrificacion mixta invalida")
        if item.electrificacion_automatica and str(item.proveedor or "").strip() not in {
            "Tarkett",
            "Offiho",
        }:
            raise ValueError("Electrificacion automatica mixta solo admite Tarkett u Offiho")
        return item.electrificacion_automatica
    return not _uses_catalog_list_prices(metadata)


def _mixed_auto_electrification_rate(metadata: dict[str, Any]) -> float:
    snapshot = metadata.get("auto_electrification_rate")
    quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
    if not isinstance(snapshot, dict) or set(snapshot) != MIXED_AUTO_RATE_FIELDS:
        raise ValueError("Tasa de electrificacion mixta incompleta")
    if snapshot.get("base_currency") != "MXN" or snapshot.get("quote_currency") != quote_currency:
        raise ValueError("Par de electrificacion mixta invalido")
    rate = _positive_num(snapshot.get("exchange_rate"))
    if rate is None or not math.isfinite(rate):
        raise ValueError("Tasa de electrificacion mixta invalida")
    return rate


def _money_format(metadata: dict[str, Any]) -> str:
    if not _uses_mixed_catalog_prices(metadata):
        return MONEY_FORMAT
    quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
    try:
        return MIXED_MONEY_FORMATS[quote_currency]
    except KeyError as exc:
        raise ValueError("Moneda mixta incompleta") from exc


def _mixed_rate_summary(metadata: dict[str, Any]) -> list[dict[str, Any]]:
    quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
    if quote_currency not in MIXED_MONEY_FORMATS:
        raise ValueError("Moneda mixta incompleta")
    raw_summary = metadata.get("rate_summary")
    if not isinstance(raw_summary, list):
        raise ValueError("Resumen de tasas mixtas invalido")
    summary: list[dict[str, Any]] = []
    seen: list[str] = []
    for raw in raw_summary:
        if not isinstance(raw, dict) or set(raw) != MIXED_RATE_FIELDS:
            raise ValueError("Resumen de tasas mixtas invalido")
        catalog = raw.get("catalog")
        if catalog not in MIXED_CATALOG_ORDER or catalog in seen:
            raise ValueError("Resumen de tasas mixtas invalido")
        seen.append(catalog)
        if seen != sorted(seen, key=MIXED_CATALOG_ORDER.index):
            raise ValueError("Resumen de tasas mixtas invalido")
        base_currency = str(raw.get("base_currency") or "").strip().upper()
        if (
            base_currency != MIXED_CATALOG_BASE_CURRENCIES[catalog]
            or raw.get("quote_currency") != quote_currency
        ):
            raise ValueError("Resumen de tasas mixtas invalido")
        rate_text = raw.get("exchange_rate")
        if not isinstance(rate_text, str) or not re.fullmatch(r"\d+\.\d{6}", rate_text):
            raise ValueError("Resumen de tasas mixtas invalido")
        rate = _positive_num(rate_text)
        source = raw.get("rate_source")
        effective_date = raw.get("rate_effective_date")
        retrieved_at = raw.get("rate_retrieved_at")
        if (
            rate is None
            or not math.isfinite(rate)
            or source not in {"identity", "saas_exchange_rates"}
            or not isinstance(effective_date, str)
            or re.fullmatch(r"\d{4}-\d{2}-\d{2}", effective_date) is None
            or not isinstance(retrieved_at, str)
            or len(retrieved_at) > 80
        ):
            raise ValueError("Resumen de tasas mixtas invalido")
        try:
            date.fromisoformat(effective_date)
            retrieved_datetime = (
                datetime.fromisoformat(retrieved_at.replace("Z", "+00:00"))
                if retrieved_at
                else None
            )
        except ValueError as exc:
            raise ValueError("Resumen de tasas mixtas invalido") from exc
        if source == "identity":
            if base_currency != quote_currency or rate_text != "1.000000" or retrieved_at != "":
                raise ValueError("Resumen de tasas mixtas invalido")
        elif not retrieved_at or retrieved_datetime is None or retrieved_datetime.tzinfo is None:
            raise ValueError("Resumen de tasas mixtas invalido")
        summary.append(raw)
    return summary


def _validate_mixed_catalog_metadata(items: list[QuoteItem], metadata: dict[str, Any]) -> None:
    if not _uses_mixed_catalog_prices(metadata):
        return
    product_items = [item for item in items if item.tipo == "producto"]
    if not product_items:
        raise ValueError("Carrito mixto sin productos")
    summary = _mixed_rate_summary(metadata)
    summaries_by_provider = {
        MIXED_CATALOG_LABELS[entry["catalog"]]: entry for entry in summary
    }
    auto_items: list[QuoteItem] = []
    seen_providers: set[str] = set()
    for item in product_items:
        _item_discount_rate(item, metadata)
        if _item_auto_electrification(item, metadata):
            auto_items.append(item)
        mode = str(item.modo_precio or "").strip().lower()
        provider = str(item.proveedor or "").strip()
        snapshot = None
        if mode != "imported":
            seen_providers.add(provider)
            snapshot = summaries_by_provider.get(provider)
            if snapshot is None:
                raise ValueError("Proveedor mixto sin tasa congelada")
        original_currency = str(item.moneda_original or "").strip().upper()
        if mode == "pending":
            if (
                provider in {"Tarkett", "Offiho"}
                or original_currency != snapshot["base_currency"]
                or item.precio_original not in {None, ""}
                or item.tipo_cambio_congelado not in {None, ""}
                or str(item.precio or "").strip() != "Por confirmar"
                or not str(item.referencia_fuente or "").strip()
            ):
                raise ValueError("Auditoria de precio pendiente invalida")
            continue
        frozen_rate = _mixed_decimal(
            item.tipo_cambio_congelado,
            "Auditoria de precio mixto invalida",
            positive=True,
        )
        original_price = _mixed_decimal(
            item.precio_original,
            "Auditoria de precio mixto invalida",
        )
        converted_price = _mixed_decimal(
            item.precio,
            "Auditoria de precio mixto invalida",
        )
        try:
            expected_converted_price = (original_price * frozen_rate).quantize(
                MIXED_MONEY_QUANTUM,
                rounding=ROUND_HALF_UP,
            )
        except InvalidOperation as exc:
            raise ValueError("Auditoria de precio mixto invalida") from exc
        if (
            (
                snapshot is not None
                and (
                    original_currency != snapshot["base_currency"]
                    or frozen_rate != Decimal(snapshot["exchange_rate"])
                )
            )
            or (mode == "imported" and original_currency not in {"MXN", "USD", "EUR"})
            or not str(item.referencia_fuente or "").strip()
        ):
            raise ValueError("Auditoria de precio mixto invalida")
        if converted_price != expected_converted_price:
            raise ValueError("Precio convertido mixto inconsistente")
    if seen_providers != set(summaries_by_provider):
        raise ValueError("Resumen de tasas mixtas inconsistente")
    if auto_items:
        _mixed_auto_electrification_rate(metadata)
        snapshot = metadata["auto_electrification_rate"]
        for provider in {str(item.proveedor or "").strip() for item in auto_items}:
            provider_snapshot = summaries_by_provider[provider]
            if any(
                provider_snapshot[field] != snapshot[field]
                for field in MIXED_AUTO_RATE_FIELDS
            ):
                raise ValueError("Tasa de electrificacion mixta inconsistente")
    elif metadata.get("auto_electrification_rate") is not None:
        raise ValueError("Tasa de electrificacion mixta inesperada")


def _mixed_rate_legend(metadata: dict[str, Any]) -> str:
    summary = _mixed_rate_summary(metadata)
    quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
    rates = "; ".join(
        f"{MIXED_CATALOG_LABELS[entry['catalog']]} "
        f"{entry['base_currency']}/{entry['quote_currency']} {entry['exchange_rate']}"
        for entry in summary
    )
    external = {
        (entry["rate_source"], entry["rate_effective_date"])
        for entry in summary
        if entry["rate_source"] != "identity"
    }
    suffix = ""
    if len(external) == 1:
        source, effective_date = next(iter(external))
        source_label = "Banco de Mexico / DOF" if source == "saas_exchange_rates" else source
        suffix = f" {source_label} {effective_date}"
    elif external:
        suffix = " | " + "; ".join(
            f"{'Banco de Mexico / DOF' if source == 'saas_exchange_rates' else source} {effective_date}"
            for source, effective_date in sorted(external)
        )
    return safe_excel_text(
        f"{quote_currency} | precios mixtos mas IVA | {rates}{suffix}"[:1000]
    )


def _write_header(ws, metadata: dict[str, Any]) -> None:
    for row in range(3, 13):
        _unmerge_row(ws, row)
    catalog_list_prices = _uses_catalog_list_prices(metadata)
    mixed_catalog_prices = _uses_mixed_catalog_prices(metadata)
    converted_catalog_prices = _uses_converted_catalog_prices(metadata)
    text = safe_excel_text if converted_catalog_prices else lambda value: value
    ws["B3"] = text(metadata.get("cotizacion", ""))
    if mixed_catalog_prices:
        ws["B4"] = _mixed_rate_legend(metadata)
    elif catalog_list_prices:
        base_currency = str(metadata.get("base_currency") or "").strip().upper()
        quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
        exchange_rate = str(metadata.get("exchange_rate") or "").strip()
        effective_date = str(metadata.get("rate_effective_date") or "").strip()
        rate_source = str(metadata.get("rate_source") or "").strip()
        source_label = "Banco de Mexico / DOF" if rate_source == "saas_exchange_rates" else rate_source
        ws["B4"] = safe_excel_text(
            f"{quote_currency} | precios netos mas IVA | Tipo de cambio "
            f"{base_currency}/{quote_currency}: {exchange_rate} | {source_label} | Fecha {effective_date}"
        )
    else:
        ws["B4"] = None
    ws["B7"] = text(metadata.get("proyecto", ""))
    ws["B8"] = text(metadata.get("cliente", ""))
    ws["B9"] = text(metadata.get("correo", ""))
    ws["B10"] = text(metadata.get("telefono", ""))
    ws["B11"] = text(metadata.get("direccion", ""))
    ws["B12"] = text(metadata.get("razon_social", ""))


def _apply_mobiliti_provider_validation(ws) -> None:
    if "Proveedores" not in ws.parent.sheetnames:
        return

    provider_ws = ws.parent["Proveedores"]
    last_provider_row = 1
    for row in range(provider_ws.max_row, 1, -1):
        if provider_ws.cell(row, 1).value:
            last_provider_row = row
            break
    if last_provider_row < 2:
        return

    formula = MOBILITI_PROVIDER_LIST_NAME
    target = f"{_sheet_name(provider_ws.title)}!$A$2:$A${last_provider_row}"
    if MOBILITI_PROVIDER_LIST_NAME in ws.parent.defined_names:
        del ws.parent.defined_names[MOBILITI_PROVIDER_LIST_NAME]
    ws.parent.defined_names.add(
        DefinedName(MOBILITI_PROVIDER_LIST_NAME, attr_text=target)
    )
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    for start_row, capacity in _mobiliti_product_ranges(ws):
        validation.add(f"F{start_row}:F{start_row + capacity - 1}")


def _apply_mobiliti_region_validation(ws) -> None:
    validation = DataValidation(type="list", formula1="Taba_Region", allow_blank=True)
    ws.add_data_validation(validation)
    for start_row, capacity in _mobiliti_product_ranges(ws):
        validation.add(f"P{start_row}:P{start_row + capacity - 1}")


def _write_fletes(ws, mobiliti_total_row: int | None = None) -> None:
    for row, (origin, destination) in FLETE_ROUTES.items():
        ws.cell(row, 1).value = origin
        ws.cell(row, 3).value = destination
    ws["D19"] = f"=Mobiliti!H{mobiliti_total_row or MOBILITI_TOTAL_ROW}"
    ws["F20"] = "=IF($D$19=0,0,D11/$D$19)"
    for row in range(6, 19):
        for col in (11, 14):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("=") and not value.startswith("=IFERROR("):
                ws.cell(row, col).value = f"=IFERROR({value[1:]},0)"
    for row in range(27, 31):
        ws.cell(row, 2).value = f"=IF($D$19=0,0,_xlfn.XLOOKUP(A{row},Taba_Region,Fletes!$D$5:$D$18)/$D$19)"
        ws.cell(row, 3).value = 0
        ws.cell(row, 4).value = 0
    ws["I8"] = "Escritorios-WorkStation"
    ws["M8"] = "Escritorios-WorkStation"


def _write_mobiliti_settings(ws, metadata: dict[str, Any]) -> None:
    if _uses_mixed_catalog_prices(metadata):
        quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
        if quote_currency not in MIXED_MONEY_FORMATS:
            raise ValueError("Moneda mixta incompleta")
        exchange_pair = f"{quote_currency}/{quote_currency}"
        exchange_rate = 1
    elif _uses_catalog_list_prices(metadata):
        exchange_rate = _positive_num(metadata.get("exchange_rate"))
        if exchange_rate is None:
            raise ValueError("Tipo de cambio congelado invalido")
        base_currency = str(metadata.get("base_currency") or "").strip().upper()
        quote_currency = str(metadata.get("quote_currency") or "").strip().upper()
        if not base_currency or not quote_currency:
            raise ValueError("Moneda de catalogo incompleta")
        exchange_pair = f"{base_currency}/{quote_currency}"
    else:
        exchange_rate = _exchange_rate(metadata)
        exchange_pair = "USD/MXN"
    delivery_place = (
        metadata.get("lugar_entrega")
        or metadata.get("delivery_place")
        or DEFAULT_DELIVERY_PLACE
    )
    ws["J6"] = exchange_pair
    ws["K6"] = exchange_rate
    ws["K8"] = delivery_place


def _write_estrategia_comercial(ws, total_row: int) -> None:
    ws["D59"] = f"=Cotizacion!H{total_row}"


def _format_product_row_text(ws, row: int) -> None:
    for col in range(1, 11):
        cell = ws.cell(row, col)
        font = copy(cell.font)
        font.sz = 18
        cell.font = font
        alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=alignment.horizontal,
            vertical="center",
            text_rotation=alignment.text_rotation,
            wrap_text=False if col == 2 else True,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent,
        )


def _align_description_top_for_category(ws, row: int, category: str) -> None:
    if category not in {"Escritorios-WorkStation", "Mesas de Juntas"}:
        return
    cell = ws.cell(row, 3)
    alignment = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal=alignment.horizontal,
        vertical="top",
        text_rotation=alignment.text_rotation,
        wrap_text=alignment.wrap_text,
        shrink_to_fit=alignment.shrink_to_fit,
        indent=alignment.indent,
    )


def _apply_warning_description_fill(cell) -> None:
    if "ADVERTENCIA:" in str(cell.value or "").upper():
        cell.fill = PatternFill(fill_type="solid", fgColor=WARNING_FILL)


def _anchor_position(img: XlsxImage) -> tuple[int, int]:
    anchor = getattr(img, "anchor", None)
    marker = getattr(anchor, "_from", None)
    return (
        int(getattr(marker, "row", 0) or 0) + 1,
        int(getattr(marker, "col", 0) or 0) + 1,
    )


def _transparent_white_logo_png(data: bytes) -> BytesIO:
    output = BytesIO()
    with PILImage.open(BytesIO(data)) as src:
        rgba = src.convert("RGBA")
        pixels = rgba.load()
        for y in range(rgba.height):
            for x in range(rgba.width):
                r, g, b, a = pixels[x, y]
                if a == 0:
                    continue
                whiteness = min(r, g, b)
                if whiteness >= 246:
                    pixels[x, y] = (r, g, b, 0)
                elif whiteness >= 224 and max(r, g, b) - min(r, g, b) <= 18:
                    alpha = int(a * (246 - whiteness) / 22)
                    pixels[x, y] = (r, g, b, alpha)
        rgba.save(output, format="PNG")
    output.seek(0)
    return output


def _normalize_cotizacion_header_logo(ws) -> None:
    replacements: list[tuple[XlsxImage, XlsxImage]] = []
    for img in list(getattr(ws, "_images", [])):
        row, col = _anchor_position(img)
        if row > 15 or col < 6:
            continue
        stream = _transparent_white_logo_png(img._data())
        replacement = XlsxImage(stream)
        replacement.width = img.width
        replacement.height = img.height
        replacement.anchor = deepcopy(img.anchor)
        replacement._mobiliti_stream = stream
        replacements.append((img, replacement))

    for old_img, new_img in replacements:
        ws._images.remove(old_img)
        ws.add_image(new_img)


def _find_authorization_row(ws, fallback_row: int) -> int:
    targets = (
        "NOMBRE, FIRMA Y FECHA DE AUTORIZACIÓN DEL CLIENTE",
        "NOMBRE, FIRMA Y FECHA DE AUTORIZACION DEL CLIENTE",
    )
    for row in range(1, ws.max_row + 1):
        for col in range(1, 11):
            value = str(ws.cell(row, col).value or "").upper()
            if any(target in value for target in targets):
                return row
    return fallback_row


def _lumbro_accessory_image_path(
    metadata: dict[str, Any],
    category: str | None = None,
    user_count: int | None = None,
) -> Path | None:
    raw_path = metadata.get("lumbro_image_path", metadata.get("imagen_lumbro"))
    if raw_path:
        path = Path(raw_path)
    elif category == "Escritorios-WorkStation" and (user_count or 0) > 1:
        path = LUMBRO_WORKSTATION_IMAGE
    else:
        path = LUMBRO_ACCESSORY_IMAGE
    return path if path.exists() else None


def _bottom_center_image_in_cell(
    img: XlsxImage,
    *,
    row: int,
    column: int,
    cell_width: float,
    cell_height: float,
    bottom_padding: float = 18,
) -> XlsxImage:
    col_offset = max(0, (cell_width - img.width) / 2)
    row_offset = max(0, cell_height - img.height - bottom_padding)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=column - 1,
            row=row - 1,
            colOff=pixels_to_EMU(col_offset),
            rowOff=pixels_to_EMU(row_offset),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)),
    )
    return img


def _write_cotizacion(
    ws,
    items: list[QuoteItem],
    row_map: dict[int, int],
    lumbro_row_map: dict[int, list[int]],
    image_map: dict[int, str],
    metadata: dict[str, Any],
) -> int:
    description_language = normalize_description_language(
        metadata.get("description_language", metadata.get("idioma_descripcion", "es"))
    )
    category_dictionary = load_category_dictionary(
        [str(item.nombre or "") for item in items if item.tipo == "producto"]
    )
    terms_start = _find_terms_row(ws)
    totals_start = _find_totals_row(ws, terms_start)
    totals_snapshot = _snapshot_rows(ws, totals_start, totals_start + 4)
    terms_end = min(ws.max_row, max(terms_start, terms_start + 136))
    terms_snapshot = _snapshot_rows(ws, terms_start, terms_end)
    base_description_fill = copy(ws.cell(17, 3).fill)

    for row in range(16, terms_end + 1):
        _clear_row(ws, row)

    current_row = 16
    first_product = None
    last_product = None
    discount_rate = _discount_rate(metadata)
    catalog_list_prices = _uses_catalog_list_prices(metadata)
    mixed_catalog_prices = _uses_mixed_catalog_prices(metadata)
    converted_catalog_prices = _uses_converted_catalog_prices(metadata)
    money_format = _money_format(metadata)
    quote_to_cot: dict[int, int] = {}
    category_by_quote_row: dict[int, str] = {}
    user_count_by_quote_row: dict[int, int | None] = {}

    for item in items:
        if item.tipo == "categoria":
            _copy_row_style(ws, 16, current_row)
            _clear_row(ws, current_row)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(current_row, 1).value = _formula("Quotation", f"A{item.row}")
            current_row += 1
            continue

        if first_product is None:
            first_product = current_row
        last_product = current_row
        quote_to_cot[item.row] = current_row
        mob_row = row_map.get(item.row)

        _copy_row_style(ws, 17, current_row)
        _clear_row(ws, current_row)
        category = classify_product_name(
            str(item.nombre or ""),
            category_dictionary,
            description=item.descripcion,
            source_category=item.categoria,
            supplier=(
                item.proveedor
                or metadata.get("catalog_supplier")
                or metadata.get("catalog_supplier_label")
            ),
        )
        category_by_quote_row[item.row] = category
        user_count_by_quote_row[item.row] = _detect_user_count(item)
        ws.cell(current_row, 1).value = _formula("Quotation", f"B{item.row}")
        code_alignment = copy(ws.cell(current_row, 1).alignment)
        ws.cell(current_row, 1).alignment = Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=code_alignment.text_rotation,
            wrap_text=code_alignment.wrap_text,
            shrink_to_fit=code_alignment.shrink_to_fit,
            indent=code_alignment.indent,
        )
        description_cell = ws.cell(current_row, 3)
        description_cell.fill = copy(base_description_fill)
        description_cell.value = build_product_description(
            item.nombre,
            item.descripcion,
            category,
            description_language,
        )
        _apply_warning_description_fill(description_cell)
        ws.cell(current_row, 4).value = _formula("Quotation", f"E{item.row}")
        if mob_row:
            ws.cell(current_row, 5).value = f"=Mobiliti!H{mob_row}"
            lumbro_rows = lumbro_row_map.get(item.row, [])
            if mixed_catalog_prices:
                if lumbro_rows:
                    price_terms = [
                        f"Mobiliti!X{mob_row}*Mobiliti!H{mob_row}",
                        *(f"Mobiliti!X{row}*Mobiliti!H{row}" for row in lumbro_rows),
                    ]
                    total_formula = "+".join(price_terms)
                    ws.cell(current_row, 6).value = (
                        f"=ROUND(IFERROR(({total_formula})/Mobiliti!H{mob_row},0),2)"
                    )
                else:
                    ws.cell(current_row, 6).value = f"=ROUND(Mobiliti!X{mob_row},2)"
            elif catalog_list_prices:
                ws.cell(current_row, 6).value = f"=ROUND(Mobiliti!X{mob_row},2)"
            elif lumbro_rows:
                price_terms = [
                    f"Mobiliti!X{mob_row}*Mobiliti!H{mob_row}",
                    *(f"Mobiliti!Y{row}" for row in lumbro_rows),
                ]
                total_formula = "+".join(price_terms)
                ws.cell(current_row, 6).value = f"=IFERROR(({total_formula})/Mobiliti!H{mob_row},0)"
            else:
                ws.cell(current_row, 6).value = f"=Mobiliti!X{mob_row}"
        else:
            ws.cell(current_row, 5).value = item.cantidad
            ws.cell(current_row, 6).value = (
                f"=ROUND({_excel_decimal(item.precio)},2)"
                if converted_catalog_prices
                else item.precio
            )
        ws.cell(current_row, 7).value = (
            discount_rate if current_row == first_product else f"=G${first_product}"
        )
        if converted_catalog_prices:
            ws.cell(current_row, 8).value = f"=ROUND(F{current_row}*G{current_row},2)"
            ws.cell(current_row, 9).value = f"=ROUND(F{current_row}-H{current_row},2)"
            ws.cell(current_row, 10).value = f"=ROUND(E{current_row}*I{current_row},2)"
        else:
            ws.cell(current_row, 8).value = f"=F{current_row}*G{current_row}"
            ws.cell(current_row, 9).value = f"=F{current_row}-H{current_row}"
            ws.cell(current_row, 10).value = f"=E{current_row}*I{current_row}"
        ws.cell(current_row, 7).number_format = PERCENT_FORMAT
        for col in [6, 8, 9, 10]:
            ws.cell(current_row, col).number_format = money_format
        _format_product_row_text(ws, current_row)
        _align_description_top_for_category(ws, current_row, category)
        current_row += 1

    if first_product is None or last_product is None:
        raise ValueError("No se encontraron productos en Quotation")

    total_labels = ["SUBTOTAL:", "COSTO DE FLETE:", "SUBTOTAL:", "IVA:", "TOTAL:"]
    if converted_catalog_prices:
        total_formulas = [
            f"=ROUND(SUM(J{first_product}:J{last_product}),2)",
            f"=ROUND(H{current_row}*12%,2)",
            f"=ROUND(H{current_row}+H{current_row + 1},2)",
            f"=ROUND(H{current_row + 2}*16%,2)",
            f"=ROUND(H{current_row + 2}+H{current_row + 3},2)",
        ]
    else:
        total_formulas = [
            f"=SUM(J{first_product}:J{last_product})",
            f"=H{current_row}*12%",
            f"=H{current_row}+H{current_row + 1}",
            f"=H{current_row + 2}*16%",
            f"=H{current_row + 2}+H{current_row + 3}",
        ]
    total_row = current_row + len(total_labels) - 1
    _restore_rows(ws, current_row, totals_snapshot)
    for offset, (label, formula) in enumerate(zip(total_labels, total_formulas)):
        row = current_row + offset
        ws.cell(row, 4).value = label
        ws.cell(row, 8).value = formula
        ws.cell(row, 8).number_format = money_format
        value_alignment = copy(ws.cell(row, 8).alignment)
        ws.cell(row, 8).alignment = Alignment(
            horizontal="right",
            vertical=value_alignment.vertical,
            text_rotation=value_alignment.text_rotation,
            wrap_text=value_alignment.wrap_text,
            shrink_to_fit=value_alignment.shrink_to_fit,
            indent=value_alignment.indent,
        )
    for row in range(current_row + len(total_labels), current_row + 7):
        ws.row_dimensions[row].height = 8
    current_row += 7
    last_terms = _restore_rows(ws, current_row, terms_snapshot)

    for q_row, image_path in image_map.items():
        cot_row = quote_to_cot.get(q_row)
        if not cot_row:
            continue
        cell = ws.cell(cot_row, 2)
        width_px = (ws.column_dimensions[get_column_letter(cell.column)].width or 18) * 7
        height_px = (ws.row_dimensions[cot_row].height or 90) * 1.33
        scale = image_scale_for_category(category_by_quote_row.get(q_row))
        img = fit_image_to_cell(image_path, width_px, height_px, scale=scale)
        center_image_in_cell(img, row=cot_row, column=cell.column, cell_width=width_px, cell_height=height_px)
        ws.add_image(img)

    for q_row in lumbro_row_map:
        lumbro_image_path = _lumbro_accessory_image_path(
            metadata,
            category_by_quote_row.get(q_row),
            user_count_by_quote_row.get(q_row),
        )
        if lumbro_image_path:
            cot_row = quote_to_cot.get(q_row)
            if not cot_row:
                continue
            cell = ws.cell(cot_row, 3)
            width_px = (ws.column_dimensions[get_column_letter(cell.column)].width or 42) * 7
            height_px = (ws.row_dimensions[cot_row].height or 90) * 1.33
            img = fit_image_to_cell(str(lumbro_image_path), width_px, height_px, scale=0.38)
            _bottom_center_image_in_cell(
                img,
                row=cot_row,
                column=cell.column,
                cell_width=width_px,
                cell_height=height_px,
            )
            ws.add_image(img)

    print_end_row = _find_authorization_row(ws, last_terms)
    ws.print_area = f"A1:J{print_end_row}"
    return total_row


def _align_image_map_to_product_rows(
    image_map: dict[int, str],
    items: list[QuoteItem],
    max_distance: int = 3,
) -> dict[int, str]:
    product_rows = [item.row for item in items if item.tipo == "producto"]
    if not product_rows:
        return image_map

    product_row_set = set(product_rows)
    aligned: dict[int, str] = {}
    for image_row, image_path in sorted(image_map.items()):
        if image_row in product_row_set:
            target_row = image_row
        else:
            nearby = [row for row in product_rows if abs(row - image_row) <= max_distance]
            if not nearby:
                aligned[image_row] = image_path
                continue
            # Excel often anchors floating images one row above the product row.
            target_row = min(nearby, key=lambda row: (abs(row - image_row), 0 if row >= image_row else 1))

        if target_row not in aligned or target_row == image_row:
            aligned[target_row] = image_path
    return aligned


def _generate_missing_dezgo_images(
    image_map: dict[int, str],
    items: list[QuoteItem],
    temp_dir: str | Path,
    metadata: dict[str, Any],
    stats: dict[str, Any] | None = None,
) -> dict[int, str]:
    requested_provider = metadata.get("image_provider", metadata.get("proveedor_imagen"))
    provider = normalize_image_provider(requested_provider or os.environ.get("IMAGE_PROVIDER"))
    if provider != "dezgo":
        return image_map

    config = dezgo_config_from_env()
    style_prompt = str(metadata.get("image_prompt", metadata.get("prompt_imagen", ""))).strip() or config.prompt
    generated_dir = Path(temp_dir) / "generated"
    result = dict(image_map)
    for item in items:
        if item.tipo != "producto" or item.row in result:
            continue
        _bump_image_stat(stats, "image_ai_missing_attempted_count")
        prompt = _dezgo_missing_image_prompt(item, style_prompt)
        output = generated_dir / f"missing_{item.row}_{hashlib.sha256(prompt.encode('utf-8')).hexdigest()[:16]}.png"
        if not output.exists():
            try:
                generate_with_dezgo(prompt, output, config)
            except Exception:
                _bump_image_stat(stats, "image_ai_missing_failed_count")
                if image_provider_failure_is_fatal(requested_provider):
                    raise
                continue
        _bump_image_stat(stats, "image_ai_generated_count")
        result[item.row] = str(output)
    return result


def _resolve_sunon_web_images(
    image_map: dict[int, str],
    items: list[QuoteItem],
    temp_dir: str | Path,
    metadata: dict[str, Any],
    stats: dict[str, Any] | None = None,
) -> dict[int, str]:
    requested_provider = metadata.get("image_provider", metadata.get("proveedor_imagen"))
    provider = normalize_image_provider(requested_provider or os.environ.get("IMAGE_PROVIDER"))
    if provider != "sunon_web" or not sunon_lookup_enabled():
        return image_map

    result = dict(image_map)
    sunon_dir = Path(temp_dir) / "sunon"
    lookup_limit, timeout_seconds, deadline = _sunon_lookup_controls()
    lookup_cache: dict[str, Path | None] = {}
    lookup_count = 0
    for item in items:
        if item.tipo != "producto":
            continue
        code = extract_product_code(str(item.nombre or ""))
        if not code:
            continue
        cache_key = normalize_sunon_code(code)
        if cache_key in lookup_cache:
            _bump_image_stat(stats, "image_sunon_cache_hit_count")
            image_path = lookup_cache[cache_key]
            if image_path:
                result[item.row] = str(image_path)
            continue
        if lookup_count >= lookup_limit or time.monotonic() >= deadline:
            _bump_image_stat(stats, "image_sunon_skipped_limit_count")
            continue
        lookup_count += 1
        _bump_image_stat(stats, "image_sunon_attempted_count")
        try:
            image_path = fetch_sunon_product_image(
                str(item.nombre or ""),
                sunon_dir,
                timeout_seconds=timeout_seconds,
            )
        except Exception:
            _bump_image_stat(stats, "image_sunon_failed_count")
            lookup_cache[cache_key] = None
            continue
        lookup_cache[cache_key] = image_path
        if not image_path:
            _bump_image_stat(stats, "image_sunon_not_found_count")
            continue
        _bump_image_stat(stats, "image_sunon_found_count")
        result[item.row] = str(image_path)
    return result


def _resolve_sunon_catalog_images(
    image_map: dict[int, str],
    items: list[QuoteItem],
    temp_dir: str | Path,
    metadata: dict[str, Any],
    stats: dict[str, Any] | None = None,
) -> dict[int, str]:
    requested_provider = metadata.get("image_provider", metadata.get("proveedor_imagen"))
    provider = normalize_image_provider(requested_provider or os.environ.get("IMAGE_PROVIDER"))
    if provider != "sunon_catalog" or not sunon_lookup_enabled():
        return image_map

    result = dict(image_map)
    sunon_dir = Path(temp_dir) / "sunon_catalog"
    lookup_limit, timeout_seconds, deadline = _sunon_lookup_controls(catalog=True)
    live_lookup_value = str(metadata.get("sunon_live_lookup", os.environ.get("SUNON_CATALOG_LIVE_LOOKUP_ENABLED", "")))
    live_lookup = live_lookup_value.strip().lower() in {"1", "true", "yes", "on"}
    lookup_cache: dict[str, Path | None] = {}
    lookup_count = 0
    for item in items:
        if item.tipo != "producto":
            continue
        code = extract_product_code(str(item.nombre or ""))
        if not code:
            _bump_image_stat(stats, "image_sunon_catalog_not_found_count")
            continue
        _entry, matched_candidate, match_type = find_sunon_catalog_match(code)
        cache_key = normalize_sunon_code(matched_candidate or code)
        if cache_key in lookup_cache:
            _bump_image_stat(stats, "image_sunon_catalog_cache_hit_count")
            image_path = lookup_cache[cache_key]
        else:
            if lookup_count >= lookup_limit or time.monotonic() >= deadline:
                _bump_image_stat(stats, "image_sunon_catalog_skipped_limit_count")
                continue
            lookup_count += 1
            _bump_image_stat(stats, "image_sunon_catalog_attempted_count")
            try:
                image_path = fetch_sunon_catalog_product_image(
                    str(item.nombre or ""),
                    sunon_dir,
                    live_lookup=live_lookup,
                    timeout_seconds=timeout_seconds,
                )
            except Exception:
                _bump_image_stat(stats, "image_sunon_catalog_failed_count")
                lookup_cache[cache_key] = None
                continue
            lookup_cache[cache_key] = image_path
        if not image_path:
            _bump_image_stat(stats, "image_sunon_catalog_not_found_count")
            _bump_image_stat(stats, "image_sunon_catalog_fallback_local_count")
            continue
        if match_type == "base_code":
            _bump_image_stat(stats, "image_sunon_catalog_base_code_count")
        else:
            _bump_image_stat(stats, "image_sunon_catalog_exact_code_count")
        result[item.row] = str(image_path)
    return result


def _sunon_lookup_controls(*, catalog: bool = False) -> tuple[int, int, float]:
    limit = max(
        1,
        min(
            100,
            int(_num(os.environ.get("SUNON_MAX_LOOKUPS_PER_JOB"), DEFAULT_SUNON_LOOKUP_LIMIT)),
        ),
    )
    timeout_name = "SUNON_CATALOG_TIMEOUT_SECONDS" if catalog else "SUNON_LOOKUP_TIMEOUT_SECONDS"
    timeout_seconds = max(
        2,
        min(
            12,
            int(_num(os.environ.get(timeout_name), DEFAULT_SUNON_LOOKUP_TIMEOUT_SECONDS)),
        ),
    )
    budget_seconds = max(
        30,
        min(
            300,
            int(_num(os.environ.get("SUNON_LOOKUP_BUDGET_SECONDS"), DEFAULT_SUNON_LOOKUP_BUDGET_SECONDS)),
        ),
    )
    return limit, timeout_seconds, time.monotonic() + budget_seconds


def _bump_image_stat(stats: dict[str, Any] | None, key: str, amount: int = 1) -> None:
    if stats is not None:
        stats[key] = int(stats.get(key, 0) or 0) + amount


def _estimate_generation_seconds(metadata: dict[str, Any], image_stats: dict[str, Any], item_count: int) -> int:
    provider = normalize_image_provider(metadata.get("image_provider", metadata.get("proveedor_imagen")))
    products = max(1, item_count)
    if provider == "dezgo":
        source_images = int(image_stats.get("image_source_count", 0) or 0)
        missing_attempts = int(image_stats.get("image_ai_missing_attempted_count", 0) or 0)
        return min(1800, max(240, 90 + products * 4 + source_images * 22 + missing_attempts * 35))
    if provider == "sunon_web":
        lookup_attempts = int(image_stats.get("image_sunon_attempted_count", 0) or 0)
        return min(900, max(120, 60 + products * 2 + lookup_attempts * 4))
    if provider == "sunon_catalog":
        lookup_attempts = int(image_stats.get("image_sunon_catalog_attempted_count", 0) or 0)
        return min(900, max(120, 60 + products * 2 + lookup_attempts * 4))
    return min(420, max(75, 45 + products * 2))


def _dezgo_missing_image_prompt(item: QuoteItem, style_prompt: str) -> str:
    details = " ".join(
        str(part or "").replace("\n", " ")
        for part in [item.nombre, item.dimension, item.descripcion]
        if part
    )
    details = " ".join(details.split())[:560]
    category = f"Category: {item.categoria}. " if item.categoria else ""
    return (
        f"{style_prompt}. Generate a catalog image for this quoted office furniture product. "
        f"{category}Product: {details}. "
        "Single item only, centered, full product visible, realistic scale, commercial furniture render, "
        "no text, no SKU labels, no watermark, no people."
    )


def _set_calc_mode(wb: Workbook) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _official_file_hash(source: Path | bytes) -> str:
    digest = hashlib.sha256()
    if isinstance(source, bytes):
        digest.update(source)
    else:
        with source.open("rb") as stream:
            while chunk := stream.read(1024 * 1024):
                digest.update(chunk)
    return digest.hexdigest()


def _official_decimal(value: Any, field_name: str, *, positive: bool = False) -> Decimal:
    if isinstance(value, bool):
        raise ValueError(f"{field_name} invalido")
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"{field_name} invalido") from exc
    if not number.is_finite() or number < 0 or (positive and number <= 0):
        raise ValueError(f"{field_name} invalido")
    return number


def _official_volume(value: Any) -> Decimal:
    """Devuelve siempre un M3 numérico; una medida descriptiva no es volumen."""
    if value is None or str(value).strip() == "":
        return Decimal("0")
    try:
        return _official_decimal(value, "M3")
    except ValueError:
        return Decimal("0")


def _official_quote_currency(metadata: dict[str, Any]) -> str:
    currency = str(metadata.get("quote_currency") or "MXN").strip().upper()
    if currency not in {"MXN", "USD", "EUR"}:
        raise ValueError("Moneda de cotizacion invalida")
    return currency


def _official_item_cost(
    item: QuoteItem,
    metadata: dict[str, Any],
) -> tuple[str, Decimal | None, Decimal | None, Decimal | None]:
    if _uses_mixed_catalog_prices(metadata):
        currency = str(item.moneda_original or "").strip().upper()
        if currency not in {"MXN", "USD", "EUR"}:
            raise ValueError("Moneda original mixta invalida")
        mode = str(item.modo_precio or "").strip().lower()
        if mode == "pending":
            if (
                item.precio_original not in {None, ""}
                or item.tipo_cambio_congelado not in {None, ""}
                or str(item.precio or "").strip() != "Por confirmar"
            ):
                raise ValueError("Precio pendiente mixto inconsistente")
            return currency, None, None, None
        original = _official_decimal(item.precio_original, "Precio original")
        rate = _official_decimal(
            item.tipo_cambio_congelado,
            "Tipo de cambio congelado",
            positive=True,
        )
        converted = _official_decimal(item.precio, "Precio convertido").quantize(
            MIXED_MONEY_QUANTUM,
            rounding=ROUND_HALF_UP,
        )
        expected = (original * rate).quantize(
            MIXED_MONEY_QUANTUM,
            rounding=ROUND_HALF_UP,
        )
        if converted != expected:
            raise ValueError("Precio convertido mixto inconsistente")
        return currency, original, rate, converted

    converted = _official_decimal(item.precio or 0, "Precio").quantize(
        MIXED_MONEY_QUANTUM,
        rounding=ROUND_HALF_UP,
    )
    return _official_quote_currency(metadata), converted, Decimal("1"), converted


def _official_source_groups(items: Sequence[QuoteItem]) -> list[tuple[str, list[QuoteItem]]]:
    groups: list[tuple[str, list[QuoteItem]]] = []
    title = "Mobiliario"
    products: list[QuoteItem] = []
    for item in items:
        if item.tipo == "categoria":
            if products:
                groups.append((title, products))
                products = []
            title = str(item.nombre or "Mobiliario").strip() or "Mobiliario"
        elif item.tipo == "producto":
            products.append(item)
    if products:
        groups.append((title, products))
    if not groups:
        raise ValueError("No se encontraron productos en Quotation")
    return groups


def _related_package_part(
    package: XlsxPackage,
    owner: str,
    relationship_id: str,
    relationship_name: str,
) -> str:
    relationships_part = relationship_part_name(owner)
    try:
        root = ET.fromstring(package.parts[relationships_part])
    except (KeyError, ET.ParseError) as error:
        raise ValueError(f"Relaciones OOXML ausentes para {owner}") from error
    matches = [
        relationship
        for relationship in root.findall(f"{{{_PKG_REL_NS}}}Relationship")
        if relationship.attrib.get("Id") == relationship_id
        and relationship.attrib.get("Type")
        in relationship_type_uris(relationship_name)
    ]
    if len(matches) != 1:
        raise ValueError(f"Relación {relationship_name} OOXML inválida")
    relationship = matches[0]
    if relationship.attrib.get("TargetMode", "").casefold() == "external":
        raise ValueError(f"TargetMode externo no permitido para {relationship_name}")
    return resolve_internal_target(owner, relationship.attrib.get("Target", ""))


def _source_product_image_payloads(
    package: XlsxPackage,
) -> dict[int, tuple[bytes, str]]:
    try:
        sheet_part = package.sheet_part("Quotation")
    except KeyError:
        return {}
    sheet = ET.fromstring(package.parts[sheet_part])
    drawings = sheet.findall(f"{{{_SHEET_NS}}}drawing")
    if not drawings:
        return {}
    if len(drawings) != 1:
        raise ValueError("Quotation contiene drawings ambiguos")
    relationship_id = drawings[0].attrib.get(f"{{{_OFFICE_REL_NS}}}id", "")
    drawing_part = _related_package_part(
        package,
        sheet_part,
        relationship_id,
        "drawing",
    )
    drawing = ET.fromstring(package.parts[drawing_part])
    relationships_part = relationship_part_name(drawing_part)
    try:
        relationships = ET.fromstring(package.parts[relationships_part])
    except (KeyError, ET.ParseError) as error:
        raise ValueError("Relaciones del drawing Quotation inválidas") from error

    embedded_ids = {
        blip.attrib.get(f"{{{_OFFICE_REL_NS}}}embed", "")
        for blip in drawing.iter(f"{{{_DRAWING_NS}}}blip")
    }
    media_by_id: dict[str, str] = {}
    for relationship in relationships.findall(f"{{{_PKG_REL_NS}}}Relationship"):
        if relationship.attrib.get("Type") not in relationship_type_uris("image"):
            continue
        relationship_id = relationship.attrib.get("Id", "")
        if relationship_id not in embedded_ids:
            continue
        if relationship.attrib.get("TargetMode", "").casefold() == "external":
            raise ValueError("TargetMode externo no permitido para image")
        if not relationship_id or relationship_id in media_by_id:
            raise ValueError("Relación de imagen Quotation ambigua")
        media_by_id[relationship_id] = resolve_internal_target(
            drawing_part,
            relationship.attrib.get("Target", ""),
        )

    images: dict[int, tuple[bytes, str]] = {}
    for anchor in drawing:
        blip = anchor.find(f".//{{{_DRAWING_NS}}}blip")
        if blip is None:
            continue
        row_text = anchor.findtext(f"{{{_XDR_NS}}}from/{{{_XDR_NS}}}row")
        relationship_id = blip.attrib.get(f"{{{_OFFICE_REL_NS}}}embed", "")
        if row_text is None or not row_text.isdigit() or relationship_id not in media_by_id:
            raise ValueError("Ancla de imagen Quotation inválida")
        row = int(row_text) + 1
        if row in images:
            raise ValueError("Más de una imagen Quotation para la misma fila")
        media_part = media_by_id[relationship_id]
        payload = package.parts[media_part]
        if payload.startswith(b"\x89PNG\r\n\x1a\n"):
            content_type = "image/png"
        elif payload.startswith(b"\xff\xd8\xff"):
            content_type = "image/jpeg"
        else:
            raise ValueError("Formato de imagen Quotation no permitido")
        images[row] = (payload, content_type)
    return images


def _source_product_images(source: Path | bytes) -> dict[int, tuple[bytes, str]]:
    package = (
        XlsxPackage.from_bytes(source)
        if isinstance(source, bytes)
        else XlsxPackage.read(source)
    )
    return _source_product_image_payloads(package)


def _official_presentation_lines(
    items: Sequence[QuoteItem],
    metadata: dict[str, Any],
    source_path: Path,
    lumbro_prices: Mapping[str, LumbroPriceRef],
    image_payloads: Mapping[int, tuple[bytes, str]],
) -> tuple[tuple[_OfficialPresentationLine, ...], tuple[SectionNeed, ...]]:
    product_items = [item for item in items if item.tipo == "producto"]
    category_dictionary = load_category_dictionary(
        [str(item.nombre or "") for item in product_items]
    )
    provider_default = (
        str(metadata.get("catalog_supplier_label") or "Sunon Inc").strip()
        or "Sunon Inc"
    )
    source_type = str(metadata.get("source_type") or "").strip().lower()
    single_catalog_origin = {
        "tarkett_cart": "tarkett",
        "offiho_cart": "offiho",
        "supplier_cart": str(metadata.get("catalog_supplier") or "").strip().lower(),
    }.get(source_type, "")
    groups = _official_source_groups(items)
    lines: list[_OfficialPresentationLine] = []
    needs: list[SectionNeed] = []

    for section_index, (raw_title, products) in enumerate(groups, start=1):
        section_id = f"section-{section_index}"
        section_title = safe_excel_text(raw_title)
        section_start = len(lines)
        for item in products:
            category = classify_product_name(
                str(item.nombre or ""),
                category_dictionary,
                description=item.descripcion,
                source_category=item.categoria,
                supplier=(
                    item.proveedor
                    or single_catalog_origin
                    or metadata.get("catalog_supplier")
                    or provider_default
                ),
            )
            provider = (
                safe_excel_text(item.proveedor)
                if _uses_mixed_catalog_prices(metadata)
                else provider_default
            )
            provider = _official_mobiliti_provider(provider or provider_default)
            currency, original, rate, converted = _official_item_cost(item, metadata)
            mode = str(item.modo_precio or "").strip().lower()
            imported = mode == "imported"
            upstream_hash = (
                hashlib.sha256(
                    str(item.referencia_fuente or f"Quotation:{item.row}").encode(
                        "utf-8"
                    )
                ).hexdigest()
                if imported
                else ""
            )
            image_payload = image_payloads.get(item.row)
            item_supplier = str(
                item.proveedor
                or single_catalog_origin
                or metadata.get("catalog_supplier")
                or ""
            ).strip().casefold()
            if (
                image_payload is None
                and str(item.nombre or "").strip().casefold() == "cub hive"
                and item_supplier == "lauco"
                and CUB_HIVE_REFERENCE_IMAGE.exists()
            ):
                image_payload = (
                    CUB_HIVE_REFERENCE_IMAGE.read_bytes(),
                    "image/png",
                )
            line = _OfficialPresentationLine(
                item_key=f"quotation:{item.row}",
                section_id=section_id,
                section_title=section_title,
                item=item,
                name=safe_excel_text(item.nombre or f"Producto {item.row}"),
                description=safe_excel_text(item.descripcion or ""),
                dimensions=safe_excel_text(item.dimension or ""),
                m3=_official_volume(item.m3),
                quantity=_official_decimal(
                    item.cantidad or 1,
                    "Cantidad",
                    positive=True,
                ),
                category=safe_excel_text(category),
                provider=safe_excel_text(provider),
                region=official_region_name(DEFAULT_MOBILITI_REGION),
                original_currency=currency,
                original_cost=original,
                frozen_rate=rate,
                converted_cost=converted,
                origin=(
                    "imported"
                    if imported
                    else single_catalog_origin or "quotation"
                ),
                source_row=item.row if imported else None,
                upstream_row_hash=upstream_hash,
                image_content=image_payload[0] if image_payload is not None else None,
                image_content_type=image_payload[1] if image_payload is not None else None,
            )
            lines.append(line)

            if not _item_auto_electrification(item, metadata):
                continue
            accessories = _lumbro_accessories_for_item(item, category)
            if not accessories:
                continue
            lumbro_rate = (
                _official_decimal(
                    _mixed_auto_electrification_rate(metadata),
                    "Tipo congelado Lumbro",
                    positive=True,
                )
                if _uses_mixed_catalog_prices(metadata)
                else Decimal("1")
            )
            lumbro_image_path = _lumbro_accessory_image_path(
                metadata,
                category,
                _detect_user_count(item),
            )
            lumbro_image_content = (
                lumbro_image_path.read_bytes()
                if lumbro_image_path is not None
                else None
            )
            lumbro_image_type = (
                "image/png"
                if lumbro_image_path is not None
                and lumbro_image_path.suffix.casefold() == ".png"
                else "image/jpeg"
                if lumbro_image_path is not None
                and lumbro_image_path.suffix.casefold() in {".jpg", ".jpeg"}
                else None
            )
            for accessory_index, (code, quantity) in enumerate(accessories, start=1):
                price_ref = lumbro_prices.get(code)
                if price_ref is None:
                    raise ValueError(f"Precio oficial Lumbro ausente: {code}")
                original_lumbro = _official_decimal(
                    price_ref.price_mxn,
                    f"Precio oficial Lumbro {code}",
                ).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
                lines.append(
                    _OfficialPresentationLine(
                        item_key=(
                            f"{line.item_key}:lumbro:{accessory_index}:{code}"
                        ),
                        section_id=section_id,
                        section_title=section_title,
                        item=None,
                        name=code,
                        description="Accesorio de electrificacion Lumbro",
                        dimensions="",
                        m3=Decimal("0"),
                        quantity=Decimal(quantity),
                        category=LUMBRO_CATEGORY,
                        provider=LUMBRO_PROVIDER,
                        region=DEFAULT_MOBILITI_REGION,
                        original_currency="MXN",
                        original_cost=original_lumbro,
                        frozen_rate=lumbro_rate,
                        converted_cost=lumbro_frozen_cost(
                            original_lumbro,
                            lumbro_rate,
                        ),
                        origin="lumbro",
                        source_row=None,
                        upstream_row_hash="",
                        parent_item_key=line.item_key,
                        image_content=(
                            lumbro_image_content if lumbro_image_type else None
                        ),
                        image_content_type=lumbro_image_type,
                    )
                )
        needs.append(
            SectionNeed(
                section_id,
                section_title,
                len(lines) - section_start,
            )
        )

    return tuple(lines), tuple(needs)


def _canonical_handoff_rows(
    value: object,
) -> tuple[QuotationDataRow, ...] | None:
    if value is _ARGUMENT_OMITTED or value is None:
        return None
    if isinstance(value, (str, bytes, bytearray)) or not isinstance(value, Sequence):
        raise TypeError("quotation_data_rows debe ser una secuencia canónica")
    rows = tuple(value)
    if not rows:
        return None
    # Serializar la hoja ejecuta la validación completa de tipo, orden y hashes.
    build_quotation_data_sheet(rows)
    return rows


def _validate_authoritative_handoff_metadata(
    items: Sequence[QuoteItem],
    metadata: dict[str, Any],
) -> None:
    if not _uses_mixed_catalog_prices(metadata):
        return
    _official_quote_currency(metadata)
    if not isinstance(metadata.get("rate_summary"), list):
        raise ValueError("Resumen de tasas mixtas invalido")
    products = [item for item in items if item.tipo == "producto"]
    if not products:
        raise ValueError("Carrito mixto sin productos")
    for item in products:
        mode = str(item.modo_precio or "").strip().lower()
        if mode == "imported":
            discount = _official_decimal(
                item.descuento,
                "Descuento mixto por linea",
            )
            if discount > Decimal("100"):
                raise ValueError("Descuento mixto por linea invalido")
            if item.electrificacion_automatica is not False:
                raise ValueError("Electrificacion imported debe ser explícitamente false")
            continue
        _mixed_item_discount_fraction(item)
        _item_auto_electrification(item, metadata)


def _official_section_needs(
    lines: Sequence[_OfficialPresentationLine],
) -> tuple[SectionNeed, ...]:
    needs: list[SectionNeed] = []
    seen: set[str] = set()
    current_id: str | None = None
    current_title = ""
    count = 0
    for line in lines:
        if line.section_id == current_id:
            if line.section_title != current_title:
                raise ValueError("Título de sección canónica inconsistente")
            count += 1
            continue
        if current_id is not None:
            needs.append(SectionNeed(current_id, current_title, count))
        if line.section_id in seen:
            raise ValueError("Sección canónica no contigua")
        seen.add(line.section_id)
        current_id = line.section_id
        current_title = line.section_title
        count = 1
    if current_id is not None:
        needs.append(SectionNeed(current_id, current_title, count))
    return tuple(needs)


def _bind_authoritative_canonical_rows(
    lines: Sequence[_OfficialPresentationLine],
    canonical_rows: Sequence[QuotationDataRow],
    source: Path | bytes,
    metadata: dict[str, Any],
) -> tuple[
    tuple[_OfficialPresentationLine, ...],
    tuple[SectionNeed, ...],
    tuple[QuotationDataRow, ...],
]:
    base_lines = tuple(line for line in lines if line.origin != "lumbro")
    if len(base_lines) != len(canonical_rows):
        raise ValueError("La cantidad canónica no coincide con la presentación")

    catalog_lines = tuple(
        line for line in base_lines if line.origin != "imported"
    )
    catalogs_by_source_hash: dict[str, str] = {}
    if catalog_lines:
        catalog_source_hashes = metadata.get("catalog_source_hashes")
        if not isinstance(catalog_source_hashes, dict) or not catalog_source_hashes:
            raise ValueError("Hashes fuente de catalogo ausentes")
        for catalog, source_hash in catalog_source_hashes.items():
            if (
                catalog not in MIXED_CATALOG_ORDER
                or not isinstance(source_hash, str)
                or re.fullmatch(r"[0-9a-f]{64}", source_hash) is None
            ):
                raise ValueError("Hashes fuente de catalogo invalidos")
            if source_hash in catalogs_by_source_hash:
                raise ValueError("Hash fuente de catalogo ambiguo")
            catalogs_by_source_hash[source_hash] = catalog

    bound_bases: dict[str, _OfficialPresentationLine] = {}
    canonical_by_source_key: dict[str, QuotationDataRow] = {}
    for position, (line, canonical) in enumerate(
        zip(base_lines, canonical_rows, strict=True),
        start=1,
    ):
        item = line.item
        if item is None or not isinstance(item.canonical_key, str) or not item.canonical_key:
            raise ValueError("Identidad técnica ausente: canonical_key")
        if not isinstance(item.source_hash, str) or not item.source_hash:
            raise ValueError("Identidad técnica ausente: source_hash")
        expected_origin = line.origin
        if expected_origin == "imported":
            if type(item.source_row) is not int or item.source_row <= 0:
                raise ValueError("Identidad técnica ausente: source_row")
            if (
                not isinstance(item.upstream_row_hash, str)
                or not item.upstream_row_hash
            ):
                raise ValueError("Identidad técnica ausente: upstream_row_hash")
            origin_matches = canonical.origin == expected_origin
        elif canonical.source_hash != item.source_hash:
            # Conserva el diagnóstico de identidad técnica antes de resolver
            # el catálogo autoritativo asociado al hash.
            origin_matches = False
        else:
            expected_origin = catalogs_by_source_hash.get(item.source_hash)
            if expected_origin is None:
                raise ValueError("Hash fuente de catalogo ausente")
            origin_matches = canonical.origin == expected_origin
        comparisons = {
            "item_key": canonical.item_key == item.canonical_key,
            "source_hash": canonical.source_hash == item.source_hash,
            "position": canonical.position == position,
            "origin": origin_matches,
            "original_currency": canonical.original_currency
            == line.original_currency,
            "original_cost": canonical.original_cost == line.original_cost,
            "frozen_rate": canonical.frozen_rate == line.frozen_rate,
            "converted_cost": canonical.converted_cost == line.converted_cost,
            "quantity": canonical.quantity == line.quantity,
            "provider": canonical.provider == line.provider,
        }
        if expected_origin == "imported":
            comparisons["source_row"] = canonical.source_row == item.source_row
            comparisons["upstream_row_hash"] = (
                canonical.upstream_row_hash == item.upstream_row_hash
            )
        mismatch = next(
            (name for name, matches in comparisons.items() if not matches),
            None,
        )
        if mismatch is not None:
            raise ValueError(f"Fila canónica no coincide: {mismatch}")
        canonical_by_source_key[line.item_key] = canonical
        bound_bases[line.item_key] = replace(
            line,
            item_key=canonical.item_key,
            section_id=canonical.section_id,
            section_title=canonical.section_title,
            region=canonical.region,
            origin=canonical.origin,
            source_row=canonical.source_row,
            upstream_row_hash=canonical.upstream_row_hash,
        )

    rebound: list[_OfficialPresentationLine] = []
    original_lines: list[_OfficialPresentationLine] = []
    seen_keys: set[str] = set()
    for line in lines:
        if line.origin != "lumbro":
            bound = bound_bases[line.item_key]
        else:
            parent_source_key = line.parent_item_key
            if parent_source_key is None or parent_source_key not in bound_bases:
                raise ValueError("Accesorio Lumbro sin producto padre canónico")
            parent = bound_bases[parent_source_key]
            prefix = f"{parent_source_key}:"
            if not line.item_key.startswith(prefix):
                raise ValueError("Identidad Lumbro derivada inconsistente")
            bound = replace(
                line,
                item_key=f"{parent.item_key}:{line.item_key[len(prefix):]}",
                parent_item_key=parent.item_key,
                section_id=parent.section_id,
                section_title=parent.section_title,
            )
        if bound.item_key in seen_keys:
            raise ValueError("Identidad canónica duplicada en la presentación")
        seen_keys.add(bound.item_key)
        original_lines.append(line)
        rebound.append(bound)

    source_hash = _official_file_hash(source)
    enriched_rows: list[QuotationDataRow] = []
    for position, (original, bound) in enumerate(
        zip(original_lines, rebound, strict=True),
        start=1,
    ):
        if original.origin == "lumbro":
            enriched = _canonical_row_from_line(bound, position, source_hash)
        else:
            enriched = _with_canonical_hash(
                replace(
                    canonical_by_source_key[original.item_key],
                    position=position,
                    row_hash="",
                )
            )
        enriched_rows.append(enriched)

    result = tuple(rebound)
    return (
        result,
        _official_section_needs(result),
        tuple(enriched_rows),
    )


def _canonical_row_from_line(
    line: _OfficialPresentationLine,
    position: int,
    source_hash: str,
) -> QuotationDataRow:
    return _with_canonical_hash(
        QuotationDataRow(
            item_key=line.item_key,
            section_id=line.section_id,
            section_title=line.section_title,
            position=position,
            origin=line.origin,
            source_row=line.source_row,
            original_currency=line.original_currency,
            original_cost=line.original_cost,
            frozen_rate=line.frozen_rate,
            converted_cost=line.converted_cost,
            quantity=line.quantity,
            provider=line.provider,
            region=line.region,
            source_hash=source_hash,
            upstream_row_hash=line.upstream_row_hash,
            row_hash="",
        )
    )


def _official_canonical_rows(
    lines: Sequence[_OfficialPresentationLine],
    source: Path | bytes,
) -> tuple[QuotationDataRow, ...]:
    source_hash = _official_file_hash(source)
    return tuple(
        _canonical_row_from_line(line, position, source_hash)
        for position, line in enumerate(lines, start=1)
    )


def _official_dimension_write(coordinate: str, value: str) -> MobilitiCellWrite:
    try:
        number = Decimal(value)
    except (InvalidOperation, ValueError):
        return MobilitiCellWrite(coordinate, "text", value)
    if not number.is_finite():
        return MobilitiCellWrite(coordinate, "text", value)
    return MobilitiCellWrite(coordinate, "number", number)


def _build_official_mobiliti(
    base: XlsxPackage,
    lines: Sequence[_OfficialPresentationLine],
    needs: Sequence[SectionNeed],
    canonical_rows: Sequence[QuotationDataRow],
    metadata: dict[str, Any],
    quotation_rows: Mapping[str, int],
    quotation_rates: Mapping[str, Decimal],
) -> MobilitiSheetMutation:
    row_map = plan_mobiliti_layout(needs)
    if len(lines) != len(row_map.item_rows):
        raise ValueError("Presentacion Mobiliti inconsistente")
    writes: list[MobilitiCellWrite] = []
    bindings: list[PricingRowBinding] = []
    for position, (line, target_row) in enumerate(
        zip(lines, row_map.item_rows, strict=True),
        start=1,
    ):
        quotation_row = quotation_rows.get(line.item_key)
        if type(quotation_row) is not int or quotation_row <= 0:
            raise ValueError(
                f"Fila Quotation ausente para {line.item_key}"
            )
        writes.extend(
            (
                MobilitiCellWrite(
                    f"D{target_row}",
                    "formula",
                    f"=Quotation!B{quotation_row}",
                ),
                MobilitiCellWrite(f"E{target_row}", "text", line.category),
                MobilitiCellWrite(f"F{target_row}", "text", line.provider),
                MobilitiCellWrite(
                    f"H{target_row}",
                    "formula",
                    f"=Quotation!H{quotation_row}",
                ),
                MobilitiCellWrite(
                    f"K{target_row}",
                    "formula",
                    f"=Quotation!I{quotation_row}",
                ),
                MobilitiCellWrite(f"P{target_row}", "text", line.region),
            )
        )
        bindings.append(
            PricingRowBinding(
                item_key=line.item_key,
                section_id=line.section_id,
                position=position,
                target_row=target_row,
                quotation_row=quotation_row,
                quotation_rate=quotation_rates.get(line.item_key, Decimal("1")),
            )
        )
    writes.extend(
        build_mobiliti_pricing_writes(
            canonical_rows,
            row_map,
            bindings=tuple(bindings),
        )
    )
    mutation = build_mobiliti_sheet(
        base.parts[base.sheet_part("Mobiliti")],
        list(needs),
        writes,
    )
    editor = WorksheetEditor.from_xml(mutation.xml)
    requested_delivery_place = safe_excel_text(
        metadata.get("lugar_entrega")
        or metadata.get("delivery_place")
        or DEFAULT_DELIVERY_PLACE
    )
    delivery_places = {
        " ".join(destination.split()).casefold(): destination
        for _origin, destination in FLETE_ROUTES.values()
    }
    delivery_place = delivery_places.get(
        " ".join(requested_delivery_place.split()).casefold()
    )
    if delivery_place is None:
        raise ValueError(
            "Lugar de entrega no soportado para Mobiliti!K8: "
            f"{requested_delivery_place}"
        )
    write_official_currency_selector(
        editor,
        _official_quote_currency(metadata),
        delivery_place,
    )
    return MobilitiSheetMutation(editor.to_xml(), mutation.row_map)


def _official_cotizacion_description(
    line: _OfficialPresentationLine,
    language: str,
) -> str:
    return (
        build_product_description(
            line.item.nombre,
            line.item.descripcion,
            line.category,
            language,
        )
        if line.item is not None
        else line.description
    )


def _quotation_description_formula(
    item_key: str,
    quotation_rows: Mapping[str, int],
) -> str:
    quotation_row = quotation_rows.get(item_key)
    if type(quotation_row) is not int or quotation_row <= 1:
        raise ValueError(f"Fila Quotation ausente para {item_key}")
    return f"=Quotation!D{quotation_row}"


def _legacy_cotizacion_product(
    line: _OfficialPresentationLine,
    target_row: int,
    *,
    language: str,
    discount: Decimal,
    quotation_rows: Mapping[str, int],
) -> CotizacionProduct:
    return CotizacionProduct(
        item_key=line.item_key,
        name=line.name,
        description=_official_cotizacion_description(line, language),
        dimensions=line.dimensions,
        quantity=line.quantity,
        mobiliti_row=target_row,
        discount=discount,
        description_formula=_quotation_description_formula(
            line.item_key,
            quotation_rows,
        ),
        image_content=line.image_content,
        image_content_type=line.image_content_type,
        price_pending=line.converted_cost is None,
    )


def _legacy_cotizacion_sections(
    lines: Sequence[_OfficialPresentationLine],
    mobiliti: MobilitiSheetMutation,
    metadata: dict[str, Any],
    quotation_rows: Mapping[str, int],
) -> tuple[CotizacionSection, ...]:
    target_by_key = {
        line.item_key: target_row
        for line, target_row in zip(lines, mobiliti.row_map.item_rows, strict=True)
    }
    language = normalize_description_language(
        metadata.get("description_language", metadata.get("idioma_descripcion", "es"))
    )
    section_order: OrderedDict[str, tuple[str, list[CotizacionProduct]]] = OrderedDict()
    discount = Decimal(str(_discount_rate(metadata)))
    for line in lines:
        title, products = section_order.setdefault(
            line.section_id,
            (line.section_title, []),
        )
        products.append(
            _legacy_cotizacion_product(
                line,
                target_by_key[line.item_key],
                language=language,
                discount=discount,
                quotation_rows=quotation_rows,
            )
        )
    return tuple(
        CotizacionSection(title=title, products=tuple(products))
        for title, products in section_order.values()
    )


def _project_context_lines(context: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    payload = context.get("normalized_project_payload")
    if not isinstance(payload, Mapping):
        raise ValueError("Contexto de Proyecto sin payload congelado")
    raw_lines = payload.get("lines")
    if not isinstance(raw_lines, list) or not raw_lines:
        raise ValueError("Contexto de Proyecto sin líneas congeladas")
    result: dict[str, Mapping[str, Any]] = {}
    for raw_line in raw_lines:
        if not isinstance(raw_line, Mapping):
            raise ValueError("Línea congelada de Proyecto inválida")
        line_id = raw_line.get("line_id")
        if not isinstance(line_id, str) or not line_id:
            raise ValueError("Identidad de línea congelada inválida")
        if line_id in result:
            raise ValueError("Identidad de línea congelada repetida")
        result[line_id] = raw_line
    return result


def _project_cotizacion_sections(
    lines: Sequence[_OfficialPresentationLine],
    mobiliti: MobilitiSheetMutation,
    metadata: dict[str, Any],
    quotation_rows: Mapping[str, int],
) -> tuple[CotizacionSection, ...]:
    context = metadata.get("project_context")
    if context is None:
        return _legacy_cotizacion_sections(
            lines,
            mobiliti,
            metadata,
            quotation_rows,
        )
    if not isinstance(context, Mapping):
        raise ValueError("Contexto de Proyecto inválido")
    compositions = context.get("compositions")
    if not isinstance(compositions, list) or not compositions:
        raise ValueError("Contexto de Proyecto sin composiciones")
    if len(lines) != len(mobiliti.row_map.item_rows):
        raise ValueError("Presentación de Proyecto inconsistente")

    by_key: dict[str, _OfficialPresentationLine] = {}
    target_by_key: dict[str, int] = {}
    for line, target_row in zip(lines, mobiliti.row_map.item_rows, strict=True):
        if line.item_key in by_key:
            raise ValueError("Identidad física de Proyecto repetida")
        by_key[line.item_key] = line
        target_by_key[line.item_key] = target_row

    base_by_key = {
        line.item_key: line for line in lines if line.parent_item_key is None
    }
    frozen_by_key = _project_context_lines(context)
    unknown_frozen = set(frozen_by_key) - set(base_by_key)
    if unknown_frozen:
        raise ValueError("Línea congelada de Proyecto ausente")
    unexpected_physical = set(base_by_key) - set(frozen_by_key)
    if unexpected_physical:
        raise ValueError("Línea física sin contexto de Proyecto")

    language = normalize_description_language(
        metadata.get(
            "description_language",
            metadata.get("idioma_descripcion", "es"),
        )
    )
    discount = Decimal(str(_discount_rate(metadata)))
    consumed: set[str] = set()
    consumed_lumbro: set[str] = set()
    accessories_by_parent: dict[str, list[_OfficialPresentationLine]] = {}
    for line in lines:
        if line.parent_item_key is None:
            continue
        if (
            line.origin != "lumbro"
            or not isinstance(line.parent_item_key, str)
            or line.parent_item_key not in base_by_key
        ):
            raise ValueError("Accesorio Lumbro sin padre de Proyecto")
        parent = base_by_key[line.parent_item_key]
        if line.section_id != parent.section_id:
            raise ValueError("Accesorio Lumbro en sección incorrecta")
        accessories_by_parent.setdefault(line.parent_item_key, []).append(line)

    section_order: OrderedDict[str, tuple[str, list[CotizacionProduct]]] = OrderedDict()
    for raw_composition in compositions:
        if not isinstance(raw_composition, Mapping):
            raise ValueError("Composición de Proyecto inválida")
        principal_id = raw_composition.get("principal_line_id")
        section_id = raw_composition.get("section_id")
        component_ids = raw_composition.get("component_line_ids")
        raw_terms = raw_composition.get("price_terms")
        if (
            not isinstance(principal_id, str)
            or not isinstance(section_id, str)
            or not isinstance(component_ids, list)
            or not component_ids
            or not isinstance(raw_terms, list)
        ):
            raise ValueError("Composición de Proyecto incompleta")
        if component_ids[0] != principal_id:
            raise ValueError("Principal de composición ausente")
        if len(component_ids) != len(raw_terms):
            raise ValueError("Términos de precio de Proyecto incompletos")
        if any(
            not isinstance(component_id, str) or not component_id
            for component_id in component_ids
        ):
            raise ValueError("Línea de Proyecto desconocida en composición")
        if len(component_ids) != len(set(component_ids)):
            raise ValueError("Línea de Proyecto repetida en composición")
        if any(line_id in consumed for line_id in component_ids):
            raise ValueError("Línea de Proyecto repetida entre composiciones")

        missing = next(
            (
                line_id
                for line_id in component_ids
                if line_id not in base_by_key
            ),
            None,
        )
        if missing is not None:
            raise ValueError("Línea de Proyecto desconocida en composición")
        principal = base_by_key[principal_id]
        frozen_principal = frozen_by_key[principal_id]
        if (
            principal.section_id != section_id
            or frozen_principal.get("role") != "principal"
            or frozen_principal.get("section_id") != section_id
            or frozen_principal.get("parent_line_id") is not None
        ):
            raise ValueError("Principal de Proyecto en sección incorrecta")

        components = [base_by_key[line_id] for line_id in component_ids]
        for component_id, component in zip(component_ids[1:], components[1:]):
            frozen_component = frozen_by_key[component_id]
            if (
                component.section_id != section_id
                or frozen_component.get("role") != "complement"
                or frozen_component.get("section_id") is not None
                or frozen_component.get("parent_line_id") != principal_id
            ):
                raise ValueError("Complemento de Proyecto en sección incorrecta")

        price_terms: list[CotizacionPriceTerm] = []
        for component_id, raw_term in zip(component_ids, raw_terms, strict=True):
            if (
                not isinstance(raw_term, Mapping)
                or raw_term.get("line_id") != component_id
            ):
                raise ValueError("Términos de precio de Proyecto incompletos")
            try:
                price_terms.append(
                    CotizacionPriceTerm(
                        target_by_key[component_id],
                        Decimal(str(raw_term.get("numerator"))),
                        Decimal(str(raw_term.get("denominator"))),
                    )
                )
            except (InvalidOperation, TypeError, ValueError) as exc:
                raise ValueError("Término de precio de Proyecto inválido") from exc

        descriptions = [
            _official_cotizacion_description(component, language)
            for component in components
        ]
        description = safe_excel_text(
            "\n".join(
                [
                    descriptions[0],
                    *(
                        f"+ {child_description}"
                        for child_description in descriptions[1:]
                        if child_description
                    ),
                ]
            )
        )
        complement_images = tuple(
            CotizacionImageSource(
                content=component.image_content,
                content_type=component.image_content_type,
            )
            for component in components[1:]
            if component.image_content is not None
        )
        _title, products = section_order.setdefault(
            section_id,
            (principal.section_title, []),
        )
        products.append(
            CotizacionProduct(
                item_key=principal_id,
                name=principal.name,
                description=description,
                dimensions=principal.dimensions,
                quantity=principal.quantity,
                mobiliti_row=target_by_key[principal_id],
                discount=discount,
                description_formula=_quotation_description_formula(
                    principal_id,
                    quotation_rows,
                ),
                image_content=principal.image_content,
                image_content_type=principal.image_content_type,
                complement_images=complement_images,
                price_terms=tuple(price_terms),
                price_pending=any(
                    component.converted_cost is None for component in components
                ),
            )
        )
        consumed.update(component_ids)

        for component_id in component_ids:
            for accessory in accessories_by_parent.get(component_id, ()):
                products.append(
                    _legacy_cotizacion_product(
                        accessory,
                        target_by_key[accessory.item_key],
                        language=language,
                        discount=discount,
                        quotation_rows=quotation_rows,
                    )
                )
                consumed_lumbro.add(accessory.item_key)

    unconsumed = set(base_by_key) - consumed
    if unconsumed:
        raise ValueError("Línea congelada de Proyecto sin consumir")
    all_lumbro = {
        line.item_key for line in lines if line.parent_item_key is not None
    }
    if all_lumbro != consumed_lumbro:
        raise ValueError("Accesorio Lumbro de Proyecto sin consumir")

    return tuple(
        CotizacionSection(title=title, products=tuple(products))
        for title, products in section_order.values()
    )


def _build_official_cotizacion(
    base: XlsxPackage,
    lines: Sequence[_OfficialPresentationLine],
    mobiliti: MobilitiSheetMutation,
    metadata: dict[str, Any],
    quotation_rows: Mapping[str, int],
):
    try:
        base.sheet_part(CDMX_LUMBRO_SHEET)
    except KeyError:
        composer_variant = "official"
    else:
        composer_variant = "sunon_cdmx_v1c"
    cotizacion_lines = _improve_official_cotizacion_images(lines, metadata)
    sections = _project_cotizacion_sections(
        cotizacion_lines,
        mobiliti,
        metadata,
        quotation_rows,
    )
    mutation = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=CotizacionMetadata(
            quotation_number=safe_excel_text(metadata.get("cotizacion", "")),
            project=safe_excel_text(metadata.get("proyecto", "")),
            client=safe_excel_text(metadata.get("cliente", "")),
            email=safe_excel_text(metadata.get("correo", "")),
            phone=safe_excel_text(metadata.get("telefono", "")),
            address=safe_excel_text(metadata.get("direccion", "")),
            business_name=safe_excel_text(metadata.get("razon_social", "")),
            quote_currency=_official_quote_currency(metadata),
        ),
        sections=sections,
        composer_variant=composer_variant,
    )
    return mutation, sections


def _improve_official_cotizacion_images(
    lines: Sequence[_OfficialPresentationLine],
    metadata: Mapping[str, Any],
) -> tuple[_OfficialPresentationLine, ...]:
    """Mejora imágenes de Quotation sólo en Cotizacion; preserva catálogos."""

    image_keys = {
        "image_provider",
        "proveedor_imagen",
        "image_cleanup_strength",
        "limpieza_imagen",
        "image_background",
        "fondo_imagen",
        "image_prompt",
        "prompt_imagen",
    }
    if not any(key in metadata for key in image_keys):
        return tuple(lines)
    background = str(
        metadata.get("image_background", metadata.get("fondo_imagen", "white"))
    ).strip().lower()
    background = {
        "blanco": "white",
        "transparente": "transparent",
    }.get(background, background)
    if background not in {"white", "transparent"}:
        background = "white"
    cleanup_strength = str(
        metadata.get(
            "image_cleanup_strength",
            metadata.get("limpieza_imagen", "balanced"),
        )
    ).strip().lower()
    if cleanup_strength not in {"normal", "balanced", "aggressive"}:
        cleanup_strength = "balanced"

    improved: list[_OfficialPresentationLine] = []
    for line in lines:
        if (
            line.origin not in {"imported", "quotation"}
            or line.image_content is None
            or line.image_content_type is None
        ):
            improved.append(line)
            continue
        try:
            content, content_type = improve_product_image_bytes(
                line.image_content,
                line.image_content_type,
                background=background,
                min_size=550,
                cleanup_strength=cleanup_strength,
                remove_shadow=True,
            )
        except (OSError, ValueError):
            improved.append(line)
            continue
        improved.append(
            replace(
                line,
                image_content=content,
                image_content_type=content_type,
            )
        )
    return tuple(improved)


def _quotation_cell_text(
    cell: ET.Element | None,
    shared_strings: Sequence[bytes],
) -> str:
    if cell is None:
        return ""
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(
            node.text or ""
            for node in cell.findall(f".//{{{_SHEET_NS}}}t")
        )
    value = cell.findtext(f"{{{_SHEET_NS}}}v")
    if cell_type == "s" and value is not None and value.isdigit():
        index = int(value)
        if index >= len(shared_strings):
            raise ValueError("Shared string fuera de rango en Quotation")
        root = ET.fromstring(shared_strings[index])
        return "".join(
            node.text or "" for node in root.findall(f".//{{{_SHEET_NS}}}t")
        )
    return value or ""


def _quotation_row_cell(row: ET.Element, column: str) -> ET.Element | None:
    row_number = row.attrib.get("r", "")
    return row.find(f"{{{_SHEET_NS}}}c[@r='{column}{row_number}']")


@dataclass(frozen=True)
class _QuotationContentLayout:
    first_row: int
    last_row: int
    last_populated_row: int
    section_rows: tuple[int, ...]
    product_rows: tuple[int, ...]
    product_sections: Mapping[int, str]


@dataclass(frozen=True)
class _QuotationColumnLayout:
    transformed: bool
    description: str
    dimension: str
    color: str
    quantity: str
    volume: str
    total_volume: str
    unit_price: str
    total_price: str
    remark: str
    visible_end: int


def _quotation_column_layout(
    rows: Mapping[int, ET.Element],
    layout: _QuotationContentLayout,
    shared_strings: Sequence[bytes],
) -> _QuotationColumnLayout:
    header = rows.get(layout.first_row - 1)
    if header is None:
        raise ValueError("Quotation sin encabezado visible")
    fourth = _quotation_cell_text(
        _quotation_row_cell(header, "D"),
        shared_strings,
    ).strip()
    fifth = _quotation_cell_text(
        _quotation_row_cell(header, "E"),
        shared_strings,
    ).strip()
    tenth = _quotation_cell_text(
        _quotation_row_cell(header, "J"),
        shared_strings,
    ).strip()
    eleventh = _quotation_cell_text(
        _quotation_row_cell(header, "K"),
        shared_strings,
    ).strip()
    description_headers = {"description", "descripcion", "descripción"}
    if (
        fourth.casefold() == "trasformacion"
        and fifth.casefold() in description_headers
    ):
        return _QuotationColumnLayout(
            transformed=True,
            description="E",
            dimension="F",
            color="G",
            quantity="H",
            volume="I",
            total_volume="J",
            unit_price="K",
            total_price="L",
            remark="M",
            visible_end=13,
        )
    if (
        fourth.casefold() in description_headers
        or (
            not fourth
            and not fifth
            and tenth.casefold() in {"unit price", "list price"}
            and eleventh.casefold() in {"tot.price", "total price"}
        )
    ):
        return _QuotationColumnLayout(
            transformed=False,
            description="D",
            dimension="E",
            color="F",
            quantity="G",
            volume="H",
            total_volume="I",
            unit_price="J",
            total_price="K",
            remark="L",
            visible_end=12,
        )
    raise ValueError("Esquema de columnas Quotation no reconocido")


def _quotation_content_rows(
    sheet_root: ET.Element,
    shared_strings: Sequence[bytes],
) -> _QuotationContentLayout:
    sheet_data = sheet_root.find(f"{{{_SHEET_NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("Quotation original sin sheetData")
    populated: list[int] = []
    section_rows: list[int] = []
    product_rows: list[int] = []
    product_sections: dict[int, str] = {}
    current_section = ""
    for row in sheet_data.findall(f"{{{_SHEET_NS}}}row"):
        row_number_text = row.attrib.get("r", "")
        if not row_number_text.isdigit():
            raise ValueError("Fila invalida en Quotation original")
        row_number = int(row_number_text)
        cells = row.findall(f"{{{_SHEET_NS}}}c")
        if any(
            cell.find(f"{{{_SHEET_NS}}}f") is not None
            or _quotation_cell_text(cell, shared_strings).strip()
            for cell in cells
        ):
            populated.append(row_number)
        first = _quotation_cell_text(
            _quotation_row_cell(row, "A"),
            shared_strings,
        ).strip()
        second = _quotation_cell_text(
            _quotation_row_cell(row, "B"),
            shared_strings,
        ).strip()
        if first.startswith("-"):
            current_section = " ".join(first.lstrip("-").split()) or "Mobiliario"
            section_rows.append(row_number)
        try:
            numeric_first = Decimal(first)
        except (InvalidOperation, ValueError):
            numeric_first = None
        if numeric_first is not None and numeric_first.is_finite() and second:
            if not current_section:
                raise ValueError("Producto sin seccion en Quotation original")
            product_rows.append(row_number)
            product_sections[row_number] = current_section
    if not populated or not section_rows or not product_rows:
        raise ValueError("Quotation original sin filas de formato reutilizables")
    content_rows = (*section_rows, *product_rows)
    return _QuotationContentLayout(
        first_row=min(content_rows),
        last_row=max(content_rows),
        last_populated_row=max(populated),
        section_rows=tuple(section_rows),
        product_rows=tuple(product_rows),
        product_sections=product_sections,
    )


def _move_quotation_row(source: ET.Element, target_row: int) -> ET.Element:
    source_row_text = source.attrib.get("r", "")
    if not source_row_text.isdigit() or target_row <= 0:
        raise ValueError("Fila invalida al desplazar Quotation")
    source_row = int(source_row_text)
    clone = deepcopy(source)
    clone.attrib["r"] = str(target_row)
    row_delta = target_row - source_row
    for cell in clone.findall(f"{{{_SHEET_NS}}}c"):
        coordinate = cell.attrib.get("r", "")
        match = re.fullmatch(r"(?P<column>[A-Z]+)(?P<row>[1-9][0-9]*)", coordinate)
        if match is None or int(match.group("row")) != source_row:
            raise ValueError("Coordenada invalida al desplazar Quotation")
        target_coordinate = f"{match.group('column')}{target_row}"
        formula = cell.find(f"{{{_SHEET_NS}}}f")
        if formula is not None and formula.text:
            translated = _translate_formula_value(
                "=" + formula.text,
                coordinate,
                target_coordinate,
            )
            if isinstance(translated, str) and translated.startswith("="):
                formula.text = translated[1:]
                cached = cell.find(f"{{{_SHEET_NS}}}v")
                if cached is not None:
                    cell.remove(cached)
            reference = formula.attrib.get("ref")
            if reference:
                try:
                    formula_range = CellRange(reference)
                    formula_range.shift(row_shift=row_delta)
                    formula.attrib["ref"] = str(formula_range)
                except (TypeError, ValueError):
                    raise ValueError(
                        "Referencia de formula invalida al desplazar Quotation"
                    ) from None
        cell.attrib["r"] = target_coordinate
    return clone


def _quotation_cell_column(cell: ET.Element) -> int:
    coordinate = cell.attrib.get("r", "")
    match = re.fullmatch(r"(?P<column>[A-Z]+)[1-9][0-9]*", coordinate)
    if match is None:
        raise ValueError("Coordenada de celda invalida en Quotation")
    return column_index_from_string(match.group("column"))


def _remove_quotation_sidecars(
    row: ET.Element,
    visible_end: int = 12,
) -> None:
    for cell in list(row.findall(f"{{{_SHEET_NS}}}c")):
        if _quotation_cell_column(cell) > visible_end:
            row.remove(cell)


def _restore_quotation_sidecars(
    sheet_data: ET.Element,
    sidecars: Mapping[int, Sequence[ET.Element]],
    source_rows: Mapping[int, ET.Element],
) -> None:
    rows = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{_SHEET_NS}}}row")
    }
    for row_number, cells in sidecars.items():
        target = rows.get(row_number)
        if target is None:
            target = ET.Element(
                f"{{{_SHEET_NS}}}row",
                dict(source_rows[row_number].attrib),
            )
            target.attrib["r"] = str(row_number)
            sheet_data.append(target)
            rows[row_number] = target
        existing = {
            cell.attrib.get("r", ""): cell
            for cell in target.findall(f"{{{_SHEET_NS}}}c")
        }
        for source_cell in cells:
            coordinate = source_cell.attrib["r"]
            if coordinate in existing:
                target.remove(existing[coordinate])
            target.append(deepcopy(source_cell))
        target[:] = sorted(
            target,
            key=lambda cell: (
                _quotation_cell_column(cell)
                if cell.tag == f"{{{_SHEET_NS}}}c"
                else 1_000_000
            ),
        )
    sheet_data[:] = sorted(
        sheet_data,
        key=lambda row: int(row.attrib["r"]),
    )


def _quotation_row_formula(row: ET.Element, column: str) -> str | None:
    cell = _quotation_row_cell(row, column)
    if cell is None:
        return None
    formula = cell.find(f"{{{_SHEET_NS}}}f")
    return None if formula is None else "=" + (formula.text or "")


def _shift_quotation_merge(reference: str, row_delta: int) -> str:
    try:
        cell_range = CellRange(reference)
        cell_range.shift(row_shift=row_delta)
    except (TypeError, ValueError):
        raise ValueError("Merge invalido al desplazar Quotation") from None
    return str(cell_range)


def _replace_quotation_row(sheet_data: ET.Element, row: ET.Element) -> None:
    target = int(row.attrib["r"])
    for existing in list(sheet_data):
        if existing.tag == f"{{{_SHEET_NS}}}row" and int(existing.attrib["r"]) == target:
            sheet_data.remove(existing)
            break
    sheet_data.append(row)
    sheet_data[:] = sorted(
        sheet_data,
        key=lambda item: int(item.attrib.get("r", "0")),
    )


def _quotation_media_extension(line: _OfficialPresentationLine) -> tuple[str, str]:
    if line.image_content_type == "image/png":
        return "png", "image/png"
    if line.image_content_type == "image/jpeg":
        return "jpeg", "image/jpeg"
    raise ValueError("Formato de imagen de Proyecto no permitido en Quotation")


def _quotation_decimal_value(
    row: ET.Element,
    column: str,
    shared_strings: Sequence[bytes],
) -> Decimal | None:
    text = _quotation_cell_text(
        _quotation_row_cell(row, column),
        shared_strings,
    ).strip()
    if not text:
        return None
    try:
        value = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return value if value.is_finite() else None


def _imported_line_matches_original(
    row: ET.Element,
    line: _OfficialPresentationLine,
    shared_strings: Sequence[bytes],
) -> bool:
    name = _quotation_cell_text(
        _quotation_row_cell(row, "B"),
        shared_strings,
    ).strip()
    dimensions = _quotation_cell_text(
        _quotation_row_cell(row, "E"),
        shared_strings,
    ).strip()
    quantity = _quotation_decimal_value(row, "G", shared_strings)
    price = _quotation_decimal_value(row, "J", shared_strings)
    return (
        name == line.name.strip()
        and (not line.dimensions.strip() or dimensions == line.dimensions.strip())
        and quantity == line.quantity
        and price == line.original_cost
    )


def _ensure_image_content_type(
    content_types: ET.Element,
    extension: str,
    content_type: str,
) -> None:
    namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    matches = [
        item
        for item in content_types.findall(f"{{{namespace}}}Default")
        if item.attrib.get("Extension", "").casefold() == extension.casefold()
    ]
    if matches:
        if any(item.attrib.get("ContentType") != content_type for item in matches):
            raise ValueError("Content type de imagen incompatible en Quotation")
        return
    ET.SubElement(
        content_types,
        f"{{{namespace}}}Default",
        {"Extension": extension, "ContentType": content_type},
    )


def _quotation_column_width_emu(
    sheet_root: ET.Element,
    column_index: int,
) -> int:
    if type(column_index) is not int or column_index < 1:
        raise ValueError("Columna Quotation inválida")
    sheet_format = sheet_root.find(f"{{{_SHEET_NS}}}sheetFormatPr")
    default_width = 8.43
    if sheet_format is not None and sheet_format.attrib.get("defaultColWidth"):
        default_width = float(sheet_format.attrib["defaultColWidth"])
    columns = sheet_root.find(f"{{{_SHEET_NS}}}cols")
    width = default_width
    if columns is not None:
        for item in columns.findall(f"{{{_SHEET_NS}}}col"):
            minimum = int(item.attrib.get("min", "0"))
            maximum = int(item.attrib.get("max", "0"))
            if minimum <= column_index <= maximum:
                width = float(item.attrib.get("width", default_width))
    if not math.isfinite(width) or width <= 0:
        raise ValueError("Ancho de columna Quotation inválido")
    pixels = (
        max(1, round(width * 12))
        if width < 1
        else max(1, math.floor(width * 7 + 5))
    )
    return int(pixels_to_EMU(pixels))


def _quotation_row_height_emu(
    sheet_root: ET.Element,
    row_number: int,
) -> int:
    if type(row_number) is not int or row_number < 1:
        raise ValueError("Fila Quotation inválida")
    sheet_format = sheet_root.find(f"{{{_SHEET_NS}}}sheetFormatPr")
    default_height = 15.0
    if sheet_format is not None and sheet_format.attrib.get("defaultRowHeight"):
        default_height = float(sheet_format.attrib["defaultRowHeight"])
    row = sheet_root.find(
        f"{{{_SHEET_NS}}}sheetData/{{{_SHEET_NS}}}row[@r='{row_number}']"
    )
    height = (
        float(row.attrib["ht"])
        if row is not None and row.attrib.get("ht")
        else default_height
    )
    if not math.isfinite(height) or height <= 0:
        raise ValueError("Altura de fila Quotation inválida")
    return max(1, round(height * 12_700))


def _quotation_image_box(
    sheet_root: ET.Element,
    row_number: int,
) -> tuple[int, int, int, int, int, int]:
    cell_width = _quotation_column_width_emu(sheet_root, 3)
    cell_height = _quotation_row_height_emu(sheet_root, row_number)
    horizontal_margin = max(1, round(cell_width * 0.04))
    vertical_margin = max(1, round(cell_height * 0.04))
    width = cell_width - 2 * horizontal_margin
    height = cell_height - 2 * vertical_margin
    if width <= 0 or height <= 0:
        raise ValueError("Celda Photo sin espacio utilizable")
    return (
        cell_width,
        cell_height,
        horizontal_margin,
        vertical_margin,
        width,
        height,
    )


def _quotation_anchor_dimensions(anchor: ET.Element) -> tuple[int, int]:
    extent = anchor.find(f"{{{_XDR_NS}}}ext")
    if extent is None:
        extent = anchor.find(
            f".//{{{_DRAWING_NS}}}xfrm/{{{_DRAWING_NS}}}ext"
        )
    if extent is None:
        raise ValueError("Imagen Quotation sin dimensiones")
    width = int(extent.attrib.get("cx", "0"))
    height = int(extent.attrib.get("cy", "0"))
    if width <= 0 or height <= 0:
        raise ValueError("Dimensiones de imagen Quotation inválidas")
    return width, height


def _contained_quotation_product_anchor(
    anchor: ET.Element,
    *,
    target_row: int,
    sheet_root: ET.Element,
) -> ET.Element:
    picture = anchor.find(f"{{{_XDR_NS}}}pic")
    if picture is None:
        raise ValueError("Ancla de producto Quotation sin imagen")
    source_width, source_height = _quotation_anchor_dimensions(anchor)
    (
        _cell_width,
        _cell_height,
        left,
        top,
        box_width,
        box_height,
    ) = _quotation_image_box(sheet_root, target_row)
    scale = min(1.0, box_width / source_width, box_height / source_height)
    width = max(1, int(source_width * scale))
    height = max(1, int(source_height * scale))
    column_offset = left + (box_width - width) // 2
    row_offset = top + (box_height - height) // 2

    normalized = ET.Element(f"{{{_XDR_NS}}}oneCellAnchor")
    marker = ET.SubElement(normalized, f"{{{_XDR_NS}}}from")
    for tag, value in (
        ("col", 2),
        ("colOff", column_offset),
        ("row", target_row - 1),
        ("rowOff", row_offset),
    ):
        ET.SubElement(marker, f"{{{_XDR_NS}}}{tag}").text = str(value)
    ET.SubElement(
        normalized,
        f"{{{_XDR_NS}}}ext",
        {"cx": str(width), "cy": str(height)},
    )
    picture_copy = deepcopy(picture)
    shape_extent = picture_copy.find(
        f".//{{{_DRAWING_NS}}}xfrm/{{{_DRAWING_NS}}}ext"
    )
    if shape_extent is not None:
        shape_extent.attrib.update({"cx": str(width), "cy": str(height)})
    normalized.append(picture_copy)
    client_data = anchor.find(f"{{{_XDR_NS}}}clientData")
    normalized.append(
        deepcopy(client_data)
        if client_data is not None
        else ET.Element(f"{{{_XDR_NS}}}clientData")
    )
    return normalized


def _augment_quotation_drawing(
    package: XlsxPackage,
    sheet_part: str,
    sheet_root: ET.Element,
    images: Sequence[tuple[int, _OfficialPresentationLine]],
    *,
    content_first_row: int | None = None,
    content_last_row: int | None = None,
    last_populated_row: int | None = None,
    content_row_map: Mapping[int, Sequence[int]] | None = None,
    footer_delta: int = 0,
) -> tuple[dict[str, bytes], dict[str, bytes]]:
    row_map = dict(content_row_map or {})
    remap_requested = content_first_row is not None
    if not images and not remap_requested:
        return {}, {}
    drawings = sheet_root.findall(f"{{{_SHEET_NS}}}drawing")
    if not drawings:
        if images and any(line.item is not None for _, line in images):
            raise ValueError(
                "Quotation original sin dibujo reutilizable para agregar imagenes"
            )
        return {}, {}
    if len(drawings) != 1:
        raise ValueError("Quotation original requiere un dibujo unico para agregar imagenes")
    if remap_requested and (
        content_last_row is None
        or last_populated_row is None
        or content_first_row > content_last_row
        or content_last_row > last_populated_row
    ):
        raise ValueError("Rango de contenido invalido al desplazar imagenes Quotation")
    sheet_rels_part = relationship_part_name(sheet_part)
    sheet_rels = ET.fromstring(package.parts[sheet_rels_part])
    drawing_id = drawings[0].attrib.get(f"{{{_OFFICE_REL_NS}}}id", "")
    drawing_relationship = next(
        (
            item
            for item in sheet_rels.findall(f"{{{_PKG_REL_NS}}}Relationship")
            if item.attrib.get("Id") == drawing_id
            and item.attrib.get("Type") in relationship_type_uris("drawing")
        ),
        None,
    )
    if drawing_relationship is None:
        raise ValueError("Relacion de dibujo invalida en Quotation original")
    drawing_part = resolve_internal_target(
        sheet_part,
        drawing_relationship.attrib["Target"],
    )
    drawing_rels_part = relationship_part_name(drawing_part)
    drawing_root = ET.fromstring(package.parts[drawing_part])
    drawing_rels = ET.fromstring(package.parts[drawing_rels_part])
    content_types = ET.fromstring(package.parts["[Content_Types].xml"])
    if remap_requested:
        assert content_first_row is not None
        assert content_last_row is not None
        assert last_populated_row is not None
        next_copied_picture_id = _next_picture_id(drawing_root)
        for anchor in list(drawing_root):
            marker = anchor.find(f"{{{_XDR_NS}}}from")
            row_node = (
                marker.find(f"{{{_XDR_NS}}}row")
                if marker is not None
                else None
            )
            if row_node is None or row_node.text is None:
                continue
            source_row = int(row_node.text) + 1
            if content_first_row <= source_row <= content_last_row:
                target_rows = row_map.get(source_row, ())
                if not target_rows:
                    drawing_root.remove(anchor)
                    continue
                if anchor.find(f"{{{_XDR_NS}}}pic") is not None:
                    index = list(drawing_root).index(anchor)
                    drawing_root.remove(anchor)
                    for copy_index, target_row in enumerate(target_rows):
                        copied_anchor = _contained_quotation_product_anchor(
                            anchor,
                            target_row=target_row,
                            sheet_root=sheet_root,
                        )
                        if copy_index:
                            non_visual = copied_anchor.find(
                                f".//{{{_XDR_NS}}}cNvPr"
                            )
                            if non_visual is None:
                                raise ValueError(
                                    "Imagen importada sin identidad Quotation"
                                )
                            base_name = (
                                non_visual.attrib.get("name", "").strip()
                                or "Imagen importada"
                            )
                            non_visual.attrib.update(
                                {
                                    "id": str(next_copied_picture_id),
                                    "name": (
                                        f"{base_name} copia "
                                        f"{next_copied_picture_id}"
                                    ),
                                }
                            )
                            next_copied_picture_id += 1
                        drawing_root.insert(index + copy_index, copied_anchor)
                    continue
                target_row = target_rows[0]
                row_delta = target_row - source_row
            elif content_last_row < source_row <= last_populated_row:
                row_delta = footer_delta
            else:
                row_delta = 0
            if row_delta:
                for marker_name in ("from", "to"):
                    target_marker = anchor.find(f"{{{_XDR_NS}}}{marker_name}")
                    target_row_node = (
                        target_marker.find(f"{{{_XDR_NS}}}row")
                        if target_marker is not None
                        else None
                    )
                    if target_row_node is None or target_row_node.text is None:
                        continue
                    shifted = int(target_row_node.text) + row_delta
                    if shifted < 0:
                        raise ValueError(
                            "Imagen fuera de rango al desplazar Quotation"
                        )
                    target_row_node.text = str(shifted)
    additions: dict[str, bytes] = {}
    used_relationship_ids = {
        item.attrib.get("Id", "")
        for item in drawing_rels.findall(f"{{{_PKG_REL_NS}}}Relationship")
    }
    picture_id = _next_picture_id(drawing_root)
    for sequence, (row, line) in enumerate(images, start=1):
        extension, content_type = _quotation_media_extension(line)
        media_part = f"xl/media/project_quotation_{sequence:04d}.{extension}"
        while media_part in package.parts or media_part in additions:
            sequence += 1
            media_part = f"xl/media/project_quotation_{sequence:04d}.{extension}"
        relationship_id = _next_relationship_id(used_relationship_ids)
        used_relationship_ids.add(relationship_id)
        ET.SubElement(
            drawing_rels,
            f"{{{_PKG_REL_NS}}}Relationship",
            {
                "Id": relationship_id,
                "Type": _stable_image_relationship_type(drawing_rels),
                "Target": posixpath.relpath(
                    media_part,
                    posixpath.dirname(drawing_part),
                ),
            },
        )
        assert line.image_content is not None
        with PILImage.open(BytesIO(line.image_content)) as image:
            width, height = image.size
            image.verify()
        (
            cell_width,
            cell_height,
            box_left,
            box_top,
            box_width,
            box_height,
        ) = _quotation_image_box(sheet_root, row)
        anchor = _product_image_anchor(
            relationship_id=relationship_id,
            picture_id=picture_id,
            target_row=row,
            width=width,
            height=height,
            name=f"Imagen de Proyecto {sequence:04d}",
            cell_width=cell_width,
            cell_height=cell_height,
            box=(box_left, box_top, box_width, box_height),
        )
        marker = anchor.find(f"{{{_XDR_NS}}}from")
        column = marker.find(f"{{{_XDR_NS}}}col") if marker is not None else None
        if column is None:
            raise ValueError("Ancla de imagen invalida para Quotation")
        column.text = "2"
        drawing_root.append(anchor)
        picture_id += 1
        additions[media_part] = line.image_content
        _ensure_image_content_type(content_types, extension, content_type)
    replacements = {
        drawing_part: ET.tostring(
            drawing_root,
            encoding="utf-8",
            xml_declaration=True,
        ),
        drawing_rels_part: ET.tostring(
            drawing_rels,
            encoding="utf-8",
            xml_declaration=True,
        ),
        "[Content_Types].xml": ET.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        ),
    }
    return replacements, additions


def _augment_original_quotation(
    source: bytes,
    lines: Sequence[_OfficialPresentationLine],
) -> tuple[bytes, dict[str, int], dict[str, Decimal]]:
    """Conserva la Quotation importada y agrega al final las líneas del catálogo."""

    if not isinstance(source, bytes):
        raise TypeError("La Quotation original debe recibirse como bytes")
    normalized = _normalized_quotation_snapshot(source)
    package = XlsxPackage.from_bytes(normalized)
    try:
        sheet_part = package.sheet_part("Quotation")
    except KeyError as error:
        raise ValueError("La fuente importada no contiene Quotation") from error
    sheet_root, sheet_namespace_prefixes = _parse_worksheet_document(
        package.parts[sheet_part]
    )
    shared_strings = package.shared_strings()
    layout = _quotation_content_rows(sheet_root, shared_strings)
    sheet_data = sheet_root.find(f"{{{_SHEET_NS}}}sheetData")
    assert sheet_data is not None
    source_rows = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{_SHEET_NS}}}row")
    }
    columns = _quotation_column_layout(source_rows, layout, shared_strings)
    absolute_sidecars = {
        row_number: tuple(
            deepcopy(cell)
            for cell in row.findall(f"{{{_SHEET_NS}}}c")
            if _quotation_cell_column(cell) > columns.visible_end
        )
        for row_number, row in source_rows.items()
        if layout.first_row <= row_number <= layout.last_populated_row
    }
    absolute_sidecars = {
        row_number: cells
        for row_number, cells in absolute_sidecars.items()
        if cells
    }
    section_style_row = source_rows[layout.section_rows[-1]]
    product_style_row = source_rows[layout.product_rows[-1]]
    quotation_rows: dict[str, int] = {}
    quotation_rates: dict[str, Decimal] = {}
    grouped_lines: OrderedDict[
        str, tuple[str, list[_OfficialPresentationLine]]
    ] = OrderedDict()
    for line in lines:
        if line.item_key in quotation_rows:
            raise ValueError("Linea de Proyecto duplicada al ampliar Quotation")
        section_key = safe_excel_text(line.section_id).strip()
        if not section_key:
            section_key = "title:" + " ".join(line.section_title.split()).casefold()
        title, grouped = grouped_lines.setdefault(
            section_key,
            (line.section_title, []),
        )
        if title != line.section_title:
            raise ValueError("Titulo de seccion inconsistente al ampliar Quotation")
        grouped.append(line)
    if not grouped_lines:
        raise ValueError("Proyecto sin lineas para ampliar Quotation")

    available_sections: dict[str, list[int]] = {}
    for row_number in layout.section_rows:
        title = _quotation_cell_text(
            _quotation_row_cell(source_rows[row_number], "A"),
            shared_strings,
        )
        key = " ".join(title.lstrip("-").split()).casefold()
        available_sections.setdefault(key, []).append(row_number)

    planned_rows: list[ET.Element] = []
    planned_merge_sources: list[tuple[int, int]] = []
    source_product_targets: dict[int, list[int]] = {}
    product_targets: list[int] = []
    product_index = 0
    cursor = layout.first_row
    added_images: list[tuple[int, _OfficialPresentationLine]] = []
    for _section_id, (section_title, grouped) in grouped_lines.items():
        title_key = " ".join(section_title.split()).casefold()
        matching_sections = available_sections.get(title_key, [])
        section_source_row = (
            matching_sections.pop(0)
            if matching_sections
            else int(section_style_row.attrib["r"])
        )
        section_target_row = cursor
        section = _move_quotation_row(
            source_rows[section_source_row],
            section_target_row,
        )
        _remove_quotation_sidecars(section, columns.visible_end)
        _set_inline_string(
            section,
            f"A{section_target_row}",
            "- " + safe_excel_text(section_title),
        )
        planned_merge_sources.append((section_source_row, section_target_row))
        planned_rows.append(section)
        cursor += 1
        first_product_row = cursor
        for line in grouped:
            product_index += 1
            source_product_row = (
                line.source_row
                if line.origin == "imported"
                and line.source_row in layout.product_rows
                else None
            )
            if source_product_row is not None:
                product_source = source_rows[source_product_row]
                source_product_targets.setdefault(source_product_row, []).append(
                    cursor
                )
            else:
                product_source = product_style_row
            product_source_row = int(product_source.attrib["r"])
            product = _move_quotation_row(product_source, cursor)
            _remove_quotation_sidecars(product, columns.visible_end)
            _set_number(product, f"A{cursor}", Decimal(product_index))
            _set_inline_string(product, f"B{cursor}", safe_excel_text(line.name))
            if source_product_row is None:
                _set_inline_string(
                    product,
                    f"{columns.description}{cursor}",
                    safe_excel_text(line.description),
                )
            _set_inline_string(
                product,
                f"{columns.dimension}{cursor}",
                safe_excel_text(line.dimensions),
            )
            _set_number(product, f"{columns.quantity}{cursor}", line.quantity)
            _set_number(product, f"{columns.volume}{cursor}", line.m3)
            _set_formula(
                product,
                f"{columns.total_volume}{cursor}",
                (
                    f"={columns.quantity}{cursor}"
                    f"*{columns.volume}{cursor}"
                ),
            )
            if line.converted_cost is None:
                _set_inline_string(
                    product,
                    f"{columns.unit_price}{cursor}",
                    "Por confirmar",
                )
                _set_inline_string(
                    product,
                    f"{columns.total_price}{cursor}",
                    "",
                )
            else:
                _set_number(
                    product,
                    f"{columns.unit_price}{cursor}",
                    line.converted_cost,
                )
                _set_formula(
                    product,
                    f"{columns.total_price}{cursor}",
                    (
                        f"={columns.quantity}{cursor}"
                        f"*{columns.unit_price}{cursor}"
                    ),
                )
            if source_product_row is None:
                for column in ("C", columns.color, columns.remark):
                    _set_inline_string(product, f"{column}{cursor}", "")
            planned_merge_sources.append((product_source_row, cursor))
            planned_rows.append(product)
            if source_product_row is None and line.image_content is not None:
                added_images.append((cursor, line))
            quotation_rows[line.item_key] = cursor
            quotation_rates[line.item_key] = Decimal("1")
            product_targets.append(cursor)
            cursor += 1
        last_product_row = cursor - 1
        _set_inline_string(
            section,
            f"{columns.total_volume}{section_target_row}",
            f"- {safe_excel_text(section_title)} Tot.Price :",
        )
        _set_formula(
            section,
            f"{columns.unit_price}{section_target_row}",
            (
                f"=SUM({columns.total_price}{first_product_row}:"
                f"{columns.total_price}{last_product_row})"
            ),
        )

    new_last_content_row = cursor - 1
    footer_delta = new_last_content_row - layout.last_row
    footer_rows: list[ET.Element] = []
    footer_targets: dict[int, ET.Element] = {}
    for source_row_number in range(
        layout.last_row + 1,
        layout.last_populated_row + 1,
    ):
        source_row = source_rows.get(source_row_number)
        if source_row is None:
            continue
        moved = _move_quotation_row(
            source_row,
            source_row_number + footer_delta,
        )
        _remove_quotation_sidecars(moved, columns.visible_end)
        footer_rows.append(moved)
        footer_targets[source_row_number] = moved

    total_source_row = next(
        (
            row_number
            for row_number in footer_targets
            if _quotation_row_formula(
                source_rows[row_number],
                columns.total_volume,
            )
            is not None
            and _quotation_row_formula(
                source_rows[row_number],
                columns.total_price,
            )
            is not None
        ),
        None,
    )
    if total_source_row is not None:
        total_row = footer_targets[total_source_row]
        target_row = total_source_row + footer_delta
        first_product_target = min(product_targets)
        last_product_target = max(product_targets)
        _set_formula(
            total_row,
            f"{columns.total_volume}{target_row}",
            (
                f"=SUM({columns.total_volume}{first_product_target}:"
                f"{columns.total_volume}{last_product_target})"
            ),
        )
        _set_formula(
            total_row,
            f"{columns.total_price}{target_row}",
            (
                f"=SUM({columns.total_price}{first_product_target}:"
                f"{columns.total_price}{last_product_target})"
            ),
        )

    planned_target_numbers = {int(row.attrib["r"]) for row in planned_rows}
    footer_target_numbers = {int(row.attrib["r"]) for row in footer_rows}
    retained_rows = [
        row
        for row in sheet_data.findall(f"{{{_SHEET_NS}}}row")
        if (
            int(row.attrib["r"]) < layout.first_row
            or int(row.attrib["r"]) > layout.last_populated_row
        )
        and int(row.attrib["r"]) not in planned_target_numbers
        and int(row.attrib["r"]) not in footer_target_numbers
    ]
    sheet_data[:] = sorted(
        (*retained_rows, *planned_rows, *footer_rows),
        key=lambda row: int(row.attrib["r"]),
    )
    _restore_quotation_sidecars(
        sheet_data,
        absolute_sidecars,
        source_rows,
    )

    merge_cells = sheet_root.find(f"{{{_SHEET_NS}}}mergeCells")
    if merge_cells is not None:
        original_merges = [
            item.attrib.get("ref", "")
            for item in merge_cells.findall(f"{{{_SHEET_NS}}}mergeCell")
        ]
        output_merges: list[str] = []
        content_merge_map: dict[int, list[int]] = {}
        for source_row_number, target_row_number in planned_merge_sources:
            content_merge_map.setdefault(source_row_number, []).append(
                target_row_number
            )
        for reference in original_merges:
            try:
                cell_range = CellRange(reference)
            except (TypeError, ValueError):
                raise ValueError("Merge invalido en Quotation original") from None
            if cell_range.max_row < layout.first_row:
                output_merges.append(reference)
                continue
            if cell_range.min_row > layout.last_row:
                if cell_range.min_row <= layout.last_populated_row:
                    if cell_range.max_row > layout.last_populated_row:
                        raise ValueError(
                            "Merge cruza el pie dinamico de Quotation"
                        )
                    output_merges.append(
                        _shift_quotation_merge(reference, footer_delta)
                    )
                else:
                    output_merges.append(reference)
                continue
            if (
                cell_range.min_row < layout.first_row
                or cell_range.max_row > layout.last_row
                or cell_range.min_row != cell_range.max_row
            ):
                raise ValueError(
                    "Merge cruza el bloque dinamico de Quotation"
                )
            for target_row in content_merge_map.get(cell_range.min_row, ()):
                output_merges.append(
                    _shift_quotation_merge(
                        reference,
                        target_row - cell_range.min_row,
                    )
                )
        unique_merges = list(dict.fromkeys(output_merges))
        for item in list(merge_cells):
            merge_cells.remove(item)
        for reference in unique_merges:
            ET.SubElement(
                merge_cells,
                f"{{{_SHEET_NS}}}mergeCell",
                {"ref": reference},
            )
        merge_cells.attrib["count"] = str(len(unique_merges))

    dimension = sheet_root.find(f"{{{_SHEET_NS}}}dimension")
    if dimension is not None:
        try:
            dimension_range = CellRange(dimension.attrib["ref"])
        except (KeyError, TypeError, ValueError):
            raise ValueError("Dimension invalida en Quotation original") from None
        target_last_populated = layout.last_populated_row + footer_delta
        if target_last_populated > dimension_range.max_row:
            dimension.attrib["ref"] = (
                f"{get_column_letter(dimension_range.min_col)}"
                f"{dimension_range.min_row}:"
                f"{get_column_letter(max(dimension_range.max_col, columns.visible_end))}"
                f"{target_last_populated}"
            )
    drawing_replacements, media_additions = _augment_quotation_drawing(
        package,
        sheet_part,
        sheet_root,
        tuple(added_images),
        content_first_row=layout.first_row,
        content_last_row=layout.last_row,
        last_populated_row=layout.last_populated_row,
        content_row_map=source_product_targets,
        footer_delta=footer_delta,
    )
    replacements = {
        sheet_part: _xml_bytes(sheet_root, sheet_namespace_prefixes),
        **drawing_replacements,
    }
    snapshot = package.to_bytes(
        PackageMutation(
            replacements=replacements,
            additions=media_additions,
        )
    )
    XlsxPackage.from_bytes(snapshot)
    return snapshot, quotation_rows, quotation_rates


_QUOTATION_TRANSFORMATION_COLUMN = 4
_QUOTATION_TRANSFORMATION_HEADER = "Trasformacion"
_QUOTATION_DRAWING_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
_A1_CELL_TOKEN = re.compile(
    r"(?P<column_absolute>\$?)(?P<column>[A-Z]+)"
    r"(?P<row_absolute>\$?)(?P<row>[1-9][0-9]*)"
)
_QUOTATION_FORMULA_COLUMN = re.compile(r"(?P<absolute>\$?)(?P<column>[A-Z]+)")


def _shift_a1_cell_for_quotation_column(cell: str) -> str:
    match = _A1_CELL_TOKEN.fullmatch(cell)
    if match is None:
        raise ValueError("Referencia A1 inválida al ampliar Quotation")
    column = column_index_from_string(match.group("column"))
    if column >= _QUOTATION_TRANSFORMATION_COLUMN:
        column += 1
    return (
        f"{match.group('column_absolute')}{get_column_letter(column)}"
        f"{match.group('row_absolute')}{match.group('row')}"
    )


def _shift_quotation_column_range(reference: str) -> str:
    try:
        ranges = []
        for token in reference.split():
            start, separator, end = token.partition(":")
            shifted_start = _shift_a1_cell_for_quotation_column(start)
            shifted_end = (
                _shift_a1_cell_for_quotation_column(end)
                if separator
                else ""
            )
            ranges.append(
                f"{shifted_start}:{shifted_end}"
                if separator
                else shifted_start
            )
        if not ranges:
            raise ValueError
    except (TypeError, ValueError):
        raise ValueError(
            "Referencia inválida al insertar columna en Quotation"
        ) from None
    return " ".join(ranges)


def _shift_quotation_formula_range(value: str) -> str:
    sheet_prefix = ""
    reference = value
    if "!" in value:
        raw_sheet, reference = value.rsplit("!", 1)
        normalized_sheet = raw_sheet
        if (
            len(normalized_sheet) >= 2
            and normalized_sheet.startswith("'")
            and normalized_sheet.endswith("'")
        ):
            normalized_sheet = normalized_sheet[1:-1].replace("''", "'")
        if (
            "[" in normalized_sheet
            or normalized_sheet.casefold() != "quotation"
        ):
            return value
        sheet_prefix = raw_sheet + "!"

    shifted_parts: list[str] = []
    for part in reference.split(":"):
        if _A1_CELL_TOKEN.fullmatch(part):
            shifted_parts.append(_shift_a1_cell_for_quotation_column(part))
            continue
        column_match = _QUOTATION_FORMULA_COLUMN.fullmatch(part)
        if column_match is None:
            return value
        try:
            column = column_index_from_string(column_match.group("column"))
        except ValueError:
            return value
        if column > 16_384:
            return value
        if column >= _QUOTATION_TRANSFORMATION_COLUMN:
            column += 1
        shifted_parts.append(
            f"{column_match.group('absolute')}{get_column_letter(column)}"
        )
    return sheet_prefix + ":".join(shifted_parts)


def _shift_quotation_formula(formula: str) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        raise ValueError("Fórmula inválida al insertar columna en Quotation")
    try:
        tokens = Tokenizer(formula).items
    except (TypeError, ValueError):
        raise ValueError(
            "Fórmula inválida al insertar columna en Quotation"
        ) from None
    return "=" + "".join(
        _shift_quotation_formula_range(token.value)
        if token.type == "OPERAND" and token.subtype == "RANGE"
        else token.value
        for token in tokens
    )


def _shift_quotation_formula_text(formula: str) -> str:
    """Traslada una fórmula OOXML conservando si incluía el signo igual."""

    if not isinstance(formula, str) or not formula.strip():
        raise ValueError("Fórmula vacía al insertar columna en Quotation")
    has_equals = formula.startswith("=")
    shifted = _shift_quotation_formula(formula if has_equals else "=" + formula)
    return shifted if has_equals else shifted[1:]


def _shift_quotation_column_definitions(sheet_root: ET.Element) -> None:
    columns = sheet_root.find(f"{{{_SHEET_NS}}}cols")
    if columns is None:
        return
    shifted: list[ET.Element] = []
    for definition in list(columns):
        if definition.tag != f"{{{_SHEET_NS}}}col":
            shifted.append(deepcopy(definition))
            continue
        try:
            minimum = int(definition.attrib["min"])
            maximum = int(definition.attrib["max"])
        except (KeyError, TypeError, ValueError):
            raise ValueError("Definición de columna inválida en Quotation") from None
        if minimum <= 0 or maximum < minimum:
            raise ValueError("Definición de columna inválida en Quotation")
        if maximum < _QUOTATION_TRANSFORMATION_COLUMN:
            shifted.append(deepcopy(definition))
            continue
        if minimum < _QUOTATION_TRANSFORMATION_COLUMN:
            before = deepcopy(definition)
            before.attrib["max"] = str(_QUOTATION_TRANSFORMATION_COLUMN - 1)
            shifted.append(before)
        inserted = deepcopy(definition)
        inserted.attrib["min"] = str(_QUOTATION_TRANSFORMATION_COLUMN)
        inserted.attrib["max"] = str(_QUOTATION_TRANSFORMATION_COLUMN)
        if minimum <= _QUOTATION_TRANSFORMATION_COLUMN <= maximum:
            shifted.append(inserted)
        moved = deepcopy(definition)
        moved.attrib["min"] = str(
            max(minimum, _QUOTATION_TRANSFORMATION_COLUMN) + 1
        )
        moved.attrib["max"] = str(maximum + 1)
        shifted.append(moved)
    columns[:] = sorted(
        shifted,
        key=lambda item: (
            int(item.attrib.get("min", "0")),
            int(item.attrib.get("max", "0")),
        ),
    )


def _shift_quotation_worksheet_references(sheet_root: ET.Element) -> None:
    reference_attributes = {
        "dimension": ("ref",),
        "mergeCell": ("ref",),
        "autoFilter": ("ref",),
        "sortState": ("ref",),
        "hyperlink": ("ref",),
        "conditionalFormatting": ("sqref",),
        "dataValidation": ("sqref",),
        "ignoredError": ("sqref",),
        "protectedRange": ("sqref",),
        "selection": ("activeCell", "sqref"),
        "pane": ("topLeftCell",),
    }
    for auto_filter in (
        element
        for element in sheet_root.iter()
        if element.tag.rsplit("}", 1)[-1] == "autoFilter"
    ):
        reference = auto_filter.attrib.get("ref")
        if not reference:
            continue
        try:
            filter_range = CellRange(reference)
        except (TypeError, ValueError):
            raise ValueError("Rango de filtro inválido en Quotation") from None
        if not (
            filter_range.min_col < _QUOTATION_TRANSFORMATION_COLUMN
            <= filter_range.max_col
        ):
            continue
        insertion_index = (
            _QUOTATION_TRANSFORMATION_COLUMN - filter_range.min_col
        )
        for filter_column in (
            element
            for element in auto_filter.iter()
            if element.tag.rsplit("}", 1)[-1] == "filterColumn"
        ):
            try:
                column_id = int(filter_column.attrib["colId"])
            except (KeyError, TypeError, ValueError):
                raise ValueError(
                    "Columna de filtro inválida en Quotation"
                ) from None
            if column_id >= insertion_index:
                filter_column.attrib["colId"] = str(column_id + 1)
    for element in sheet_root.iter():
        local_name = element.tag.rsplit("}", 1)[-1]
        namespace = (
            element.tag[1:].split("}", 1)[0]
            if element.tag.startswith("{")
            else ""
        )
        for attribute in reference_attributes.get(local_name, ()):
            value = element.attrib.get(attribute)
            if value:
                element.attrib[attribute] = _shift_quotation_column_range(value)
        if local_name == "sqref" and element.text and element.text.strip():
            element.text = _shift_quotation_column_range(element.text)
        if (
            local_name in {"formula", "formula1", "formula2"}
            and element.text
            and element.text.strip()
        ):
            element.text = _shift_quotation_formula_text(element.text)
        elif (
            local_name == "f"
            and namespace != _SHEET_NS
            and element.text
            and element.text.strip()
        ):
            element.text = _shift_quotation_formula_text(element.text)


def _shift_quotation_formula_cell(cell: ET.Element) -> None:
    coordinate = cell.attrib.get("r", "")
    match = re.fullmatch(
        r"(?P<column>[A-Z]+)(?P<row>[1-9][0-9]*)",
        coordinate,
    )
    if match is None:
        raise ValueError("Coordenada inválida al insertar columna en Quotation")
    column = column_index_from_string(match.group("column"))
    formula = cell.find(f"{{{_SHEET_NS}}}f")
    if formula is not None:
        if formula.text:
            formula.text = _shift_quotation_formula("=" + formula.text)[1:]
        reference = formula.attrib.get("ref")
        if reference:
            formula.attrib["ref"] = _shift_quotation_column_range(reference)
        cached = cell.find(f"{{{_SHEET_NS}}}v")
        if cached is not None:
            cell.remove(cached)
    if column >= _QUOTATION_TRANSFORMATION_COLUMN:
        cell.attrib["r"] = (
            f"{get_column_letter(column + 1)}{match.group('row')}"
        )


def _insert_quotation_text_cell(
    row: ET.Element,
    coordinate: str,
    value: str,
) -> None:
    row_number = row.attrib.get("r", "")
    if not row_number.isdigit() or not coordinate.endswith(row_number):
        raise ValueError("Fila inválida al escribir transformación Quotation")
    style_source = _quotation_row_cell(row, "E")
    if style_source is not None and _quotation_row_cell(row, "D") is None:
        clone = deepcopy(style_source)
        clone.attrib["r"] = coordinate
        for child in list(clone):
            clone.remove(child)
        clone.attrib.pop("t", None)
        row.append(clone)
    _set_inline_string(row, coordinate, safe_excel_text(value))
    row[:] = sorted(
        row,
        key=lambda item: (
            _quotation_cell_column(item)
            if item.tag == f"{{{_SHEET_NS}}}c"
            else 1_000_000
        ),
    )


def _shift_quotation_related_parts(
    package: XlsxPackage,
    sheet_part: str,
    sheet_root: ET.Element,
) -> dict[str, bytes]:
    rels_part = relationship_part_name(sheet_part)
    if rels_part not in package.parts:
        return {}
    rels_root = ET.fromstring(package.parts[rels_part])
    relationships = {
        item.attrib.get("Id"): item
        for item in rels_root.findall(f"{{{_PKG_REL_NS}}}Relationship")
    }
    replacements: dict[str, bytes] = {}
    for drawing in sheet_root.findall(f"{{{_SHEET_NS}}}drawing"):
        relationship_id = drawing.attrib.get(f"{{{_OFFICE_REL_NS}}}id")
        relationship = relationships.get(relationship_id)
        if relationship is None:
            raise ValueError("Dibujo Quotation sin relación")
        drawing_part = resolve_internal_target(
            sheet_part,
            relationship.attrib.get("Target", ""),
        )
        drawing_root = ET.fromstring(package.parts[drawing_part])
        for marker_name in ("from", "to"):
            for marker in drawing_root.findall(
                f".//{{{_QUOTATION_DRAWING_NS}}}{marker_name}"
            ):
                column = marker.find(f"{{{_QUOTATION_DRAWING_NS}}}col")
                if column is None or column.text is None:
                    raise ValueError("Ancla de dibujo inválida en Quotation")
                try:
                    index = int(column.text)
                except ValueError:
                    raise ValueError(
                        "Ancla de dibujo inválida en Quotation"
                    ) from None
                if index >= _QUOTATION_TRANSFORMATION_COLUMN - 1:
                    column.text = str(index + 1)
        replacements[drawing_part] = ET.tostring(
            drawing_root,
            encoding="utf-8",
            xml_declaration=True,
        )
    for relationship in relationships.values():
        relationship_type = relationship.attrib.get("Type")
        if relationship_type in relationship_type_uris("comments"):
            comments_part = resolve_internal_target(
                sheet_part,
                relationship.attrib.get("Target", ""),
            )
            comments_root = ET.fromstring(package.parts[comments_part])
            for comment in comments_root.iter():
                if comment.tag.rsplit("}", 1)[-1] != "comment":
                    continue
                reference = comment.attrib.get("ref")
                if reference:
                    comment.attrib["ref"] = _shift_quotation_column_range(
                        reference
                    )
            replacements[comments_part] = ET.tostring(
                comments_root,
                encoding="utf-8",
                xml_declaration=True,
            )
            continue
        if relationship_type in relationship_type_uris("vmlDrawing"):
            vml_part = resolve_internal_target(
                sheet_part,
                relationship.attrib.get("Target", ""),
            )
            vml_root = ET.fromstring(package.parts[vml_part])
            for element in vml_root.iter():
                local_name = element.tag.rsplit("}", 1)[-1]
                if local_name == "Column" and element.text is not None:
                    try:
                        column_index = int(element.text)
                    except ValueError:
                        raise ValueError(
                            "Columna VML inválida en Quotation"
                        ) from None
                    if column_index >= _QUOTATION_TRANSFORMATION_COLUMN - 1:
                        element.text = str(column_index + 1)
                elif local_name == "Anchor" and element.text:
                    parts = [part.strip() for part in element.text.split(",")]
                    if len(parts) != 8:
                        raise ValueError("Ancla VML inválida en Quotation")
                    try:
                        coordinates = [int(part) for part in parts]
                    except ValueError:
                        raise ValueError(
                            "Ancla VML inválida en Quotation"
                        ) from None
                    for position in (0, 4):
                        if (
                            coordinates[position]
                            >= _QUOTATION_TRANSFORMATION_COLUMN - 1
                        ):
                            coordinates[position] += 1
                    element.text = ", ".join(str(value) for value in coordinates)
            replacements[vml_part] = ET.tostring(
                vml_root,
                encoding="utf-8",
                xml_declaration=True,
            )
            continue
        if relationship_type not in relationship_type_uris("table"):
            continue
        table_part = resolve_internal_target(
            sheet_part,
            relationship.attrib.get("Target", ""),
        )
        table_root = ET.fromstring(package.parts[table_part])
        table_reference = table_root.attrib.get("ref")
        try:
            table_range = (
                CellRange(table_reference)
                if table_reference
                else None
            )
        except (TypeError, ValueError):
            raise ValueError("Rango de tabla inválido en Quotation") from None
        for element in table_root.iter():
            if element.tag.rsplit("}", 1)[-1] in {
                "table",
                "autoFilter",
                "sortState",
                "sortCondition",
            }:
                reference = element.attrib.get("ref")
                if reference:
                    element.attrib["ref"] = _shift_quotation_column_range(
                        reference
                    )
            local_name = element.tag.rsplit("}", 1)[-1]
            if (
                local_name in {"calculatedColumnFormula", "totalsRowFormula"}
                and element.text
                and element.text.strip()
            ):
                element.text = _shift_quotation_formula_text(element.text)
        if (
            table_range is not None
            and table_range.min_col < _QUOTATION_TRANSFORMATION_COLUMN
            <= table_range.max_col
        ):
            insertion_index = (
                _QUOTATION_TRANSFORMATION_COLUMN - table_range.min_col
            )
            for filter_column in (
                element
                for element in table_root.iter()
                if element.tag.rsplit("}", 1)[-1] == "filterColumn"
            ):
                try:
                    column_id = int(filter_column.attrib["colId"])
                except (KeyError, TypeError, ValueError):
                    raise ValueError(
                        "Columna de filtro de tabla inválida"
                    ) from None
                if column_id >= insertion_index:
                    filter_column.attrib["colId"] = str(column_id + 1)
            table_columns = next(
                (
                    element
                    for element in table_root
                    if element.tag.rsplit("}", 1)[-1] == "tableColumns"
                ),
                None,
            )
            if table_columns is None:
                raise ValueError("Tabla Quotation sin definición de columnas")
            columns = [
                element
                for element in table_columns
                if element.tag.rsplit("}", 1)[-1] == "tableColumn"
            ]
            expected_columns = table_range.max_col - table_range.min_col + 1
            if len(columns) != expected_columns:
                raise ValueError("Definición de tabla Quotation inconsistente")
            column_ids: list[int] = []
            for column in columns:
                try:
                    column_id = int(column.attrib["id"])
                except (KeyError, TypeError, ValueError):
                    raise ValueError(
                        "Identificador de columna de tabla inválido"
                    ) from None
                if column_id <= 0 or column_id in column_ids:
                    raise ValueError(
                        "Identificador de columna de tabla inválido"
                    )
                column_ids.append(column_id)
            namespace = table_columns.tag.rsplit("}", 1)[0].lstrip("{")
            inserted = ET.Element(
                f"{{{namespace}}}tableColumn",
                {
                    "id": str(max(column_ids, default=0) + 1),
                    "name": _QUOTATION_TRANSFORMATION_HEADER,
                },
            )
            table_columns.insert(insertion_index, inserted)
            table_columns.attrib["count"] = str(len(columns) + 1)
        replacements[table_part] = ET.tostring(
            table_root,
            encoding="utf-8",
            xml_declaration=True,
        )
    return replacements


def _shift_quotation_defined_names(package: XlsxPackage) -> bytes | None:
    root = ET.fromstring(package.parts["xl/workbook.xml"])
    changed = False

    for defined_name in root.findall(
        f"{{{_SHEET_NS}}}definedNames/{{{_SHEET_NS}}}definedName"
    ):
        if defined_name.text:
            shifted = _shift_quotation_formula_text(defined_name.text)
            if shifted != defined_name.text:
                defined_name.text = shifted
                changed = True
    if not changed:
        return None
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _insert_quotation_transformation_column(
    source: bytes,
    quotation_rows: Mapping[str, int],
    processed_descriptions: Mapping[str, str],
) -> bytes:
    if not isinstance(source, bytes):
        raise TypeError("La Quotation ampliada debe recibirse como bytes")
    if set(quotation_rows) != set(processed_descriptions):
        raise ValueError("Descripciones procesadas de Quotation incompletas")
    package = XlsxPackage.from_bytes(source)
    sheet_part = package.sheet_part("Quotation")
    sheet_root, namespace_prefixes = _parse_worksheet_document(
        package.parts[sheet_part]
    )
    sheet_data = sheet_root.find(f"{{{_SHEET_NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("Quotation ampliada sin sheetData")
    shared_strings = package.shared_strings()
    layout = _quotation_content_rows(sheet_root, shared_strings)
    rows = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{_SHEET_NS}}}row")
    }
    column_layout = _quotation_column_layout(rows, layout, shared_strings)

    def write_processed_descriptions() -> None:
        used_rows: set[int] = set()
        for item_key, row_number in quotation_rows.items():
            if (
                type(row_number) is not int
                or row_number not in rows
                or row_number in used_rows
            ):
                raise ValueError("Fila de transformación Quotation inválida")
            used_rows.add(row_number)
            _insert_quotation_text_cell(
                rows[row_number],
                f"D{row_number}",
                processed_descriptions[item_key],
            )

    if column_layout.transformed:
        write_processed_descriptions()
        result = package.to_bytes(
            PackageMutation(
                replacements={
                    sheet_part: _xml_bytes(sheet_root, namespace_prefixes),
                }
            )
        )
        XlsxPackage.from_bytes(result)
        return result

    for row in rows.values():
        for cell in reversed(row.findall(f"{{{_SHEET_NS}}}c")):
            _shift_quotation_formula_cell(cell)
        row[:] = sorted(
            row,
            key=lambda item: (
                _quotation_cell_column(item)
                if item.tag == f"{{{_SHEET_NS}}}c"
                else 1_000_000
            ),
        )
    header_row = rows.get(layout.first_row - 1)
    if header_row is None:
        raise ValueError("Quotation sin encabezado visible para transformación")
    _insert_quotation_text_cell(
        header_row,
        f"D{layout.first_row - 1}",
        _QUOTATION_TRANSFORMATION_HEADER,
    )
    write_processed_descriptions()
    _shift_quotation_column_definitions(sheet_root)
    _shift_quotation_worksheet_references(sheet_root)
    replacements = {
        sheet_part: _xml_bytes(sheet_root, namespace_prefixes),
        **_shift_quotation_related_parts(package, sheet_part, sheet_root),
    }
    workbook_xml = _shift_quotation_defined_names(package)
    if workbook_xml is not None:
        replacements["xl/workbook.xml"] = workbook_xml
    result = package.to_bytes(PackageMutation(replacements=replacements))
    XlsxPackage.from_bytes(result)
    return result


def _normalized_quotation_source(path: Path) -> bytes:
    """Devuelve un snapshot OOXML auditado sin mutar el XLSX recibido."""

    package = XlsxPackage.read(Path(path), audit=False)
    return _normalized_quotation_package(package)


def _normalized_quotation_snapshot(content: bytes) -> bytes:
    """Normaliza un snapshot en memoria producido al ampliar Quotation."""

    package = XlsxPackage.from_bytes(content, audit=False)
    return _normalized_quotation_package(package)


def _normalized_quotation_package(package: XlsxPackage) -> bytes:
    sanitized_parts = _strip_external_image_links_with_embedded_fallback(
        package.parts
    )
    names = set(sanitized_parts)
    replacements: dict[str, bytes] = {
        name: content
        for name, content in sanitized_parts.items()
        if content != package.parts[name]
    }
    for name, bounded_payload in sanitized_parts.items():
        payload = bounded_payload
        entry_changed = False
        if name.endswith(".rels"):
            root = ET.fromstring(payload)
            owner = _engine_relationship_owner(name)
            owner_directory = posixpath.dirname(owner) if owner else ""
            for relationship in root.findall(
                f"{{{_PKG_REL_NS}}}Relationship"
            ):
                if relationship.attrib.get("TargetMode", "").casefold() == "external":
                    continue
                relationship_target = relationship.attrib.get("Target", "")
                if not relationship_target.startswith("/"):
                    continue
                package_part = relationship_target[1:]
                if (
                    not package_part
                    or package_part.startswith("/")
                    or "\\" in package_part
                    or posixpath.normpath(package_part) != package_part
                    or package_part not in names
                ):
                    raise ValueError("Target OOXML package-rooted invalido")
                relationship.attrib["Target"] = posixpath.relpath(
                    package_part,
                    owner_directory,
                )
                entry_changed = True
            if entry_changed:
                payload = ET.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
        elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
            root = ET.fromstring(payload)
            for cell in root.findall(
                ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"
                "[@t='inlineStr']"
            ):
                inline_strings = cell.findall(
                    "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is"
                )
                if inline_strings:
                    if len(inline_strings) != 1:
                        raise ValueError("Celda inlineStr ambigua")
                    continue
                if list(cell):
                    raise ValueError("Celda inlineStr invalida")
                ET.SubElement(
                    cell,
                    "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is",
                )
                entry_changed = True
            if entry_changed:
                payload = ET.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
        if entry_changed:
            replacements[name] = payload
    if not replacements:
        package.audit()
    snapshot = package.to_bytes(PackageMutation(replacements=replacements))
    XlsxPackage.from_bytes(snapshot)
    return snapshot


def _engine_relationship_owner(rels_name: str) -> str | None:
    if rels_name == "_rels/.rels":
        return None
    marker = "/_rels/"
    if marker not in rels_name or not rels_name.endswith(".rels"):
        raise ValueError(f"Ruta de relaciones OOXML invalida: {rels_name}")
    directory, filename = rels_name.split(marker, 1)
    return posixpath.join(directory, filename.removesuffix(".rels"))


def generate_quote(
    source_path: str | Path,
    output_path: str | Path,
    metadata: dict[str, Any] | None = None,
    template_path: str | Path | None = None,
    *,
    original_quotation_path: object = _ARGUMENT_OMITTED,
    quotation_data_rows: object = _ARGUMENT_OMITTED,
) -> Path:
    metadata = dict(metadata or {})
    source_path = Path(source_path).resolve(strict=True)
    output_path = Path(output_path)
    lexical_output = output_path if output_path.is_absolute() else Path.cwd() / output_path
    official_template = Path(template_path or OFFICIAL_TEMPLATE_PATH).resolve(strict=True)
    _verified_template, resolved_output = _validate_compose_paths(
        official_template,
        lexical_output,
    )

    normalized_source = _normalized_quotation_source(source_path)
    items, _column_map = read_items(BytesIO(normalized_source))
    handed_off_rows = _canonical_handoff_rows(quotation_data_rows)
    if handed_off_rows is None:
        _validate_mixed_catalog_metadata(items, metadata)
    else:
        _validate_authoritative_handoff_metadata(items, metadata)
    contract = load_template_contract(
        _contract_path_for_template(official_template)
    )
    base = XlsxPackage.read(official_template)
    lumbro_prices = _load_lumbro_prices(official_template)
    if original_quotation_path is _ARGUMENT_OMITTED:
        original_source: Path | None = source_path
        normalized_original: bytes | None = normalized_source
    elif original_quotation_path is None:
        original_source = None
        normalized_original = None
    elif isinstance(original_quotation_path, (str, Path)):
        original_source = Path(original_quotation_path).resolve(strict=True)
        normalized_original = (
            normalized_source
            if original_source == source_path
            else _normalized_quotation_source(original_source)
        )
    else:
        raise TypeError("original_quotation_path debe ser Path, str o None")
    image_payloads = _source_product_images(normalized_source)
    lines, needs = _official_presentation_lines(
        items,
        metadata,
        source_path,
        lumbro_prices,
        image_payloads,
    )
    if handed_off_rows is None:
        canonical_rows = _official_canonical_rows(lines, normalized_source)
    else:
        lines, needs, canonical_rows = _bind_authoritative_canonical_rows(
            lines,
            handed_off_rows,
            normalized_source,
            metadata,
        )
    quotation_snapshot, quotation_rows, quotation_rates = _augment_original_quotation(
        normalized_original or normalized_source,
        lines,
    )
    mobiliti = _build_official_mobiliti(
        base,
        lines,
        needs,
        canonical_rows,
        metadata,
        quotation_rows,
        quotation_rates,
    )
    cotizacion, cotizacion_sections = _build_official_cotizacion(
        base,
        lines,
        mobiliti,
        metadata,
        quotation_rows,
    )
    description_language = normalize_description_language(
        metadata.get(
            "description_language",
            metadata.get("idioma_descripcion", "es"),
        )
    )
    processed_descriptions_by_key = {
        line.item_key: _official_cotizacion_description(
            line,
            description_language,
        )
        for line in lines
    }
    for section in cotizacion_sections:
        for product in section.products:
            processed_descriptions_by_key[product.item_key] = product.description
    quotation_snapshot = _insert_quotation_transformation_column(
        quotation_snapshot,
        quotation_rows,
        processed_descriptions_by_key,
    )
    quotation = transplant_quotation(quotation_snapshot, base)
    compose_official_quote(
        ComposeRequest(
            template=official_template,
            output=resolved_output,
            mobiliti=mobiliti,
            cotizacion=cotizacion,
            quotation=quotation,
            quotation_data=build_quotation_data_sheet(canonical_rows),
            contract=contract,
            project_composition=metadata.get("project_context") is not None,
        )
    )
    return output_path
