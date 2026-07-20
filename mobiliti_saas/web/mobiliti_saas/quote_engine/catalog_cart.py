from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
import ipaddress
import os
import re
import socket
import tempfile
import unicodedata
import urllib.request
import warnings
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError


MAX_IMAGE_BYTES = 8 * 1024 * 1024
MAX_CATALOG_IMAGE_PIXELS = 40_000_000
MAX_EXCEL_CELL_TEXT_LENGTH = 32_767
CATALOG_IMAGE_FORMATS = {
    "PNG": ("image/png", ".png"),
    "JPEG": ("image/jpeg", ".jpg"),
    "WEBP": ("image/webp", ".webp"),
}
MAX_COMMERCIAL_QUANTITY = Decimal("1000000")
DEFAULT_QUANTITY_DECIMAL_PLACES = 3
MAX_QUANTITY_DECIMAL_PLACES = 6
MAX_QUANTITY_INTEGER_DIGITS = 7
MAX_QUANTITY_TEXT_LENGTH = 64
WARNING_FILL = "FFF2CC"
OFFICIAL_IMAGE_HOSTS = {
    "offiho_cart": frozenset(
        {
            "offiho.com",
            "www.offiho.com",
            "offiho.com.mx",
            "offihoblack.com",
            "www.offihoblack.com",
            "web-lemon-one-45.vercel.app",
        }
    ),
    "tarkett_cart": frozenset(
        {
            "tarkett.com.mx",
            "www.tarkett.com.mx",
            "tarkett.com.ar",
            "www.tarkett.com.ar",
            "profesional.tarkett.es",
            "media.tarkett-image.com",
            "tarkettnet.com.mx",
            "www.tarkettnet.com.mx",
        }
    ),
}


def create_catalog_quotation_workbook(
    payload: dict[str, Any],
    output_path: str | Path,
    *,
    source_type: str,
    category_label: str,
    image_dir: str | Path | None = None,
    text_transform: Callable[[object], str] | None = None,
) -> Path:
    if payload.get("source_type") != source_type:
        raise ValueError(f"Payload {category_label} invalido")
    items = list(payload.get("items") or [])
    if not items:
        raise ValueError(f"Payload {category_label} sin productos")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    tmp_context = None
    if image_dir is None:
        tmp_context = tempfile.TemporaryDirectory(prefix="catalog_images_")
        images_root = Path(tmp_context.name)
    else:
        images_root = Path(image_dir)
        images_root.mkdir(parents=True, exist_ok=True)

    wb = None
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        write_catalog_quotation_headers(ws)
        ws.cell(8, 1).value = f"- {category_label}"
        ws.cell(8, 1).font = Font(bold=True)

        transform = text_transform or (lambda value: str(value or ""))
        for index, item in enumerate(items, start=1):
            row = index + 8
            write_catalog_quotation_item(
                ws,
                row=row,
                index=index,
                item=item,
                source_type=source_type,
                images_root=images_root,
                text_transform=transform,
            )

        _set_column_widths(ws)
        wb.save(output)
    finally:
        try:
            if wb is not None:
                wb.close()
        finally:
            if tmp_context is not None:
                tmp_context.cleanup()
    return output


def write_catalog_quotation_headers(
    ws, extra_headers: dict[int, str] | None = None
) -> None:
    headers = {
        1: "No.",
        2: "Item",
        3: "Image",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
        11: "URL",
    }
    headers.update(extra_headers or {})
    for col, title in headers.items():
        cell = ws.cell(7, col)
        cell.value = title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2F6B")


def write_catalog_quotation_item(
    ws,
    *,
    row: int,
    index: int,
    item: dict[str, Any],
    source_type: str,
    images_root: Path,
    text_transform: Callable[[object], str],
    image_file_key: str | None = None,
    extra_description_parts: tuple[str, ...] = (),
) -> None:
    code = str(item.get("code") or item.get("sku") or "").strip()
    name = str(item.get("name", "")).strip()
    unit = str(item.get("unit", "")).strip()
    attributes = (
        item.get("attributes") if isinstance(item.get("attributes"), dict) else {}
    )
    dimensions = str(attributes.get("dimensions") or "").strip()
    url = str(item.get("product_url", "") or "").strip()
    description, warning, quantity = catalog_quotation_item_text(
        item,
        index=index,
        source_type=source_type,
        extra_description_parts=extra_description_parts,
    )
    ws.cell(row, 1).value = index
    ws.cell(row, 2).value = text_transform(name)
    ws.cell(row, 4).value = text_transform(description)
    ws.cell(row, 5).value = text_transform(
        dimensions if source_type == "supplier_cart" and dimensions else unit
    )
    ws.cell(row, 7).value = _excel_number(quantity)
    ws.cell(row, 10).value = _excel_number(_decimal(item.get("unit_price", 0)))
    ws.cell(row, 11).value = text_transform(url)
    if warning:
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=WARNING_FILL)
    ws.row_dimensions[row].height = 72
    _add_catalog_image(
        ws,
        row,
        item.get("image_url"),
        images_root,
        code,
        source_type,
        destination_key=image_file_key,
    )


def catalog_quotation_item_text(
    item: dict[str, Any],
    *,
    index: int,
    source_type: str,
    extra_description_parts: tuple[str, ...] = (),
) -> tuple[str, str, Decimal]:
    code = str(item.get("code") or item.get("sku") or "").strip()
    name = str(item.get("name", "")).strip()
    unit = str(item.get("unit", "")).strip()
    url = str(item.get("product_url", "") or "").strip()
    if source_type == "supplier_cart":
        quantity_precision = 6 if _is_square_meter_unit(unit) else 0
    elif source_type == "tarkett_cart":
        quantity_precision = 6
    else:
        quantity_precision = 3
    quantity = parse_commercial_quantity(
        item.get("quantity", 0),
        item_label=code or name or str(index),
        max_decimal_places=quantity_precision,
    )
    description, warning = _description_for_item(
        item,
        code,
        url,
        quantity,
        extra_description_parts=extra_description_parts,
    )
    return description, warning, quantity


def _description_for_item(
    item: dict[str, Any],
    code: str,
    url: str,
    quantity: Decimal,
    *,
    extra_description_parts: tuple[str, ...] = (),
) -> tuple[str, str]:
    code_label = "SKU" if "sku" in item else "Clave"
    parts = [
        str(item.get("description", "") or "").strip(),
        *(str(value or "").strip() for value in extra_description_parts),
        f"{code_label}: {code}" if code else "",
    ]
    configuration = str(item.get("configuration") or "").strip()
    variant = str(item.get("variant") or "").strip()
    if configuration:
        parts.append(configuration)
    elif variant:
        parts.append(f"Variante: {variant}")
    attributes = item.get("attributes") if isinstance(item.get("attributes"), dict) else {}
    color = str(attributes.get("color") or "").strip()
    if color:
        parts.append(f"Color: {color}")
    warranty = str(attributes.get("warranty") or "").strip()
    if warranty:
        parts.append(f"Garantia: {warranty}")
    product_notes = attributes.get("product_notes")
    if isinstance(product_notes, list):
        notes = [str(value).strip() for value in product_notes if str(value).strip()]
        if notes:
            parts.append(f"Notas: {'; '.join(notes)}")
    lead_time = str(item.get("lead_time", "") or "").strip()
    availability_type = str(item.get("availability_type", "") or "").strip()
    if lead_time:
        parts.append(f"Entrega: {lead_time}")
    elif availability_type == "made_to_order":
        parts.append("Entrega: Sobre pedido")
    if availability_type == "unknown":
        parts.append("Disponibilidad: por confirmar")
    availability = _availability_bucket_summary(item, attributes)
    if availability:
        parts.append(f"Disponibilidad: {availability}")
    if url:
        parts.append(f"URL: {url}")
    derived_warnings = [
        warning
        for warning in (
            "Codigo por verificar" if item.get("code_status") == "needs_review" else "",
            "Imagen de referencia" if item.get("image_kind") == "generated_reference" else "",
            _stock_warning(item, quantity),
            _price_warning(item),
        )
        if warning
    ]
    warnings = _merge_catalog_warnings(derived_warnings, item.get("warnings"))
    parts.extend(warnings)
    return " | ".join(part for part in parts if part), " | ".join(warnings)


def _catalog_warning_key(value: object) -> str:
    text = " ".join(str(value or "").strip().casefold().split())
    normalized = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    for category in (
        "precio por confirmar",
        "imagen de referencia",
        "codigo por verificar",
        "existencia insuficiente",
        "producto agotado",
        "agotado",
    ):
        if category in normalized:
            return "agotado" if category in {"producto agotado", "agotado"} else category
    return normalized


def _merge_catalog_warnings(derived: list[str], raw: object) -> list[str]:
    candidates = [*derived]
    if isinstance(raw, list):
        candidates.extend(str(value).strip() for value in raw if str(value).strip())
    result = []
    seen = set()
    for warning in candidates:
        key = _catalog_warning_key(warning)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(warning)
    return result


def _availability_bucket_summary(item: dict[str, Any], attributes: dict[str, Any]) -> str:
    buckets = attributes.get("availability_buckets")
    if not isinstance(buckets, list):
        return ""
    unit = str(item.get("unit") or "").strip()
    grouped: dict[str, Decimal] = {}
    for bucket in buckets:
        if not isinstance(bucket, dict):
            continue
        quantity = str(bucket.get("quantity") or "").strip()
        lead_time = str(bucket.get("lead_time") or "").strip()
        if quantity and lead_time:
            grouped[lead_time] = grouped.get(lead_time, Decimal(0)) + _decimal(quantity)
    parts = [
        f"{_excel_number(quantity)}{f' {unit}' if unit else ''} ({lead_time})"
        for lead_time, quantity in grouped.items()
    ]
    return "; ".join(parts)


def _stock_warning(item: dict[str, Any], quantity: Decimal) -> str:
    status = str(item.get("stock_status", "") or "").strip()
    availability_type = str(item.get("availability_type", "") or "").strip()
    if availability_type == "made_to_order":
        return ""
    available_value = item.get(
        "available_after_reservations",
        item.get("available_quantity", item.get("stock")),
    )
    available = _decimal(available_value) if available_value is not None else None
    if status in {"out_of_stock", "out", "exhausted"} or (
        availability_type == "stocked" and available is not None and available <= 0
    ):
        return "ADVERTENCIA: PRODUCTO AGOTADO"
    if status in {"insufficient_stock", "insufficient"} or (
        availability_type == "stocked" and available is not None and quantity > available
    ):
        quantity_number = _excel_number(quantity)
        available_number = _excel_number(available or Decimal(0))
        return (
            "ADVERTENCIA: EXISTENCIA INSUFICIENTE"
            f" - SOLICITADO {quantity_number} - DISPONIBLE {available_number}"
        )
    return ""


def _price_warning(item: dict[str, Any]) -> str:
    if str(item.get("price_source", "") or "").strip().casefold() == "missing":
        return "ADVERTENCIA: PRECIO POR CONFIRMAR"
    return ""


def _set_column_widths(ws) -> None:
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["K"].width = 42


def _add_catalog_image(
    ws,
    row: int,
    image_url: Any,
    image_dir: Path,
    code: str,
    source_type: str,
    destination_key: str | None = None,
) -> None:
    clean_url = str(image_url or "").strip()
    if not clean_url:
        return
    if destination_key is None:
        image_path = _download_catalog_image(image_url, image_dir, code, source_type)
    else:
        cache = getattr(ws, "_mobiliti_catalog_image_cache", None)
        if cache is None:
            cache = {}
            setattr(ws, "_mobiliti_catalog_image_cache", cache)
        cache_key = (source_type, clean_url)
        if cache_key in cache:
            image_path = cache[cache_key]
        else:
            image_path = _download_catalog_image(
                image_url,
                image_dir,
                code,
                source_type,
                destination_key=destination_key,
            )
            cache[cache_key] = image_path
    if not image_path:
        return
    try:
        image = XlsxImage(str(image_path))
        if image.width and image.height:
            scale = min(90 / image.width, 66 / image.height)
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
        image.anchor = f"C{row}"
        ws.add_image(image)
    except Exception:
        return


class _OfficialRedirectHandler(urllib.request.HTTPRedirectHandler):
    def __init__(self, allowed_hosts: frozenset[str]):
        super().__init__()
        self._allowed_hosts = allowed_hosts

    def redirect_request(self, req, fp, code, msg, headers, newurl):
        _validate_connected_peer(fp)
        _validate_official_https_url(newurl, self._allowed_hosts)
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def _download_catalog_image(
    url: Any,
    image_dir: Path,
    code: str,
    source_type: str,
    destination_key: str | None = None,
) -> Path | None:
    clean_url = str(url or "").strip()
    try:
        allowed_hosts = _allowed_image_hosts(source_type)
        if not allowed_hosts:
            return None
        _validate_official_https_url(clean_url, allowed_hosts)
        request = urllib.request.Request(clean_url, headers={"User-Agent": "Mobiliti Official Catalog/1.0"})
        opener = urllib.request.build_opener(
            urllib.request.ProxyHandler({}),
            _OfficialRedirectHandler(allowed_hosts),
        )
        with opener.open(request, timeout=18) as response:
            _validate_connected_peer(response)
            content_type = response.headers.get("content-type", "").split(";", 1)[0].strip().lower()
            content_length = response.headers.get("content-length")
            if not content_type.startswith("image/"):
                return None
            if content_length and int(content_length) > MAX_IMAGE_BYTES:
                return None
            data = response.read(MAX_IMAGE_BYTES + 1)
        if not data or len(data) > MAX_IMAGE_BYTES:
            return None
        suffix = _validated_catalog_image_suffix(data, content_type)
        safe_key = re.sub(
            r"[^A-Za-z0-9_-]+", "_", destination_key or code or "producto"
        )
        destination = image_dir / f"{safe_key}{suffix}"
        destination.write_bytes(data)
        return destination
    except Exception:
        return None


def _validated_catalog_image_suffix(data: bytes, content_type: str) -> str:
    with warnings.catch_warnings():
        warnings.simplefilter("error", Image.DecompressionBombWarning)
        try:
            with Image.open(BytesIO(data)) as image:
                image_format = str(image.format or "").upper()
                expected = CATALOG_IMAGE_FORMATS.get(image_format)
                if expected is None or content_type != expected[0]:
                    raise ValueError("Formato de imagen no coincide con MIME")
                width, height = image.size
                if width <= 0 or height <= 0 or width * height > MAX_CATALOG_IMAGE_PIXELS:
                    raise ValueError("Dimensiones de imagen invalidas")
                image.verify()
            with Image.open(BytesIO(data)) as decoded:
                if str(decoded.format or "").upper() != image_format:
                    raise ValueError("Formato de imagen inconsistente")
                decoded.load()
                width, height = decoded.size
                if width <= 0 or height <= 0 or width * height > MAX_CATALOG_IMAGE_PIXELS:
                    raise ValueError("Dimensiones de imagen invalidas")
        except (Image.DecompressionBombWarning, UnidentifiedImageError):
            raise
    return CATALOG_IMAGE_FORMATS[image_format][1]


def _allowed_image_hosts(source_type: str) -> frozenset[str]:
    if source_type != "supplier_cart":
        return OFFICIAL_IMAGE_HOSTS.get(source_type, frozenset())
    hosts: set[str] = set()
    for variable in ("CATALOG_ASSET_PUBLIC_BASE_URL", "SUPABASE_URL"):
        value = os.environ.get(variable, "").strip()
        if not value:
            continue
        try:
            parsed = urlsplit(value)
            host = (parsed.hostname or "").lower().rstrip(".")
            port = parsed.port
        except (UnicodeError, ValueError):
            return frozenset()
        labels = host.split(".")
        if (
            not host.isascii()
            or len(host) > 253
            or any(
                not label
                or len(label) > 63
                or label.startswith("-")
                or label.endswith("-")
                or re.fullmatch(r"[a-z0-9-]+", label) is None
                for label in labels
            )
        ):
            return frozenset()
        if (
            parsed.scheme.lower() == "https"
            and host
            and host != "kundesign.com"
            and not host.endswith(".kundesign.com")
            and not parsed.username
            and not parsed.password
            and port in (None, 443)
        ):
            hosts.add(host)
    return frozenset(hosts)


def _validate_official_https_url(url: str, allowed_hosts: frozenset[str]) -> None:
    parsed = urlsplit(url)
    host = (parsed.hostname or "").lower().rstrip(".")
    if (
        parsed.scheme.lower() != "https"
        or not host
        or parsed.username
        or parsed.password
        or parsed.port not in (None, 443)
        or host not in allowed_hosts
    ):
        raise ValueError("URL de imagen no es HTTPS oficial")
    _resolve_public_host(host)


def _resolve_public_host(host: str) -> None:
    addresses = {record[4][0] for record in socket.getaddrinfo(host, 443, type=socket.SOCK_STREAM)}
    if not addresses or any(not ipaddress.ip_address(address).is_global for address in addresses):
        raise ValueError("Host de imagen no resuelve a una direccion publica")


def _validate_connected_peer(response: Any) -> None:
    socket_paths = (
        ("fp", "raw", "_sock"),
        ("fp", "raw", "_socket"),
        ("fp", "_sock"),
        ("raw", "_sock"),
        ("_sock",),
    )
    connected_socket = None
    for path in socket_paths:
        candidate = response
        for attribute in path:
            candidate = getattr(candidate, attribute, None)
            if candidate is None:
                break
        if candidate is not None and callable(getattr(candidate, "getpeername", None)):
            connected_socket = candidate
            break
    if connected_socket is None:
        raise ValueError("No se pudo inspeccionar la IP conectada")
    try:
        peer = connected_socket.getpeername()
        address = str(peer[0]).split("%", 1)[0]
        peer_ip = ipaddress.ip_address(address)
    except (IndexError, OSError, TypeError, ValueError) as exc:
        raise ValueError("No se pudo inspeccionar la IP conectada") from exc
    if not peer_ip.is_global:
        raise ValueError("La IP conectada no es publica")


def parse_commercial_quantity(
    value: Any,
    *,
    item_label: str,
    max_decimal_places: int = DEFAULT_QUANTITY_DECIMAL_PLACES,
) -> Decimal:
    if not isinstance(max_decimal_places, int) or not 0 <= max_decimal_places <= MAX_QUANTITY_DECIMAL_PLACES:
        raise ValueError("Precision de cantidad invalida")
    text = str(value).replace(",", "").strip()
    if not text or len(text) > MAX_QUANTITY_TEXT_LENGTH:
        raise ValueError(f"Cantidad invalida para {item_label}")
    try:
        quantity = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Cantidad invalida para {item_label}") from None
    if not quantity.is_finite() or quantity <= 0:
        raise ValueError(f"Cantidad invalida para {item_label}")

    _, digits_tuple, exponent = quantity.as_tuple()
    digits = list(digits_tuple)
    while digits and digits[-1] == 0:
        digits.pop()
        exponent += 1
    decimal_places = max(-exponent, 0)
    if (
        len(digits) > MAX_QUANTITY_INTEGER_DIGITS + max_decimal_places
        or decimal_places > max_decimal_places
        or quantity > MAX_COMMERCIAL_QUANTITY
    ):
        raise ValueError(f"Cantidad invalida para {item_label}")
    return quantity


def _is_square_meter_unit(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return "".join(normalized.split()) in {"m2", "m^2"}


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def _excel_number(value: Decimal) -> int | float:
    if value == value.to_integral():
        return int(value)
    return float(value)
