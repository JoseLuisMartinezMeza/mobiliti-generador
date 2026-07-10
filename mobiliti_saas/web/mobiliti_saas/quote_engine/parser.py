from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


Q_HEADER_ROW = 7


@dataclass
class QuoteItem:
    tipo: str
    row: int
    nombre: Any = ""
    descripcion: Any = ""
    dimension: Any = ""
    cantidad: Any = None
    precio: Any = None
    categoria: str = ""


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).lower().strip()
    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _find_header_column(ws, terms: list[str], row: int = Q_HEADER_ROW) -> str | None:
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=row, column=col).value)
        if not header:
            continue
        tokens = set(header.split())
        if any(term in header or header in term or term in tokens for term in terms):
            return get_column_letter(col)
    return None


def detect_columns(ws) -> dict[str, str]:
    keywords = {
        "cantidad": ["qty", "quantity", "cantidad"],
        "unit_price": ["unit price", "unitprice", "price unit", "precio unitario"],
        "total_price": ["tot price", "total price", "totprice", "amount", "precio total"],
        "list_price": ["list price", "listprice", "price list"],
        "descripcion": ["description", "desc", "descripcion"],
        "dimension": ["dimension", "dimensions", "size", "medida"],
    }
    column_map: dict[str, str] = {}

    vol_col = _find_header_column(ws, ["vol", "volumen", "volume"])
    if vol_col:
        column_map["m3"] = vol_col

    for key, terms in keywords.items():
        found = _find_header_column(ws, terms)
        if found:
            column_map[key] = found

    if "m3" not in column_map:
        column_map["m3"] = column_map.get("dimension", "E")
    if "unit_price" not in column_map and "list_price" not in column_map:
        for col in range(1, ws.max_column + 1):
            if "price" in normalize_header(ws.cell(row=Q_HEADER_ROW, column=col).value):
                column_map["unit_price"] = get_column_letter(col)
                break
    return column_map


def col_index(column_map: dict[str, str], key: str, fallback: str) -> int:
    return column_index_from_string(column_map.get(key, fallback))


def read_items(source_path: str | Path) -> tuple[list[QuoteItem], dict[str, str]]:
    wb = load_workbook(source_path, data_only=True)
    if "Quotation" not in wb.sheetnames:
        wb.close()
        raise ValueError("El archivo no contiene hoja Quotation")

    ws = wb["Quotation"]
    column_map = detect_columns(ws)
    desc_col = col_index(column_map, "descripcion", "D")
    dim_col = col_index(column_map, "dimension", "E")
    qty_col = col_index(column_map, "cantidad", "G")
    price_col = col_index(column_map, "list_price", column_map.get("unit_price", "J"))

    last_row = ws.max_row
    for row in range(last_row, 0, -1):
        if ws.cell(row=row, column=1).value is not None:
            last_row = row
            break

    current_category = ""
    items: list[QuoteItem] = []
    for row in range(Q_HEADER_ROW + 1, last_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        descripcion = ws.cell(row=row, column=desc_col).value
        dimension = ws.cell(row=row, column=dim_col).value

        if isinstance(no_val, str) and no_val.startswith("-"):
            current_category = no_val.strip("- ").strip()
            items.append(QuoteItem(tipo="categoria", row=row, nombre=current_category))
            continue
        if (item_name is None or item_name == "") and (no_val is None or no_val == ""):
            continue
        if isinstance(no_val, (int, float)):
            items.append(
                QuoteItem(
                    tipo="producto",
                    row=row,
                    nombre=item_name,
                    descripcion=descripcion,
                    dimension=dimension,
                    cantidad=ws.cell(row=row, column=qty_col).value,
                    precio=ws.cell(row=row, column=price_col).value,
                    categoria=current_category,
                )
            )
        elif (no_val is None or no_val == "") and item_name:
            current_category = str(item_name).strip()
            items.append(QuoteItem(tipo="categoria", row=row, nombre=current_category))

    wb.close()
    return items, column_map
