"""Composición OOXML allowlist sobre la plantilla oficial de Mobiliti.

El módulo nunca abre ni guarda el libro completo con OpenPyXL.  Parte de los
bytes oficiales, reemplaza únicamente las superficies declaradas por el
contrato y publica el ZIP sólo después de auditar el archivo candidato.
"""

from __future__ import annotations

from copy import deepcopy
import ctypes
from dataclasses import dataclass, field
from decimal import Decimal
import errno
from io import BytesIO
import math
import os
from pathlib import Path
import posixpath
import re
import shutil
import stat
import subprocess
import sys
from types import MappingProxyType
from typing import Mapping, Sequence
import uuid
from xml.etree import ElementTree as ET

from openpyxl.formula.tokenizer import Token, Tokenizer, TokenizerError
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from .mobiliti_layout import MobilitiRowMap, SectionNeed
from .ooxml_formula import translate_formula
from .ooxml_package import (
    OFFICE_DOCUMENT_RELATIONSHIPS,
    PACKAGE_RELATIONSHIPS,
    PackageAudit,
    PackageMutation,
    XlsxPackage,
    assert_packages_preserved,
    relationship_owner,
    relationship_part_name,
    relationship_type_uris,
    part_name_has_ascii_prefix,
    resolve_internal_target,
    validate_part_name,
)
from .ooxml_worksheet import (
    MobilitiCellWrite,
    MobilitiSheetMutation,
    WorksheetEditor,
    build_mobiliti_sheet,
)
from .official_template import TemplateContract, verify_official_template
from .quotation_sheets import (
    SheetAddition,
    _parse_xml_document,
    _xml_bytes,
)


MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = OFFICE_DOCUMENT_RELATIONSHIPS
PKG_REL = PACKAGE_RELATIONSHIPS
CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING = "http://schemas.openxmlformats.org/drawingml/2006/main"
CHART = "http://schemas.openxmlformats.org/drawingml/2006/chart"
XML = "http://www.w3.org/XML/1998/namespace"
WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
TABLE_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
)
CHART_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
)
STYLES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
)
RELATIONSHIPS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-package.relationships+xml"
)
PNG_CONTENT_TYPE = "image/png"
JPEG_CONTENT_TYPE = "image/jpeg"
MAX_CELL_TEXT = 32_767
MAX_IMAGE_BYTES = 10 * 1024 * 1024
MAX_IMAGE_PIXELS = 40_000_000
MAX_EXCEL_FORMULA_LENGTH = 8_192
XLSX_MAX_ROWS = 1_048_576
CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW = 16
CANONICAL_COTIZACION_FIRST_PRODUCT_ROW = 17
CANONICAL_COTIZACION_TOTAL_START = 20
CANONICAL_COTIZACION_TOTAL_ROW = 24
CANONICAL_COTIZACION_TOTAL_SPAN = (
    CANONICAL_COTIZACION_TOTAL_ROW - CANONICAL_COTIZACION_TOTAL_START
)
CANONICAL_COTIZACION_TERMS_START = 28
CANONICAL_COTIZACION_CURRENCY_TERM_ROW = 33
CANONICAL_COTIZACION_PRINT_END = 76
CANONICAL_MOBILITI_TOTAL_ROW = 573
CANONICAL_MOBILITI_AUX_START = 574
CANONICAL_MOBILITI_AUX_END = 610
FLETE_ROUTES = {
    5: ("Guadalajara", "Monterrey"),
    7: ("Guadalajara", "Mexico City"),
    9: ("Guadalajara", "Tijuana"),
    11: ("Guadalajara", "Mérida, Yucatán"),
    13: ("Guadalajara", "Querétaro"),
    15: ("Guadalajara", "State of Mexico"),
    17: ("Guadalajara", "Guadalajara"),
}
FLETE_CATEGORY_ROWS = {
    16: ("Bancos", Decimal("0.2"), Decimal("56")),
    17: ("Cocineta", Decimal("0.2"), Decimal("1790")),
    18: ("Pizarrones", Decimal("0.2"), Decimal("210")),
}
FLETE_CATEGORY_LABEL_ROWS = {8: "Escritorios-WorkStation"}
_COTIZACION_IMAGE_CELL_WIDTH_EMU = 5_019_675
_COTIZACION_IMAGE_CELL_HEIGHT_EMU = 4_408_170
CANONICAL_MOBILITI_SECTION_COUNT = 16
CANONICAL_MOBILITI_BLOCK_HEIGHT = 35
CANONICAL_MOBILITI_FIRST_SECTION_ROW = 13
CANONICAL_MOBILITI_PRODUCT_CAPACITY = 33
CDMX_LUMBRO_SHEET = "Cantidades Lumbro "
CDMX_LUMBRO_FIRST_PRODUCT_ROW = 4
CDMX_LUMBRO_MIN_END_ROW = 28
CDMX_LUMBRO_LAST_COLUMN = 16
_CELL = re.compile(r"(?P<column>[A-Z]{1,3})(?P<row>[1-9][0-9]*)\Z")
_SAFE_IMAGE_NAME = re.compile(r"[A-Za-z0-9][A-Za-z0-9_. -]{0,180}\Z")
_DANGEROUS_IMPORTED_FUNCTIONS = frozenset(
    {
        "CALL",
        "DDE",
        "ENCODEURL",
        "EXEC",
        "FILTERXML",
        "HYPERLINK",
        "REGISTER.ID",
        "RTD",
        "WEBSERVICE",
    }
)
_FORMULA_PART_CONTENT_TYPES = {
    "worksheet": WORKSHEET_CONTENT_TYPE,
    "table": TABLE_CONTENT_TYPE,
    "chart": CHART_CONTENT_TYPE,
}
_FORMULA_PART_KIND_BY_CONTENT_TYPE = {
    content_type: kind
    for kind, content_type in _FORMULA_PART_CONTENT_TYPES.items()
}
_FORMULA_PART_ROOTS = {
    "worksheet": f"{{{MAIN}}}worksheet",
    "table": f"{{{MAIN}}}table",
    "chart": f"{{{CHART}}}chartSpace",
}
_FORMULA_RELATIONSHIP_KIND_BY_TYPE = {
    relationship_type: kind
    for kind in ("table", "chart")
    for relationship_type in relationship_type_uris(kind)
}
_FORMULA_ELEMENT_PATHS = {
    "worksheet": (
        ("cell", f".//{{{MAIN}}}c/{{{MAIN}}}f", True),
        (
            "dataValidation/formula1",
            f".//{{{MAIN}}}dataValidation/{{{MAIN}}}formula1",
            False,
        ),
        (
            "dataValidation/formula2",
            f".//{{{MAIN}}}dataValidation/{{{MAIN}}}formula2",
            False,
        ),
        (
            "conditionalFormatting/formula",
            f".//{{{MAIN}}}cfRule/{{{MAIN}}}formula",
            False,
        ),
    ),
    "table": (
        (
            "calculatedColumnFormula",
            f".//{{{MAIN}}}calculatedColumnFormula",
            False,
        ),
        (
            "totalsRowFormula",
            f".//{{{MAIN}}}totalsRowFormula",
            False,
        ),
    ),
    "chart": (("range", f".//{{{CHART}}}f", False),),
}
_APPEND_ONLY_STYLE_SECTIONS = {
    "numFmts": "numFmt",
    "fonts": "font",
    "fills": "fill",
    "borders": "border",
    "cellStyleXfs": "xf",
    "cellXfs": "xf",
    "cellStyles": "cellStyle",
    "dxfs": "dxf",
    "tableStyles": "tableStyle",
}

for prefix, namespace in (
    ("", MAIN),
    ("r", REL),
    ("xdr", XDR),
    ("a", DRAWING),
):
    ET.register_namespace(prefix, namespace)


@dataclass(frozen=True)
class CotizacionMetadata:
    """Valores visibles permitidos en el encabezado oficial."""

    quotation_number: str = ""
    project: str = ""
    client: str = ""
    email: str = ""
    phone: str = ""
    address: str = ""
    business_name: str = ""
    quote_currency: str = ""

    def __post_init__(self) -> None:
        for value in (
            self.quotation_number,
            self.project,
            self.client,
            self.email,
            self.phone,
            self.address,
            self.business_name,
        ):
            _validate_text(value, allow_empty=True)
        if self.quote_currency not in {"", "MXN", "USD", "EUR"}:
            raise ValueError("Moneda de Cotizacion inválida")


@dataclass(frozen=True)
class CotizacionPriceTerm:
    """Término racional que enlaza un precio visible con una fila Mobiliti."""

    mobiliti_row: int
    numerator: Decimal = Decimal("1")
    denominator: Decimal = Decimal("1")

    def __post_init__(self) -> None:
        if type(self.mobiliti_row) is not int or not (
            1 <= self.mobiliti_row <= XLSX_MAX_ROWS
        ):
            raise ValueError("Fila Mobiliti del término inválida")
        numerator = _decimal(self.numerator, "numerador")
        denominator = _decimal(self.denominator, "denominador")
        if numerator <= 0 or denominator <= 0:
            raise ValueError("Factor de precio compuesto inválido")
        object.__setattr__(self, "numerator", numerator)
        object.__setattr__(self, "denominator", denominator)


@dataclass(frozen=True)
class CotizacionFormulaContract:
    """Contrato tipado para enlazar F con Mobiliti y conservar la fórmula I."""

    def product_formulas(
        self,
        *,
        target_row: int,
        price_terms: tuple[CotizacionPriceTerm, ...],
    ) -> dict[str, str]:
        if type(target_row) is not int or not 1 <= target_row <= XLSX_MAX_ROWS:
            raise ValueError("Fila Cotizacion del contrato F/I inválida")
        if (
            not isinstance(price_terms, tuple)
            or not price_terms
            or not all(isinstance(term, CotizacionPriceTerm) for term in price_terms)
        ):
            raise TypeError("Términos de precio Cotizacion inválidos")

        parts = ["="]
        formula_length = 1
        expected_f_tokens: list[tuple[str, str, str]] = []
        composed = len(price_terms) > 1
        principal_row = price_terms[0].mobiliti_row
        if composed and (
            price_terms[0].numerator != 1
            or price_terms[0].denominator != 1
        ):
            raise ValueError("Principal de precio compuesto inválido")
        for index, term in enumerate(price_terms):
            if index:
                formula_length = _append_bounded_formula(
                    parts,
                    "+",
                    formula_length,
                )
                expected_f_tokens.append((Token.OP_IN, "", "+"))
            price_column = "X"
            reference = f"Mobiliti!{price_column}{term.mobiliti_row}"
            formula_length = _append_bounded_formula(
                parts,
                reference,
                formula_length,
            )
            expected_f_tokens.append((Token.OPERAND, Token.RANGE, reference))
            if composed and index:
                quantity_reference = f"Mobiliti!H{term.mobiliti_row}"
                formula_length = _append_bounded_formula(
                    parts,
                    f"*{quantity_reference}",
                    formula_length,
                )
                expected_f_tokens.extend(
                    (
                        (Token.OP_IN, "", "*"),
                        (Token.OPERAND, Token.RANGE, quantity_reference),
                    )
                )
                principal_quantity_reference = f"Mobiliti!H{principal_row}"
                formula_length = _append_bounded_formula(
                    parts,
                    f"/{principal_quantity_reference}",
                    formula_length,
                )
                expected_f_tokens.extend(
                    (
                        (Token.OP_IN, "", "/"),
                        (
                            Token.OPERAND,
                            Token.RANGE,
                            principal_quantity_reference,
                        ),
                    )
                )
            elif term.numerator != 1:
                numerator = _excel_decimal(term.numerator)
                formula_length = _append_bounded_formula(
                    parts,
                    f"*{numerator}",
                    formula_length,
                )
                expected_f_tokens.extend(
                    (
                        (Token.OP_IN, "", "*"),
                        (Token.OPERAND, Token.NUMBER, numerator),
                    )
                )
            if not composed and term.denominator != 1:
                denominator = _excel_decimal(term.denominator)
                formula_length = _append_bounded_formula(
                    parts,
                    f"/{denominator}",
                    formula_length,
                )
                expected_f_tokens.extend(
                    (
                        (Token.OP_IN, "", "/"),
                        (Token.OPERAND, Token.NUMBER, denominator),
                    )
                )
        formulas = {
            "F": "".join(parts),
            "I": f"=F{target_row}-H{target_row}",
        }
        _validate_formula_token_contract(
            formulas["F"],
            tuple(expected_f_tokens),
        )
        _validate_formula_token_contract(
            formulas["I"],
            (
                (Token.OPERAND, Token.RANGE, f"F{target_row}"),
                (Token.OP_IN, "", "-"),
                (Token.OPERAND, Token.RANGE, f"H{target_row}"),
            ),
        )
        return formulas


@dataclass(frozen=True)
class CotizacionImageSource:
    """Fuente binaria o local de una imagen visible de producto."""

    path: Path | None = None
    content: bytes | None = None
    content_type: str | None = None

    def __post_init__(self) -> None:
        if self.path is not None:
            object.__setattr__(self, "path", Path(self.path))
        if self.content is not None:
            if not isinstance(self.content, bytes):
                raise TypeError("Bytes de imagen Cotizacion inválidos")
            object.__setattr__(self, "content", bytes(self.content))
        if (self.path is None) == (self.content is None):
            raise ValueError("La imagen Cotizacion requiere una fuente única")
        if self.content is not None and self.content_type not in {
            PNG_CONTENT_TYPE,
            JPEG_CONTENT_TYPE,
        }:
            raise ValueError("Content type de imagen Cotizacion inválido")
        if self.path is not None and self.content_type is not None:
            raise ValueError("La ruta de imagen no declara content type anticipado")


@dataclass(frozen=True)
class CotizacionProduct:
    """Una línea visible de ``Cotizacion`` enlazada a una fila Mobiliti."""

    item_key: str
    name: str
    description: str
    dimensions: str
    quantity: Decimal
    mobiliti_row: int
    discount: Decimal = Decimal("0")
    description_formula: str | None = None
    image_path: Path | None = None
    image_content: bytes | None = None
    image_content_type: str | None = None
    complement_images: tuple[CotizacionImageSource, ...] = ()
    price_terms: tuple[CotizacionPriceTerm, ...] = ()
    price_pending: bool = False

    def __post_init__(self) -> None:
        _validate_text(self.item_key)
        _validate_text(self.name)
        _validate_text(self.description, allow_empty=True)
        _validate_text(self.dimensions, allow_empty=True)
        quantity = _decimal(self.quantity, "cantidad")
        discount = _decimal(self.discount, "descuento")
        if quantity <= 0:
            raise ValueError("La cantidad de Cotizacion debe ser positiva")
        if discount < 0 or discount > 1:
            raise ValueError("El descuento de Cotizacion debe estar entre 0 y 1")
        if type(self.mobiliti_row) is not int or not 1 <= self.mobiliti_row <= XLSX_MAX_ROWS:
            raise ValueError("Fila Mobiliti de Cotizacion inválida")
        if not isinstance(self.price_terms, tuple) or not all(
            isinstance(term, CotizacionPriceTerm) for term in self.price_terms
        ):
            raise TypeError("Términos de precio Cotizacion inválidos")
        if type(self.price_pending) is not bool:
            raise TypeError("Estado de precio Cotizacion inválido")
        if not isinstance(self.complement_images, tuple) or not all(
            isinstance(image, CotizacionImageSource)
            for image in self.complement_images
        ):
            raise TypeError("Imágenes de complemento Cotizacion inválidas")
        if self.description_formula is not None:
            _validate_cotizacion_description_formula(self.description_formula)
        price_terms = self.price_terms or (CotizacionPriceTerm(self.mobiliti_row),)
        mobiliti_rows = tuple(term.mobiliti_row for term in price_terms)
        if len(mobiliti_rows) != len(set(mobiliti_rows)):
            raise ValueError("Fila Mobiliti duplicada en precio compuesto")
        object.__setattr__(self, "quantity", quantity)
        object.__setattr__(self, "discount", discount)
        object.__setattr__(self, "price_terms", price_terms)
        if self.image_path is not None:
            object.__setattr__(self, "image_path", Path(self.image_path))
        if self.image_content is not None:
            if not isinstance(self.image_content, bytes):
                raise TypeError("Bytes de imagen Cotizacion inválidos")
            object.__setattr__(self, "image_content", bytes(self.image_content))
        if (self.image_path is None) == (self.image_content is None):
            if self.image_path is not None:
                raise ValueError("Producto Cotizacion contiene dos fuentes de imagen")
            if self.image_content_type is not None:
                raise ValueError("Content type sin imagen Cotizacion")
        elif self.image_content is not None and self.image_content_type not in {
            PNG_CONTENT_TYPE,
            JPEG_CONTENT_TYPE,
        }:
            raise ValueError("Content type de imagen Cotizacion inválido")


@dataclass(frozen=True)
class CotizacionSection:
    """Categoría de presentación y sus productos, en orden autoritativo."""

    title: str
    products: tuple[CotizacionProduct, ...]

    def __post_init__(self) -> None:
        _validate_text(self.title)
        if not isinstance(self.products, tuple) or not all(
            isinstance(item, CotizacionProduct) for item in self.products
        ):
            raise TypeError("Productos de sección Cotizacion inválidos")


@dataclass(frozen=True)
class CotizacionProductImage:
    path: Path | None
    target_row: int
    content: bytes | None = None
    content_type: str | None = None
    position: int = 0

    def __post_init__(self) -> None:
        if type(self.target_row) is not int or not 1 <= self.target_row <= XLSX_MAX_ROWS:
            raise ValueError("Fila de imagen Cotizacion inválida")
        if type(self.position) is not int or self.position < 0:
            raise ValueError("Posición de imagen Cotizacion inválida")
        if self.path is not None:
            object.__setattr__(self, "path", Path(self.path))
        if self.content is not None:
            if not isinstance(self.content, bytes):
                raise TypeError("Bytes de imagen Cotizacion inválidos")
            object.__setattr__(self, "content", bytes(self.content))
        if (self.path is None) == (self.content is None):
            raise ValueError("La imagen Cotizacion requiere una fuente única")
        if self.content is not None and self.content_type not in {
            PNG_CONTENT_TYPE,
            JPEG_CONTENT_TYPE,
        }:
            raise ValueError("Content type de imagen Cotizacion inválido")
        if self.path is not None and self.content_type is not None:
            raise ValueError("La ruta de imagen no declara content type anticipado")


@dataclass(frozen=True)
class CotizacionSheetMutation:
    """Mutación de hoja y partes de dibujo explícitamente relacionadas."""

    xml: bytes
    related_parts: Mapping[str, bytes] = field(default_factory=dict)
    total_row: int = CANONICAL_COTIZACION_TOTAL_ROW
    related_additions: Mapping[str, bytes] = field(default_factory=dict)
    related_content_types: Mapping[str, str] = field(default_factory=dict)
    images: tuple[CotizacionProductImage, ...] = ()
    terms_row_delta: int = 0
    product_rows: tuple[int, ...] = ()
    product_mobiliti_rows: tuple[tuple[int, ...], ...] = ()
    composer_variant: str = "official"
    section_subtotal_rows: tuple[int, ...] = ()
    quote_currency: str = ""

    def __post_init__(self) -> None:
        if not isinstance(self.xml, bytes):
            raise TypeError("XML de Cotizacion inválido")
        if type(self.total_row) is not int or self.total_row < 1:
            raise ValueError("Fila total de Cotizacion inválida")
        if type(self.terms_row_delta) is not int:
            raise TypeError("Desplazamiento de términos inválido")
        if not isinstance(self.images, tuple) or not all(
            isinstance(item, CotizacionProductImage) for item in self.images
        ):
            raise TypeError("Imágenes de Cotizacion inválidas")
        if not isinstance(self.product_rows, tuple) or not all(
            type(item) is int and item >= 1 for item in self.product_rows
        ):
            raise TypeError("Filas de producto Cotizacion inválidas")
        if (
            not isinstance(self.product_mobiliti_rows, tuple)
            or len(self.product_mobiliti_rows) != len(self.product_rows)
            or not all(
                isinstance(rows, tuple)
                and rows
                and all(type(row) is int and 1 <= row <= XLSX_MAX_ROWS for row in rows)
                and len(rows) == len(set(rows))
                for rows in self.product_mobiliti_rows
            )
        ):
            raise TypeError("Filas Mobiliti por producto Cotizacion inválidas")
        if self.composer_variant not in {"official", "sunon_cdmx_v1c"}:
            raise ValueError("Variante de compositor Cotizacion inválida")
        if self.quote_currency not in {"", "MXN", "USD", "EUR"}:
            raise ValueError("Moneda de mutación Cotizacion inválida")
        if (
            not isinstance(self.section_subtotal_rows, tuple)
            or not all(
                type(item) is int and item >= 1
                for item in self.section_subtotal_rows
            )
            or tuple(sorted(set(self.section_subtotal_rows)))
            != self.section_subtotal_rows
        ):
            raise TypeError("Filas de subtotal por sección inválidas")
        if self.composer_variant == "official" and self.section_subtotal_rows:
            raise ValueError("La plantilla oficial no admite subtotales CDMX")
        object.__setattr__(
            self,
            "related_parts",
            _immutable_bytes_mapping(self.related_parts, "Partes relacionadas"),
        )
        object.__setattr__(
            self,
            "related_additions",
            _immutable_bytes_mapping(self.related_additions, "Adiciones relacionadas"),
        )
        object.__setattr__(
            self,
            "related_content_types",
            MappingProxyType(dict(self.related_content_types)),
        )


@dataclass(frozen=True)
class ComposeRequest:
    template: Path
    output: Path
    mobiliti: MobilitiSheetMutation
    cotizacion: CotizacionSheetMutation
    quotation: SheetAddition | None
    quotation_data: SheetAddition
    contract: TemplateContract
    project_composition: bool = False

    def __post_init__(self) -> None:
        object.__setattr__(self, "template", Path(self.template))
        object.__setattr__(self, "output", Path(self.output))
        if not isinstance(self.mobiliti, MobilitiSheetMutation):
            raise TypeError("Mutación Mobiliti inválida")
        if not isinstance(self.cotizacion, CotizacionSheetMutation):
            raise TypeError("Mutación Cotizacion inválida")
        if self.quotation is not None and not isinstance(self.quotation, SheetAddition):
            raise TypeError("Adición Quotation inválida")
        if not isinstance(self.quotation_data, SheetAddition):
            raise TypeError("Adición Quotation_Data inválida")
        if self.quotation_data.name != "Quotation_Data":
            raise ValueError("La adición obligatoria debe ser Quotation_Data")
        if not isinstance(self.contract, TemplateContract):
            raise TypeError("Contrato oficial inválido")
        if type(self.project_composition) is not bool:
            raise TypeError("Señal de composición de Proyecto inválida")

    def with_output(self, output: Path) -> "ComposeRequest":
        return ComposeRequest(
            template=self.template,
            output=Path(output),
            mobiliti=self.mobiliti,
            cotizacion=self.cotizacion,
            quotation=self.quotation,
            quotation_data=self.quotation_data,
            contract=self.contract,
            project_composition=self.project_composition,
        )


class CotizacionSheetEditor:
    """Editor estrecho que clona los bloques visibles del XML oficial."""

    def __init__(self, root: ET.Element):
        if root.tag != f"{{{MAIN}}}worksheet":
            raise ValueError("La parte indicada no es Cotizacion SpreadsheetML")
        sheet_data = root.find(f"{{{MAIN}}}sheetData")
        if sheet_data is None:
            raise ValueError("Cotizacion no contiene sheetData")
        self.root = root
        self.sheet_data = sheet_data

    @classmethod
    def from_xml(cls, payload: bytes) -> "CotizacionSheetEditor":
        try:
            root = ET.fromstring(payload)
        except (ET.ParseError, TypeError) as error:
            raise ValueError("El XML de Cotizacion no es válido") from error
        return cls(root)

    def compose(
        self,
        *,
        metadata: CotizacionMetadata,
        sections: Sequence[CotizacionSection],
        composer_variant: str = "official",
    ) -> CotizacionSheetMutation:
        """Reemplaza únicamente encabezado y bloques dinámicos A:J."""

        if not isinstance(metadata, CotizacionMetadata):
            raise TypeError("Metadata de Cotizacion inválida")
        if isinstance(sections, (str, bytes)) or not isinstance(sections, Sequence):
            raise TypeError("Secciones Cotizacion inválidas")
        if composer_variant not in {"official", "sunon_cdmx_v1c"}:
            raise ValueError("Variante de compositor Cotizacion inválida")
        frozen_sections = tuple(sections)
        if not all(isinstance(section, CotizacionSection) for section in frozen_sections):
            raise TypeError("Secciones Cotizacion inválidas")
        products = tuple(
            product for section in frozen_sections for product in section.products
        )
        if not products:
            raise ValueError("Cotizacion requiere al menos un producto")
        keys = tuple(product.item_key for product in products)
        if len(keys) != len(set(keys)):
            raise ValueError("item_key duplicado en Cotizacion")

        rows = {int(row.attrib["r"]): row for row in self.sheet_data.findall(f"{{{MAIN}}}row")}
        try:
            category_template = rows[16]
            product_template = rows[17]
            total_templates = tuple(rows[number] for number in range(20, 25))
            gap_templates = tuple(rows[number] for number in range(25, 28))
        except KeyError as error:
            raise ValueError("Bloques oficiales de Cotizacion incompletos") from error
        product_h_formula = _formula_text(product_template, "H17")
        product_j_formula = _formula_text(product_template, "J17")
        subtotal_formula = _formula_text(total_templates[0], "H20")
        if None in {product_h_formula, product_j_formula, subtotal_formula}:
            raise ValueError("Fórmulas oficiales de Cotizacion incompletas")
        section_subtotal_template: ET.Element | None = None
        if composer_variant == "sunon_cdmx_v1c":
            try:
                section_subtotal_template = rows[18]
            except KeyError as error:
                raise ValueError(
                    "Prototipo de subtotal por área CDMX ausente"
                ) from error
            if _formula_text(
                section_subtotal_template,
                "J18",
            ) != "=SUM(J13:J15)":
                raise ValueError(
                    "Firma del subtotal por área CDMX inesperada"
                )
        official_f = _require_cell(product_template, "F17")
        official_i = _require_cell(product_template, "I17")
        if (
            _formula_text(product_template, "F17") != "=Mobiliti!X14"
            or official_f.findtext(f"{{{MAIN}}}v") != "0"
            or _formula_text(product_template, "I17") is not None
            or official_i.find(f"{{{MAIN}}}v") is not None
        ):
            raise ValueError("Firma oficial F/I inesperada")
        formula_contract = CotizacionFormulaContract()

        self._write_metadata(metadata)
        preserved_headers = [row for number, row in sorted(rows.items()) if number < 16]
        dynamic_rows: list[ET.Element] = []
        dynamic_merges: list[str] = []
        product_rows: list[int] = []
        section_subtotal_rows: list[int] = []
        images: list[CotizacionProductImage] = []
        cursor = CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW
        first_discount_row: int | None = None

        for section in frozen_sections:
            if not section.products:
                continue
            header = _clone_row_region(category_template, 16, cursor, last_column=10)
            _set_inline_string(header, f"A{cursor}", section.title)
            dynamic_rows.append(header)
            dynamic_merges.append(f"A{cursor}:J{cursor}")
            cursor += 1
            section_first_product_row = cursor
            for product in section.products:
                target_row = cursor
                if first_discount_row is None:
                    first_discount_row = target_row
                row = _clone_row_region(product_template, 17, target_row, last_column=10)
                _clear_cell_value(_require_cell(row, f"B{target_row}"))
                if product.description_formula is None:
                    _set_inline_string(row, f"A{target_row}", product.name)
                    _set_inline_string(row, f"C{target_row}", product.description)
                    _set_inline_string(row, f"D{target_row}", product.dimensions)
                    _set_number(row, f"E{target_row}", product.quantity)
                else:
                    quotation_row = _validate_cotizacion_description_formula(
                        product.description_formula
                    )
                    _set_formula(
                        row,
                        f"A{target_row}",
                        f"=Mobiliti!D{product.mobiliti_row}",
                    )
                    _set_formula(
                        row,
                        f"C{target_row}",
                        product.description_formula,
                    )
                    _set_formula(
                        row,
                        f"D{target_row}",
                        f"=Quotation!F{quotation_row}",
                    )
                    _set_formula(
                        row,
                        f"E{target_row}",
                        f"=Mobiliti!H{product.mobiliti_row}",
                    )
                contract_formulas = None
                if product.price_pending:
                    _set_inline_string(row, f"F{target_row}", "Por confirmar")
                else:
                    contract_formulas = formula_contract.product_formulas(
                        price_terms=product.price_terms,
                        target_row=target_row,
                    )
                    _set_formula(row, f"F{target_row}", contract_formulas["F"])
                if target_row == first_discount_row:
                    _set_number(row, f"G{target_row}", product.discount)
                else:
                    _set_formula(
                        row,
                        f"G{target_row}",
                        f"=$G${first_discount_row}",
                    )
                if product.price_pending:
                    _set_inline_string(row, f"H{target_row}", "")
                    _set_inline_string(row, f"I{target_row}", "Por confirmar")
                    _set_inline_string(row, f"J{target_row}", "")
                else:
                    _set_formula(
                        row,
                        f"H{target_row}",
                        translate_formula(
                            product_h_formula,
                            origin="H17",
                            target=f"H{target_row}",
                            sheet="Cotizacion",
                        ),
                    )
                    assert contract_formulas is not None
                    _set_formula(row, f"I{target_row}", contract_formulas["I"])
                    _set_formula(
                        row,
                        f"J{target_row}",
                        translate_formula(
                            product_j_formula,
                            origin="J17",
                            target=f"J{target_row}",
                            sheet="Cotizacion",
                        ),
                    )
                dynamic_rows.append(row)
                product_rows.append(target_row)
                image_sources: list[CotizacionImageSource] = []
                if product.image_path is not None or product.image_content is not None:
                    image_sources.append(
                        CotizacionImageSource(
                            path=product.image_path,
                            content=product.image_content,
                            content_type=product.image_content_type,
                        )
                    )
                image_sources.extend(product.complement_images)
                for position, source in enumerate(image_sources):
                    images.append(
                        CotizacionProductImage(
                            source.path,
                            target_row,
                            content=source.content,
                            content_type=source.content_type,
                            position=position,
                        )
                    )
                cursor += 1
            if section_subtotal_template is not None:
                section_subtotal_row = cursor
                footer = _clone_row_region(
                    section_subtotal_template,
                    18,
                    section_subtotal_row,
                    last_column=10,
                )
                _set_inline_string(
                    footer,
                    f"I{section_subtotal_row}",
                    "SUBTOTAL AREA",
                )
                _set_formula(
                    footer,
                    f"J{section_subtotal_row}",
                    (
                        f"=SUM(J{section_first_product_row}:"
                        f"J{section_subtotal_row - 1})"
                    ),
                )
                dynamic_rows.append(footer)
                section_subtotal_rows.append(section_subtotal_row)
                cursor += 1

        total_start = cursor
        total_delta = total_start - CANONICAL_COTIZACION_TOTAL_START
        totals = [
            _clone_row_region(row, source, source + total_delta, last_column=10)
            for source, row in zip(range(20, 25), total_templates, strict=True)
        ]
        subtotal_row = total_start
        total_row = CANONICAL_COTIZACION_TOTAL_ROW + total_delta
        if total_row + (CANONICAL_COTIZACION_PRINT_END - CANONICAL_COTIZACION_TOTAL_ROW) > XLSX_MAX_ROWS:
            raise ValueError("Cotizacion excede la capacidad física de XLSX")
        if composer_variant == "sunon_cdmx_v1c":
            subtotal_references = ",".join(
                f"J{row}" for row in section_subtotal_rows
            )
            if not subtotal_references:
                raise ValueError("Cotizacion CDMX requiere subtotales por área")
            _set_formula(
                totals[0],
                f"H{subtotal_row}",
                f"=SUM({subtotal_references})",
            )
        else:
            translated_subtotal = translate_formula(
                subtotal_formula,
                origin="H20",
                target=f"H{subtotal_row}",
                sheet="Cotizacion",
            )
            subtotal_ranges = [
                token
                for token in _formula_range_tokens(translated_subtotal[1:])
                if re.fullmatch(r"J[1-9][0-9]*:J[1-9][0-9]*", token)
            ]
            if len(subtotal_ranges) != 1:
                raise ValueError("Rango subtotal oficial Cotizacion inesperado")
            translated_subtotal = translate_formula(
                subtotal_formula,
                origin="H20",
                target=f"H{subtotal_row}",
                range_overrides={
                    subtotal_ranges[0]: f"J{product_rows[0]}:J{product_rows[-1]}"
                },
                sheet="Cotizacion",
            )
            _set_formula(
                totals[0],
                f"H{subtotal_row}",
                translated_subtotal,
                attributes={"t": "array", "ref": f"H{subtotal_row}"},
            )
        for offset, row in enumerate(totals):
            target = total_start + offset
            if offset != 0:
                source = CANONICAL_COTIZACION_TOTAL_START + offset
                source_formula = _formula_text(total_templates[offset], f"H{source}")
                if source_formula is not None:
                    if offset == 2:
                        source_formula = (
                            f"=H{CANONICAL_COTIZACION_TOTAL_START}"
                            f"+H{CANONICAL_COTIZACION_TOTAL_START + 1}"
                        )
                    _set_formula(
                        row,
                        f"H{target}",
                        translate_formula(
                            source_formula,
                            origin=f"H{source}",
                            target=f"H{target}",
                            sheet="Cotizacion",
                        ),
                    )
        dynamic_rows.extend(totals)
        dynamic_rows.extend(
            _clone_row_region(row, source, source + total_delta, last_column=10)
            for source, row in zip(range(25, 28), gap_templates, strict=True)
        )

        terms_rows = [
            _clone_row_region(row, number, number + total_delta, last_column=10)
            for number, row in sorted(rows.items())
            if number >= CANONICAL_COTIZACION_TERMS_START
        ]
        if metadata.quote_currency:
            currency_name = {
                "MXN": "Pesos Mexicanos",
                "USD": "Dólares Estadounidenses",
                "EUR": "Euros",
            }[metadata.quote_currency]
            target = CANONICAL_COTIZACION_CURRENCY_TERM_ROW + total_delta
            currency_row = next(
                row for row in terms_rows if int(row.attrib["r"]) == target
            )
            _set_inline_string(
                currency_row,
                f"A{target}",
                (
                    f"a. Esta cotización se encuentra en {currency_name}, "
                    f"{metadata.quote_currency}. En caso de que el pago se efectúe "
                    "en una moneda diferente a la de este documento, se tomará como "
                    "referencia el tipo de cambio bancario."
                ),
            )
        sidecars = _sidecar_cells(rows, first_row=16, first_sidecar_column=11)
        rebuilt = preserved_headers + dynamic_rows + terms_rows
        _apply_sidecars(rebuilt, sidecars)
        rebuilt.sort(key=lambda row: int(row.attrib["r"]))
        self.sheet_data[:] = rebuilt

        _rebuild_cotizacion_merges(
            self.root,
            dynamic_merges=dynamic_merges,
            total_delta=total_delta,
        )
        _update_dimension(self.root, total_delta)
        return CotizacionSheetMutation(
            xml=ET.tostring(self.root, encoding="utf-8", xml_declaration=True),
            total_row=total_row,
            images=tuple(images),
            terms_row_delta=total_delta,
            product_rows=tuple(product_rows),
            product_mobiliti_rows=tuple(
                tuple(term.mobiliti_row for term in product.price_terms)
                for product in products
            ),
            composer_variant=composer_variant,
            section_subtotal_rows=tuple(section_subtotal_rows),
            quote_currency=metadata.quote_currency,
        )

    def _write_metadata(self, metadata: CotizacionMetadata) -> None:
        rows = {
            int(row.attrib["r"]): row
            for row in self.sheet_data.findall(f"{{{MAIN}}}row")
        }
        values = {
            "B3": metadata.quotation_number,
            "B7": metadata.project,
            "B8": metadata.client,
            "B9": metadata.email,
            "B10": metadata.phone,
            "B11": metadata.address,
            "B12": metadata.business_name,
        }
        for coordinate, value in values.items():
            row_number = int(_CELL.fullmatch(coordinate).group("row"))
            if row_number not in rows:
                raise ValueError(f"Encabezado oficial ausente: {coordinate}")
            _set_inline_string(rows[row_number], coordinate, value)


def compose_official_quote(request: ComposeRequest) -> PackageAudit:
    """Compone, audita y publica una cotización oficial de forma atómica."""

    if not isinstance(request, ComposeRequest):
        raise TypeError("Solicitud de composición inválida")
    template, output = _validate_compose_paths(request.template, request.output)
    verify_official_template(template, request.contract)
    base = XlsxPackage.read(template)
    mutation = build_allowlisted_mutation(base, request)
    payload = base.to_bytes(mutation)
    candidate_package = XlsxPackage.from_bytes(payload)
    audit = assert_packages_preserved(base, candidate_package, mutation.allowed_parts)
    expected_defined_names = _defined_name_signatures(
        mutation.replacements["xl/workbook.xml"]
    )
    verify_output_contract(
        candidate_package,
        request.contract,
        request.mobiliti.row_map,
        cotizacion_total_row=request.cotizacion.total_row,
        expected_defined_names=expected_defined_names,
    )
    candidate = _write_candidate(output, payload)
    try:
        _atomic_publish_no_replace(candidate, output)
    except BaseException:
        if candidate.exists():
            _recycle_candidate(candidate)
        raise
    return audit


def _serialize_xml_like_source(
    candidate: bytes,
    source: bytes,
    label: str,
) -> bytes:
    """Serializa conservando los prefijos MC declarados por la parte oficial."""

    if not isinstance(candidate, bytes) or not isinstance(source, bytes):
        raise TypeError(f"XML {label} invalido")
    _source_root, namespace_prefixes = _parse_xml_document(
        source,
        f"XML oficial {label} invalido",
    )
    try:
        root = ET.fromstring(candidate)
    except ET.ParseError as error:
        raise ValueError(f"XML mutado {label} invalido") from error
    return _xml_bytes(root, namespace_prefixes)


def build_allowlisted_mutation(
    base: XlsxPackage,
    request: ComposeRequest,
) -> PackageMutation:
    """Materializa la mutación concreta y rechaza superficies no declaradas."""

    if not isinstance(base, XlsxPackage) or not isinstance(request, ComposeRequest):
        raise TypeError("Entrada del compositor inválida")
    mutable_sheets = frozenset(request.contract.mutable_sheets)
    official_mutable_sheets = frozenset({"Mobiliti", "Cotizacion"})
    cdmx_mutable_sheets = frozenset(
        {"Mobiliti", "Cotizacion", CDMX_LUMBRO_SHEET}
    )
    if mutable_sheets not in {official_mutable_sheets, cdmx_mutable_sheets}:
        raise ValueError("Contrato de hojas mutables incompatible")
    _validate_declared_sheet_surfaces(base, request)

    cotizacion = merge_cotizacion_product_images(base, request.cotizacion)
    mobiliti_part = base.sheet_part("Mobiliti")
    cotizacion_part = base.sheet_part("Cotizacion")
    replacements: dict[str, bytes] = {
        mobiliti_part: _serialize_xml_like_source(
            request.mobiliti.xml,
            base.parts[mobiliti_part],
            "Mobiliti",
        ),
        cotizacion_part: _serialize_xml_like_source(
            cotizacion.xml,
            base.parts[cotizacion_part],
            "Cotizacion",
        ),
    }
    cantidades_lumbro_end_row: int | None = None
    cantidades_lumbro_part: str | None = None
    if CDMX_LUMBRO_SHEET in mutable_sheets:
        cantidades_lumbro_part = base.sheet_part(CDMX_LUMBRO_SHEET)
        cantidades_lumbro_xml, cantidades_lumbro_end_row = (
            _build_cantidades_lumbro_sheet(
                base.parts[cantidades_lumbro_part],
                _count_lumbro_items(request.mobiliti),
            )
        )
        replacements[cantidades_lumbro_part] = cantidades_lumbro_xml
    mobiliti_drawing_replacements = _relocate_mobiliti_auxiliary_drawing(
        base,
        request.mobiliti.row_map,
    )
    _merge_disjoint(
        replacements,
        mobiliti_drawing_replacements,
        "dibujo auxiliar Mobiliti",
    )
    additions: dict[str, bytes] = {}
    _merge_disjoint(replacements, cotizacion.related_parts, "partes Cotizacion")
    _merge_disjoint(additions, cotizacion.related_additions, "adiciones Cotizacion")

    fletes_part = base.sheet_part("Fletes")
    estrategia_part = base.sheet_part("Estrategia Comercial ")
    replacements[fletes_part] = _serialize_xml_like_source(
        _translate_fletes(base.parts[fletes_part], request.mobiliti.row_map),
        base.parts[fletes_part],
        "Fletes",
    )
    replacements[estrategia_part] = _serialize_xml_like_source(
        _translate_estrategia(
            base.parts[estrategia_part],
            request.mobiliti.row_map,
            request.cotizacion.total_row,
        ),
        base.parts[estrategia_part],
        "Estrategia Comercial",
    )

    sheet_additions = tuple(
        item
        for item in (request.quotation, request.quotation_data)
        if item is not None
    )
    validated_addition_replacements = _validate_sheet_addition_replacements(
        base,
        sheet_additions,
    )
    _validate_imported_formula_surfaces(sheet_additions)
    workbook_xml, workbook_rels, content_types, added_parts, extra_replacements = (
        _add_workbook_sheets(
            base,
            sheet_additions,
            cotizacion_terms_delta=request.cotizacion.terms_row_delta,
            extra_additions=cotizacion.related_additions,
            extra_content_types=cotizacion.related_content_types,
        )
    )
    if cantidades_lumbro_end_row is not None:
        workbook_xml = _set_sheet_print_area(
            workbook_xml,
            base=base,
            sheet_name=CDMX_LUMBRO_SHEET,
            first_column="H",
            first_row=1,
            last_column="P",
            last_row=cantidades_lumbro_end_row,
        )
    replacements["xl/workbook.xml"] = workbook_xml
    replacements["xl/_rels/workbook.xml.rels"] = workbook_rels
    replacements["[Content_Types].xml"] = content_types
    _merge_replacements(replacements, extra_replacements)
    _merge_disjoint(additions, added_parts, "hojas agregadas")

    calc_chain_part = _calc_chain_part(base)
    if calc_chain_part is not None:
        replacements[calc_chain_part] = _translate_official_calc_chain(
            base,
            request.mobiliti,
            request.cotizacion,
        )

    fixed_replacements = {
        base.sheet_part("Mobiliti"),
        base.sheet_part("Cotizacion"),
        fletes_part,
        estrategia_part,
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "[Content_Types].xml",
        *cotizacion.related_parts,
        *mobiliti_drawing_replacements,
        *validated_addition_replacements,
    }
    if cantidades_lumbro_part is not None:
        fixed_replacements.add(cantidades_lumbro_part)
    if calc_chain_part is not None:
        fixed_replacements.add(calc_chain_part)
    unexpected_replacements = set(replacements) - fixed_replacements
    if unexpected_replacements:
        raise ValueError(
            f"Reemplazo fuera del contrato fijo: {sorted(unexpected_replacements)}"
        )

    protected = tuple(request.contract.protected_prefixes)
    illegal = {
        name
        for name in (*replacements, *additions)
        if part_name_has_ascii_prefix(name, protected)
    }
    if illegal:
        raise ValueError(f"Mutación de parte protegida: {sorted(illegal)}")
    return PackageMutation(replacements=replacements, additions=additions)


def verify_output_contract(
    output: Path | XlsxPackage,
    contract: TemplateContract,
    row_map: MobilitiRowMap,
    *,
    cotizacion_total_row: int,
    expected_defined_names: Sequence[
        tuple[tuple[tuple[str, str], ...], str]
    ],
) -> None:
    """Verifica invariantes oficiales antes de publicar el candidato."""

    package = output if isinstance(output, XlsxPackage) else XlsxPackage.read(Path(output))
    for part_name, payload in package.parts.items():
        if part_name.endswith(".xml"):
            _parse_xml_document(
                payload,
                f"XML de salida invalido: {part_name}",
            )
    rows = package._sheet_rows()
    names = [name for name, _state, _index, _part in rows]
    if len(names) != len({name.casefold() for name in names}):
        raise ValueError("La salida contiene hojas duplicadas")
    states = {name: state for name, state, _index, _part in rows}
    for name, state in contract.sheet_states.items():
        if states.get(name) != state:
            raise ValueError(f"Estado oficial de hoja alterado: {name}")
    if states.get("Quotation_Data") != "veryHidden":
        raise ValueError("Quotation_Data no quedó veryHidden")

    external_parts = sum(
        part_name_has_ascii_prefix(name, ("xl/externalLinks/",))
        for name in package.parts
    )
    if external_parts != contract.external_link_parts:
        raise ValueError("Cantidad de externalLinks alterada")
    spec_formulas = 0
    for name, _state, _index, part in rows:
        if name.strip().casefold().startswith("spec"):
            spec_formulas += len(
                ET.fromstring(package.parts[part]).findall(f".//{{{MAIN}}}f")
            )
    if spec_formulas != contract.spec_formula_count:
        raise ValueError("Fórmulas SPEC alteradas")

    workbook = ET.fromstring(package.parts["xl/workbook.xml"])
    calc_properties = workbook.find(f"{{{MAIN}}}calcPr")
    if calc_properties is None or any(
        calc_properties.attrib.get(name) != value
        for name, value in {
            "calcMode": "auto",
            "fullCalcOnLoad": "1",
            "forceFullCalc": "1",
        }.items()
    ):
        raise ValueError("La salida no fuerza el recalculo seguro de Excel")
    calc_chain_part = _calc_chain_part(package)
    if calc_chain_part is None:
        raise ValueError("La salida perdio calcChain")
    for sheet_name in ("Mobiliti", "Cotizacion"):
        _validate_calc_chain_sheet_coordinates(
            package.parts[calc_chain_part],
            sheet_id=_workbook_sheet_id(package, sheet_name),
            expected=set(),
        )
    defined_names = workbook.findall(f"{{{MAIN}}}definedNames/{{{MAIN}}}definedName")
    actual_defined_names = _defined_name_signatures(package.parts["xl/workbook.xml"])
    normalized_expected_names = tuple(sorted(tuple(expected_defined_names)))
    if (
        len(normalized_expected_names) < contract.defined_name_count
        or actual_defined_names != normalized_expected_names
        or len(defined_names) != len(normalized_expected_names)
    ):
        raise ValueError("La salida alteró el conjunto exacto de nombres definidos")
    if _formula_at(package, "Fletes", "D19") != f"Mobiliti!H{row_map.total_row}":
        raise ValueError("Referencia Fletes!D19 desactualizada")
    if _formula_at(package, "Estrategia Comercial ", "D59") != (
        f"Cotizacion!H{cotizacion_total_row}"
    ):
        raise ValueError("Referencia Estrategia!D59 desactualizada")
    estrategia = ET.fromstring(package.parts[package.sheet_part("Estrategia Comercial ")])
    for row in range(7, 38):
        for column in ("B", "C"):
            formula = _formula_in_root(estrategia, f"{column}{row}")
            if formula is None or f"${row_map.last_product_row}" not in formula:
                raise ValueError("Rangos de Estrategia Comercial desactualizados")

    mobiliti = ET.fromstring(package.parts[package.sheet_part("Mobiliti")])
    if _formula_in_root(mobiliti, "K6") is None:
        raise ValueError("La fórmula oficial Mobiliti!K6 fue eliminada")
    first_product_row = row_map.sections[0].product_start
    last_product_row = row_map.last_product_row
    for target_row in row_map.item_rows:
        for column in ("W", "X", "Y"):
            if _formula_in_root(mobiliti, f"{column}{target_row}") is None:
                raise ValueError(f"Fórmula oficial Mobiliti!{column}{target_row} ausente")
        x_formula = _formula_in_root(mobiliti, f"X{target_row}") or ""
        expected_x_formula = (
            f"_xlfn.MINIFS($W${first_product_row}:$W${last_product_row},"
            f"$D${first_product_row}:$D${last_product_row},D{target_row},"
            f"$H${first_product_row}:$H${last_product_row},"
            f"_xlfn.MAXIFS($H${first_product_row}:$H${last_product_row},"
            f"$D${first_product_row}:$D${last_product_row},D{target_row}))"
        )
        if x_formula != expected_x_formula:
            raise ValueError(
                f"Precio uniforme Mobiliti!X{target_row} desactualizado"
            )
        if _formula_in_root(mobiliti, f"Y{target_row}") != (
            f"(X{target_row}*H{target_row})"
        ):
            raise ValueError(f"Importe uniforme Mobiliti!Y{target_row} desactualizado")
        price = _cell_in_root(mobiliti, f"J{target_row}")
        if price is None:
            raise ValueError(f"Mobiliti!J{target_row} no contiene un costo")
        price_formula = price.find(f"{{{MAIN}}}f")
        if price_formula is not None and re.fullmatch(
            r"(?:Quotation!K[1-9][0-9]*|"
            r"ROUND\(Quotation!K[1-9][0-9]*"
            r"\*[0-9]+(?:\.[0-9]+)?,2\))",
            price_formula.text or "",
        ) is None:
            raise ValueError(
                f"Mobiliti!J{target_row} no referencia un costo de Quotation"
            )


def merge_cotizacion_product_images(
    base: XlsxPackage,
    mutation: CotizacionSheetMutation,
) -> CotizacionSheetMutation:
    """Preserva dibujo estático, limpia anclas contaminadas y agrega imágenes."""

    if mutation.related_parts or mutation.related_additions:
        raise ValueError("Cotizacion ya contiene partes relacionadas no verificadas")
    image_rows = tuple(image.target_row for image in mutation.images)
    if not set(image_rows).issubset(mutation.product_rows):
        raise ValueError("Las imágenes Cotizacion deben pertenecer a filas de producto")
    images_by_row: dict[int, list[CotizacionProductImage]] = {}
    for image in mutation.images:
        images_by_row.setdefault(image.target_row, []).append(image)
    for images in images_by_row.values():
        positions = tuple(image.position for image in images)
        if positions != tuple(range(len(images))):
            raise ValueError(
                "Las posiciones de imágenes Cotizacion deben ser únicas y consecutivas"
            )
    sheet_part = base.sheet_part("Cotizacion")
    sheet_root = ET.fromstring(base.parts[sheet_part])
    drawing_nodes = sheet_root.findall(f"{{{MAIN}}}drawing")
    if len(drawing_nodes) != 1:
        if mutation.images:
            raise ValueError("Cotizacion oficial no contiene un dibujo único")
        return mutation
    relationship_id = drawing_nodes[0].attrib.get(f"{{{REL}}}id")
    sheet_rels_part = relationship_part_name(sheet_part)
    rels = ET.fromstring(base.parts[sheet_rels_part])
    relationship = next(
        (
            item
            for item in rels.findall(f"{{{PKG_REL}}}Relationship")
            if item.attrib.get("Id") == relationship_id
            and item.attrib.get("Type") in relationship_type_uris("drawing")
        ),
        None,
    )
    if relationship is None:
        raise ValueError("Relación de dibujo Cotizacion inválida")
    drawing_part = resolve_internal_target(sheet_part, relationship.attrib["Target"])
    drawing_rels_part = relationship_part_name(drawing_part)
    drawing_root = ET.fromstring(base.parts[drawing_part])
    drawing_rels = ET.fromstring(base.parts[drawing_rels_part])

    for anchor in list(drawing_root):
        marker = anchor.find(f"{{{XDR}}}from")
        if marker is None:
            continue
        row_node = marker.find(f"{{{XDR}}}row")
        column_node = marker.find(f"{{{XDR}}}col")
        if row_node is None or column_node is None or row_node.text is None or column_node.text is None:
            raise ValueError("Ancla de dibujo Cotizacion inválida")
        worksheet_row = int(row_node.text) + 1
        worksheet_column = int(column_node.text) + 1
        if 16 <= worksheet_row <= 18 and 1 <= worksheet_column <= 10:
            drawing_root.remove(anchor)
            continue
        if worksheet_row >= CANONICAL_COTIZACION_TERMS_START:
            _shift_anchor_rows(anchor, mutation.terms_row_delta)

    related_additions: dict[str, bytes] = {}
    content_types: dict[str, str] = {}
    next_picture_id = _next_picture_id(drawing_root)
    used_relationship_ids = {
        item.attrib.get("Id", "")
        for item in drawing_rels.findall(f"{{{PKG_REL}}}Relationship")
    }
    row_sequences = {
        row: sequence
        for sequence, row in enumerate(
            (row for row in mutation.product_rows if row in images_by_row),
            start=1,
        )
    }
    for sequence, image in enumerate(mutation.images, start=1):
        content, extension, content_type, width, height = _read_product_image(image)
        media_part = _allocate_media_part(base, related_additions, sequence, extension)
        rel_id = _next_relationship_id(used_relationship_ids)
        used_relationship_ids.add(rel_id)
        ET.SubElement(
            drawing_rels,
            f"{{{PKG_REL}}}Relationship",
            {
                "Id": rel_id,
                "Type": _stable_image_relationship_type(drawing_rels),
                "Target": posixpath.relpath(media_part, posixpath.dirname(drawing_part)),
            },
        )
        row_images = images_by_row[image.target_row]
        row_sequence = row_sequences[image.target_row]
        if len(row_images) == 1:
            image_name = f"Imagen de producto {sequence:04d}"
            image_box = None
        elif image.position == 0:
            image_name = f"Imagen principal {row_sequence:04d}"
            image_box = _cotizacion_image_box(image.position, len(row_images))
        else:
            image_name = (
                f"Imagen complemento {row_sequence:04d}-{image.position:02d}"
            )
            image_box = _cotizacion_image_box(image.position, len(row_images))
        drawing_root.append(
            _product_image_anchor(
                relationship_id=rel_id,
                picture_id=next_picture_id,
                target_row=image.target_row,
                width=width,
                height=height,
                name=image_name,
                cell_width=_COTIZACION_IMAGE_CELL_WIDTH_EMU,
                cell_height=_COTIZACION_IMAGE_CELL_HEIGHT_EMU,
                box=image_box,
            )
        )
        next_picture_id += 1
        related_additions[media_part] = content
        content_types[media_part] = content_type

    related_parts = {
        drawing_part: ET.tostring(drawing_root, encoding="utf-8", xml_declaration=True),
        drawing_rels_part: ET.tostring(
            drawing_rels, encoding="utf-8", xml_declaration=True
        ),
    }
    return CotizacionSheetMutation(
        xml=mutation.xml,
        related_parts=related_parts,
        total_row=mutation.total_row,
        related_additions=related_additions,
        related_content_types=content_types,
        images=mutation.images,
        terms_row_delta=mutation.terms_row_delta,
        product_rows=mutation.product_rows,
        product_mobiliti_rows=mutation.product_mobiliti_rows,
        composer_variant=mutation.composer_variant,
        section_subtotal_rows=mutation.section_subtotal_rows,
        quote_currency=mutation.quote_currency,
    )


def _relocate_mobiliti_auxiliary_drawing(
    base: XlsxPackage,
    row_map: MobilitiRowMap,
) -> dict[str, bytes]:
    """Desplaza sólo los objetos ubicados bajo la tabla canónica de Mobiliti."""

    delta = row_map.total_row - CANONICAL_MOBILITI_TOTAL_ROW
    if delta == 0:
        return {}

    sheet_part = base.sheet_part("Mobiliti")
    sheet_root = ET.fromstring(base.parts[sheet_part])
    drawing_nodes = sheet_root.findall(f"{{{MAIN}}}drawing")
    if len(drawing_nodes) != 1:
        raise ValueError("Mobiliti oficial no contiene un dibujo único")
    relationship_id = drawing_nodes[0].attrib.get(f"{{{REL}}}id")
    relationships_part = relationship_part_name(sheet_part)
    relationships = ET.fromstring(base.parts[relationships_part])
    relationship = next(
        (
            item
            for item in relationships.findall(f"{{{PKG_REL}}}Relationship")
            if item.attrib.get("Id") == relationship_id
            and item.attrib.get("Type") in relationship_type_uris("drawing")
        ),
        None,
    )
    if relationship is None:
        raise ValueError("Relación de dibujo Mobiliti inválida")
    drawing_part = resolve_internal_target(
        sheet_part,
        relationship.attrib["Target"],
    )
    source = base.parts.get(drawing_part)
    if source is None:
        raise ValueError("Parte de dibujo Mobiliti ausente")
    drawing_root, namespace_prefixes = _parse_xml_document(
        source,
        "Dibujo oficial Mobiliti inválido",
    )

    shifted = 0
    for anchor in drawing_root:
        marker = anchor.find(f"{{{XDR}}}from")
        if marker is None:
            continue
        row = marker.find(f"{{{XDR}}}row")
        if row is None or row.text is None:
            raise ValueError("Ancla de dibujo Mobiliti inválida")
        if int(row.text) + 1 >= CANONICAL_MOBILITI_AUX_START:
            _shift_anchor_rows(anchor, delta)
            shifted += 1
    if shifted == 0:
        raise ValueError("Dibujo auxiliar Mobiliti no encontrado")
    return {drawing_part: _xml_bytes(drawing_root, namespace_prefixes)}


def _stable_image_relationship_type(drawing_rels: ET.Element) -> str:
    allowed = relationship_type_uris("image")
    for relationship in drawing_rels.findall(f"{{{PKG_REL}}}Relationship"):
        relationship_type = relationship.attrib.get("Type", "")
        if relationship_type in allowed:
            return relationship_type
    return f"{OFFICE_DOCUMENT_RELATIONSHIPS}/image"


def _translate_fletes(payload: bytes, row_map: MobilitiRowMap) -> bytes:
    root = _worksheet_root(payload, "Fletes")
    sheet_data = root.find(f"{{{MAIN}}}sheetData")
    if sheet_data is None:
        raise ValueError("Fletes no contiene sheetData")
    rows = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{MAIN}}}row")
    }
    required_rows = (
        set(FLETE_ROUTES)
        | set(FLETE_CATEGORY_ROWS)
        | set(FLETE_CATEGORY_LABEL_ROWS)
    )
    if not required_rows.issubset(rows):
        raise ValueError("Filas oficiales de Fletes incompletas")
    for row, (origin, destination) in FLETE_ROUTES.items():
        _set_inline_string(rows[row], f"A{row}", origin)
        _set_inline_string(rows[row], f"C{row}", destination)
    for row, category in FLETE_CATEGORY_LABEL_ROWS.items():
        _set_inline_string(rows[row], f"I{row}", category)
        _set_inline_string(rows[row], f"M{row}", category)
    for row, (category, factor, installation) in FLETE_CATEGORY_ROWS.items():
        _set_inline_string(rows[row], f"I{row}", category)
        _set_number(rows[row], f"J{row}", factor)
        _set_formula(
            rows[row],
            f"K{row}",
            (
                f"=IF(Mobiliti!$K$4=TRUE,"
                f"((J{row}/$J$21)*Mobiliti!$P$9)/Mobiliti!$K$6,"
                f"(J{row}/$J$21)*Mobiliti!$P$9)"
            ),
        )
        _set_inline_string(rows[row], f"M{row}", category)
        installation_text = format(installation, "f")
        _set_formula(
            rows[row],
            f"N{row}",
            (
                f"=IF(Mobiliti!$K$4=TRUE,"
                f"({installation_text}/Mobiliti!$K$6),"
                f"{installation_text})"
            ),
        )
    source = _formula_in_root(root, "D19")
    if source is None or _formula_range_tokens(source).count("Mobiliti!H573") != 1:
        raise ValueError("Fórmula oficial inesperada: Fletes!D19")
    translated = _translate_formula_with_overrides(
        source,
        coordinate="D19",
        sheet="Fletes",
        overrides={"Mobiliti!H573": f"Mobiliti!H{row_map.total_row}"},
    )
    _replace_formula(root, "D19", "=" + translated)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _validate_declared_sheet_surfaces(
    base: XlsxPackage,
    request: ComposeRequest,
) -> None:
    """Reconstruye la allowlist tipada y compara la estructura completa."""

    _validate_exact_mobiliti_surface(base, request.mobiliti)
    _validate_exact_cotizacion_surface(
        base,
        request.cotizacion,
        request.mobiliti.row_map,
        project_composition=request.project_composition,
    )


def _count_lumbro_items(mutation: MobilitiSheetMutation) -> int:
    """Cuenta únicamente renglones reales cuyo proveedor pertenece a Lumbro."""

    if not isinstance(mutation, MobilitiSheetMutation):
        raise TypeError("Mutación Mobiliti inválida para Cantidades Lumbro")
    root = _worksheet_root(mutation.xml, "Mobiliti")
    count = 0
    for row in mutation.row_map.item_rows:
        cell = _cell_in_root(root, f"F{row}")
        if cell is None:
            continue
        if cell.attrib.get("t") != "inlineStr":
            raise ValueError(f"Proveedor Mobiliti!F{row} no es texto canónico")
        value = "".join(
            node.text or ""
            for node in cell.findall(f".//{{{MAIN}}}t")
        )
        if "lumbro" in value.casefold():
            count += 1
    return count


def _build_cantidades_lumbro_sheet(
    payload: bytes,
    item_count: int,
) -> tuple[bytes, int]:
    """Ajusta el lienzo del reporte Lumbro sin duplicar fórmulas de negocio."""

    if not isinstance(payload, bytes):
        raise TypeError("Worksheet Cantidades Lumbro inválido")
    if type(item_count) is not int or item_count < 0:
        raise ValueError("Cantidad de productos Lumbro inválida")
    root, prefixes = _parse_xml_document(
        payload,
        "Worksheet Cantidades Lumbro inválido",
    )
    if root.tag != f"{{{MAIN}}}worksheet":
        raise ValueError("Parte Cantidades Lumbro no es worksheet")
    sheet_data = root.find(f"{{{MAIN}}}sheetData")
    if sheet_data is None:
        raise ValueError("Cantidades Lumbro no contiene sheetData")
    rows = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{MAIN}}}row")
        if row.attrib.get("r", "").isdigit()
    }
    source = rows.get(CDMX_LUMBRO_FIRST_PRODUCT_ROW)
    if source is None:
        raise ValueError("Fila modelo Cantidades Lumbro ausente")

    end_row = max(
        CDMX_LUMBRO_MIN_END_ROW,
        CDMX_LUMBRO_FIRST_PRODUCT_ROW - 1 + item_count,
    )
    for row in list(sheet_data.findall(f"{{{MAIN}}}row")):
        row_number = int(row.attrib.get("r", "0") or "0")
        if row_number > CDMX_LUMBRO_FIRST_PRODUCT_ROW:
            sheet_data.remove(row)
    for target_row in range(CDMX_LUMBRO_FIRST_PRODUCT_ROW + 1, end_row + 1):
        clone = _clone_row_region(
            source,
            CDMX_LUMBRO_FIRST_PRODUCT_ROW,
            target_row,
            last_column=CDMX_LUMBRO_LAST_COLUMN,
        )
        for cell in clone.findall(f"{{{MAIN}}}c"):
            _clear_cell_value(cell)
        sheet_data.append(clone)

    dimension = root.find(f"{{{MAIN}}}dimension")
    if dimension is None:
        dimension = ET.Element(f"{{{MAIN}}}dimension")
        root.insert(0, dimension)
    dimension.attrib["ref"] = f"A1:P{end_row}"
    return _xml_bytes(root, prefixes), end_row


def _set_sheet_print_area(
    workbook_payload: bytes,
    *,
    base: XlsxPackage,
    sheet_name: str,
    first_column: str,
    first_row: int,
    last_column: str,
    last_row: int,
) -> bytes:
    """Actualiza un Print_Area predeclarado para una hoja del template."""

    if not isinstance(base, XlsxPackage):
        raise TypeError("Paquete base inválido para Print_Area")
    for column in (first_column, last_column):
        if re.fullmatch(r"[A-Z]{1,3}", column) is None:
            raise ValueError("Columna de Print_Area inválida")
    if min(first_row, last_row) < 1 or first_row > last_row:
        raise ValueError("Filas de Print_Area inválidas")
    root, prefixes = _parse_xml_document(
        workbook_payload,
        "Workbook inválido al ajustar Print_Area",
    )
    defined_names = root.find(f"{{{MAIN}}}definedNames")
    if defined_names is None:
        raise ValueError("Workbook sin nombres definidos")
    local_sheet_id = str(base.sheet_index(sheet_name))
    matches = [
        node
        for node in defined_names.findall(f"{{{MAIN}}}definedName")
        if node.attrib.get("name") == "_xlnm.Print_Area"
        and node.attrib.get("localSheetId") == local_sheet_id
    ]
    if len(matches) != 1:
        raise ValueError(
            f"Print_Area de {sheet_name.rstrip()} no está declarado exactamente una vez"
        )
    matches[0].text = (
        f"'{sheet_name}'!${first_column}${first_row}:"
        f"${last_column}${last_row}"
    )
    return _xml_bytes(root, prefixes)


def _validate_exact_mobiliti_surface(
    base: XlsxPackage,
    mutation: MobilitiSheetMutation,
) -> None:
    candidate = _worksheet_root(mutation.xml, "Mobiliti")
    _validate_exact_inline_strings(candidate, "Mobiliti")
    row_map = mutation.row_map
    needs = [
        SectionNeed(section.id, section.title, section.item_count)
        for section in row_map.sections
    ]
    input_kinds = {
        "D": ("text", "formula"),
        "E": ("text",),
        "F": ("text",),
        "H": ("number", "formula"),
        "J": ("number", "formula", "text"),
        "K": ("text", "number", "formula"),
        "P": ("text",),
    }
    writes: list[MobilitiCellWrite] = []
    for row in row_map.item_rows:
        for column, allowed_kinds in input_kinds.items():
            coordinate = f"{column}{row}"
            cell = _cell_in_root(candidate, coordinate)
            if cell is None:
                continue
            kind, value = _exact_allowed_typed_value(
                cell,
                allowed_kinds,
                allow_blank=True,
            )
            if column == "J" and kind == "text" and value != "Por confirmar":
                raise ValueError("Texto de precio Mobiliti inválido")
            if kind is not None:
                writes.append(MobilitiCellWrite(coordinate, kind, value))

    official_payload = base.parts[base.sheet_part("Mobiliti")]
    expected = build_mobiliti_sheet(official_payload, needs, writes)
    if expected.row_map != row_map:
        raise ValueError("Mobiliti no cumple el contrato exacto de layout")
    editor = WorksheetEditor.from_xml(expected.xml)
    selector_writes: list[MobilitiCellWrite] = []
    for coordinate, kind in (("K4", "boolean"), ("K8", "text")):
        cell = _cell_in_root(candidate, coordinate)
        if cell is None:
            raise ValueError(f"Mobiliti no cumple el contrato exacto: {coordinate}")
        value = _exact_typed_value(cell, kind, allow_blank=False)
        selector_writes.append(MobilitiCellWrite(coordinate, kind, value))
    editor.set_typed_values(selector_writes)
    _assert_exact_worksheet(editor.root, candidate, "Mobiliti")


_COTIZACION_POSITIVE_DECIMAL = (
    r"(?:[1-9][0-9]*(?:\.[0-9]*[1-9])?|0\.[0-9]*[1-9])"
)
_COTIZACION_PRICE_TERM = re.compile(
    rf"Mobiliti!X(?P<row>[1-9][0-9]*)"
    rf"(?:\*(?P<numerator>{_COTIZACION_POSITIVE_DECIMAL}))?"
    rf"(?:/(?P<denominator>{_COTIZACION_POSITIVE_DECIMAL}))?"
)
_COTIZACION_COMPOSED_PRINCIPAL = re.compile(
    r"Mobiliti!X(?P<row>[1-9][0-9]*)"
)
_COTIZACION_COMPOSED_COMPLEMENT = re.compile(
    r"Mobiliti!X(?P<row>[1-9][0-9]*)"
    r"\*Mobiliti!H(?P<quantity_row>[1-9][0-9]*)"
    r"/Mobiliti!H(?P<principal_row>[1-9][0-9]*)"
)
_COTIZACION_DESCRIPTION_FORMULA = re.compile(
    r"=Quotation!D(?P<row>[1-9][0-9]*)\Z"
)


def _validate_cotizacion_description_formula(formula: str) -> int:
    match = (
        None
        if (
        not isinstance(formula, str)
        or len(formula) > MAX_EXCEL_FORMULA_LENGTH
        )
        else _COTIZACION_DESCRIPTION_FORMULA.fullmatch(formula)
    )
    if match is None:
        raise ValueError("Fórmula de descripción Cotizacion fuera de contrato")
    row = int(match.group("row"))
    if row > XLSX_MAX_ROWS:
        raise ValueError("Fórmula de descripción Cotizacion fuera de contrato")
    return row


def _cotizacion_price_terms_from_formula(
    formula: str,
) -> tuple[CotizacionPriceTerm, ...]:
    if not isinstance(formula, str) or not formula:
        raise ValueError("Cotizacion no cumple el contrato exacto de fórmulas")
    raw_terms = formula.split("+")
    terms: list[CotizacionPriceTerm] = []
    principal_match = (
        _COTIZACION_COMPOSED_PRINCIPAL.fullmatch(raw_terms[0])
        if len(raw_terms) > 1
        else None
    )
    if principal_match is not None:
        principal_row = int(principal_match.group("row"))
        if len(raw_terms) == 1:
            raise ValueError("Cotizacion no cumple el contrato exacto de fórmulas")
        terms.append(CotizacionPriceTerm(principal_row))
        for raw_term in raw_terms[1:]:
            match = _COTIZACION_COMPOSED_COMPLEMENT.fullmatch(raw_term)
            if (
                match is None
                or match.group("row") != match.group("quantity_row")
                or int(match.group("principal_row")) != principal_row
            ):
                raise ValueError(
                    "Cotizacion no cumple el contrato exacto de fórmulas"
                )
            terms.append(CotizacionPriceTerm(int(match.group("row"))))
    else:
        for raw_term in raw_terms:
            match = _COTIZACION_PRICE_TERM.fullmatch(raw_term)
            if match is None:
                raise ValueError(
                    "Cotizacion no cumple el contrato exacto de fórmulas"
                )
            terms.append(
                CotizacionPriceTerm(
                    int(match.group("row")),
                    Decimal(match.group("numerator") or "1"),
                    Decimal(match.group("denominator") or "1"),
                )
            )
    result = tuple(terms)
    expected = CotizacionFormulaContract().product_formulas(
        target_row=1,
        price_terms=result,
    )["F"][1:]
    if formula != expected:
        raise ValueError("Cotizacion no cumple el contrato exacto de fórmulas")
    return result


def _cotizacion_composer_variant(base: XlsxPackage) -> str:
    """Selecciona la única variante permitida a partir del template firmado."""

    if not isinstance(base, XlsxPackage):
        raise TypeError("Paquete base Cotizacion inválido")
    try:
        base.sheet_part(CDMX_LUMBRO_SHEET)
    except KeyError:
        return "official"
    return "sunon_cdmx_v1c"


def _validate_exact_cotizacion_surface(
    base: XlsxPackage,
    mutation: CotizacionSheetMutation,
    row_map: MobilitiRowMap,
    *,
    project_composition: bool = False,
) -> None:
    if type(project_composition) is not bool:
        raise TypeError("Señal de composición de Proyecto inválida")
    candidate = _worksheet_root(mutation.xml, "Cotizacion")
    _validate_exact_inline_strings(candidate, "Cotizacion")
    product_rows = mutation.product_rows
    product_count_invalid = (
        len(product_rows) > len(row_map.item_rows)
        if project_composition
        else len(product_rows) != len(row_map.item_rows)
    )
    if (
        not product_rows
        or tuple(sorted(set(product_rows))) != product_rows
        or product_count_invalid
    ):
        raise ValueError("Cotizacion no cumple el contrato exacto de filas de producto")
    if mutation.terms_row_delta != mutation.total_row - CANONICAL_COTIZACION_TOTAL_ROW:
        raise ValueError("Cotizacion no cumple el contrato exacto de términos")
    subtotal_row = mutation.total_row - CANONICAL_COTIZACION_TOTAL_SPAN
    if subtotal_row <= CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW:
        raise ValueError("Cotizacion no cumple el contrato exacto de totales")
    expected_variant = _cotizacion_composer_variant(base)
    if mutation.composer_variant != expected_variant:
        raise ValueError("Cotizacion no cumple la variante exacta del template")
    section_subtotal_rows = mutation.section_subtotal_rows
    if (
        any(
            row <= CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW
            or row >= subtotal_row
            for row in section_subtotal_rows
        )
        or set(section_subtotal_rows) & set(product_rows)
        or (
            expected_variant == "sunon_cdmx_v1c"
            and not section_subtotal_rows
        )
    ):
        raise ValueError("Cotizacion no cumple los subtotales exactos por sección")
    if any(
        row <= CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW or row >= subtotal_row
        for row in product_rows
    ):
        raise ValueError("Cotizacion no cumple el contrato exacto de filas de producto")

    metadata_coordinates = {
        "quotation_number": "B3",
        "project": "B7",
        "client": "B8",
        "email": "B9",
        "phone": "B10",
        "address": "B11",
        "business_name": "B12",
    }
    metadata_values = {
        field: _exact_inline_text(_require_root_cell(candidate, coordinate), coordinate)
        for field, coordinate in metadata_coordinates.items()
    }
    metadata = CotizacionMetadata(
        **metadata_values,
        quote_currency=mutation.quote_currency,
    )

    sections: list[CotizacionSection] = []
    current_title: str | None = None
    current_products: list[CotizacionProduct] = []
    current_product_rows: list[int] = []
    product_set = set(product_rows)
    section_subtotal_set = set(section_subtotal_rows)
    first_discount: Decimal | None = None
    mobiliti_rows: list[int] = []
    product_sequence = 0
    for worksheet_row in range(CANONICAL_COTIZACION_FIRST_DYNAMIC_ROW, subtotal_row):
        if worksheet_row in section_subtotal_set:
            if (
                expected_variant != "sunon_cdmx_v1c"
                or current_title is None
                or not current_products
                or not current_product_rows
            ):
                raise ValueError(
                    "Cotizacion no cumple los subtotales exactos por sección"
                )
            if _exact_inline_text(
                _require_root_cell(candidate, f"I{worksheet_row}"),
                f"I{worksheet_row}",
            ) != "SUBTOTAL AREA":
                raise ValueError(
                    "Cotizacion no cumple los subtotales exactos por sección"
                )
            expected_section_formula = (
                f"SUM(J{current_product_rows[0]}:J{current_product_rows[-1]})"
            )
            if _exact_formula_text(
                _require_root_cell(candidate, f"J{worksheet_row}"),
                f"J{worksheet_row}",
            ) != expected_section_formula:
                raise ValueError(
                    "Cotizacion no cumple los subtotales exactos por sección"
                )
            sections.append(CotizacionSection(current_title, tuple(current_products)))
            current_title = None
            current_products = []
            current_product_rows = []
            continue
        if worksheet_row not in product_set:
            if current_title is not None:
                if not current_products:
                    raise ValueError("Cotizacion no cumple el contrato exacto de secciones")
                if expected_variant == "sunon_cdmx_v1c":
                    raise ValueError(
                        "Cotizacion CDMX omite el subtotal de una sección"
                    )
                sections.append(
                    CotizacionSection(current_title, tuple(current_products))
                )
            current_title = _exact_inline_text(
                _require_root_cell(candidate, f"A{worksheet_row}"),
                f"A{worksheet_row}",
            )
            current_products = []
            current_product_rows = []
            continue
        if current_title is None:
            raise ValueError("Cotizacion no cumple el contrato exacto de secciones")
        product_sequence += 1
        declared_mobiliti_rows = mutation.product_mobiliti_rows[
            product_sequence - 1
        ]
        if first_discount is None:
            first_discount = _exact_typed_value(
                _require_root_cell(candidate, f"G{worksheet_row}"),
                "number",
                allow_blank=False,
            )
        price_cell = _require_root_cell(candidate, f"F{worksheet_row}")
        price_pending = price_cell.attrib.get("t") == "inlineStr"
        if price_pending:
            if (
                _exact_inline_text(price_cell, f"F{worksheet_row}")
                != "Por confirmar"
                or _exact_inline_text(
                    _require_root_cell(candidate, f"H{worksheet_row}"),
                    f"H{worksheet_row}",
                ) != ""
                or _exact_inline_text(
                    _require_root_cell(candidate, f"I{worksheet_row}"),
                    f"I{worksheet_row}",
                ) != "Por confirmar"
                or _exact_inline_text(
                    _require_root_cell(candidate, f"J{worksheet_row}"),
                    f"J{worksheet_row}",
                ) != ""
            ):
                raise ValueError("Estado de precio Cotizacion inválido")
            name_formula = _exact_formula_text(
                _require_root_cell(candidate, f"A{worksheet_row}"),
                f"A{worksheet_row}",
            )
            match = re.fullmatch(r"Mobiliti!D([1-9][0-9]*)", name_formula)
            if (
                match is None
                or int(match.group(1)) != declared_mobiliti_rows[0]
            ):
                raise ValueError("Referencia pendiente Cotizacion inválida")
            price_terms = tuple(
                CotizacionPriceTerm(row) for row in declared_mobiliti_rows
            )
        else:
            formula = _exact_formula_text(price_cell, f"F{worksheet_row}")
            match = re.fullmatch(r"Mobiliti!X([1-9][0-9]*)", formula)
            if match is None:
                price_terms = _cotizacion_price_terms_from_formula(formula)
            else:
                price_terms = (CotizacionPriceTerm(int(match.group(1))),)
            if tuple(term.mobiliti_row for term in price_terms) != declared_mobiliti_rows:
                raise ValueError(
                    "Cotizacion no cumple el contrato exacto de filas Mobiliti"
                )
        mobiliti_row = price_terms[0].mobiliti_row
        mobiliti_rows.extend(term.mobiliti_row for term in price_terms)
        description_cell = _require_root_cell(candidate, f"C{worksheet_row}")
        description_node = description_cell.find(f"{{{MAIN}}}f")
        if description_node is None:
            description = _exact_inline_text(
                description_cell,
                f"C{worksheet_row}",
            )
            description_formula = None
            name = _exact_inline_text(
                _require_root_cell(candidate, f"A{worksheet_row}"),
                f"A{worksheet_row}",
            )
            dimensions = _exact_inline_text(
                _require_root_cell(candidate, f"D{worksheet_row}"),
                f"D{worksheet_row}",
            )
            quantity = _exact_typed_value(
                _require_root_cell(candidate, f"E{worksheet_row}"),
                "number",
                allow_blank=False,
            )
        else:
            description_formula = (
                "="
                + _exact_formula_text(
                    description_cell,
                    f"C{worksheet_row}",
                )
            )
            quotation_row = _validate_cotizacion_description_formula(
                description_formula
            )
            if (
                "="
                + _exact_formula_text(
                    _require_root_cell(candidate, f"A{worksheet_row}"),
                    f"A{worksheet_row}",
                )
                != f"=Mobiliti!D{mobiliti_row}"
                or "="
                + _exact_formula_text(
                    _require_root_cell(candidate, f"D{worksheet_row}"),
                    f"D{worksheet_row}",
                )
                != f"=Quotation!F{quotation_row}"
                or "="
                + _exact_formula_text(
                    _require_root_cell(candidate, f"E{worksheet_row}"),
                    f"E{worksheet_row}",
                )
                != f"=Mobiliti!H{mobiliti_row}"
            ):
                raise ValueError(
                    "Cotizacion no cumple el contrato exacto de referencias"
                )
            description = ""
            name = "referenced"
            dimensions = ""
            quantity = Decimal("1")
        current_products.append(
            CotizacionProduct(
                item_key=f"contract-{product_sequence}",
                name=name,
                description=description,
                dimensions=dimensions,
                quantity=quantity,
                mobiliti_row=mobiliti_row,
                discount=first_discount,
                description_formula=description_formula,
                price_terms=price_terms,
                price_pending=price_pending,
            )
        )
        current_product_rows.append(worksheet_row)
    if expected_variant == "sunon_cdmx_v1c":
        if (
            current_title is not None
            or current_products
            or current_product_rows
            or len(sections) != len(section_subtotal_rows)
        ):
            raise ValueError(
                "Cotizacion no cumple los subtotales exactos por sección"
            )
    else:
        if current_title is None or not current_products:
            raise ValueError("Cotizacion no cumple el contrato exacto de secciones")
        sections.append(CotizacionSection(current_title, tuple(current_products)))
    if project_composition:
        mobiliti_contract_invalid = (
            len(mobiliti_rows) != len(set(mobiliti_rows))
            or set(mobiliti_rows) != set(row_map.item_rows)
        )
    else:
        mobiliti_contract_invalid = tuple(mobiliti_rows) != row_map.item_rows
    if mobiliti_contract_invalid:
        raise ValueError("Cotizacion no cumple el contrato exacto de filas Mobiliti")

    expected = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=metadata,
        sections=tuple(sections),
        composer_variant=expected_variant,
    )
    if (
        expected.total_row != mutation.total_row
        or expected.terms_row_delta != mutation.terms_row_delta
        or expected.product_rows != mutation.product_rows
        or expected.composer_variant != mutation.composer_variant
        or expected.section_subtotal_rows != mutation.section_subtotal_rows
        or expected.quote_currency != mutation.quote_currency
    ):
        raise ValueError("Cotizacion no cumple el contrato exacto de layout")
    expected_root = _worksheet_root(expected.xml, "Cotizacion")
    _assert_exact_worksheet(expected_root, candidate, "Cotizacion")


def _require_root_cell(root: ET.Element, coordinate: str) -> ET.Element:
    cell = _cell_in_root(root, coordinate)
    if cell is None:
        raise ValueError(f"Celda requerida ausente: {coordinate}")
    return cell


def _validate_exact_inline_strings(root: ET.Element, sheet: str) -> None:
    for cell in root.findall(f".//{{{MAIN}}}c[@t='inlineStr']"):
        coordinate = cell.attrib.get("r", "?")
        _exact_inline_text(cell, f"{sheet}!{coordinate}")


def _exact_inline_text(cell: ET.Element, coordinate: str) -> str:
    children = list(cell)
    if cell.attrib.get("t") != "inlineStr" or len(children) != 1:
        raise ValueError(f"inlineStr exacto inválido: {coordinate}")
    inline = children[0]
    if inline.tag != f"{{{MAIN}}}is" or inline.attrib or len(inline) != 1:
        raise ValueError(f"inlineStr exacto inválido: {coordinate}")
    text = inline[0]
    allowed_text_attributes = {f"{{{XML}}}space"}
    if (
        text.tag != f"{{{MAIN}}}t"
        or set(text.attrib) - allowed_text_attributes
        or text.attrib.get(f"{{{XML}}}space") not in {None, "preserve"}
        or len(text)
    ):
        raise ValueError(f"inlineStr exacto inválido: {coordinate}")
    return text.text or ""


def _exact_typed_value(
    cell: ET.Element,
    kind: str,
    *,
    allow_blank: bool,
) -> Decimal | str | bool | None:
    children = list(cell)
    if not children and cell.attrib.get("t") is None and allow_blank:
        return None
    coordinate = cell.attrib.get("r", "?")
    if kind == "text":
        return _exact_inline_text(cell, coordinate)
    if kind == "boolean":
        if (
            cell.attrib.get("t") != "b"
            or len(children) != 1
            or children[0].tag != f"{{{MAIN}}}v"
            or children[0].attrib
            or children[0].text not in {"0", "1"}
        ):
            raise ValueError(f"Booleano tipado inválido: {coordinate}")
        return children[0].text == "1"
    if kind == "number":
        if (
            cell.attrib.get("t") is not None
            or len(children) != 1
            or children[0].tag != f"{{{MAIN}}}v"
            or children[0].attrib
            or children[0].text is None
        ):
            raise ValueError(f"Número tipado inválido: {coordinate}")
        try:
            value = Decimal(children[0].text)
        except Exception as error:
            raise ValueError(f"Número tipado inválido: {coordinate}") from error
        if not value.is_finite():
            raise ValueError(f"Número tipado inválido: {coordinate}")
        return value
    if kind == "formula":
        return "=" + _exact_formula_text(cell, coordinate)
    raise ValueError(f"Tipo de celda no permitido: {kind}")


def _exact_allowed_typed_value(
    cell: ET.Element,
    allowed_kinds: Sequence[str],
    *,
    allow_blank: bool,
) -> tuple[str | None, Decimal | str | bool | None]:
    children = list(cell)
    if not children and cell.attrib.get("t") is None and allow_blank:
        return None, None
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        kind = "text"
    elif cell_type == "b":
        kind = "boolean"
    elif (
        cell_type is None
        and len(children) == 1
        and children[0].tag == f"{{{MAIN}}}f"
    ):
        kind = "formula"
    elif cell_type is None:
        kind = "number"
    else:
        raise ValueError(f"Tipo de celda tipada inválido: {cell.attrib.get('r', '?')}")
    if kind not in allowed_kinds:
        raise ValueError(f"Tipo de celda fuera de allowlist: {cell.attrib.get('r', '?')}")
    return kind, _exact_typed_value(cell, kind, allow_blank=allow_blank)


def _exact_formula_text(cell: ET.Element, coordinate: str) -> str:
    children = list(cell)
    if (
        cell.attrib.get("t") is not None
        or len(children) != 1
        or children[0].tag != f"{{{MAIN}}}f"
        or children[0].attrib
        or children[0].text is None
    ):
        raise ValueError(f"Fórmula exacta inválida: {coordinate}")
    return children[0].text


def _assert_exact_worksheet(
    expected: ET.Element,
    candidate: ET.Element,
    sheet: str,
) -> None:
    if _xml_signature(expected) != _xml_signature(candidate):
        raise ValueError(
            f"{sheet} no cumple el contrato exacto de mutación: "
            "escritura fuera de la superficie mutable"
        )


def _xml_signature(node: ET.Element) -> tuple:
    text = node.text
    if text is not None and not text.strip() and node.tag != f"{{{MAIN}}}t":
        text = None
    return (
        node.tag,
        tuple(sorted(node.attrib.items())),
        text,
        tuple(_xml_signature(child) for child in node),
    )


def _translate_estrategia(
    payload: bytes,
    row_map: MobilitiRowMap,
    cotizacion_total_row: int,
) -> bytes:
    root = _worksheet_root(payload, "Estrategia Comercial ")
    for row in range(7, 38):
        for column in ("B", "C"):
            coordinate = f"{column}{row}"
            formula = _formula_in_root(root, coordinate)
            if formula is None:
                raise ValueError(f"Fórmula oficial ausente: Estrategia!{coordinate}")
            references = []
            overrides: dict[str, str] = {}
            for token in _formula_range_tokens(formula):
                match = re.fullmatch(
                    r"Mobiliti!(?P<first>\$?[A-Z]{1,3}\$14):"
                    r"(?P<column>\$?[A-Z]{1,3}\$)571",
                    token,
                )
                if match is None:
                    continue
                references.append(token)
                overrides[token] = (
                    f"Mobiliti!{match.group('first')}:"
                    f"{match.group('column')}{row_map.last_product_row}"
                )
            if len(references) != 2:
                raise ValueError(f"Rangos oficiales inesperados: Estrategia!{coordinate}")
            translated = _translate_formula_with_overrides(
                formula,
                coordinate=coordinate,
                sheet="Estrategia Comercial ",
                overrides=overrides,
            )
            _replace_formula(root, coordinate, "=" + translated)
    total_formula = _formula_in_root(root, "D59")
    official_total = f"Cotizacion!H{CANONICAL_COTIZACION_TOTAL_ROW}"
    if total_formula is None or _formula_range_tokens(total_formula).count(
        official_total
    ) != 1:
        raise ValueError("Referencia oficial inesperada: Estrategia!D59")
    translated_total = _translate_formula_with_overrides(
        total_formula,
        coordinate="D59",
        sheet="Estrategia Comercial ",
        overrides={official_total: f"Cotizacion!H{cotizacion_total_row}"},
    )
    _replace_formula(root, "D59", "=" + translated_total)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _formula_range_tokens(formula: str) -> list[str]:
    try:
        tokens = Tokenizer("=" + formula).items
    except Exception as error:
        raise ValueError("Fórmula oficial no tokenizable") from error
    return [
        token.value
        for token in tokens
        if token.type == "OPERAND" and token.subtype == "RANGE"
    ]


def _validate_formula_token_contract(
    formula: str,
    expected: tuple[tuple[str, str, str], ...],
) -> None:
    """Valida el contrato F/I por tokens exactos, nunca por sustitución textual."""

    if not isinstance(formula, str) or not formula.startswith("="):
        raise ValueError("Fórmula del contrato F/I inválida")
    try:
        items = tuple(Tokenizer(formula).items)
    except (TokenizerError, IndexError, TypeError, ValueError) as error:
        raise ValueError("Fórmula del contrato F/I no tokenizable") from error
    actual = tuple((item.type, item.subtype, item.value) for item in items)
    if "".join(item.value for item in items) != formula[1:] or actual != expected:
        raise ValueError("Fórmula del contrato F/I fuera de contrato")


def _append_bounded_formula(
    parts: list[str],
    fragment: str,
    current_length: int,
) -> int:
    """Agrega un fragmento sin materializar una fórmula mayor al límite XLSX."""

    next_length = current_length + len(fragment)
    if next_length > MAX_EXCEL_FORMULA_LENGTH:
        raise ValueError("Fórmula Cotizacion excede 8,192 caracteres")
    parts.append(fragment)
    return next_length


def _validate_imported_formula_surfaces(
    sheet_additions: Sequence[SheetAddition],
) -> None:
    for addition in sheet_additions:
        for part, payload, kind in _imported_formula_xml_payloads(addition):
            try:
                root = ET.fromstring(payload)
            except ET.ParseError as error:
                raise ValueError(f"XML importado inválido: {part}") from error
            if root.tag != _FORMULA_PART_ROOTS[kind]:
                raise ValueError(f"Raíz XML de {kind} importado inválida: {part}")
            for surface, path, allow_empty_shared in _FORMULA_ELEMENT_PATHS[kind]:
                for index, formula in enumerate(root.findall(path), start=1):
                    text = formula.text or ""
                    if not text:
                        if (
                            allow_empty_shared
                            and formula.attrib.get("t") == "shared"
                            and formula.attrib.get("si", "").isdigit()
                        ):
                            continue
                        raise ValueError(
                            "Fórmula importada no permitida: fórmula vacía"
                        )
                    _validate_imported_formula(
                        text,
                        f"{addition.name}:{part}:{surface}:{index}",
                    )
        for defined_name in addition.defined_names:
            _validate_imported_formula(
                defined_name.text,
                f"{addition.name}:definedName:{defined_name.name}",
            )


def _imported_formula_xml_payloads(
    addition: SheetAddition,
) -> tuple[tuple[str, bytes, str], ...]:
    """Selecciona XML con fórmulas por ContentType y relaciones OPC exactas."""

    if addition.relationship_type not in relationship_type_uris("worksheet"):
        raise ValueError(f"Relación de worksheet importado inválida: {addition.name}")
    payloads = {**addition.parts, **addition.replacements}
    content_types = {
        **addition.content_types,
        **addition.replacement_content_types,
    }
    if addition.sheet_part is None:
        selected: list[tuple[str, bytes, str]] = [
            (f"<hoja:{addition.name}>", addition.xml, "worksheet")
        ]
    else:
        main_content_type = content_types.get(addition.sheet_part)
        if main_content_type != WORKSHEET_CONTENT_TYPE:
            raise ValueError(
                f"Content type de worksheet importado inválido: {addition.sheet_part}"
            )
        selected = [(addition.sheet_part, addition.xml, "worksheet")]

    related_formula_parts: dict[str, str] = {}
    for rels_part in sorted(payloads):
        if content_types[rels_part] != RELATIONSHIPS_CONTENT_TYPE:
            continue
        try:
            relationships = ET.fromstring(payloads[rels_part])
        except ET.ParseError as error:
            raise ValueError(
                f"Relaciones importadas inválidas: {rels_part}"
            ) from error
        if relationships.tag != f"{{{PKG_REL}}}Relationships":
            raise ValueError(f"Raíz de relaciones importada inválida: {rels_part}")
        owner = relationship_owner(rels_part)
        for relationship in relationships:
            if relationship.tag != f"{{{PKG_REL}}}Relationship":
                raise ValueError(f"Relación importada inválida: {rels_part}")
            kind = _FORMULA_RELATIONSHIP_KIND_BY_TYPE.get(
                relationship.attrib.get("Type", "")
            )
            if kind is None:
                continue
            if owner is None:
                raise ValueError(f"Relación de fórmula sin propietario: {rels_part}")
            target_mode = relationship.attrib.get("TargetMode", "").casefold()
            if target_mode not in {"", "internal"}:
                raise ValueError(f"Relación de fórmula externa: {rels_part}")
            target = resolve_internal_target(
                owner,
                relationship.attrib.get("Target", ""),
            )
            if target not in payloads:
                raise ValueError(f"Clausura de fórmula incompleta: {target}")
            previous = related_formula_parts.setdefault(target, kind)
            if previous != kind:
                raise ValueError(f"Perfil de fórmula ambiguo: {target}")

    for part, kind in related_formula_parts.items():
        if content_types[part] != _FORMULA_PART_CONTENT_TYPES[kind]:
            raise ValueError(f"Content type de fórmula inconsistente: {part}")

    for part in sorted(payloads):
        kind = _FORMULA_PART_KIND_BY_CONTENT_TYPE.get(content_types[part])
        if kind is None or part == addition.sheet_part:
            continue
        if kind != "worksheet":
            related_kind = related_formula_parts.get(part)
            if related_kind is None:
                raise ValueError(f"Parte de fórmula sin relación: {part}")
            if related_kind != kind:
                raise ValueError(f"Content type de fórmula inconsistente: {part}")
        selected.append((part, payloads[part], kind))
    return tuple(selected)


def _validate_imported_formula(formula: str, context: str) -> None:
    if (
        not isinstance(formula, str)
        or not formula
        or len(formula) > MAX_EXCEL_FORMULA_LENGTH
    ):
        raise ValueError(f"Fórmula importada no permitida: {context}")
    body = formula[1:] if formula.startswith("=") else formula
    try:
        items = tuple(Tokenizer("=" + body).items)
    except (TokenizerError, IndexError, TypeError, ValueError) as error:
        raise ValueError(f"Fórmula importada no permitida: {context}") from error
    if "".join(item.value for item in items) != body:
        raise ValueError(f"Fórmula importada no permitida: {context}")
    for position, item in enumerate(items):
        function_name: str | None = None
        if item.type == Token.FUNC and item.subtype == Token.OPEN:
            function_name = item.value[:-1]
        elif item.type == Token.PAREN and item.subtype == Token.OPEN:
            previous = position - 1
            while previous >= 0 and items[previous].type == Token.WSPACE:
                previous -= 1
            if (
                previous >= 0
                and items[previous].type == Token.OPERAND
                and items[previous].subtype == Token.RANGE
            ):
                function_name = items[previous].value
        if function_name is not None:
            normalized_name = function_name.strip().upper()
            while True:
                previous_name = normalized_name
                normalized_name = normalized_name.lstrip("@")
                for prefix in ("_XLFN.", "_XLWS."):
                    if normalized_name.startswith(prefix):
                        normalized_name = normalized_name[len(prefix) :]
                        break
                if normalized_name == previous_name:
                    break
            if normalized_name in _DANGEROUS_IMPORTED_FUNCTIONS:
                raise ValueError(f"Fórmula importada no permitida: {context}")
        if item.type == Token.OPERAND and item.subtype == Token.RANGE:
            value = item.value
            bang = value.find("!")
            qualifier = value[:bang] if bang >= 0 else ""
            folded = value.casefold()
            if (
                "|" in value
                or ("[" in qualifier and "]" in qualifier)
                or folded.startswith(("http:", "https:", "ftp:", "file:", "\\\\"))
            ):
                raise ValueError(f"Fórmula importada no permitida: {context}")
        if item.type == Token.OPERAND and item.subtype == Token.TEXT:
            folded = item.value.casefold()
            if any(
                marker in folded
                for marker in ("http://", "https://", "ftp://", "file://", "\\\\")
            ):
                raise ValueError(f"Fórmula importada no permitida: {context}")


def _validate_sheet_addition_replacements(
    base: XlsxPackage,
    sheet_additions: Sequence[SheetAddition],
) -> frozenset[str]:
    """Valida reemplazos de hojas contra un contrato fijo por tipo."""

    validated: set[str] = set()
    styles_part = base.workbook_related_part("styles")
    for addition in sheet_additions:
        if addition.name == "Quotation_Data":
            if addition.replacements:
                raise ValueError("Quotation_Data no permite reemplazos")
            continue
        if addition.name != "Quotation":
            raise ValueError("Hoja no permitida por el compositor")
        allowed = {styles_part} if styles_part is not None else set()
        unexpected = set(addition.replacements) - allowed
        if unexpected:
            raise ValueError(
                f"Reemplazo de Quotation no permitido: {sorted(unexpected)}"
            )
        if not addition.replacements:
            continue
        if styles_part is None:
            raise ValueError("Styles oficial ausente para Quotation")
        expected_content_type = base.content_types_for({styles_part})[styles_part]
        if (
            expected_content_type != STYLES_CONTENT_TYPE
            or addition.replacement_content_types.get(styles_part)
            != expected_content_type
        ):
            raise ValueError("Content type de styles de Quotation invalido")
        _validate_append_only_styles(
            base.parts[styles_part],
            addition.replacements[styles_part],
        )
        validated.add(styles_part)
    return frozenset(validated)


def _validate_append_only_styles(base_payload: bytes, candidate_payload: bytes) -> None:
    """Exige que styles preserve el arbol oficial y solo agregue entradas."""

    try:
        base = ET.fromstring(base_payload)
        candidate = ET.fromstring(candidate_payload)
    except (ET.ParseError, TypeError) as error:
        raise ValueError("Styles de Quotation no es append-only") from error
    style_sheet_tag = f"{{{MAIN}}}styleSheet"
    if (
        base.tag != style_sheet_tag
        or candidate.tag != style_sheet_tag
        or tuple(sorted(base.attrib.items()))
        != tuple(sorted(candidate.attrib.items()))
    ):
        raise ValueError("Styles de Quotation no preserva la raiz oficial")
    base_sections = list(base)
    candidate_sections = list(candidate)
    if [item.tag for item in candidate_sections] != [
        item.tag for item in base_sections
    ]:
        raise ValueError("Styles de Quotation no preserva las secciones oficiales")

    for official_section, candidate_section in zip(
        base_sections,
        candidate_sections,
    ):
        section_name = official_section.tag.rsplit("}", 1)[-1]
        child_name = _APPEND_ONLY_STYLE_SECTIONS.get(section_name)
        if child_name is None:
            if _semantic_xml_signature(candidate_section) != _semantic_xml_signature(
                official_section
            ):
                raise ValueError(
                    f"Styles de Quotation no preserva la seccion {section_name}"
                )
            continue
        expected_child_tag = f"{{{MAIN}}}{child_name}"
        if any(
            child.tag != expected_child_tag
            for child in (*official_section, *candidate_section)
        ):
            raise ValueError(
                f"Styles de Quotation contiene una seccion invalida: {section_name}"
            )
        official_attributes = {
            name: value
            for name, value in official_section.attrib.items()
            if name != "count"
        }
        candidate_attributes = {
            name: value
            for name, value in candidate_section.attrib.items()
            if name != "count"
        }
        if (
            official_attributes != candidate_attributes
            or official_section.attrib.get("count") != str(len(official_section))
            or candidate_section.attrib.get("count") != str(len(candidate_section))
            or len(candidate_section) < len(official_section)
        ):
            raise ValueError(
                f"Styles de Quotation no es append-only en {section_name}"
            )
        for index, official_child in enumerate(official_section):
            if _semantic_xml_signature(
                candidate_section[index]
            ) != _semantic_xml_signature(official_child):
                raise ValueError(
                    f"Styles de Quotation no preserva {section_name}[{index}]"
                )


def _semantic_xml_signature(element: ET.Element) -> tuple:
    text = element.text or ""
    if not text.strip():
        text = ""
    return (
        element.tag,
        tuple(sorted(element.attrib.items())),
        text,
        tuple(_semantic_xml_signature(child) for child in element),
    )


def _defined_name_signatures(
    workbook_payload: bytes,
) -> tuple[tuple[tuple[tuple[str, str], ...], str], ...]:
    try:
        workbook = ET.fromstring(workbook_payload)
    except (ET.ParseError, TypeError) as error:
        raise ValueError("Workbook inválido al verificar nombres definidos") from error
    signatures = (
        (
            tuple(sorted(item.attrib.items())),
            item.text or "",
        )
        for item in workbook.findall(f"{{{MAIN}}}definedNames/{{{MAIN}}}definedName")
    )
    return tuple(sorted(signatures))


def _translate_formula_with_overrides(
    formula: str,
    *,
    coordinate: str,
    sheet: str,
    overrides: Mapping[str, str],
) -> str:
    return translate_formula(
        "=" + formula,
        origin=coordinate,
        target=coordinate,
        range_overrides=overrides,
        sheet=sheet,
    )[1:]


def _add_workbook_sheets(
    base: XlsxPackage,
    sheet_additions: Sequence[SheetAddition],
    *,
    cotizacion_terms_delta: int,
    extra_additions: Mapping[str, bytes],
    extra_content_types: Mapping[str, str],
) -> tuple[bytes, bytes, bytes, dict[str, bytes], dict[str, bytes]]:
    workbook, workbook_prefixes = _parse_xml_document(
        base.parts["xl/workbook.xml"],
        "Workbook oficial invalido",
    )
    workbook_rels = ET.fromstring(base.parts["xl/_rels/workbook.xml.rels"])
    content_types = ET.fromstring(base.parts["[Content_Types].xml"])
    sheets = workbook.find(f"{{{MAIN}}}sheets")
    if sheets is None:
        raise ValueError("Workbook oficial sin sheets")
    existing_names = {item.attrib["name"].casefold() for item in sheets}
    used_ids = {int(item.attrib["sheetId"]) for item in sheets}
    used_rel_ids = {
        item.attrib.get("Id", "")
        for item in workbook_rels.findall(f"{{{PKG_REL}}}Relationship")
    }
    occupied = set(base.parts) | set(extra_additions)
    added_parts: dict[str, bytes] = {}
    replacements: dict[str, bytes] = {}
    additions_with_parts: list[tuple[SheetAddition, str, int]] = []

    for addition in sheet_additions:
        if addition.name.casefold() in existing_names:
            raise ValueError(f"La hoja ya existe: {addition.name}")
        if addition.name not in {"Quotation", "Quotation_Data"}:
            raise ValueError("Hoja no permitida por el compositor")
        part = addition.sheet_part
        if part is None:
            part = _allocate_worksheet_part(occupied | set(added_parts), addition.name)
            parts = {part: addition.xml}
            part_content_types = {part: WORKSHEET_CONTENT_TYPE}
        else:
            parts = dict(addition.parts)
            part_content_types = dict(addition.content_types)
            if parts.get(part) != addition.xml:
                raise ValueError("Parte principal de hoja agregada inconsistente")
        overlap = set(parts) & (occupied | set(added_parts))
        if overlap:
            raise ValueError(f"Colisión de partes de hoja: {sorted(overlap)}")
        _merge_disjoint(added_parts, parts, f"partes {addition.name}")
        occupied.update(parts)
        _merge_replacements(replacements, addition.replacements)

        relationship_id = _next_relationship_id(used_rel_ids)
        used_rel_ids.add(relationship_id)
        sheet_id = max(used_ids, default=0) + 1
        used_ids.add(sheet_id)
        attributes = {
            "name": addition.name,
            "sheetId": str(sheet_id),
            f"{{{REL}}}id": relationship_id,
        }
        if addition.state != "visible":
            attributes["state"] = addition.state
        ET.SubElement(sheets, f"{{{MAIN}}}sheet", attributes)
        ET.SubElement(
            workbook_rels,
            f"{{{PKG_REL}}}Relationship",
            {
                "Id": relationship_id,
                "Type": addition.relationship_type,
                "Target": posixpath.relpath(part, "xl"),
            },
        )
        additions_with_parts.append((addition, part, len(sheets) - 1))
        existing_names.add(addition.name.casefold())
        for name, content_type in part_content_types.items():
            _ensure_content_type(content_types, name, content_type)
        for name, content_type in addition.replacement_content_types.items():
            actual = base.content_types_for({name})[name]
            if actual != content_type:
                raise ValueError(f"Content type de reemplazo inconsistente: {name}")

    for name, content_type in extra_content_types.items():
        _ensure_content_type(content_types, name, content_type)

    defined_names = workbook.find(f"{{{MAIN}}}definedNames")
    if defined_names is None:
        defined_names = ET.Element(f"{{{MAIN}}}definedNames")
        insertion = list(workbook).index(sheets) + 1
        workbook.insert(insertion, defined_names)
    print_area = next(
        (
            item
            for item in defined_names.findall(f"{{{MAIN}}}definedName")
            if item.attrib.get("name") == "_xlnm.Print_Area"
            and item.attrib.get("localSheetId") == "0"
        ),
        None,
    )
    official_print_area = (
        f"Cotizacion!$A$1:$J${CANONICAL_COTIZACION_PRINT_END}"
    )
    if print_area is None or print_area.text != official_print_area:
        raise ValueError("Print_Area oficial de Cotizacion inesperada")
    print_area.text = (
        f"Cotizacion!$A$1:$J${CANONICAL_COTIZACION_PRINT_END + cotizacion_terms_delta}"
    )
    for addition, _part, sheet_index in additions_with_parts:
        for defined_name in addition.defined_names:
            defined_names.append(ET.fromstring(defined_name.xml_for_sheet_index(sheet_index)))

    calc_properties = workbook.find(f"{{{MAIN}}}calcPr")
    if calc_properties is None:
        raise ValueError("Workbook oficial sin calcPr")
    calc_properties.attrib.update(
        {
            "calcMode": "auto",
            "fullCalcOnLoad": "1",
            "forceFullCalc": "1",
        }
    )

    return (
        _xml_bytes(workbook, workbook_prefixes),
        ET.tostring(workbook_rels, encoding="utf-8", xml_declaration=True),
        ET.tostring(content_types, encoding="utf-8", xml_declaration=True),
        added_parts,
        replacements,
    )


def _translate_official_calc_chain(
    base: XlsxPackage,
    mobiliti: MobilitiSheetMutation,
    cotizacion: CotizacionSheetMutation,
) -> bytes:
    calc_part = _calc_chain_part(base)
    if calc_part is None:
        raise ValueError("calcChain oficial ausente")
    if not isinstance(mobiliti, MobilitiSheetMutation) or not isinstance(
        cotizacion, CotizacionSheetMutation
    ):
        raise TypeError("Mutacion de calcChain invalida")
    return _prune_calc_chain_sheets(
        base.parts[calc_part],
        {
            _workbook_sheet_id(base, "Mobiliti"),
            _workbook_sheet_id(base, "Cotizacion"),
        },
    )


def _prune_calc_chain_sheets(payload: bytes, sheet_ids: set[int]) -> bytes:
    """Conserva la cadena oficial salvo hojas que Excel debe recalcular."""

    if not isinstance(payload, bytes) or not isinstance(sheet_ids, set) or not all(
        type(sheet_id) is int and sheet_id > 0 for sheet_id in sheet_ids
    ):
        raise TypeError("Entrada de poda calcChain invalida")
    try:
        root = ET.fromstring(payload)
    except ET.ParseError as error:
        raise ValueError("calcChain oficial invalido") from error
    if root.tag != f"{{{MAIN}}}calcChain":
        raise ValueError("Raiz calcChain oficial invalida")

    excluded = {str(sheet_id) for sheet_id in sheet_ids}
    source_sheet_id: str | None = None
    output_sheet_id: str | None = None
    preserved: list[ET.Element] = []
    for cell in root:
        if cell.tag != f"{{{MAIN}}}c":
            raise ValueError("Hijo calcChain oficial invalido")
        if "i" in cell.attrib:
            source_sheet_id = cell.attrib["i"]
        if source_sheet_id is None:
            raise ValueError("calcChain oficial sin hoja efectiva")
        if source_sheet_id in excluded:
            continue
        clone = deepcopy(cell)
        if output_sheet_id != source_sheet_id:
            clone.attrib["i"] = source_sheet_id
        preserved.append(clone)
        output_sheet_id = source_sheet_id

    if not preserved:
        raise ValueError("calcChain oficial quedo vacio")
    root[:] = preserved
    result = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    for sheet_id in sheet_ids:
        _validate_calc_chain_sheet_coordinates(
            result,
            sheet_id=sheet_id,
            expected=set(),
        )
    return result


def _rebuild_mobiliti_calc_chain(
    payload: bytes,
    *,
    sheet_id: int,
    mobiliti: MobilitiSheetMutation,
) -> bytes:
    """Reconstruye Mobiliti desde sus fórmulas finales, sin entradas heredadas."""

    root = ET.fromstring(payload)
    formula_root = _worksheet_root(mobiliti.xml, "Mobiliti")
    formula_cells: list[tuple[str, ET.Element]] = []
    formula_coordinates: set[str] = set()
    for cell in formula_root.findall(f".//{{{MAIN}}}c"):
        formula = cell.find(f"{{{MAIN}}}f")
        if formula is None:
            continue
        coordinate = cell.attrib.get("r", "")
        if _CELL.fullmatch(coordinate) is None:
            raise ValueError("Formula Mobiliti sin coordenada valida")
        if coordinate in formula_coordinates:
            raise ValueError(f"Formula Mobiliti duplicada: {coordinate}")
        formula_coordinates.add(coordinate)
        formula_cells.append((coordinate, formula))
    formula_cells.sort(key=lambda item: _cell_coordinate_sort_key(item[0]))

    target_sheet_id = str(sheet_id)
    effective_sheet_id: str | None = None
    preserved: list[ET.Element] = []
    source_metadata: dict[str, list[dict[str, str]]] = {}
    for cell in root:
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id != target_sheet_id:
            preserved.append(cell)
            continue
        coordinate = cell.attrib["r"]
        source_metadata.setdefault(coordinate, []).append(
            {
                name: value
                for name, value in cell.attrib.items()
                if name not in {"r", "i"}
            }
        )

    for coordinate, formula in formula_cells:
        attributes = {"r": coordinate, "i": target_sheet_id}
        source_coordinate = _mobiliti_calc_metadata_source(
            coordinate,
            mobiliti.row_map,
            source_metadata,
        )
        if source_coordinate is not None:
            attributes.update(
                _select_calc_chain_metadata(
                    source_metadata[source_coordinate],
                    formula,
                )
            )
        preserved.append(ET.Element(f"{{{MAIN}}}c", attributes))
    root[:] = preserved
    rebuilt = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    _validate_calc_chain_sheet_coordinates(
        rebuilt,
        sheet_id=sheet_id,
        expected=formula_coordinates,
    )
    return rebuilt


def _mobiliti_calc_metadata_source(
    coordinate: str,
    row_map: MobilitiRowMap,
    source_metadata: Mapping[str, Sequence[Mapping[str, str]]],
) -> str | None:
    match = _CELL.fullmatch(coordinate)
    if match is None:
        raise ValueError("Coordenada Mobiliti invalida")
    column = match.group("column")
    row = int(match.group("row"))
    table_column = column_index_from_string(column) <= 34

    if table_column:
        for index, section in enumerate(row_map.sections):
            if section.product_start <= row < section.product_start + section.capacity:
                source = f"{column}{CANONICAL_MOBILITI_FIRST_SECTION_ROW + 1}"
                return source if source in source_metadata else None
            if row == section.section_row:
                source = f"{column}{13 if index == 0 else 48}"
                return source if source in source_metadata else None
            if row == section.subtotal_row:
                source = f"{column}{47 if index == 0 else 82}"
                return source if source in source_metadata else None

    if row == row_map.total_row:
        source = f"{column}{CANONICAL_MOBILITI_TOTAL_ROW}"
        return source if source in source_metadata else None
    auxiliary_offset = row - row_map.total_row
    if 1 <= auxiliary_offset <= CANONICAL_MOBILITI_AUX_END - CANONICAL_MOBILITI_TOTAL_ROW:
        source = f"{column}{CANONICAL_MOBILITI_TOTAL_ROW + auxiliary_offset}"
        return source if source in source_metadata else None
    return coordinate if coordinate in source_metadata else None


def _rebuild_cotizacion_calc_chain(
    payload: bytes,
    *,
    sheet_id: int,
    cotizacion: CotizacionSheetMutation,
) -> bytes:
    try:
        root = ET.fromstring(payload)
    except (ET.ParseError, TypeError) as error:
        raise ValueError("calcChain traducido invalido") from error
    if root.tag != f"{{{MAIN}}}calcChain":
        raise ValueError("Raiz calcChain traducida invalida")

    formula_root = _worksheet_root(cotizacion.xml, "Cotizacion")
    formula_cells: list[tuple[str, ET.Element]] = []
    formula_coordinates: set[str] = set()
    for cell in formula_root.findall(f".//{{{MAIN}}}c"):
        formula = cell.find(f"{{{MAIN}}}f")
        if formula is None:
            continue
        coordinate = cell.attrib.get("r", "")
        if _CELL.fullmatch(coordinate) is None:
            raise ValueError("Formula Cotizacion sin coordenada valida")
        if coordinate in formula_coordinates:
            raise ValueError(f"Formula Cotizacion duplicada: {coordinate}")
        formula_coordinates.add(coordinate)
        formula_cells.append((coordinate, formula))
    formula_cells.sort(key=lambda item: _cell_coordinate_sort_key(item[0]))

    target_sheet_id = str(sheet_id)
    effective_sheet_id: str | None = None
    preserved: list[ET.Element] = []
    source_metadata: dict[str, list[dict[str, str]]] = {}
    for cell in root:
        if cell.tag != f"{{{MAIN}}}c":
            raise ValueError("Hijo calcChain traducido invalido")
        coordinate = cell.attrib.get("r", "")
        if _CELL.fullmatch(coordinate) is None:
            raise ValueError("Coordenada calcChain traducida invalida")
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id != target_sheet_id:
            preserved.append(cell)
            continue
        source_metadata.setdefault(coordinate, []).append(
            {
                name: value
                for name, value in cell.attrib.items()
                if name not in {"r", "i"}
            }
        )

    for coordinate, formula in formula_cells:
        attributes = {"r": coordinate, "i": target_sheet_id}
        source_coordinate = _cotizacion_calc_metadata_source(
            coordinate,
            cotizacion,
            source_metadata,
        )
        if source_coordinate is not None:
            attributes.update(
                _select_calc_chain_metadata(
                    source_metadata[source_coordinate],
                    formula,
                )
            )
        preserved.append(ET.Element(f"{{{MAIN}}}c", attributes))
    root[:] = preserved
    rebuilt = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    _validate_calc_chain_sheet_coordinates(
        rebuilt,
        sheet_id=sheet_id,
        expected=formula_coordinates,
    )
    return rebuilt


def _cotizacion_calc_metadata_source(
    coordinate: str,
    mutation: CotizacionSheetMutation,
    source_metadata: Mapping[str, Sequence[Mapping[str, str]]],
) -> str | None:
    match = _CELL.fullmatch(coordinate)
    if match is None:
        raise ValueError("Coordenada Cotizacion invalida")
    column = match.group("column")
    row = int(match.group("row"))
    editable_column = column_index_from_string(column) <= 10
    if editable_column and row in mutation.section_subtotal_rows:
        template = f"{column}18"
        return template if template in source_metadata else None
    if editable_column and row in mutation.product_rows:
        template = f"{column}{CANONICAL_COTIZACION_FIRST_PRODUCT_ROW}"
        return template if template in source_metadata else None

    subtotal_row = mutation.total_row - CANONICAL_COTIZACION_TOTAL_SPAN
    if editable_column and subtotal_row <= row <= mutation.total_row:
        template = f"{column}{CANONICAL_COTIZACION_TOTAL_START + row - subtotal_row}"
        return template if template in source_metadata else None

    shifted_terms_start = CANONICAL_COTIZACION_TERMS_START + mutation.terms_row_delta
    if editable_column and row >= shifted_terms_start:
        template = f"{column}{row - mutation.terms_row_delta}"
        return template if template in source_metadata else None
    return coordinate if coordinate in source_metadata else None


def _select_calc_chain_metadata(
    candidates: Sequence[Mapping[str, str]],
    formula: ET.Element,
) -> dict[str, str]:
    if not candidates:
        return {}
    is_array = formula.attrib.get("t") == "array"
    return dict(
        min(
            enumerate(candidates),
            key=lambda item: (
                item[1].get("a") != "1" if is_array else item[1].get("a") == "1",
                item[0],
            ),
        )[1]
    )


def _cell_coordinate_sort_key(coordinate: str) -> tuple[int, int]:
    match = _CELL.fullmatch(coordinate)
    if match is None:
        raise ValueError("Coordenada de celda invalida")
    return (
        int(match.group("row")),
        column_index_from_string(match.group("column")),
    )


def _validate_calc_chain_sheet_coordinates(
    payload: bytes,
    *,
    sheet_id: int,
    expected: set[str],
) -> None:
    root = ET.fromstring(payload)
    target_sheet_id = str(sheet_id)
    effective_sheet_id: str | None = None
    actual: list[str] = []
    for cell in root:
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id == target_sheet_id:
            actual.append(cell.attrib.get("r", ""))
    if len(actual) != len(set(actual)):
        raise ValueError("calcChain Cotizacion contiene coordenadas duplicadas")
    if set(actual) != expected:
        raise ValueError("calcChain Cotizacion no coincide con sus formulas finales")


def _validate_compose_paths(template: Path, output: Path) -> tuple[Path, Path]:
    for label, path in (("plantilla", template), ("salida", output)):
        if not path.is_absolute():
            raise ValueError(f"La ruta de {label} debe ser absoluta")
        if ".." in path.parts:
            raise ValueError(f"La ruta de {label} contiene segmentos léxicos inseguros")
        for component in (path, *path.parents):
            if _path_is_reparse_point(component):
                raise ValueError(
                    f"La ruta de {label} no puede atravesar un reparse point"
                )
        if any(
            ":" in component
            for index, component in enumerate(path.parts)
            if index > 0
        ):
            raise ValueError(f"La ruta de {label} contiene un alias de dispositivo")
        if any(parent.exists() and parent.is_symlink() for parent in (path, *path.parents)):
            raise ValueError(f"La ruta de {label} no puede atravesar symlinks")
    if not template.exists() or not template.is_file():
        raise FileNotFoundError(f"Plantilla inexistente: {template}")
    if output.exists():
        raise FileExistsError(f"La salida ya existe: {output}")
    if not output.parent.exists() or not output.parent.is_dir():
        raise FileNotFoundError(f"Directorio de salida inexistente: {output.parent}")
    if template.resolve(strict=True) == output.resolve(strict=False):
        raise ValueError("La salida no puede reemplazar la plantilla")
    if output.suffix.casefold() != ".xlsx":
        raise ValueError("La salida debe ser un archivo .xlsx")
    return template.resolve(strict=True), output.resolve(strict=False)


def _path_is_reparse_point(path: Path) -> bool:
    try:
        metadata = os.lstat(path)
    except FileNotFoundError:
        return False
    attributes = getattr(metadata, "st_file_attributes", 0)
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    return stat.S_ISLNK(metadata.st_mode) or bool(attributes & reparse_flag)


def _write_candidate(output: Path, payload: bytes) -> Path:
    for index in range(1, 10_001):
        candidate = output.with_name(f".{output.name}.compose-{index}.tmp")
        try:
            stream = candidate.open("xb")
        except FileExistsError:
            continue
        try:
            with stream:
                stream.write(payload)
                stream.flush()
                os.fsync(stream.fileno())
        except BaseException:
            if candidate.exists():
                _recycle_candidate(candidate)
            raise
        return candidate
    raise FileExistsError("No se pudo reservar un candidato OOXML")


def _atomic_publish_no_replace(candidate: Path, output: Path) -> None:
    """Publica atómicamente sin reemplazar un nombre creado por una carrera."""

    if os.name == "nt":
        os.rename(candidate, output)
        return
    if sys.platform.startswith("linux") and _linux_rename_no_replace(candidate, output):
        return
    if sys.platform == "darwin" and _darwin_rename_no_replace(candidate, output):
        return
    try:
        os.link(candidate, output)
    except FileExistsError as error:
        raise FileExistsError(f"La salida ya existe: {output}") from error
    _recycle_candidate(candidate)


def _linux_rename_no_replace(candidate: Path, output: Path) -> bool:
    try:
        function = ctypes.CDLL(None, use_errno=True).renameat2
    except AttributeError:
        return False
    function.argtypes = (
        ctypes.c_int,
        ctypes.c_char_p,
        ctypes.c_int,
        ctypes.c_char_p,
        ctypes.c_uint,
    )
    function.restype = ctypes.c_int
    result = function(
        -100,
        os.fsencode(candidate),
        -100,
        os.fsencode(output),
        1,
    )
    if result == 0:
        return True
    error_number = ctypes.get_errno()
    if error_number == errno.EEXIST:
        raise FileExistsError(f"La salida ya existe: {output}")
    if error_number in {errno.ENOSYS, errno.EINVAL, getattr(errno, "ENOTSUP", 95)}:
        return False
    raise OSError(error_number, os.strerror(error_number), str(output))


def _darwin_rename_no_replace(candidate: Path, output: Path) -> bool:
    try:
        function = ctypes.CDLL(None, use_errno=True).renamex_np
    except AttributeError:
        return False
    function.argtypes = (ctypes.c_char_p, ctypes.c_char_p, ctypes.c_uint)
    function.restype = ctypes.c_int
    result = function(os.fsencode(candidate), os.fsencode(output), 0x00000004)
    if result == 0:
        return True
    error_number = ctypes.get_errno()
    if error_number == errno.EEXIST:
        raise FileExistsError(f"La salida ya existe: {output}")
    if error_number in {errno.ENOSYS, errno.EINVAL, getattr(errno, "ENOTSUP", 45)}:
        return False
    raise OSError(error_number, os.strerror(error_number), str(output))


def _recycle_candidate(candidate: Path) -> Path | None:
    """Retira un candidato únicamente mediante Trash/Recycle Bin o cuarentena."""

    candidate = Path(candidate)
    if _lstat_or_none(candidate) is None:
        return None
    command: list[str] | None = None
    if os.name == "nt":
        profile = Path(os.environ.get("USERPROFILE", ""))
        script = profile / ".codex" / "bin" / "Send-ToRecycleBin.ps1"
        if script.is_file():
            command = [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script),
                "-LiteralPath",
                str(candidate),
            ]
    elif sys.platform == "darwin" and Path("/usr/bin/trash").is_file():
        command = ["/usr/bin/trash", str(candidate)]
    else:
        gio = shutil.which("gio")
        trash_put = shutil.which("trash-put")
        if gio:
            command = [gio, "trash", str(candidate)]
        elif trash_put:
            command = [trash_put, str(candidate)]
    if command is not None:
        try:
            subprocess.run(
                command,
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=30,
            )
            if _lstat_or_none(candidate) is None:
                return None
        except (OSError, subprocess.SubprocessError):
            pass
    return _quarantine_candidate(candidate)


def _quarantine_candidate(candidate: Path) -> Path:
    for _attempt in range(100):
        quarantine_directory = candidate.parent / (
            f".mobiliti-recovery-{uuid.uuid4().hex}"
        )
        try:
            os.mkdir(quarantine_directory, mode=0o700)
        except FileExistsError:
            continue
        destination = quarantine_directory / candidate.name
        _move_candidate_no_replace(candidate, destination)
        if _lstat_or_none(candidate) is not None or _lstat_or_none(destination) is None:
            raise RuntimeError("La cuarentena no confirmó el movimiento del candidato")
        return destination
    raise RuntimeError(
        f"No se pudo mover el candidato a una cuarentena recuperable: {candidate}"
    )


def _move_candidate_no_replace(candidate: Path, destination: Path) -> None:
    if os.name == "nt":
        os.rename(candidate, destination)
        return
    if sys.platform.startswith("linux") and _linux_rename_no_replace(candidate, destination):
        return
    if sys.platform == "darwin" and _darwin_rename_no_replace(candidate, destination):
        return
    if _lstat_or_none(destination) is not None:
        raise FileExistsError(f"La cuarentena ya existe: {destination}")
    os.rename(candidate, destination)


def _lstat_or_none(path: Path) -> os.stat_result | None:
    try:
        return os.lstat(path)
    except FileNotFoundError:
        return None


def _worksheet_root(payload: bytes, name: str) -> ET.Element:
    try:
        root = ET.fromstring(payload)
    except (ET.ParseError, TypeError) as error:
        raise ValueError(f"XML de {name} inválido") from error
    if root.tag != f"{{{MAIN}}}worksheet":
        raise ValueError(f"Parte de {name} no es worksheet")
    return root


def _replace_formula(root: ET.Element, coordinate: str, formula: str) -> None:
    cell = _cell_in_root(root, coordinate)
    if cell is None:
        raise ValueError(f"Celda oficial ausente: {coordinate}")
    _set_formula_element(cell, formula)


def _formula_at(package: XlsxPackage, sheet: str, coordinate: str) -> str | None:
    root = ET.fromstring(package.parts[package.sheet_part(sheet)])
    return _formula_in_root(root, coordinate)


def _formula_in_root(root: ET.Element, coordinate: str) -> str | None:
    cell = _cell_in_root(root, coordinate)
    if cell is None:
        return None
    formula = cell.find(f"{{{MAIN}}}f")
    return None if formula is None else formula.text or ""


def _cell_in_root(root: ET.Element, coordinate: str) -> ET.Element | None:
    return root.find(f".//{{{MAIN}}}c[@r='{coordinate}']")


def _clone_row_region(
    source: ET.Element,
    source_row: int,
    target_row: int,
    *,
    last_column: int,
) -> ET.Element:
    clone = deepcopy(source)
    clone.attrib["r"] = str(target_row)
    for cell in list(clone.findall(f"{{{MAIN}}}c")):
        match = _CELL.fullmatch(cell.attrib.get("r", ""))
        if match is None:
            raise ValueError("Coordenada inválida en Cotizacion oficial")
        column = match.group("column")
        if column_index_from_string(column) > last_column:
            clone.remove(cell)
            continue
        cell.attrib["r"] = f"{column}{target_row}"
    return clone


def _require_cell(row: ET.Element, coordinate: str) -> ET.Element:
    cell = next(
        (item for item in row.findall(f"{{{MAIN}}}c") if item.attrib.get("r") == coordinate),
        None,
    )
    if cell is None:
        match = _CELL.fullmatch(coordinate)
        if match is None or int(match.group("row")) != int(row.attrib["r"]):
            raise ValueError(f"Coordenada Cotizacion inválida: {coordinate}")
        cell = ET.Element(f"{{{MAIN}}}c", {"r": coordinate})
        row.append(cell)
        _sort_cells(row)
    return cell


def _set_inline_string(row: ET.Element, coordinate: str, value: str) -> None:
    _validate_text(value, allow_empty=True)
    cell = _require_cell(row, coordinate)
    _clear_cell_value(cell)
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{MAIN}}}is")
    text = ET.SubElement(inline, f"{{{MAIN}}}t")
    if value != value.strip():
        text.attrib[f"{{{XML}}}space"] = "preserve"
    text.text = value


def _set_number(row: ET.Element, coordinate: str, value: Decimal) -> None:
    number = _decimal(value, coordinate)
    cell = _require_cell(row, coordinate)
    _clear_cell_value(cell)
    cell.attrib.pop("t", None)
    ET.SubElement(cell, f"{{{MAIN}}}v").text = format(number, "f")


def _set_formula(
    row: ET.Element,
    coordinate: str,
    formula: str,
    *,
    attributes: Mapping[str, str] | None = None,
) -> None:
    cell = _require_cell(row, coordinate)
    _set_formula_element(cell, formula, attributes=attributes)


def _set_formula_element(
    cell: ET.Element,
    formula: str,
    *,
    attributes: Mapping[str, str] | None = None,
) -> None:
    if not isinstance(formula, str) or not formula.startswith("="):
        raise ValueError("Fórmula Cotizacion inválida")
    _clear_cell_value(cell)
    cell.attrib.pop("t", None)
    formula_node = ET.SubElement(cell, f"{{{MAIN}}}f", dict(attributes or {}))
    formula_node.text = formula[1:]


def _clear_cell_value(cell: ET.Element) -> None:
    for child in list(cell):
        if child.tag in {f"{{{MAIN}}}f", f"{{{MAIN}}}v", f"{{{MAIN}}}is"}:
            cell.remove(child)
    cell.attrib.pop("t", None)
    cell.attrib.pop("cm", None)
    cell.attrib.pop("vm", None)


def _formula_text(row: ET.Element, coordinate: str) -> str | None:
    cell = next(
        (item for item in row.findall(f"{{{MAIN}}}c") if item.attrib.get("r") == coordinate),
        None,
    )
    if cell is None:
        return None
    formula = cell.find(f"{{{MAIN}}}f")
    return None if formula is None else "=" + (formula.text or "")


def _sidecar_cells(
    rows: Mapping[int, ET.Element],
    *,
    first_row: int,
    first_sidecar_column: int,
) -> dict[int, tuple[ET.Element, ...]]:
    result: dict[int, tuple[ET.Element, ...]] = {}
    for number, row in rows.items():
        if number < first_row:
            continue
        cells = tuple(
            deepcopy(cell)
            for cell in row.findall(f"{{{MAIN}}}c")
            if column_index_from_string(_CELL.fullmatch(cell.attrib["r"]).group("column"))
            >= first_sidecar_column
        )
        if cells:
            result[number] = cells
    return result


def _apply_sidecars(
    rows: list[ET.Element],
    sidecars: Mapping[int, tuple[ET.Element, ...]],
) -> None:
    by_number = {int(row.attrib["r"]): row for row in rows}
    for number, cells in sidecars.items():
        row = by_number.get(number)
        if row is None:
            row = ET.Element(f"{{{MAIN}}}row", {"r": str(number)})
            rows.append(row)
            by_number[number] = row
        existing = {cell.attrib["r"] for cell in row.findall(f"{{{MAIN}}}c")}
        for cell in cells:
            if cell.attrib["r"] not in existing:
                row.append(deepcopy(cell))
        _sort_cells(row)


def _rebuild_cotizacion_merges(
    root: ET.Element,
    *,
    dynamic_merges: Sequence[str],
    total_delta: int,
) -> None:
    merges = root.find(f"{{{MAIN}}}mergeCells")
    if merges is None:
        raise ValueError("Cotizacion oficial no contiene mergeCells")
    preserved: list[ET.Element] = []
    for merge in merges.findall(f"{{{MAIN}}}mergeCell"):
        bounds = _range_bounds(merge.attrib.get("ref", ""))
        if bounds is None:
            raise ValueError("Merge Cotizacion inválido")
        min_col, min_row, max_col, max_row = bounds
        if min_col > 10:
            preserved.append(deepcopy(merge))
        elif max_col <= 10 and max_row < 16:
            preserved.append(deepcopy(merge))
        elif max_col <= 10 and 19 <= min_row <= 24:
            preserved.append(_shift_merge(merge, total_delta))
        elif max_col <= 10 and min_row >= 28:
            preserved.append(_shift_merge(merge, total_delta))
        elif max_col > 10:
            preserved.append(deepcopy(merge))
    preserved.extend(
        ET.Element(f"{{{MAIN}}}mergeCell", {"ref": reference})
        for reference in dynamic_merges
    )
    preserved.sort(key=lambda item: _range_bounds(item.attrib["ref"]) or (0, 0, 0, 0))
    merges[:] = preserved
    merges.attrib["count"] = str(len(preserved))


def _range_bounds(reference: str) -> tuple[int, int, int, int] | None:
    parts = reference.split(":")
    if len(parts) not in {1, 2}:
        return None
    first = _CELL.fullmatch(parts[0].replace("$", ""))
    last = _CELL.fullmatch(parts[-1].replace("$", ""))
    if first is None or last is None:
        return None
    return (
        column_index_from_string(first.group("column")),
        int(first.group("row")),
        column_index_from_string(last.group("column")),
        int(last.group("row")),
    )


def _shift_merge(merge: ET.Element, delta: int) -> ET.Element:
    clone = deepcopy(merge)
    bounds = _range_bounds(clone.attrib["ref"])
    if bounds is None:
        raise ValueError("Merge Cotizacion inválido")
    min_col, min_row, max_col, max_row = bounds
    clone.attrib["ref"] = (
        f"{get_column_letter(min_col)}{min_row + delta}:"
        f"{get_column_letter(max_col)}{max_row + delta}"
    )
    return clone


def _update_dimension(root: ET.Element, delta: int) -> None:
    dimension = root.find(f"{{{MAIN}}}dimension")
    if dimension is None:
        return
    reference = dimension.attrib.get("ref", "")
    if ":" not in reference:
        return
    first, last = reference.split(":", 1)
    match = _CELL.fullmatch(last.replace("$", ""))
    if match is None:
        raise ValueError("Dimensión oficial Cotizacion inválida")
    dimension.attrib["ref"] = (
        f"{first}:{match.group('column')}{int(match.group('row')) + delta}"
    )


def _sort_cells(row: ET.Element) -> None:
    cells = list(row.findall(f"{{{MAIN}}}c"))
    cells.sort(
        key=lambda cell: column_index_from_string(
            _CELL.fullmatch(cell.attrib["r"]).group("column")
        )
    )
    row[:] = cells


def _shift_anchor_rows(anchor: ET.Element, delta: int) -> None:
    if delta == 0:
        return
    for marker_name in ("from", "to"):
        marker = anchor.find(f"{{{XDR}}}{marker_name}")
        if marker is None:
            continue
        row = marker.find(f"{{{XDR}}}row")
        if row is not None and row.text is not None:
            target = int(row.text) + delta
            if target < 0:
                raise ValueError("Ancla Cotizacion fuera de rango")
            row.text = str(target)


def _next_picture_id(root: ET.Element) -> int:
    values: list[int] = []
    for node in root.findall(f".//{{{XDR}}}cNvPr"):
        raw = node.attrib.get("id", "")
        if not raw.isdigit() or int(raw) <= 0:
            raise ValueError("ID cNvPr Cotizacion inválido")
        values.append(int(raw))
    if len(values) != len(set(values)):
        raise ValueError("IDs cNvPr Cotizacion duplicados")
    return max(values, default=0) + 1


def _read_product_image(
    source: CotizacionProductImage | Path,
) -> tuple[bytes, str, str, int, int]:
    if isinstance(source, CotizacionProductImage):
        expected_content_type = source.content_type
        content = (
            _bounded_read_product_image_path(source.path)
            if source.path is not None
            else bytes(source.content or b"")
        )
    else:
        expected_content_type = None
        content = _bounded_read_product_image_path(Path(source))
    if not content or len(content) > MAX_IMAGE_BYTES:
        raise ValueError("Tamaño de imagen de producto inválido")
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        extension, content_type = ".png", PNG_CONTENT_TYPE
    elif content.startswith(b"\xff\xd8\xff"):
        extension, content_type = ".jpeg", JPEG_CONTENT_TYPE
    else:
        raise ValueError("Formato de imagen de producto no permitido")
    if expected_content_type is not None and expected_content_type != content_type:
        raise ValueError("Content type de imagen de producto inconsistente")
    try:
        from PIL import Image

        with Image.open(BytesIO(content)) as image:
            width, height = image.size
            image.verify()
    except Exception as error:
        raise ValueError("Imagen de producto inválida") from error
    if width <= 0 or height <= 0 or width * height > MAX_IMAGE_PIXELS:
        raise ValueError("Dimensiones de imagen de producto inválidas")
    return content, extension, content_type, width, height


def _bounded_read_product_image_path(path: Path | None) -> bytes:
    if path is None:
        raise ValueError("Ruta de imagen de producto ausente")
    path = Path(path)
    if not path.is_absolute():
        raise ValueError("La imagen de producto debe usar ruta absoluta")
    if not _SAFE_IMAGE_NAME.fullmatch(path.name):
        raise ValueError("Nombre de imagen de producto inválido")
    for component in (path, *path.parents):
        if _path_is_reparse_point(component):
            raise ValueError("La imagen de producto no puede atravesar un reparse point")
    try:
        before = os.lstat(path)
    except FileNotFoundError as error:
        raise FileNotFoundError(f"Imagen de producto inexistente: {path}") from error
    if not stat.S_ISREG(before.st_mode):
        raise ValueError("La imagen de producto debe ser un archivo regular")
    if before.st_size <= 0 or before.st_size > MAX_IMAGE_BYTES:
        raise ValueError("Tamaño de imagen de producto inválido")
    flags = os.O_RDONLY | getattr(os, "O_BINARY", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(path, flags)
    try:
        opened = os.fstat(descriptor)
        if (
            not stat.S_ISREG(opened.st_mode)
            or (before.st_dev, before.st_ino) != (opened.st_dev, opened.st_ino)
            or opened.st_size != before.st_size
        ):
            raise ValueError("La ruta de imagen cambió durante la apertura")
        chunks: list[bytes] = []
        total = 0
        while True:
            chunk = os.read(descriptor, min(1024 * 1024, MAX_IMAGE_BYTES + 1 - total))
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_IMAGE_BYTES:
                raise ValueError("Tamaño de imagen de producto inválido")
            chunks.append(chunk)
        if total != opened.st_size:
            raise ValueError("Tamaño de imagen cambió durante la lectura")
        return b"".join(chunks)
    finally:
        os.close(descriptor)


def _allocate_media_part(
    base: XlsxPackage,
    additions: Mapping[str, bytes],
    sequence: int,
    extension: str,
) -> str:
    suffix = sequence
    while True:
        candidate = f"xl/media/quote_product_{suffix:04d}{extension}"
        if candidate not in base.parts and candidate not in additions:
            return candidate
        suffix += 1


def _cotizacion_image_box(
    position: int,
    image_count: int,
) -> tuple[int, int, int, int]:
    """Devuelve un rectángulo interno para principal y miniaturas."""

    if (
        type(position) is not int
        or type(image_count) is not int
        or image_count < 2
        or not 0 <= position < image_count
    ):
        raise ValueError("Distribución de imágenes Cotizacion inválida")
    margin = 120_000
    gap = 90_000
    strip_width = 1_350_000
    content_height = _COTIZACION_IMAGE_CELL_HEIGHT_EMU - 2 * margin
    strip_left = _COTIZACION_IMAGE_CELL_WIDTH_EMU - margin - strip_width
    if position == 0:
        width = strip_left - gap - margin
        if width <= 0 or content_height <= 0:
            raise ValueError("Área principal de imagen Cotizacion inválida")
        return margin, margin, width, content_height

    complement_count = image_count - 1
    columns = 1 if complement_count <= 3 else 2
    rows = math.ceil(complement_count / columns)
    box_width = (strip_width - gap * (columns - 1)) // columns
    box_height = (content_height - gap * (rows - 1)) // rows
    complement_index = position - 1
    column = complement_index % columns
    row = complement_index // columns
    if box_width <= 0 or box_height <= 0:
        raise ValueError("Área de complemento Cotizacion insuficiente")
    return (
        strip_left + column * (box_width + gap),
        margin + row * (box_height + gap),
        box_width,
        box_height,
    )


def _product_image_anchor(
    *,
    relationship_id: str,
    picture_id: int,
    target_row: int,
    width: int,
    height: int,
    name: str,
    cell_width: int | None = None,
    cell_height: int | None = None,
    box: tuple[int, int, int, int] | None = None,
) -> ET.Element:
    max_width = 3_200_000
    max_height = 3_000_000
    if (cell_width is None) != (cell_height is None):
        raise ValueError("Geometría parcial de celda para imagen Cotizacion")
    if box is not None:
        if cell_width is None or cell_height is None:
            raise ValueError("Rectángulo de imagen sin geometría de celda")
        if (
            not isinstance(box, tuple)
            or len(box) != 4
            or not all(type(value) is int for value in box)
        ):
            raise TypeError("Rectángulo de imagen Cotizacion inválido")
        box_left, box_top, box_width, box_height = box
        if (
            box_left < 0
            or box_top < 0
            or box_width <= 0
            or box_height <= 0
            or box_left + box_width > cell_width
            or box_top + box_height > cell_height
        ):
            raise ValueError("Rectángulo de imagen fuera de la celda Cotizacion")
        scale = min(box_width / width, box_height / height)
        cx = max(1, int(width * scale))
        cy = max(1, int(height * scale))
        column_offset = box_left + (box_width - cx) // 2
        row_offset = box_top + (box_height - cy) // 2
    else:
        if cell_width is not None:
            max_width = min(max_width, cell_width)
            max_height = min(max_height, cell_height)
        scale = min(max_width / width, max_height / height)
        cx = max(1, int(width * scale))
        cy = max(1, int(height * scale))
    if box is None and cell_width is None:
        column_offset = 0
        row_offset = 0
    elif box is None:
        if (
            type(cell_width) is not int
            or type(cell_height) is not int
            or cell_width < cx
            or cell_height < cy
        ):
            raise ValueError("Geometría de celda insuficiente para imagen Cotizacion")
        column_offset = (cell_width - cx) // 2
        row_offset = (cell_height - cy) // 2
    anchor = ET.Element(f"{{{XDR}}}oneCellAnchor")
    marker = ET.SubElement(anchor, f"{{{XDR}}}from")
    for tag, text in (
        ("col", "1"),
        ("colOff", str(column_offset)),
        ("row", str(target_row - 1)),
        ("rowOff", str(row_offset)),
    ):
        ET.SubElement(marker, f"{{{XDR}}}{tag}").text = text
    ET.SubElement(anchor, f"{{{XDR}}}ext", {"cx": str(cx), "cy": str(cy)})
    picture = ET.SubElement(anchor, f"{{{XDR}}}pic")
    non_visual = ET.SubElement(picture, f"{{{XDR}}}nvPicPr")
    ET.SubElement(
        non_visual,
        f"{{{XDR}}}cNvPr",
        {"id": str(picture_id), "name": name},
    )
    ET.SubElement(non_visual, f"{{{XDR}}}cNvPicPr")
    fill = ET.SubElement(picture, f"{{{XDR}}}blipFill")
    ET.SubElement(fill, f"{{{DRAWING}}}blip", {f"{{{REL}}}embed": relationship_id})
    stretch = ET.SubElement(fill, f"{{{DRAWING}}}stretch")
    ET.SubElement(stretch, f"{{{DRAWING}}}fillRect")
    shape = ET.SubElement(picture, f"{{{XDR}}}spPr")
    transform = ET.SubElement(shape, f"{{{DRAWING}}}xfrm")
    ET.SubElement(transform, f"{{{DRAWING}}}off", {"x": "0", "y": "0"})
    ET.SubElement(transform, f"{{{DRAWING}}}ext", {"cx": str(cx), "cy": str(cy)})
    geometry = ET.SubElement(shape, f"{{{DRAWING}}}prstGeom", {"prst": "rect"})
    ET.SubElement(geometry, f"{{{DRAWING}}}avLst")
    ET.SubElement(anchor, f"{{{XDR}}}clientData")
    return anchor


def _allocate_worksheet_part(occupied: set[str], name: str) -> str:
    stem = "quotation_data" if name == "Quotation_Data" else "quotation"
    for index in range(1, 1_000_001):
        candidate = f"xl/worksheets/{stem}{index}.xml"
        if candidate not in occupied:
            return candidate
    raise ValueError("No se pudo asignar una parte worksheet")


def _ensure_content_type(root: ET.Element, part: str, content_type: str) -> None:
    validate_part_name(part)
    if not isinstance(content_type, str) or not content_type:
        raise ValueError("Content type inválido")
    part_name = "/" + part
    existing = next(
        (
            item
            for item in root.findall(f"{{{CONTENT_TYPES}}}Override")
            if item.attrib.get("PartName") == part_name
        ),
        None,
    )
    if existing is not None:
        if existing.attrib.get("ContentType") != content_type:
            raise ValueError(f"Content type en conflicto: {part}")
        return
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES}}}Override",
        {"PartName": part_name, "ContentType": content_type},
    )


def _next_relationship_id(used: set[str]) -> str:
    for index in range(1, 1_000_001):
        candidate = f"rId{index}"
        if candidate not in used:
            return candidate
    raise ValueError("No se pudo asignar un relationship Id")


def _merge_disjoint(target: dict[str, bytes], source: Mapping[str, bytes], label: str) -> None:
    overlap = set(target) & set(source)
    if overlap:
        raise ValueError(f"Colisión de {label}: {sorted(overlap)}")
    target.update(source)


def _merge_replacements(target: dict[str, bytes], source: Mapping[str, bytes]) -> None:
    for name, payload in source.items():
        if name in target and target[name] != payload:
            raise ValueError(f"Reemplazos incompatibles: {name}")
        target[name] = payload


def _immutable_bytes_mapping(value: Mapping[str, bytes], label: str) -> Mapping[str, bytes]:
    copied: dict[str, bytes] = {}
    for name, payload in value.items():
        validate_part_name(name)
        if not isinstance(payload, bytes):
            raise TypeError(f"{label} contiene bytes inválidos")
        copied[name] = payload
    return MappingProxyType(copied)


def _calc_chain_part(base: XlsxPackage) -> str | None:
    rels = ET.fromstring(base.parts["xl/_rels/workbook.xml.rels"])
    matches = [
        item
        for item in rels.findall(f"{{{PKG_REL}}}Relationship")
        if item.attrib.get("Type") in relationship_type_uris("calcChain")
    ]
    if not matches:
        return None
    if len(matches) != 1 or matches[0].attrib.get("TargetMode", "").casefold() == "external":
        raise ValueError("Relación calcChain inválida")
    return resolve_internal_target("xl/workbook.xml", matches[0].attrib["Target"])


def _workbook_sheet_id(base: XlsxPackage, name: str) -> int:
    workbook = ET.fromstring(base.parts["xl/workbook.xml"])
    matches = [
        item
        for item in workbook.findall(f"{{{MAIN}}}sheets/{{{MAIN}}}sheet")
        if item.attrib.get("name") == name
    ]
    if len(matches) != 1:
        raise ValueError(f"Hoja oficial inválida: {name}")
    return int(matches[0].attrib["sheetId"])


def _formula_coordinates_by_row(root: ET.Element) -> dict[int, tuple[str, ...]]:
    result: dict[int, list[str]] = {}
    for cell in root.findall(f".//{{{MAIN}}}c"):
        if cell.find(f"{{{MAIN}}}f") is None:
            continue
        coordinate = cell.attrib.get("r", "")
        match = _CELL.fullmatch(coordinate)
        if match is None:
            raise ValueError("Coordenada de fórmula oficial inválida")
        result.setdefault(int(match.group("row")), []).append(coordinate)
    return {row: tuple(coordinates) for row, coordinates in result.items()}


def _validate_text(value: object, *, allow_empty: bool = False) -> None:
    if not isinstance(value, str):
        raise TypeError("Texto de Cotizacion inválido")
    if not allow_empty and not value:
        raise ValueError("Texto de Cotizacion vacío")
    if len(value) > MAX_CELL_TEXT or any(not _is_xml_character(ord(char)) for char in value):
        raise ValueError("Texto de Cotizacion no representable en XLSX")


def _decimal(value: object, field_name: str) -> Decimal:
    if isinstance(value, bool):
        raise TypeError(f"{field_name} inválido")
    try:
        result = value if isinstance(value, Decimal) else Decimal(str(value))
    except Exception as error:
        raise TypeError(f"{field_name} inválido") from error
    if not result.is_finite():
        raise ValueError(f"{field_name} no finito")
    return result


def _excel_decimal(value: Decimal) -> str:
    """Serializa un decimal exacto como constante numérica estable de Excel."""

    if not isinstance(value, Decimal) or not value.is_finite() or value <= 0:
        raise ValueError("Constante decimal de fórmula inválida")
    sign, raw_digits, raw_exponent = value.as_tuple()
    if sign or not isinstance(raw_exponent, int):
        raise ValueError("Constante decimal de fórmula inválida")

    digits = raw_digits
    exponent = raw_exponent
    removable_zeros = 0
    if exponent < 0:
        for digit in reversed(digits):
            if digit != 0 or removable_zeros >= -exponent:
                break
            removable_zeros += 1
    if removable_zeros:
        digits = digits[:-removable_zeros]
        exponent += removable_zeros

    digit_count = len(digits)
    if exponent >= 0:
        literal_length = digit_count + exponent
    else:
        decimal_position = digit_count + exponent
        literal_length = (
            digit_count + 1
            if decimal_position > 0
            else 2 - decimal_position + digit_count
        )
    if literal_length > MAX_EXCEL_FORMULA_LENGTH:
        raise ValueError("El literal decimal Excel excede 8,192 caracteres")

    digit_text = "".join(str(digit) for digit in digits)
    if exponent >= 0:
        return digit_text + ("0" * exponent)
    decimal_position = digit_count + exponent
    if decimal_position > 0:
        return f"{digit_text[:decimal_position]}.{digit_text[decimal_position:]}"
    return "0." + ("0" * -decimal_position) + digit_text


def _is_xml_character(codepoint: int) -> bool:
    return (
        codepoint in {0x9, 0xA, 0xD}
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )
