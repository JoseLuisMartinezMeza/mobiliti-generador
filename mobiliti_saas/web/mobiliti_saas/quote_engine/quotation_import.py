from __future__ import annotations

from collections.abc import Sequence
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import PurePath
import hashlib
import json
import mimetypes
import posixpath
import re
import unicodedata
import uuid
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from .supplier_catalog import resolve_conversion_rate


ALLOWED_IMPORT_CURRENCIES = frozenset({"MXN", "USD", "EUR"})
XLSX_MAX_ROWS = 1_048_576
MOBILITI_FIRST_SECTION_ROW = 13
MOBILITI_BASE_SECTIONS = 16
MOBILITI_BASE_PRODUCTS = 33
MOBILITI_RESERVED_ROWS_AFTER_TOTAL = 64
MAX_QUOTE_REQUEST_BYTES = 25 * 1024 * 1024
MAX_IMPORTED_LINES = XLSX_MAX_ROWS - MOBILITI_RESERVED_ROWS_AFTER_TOTAL
MAX_TEXT_LENGTH = 1_000
MAX_DESCRIPTION_LENGTH = 10_000
MAX_FILENAME_LENGTH = 255
MAX_MONEY = Decimal("1000000000")
MAX_QUANTITY = Decimal("1000000")
MAX_XLSX_INPUT_BYTES = MAX_QUOTE_REQUEST_BYTES
MAX_ZIP_ENTRIES = 5_000
MAX_ZIP_MEMBER_UNCOMPRESSED = 100 * 1024 * 1024
MAX_ZIP_TOTAL_UNCOMPRESSED = 200 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 200
SIX_PLACES = Decimal("0.000001")
TWO_PLACES = Decimal("0.01")
Q_HEADER_ROW = 7
WIRE_QUANTITY_PATTERN = re.compile(r"^(?:0|[1-9][0-9]{0,6})(?:\.[0-9]{1,6})?$")
WIRE_PRICE_PATTERN = re.compile(r"^(?:0|[1-9][0-9]{0,9})(?:\.[0-9]{1,6})?$")

MANIFEST_FIELDS = {
    "schema_version", "import_id", "source_hash", "original_filename", "provider",
    "source_currency", "currency_status", "columns", "sections", "items",
}
MANIFEST_ITEM_FIELDS = {
    "key", "source_row", "category", "name", "description", "dimension",
    "quantity", "unit_price", "source_currency", "row_hash", "source_reference",
}
MANIFEST_SECTION_FIELDS = {"id", "title", "item_keys"}
RAW_ITEM_FIELDS = {
    "kind", "import_id", "source_row", "source_currency", "quantity", "overrides",
}
OVERRIDE_FIELDS = {"name", "description", "dimension", "unit_price", "provider"}

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rels": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def required_mobiliti_rows(section_counts: Sequence[int]) -> int:
    counts = list(section_counts)
    if any(type(count) is not int or count < 0 for count in counts):
        raise ValueError("Cantidad de productos por seccion invalida")
    visible = counts + [0] * max(0, MOBILITI_BASE_SECTIONS - len(counts))
    return MOBILITI_FIRST_SECTION_ROW + sum(
        max(MOBILITI_BASE_PRODUCTS, count) + 2 for count in visible
    )


def validate_quote_size(*, section_counts: Sequence[int], encoded_bytes: int) -> None:
    counts = list(section_counts)
    if not counts:
        raise ValueError("La cotizacion debe contener al menos una linea")
    required_rows = required_mobiliti_rows(counts)
    if sum(counts) < 1:
        raise ValueError("La cotizacion debe contener al menos una linea")
    if type(encoded_bytes) is not int or encoded_bytes < 0:
        raise ValueError("Tamano codificado de cotizacion invalido")
    final_row = required_rows + MOBILITI_RESERVED_ROWS_AFTER_TOTAL
    if final_row > XLSX_MAX_ROWS:
        raise ValueError(
            f"La cotizacion requiere la fila {final_row}; "
            f"XLSX permite hasta {XLSX_MAX_ROWS} filas"
        )
    if encoded_bytes > MAX_QUOTE_REQUEST_BYTES:
        raise ValueError(
            f"La cotizacion tiene {encoded_bytes} bytes y excede el limite "
            f"de {MAX_QUOTE_REQUEST_BYTES} bytes"
        )


def build_import_manifest(
    source_bytes: bytes,
    import_id: str,
    original_filename: str,
) -> tuple[dict, dict[int, tuple[bytes, str]]]:
    source = _source_bytes(source_bytes)
    canonical_import_id = str(uuid.UUID(import_id))
    filename = safe_filename(original_filename)
    source_hash = hashlib.sha256(source).hexdigest()
    items, columns = read_items_from_bytes(source)
    products = [item for item in items if item["tipo"] == "producto"]
    if not products:
        raise ValueError("La quotation debe contener al menos un producto")
    sections = _manifest_sections(items, canonical_import_id)
    validate_quote_size(
        section_counts=[len(section["item_keys"]) for section in sections],
        encoded_bytes=len(source),
    )

    rows = [
        _manifest_item(item, canonical_import_id, filename)
        for item in products
    ]
    manifest = {
        "schema_version": 1,
        "import_id": canonical_import_id,
        "source_hash": source_hash,
        "original_filename": filename,
        "provider": _provider_from_workbook_bytes(source),
        "source_currency": _uniform_currency(rows),
        "currency_status": "detected" if all(item["source_currency"] for item in rows) else "required",
        "columns": columns,
        "sections": sections,
        "items": rows,
    }
    checked = validate_import_manifest(manifest)
    product_rows = {item["source_row"] for item in rows}
    image_map = {
        row: image
        for row, image in extract_images_from_bytes(source).items()
        if row in product_rows
    }
    return checked, image_map


def normalize_imported_items(
    raw_items: list[dict],
    manifest: dict,
    source_currency: str | None,
    quote_currency: str,
    rate_rows: list[dict],
    discount_percent: str,
) -> list[dict]:
    checked = validate_import_manifest(manifest)
    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError("Items importados invalidos")
    validate_quote_size(section_counts=[len(raw_items)], encoded_bytes=0)

    fallback_currency = _currency(source_currency, "Moneda de origen requerida", allow_none=True)
    destination = _currency(quote_currency, "Moneda de cotizacion invalida")
    discount = _decimal(discount_percent, "Descuento invalido", minimum=Decimal(0), maximum=Decimal(100))
    originals = {item["source_row"]: item for item in checked["items"]}
    normalized: list[dict] = []
    seen: set[int] = set()
    for raw in raw_items:
        row = _source_row(raw)
        if row in seen or row not in originals:
            raise ValueError("Fila importada invalida")
        seen.add(row)
        _raw_item(raw, checked["import_id"])
        original = originals[row]
        overrides = _import_overrides(raw["overrides"])
        explicit_currency = original.get("source_currency")
        payload_currency = _currency(
            raw["source_currency"], "Moneda de origen requerida", allow_none=True
        )
        if explicit_currency:
            if payload_currency is not None and payload_currency != explicit_currency:
                raise ValueError("Moneda de origen explicita no coincide con la fila importada")
            currency = explicit_currency
        else:
            currency = _currency(
                payload_currency or fallback_currency,
                "Moneda de origen requerida",
            )
        quantity = _wire_decimal(
            raw["quantity"],
            "Cantidad importada invalida",
            pattern=WIRE_QUANTITY_PATTERN,
            minimum=Decimal("0.000001"),
            maximum=MAX_QUANTITY,
        )
        original_price = _decimal(
            overrides["unit_price"],
            "Precio importado invalido",
            minimum=Decimal(0),
        )
        rate = _conversion_rate(currency, destination, rate_rows)
        converted = (original_price * rate).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
        normalized.append(
            {
                "kind": "imported",
                "key": original["key"],
                "import_id": checked["import_id"],
                "source_row": row,
                "category": original["category"],
                "name": overrides["name"],
                "description": overrides["description"],
                "dimension": overrides["dimension"],
                "provider": overrides["provider"],
                "quantity": _plain_decimal(quantity),
                "original_unit_price": _fixed(original_price, 6),
                "original_currency": currency,
                "unit_price": _fixed(converted, 2),
                "frozen_exchange_rate": _fixed(rate, 6),
                "discount_percent": _fixed(discount, 6),
                "source_hash": checked["source_hash"],
                "row_hash": original["row_hash"],
                "source_reference": original["source_reference"],
            }
        )
    return normalized


def validate_import_manifest(manifest: dict) -> dict:
    if not isinstance(manifest, dict) or set(manifest) != MANIFEST_FIELDS:
        raise ValueError("Manifiesto de importacion invalido")
    if manifest["schema_version"] != 1:
        raise ValueError("Version de manifiesto invalida")
    try:
        import_id = str(uuid.UUID(manifest["import_id"]))
    except (AttributeError, TypeError, ValueError):
        raise ValueError("import_id invalido") from None
    if import_id != manifest["import_id"]:
        raise ValueError("import_id invalido")
    source_hash = _sha256(manifest["source_hash"], "source_hash")
    filename = safe_filename(manifest["original_filename"])
    if filename != manifest["original_filename"]:
        raise ValueError("Nombre de archivo invalido")
    provider = _text(manifest["provider"], "Proveedor", allow_empty=True)
    source_currency = _currency(manifest["source_currency"], "Moneda de origen invalida", allow_none=True)
    if manifest["currency_status"] not in {"detected", "required"}:
        raise ValueError("Estado de moneda invalido")
    columns = _columns(manifest["columns"])
    rows = manifest["items"]
    if not isinstance(rows, list) or not rows:
        raise ValueError("La quotation debe contener al menos un producto")
    if len(rows) > MAX_IMPORTED_LINES:
        validate_quote_size(section_counts=[len(rows)], encoded_bytes=0)

    checked_rows = [_validate_manifest_item(item, import_id, filename) for item in rows]
    row_numbers = [item["source_row"] for item in checked_rows]
    if len(set(row_numbers)) != len(row_numbers):
        raise ValueError("Filas de manifiesto duplicadas")
    if source_currency != _uniform_currency(checked_rows):
        raise ValueError("Moneda de origen inconsistente")
    expected_status = "detected" if all(item["source_currency"] for item in checked_rows) else "required"
    if manifest["currency_status"] != expected_status:
        raise ValueError("Estado de moneda inconsistente")

    sections = _validate_manifest_sections(manifest["sections"], import_id, checked_rows)
    validate_quote_size(
        section_counts=[len(section["item_keys"]) for section in sections],
        encoded_bytes=0,
    )
    return {
        "schema_version": 1,
        "import_id": import_id,
        "source_hash": source_hash,
        "original_filename": filename,
        "provider": provider,
        "source_currency": source_currency,
        "currency_status": expected_status,
        "columns": columns,
        "sections": sections,
        "items": checked_rows,
    }


def read_items_from_bytes(source_bytes: bytes) -> tuple[list[dict], dict[str, str]]:
    source = _source_bytes(source_bytes)
    try:
        workbook = load_workbook(BytesIO(source), data_only=False)
    except Exception as exc:
        raise ValueError("El archivo .xlsx es invalido") from exc
    try:
        if "Quotation" not in workbook.sheetnames:
            raise ValueError("El archivo no contiene hoja Quotation")
        sheet = workbook["Quotation"]
        columns = _detect_columns(sheet)
        description_column = _column_index(columns, "descripcion", "D")
        dimension_column = _column_index(columns, "dimension", "E")
        quantity_column = _column_index(columns, "cantidad", "G")
        price_column = _column_index(columns, "list_price", columns.get("unit_price", "J"))
        last_row = _last_data_row(sheet)
        current_category = ""
        items: list[dict] = []
        for row in range(Q_HEADER_ROW + 1, last_row + 1):
            number = sheet.cell(row=row, column=1).value
            name = sheet.cell(row=row, column=2).value
            if isinstance(number, str) and number.startswith("-"):
                current_category = str(number).strip("- ").strip()
                items.append({"tipo": "categoria", "row": row, "nombre": current_category})
                continue
            if (name is None or name == "") and (number is None or number == ""):
                continue
            if isinstance(number, (int, float)) and not isinstance(number, bool):
                items.append(
                    {
                        "tipo": "producto",
                        "row": row,
                        "nombre": name,
                        "descripcion": sheet.cell(row=row, column=description_column).value,
                        "dimension": sheet.cell(row=row, column=dimension_column).value,
                        "cantidad": sheet.cell(row=row, column=quantity_column).value,
                        "precio": sheet.cell(row=row, column=price_column).value,
                        "categoria": current_category,
                        "moneda_original": _optional_cell(sheet, row, columns, "moneda_original"),
                    }
                )
            elif (number is None or number == "") and name:
                current_category = str(name).strip()
                items.append({"tipo": "categoria", "row": row, "nombre": current_category})
        return items, columns
    finally:
        workbook.close()


def extract_images_from_bytes(source_bytes: bytes) -> dict[int, tuple[bytes, str]]:
    source = _source_bytes(source_bytes)
    try:
        with zipfile.ZipFile(BytesIO(source), "r") as archive:
            worksheet_path = _worksheet_path_for_sheet(archive, "Quotation")
            relationship_path = _rels_path_for_part(worksheet_path)
            if relationship_path not in archive.namelist():
                return {}
            worksheet_rels = ET.fromstring(archive.read(relationship_path))
            drawing_path = next(
                (
                    _resolve_zip_part(worksheet_path, relation.get("Target") or "")
                    for relation in worksheet_rels.findall(".//rels:Relationship", _NS)
                    if "drawing" in (relation.get("Target") or "").lower()
                    and "vml" not in (relation.get("Target") or "").lower()
                ),
                None,
            )
            if not drawing_path or drawing_path not in archive.namelist():
                return {}
            drawing_rels_path = f"xl/drawings/_rels/{posixpath.basename(drawing_path)}.rels"
            relation_to_media: dict[str, str] = {}
            if drawing_rels_path in archive.namelist():
                drawing_rels = ET.fromstring(archive.read(drawing_rels_path))
                for relation in drawing_rels.findall(".//rels:Relationship", _NS):
                    if relation.get("TargetMode") != "External":
                        relation_to_media[relation.get("Id") or ""] = _resolve_zip_part(
                            drawing_path, relation.get("Target") or ""
                        )
            drawing = ET.fromstring(archive.read(drawing_path))
            images: dict[int, tuple[bytes, str]] = {}
            anchors = [
                *drawing.findall(".//xdr:twoCellAnchor", _NS),
                *drawing.findall(".//xdr:oneCellAnchor", _NS),
            ]
            for anchor in anchors:
                blip = anchor.find(".//a:blip", _NS)
                row_node = anchor.find("xdr:from/xdr:row", _NS)
                if blip is None or row_node is None:
                    continue
                relationship_id = blip.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                )
                media_path = relation_to_media.get(relationship_id or "")
                if not media_path or media_path not in archive.namelist():
                    continue
                media_type = mimetypes.guess_type(media_path)[0] or "application/octet-stream"
                if not media_type.startswith("image/"):
                    continue
                images[int(row_node.text or "0") + 1] = (archive.read(media_path), media_type)
            return images
    except (ET.ParseError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("No se pudieron extraer imagenes de la quotation") from exc


def safe_filename(value: object) -> str:
    if not isinstance(value, str):
        raise ValueError("Nombre de archivo invalido")
    filename = PurePath(value.replace("\\", "/")).name.strip()
    if not filename or len(filename) > MAX_FILENAME_LENGTH or _contains_control(filename):
        raise ValueError("Nombre de archivo invalido")
    return filename


def _manifest_item(item: dict, import_id: str, filename: str) -> dict:
    row = item["row"]
    name = _workbook_text(item["nombre"], "Nombre")
    description = _workbook_text(
        item["descripcion"], "Descripcion", allow_empty=True,
        maximum=MAX_DESCRIPTION_LENGTH,
    )
    dimension = _workbook_text(
        item["dimension"], "Dimension", allow_empty=True, allow_none=True
    )
    quantity = _quantity(item["cantidad"])
    unit_price = _decimal(item["precio"], "Precio unitario invalido", minimum=Decimal(0))
    currency = _currency(item.get("moneda_original"), "Moneda de origen invalida", allow_none=True)
    row_data = {
        "key": f"import:{import_id}:{row}",
        "source_row": row,
        "category": _workbook_text(item.get("categoria", ""), "Categoria", allow_empty=True),
        "name": name,
        "description": description,
        "dimension": dimension,
        "quantity": _plain_decimal(quantity),
        "unit_price": _plain_decimal(unit_price),
        "source_currency": currency,
        "source_reference": f"{filename}#Quotation!{row}",
    }
    row_data["row_hash"] = _row_hash(row_data, None)
    return row_data


def _manifest_sections(items: list[dict], import_id: str) -> list[dict]:
    sections: list[dict] = []
    current: dict | None = None
    for item in items:
        if item["tipo"] == "categoria":
            current = {
                "id": f"import-section-{len(sections) + 1}",
                "title": _workbook_text(item["nombre"], "Categoria"),
                "item_keys": [],
            }
            sections.append(current)
        elif item["tipo"] == "producto":
            if current is None:
                current = {
                    "id": f"import-section-{len(sections) + 1}",
                    "title": "Sin categoria",
                    "item_keys": [],
                }
                sections.append(current)
            current["item_keys"].append(f"import:{import_id}:{item['row']}")
    return [section for section in sections if section["item_keys"]]


def _validate_manifest_item(item: object, import_id: str, filename: str) -> dict:
    if not isinstance(item, dict) or set(item) != MANIFEST_ITEM_FIELDS:
        raise ValueError("Item de manifiesto invalido")
    row = item["source_row"]
    if type(row) is not int or row <= Q_HEADER_ROW:
        raise ValueError("Fila de manifiesto invalida")
    expected_key = f"import:{import_id}:{row}"
    if item["key"] != expected_key:
        raise ValueError("Clave de manifiesto invalida")
    checked = {
        "key": expected_key,
        "source_row": row,
        "category": _text(item["category"], "Categoria", allow_empty=True),
        "name": _text(item["name"], "Nombre"),
        "description": _text(item["description"], "Descripcion", allow_empty=True, maximum=MAX_DESCRIPTION_LENGTH),
        "dimension": _text(item["dimension"], "Dimension", allow_empty=True),
        "quantity": _plain_decimal(_quantity(item["quantity"])),
        "unit_price": _plain_decimal(_decimal(item["unit_price"], "Precio unitario invalido", minimum=Decimal(0))),
        "source_currency": _currency(item["source_currency"], "Moneda de origen invalida", allow_none=True),
        "source_reference": f"{filename}#Quotation!{row}",
    }
    if item["source_reference"] != checked["source_reference"]:
        raise ValueError("Referencia de fuente invalida")
    expected_hash = _row_hash(checked, None)
    supplied_hash = _sha256(item["row_hash"], "row_hash")
    if supplied_hash != expected_hash:
        raise ValueError("row_hash invalido")
    checked["row_hash"] = supplied_hash
    return checked


def _validate_manifest_sections(sections: object, import_id: str, rows: list[dict]) -> list[dict]:
    if not isinstance(sections, list) or not sections:
        raise ValueError("Secciones de manifiesto invalidas")
    expected_keys = {item["key"] for item in rows}
    seen_keys: list[str] = []
    checked: list[dict] = []
    for position, section in enumerate(sections, start=1):
        if not isinstance(section, dict) or set(section) != MANIFEST_SECTION_FIELDS:
            raise ValueError("Seccion de manifiesto invalida")
        if section["id"] != f"import-section-{position}":
            raise ValueError("Identificador de seccion invalido")
        item_keys = section["item_keys"]
        if not isinstance(item_keys, list) or not item_keys:
            raise ValueError("Items de seccion invalidos")
        for key in item_keys:
            if not isinstance(key, str) or not re.fullmatch(rf"import:{re.escape(import_id)}:[1-9]\d*", key):
                raise ValueError("Clave de seccion invalida")
            seen_keys.append(key)
        checked.append(
            {
                "id": section["id"],
                "title": _text(section["title"], "Titulo de seccion"),
                "item_keys": list(item_keys),
            }
        )
    if len(seen_keys) != len(set(seen_keys)) or set(seen_keys) != expected_keys:
        raise ValueError("Cobertura de secciones invalida")
    return checked


def _raw_item(raw: object, import_id: str) -> None:
    if not isinstance(raw, dict) or set(raw) != RAW_ITEM_FIELDS:
        raise ValueError("Item importado invalido")
    if raw["kind"] != "imported" or raw["import_id"] != import_id:
        raise ValueError("Item importado invalido")


def _source_row(raw: object) -> int:
    if not isinstance(raw, dict):
        raise ValueError("Fila importada invalida")
    row = raw.get("source_row")
    if type(row) is not int or row <= Q_HEADER_ROW:
        raise ValueError("Fila importada invalida")
    return row


def _import_overrides(value: object) -> dict[str, str]:
    if not isinstance(value, dict) or set(value) != OVERRIDE_FIELDS:
        raise ValueError("Overrides importados invalidos")
    return {
        "name": _text(value["name"], "Nombre"),
        "description": _text(value["description"], "Descripcion", allow_empty=True, maximum=MAX_DESCRIPTION_LENGTH),
        "dimension": _text(value["dimension"], "Dimension", allow_empty=True),
        "unit_price": _plain_decimal(_wire_decimal(
            value["unit_price"],
            "Precio importado invalido",
            pattern=WIRE_PRICE_PATTERN,
            minimum=Decimal(0),
            maximum=MAX_MONEY,
        )),
        "provider": _text(value["provider"], "Proveedor"),
    }


def _conversion_rate(base_currency: str, quote_currency: str, rate_rows: object) -> Decimal:
    if not isinstance(rate_rows, list):
        raise ValueError("Filas de tasa invalidas")
    normalized_rows = []
    for raw in rate_rows:
        if not isinstance(raw, dict) or set(raw) not in (
            {"currency", "mxn_per_unit", "effective_date"},
            {"currency", "mxn_per_unit", "effective_date", "retrieved_at"},
        ):
            raise ValueError("Fila de tasa invalida")
        row = dict(raw)
        row.setdefault("retrieved_at", f"{row['effective_date']}T00:00:00Z")
        normalized_rows.append(row)
    return resolve_conversion_rate(base_currency, quote_currency, normalized_rows, date.today()).exchange_rate


def _provider_from_workbook_bytes(source_bytes: bytes) -> str:
    try:
        workbook = load_workbook(BytesIO(_source_bytes(source_bytes)), data_only=False, read_only=True)
        try:
            if "Quotation" not in workbook.sheetnames:
                raise ValueError("El archivo no contiene hoja Quotation")
            return _workbook_text(
                workbook["Quotation"]["A1"].value,
                "Proveedor",
                allow_empty=True,
                allow_none=True,
            )
        finally:
            workbook.close()
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("El archivo .xlsx es invalido") from exc


def _detect_columns(sheet) -> dict[str, str]:
    keywords = {
        "cantidad": ["qty", "quantity", "cantidad"],
        "unit_price": ["unit price", "unitprice", "price unit", "precio unitario"],
        "total_price": ["tot price", "total price", "totprice", "amount", "precio total"],
        "list_price": ["list price", "listprice", "price list"],
        "descripcion": ["description", "desc", "descripcion"],
        "dimension": ["dimension", "dimensions", "size", "medida"],
        "moneda_original": ["original currency", "base currency", "moneda original"],
    }
    result: dict[str, str] = {}
    volume = _find_header_column(sheet, ["vol", "volumen", "volume"])
    if volume:
        result["m3"] = volume
    for key, terms in keywords.items():
        found = _find_header_column(sheet, terms)
        if found:
            result[key] = found
    if "m3" not in result:
        result["m3"] = result.get("dimension", "E")
    if "unit_price" not in result and "list_price" not in result:
        for column in range(1, sheet.max_column + 1):
            if "price" in _normalize_header(sheet.cell(Q_HEADER_ROW, column).value):
                result["unit_price"] = get_column_letter(column)
                break
    return result


def _find_header_column(sheet, terms: list[str]) -> str | None:
    for column in range(1, sheet.max_column + 1):
        header = _normalize_header(sheet.cell(Q_HEADER_ROW, column).value)
        if not header:
            continue
        tokens = set(header.split())
        if any(term in header or header in term or term in tokens for term in terms):
            return get_column_letter(column)
    return None


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value).lower().strip()
    text = "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _optional_cell(sheet, row: int, columns: dict[str, str], key: str) -> object:
    column = columns.get(key)
    return sheet.cell(row=row, column=column_index_from_string(column)).value if column else None


def _column_index(columns: dict[str, str], key: str, fallback: str) -> int:
    return column_index_from_string(columns.get(key, fallback))


def _last_data_row(sheet) -> int:
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=1).value is not None:
            return row
    return 0


def _worksheet_path_for_sheet(archive: zipfile.ZipFile, sheet_name: str) -> str:
    default = "xl/worksheets/sheet1.xml"
    if "xl/workbook.xml" not in archive.namelist() or "xl/_rels/workbook.xml.rels" not in archive.namelist():
        return default
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        relation.get("Id") or "": relation.get("Target") or ""
        for relation in relationships.findall(".//rels:Relationship", _NS)
    }
    for sheet in workbook.findall(".//main:sheet", _NS):
        if str(sheet.get("name") or "").strip().lower() == sheet_name.lower():
            relation_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            return _resolve_zip_part("xl/workbook.xml", targets.get(relation_id or "", ""))
    return default


def _rels_path_for_part(part_path: str) -> str:
    return posixpath.join(posixpath.dirname(part_path), "_rels", f"{posixpath.basename(part_path)}.rels")


def _resolve_zip_part(base_part: str, target: str) -> str:
    clean_target = target.replace("\\", "/")
    if clean_target.startswith("/"):
        return clean_target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), clean_target))


def _source_bytes(value: object) -> bytes:
    if not isinstance(value, bytes) or not value:
        raise ValueError("Archivo de quotation invalido")
    if len(value) > MAX_XLSX_INPUT_BYTES:
        raise ValueError(
            f"El archivo .xlsx tiene {len(value)} bytes y excede el limite "
            f"de {MAX_XLSX_INPUT_BYTES} bytes"
        )
    _preflight_xlsx_zip(value)
    return value


def _preflight_xlsx_zip(source: bytes) -> None:
    """Valida el directorio central del XLSX sin descomprimir ninguno de sus miembros."""
    try:
        with zipfile.ZipFile(BytesIO(source), "r") as archive:
            members = archive.infolist()
    except (zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise ValueError("El archivo .xlsx inseguro no es un ZIP valido") from exc

    if not members or len(members) > MAX_ZIP_ENTRIES:
        raise ValueError("El archivo .xlsx inseguro contiene demasiados miembros")

    names: set[str] = set()
    total_uncompressed = 0
    for member in members:
        name = member.filename
        if not _safe_zip_member_name(name) or name in names:
            raise ValueError("El archivo .xlsx inseguro contiene nombres de miembros anómalos")
        names.add(name)
        if member.file_size > MAX_ZIP_MEMBER_UNCOMPRESSED:
            raise ValueError("El archivo .xlsx inseguro excede el limite por miembro")
        total_uncompressed += member.file_size
        if total_uncompressed > MAX_ZIP_TOTAL_UNCOMPRESSED:
            raise ValueError("El archivo .xlsx inseguro excede el limite descomprimido")
        if member.file_size and (
            not member.compress_size
            or member.file_size > member.compress_size * MAX_ZIP_COMPRESSION_RATIO
        ):
            raise ValueError("El archivo .xlsx inseguro excede el ratio de compresion")


def _safe_zip_member_name(name: object) -> bool:
    if not isinstance(name, str) or not name or len(name) > 1_024 or "\x00" in name or "\\" in name:
        return False
    if name.startswith("/") or re.match(r"^[A-Za-z]:", name):
        return False
    parts = name.split("/")
    return all(part not in {"", ".", ".."} for part in parts)


def _columns(value: object) -> dict[str, str]:
    if not isinstance(value, dict):
        raise ValueError("Columnas de manifiesto invalidas")
    result = {}
    for key, column in value.items():
        if not isinstance(key, str) or not isinstance(column, str) or not re.fullmatch(r"[A-Z]{1,3}", column):
            raise ValueError("Columnas de manifiesto invalidas")
        result[key] = column
    return result


def _currency(value: object, error: str, *, allow_none: bool = False) -> str | None:
    if value is None and allow_none:
        return None
    if not isinstance(value, str) or value.strip().upper() not in ALLOWED_IMPORT_CURRENCIES:
        raise ValueError(error)
    return value.strip().upper()


def _quantity(value: object) -> Decimal:
    return _decimal(value, "Cantidad importada invalida", minimum=Decimal("0.000001"), maximum=MAX_QUANTITY)


def _wire_decimal(
    value: object,
    error: str,
    *,
    pattern: re.Pattern[str],
    minimum: Decimal,
    maximum: Decimal,
) -> Decimal:
    if not isinstance(value, str):
        raise ValueError(error)
    text = value
    if not pattern.fullmatch(text):
        raise ValueError(error)
    number = Decimal(text)
    if number < minimum or number > maximum:
        raise ValueError(error)
    return number


def _decimal(value: object, error: str, *, minimum: Decimal, maximum: Decimal = MAX_MONEY) -> Decimal:
    if isinstance(value, bool) or not isinstance(value, (str, int, float, Decimal)):
        raise ValueError(error)
    text = str(value).strip()
    if not text or len(text) > 64:
        raise ValueError(error)
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(error) from None
    if not number.is_finite() or number < minimum or number > maximum or max(-number.as_tuple().exponent, 0) > 6:
        raise ValueError(error)
    return number


def _text(value: object, error: str, *, allow_empty: bool = False, maximum: int = MAX_TEXT_LENGTH) -> str:
    if not isinstance(value, str):
        raise ValueError(error)
    text = unicodedata.normalize("NFC", value).strip()
    if len(text) > maximum or _contains_control(text) or text.lstrip()[:1] in {"=", "+", "-", "@"}:
        raise ValueError(error)
    if not text and not allow_empty:
        raise ValueError(error)
    return text


def _workbook_text(
    value: object,
    error: str,
    *,
    allow_empty: bool = False,
    allow_none: bool = False,
    maximum: int = MAX_TEXT_LENGTH,
) -> str:
    if value is None and allow_none:
        value = ""
    if isinstance(value, str):
        value = re.sub(r"[\t\r\n]+", " ", value)
    return _text(value, error, allow_empty=allow_empty, maximum=maximum)


def _contains_control(value: str) -> bool:
    return any(unicodedata.category(character) == "Cc" for character in value)


def _sha256(value: object, error: str) -> str:
    if not isinstance(value, str) or not re.fullmatch(r"[0-9a-f]{64}", value):
        raise ValueError(error)
    return value


def _row_hash(row: dict, source_hash: str | None) -> str:
    payload = {key: value for key, value in row.items() if key != "row_hash"}
    if source_hash is not None:
        payload["source_hash"] = source_hash
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _uniform_currency(rows: list[dict]) -> str | None:
    currencies = {item["source_currency"] for item in rows}
    return next(iter(currencies)) if len(currencies) == 1 and None not in currencies else None


def _plain_decimal(value: Decimal) -> str:
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _fixed(value: Decimal, places: int) -> str:
    return f"{value:.{places}f}"
