from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import json
import math
from pathlib import Path
import re
import tempfile
from typing import Any
import unicodedata
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.styles import Font

from .catalog_cart import (
    MAX_EXCEL_CELL_TEXT_LENGTH,
    _set_column_widths,
    catalog_quotation_item_text,
    write_catalog_quotation_headers,
    write_catalog_quotation_item,
)
from .offiho_catalog import build_offiho_cart_payload
from .supplier_catalog import (
    MAX_ATTRIBUTES_DEPTH,
    MAX_ATTRIBUTES_JSON_BYTES,
    _validate_attributes,
    build_supplier_cart_payload,
    resolve_conversion_rate,
    safe_excel_text,
)
from .tarkett_catalog import build_tarkett_cart_payload


MIXED_CATALOG_CART_SOURCE_TYPE = "mixed_catalog_cart"
MIXED_CATALOG_ORDER = ("tarkett", "offiho", "cr-global", "sonara", "sunon", "alma", "lumbro")
MIXED_CATALOG_LABELS = {
    "tarkett": "Tarkett", "offiho": "Offiho", "cr-global": "CR Global",
    "sonara": "Sonara", "sunon": "Sunon", "alma": "ALMA", "lumbro": "Lumbro",
}
MIXED_GROUP_SOURCE_TYPES = {
    "tarkett": "tarkett_cart",
    "offiho": "offiho_cart",
    "cr-global": "supplier_cart",
    "sonara": "supplier_cart",
    "sunon": "supplier_cart",
    "alma": "supplier_cart",
    "lumbro": "supplier_cart",
}
MIXED_EXPECTED_BASE_CURRENCY = {
    "tarkett": "MXN", "offiho": "MXN", "cr-global": "MXN", "sonara": "MXN",
    "sunon": "USD", "alma": "USD", "lumbro": "MXN",
}
MIXED_QUOTE_CURRENCIES = frozenset({"MXN", "USD", "EUR"})
MAX_MIXED_CATALOG_LINES = 500
MAX_MIXED_REQUEST_BYTES = 1_000_000
MAX_MIXED_PAYLOAD_BYTES = 5_000_000
MAX_MIXED_TEXT = 2_000
MAX_MIXED_URL = 2_048
MAX_MIXED_WARNINGS = 50
MAX_MIXED_IDENTITY = 1_000
MAX_MIXED_OPTIONS_PER_LINE = 200
MIXED_ALLOWED_FIELDS = {
    "tarkett": frozenset({"catalog", "code", "quantity"}),
    "offiho": frozenset({"catalog", "inventory_key", "quantity"}),
    "supplier": frozenset({"catalog", "internal_id", "quantity", "base_option_id", "add_on_option_ids"}),
}
MIXED_REQUIRED_FIELDS = {
    "tarkett": frozenset({"catalog", "code", "quantity"}),
    "offiho": frozenset({"catalog", "inventory_key", "quantity"}),
    "supplier": frozenset({"catalog", "internal_id", "quantity"}),
}
MIXED_LINE_FIELDS = frozenset({
    "canonical_key", "catalog", "supplier", "code", "name", "description", "unit",
    "quantity", "unit_price", "discount_percent", "original_currency",
    "original_unit_price", "frozen_exchange_rate", "source_reference", "price_mode",
    "auto_electrification", "tax_rate", "image_url", "product_url", "warnings",
    "code_status", "configuration", "attributes", "variant", "availability_type",
    "available_quantity", "stock", "lead_time", "price_source", "stock_status",
    "image_kind", "reservation",
})
MIXED_RESERVATION_RESULT_FIELDS = frozenset({"reserved_quantity", "available_after_reservations", "reserved_by_others"})
MIXED_GROUP_FIELDS = frozenset({
    "catalog", "catalog_source_hash", "base_currency", "quote_currency", "exchange_rate",
    "rate_source", "rate_effective_date", "rate_retrieved_at", "items",
})
AUTO_ELECTRIFICATION_RATE_FIELDS = (
    "base_currency", "quote_currency", "exchange_rate", "rate_source", "rate_effective_date", "rate_retrieved_at",
)
SIX_PLACES = Decimal("0.000001")
TWO_PLACES = Decimal("0.01")
EXCEL_ILLEGAL_CONTROL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _field_family(catalog: str) -> str:
    return catalog if catalog in {"tarkett", "offiho"} else "supplier"


def _identity_text(value: object, field: str, *, allow_empty: bool = False, limit: int = MAX_MIXED_IDENTITY) -> str:
    if not isinstance(value, str):
        raise ValueError(f"{field} invalido")
    text = value.strip()
    if (not text and not allow_empty) or len(text) > limit or any(unicodedata.category(char) in {"Cc", "Cf", "Cs"} for char in text):
        raise ValueError(f"{field} invalido")
    return text


def _validate_browser_row(raw: object) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raise ValueError("Cada producto mixto debe ser un objeto")
    catalog = raw.get("catalog")
    if not isinstance(catalog, str) or catalog not in MIXED_CATALOG_ORDER:
        raise ValueError("Catalogo mixto no soportado")
    family = _field_family(catalog)
    unexpected = set(raw) - MIXED_ALLOWED_FIELDS[family]
    if unexpected:
        raise ValueError(f"Campo mixto no permitido: {min(map(str, unexpected))}")
    missing = MIXED_REQUIRED_FIELDS[family] - set(raw)
    if missing:
        raise ValueError(f"Campo mixto requerido: {min(missing)}")
    normalized = dict(raw)
    identity_field = {"tarkett": "code", "offiho": "inventory_key", "supplier": "internal_id"}[family]
    normalized[identity_field] = _identity_text(normalized[identity_field], identity_field)
    quantity = normalized.get("quantity")
    if type(quantity) not in {str, int, float}:
        raise ValueError("quantity invalida")
    if isinstance(quantity, float) and not math.isfinite(quantity):
        raise ValueError("quantity invalida")
    quantity_text = str(quantity)
    if len(quantity_text) > 64 or any(unicodedata.category(char) in {"Cc", "Cf", "Cs"} for char in quantity_text):
        raise ValueError("quantity invalida")
    if family == "supplier":
        normalized.setdefault("base_option_id", "")
        normalized.setdefault("add_on_option_ids", [])
        normalized["base_option_id"] = _identity_text(normalized["base_option_id"], "base_option_id", allow_empty=True, limit=500)
        values = normalized["add_on_option_ids"]
        if not isinstance(values, list) or len(values) > MAX_MIXED_OPTIONS_PER_LINE:
            raise ValueError("add_on_option_ids debe ser una lista de textos")
        cleaned = []
        for value in values:
            if not isinstance(value, str) or not value.strip() or len(value.strip()) > 500 or any(unicodedata.category(char) in {"Cc", "Cf", "Cs"} for char in value):
                raise ValueError("add_on_option_ids debe ser una lista de textos")
            cleaned.append(value.strip())
        if len(cleaned) != len(set(cleaned)):
            raise ValueError("add_on_option_ids contiene duplicados")
        normalized["add_on_option_ids"] = sorted(cleaned)
    return normalized


def preflight_mixed_catalog_items(raw_items: object) -> list[dict[str, Any]]:
    if not isinstance(raw_items, list) or not 1 <= len(raw_items) <= MAX_MIXED_CATALOG_LINES:
        raise ValueError("El carrito mixto debe contener entre 1 y 500 filas")
    normalized = [_validate_browser_row(candidate) for candidate in raw_items]
    try:
        encoded = json.dumps(normalized, ensure_ascii=False, separators=(",", ":"), allow_nan=False).encode("utf-8")
    except (TypeError, ValueError, UnicodeError) as exc:
        raise ValueError("Carrito mixto invalido") from exc
    if len(encoded) > MAX_MIXED_REQUEST_BYTES:
        raise ValueError("Carrito mixto excede el limite de bytes")
    return normalized


def mixed_cart_key(raw: dict[str, Any]) -> str:
    catalog = str(raw["catalog"])
    if catalog == "tarkett":
        return f"tarkett:{str(raw.get('code') or '').strip()}"
    if catalog == "offiho":
        return f"offiho:{str(raw.get('inventory_key') or '').strip()}"
    identity = [str(raw.get("internal_id") or "").strip(), str(raw.get("base_option_id") or "").strip(), list(raw.get("add_on_option_ids", []))]
    return f"{catalog}:{json.dumps(identity, ensure_ascii=False, separators=(',', ':'))}"


def _group_browser_rows(raw_items: list[dict[str, object]]) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    seen: set[str] = set()
    for raw in preflight_mixed_catalog_items(raw_items):
        key = mixed_cart_key(raw)
        if key in seen:
            raise ValueError(f"Clave mixta duplicada: {key}")
        seen.add(key)
        groups.setdefault(str(raw["catalog"]), []).append(raw)
    return groups


def _commercial_discount_percent(value: object) -> Decimal:
    try:
        discount = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Descuento comercial invalido") from exc
    if not discount.is_finite() or discount < 0 or discount > 100:
        raise ValueError("Descuento comercial debe estar entre 0 y 100")
    return discount.quantize(SIX_PLACES, rounding=ROUND_HALF_UP)


def _six(value: object) -> str:
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Decimal mixto invalido") from exc
    if not number.is_finite():
        raise ValueError("Decimal mixto invalido")
    return f"{number.quantize(SIX_PLACES, rounding=ROUND_HALF_UP):.6f}"


def _stable_warnings(*groups: list[object], derived: list[str] | None = None) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in [*(item for group in groups for item in group), *(derived or [])]:
        text = str(value or "").strip()
        key = _warning_key(text)
        if text and key not in seen:
            seen.add(key)
            result.append(text)
    return result


def _warning_key(text: str) -> str:
    return " ".join("".join(char for char in unicodedata.normalize("NFKD", text.casefold()) if not unicodedata.combining(char)).split())


def _rate_payload(rate: Any) -> dict[str, str]:
    return {
        "base_currency": rate.base_currency, "quote_currency": rate.quote_currency,
        "exchange_rate": _six(rate.exchange_rate), "rate_source": rate.rate_source,
        "rate_effective_date": rate.rate_effective_date.isoformat(), "rate_retrieved_at": rate.rate_retrieved_at,
    }


def _warnings(line: dict[str, Any], *, catalog: str) -> list[str]:
    derived: list[str] = []
    if line.get("code_status") == "needs_review":
        derived.append("Codigo por verificar")
    if line.get("image_kind") == "generated_reference":
        derived.append("Imagen de referencia")
    if line.get("price_source") == "missing":
        derived.append("Precio por confirmar")
    return _stable_warnings(list(line.get("warnings") or []), derived=derived)


def _common_line(raw: dict[str, Any], *, catalog: str, rate: dict[str, str], discount: Decimal, source_hash: str) -> dict[str, Any]:
    code = str(raw.get("code") or "").strip()
    stock = _six(raw["available_quantity"])
    image_url = str(raw.get("image_url") or "")
    line = {
        "canonical_key": f"{catalog}:{code}" if catalog == "tarkett" else f"offiho:{raw['inventory_key']}",
        "catalog": catalog, "supplier": MIXED_CATALOG_LABELS[catalog], "code": code or str(raw.get("inventory_key") or ""),
        "name": str(raw["name"]), "description": "" if catalog == "tarkett" else str(raw.get("description") or ""), "unit": str(raw["unit"]),
        "quantity": _six(raw["quantity"]), "discount_percent": f"{discount:.6f}", "original_currency": "MXN",
        "original_unit_price": _six(raw["unit_price"]), "frozen_exchange_rate": rate["exchange_rate"],
        "source_reference": f"{catalog}:{source_hash}:{code if catalog == 'tarkett' else raw['inventory_key']}",
        "price_mode": "list", "auto_electrification": True, "tax_rate": "0.160000",
        "image_url": image_url, "product_url": str(raw.get("product_url") or ""), "code_status": "verified",
        "configuration": "", "attributes": {}, "variant": "" if catalog == "tarkett" else str(raw.get("variant") or ""),
        "availability_type": "stocked", "available_quantity": stock, "stock": stock, "lead_time": "",
        "price_source": str(raw.get("price_source") or "missing"),
        "stock_status": "available" if catalog == "tarkett" else str(raw.get("stock_status") or "available"),
        "image_kind": "official" if image_url else "placeholder",
    }
    original = Decimal(line["original_unit_price"])
    line["unit_price"] = f"{(original * Decimal(rate['exchange_rate'])).quantize(TWO_PLACES, rounding=ROUND_HALF_UP):.2f}"
    line["warnings"] = _warnings(line, catalog=catalog)
    identity = code if catalog == "tarkett" else str(raw["inventory_key"])
    line["reservation"] = {"identity": identity, "sku": line["code"] or identity, "quantity": line["quantity"], "stock": line["stock"]}
    return line


def _supplier_line(raw: dict[str, Any], browser: dict[str, Any], *, catalog: str, payload: dict[str, Any]) -> dict[str, Any]:
    availability = str(raw["availability_type"])
    stock = _six(raw["stock"]) if availability == "stocked" else None
    quantity = _six(raw["quantity"])
    line = {
        "canonical_key": mixed_cart_key(browser), "catalog": catalog, "supplier": MIXED_CATALOG_LABELS[catalog],
        "code": str(raw.get("sku") or ""), "name": str(raw["name"]), "description": str(raw.get("description") or ""), "unit": str(raw["unit"]),
        "quantity": quantity, "unit_price": str(raw["unit_price"]), "discount_percent": "0.000000",
        "original_currency": str(raw["base_currency"]), "original_unit_price": str(raw["unit_price_base"]),
        "frozen_exchange_rate": str(payload["exchange_rate"]), "source_reference": str(raw["source_reference"]),
        "price_mode": "net", "auto_electrification": False, "tax_rate": str(raw["tax_rate"]),
        "image_url": str(raw.get("image_url") or ""), "product_url": str(raw.get("product_url") or ""),
        "code_status": str(raw["code_status"]), "configuration": str(raw.get("configuration") or ""),
        "attributes": deepcopy(raw.get("attributes") or {}), "variant": "", "availability_type": availability,
        "available_quantity": stock, "stock": stock, "lead_time": str(raw.get("lead_time") or ""),
        "price_source": "catalog", "stock_status": "",
        "image_kind": str(raw.get("image_kind") or "placeholder"),
        "warnings": list(raw.get("warnings") or []),
    }
    if availability == "stocked":
        available = Decimal(stock or "0")
        line["stock_status"] = "out_of_stock" if available <= 0 else "insufficient_stock" if Decimal(quantity) > available else "available"
        line["reservation"] = {"identity": str(raw["internal_id"]), "sku": line["code"], "quantity": quantity, "stock": stock}
    else:
        line["reservation"] = None
    line["warnings"] = _warnings(line, catalog=catalog)
    return line


def build_mixed_catalog_cart_payload(raw_items: list[dict[str, object]], *, catalogs: dict[str, dict], rate_rows: list[dict], quote_currency: str, commercial_discount_percent: object, today: date | None = None) -> dict:
    if not isinstance(quote_currency, str) or quote_currency not in MIXED_QUOTE_CURRENCIES:
        raise ValueError("Grupos mixtos invalidos")
    if not isinstance(catalogs, dict):
        raise ValueError("Catalogos mixtos invalidos")
    discount = _commercial_discount_percent(commercial_discount_percent)
    effective_today = today or date.today()
    rows_by_catalog = _group_browser_rows(raw_items)
    normalized_groups: list[dict[str, Any]] = []
    for catalog in MIXED_CATALOG_ORDER:
        rows = rows_by_catalog.get(catalog)
        if not rows:
            continue
        source_catalog = catalogs.get(catalog)
        if not isinstance(source_catalog, dict):
            raise ValueError(f"Catalogo mixto invalido: {catalog}")
        if catalog == "tarkett":
            source = build_tarkett_cart_payload([{key: value for key, value in row.items() if key != "catalog"} for row in rows], catalog=source_catalog)
            rate = _rate_payload(resolve_conversion_rate("MXN", quote_currency, rate_rows, effective_today))
            items = [_common_line(line, catalog=catalog, rate=rate, discount=discount, source_hash=source["catalog_source_hash"]) for line in source["items"]]
        elif catalog == "offiho":
            source = build_offiho_cart_payload([{key: value for key, value in row.items() if key != "catalog"} for row in rows], catalog=source_catalog)
            rate = _rate_payload(resolve_conversion_rate("MXN", quote_currency, rate_rows, effective_today))
            items = [_common_line(line, catalog=catalog, rate=rate, discount=discount, source_hash=source["catalog_source_hash"]) for line in source["items"]]
        else:
            builder_rows = []
            for row in rows:
                builder_row = {key: value for key, value in row.items() if key != "catalog"}
                if not builder_row.get("base_option_id"):
                    builder_row.pop("base_option_id", None)
                builder_rows.append(builder_row)
            try:
                source = build_supplier_cart_payload(builder_rows, source_catalog, quote_currency, rate_rows, today=effective_today)
            except ValueError as exc:
                if "moneda base" in str(exc).casefold():
                    raise ValueError(f"Moneda base mixta invalida: {catalog}") from exc
                raise
            items = [_supplier_line(line, row, catalog=catalog, payload=source) for line, row in zip(source["items"], rows, strict=True)]
            rate = {field: source[field] for field in AUTO_ELECTRIFICATION_RATE_FIELDS}
        if rate["base_currency"] != MIXED_EXPECTED_BASE_CURRENCY[catalog]:
            raise ValueError(f"Moneda base mixta invalida: {catalog}")
        if any(line["original_currency"] != MIXED_EXPECTED_BASE_CURRENCY[catalog] for line in items):
            raise ValueError(f"Moneda original mixta invalida: {catalog}")
        if any(Decimal(line["tax_rate"]) != Decimal("0.160000") for line in items):
            bad = next(line["canonical_key"] for line in items if Decimal(line["tax_rate"]) != Decimal("0.160000"))
            raise ValueError(f"{catalog}:{bad}: IVA 16 requerido")
        normalized_groups.append({"catalog": catalog, "catalog_source_hash": source["catalog_source_hash"], **rate, "items": items})
    eligible = [{field: group[field] for field in AUTO_ELECTRIFICATION_RATE_FIELDS} for group in normalized_groups if any(line["auto_electrification"] for line in group["items"])]
    if eligible and any(snapshot != eligible[0] for snapshot in eligible[1:]):
        raise ValueError("Tasa de electrificacion mixta inconsistente")
    payload = {
        "source_type": MIXED_CATALOG_CART_SOURCE_TYPE, "quote_currency": quote_currency,
        "created_at": datetime.now(timezone.utc).isoformat(), "groups": normalized_groups,
        "item_count": sum(len(group["items"]) for group in normalized_groups),
        "auto_electrification_rate": eligible[0] if eligible else None,
        "rate_summary": [{key: group[key] for key in ("catalog", *AUTO_ELECTRIFICATION_RATE_FIELDS)} for group in normalized_groups],
    }
    return validate_mixed_catalog_payload(payload)


def _decimal_text(value: object, field: str, *, nonnegative: bool = False, positive: bool = False, places: int | None = None, maximum: Decimal | None = None) -> Decimal:
    if not isinstance(value, str) or not value or len(value) > 64:
        raise ValueError(f"{field} mixto invalido")
    try:
        number = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError(f"{field} mixto invalido") from exc
    if not number.is_finite() or (nonnegative and number < 0) or (positive and number <= 0) or (maximum is not None and number > maximum):
        raise ValueError(f"{field} mixto invalido")
    exponent = -number.as_tuple().exponent
    if places is not None and (exponent != places):
        raise ValueError(f"{field} mixto invalido")
    if places is None and exponent > 6:
        raise ValueError(f"{field} mixto invalido")
    return number


def _bounded(value: object, field: str, *, required: bool = False, limit: int = MAX_MIXED_TEXT) -> str:
    if not isinstance(value, str) or len(value) > limit or any(unicodedata.category(char) in {"Cc", "Cs"} for char in value):
        raise ValueError(f"{field} mixto invalido")
    if required and not value.strip():
        raise ValueError(f"{field} mixto invalido")
    return value


def _iso_timestamp(value: object, field: str, *, allow_empty: bool = False) -> str:
    if allow_empty and value == "":
        return ""
    text = _bounded(value, field, required=True)
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError(f"{field} mixto invalido") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise ValueError(f"{field} mixto invalido")
    return text


def _https_url(value: object, field: str) -> str:
    text = _bounded(value, field, limit=MAX_MIXED_URL)
    if not text:
        return ""
    try:
        parsed = urlsplit(text)
        port = parsed.port
    except ValueError as exc:
        raise ValueError(f"{field} mixto invalido") from exc
    if parsed.scheme.lower() != "https" or not parsed.hostname or parsed.username is not None or parsed.password is not None or port not in (None, 443):
        raise ValueError(f"{field} mixto invalido")
    return text


def _require_warning(warnings: list[str], expected: str, field: str) -> None:
    if sum(_warning_key(value) == _warning_key(expected) for value in warnings) != 1:
        raise ValueError(f"{field} mixto invalido")


def _validate_mixed_attribute_controls(value: object) -> None:
    if isinstance(value, str):
        if EXCEL_ILLEGAL_CONTROL_RE.search(value):
            raise ValueError("attributes mixtos contiene caracteres de control ilegales")
        return
    if isinstance(value, dict):
        for key, nested in value.items():
            _validate_mixed_attribute_controls(key)
            _validate_mixed_attribute_controls(nested)
        return
    if isinstance(value, list):
        for nested in value:
            _validate_mixed_attribute_controls(nested)


def _validate_mixed_excel_text_cells(
    line: dict[str, Any],
    *,
    catalog: str,
    source_hash: str,
    index: int,
) -> None:
    extra_description_parts = (
        f"Fuente: {line['source_reference']}",
        f"Hash fuente: {source_hash}",
    )
    description, _, _ = catalog_quotation_item_text(
        line,
        index=index,
        source_type=MIXED_GROUP_SOURCE_TYPES[catalog],
        extra_description_parts=extra_description_parts,
    )
    attributes = line["attributes"]
    dimensions = str(attributes.get("dimensions") or "").strip()
    dimension_cell = (
        dimensions
        if MIXED_GROUP_SOURCE_TYPES[catalog] == "supplier_cart" and dimensions
        else line["unit"]
    )
    text_cells = (
        line["name"],
        description,
        dimension_cell,
        line["product_url"],
        line["supplier"],
        line["original_currency"],
        line["source_reference"],
        line["price_mode"],
    )
    if any(
        len(safe_excel_text(value)) > MAX_EXCEL_CELL_TEXT_LENGTH
        for value in text_cells
    ):
        raise ValueError("Texto mixto excede limite de Excel de 32767 caracteres")


def _validate_reservation(line: dict[str, Any], catalog: str) -> None:
    reservation = line["reservation"]
    stocked = line["availability_type"] == "stocked"
    optional = MIXED_RESERVATION_RESULT_FIELDS & set(line)
    if reservation is None:
        if stocked or optional:
            raise ValueError("Reserva mixta invalida")
        return
    if not stocked or not isinstance(reservation, dict) or set(reservation) != {"identity", "sku", "quantity", "stock"}:
        raise ValueError("Reserva mixta invalida")
    identity = _identity_text(reservation["identity"], "identity")
    _identity_text(
        reservation["sku"], "sku",
        allow_empty=line["code_status"] == "needs_review" and catalog in {"sonara", "lumbro"},
    )
    if reservation["quantity"] != line["quantity"] or reservation["stock"] != line["stock"]:
        raise ValueError("Reserva mixta invalida")
    if catalog == "tarkett" and identity != line["code"]:
        raise ValueError("Reserva mixta invalida")
    if catalog == "offiho" and identity != line["canonical_key"].split(":", 1)[1]:
        raise ValueError("Reserva mixta invalida")
    if catalog not in {"tarkett", "offiho"}:
        try:
            identity_tuple = json.loads(line["canonical_key"].split(":", 1)[1])
        except (IndexError, json.JSONDecodeError) as exc:
            raise ValueError("Reserva mixta invalida") from exc
        if not isinstance(identity_tuple, list) or not identity_tuple or identity != identity_tuple[0]:
            raise ValueError("Reserva mixta invalida")
    if optional and optional != MIXED_RESERVATION_RESULT_FIELDS:
        raise ValueError("Reserva mixta invalida")
    if optional:
        reserved = _decimal_text(line["reserved_quantity"], "Reserva", nonnegative=True, places=6)
        after = _decimal_text(line["available_after_reservations"], "Reserva", nonnegative=True, places=6)
        if type(line["reserved_by_others"]) is not bool or after != max(Decimal(reservation["stock"]) - reserved, Decimal(0)):
            raise ValueError("Reserva mixta invalida")
        insufficient = Decimal(line["quantity"]) > after
        if catalog == "tarkett" and insufficient:
            raise ValueError("Reserva mixta invalida")
        if catalog != "tarkett" and insufficient:
            _require_warning(line["warnings"], "Existencia insuficiente; verificar disponibilidad.", "Reserva")


def _validate_mixed_catalog_payload(payload: object) -> dict:
    if not isinstance(payload, dict) or set(payload) != {"source_type", "quote_currency", "created_at", "groups", "item_count", "auto_electrification_rate", "rate_summary"}:
        raise ValueError("Grupos mixtos invalidos")
    try:
        encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), allow_nan=False).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise ValueError("Grupos mixtos invalidos") from exc
    if len(encoded) > MAX_MIXED_PAYLOAD_BYTES or payload["source_type"] != MIXED_CATALOG_CART_SOURCE_TYPE or not isinstance(payload["quote_currency"], str) or payload["quote_currency"] not in MIXED_QUOTE_CURRENCIES:
        raise ValueError("Grupos mixtos invalidos")
    _iso_timestamp(payload["created_at"], "created_at")
    groups = payload["groups"]
    if not isinstance(groups, list) or not 1 <= len(groups) <= len(MIXED_CATALOG_ORDER):
        raise ValueError("Grupos mixtos invalidos")
    seen_keys: set[str] = set()
    total = 0
    seen_catalogs: list[str] = []
    eligible: list[dict[str, str]] = []
    product_index = 1
    for group in groups:
        if not isinstance(group, dict) or set(group) != MIXED_GROUP_FIELDS:
            raise ValueError("Grupos mixtos invalidos")
        catalog = group.get("catalog")
        if catalog not in MIXED_CATALOG_ORDER or catalog in seen_catalogs:
            raise ValueError("Grupos mixtos invalidos")
        seen_catalogs.append(catalog)
        if seen_catalogs != sorted(seen_catalogs, key=MIXED_CATALOG_ORDER.index):
            raise ValueError("Grupos mixtos invalidos")
        source_hash = group.get("catalog_source_hash")
        if not isinstance(source_hash, str) or len(source_hash) != 64 or any(char not in "0123456789abcdef" for char in source_hash):
            raise ValueError("Grupos mixtos invalidos")
        if group["base_currency"] != MIXED_EXPECTED_BASE_CURRENCY[catalog] or group["quote_currency"] != payload["quote_currency"]:
            raise ValueError("Grupos mixtos invalidos")
        rate = _decimal_text(group["exchange_rate"], "Tasa", positive=True, places=6)
        _bounded(group["rate_source"], "Fuente de tasa", required=True)
        try:
            date.fromisoformat(_bounded(group["rate_effective_date"], "Fecha de tasa", required=True))
        except ValueError as exc:
            raise ValueError("Grupos mixtos invalidos") from exc
        if group["rate_source"] == "identity":
            if group["base_currency"] != group["quote_currency"] or rate != Decimal("1.000000") or group["rate_retrieved_at"] != "":
                raise ValueError("Grupos mixtos invalidos")
        else:
            _iso_timestamp(group["rate_retrieved_at"], "rate_retrieved_at")
        items = group["items"]
        if not isinstance(items, list) or not items:
            raise ValueError("Grupos mixtos invalidos")
        total += len(items)
        if total > MAX_MIXED_CATALOG_LINES:
            raise ValueError("Grupos mixtos invalidos")
        for line in items:
            if not isinstance(line, dict) or set(line) not in {MIXED_LINE_FIELDS, MIXED_LINE_FIELDS | MIXED_RESERVATION_RESULT_FIELDS}:
                raise ValueError("Grupos mixtos invalidos")
            if line["catalog"] != catalog or line["supplier"] != MIXED_CATALOG_LABELS[catalog]:
                raise ValueError("Grupos mixtos invalidos")
            key = _identity_text(line["canonical_key"], "Clave")
            if key in seen_keys or not key.startswith(f"{catalog}:"):
                raise ValueError("Grupos mixtos invalidos")
            seen_keys.add(key)
            for field, required in (("name", True), ("description", False), ("unit", True), ("source_reference", True), ("configuration", False), ("variant", False), ("lead_time", False), ("price_source", True)):
                _bounded(line[field], field, required=required)
            if line["code_status"] not in {"verified", "needs_review"} or (not line["code"] and line["code_status"] != "needs_review"):
                raise ValueError("Grupos mixtos invalidos")
            if line["code_status"] == "needs_review" and catalog not in {"sonara", "lumbro"}:
                raise ValueError("Grupos mixtos invalidos")
            if catalog in {"tarkett", "offiho"} and line["code_status"] != "verified":
                raise ValueError("Grupos mixtos invalidos")
            _bounded(line["code"], "code")
            if not isinstance(line["warnings"], list) or len(line["warnings"]) > MAX_MIXED_WARNINGS:
                raise ValueError("Grupos mixtos invalidos")
            for warning in line["warnings"]:
                _bounded(warning, "warning", required=True)
            _https_url(line["image_url"], "image_url")
            _https_url(line["product_url"], "product_url")
            if not isinstance(line["attributes"], dict):
                raise ValueError("Grupos mixtos invalidos")
            _validate_mixed_attribute_controls(line["attributes"])
            _validate_attributes(line["attributes"])
            if line["availability_type"] not in {"stocked", "made_to_order", "unknown"} or line["image_kind"] not in {"official", "generated_reference", "placeholder"} or line["stock_status"] not in {"", "available", "out_of_stock", "insufficient_stock"}:
                raise ValueError("Grupos mixtos invalidos")
            if line["availability_type"] == "stocked":
                available = _decimal_text(line["available_quantity"], "Stock", nonnegative=True, places=6)
                if line["stock"] != line["available_quantity"]:
                    raise ValueError("Grupos mixtos invalidos")
                expected_status = "out_of_stock" if available <= 0 else "insufficient_stock" if Decimal(line["quantity"]) > available else "available"
                if line["stock_status"] != expected_status:
                    raise ValueError("Grupos mixtos invalidos")
            else:
                if line["available_quantity"] is not None or line["stock"] is not None or line["stock_status"] != "":
                    raise ValueError("Grupos mixtos invalidos")
                available = None
            if catalog in {"tarkett", "offiho"} and line["availability_type"] != "stocked":
                raise ValueError("Grupos mixtos invalidos")
            if catalog == "tarkett" and line["stock_status"] != "available":
                raise ValueError("Grupos mixtos invalidos")
            if line["image_kind"] == "generated_reference":
                _require_warning(line["warnings"], "Imagen de referencia", "warning")
            if line["price_source"] == "missing":
                _require_warning(line["warnings"], "Precio por confirmar", "warning")
            if line["code_status"] == "needs_review":
                _require_warning(line["warnings"], "Codigo por verificar", "warning")
            quantity = _decimal_text(line["quantity"], "Cantidad", positive=True, maximum=Decimal("1000000"))
            original = _decimal_text(line["original_unit_price"], "Precio original", nonnegative=True, places=6)
            unit = _decimal_text(line["unit_price"], "Precio", nonnegative=True, places=2)
            frozen = _decimal_text(line["frozen_exchange_rate"], "Tasa congelada", positive=True, places=6)
            if line["original_currency"] != group["base_currency"] or frozen != rate or _decimal_text(line["tax_rate"], "IVA", nonnegative=True, places=6) != Decimal("0.160000") or unit != (original * frozen).quantize(TWO_PLACES, rounding=ROUND_HALF_UP):
                raise ValueError("Grupos mixtos invalidos")
            discount = _decimal_text(line["discount_percent"], "Descuento", nonnegative=True, places=6)
            is_legacy = catalog in {"tarkett", "offiho"}
            if is_legacy:
                if line["price_mode"] != "list" or not 0 <= discount <= 100 or type(line["auto_electrification"]) is not bool or line["auto_electrification"] is not True:
                    raise ValueError("Grupos mixtos invalidos")
            elif line["price_mode"] != "net" or discount != 0 or type(line["auto_electrification"]) is not bool or line["auto_electrification"] is not False:
                raise ValueError("Grupos mixtos invalidos")
            _validate_reservation(line, catalog)
            _validate_mixed_excel_text_cells(
                line,
                catalog=catalog,
                source_hash=source_hash,
                index=product_index,
            )
            product_index += 1
            if line["auto_electrification"]:
                eligible.append({field: group[field] for field in AUTO_ELECTRIFICATION_RATE_FIELDS})
    if type(payload["item_count"]) is not int or payload["item_count"] != total:
        raise ValueError("Conteo mixto inconsistente")
    expected_summary = [{key: group[key] for key in ("catalog", *AUTO_ELECTRIFICATION_RATE_FIELDS)} for group in groups]
    if payload["rate_summary"] != expected_summary:
        raise ValueError("Resumen de tasas mixtas inconsistente")
    automatic = payload["auto_electrification_rate"]
    if eligible:
        if not isinstance(automatic, dict) or set(automatic) != set(AUTO_ELECTRIFICATION_RATE_FIELDS) or any(automatic != snapshot for snapshot in eligible):
            raise ValueError("Tasa de electrificacion mixta invalida")
    elif automatic is not None:
        raise ValueError("Tasa de electrificacion mixta invalida")
    return payload


def validate_mixed_catalog_payload(payload: object) -> dict:
    try:
        return _validate_mixed_catalog_payload(payload)
    except ValueError as exc:
        message = str(exc)
        if message in {
            "Grupos mixtos invalidos", "Conteo mixto inconsistente",
            "Resumen de tasas mixtas inconsistente", "Tasa de electrificacion mixta invalida",
        } or message.startswith("Tasa de electrificacion mixta"):
            raise
        raise ValueError(f"Grupos mixtos invalidos: {message}") from exc


def build_mixed_reservation_groups(payload: dict) -> list[dict]:
    groups: list[dict] = []
    for group in payload["groups"]:
        aggregated: dict[str, dict] = {}
        for line in group["items"]:
            reservation = deepcopy(line["reservation"])
            if reservation is None:
                continue
            identity = reservation["identity"]
            existing = aggregated.get(identity)
            if existing is None:
                aggregated[identity] = reservation
                continue
            if existing["sku"] != reservation["sku"] or existing["stock"] != reservation["stock"]:
                raise ValueError(f"Reserva mixta incompatible: {group['catalog']}:{identity}")
            existing["quantity"] = f"{Decimal(existing['quantity']) + Decimal(reservation['quantity']):.6f}"
        if aggregated:
            groups.append({"catalog": group["catalog"], "items": [aggregated[key] for key in sorted(aggregated)]})
    return groups


def create_mixed_catalog_quotation_workbook(
    payload: dict[str, Any],
    output_path: str | Path,
    *,
    image_dir: str | Path | None = None,
) -> Path:
    payload = validate_mixed_catalog_payload(payload)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    tmp_context = None
    if image_dir is None:
        tmp_context = tempfile.TemporaryDirectory(prefix="mixed_catalog_images_")
        images_root = Path(tmp_context.name)
    else:
        images_root = Path(image_dir)
        images_root.mkdir(parents=True, exist_ok=True)

    wb = None
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        write_catalog_quotation_headers(
            ws,
            {
                12: safe_excel_text("Supplier"),
                13: safe_excel_text("Discount Percent"),
                14: safe_excel_text("Original Currency"),
                15: safe_excel_text("Original Unit Price"),
                16: safe_excel_text("Frozen Exchange Rate"),
                17: safe_excel_text("Source Reference"),
                18: safe_excel_text("Price Mode"),
                19: safe_excel_text("Auto Electrification"),
            },
        )

        row = 8
        product_index = 1
        for group in payload["groups"]:
            ws.cell(row, 1).value = "- " + safe_excel_text(
                MIXED_CATALOG_LABELS[group["catalog"]]
            )
            ws.cell(row, 1).font = Font(bold=True)
            row += 1
            for item in group["items"]:
                write_catalog_quotation_item(
                    ws,
                    row=row,
                    index=product_index,
                    item=item,
                    source_type=MIXED_GROUP_SOURCE_TYPES[group["catalog"]],
                    images_root=images_root,
                    text_transform=safe_excel_text,
                    image_file_key=(
                        f"{group['catalog']}-{row}-"
                        f"{hashlib.sha256(item['canonical_key'].encode('utf-8')).hexdigest()[:16]}"
                    ),
                    extra_description_parts=(
                        f"Fuente: {item['source_reference']}",
                        f"Hash fuente: {group['catalog_source_hash']}",
                    ),
                )
                ws.cell(row, 12).value = safe_excel_text(item["supplier"])
                ws.cell(row, 13).value = float(Decimal(item["discount_percent"]))
                ws.cell(row, 13).number_format = "0.000000"
                ws.cell(row, 14).value = safe_excel_text(item["original_currency"])
                ws.cell(row, 15).value = float(Decimal(item["original_unit_price"]))
                ws.cell(row, 16).value = float(Decimal(item["frozen_exchange_rate"]))
                ws.cell(row, 17).value = safe_excel_text(item["source_reference"])
                ws.cell(row, 18).value = safe_excel_text(item["price_mode"])
                if not isinstance(item["auto_electrification"], bool):
                    raise ValueError("Auto Electrification mixto debe ser booleano")
                ws.cell(row, 19).value = item["auto_electrification"]
                row += 1
                product_index += 1

        _set_column_widths(ws)
        for column, width in {
            "L": 20,
            "M": 20,
            "N": 20,
            "O": 22,
            "P": 22,
            "Q": 42,
            "R": 18,
            "S": 22,
        }.items():
            ws.column_dimensions[column].width = width
        wb.save(output)
    finally:
        try:
            if wb is not None:
                wb.close()
        finally:
            if tmp_context is not None:
                tmp_context.cleanup()
    return output
