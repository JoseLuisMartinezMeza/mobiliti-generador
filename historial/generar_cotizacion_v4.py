#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti v4 - DISEÑO PIXEL-PERFECT
=============================================================
Enfoque absoluto en el diseño visual. El output debe verse idéntico
al ejemplo objetivo: Cotízación-IZA-Monterrey-BH-Cowork.xlsx

Flujo:
1. Lee Quotation del proveedor (items + imágenes)
2. Copia template como base (preserva formato visual)
3. Copia hoja Quotation completa del source
4. Genera Mobiliti con formato exacto (categorías marrón, productos amarillo)
5. Genera Cotizacion con formato exacto (categorías azul, productos blanco, imágenes)
6. Preserva términos y condiciones

Uso:
    python generar_cotizacion_v4.py \
        --source "Quotation.xlsx" \
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \
        --output "Cotizacion.xlsx" \
        --cotizacion "100-99999" \
        --proyecto "Proyecto" \
        --cliente "Cliente"
"""

import argparse
import copy
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


Q_HEADER_ROW = 7

# ─── FORMATOS VISUALES EXTRAÍDOS DEL EJEMPLO OBJETIVO ───

# Cotizacion - Categoría (ej. fila 16 del ejemplo)
COT_CAT_FONT = Font(name='Roboto', size=16, bold=True, color=None)
COT_CAT_FILL = PatternFill(start_color='FF73A9DB', end_color='FF73A9DB', fill_type='solid')

# Cotizacion - Producto (ej. fila 17 del ejemplo)
COT_PROD_FONT_BOLD = Font(name='Roboto', size=16, bold=True, color=None)
COT_PROD_FONT_NORMAL = Font(name='Roboto', size=16, bold=False, color=None)
COT_PROD_FILL = PatternFill(fill_type=None)

# Cotizacion - Headers de tabla (fila 15)
COT_HDR_FONT = Font(name='Roboto', size=14, bold=True, color='FF787A79')
COT_HDR_ALIGN = Alignment(horizontal='center')

# Mobiliti - Categoría (ej. fila 17 del ejemplo Mobiliti)
MOB_CAT_FONT = Font(name='Calibri', size=20, bold=True, color=None)
MOB_CAT_FILL = PatternFill(start_color='FF3E2500', end_color='FF3E2500', fill_type='solid')

# Mobiliti - Producto (ej. fila 18 del ejemplo Mobiliti)
MOB_PROD_FONT_BOLD = Font(name='Century Gothic', size=11, bold=True, color=None)
MOB_PROD_FONT_NORMAL = Font(name='Century Gothic', size=11, bold=False, color=None)
MOB_PROD_FILL = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')

# Mobiliti - Subtotales
MOB_SUB_FONT = Font(name='Calibri', size=11, bold=True, color=None)

# Bordes comunes
THIN_BOTTOM = Border(bottom=Side(style='thin', color='FF000000'))
THIN_BORDER = Border(
    bottom=Side(style='thin', color='FF000000'),
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
)


def limpiar_excel():
    """Cierra procesos Excel colgados."""
    try:
        subprocess.run(['taskkill', '//F', '//IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes_source(source_path):
    """Extrae imágenes del source y crea mapeo row->imagen_path."""
    temp_dir = tempfile.mkdtemp(prefix="excel_images_")
    image_map = {}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        media_files = [f for f in z.namelist() if 'media' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
    
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    for img in ws._images:
        path = img.path
        filename = os.path.basename(path)
        filepath = os.path.join(temp_dir, filename)
        
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1
        else:
            row = img.anchor.row + 1
        
        if os.path.exists(filepath):
            image_map[row] = filepath
    
    wb.close()
    return image_map, temp_dir


def leer_datos_quotation(source_path):
    """Lee items de Quotation del proveedor."""
    print("[1/8] Leyendo datos del Quotation del proveedor...")
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    items = []
    last_row = get_last_used_row(ws)
    for row in range(Q_HEADER_ROW + 1, last_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        if no_val is None and item_name is None:
            continue
        
        # Categoría: texto que empieza con "-" o no tiene número pero tiene texto
        if isinstance(no_val, str) and no_val.startswith('-'):
            items.append({
                'tipo': 'categoria',
                'nombre': no_val.strip('- ').strip(),
                'row': row
            })
        elif isinstance(no_val, (int, float)) and item_name:
            desc = ws.cell(row=row, column=4).value
            if desc and isinstance(desc, str):
                lines = [l.strip() for l in desc.split('\n') if l.strip()]
                desc = lines[0][:200] if lines else desc[:200]
            
            items.append({
                'tipo': 'producto',
                'no': int(no_val),
                'item_name': item_name,
                'description': desc,
                'dimension': ws.cell(row=row, column=5).value,
                'color': ws.cell(row=row, column=6).value,
                'qty': ws.cell(row=row, column=7).value,
                'unit_price': ws.cell(row=row, column=10).value,
                'row': row
            })
        elif item_name and isinstance(item_name, str) and not no_val:
            items.append({
                'tipo': 'categoria',
                'nombre': item_name.strip(),
                'row': row
            })
    
    print(f"      [OK] {len(items)} items leídos ({sum(1 for i in items if i['tipo']=='categoria')} categorías, {sum(1 for i in items if i['tipo']=='producto')} productos)")
    wb.close()
    return items


def get_last_used_row(ws):
    """Encuentra la última fila con datos reales (evita max_row falso de 65536)."""
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=1).value is not None:
            return row
    return 1


def copiar_hoja_completa(wb_dest, wb_src, sheet_name, new_name=None):
    """Copia una hoja completa incluyendo valores, estilos, merged cells e imágenes."""
    if new_name is None:
        new_name = sheet_name
    
    if new_name in wb_dest.sheetnames:
        del wb_dest[new_name]
    
    ws_src = wb_src[sheet_name]
    ws_dest = wb_dest.create_sheet(new_name)
    
    # Copiar dimensiones
    for col_letter, col_dim in ws_src.column_dimensions.items():
        if col_dim.width:
            ws_dest.column_dimensions[col_letter].width = col_dim.width
    
    for row_num, row_dim in ws_src.row_dimensions.items():
        if row_dim.height:
            ws_dest.row_dimensions[row_num].height = row_dim.height
    
    # Copiar celdas con valores y estilos (limitado a filas reales)
    last_row = get_last_used_row(ws_src)
    for row in ws_src.iter_rows(min_row=1, max_row=last_row):
        for cell_src in row:
            cell_dest = ws_dest.cell(row=cell_src.row, column=cell_src.column)
            cell_dest.value = cell_src.value
            if cell_src.has_style:
                cell_dest.font = copy.copy(cell_src.font)
                cell_dest.border = copy.copy(cell_src.border)
                cell_dest.fill = copy.copy(cell_src.fill)
                cell_dest.number_format = copy.copy(cell_src.number_format)
                cell_dest.protection = copy.copy(cell_src.protection)
                cell_dest.alignment = copy.copy(cell_src.alignment)
    
    # Copiar merged cells (solo las que están dentro del rango)
    for merged_range in ws_src.merged_cells.ranges:
        if merged_range.max_row <= last_row:
            ws_dest.merge_cells(str(merged_range))
    
    # Copiar imágenes
    for img in ws_src._images:
        ws_dest._images.append(img)
    
    return ws_dest


def aplicar_formato_categoria_cotizacion(ws, row):
    """Aplica formato de categoría a toda la fila en Cotizacion."""
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.font = COT_CAT_FONT
        cell.fill = COT_CAT_FILL


def aplicar_formato_producto_cotizacion(ws, row):
    """Aplica formato de producto a toda la fila en Cotizacion."""
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        # Columnas B y C usan font normal, el resto bold
        if col in (2, 3):
            cell.font = COT_PROD_FONT_NORMAL
        else:
            cell.font = COT_PROD_FONT_BOLD
        cell.fill = COT_PROD_FILL


def aplicar_formato_categoria_mobiliti(ws, row):
    """Aplica formato de categoría a toda la fila en Mobiliti."""
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.font = MOB_CAT_FONT
        cell.fill = MOB_CAT_FILL


def aplicar_formato_producto_mobiliti(ws, row):
    """Aplica formato de producto a toda la fila en Mobiliti."""
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        # A, B, D, E = bold; C, F, G = normal
        if col in (1, 2, 4, 5):
            cell.font = MOB_PROD_FONT_BOLD
        else:
            cell.font = MOB_PROD_FONT_NORMAL
        cell.fill = MOB_PROD_FILL


def aplicar_formato_subtotal_mobiliti(ws, row):
    """Aplica formato de subtotales en Mobiliti."""
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.font = MOB_SUB_FONT


def unmerge_cells_in_rows(ws, start_row, end_row):
    """Des-mergea celdas que intersectan el rango de filas especificado."""
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        # Des-mergear si el rango intersecta con [start_row, end_row]
        if not (merged_range.max_row < start_row or merged_range.min_row > end_row):
            to_unmerge.append(str(merged_range))
    for merge in to_unmerge:
        try:
            ws.unmerge_cells(merge)
        except (KeyError, ValueError):
            pass


def llenar_encabezado_cotizacion(ws, args):
    """Llena datos del cliente en Cotizacion (filas 3-12)."""
    print("[3/8] Llenando encabezado de Cotizacion...")
    
    # Des-mergear celdas del encabezado para poder escribir
    unmerge_cells_in_rows(ws, 3, 12)
    
    campos = {
        3: ('COTIZACION #  ', args.cotizacion, 2),
        4: (None, '=TODAY()', 2),
        7: ('PROYECTO:   ', args.proyecto, 2),
        8: ('Nombre:', args.cliente, 2),
        9: ('Correo:', args.correo, 2),
        10: ('Telefono:', args.telefono, 2),
        11: ('Direccion:', args.direccion, 2),
        12: ('Razon Social:', args.razon_social, 2),
    }
    
    for row, (label, valor, col) in campos.items():
        if valor:
            ws.cell(row=row, column=col, value=valor)
            if label:
                ws.cell(row=row, column=1, value=label)
    
    print(f"      [OK] Cliente: {args.cliente}")


def generar_mobiliti(ws, items):
    """Genera hoja Mobiliti con formato exacto del ejemplo."""
    print("[5/8] Generando Mobiliti con formato visual...")
    
    # Des-mergear celdas del área de datos
    unmerge_cells_in_rows(ws, 17, 200)
    
    # Limpiar área de datos (filas 17 en adelante)
    for row in range(17, 200):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
    
    current_row = 17
    section_num = 1
    section_start_row = current_row
    mobiliti_row_map = {}  # Mapeo: q_row -> mobiliti_row (para Cotizacion)
    
    for item in items:
        if item['tipo'] == 'categoria':
            # Subtotales de sección anterior
            if current_row > section_start_row:
                ws.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
                ws.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
                aplicar_formato_subtotal_mobiliti(ws, current_row)
                current_row += 1
                section_num += 1
            
            # Fila de categoría
            ws.cell(row=current_row, column=1, value=f"=Quotation!A{item['row']}")
            aplicar_formato_categoria_mobiliti(ws, current_row)
            current_row += 1
            section_start_row = current_row
        else:
            # Fila de producto
            q_row = item['row']
            ws.cell(row=current_row, column=1, value=f"=Quotation!B{q_row}")
            ws.cell(row=current_row, column=2, value=f"=Quotation!D{q_row}")
            ws.cell(row=current_row, column=3, value="Sunon Inc")
            ws.cell(row=current_row, column=4, value=f'=IFERROR(VLOOKUP(C{current_row},Proveedores!A:B,2,0)," ")')
            ws.cell(row=current_row, column=5, value=f"=Quotation!G{q_row}")
            ws.cell(row=current_row, column=6, value=f"=Quotation!J{q_row}")
            ws.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}")
            aplicar_formato_producto_mobiliti(ws, current_row)
            
            mobiliti_row_map[q_row] = current_row
            current_row += 1
    
    # Subtotales de última sección
    if current_row > section_start_row:
        ws.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
        ws.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
        aplicar_formato_subtotal_mobiliti(ws, current_row)
    
    print(f"      [OK] {section_num} secciones, {len(mobiliti_row_map)} productos mapeados")
    return mobiliti_row_map


def generar_cotizacion(ws, items, image_map, mobiliti_row_map):
    """Genera hoja Cotizacion con formato exacto del ejemplo."""
    print("[6/8] Generando Cotizacion con formato visual...")
    
    # Limpiar área de datos (filas 16 en adelante, hasta antes de términos)
    # Primero, encontrar dónde empiezan los términos en el template
    terminos_start_row = None
    for row in range(16, 100):
        val = ws.cell(row=row, column=1).value
        if val and isinstance(val, str) and 'CONDICIONES' in val:
            terminos_start_row = row
            break
    
    if not terminos_start_row:
        terminos_start_row = 32  # Default del template
    
    # Des-mergear TODAS las celdas del área de datos (productos NO deben estar mergeados)
    unmerge_cells_in_rows(ws, 16, terminos_start_row + 20)
    
    # Limpiar filas de ejemplo (16 hasta terminos_start_row - 1)
    for row in range(16, terminos_start_row):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
    
    # Calcular cuántas filas necesitamos
    num_items = len(items)
    filas_ejemplo = terminos_start_row - 16  # Filas disponibles en template
    
    # Insertar filas adicionales si necesitamos más espacio
    if num_items > filas_ejemplo:
        filas_extra = num_items - filas_ejemplo
        print(f"      Insertando {filas_extra} filas adicionales...")
        ws.insert_rows(terminos_start_row, filas_extra)
        # Recalcular dónde están los términos después de insertar
        terminos_start_row += filas_extra
        # Re-detectar términos por si acaso
        for row in range(terminos_start_row, terminos_start_row + 10):
            val = ws.cell(row=row, column=1).value
            if val and isinstance(val, str) and 'CONDICIONES' in val:
                terminos_start_row = row
                break
    
    # Escribir items desde fila 16
    current_row = 16
    first_data_row = None
    last_data_row = None
    
    # Establecer descuento base en fila 19 (o la primera fila de producto + 2)
    descuento_row = None
    
    for item in items:
        if item['tipo'] == 'categoria':
            ws.cell(row=current_row, column=1, value=f"=Quotation!A{item['row']}")
            aplicar_formato_categoria_cotizacion(ws, current_row)
            # Mergear toda la fila de categoría (como en el ejemplo)
            try:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            except Exception:
                pass
            current_row += 1
        else:
            if first_data_row is None:
                first_data_row = current_row
                descuento_row = current_row + 2  # Fila para el descuento base
            
            q_row = item['row']
            mob_row = mobiliti_row_map.get(q_row)
            
            # Código
            ws.cell(row=current_row, column=1, value=f"=Quotation!B{q_row}")
            # Imagen (vacía, se inserta después)
            ws.cell(row=current_row, column=2, value=None)
            # Descripción
            ws.cell(row=current_row, column=3, value=f"=Quotation!D{q_row}")
            # Medidas
            ws.cell(row=current_row, column=4, value=f"=Quotation!E{q_row}")
            # Cantidad
            ws.cell(row=current_row, column=5, value=f"=Quotation!G{q_row}")
            # P. UNIT - referencia a Mobiliti si existe, sino Quotation
            if mob_row:
                ws.cell(row=current_row, column=6, value=f"=Mobiliti!F{mob_row}")
            else:
                ws.cell(row=current_row, column=6, value=f"=Quotation!J{q_row}")
            # % DESC
            ws.cell(row=current_row, column=7, value=f"=G${descuento_row}")
            # DESCUENTO
            ws.cell(row=current_row, column=8, value=f"=F{current_row}*G{current_row}")
            # SUBTOTAL
            ws.cell(row=current_row, column=9, value=f"=F{current_row}-H{current_row}")
            # TOTAL
            ws.cell(row=current_row, column=10, value=f"=I{current_row}*E{current_row}")
            
            aplicar_formato_producto_cotizacion(ws, current_row)
            current_row += 1
    
    last_data_row = current_row - 1
    
    # Establecer descuento base
    if descuento_row and descuento_row <= last_data_row:
        ws.cell(row=descuento_row, column=7, value=0.7)
    
    # LIMPIEZA FINAL: des-mergear TODO el área de datos y re-mergear solo categorías
    # Primero guardar las filas de categoría (reiniciar current_row)
    cat_rows = set()
    check_row = 16
    for item in items:
        if item['tipo'] == 'categoria':
            cat_rows.add(check_row)
            check_row += 1
        else:
            check_row += 1
    
    # Des-mergear TODO en el rango de datos
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 16 and merged_range.max_row <= terminos_start_row - 1:
            try:
                ws.unmerge_cells(str(merged_range))
            except (KeyError, ValueError):
                pass
    
    # Re-mergear solo las filas de categoría
    for cat_row in sorted(cat_rows):
        try:
            ws.merge_cells(start_row=cat_row, start_column=1, end_row=cat_row, end_column=10)
        except (KeyError, ValueError):
            pass
    
    print(f"      [OK] Items en filas {first_data_row or 'N/A'}-{last_data_row or 'N/A'}")
    return first_data_row, last_data_row, descuento_row


def agregar_totales_cotizacion(ws, first_row, last_row, descuento_row, terminos_start_row):
    """Agrega totales después de los productos."""
    if not first_row or not last_row:
        return
    
    print("[7/8] Agregando totales...")
    
    # Los totales van justo antes de los términos
    row_sub = terminos_start_row - 5
    
    # Asegurar que hay espacio
    if row_sub <= last_row + 1:
        # Insertar espacio si es necesario
        ws.insert_rows(last_row + 2, 5)
        row_sub = last_row + 2
        terminos_start_row = row_sub + 5
    
    # Des-mergear celdas del área de totales
    unmerge_cells_in_rows(ws, row_sub, row_sub + 4)
    
    ws.cell(row=row_sub, column=4, value="SUBTOTAL:")
    ws.cell(row=row_sub, column=7, value=f"=SUM(J{first_row}:J{last_row})")
    ws.cell(row=row_sub, column=4).font = Font(bold=True)
    ws.cell(row=row_sub, column=7).font = Font(bold=True)
    
    row_flete = row_sub + 1
    ws.cell(row=row_flete, column=4, value="COSTO DE FLETE E INSTALACION:")
    ws.cell(row=row_flete, column=7, value=f"=G{row_sub}*12%")
    ws.cell(row=row_flete, column=4).font = Font(bold=True)
    ws.cell(row=row_flete, column=7).font = Font(bold=True)
    
    row_sub2 = row_flete + 1
    ws.cell(row=row_sub2, column=4, value="SUBTOTAL:")
    ws.cell(row=row_sub2, column=7, value=f"=G{row_sub}+G{row_flete}")
    ws.cell(row=row_sub2, column=4).font = Font(bold=True)
    ws.cell(row=row_sub2, column=7).font = Font(bold=True)
    
    row_iva = row_sub2 + 1
    ws.cell(row=row_iva, column=4, value="IVA:")
    ws.cell(row=row_iva, column=7, value=f"=G{row_sub2}*16%")
    ws.cell(row=row_iva, column=4).font = Font(bold=True)
    ws.cell(row=row_iva, column=7).font = Font(bold=True)
    
    row_total = row_iva + 1
    ws.cell(row=row_total, column=4, value="TOTAL:")
    ws.cell(row=row_total, column=7, value=f"=G{row_sub2}+G{row_iva}")
    ws.cell(row=row_total, column=4).font = Font(bold=True)
    ws.cell(row=row_total, column=7).font = Font(bold=True)
    
    print(f"      [OK] Totales en filas {row_sub}-{row_total}")
    return row_total


def insertar_imagenes_cotizacion(ws, items, image_map, descuento_row):
    """Inserta imágenes de productos en columna B de Cotizacion."""
    print("      Insertando imágenes...")
    
    current_row = 16
    count = 0
    
    for item in items:
        if item['tipo'] == 'categoria':
            current_row += 1
            continue
        
        q_row = item['row']
        if image_map and q_row in image_map:
            try:
                img_path = image_map[q_row]
                img = XLImage(img_path)
                img.width = 158
                img.height = 118
                ws.add_image(img, f'B{current_row}')
                count += 1
            except Exception as e:
                pass
        
        current_row += 1
    
    print(f"      [OK] {count} imágenes insertadas")


def generar_cotizacion_completa(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI v4 - DISEÑO PERFECTO")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source)
    template_path = Path(args.template)
    
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx")
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template no encontrado: {template_path}")
        sys.exit(1)
    
    temp_dir = None
    wb_src = None
    
    try:
        # PASO 1: Extraer imágenes
        image_map, temp_dir = extraer_imagenes_source(source_path)
        
        # PASO 2: Leer datos
        items = leer_datos_quotation(source_path)
        
        # PASO 3: Abrir template y copiar Quotation
        print("[2/8] Preparando workbook...")
        wb_src = load_workbook(str(source_path), data_only=False)
        wb_dest = load_workbook(str(template_path), data_only=False)
        
        # Copiar Quotation completa
        print("[4/8] Copiando hoja Quotation...")
        copiar_hoja_completa(wb_dest, wb_src, 'Quotation', 'Quotation')
        print("      [OK] Quotation copiada")
        
        # PASO 4: Llenar encabezado
        ws_cot = wb_dest['Cotizacion']
        llenar_encabezado_cotizacion(ws_cot, args)
        
        # PASO 5: Generar Mobiliti
        ws_mob = wb_dest['Mobiliti']
        mobiliti_row_map = generar_mobiliti(ws_mob, items)
        
        # PASO 6: Generar Cotizacion
        first_row, last_row, descuento_row = generar_cotizacion(
            ws_cot, items, image_map, mobiliti_row_map
        )
        
        # PASO 7: Agregar totales
        terminos_start = 32
        # Ajustar si insertamos filas
        if last_row and last_row >= terminos_start - 5:
            terminos_start = last_row + 6
        
        agregar_totales_cotizacion(ws_cot, first_row, last_row, descuento_row, terminos_start)
        
        # PASO 8: Insertar imágenes
        insertar_imagenes_cotizacion(ws_cot, items, image_map, descuento_row)
        
        # Guardar
        print("[8/8] Guardando archivo...")
        wb_dest.save(str(output_path))
        wb_dest.close()
        
        print()
        print("=" * 60)
        print("[OK] COTIZACIÓN GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print(f"Items: {len(items)}")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if wb_src:
            wb_src.close()
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti v4 - Diseño Perfecto',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_v4.py \\
    --source "Quotation.xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotización 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    
    args = parser.parse_args()
    generar_cotizacion_completa(args)


if __name__ == '__main__':
    main()
