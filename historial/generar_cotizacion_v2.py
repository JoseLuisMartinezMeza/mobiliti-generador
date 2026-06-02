#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti v2
=====================================
Flujo: Source -> Quotation -> Mobiliti -> Cotizacion

Uso:
    python generar_cotizacion_v2.py \\
        --source "IZA MONTERREY BH-Quotation Sheet - V1.xlsx" \\
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \\
        --output "Cotizacion-Generada.xlsx" \\
        --cotizacion "100-99999" \\
        --proyecto "NOMBRE PROYECTO" \\
        --cliente "Nombre Cliente" \\
        --correo "cliente@email.com" \\
        --telefono "3312345678" \\
        --direccion "Direccion completa" \\
        --razon_social "Razon Social S.A. de C.V."
"""

import argparse
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


Q_HEADER_ROW = 7


def limpiar_excel():
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes_source(source_path):
    """Extrae imagenes del source y crea mapeo row->imagen_path."""
    temp_dir = tempfile.mkdtemp(prefix="excel_images_")
    image_map = {}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        media_files = [f for f in z.namelist() if 'media' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
    
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    for img in ws._images:
        path = img.path
        filename = os.path.basename(path)
        filepath = os.path.join(temp_dir, filename)
        
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row
        else:
            row = img.anchor.row
        
        if os.path.exists(filepath):
            image_map[row] = filepath
    
    wb.close()
    return image_map, temp_dir


def leer_datos_quotation(source_path):
    """Lee items de Quotation."""
    print("[2/8] Leyendo datos del archivo fuente...")
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    items = []
    for row in range(Q_HEADER_ROW + 1, ws.max_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        if no_val is None and item_name is None:
            continue
        
        if isinstance(no_val, str) and no_val.startswith('-'):
            items.append({
                'tipo': 'categoria',
                'nombre': no_val.strip('- ').strip(),
                'row': row
            })
        elif isinstance(no_val, (int, float)) and item_name:
            desc = ws.cell(row=row, column=4).value
            # Truncar descripcion a primera linea o max 100 chars
            if desc and isinstance(desc, str):
                lines = [l.strip() for l in desc.split('\n') if l.strip()]
                if lines:
                    desc = lines[0][:100]
                else:
                    desc = desc[:100]
            
            items.append({
                'tipo': 'producto',
                'no': int(no_val),
                'item_name': item_name,
                'description': desc,
                'dimension': ws.cell(row=row, column=5).value,
                'color': ws.cell(row=row, column=6).value,
                'qty': ws.cell(row=row, column=7).value,
                'unit_price': ws.cell(row=row, column=10).value,
                'row': row
            })
        elif item_name and isinstance(item_name, str) and not no_val:
            items.append({
                'tipo': 'categoria',
                'nombre': item_name.strip(),
                'row': row
            })
    
    print(f"      [OK] {len(items)} items leidos")
    wb.close()
    return items


def copiar_template_completo(template_path, output_path):
    """Copia el template completo incluyendo todas las hojas."""
    print("[3/8] Copiando template completo...")
    shutil.copy(str(template_path), str(output_path))
    print("      [OK] Template copiado")


def crear_quotation_en_destino(wb_dest, items, source_path):
    """Crea hoja Quotation en destino con datos del source."""
    print("[4/8] Creando hoja Quotation...")
    
    if 'Quotation' in wb_dest.sheetnames:
        del wb_dest['Quotation']
    
    ws_q = wb_dest.create_sheet('Quotation')
    
    wb_src = load_workbook(str(source_path), data_only=False)
    ws_src = wb_src['Quotation']
    
    max_row = min(ws_src.max_row, 100)
    max_col = ws_src.max_column
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            src_cell = ws_src.cell(row=row, column=col)
            if src_cell.value is not None:
                ws_q.cell(row=row, column=col, value=src_cell.value)
    
    wb_src.close()
    print("      [OK] Quotation creada")


def llenar_mobiliti(wb_dest, items):
    """Llena la hoja Mobiliti con datos de Quotation."""
    print("[5/8] Llenando hoja Mobiliti...")
    
    if 'Mobiliti' not in wb_dest.sheetnames:
        print("      [SKIP] Hoja Mobiliti no existe en template")
        return
    
    ws_m = wb_dest['Mobiliti']
    
    # Descombinar y limpiar area de datos en Mobiliti
    unmerge_cells_in_rows(ws_m, 14, 200)
    for row in range(14, 200):
        for col in range(1, 8):
            cell = ws_m.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
    
    # Escribir productos organizados por secciones
    current_row = 14
    section_num = 1
    section_start_row = current_row
    
    for item in items:
        if item['tipo'] == 'categoria':
            # Si ya hay productos en la seccion actual, agregar subtotal
            if current_row > section_start_row + 1:
                ws_m.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
                ws_m.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
                current_row += 1
                section_num += 1
            
            # Fila de categoria
            ws_m.cell(row=current_row, column=1, value=f"=Quotation!A{item['row']}")
            current_row += 1
            section_start_row = current_row
        else:
            # Producto
            q_row = item['row']
            ws_m.cell(row=current_row, column=1, value=f"=Quotation!B{q_row}")
            ws_m.cell(row=current_row, column=2, value=f"=Quotation!D{q_row}")
            ws_m.cell(row=current_row, column=3, value="Sunon Inc")
            ws_m.cell(row=current_row, column=4, value="=IFERROR(VLOOKUP(C{row},Proveedores!A:B,2,0),\" \")".format(row=current_row))
            ws_m.cell(row=current_row, column=5, value=f"=Quotation!G{q_row}")
            ws_m.cell(row=current_row, column=6, value=f"=Quotation!J{q_row}")
            ws_m.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}")
            current_row += 1
    
    # Subtotal de la ultima seccion
    if current_row > section_start_row:
        ws_m.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
        ws_m.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
    
    print(f"      [OK] Mobiliti llenada ({section_num} secciones)")


def unmerge_cells_in_rows(ws, start_row, end_row):
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            to_unmerge.append(str(merged_range))
    for merge in to_unmerge:
        ws.unmerge_cells(merge)


def llenar_encabezado_cliente(ws_cot, args):
    print("[6/8] Llenando datos del cliente...")
    unmerge_cells_in_rows(ws_cot, 3, 12)
    
    campos = {
        3: ("COTIZACION #", args.cotizacion, 2),
        4: (None, datetime.now(), 1),
        7: ("PROYECTO", args.proyecto, 2),
        8: ("Nombre", args.cliente, 2),
        9: ("Correo", args.correo, 2),
        10: ("Telefono", args.telefono, 2),
        11: ("Direccion", args.direccion, 2),
        12: ("Razon Social", args.razon_social, 2),
    }
    
    for row, (label, valor, col) in campos.items():
        if valor:
            ws_cot.cell(row=row, column=col, value=valor)
            if label:
                ws_cot.cell(row=row, column=1, value=f"{label}:")
    
    print(f"      [OK] Cliente: {args.cliente}")


def extraer_terminos_template(ws_cot):
    """Extrae las filas de terminos y condiciones del template (fila 29 en adelante)."""
    START_ROW = 29  # Empezar desde SHOWROOM, no desde TOTAL
    terminos = []
    merges = []
    
    # Encontrar merged cells en el area de terminos (fila 29+)
    for merged_range in list(ws_cot.merged_cells.ranges):
        if merged_range.min_row >= START_ROW:
            merges.append({
                'min_row': merged_range.min_row,
                'min_col': merged_range.min_col,
                'max_row': merged_range.max_row,
                'max_col': merged_range.max_col,
            })
    
    # Extraer valores de filas 29-70
    for row in range(START_ROW, 71):
        row_data = {}
        has_value = False
        for col in range(1, 11):
            cell = ws_cot.cell(row=row, column=col)
            if cell.value is not None:
                row_data[col] = cell.value
                has_value = True
        if has_value or any(m['min_row'] <= row <= m['max_row'] for m in merges):
            terminos.append({
                'rel_row': row - START_ROW,
                'values': row_data,
            })
    
    return {'rows': terminos, 'merges': merges}


def insertar_terminos(ws_cot, terminos_data, start_row):
    """Inserta los terminos y condiciones a partir de start_row."""
    if not terminos_data or not terminos_data['rows']:
        return
    
    offset = start_row - 25
    
    # Copiar valores
    for row_info in terminos_data['rows']:
        dest_row = 25 + row_info['rel_row'] + offset
        for col, value in row_info['values'].items():
            ws_cot.cell(row=dest_row, column=col, value=value)
    
    # Recrear merged cells con nuevo offset
    START_ROW = 29
    for merge in terminos_data['merges']:
        new_min_row = merge['min_row'] - START_ROW + start_row
        new_max_row = merge['max_row'] - START_ROW + start_row
        try:
            ws_cot.merge_cells(
                start_row=new_min_row,
                start_column=merge['min_col'],
                end_row=new_max_row,
                end_column=merge['max_col']
            )
        except Exception:
            pass  # Ignorar si ya existe o no es valido
    
    print(f"      [OK] Terminos insertados en filas {start_row}-{start_row + 40}")


def limpiar_cotizacion(ws_cot):
    """Limpia el area de datos de Cotizacion, incluyendo imagenes."""
    # Guardar imagenes del header (filas 1-14) y eliminar las del area de datos
    header_images = []
    for img in ws_cot._images:
        row = img.anchor._from.row if hasattr(img.anchor, '_from') else img.anchor.row
        if row < 15:
            header_images.append(img)
    ws_cot._images = header_images
    
    unmerge_cells_in_rows(ws_cot, 16, 200)
    for row in range(16, 200):
        for col in range(1, 11):
            cell = ws_cot.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None


def crear_fila_categoria(ws_cot, row_dest, nombre_categoria):
    ws_cot.cell(row=row_dest, column=1, value=nombre_categoria)


def crear_fila_producto(ws_cot, row_dest, item, image_map=None):
    q_row = item['row']
    
    ws_cot.cell(row=row_dest, column=1, value=f"=Quotation!B{q_row}")
    ws_cot.cell(row=row_dest, column=3, value=f"=Quotation!D{q_row}")
    ws_cot.cell(row=row_dest, column=4, value=f"=Quotation!E{q_row}")
    ws_cot.cell(row=row_dest, column=5, value=f"=Quotation!G{q_row}")
    ws_cot.cell(row=row_dest, column=6, value=f"=Quotation!J{q_row}")
    ws_cot.cell(row=row_dest, column=7, value="=G$19")
    ws_cot.cell(row=row_dest, column=8, value=f"=F{row_dest}*G{row_dest}")
    ws_cot.cell(row=row_dest, column=9, value=f"=F{row_dest}-H{row_dest}")
    ws_cot.cell(row=row_dest, column=10, value=f"=I{row_dest}*E{row_dest}")
    
    # Insertar imagen si existe
    if image_map and q_row in image_map:
        try:
            img_path = image_map[q_row]
            img = XLImage(img_path)
            # Tamaño proporcional para celda (157x118 px)
            img.width = 158  # pixels
            img.height = 118  # pixels
            ws_cot.add_image(img, f'B{row_dest}')
        except Exception as e:
            print(f"      Advertencia: No se pudo insertar imagen fila {row_dest}: {e}")


def procesar_items_cotizacion(ws_cot, items, image_map=None):
    print("[7/8] Procesando items en Cotizacion...")
    
    current_row = 16
    first_data_row = None
    
    ws_cot.cell(row=19, column=7, value=0.7)
    
    for item in items:
        if item['tipo'] == 'categoria':
            crear_fila_categoria(ws_cot, current_row, item['nombre'])
            current_row += 1
        else:
            if first_data_row is None:
                first_data_row = current_row
            crear_fila_producto(ws_cot, current_row, item, image_map)
            current_row += 1
    
    last_data_row = current_row - 1
    print(f"      [OK] {len(items)} items, filas {first_data_row}-{last_data_row}")
    return first_data_row, last_data_row


def agregar_totales(ws_cot, row_inicio, row_fin):
    print("[8/8] Agregando totales...")
    
    row_sub = row_fin + 2
    ws_cot.cell(row=row_sub, column=4, value="SUBTOTAL:")
    ws_cot.cell(row=row_sub, column=7, value=f"=SUM(J{row_inicio}:J{row_fin})")
    
    row_flete = row_sub + 1
    ws_cot.cell(row=row_flete, column=4, value="COSTO DE FLETE E INSTALACION:")
    ws_cot.cell(row=row_flete, column=7, value=f"=G{row_sub}*12%")
    
    row_sub2 = row_flete + 1
    ws_cot.cell(row=row_sub2, column=4, value="SUBTOTAL:")
    ws_cot.cell(row=row_sub2, column=7, value=f"=G{row_sub}+G{row_flete}")
    
    row_iva = row_sub2 + 1
    ws_cot.cell(row=row_iva, column=4, value="IVA:")
    ws_cot.cell(row=row_iva, column=7, value=f"=G{row_sub2}*16%")
    
    row_total = row_iva + 1
    ws_cot.cell(row=row_total, column=4, value="TOTAL:")
    ws_cot.cell(row=row_total, column=7, value=f"=G{row_sub2}+G{row_iva}")
    
    print(f"      [OK] Totales en filas {row_sub}-{row_total}")
    return row_total


def generar_cotizacion(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI v2")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source)
    template_path = Path(args.template)
    
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx")
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template no encontrado: {template_path}")
        sys.exit(1)
    
    temp_dir = None
    try:
        # PASO 1: Extraer imagenes
        image_map, temp_dir = extraer_imagenes_source(source_path)
        
        # PASO 2: Leer datos
        items = leer_datos_quotation(source_path)
        
        # PASO 3: Copiar template completo
        copiar_template_completo(template_path, output_path)
        
        # PASO 4: Abrir destino
        wb_dest = load_workbook(str(output_path), data_only=False)
        
        # PASO 5: Crear Quotation
        crear_quotation_en_destino(wb_dest, items, source_path)
        
        # PASO 6: Llenar Mobiliti
        llenar_mobiliti(wb_dest, items)
        
        # PASO 7: Llenar Cotizacion
        ws_cot = wb_dest['Cotizacion']
        
        # Extraer terminos del template ANTES de limpiar
        terminos_data = extraer_terminos_template(ws_cot)
        
        limpiar_cotizacion(ws_cot)
        llenar_encabezado_cliente(ws_cot, args)
        first_row, last_row = procesar_items_cotizacion(ws_cot, items, image_map)
        
        # PASO 8: Totales
        row_total = None
        if first_row and last_row:
            row_total = agregar_totales(ws_cot, first_row, last_row)
        
        # PASO 9: Insertar terminos y condiciones despues de los totales
        if row_total and terminos_data:
            print("[9/9] Restaurando terminos y condiciones...")
            insertar_terminos(ws_cot, terminos_data, row_total + 2)
        
        # Guardar
        wb_dest.save(str(output_path))
        wb_dest.close()
        
        print()
        print("=" * 60)
        print("[OK] COTIZACION GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print(f"Items: {len(items)}")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti v2',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_v2.py \\
    --source "Quotation.xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotizacion 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    
    args = parser.parse_args()
    generar_cotizacion(args)


if __name__ == '__main__':
    main()
