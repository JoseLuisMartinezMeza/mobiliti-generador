"""
Smoke test SaaS: login -> upload -> submit -> wait -> download.

Uses only stdlib plus openpyxl to create a small Quotation.xlsx fixture.
Env:
  MOBILITI_SMOKE_API_URL, MOBILITI_SMOKE_EMAIL, MOBILITI_SMOKE_PASSWORD
  SUPABASE_URL/VITE_SUPABASE_URL + SUPABASE_ANON_KEY/VITE_SUPABASE_ANON_KEY only for legacy Supabase signed-token upload
"""

from __future__ import annotations

import argparse
import json
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook


def _json_request(method: str, url: str, payload: dict | None = None, token: str | None = None) -> dict:
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Content-Type", "application/json")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8")
            return json.loads(body) if body else {}
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8")
        raise RuntimeError(f"{method} {url} -> {exc.code}: {body}") from exc


def _multipart_upload(url: str, file_path: Path, token: str) -> dict:
    boundary = "----mobilitiSmokeBoundary"
    content = file_path.read_bytes()
    body = b"".join(
        [
            f"--{boundary}\r\n".encode(),
            b'Content-Disposition: form-data; name="file"; filename="quotation.xlsx"\r\n',
            b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n",
            content,
            f"\r\n--{boundary}--\r\n".encode(),
        ]
    )
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    req.add_header("Authorization", f"Bearer {token}")
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _signed_storage_upload(supabase_url: str, anon_key: str, bucket: str, path: str, token: str, file_path: Path) -> dict:
    encoded_path = urllib.parse.quote(path, safe="/")
    encoded_bucket = urllib.parse.quote(bucket, safe="")
    url = f"{supabase_url.rstrip('/')}/storage/v1/object/upload/sign/{encoded_bucket}/{encoded_path}?token={urllib.parse.quote(token, safe='')}"
    req = urllib.request.Request(url, data=file_path.read_bytes(), method="PUT")
    req.add_header("apikey", anon_key)
    req.add_header("Authorization", f"Bearer {anon_key}")
    req.add_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    req.add_header("x-upsert", "true")
    with urllib.request.urlopen(req, timeout=90) as resp:
        body = resp.read().decode("utf-8")
        return json.loads(body) if body else {}


def _direct_signed_upload(url: str, file_path: Path) -> dict:
    req = urllib.request.Request(url, data=file_path.read_bytes(), method="PUT")
    req.add_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with urllib.request.urlopen(req, timeout=120) as resp:
        body = resp.read().decode("utf-8")
        return json.loads(body) if body else {}


def _download_head(url: str) -> bytes:
    with urllib.request.urlopen(url, timeout=60) as resp:
        return resp.read(4)


def _download_bytes(url: str, token: str | None = None) -> bytes:
    req = urllib.request.Request(url, method="GET")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def _sample_quotation(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    for col, value in {1: "No.", 2: "Item", 4: "Description", 5: "Dimension", 7: "Qty", 10: "List Price"}.items():
        ws.cell(row=7, column=col, value=value)
    ws.cell(row=8, column=1, value="- Sillas")
    ws.cell(row=9, column=1, value=1)
    ws.cell(row=9, column=2, value="Silla Smoke")
    ws.cell(row=9, column=4, value="Silla operativa")
    ws.cell(row=9, column=5, value="600 x 600")
    ws.cell(row=9, column=7, value=2)
    ws.cell(row=9, column=10, value=250)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Smoke test Mobiliti SaaS")
    parser.add_argument("--api-url", default=None)
    parser.add_argument("--email", default=None)
    parser.add_argument("--password", default=None)
    parser.add_argument("--timeout", type=int, default=90)
    parser.add_argument("--source", type=Path, default=None, help="Quotation.xlsx real. Si falta, crea fixture pequeno.")
    parser.add_argument("--output", type=Path, default=None, help="Ruta donde guardar XLSX descargado.")
    args = parser.parse_args()

    import os

    api_url = (args.api_url or os.environ.get("MOBILITI_SMOKE_API_URL") or "http://127.0.0.1:8000").rstrip("/")
    email = args.email or os.environ.get("MOBILITI_SMOKE_EMAIL") or "dev@mobiliti.local"
    password = args.password or os.environ.get("MOBILITI_SMOKE_PASSWORD") or "dev12345"

    source = args.source
    if source is None:
        source = Path(tempfile.gettempdir()) / "mobiliti_smoke_quotation.xlsx"
        _sample_quotation(source)
    if not source.exists():
        raise SystemExit(f"Source no existe: {source}")

    login = _json_request("POST", f"{api_url}/login", {"email": email, "password": password})
    token = login["access_token"]
    init = _json_request(
        "POST",
        f"{api_url}/cotizaciones/init-upload",
        {"filename": source.name, "size": source.stat().st_size, "template": "online"},
        token,
    )

    if init.get("upload_url"):
        upload_url = init["upload_url"]
        if upload_url.startswith("/"):
            upload_url = f"{api_url}{upload_url}"
        _multipart_upload(upload_url, source, token)
    elif init.get("signed_upload_url"):
        _direct_signed_upload(init["signed_upload_url"], source)
    else:
        supabase_url = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
        anon_key = os.environ.get("SUPABASE_ANON_KEY") or os.environ.get("VITE_SUPABASE_ANON_KEY")
        if not supabase_url or not anon_key:
            raise RuntimeError("Faltan SUPABASE_URL/VITE_SUPABASE_URL y SUPABASE_ANON_KEY/VITE_SUPABASE_ANON_KEY para upload prod")
        _signed_storage_upload(supabase_url, anon_key, init["bucket"], init["path"], init["token"], source)

    _json_request(
        "POST",
        f"{api_url}/cotizaciones/{init['job_id']}/submit",
        {
            "cotizacion": "SMOKE-001",
            "proyecto": "Smoke Test",
            "cliente": "Cliente Smoke",
            "correo": "smoke@example.com",
            "telefono": "555",
            "direccion": "Direccion Smoke",
            "razon_social": "Empresa Smoke",
            "template": "online",
        },
        token,
    )

    deadline = time.monotonic() + args.timeout
    job = {}
    while time.monotonic() < deadline:
        job = _json_request("GET", f"{api_url}/cotizaciones/{init['job_id']}", token=token)["job"]
        if job["status"] in {"completed", "failed"}:
            break
        time.sleep(2)
    if job.get("status") != "completed":
        raise SystemExit(f"Job no completo: {json.dumps(job, ensure_ascii=False)}")

    download = _json_request("GET", f"{api_url}/cotizaciones/{init['job_id']}/download", token=token)
    url = download["download_url"]
    if url.startswith("/"):
        url = f"{api_url}{url}"
    magic = _download_head(url)
    if magic != b"PK\x03\x04":
        raise SystemExit(f"Descarga no parece XLSX: {magic!r}")

    saved = None
    if args.output:
        content = _download_bytes(f"{api_url}/cotizaciones/{init['job_id']}/file", token)
        if not content.startswith(b"PK\x03\x04"):
            raise SystemExit(f"Descarga directa no parece XLSX: {content[:4]!r}")
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_bytes(content)
        saved = str(args.output)

    print(json.dumps({"ok": True, "job_id": init["job_id"], "status": job["status"], "saved": saved}, ensure_ascii=False))


if __name__ == "__main__":
    main()
