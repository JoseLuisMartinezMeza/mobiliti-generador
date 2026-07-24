from __future__ import annotations

import ast
from copy import deepcopy
from dataclasses import dataclass
from collections import Counter
from decimal import Decimal
import hashlib
from pathlib import Path
import sys
from typing import Sequence
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
import pytest

from mobiliti_saas.quote_engine.mobiliti_layout import SectionNeed, plan_mobiliti_layout
from mobiliti_saas.quote_engine.ooxml_package import (
    XlsxPackage,
    assert_package_preserved,
)
from mobiliti_saas.quote_engine.quotation_sheets import (
    QuotationDataRow,
    _with_canonical_hash,
)
from mobiliti_saas.quote_engine import engine as quote_engine


ROOT = Path(__file__).resolve().parents[1]
WORKER_DIR = ROOT / "mobiliti_saas" / "worker"
OFFICIAL_TEMPLATE = (
    WORKER_DIR / "templates" / "Formato Cotizacion 2026 Oficial.xlsx"
)
if str(WORKER_DIR) not in sys.path:
    sys.path.insert(0, str(WORKER_DIR))

import quote_worker  # noqa: E402


SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICIAL_ALLOWED_PARTS = frozenset(
    {
        "[Content_Types].xml",
        "xl/_rels/workbook.xml.rels",
        "xl/calcChain.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/drawings/drawing1.xml",
        "xl/styles.xml",
        "xl/workbook.xml",
        "xl/worksheets/quotation_data1.xml",
        "xl/worksheets/quotation_original1.xml",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet3.xml",
        "xl/worksheets/sheet4.xml",
        "xl/worksheets/sheet5.xml",
    }
)
SEVEN_CATALOGS = (
    "tarkett",
    "offiho",
    "cr-global",
    "sonara",
    "sunon",
    "alma",
    "lumbro",
)
CATALOG_LABELS = {
    "tarkett": "Tarkett",
    "offiho": "Offiho",
    "cr-global": "CR Global",
    "sonara": "Sonara",
    "sunon": "Sunon",
    "alma": "ALMA",
    "lumbro": "Lumbro",
}


@dataclass(frozen=True)
class QuoteShape:
    section_counts: tuple[int, ...]

    def __init__(self, section_counts: Sequence[int]):
        object.__setattr__(self, "section_counts", tuple(section_counts))

    @property
    def total_items(self) -> int:
        return sum(self.section_counts)


STRESS_SHAPES = (
    QuoteShape([34]),
    QuoteShape([100]),
    QuoteShape([1] * 17),
    QuoteShape([1] * 20),
    QuoteShape([40] * 20),
    QuoteShape([100] * 10),
)


@dataclass(frozen=True)
class SyntheticMixedRequest:
    source: Path
    original_quotation: Path
    quotation_data: tuple[QuotationDataRow, ...]
    metadata: dict[str, object]
    names: tuple[str, ...]


def as_persistent_project_request(
    request: SyntheticMixedRequest,
) -> SyntheticMixedRequest:
    """Agrega al stress mixto el contexto inmutable de un Proyecto guardado.

    Las filas físicas se conservan intactas. Las tres primeras forman una
    composición (principal, por unidad y fija); las demás son principales
    independientes. Dos ocurrencias de catálogo comparten identidad canónica
    deliberadamente para cubrir reemplazos/duplicados sin colapsar sus UUID de
    ocurrencia.
    """

    frozen_lines: list[dict[str, object]] = []
    compositions: list[dict[str, object]] = []
    rows_by_section: dict[str, list[QuotationDataRow]] = {}
    for row in request.quotation_data:
        rows_by_section.setdefault(row.section_id, []).append(row)

    for section_rows in rows_by_section.values():
        first_id = section_rows[0].item_key
        for index, row in enumerate(section_rows):
            is_first_composition_child = (
                row.section_id == request.quotation_data[0].section_id
                and index in {1, 2}
            )
            if is_first_composition_child:
                frozen_lines.append(
                    {
                        "line_id": row.item_key,
                        "role": "complement",
                        "section_id": None,
                        "parent_line_id": first_id,
                        "identity": {
                            "internal_id": "sunon:duplicate-component",
                            "base_option_id": "",
                            "add_on_option_ids": [],
                        },
                    }
                )
                continue
            frozen_lines.append(
                {
                    "line_id": row.item_key,
                    "role": "principal",
                    "section_id": row.section_id,
                    "parent_line_id": None,
                    "identity": {
                        "internal_id": (
                            "sunon:duplicate-principal"
                            if index in {3, 4}
                            else f"sunon:{row.item_key}"
                        ),
                        "base_option_id": "",
                        "add_on_option_ids": [],
                    },
                }
            )

        if section_rows[0] is request.quotation_data[0]:
            component_rows = section_rows[:3]
            compositions.append(
                {
                    "principal_line_id": component_rows[0].item_key,
                    "section_id": component_rows[0].section_id,
                    "component_line_ids": [
                        component.item_key for component in component_rows
                    ],
                    "price_terms": [
                        {
                            "line_id": component_rows[0].item_key,
                            "numerator": "1",
                            "denominator": "1",
                        },
                        {
                            "line_id": component_rows[1].item_key,
                            "numerator": "1",
                            "denominator": "1",
                        },
                        {
                            "line_id": component_rows[2].item_key,
                            "numerator": "1",
                            "denominator": "1",
                        },
                    ],
                }
            )
            independent_rows = section_rows[3:]
        else:
            independent_rows = section_rows
        compositions.extend(
            {
                "principal_line_id": row.item_key,
                "section_id": row.section_id,
                "component_line_ids": [row.item_key],
                "price_terms": [
                    {
                        "line_id": row.item_key,
                        "numerator": "1",
                        "denominator": "1",
                    }
                ],
            }
            for row in independent_rows
        )

    metadata = deepcopy(request.metadata)
    metadata["catalog_source_hashes"] = {
        catalog: hashlib.sha256(f"stress:{catalog}".encode()).hexdigest()
        for catalog in SEVEN_CATALOGS
    }
    metadata["project_context"] = {
        "project_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
        "project_revision": 7,
        "project_payload_hash": "a" * 64,
        "normalized_project_payload": {
            "sections": [
                {
                    "section_id": section_id,
                    "concept": section_rows[0].section_title,
                    "position": position,
                }
                for position, (section_id, section_rows) in enumerate(
                    rows_by_section.items()
                )
            ],
            "lines": frozen_lines,
        },
        "compositions": compositions,
    }
    return SyntheticMixedRequest(
        source=request.source,
        original_quotation=request.original_quotation,
        quotation_data=request.quotation_data,
        metadata=metadata,
        names=request.names,
    )


def excel_desktop_unavailable_reason() -> str | None:
    """Devuelve una razón de skip únicamente cuando Excel COM no está disponible."""

    if sys.platform != "win32":
        return "Excel de escritorio COM solo está disponible en Windows"
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        return f"pywin32 no está disponible: {exc}"

    pythoncom.CoInitialize()
    application = None
    try:
        application = win32com.client.DispatchEx("Excel.Application")
        application.Visible = False
        application.DisplayAlerts = False
        application.Quit()
        application = None
    except Exception as exc:
        return f"Excel de escritorio no está disponible: {exc}"
    finally:
        if application is not None:
            try:
                application.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
    return None


def excel_com_roundtrip(
    source: Path,
    destination: Path,
    *,
    formula_expectations: dict[tuple[str, str], str],
    inspected_cells: Sequence[tuple[str, str]],
) -> Path:
    """Valida, recalcula y reabre una copia sin modificar el XLSX del worker."""

    import pythoncom
    import pywintypes
    import tempfile
    import time
    import win32com.client

    def retry_com(action, *, timeout_seconds: float = 180.0):
        deadline = time.monotonic() + timeout_seconds
        while True:
            try:
                return action()
            except (AttributeError, TypeError, pywintypes.com_error):
                if time.monotonic() >= deadline:
                    raise
                pythoncom.PumpWaitingMessages()
                time.sleep(0.25)

    def wait_until_ready(application) -> None:
        deadline = time.monotonic() + 180.0
        while not retry_com(lambda: bool(application.Ready)):
            if time.monotonic() >= deadline:
                raise TimeoutError("Excel no terminó de procesar el libro")
            pythoncom.PumpWaitingMessages()
            time.sleep(0.25)

    def quit_excel(application) -> None:
        retry_com(
            lambda: application._oleobj_.InvokeTypes(
                0x12E,
                0,
                pythoncom.DISPATCH_METHOD,
                (pythoncom.VT_EMPTY, 0),
                (),
            )
        )

    def close_workbook(workbook, *, save: bool = False) -> None:
        retry_com(lambda: workbook.Close(SaveChanges=save))

    def configure(application) -> None:
        application.Visible = False
        application.DisplayAlerts = False
        application.AskToUpdateLinks = False
        application.AutomationSecurity = 3

    def repair_log_snapshot() -> dict[str, tuple[int, int]]:
        temp_dir = Path(tempfile.gettempdir())
        candidates = {
            *temp_dir.glob("error*.xml"),
            *temp_dir.glob("recover*.xml"),
            *temp_dir.glob("*repair*.xml"),
        }
        return {
            str(path.resolve()): (path.stat().st_mtime_ns, path.stat().st_size)
            for path in candidates
            if path.is_file()
        }

    def open_workbook(application, path: Path, *, read_only: bool):
        return retry_com(
            lambda: application.Workbooks.Open(
                str(path),
                UpdateLinks=0,
                ReadOnly=read_only,
                IgnoreReadOnlyRecommended=True,
                CorruptLoad=0,
            )
        )

    def assert_no_repair_log(before: dict[str, tuple[int, int]]) -> None:
        after = repair_log_snapshot()
        changed = {
            path
            for path, signature in after.items()
            if before.get(path) != signature
        }
        assert changed == set(), (
            "Excel produjo un registro de recuperación/reparación: "
            f"{sorted(changed)}"
        )

    def cell_text(cell) -> str:
        return str(retry_com(lambda: cell.Text) or "")

    def workbook_cell(workbook, sheet_name: str, coordinate: str):
        column_name, row = coordinate_from_string(coordinate)
        column = column_index_from_string(column_name)
        return retry_com(
            lambda: workbook.Worksheets.Item(sheet_name).Cells.Item(row, column)
        )

    def assert_dynamic_surface(workbook) -> None:
        for (sheet_name, coordinate), expected in formula_expectations.items():
            cell = workbook_cell(workbook, sheet_name, coordinate)
            actual = str(retry_com(lambda: cell.Formula) or "")
            canonical_actual = actual.replace("_xlfn.", "")
            canonical_expected = expected.replace("_xlfn.", "")
            assert canonical_actual == canonical_expected, (
                f"Fórmula alterada en {sheet_name}!{coordinate}: "
                f"{actual!r} != {expected!r}"
            )
            assert "[" not in actual, (
                f"Vínculo externo en fórmula dinámica {sheet_name}!{coordinate}"
            )
            rendered = cell_text(cell).upper()
            assert "#REF!" not in rendered
            assert "#VALUE!" not in rendered
        for sheet_name, coordinate in inspected_cells:
            cell = workbook_cell(workbook, sheet_name, coordinate)
            rendered = cell_text(cell).upper()
            assert "#REF!" not in rendered
            assert "#VALUE!" not in rendered

    source_path = Path(source).resolve()
    destination_path = Path(destination).resolve()
    assert source_path != destination_path, "El roundtrip debe usar una copia distinta"
    assert not destination_path.exists(), (
        f"El destino de validación ya existe: {destination_path}"
    )
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()

    pythoncom.CoInitialize()
    application = None
    workbook = None
    try:
        # La primera apertura detecta reparación antes de crear o guardar la copia.
        application = win32com.client.DispatchEx("Excel.Application")
        configure(application)
        repair_logs_before_open = repair_log_snapshot()
        workbook = open_workbook(application, source_path, read_only=True)
        wait_until_ready(application)
        assert_no_repair_log(repair_logs_before_open)
        assert_dynamic_surface(workbook)
        close_workbook(workbook)
        workbook = None
        quit_excel(application)
        application = None

        destination_path.write_bytes(source_path.read_bytes())
        application = win32com.client.DispatchEx("Excel.Application")
        configure(application)
        repair_logs_before_open = repair_log_snapshot()
        workbook = open_workbook(application, destination_path, read_only=False)
        wait_until_ready(application)
        assert_no_repair_log(repair_logs_before_open)
        assert_dynamic_surface(workbook)
        retry_com(application.CalculateFullRebuild)
        wait_until_ready(application)
        assert_dynamic_surface(workbook)
        retry_com(workbook.Save)
        wait_until_ready(application)
        close_workbook(workbook)
        workbook = None
        quit_excel(application)
        application = None

        application = win32com.client.DispatchEx("Excel.Application")
        configure(application)
        repair_logs_before_open = repair_log_snapshot()
        workbook = open_workbook(application, destination_path, read_only=True)
        wait_until_ready(application)
        assert_no_repair_log(repair_logs_before_open)
        assert_dynamic_surface(workbook)
        close_workbook(workbook)
        workbook = None
        quit_excel(application)
        application = None
    finally:
        if workbook is not None:
            try:
                close_workbook(workbook)
            except Exception:
                pass
        if application is not None:
            try:
                quit_excel(application)
            except Exception:
                pass
        pythoncom.CoUninitialize()

    assert hashlib.sha256(source_path.read_bytes()).hexdigest() == source_hash
    assert ZipFile(destination_path).testzip() is None
    package = XlsxPackage.read(destination_path)
    suspicious_parts = {
        name
        for name in package.parts
        if "recover" in name.lower() or "repair" in name.lower()
    }
    assert suspicious_parts == set()
    return destination_path


def _write_original_quotation(path: Path, imported_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Quotation"
    worksheet.merge_cells("A1:N1")
    worksheet["A1"] = "Quotation importada original"
    worksheet["A1"].font = Font(bold=True)
    for column, value in {
        1: "No.",
        2: "Item",
        4: "Description",
        7: "Qty",
        10: "List Price",
    }.items():
        worksheet.cell(7, column).value = value
    worksheet["A8"] = "- Importados"
    worksheet["A9"] = 1
    worksheet["B9"] = imported_name
    worksheet["D9"] = "Fila original preservada"
    worksheet["G9"] = 1
    worksheet["J9"] = 11
    worksheet["N9"] = "=G9*J9"
    worksheet.row_dimensions[12].hidden = True
    worksheet.print_area = "A1:N20"
    workbook.save(path)
    workbook.close()


def _source_headers(worksheet) -> None:
    for column, value in {
        1: "No.",
        2: "Item",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
        12: "Supplier",
        13: "Discount Percent",
        14: "Original Currency",
        15: "Original Unit Price",
        16: "Frozen Exchange Rate",
        17: "Source Reference",
        18: "Price Mode",
        19: "Auto Electrification",
        20: "Canonical Key",
        21: "Source Hash",
        22: "Original Source Row",
        23: "Upstream Row Hash",
    }.items():
        worksheet.cell(7, column).value = value


def synthetic_mixed_request(
    tmp_path: Path,
    shape: QuoteShape,
    *,
    include_imported: bool,
    catalogs: Sequence[str],
) -> SyntheticMixedRequest:
    if not include_imported or tuple(catalogs) != SEVEN_CATALOGS:
        raise ValueError("La matriz stress exige importado y los siete catalogos")
    imported_name = "Stress Item 0001 [imported]"
    original = tmp_path / "original-imported.xlsx"
    _write_original_quotation(original, imported_name)
    original_hash = hashlib.sha256(original.read_bytes()).hexdigest()

    source = tmp_path / "synthetic-mixed.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Quotation"
    _source_headers(worksheet)
    row = 8
    position = 0
    rows: list[QuotationDataRow] = []
    names: list[str] = []
    for section_index, item_count in enumerate(shape.section_counts, start=1):
        section_id = f"stress-section-{section_index}"
        section_title = f"Stress Section {section_index}"
        worksheet.cell(row, 1).value = f"- {section_title}"
        row += 1
        for _offset in range(item_count):
            position += 1
            imported = position == 1
            origin = "imported" if imported else catalogs[(position - 2) % len(catalogs)]
            provider = "Imported Provider" if imported else CATALOG_LABELS[origin]
            item_key = (
                f"import:stress:{row}"
                if imported
                else f"{origin}:stress-{position:04d}"
            )
            name = f"Stress Item {position:04d} [{origin}]"
            cost = Decimal((position % 97) + 1).quantize(Decimal("0.01"))
            source_hash = (
                original_hash
                if imported
                else hashlib.sha256(f"stress:{origin}".encode()).hexdigest()
            )
            reference = f"{origin}:stress:{position}"
            upstream_hash = (
                hashlib.sha256(reference.encode()).hexdigest() if imported else ""
            )
            worksheet.cell(row, 1).value = position
            worksheet.cell(row, 2).value = name
            worksheet.cell(row, 4).value = f"Descripcion stress {position}"
            worksheet.cell(row, 5).value = "60 x 60 cm"
            worksheet.cell(row, 7).value = 1
            worksheet.cell(row, 10).value = float(cost)
            worksheet.cell(row, 12).value = provider
            worksheet.cell(row, 13).value = 0
            worksheet.cell(row, 14).value = "MXN"
            worksheet.cell(row, 15).value = float(cost)
            worksheet.cell(row, 16).value = 1
            worksheet.cell(row, 17).value = reference
            worksheet.cell(row, 18).value = "imported" if imported else "net"
            worksheet.cell(row, 19).value = False
            worksheet.cell(row, 20).value = item_key
            worksheet.cell(row, 21).value = source_hash
            worksheet.cell(row, 22).value = row if imported else None
            worksheet.cell(row, 23).value = upstream_hash or None
            rows.append(
                _with_canonical_hash(
                    QuotationDataRow(
                        item_key=item_key,
                        section_id=section_id,
                        section_title=section_title,
                        position=position,
                        origin=origin,
                        source_row=row if imported else None,
                        original_currency="MXN",
                        original_cost=cost,
                        frozen_rate=Decimal("1"),
                        converted_cost=cost,
                        quantity=Decimal("1"),
                        provider=provider,
                        region="imported" if imported else "Centro",
                        source_hash=source_hash,
                        upstream_row_hash=upstream_hash,
                        row_hash="",
                    )
                )
            )
            names.append(name)
            row += 1
    workbook.save(source)
    workbook.close()
    return SyntheticMixedRequest(
        source=source,
        original_quotation=original,
        quotation_data=tuple(rows),
        metadata={
            "catalog_price_mode": "mixed_catalog_converted",
            "quote_currency": "MXN",
            "descuento": 30,
            "rate_summary": [],
            "auto_electrification_rate": None,
            "catalog_source_hashes": {
                catalog: hashlib.sha256(
                    f"stress:{catalog}".encode()
                ).hexdigest()
                for catalog in SEVEN_CATALOGS
            },
            "cotizacion": "100-STRESS",
            "proyecto": "Stress local",
            "cliente": "Cliente stress",
        },
        names=tuple(names),
    )


def run_local_worker_job(
    tmp_path: Path,
    request: SyntheticMixedRequest,
    monkeypatch: pytest.MonkeyPatch,
) -> Path:
    output = tmp_path / "stress-output.xlsx"
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")
    monkeypatch.setattr(quote_worker, "_template_path", lambda: str(OFFICIAL_TEMPLATE))
    quote_worker._run_generator(
        {"metadata": dict(request.metadata)},
        quote_worker.PreparedGeneratorInput(
            parser_source=request.source,
            original_quotation=request.original_quotation,
            quotation_data=request.quotation_data,
        ),
        output,
    )
    assert output.is_file()
    return output


def _worksheet_root(package: XlsxPackage, sheet_name: str) -> ET.Element:
    return ET.fromstring(package.parts[package.sheet_part(sheet_name)])


def _cell_map(package: XlsxPackage, sheet_name: str) -> dict[str, ET.Element]:
    root = _worksheet_root(package, sheet_name)
    return {
        cell.attrib["r"]: cell
        for cell in root.findall(f".//{{{SHEET_NS}}}c")
        if "r" in cell.attrib
    }


def _cell_text(cell: ET.Element | None) -> str:
    if cell is None:
        return ""
    return "".join(
        node.text or "" for node in cell.findall(f".//{{{SHEET_NS}}}t")
    )


def _formula(cell: ET.Element | None) -> str:
    if cell is None:
        return ""
    formula = cell.find(f"{{{SHEET_NS}}}f")
    return "" if formula is None else (formula.text or "")


def _quotation_data_keys(package: XlsxPackage) -> list[str]:
    cells = _cell_map(package, "Quotation_Data")
    return [
        _cell_text(cells[f"A{row}"])
        for row in range(2, len(cells) + 2)
        if f"A{row}" in cells
    ]


def _formulas_containing(package: XlsxPackage, token: str) -> Counter[str]:
    formulas: list[str] = []
    for name, payload in package.parts.items():
        if not name.endswith(".xml"):
            continue
        try:
            root = ET.fromstring(payload)
        except ET.ParseError:
            continue
        formulas.extend(
            formula.text or ""
            for formula in root.findall(f".//{{{SHEET_NS}}}f")
            if token in (formula.text or "")
        )
        formulas.extend(
            defined_name.text or ""
            for defined_name in root.findall(f".//{{{SHEET_NS}}}definedName")
            if token in (defined_name.text or "")
        )
    return Counter(formulas)


def _assert_calc_chain_targets_formula_cells(
    package: XlsxPackage,
    required_mobiliti_rows: Sequence[int] = (),
) -> None:
    workbook = ET.fromstring(package.parts["xl/workbook.xml"])
    sheet_parts_by_id = {
        sheet.attrib["sheetId"]: package.sheet_part(sheet.attrib["name"])
        for sheet in workbook.findall(f"{{{SHEET_NS}}}sheets/{{{SHEET_NS}}}sheet")
    }
    formula_coordinates = {
        part: {
            cell.attrib["r"]
            for cell in ET.fromstring(package.parts[part]).findall(
                f".//{{{SHEET_NS}}}c"
            )
            if cell.find(f"{{{SHEET_NS}}}f") is not None
        }
        for part in sheet_parts_by_id.values()
    }
    current_sheet_id: str | None = None
    dangling: list[tuple[str | None, str]] = []
    mobiliti_sheet_id = next(
        sheet_id
        for sheet_id, part in sheet_parts_by_id.items()
        if part == package.sheet_part("Mobiliti")
    )
    mobiliti_calc_coordinates: list[str] = []
    calc_chain = ET.fromstring(package.parts["xl/calcChain.xml"])
    for entry in calc_chain.findall(f"{{{SHEET_NS}}}c"):
        current_sheet_id = entry.attrib.get("i", current_sheet_id)
        coordinate = entry.attrib["r"]
        part = sheet_parts_by_id.get(current_sheet_id or "")
        if part is None or coordinate not in formula_coordinates[part]:
            dangling.append((current_sheet_id, coordinate))
        if current_sheet_id == mobiliti_sheet_id:
            column = coordinate.rstrip("0123456789")
            if column_index_from_string(column) <= 34:
                mobiliti_calc_coordinates.append(coordinate)
    assert dangling == []

    mobiliti_calc_set = set(mobiliti_calc_coordinates)
    assert len(mobiliti_calc_coordinates) == len(mobiliti_calc_set)
    calc_properties = workbook.find(f"{{{SHEET_NS}}}calcPr")
    assert calc_properties is not None
    assert {
        key: calc_properties.attrib.get(key)
        for key in ("calcMode", "fullCalcOnLoad", "forceFullCalc")
    } == {
        "calcMode": "auto",
        "fullCalcOnLoad": "1",
        "forceFullCalc": "1",
    }


def _assert_subtotals_cover_all_items(
    package: XlsxPackage,
    layout,
    cotizacion_rows: Sequence[int],
) -> None:
    mobiliti = _cell_map(package, "Mobiliti")
    for section in layout.sections:
        if section.item_count == 0:
            continue
        end = section.product_start + section.item_count - 1
        assert _formula(mobiliti[f"H{section.subtotal_row}"]) == (
            f"SUM(IFERROR(H{section.product_start}:H{end},0))"
        )
    total_formula = _formula(mobiliti[f"H{layout.total_row}"])
    assert all(
        f"H{section.subtotal_row}" in total_formula
        for section in layout.sections
        if section.item_count
    )
    cotizacion = _cell_map(package, "Cotizacion")
    expected = (
        f"SUM(IFERROR(J{min(cotizacion_rows)}:J{max(cotizacion_rows)},0))"
    )
    assert expected in {_formula(cell) for cell in cotizacion.values()}


@pytest.mark.parametrize(
    "shape",
    STRESS_SHAPES,
    ids=("one-34", "one-100", "sections-17", "sections-20", "20x40", "10x100"),
)
def test_large_quotes_preserve_every_line_and_official_contract(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    shape: QuoteShape,
) -> None:
    request = synthetic_mixed_request(
        tmp_path,
        shape,
        include_imported=True,
        catalogs=SEVEN_CATALOGS,
    )
    output = run_local_worker_job(tmp_path, request, monkeypatch)
    package = XlsxPackage.read(output)
    layout = plan_mobiliti_layout(
        [
            SectionNeed(f"stress-section-{index}", f"Stress Section {index}", count)
            for index, count in enumerate(shape.section_counts, start=1)
        ]
    )

    keys = _quotation_data_keys(package)
    assert keys == [row.item_key for row in request.quotation_data]
    assert len(keys) == shape.total_items
    assert len(set(keys)) == shape.total_items

    mobiliti = _cell_map(package, "Mobiliti")
    assert [_cell_text(mobiliti[f"D{row}"]) for row in layout.item_rows] == list(
        request.names
    )
    cotizacion = _cell_map(package, "Cotizacion")
    cotizacion_names = [
        _cell_text(cell)
        for coordinate, cell in cotizacion.items()
        if coordinate.startswith("A") and _cell_text(cell) in set(request.names)
    ]
    assert cotizacion_names == list(request.names)
    cotizacion_rows = sorted(
        int(coordinate[1:])
        for coordinate, cell in cotizacion.items()
        if coordinate.startswith("A") and _cell_text(cell) in set(request.names)
    )
    official = XlsxPackage.read(OFFICIAL_TEMPLATE)
    # El formato oficial contiene referencias historicas en superficies
    # protegidas. La generacion no puede inventar ninguna adicional.
    assert not (
        _formulas_containing(package, "#REF!")
        - _formulas_containing(official, "#REF!")
    )
    assert all(
        "#REF!" not in _formula(mobiliti[f"{column}{row}"])
        for row in layout.item_rows
        for column in ("W", "X")
    )
    assert all(
        mobiliti[f"J{row}"].find(f"{{{SHEET_NS}}}f") is None
        and Decimal(mobiliti[f"J{row}"].findtext(f"{{{SHEET_NS}}}v"))
        == canonical.converted_cost
        for row, canonical in zip(
            layout.item_rows, request.quotation_data, strict=True
        )
    )
    _assert_subtotals_cover_all_items(package, layout, cotizacion_rows)
    _assert_calc_chain_targets_formula_cells(package, layout.item_rows)
    assert_package_preserved(
        OFFICIAL_TEMPLATE,
        output,
        allowed_parts=set(OFFICIAL_ALLOWED_PARTS),
    )

    quotation = _cell_map(package, "Quotation")
    visible_values = {_cell_text(cell) for cell in quotation.values()}
    assert request.names[0] in visible_values
    assert request.names[1] not in visible_values
    assert _formula(quotation["N9"]) == "G9*J9"

FORBIDDEN_LEGACY_ENGINE_SYMBOLS = {
    "_sanitize_template_workbook",
    "_default_template",
    "_ensure_mobiliti_formula_layout",
    "_ensure_mobiliti_capacity_legacy",
    "_write_mobiliti_row_formulas",
    "_normalize_mobiliti_row_formulas",
    "_set_mobiliti_subtotal_formulas",
    "_copy_source_sheet",
    "_patch_quotation_drawing_from_source",
    "_sanitize_output_xlsx_for_excel",
    "_prepare_openpyxl_mobiliti_columns",
    "_populate_openpyxl_mobiliti_formulas",
    "_refresh_openpyxl_mobiliti_formulas",
    "_populate_openpyxl_mobiliti_subtotal",
    "_ensure_mobiliti_capacity",
    "_write_mobiliti",
    "_prepare_unused_mobiliti_product_rows",
}


def test_engine_no_longer_exposes_destructive_legacy_writers() -> None:
    tree = ast.parse(Path(quote_engine.__file__).read_text(encoding="utf-8"))
    definitions = {
        node.name
        for node in ast.walk(tree)
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
    }
    calls = {
        node.func.id
        for node in ast.walk(tree)
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
    }

    assert FORBIDDEN_LEGACY_ENGINE_SYMBOLS.isdisjoint(definitions)
    assert FORBIDDEN_LEGACY_ENGINE_SYMBOLS.isdisjoint(calls)
