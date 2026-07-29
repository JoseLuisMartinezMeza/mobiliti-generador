from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from io import BytesIO
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mobiliti_saas.quote_engine.mobiliti_layout import (  # noqa: E402
    SectionNeed,
    plan_mobiliti_layout,
)
from mobiliti_saas.quote_engine import generate_quote  # noqa: E402
import mobiliti_saas.quote_engine.engine as engine_module  # noqa: E402
import mobiliti_saas.quote_engine.official_composer as official_composer_module  # noqa: E402
import mobiliti_saas.quote_engine.ooxml_package as ooxml_package_module  # noqa: E402
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage  # noqa: E402
from mobiliti_saas.quote_engine.ooxml_package import (  # noqa: E402
    PackageMutation,
    relationship_part_name,
    resolve_internal_target,
)
from mobiliti_saas.quote_engine.ooxml_worksheet import (  # noqa: E402
    MobilitiCellWrite,
    MobilitiSheetMutation,
    WorksheetEditor,
    build_mobiliti_sheet,
)
from mobiliti_saas.quote_engine.mobiliti_pricing import (  # noqa: E402
    PricingRowBinding,
    build_mobiliti_pricing_writes,
    lumbro_frozen_cost,
    write_official_currency_selector,
)
from mobiliti_saas.quote_engine.official_composer import (  # noqa: E402
    ComposeRequest,
    CotizacionImageSource,
    CotizacionMetadata,
    CotizacionPriceTerm,
    CotizacionProduct,
    CotizacionProductImage,
    CotizacionSection,
    CotizacionSheetEditor,
    _translate_estrategia,
    build_allowlisted_mutation,
    compose_official_quote,
    verify_output_contract,
)
from mobiliti_saas.quote_engine.official_template import (  # noqa: E402
    load_template_contract,
)
from mobiliti_saas.quote_engine.quotation_sheets import (  # noqa: E402
    LocalDefinedName,
    QuotationDataRow,
    SheetAddition,
    _parse_xml_document,
    _with_canonical_hash,
    build_quotation_data_sheet,
    transplant_quotation,
)


OFFICIAL_TEMPLATE = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 Oficial.xlsx"
)
CONTRACT_PATH = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "formato-cotizacion-2026-oficial.contract.json"
)
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
CHART = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def _write_engine_source(
    path: Path,
    products: tuple[dict[str, object], ...],
    *,
    category: str = "Sillas",
    image_path: Path | None = None,
) -> None:
    workbook = Workbook()
    quotation = workbook.active
    quotation.title = "Quotation"
    headers = {
        1: "No.",
        2: "Item",
        4: "Description",
        5: "Dimension",
        6: "Volume",
        7: "Qty",
        10: "List Price",
        12: "Supplier",
        13: "Discount Percent",
        14: "Original Currency",
        15: "Original Unit Price",
        16: "Frozen Exchange Rate",
        17: "Source Reference",
        18: "Price Mode",
        19: "Auto Electrification",
        20: "Canonical Key",
        21: "Source Hash",
        22: "Original Source Row",
        23: "Upstream Row Hash",
    }
    for column, value in headers.items():
        quotation.cell(7, column).value = value
    quotation["A8"] = f"- {category}"
    for offset, product in enumerate(products, start=9):
        quotation.cell(offset, 1).value = offset - 8
        quotation.cell(offset, 2).value = product["name"]
        quotation.cell(offset, 4).value = product.get("description", "")
        quotation.cell(offset, 5).value = product.get("dimension", "")
        quotation.cell(offset, 6).value = product.get("m3")
        quotation.cell(offset, 7).value = product.get("quantity", 1)
        quotation.cell(offset, 10).value = product.get("price", 0)
        quotation.cell(offset, 12).value = product.get("provider", "")
        quotation.cell(offset, 13).value = product.get("discount")
        quotation.cell(offset, 14).value = product.get("currency", "")
        quotation.cell(offset, 15).value = product.get("original_price")
        quotation.cell(offset, 16).value = product.get("rate")
        quotation.cell(offset, 17).value = product.get("reference", "")
        quotation.cell(offset, 18).value = product.get("mode", "")
        quotation.cell(offset, 19).value = product.get("auto")
        quotation.cell(offset, 20).value = product.get(
            "canonical_key",
            product.get("reference", ""),
        )
        quotation.cell(offset, 21).value = product.get(
            "source_hash",
            "a" * 64 if product.get("reference") else "",
        )
        quotation.cell(offset, 22).value = product.get(
            "source_row",
            offset if product.get("mode") == "imported" else None,
        )
        quotation.cell(offset, 23).value = product.get(
            "upstream_row_hash",
            "b" * 64 if product.get("mode") == "imported" else "",
        )
    if image_path is not None:
        quotation.add_image(OpenpyxlImage(image_path), "B9")
    workbook.save(path)
    workbook.close()


def _canonical_row(
    *,
    item_key: str,
    section_id: str,
    section_title: str,
    origin: str,
    source_row: int | None,
    original_cost: str,
    frozen_rate: str,
    converted_cost: str,
    quantity: str,
    provider: str,
    region: str,
) -> QuotationDataRow:
    return _with_canonical_hash(
        QuotationDataRow(
            item_key=item_key,
            section_id=section_id,
            section_title=section_title,
            position=1,
            origin=origin,
            source_row=source_row,
            original_currency="MXN",
            original_cost=Decimal(original_cost),
            frozen_rate=Decimal(frozen_rate),
            converted_cost=Decimal(converted_cost),
            quantity=Decimal(quantity),
            provider=provider,
            region=region,
            source_hash="a" * 64,
            upstream_row_hash="b" * 64 if origin == "imported" else "",
            row_hash="",
        )
    )


def _mixed_metadata(*, imported_only: bool) -> dict[str, object]:
    return {
        "catalog_price_mode": "mixed_catalog_converted",
        "catalog_source_hashes": {} if imported_only else {"tarkett": "a" * 64},
        "quote_currency": "MXN",
        "descuento": 30,
        "auto_electrification_rate": None,
        "rate_summary": []
        if imported_only
        else [
            {
                "catalog": "tarkett",
                "base_currency": "MXN",
                "quote_currency": "MXN",
                "exchange_rate": "1.000000",
                "rate_source": "identity",
                "rate_effective_date": "2026-07-22",
                "rate_retrieved_at": "",
            }
        ],
    }


def _cell_formula(package: XlsxPackage, sheet_name: str, coordinate: str) -> str | None:
    root = ET.fromstring(package.parts[package.sheet_part(sheet_name)])
    cell = root.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
    if cell is None:
        return None
    formula = cell.find(f"{{{MAIN}}}f")
    return None if formula is None else f"={formula.text or ''}"


def _cell(root: ET.Element, coordinate: str) -> ET.Element:
    result = root.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
    assert result is not None
    return result


def _cell_scalar(root: ET.Element, coordinate: str) -> str | None:
    cell = _cell(root, coordinate)
    if cell.attrib.get("t") == "inlineStr":
        return "".join(
            node.text or ""
            for node in cell.findall(f".//{{{MAIN}}}t")
        )
    return cell.findtext(f"{{{MAIN}}}v")


def _cell_formula_in_root(root: ET.Element, coordinate: str) -> str | None:
    formula = _cell(root, coordinate).find(f"{{{MAIN}}}f")
    return None if formula is None else f"={formula.text or ''}"


def _mobiliti_product_name(workbook, row: int):
    value = workbook["Mobiliti"].cell(row, 4).value
    match = re.fullmatch(r"=Quotation!B([1-9][0-9]*)", str(value or ""))
    if match is None:
        return value
    return workbook["Quotation"].cell(int(match.group(1)), 2).value


def _terms_signature(root: ET.Element, *, start: int, end: int) -> tuple:
    signature = []
    for row_number in range(start, end + 1):
        row = root.find(f"{{{MAIN}}}sheetData/{{{MAIN}}}row[@r='{row_number}']")
        assert row is not None
        row_attributes = tuple(
            sorted((key, value) for key, value in row.attrib.items() if key != "r")
        )
        cells = []
        for cell in row.findall(f"{{{MAIN}}}c"):
            coordinate = cell.attrib["r"]
            column = "".join(character for character in coordinate if character.isalpha())
            if column > "J" or len(column) > 1:
                continue
            formula = cell.find(f"{{{MAIN}}}f")
            value = cell.find(f"{{{MAIN}}}v")
            text = cell.find(f"{{{MAIN}}}is/{{{MAIN}}}t")
            cells.append(
                (
                    column,
                    tuple(sorted((key, val) for key, val in cell.attrib.items() if key != "r")),
                    None if formula is None else (formula.text, tuple(sorted(formula.attrib.items()))),
                    None if value is None else value.text,
                    None if text is None else text.text,
                )
            )
        signature.append((row_attributes, tuple(cells)))
    return tuple(signature)


def _sheet_drawing_parts(
    package: XlsxPackage,
    sheet_name: str,
) -> tuple[str, str]:
    sheet_part = package.sheet_part(sheet_name)
    sheet = ET.fromstring(package.parts[sheet_part])
    drawing = sheet.find(f"{{{MAIN}}}drawing")
    assert drawing is not None
    relationship_id = drawing.attrib[f"{{{OFFICE_REL}}}id"]
    sheet_relationships = ET.fromstring(
        package.parts[relationship_part_name(sheet_part)]
    )
    relationship = next(
        item
        for item in sheet_relationships.findall(f"{{{PACKAGE_REL}}}Relationship")
        if item.attrib["Id"] == relationship_id
    )
    drawing_part = resolve_internal_target(sheet_part, relationship.attrib["Target"])
    return drawing_part, relationship_part_name(drawing_part)


def _cotizacion_drawing_parts(package: XlsxPackage) -> tuple[str, str]:
    return _sheet_drawing_parts(package, "Cotizacion")


def _calc_chain_coordinates_for_sheet(
    payload: bytes,
    sheet_id: int,
) -> list[str]:
    root = ET.fromstring(payload)
    effective_sheet_id: str | None = None
    coordinates: list[str] = []
    for cell in root.findall(f"{{{MAIN}}}c"):
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id == str(sheet_id):
            coordinates.append(cell.attrib["r"])
    return coordinates


def _worksheet_formula_coordinates(payload: bytes) -> list[str]:
    root = ET.fromstring(payload)
    return [
        cell.attrib["r"]
        for cell in root.findall(f".//{{{MAIN}}}c")
        if cell.find(f"{{{MAIN}}}f") is not None
    ]


def _calc_chain_entries_except_sheet(
    payload: bytes,
    sheet_id: int,
) -> tuple[tuple[str | None, bytes], ...]:
    root = ET.fromstring(payload)
    effective_sheet_id: str | None = None
    entries: list[tuple[str | None, bytes]] = []
    for cell in root.findall(f"{{{MAIN}}}c"):
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id != str(sheet_id):
            entries.append((effective_sheet_id, ET.tostring(cell)))
    return tuple(entries)


def _calc_chain_entries_except_sheets(
    payload: bytes,
    sheet_ids: set[int],
) -> tuple[tuple[str | None, bytes], ...]:
    root = ET.fromstring(payload)
    excluded = {str(sheet_id) for sheet_id in sheet_ids}
    effective_sheet_id: str | None = None
    entries: list[tuple[str | None, bytes]] = []
    for cell in root.findall(f"{{{MAIN}}}c"):
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id not in excluded:
            entries.append((effective_sheet_id, ET.tostring(cell)))
    return tuple(entries)


def _calc_chain_metadata_for_sheet(
    payload: bytes,
    sheet_id: int,
) -> dict[str, tuple[tuple[str, str], ...]]:
    root = ET.fromstring(payload)
    effective_sheet_id: str | None = None
    metadata: dict[str, tuple[tuple[str, str], ...]] = {}
    for cell in root.findall(f"{{{MAIN}}}c"):
        if "i" in cell.attrib:
            effective_sheet_id = cell.attrib["i"]
        if effective_sheet_id != str(sheet_id):
            continue
        coordinate = cell.attrib["r"]
        if coordinate in metadata:
            continue
        metadata[coordinate] = tuple(
            sorted(
                (name, value)
                for name, value in cell.attrib.items()
                if name not in {"r", "i"}
            )
        )
    return metadata


def _minimal_request(output: Path) -> ComposeRequest:
    return _request_for_sections(output, (1,))


def _request_for_sections(output: Path, section_sizes: tuple[int, ...]) -> ComposeRequest:
    contract = load_template_contract(CONTRACT_PATH)
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    needs = [
        SectionNeed(f"section-{index}", f"Sección {index}", size)
        for index, size in enumerate(section_sizes, start=1)
    ]
    planned = plan_mobiliti_layout(needs)
    writes: list[MobilitiCellWrite] = []
    cotizacion_sections: list[CotizacionSection] = []
    global_position = 0
    for need, layout in zip(needs, planned.sections, strict=False):
        products: list[CotizacionProduct] = []
        for offset in range(need.item_count):
            global_position += 1
            target_row = layout.product_start + offset
            name = f"Producto {global_position}"
            writes.extend(
                (
                    MobilitiCellWrite(f"D{target_row}", "text", name),
                    MobilitiCellWrite(f"E{target_row}", "text", "Silla"),
                    MobilitiCellWrite(f"F{target_row}", "text", "Proveedor"),
                    MobilitiCellWrite(f"H{target_row}", "number", Decimal("1")),
                    MobilitiCellWrite(
                        f"J{target_row}",
                        "number",
                        Decimal(global_position).quantize(Decimal("0.01")),
                    ),
                    MobilitiCellWrite(f"P{target_row}", "text", "Centro"),
                )
            )
            products.append(
                CotizacionProduct(
                    item_key=f"item-{global_position}",
                    name=name,
                    description=f"Descripción {global_position}",
                    dimensions="60 x 60 cm",
                    quantity=Decimal("1"),
                    mobiliti_row=target_row,
                    discount=Decimal("0.30"),
                )
            )
        cotizacion_sections.append(
            CotizacionSection(title=need.title, products=tuple(products))
        )
    mobiliti = build_mobiliti_sheet(
        base.parts[base.sheet_part("Mobiliti")],
        needs,
        writes,
    )
    selector = WorksheetEditor.from_xml(mobiliti.xml)
    write_official_currency_selector(selector, "MXN", "Guadalajara")
    mobiliti = MobilitiSheetMutation(selector.to_xml(), mobiliti.row_map)

    cotizacion = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=CotizacionMetadata(
            quotation_number="100-00001",
            project="Proyecto compositor",
            client="Cliente de prueba",
            email="cliente@example.test",
            phone="33 0000 0000",
            address="Guadalajara, Jalisco",
            business_name="Cliente SA de CV",
        ),
        sections=tuple(cotizacion_sections),
    )
    return ComposeRequest(
        template=OFFICIAL_TEMPLATE.resolve(),
        output=output.resolve(),
        mobiliti=mobiliti,
        cotizacion=cotizacion,
        quotation=None,
        quotation_data=build_quotation_data_sheet(()),
        contract=contract,
    )


@pytest.mark.parametrize("product_count", (1, 2, 9))
def test_calc_chain_prunes_mutated_sheets_and_forces_full_recalculation(
    tmp_path: Path,
    product_count: int,
) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _request_for_sections(
        tmp_path / f"calc-chain-{product_count}.xlsx",
        (product_count,),
    )

    mutation = build_allowlisted_mutation(base, request)

    calc_part = official_composer_module._calc_chain_part(base)
    assert calc_part is not None
    cotizacion_sheet_id = official_composer_module._workbook_sheet_id(
        base, "Cotizacion"
    )
    mobiliti_sheet_id = official_composer_module._workbook_sheet_id(base, "Mobiliti")
    cotizacion_part = base.sheet_part("Cotizacion")
    formula_coordinates = _worksheet_formula_coordinates(
        mutation.replacements[cotizacion_part]
    )
    assert len(formula_coordinates) == len(set(formula_coordinates))
    assert {"P21", "L121"}.issubset(formula_coordinates)
    assert _calc_chain_coordinates_for_sheet(
        mutation.replacements[calc_part],
        cotizacion_sheet_id,
    ) == []
    assert _calc_chain_coordinates_for_sheet(
        mutation.replacements[calc_part],
        mobiliti_sheet_id,
    ) == []
    mutated_sheet_ids = {cotizacion_sheet_id, mobiliti_sheet_id}
    assert _calc_chain_entries_except_sheets(
        mutation.replacements[calc_part],
        mutated_sheet_ids,
    ) == _calc_chain_entries_except_sheets(
        base.parts[calc_part],
        mutated_sheet_ids,
    )

    before_workbook = ET.fromstring(base.parts["xl/workbook.xml"])
    after_workbook = ET.fromstring(mutation.replacements["xl/workbook.xml"])
    before_calc = before_workbook.find(f"{{{MAIN}}}calcPr")
    after_calc = after_workbook.find(f"{{{MAIN}}}calcPr")
    assert before_calc is not None and after_calc is not None
    expected_attributes = dict(before_calc.attrib)
    expected_attributes.update(
        {"calcMode": "auto", "fullCalcOnLoad": "1", "forceFullCalc": "1"}
    )
    assert after_calc.attrib == expected_attributes


def test_mutated_xml_preserves_markup_compatibility_prefixes(tmp_path: Path) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "mc-prefixes.xlsx")

    mutation = build_allowlisted_mutation(base, request)

    parts = {
        "xl/workbook.xml",
        base.sheet_part("Cotizacion"),
        base.sheet_part("Mobiliti"),
        base.sheet_part("Fletes"),
        base.sheet_part("Estrategia Comercial "),
    }
    for part in parts:
        _parse_xml_document(
            mutation.replacements[part],
            f"XML mutado invalido: {part}",
        )


def test_composer_preserves_protected_official_package_and_updates_dependents(
    tmp_path: Path,
) -> None:
    output = tmp_path / "quote.xlsx"
    request = _minimal_request(output)

    audit = compose_official_quote(request)

    assert audit.unexpected_changed_parts == frozenset()
    result = XlsxPackage.read(output)
    assert result.sheet_state("Fletes") == "visible"
    assert result.sheet_state("Quotation_Data") == "veryHidden"
    with pytest.raises(KeyError):
        result.sheet_state("sheep")
    assert sum(name.startswith("xl/externalLinks/") for name in result.parts) == 12
    assert sum(
        len(
            ET.fromstring(result.parts[part]).findall(f".//{{{MAIN}}}f")
        )
        for name, _state, _index, part in result._sheet_rows()
        if name.strip().casefold().startswith("spec")
    ) == 1314
    workbook = ET.fromstring(result.parts["xl/workbook.xml"])
    assert len(
        workbook.findall(f"{{{MAIN}}}definedNames/{{{MAIN}}}definedName")
    ) == 29
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    for prefix in request.contract.protected_prefixes:
        for name, payload in base.parts.items():
            if name.startswith(prefix):
                assert result.parts[name] == payload
    assert _cell_formula(result, "Fletes", "D19") == (
        f"=Mobiliti!H{request.mobiliti.row_map.total_row}"
    )
    assert _cell_formula(result, "Estrategia Comercial ", "D59") == (
        f"=Cotizacion!H{request.cotizacion.total_row}"
    )


def test_composer_restores_flete_routes_from_flete_routes(
    tmp_path: Path,
) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "flete-routes.xlsx")

    mutation = build_allowlisted_mutation(base, request)

    fletes = ET.fromstring(
        mutation.replacements[base.sheet_part("Fletes")]
    )
    for row, (origin, destination) in engine_module.FLETE_ROUTES.items():
        assert _cell_scalar(fletes, f"A{row}") == origin
        assert _cell_scalar(fletes, f"C{row}") == destination


def test_composer_populates_missing_flete_and_installation_categories(
    tmp_path: Path,
) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "flete-categories.xlsx")

    mutation = build_allowlisted_mutation(base, request)

    fletes = ET.fromstring(
        mutation.replacements[base.sheet_part("Fletes")]
    )
    assert _cell_scalar(fletes, "I8") == "Escritorios-WorkStation"
    assert _cell_scalar(fletes, "M8") == "Escritorios-WorkStation"
    expected = {
        16: ("Bancos", Decimal("0.2"), Decimal("56")),
        17: ("Cocineta", Decimal("0.2"), Decimal("1790")),
        18: ("Pizarrones", Decimal("0.2"), Decimal("210")),
    }
    for row, (category, factor, installation) in expected.items():
        assert _cell_scalar(fletes, f"I{row}") == category
        assert Decimal(_cell_scalar(fletes, f"J{row}")) == factor
        assert _cell_formula_in_root(fletes, f"K{row}") == (
            f"=IF(Mobiliti!$K$4=TRUE,"
            f"((J{row}/$J$21)*Mobiliti!$P$9)/Mobiliti!$K$6,"
            f"(J{row}/$J$21)*Mobiliti!$P$9)"
        )
        assert _cell_scalar(fletes, f"M{row}") == category
        assert _cell_formula_in_root(fletes, f"N{row}") == (
            f"=IF(Mobiliti!$K$4=TRUE,"
            f"({format(installation, 'f')}/Mobiliti!$K$6),"
            f"{format(installation, 'f')})"
        )


def test_cotizacion_clears_contamination_uses_master_discount_and_keeps_terms(
    tmp_path: Path,
) -> None:
    request = _request_for_sections(tmp_path / "two-products.xlsx", (2,))
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    official = ET.fromstring(base.parts[base.sheet_part("Cotizacion")])
    candidate = ET.fromstring(request.cotizacion.xml)
    first_row, second_row = request.cotizacion.product_rows

    first_discount = _cell(candidate, f"G{first_row}")
    second_discount = _cell(candidate, f"G{second_row}")
    assert first_discount.find(f"{{{MAIN}}}f") is None
    assert Decimal(first_discount.findtext(f"{{{MAIN}}}v")) == Decimal("0.30")
    assert second_discount.findtext(f"{{{MAIN}}}f") == f"$G${first_row}"
    for row in (first_row, second_row):
        assert _cell(candidate, f"A{row}").attrib["t"] == "inlineStr"
        assert _cell(candidate, f"C{row}").attrib["t"] == "inlineStr"
        assert _cell(candidate, f"D{row}").attrib["t"] == "inlineStr"
        assert "#REF!" not in ET.tostring(
            candidate.find(f"{{{MAIN}}}sheetData/{{{MAIN}}}row[@r='{row}']"),
            encoding="unicode",
        )

    delta = request.cotizacion.terms_row_delta
    assert _terms_signature(official, start=28, end=76) == _terms_signature(
        candidate,
        start=28 + delta,
        end=76 + delta,
    )


def test_composer_handles_twenty_sections_and_one_hundred_product_section(
    tmp_path: Path,
) -> None:
    sizes = (100, *(1 for _ in range(19)))
    output = tmp_path / "large-dynamic.xlsx"
    request = _request_for_sections(output, sizes)

    compose_official_quote(request)

    package = XlsxPackage.read(output)
    assert len(request.mobiliti.row_map.sections) == 20
    assert request.mobiliti.row_map.sections[0].capacity == 100
    assert _cell_formula(package, "Fletes", "D19") == (
        f"=Mobiliti!H{request.mobiliti.row_map.total_row}"
    )
    estrategia = ET.fromstring(
        package.parts[package.sheet_part("Estrategia Comercial ")]
    )
    assert f"${request.mobiliti.row_map.last_product_row}" in (
        _cell(estrategia, "B7").findtext(f"{{{MAIN}}}f") or ""
    )
    assert _cell_formula(package, "Estrategia Comercial ", "D59") == (
        f"=Cotizacion!H{request.cotizacion.total_row}"
    )


def test_composer_moves_only_mobiliti_auxiliary_shapes_after_section_sixteen(
    tmp_path: Path,
) -> None:
    output = tmp_path / "seventeen-sections-drawing.xlsx"
    request = _request_for_sections(output, tuple(1 for _ in range(17)))
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    base_drawing_part, _base_relationships = _sheet_drawing_parts(base, "Mobiliti")
    base_drawing = ET.fromstring(base.parts[base_drawing_part])

    compose_official_quote(request)

    result = XlsxPackage.read(output)
    result_drawing_part, _result_relationships = _sheet_drawing_parts(
        result,
        "Mobiliti",
    )
    result_drawing = ET.fromstring(result.parts[result_drawing_part])
    delta = request.mobiliti.row_map.total_row - 573

    def anchors_by_name(root: ET.Element) -> dict[str, tuple[int, int]]:
        rows: dict[str, tuple[int, int]] = {}
        for anchor in root:
            properties = anchor.find(f".//{{{XDR}}}cNvPr")
            if properties is None:
                continue
            start = anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}row")
            end = anchor.findtext(f"{{{XDR}}}to/{{{XDR}}}row")
            if start is not None and end is not None:
                rows[properties.attrib["name"]] = (int(start), int(end))
        return rows

    before = anchors_by_name(base_drawing)
    after = anchors_by_name(result_drawing)
    for name in ("Straight Arrow Connector 1", "Straight Arrow Connector 7"):
        assert after[name] == tuple(value + delta for value in before[name])
    for name in ("Straight Arrow Connector 9", "Imagen 2"):
        assert after[name] == before[name]


def test_composer_preserves_static_drawing_and_adds_safe_product_png(
    tmp_path: Path,
) -> None:
    image_path = tmp_path / "product.png"
    Image.new("RGB", (80, 40), (25, 120, 210)).save(image_path)
    output = tmp_path / "with-image.xlsx"
    request = _minimal_request(output)
    cotizacion = replace(
        request.cotizacion,
        images=(
            CotizacionProductImage(
                image_path.resolve(),
                request.cotizacion.product_rows[0],
            ),
        ),
    )

    compose_official_quote(replace(request, cotizacion=cotizacion))

    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    result = XlsxPackage.read(output)
    base_drawing, base_rels = _cotizacion_drawing_parts(base)
    result_drawing, result_rels = _cotizacion_drawing_parts(result)
    assert result_drawing == base_drawing
    base_root = ET.fromstring(base.parts[base_drawing])
    result_root = ET.fromstring(result.parts[result_drawing])
    assert len(list(base_root)) == 5
    assert len(list(result_root)) == len(list(base_root)) + 1
    product_anchor = result_root.findall(f"{{{XDR}}}oneCellAnchor")[-1]
    assert product_anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}row") == str(
        request.cotizacion.product_rows[0] - 1
    )

    before_relationships = {
        item.attrib["Id"]: dict(item.attrib)
        for item in ET.fromstring(base.parts[base_rels]).findall(
            f"{{{PACKAGE_REL}}}Relationship"
        )
    }
    after_relationships = {
        item.attrib["Id"]: dict(item.attrib)
        for item in ET.fromstring(result.parts[result_rels]).findall(
            f"{{{PACKAGE_REL}}}Relationship"
        )
    }
    assert all(after_relationships[key] == value for key, value in before_relationships.items())
    new_relationships = set(after_relationships) - set(before_relationships)
    assert len(new_relationships) == 1
    new_target = after_relationships[new_relationships.pop()]["Target"]
    new_media = resolve_internal_target(result_drawing, new_target)
    assert result.parts[new_media].startswith(b"\x89PNG\r\n\x1a\n")


def test_composer_uses_task7_numeric_price_once_for_import_catalog_and_lumbro(
    tmp_path: Path,
) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    need = SectionNeed("section-prices", "Precios", 3)
    row_map = plan_mobiliti_layout((need,))
    definitions = (
        ("catalog", None, "", Decimal("100"), Decimal("1"), Decimal("100.00")),
        ("imported", 9, "b" * 64, Decimal("10"), Decimal("18.5"), Decimal("185.00")),
        (
            "lumbro",
            None,
            "",
            Decimal("100"),
            Decimal("0.054054"),
            lumbro_frozen_cost(Decimal("100"), Decimal("0.054054")),
        ),
    )
    canonical_rows = tuple(
        _with_canonical_hash(
            QuotationDataRow(
                item_key=f"item-{position}",
                section_id=need.id,
                section_title=need.title,
                position=position,
                origin=origin,
                source_row=source_row,
                original_currency="USD" if origin != "lumbro" else "MXN",
                original_cost=original,
                frozen_rate=rate,
                converted_cost=converted,
                quantity=Decimal("1"),
                provider="Proveedor",
                region="Centro",
                source_hash="a" * 64,
                upstream_row_hash=upstream,
                row_hash="",
            )
        )
        for position, (origin, source_row, upstream, original, rate, converted) in enumerate(
            definitions,
            start=1,
        )
    )
    bindings = tuple(
        PricingRowBinding(
            item_key=row.item_key,
            section_id=row.section_id,
            position=row.position,
            target_row=target_row,
        )
        for row, target_row in zip(canonical_rows, row_map.item_rows, strict=True)
    )
    input_writes: list[MobilitiCellWrite] = []
    for row, target_row in zip(canonical_rows, row_map.item_rows, strict=True):
        input_writes.extend(
            (
                MobilitiCellWrite(f"D{target_row}", "text", row.item_key),
                MobilitiCellWrite(f"E{target_row}", "text", "Silla"),
                MobilitiCellWrite(f"F{target_row}", "text", row.provider),
                MobilitiCellWrite(f"H{target_row}", "number", row.quantity),
                MobilitiCellWrite(f"P{target_row}", "text", row.region),
            )
        )
    input_writes.extend(
        build_mobiliti_pricing_writes(
            canonical_rows,
            row_map,
            bindings=bindings,
        )
    )
    mobiliti = build_mobiliti_sheet(
        base.parts[base.sheet_part("Mobiliti")],
        [need],
        input_writes,
    )
    selector = WorksheetEditor.from_xml(mobiliti.xml)
    write_official_currency_selector(selector, "USD", "Monterrey")
    mobiliti = MobilitiSheetMutation(selector.to_xml(), mobiliti.row_map)
    cotizacion = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=CotizacionMetadata(quotation_number="PRICE-001"),
        sections=(
            CotizacionSection(
                title=need.title,
                products=tuple(
                    CotizacionProduct(
                        item_key=row.item_key,
                        name=row.item_key,
                        description=row.origin,
                        dimensions="",
                        quantity=row.quantity,
                        mobiliti_row=target_row,
                    )
                    for row, target_row in zip(
                        canonical_rows,
                        row_map.item_rows,
                        strict=True,
                    )
                ),
            ),
        ),
    )
    output = tmp_path / "all-price-origins.xlsx"
    compose_official_quote(
        ComposeRequest(
            template=OFFICIAL_TEMPLATE.resolve(),
            output=output.resolve(),
            mobiliti=mobiliti,
            cotizacion=cotizacion,
            quotation=None,
            quotation_data=build_quotation_data_sheet(canonical_rows),
            contract=load_template_contract(CONTRACT_PATH),
        )
    )

    result = XlsxPackage.read(output)
    official_mobiliti = ET.fromstring(base.parts[base.sheet_part("Mobiliti")])
    result_mobiliti = ET.fromstring(result.parts[result.sheet_part("Mobiliti")])
    for row, target_row in zip(canonical_rows, row_map.item_rows, strict=True):
        price = _cell(result_mobiliti, f"J{target_row}")
        assert price.find(f"{{{MAIN}}}f") is None
        assert Decimal(price.findtext(f"{{{MAIN}}}v")) == row.converted_cost
        assert _cell(result_mobiliti, f"W{target_row}").find(f"{{{MAIN}}}f") is not None
        assert _cell(result_mobiliti, f"X{target_row}").find(f"{{{MAIN}}}f") is not None
    assert ET.tostring(_cell(result_mobiliti, "K6")) == ET.tostring(
        _cell(official_mobiliti, "K6")
    )


def test_composer_rejects_a_mobiliti_write_outside_the_declared_surface(
    tmp_path: Path,
) -> None:
    output = tmp_path / "tampered.xlsx"
    request = _minimal_request(output)
    root = ET.fromstring(request.mobiliti.xml)
    cell = root.find(f".//{{{MAIN}}}c[@r='A1']")
    if cell is None:
        sheet_data = root.find(f"{{{MAIN}}}sheetData")
        assert sheet_data is not None
        row = sheet_data.find(f"{{{MAIN}}}row[@r='1']")
        if row is None:
            row = ET.SubElement(sheet_data, f"{{{MAIN}}}row", {"r": "1"})
        cell = ET.SubElement(row, f"{{{MAIN}}}c", {"r": "A1"})
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{MAIN}}}is")
    ET.SubElement(inline, f"{{{MAIN}}}t").text = "alteración fuera de contrato"
    tampered = MobilitiSheetMutation(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        request.mobiliti.row_map,
    )

    with pytest.raises(ValueError, match="fuera de la superficie mutable"):
        compose_official_quote(replace(request, mobiliti=tampered))

    assert not output.exists()


def test_composer_rejects_nonofficial_template_before_creating_output(
    tmp_path: Path,
) -> None:
    template = tmp_path / "tampered-template.xlsx"
    shutil.copyfile(OFFICIAL_TEMPLATE, template)
    with template.open("ab") as stream:
        stream.write(b"tamper")
    output = tmp_path / "must-not-exist.xlsx"
    request = replace(_minimal_request(output), template=template.resolve())

    with pytest.raises(ValueError, match="Plantilla oficial incompatible: sha256"):
        compose_official_quote(request)

    assert not output.exists()


def test_composer_rejects_existing_or_relative_output_without_overwrite(
    tmp_path: Path,
) -> None:
    existing = tmp_path / "reserved.xlsx"
    existing.write_bytes(b"contenido del usuario")

    with pytest.raises(FileExistsError, match="salida ya existe"):
        compose_official_quote(_minimal_request(existing))
    assert existing.read_bytes() == b"contenido del usuario"

    relative_request = replace(
        _minimal_request(tmp_path / "absolute.xlsx"),
        output=Path("relative.xlsx"),
    )
    with pytest.raises(ValueError, match="ruta de salida debe ser absoluta"):
        compose_official_quote(relative_request)
    assert not (tmp_path / "absolute.xlsx").exists()


def test_composer_rejects_unsafe_product_image_before_creating_output(
    tmp_path: Path,
) -> None:
    image = tmp_path / "product.svg"
    image.write_text("<svg xmlns='http://www.w3.org/2000/svg'/>", encoding="utf-8")
    output = tmp_path / "unsafe-image.xlsx"
    request = _minimal_request(output)
    cotizacion = replace(
        request.cotizacion,
        images=(CotizacionProductImage(image.resolve(), 17),),
    )

    with pytest.raises(ValueError, match="Formato de imagen.*no permitido"):
        compose_official_quote(replace(request, cotizacion=cotizacion))

    assert not output.exists()


def test_cotizacion_product_requires_nonempty_identity_fields() -> None:
    with pytest.raises(ValueError, match="Texto de Cotizacion vacío"):
        CotizacionProduct(
            item_key="",
            name="Silla",
            description="",
            dimensions="",
            quantity=Decimal("1"),
            mobiliti_row=14,
        )


@pytest.mark.parametrize(
    "formula",
    (
        "=Quotation!$D$9",
        "=Quotation!E9",
        "=Quotation!D1048577",
        "=[externo.xlsx]Quotation!D9",
        '=HYPERLINK("https://invalid.example","abrir")',
    ),
)
def test_cotizacion_product_rejects_description_formulas_outside_quotation(
    formula: str,
) -> None:
    with pytest.raises(ValueError, match="fuera de contrato"):
        CotizacionProduct(
            item_key="item-1",
            name="Silla",
            description="Descripción procesada",
            dimensions="600 x 600 mm",
            quantity=Decimal("1"),
            mobiliti_row=14,
            description_formula=formula,
        )

    valid = CotizacionProduct(
        item_key="item-1",
        name="Silla",
        description="Descripción procesada",
        dimensions="600 x 600 mm",
        quantity=Decimal("1"),
        mobiliti_row=14,
        description_formula="=Quotation!D9",
    )
    assert valid.description_formula == "=Quotation!D9"


def test_active_engine_routes_through_official_composer_without_legacy_writers(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    quotation = workbook.active
    quotation.title = "Quotation"
    for column, value in {
        1: "No.",
        2: "Item",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
    }.items():
        quotation.cell(7, column).value = value
    quotation["A8"] = "- Sillas"
    quotation["A9"] = 1
    quotation["B9"] = "Silla de prueba"
    quotation["D9"] = "Silla operativa"
    quotation["E9"] = "60 x 60 cm"
    quotation["G9"] = 2
    quotation["J9"] = 125.50
    workbook.save(source)
    workbook.close()

    output = tmp_path / "engine.xlsx"
    result = generate_quote(
        source,
        output,
        {"cotizacion": "100-ENGINE", "proyecto": "Ruta oficial"},
        OFFICIAL_TEMPLATE,
    )

    assert result == output
    package = XlsxPackage.read(output)
    assert package.sheet_state("Quotation_Data") == "veryHidden"
    assert package.sheet_part("Quotation")


def test_active_engine_rejects_delivery_place_missing_from_flete_routes(
    tmp_path: Path,
) -> None:
    source = tmp_path / "unsupported-delivery.xlsx"
    _write_engine_source(
        source,
        (
            {
                "name": "Silla de prueba",
                "description": "Silla operativa",
                "quantity": 1,
                "price": 125.50,
            },
        ),
    )
    output = tmp_path / "unsupported-delivery-output.xlsx"

    with pytest.raises(ValueError, match="Lugar de entrega no soportado"):
        generate_quote(
            source,
            output,
            {
                "cotizacion": "100-DESTINO",
                "lugar_entrega": "Destino inexistente",
            },
            OFFICIAL_TEMPLATE,
        )

    assert not output.exists()


def test_active_engine_references_visible_quotation_and_mobiliti_fields(
    tmp_path: Path,
) -> None:
    source = tmp_path / "m3-source.xlsx"
    original_description = (
        "Mesa operativa original del proveedor | Fuente: celda D9 "
        "| Hash fuente: abc123 | URL: https://example.com/producto"
    )
    source_workbook = Workbook()
    source_quotation = source_workbook.active
    source_quotation.title = "Quotation"
    headers = (
        "No.",
        "Item Name",
        "Photo",
        "Description",
        "Dimension",
        "Color",
        "Q'ty",
        "Vol.",
        "Tot.Vol.",
        "Unit Price",
        " Tot.Price",
        "Remark",
    )
    for column, header in enumerate(headers, start=1):
        source_quotation.cell(7, column).value = header
    source_quotation["A8"] = "- Sillas"
    source_quotation["A9"] = 1
    source_quotation["B9"] = "Mesa con volumen"
    source_quotation["D9"] = original_description
    source_quotation["E9"] = "240 x 120 cm"
    source_quotation["F9"] = "Nogal"
    source_quotation["G9"] = 3
    source_quotation["H9"] = 0.42
    source_quotation["I9"] = "=G9*H9"
    source_quotation["J9"] = 1000
    source_quotation["K9"] = "=G9*J9"
    source_quotation["L9"] = "Entrega 8 semanas"
    source_quotation["A10"] = 2
    source_quotation["B10"] = "Silla sin volumen"
    source_quotation["D10"] = "Silla operativa original"
    source_quotation["E10"] = "60 x 60 x 90 cm"
    source_quotation["G10"] = 1
    source_quotation["H10"] = 0
    source_quotation["I10"] = "=G10*H10"
    source_quotation["J10"] = 250
    source_quotation["K10"] = "=G10*J10"
    source_workbook.save(source)
    source_workbook.close()
    output = tmp_path / "m3-official.xlsx"

    generate_quote(source, output, {"cotizacion": "M3-E2E"}, OFFICIAL_TEMPLATE)

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        quotation = workbook["Quotation"]
        mobiliti = workbook["Mobiliti"]
        cotizacion = workbook["Cotizacion"]
        quotation_data = workbook["Quotation_Data"]
        assert tuple(
            quotation.cell(7, column).value for column in range(4, 14)
        ) == (
            "Trasformacion",
            "Description",
            "Dimension",
            "Color",
            "Q'ty",
            "Vol.",
            "Tot.Vol.",
            "Unit Price",
            " Tot.Price",
            "Remark",
        )
        assert "Mesa operativa original del proveedor" in quotation["D9"].value
        assert "Fuente:" not in quotation["D9"].value
        assert "Hash fuente:" not in quotation["D9"].value
        assert "https://" not in quotation["D9"].value
        assert quotation["E9"].value == original_description
        assert quotation["F9"].value == "240 x 120 cm"
        assert quotation["H9"].value == 3
        assert quotation["I9"].value == pytest.approx(0.42)
        assert quotation["J9"].value == "=H9*I9"
        assert quotation["K9"].value == 1000
        assert quotation["L9"].value == "=H9*K9"
        assert quotation_data.max_column == 16
        assert mobiliti["D14"].value == "=Quotation!B9"
        assert mobiliti["H14"].value == "=Quotation!H9"
        assert mobiliti["J14"].value == "=Quotation!K9"
        assert mobiliti["K14"].value == "=Quotation!I9"
        assert mobiliti["D15"].value == "=Quotation!B10"
        assert mobiliti["H15"].value == "=Quotation!H10"
        assert mobiliti["J15"].value == "=Quotation!K10"
        assert mobiliti["K15"].value == "=Quotation!I10"
        assert cotizacion["A17"].value == "=Mobiliti!D14"
        assert cotizacion["A18"].value == "=Mobiliti!D15"
        assert cotizacion["C17"].value == "=Quotation!D9"
        assert cotizacion["C18"].value == "=Quotation!D10"
        assert cotizacion["D17"].value == "=Quotation!F9"
        assert cotizacion["D18"].value == "=Quotation!F10"
        assert cotizacion["E17"].value == "=Mobiliti!H14"
        assert cotizacion["E18"].value == "=Mobiliti!H15"
    finally:
        workbook.close()


def test_active_engine_renders_each_lumbro_accessory_once_and_includes_its_cost(
    tmp_path: Path,
) -> None:
    source = tmp_path / "workstation.xlsx"
    _write_engine_source(
        source,
        (
            {
                "name": "Workstation 2 pax",
                "description": "Estación para dos usuarios",
                "dimension": "240 x 120 cm",
                "quantity": 1,
                "price": 1000,
            },
        ),
        category="Workstations",
    )
    output = tmp_path / "workstation-official.xlsx"

    generate_quote(source, output, {"cotizacion": "LUMBRO-E2E"}, OFFICIAL_TEMPLATE)

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        cotizacion = workbook["Cotizacion"]
        expected_names = (
            "Workstation 2 pax",
            "LIDO.OP-INT",
            "JUMP-1.5M",
            "CAJA-FUS",
        )
        rows_by_name = {name: [] for name in expected_names}
        for row in range(16, cotizacion.max_row + 1):
            product_formula = cotizacion.cell(row, 1).value
            if (
                not isinstance(product_formula, str)
                or re.fullmatch(r"=Mobiliti!D[1-9][0-9]*", product_formula) is None
            ):
                continue
            mobiliti_row = int(product_formula.removeprefix("=Mobiliti!D"))
            product_name = _mobiliti_product_name(workbook, mobiliti_row)
            if product_name in rows_by_name:
                rows_by_name[product_name].append(row)
        assert all(len(rows) == 1 for rows in rows_by_name.values())

        visible_rows = tuple(rows_by_name[name][0] for name in expected_names)
        mobiliti_rows = []
        for row in visible_rows:
            formula = cotizacion.cell(row, 6).value
            assert isinstance(formula, str) and formula.startswith("=Mobiliti!X")
            mobiliti_rows.append(int(formula.removeprefix("=Mobiliti!X")))
            assert cotizacion.cell(row, 10).value == f"=E{row}*I{row}"
        assert len(set(mobiliti_rows)) == len(expected_names)

        total_row = next(
            row
            for row in range(max(visible_rows) + 1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 4).value == "TOTAL:"
        )
        subtotal_value = cotizacion.cell(total_row - 4, 8).value
        subtotal_formula = getattr(subtotal_value, "text", subtotal_value)
        assert subtotal_formula == (
            f"=SUM(IFERROR(J{min(visible_rows)}:J{max(visible_rows)},0))"
        )

        audit = workbook["Quotation_Data"]
        canonical = [
            tuple(audit.cell(row, column).value for column in range(1, 14))
            for row in range(2, audit.max_row + 1)
        ]
        assert len(canonical) == len(expected_names)
        parent_key = str(canonical[0][0])
        for code, row, mobiliti_row in zip(
            expected_names[1:],
            canonical[1:],
            mobiliti_rows[1:],
            strict=True,
        ):
            assert parent_key in str(row[0])
            assert code in str(row[0])
            assert row[1] == canonical[0][1]
            cost_formula = workbook["Mobiliti"].cell(mobiliti_row, 10).value
            assert isinstance(cost_formula, str)
            assert re.fullmatch(r"=Quotation!K[1-9][0-9]*", cost_formula)
            quotation_row = int(cost_formula.removeprefix("=Quotation!K"))
            assert workbook["Quotation"].cell(
                quotation_row,
                11,
            ).value == pytest.approx(float(row[9]))
    finally:
        workbook.close()


def test_active_engine_uses_authoritative_catalog_canonical_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "catalog-source.xlsx"
    _write_engine_source(
        source,
        (
            {
                "name": "Silla catalogada",
                "description": "Silla de catálogo",
                "dimension": "60 x 60 cm",
                "quantity": 2,
                "price": 125.50,
                "provider": "Tarkett",
                "discount": 0,
                "currency": "MXN",
                "original_price": 125.50,
                "rate": 1,
                "reference": "catalog:tarkett:silla-1",
                "mode": "net",
                "auto": False,
            },
        ),
    )
    canonical = _canonical_row(
        item_key="catalog:tarkett:silla-1",
        section_id="catalog-section:chairs",
        section_title="Sillas canónicas",
        origin="tarkett",
        source_row=None,
        original_cost="125.50",
        frozen_rate="1",
        converted_cost="125.50",
        quantity="2",
        provider="Tarkett MX",
        region="tarkett",
    )
    output = tmp_path / "catalog-official.xlsx"

    def legacy_validator_forbidden(*_args, **_kwargs):
        raise AssertionError("el handoff canónico no depende del validador legacy")

    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine._validate_mixed_catalog_metadata",
        legacy_validator_forbidden,
    )

    generate_quote(
        source,
        output,
        _mixed_metadata(imported_only=False),
        OFFICIAL_TEMPLATE,
        quotation_data_rows=(canonical,),
    )

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        audit = workbook["Quotation_Data"]
        assert audit["A2"].value == canonical.item_key
        assert audit["B2"].value == canonical.section_id
        assert audit["C2"].value == canonical.section_title
        assert audit["P2"].value == canonical.row_hash
        cotizacion = workbook["Cotizacion"]
        assert any(
            cotizacion.cell(row, 1).value == canonical.section_title
            for row in range(16, cotizacion.max_row + 1)
        )
    finally:
        workbook.close()


def test_active_engine_uses_authoritative_imported_canonical_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "imported-source.xlsx"
    _write_engine_source(
        source,
        (
            {
                "name": "Producto importado",
                "description": "Línea importada",
                "dimension": "pieza",
                "quantity": 3,
                "price": 200.25,
                "provider": "Proveedor Externo",
                "discount": 0,
                "currency": "MXN",
                "original_price": 200.25,
                "rate": 1,
                "reference": "import:quote-77:9",
                "mode": "imported",
                "auto": False,
            },
        ),
        category="Importados",
    )
    canonical = _canonical_row(
        item_key="import:quote-77:9",
        section_id="import-section:quote-77",
        section_title="Importación auditada",
        origin="imported",
        source_row=9,
        original_cost="200.25",
        frozen_rate="1",
        converted_cost="200.25",
        quantity="3",
        provider="Proveedor Externo",
        region="imported",
    )
    output = tmp_path / "imported-official.xlsx"

    def legacy_validator_forbidden(*_args, **_kwargs):
        raise AssertionError("el handoff imported no depende del validador legacy")

    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine._validate_mixed_catalog_metadata",
        legacy_validator_forbidden,
    )

    generate_quote(
        source,
        output,
        _mixed_metadata(imported_only=True),
        OFFICIAL_TEMPLATE,
        quotation_data_rows=(canonical,),
    )

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        audit = workbook["Quotation_Data"]
        assert tuple(audit.cell(2, column).value for column in (1, 2, 3, 6, 16)) == (
            canonical.item_key,
            canonical.section_id,
            canonical.section_title,
            canonical.source_row,
            canonical.row_hash,
        )
        assert workbook["Quotation"]["E9"].value == "Línea importada"
        assert "Producto importado" in workbook["Quotation"]["D9"].value
        assert "Línea importada" in workbook["Quotation"]["D9"].value
        cotizacion = workbook["Cotizacion"]
        product_row = next(
            row
            for row in range(16, cotizacion.max_row + 1)
            if cotizacion.cell(row, 1).value == "=Mobiliti!D14"
        )
        assert cotizacion.cell(product_row, 3).value == "=Quotation!D9"
    finally:
        workbook.close()


def test_active_engine_rejects_nonempty_canonical_mismatch_without_output(
    tmp_path: Path,
) -> None:
    source = tmp_path / "mismatch-source.xlsx"
    _write_engine_source(
        source,
        (
            {
                "name": "Silla",
                "quantity": 1,
                "price": 100,
                "reference": "catalog:sunon:item-1",
            },
        ),
    )
    canonical = _canonical_row(
        item_key="catalog:item-1",
        section_id="catalog-section:1",
        section_title="Sillas",
        origin="sunon",
        source_row=None,
        original_cost="100",
        frozen_rate="1",
        converted_cost="100.00",
        quantity="99",
        provider="Sunon Inc",
        region="sunon",
    )
    output = tmp_path / "must-not-exist.xlsx"

    with pytest.raises(ValueError, match="canónic.*no coincide|canónica.*inconsistente"):
        generate_quote(
            source,
            output,
            {"catalog_source_hashes": {"sunon": "a" * 64}},
            OFFICIAL_TEMPLATE,
            quotation_data_rows=(canonical,),
        )

    assert not output.exists()


def test_active_engine_explicit_empty_canonical_rows_use_compatible_fallback(
    tmp_path: Path,
) -> None:
    source = tmp_path / "empty-canonical-source.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla fallback", "quantity": 1, "price": 100},),
    )
    output = tmp_path / "empty-canonical-output.xlsx"

    generate_quote(
        source,
        output,
        {},
        OFFICIAL_TEMPLATE,
        quotation_data_rows=(),
    )

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert workbook["Quotation_Data"]["A2"].value == "quotation:9"
    finally:
        workbook.close()


def test_active_engine_omitted_original_keeps_legacy_visible_quotation(
    tmp_path: Path,
) -> None:
    source = tmp_path / "omitted-original.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla visible", "quantity": 1, "price": 100},),
    )
    output = tmp_path / "omitted-original-output.xlsx"

    generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    package = XlsxPackage.read(output)
    assert package.sheet_part("Quotation")


def test_active_engine_explicit_none_keeps_visible_quotation_for_live_formulas(
    tmp_path: Path,
) -> None:
    source = tmp_path / "none-original.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla sin original", "quantity": 1, "price": 100},),
    )
    output = tmp_path / "none-original-output.xlsx"

    generate_quote(
        source,
        output,
        {},
        OFFICIAL_TEMPLATE,
        original_quotation_path=None,
    )

    package = XlsxPackage.read(output)
    assert package.sheet_part("Quotation")
    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert workbook["Mobiliti"]["D14"].value == "=Quotation!B9"
        assert workbook["Mobiliti"]["J14"].value == "=Quotation!K9"
    finally:
        workbook.close()


def test_active_engine_explicit_original_transplants_that_workbook(
    tmp_path: Path,
) -> None:
    source = tmp_path / "generated-source.xlsx"
    original = tmp_path / "customer-original.xlsx"
    _write_engine_source(
        source,
        ({"name": "Producto generado", "quantity": 1, "price": 100},),
    )
    _write_engine_source(
        original,
        ({"name": "Producto original visible", "quantity": 7, "price": 999},),
        category="Original cliente",
    )
    original_workbook = load_workbook(original)
    try:
        original_workbook["Quotation"]["A1"] = "Encabezado exclusivo del cliente"
        original_workbook.save(original)
    finally:
        original_workbook.close()
    output = tmp_path / "explicit-original-output.xlsx"

    generate_quote(
        source,
        output,
        {},
        OFFICIAL_TEMPLATE,
        original_quotation_path=original,
    )

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert (
            workbook["Quotation"]["A1"].value
            == "Encabezado exclusivo del cliente"
        )
        assert workbook["Quotation"]["B9"].value == "Producto generado"
        cotizacion_formula = next(
            workbook["Cotizacion"].cell(row, 1).value
            for row in range(16, workbook["Cotizacion"].max_row + 1)
            if isinstance(workbook["Cotizacion"].cell(row, 1).value, str)
            and re.fullmatch(
                r"=Mobiliti!D[1-9][0-9]*",
                workbook["Cotizacion"].cell(row, 1).value,
            )
            is not None
        )
        mobiliti_row = int(cotizacion_formula.removeprefix("=Mobiliti!D"))
        assert _mobiliti_product_name(workbook, mobiliti_row) == "Producto generado"
    finally:
        workbook.close()


def test_active_engine_embedded_source_image_reaches_cotizacion_anchor(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    image = tmp_path / "source-product.png"
    Image.new("RGBA", (32, 24), (20, 80, 160, 255)).save(image)
    source = tmp_path / "source-with-image.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla con imagen", "quantity": 1, "price": 100},),
        image_path=image,
    )
    output = tmp_path / "output-with-product-image.xlsx"
    forbidden_cache = tmp_path / "forbidden-image-cache"
    monkeypatch.setattr(
        engine_module,
        "_IMAGE_CACHE_ROOT",
        forbidden_cache,
        raising=False,
    )

    generate_quote(
        source,
        output,
        {},
        OFFICIAL_TEMPLATE,
        original_quotation_path=None,
    )

    package = XlsxPackage.read(output)
    cotizacion = ET.fromstring(package.parts[package.sheet_part("Cotizacion")])
    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        product_row = next(
            row
            for row in range(16, workbook["Cotizacion"].max_row + 1)
            if (
                isinstance(
                    product_formula := workbook["Cotizacion"].cell(row, 1).value,
                    str,
                )
                and re.fullmatch(r"=Mobiliti!D[1-9][0-9]*", product_formula)
                is not None
                and _mobiliti_product_name(
                    workbook,
                    int(product_formula.removeprefix("=Mobiliti!D")),
                )
                == "Silla con imagen"
            )
        )
    finally:
        workbook.close()
    drawing_part, _drawing_rels_part = _cotizacion_drawing_parts(package)
    drawing = ET.fromstring(package.parts[drawing_part])
    product_anchors = [
        anchor
        for anchor in drawing.findall(f"{{{XDR}}}oneCellAnchor")
        if anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}row") == str(product_row - 1)
        and anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}col") == "1"
    ]
    assert len(product_anchors) == 1
    assert not forbidden_cache.exists()
    assert any(
        name.startswith("xl/media/quote_product_")
        and package.parts[name].startswith(b"\x89PNG\r\n\x1a\n")
        for name in package.parts
    )


def test_active_engine_uses_one_audited_source_snapshot_after_preflight(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original_image = tmp_path / "snapshot-original.png"
    replacement_image = tmp_path / "snapshot-replacement.png"
    Image.new("RGB", (32, 24), (15, 90, 170)).save(original_image)
    Image.new("RGB", (32, 24), (210, 30, 45)).save(replacement_image)
    source = tmp_path / "snapshot-source.xlsx"
    replacement = tmp_path / "snapshot-race-replacement.xlsx"
    _write_engine_source(
        source,
        ({"name": "Producto snapshot original", "quantity": 1, "price": 100},),
        image_path=original_image,
    )
    _write_engine_source(
        replacement,
        ({"name": "Producto carrera hostil", "quantity": 1, "price": 900},),
        image_path=replacement_image,
    )
    for path in (source, replacement):
        normalized = engine_module._normalized_quotation_source(path)
        assert isinstance(normalized, bytes)
        path.write_bytes(normalized)
    original_package = XlsxPackage.read(source)
    replacement_package = XlsxPackage.read(replacement)
    original_media = next(
        payload
        for name, payload in original_package.parts.items()
        if name.startswith("xl/media/")
    )
    replacement_media = next(
        payload
        for name, payload in replacement_package.parts.items()
        if name.startswith("xl/media/")
    )
    assert original_media != replacement_media
    real_read_items = engine_module.read_items
    calls = 0

    def race_after_preflight(audited_source):
        nonlocal calls
        calls += 1
        source.write_bytes(replacement.read_bytes())
        return real_read_items(audited_source)

    monkeypatch.setattr(engine_module, "read_items", race_after_preflight)
    output = tmp_path / "snapshot-output.xlsx"

    generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    assert calls == 1
    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert workbook["Quotation"]["B9"].value == "Producto snapshot original"
        mobiliti_rows = [
            row
            for row in range(1, workbook["Mobiliti"].max_row + 1)
            if _mobiliti_product_name(workbook, row)
            == "Producto snapshot original"
        ]
        assert len(mobiliti_rows) == 1
        assert any(
            workbook["Cotizacion"].cell(row, 1).value
            == f"=Mobiliti!D{mobiliti_rows[0]}"
            for row in range(16, workbook["Cotizacion"].max_row + 1)
        )
        assert all(
            _mobiliti_product_name(workbook, row) != "Producto carrera hostil"
            for row in range(1, workbook["Mobiliti"].max_row + 1)
        )
    finally:
        workbook.close()
    result = XlsxPackage.read(output)
    product_media = [
        payload
        for name, payload in result.parts.items()
        if name.startswith("xl/media/quote_product_")
    ]
    assert product_media == [original_media]
    assert replacement_media not in product_media


@pytest.mark.parametrize(
    "surface",
    ("mobiliti_formula", "cotizacion_terms", "cotizacion_style", "row_height", "merge"),
)
def test_composer_rejects_any_change_outside_the_exact_sheet_allowlist(
    tmp_path: Path,
    surface: str,
) -> None:
    output = tmp_path / f"exact-{surface}.xlsx"
    request = _minimal_request(output)
    if surface == "mobiliti_formula":
        root = ET.fromstring(request.mobiliti.xml)
        coordinate = f"W{request.mobiliti.row_map.item_rows[0]}"
        formula = _cell(root, coordinate).find(f"{{{MAIN}}}f")
        assert formula is not None
        formula.text = "1+1"
        request = replace(
            request,
            mobiliti=replace(
                request.mobiliti,
                xml=ET.tostring(root, encoding="utf-8", xml_declaration=True),
            ),
        )
    else:
        root = ET.fromstring(request.cotizacion.xml)
        product_row = request.cotizacion.product_rows[0]
        if surface == "cotizacion_terms":
            terms_row = 28 + request.cotizacion.terms_row_delta
            term = _cell(root, f"A{terms_row}")
            value = term.find(f"{{{MAIN}}}v")
            assert value is not None
            value.text = "999999"
        elif surface == "cotizacion_style":
            _cell(root, f"A{product_row}").attrib["s"] = "0"
        elif surface == "row_height":
            row = root.find(f"{{{MAIN}}}sheetData/{{{MAIN}}}row[@r='{product_row}']")
            assert row is not None
            row.attrib["ht"] = "999"
        else:
            merge = root.find(f"{{{MAIN}}}mergeCells/{{{MAIN}}}mergeCell[@ref='A16:J16']")
            assert merge is not None
            merge.attrib["ref"] = "A16:I16"
        request = replace(
            request,
            cotizacion=replace(
                request.cotizacion,
                xml=ET.tostring(root, encoding="utf-8", xml_declaration=True),
            ),
        )

    with pytest.raises(ValueError, match="contrato exacto"):
        compose_official_quote(request)

    assert not output.exists()


def test_composer_rejects_product_images_outside_product_rows(
    tmp_path: Path,
) -> None:
    image = tmp_path / "safe.png"
    Image.new("RGB", (16, 16), (10, 20, 30)).save(image)
    output = tmp_path / "bad-image-row.xlsx"
    request = _minimal_request(output)
    cotizacion = replace(
        request.cotizacion,
        images=(CotizacionProductImage(image.resolve(), 999),),
    )

    with pytest.raises(ValueError, match="filas de producto"):
        compose_official_quote(replace(request, cotizacion=cotizacion))

    assert not output.exists()


def test_composer_rejects_duplicate_image_positions_in_one_product_row(
    tmp_path: Path,
) -> None:
    image = tmp_path / "safe.png"
    Image.new("RGB", (16, 16), (10, 20, 30)).save(image)
    output = tmp_path / "duplicate-image-position.xlsx"
    request = _minimal_request(output)
    product_row = request.cotizacion.product_rows[0]
    cotizacion = replace(
        request.cotizacion,
        images=(
            CotizacionProductImage(image.resolve(), product_row, position=0),
            CotizacionProductImage(image.resolve(), product_row, position=0),
        ),
    )

    with pytest.raises(ValueError, match="posiciones"):
        compose_official_quote(replace(request, cotizacion=cotizacion))

    assert not output.exists()


def test_cotizacion_places_principal_and_complements_as_separate_images(
    tmp_path: Path,
) -> None:
    def png_payload(color: tuple[int, int, int]) -> bytes:
        stream = BytesIO()
        Image.new("RGB", (120, 90), color).save(stream, format="PNG")
        return stream.getvalue()

    output = tmp_path / "separate-complement-images.xlsx"
    request = _request_for_sections(output, (1,))
    product_row = request.cotizacion.product_rows[0]
    payloads = (
        png_payload((20, 80, 180)),
        png_payload((20, 170, 90)),
        png_payload((190, 90, 30)),
    )
    cotizacion = replace(
        request.cotizacion,
        images=tuple(
            CotizacionProductImage(
                None,
                product_row,
                content=payload,
                content_type="image/png",
                position=position,
            )
            for position, payload in enumerate(payloads)
        ),
    )

    compose_official_quote(replace(request, cotizacion=cotizacion))

    package = XlsxPackage.read(output)
    drawing_part, drawing_rels_part = _cotizacion_drawing_parts(package)
    drawing = ET.fromstring(package.parts[drawing_part])
    relationships = {
        node.attrib["Id"]: node.attrib
        for node in ET.fromstring(package.parts[drawing_rels_part]).findall(
            f"{{{PACKAGE_REL}}}Relationship"
        )
    }
    anchors = [
        anchor
        for anchor in drawing.findall(f"{{{XDR}}}oneCellAnchor")
        if anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}row")
        == str(product_row - 1)
        and (
            (name := anchor.find(f".//{{{XDR}}}cNvPr")) is not None
            and name.attrib["name"].startswith(("Imagen principal", "Imagen complemento"))
        )
    ]
    assert [
        anchor.find(f".//{{{XDR}}}cNvPr").attrib["name"]
        for anchor in anchors
    ] == [
        "Imagen principal 0001",
        "Imagen complemento 0001-01",
        "Imagen complemento 0001-02",
    ]
    extents = [
        (
            int(anchor.find(f"{{{XDR}}}ext").attrib["cx"]),
            int(anchor.find(f"{{{XDR}}}ext").attrib["cy"]),
        )
        for anchor in anchors
    ]
    assert extents[0][0] * extents[0][1] > extents[1][0] * extents[1][1]
    assert extents[1] == extents[2]
    for anchor in anchors:
        marker = anchor.find(f"{{{XDR}}}from")
        extent = anchor.find(f"{{{XDR}}}ext")
        assert marker is not None
        assert extent is not None
        assert int(marker.findtext(f"{{{XDR}}}colOff")) >= 0
        assert int(marker.findtext(f"{{{XDR}}}rowOff")) >= 0
        assert (
            int(marker.findtext(f"{{{XDR}}}colOff"))
            + int(extent.attrib["cx"])
            <= official_composer_module._COTIZACION_IMAGE_CELL_WIDTH_EMU
        )
        assert (
            int(marker.findtext(f"{{{XDR}}}rowOff"))
            + int(extent.attrib["cy"])
            <= official_composer_module._COTIZACION_IMAGE_CELL_HEIGHT_EMU
        )
    media_payloads = []
    for anchor in anchors:
        relationship_id = anchor.find(
            ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
        ).attrib[f"{{{OFFICE_REL}}}embed"]
        media_part = resolve_internal_target(
            drawing_part,
            relationships[relationship_id]["Target"],
        )
        media_payloads.append(package.parts[media_part])
    assert media_payloads == list(payloads)


def test_cotizacion_sheet_editor_expands_complement_sources_in_order() -> None:
    def png_payload(color: tuple[int, int, int]) -> bytes:
        stream = BytesIO()
        Image.new("RGB", (40, 30), color).save(stream, format="PNG")
        return stream.getvalue()

    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    principal = png_payload((20, 80, 180))
    first = png_payload((20, 170, 90))
    second = png_payload((190, 90, 30))
    mutation = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=CotizacionMetadata(),
        sections=(
            CotizacionSection(
                "Recepción",
                (
                    CotizacionProduct(
                        item_key="principal",
                        name="Producto principal",
                        description="Principal + complementos",
                        dimensions="600 x 600 mm",
                        quantity=Decimal("1"),
                        mobiliti_row=14,
                        image_content=principal,
                        image_content_type="image/png",
                        complement_images=(
                            CotizacionImageSource(
                                content=first,
                                content_type="image/png",
                            ),
                            CotizacionImageSource(
                                content=second,
                                content_type="image/png",
                            ),
                        ),
                    ),
                ),
            ),
        ),
    )

    assert [
        (image.target_row, image.position, image.content)
        for image in mutation.images
    ] == [
        (17, 0, principal),
        (17, 1, first),
        (17, 2, second),
    ]


def test_product_picture_ids_are_unique_and_names_are_deterministic(tmp_path: Path) -> None:
    first = tmp_path / "customer-name.png"
    second = tmp_path / "another-customer-name.png"
    Image.new("RGB", (16, 16), (10, 20, 30)).save(first)
    Image.new("RGB", (16, 16), (40, 50, 60)).save(second)
    output = tmp_path / "deterministic-pictures.xlsx"
    request = _request_for_sections(output, (2,))
    product_rows = request.cotizacion.product_rows
    cotizacion = replace(
        request.cotizacion,
        images=(
            CotizacionProductImage(first.resolve(), product_rows[0]),
            CotizacionProductImage(second.resolve(), product_rows[1]),
        ),
    )

    compose_official_quote(replace(request, cotizacion=cotizacion))

    package = XlsxPackage.read(output)
    drawing_part, _drawing_rels = _cotizacion_drawing_parts(package)
    drawing = ET.fromstring(package.parts[drawing_part])
    ids = [node.attrib["id"] for node in drawing.findall(f".//{{{XDR}}}cNvPr")]
    assert len(ids) == len(set(ids))
    product_names = []
    for anchor in drawing.findall(f"{{{XDR}}}oneCellAnchor"):
        row = anchor.findtext(f"{{{XDR}}}from/{{{XDR}}}row")
        if row in {str(number - 1) for number in product_rows}:
            node = anchor.find(f".//{{{XDR}}}cNvPr")
            assert node is not None
            product_names.append(node.attrib["name"])
    assert product_names == ["Imagen de producto 0001", "Imagen de producto 0002"]


def test_cotizacion_centers_product_images_and_preserves_original_payloads(
    tmp_path: Path,
) -> None:
    def png_payload(size: tuple[int, int], color: tuple[int, int, int]) -> bytes:
        stream = BytesIO()
        Image.new("RGB", size, color).save(stream, format="PNG")
        return stream.getvalue()

    output = tmp_path / "centered-project-images.xlsx"
    request = _request_for_sections(output, (2,))
    panoramic = png_payload((500, 100), (20, 80, 180))
    portrait = png_payload((100, 500), (20, 170, 90))
    first_row, second_row = request.cotizacion.product_rows
    cotizacion = replace(
        request.cotizacion,
        images=(
            CotizacionProductImage(
                None,
                first_row,
                content=panoramic,
                content_type="image/png",
            ),
            CotizacionProductImage(
                None,
                second_row,
                content=portrait,
                content_type="image/png",
            ),
        ),
    )

    compose_official_quote(replace(request, cotizacion=cotizacion))

    package = XlsxPackage.read(output)
    drawing_part, drawing_rels_part = _cotizacion_drawing_parts(package)
    drawing = ET.fromstring(package.parts[drawing_part])
    relationships = {
        node.attrib["Id"]: node.attrib
        for node in ET.fromstring(package.parts[drawing_rels_part]).findall(
            f"{{{PACKAGE_REL}}}Relationship"
        )
    }
    anchors = {
        node.find(f".//{{{XDR}}}cNvPr").attrib["name"]: node
        for node in drawing.findall(f"{{{XDR}}}oneCellAnchor")
        if node.find(f".//{{{XDR}}}cNvPr") is not None
        and node.find(f".//{{{XDR}}}cNvPr")
        .attrib["name"]
        .startswith("Imagen de producto")
    }
    expected = {
        "Imagen de producto 0001": (
            first_row,
            panoramic,
            3_200_000,
            640_000,
            909_837,
            1_884_085,
        ),
        "Imagen de producto 0002": (
            second_row,
            portrait,
            600_000,
            3_000_000,
            2_209_837,
            704_085,
        ),
    }
    assert set(anchors) == set(expected)
    for name, (
        target_row,
        original_payload,
        width,
        height,
        column_offset,
        row_offset,
    ) in expected.items():
        anchor = anchors[name]
        marker = anchor.find(f"{{{XDR}}}from")
        extent = anchor.find(f"{{{XDR}}}ext")
        assert marker is not None
        assert extent is not None
        assert marker.findtext(f"{{{XDR}}}col") == "1"
        assert marker.findtext(f"{{{XDR}}}row") == str(target_row - 1)
        assert int(marker.findtext(f"{{{XDR}}}colOff")) == column_offset
        assert int(marker.findtext(f"{{{XDR}}}rowOff")) == row_offset
        assert extent.attrib == {"cx": str(width), "cy": str(height)}
        shape_extent = anchor.find(
            ".//{http://schemas.openxmlformats.org/drawingml/2006/main}ext"
        )
        assert shape_extent is not None
        assert shape_extent.attrib == {"cx": str(width), "cy": str(height)}
        relationship_id = anchor.find(
            ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
        ).attrib[f"{{{OFFICE_REL}}}embed"]
        media_part = resolve_internal_target(
            drawing_part,
            relationships[relationship_id]["Target"],
        )
        assert package.parts[media_part] == original_payload


def test_composer_rejects_inline_string_value_smuggling(tmp_path: Path) -> None:
    output = tmp_path / "inline-smuggling.xlsx"
    request = _minimal_request(output)
    root = ET.fromstring(request.cotizacion.xml)
    header = _cell(root, "B3")
    ET.SubElement(header, f"{{{MAIN}}}v").text = "contenido oculto"
    cotizacion = replace(
        request.cotizacion,
        xml=ET.tostring(root, encoding="utf-8", xml_declaration=True),
    )

    with pytest.raises(ValueError, match="inlineStr exacto"):
        compose_official_quote(replace(request, cotizacion=cotizacion))

    assert not output.exists()


def test_audit_failure_never_leaves_output_or_compose_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "audit-failure.xlsx"

    def fail_contract(*_args, **_kwargs) -> None:
        raise ValueError("fallo de auditoría inyectado")

    monkeypatch.setattr(official_composer_module, "verify_output_contract", fail_contract)

    with pytest.raises(ValueError, match="fallo de auditoría"):
        compose_official_quote(_minimal_request(output))

    assert not output.exists()
    assert not tuple(tmp_path.glob(".*.compose-*.tmp"))


def test_publication_race_preserves_existing_output_and_recovers_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "race.xlsx"
    recovered = tmp_path / "recovered"
    recovered.mkdir()
    real_rename = official_composer_module.os.rename

    def race_publish(source: Path, destination: Path) -> None:
        Path(destination).write_bytes(b"archivo del usuario")
        raise FileExistsError("carrera EEXIST simulada")

    def recover_candidate(candidate: Path) -> None:
        real_rename(candidate, recovered / candidate.name)

    monkeypatch.setattr(
        official_composer_module,
        "_atomic_publish_no_replace",
        race_publish,
        raising=False,
    )
    monkeypatch.setattr(
        official_composer_module,
        "_recycle_candidate",
        recover_candidate,
        raising=False,
    )

    with pytest.raises(FileExistsError):
        compose_official_quote(_minimal_request(output))

    assert output.read_bytes() == b"archivo del usuario"
    assert not tuple(tmp_path.glob(".*.compose-*.tmp"))
    assert len(tuple(recovered.iterdir())) == 1


def test_compose_paths_reject_lexical_parent_segments_before_resolution(
    tmp_path: Path,
) -> None:
    nested = tmp_path / "existing"
    nested.mkdir()
    output = nested / ".." / "lexical.xlsx"
    request = replace(_minimal_request(tmp_path / "safe.xlsx"), output=output)

    with pytest.raises(ValueError, match="segmentos léxicos"):
        compose_official_quote(request)

    assert not (tmp_path / "lexical.xlsx").exists()


def test_compose_paths_reject_windows_reparse_points(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "reparse.xlsx"
    monkeypatch.setattr(
        official_composer_module,
        "_path_is_reparse_point",
        lambda path: Path(path) == tmp_path,
        raising=False,
    )

    with pytest.raises(ValueError, match="reparse point"):
        compose_official_quote(_minimal_request(output))

    assert not output.exists()


def test_estrategia_translation_never_rewrites_reference_like_string_literals() -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    part = base.sheet_part("Estrategia Comercial ")
    root = ET.fromstring(base.parts[part])
    formula = _cell(root, "B7").find(f"{{{MAIN}}}f")
    assert formula is not None
    formula.text = (
        'IF("Mobiliti!$G$14:$G$571"="literal",0,'
        'SUMIF(Mobiliti!$G$14:$G$571,"P00500",Mobiliti!$AD$14:$AD$571))'
    )
    row_map = plan_mobiliti_layout([SectionNeed("large", "Large", 100)])

    translated = ET.fromstring(
        _translate_estrategia(
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
            row_map,
            30,
        )
    )
    result = _cell(translated, "B7").findtext(f"{{{MAIN}}}f")
    assert '"Mobiliti!$G$14:$G$571"' in result
    assert f"Mobiliti!$G$14:$G${row_map.last_product_row}" in result
    assert f"Mobiliti!$AD$14:$AD${row_map.last_product_row}" in result


def test_estrategia_subtotal_translation_only_changes_range_tokens() -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    part = base.sheet_part("Estrategia Comercial ")
    root = ET.fromstring(base.parts[part])
    formula = _cell(root, "D59").find(f"{{{MAIN}}}f")
    assert formula is not None
    formula.text = (
        'IF("Cotizacion!H25"="literal",0,'
        'IF(B61=0,C61,B61*Cotizacion!H24))'
    )
    row_map = plan_mobiliti_layout([SectionNeed("large", "Large", 100)])

    translated = ET.fromstring(
        _translate_estrategia(
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
            row_map,
            30,
        )
    )
    result = _cell(translated, "D59").findtext(f"{{{MAIN}}}f")
    assert '"Cotizacion!H25"' in result
    assert "B61*Cotizacion!H30" in result


def test_composed_zip_bytes_ignore_output_name_image_name_and_image_mtime(
    tmp_path: Path,
) -> None:
    first_image = tmp_path / "first-customer-name.png"
    second_image = tmp_path / "second-customer-name.png"
    Image.new("RGB", (24, 16), (90, 40, 10)).save(first_image)
    second_image.write_bytes(first_image.read_bytes())
    os.utime(first_image, (1_000_000_000, 1_000_000_000))
    os.utime(second_image, (1_700_000_000, 1_700_000_000))
    first_output = tmp_path / "first-output.xlsx"
    second_output = tmp_path / "different-output-name.xlsx"

    first_request = _minimal_request(first_output)
    second_request = _minimal_request(second_output)
    first_cotizacion = replace(
        first_request.cotizacion,
        images=(
            CotizacionProductImage(
                first_image.resolve(),
                first_request.cotizacion.product_rows[0],
            ),
        ),
    )
    second_cotizacion = replace(
        second_request.cotizacion,
        images=(
            CotizacionProductImage(
                second_image.resolve(),
                second_request.cotizacion.product_rows[0],
            ),
        ),
    )

    compose_official_quote(replace(first_request, cotizacion=first_cotizacion))
    compose_official_quote(replace(second_request, cotizacion=second_cotizacion))

    assert first_output.read_bytes() == second_output.read_bytes()


@pytest.mark.parametrize(
    ("limit_name", "limit", "expected"),
    (
        ("MAX_ZIP_ENTRIES", 1, "límite de entradas"),
        ("MAX_ZIP_PART_BYTES", 8, "límite por parte"),
    ),
)
def test_generate_quote_preflights_source_before_openpyxl_or_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    limit_name: str,
    limit: int,
    expected: str,
) -> None:
    source = tmp_path / f"preflight-{limit_name}.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla segura", "qty": 1, "price": 100},),
    )
    output = tmp_path / f"out-{limit_name}.xlsx"
    cache = tmp_path / "forbidden-image-cache"
    monkeypatch.setattr(ooxml_package_module, limit_name, limit)
    monkeypatch.setattr(engine_module, "_IMAGE_CACHE_ROOT", cache, raising=False)
    monkeypatch.setattr(
        engine_module,
        "read_items",
        lambda _source: (_ for _ in ()).throw(
            AssertionError("OpenPyXL fue invocado antes del preflight")
        ),
    )

    with pytest.raises(ValueError, match=expected):
        generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    assert not output.exists()
    assert not cache.exists()


def test_generate_quote_rejects_compression_bomb_before_output_or_cache(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "compression-bomb.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla segura", "qty": 1, "price": 100},),
    )
    with ZipFile(source, "a", ZIP_DEFLATED) as archive:
        archive.writestr("custom/compression-bomb.bin", b"0" * (2 * 1024 * 1024))
    output = tmp_path / "compression-bomb-output.xlsx"
    cache = tmp_path / "forbidden-image-cache"
    monkeypatch.setattr(engine_module, "_IMAGE_CACHE_ROOT", cache, raising=False)
    monkeypatch.setattr(
        engine_module,
        "read_items",
        lambda _source: (_ for _ in ()).throw(
            AssertionError("OpenPyXL fue invocado antes del preflight")
        ),
    )

    with pytest.raises(ValueError, match="ratio de compresión"):
        generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    assert not output.exists()
    assert not cache.exists()


def test_generate_quote_missing_output_parent_has_zero_filesystem_effects(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla segura", "qty": 1, "price": 100},),
    )
    missing_parent = tmp_path / "missing" / "nested"
    output = missing_parent / "quote.xlsx"

    with pytest.raises(FileNotFoundError, match="Directorio de salida inexistente"):
        generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    assert not missing_parent.parent.exists()
    assert not output.exists()


def test_generate_quote_rejects_lexical_parent_before_any_effect(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    _write_engine_source(
        source,
        ({"name": "Silla segura", "qty": 1, "price": 100},),
    )
    existing = tmp_path / "existing"
    existing.mkdir()
    output = existing / ".." / "lexical-output.xlsx"

    with pytest.raises(ValueError, match="segmentos léxicos"):
        generate_quote(source, output, {}, OFFICIAL_TEMPLATE)

    assert not (tmp_path / "lexical-output.xlsx").exists()


def _imported_formula_addition(
    formula: str,
    *,
    defined_names: tuple[LocalDefinedName, ...] = (),
) -> SheetAddition:
    worksheet = ET.Element(f"{{{MAIN}}}worksheet")
    sheet_data = ET.SubElement(worksheet, f"{{{MAIN}}}sheetData")
    row = ET.SubElement(sheet_data, f"{{{MAIN}}}row", {"r": "1"})
    cell = ET.SubElement(row, f"{{{MAIN}}}c", {"r": "A1"})
    ET.SubElement(cell, f"{{{MAIN}}}f").text = formula
    ET.SubElement(cell, f"{{{MAIN}}}v").text = "0"
    payload = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)
    part = "xl/worksheets/quotation_security_test.xml"
    return SheetAddition(
        name="Quotation",
        state="visible",
        xml=payload,
        parts={part: payload},
        content_types={
            part: "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        },
        sheet_part=part,
        defined_names=defined_names,
    )


def _worksheet_formula_surface_addition(
    surface: str,
    formula: str,
) -> SheetAddition:
    addition = _imported_formula_addition("SUM(A2:A3)")
    assert addition.sheet_part is not None
    worksheet = ET.fromstring(addition.xml)
    if surface in {"data_validation", "data_validation_2"}:
        validations = ET.SubElement(
            worksheet,
            f"{{{MAIN}}}dataValidations",
            {"count": "1"},
        )
        validation = ET.SubElement(
            validations,
            f"{{{MAIN}}}dataValidation",
            {"sqref": "A1"},
        )
        tag = "formula2" if surface == "data_validation_2" else "formula1"
        ET.SubElement(validation, f"{{{MAIN}}}{tag}").text = formula
    elif surface == "conditional_formatting":
        formatting = ET.SubElement(
            worksheet,
            f"{{{MAIN}}}conditionalFormatting",
            {"sqref": "A1"},
        )
        rule = ET.SubElement(
            formatting,
            f"{{{MAIN}}}cfRule",
            {"type": "expression", "priority": "1"},
        )
        ET.SubElement(rule, f"{{{MAIN}}}formula").text = formula
    else:
        raise AssertionError(f"Superficie worksheet desconocida: {surface}")
    payload = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)
    return replace(
        addition,
        xml=payload,
        parts={addition.sheet_part: payload},
    )


def _related_formula_surface_addition(
    surface: str,
    formula: str,
) -> SheetAddition:
    addition = _imported_formula_addition("SUM(A2:A3)")
    assert addition.sheet_part is not None
    sheet_rels_part = relationship_part_name(addition.sheet_part)
    sheet_rels = ET.Element(f"{{{PACKAGE_REL}}}Relationships")
    parts = dict(addition.parts)
    content_types = dict(addition.content_types)

    if surface in {"table", "table_totals"}:
        target_part = "xl/tables/quotation_security_table.xml"
        ET.SubElement(
            sheet_rels,
            f"{{{PACKAGE_REL}}}Relationship",
            {
                "Id": "rIdFormulaSurface",
                "Type": f"{OFFICE_REL}/table",
                "Target": "../tables/quotation_security_table.xml",
            },
        )
        table = ET.Element(
            f"{{{MAIN}}}table",
            {"id": "1", "name": "SecurityTable", "displayName": "SecurityTable"},
        )
        columns = ET.SubElement(table, f"{{{MAIN}}}tableColumns", {"count": "1"})
        column = ET.SubElement(
            columns,
            f"{{{MAIN}}}tableColumn",
            {"id": "1", "name": "Amount"},
        )
        formula_tag = (
            "totalsRowFormula"
            if surface == "table_totals"
            else "calculatedColumnFormula"
        )
        ET.SubElement(column, f"{{{MAIN}}}{formula_tag}").text = formula
        parts[target_part] = ET.tostring(table, encoding="utf-8", xml_declaration=True)
        content_types[target_part] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
        )
    elif surface == "chart":
        drawing_part = "xl/drawings/quotation_security_drawing.xml"
        drawing_rels_part = relationship_part_name(drawing_part)
        target_part = "xl/charts/quotation_security_chart.xml"
        ET.SubElement(
            sheet_rels,
            f"{{{PACKAGE_REL}}}Relationship",
            {
                "Id": "rIdFormulaSurfaceDrawing",
                "Type": f"{OFFICE_REL}/drawing",
                "Target": "../drawings/quotation_security_drawing.xml",
            },
        )
        drawing_rels = ET.Element(f"{{{PACKAGE_REL}}}Relationships")
        ET.SubElement(
            drawing_rels,
            f"{{{PACKAGE_REL}}}Relationship",
            {
                "Id": "rIdFormulaSurfaceChart",
                "Type": f"{OFFICE_REL}/chart",
                "Target": "../charts/quotation_security_chart.xml",
            },
        )
        chart = ET.Element(f"{{{CHART}}}chartSpace")
        chart_node = ET.SubElement(chart, f"{{{CHART}}}chart")
        plot = ET.SubElement(chart_node, f"{{{CHART}}}plotArea")
        series = ET.SubElement(plot, f"{{{CHART}}}ser")
        reference = ET.SubElement(series, f"{{{CHART}}}numRef")
        ET.SubElement(reference, f"{{{CHART}}}f").text = formula
        parts[drawing_part] = ET.tostring(
            ET.Element(f"{{{XDR}}}wsDr"),
            encoding="utf-8",
            xml_declaration=True,
        )
        parts[drawing_rels_part] = ET.tostring(
            drawing_rels,
            encoding="utf-8",
            xml_declaration=True,
        )
        parts[target_part] = ET.tostring(chart, encoding="utf-8", xml_declaration=True)
        content_types[drawing_part] = (
            "application/vnd.openxmlformats-officedocument.drawing+xml"
        )
        content_types[drawing_rels_part] = (
            "application/vnd.openxmlformats-package.relationships+xml"
        )
        content_types[target_part] = (
            "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
        )
    else:
        raise AssertionError(f"Superficie relacionada desconocida: {surface}")

    parts[sheet_rels_part] = ET.tostring(
        sheet_rels,
        encoding="utf-8",
        xml_declaration=True,
    )
    content_types[sheet_rels_part] = (
        "application/vnd.openxmlformats-package.relationships+xml"
    )
    return replace(addition, parts=parts, content_types=content_types)


def _addition_with_base_replacement(
    addition: SheetAddition,
    base: XlsxPackage,
    part: str,
    payload: bytes | None = None,
) -> SheetAddition:
    return replace(
        addition,
        replacements={part: base.parts[part] if payload is None else payload},
        replacement_content_types=base.content_types_for({part}),
    )


@pytest.mark.parametrize(
    "formula",
    (
        "WEBSERVICE(\"https://example.test/data\")",
        "HYPERLINK(\"https://example.test/attack\",\"abrir\")",
        "[evil.xlsx]Sheet1!A1",
        "cmd|'/C calc'!A0",
    ),
)
def test_composer_rejects_active_or_external_imported_formulas(
    tmp_path: Path,
    formula: str,
) -> None:
    request = _minimal_request(tmp_path / "forbidden-formula.xlsx")
    request = replace(request, quotation=_imported_formula_addition(formula))

    with pytest.raises(ValueError, match="Fórmula importada no permitida"):
        build_allowlisted_mutation(XlsxPackage.read(OFFICIAL_TEMPLATE), request)


@pytest.mark.parametrize(
    "formula",
    (
        "WEBSERVICE (A1)",
        "_xlfn.WEBSERVICE (A1)",
        "_xlws.RTD (A1)",
    ),
)
def test_imported_formula_rejects_spaced_and_prefixed_dangerous_functions(
    formula: str,
) -> None:
    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        official_composer_module._validate_imported_formula(
            formula,
            "token-sequence-test",
        )


@pytest.mark.parametrize(
    "formula",
    (
        "SUM (A1:A2)",
        "(A1+A2)",
        "IF (A1, SUM (A1:A2), 0)",
    ),
)
def test_imported_formula_keeps_safe_whitespace_and_grouping(formula: str) -> None:
    official_composer_module._validate_imported_formula(
        formula,
        "safe-token-sequence-test",
    )


@pytest.mark.parametrize(
    "formula",
    (
        "@WEBSERVICE(A1)",
        "@WEBSERVICE (A1)",
        "@_xlfn.WEBSERVICE (A1)",
        "_xlfn.@WEBSERVICE (A1)",
        "@_xlfn.@_xlws.RTD (A1)",
    ),
)
def test_imported_formula_rejects_implicit_intersection_function_prefixes(
    formula: str,
) -> None:
    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        official_composer_module._validate_imported_formula(
            formula,
            "implicit-intersection-test",
        )


@pytest.mark.parametrize("formula", ("@A1", "(@A1+A2)"))
def test_imported_formula_keeps_safe_implicit_reference_and_grouping(
    formula: str,
) -> None:
    official_composer_module._validate_imported_formula(
        formula,
        "safe-implicit-intersection-test",
    )


@pytest.mark.parametrize(
    "addition",
    (
        _imported_formula_addition("@WEBSERVICE(A1)"),
        _worksheet_formula_surface_addition(
            "data_validation",
            "@_xlfn.WEBSERVICE (A1)",
        ),
        _worksheet_formula_surface_addition(
            "conditional_formatting",
            "_xlfn.@EXEC (A1)",
        ),
        _related_formula_surface_addition(
            "table",
            "@_xlws.CALL (A1)",
        ),
        _related_formula_surface_addition(
            "chart",
            "_xlws.@RTD (A1)",
        ),
        _imported_formula_addition(
            "SUM(A1:A2)",
            defined_names=(
                LocalDefinedName(
                    name="ImplicitDanger",
                    text="@_xlfn.WEBSERVICE (A1)",
                ),
            ),
        ),
    ),
    ids=("worksheet", "data-validation", "conditional-formatting", "table", "chart", "defined-name"),
)
def test_imported_formula_surfaces_reject_implicit_dangerous_function(
    addition: SheetAddition,
) -> None:
    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        official_composer_module._validate_imported_formula_surfaces((addition,))


@pytest.mark.parametrize(
    ("surface", "formula"),
    (
        ("data_validation", "EXEC (A1)"),
        ("data_validation_2", "[evil.xlsx]Sheet1!A1"),
        ("conditional_formatting", "[evil.xlsx]Sheet1!A1"),
    ),
)
def test_imported_worksheet_formula_surfaces_reject_dangerous_or_external(
    surface: str,
    formula: str,
) -> None:
    addition = _worksheet_formula_surface_addition(surface, formula)

    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        official_composer_module._validate_imported_formula_surfaces((addition,))


@pytest.mark.parametrize(
    ("surface", "formula"),
    (
        ("table", "_xlfn.CALL (A1)"),
        ("table", "[evil.xlsx]Sheet1!A1"),
        ("table_totals", "[evil.xlsx]Sheet1!A1"),
        ("chart", "_xlws.RTD (A1)"),
        ("chart", "[evil.xlsx]Sheet1!$A$1:$A$2"),
    ),
)
def test_imported_related_formula_surfaces_reject_dangerous_or_external(
    surface: str,
    formula: str,
) -> None:
    addition = _related_formula_surface_addition(surface, formula)

    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        official_composer_module._validate_imported_formula_surfaces((addition,))


@pytest.mark.parametrize(
    "addition",
    (
        _worksheet_formula_surface_addition(
            "data_validation",
            "SUM (A1:A2)",
        ),
        _related_formula_surface_addition(
            "table",
            "SUM([Amount])",
        ),
    ),
    ids=("internal-data-validation", "internal-table-formula"),
)
def test_imported_extended_formula_surfaces_keep_safe_internal_formula(
    addition: SheetAddition,
) -> None:
    before = dict(addition.parts)

    official_composer_module._validate_imported_formula_surfaces((addition,))

    assert dict(addition.parts) == before


def test_imported_formula_surfaces_ignore_homonymous_foreign_namespace() -> None:
    addition = _imported_formula_addition("SUM(A1:A2)")
    assert addition.sheet_part is not None
    worksheet = ET.fromstring(addition.xml)
    ET.SubElement(worksheet, "{urn:mobiliti:not-a-formula-surface}formula").text = (
        "EXEC (A1)"
    )
    payload = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)
    addition = replace(
        addition,
        xml=payload,
        parts={addition.sheet_part: payload},
    )

    official_composer_module._validate_imported_formula_surfaces((addition,))


def test_composer_rejects_active_imported_defined_name_formula(tmp_path: Path) -> None:
    request = _minimal_request(tmp_path / "forbidden-defined-name.xlsx")
    addition = _imported_formula_addition(
        "SUM(A1:A2)",
        defined_names=(
            LocalDefinedName(
                name="DangerousName",
                text='WEBSERVICE("https://example.test/data")',
            ),
        ),
    )

    with pytest.raises(ValueError, match="Fórmula importada no permitida"):
        build_allowlisted_mutation(
            XlsxPackage.read(OFFICIAL_TEMPLATE),
            replace(request, quotation=addition),
        )


def test_composer_keeps_safe_internal_imported_formula(tmp_path: Path) -> None:
    request = _minimal_request(tmp_path / "safe-formula.xlsx")
    addition = _imported_formula_addition('IF(A2="",0,SUM(A2:A3))')

    mutation = build_allowlisted_mutation(
        XlsxPackage.read(OFFICIAL_TEMPLATE),
        replace(request, quotation=addition),
    )

    assert mutation.additions[addition.sheet_part] == addition.xml


def test_imported_formula_validation_rejects_malformed_declared_worksheet(
    tmp_path: Path,
) -> None:
    request = _minimal_request(tmp_path / "malformed-imported-worksheet.xlsx")
    addition = _imported_formula_addition("SUM(A2:A3)")
    assert addition.sheet_part is not None
    malformed = b"<worksheet"
    addition = replace(
        addition,
        xml=malformed,
        parts={addition.sheet_part: malformed},
    )

    with pytest.raises(ValueError, match="XML importado inv.lido"):
        build_allowlisted_mutation(
            XlsxPackage.read(OFFICIAL_TEMPLATE),
            replace(request, quotation=addition),
        )


def test_imported_formula_validation_uses_content_type_not_part_extension(
    tmp_path: Path,
) -> None:
    request = _minimal_request(tmp_path / "misnamed-dangerous-worksheet.xlsx")
    addition = _imported_formula_addition(
        'WEBSERVICE("https://example.test/data")'
    )
    disguised_part = "xl/media/quotation_security_test.bin"
    addition = replace(
        addition,
        parts={disguised_part: addition.xml},
        content_types={
            disguised_part: official_composer_module.WORKSHEET_CONTENT_TYPE
        },
        sheet_part=disguised_part,
    )

    with pytest.raises(ValueError, match="F.rmula importada no permitida"):
        build_allowlisted_mutation(
            XlsxPackage.read(OFFICIAL_TEMPLATE),
            replace(request, quotation=addition),
        )


def test_imported_formula_validation_skips_bounded_non_xml_media(tmp_path: Path) -> None:
    request = _minimal_request(tmp_path / "safe-formula-with-media.xlsx")
    addition = _imported_formula_addition("SUM(A2:A3)")
    media_parts = {
        "xl/media/quotation_security_test.png": _png_payload(),
        "xl/media/quotation_security_test.wmf": b"\xd7\xcd\xc6\x9aWMF-binary-payload",
    }
    addition = replace(
        addition,
        parts={**addition.parts, **media_parts},
        content_types={
            **addition.content_types,
            "xl/media/quotation_security_test.png": "image/png",
            "xl/media/quotation_security_test.wmf": "image/x-wmf",
        },
    )

    mutation = build_allowlisted_mutation(
        XlsxPackage.read(OFFICIAL_TEMPLATE),
        replace(request, quotation=addition),
    )

    for media_part, expected in media_parts.items():
        assert mutation.additions[media_part] == expected


def test_quotation_rejects_non_append_only_styles_replacement(tmp_path: Path) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "malicious-styles-replacement.xlsx")
    styles_part = base.workbook_related_part("styles")
    assert styles_part is not None
    arbitrary_styles = (
        f'<styleSheet xmlns="{MAIN}"><fonts count="0"/></styleSheet>'
    ).encode("utf-8")
    addition = _addition_with_base_replacement(
        _imported_formula_addition("SUM(A1:A2)"),
        base,
        styles_part,
        arbitrary_styles,
    )

    with pytest.raises(ValueError, match="(?i)styles.*append|preserv"):
        build_allowlisted_mutation(
            base,
            replace(request, quotation=addition),
        )


@pytest.mark.parametrize(
    "part",
    (
        "xl/theme/theme1.xml",
        "xl/workbook.xml",
        "xl/richData/rdrichvalue.xml",
        "xl/externalLinks/externalLink1.xml",
        "[Content_Types].xml",
    ),
)
def test_quotation_rejects_unrelated_base_replacement(
    tmp_path: Path,
    part: str,
) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "unrelated-replacement.xlsx")
    addition = _addition_with_base_replacement(
        _imported_formula_addition("SUM(A1:A2)"),
        base,
        part,
    )

    with pytest.raises(ValueError, match="Reemplazo de Quotation no permitido"):
        build_allowlisted_mutation(
            base,
            replace(request, quotation=addition),
        )


def test_quotation_data_rejects_every_replacement(tmp_path: Path) -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    request = _minimal_request(tmp_path / "quotation-data-replacement.xlsx")
    assert request.quotation_data is not None
    data_addition = _addition_with_base_replacement(
        request.quotation_data,
        base,
        "xl/styles.xml",
    )

    with pytest.raises(ValueError, match="Quotation_Data no permite reemplazos"):
        build_allowlisted_mutation(
            base,
            replace(request, quotation_data=data_addition),
        )


def test_real_transplanted_quotation_keeps_valid_append_only_styles(
    tmp_path: Path,
) -> None:
    source = tmp_path / "legitimate-transplant.xlsx"
    _write_engine_source(
        source,
        ({"name": "Producto legítimo", "quantity": 1, "price": 100},),
    )
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    snapshot = engine_module._normalized_quotation_source(source)
    addition = transplant_quotation(snapshot, base)
    assert addition is not None
    assert set(addition.replacements) == {"xl/styles.xml"}
    request = replace(
        _minimal_request(tmp_path / "legitimate-transplant-output.xlsx"),
        quotation=addition,
    )

    mutation = build_allowlisted_mutation(base, request)

    assert mutation.replacements["xl/styles.xml"] == addition.replacements["xl/styles.xml"]


def test_output_contract_rejects_unexpected_defined_name(tmp_path: Path) -> None:
    request = _minimal_request(tmp_path / "unused.xlsx")
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    mutation = build_allowlisted_mutation(base, request)
    normal = XlsxPackage.from_bytes(base.to_bytes(mutation))
    workbook = ET.fromstring(normal.parts["xl/workbook.xml"])
    expected = tuple(
        (
            tuple(sorted(item.attrib.items())),
            item.text or "",
        )
        for item in workbook.findall(
            f"{{{MAIN}}}definedNames/{{{MAIN}}}definedName"
        )
    )
    container = workbook.find(f"{{{MAIN}}}definedNames")
    assert container is not None
    ET.SubElement(container, f"{{{MAIN}}}definedName", {"name": "Injected"}).text = "1"
    tampered = XlsxPackage.from_bytes(
        normal.to_bytes(
            PackageMutation(
                replacements={
                    "xl/workbook.xml": ET.tostring(
                        workbook,
                        encoding="utf-8",
                        xml_declaration=True,
                    )
                }
            )
        )
    )

    with pytest.raises(ValueError, match="conjunto exacto de nombres definidos"):
        verify_output_contract(
            tampered,
            request.contract,
            request.mobiliti.row_map,
            cotizacion_total_row=request.cotizacion.total_row,
            expected_defined_names=expected,
        )


def test_cotizacion_declares_audited_formula_contract_for_official_f_i() -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    source = ET.fromstring(base.parts[base.sheet_part("Cotizacion")])
    assert _cell(source, "F17").findtext(f"{{{MAIN}}}f") == "Mobiliti!X14"
    assert _cell(source, "F17").findtext(f"{{{MAIN}}}v") == "0"
    assert _cell(source, "I17").find(f"{{{MAIN}}}f") is None
    assert _cell(source, "I17").find(f"{{{MAIN}}}v") is None

    contract = official_composer_module.CotizacionFormulaContract()
    product = CotizacionProduct(
        item_key="legacy-price-term",
        name="Silla",
        description="",
        dimensions="",
        quantity=Decimal("1"),
        mobiliti_row=14,
    )
    formulas = contract.product_formulas(
        price_terms=product.price_terms,
        target_row=17,
    )

    assert formulas == {
        "F": "=Mobiliti!X14",
        "I": "=F17-H17",
    }
    assert all("#REF!" not in formula for formula in formulas.values())


def test_composed_product_formula_uses_exact_mobiliti_terms() -> None:
    terms = (
        CotizacionPriceTerm(14, Decimal("1"), Decimal("1")),
        CotizacionPriceTerm(15, Decimal("2"), Decimal("1")),
        CotizacionPriceTerm(16, Decimal("3"), Decimal("10")),
    )

    formulas = official_composer_module.CotizacionFormulaContract().product_formulas(
        price_terms=terms,
        target_row=17,
    )

    assert formulas == {
        "F": "=Mobiliti!X14+Mobiliti!X15*2+Mobiliti!X16*3/10",
        "I": "=F17-H17",
    }


def test_composed_product_formula_serializes_decimal_factors_without_rounding() -> None:
    formulas = official_composer_module.CotizacionFormulaContract().product_formulas(
        price_terms=(
            CotizacionPriceTerm(14, Decimal("2.5000"), Decimal("0.1250")),
        ),
        target_row=17,
    )

    assert formulas["F"] == "=Mobiliti!X14*2.5/0.125"


def test_composed_product_formula_accepts_exact_excel_length_boundary() -> None:
    numerator = Decimal((0, (1,), 8177))

    formulas = official_composer_module.CotizacionFormulaContract().product_formulas(
        price_terms=(CotizacionPriceTerm(14, numerator),),
        target_row=17,
    )

    assert len(formulas["F"]) == 8192


def test_composed_product_formula_rejects_above_excel_length_boundary() -> None:
    numerator = Decimal((0, (1,), 8178))

    with pytest.raises(ValueError, match="excede.*8,192"):
        official_composer_module.CotizacionFormulaContract().product_formulas(
            price_terms=(CotizacionPriceTerm(14, numerator),),
            target_row=17,
        )


def test_composed_product_formula_rejects_592_terms_before_writing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    terms = tuple(CotizacionPriceTerm(row) for row in range(14, 14 + 592))
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    source = ET.fromstring(base.parts[base.sheet_part("Cotizacion")])
    formulas_written: list[str] = []
    original_set_formula = official_composer_module._set_formula

    def recording_set_formula(*args: object, **kwargs: object) -> None:
        formula = args[2] if len(args) >= 3 else kwargs["formula"]
        assert isinstance(formula, str)
        formulas_written.append(formula)
        original_set_formula(*args, **kwargs)

    monkeypatch.setattr(
        official_composer_module,
        "_set_formula",
        recording_set_formula,
    )
    product = CotizacionProduct(
        item_key="oversized-composition",
        name="Composición grande",
        description="",
        dimensions="",
        quantity=Decimal("1"),
        mobiliti_row=14,
        price_terms=terms,
    )

    with pytest.raises(ValueError, match="excede.*8,192"):
        CotizacionSheetEditor(source).compose(
            metadata=CotizacionMetadata(),
            sections=(
                CotizacionSection(
                    title="Proyecto",
                    products=(product,),
                ),
            ),
        )
    assert formulas_written == []


@pytest.mark.parametrize("value", (Decimal("1E+9000"), Decimal("1E-9000")))
def test_excel_decimal_rejects_unrepresentable_exponent_without_formatting(
    value: Decimal,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def forbidden_format(*_args: object, **_kwargs: object) -> str:
        raise AssertionError("No debe materializar el decimal antes de acotarlo")

    monkeypatch.setattr(official_composer_module, "format", forbidden_format, raising=False)

    with pytest.raises(ValueError, match="literal.*excede.*8,192"):
        official_composer_module._excel_decimal(value)


def test_composed_product_formula_rejects_tokens_outside_declared_contract(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        official_composer_module,
        "_excel_decimal",
        lambda _value: "2+Mobiliti!X99",
    )

    with pytest.raises(ValueError, match="fuera de contrato"):
        official_composer_module.CotizacionFormulaContract().product_formulas(
            price_terms=(CotizacionPriceTerm(14, Decimal("2")),),
            target_row=17,
        )


def test_composed_product_rejects_duplicate_mobiliti_rows() -> None:
    with pytest.raises(ValueError, match="duplicada"):
        CotizacionProduct(
            item_key="duplicate-price-term",
            name="Silla",
            description="",
            dimensions="",
            quantity=Decimal("1"),
            mobiliti_row=14,
            price_terms=(
                CotizacionPriceTerm(14),
                CotizacionPriceTerm(14, Decimal("2")),
            ),
        )


@pytest.mark.parametrize(
    ("numerator", "denominator"),
    (
        (Decimal("0"), Decimal("1")),
        (Decimal("-1"), Decimal("1")),
        (Decimal("1"), Decimal("0")),
        (Decimal("1"), Decimal("-1")),
    ),
)
def test_cotizacion_price_term_requires_positive_factors(
    numerator: Decimal,
    denominator: Decimal,
) -> None:
    with pytest.raises(ValueError, match="Factor de precio compuesto"):
        CotizacionPriceTerm(14, numerator, denominator)


@pytest.mark.parametrize("mobiliti_row", (0, 1_048_577, True))
def test_cotizacion_price_term_requires_valid_xlsx_row(
    mobiliti_row: object,
) -> None:
    with pytest.raises(ValueError, match="Fila Mobiliti del término"):
        CotizacionPriceTerm(mobiliti_row)  # type: ignore[arg-type]


def test_cotizacion_translation_preserves_reference_like_formula_literals() -> None:
    base = XlsxPackage.read(OFFICIAL_TEMPLATE)
    source = ET.fromstring(base.parts[base.sheet_part("Cotizacion")])
    _cell(source, "H17").find(f"{{{MAIN}}}f").text = (
        'IF("F17*G17"="literal",F17*G17,0)'
    )
    _cell(source, "J17").find(f"{{{MAIN}}}f").text = (
        'IF("E17*I17"="literal",E17*I17,0)'
    )
    section = CotizacionSection(
        title="Sillas",
        products=(
            CotizacionProduct(
                item_key="safe-item",
                name="Silla",
                description="Descripción",
                dimensions="60 x 60 cm",
                quantity=Decimal("1"),
                mobiliti_row=14,
            ),
        ),
    )

    result = CotizacionSheetEditor(source).compose(
        metadata=CotizacionMetadata(),
        sections=(section,),
    )
    worksheet = ET.fromstring(result.xml)

    assert '"F17*G17"' in _cell(worksheet, "H17").findtext(f"{{{MAIN}}}f")
    assert '"E17*I17"' in _cell(worksheet, "J17").findtext(f"{{{MAIN}}}f")


def _png_payload(color: tuple[int, int, int] = (10, 20, 30)) -> bytes:
    stream = BytesIO()
    Image.new("RGB", (16, 12), color).save(stream, format="PNG")
    return stream.getvalue()


def test_cotizacion_accepts_in_memory_image_bytes_without_path(tmp_path: Path) -> None:
    output = tmp_path / "in-memory-image.xlsx"
    request = _minimal_request(output)
    payload = _png_payload()
    cotizacion = replace(
        request.cotizacion,
        images=(
            CotizacionProductImage(
                path=None,
                target_row=request.cotizacion.product_rows[0],
                content=payload,
                content_type="image/png",
            ),
        ),
    )

    compose_official_quote(replace(request, cotizacion=cotizacion))

    package = XlsxPackage.read(output)
    assert payload in package.parts.values()


def test_product_image_path_rejects_reparse_points_before_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    image = (tmp_path / "reparse.png").resolve()
    image.write_bytes(_png_payload())
    monkeypatch.setattr(
        official_composer_module,
        "_path_is_reparse_point",
        lambda path: Path(path) == image,
    )

    with pytest.raises(ValueError, match="reparse point"):
        official_composer_module._read_product_image(image)


def test_product_image_path_is_read_once_and_pil_uses_same_bytes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    image = (tmp_path / "single-read.png").resolve()
    payload = _png_payload((30, 40, 50))
    image.write_bytes(payload)
    real_open = Image.open

    def bytes_only_open(source, *args, **kwargs):
        if isinstance(source, (str, os.PathLike)):
            raise AssertionError("PIL reabrió la ruta después de leer sus bytes")
        return real_open(source, *args, **kwargs)

    monkeypatch.setattr(Image, "open", bytes_only_open)

    content, _extension, _content_type, width, height = (
        official_composer_module._read_product_image(image)
    )

    assert content == payload
    assert (width, height) == (16, 12)


def test_in_memory_product_image_enforces_byte_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _png_payload()
    monkeypatch.setattr(official_composer_module, "MAX_IMAGE_BYTES", len(payload) - 1)
    descriptor = CotizacionProductImage(
        path=None,
        target_row=17,
        content=payload,
        content_type="image/png",
    )

    with pytest.raises(ValueError, match="Tamaño de imagen"):
        official_composer_module._read_product_image(descriptor)


def test_image_relationship_serialization_is_hashseed_deterministic(
    tmp_path: Path,
) -> None:
    image = tmp_path / "seed-image.png"
    image.write_bytes(_png_payload())
    script = r'''
from dataclasses import replace
from decimal import Decimal
from pathlib import Path
import sys
from xml.etree import ElementTree as ET

import mobiliti_saas.quote_engine.official_composer as composer
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage

template = Path(sys.argv[1])
image = Path(sys.argv[2])
base = XlsxPackage.read(template)
section = composer.CotizacionSection(
    title="Sillas",
    products=(composer.CotizacionProduct(
        item_key="seed-item",
        name="Silla",
        description="Descripción",
        dimensions="60 x 60 cm",
        quantity=Decimal("1"),
        mobiliti_row=14,
    ),),
)
mutation = composer.CotizacionSheetEditor.from_xml(
    base.parts[base.sheet_part("Cotizacion")]
).compose(metadata=composer.CotizacionMetadata(), sections=(section,))
mutation = replace(
    mutation,
    images=(composer.CotizacionProductImage(image.resolve(), 17),),
)
real_relationship_types = composer.relationship_type_uris
def seeded_relationship_types(name):
    if name == "image":
        return frozenset(("http://example.test/a-image", "http://example.test/b-image"))
    return real_relationship_types(name)
composer.relationship_type_uris = seeded_relationship_types
merged = composer.merge_cotizacion_product_images(base, mutation)
rels_payload = next(
    payload for name, payload in merged.related_parts.items() if name.endswith(".rels")
)
root = ET.fromstring(rels_payload)
relationship = next(
    item for item in root if "quote_product_" in item.attrib.get("Target", "")
)
print(relationship.attrib["Type"])
'''
    values = []
    for seed in ("1", "5"):
        environment = dict(os.environ, PYTHONHASHSEED=seed)
        completed = subprocess.run(
            [
                sys.executable,
                "-c",
                script,
                str(OFFICIAL_TEMPLATE),
                str(image),
            ],
            cwd=ROOT,
            env=environment,
            check=True,
            capture_output=True,
            text=True,
            timeout=120,
        )
        values.append(completed.stdout.strip())

    assert values[0] == values[1]


@pytest.mark.parametrize("trash_mode", ("no_op", "missing"))
def test_candidate_recovery_confirms_trash_or_returns_quarantine(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    trash_mode: str,
) -> None:
    output = tmp_path / "quote.xlsx"
    output.write_bytes(b"archivo del usuario")
    candidate = tmp_path / ".quote.xlsx.compose-1.tmp"
    candidate.write_bytes(b"candidato recuperable")
    profile = tmp_path / "profile"
    monkeypatch.setenv("USERPROFILE", str(profile))
    if trash_mode == "no_op":
        script = profile / ".codex" / "bin" / "Send-ToRecycleBin.ps1"
        script.parent.mkdir(parents=True)
        script.write_text("# herramienta simulada", encoding="utf-8")
        monkeypatch.setattr(
            official_composer_module.subprocess,
            "run",
            lambda *_args, **_kwargs: subprocess.CompletedProcess([], 0),
        )
    else:
        monkeypatch.setattr(official_composer_module.shutil, "which", lambda _name: None)

    quarantine = official_composer_module._recycle_candidate(candidate)

    assert quarantine is not None
    assert quarantine.exists()
    assert quarantine.read_bytes() == b"candidato recuperable"
    assert not candidate.exists()
    assert not tuple(tmp_path.glob(".*.compose-*.tmp"))
    assert output.read_bytes() == b"archivo del usuario"
