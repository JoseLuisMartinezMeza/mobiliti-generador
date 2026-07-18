from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from mobiliti_saas.quote_engine.engine import _sanitize_output_xlsx_for_excel, _write_header
from mobiliti_saas.quote_engine.supplier_catalog import create_supplier_quotation_workbook


def test_output_sanitizer_removes_empty_color_changes_and_safe_quotation_view(tmp_path: Path):
    workbook_path = tmp_path / "bad.xlsx"
    with ZipFile(workbook_path, "w", ZIP_DEFLATED) as zf:
        zf.writestr(
            "xl/workbook.xml",
            (
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                '<sheets><sheet name="Quotation" sheetId="13" r:id="rId13"/></sheets></workbook>'
            ),
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            (
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId13" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                'Target="worksheets/sheet13.xml"/></Relationships>'
            ),
        )
        zf.writestr(
            "xl/worksheets/sheet13.xml",
            (
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<sheetViews><sheetView showGridLines="0" zoomScale="110" zoomScaleNormal="110" '
                'workbookViewId="0"><selection pane="bottomLeft" activeCell="A99" '
                'sqref="A99:XFD99"/></sheetView></sheetViews><sheetData/></worksheet>'
            ),
        )
        zf.writestr(
            "xl/drawings/drawing1.xml",
            (
                '<wsDr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                "<clrChange><clrFrom/><clrTo/></clrChange>"
                "<a:clrChange><a:clrFrom/><a:clrTo/></a:clrChange>"
                "</wsDr>"
            ),
        )

    _sanitize_output_xlsx_for_excel(workbook_path)

    with ZipFile(workbook_path) as zf:
        drawing = zf.read("xl/drawings/drawing1.xml").decode("utf-8")
        quotation = zf.read("xl/worksheets/sheet13.xml").decode("utf-8")

    assert "clrChange" not in drawing
    assert 'activeCell="A1"' in quotation
    assert 'sqref="A1"' in quotation
    assert "XFD99" not in quotation


def test_supplier_intermediate_workbook_neutralizes_formula_injection(tmp_path: Path):
    output = tmp_path / "supplier-formula-injection.xlsx"
    payload = {
        "source_type": "supplier_cart",
        "supplier": "alma",
        "catalog_source_hash": "a" * 64,
        "base_currency": "USD",
        "quote_currency": "USD",
        "exchange_rate": "1.000000",
        "rate_source": "identity",
        "rate_effective_date": "2026-07-15",
        "rate_retrieved_at": "",
        "items": [
            {
                "internal_id": "alma:unsafe",
                "supplier": "alma",
                "product_key": "unsafe",
                "sku": "@BAD-SKU",
                "code_status": "verified",
                "brand": "KUN",
                "collection": "CLOGS",
                "name": '=HYPERLINK("https://evil.test","click")',
                "description": '=HYPERLINK("https://evil.test","click")',
                "unit": "+PZA",
                "availability_type": "made_to_order",
                "stock": None,
                "lead_time": "Sobre pedido",
                "quantity": "1",
                "base_option_id": None,
                "add_on_option_ids": [],
                "configuration": "-SUM(1,1)",
                "base_currency": "USD",
                "base_price": "10.000000",
                "unit_price_base": "10.000000",
                "unit_price": "10.00",
                "line_total": "10.00",
                "tax_rate": "0.160000",
                "attributes": {},
                "image_url": "",
                "image_kind": "official",
                "product_url": '=HYPERLINK("https://evil.test","url")',
                "warnings": [],
                "source_reference": "SPEC.xlsx:A1",
            }
        ],
    }

    create_supplier_quotation_workbook(payload, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]
    for reference in ("B9", "D9", "E9", "K9"):
        assert ws[reference].data_type != "f"
        assert not str(ws[reference].value or "").lstrip().startswith("=")
    assert str(ws["B9"].value).startswith("'=")
    assert "HYPERLINK" in str(ws["D9"].value)
    assert str(ws["K9"].value).startswith("'=")
    wb.close()


def test_supplier_final_header_neutralizes_formula_injection():
    wb = Workbook()
    ws = wb.active
    metadata = {
        "catalog_price_mode": "list_price_net",
        "cotizacion": "=1+1",
        "proyecto": '=HYPERLINK("https://evil.test","click")',
        "cliente": "+SUM(1,1)",
        "correo": "-2+3",
        "telefono": "@command",
        "direccion": "=cmd|' /C calc'!A0",
        "razon_social": "Cliente ordinario",
        "base_currency": "USD",
        "quote_currency": "MXN",
        "exchange_rate": "18.500000",
        "rate_source": "saas_exchange_rates",
        "rate_effective_date": "2026-07-15",
    }

    _write_header(ws, metadata)

    for reference in ("B3", "B7", "B8", "B9", "B10", "B11", "B12"):
        assert ws[reference].data_type != "f"
    for reference in ("B3", "B7", "B8", "B9", "B10", "B11"):
        assert str(ws[reference].value).startswith("'")
    assert ws["B12"].value == "Cliente ordinario"
    wb.close()
