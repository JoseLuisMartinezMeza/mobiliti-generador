from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from datetime import date
import hashlib
from io import BytesIO
import os
from pathlib import Path
import re
import time
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string
from PIL import Image
import pytest

from mobiliti_saas.quote_engine.engine import generate_quote
from mobiliti_saas.quote_engine.mixed_catalog import (
    build_mixed_catalog_cart_payload,
    create_mixed_catalog_quotation_workbook,
)
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from mobiliti_saas.quote_engine.quotation_import import build_import_manifest
from mobiliti_saas.quote_engine.quotation_sheets import quotation_data_rows
from mobiliti_saas.quote_engine.template_profiles import (
    OFFICIAL_TEMPLATE_PROFILE_ID,
    SUNON_CDMX_TEMPLATE_PROFILE_ID,
    resolve_template_profile,
)


ROOT = Path(__file__).resolve().parents[1]
REAL_SOURCE = Path.home() / "Downloads" / "IZA REFORMA-Quotation Sheet - V1.xlsx"
ARTIFACT_DIR = ROOT / "artifacts" / "template-variant-e2e"
OFFICIAL_OUTPUT = (
    ARTIFACT_DIR / "official-project-import-output-cdmx-subtotals-20260730.xlsx"
)
CDMX_OUTPUT = (
    ARTIFACT_DIR
    / "sunon-cdmx-v1c-project-import-output-cdmx-subtotals-20260730.xlsx"
)
IMPORT_ID = "11111111-1111-4111-8111-111111111111"
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
METADATA = {
    "cotizacion": "E2E-PARIDAD-20260730",
    "proyecto": "IZA REFORMA - PARIDAD",
    "cliente": "QA Local",
    "correo": "qa@example.test",
    "telefono": "000",
    "direccion": "Guadalajara",
    "razon_social": "Mobiliti QA",
    "descuento": 40,
    "tipo_cambio": 18.5,
}


@dataclass(frozen=True)
class PairedOutputs:
    official: Path
    cdmx: Path
    durations: dict[str, float]


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _valid_reusable_output(path: Path, expected_profile: str) -> bool:
    if not path.is_file():
        return False
    try:
        package = XlsxPackage.read(path)
        if package.sheet_state("Quotation_Data") != "veryHidden":
            return False
        package.sheet_part("Quotation")
        package.sheet_part("Cotizacion")
        package.sheet_part("Mobiliti")
        if expected_profile == SUNON_CDMX_TEMPLATE_PROFILE_ID:
            package.sheet_part("Cantidades Lumbro ")
            try:
                package.sheet_part("COSTO LUMBRO ")
            except KeyError:
                pass
            else:
                return False
        origins = _quotation_data_column(package, "E")
        item_keys = tuple(str(value) for value in _quotation_data_column(package, "A"))
        if (
            len(origins) != 33
            or set(origins) != {"imported"}
            or any(":lumbro:" in item_key for item_key in item_keys)
        ):
            return False
        return True
    except (KeyError, OSError, ValueError):
        return False


def _project_import_handoff(work_root: Path) -> tuple[Path, dict, tuple]:
    """Replica el handoff canónico de Previsualizar e importar al Proyecto."""

    manifest, _images = build_import_manifest(
        REAL_SOURCE.read_bytes(),
        import_id=IMPORT_ID,
        original_filename=REAL_SOURCE.name,
    )
    line_id_by_key: dict[str, str] = {}
    imported_items: list[dict] = []
    for position, item in enumerate(manifest["items"], start=1):
        line_id = f"import-line-{position}"
        line_id_by_key[item["key"]] = line_id
        imported_items.append(
            {
                "line_id": line_id,
                "kind": "imported",
                "import_id": manifest["import_id"],
                "source_row": item["source_row"],
                "source_currency": "USD",
                "quantity": item["quantity"],
                "overrides": {
                    "name": item["name"],
                    "description": item["description"],
                    "dimension": item["dimension"],
                    "unit_price": item["unit_price"],
                    "provider": item["provider"] or manifest["provider"],
                },
            }
        )
    sections = [
        {
            "id": f"section-{position}",
            "title": section["title"],
            "line_ids": [
                line_id_by_key[item_key] for item_key in section["item_keys"]
            ],
        }
        for position, section in enumerate(manifest["sections"], start=1)
    ]
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[],
        quote_currency="USD",
        commercial_discount_percent="40",
        presentation_sections=sections,
        imported_source={
            "manifest": manifest,
            "items": imported_items,
            "source_currency": "USD",
        },
        today=date(2026, 7, 30),
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        work_root / "project-import-source.xlsx",
        image_dir=work_root / "project-import-images",
        imported_source_path=REAL_SOURCE,
    )
    metadata = {
        **METADATA,
        "catalog_price_mode": "mixed_catalog_converted",
        "catalog_source_hashes": {},
        "quote_currency": "USD",
        "rate_summary": deepcopy(payload["rate_summary"]),
        "auto_electrification_rate": deepcopy(
            payload["auto_electrification_rate"]
        ),
        "image_provider": "pillow",
        "image_cleanup_strength": "balanced",
        "image_background": "white",
        "image_prompt": (
            "Mejora la calidad de imagen y que este en fondo blanco"
        ),
    }
    return parser_source, metadata, quotation_data_rows(payload)


@pytest.fixture(scope="module")
def paired_outputs(tmp_path_factory) -> PairedOutputs:
    """Reutiliza artefactos auditados; regenera solo por solicitud o ausencia."""

    if not REAL_SOURCE.is_file():
        pytest.fail(f"Falta el input real obligatorio: {REAL_SOURCE}")

    work_root = tmp_path_factory.mktemp("template-variant-project-import")
    parser_source, metadata, canonical_rows = _project_import_handoff(work_root)
    force = os.getenv("MOBILITI_FORCE_TEMPLATE_VARIANT_E2E") == "1"
    durations: dict[str, float] = {}
    profiles = (
        (OFFICIAL_TEMPLATE_PROFILE_ID, OFFICIAL_OUTPUT),
        (SUNON_CDMX_TEMPLATE_PROFILE_ID, CDMX_OUTPUT),
    )
    for profile_id, output in profiles:
        if not force and _valid_reusable_output(output, profile_id):
            durations[profile_id] = 0.0
            continue
        if output.exists():
            pytest.fail(
                "El artefacto existente no es reutilizable y no se sobrescribe "
                f"automáticamente: {output}"
            )
        output.parent.mkdir(parents=True, exist_ok=True)
        profile = resolve_template_profile(profile_id)
        started = time.perf_counter()
        generate_quote(
            parser_source,
            output,
            dict(metadata),
            profile.template_path,
            original_quotation_path=REAL_SOURCE,
            quotation_data_rows=canonical_rows,
        )
        durations[profile_id] = time.perf_counter() - started

    return PairedOutputs(
        official=OFFICIAL_OUTPUT,
        cdmx=CDMX_OUTPUT,
        durations=durations,
    )


def _shared_strings(package: XlsxPackage) -> tuple[str, ...]:
    part = package.workbook_related_part("sharedStrings")
    if part is None:
        return ()
    root = ET.fromstring(package.parts[part])
    return tuple(
        "".join(node.text or "" for node in item.iter(f"{{{MAIN}}}t"))
        for item in root.findall(f"{{{MAIN}}}si")
    )


def _cell_semantics(
    package: XlsxPackage,
    sheet_name: str,
    *,
    allowed_columns: frozenset[str] | None = None,
    max_row: int | None = None,
) -> tuple[tuple[object, ...], ...]:
    """Firma de valores/fórmulas que ignora estilos y cachés de cálculo."""

    shared = _shared_strings(package)
    root = ET.fromstring(package.parts[package.sheet_part(sheet_name)])
    result: list[tuple[object, ...]] = []
    for cell in root.findall(f".//{{{MAIN}}}c"):
        coordinate = cell.attrib.get("r", "")
        column = "".join(character for character in coordinate if character.isalpha())
        row_text = "".join(character for character in coordinate if character.isdigit())
        row = int(row_text or 0)
        if allowed_columns is not None and column not in allowed_columns:
            continue
        if max_row is not None and row > max_row:
            continue

        formula = cell.find(f"{{{MAIN}}}f")
        if formula is not None:
            result.append(
                (
                    coordinate,
                    "formula",
                    formula.text or "",
                    formula.attrib.get("t", ""),
                    formula.attrib.get("ref", ""),
                )
            )
            continue

        cell_type = cell.attrib.get("t", "")
        value_node = cell.find(f"{{{MAIN}}}v")
        if cell_type == "s" and value_node is not None:
            value: object = shared[int(value_node.text or "0")]
        elif cell_type == "inlineStr":
            value = "".join(
                node.text or "" for node in cell.iter(f"{{{MAIN}}}t")
            )
        elif value_node is not None:
            value = value_node.text or ""
        else:
            inline = "".join(
                node.text or "" for node in cell.iter(f"{{{MAIN}}}t")
            )
            value = inline
        if value != "":
            result.append((coordinate, cell_type, value))
    return tuple(result)


def _sheet_names(package: XlsxPackage) -> tuple[str, ...]:
    root = ET.fromstring(package.parts["xl/workbook.xml"])
    return tuple(
        sheet.attrib["name"]
        for sheet in root.findall(f".//{{{MAIN}}}sheet")
    )


def _column_number(coordinate: str) -> int:
    column = "".join(character for character in coordinate if character.isalpha())
    return column_index_from_string(column)


def _ref_formula_baseline(package: XlsxPackage) -> tuple[str, ...]:
    formulas: list[str] = []
    for part_name, content in package.parts.items():
        if not part_name.startswith("xl/worksheets/") or not part_name.endswith(
            ".xml"
        ):
            continue
        root = ET.fromstring(content)
        formulas.extend(
            formula.text or ""
            for formula in root.findall(f".//{{{MAIN}}}f")
            if "#REF!" in (formula.text or "").upper()
        )
    return tuple(sorted(formulas))


def _ref_defined_name_baseline(package: XlsxPackage) -> tuple[tuple[str, str], ...]:
    root = ET.fromstring(package.parts["xl/workbook.xml"])
    return tuple(
        sorted(
            (
                node.attrib.get("name", ""),
                node.text or "",
            )
            for node in root.findall(f".//{{{MAIN}}}definedName")
            if "#REF!" in (node.text or "").upper()
        )
    )


def _media_audit(package: XlsxPackage) -> dict[str, int]:
    media = {
        name: len(content)
        for name, content in package.parts.items()
        if name.startswith("xl/media/")
    }
    assert media, "El libro final debe conservar imágenes"
    assert all(size > 0 for size in media.values())
    drawings = [
        name
        for name in package.parts
        if name.startswith("xl/drawings/") and name.endswith(".xml")
    ]
    assert drawings, "El libro final debe conservar drawings"
    return media


def _quotation_data_column(
    package: XlsxPackage,
    column: str,
) -> tuple[object, ...]:
    return tuple(
        entry[-1]
        for entry in _cell_semantics(package, "Quotation_Data")
        if entry[0].startswith(column)
        and int("".join(character for character in entry[0] if character.isdigit()))
        > 1
    )


def _cotizacion_product_reference_signature(
    package: XlsxPackage,
) -> tuple[tuple[str, ...], ...]:
    semantics = _cell_semantics(
        package,
        "Cotizacion",
        allowed_columns=frozenset("ACDEFGHIJ"),
    )
    formulas = {
        str(entry[0]): str(entry[2])
        for entry in semantics
        if len(entry) >= 3 and entry[1] == "formula"
    }
    product_rows = sorted(
        int("".join(character for character in coordinate if character.isdigit()))
        for coordinate, formula in formulas.items()
        if coordinate.startswith("A") and formula.startswith("Mobiliti!D")
    )
    return tuple(
        tuple(formulas.get(f"{column}{row}", "") for column in "ACDEF")
        for row in product_rows
    )


def _cdmx_section_subtotal_audit(
    package: XlsxPackage,
) -> tuple[tuple[int, ...], tuple[int, ...], str]:
    semantics = _cell_semantics(
        package,
        "Cotizacion",
        allowed_columns=frozenset("AHIJ"),
    )
    values = {
        str(entry[0]): str(entry[2])
        for entry in semantics
        if len(entry) >= 3 and entry[1] != "formula"
    }
    formulas = {
        str(entry[0]): str(entry[2])
        for entry in semantics
        if len(entry) >= 3 and entry[1] == "formula"
    }
    product_rows = tuple(
        sorted(
            int(
                "".join(
                    character
                    for character in coordinate
                    if character.isdigit()
                )
            )
            for coordinate, formula in formulas.items()
            if coordinate.startswith("A") and formula.startswith("Mobiliti!D")
        )
    )
    subtotal_rows = tuple(
        sorted(
            int(
                "".join(
                    character
                    for character in coordinate
                    if character.isdigit()
                )
            )
            for coordinate, value in values.items()
            if coordinate.startswith("I") and value == "SUBTOTAL AREA"
        )
    )
    covered_product_rows: list[int] = []
    for subtotal_row in subtotal_rows:
        formula = formulas.get(f"J{subtotal_row}", "")
        match = re.fullmatch(r"SUM\(J([1-9][0-9]*):J([1-9][0-9]*)\)", formula)
        assert match is not None, (
            f"Subtotal CDMX J{subtotal_row} no suma su sección completa: "
            f"{formula!r}"
        )
        first_product_row, last_product_row = map(int, match.groups())
        assert last_product_row == subtotal_row - 1
        covered_product_rows.extend(range(first_product_row, last_product_row + 1))

    expected_global = "SUM(" + ",".join(
        f"J{subtotal_row}" for subtotal_row in subtotal_rows
    ) + ")"
    global_subtotals = tuple(
        formula
        for coordinate, formula in formulas.items()
        if coordinate.startswith("H") and formula == expected_global
    )
    assert global_subtotals == (expected_global,)
    return product_rows, tuple(covered_product_rows), expected_global


def test_real_input_uses_project_import_semantics_without_legacy_lumbro(
    paired_outputs: PairedOutputs,
) -> None:
    for output in (paired_outputs.official, paired_outputs.cdmx):
        package = XlsxPackage.read(output)
        origins = _quotation_data_column(package, "E")
        item_keys = tuple(str(value) for value in _quotation_data_column(package, "A"))

        assert len(origins) == 33
        assert set(origins) == {"imported"}
        assert not any(":lumbro:" in item_key for item_key in item_keys)


def test_real_input_uses_the_current_processed_image_pipeline(
    paired_outputs: PairedOutputs,
) -> None:
    for output in (paired_outputs.official, paired_outputs.cdmx):
        with ZipFile(output) as archive:
            names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/media/quote_product_")
                and name.endswith(".png")
            )
            assert len(names) == 33
            for name in names:
                with Image.open(BytesIO(archive.read(name))) as image:
                    assert image.mode == "RGB"
                    assert image.width >= 900
                    assert image.height >= 900


def test_real_input_preserves_core_semantics_and_adds_cdmx_section_totals(
    paired_outputs: PairedOutputs,
) -> None:
    official = XlsxPackage.read(paired_outputs.official)
    cdmx = XlsxPackage.read(paired_outputs.cdmx)

    official_names = set(_sheet_names(official))
    cdmx_names = set(_sheet_names(cdmx))
    assert official_names < cdmx_names
    assert cdmx_names - official_names == {"Cantidades Lumbro "}

    official_audit = official.parts[official.sheet_part("Quotation_Data")]
    cdmx_audit = cdmx.parts[cdmx.sheet_part("Quotation_Data")]
    assert cdmx_audit == official_audit

    assert _cell_semantics(cdmx, "Quotation") == _cell_semantics(
        official, "Quotation"
    )
    assert _cell_semantics(cdmx, "Mobiliti") == _cell_semantics(
        official, "Mobiliti"
    )
    assert _cotizacion_product_reference_signature(
        cdmx
    ) == _cotizacion_product_reference_signature(official)

    product_rows, covered_product_rows, global_subtotal = (
        _cdmx_section_subtotal_audit(cdmx)
    )
    assert len(product_rows) == 33
    assert len(set(covered_product_rows)) == len(covered_product_rows)
    assert tuple(sorted(covered_product_rows)) == product_rows
    assert global_subtotal.count("J") == 9

    cdmx_text = "\n".join(
        str(entry[2])
        for entry in _cell_semantics(cdmx, "Cotizacion")
        if len(entry) >= 3 and entry[1] != "formula"
    )
    assert "COMERCIALIZADORA VICARJOFRAA DE OCCIDENTE" in cdmx_text
    assert (
        "Pago 70% Anticipo + 20% Contra Aviso de Embarque "
        "+ 10% Contra Entrega"
    ) in cdmx_text
    assert "10-12 SEMANAS" in cdmx_text
    assert "Ciudad de México, CDMX" in cdmx_text
    assert "Pago 60% Anticipo" not in cdmx_text


def test_real_outputs_are_valid_ooxml_with_no_new_ref_damage(
    paired_outputs: PairedOutputs,
) -> None:
    official = XlsxPackage.read(paired_outputs.official)
    cdmx = XlsxPackage.read(paired_outputs.cdmx)

    assert official.sheet_state("Quotation_Data") == "veryHidden"
    assert cdmx.sheet_state("Quotation_Data") == "veryHidden"
    assert cdmx.sheet_state("Cantidades Lumbro ") == "visible"
    with pytest.raises(KeyError):
        cdmx.sheet_part("COSTO LUMBRO ")
    assert _ref_formula_baseline(cdmx) == _ref_formula_baseline(official)
    assert _ref_defined_name_baseline(cdmx) == _ref_defined_name_baseline(
        official
    )
    assert len(_media_audit(official)) > 0
    assert len(_media_audit(cdmx)) > 0


def _write_lumbro_stress_source(path: Path, count: int = 30) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    headers = {
        1: "No.",
        2: "Item Name",
        4: "Description",
        5: "Dimension",
        6: "Provider",
        7: "Qty",
        10: "Unit Price",
        14: "Original Currency",
    }
    for column, value in headers.items():
        sheet.cell(row=7, column=column, value=value)
    sheet.cell(row=8, column=1, value="- LUMBRO STRESS")
    for offset in range(count):
        row = 9 + offset
        sheet.cell(row=row, column=1, value=offset + 1)
        sheet.cell(
            row=row,
            column=2,
            value=f"MULTICONTACTO LUMBRO E2E {offset + 1:02d}",
        )
        sheet.cell(
            row=row,
            column=4,
            value=f"Multicontacto de validación {offset + 1:02d}",
        )
        sheet.cell(row=row, column=5, value="245 x 102 x 60 mm")
        sheet.cell(row=row, column=6, value="Lumbro CH")
        sheet.cell(row=row, column=7, value=offset + 1)
        sheet.cell(row=row, column=10, value=1000 + offset)
        sheet.cell(row=row, column=14, value="MXN")
    workbook.save(path)


def test_cdmx_lumbro_surface_expands_to_thirty_live_rows(
    tmp_path: Path,
) -> None:
    source = tmp_path / "lumbro-30-source.xlsx"
    output = tmp_path / "lumbro-30-output.xlsx"
    _write_lumbro_stress_source(source)
    profile = resolve_template_profile(SUNON_CDMX_TEMPLATE_PROFILE_ID)

    generate_quote(
        source,
        output,
        {
            **METADATA,
            "cotizacion": "E2E-LUMBRO-30",
            "proyecto": "CAPACIDAD LUMBRO 30",
            "catalog_supplier_label": "Lumbro CH",
        },
        profile.template_path,
    )

    package = XlsxPackage.read(output)
    quotation_data = _cell_semantics(package, "Quotation_Data")
    assert sum(
        1
        for cell in quotation_data
        if cell[0].startswith("E") and cell[-1] == "quotation"
    ) >= 30

    lumbro = dict(
        (entry[0], entry[2])
        for entry in _cell_semantics(package, "Cantidades Lumbro ")
        if len(entry) >= 3 and entry[1] == "formula"
    )
    assert "_xlws.FILTER" in lumbro["H4"]
    assert "Mobiliti!$D$14:$D$5000" in lumbro["H4"]
    assert "_xlws.FILTER" in lumbro["I4"]
    assert "Mobiliti!$H$14:$H$5000" in lumbro["I4"]
    assert "_xlws.FILTER" in lumbro["L4"]
    assert "Mobiliti!$J$14:$J$5000" in lumbro["L4"]

    root = ET.fromstring(
        package.parts[package.sheet_part("Cantidades Lumbro ")]
    )
    dimension = root.find(f"{{{MAIN}}}dimension")
    assert dimension is not None
    assert dimension.attrib["ref"] == "A1:P33"
    formula_coordinates = {
        cell.attrib["r"]
        for cell in root.findall(f".//{{{MAIN}}}c")
        if cell.find(f"{{{MAIN}}}f") is not None
    }
    assert formula_coordinates == {
        "M3",
        "H4",
        "I4",
        "J4",
        "K4",
        "L4",
        "M4",
        "N4",
        "P4",
    }
    row_styles = {
        int(row.attrib["r"]): tuple(
            cell.attrib.get("s", "")
            for cell in row.findall(f"{{{MAIN}}}c")
            if 8
            <= _column_number(cell.attrib.get("r", ""))
            <= 16
        )
        for row in root.findall(f".//{{{MAIN}}}row")
        if 4 <= int(row.attrib.get("r", "0")) <= 33
    }
    assert set(row_styles) == set(range(4, 34))
    assert all(styles == row_styles[4] for styles in row_styles.values())

    workbook = ET.fromstring(package.parts["xl/workbook.xml"])
    print_areas = [
        item.text
        for item in workbook.findall(
            f"{{{MAIN}}}definedNames/{{{MAIN}}}definedName"
        )
        if item.attrib.get("name") == "_xlnm.Print_Area"
        and item.attrib.get("localSheetId")
        == str(package.sheet_index("Cantidades Lumbro "))
    ]
    assert print_areas == ["'Cantidades Lumbro '!$H$1:$P$33"]


def test_excel_com_open_calculate_and_reopen_when_explicitly_enabled(
    paired_outputs: PairedOutputs,
) -> None:
    if os.getenv("MOBILITI_RUN_EXCEL_COM_E2E") != "1":
        pytest.skip(
            "COM es opt-in para evitar bloqueo con otras automatizaciones de Excel"
        )

    import pythoncom
    import win32com.client

    totals: list[float] = []
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    try:
        for path in (paired_outputs.official, paired_outputs.cdmx):
            workbook = excel.Workbooks.Open(
                str(path.resolve()),
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                CorruptLoad=0,
            )
            try:
                excel.CalculateFullRebuild()
                deadline = time.monotonic() + 600
                while excel.CalculationState != 0:
                    if time.monotonic() >= deadline:
                        pytest.fail(f"Excel no terminó de calcular en 600 s: {path}")
                    time.sleep(0.25)
                total = workbook.Worksheets("Cotizacion").Range("H71").Value2
                assert isinstance(total, (int, float))
                totals.append(float(total))
            finally:
                workbook.Close(SaveChanges=False)
        assert totals[0] == pytest.approx(totals[1], rel=0, abs=0.01)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


def test_artifact_hashes_are_stable_and_reportable(
    paired_outputs: PairedOutputs,
) -> None:
    assert _sha256(REAL_SOURCE) == (
        "ea40b9b7d16cec7765a13227b02b65a444938cc46b22bd11cad76e76a0d67f22"
    )
    assert len(_sha256(paired_outputs.official)) == 64
    assert len(_sha256(paired_outputs.cdmx)) == 64
