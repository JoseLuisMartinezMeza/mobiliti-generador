from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import hashlib
import json
from pathlib import Path
import re
import subprocess
import sys
import uuid
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from mobiliti_saas.quote_engine import engine
from mobiliti_saas.quote_engine.mobiliti_layout import (
    SectionNeed,
    plan_mobiliti_layout,
)
from mobiliti_saas.quote_engine.ooxml_formula import translate_formula
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from mobiliti_saas.quote_engine.ooxml_package import assert_package_preserved
from mobiliti_saas.quote_engine.quotation_import import (
    MAX_QUOTE_REQUEST_BYTES,
    MOBILITI_RESERVED_ROWS_AFTER_TOTAL,
    XLSX_MAX_ROWS,
    build_import_manifest,
    required_mobiliti_rows,
)
from mobiliti_saas.web.api import index as web_api
from quotation_import_fixtures import write_import_fixture
from test_official_quote_stress import (
    OFFICIAL_ALLOWED_PARTS,
    _cell_map,
    _cell_text,
    _formula,
    quote_worker,
)
from test_official_template_contract import assert_official_template_contract
from test_project_quote_engine import (
    FIXED_ID,
    PER_UNIT_ID,
    PRINCIPAL_ID,
    _project_payload,
    _supplier_item,
)
from test_quotation_sheet_transplant import quotation_semantic_signature
from test_quotation_sheet_transplant import _canonical


ROOT = Path(__file__).resolve().parents[1]
OFFICIAL_TEMPLATE = engine.OFFICIAL_TEMPLATE_PATH
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _formula_semantic_xml(content: bytes) -> tuple:
    root = ET.fromstring(content)
    for cell in root.findall(f".//{{{MAIN}}}c"):
        if cell.find(f"{{{MAIN}}}f") is None:
            continue
        cached = cell.find(f"{{{MAIN}}}v")
        if cached is not None:
            cell.remove(cached)
    return _canonical(root)


def _auth_headers() -> dict[str, str]:
    token = web_api.create_access_token(
        {"sub": "1", "email": web_api.DEV_USER_EMAIL}
    )
    return {"Authorization": f"Bearer {token}"}


def _catalog_snapshot() -> dict:
    return {
        "supplier": "sunon",
        "source_hash": hashlib.sha256(b"task-8-persisted-sunon").hexdigest(),
        "generated_at": "2026-07-23T00:00:00+00:00",
        "items": [
            _supplier_item("sunon:main-1", "MAIN-1", "Principal"),
            _supplier_item("sunon:per-1", "PER-1", "Complemento por unidad"),
            _supplier_item("sunon:fixed-1", "FIXED-1", "Complemento fijo"),
        ],
    }


@dataclass(frozen=True)
class PersistedQuoteResult:
    output: Path
    frozen_payload: dict
    project: dict
    job: dict
    layout: object
    original_quotation: Path | None
    claim_events: tuple[str, ...]


def _configure_persisted_runtime(
    monkeypatch: pytest.MonkeyPatch,
    store_dir: Path,
) -> None:
    monkeypatch.setattr(web_api, "JWT_SECRET_KEY", "task-8-persisted-secret")
    monkeypatch.setattr(web_api, "DEV_MODE", True)
    monkeypatch.setattr(web_api, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(quote_worker, "DEV_MODE", True)
    monkeypatch.setattr(quote_worker, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")
    monkeypatch.setattr(
        web_api,
        "db_get_usuario_by_id",
        lambda user_id: {
            "id": int(user_id),
            "email": web_api.DEV_USER_EMAIL,
            "activo": True,
            "es_admin": True,
        },
    )
    monkeypatch.setattr(web_api, "_require_active_subscription", lambda _user_id: None)
    monkeypatch.setattr(web_api, "_require_enabled_catalog_supplier", lambda value: value)
    monkeypatch.setattr(
        web_api,
        "_load_supplier_catalog_cached",
        lambda supplier: (
            _catalog_snapshot()
            if supplier == "sunon"
            else (_ for _ in ()).throw(
                AssertionError(f"Catalogo inesperado: {supplier}")
            )
        ),
    )
    effective = (date.today() - timedelta(days=1)).isoformat()
    monkeypatch.setattr(
        web_api,
        "db_list_exchange_rates",
        lambda: [
            {
                "currency": "USD",
                "effective_date": effective,
                "mxn_per_unit": "18.500000",
                "retrieved_at": f"{effective}T20:00:00Z",
            }
        ],
    )
    monkeypatch.setattr(web_api, "_next_quote_number_for_user", lambda _user: None)
    monkeypatch.setattr(
        web_api, "_enforce_active_quote_limit", lambda *_args, **_kwargs: None
    )
    monkeypatch.setattr(web_api, "_wake_worker", lambda: None)
    # Los artefactos de aceptación son evidencia; no se eliminan al completar.
    monkeypatch.setattr(quote_worker, "_delete_job_input", lambda *_args: None)
    monkeypatch.setattr(
        quote_worker, "_cleanup_completed_import_source", lambda *_args: True
    )


def _seed_import_job(source: Path) -> tuple[str, dict]:
    import_id = str(uuid.uuid4())
    source_bytes = source.read_bytes()
    manifest, images = build_import_manifest(source_bytes, import_id, source.name)
    # El source de esta aceptación no aporta una imagen nueva a Cotizacion;
    # así el conjunto de adiciones OOXML queda cerrado exactamente por las dos
    # hojas canónicas que este contrato pretende auditar.
    preview_rows = [row for row in sorted(images) if row != 11]
    prefix = f"users/1/jobs/{import_id}/"
    preview_prefix = f"{prefix}preview/{manifest['source_hash'][:16]}"
    preview_paths = {
        str(row): f"{preview_prefix}/row-{row}.png"
        for row in preview_rows
    }
    input_path = f"{prefix}input.xlsx"
    manifest_path = f"{preview_prefix}/manifest.json"
    metadata = {
        "original_filename": source.name,
        "import_manifest_path": manifest_path,
        "import_preview_paths": preview_paths,
        "import_source_hash": manifest["source_hash"],
        "import_item_count": len(manifest["items"]),
    }
    web_api.db_create_quote_job(
        1,
        "import-preview",
        metadata,
        input_path,
        job_id=import_id,
    )
    web_api._storage_upload_bytes(
        input_path,
        source_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    for row in preview_rows:
        content, _image_type = images[row]
        web_api._storage_upload_bytes(preview_paths[str(row)], content, "image/png")
    web_api._storage_upload_bytes(
        manifest_path,
        json.dumps(
            {**manifest, "preview_image_paths": preview_paths},
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8"),
        "application/json",
    )
    return import_id, manifest


def _imported_line(
    *,
    line_id: str,
    section_id: str,
    position: int,
    manifest: dict,
    promotion: dict,
) -> dict:
    item = next(row for row in manifest["items"] if row["source_row"] == 11)
    return {
        "line_id": line_id,
        "role": "principal",
        "section_id": section_id,
        "parent_line_id": None,
        "position": position,
        "quantity": "1",
        "source": "imported",
        "import_id": manifest["import_id"],
        "source_row": item["source_row"],
        "source_currency": "USD",
        "official_code": item.get("official_code") or "",
        "provider": manifest["provider"],
        "name": item["name"],
        "description": item["description"],
        "dimension": item["dimension"],
        "unit_price": item["unit_price"],
        "image_asset_key": promotion["image_asset_keys"].get(
            str(item["source_row"]), ""
        ),
        "source_asset_key": promotion["source_asset_key"],
        "display_cache": {
            "name": item["name"],
            "code": item.get("official_code") or "",
            "image_url": "",
        },
    }


def _stress_project_payload() -> dict:
    payload = _project_payload()
    payload["quote_fields"]["proyecto"] = "Proyecto stress persistido"
    section_ids = (
        "section-1",
        *(f"section-{section}" for section in range(5, 24)),
    )
    payload["sections"] = [
        {
            "section_id": section_id,
            "concept": f"Stress Section {position}",
            "position": position - 1,
        }
        for position, section_id in enumerate(section_ids, start=1)
    ]
    lines: list[dict] = []
    principal_counter = 0
    for section_position, section_id in enumerate(section_ids, start=1):
        for offset in range(35):
            line_id = str(uuid.uuid4())
            if section_position == 1 and offset == 0:
                lines.append(
                    {
                        "line_id": line_id,
                        "role": "import-placeholder",
                        "section_id": section_id,
                        "parent_line_id": None,
                        "position": 0,
                    }
                )
                parent_id = line_id
                continue
            if section_position == 1 and offset in {1, 2}:
                lines.append(
                    {
                        "line_id": line_id,
                        "role": "complement",
                        "section_id": None,
                        "parent_line_id": parent_id,
                        "position": offset - 1,
                        "quantity": "1",
                        "quantity_mode": (
                            "per_parent_unit"
                            if offset == 1
                            else "fixed_project"
                        ),
                        "source": "catalog",
                        "catalog": "sunon",
                        "official_code": "PER-1" if offset == 1 else "FIXED-1",
                        "identity": {
                            "internal_id": (
                                "sunon:per-1"
                                if offset == 1
                                else "sunon:fixed-1"
                            ),
                            "base_option_id": "",
                            "add_on_option_ids": [],
                        },
                        "display_cache": {
                            "name": "Complemento",
                            "code": "PER-1" if offset == 1 else "FIXED-1",
                            "image_url": "",
                        },
                    }
                )
                continue
            principal_counter += 1
            lines.append(
                {
                    "line_id": line_id,
                    "role": "principal",
                    "section_id": section_id,
                    "parent_line_id": None,
                    "position": offset - 2 if section_position == 1 else offset,
                    "quantity": "1",
                    "source": "catalog",
                    "catalog": "sunon",
                    "official_code": "MAIN-1",
                    "identity": {
                        "internal_id": "sunon:main-1",
                        "base_option_id": "",
                        "add_on_option_ids": [],
                    },
                    "display_cache": {
                        "name": f"Principal {principal_counter}",
                        "code": "MAIN-1",
                        "image_url": "",
                    },
                }
            )
    payload["lines"] = lines
    return payload


def _run_persisted_project_case(
    case_dir: Path,
    payload: dict,
    *,
    imported_source: Path | None = None,
) -> PersistedQuoteResult:
    monkeypatch = pytest.MonkeyPatch()
    _configure_persisted_runtime(monkeypatch, case_dir / "dev-store")
    client = TestClient(web_api.app)
    claim_events: list[str] = []
    try:
        initial_payload = (
            _project_payload() if imported_source is not None else deepcopy(payload)
        )
        initial_payload["quote_fields"]["quote_currency"] = payload["quote_fields"][
            "quote_currency"
        ]
        created = client.post(
            "/projects",
            headers=_auth_headers(),
            json={"name": case_dir.name, "payload": initial_payload},
        )
        assert created.status_code == 201, created.json()
        project = created.json()["project"]

        if imported_source is not None:
            import_id, manifest = _seed_import_job(imported_source)
            promoted = client.post(
                f"/projects/{project['id']}/imports/{import_id}",
                headers=_auth_headers(),
            )
            assert promoted.status_code == 200, promoted.json()
            final_payload = deepcopy(payload)
            placeholder = next(
                line
                for line in final_payload["lines"]
                if line["role"] == "import-placeholder"
            )
            final_payload["lines"][
                final_payload["lines"].index(placeholder)
            ] = _imported_line(
                line_id=placeholder["line_id"],
                section_id=placeholder["section_id"],
                position=placeholder["position"],
                manifest=manifest,
                promotion=promoted.json(),
            )
            patched = client.patch(
                f"/projects/{project['id']}",
                headers=_auth_headers(),
                json={
                    "name": project["name"],
                    "payload": final_payload,
                    "expected_revision": project["revision"],
                    "operation_id": str(uuid.uuid4()),
                },
            )
            assert patched.status_code == 200, patched.json()
            project = patched.json()["project"]

        with TestClient(web_api.app) as reloaded_client:
            reloaded = reloaded_client.get(
                f"/projects/{project['id']}", headers=_auth_headers()
            )
        assert reloaded.status_code == 200, reloaded.json()
        assert reloaded.json()["project"]["payload"] == project["payload"]

        quoted = client.post(
            f"/projects/{project['id']}/quote",
            headers=_auth_headers(),
            json={"expected_revision": project["revision"]},
        )
        assert quoted.status_code == 202, quoted.json()
        queued_job = quoted.json()["job"]
        frozen_payload = json.loads(
            web_api._storage_download_bytes(queued_job["input_path"])
        )
        assert frozen_payload["project_context"]["project_id"] == project["id"]
        assert frozen_payload["project_context"]["project_revision"] == project[
            "revision"
        ]

        original_claim = quote_worker.claim_job

        def tracked_claim(worker_client, job):
            claim_events.append(str(job["id"]))
            return original_claim(worker_client, job)

        monkeypatch.setattr(quote_worker, "claim_job", tracked_claim)
        assert quote_worker.run_once() is True
        completed = web_api.db_get_quote_job(queued_job["id"])
        assert completed["status"] == "completed", completed
        assert claim_events == [queued_job["id"]]
        output_bytes = web_api._storage_download_bytes(completed["output_path"])
        output = case_dir / "worker-output.xlsx"
        output.write_bytes(output_bytes)
        assert ZipFile(output).testzip() is None
        XlsxPackage.read(output)
        layout = plan_mobiliti_layout(
            [
                SectionNeed(
                    section["id"],
                    section["title"],
                    len(section["line_ids"]),
                )
                for section in frozen_payload["sections"]
            ]
        )
        return PersistedQuoteResult(
            output=output,
            frozen_payload=frozen_payload,
            project=project,
            job=completed,
            layout=layout,
            original_quotation=imported_source,
            claim_events=tuple(claim_events),
        )
    finally:
        client.close()
        monkeypatch.undo()


@pytest.fixture(scope="module")
def persisted_project_outputs(
    tmp_path_factory: pytest.TempPathFactory,
) -> dict[str, PersistedQuoteResult]:
    root = tmp_path_factory.mktemp("task8-persisted-project")
    source = write_import_fixture(root / "original-quotation.xlsx", currency="USD")
    source_workbook = load_workbook(source)
    source_workbook["Quotation"]._images = []
    source_workbook.save(source)
    source_workbook.close()
    mxn_payload = _project_payload()
    mxn_payload["quote_fields"]["quote_currency"] = "MXN"
    usd_payload = _project_payload()
    usd_payload["quote_fields"]["quote_currency"] = "USD"
    imported_payload = _stress_project_payload()
    imported_payload["sections"] = imported_payload["sections"][:1]
    imported_payload["lines"] = imported_payload["lines"][:8]
    stress_payload = _stress_project_payload()
    return {
        "MXN": _run_persisted_project_case(root / "mxn", mxn_payload),
        "USD": _run_persisted_project_case(root / "usd", usd_payload),
        "imported": _run_persisted_project_case(
            root / "imported",
            imported_payload,
            imported_source=source,
        ),
        "stress": _run_persisted_project_case(
            root / "stress",
            stress_payload,
            imported_source=source,
        ),
    }


@pytest.mark.parametrize("quote_currency", ("MXN", "USD"))
def test_project_quote_opens_without_repair_and_totals_equal_components(
    persisted_project_outputs: dict[str, PersistedQuoteResult],
    quote_currency: str,
) -> None:
    result = persisted_project_outputs[quote_currency]
    output = result.output
    payload = result.frozen_payload
    expected_unit_cost = Decimal("1850") if quote_currency == "MXN" else Decimal("100")
    physical_quantities = (Decimal("10"), Decimal("20"), Decimal("3"))
    component_totals = tuple(
        expected_unit_cost * quantity for quantity in physical_quantities
    )
    visible_principal_total = (
        expected_unit_cost
        + expected_unit_cost * Decimal("2")
        + expected_unit_cost * Decimal("3") / Decimal("10")
    ) * Decimal("10")

    assert visible_principal_total == sum(component_totals)
    assert [Decimal(item["unit_price"]) for item in payload["groups"][0]["items"]] == [
        expected_unit_cost,
        expected_unit_cost,
        expected_unit_cost,
    ]
    assert result.claim_events == (result.job["id"],)
    assert result.job["metadata"]["project_id"] == result.project["id"]

    workbook = load_workbook(output, data_only=False, read_only=False)
    try:
        mobiliti = workbook["Mobiliti"]
        cotizacion = workbook["Cotizacion"]
        quotation = workbook["Quotation"]
        quotation_rows = []
        for row in (14, 15, 16):
            formula = str(mobiliti.cell(row, 10).value)
            assert formula.startswith("=Quotation!K")
            quotation_rows.append(int(formula.removeprefix("=Quotation!K")))
        assert [
            Decimal(str(quotation.cell(row, 11).value))
            for row in quotation_rows
        ] == [expected_unit_cost, expected_unit_cost, expected_unit_cost]
        assert [mobiliti.cell(row, 8).value for row in (14, 15, 16)] == [
            f"=Quotation!H{quotation_row}"
            for quotation_row in quotation_rows
        ]
        assert [
            Decimal(str(quotation.cell(row, 8).value))
            for row in quotation_rows
        ] == list(physical_quantities)
        assert cotizacion["F17"].value == (
            "=Mobiliti!AA14"
            "+Mobiliti!AA15*Mobiliti!H15/Mobiliti!H14"
            "+Mobiliti!AA16*Mobiliti!H16/Mobiliti!H14"
        )
        assert cotizacion["A17"].value == "=Mobiliti!D14"
        assert cotizacion["C17"].value == f"=Quotation!D{quotation_rows[0]}"
        assert cotizacion["D17"].value == f"=Quotation!F{quotation_rows[0]}"
        assert cotizacion["E17"].value == "=Mobiliti!H14"
        assert quotation.cell(quotation_rows[0], 4).value.count("\n+ ") == 2
    finally:
        workbook.close()


def _exact_generated_additions(
    template_package: XlsxPackage,
    output_package: XlsxPackage,
) -> frozenset[str]:
    template_parts = set(template_package.parts)
    actual = set(output_package.parts) - template_parts
    expected = (
        set(
            output_package.relationship_closure(
                output_package.sheet_part("Quotation")
            )
        )
        | set(
            output_package.relationship_closure(
                output_package.sheet_part("Quotation_Data")
            )
        )
    ) - template_parts
    assert actual == expected
    return frozenset(expected)


def test_project_quote_preserves_original_quotation_and_template_contract(
    persisted_project_outputs: dict[str, PersistedQuoteResult],
) -> None:
    result = persisted_project_outputs["imported"]
    output = result.output
    imported_source = result.original_quotation
    assert imported_source is not None
    source_sha = hashlib.sha256(imported_source.read_bytes()).hexdigest()
    template_sha = hashlib.sha256(OFFICIAL_TEMPLATE.read_bytes()).hexdigest()

    source_signature = quotation_semantic_signature(imported_source)
    output_signature = quotation_semantic_signature(output)
    source_workbook = load_workbook(imported_source, data_only=False)
    try:
        source_sheet = source_workbook["Quotation"]
        source_content_last_row = max(
            row
            for row in range(1, source_sheet.max_row + 1)
            if any(
                source_sheet.cell(row, column).value is not None
                for column in range(1, 13)
            )
        )
        source_content_first_row = min(
            row
            for row in range(1, source_content_last_row + 1)
            if (
                str(source_sheet.cell(row, 1).value or "").startswith("-")
                or (
                    isinstance(source_sheet.cell(row, 1).value, (int, float))
                    and not isinstance(source_sheet.cell(row, 1).value, bool)
                )
            )
        )
        source_product_rows = [
            row
            for row in range(source_content_first_row, source_content_last_row + 1)
            if (
                isinstance(source_sheet.cell(row, 1).value, (int, float))
                and not isinstance(source_sheet.cell(row, 1).value, bool)
            )
        ]
    finally:
        source_workbook.close()
    row_keys = {f"source-{row}": row for row in source_product_rows}
    expected_snapshot = engine._insert_quotation_transformation_column(
        engine._normalized_quotation_source(imported_source),
        row_keys,
        {key: "Descripcion procesada esperada" for key in row_keys},
    )
    expected_signature = quotation_semantic_signature(BytesIO(expected_snapshot))
    assert output_signature[0] == expected_signature[0]
    expected_cells = {
        coordinate: signature
        for coordinate, *signature in expected_signature[1]
        if (
            int("".join(character for character in coordinate if character.isdigit()))
            < source_content_first_row
            or column_index_from_string(
                "".join(character for character in coordinate if character.isalpha())
            )
            > 13
        )
    }
    output_cells = {
        coordinate: signature
        for coordinate, *signature in output_signature[1]
    }
    assert {
        coordinate: output_cells[coordinate]
        for coordinate in expected_cells
    } == expected_cells
    assert output_signature[3:5] == expected_signature[3:5]
    assert output_signature[6] == expected_signature[6]
    output_workbook = load_workbook(output, data_only=False)
    try:
        output_sheet = output_workbook["Quotation"]
        output_product_count = sum(
            isinstance(output_sheet.cell(row, 1).value, (int, float))
            and not isinstance(output_sheet.cell(row, 1).value, bool)
            for row in range(1, output_sheet.max_row + 1)
        )
        assert output_product_count == result.frozen_payload["item_count"]
    finally:
        output_workbook.close()
    source_package = XlsxPackage.from_bytes(
        engine._normalized_quotation_source(imported_source)
    )
    output_package = XlsxPackage.read(output)
    source_media = {
        hashlib.sha256(payload).hexdigest()
        for name, payload in source_package.parts.items()
        if name.startswith("xl/media/")
    }
    output_media = {
        hashlib.sha256(payload).hexdigest()
        for name, payload in output_package.parts.items()
        if name.startswith("xl/media/")
    }
    assert source_media.issubset(output_media)
    assert hashlib.sha256(imported_source.read_bytes()).hexdigest() == source_sha
    assert hashlib.sha256(OFFICIAL_TEMPLATE.read_bytes()).hexdigest() == template_sha
    assert_official_template_contract()
    template_package = XlsxPackage.read(OFFICIAL_TEMPLATE)
    exact_additions = _exact_generated_additions(template_package, output_package)
    fletes_part = template_package.sheet_part("Fletes")
    template_fletes = _cell_map(template_package, "Fletes")
    output_fletes = _cell_map(output_package, "Fletes")
    template_formulas = {
        coordinate: _formula(cell)
        for coordinate, cell in template_fletes.items()
        if _formula(cell)
    }
    output_formulas = {
        coordinate: _formula(cell)
        for coordinate, cell in output_fletes.items()
        if _formula(cell)
    }
    guarded_fletes = {"B66", "E61"}
    assert {
        coordinate: output_formulas[coordinate]
        for coordinate in template_formulas
        if coordinate not in guarded_fletes
    } == {
        coordinate: formula
        for coordinate, formula in template_formulas.items()
        if coordinate not in guarded_fletes
    }
    assert output_formulas["E61"] == "IF(B67=0,0,MIN(1,B67/(B62*B71+B63*B74)))"
    assert output_formulas["B66"] == (
        'IF(E60="MANUAL",E63,IF(B61=0,0,IF(E60="PRORRATEADO",'
        "(B61*B65+E62+B78)/B61,(B61*B65+B64+B78)/B61)))"
    )
    expected_category_formulas = {
        "N18": "IF(Mobiliti!$P$4=TRUE,(56/Mobiliti!$P$6),56)",
        "N19": "IF(Mobiliti!$P$4=TRUE,(1790/Mobiliti!$P$6),1790)",
        "N20": "IF(Mobiliti!$P$4=TRUE,(210/Mobiliti!$P$6),210)",
        "N21": "IF(Mobiliti!$P$4=TRUE,(980/Mobiliti!$P$6),980)",
    }
    assert {
        coordinate: output_formulas[coordinate]
        for coordinate in expected_category_formulas
    } == expected_category_formulas
    assert set(output_formulas) - set(template_formulas) == set(
        expected_category_formulas
    )
    assert_package_preserved(
        OFFICIAL_TEMPLATE,
        output,
        allowed_parts=(
            set(OFFICIAL_ALLOWED_PARTS)
            | set(exact_additions)
            | {fletes_part}
        ),
    )
    XlsxPackage.read(output)
    assert ZipFile(output).testzip() is None


def test_project_quote_expands_past_16_sections_and_33_components(
    persisted_project_outputs: dict[str, PersistedQuoteResult],
) -> None:
    result = persisted_project_outputs["stress"]
    output = result.output
    request = result.frozen_payload
    package = XlsxPackage.read(output)

    assert ZipFile(output).testzip() is None
    assert request["item_count"] == 700
    assert len(request["sections"]) == 20
    assert [section["id"] for section in request["sections"]] == [
        "section-1",
        *(f"section-{section}" for section in range(5, 24)),
    ]
    assert [
        section["section_id"]
        for section in request["project_context"]["normalized_project_payload"][
            "sections"
        ]
    ] == [section["id"] for section in request["sections"]]
    assert request["imported_source"]["items"][0]["source_row"] == 11
    identities = [
        line["identity"]["internal_id"]
        for line in request["project_context"]["normalized_project_payload"]["lines"]
        if "identity" in line
    ]
    assert identities.count("sunon:main-1") >= 2

    quotation_data = _cell_map(package, "Quotation_Data")
    assert sum(
        1
        for coordinate, cell in quotation_data.items()
        if coordinate.startswith("A")
        and coordinate != "A1"
        and _cell_text(cell)
    ) == 700
    cotizacion = _cell_map(package, "Cotizacion")
    visible_formula_rows = sorted(
        int(coordinate[1:])
        for coordinate, cell in cotizacion.items()
        if coordinate.startswith("F")
        and "Mobiliti!AA" in _formula(cell)
    )
    assert len(visible_formula_rows) == 698
    first_visible_row = visible_formula_rows[0]
    assert _formula(cotizacion[f"F{first_visible_row}"]) == (
        "Mobiliti!AA14"
        "+Mobiliti!AA15*Mobiliti!H15/Mobiliti!H14"
        "+Mobiliti!AA16*Mobiliti!H16/Mobiliti!H14"
    )


def _excel_acceptance_surface(
    result: PersistedQuoteResult,
) -> tuple[dict[tuple[str, str], str], tuple[tuple[str, str], ...]]:
    package = XlsxPackage.read(result.output)
    official = XlsxPackage.read(OFFICIAL_TEMPLATE)
    mobiliti = _cell_map(package, "Mobiliti")
    official_mobiliti = _cell_map(official, "Mobiliti")
    selected_rows: set[int] = set()
    for section in result.layout.sections:
        if section.item_count:
            selected_rows.add(section.product_start)
            selected_rows.add(section.product_start + section.item_count - 1)
        if section.item_count < section.capacity:
            selected_rows.add(section.product_start + section.item_count)

    expectations: dict[tuple[str, str], str] = {}
    inspected: set[tuple[str, str]] = set()
    for row in sorted(selected_rows):
        for column in ("W", "X", "AA"):
            coordinate = f"{column}{row}"
            if column == "AA":
                last_row = result.layout.last_product_row
                expected = (
                    f"=IF(Z{row}>=Y{row},"
                    f"_xlfn.MINIFS($Z$14:$Z${last_row},"
                    f"$D$14:$D${last_row},D{row},"
                    f"$H$14:$H${last_row},"
                    f"_xlfn.MAXIFS($H$14:$H${last_row},"
                    f"$D$14:$D${last_row},D{row})),"
                    '"NO SE ESTA RESPETANDO EL MARGEN")'
                )
            else:
                expected = translate_formula(
                    f"={_formula(official_mobiliti[f'{column}14'])}",
                    origin=f"{column}14",
                    target=coordinate,
                    sheet="Mobiliti",
                )
            actual = f"={_formula(mobiliti[coordinate])}"
            assert actual == expected
            expectations[("Mobiliti", coordinate)] = actual
        y_coordinate = f"Y{row}"
        y_formula = _formula(mobiliti[y_coordinate])
        assert y_formula
        expectations[("Mobiliti", y_coordinate)] = f"={y_formula}"

    cotizacion = _cell_map(package, "Cotizacion")
    cotizacion_formula_coordinates = sorted(
        (
            coordinate
            for coordinate, cell in cotizacion.items()
            if coordinate.startswith(("F", "J")) and _formula(cell)
        ),
        key=lambda coordinate: (coordinate[0], int(coordinate[1:])),
    )
    for coordinate in (
        cotizacion_formula_coordinates[0],
        cotizacion_formula_coordinates[len(cotizacion_formula_coordinates) // 2],
        cotizacion_formula_coordinates[-1],
    ):
        expectations[("Cotizacion", coordinate)] = (
            f"={_formula(cotizacion[coordinate])}"
        )
    return expectations, tuple(sorted(inspected))


def _excel_desktop_static_unavailable_reason() -> str | None:
    if sys.platform != "win32":
        return "La aceptación de Excel requiere Windows"
    try:
        import pythoncom  # noqa: F401
        import win32com.client  # noqa: F401
        import winreg
    except ImportError as exc:
        return f"pywin32 no está disponible: {exc}"
    try:
        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CLSID"):
            pass
    except OSError as exc:
        return f"Excel de escritorio no está registrado: {exc}"
    return None


def _excel_com_roundtrip_isolated(
    source: Path,
    destination: Path,
    *,
    formula_expectations: dict[tuple[str, str], str],
    inspected_cells: tuple[tuple[str, str], ...],
    specification: Path,
) -> Path:
    specification.write_text(
        json.dumps(
            {
                "source": str(source),
                "destination": str(destination),
                "formula_expectations": [
                    {
                        "sheet": sheet,
                        "coordinate": coordinate,
                        "formula": formula,
                    }
                    for (sheet, coordinate), formula in formula_expectations.items()
                ],
                "inspected_cells": [
                    {"sheet": sheet, "coordinate": coordinate}
                    for sheet, coordinate in inspected_cells
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    child_script = """
import json
from pathlib import Path
import sys

sys.path.insert(0, str(Path.cwd() / "tests"))
from test_official_quote_stress import excel_com_roundtrip

payload = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
expectations = {
    (entry["sheet"], entry["coordinate"]): entry["formula"]
    for entry in payload["formula_expectations"]
}
inspected = tuple(
    (entry["sheet"], entry["coordinate"])
    for entry in payload["inspected_cells"]
)
excel_com_roundtrip(
    Path(payload["source"]),
    Path(payload["destination"]),
    formula_expectations=expectations,
    inspected_cells=inspected,
)
"""
    completed = subprocess.run(
        [sys.executable, "-c", child_script, str(specification)],
        cwd=ROOT,
        capture_output=True,
        text=True,
        timeout=900,
        check=False,
    )
    assert completed.returncode == 0, (
        "Excel COM aislado no completó el roundtrip.\n"
        f"STDOUT:\n{completed.stdout}\nSTDERR:\n{completed.stderr}"
    )
    assert destination.exists()
    return destination


def test_project_quote_excel_desktop_acceptance_for_four_persisted_cases(
    persisted_project_outputs: dict[str, PersistedQuoteResult],
    tmp_path: Path,
) -> None:
    unavailable = _excel_desktop_static_unavailable_reason()
    if unavailable is not None:
        pytest.skip(unavailable)

    for case_name in ("MXN", "USD", "imported", "stress"):
        result = persisted_project_outputs[case_name]
        source_hash = hashlib.sha256(result.output.read_bytes()).hexdigest()
        expectations, inspected_cells = _excel_acceptance_surface(result)
        validated = _excel_com_roundtrip_isolated(
            result.output,
            tmp_path / f"{case_name}-excel-roundtrip.xlsx",
            formula_expectations=expectations,
            inspected_cells=inspected_cells,
            specification=tmp_path / f"{case_name}-excel-roundtrip.json",
        )
        assert hashlib.sha256(result.output.read_bytes()).hexdigest() == source_hash
        assert validated != result.output
        assert ZipFile(validated).testzip() is None
        XlsxPackage.read(validated)


class _ProjectedLineIds(list):
    def __init__(self, reported_count: int):
        super().__init__()
        self.reported_count = reported_count

    def __len__(self) -> int:
        return self.reported_count


def _capacity_projection(component_count: int, *, padding: str = "") -> dict:
    project = _project_payload()
    return {
        "source_type": "mixed_catalog_cart",
        "quote_currency": "MXN",
        "created_at": "2026-07-23T00:00:00+00:00",
        "groups": [],
        "imported_source": None,
        "sections": [
            {
                "id": "section-1",
                "title": "Boundary",
                "line_ids": _ProjectedLineIds(component_count),
            }
        ],
        "item_count": component_count,
        "auto_electrification_rate": None,
        "rate_summary": [],
        "project_context": {
            "project_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
            "project_revision": 3,
            "project_payload_hash": "a" * 64,
            "normalized_project_payload": project,
            "compositions": [],
        },
        **({"padding": padding} if padding else {}),
    }


def _capacity_endpoint(
    monkeypatch: pytest.MonkeyPatch,
    projection: dict,
) -> tuple[TestClient, list[str]]:
    project = {
        "id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
        "usuario_id": 1,
        "name": "Boundary",
        "status": "active",
        "revision": 3,
        "schema_version": 1,
        "payload": _project_payload(),
    }
    events: list[str] = []
    monkeypatch.setattr(web_api, "JWT_SECRET_KEY", "task-8-capacity-secret")
    monkeypatch.setattr(
        web_api,
        "db_get_usuario_by_id",
        lambda user_id: {"id": int(user_id), "activo": True, "es_admin": False},
    )
    monkeypatch.setattr(web_api, "_require_active_subscription", lambda _user_id: None)
    monkeypatch.setattr(web_api, "db_get_project", lambda *_args: deepcopy(project))
    monkeypatch.setattr(
        web_api,
        "_build_saved_project_quote_payload",
        lambda *_args: (projection, None, None),
    )
    monkeypatch.setattr(web_api, "_next_quote_number_for_user", lambda _user: None)
    monkeypatch.setattr(
        web_api, "_enforce_active_quote_limit", lambda *_args, **_kwargs: None
    )
    monkeypatch.setattr(web_api, "_wake_worker", lambda: events.append("wake"))
    monkeypatch.setattr(
        web_api,
        "db_create_quote_job",
        lambda *_args, **_kwargs: events.append("create"),
    )
    monkeypatch.setattr(
        web_api,
        "db_reserve_mixed_cart",
        lambda *_args, **_kwargs: events.append("reserve") or [],
    )
    monkeypatch.setattr(
        web_api,
        "_storage_upload_bytes",
        lambda *_args, **_kwargs: events.append("upload"),
    )
    monkeypatch.setattr(
        web_api,
        "db_queue_mixed_quote_job",
        lambda job_id, metadata: (
            events.append("queue")
            or {
                "id": job_id,
                "usuario_id": 1,
                "status": "queued",
                "metadata": metadata,
            }
        ),
    )

    def enqueue(**kwargs):
        events.append("enqueue")
        job_id = "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb"
        web_api.db_create_quote_job(
            1,
            kwargs["template"],
            kwargs["metadata"],
            "users/1/jobs/boundary/input.json",
            job_id=job_id,
        )
        web_api.db_reserve_mixed_cart(1, job_id, [])
        web_api._storage_upload_bytes(
            "users/1/jobs/boundary/input.json",
            b"{}",
            "application/json",
        )
        return web_api.db_queue_mixed_quote_job(job_id, kwargs["metadata"])

    monkeypatch.setattr(web_api, "_enqueue_mixed_payload", enqueue)
    return TestClient(web_api.app), events


def test_project_quote_rejects_only_after_physical_xlsx_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixed_rows = required_mobiliti_rows([34]) - 34
    maximum_components = (
        XLSX_MAX_ROWS
        - MOBILITI_RESERVED_ROWS_AFTER_TOTAL
        - fixed_rows
    )
    assert required_mobiliti_rows([maximum_components]) + (
        MOBILITI_RESERVED_ROWS_AFTER_TOTAL
    ) == XLSX_MAX_ROWS

    projection = _capacity_projection(maximum_components)
    client, events = _capacity_endpoint(monkeypatch, projection)
    accepted = client.post(
        "/projects/aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa/quote",
        headers=_auth_headers(),
        json={"expected_revision": 3},
    )
    assert accepted.status_code == 202, accepted.json()
    assert events == ["enqueue", "create", "reserve", "upload", "queue", "wake"]

    events.clear()
    projection["sections"][0]["line_ids"] = _ProjectedLineIds(
        maximum_components + 1
    )
    rejected = client.post(
        "/projects/aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa/quote",
        headers=_auth_headers(),
        json={"expected_revision": 3},
    )
    assert rejected.status_code == 400
    assert "XLSX permite hasta" in rejected.json()["detail"]
    assert events == []
    client.close()


def test_project_quote_rejects_byte_limit_before_enqueue(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    projection = _capacity_projection(
        1,
        padding="x" * (MAX_QUOTE_REQUEST_BYTES + 1),
    )
    client, events = _capacity_endpoint(monkeypatch, projection)
    rejected = client.post(
        "/projects/aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa/quote",
        headers=_auth_headers(),
        json={"expected_revision": 3},
    )
    assert rejected.status_code == 400
    assert "bytes" in rejected.json()["detail"]
    assert events == []
    client.close()
