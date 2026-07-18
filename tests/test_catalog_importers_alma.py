import hashlib
import importlib
import importlib.util
import io
import json
import zipfile
from dataclasses import dataclass, replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image

from mobiliti_saas.quote_engine.supplier_catalog import (
    PUBLIC_ITEM_FIELDS,
    build_supplier_cart_payload,
    load_supplier_catalog_data,
)
import mobiliti_saas.worker.catalog_sync.importers.common as common_module
from mobiliti_saas.worker.catalog_sync.importers.common import SourceSafetyError


MODULE = "mobiliti_saas.worker.catalog_sync.importers.alma"
MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
KUN_PATH = "SPEC Guide-Alma-KUN.xlsx"
KUN_PRICE_PATH = "SPEC GUIDES 2026/ALMA/Spec guide-Alma-KUN Design.xlsx"
MONDECASA_PATH = "SPEC Guide-Alma-Mondecasa.xlsx"
REAL_HASHES = {
    KUN_PATH: "1c17db827c5ed308afe76c145b5e70ca04e5c7348de0bf18e596e56e42e5a613",
    KUN_PRICE_PATH: "e01f1bedb1871909fa23f2990a3c5b1cf934ae6004002f7028aaafb37db3bb68",
    MONDECASA_PATH: "815fa963b7a5e1832871c9118bc6225574150b8c813d9646e7f1e9f40e301c13",
}


@dataclass(frozen=True)
class AdapterFile:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path | None


def _alma():
    return importlib.import_module(MODULE)


def _sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _image(color):
    buffer = io.BytesIO()
    Image.new("RGB", (12, 8), color).save(buffer, "PNG")
    buffer.seek(0)
    return WorksheetImage(buffer)


def _write_kun(path):
    workbook = Workbook()
    design = workbook.active
    design.title = "KUN DESIGN"
    headers = (
        "C\u00f3digo",
        "Imagen",
        "Descripci\u00f3n",
        "Dimensiones",
        "Sin coj\u00edn\nAluminio: Recubrimiento en polvo",
        "Sin coj\u00edn\nAluminio: Aspecto de teca",
        "Solo Coj\u00edn\nCalidad: Tela A\nEspuma Normal\n-----------\nCeramica A",
        "Solo Coj\u00edn\nCalidad: Tela A+\nEspuma Normal\n-----------\nCeramica A+",
        "Solo Coj\u00edn\nCalidad: Tela A++\nEspuma Normal\n-----------\nCeramica A++",
    )
    for column, value in enumerate(headers, 1):
        design.cell(6, column, value)
    design.merge_cells("A7:I7")
    design["A7"] = "TEST COLLECTION"

    for column in "ABCDEF":
        design.merge_cells(f"{column}8:{column}9")
    for cell, value in {
        "A8": "KUN-1",
        "C8": "Configurable chair",
        "D8": "60*60cm",
        "E8": 100,
        "F8": 120,
        "G8": 10,
        "H8": 20,
        "I8": 30,
        "G9": 11,
        "H9": 21,
        "I9": 31,
        "A10": "KUN-SINGLE",
        "C10": "Single base table",
        "D10": "80*80cm",
        "E10": 50,
        "F10": "\u2014",
        "A11": "KUN-LOW",
        "C11": "Low alternate base",
        "D11": "90*90cm",
        "E11": 1000,
        "F11": 100,
        "A12": "KUN-BAD",
        "C12": "Invalid prices",
        "D12": "70*70cm",
        "E12": "=1+1",
        "F12": "\u2014",
        "G12": 1_000_000_001,
        "H12": "not money",
        "I12": -1,
        "G13": 999,
        "A14": "SHARED-1",
        "C14": "KUN shared code",
        "D14": "100*100cm",
        "E14": 75,
    }.items():
        design[cell] = value
    design.add_image(_image("red"), "B8")

    pavilion = workbook.create_sheet("PAVILION ")
    for cell, value in {
        "A1": "Picture",
        "B1": "Item no.",
        "C1": "Description",
        "D1": "Dimensions",
        "E1": "FOB Price",
        "A2": "PAVILION",
        "B3": "KUN-PAV",
        "C3": "Configured pavilion",
        "D3": "300*300cm",
        "E3": 250,
    }.items():
        pavilion[cell] = value
    pavilion.merge_cells("A2:O2")
    pavilion.add_image(_image("green"), "A3")
    workbook.save(path)
    workbook.close()


def _write_kun_prices(path):
    workbook = Workbook()
    sales = workbook.active
    sales.title = "SPEC Alma"
    costs = workbook.create_sheet("Costo Alma")
    for sheet, price_header in ((sales, "Precio Venta"), (costs, "P. Unitario.")):
        headers = ("Imagen.", "Cod.", "Descripcion.", *(price_header for _ in range(5)))
        for column, value in enumerate(headers, 1):
            sheet.cell(8, column, value)
    sales["I8"] = "Moneda."
    costs["I8"] = "Tipo de Cambio"
    costs["J8"] = "LAB Cedis"
    rows = (
        (9, "KUN-1", "Configurable chair", (100, 120, 10, 20, 30)),
        (10, "KUN-1", "Configurable chair", (0, 0, 11, 21, 31)),
        (11, "KUN-SINGLE", "Single base table", (50, 0, 0, 0, 0)),
        (12, "KUN-LOW", "Low alternate base", (1000, 100, 0, 0, 0)),
        (14, "SHARED-1", "KUN shared code", (75, 0, 0, 0, 0)),
    )
    sales.merge_cells("B9:B10")
    sales.merge_cells("C9:C10")
    costs.merge_cells("B9:B10")
    costs.merge_cells("C9:C10")
    for row, code, description, prices in rows:
        if row != 10:
            sales.cell(row, 2, code)
            sales.cell(row, 3, description)
            costs.cell(row, 2, code)
            costs.cell(row, 3, description)
        for offset, cost in enumerate(prices, 4):
            costs.cell(row, offset, cost)
            sales.cell(row, offset, cost / 0.3 / 0.5 if cost else 0)
        sales.cell(row, 9, "USD")
        costs.cell(row, 9, 0.3)
        costs.cell(row, 10, 0.5)
    workbook.save(path)
    workbook.close()


def _mondecasa_header(sheet, row, collection):
    values = (
        "Cod.",
        collection,
        "Descripcion.",
        "Dimensiones",
        "Costo SIN Coj\u00edn",
        "Costo del coj\u00edn solamente\n(Tela china + espuma normal)",
        "Precio del coj\u00edn solamente\n(Tela con protecci\u00f3n solar + espuma normal)",
        "Tejidos Sunbrella 5461 /Vita + espuma normal",
    )
    for column, value in enumerate(values, 1):
        sheet.cell(row, column, value)


def _write_mondecasa(path):
    workbook = Workbook()
    products = workbook.active
    products.title = "MONDECASA"
    _mondecasa_header(products, 8, "FIRST")
    for cell, value in {
        "A9": "MON-1",
        "C9": "Mondecasa lounge chair",
        "D9": "90*90cm",
        "E9": 100,
        "F9": 10,
        "G9": 20,
        "H9": 30,
        "A10": "MON-TABLE",
        "C10": "Mondecasa table",
        "D10": "120*80cm",
        "E10": 50,
        "A11": "SHARED-1",
        "C11": "Different shared product",
        "D11": "200*100cm",
        "E11": 80,
        "F11": 8,
        "C12": "Missing unmerged code",
        "D12": "40*40cm",
        "E12": 60,
    }.items():
        products[cell] = value
    products.add_image(_image("blue"), "B9")
    _mondecasa_header(products, 13, "SECOND")
    for cell, value in {
        "A14": "MON-2",
        "C14": "Second collection table",
        "D14": "50*50cm",
        "E14": 70,
        "A15": "MON-MERGED",
        "C15": "Configurable Mondecasa table",
        "D15": "80*80cm",
        "E15": 90,
        "G15": 9,
        "D16": "120*80cm",
        "E16": 110,
        "G16": 11,
    }.items():
        products[cell] = value
    for column in "ABC":
        products.merge_cells(f"{column}15:{column}16")

    pavilions = workbook.create_sheet("PAVILIONS")
    for cell, value in {
        "A5": "Item no.",
        "B5": "Photo",
        "C5": "Description",
        "D5": "Dimensions",
        "L5": "Qty/ 40'HQ",
        "M5": "Qty / 20FT",
        "E6": "Main frame only",
        "F6": "Polar wood floor only",
        "G6": "Alu. panel per side",
        "H6": "Alu. panel for the roof",
        "I6": "PC water proof roof",
        "J6": "Sliding curtain for the roof",
        "K6": "Curtain per side (Chinese fabric)",
        "A7": "PAV-A",
        "C7": "Large pavilion",
        "D7": "400*400cm",
        "E7": 100,
        "F7": 10,
        "G7": 20,
        "H7": 30,
        "I7": 40,
        "J7": 50,
        "K7": 60,
        "L7": 999,
        "M7": 888,
        "A10": "Item no.",
        "B10": "Photo",
        "C10": "Description",
        "D10": "Dimensions",
        "E11": "Main frame only",
        "F11": "Alu. panel per side",
        "G11": "Sliding curtain for the roof",
        "H11": "PC roof",
        "I11": "Sling roof (based on sling 0020)",
        "J11": "Adjustable shade (sling)",
        "K11": "(indonesian fabrics-L0767ABC/Dwhite&L0767ABC/BLight grey)",
        "L11": "Cushion price only punpoly fabric + normal foam)",
        "M11": "Cushion price only Sunbrella natural 5404 + quick dry foam)",
        "A12": "PAV-B",
        "C12": "Small pavilion",
        "D12": "200*200cm",
        "E12": 200,
        "F12": 11,
        "G12": 21,
        "H12": 31,
        "I12": 41,
        "J12": 51,
        "K12": 61,
        "L12": 71,
        "M12": 81,
        "A17": "Item no.",
        "B17": "Photo",
        "C17": "Description",
        "D17": "Dimensions",
        "E17": "FOB Price W/O cushion",
        "F17": "Roof Only",
        "G17": "Cushion price only Chinese fabric + normal foam",
        "H17": "Cushion price only Chinese fabric + quick dry foam",
        "I17": "Cushion price only Sunbrella 5404+normal foam",
        "J17": "Cushion price only Sunbrella 5404+quick dry foam",
        "A18": "PAV-C",
        "C18": "Day bed",
        "D18": "200*180cm",
        "E18": 300,
        "F18": 12,
        "G18": 22,
        "H18": 32,
        "I18": 42,
        "J18": 52,
        "N17": "Side blind",
        "N18": 62,
        "A19": "PAV-BAD",
        "C19": "Bad pavilion prices",
        "D19": "100*100cm",
        "E19": "\u2014",
        "F19": "=1+1",
        "G19": -1,
        "H19": "bad",
        "I19": 1_000_000_001,
        "A20": "PAV-MERGED",
        "C20": "Configurable pavilion",
        "D20": "300*300cm",
        "E20": 400,
        "F20": 40,
        "D21": "400*400cm",
        "E21": 500,
        "F21": 50,
    }.items():
        pavilions[cell] = value
    for column in "ABC":
        pavilions.merge_cells(f"{column}20:{column}21")
    pavilions.add_image(_image("yellow"), "B7")
    workbook.save(path)
    workbook.close()


@pytest.fixture
def source_bundle(tmp_path, monkeypatch):
    module = _alma()
    monkeypatch.setattr(module, "_KUN_DESIGN_COUNT", 4)
    monkeypatch.setattr(module, "_KUN_PAVILION_COUNT", 1)
    monkeypatch.setattr(module, "_KUN_MIN_IDENTITY_MATCHES", 4)
    sources = (
        (KUN_PATH, "KUN", tmp_path / "kun.xlsx", _write_kun),
        (KUN_PRICE_PATH, "KUN", tmp_path / "kun-prices.xlsx", _write_kun_prices),
        (MONDECASA_PATH, "Mondecasa", tmp_path / "mondecasa.xlsx", _write_mondecasa),
    )
    rows = []
    for logical_path, brand, local_path, writer in sources:
        writer(local_path)
        digest = _sha256(local_path)
        rows.append(AdapterFile(logical_path, "spec_guide", brand, digest, MIME, local_path))
    return tuple(rows)


def _snapshot(files):
    return _alma().build_alma_snapshot(files)


def _without_generated_at(snapshot):
    return {key: value for key, value in snapshot.items() if key != "generated_at"}


def _item(snapshot, sku):
    return next(
        item for item in snapshot["items"]
        if item["attributes"].get("source_code") == sku
    )


def _edit(bundle, logical_path, mutation):
    row = next(candidate for candidate in bundle if candidate.path == logical_path)
    workbook = load_workbook(row.local_path)
    mutation(workbook)
    workbook.save(row.local_path)
    workbook.close()
    digest = _sha256(row.local_path)
    updated = replace(row, sha256=digest)
    return tuple(updated if candidate is row else candidate for candidate in bundle)


def _rewrite_zip(source, target, transform):
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(
        target, "w", zipfile.ZIP_DEFLATED
    ) as output:
        for info in archive.infolist():
            name, data = transform(info.filename, archive.read(info))
            output.writestr(name, data)


def test_alma_adapter_module_exists():
    assert importlib.util.find_spec(MODULE) is not None


def test_snapshot_contract_hash_and_order_are_deterministic(source_bundle):
    first = _snapshot(source_bundle)
    second = _snapshot(tuple(reversed(source_bundle)))
    loaded = load_supplier_catalog_data(first, expected_supplier="alma")
    alma = _alma()
    expected_hash = alma._source_hash(
        source_bundle,
        alma.load_kundesign_link_index(),
    )

    assert set(first) == {"supplier", "source_hash", "generated_at", "items"}
    assert first["source_hash"] == expected_hash
    assert _without_generated_at(first) == _without_generated_at(second)
    assert len(loaded["items"]) == 16
    assert all(set(item) == set(PUBLIC_ITEM_FIELDS) for item in first["items"])
    assert all(
        item["supplier"] == "alma"
        and item["unit"] == "PZA"
        and item["availability_type"] == "made_to_order"
        and item["stock"] is None
        and item["lead_time"] == "Sobre pedido"
        and item["base_currency"] == "USD"
        and item["tax_rate"] == "0.160000"
        and item["image_url"] == ""
        and item["image_kind"] == "placeholder"
        for item in first["items"]
    )


def test_kun_merged_block_splits_bases_from_one_aggregated_cushion_family(source_bundle):
    item = _item(_snapshot(source_bundle), "KUN-1")

    assert [option["price_net"] for option in item["base_price_options"]] == [
        "666.666667",
        "800.000000",
    ]
    assert all(option["available"] for option in item["base_price_options"])
    assert [option["price_net"] for option in item["add_on_options"]] == [
        "140.000000",
        "273.333333",
        "406.666667",
    ]
    assert {option["family"] for option in item["add_on_options"]} == {"cushion"}
    assert all(option["available"] for option in item["add_on_options"])
    assert "999.000000" not in {
        evidence["price_net"] for evidence in item["attributes"]["price_evidence"]
    }
    image = item["attributes"]["source_images"][0]
    assert len(image["sha256"]) == 64
    assert image["width"] == 12 and image["height"] == 8
    assert image["source"]["cell_or_bbox"] == "B8"


def test_kun_single_base_and_materially_lower_alternative(source_bundle):
    snapshot = _snapshot(source_bundle)
    single = _item(snapshot, "KUN-SINGLE")
    low = _item(snapshot, "KUN-LOW")

    assert len(single["base_price_options"]) == 1
    assert single["base_price_options"][0]["price_net"] == "333.333333"
    assert [option["price_net"] for option in low["base_price_options"]] == [
        "6666.666667",
        "666.666667",
    ]


def test_kun_pavilion_complete_row_uses_direct_price(source_bundle):
    item = _item(_snapshot(source_bundle), "KUN-PAV")

    assert item["brand"] == "KUN"
    assert item["collection"] == "PAVILION"
    assert item["price_net"] == "250.000000"
    assert item["base_price_options"] == []
    assert item["add_on_options"] == []
    assert item["attributes"]["source_images"][0]["source"]["cell_or_bbox"] == "A3"


def test_mondecasa_base_and_one_cushion_family(source_bundle):
    snapshot = _snapshot(source_bundle)
    chair = _item(snapshot, "MON-1")
    table = _item(snapshot, "MON-TABLE")
    second = _item(snapshot, "MON-2")

    assert chair["price_net"] == "100.000000"
    assert chair["base_price_options"] == []
    assert [option["price_net"] for option in chair["add_on_options"]] == [
        "10.000000",
        "20.000000",
        "30.000000",
    ]
    assert {option["family"] for option in chair["add_on_options"]} == {"cushion"}
    assert all(option["available"] for option in chair["add_on_options"])
    assert chair["product_url"] == "https://www.mondecasa.com/products"
    assert chair["attributes"]["product_url_match"]["status"] == "catalog_fallback"
    assert table["add_on_options"] == []
    assert second["collection"] == "SECOND"


def test_vertical_merges_create_one_configurable_item_instead_of_duplicates(source_bundle):
    snapshot = _snapshot(source_bundle)
    mondecasa = _item(snapshot, "MON-MERGED")
    pavilion = _item(snapshot, "PAV-MERGED")

    for item, expected_prices in (
        (mondecasa, ["90.000000", "110.000000"]),
        (pavilion, ["400.000000", "500.000000"]),
    ):
        assert item["code_status"] == "verified"
        assert item["price_net"] == "0.000000"
        assert [option["price_net"] for option in item["base_price_options"]] == expected_prices
        assert all(option["available"] for option in item["base_price_options"])
        assert all(
            option.get("compatible_base_option_ids")
            for option in item["add_on_options"]
        )


def test_alma_base_and_add_on_prices_flow_into_supplier_cart(source_bundle):
    snapshot = _snapshot(source_bundle)
    kun = _item(snapshot, "KUN-1")
    mondecasa = _item(snapshot, "MON-1")

    payload = build_supplier_cart_payload(
        [
            {
                "internal_id": kun["internal_id"],
                "quantity": "2",
                "base_option_id": kun["base_price_options"][1]["id"],
                "add_on_option_ids": [kun["add_on_options"][0]["id"]],
            },
            {
                "internal_id": mondecasa["internal_id"],
                "quantity": "1",
                "add_on_option_ids": [mondecasa["add_on_options"][0]["id"]],
            },
        ],
        snapshot,
        "USD",
        [],
    )

    assert payload["items"][0]["unit_price_base"] == "940.000000"
    assert payload["items"][0]["line_total"] == "1880.00"
    assert payload["items"][1]["unit_price_base"] == "110.000000"
    assert payload["items"][1]["line_total"] == "110.00"


def test_pavilions_classifies_only_explicit_headers_and_excludes_packaging(source_bundle):
    snapshot = _snapshot(source_bundle)
    first = _item(snapshot, "PAV-A")
    second = _item(snapshot, "PAV-B")
    third = _item(snapshot, "PAV-C")

    assert first["price_net"] == "100.000000"
    assert len(first["add_on_options"]) == 6
    assert {option["family"] for option in first["add_on_options"]} == {
        "floor",
        "panel-alu-panel-per-side",
        "roof",
        "side-curtain",
    }
    all_prices = {
        evidence["price_net"]
        for item in (first, second, third)
        for evidence in item["attributes"]["price_evidence"]
    }
    assert "999.000000" not in all_prices and "888.000000" not in all_prices

    assert len(second["add_on_options"]) == 8
    unknown = next(option for option in second["add_on_options"] if not option["available"])
    assert unknown["name"] == "Agregado por verificar"
    assert sum(option["family"] == "roof" for option in second["add_on_options"]) == 4
    assert sum(option["family"] == "cushion" for option in second["add_on_options"]) == 2
    assert len(third["add_on_options"]) == 6
    assert sum(option["family"] == "cushion" for option in third["add_on_options"]) == 4
    assert sum(option["family"] == "side-curtain" for option in third["add_on_options"]) == 1


def test_duplicate_codes_stay_separate_and_missing_code_stays_blocked(source_bundle):
    snapshot = _snapshot(source_bundle)
    shared = [
        item
        for item in snapshot["items"]
        if item["attributes"].get("source_code") == "SHARED-1"
    ]
    missing = next(
        item
        for item in snapshot["items"]
        if item["name"] == "Missing unmerged code"
    )

    assert len(shared) == 2
    assert len({item["internal_id"] for item in shared}) == 2
    assert {item["brand"] for item in shared} == {"KUN", "Mondecasa"}
    assert all(item["sku"] and item["code_status"] == "verified" for item in shared)
    assert len({item["sku"] for item in shared}) == 2
    assert missing["sku"] == "" and missing["price_net"] == "0.000000"
    assert any("c\u00f3digo" in warning.casefold() for warning in missing["warnings"])

    with pytest.raises(ValueError, match="verificar"):
        build_supplier_cart_payload(
            [{"internal_id": missing["internal_id"], "quantity": "1"}],
            snapshot,
            "USD",
            [],
        )


def test_formula_dash_blank_malformed_extreme_and_negative_prices_fail_closed(source_bundle):
    snapshot = _snapshot(source_bundle)
    pavilion = next(item for item in snapshot["items"] if item["attributes"]["source_code"] == "PAV-BAD")

    assert not any(item["attributes"]["source_code"] == "KUN-BAD" for item in snapshot["items"])
    assert pavilion["price_net"] == "0.000000"
    assert pavilion["add_on_options"] == []
    assert pavilion["code_status"] == "needs_review"
    assert pavilion["warnings"]
    pavilion_refs = {row["cell_or_bbox"] for row in json.loads(pavilion["source_reference"])}
    assert {"F19", "G19", "H19", "I19"}.issubset(pavilion_refs)

    with pytest.raises(ValueError, match="verificar"):
        build_supplier_cart_payload(
            [{"internal_id": pavilion["internal_id"], "quantity": "1"}],
            snapshot,
            "USD",
            [],
        )


def test_sub_quantum_direct_price_fails_closed(source_bundle):
    changed = _edit(
        source_bundle,
        KUN_PATH,
        lambda workbook: setattr(workbook["PAVILION "]["E3"], "value", 0.0000004),
    )
    with pytest.raises(ValueError, match="ALMA_KUN_COUNT"):
        _snapshot(changed)


@pytest.mark.parametrize(
    "mutation",
    (
        lambda rows: rows[:-1],
        lambda rows: rows + (rows[0],),
        lambda rows: tuple(
            replace(row, path="unexpected.xlsx") if row.path == KUN_PATH else row
            for row in rows
        ),
        lambda rows: tuple(
            replace(row, kind="catalog") if row.path == KUN_PATH else row for row in rows
        ),
        lambda rows: tuple(
            replace(row, brand="Wrong") if row.path == KUN_PATH else row for row in rows
        ),
        lambda rows: tuple(
            replace(row, mime_type="application/octet-stream")
            if row.path == KUN_PATH
            else row
            for row in rows
        ),
        lambda rows: tuple(
            replace(row, sha256="0" * 64) if row.path == KUN_PATH else row
            for row in rows
        ),
        lambda rows: tuple(
            replace(row, local_path=None) if row.path == KUN_PATH else row for row in rows
        ),
    ),
)
def test_exact_bundle_path_kind_brand_mime_hash_and_shape_fail_closed(source_bundle, mutation):
    with pytest.raises((ValueError, SourceSafetyError)):
        _snapshot(mutation(source_bundle))


def test_sheet_and_header_contract_fail_closed(source_bundle):
    wrong_sheet = _edit(
        source_bundle,
        KUN_PATH,
        lambda workbook: setattr(workbook["PAVILION "], "title", "PAVILION"),
    )
    with pytest.raises(ValueError, match="ALMA_SHEETS"):
        _snapshot(wrong_sheet)

    def restore_sheet_and_break_header(workbook):
        workbook["PAVILION"].title = "PAVILION "
        workbook["KUN DESIGN"]["E6"] = "Wrong"

    source_bundle = _edit(source_bundle, KUN_PATH, restore_sheet_and_break_header)
    with pytest.raises(ValueError, match="ALMA_HEADER"):
        _snapshot(source_bundle)


def test_unsafe_source_fails_closed(source_bundle):
    unsafe = _edit(
        source_bundle,
        KUN_PATH,
        lambda workbook: setattr(
            workbook["KUN DESIGN"]["C10"],
            "value",
            "='[outside.xlsx]Sheet1'!A1",
        ),
    )
    with pytest.raises(SourceSafetyError):
        _snapshot(unsafe)


def test_parses_the_exact_validated_bytes(source_bundle, monkeypatch):
    original = _alma().read_validated_source

    def validate_then_change(path, extension):
        validated, data = original(path, extension)
        path.write_bytes(b"changed after validation")
        return validated, data

    monkeypatch.setattr(_alma(), "read_validated_source", validate_then_change)
    assert len(_snapshot(source_bundle)["items"]) == 16


def test_passive_emf_never_reaches_decoder(source_bundle, tmp_path, monkeypatch):
    row = next(candidate for candidate in source_bundle if candidate.path == MONDECASA_PATH)
    mixed = tmp_path / "mondecasa-emf.xlsx"

    def make_second_image_emf(name, data):
        if name == "[Content_Types].xml":
            data = data.replace(
                b"</Types>",
                b'<Default Extension="emf" ContentType="image/x-emf"/></Types>',
            )
        elif name.startswith("xl/drawings/_rels/") and name.endswith(".rels"):
            data = data.replace(b"image2.png", b"image2.emf")
        elif name == "xl/media/image2.png":
            return "xl/media/image2.emf", b"inert-emf-must-not-be-decoded"
        return name, data

    _rewrite_zip(row.local_path, mixed, make_second_image_emf)
    digest = _sha256(mixed)
    changed = tuple(
        replace(candidate, local_path=mixed, sha256=digest) if candidate is row else candidate
        for candidate in source_bundle
    )
    decoded = []
    real_normalize = common_module._normalize_image

    def audited_normalize(data):
        decoded.append(data)
        return real_normalize(data)

    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)
    assert len(_snapshot(changed)["items"]) == 16
    assert all(raw != b"inert-emf-must-not-be-decoded" for raw in decoded)


def test_ignored_real_sources_report_coverage_metrics():
    root = Path(".cache/catalog_sources/alma/sharepoint_2026-07-17")
    rows = (
        AdapterFile(
            KUN_PATH,
            "spec_guide",
            "KUN",
            REAL_HASHES[KUN_PATH],
            MIME,
            root / "SPEC Guide-Alma-KUN.root.xlsx",
        ),
        AdapterFile(
            KUN_PRICE_PATH,
            "spec_guide",
            "KUN",
            REAL_HASHES[KUN_PRICE_PATH],
            MIME,
            root / "Spec guide-Alma-KUN Design.current.xlsx",
        ),
        AdapterFile(
            MONDECASA_PATH,
            "spec_guide",
            "Mondecasa",
            REAL_HASHES[MONDECASA_PATH],
            MIME,
            root / "SPEC Guide-Alma-Mondecasa.current.xlsx",
        ),
    )
    if any(not row.local_path.exists() for row in rows):
        pytest.skip("ignored ALMA source workbooks are not available")
    assert all(_sha256(row.local_path) == row.sha256 for row in rows)

    snapshot = _snapshot(rows)
    items = load_supplier_catalog_data(snapshot, expected_supplier="alma")["items"]
    kun = [item for item in items if item["brand"] == "KUN"]
    metrics = {
        "items": len(items),
        "kun": len(kun),
        "kun_verified": sum(item["code_status"] == "verified" for item in kun),
        "kun_needs_review": sum(item["code_status"] == "needs_review" for item in kun),
        "kun_pavilion": sum(item["collection"] == "PAVILION" for item in kun),
        "kun_derived": sum(
            item["attributes"]["price_reconciliation"]["method"]
            == "derived_from_identity_cost"
            for item in kun
        ),
    }
    source_rows = {}
    emf_parts = 0
    for row in rows:
        _, data = common_module.read_validated_source(row.local_path, ".xlsx")
        workbook = common_module.open_xlsx_data_only_from_bytes(data)
        source_rows.update({sheet.title: sheet.max_row for sheet in workbook.worksheets})
        workbook.close()
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            emf_parts += sum(name.casefold().endswith(".emf") for name in archive.namelist())
    report = {"rows": source_rows, "emf_parts": emf_parts, **metrics}
    print("ALMA_REAL_METRICS=" + json.dumps(report, sort_keys=True))

    assert source_rows == {
        "KUN DESIGN": 410,
        "PAVILION ": 5,
        "SPEC Alma": 360,
        "Costo Alma": 358,
        "MONDECASA": 414,
        "PAVILIONS": 192,
    }
    assert emf_parts == 71
    assert metrics == {
        "items": 654,
        "kun": 310,
        "kun_verified": 262,
        "kun_needs_review": 48,
        "kun_pavilion": 3,
        "kun_derived": 2,
    }
