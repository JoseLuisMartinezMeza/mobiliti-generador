"""E2E local de los catálogos IDÉLIKA y Conceptos.

Este contrato usa únicamente los artefactos locales validados, un repositorio
en memoria y un almacenamiento de assets falso. No inicializa red ni base de
datos remota.
"""

from __future__ import annotations

import hashlib
import json
import subprocess
from datetime import date
from decimal import Decimal
from pathlib import Path
from uuid import UUID

import pytest
from openpyxl import load_workbook

from mobiliti_saas.quote_engine import catalog_cart, generate_quote
from mobiliti_saas.quote_engine.catalog_search import search_catalog_products
from mobiliti_saas.quote_engine.mixed_catalog import build_mixed_catalog_cart_payload
from mobiliti_saas.quote_engine.project_model import normalize_project_payload
from mobiliti_saas.quote_engine.supplier_catalog import load_supplier_catalog_data
from mobiliti_saas.quote_engine.quotation_sheets import (
    QUOTATION_DATA_HEADERS,
    quotation_data_rows,
)
from mobiliti_saas.worker.catalog_sync.graph import DeltaResult, DownloadedFile, GraphItem
from mobiliti_saas.worker.catalog_sync.repository import SourceFileRecord, SourceRecord
from mobiliti_saas.worker.catalog_sync import load_source_config
from mobiliti_saas.worker.catalog_sync.importers.idelika import (
    build_idelika_spec_artifact,
)
from mobiliti_saas.worker.catalog_sync.service import ADAPTERS, run_supplier_sync
from mobiliti_saas.worker.quote_worker import _convert_mixed_catalog_cart_to_quotation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "019f7907-1ecc-7001-b3f3-8eb209086fa8"
IDELIKA_SOURCES = {
    "IDELIKA/1 CATALOGO FABRICACION 2026B.pdf": OUTPUT / "sources" / "1 CATALOGO FABRICACION 2026B.pdf",
    "IDELIKA/2 CATALOGO STOCK 2026.pdf": OUTPUT / "sources" / "2 CATALOGO STOCK 2026.pdf",
    "IDELIKA/4 SCHOOL SERIES 2026.pdf": OUTPUT / "sources" / "4 SCHOOL SERIES 2026.pdf",
}
CONCEPTOS_SPEC = (
    OUTPUT / "sources" / "Spec guide - Conceptos - Sofas - CdMx - Gdl - Qro - 2021.xlsx"
)
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SOURCES_CONFIG = ROOT / "mobiliti_saas" / "worker" / "catalog_sync" / "sources.json"
WORKER_TEMPLATE = ROOT / "mobiliti_saas" / "worker" / "templates" / "Formato Cotizacion 2026 Oficial.xlsx"
EXPECTED_FILE_HASHES = {
    "IDELIKA/1 CATALOGO FABRICACION 2026B.pdf": "b3bfe5761f850fd2203e0f8054d801b42d3b4e7c6c8d349ac8d30bc28fb64cfb",
    "IDELIKA/2 CATALOGO STOCK 2026.pdf": "a40a53e8599615a0bd2fc21fe5a1658a32a2cc83ad2ee5e89ce8bb1d9d372f76",
    "IDELIKA/4 SCHOOL SERIES 2026.pdf": "6fcafbb6c973e961510eead9fcf3aacdcf94ac59d48d6d2bcf7b38c7a710f2ec",
    "conceptos": "7d2380cb3cde7ec70ee25832c06c351c6f65dc757f793713a86524bcd7a06c70",
}
IDELIKA_GENERATED_SPEC_HASH = "5f5325642ec5a51649e77129ae1bf8e819d922f2027f32f9f0fd89115e5fbd89"
CONCEPTOS_SOURCE_HASH = "669c3b47d97756bbd6c1fd97d882832cb47aef0096fe0e193c8a8d7b0782d3b5"
RUN_ID = UUID("90000000-0000-4000-8000-000000000009")
CANDIDATE_IDS = {
    "idelika": UUID("90000000-0000-4000-8000-000000000101"),
    "conceptos": UUID("90000000-0000-4000-8000-000000000102"),
}


class RepositorioLocalDesechable:
    """Repositorio/base en memoria que conserva únicamente el estado del E2E."""

    def __init__(self, supplier: str, label: str, adapter: str, root_path: str):
        self.source = SourceRecord(
            UUID("90000000-0000-4000-8000-000000000001"),
            supplier,
            label,
            adapter,
            "drive-local",
            "root-local",
            None,
            True,
            None,
        )
        self.active: dict[str, SourceFileRecord] = {}
        self.raw: dict[str, bytes] = {}
        self.assets: dict[str, bytes] = {}
        self.staged: list[dict] = []
        self.failed: list[str] = []

    def get_source(self, supplier):
        assert supplier == self.source.supplier
        return self.source

    def start_run(self, source_id, trigger, requested_by):
        assert source_id == self.source.id and trigger == "manual" and requested_by == 9
        return RUN_ID

    def get_published_snapshot(self, source):
        assert source == self.source
        return None

    def find_file(self, source_id, drive_item_id, e_tag):
        return None

    def list_latest_files(self, source_id, allowed_paths):
        assert source_id == self.source.id
        return tuple(sorted(self.active.values(), key=lambda row: row.path))

    def store_raw_if_absent(self, local_path, sha256, extension, mime_type):
        self.raw[sha256] = Path(local_path).read_bytes()
        return f"catalog-sources/{sha256}.{extension.removeprefix('.')}"

    def materialize_raw_if_present(self, row, destination):
        content = self.raw[row.sha256]
        Path(destination).write_bytes(content)
        return DownloadedFile(Path(destination), len(content), row.sha256)

    def record_source_file(self, source_id, graph_row, downloaded, object_path, run_id, validation):
        row = SourceFileRecord(
            UUID(int=len(self.active) + 1),
            source_id,
            graph_row.id,
            graph_row.path,
            graph_row.e_tag,
            graph_row.c_tag,
            downloaded.size,
            downloaded.sha256,
            graph_row.mime_type,
            object_path,
            validation["status"],
            validation["summary"],
            run_id,
            False,
            None,
            None,
            date(2026, 8, 3),
            None,
        )
        self.active[graph_row.id] = row
        return row

    def mark_file_deleted(self, source_id, drive_item_id, run_id):
        self.active.pop(drive_item_id, None)

    def store_catalog_asset_if_absent(self, object_name, content, content_type):
        assert content_type == "image/png"
        self.assets.setdefault(object_name, content)
        return object_name

    def catalog_asset_matches(
        self, object_name, expected_sha256, expected_size, expected_mime
    ):
        content = self.assets.get(object_name)
        if content is None:
            return None
        return (
            expected_mime == "image/png"
            and len(content) == expected_size
            and hashlib.sha256(content).hexdigest() == expected_sha256
        )

    def stage_candidate(self, run_id, snapshot, metrics, delta_link):
        assert run_id == RUN_ID and delta_link == "local-e2e-delta"
        stored = dict(snapshot)
        stored["generated_at"] = snapshot["generated_at"].isoformat()
        self.staged.append(stored)
        return CANDIDATE_IDS[self.source.supplier]

    def auto_publish_candidate(self, candidate_id):
        return candidate_id

    def finish_no_changes(self, run_id, metrics, delta_link):
        raise AssertionError("el primer sync local no debe ser no_changes")

    def finish_failed(self, run_id, error_code, metrics):
        self.failed.append(error_code)


class GrafoLocalDeFuentes:
    """Graph local que entrega sólo archivos declarados en sources.json."""

    def __init__(self, config, files: dict[str, Path]):
        self.files = {path: source.read_bytes() for path, source in files.items()}
        self.config = config
        self.download_destinations: list[Path] = []

    def iter_delta(self, drive_id, root_id, delta_link=None):
        assert (drive_id, root_id, delta_link) == ("drive-local", "root-local", None)
        items = []
        for configured in self.config.files:
            path = configured.path
            content = self.files[path]
            parent, name = path.rsplit("/", 1) if "/" in path else ("", path)
            items.append(GraphItem(
                configured.drive_item_id,
                name,
                f"/drives/drive-local/root:/{self.config.root_path}/{parent}".rstrip("/"),
                len(content),
                f'"local-{configured.drive_item_id}"',
                None,
                configured.mime_type,
                False,
                None,
            ))
        return DeltaResult(tuple(items), "local-e2e-delta")

    def download_content(self, drive_id, graph_row, destination, max_bytes):
        configured = next(
            row for row in self.config.files if row.drive_item_id == graph_row.id
        )
        content = self.files[configured.path]
        self.download_destinations.append(Path(destination))
        Path(destination).write_bytes(content)
        return DownloadedFile(
            Path(destination), len(content), hashlib.sha256(content).hexdigest()
        )


def _sync_locales():
    configs = {row.supplier: row for row in load_source_config(SOURCES_CONFIG)}
    fixtures = {
        "idelika": IDELIKA_SOURCES,
        "conceptos": {"SPEC GUIDES 2026/CONCEPTOS/Spec guide - Conceptos - Sofas - CdMx - Gdl - Qro - 2021.xlsx": CONCEPTOS_SPEC},
    }
    repositories = {}
    snapshots = {}
    graphs = {}
    for supplier, files in fixtures.items():
        config = configs[supplier]
        repository = RepositorioLocalDesechable(
            supplier, config.label, config.adapter, config.root_path,
        )
        graph = GrafoLocalDeFuentes(config, files)
        result = run_supplier_sync(
            supplier,
            "manual",
            9,
            False,
            repository=repository,
            graph_client=graph,
            adapters=ADAPTERS,
            source_config_path=SOURCES_CONFIG,
        )
        assert result.status == "awaiting_approval", (supplier, result.error_code, repository.failed)
        assert repository.failed == []
        assert len(repository.staged) == 1
        repositories[supplier] = repository
        snapshots[supplier] = repository.staged[0]
        graphs[supplier] = graph
    return repositories, snapshots, graphs


@pytest.fixture(scope="module")
def sincronizados_locales():
    """Un único sync caro; cada prueba sólo consume snapshots inmutables."""

    return _sync_locales()


def _alma_control() -> dict:
    return load_supplier_catalog_data({
        "supplier": "alma",
        "source_hash": "a" * 64,
        "generated_at": "2026-08-03T00:00:00+00:00",
        "items": [{
            "internal_id": "alma:control-local-e2e",
            "supplier": "alma",
            "product_key": "control-local-e2e",
            "sku": "ALMA-CONTROL",
            "code_status": "verified",
            "brand": "ALMA",
            "collection": "Control",
            "name": "Control catálogo existente",
            "description": "Control estable sin cambios.",
            "unit": "PZA",
            "availability_type": "stocked",
            "stock": "9.000000",
            "lead_time": "",
            "base_price_options": [],
            "add_on_options": [],
            "base_currency": "USD",
            "price_net": "100.000000",
            "tax_rate": "0.160000",
            "attributes": {},
            "image_url": "",
            "image_kind": "placeholder",
            "product_url": "",
            "warnings": [],
            "source_reference": "alma:control-local-e2e",
        }],
    }, expected_supplier="alma")


def _selecciones_de_busqueda(catalogs: dict, priced: dict, pending: dict, configurable: dict) -> dict:
    def one(query: str, supplier: str) -> dict:
        result = search_catalog_products(
            catalogs, query=query, supplier=supplier, offset=0, limit=50,
        )
        assert result["total"] >= 1
        return result["items"][0]

    results = {
        "priced": one(priced["sku"] or priced["name"], "idelika"),
        "pending": one(pending["name"], "idelika"),
        "conceptos": one(configurable["sku"], "conceptos"),
        "alma": one("ALMA-CONTROL", "alma"),
    }
    assert results["priced"]["identity"]["internal_id"] == priced["internal_id"]
    assert results["pending"]["identity"]["internal_id"] == pending["internal_id"]
    assert results["pending"]["official_code"] == ""
    assert results["pending"]["display_key"] == pending["internal_id"]
    assert "Código por verificar" in results["pending"]["snapshot"]["warnings"]
    assert "Precio por confirmar" in results["pending"]["snapshot"]["warnings"]
    assert results["conceptos"]["identity"]["internal_id"] == configurable["internal_id"]
    assert results["alma"]["identity"]["internal_id"] == "alma:control-local-e2e"
    return results


def _proyecto_desde_picker_real(selections: dict, replacement: dict, option_id: str) -> dict:
    replacements = {
        "internal_id": replacement["internal_id"],
        "catalog": "idelika",
        "official_code": replacement["sku"],
        "display_key": replacement["internal_id"],
        "identity": {
            "internal_id": replacement["internal_id"],
            "base_option_id": "",
            "add_on_option_ids": [],
        },
        "snapshot": {"name": replacement["name"], "image_url": "", "warnings": []},
    }
    script = f"""
import {{ createCanonicalProductSelection }} from './mobiliti_saas/web/src/productPicker.js';
import {{ createProjectPickerTarget }} from './mobiliti_saas/web/src/projectWorkspace.js';
import {{ createMixedCartLine, replaceProjectLine, serializeProject, toMixedQuoteItem }} from './mobiliti_saas/web/src/mixedCart.js';
const selected = {json.dumps(selections, ensure_ascii=False)};
const replacement = {json.dumps(replacements, ensure_ascii=False)};
const target = (row, base = undefined) => createProjectPickerTarget(
  createCanonicalProductSelection(row, base, [])
);
let lines = [
  createMixedCartLine({{...target(selected.priced), sectionId: 'section-1', role: 'principal', parentLineId: null, position: 0}}),
  createMixedCartLine({{...target(selected.pending), sectionId: 'section-1', role: 'principal', parentLineId: null, position: 1}}),
  createMixedCartLine({{...target(selected.conceptos, {json.dumps(option_id)}), sectionId: 'section-1', role: 'principal', parentLineId: null, position: 2}}),
  createMixedCartLine({{...target(selected.alma), sectionId: 'section-1', role: 'principal', parentLineId: null, position: 3}}),
];
lines[0] = {{...lines[0], quantity: '2'}};
lines = replaceProjectLine(lines, lines[0].lineId, target(replacement)).lines;
const quoteFields = {{proyecto: 'E2E local', cliente: 'Cliente', correo: 'cliente@example.test', telefono: '3330000000', direccion: 'GDL', razon_social: 'Cliente SA', quote_currency: 'USD', descuento: '40', template: 'official_2026_gdl', description_language: 'es'}};
const project = serializeProject({{quoteFields, sections: [{{id: 'section-1', concept: 'Principal'}}], lines}});
console.log(JSON.stringify({{project, items: lines.map(toMixedQuoteItem)}}));
"""
    result = subprocess.run(
        ["node", "--input-type=module", "-e", script],
        cwd=ROOT,
        check=False,
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert result.returncode == 0, result.stderr
    return json.loads(result.stdout)


def test_fuentes_locales_sincronizan_a_snapshots_con_assets_desechables(sincronizados_locales):
    repositories, snapshots, graphs = sincronizados_locales

    assert {
        path: hashlib.sha256(file.read_bytes()).hexdigest()
        for path, file in IDELIKA_SOURCES.items()
    } == {path: EXPECTED_FILE_HASHES[path] for path in IDELIKA_SOURCES}
    assert hashlib.sha256(CONCEPTOS_SPEC.read_bytes()).hexdigest() == EXPECTED_FILE_HASHES["conceptos"]
    assert snapshots["idelika"]["source_hash"] == IDELIKA_GENERATED_SPEC_HASH
    assert snapshots["conceptos"]["source_hash"] == CONCEPTOS_SOURCE_HASH
    assert len(snapshots["idelika"]["items"]) == 220
    priced = [item for item in snapshots["idelika"]["items"] if item["price_net"] is not None]
    pending = [item for item in snapshots["idelika"]["items"] if item["price_net"] is None]
    assert len(priced) == 195
    assert len(pending) == 25
    assert sum(item["collection"] == "School Series" for item in pending) == 20
    assert sum(item["collection"] != "School Series" for item in pending) == 5
    assert all("price_pending" in item["warnings"] for item in pending)
    assert all(item["attributes"]["quotable"] is True for item in pending)
    assert len(snapshots["conceptos"]["items"]) == 40
    assert sum(
        len(item["base_price_options"]) + len(item["add_on_options"])
        for item in snapshots["conceptos"]["items"]
    ) == 171
    assert all(item["base_price_options"] for item in snapshots["conceptos"]["items"])
    assert len(repositories["conceptos"].assets) == 39
    assert len(repositories["idelika"].assets) >= 150
    assert all(
        item["image_kind"] == "official"
        for supplier in ("idelika", "conceptos")
        for item in snapshots[supplier]["items"]
    )
    assert all(
        destination.name.startswith("mobiliti-catalog-")
        for destination in graphs["idelika"].download_destinations
    )
    assert {
        item["attributes"]["provenance"]["file"]
        for item in snapshots["idelika"]["items"]
    } == {Path(path).name for path in IDELIKA_SOURCES}


def test_builder_conserva_spec_y_sidecar_con_nombres_logicos_oficiales():
    config = next(row for row in load_source_config(SOURCES_CONFIG) if row.supplier == "idelika")
    documents = tuple(
        type("DocumentoLocal", (), {
            "path": source.path,
            "kind": source.kind,
            "mime_type": source.mime_type,
            "local_path": IDELIKA_SOURCES[source.path],
        })()
        for source in config.files
    )

    artifact = build_idelika_spec_artifact(documents)

    assert artifact.is_file()
    assert artifact.with_suffix(".validation.json").is_file()


def test_proyecto_mixto_local_cubre_busqueda_configuracion_reemplazo_y_conversion(
    sincronizados_locales,
    monkeypatch,
    tmp_path: Path,
):
    _, synchronized, _ = sincronizados_locales
    catalogs = {**synchronized, "alma": _alma_control()}
    idelika_items = synchronized["idelika"]["items"]
    priced = next(item for item in idelika_items if item["price_net"] is not None)
    replacement = next(
        item for item in idelika_items
        if item["price_net"] is not None and item["internal_id"] != priced["internal_id"]
    )
    pending = next(
        item for item in idelika_items
        if item["collection"] == "School Series" and item["price_net"] is None
    )
    configurable = next(
        item for item in synchronized["conceptos"]["items"]
        if len(item["base_price_options"]) > 1
    )
    option = configurable["base_price_options"][1]

    selections = _selecciones_de_busqueda(catalogs, priced, pending, configurable)
    selected = _proyecto_desde_picker_real(selections, replacement, option["id"])
    project = normalize_project_payload(selected["project"])
    rows = selected["items"]
    rows[2]["quantity"] = "3"
    assert project["lines"][0]["identity"]["internal_id"] == replacement["internal_id"]
    assert project["lines"][0]["quantity"] == "2"
    assert project["lines"][2]["identity"]["base_option_id"] == option["id"]
    payload = build_mixed_catalog_cart_payload(
        rows,
        catalogs=catalogs,
        rate_rows=[{
            "currency": "USD", "effective_date": "2026-08-03",
            "mxn_per_unit": "20.000000", "retrieved_at": "2026-08-03T00:00:00+00:00",
        }],
        quote_currency="USD",
        commercial_discount_percent="40",
        today=date(2026, 8, 3),
    )
    lines = [line for group in payload["groups"] for line in group["items"]]
    by_catalog = {line["canonical_key"]: line for line in lines}
    pending_line = next(line for line in lines if line["canonical_key"].startswith("idelika:") and line["original_unit_price"] is None)
    conceptos_line = next(line for line in lines if line["canonical_key"].startswith("conceptos:"))
    alma_line = next(line for line in lines if line["canonical_key"].startswith("alma:"))
    priced_line = next(line for line in lines if line["canonical_key"].startswith("idelika:") and line["original_unit_price"] is not None)

    assert len(lines) == 4
    assert priced_line["quantity"] == "2.000000"
    assert priced_line["original_unit_price"] == replacement["price_net"]
    assert priced_line["frozen_exchange_rate"] == "0.050000"
    assert priced_line["unit_price"] == format(
        (Decimal(replacement["price_net"]) / Decimal("20")).quantize(Decimal("0.01")), "f"
    )
    assert pending_line["unit_price"] is None
    assert pending_line["frozen_exchange_rate"] is None
    assert pending_line["original_unit_price"] is None
    assert "Precio por confirmar" in pending_line["warnings"]
    assert conceptos_line["quantity"] == "3.000000"
    assert conceptos_line["original_unit_price"] == option["price_net"]
    assert alma_line["original_currency"] == "USD"
    assert alma_line["frozen_exchange_rate"] == "1.000000"
    assert by_catalog
    canonical_rows = quotation_data_rows(payload)

    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    intermediate = tmp_path / "idelika-conceptos-real-worker.xlsx"
    _convert_mixed_catalog_cart_to_quotation(
        tmp_path / "idelika-conceptos-real.json",
        intermediate,
        payload,
    )
    parser_book = load_workbook(intermediate, data_only=False)
    try:
        quotation = parser_book["Quotation"]
        pending_source_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == pending_line["name"]
        )
        assert quotation.cell(pending_source_row, 10).value == "Por confirmar"
        assert quotation.cell(pending_source_row, 15).value is None
        assert quotation.cell(pending_source_row, 16).value is None
        conceptos_source_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == conceptos_line["name"]
        )
        assert option["name"] in str(quotation.cell(conceptos_source_row, 4).value)
        assert Decimal(str(quotation.cell(conceptos_source_row, 7).value)) == Decimal("3")
    finally:
        parser_book.close()

    official = tmp_path / "idelika-conceptos-real-official.xlsx"
    generate_quote(
        intermediate,
        official,
        {
            "catalog_price_mode": "mixed_catalog_converted",
            "catalog_source_hashes": {
                group["catalog"]: group["catalog_source_hash"]
                for group in payload["groups"]
            },
            "quote_currency": "USD",
            "rate_summary": payload["rate_summary"],
            "auto_electrification_rate": payload["auto_electrification_rate"],
            "descuento": 40,
            "cotizacion": "E2E-IDELIKA-CONCEPTOS",
            "proyecto": "E2E local",
            "cliente": "Cliente",
            "description_language": "es",
        },
        WORKER_TEMPLATE,
        original_quotation_path=None,
        quotation_data_rows=canonical_rows,
    )
    workbook = load_workbook(official, data_only=False)
    try:
        quotation = workbook["Quotation"]
        audit = workbook["Quotation_Data"]
        assert audit.sheet_state == "veryHidden"
        assert tuple(
            audit.cell(1, column).value
            for column in range(1, len(QUOTATION_DATA_HEADERS) + 1)
        ) == QUOTATION_DATA_HEADERS
        assert audit.max_row == len(canonical_rows) + 1 == 5
        audit_rows = {
            audit.cell(row, 1).value: row
            for row in range(2, audit.max_row + 1)
        }
        assert set(audit_rows) == {row.item_key for row in canonical_rows}
        decimal_fields = {"original_cost", "frozen_rate", "converted_cost", "quantity"}
        for expected in canonical_rows:
            audit_row = audit_rows[expected.item_key]
            for column, field in enumerate(QUOTATION_DATA_HEADERS, start=1):
                actual = audit.cell(audit_row, column).value
                wanted = getattr(expected, field)
                if wanted is None:
                    assert actual is None
                elif field in decimal_fields:
                    assert Decimal(str(actual)) == wanted
                else:
                    assert actual == wanted

        pending_audit_row = audit_rows[pending_line["line_id"]]
        assert tuple(audit.cell(pending_audit_row, column).value for column in (8, 9, 10)) == (
            None,
            None,
            None,
        )
        for numeric_line in (priced_line, conceptos_line, alma_line):
            numeric_audit_row = audit_rows[numeric_line["line_id"]]
            assert Decimal(str(audit.cell(numeric_audit_row, 8).value)) == Decimal(
                numeric_line["original_unit_price"]
            )
            assert Decimal(str(audit.cell(numeric_audit_row, 9).value)) == Decimal(
                numeric_line["frozen_exchange_rate"]
            )
            assert Decimal(str(audit.cell(numeric_audit_row, 10).value)) == Decimal(
                numeric_line["unit_price"]
            )

        pending_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == pending_line["name"]
        )
        assert quotation.cell(pending_row, 11).value == "Por confirmar"
        assert quotation.cell(pending_row, 12).value in {None, ""}
        assert quotation.cell(pending_row, 12).value != 0
        priced_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == priced_line["name"]
        )
        assert Decimal(str(quotation.cell(priced_row, 11).value)) == Decimal(priced_line["unit_price"])
        alma_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == alma_line["name"]
        )
        assert Decimal(str(quotation.cell(alma_row, 11).value)) == Decimal("100")
        conceptos_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == conceptos_line["name"]
        )
        quotation_rows = {
            pending_line["line_id"]: pending_row,
            priced_line["line_id"]: priced_row,
            conceptos_line["line_id"]: conceptos_row,
            alma_line["line_id"]: alma_row,
        }
        for line in (pending_line, priced_line, conceptos_line, alma_line):
            source_row = quotation_rows[line["line_id"]]
            assert Decimal(str(quotation.cell(source_row, 8).value)) == Decimal(line["quantity"])
        pending_technical_text = " ".join(
            str(quotation.cell(pending_row, column).value or "")
            for column in (4, 5)
        )
        assert "SKU:" not in pending_technical_text
        assert "Clave:" not in pending_technical_text
        assert option["name"] in str(quotation.cell(conceptos_row, 4).value)
        assert f"Clave: {alma_line['code']}" in str(quotation.cell(alma_row, 5).value)
        if priced_line["code"]:
            assert priced_line["code"] in " ".join(
                str(quotation.cell(priced_row, column).value or "")
                for column in (4, 5)
            )
        for numeric_line in (priced_line, conceptos_line, alma_line):
            source_row = quotation_rows[numeric_line["line_id"]]
            assert Decimal(str(quotation.cell(source_row, 11).value)) == Decimal(
                numeric_line["unit_price"]
            )
            assert quotation.cell(source_row, 12).value == f"=H{source_row}*K{source_row}"

        mobiliti = workbook["Mobiliti"]
        assert mobiliti["AD14"].value == 0.4
        assert mobiliti["P4"].value is True
        assert getattr(mobiliti["P6"].value, "text", None) == (
            '=IF(P4=TRUE,_FV(J6,"Price"),0)'
        )
        pending_mobiliti_row = next(
            row for row in range(1, mobiliti.max_row + 1)
            if mobiliti.cell(row, 4).value == f"=Quotation!B{pending_row}"
        )
        assert mobiliti.cell(pending_mobiliti_row, 10).value == "Por confirmar"
        mobiliti_rows = {pending_line["line_id"]: pending_mobiliti_row}
        for numeric_line in (priced_line, conceptos_line, alma_line):
            source_row = quotation_rows[numeric_line["line_id"]]
            mobiliti_row = next(
                row for row in range(1, mobiliti.max_row + 1)
                if mobiliti.cell(row, 4).value == f"=Quotation!B{source_row}"
            )
            mobiliti_rows[numeric_line["line_id"]] = mobiliti_row
            assert mobiliti.cell(mobiliti_row, 8).value == f"=Quotation!H{source_row}"
            assert mobiliti.cell(mobiliti_row, 10).value == f"=Quotation!K{source_row}"
            assert mobiliti.cell(mobiliti_row, 11).value == (
                f'=IFERROR(VLOOKUP(TRIM(F{mobiliti_row}), ProveedoreS_TC, 5, FALSE), "Not Found")'
            )

        cotizacion = workbook["Cotizacion"]
        pending_cotizacion_row = next(
            row for row in range(1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 1).value == f"=Mobiliti!D{pending_mobiliti_row}"
        )
        assert cotizacion.cell(pending_cotizacion_row, 6).value == "Por confirmar"
        assert cotizacion.cell(pending_cotizacion_row, 8).value in {None, ""}
        assert cotizacion.cell(pending_cotizacion_row, 9).value == "Por confirmar"
        assert cotizacion.cell(pending_cotizacion_row, 10).value in {None, ""}
        assert cotizacion.cell(pending_cotizacion_row, 10).value != 0
        for numeric_line in (priced_line, conceptos_line, alma_line):
            mobiliti_row = mobiliti_rows[numeric_line["line_id"]]
            cotizacion_row = next(
                row for row in range(1, cotizacion.max_row + 1)
                if cotizacion.cell(row, 1).value == f"=Mobiliti!D{mobiliti_row}"
            )
            assert cotizacion.cell(cotizacion_row, 3).value == (
                f"=Quotation!D{quotation_rows[numeric_line['line_id']]}"
            )
            assert cotizacion.cell(cotizacion_row, 5).value == f"=Mobiliti!H{mobiliti_row}"
            assert cotizacion.cell(cotizacion_row, 6).value == f"=Mobiliti!AA{mobiliti_row}"
            assert cotizacion.cell(cotizacion_row, 7).value == "=ROUND(Mobiliti!$AD$14,2)"
            assert cotizacion.cell(cotizacion_row, 10).value == (
                f"=E{cotizacion_row}*I{cotizacion_row}"
            )
    finally:
        workbook.close()
