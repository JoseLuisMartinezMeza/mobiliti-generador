from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from mobiliti_saas.quote_engine.supplier_catalog import PUBLIC_ITEM_FIELDS
from mobiliti_saas.worker.catalog_sync.importers import idelika as idelika_importer
from mobiliti_saas.worker.catalog_sync.importers.common import ImageAsset
from mobiliti_saas.worker.catalog_sync.importers.idelika import (
    IdelikaSpecValidationError,
    _normalized_spec_digest,
    build_idelika_snapshot_with_assets,
    load_validated_idelika_spec,
)


ROOT = Path(__file__).resolve().parents[1]
SPEC = (
    ROOT
    / "outputs"
    / "019f7907-1ecc-7001-b3f3-8eb209086fa8"
    / "Spec guide-IDELIKA-2026.xlsx"
)
SIDECAR = SPEC.with_suffix(".validation.json")
SOURCE_ROOT = SPEC.parent / "sources"
OFFICIAL_PDFS = (
    (
        "IDELIKA/1 CATALOGO FABRICACION 2026B.pdf",
        "catalog",
        SOURCE_ROOT / "1 CATALOGO FABRICACION 2026B.pdf",
    ),
    (
        "IDELIKA/2 CATALOGO STOCK 2026.pdf",
        "inventory",
        SOURCE_ROOT / "2 CATALOGO STOCK 2026.pdf",
    ),
    (
        "IDELIKA/4 SCHOOL SERIES 2026.pdf",
        "catalog",
        SOURCE_ROOT / "4 SCHOOL SERIES 2026.pdf",
    ),
)


def _copy_contract(tmp_path: Path) -> Path:
    target = tmp_path / SPEC.name
    shutil.copy2(SPEC, target)
    shutil.copy2(SIDECAR, target.with_suffix(".validation.json"))
    return target


def test_load_validated_idelika_spec_enforces_two_stage_contract() -> None:
    rows = load_validated_idelika_spec(SPEC)

    assert len(rows) == 220
    assert sum(row.subcatalog == "Fabricacion" for row in rows) == 138
    assert sum(row.subcatalog == "Stock" for row in rows) == 62
    assert sum(row.subcatalog == "School Series" for row in rows) == 20
    assert all(row.source_file.lower().endswith(".pdf") for row in rows)
    assert all(row.source_page > 0 for row in rows)


def test_build_idelika_snapshot_uses_cost_mxn_and_preserves_provenance() -> None:
    build = build_idelika_snapshot_with_assets(
        SPEC,
        synced_at=datetime(2026, 8, 2, 12, 0, tzinfo=timezone.utc),
    )
    snapshot = build.snapshot

    assert snapshot["supplier"] == "idelika"
    assert len(snapshot["items"]) == 220
    assert build.assets_by_sha256 == {}
    assert build.bindings == ()
    assert all(set(item) == set(PUBLIC_ITEM_FIELDS) for item in snapshot["items"])

    priced_items = [item for item in snapshot["items"] if item["price_net"] is not None]
    pending_items = [item for item in snapshot["items"] if item["price_net"] is None]
    assert len(priced_items) == 195
    assert len(pending_items) == 25

    priced = priced_items[0]
    assert priced["base_currency"] == "MXN"
    assert priced["price_net"]
    assert priced["attributes"]["reference_price_mxn"] is not None
    assert priced["attributes"]["provenance"]["file"].lower().endswith(".pdf")
    assert priced["attributes"]["provenance"]["page"] > 0

    pending = pending_items[0]
    assert "price_pending" in pending["warnings"]
    assert pending["attributes"]["quotable"] is True
    assert "confirmar" in pending["attributes"]["price_status"]
    assert all(item["attributes"]["quotable"] is True for item in pending_items)

    missing_code = next(item for item in snapshot["items"] if not item["sku"])
    assert missing_code["code_status"] == "needs_review"
    assert "missing_code" in missing_code["warnings"]


def test_loader_rejects_workbook_tampering_with_stable_error(tmp_path: Path) -> None:
    target = _copy_contract(tmp_path)
    workbook = load_workbook(target)
    workbook["Consolidado"]["H2"] = "Producto alterado"
    workbook.save(target)
    workbook.close()

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:",
    ):
        load_validated_idelika_spec(target)


def test_loader_rejects_coherently_forged_workbook_and_sidecar(
    tmp_path: Path,
) -> None:
    target = _copy_contract(tmp_path)
    workbook = load_workbook(target)
    workbook["Consolidado"]["H2"] = "Producto alterado con sidecar coherente"
    workbook.save(target)
    workbook.close()

    forged_digest = _normalized_spec_digest(target)
    sidecar_path = target.with_suffix(".validation.json")
    sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
    sidecar["determinism"]["first_normalized_sha256"] = forged_digest
    sidecar["determinism"]["second_normalized_sha256"] = forged_digest
    sidecar_path.write_text(json.dumps(sidecar), encoding="utf-8")

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:workbook_digest_mismatch$",
    ):
        load_validated_idelika_spec(target)


def test_loader_rejects_sidecar_count_tampering_with_stable_error(
    tmp_path: Path,
) -> None:
    target = _copy_contract(tmp_path)
    sidecar_path = target.with_suffix(".validation.json")
    sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
    sidecar["counts"]["Fabricacion"] += 1
    sidecar_path.write_text(json.dumps(sidecar), encoding="utf-8")

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:",
    ):
        load_validated_idelika_spec(target)


def test_loader_rejects_sidecar_source_hash_tampering(
    tmp_path: Path,
) -> None:
    target = _copy_contract(tmp_path)
    sidecar_path = target.with_suffix(".validation.json")
    sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
    sidecar["source_hashes"]["1 CATALOGO FABRICACION 2026B.pdf"] = "0" * 64
    sidecar_path.write_text(json.dumps(sidecar), encoding="utf-8")

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:sidecar_source_hashes_mismatch$",
    ):
        load_validated_idelika_spec(target)


def test_loader_rejects_sidecar_digest_tampering(
    tmp_path: Path,
) -> None:
    target = _copy_contract(tmp_path)
    sidecar_path = target.with_suffix(".validation.json")
    sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
    sidecar["determinism"]["first_normalized_sha256"] = "0" * 64
    sidecar_path.write_text(json.dumps(sidecar), encoding="utf-8")

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:sidecar_digest_mismatch$",
    ):
        load_validated_idelika_spec(target)


@pytest.mark.parametrize(
    ("manifest_text", "reason"),
    [
        (None, "trust_manifest_missing"),
        ("{}", "trust_manifest_invalid"),
    ],
)
def test_loader_fails_closed_when_trust_manifest_is_unavailable_or_invalid(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    manifest_text: str | None,
    reason: str,
) -> None:
    manifest_path = tmp_path / "idelika_spec_contract.v1.json"
    if manifest_text is not None:
        manifest_path.write_text(manifest_text, encoding="utf-8")
    monkeypatch.setattr(
        idelika_importer,
        "_IDELIKA_TRUST_MANIFEST_PATH",
        manifest_path,
    )

    with pytest.raises(
        IdelikaSpecValidationError,
        match=rf"^IDELIKA_SPEC_VALIDATION_FAILED:{reason}$",
    ):
        load_validated_idelika_spec(SPEC)


def test_pdf_bundle_rechaza_mime_o_nombre_fuera_del_contrato() -> None:
    documents = (
        SimpleNamespace(
            path="IDELIKA/1 CATALOGO FABRICACION 2026B.pdf",
            kind="catalog",
            mime_type="application/octet-stream",
            local_path=SPEC,
        ),
        SimpleNamespace(
            path="IDELIKA/2 CATALOGO STOCK 2026.pdf",
            kind="inventory",
            mime_type="application/pdf",
            local_path=SPEC,
        ),
        SimpleNamespace(
            path="IDELIKA/4 SCHOOL SERIES 2026.pdf",
            kind="catalog",
            mime_type="application/pdf",
            local_path=SPEC,
        ),
    )

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:invalid_source_bundle$",
    ):
        build_idelika_snapshot_with_assets(documents)


def test_pdf_bundle_rejects_source_bytes_outside_the_versioned_contract(
    tmp_path: Path,
) -> None:
    documents = []
    for path, kind, _local_path in OFFICIAL_PDFS:
        local_path = tmp_path / Path(path).name
        local_path.write_bytes(b"not-the-reviewed-pdf")
        documents.append(
            SimpleNamespace(
                path=path,
                kind=kind,
                mime_type="application/pdf",
                local_path=local_path,
            )
        )

    with pytest.raises(
        IdelikaSpecValidationError,
        match=r"^IDELIKA_SPEC_VALIDATION_FAILED:source_hash_mismatch:",
    ):
        build_idelika_snapshot_with_assets(tuple(documents))


def test_pdf_bundle_publica_recorte_oficial_exacto_para_cada_item() -> None:
    documents = tuple(
        SimpleNamespace(
            path=path,
            kind=kind,
            mime_type="application/pdf",
            local_path=local_path,
        )
        for path, kind, local_path in OFFICIAL_PDFS
    )

    build = build_idelika_snapshot_with_assets(
        documents,
        synced_at=datetime(2026, 8, 4, 18, 0, tzinfo=timezone.utc),
    )

    items = build.snapshot["items"]
    assert len(items) == 220
    assert len(build.bindings) == 220
    assert len({binding.internal_id for binding in build.bindings}) == 220
    assert {binding.internal_id for binding in build.bindings} == {
        item["internal_id"] for item in items
    }
    assert len(build.assets_by_sha256) >= 150
    assert all(binding.match_status == "exact_pdf" for binding in build.bindings)
    assert all(len(binding.source_references) == 1 for binding in build.bindings)
    assert all(item["image_kind"] == "official" for item in items)
    assert all(item["attributes"]["image_match"]["status"] == "exact_pdf" for item in items)
    assert all(
        item["attributes"]["approved_asset"]["approved"] is True
        for item in items
    )


def test_pdf_bundle_uses_bundled_validated_spec_without_local_node_builder(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    documents = tuple(
        SimpleNamespace(
            path=path,
            kind=kind,
            mime_type="application/pdf",
            local_path=local_path,
        )
        for path, kind, local_path in OFFICIAL_PDFS
    )
    asset = ImageAsset(b"png", "image/png", 1, 1, "a" * 64)

    def fail_if_local_builder_runs(_documents):
        raise AssertionError("production must not require the local Node builder")

    def fake_assets(rows, _sources):
        return {asset.sha256: asset}, {
            row.stable_key: idelika_importer._PdfAssetMatch(
                asset,
                {"file": row.source_file, "page": row.source_page},
            )
            for row in rows
        }

    monkeypatch.setattr(
        idelika_importer,
        "build_idelika_spec_artifact",
        fail_if_local_builder_runs,
    )
    monkeypatch.setattr(idelika_importer, "_build_idelika_pdf_assets", fake_assets)

    build = build_idelika_snapshot_with_assets(documents)

    assert len(build.snapshot["items"]) == 220
    assert len(build.bindings) == 220
    assert build.snapshot["source_hash"] == (
        "5f5325642ec5a51649e77129ae1bf8e819d922f2027f32f9f0fd89115e5fbd89"
    )
