import inspect
import io
import os
import re
import time
import types
import warnings
import zipfile
from copy import deepcopy
from dataclasses import FrozenInstanceError
from pathlib import Path
from typing import Iterator, get_type_hints
from xml.etree import ElementTree

import fitz
import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image, PngImagePlugin

import mobiliti_saas.worker.catalog_sync.importers.common as common_module
from mobiliti_saas.worker.catalog_sync.importers.common import (
    MAX_FILE_BYTES,
    MAX_IMAGE_BYTES,
    MAX_PDF_PAGES,
    MAX_ZIP_ENTRIES,
    CellRef,
    PdfPage,
    SourceSafetyError,
    extract_xlsx_images,
    extract_xlsx_images_from_bytes,
    iter_pdf_pages,
    neutralize_spreadsheet_text,
    open_xlsx_data_only,
    open_xlsx_data_only_from_bytes,
    read_validated_source,
    source_ref,
    validate_source_file,
)


def xlsx(path, *, formula=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet["A1"] = formula if formula is not None else "SKU-1"
    workbook.save(path)


def pdf(path, *, pages=1):
    document = fitz.open()
    for index in range(pages):
        page = document.new_page()
        page.insert_text((72, 72), f"Page {index + 1}")
    document.save(path)
    document.close()


def rewrite_zip(source, target, transform):
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as output:
        for info in archive.infolist():
            name, data = transform(info.filename, archive.read(info))
            output.writestr(name, data)


def repeat_first_xlsx_image_anchor(source, target, count, *, same_cell=False):
    drawing_namespace = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"

    def repeat(name, data):
        if name.startswith("xl/drawings/drawing") and name.endswith(".xml"):
            root = ElementTree.fromstring(data)
            original = root[0]
            root[:] = []
            for index in range(count):
                anchor = deepcopy(original)
                anchor.find(f"{drawing_namespace}from/{drawing_namespace}row").text = (
                    "0" if same_cell else str(index)
                )
                root.append(anchor)
            data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        return name, data

    rewrite_zip(source, target, repeat)


def assert_code(code, action):
    with pytest.raises(SourceSafetyError) as caught:
        action()
    assert str(caught.value) == code
    assert caught.value.__cause__ is None
    assert caught.value.__context__ is None


class FakePipe:
    def __init__(self, *, poll_result=True, receive=None, close_error=None):
        self.poll_result = poll_result
        self.receive = receive
        self.close_error = close_error
        self.closed = False
        self.sent = []
        self.poll_timeouts = []

    def close(self):
        self.closed = True
        if self.close_error is not None:
            raise self.close_error

    def poll(self, timeout):
        self.poll_timeouts.append(timeout)
        return self.poll_result

    def recv_bytes(self, _maximum):
        if isinstance(self.receive, BaseException):
            raise self.receive
        return self.receive

    def send_bytes(self, value):
        self.sent.append(value)


class FakePopen:
    def __init__(self):
        self.closed = False

    def close(self):
        self.closed = True


class FakeProcess:
    def __init__(self, *, alive=True, terminate_stops=False, kill_stops=True, exitcode=None):
        self.pid = 123
        self.exitcode = exitcode
        self.alive = alive
        self.terminate_stops = terminate_stops
        self.kill_stops = kill_stops
        self.terminate_calls = 0
        self.kill_calls = 0
        self.joins = []
        self.closed = False
        self._popen = FakePopen()

    def start(self):
        pass

    def is_alive(self):
        return self.alive

    def terminate(self):
        self.terminate_calls += 1
        if self.terminate_stops:
            self.alive = False

    def kill(self):
        self.kill_calls += 1
        if self.kill_stops:
            self.alive = False

    def join(self, timeout=None):
        self.joins.append(timeout)

    def close(self):
        self.closed = True


class FakeContext:
    def __init__(self, process, output_read):
        self.process = process
        self.control_read = FakePipe()
        self.control_write = FakePipe()
        self.output_read = output_read
        self.output_write = FakePipe()
        self.pipes = [
            (self.control_read, self.control_write),
            (self.output_read, self.output_write),
        ]

    def Pipe(self, *, duplex):
        assert duplex is False
        return self.pipes.pop(0)

    def Process(self, **_kwargs):
        return self.process


def test_valid_minimal_xlsx_pdf_and_png_pass(tmp_path):
    book = tmp_path / "catalog.xlsx"
    document = tmp_path / "catalog.pdf"
    picture = tmp_path / "catalog.png"
    xlsx(book)
    pdf(document)
    Image.new("RGB", (2, 3), "red").save(picture)

    assert validate_source_file(book, ".xlsx").extension == ".xlsx"
    assert validate_source_file(document, "pdf").extension == ".pdf"
    assert validate_source_file(picture, ".png").extension == ".png"
    assert validate_source_file(book, ".xlsx").sha256 == validate_source_file(book, ".xlsx").sha256


def test_source_rejects_extension_magic_symlink_size_and_identity_change(monkeypatch, tmp_path):
    book = tmp_path / "catalog.xlsx"
    xlsx(book)
    false_extension = tmp_path / "catalog.pdf"
    false_extension.write_bytes(book.read_bytes())
    assert_code("SOURCE_TYPE", lambda: validate_source_file(false_extension, ".pdf"))
    assert_code("SOURCE_TYPE", lambda: validate_source_file(book, ".pdf"))
    assert_code("SOURCE_SIZE", lambda: validate_source_file(book, ".xlsx", max_bytes=1))

    link = tmp_path / "link.xlsx"
    try:
        link.symlink_to(book)
    except OSError:
        pass
    else:
        assert_code("SOURCE_FILE", lambda: validate_source_file(link, ".xlsx"))

    real_fstat = os.fstat
    calls = 0

    def changed(descriptor):
        nonlocal calls
        result = real_fstat(descriptor)
        calls += 1
        if calls < 2:
            return result
        return types.SimpleNamespace(
            st_mode=result.st_mode,
            st_dev=result.st_dev,
            st_ino=result.st_ino,
            st_size=result.st_size + 1,
            st_mtime_ns=result.st_mtime_ns,
        )

    monkeypatch.setattr("mobiliti_saas.worker.catalog_sync.importers.common.os.fstat", changed)
    assert_code("SOURCE_CHANGED", lambda: validate_source_file(book, ".xlsx"))


@pytest.mark.parametrize(
    "member",
    ["../escape.xml", "/absolute.xml", "C:/absolute.xml"],
    ids=["traversal", "posix-absolute", "windows-absolute"],
)
def test_xlsx_rejects_unsafe_member_names(tmp_path, member):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "unsafe.xlsx"
    xlsx(source)
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(target, "w") as output:
        for info in archive.infolist():
            output.writestr(info.filename, archive.read(info))
        output.writestr(member, b"x")
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(target))


def test_xlsx_rejects_duplicate_encrypted_macro_ole_and_external_link(tmp_path):
    source = tmp_path / "source.xlsx"
    xlsx(source)

    duplicate = tmp_path / "duplicate.xlsx"
    duplicate.write_bytes(source.read_bytes())
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(duplicate, "a") as archive:
            archive.writestr("docProps/app.xml", b"duplicate")
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(duplicate))

    encrypted = tmp_path / "encrypted.xlsx"
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(encrypted, "w") as output:
        for index, info in enumerate(archive.infolist()):
            cloned = zipfile.ZipInfo(info.filename)
            cloned.compress_type = info.compress_type
            cloned.flag_bits = 1 if index == 0 else 0
            output.writestr(cloned, archive.read(info))
    raw = bytearray(encrypted.read_bytes())
    first = raw.find(b"PK\x03\x04")
    raw[first + 6 : first + 8] = (1).to_bytes(2, "little")
    encrypted.write_bytes(raw)
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(encrypted))

    for label, name, data in (
        ("macro", "xl/vbaProject.bin", b"macro"),
        ("ole", "xl/embeddings/oleObject1.bin", b"ole"),
        (
            "external",
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" Target="https://secret.invalid/x" TargetMode="External"/>'
            b"</Relationships>",
        ),
    ):
        target = tmp_path / f"{label}.xlsx"
        with zipfile.ZipFile(source) as archive, zipfile.ZipFile(target, "w") as output:
            for info in archive.infolist():
                output.writestr(info.filename, archive.read(info))
            output.writestr(name, data)
        assert_code("XLSX_UNSAFE", lambda target=target: open_xlsx_data_only(target))


def test_xlsx_rejects_relationship_to_missing_package_part(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "missing-part.xlsx"
    xlsx(source)

    def missing_part(name, data):
        if name == "_rels/.rels":
            data = data.replace(
                b"</Relationships>",
                b'<Relationship Id="missing" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml" Target="missing.xml"/></Relationships>',
            )
        return name, data

    rewrite_zip(source, target, missing_part)
    assert_code("XLSX_INVALID", lambda: open_xlsx_data_only(target))


def test_xlsx_rejects_xlm_macrosheet_and_false_microsoft_relationship(tmp_path):
    source = tmp_path / "source.xlsx"
    xlsx(source)

    macrosheet = tmp_path / "macrosheet.xlsx"
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(macrosheet, "w", zipfile.ZIP_DEFLATED) as output:
        for info in archive.infolist():
            data = archive.read(info)
            if info.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/macrosheets/sheet1.xml" ContentType="application/vnd.ms-excel.macrosheet+xml"/></Types>',
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace(
                    b"</Relationships>",
                    b'<Relationship Id="xlm" Type="http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet" Target="macrosheets/sheet1.xml"/></Relationships>',
                )
            output.writestr(info.filename, data)
        output.writestr("xl/macrosheets/sheet1.xml", b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>')
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(macrosheet))

    unknown = tmp_path / "unknown-active.xlsx"
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(unknown, "w", zipfile.ZIP_DEFLATED) as output:
        for info in archive.infolist():
            data = archive.read(info)
            if info.filename == "_rels/.rels":
                data = data.replace(
                    b"</Relationships>",
                    b'<Relationship Id="active" Type="http://schemas.microsoft.com/office/2006/relationships/worksheet" Target="active.xml"/></Relationships>',
                )
            output.writestr(info.filename, data)
        output.writestr("active.xml", b"<active/>")
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(unknown))


def test_xlsx_structural_allowlist_accepts_passive_comments(tmp_path):
    source = tmp_path / "comment.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "ordinary"
    workbook.active["A1"].comment = Comment("reviewed", "Mobiliti")
    workbook.save(source)

    validated = open_xlsx_data_only(source)
    assert validated.active["A1"].value == "ordinary"
    validated.close()


def test_xlsx_accepts_passive_hyperlinks_without_following_them(tmp_path):
    source = tmp_path / "hyperlink.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "Mobiliti"
    workbook.active["A1"].hyperlink = "http://www.mobiliti.mx/"
    workbook.save(source)

    validated = open_xlsx_data_only(source)
    assert validated.active["A1"].value == "Mobiliti"
    assert validated.active["A1"].hyperlink is None
    validated.close()


def test_xlsx_ignores_null_external_image_marker_but_rejects_remote_images(tmp_path):
    picture = tmp_path / "picture.png"
    source = tmp_path / "source.xlsx"
    Image.new("RGB", (2, 3), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.save(source)

    def with_external_image(target):
        output = tmp_path / ("external-null.xlsx" if target == "NULL" else "external-remote.xlsx")

        def transform(name, data):
            if name.startswith("xl/drawings/_rels/") and name.endswith(".rels"):
                root = ElementTree.fromstring(data)
                for relationship in root:
                    if relationship.get("Type", "").endswith("/image"):
                        relationship.set("Target", target)
                        relationship.set("TargetMode", "External")
                data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            return name, data

        rewrite_zip(source, output, transform)
        return output

    null_marker = with_external_image("NULL")
    assert validate_source_file(null_marker, ".xlsx").extension == ".xlsx"
    assert extract_xlsx_images(null_marker) == {}
    validated = open_xlsx_data_only(null_marker)
    validated.close()
    assert_code(
        "XLSX_UNSAFE",
        lambda: validate_source_file(with_external_image("https://example.invalid/image.png"), ".xlsx"),
    )


def test_xlsx_skips_passive_emf_without_decoding_it(tmp_path, monkeypatch):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    source = tmp_path / "source.xlsx"
    mixed = tmp_path / "mixed-images.xlsx"
    Image.new("RGB", (2, 3), "red").save(first)
    Image.new("RGB", (3, 4), "blue").save(second)
    workbook = Workbook()
    workbook.active["C1"] = "ordinary"
    workbook.active.add_image(XlsxImage(first), "A1")
    workbook.active.add_image(XlsxImage(second), "B2")
    workbook.save(source)

    def make_second_image_emf(name, data):
        if name == "[Content_Types].xml":
            data = data.replace(
                b"</Types>",
                b'<Default Extension="emf" ContentType="image/x-emf"/></Types>',
            )
        elif name.startswith("xl/drawings/_rels/") and name.endswith(".rels"):
            data = data.replace(b"image2.png", b"image2.emf")
        elif name == "xl/media/image2.png":
            return "xl/media/image2.emf", b"inert-emf-must-not-be-decoded"
        return name, data

    rewrite_zip(source, mixed, make_second_image_emf)
    decoded = []
    real_normalize = common_module._normalize_image

    def audited_normalize(data):
        decoded.append(data)
        return real_normalize(data)

    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)

    validated, data = read_validated_source(mixed, ".xlsx")
    loaded = open_xlsx_data_only_from_bytes(data)
    assets = extract_xlsx_images_from_bytes(data)

    assert validated.extension == ".xlsx"
    assert loaded.active["C1"].value == "ordinary"
    assert list(assets) == [CellRef("Sheet", "A1")]
    assert all(raw != b"inert-emf-must-not-be-decoded" for raw in decoded)
    loaded.close()


@pytest.mark.parametrize(
    ("media_name", "relationship_target"),
    [
        ("image2.png", "../media/image2.png"),
        ("image2.emf", "../media/image2.emf/."),
    ],
)
def test_xlsx_skips_emf_by_declared_content_type_and_canonical_target(
    tmp_path, monkeypatch, media_name, relationship_target
):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    source = tmp_path / "source.xlsx"
    mixed = tmp_path / "mixed-images.xlsx"
    Image.new("RGB", (2, 3), "red").save(first)
    Image.new("RGB", (3, 4), "blue").save(second)
    workbook = Workbook()
    workbook.active["C1"] = "ordinary"
    workbook.active.add_image(XlsxImage(first), "A1")
    workbook.active.add_image(XlsxImage(second), "B2")
    workbook.save(source)

    def disguise_second_image(name, data):
        if name == "[Content_Types].xml":
            root = ElementTree.fromstring(data)
            namespace = root.tag.removesuffix("Types")
            root.append(
                ElementTree.Element(
                    f"{namespace}Override",
                    PartName=f"/xl/media/{media_name}",
                    ContentType="image/x-emf",
                )
            )
            data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        elif name.startswith("xl/drawings/_rels/") and name.endswith(".rels"):
            root = ElementTree.fromstring(data)
            image_relationships = [
                relationship for relationship in root if relationship.get("Type", "").endswith("/image")
            ]
            image_relationships[1].set("Target", relationship_target)
            data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        elif name == "xl/media/image2.png":
            return f"xl/media/{media_name}", b"inert-emf-must-not-be-decoded"
        return name, data

    rewrite_zip(source, mixed, disguise_second_image)
    decoded = []
    opened = []
    real_normalize = common_module._normalize_image
    real_image_open = Image.open

    def audited_normalize(data):
        decoded.append(data)
        return real_normalize(data)

    def audited_image_open(source, *args, **kwargs):
        if isinstance(source, io.BytesIO):
            opened.append(source.getvalue())
        return real_image_open(source, *args, **kwargs)

    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)
    monkeypatch.setattr(Image, "open", audited_image_open)

    validated, data = read_validated_source(mixed, ".xlsx")
    loaded = open_xlsx_data_only_from_bytes(data)
    assets = extract_xlsx_images_from_bytes(data)

    assert validated.extension == ".xlsx"
    assert loaded.active["C1"].value == "ordinary"
    assert list(assets) == [CellRef("Sheet", "A1")]
    assert all(raw != b"inert-emf-must-not-be-decoded" for raw in decoded)
    assert all(raw != b"inert-emf-must-not-be-decoded" for raw in opened)
    loaded.close()


def test_validated_xlsx_bytes_are_the_bytes_parsed(tmp_path):
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "first"
    workbook.save(source)
    validated, data = read_validated_source(source, ".xlsx")

    workbook = Workbook()
    workbook.active["A1"] = "second"
    workbook.save(source)

    loaded = open_xlsx_data_only_from_bytes(data)
    assert loaded.active["A1"].value == "first"
    assert extract_xlsx_images_from_bytes(data) == {}
    assert validated.sha256 != validate_source_file(source, ".xlsx").sha256
    loaded.close()


def test_xlsx_omits_ambiguous_images_anchored_to_the_same_cell(tmp_path):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    source = tmp_path / "ambiguous-images.xlsx"
    Image.new("RGB", (2, 3), "red").save(first)
    Image.new("RGB", (2, 3), "blue").save(second)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(first), "A1")
    workbook.active.add_image(XlsxImage(second), "A1")
    workbook.save(source)

    assert extract_xlsx_images(source) == {}


def test_xlsx_rejects_entry_expansion_ratio_and_dde_limits(monkeypatch, tmp_path):
    source = tmp_path / "source.xlsx"
    xlsx(source)

    monkeypatch.setattr("mobiliti_saas.worker.catalog_sync.importers.common.MAX_ZIP_ENTRIES", 1)
    assert_code("XLSX_LIMIT", lambda: open_xlsx_data_only(source))
    monkeypatch.undo()

    monkeypatch.setattr("mobiliti_saas.worker.catalog_sync.importers.common.MAX_WORKBOOK_CELLS", 0)
    assert_code("XLSX_LIMIT", lambda: open_xlsx_data_only(source))
    monkeypatch.undo()

    bomb = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", b"0" * 100_000)
    assert_code("XLSX_LIMIT", lambda: open_xlsx_data_only(bomb))

    dde = tmp_path / "dde.xlsx"
    xlsx(dde, formula="=cmd|' /C secret'!A0")
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(dde))


def test_xlsx_accepts_bounded_merged_ranges(tmp_path):
    source = tmp_path / "bounded-merges.xlsx"
    workbook = Workbook()
    workbook.active.merge_cells("A1:C2")
    workbook.active.merge_cells("E1:F3")
    workbook.save(source)

    loaded = open_xlsx_data_only(source)
    assert {str(merged) for merged in loaded.active.merged_cells.ranges} == {"A1:C2", "E1:F3"}
    loaded.close()


@pytest.mark.parametrize(
    ("limit_name", "limit"),
    [
        ("MAX_WORKBOOK_MERGED_RANGES", 1),
        ("MAX_MERGED_RANGE_CELLS", 3),
        ("MAX_WORKBOOK_MERGED_CELLS", 7),
    ],
)
def test_xlsx_merge_limits_run_before_openpyxl(monkeypatch, tmp_path, limit_name, limit):
    source = tmp_path / f"{limit_name}.xlsx"
    workbook = Workbook()
    workbook.active.merge_cells("A1:B2")
    workbook.active.merge_cells("D1:E2")
    workbook.save(source)

    monkeypatch.setattr(common_module, limit_name, limit, raising=False)
    monkeypatch.setattr(
        common_module,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("openpyxl ran before merge preflight"),
    )

    assert_code("XLSX_LIMIT", lambda: open_xlsx_data_only(source))


def test_formula_coordinates_are_sorted_and_cached_values_are_removed(tmp_path):
    source = tmp_path / "formula-source.xlsx"
    target = tmp_path / "formula.xlsx"
    xlsx(source, formula="=1+1")

    def cached_value(name, data):
        if name == "xl/worksheets/sheet1.xml":
            data = data.replace(b"<v></v>", b"<v>2</v>")
        return name, data

    rewrite_zip(source, target, cached_value)
    workbook = open_xlsx_data_only(target)
    assert workbook.formula_cells == (CellRef("Catalog", "A1"),)
    assert workbook["Catalog"]["A1"].value is None
    workbook.close()


def test_defined_names_cannot_hide_dde(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "defined-name.xlsx"
    xlsx(source)

    def add_dde(name, data):
        if name == "xl/workbook.xml":
            data = data.replace(
                b"</workbook>",
                b"<definedNames><definedName name=\"Bad\">cmd|' /C secret'!A0</definedName></definedNames></workbook>",
            )
        return name, data

    rewrite_zip(source, target, add_dde)
    assert_code("XLSX_UNSAFE", lambda: open_xlsx_data_only(target))


def test_pdf_pages_are_bounded_and_deterministic(tmp_path, monkeypatch):
    source = tmp_path / "catalog.pdf"
    pdf(source, pages=2)
    pages = list(iter_pdf_pages(source))
    assert [(page.page_number, page.text.strip()) for page in pages] == [(1, "Page 1"), (2, "Page 2")]

    monkeypatch.setattr("mobiliti_saas.worker.catalog_sync.importers.common.MAX_PDF_PAGES", 1)
    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_enforces_aggregate_expanded_stream_bound(tmp_path, monkeypatch):
    source = tmp_path / "aggregate.pdf"
    pdf(source)
    monkeypatch.setattr(
        "mobiliti_saas.worker.catalog_sync.importers.common.MAX_PDF_STREAM_EXPANDED_BYTES",
        1,
    )
    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_per_call_profile_preserves_default_and_uses_isolated_extraction(
    tmp_path, monkeypatch
):
    source = tmp_path / "profile.pdf"
    pdf(source, pages=2)
    monkeypatch.setattr(common_module, "MAX_PDF_STREAM_EXPANDED_BYTES", 1)

    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))
    monkeypatch.setattr(
        fitz.Page,
        "get_text",
        lambda *_args, **_kwargs: pytest.fail("in-process extractor called"),
    )

    pages = list(
        iter_pdf_pages(
            source,
            pdf_max_pages=80,
            pdf_max_stream_expanded_bytes=384 * 1024 * 1024,
        )
    )

    assert [page.text.strip() for page in pages] == ["Page 1", "Page 2"]


@pytest.mark.parametrize(
    ("pages", "expanded"),
    [
        (81, 384 * 1024 * 1024),
        (80, 384 * 1024 * 1024 + 1),
        (0, 384 * 1024 * 1024),
        (80, 0),
        (True, 384 * 1024 * 1024),
        (80, True),
        (80, None),
        (None, 384 * 1024 * 1024),
    ],
)
def test_pdf_per_call_profile_rejects_invalid_or_over_hard_cap(
    tmp_path, pages, expanded
):
    source = tmp_path / "profile-invalid.pdf"
    pdf(source)

    assert_code(
        "SOURCE_ARGUMENT",
        lambda: list(
            iter_pdf_pages(
                source,
                pdf_max_pages=pages,
                pdf_max_stream_expanded_bytes=expanded,
            )
        ),
    )


def test_pdf_per_call_profile_keeps_page_stream_and_active_content_guards(
    tmp_path,
):
    too_many_pages = tmp_path / "too-many-pages.pdf"
    pdf(too_many_pages, pages=81)
    assert_code(
        "PDF_LIMIT",
        lambda: list(
            iter_pdf_pages(
                too_many_pages,
                pdf_max_pages=80,
                pdf_max_stream_expanded_bytes=384 * 1024 * 1024,
            )
        ),
    )

    active = tmp_path / "active-profile.pdf"
    document = fitz.open()
    document.new_page()
    document.xref_set_key(
        document.pdf_catalog(),
        "OpenAction",
        "<</S/Launch/F(https://secret.invalid)>>",
    )
    document.save(active)
    document.close()
    assert_code(
        "PDF_UNSAFE",
        lambda: list(
            iter_pdf_pages(
                active,
                pdf_max_pages=80,
                pdf_max_stream_expanded_bytes=384 * 1024 * 1024,
            )
        ),
    )

    tiny_stream_budget = tmp_path / "tiny-stream-budget.pdf"
    pdf(tiny_stream_budget)
    assert_code(
        "PDF_LIMIT",
        lambda: list(
            iter_pdf_pages(
                tiny_stream_budget,
                pdf_max_pages=80,
                pdf_max_stream_expanded_bytes=1,
            )
        ),
    )


def test_pdf_accepts_passive_web_links_and_bounds_highly_compressed_images(tmp_path, monkeypatch):
    picture = tmp_path / "flat.png"
    Image.new("RGB", (1200, 1200), "white").save(picture)
    source = tmp_path / "passive.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_image(fitz.Rect(0, 0, 120, 120), filename=picture)
    page.insert_link(
        {
            "kind": fitz.LINK_URI,
            "from": fitz.Rect(0, 130, 120, 150),
            "uri": "https://www.crglobal.mx/",
        }
    )
    document.save(source, deflate=True)
    document.close()

    pages = list(iter_pdf_pages(source))
    assert len(pages) == 1
    assert pages[0].image_count == 1

    monkeypatch.setattr(
        "mobiliti_saas.worker.catalog_sync.importers.common.MAX_PDF_STREAM_DECODED_BYTES", 1_000_000
    )
    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_accepts_bounded_indexed_color_images(tmp_path):
    picture = tmp_path / "indexed.png"
    image = Image.new("P", (32, 32))
    image.putpalette([0, 0, 0, 255, 255, 255] + [0, 0, 0] * 254)
    image.save(picture)
    source = tmp_path / "indexed.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_image(fitz.Rect(0, 0, 32, 32), filename=picture)
    image_xref = page.get_images(full=True)[0][0]
    palette_xref = document.get_new_xref()
    document.update_object(palette_xref, "<<>>")
    document.update_stream(palette_xref, bytes(range(256)) * 3)
    color_space_xref = document.get_new_xref()
    document.update_object(
        color_space_xref,
        f"[/Indexed /DeviceRGB 255 {palette_xref} 0 R]",
    )
    document.xref_set_key(image_xref, "ColorSpace", f"{color_space_xref} 0 R")
    document.save(source, deflate=True)
    document.close()

    pages = list(iter_pdf_pages(source))
    assert len(pages) == 1
    assert pages[0].image_count == 1

    truncated = tmp_path / "truncated-palette.pdf"
    document = fitz.open(source)
    image_xref = document[0].get_images(full=True)[0][0]
    color_space_xref = int(document.xref_get_key(image_xref, "ColorSpace")[1].split()[0])
    palette_xref = int(re.findall(r"([1-9][0-9]*) 0 R", document.xref_object(color_space_xref))[-1])
    document.update_stream(palette_xref, b"\0")
    document.save(truncated)
    document.close()
    assert_code("PDF_INVALID", lambda: list(iter_pdf_pages(truncated)))


def test_pdf_counts_byte_aligned_image_rows_in_aggregate_bound(tmp_path, monkeypatch):
    picture = tmp_path / "narrow.png"
    Image.new("1", (1, 32), 1).save(picture)
    source = tmp_path / "narrow.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_image(fitz.Rect(0, 0, 1, 32), filename=picture)
    image_xref = page.get_images(full=True)[0][0]
    document.xref_set_key(image_xref, "Width", "1")
    document.xref_set_key(image_xref, "Height", "16384")
    document.xref_set_key(image_xref, "BitsPerComponent", "1")
    document.xref_set_key(image_xref, "ColorSpace", "/DeviceGray")
    document.save(source, deflate=True)
    document.close()
    monkeypatch.setattr(
        "mobiliti_saas.worker.catalog_sync.importers.common.MAX_PDF_STREAM_EXPANDED_BYTES",
        10_000,
    )
    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_rejects_passive_uri_with_chained_action(tmp_path):
    source = tmp_path / "chained-action.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_link(
        {
            "kind": fitz.LINK_URI,
            "from": fitz.Rect(0, 0, 120, 20),
            "uri": "https://www.crglobal.mx/",
        }
    )
    annotation = page.annot_xrefs()[0][0]
    document.xref_set_key(
        annotation,
        "A",
        "<</S/URI/URI(https://www.crglobal.mx/)/Next<</S/Named/N/Print>>>>",
    )
    document.save(source)
    document.close()

    assert_code("PDF_UNSAFE", lambda: list(iter_pdf_pages(source)))


def test_pdf_accepts_passive_internal_open_action(tmp_path):
    source = tmp_path / "internal-open-action.pdf"
    document = fitz.open()
    page = document.new_page()
    action = document.get_new_xref()
    document.update_object(action, f"<</S/GoTo/D[{page.xref} 0 R/Fit]>>")
    document.xref_set_key(document.pdf_catalog(), "OpenAction", f"{action} 0 R")
    document.save(source)
    document.close()

    pages = list(iter_pdf_pages(source))
    assert len(pages) == 1


def test_pdf_rejects_bad_magic_malformed_encrypted_and_active_content(tmp_path):
    bad_magic = tmp_path / "bad.pdf"
    bad_magic.write_bytes(b"not a pdf")
    assert_code("SOURCE_TYPE", lambda: list(iter_pdf_pages(bad_magic)))

    malformed = tmp_path / "malformed.pdf"
    malformed.write_bytes(b"%PDF-this is not valid")
    assert_code("PDF_INVALID", lambda: list(iter_pdf_pages(malformed)))

    plain = fitz.open()
    plain.new_page()
    encrypted = tmp_path / "encrypted.pdf"
    plain.save(encrypted, encryption=fitz.PDF_ENCRYPT_AES_256, owner_pw="owner", user_pw="secret")
    plain.close()
    assert_code("PDF_UNSAFE", lambda: list(iter_pdf_pages(encrypted)))

    active = tmp_path / "active.pdf"
    document = fitz.open()
    document.new_page()
    document.xref_set_key(document.pdf_catalog(), "OpenAction", "<</S/Launch/F(https://secret.invalid)>>")
    document.save(active)
    document.close()
    assert_code("PDF_UNSAFE", lambda: list(iter_pdf_pages(active)))

    embedded = tmp_path / "embedded.pdf"
    document = fitz.open()
    document.new_page()
    document.embfile_add("payload.bin", b"secret")
    document.save(embedded)
    document.close()
    assert_code("PDF_UNSAFE", lambda: list(iter_pdf_pages(embedded)))


def test_pdf_rejects_widget_button_named_print_action(tmp_path):
    source = tmp_path / "button.pdf"
    document = fitz.open()
    page = document.new_page()
    widget = fitz.Widget()
    widget.field_name = "print-button"
    widget.field_type = fitz.PDF_WIDGET_TYPE_BUTTON
    widget.rect = fitz.Rect(72, 72, 160, 100)
    page.add_widget(widget)
    document.xref_set_key(page.first_widget.xref, "A", "<</S/Named/N/Print>>")
    document.save(source)
    document.close()

    assert_code("PDF_UNSAFE", lambda: list(iter_pdf_pages(source)))


def test_pdf_rejects_compressed_stream_before_text_extraction(monkeypatch, tmp_path):
    source = tmp_path / "compressed-stream.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "small visible text")
    content_xref = page.get_contents()[0]
    document.update_stream(content_xref, b"%" + b"A" * 4096, compress=True)
    document.save(source)
    document.close()

    monkeypatch.setattr(common_module, "MAX_PDF_STREAM_DECODED_BYTES", 1024, raising=False)
    monkeypatch.setattr(common_module, "MAX_PDF_STREAM_RATIO", 4, raising=False)
    monkeypatch.setattr(fitz.Page, "get_text", lambda *_args, **_kwargs: pytest.fail("text decoded before preflight"))
    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_text_extraction_never_runs_in_process(monkeypatch, tmp_path):
    source = tmp_path / "isolated.pdf"
    pdf(source)
    monkeypatch.setattr(fitz.Page, "get_text", lambda *_args, **_kwargs: pytest.fail("in-process extractor called"))

    assert list(iter_pdf_pages(source))[0].text.strip() == "Page 1"


def test_pdf_isolated_text_output_is_rejected_at_limit(monkeypatch, tmp_path):
    source = tmp_path / "oversized-text.pdf"
    document = fitz.open()
    page = document.new_page(width=20_000, height=500)
    page.insert_text((20, 72), "A" * 20_000)
    document.save(source)
    document.close()
    monkeypatch.setattr(common_module, "MAX_PDF_TEXT_BYTES", 64)
    monkeypatch.setattr(common_module, "MAX_PDF_STREAM_RATIO", 1_000_000)
    monkeypatch.setattr(fitz.Page, "get_text", lambda *_args, **_kwargs: pytest.fail("in-process extractor called"))

    assert_code("PDF_LIMIT", lambda: list(iter_pdf_pages(source)))


def test_pdf_worker_timeout_shutdown_is_bounded_and_closes_resources(monkeypatch):
    process = FakeProcess(alive=True, terminate_stops=False, kill_stops=True)
    context = FakeContext(process, FakePipe(poll_result=False))
    closed_jobs = []
    monkeypatch.setattr(common_module.multiprocessing, "get_context", lambda _method: context)
    monkeypatch.setattr(common_module, "_windows_text_job", lambda _process: 77)
    monkeypatch.setattr(common_module, "_close_windows_handle", closed_jobs.append)
    monkeypatch.setattr(common_module, "MAX_PDF_TEXT_SECONDS", 0.01)
    monkeypatch.setattr(common_module, "MAX_PDF_WORKER_SHUTDOWN_SECONDS", 0.01, raising=False)

    started = time.monotonic()
    assert_code("PDF_LIMIT", lambda: common_module._pdf_text_isolated(b"%PDF-test", 1))

    assert time.monotonic() - started < 0.5
    assert process.terminate_calls == 1
    assert process.kill_calls == 1
    assert process.joins and all(timeout is not None for timeout in process.joins)
    assert process.closed
    assert closed_jobs == [77]
    assert all(pipe.closed for pipe in (
        context.control_read,
        context.control_write,
        context.output_read,
        context.output_write,
    ))


def test_pdf_worker_unexpected_death_is_stable_and_bounded(monkeypatch):
    process = FakeProcess(alive=False, exitcode=9)
    context = FakeContext(process, FakePipe(receive=EOFError()))
    monkeypatch.setattr(common_module.multiprocessing, "get_context", lambda _method: context)
    monkeypatch.setattr(common_module, "_windows_text_job", lambda _process: 88)
    monkeypatch.setattr(common_module, "_close_windows_handle", lambda _handle: None)

    started = time.monotonic()
    assert_code("PDF_LIMIT", lambda: common_module._pdf_text_isolated(b"%PDF-test", 1))

    assert time.monotonic() - started < 0.5
    assert process.terminate_calls == process.kill_calls == 0
    assert process.joins and all(timeout is not None for timeout in process.joins)
    assert process.closed
    assert all(pipe.closed for pipe in (
        context.control_read,
        context.control_write,
        context.output_read,
        context.output_write,
    ))


def test_pdf_worker_that_survives_kill_returns_without_blocking(monkeypatch):
    process = FakeProcess(alive=True, terminate_stops=False, kill_stops=False)
    context = FakeContext(process, FakePipe(poll_result=False))
    closed_jobs = []
    monkeypatch.setattr(common_module.multiprocessing, "get_context", lambda _method: context)
    monkeypatch.setattr(common_module, "_windows_text_job", lambda _process: 99)
    monkeypatch.setattr(common_module, "_close_windows_handle", closed_jobs.append)
    monkeypatch.setattr(common_module, "MAX_PDF_TEXT_SECONDS", 0.01)
    monkeypatch.setattr(common_module, "MAX_PDF_WORKER_SHUTDOWN_SECONDS", 0.01, raising=False)

    started = time.monotonic()
    assert_code("PDF_LIMIT", lambda: common_module._pdf_text_isolated(b"%PDF-test", 1))

    assert time.monotonic() - started < 0.5
    assert process.terminate_calls == process.kill_calls == 1
    assert process.joins and all(timeout is not None for timeout in process.joins)
    assert closed_jobs == [99]
    assert process._popen.closed


def test_pdf_worker_pipe_close_failures_are_redacted_and_best_effort(monkeypatch):
    process = FakeProcess(alive=False, exitcode=9)
    context = FakeContext(process, FakePipe(receive=EOFError()))
    context.output_read.close_error = RuntimeError("SECRET pipe close")
    monkeypatch.setattr(common_module.multiprocessing, "get_context", lambda _method: context)
    monkeypatch.setattr(common_module, "_windows_text_job", lambda _process: 101)
    monkeypatch.setattr(common_module, "_close_windows_handle", lambda _handle: None)

    assert_code("PDF_LIMIT", lambda: common_module._pdf_text_isolated(b"%PDF-test", 1))
    assert all(pipe.closed for pipe in (
        context.control_read,
        context.control_write,
        context.output_read,
        context.output_write,
    ))


def test_xlsx_images_are_verified_normalized_bounded_and_sorted(tmp_path, monkeypatch):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    Image.new("RGB", (3, 2), "red").save(first)
    Image.new("RGB", (2, 4), "blue").save(second)
    book = tmp_path / "images.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pictures"
    sheet.add_image(XlsxImage(second), "B2")
    sheet.add_image(XlsxImage(first), "A1")
    workbook.save(book)

    assets = extract_xlsx_images(book)
    assert list(assets) == [CellRef("Pictures", "A1"), CellRef("Pictures", "B2")]
    assert [(asset.width, asset.height, asset.media_type) for asset in assets.values()] == [
        (3, 2, "image/png"),
        (2, 4, "image/png"),
    ]
    assert all(asset.sha256 and asset.data.startswith(b"\x89PNG\r\n\x1a\n") for asset in assets.values())

    monkeypatch.setattr("mobiliti_saas.worker.catalog_sync.importers.common.MAX_IMAGE_BYTES", 1)
    assert_code("IMAGE_LIMIT", lambda: extract_xlsx_images(book))


def test_xlsx_same_media_thousands_of_anchors_is_normalized_once(tmp_path, monkeypatch):
    picture = tmp_path / "picture.png"
    source = tmp_path / "single-anchor.xlsx"
    repeated = tmp_path / "repeated-anchor.xlsx"
    Image.new("RGB", (2, 3), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.save(source)
    repeat_first_xlsx_image_anchor(source, repeated, 2_000)

    calls = []
    real_normalize = common_module._normalize_image

    def audited_normalize(raw):
        calls.append(raw)
        return real_normalize(raw)

    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)
    assets = extract_xlsx_images_from_bytes(repeated.read_bytes())

    assert len(assets) == 2_000
    assert len(calls) == 1
    assert len({id(asset) for asset in assets.values()}) == 1


def test_xlsx_identical_raw_media_parts_are_deduplicated_before_normalizing(
    tmp_path, monkeypatch
):
    picture = tmp_path / "picture.png"
    source = tmp_path / "duplicate-media.xlsx"
    Image.new("RGB", (2, 3), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.active.add_image(XlsxImage(picture), "B2")
    workbook.save(source)

    calls = []
    real_normalize = common_module._normalize_image

    def audited_normalize(raw):
        calls.append(raw)
        return real_normalize(raw)

    monkeypatch.setattr(common_module, "MAX_WORKBOOK_UNIQUE_IMAGES", 1, raising=False)
    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)
    assets = extract_xlsx_images_from_bytes(source.read_bytes())

    assert list(assets) == [CellRef("Sheet", "A1"), CellRef("Sheet", "B2")]
    assert len(calls) == 1
    assert assets[CellRef("Sheet", "A1")] is assets[CellRef("Sheet", "B2")]


def test_xlsx_rejects_excess_anchors_per_cell_before_normalizing(tmp_path, monkeypatch):
    picture = tmp_path / "picture.png"
    source = tmp_path / "single-anchor.xlsx"
    repeated = tmp_path / "same-cell.xlsx"
    Image.new("RGB", (2, 3), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.save(source)
    repeat_first_xlsx_image_anchor(source, repeated, 2, same_cell=True)

    monkeypatch.setattr(common_module, "MAX_WORKBOOK_IMAGES_PER_CELL", 1, raising=False)
    monkeypatch.setattr(
        common_module,
        "_normalize_image",
        lambda _raw: pytest.fail("image normalized before anchor preflight"),
    )

    assert_code("XLSX_LIMIT", lambda: extract_xlsx_images_from_bytes(repeated.read_bytes()))


def test_xlsx_rejects_excess_unique_raw_media_before_normalizing(tmp_path, monkeypatch):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    source = tmp_path / "unique-media.xlsx"
    Image.new("RGB", (2, 3), "red").save(first)
    Image.new("RGB", (2, 3), "blue").save(second)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(first), "A1")
    workbook.active.add_image(XlsxImage(second), "B2")
    workbook.save(source)

    monkeypatch.setattr(common_module, "MAX_WORKBOOK_UNIQUE_IMAGES", 1, raising=False)
    monkeypatch.setattr(
        common_module,
        "_normalize_image",
        lambda _raw: pytest.fail("image normalized before unique-media preflight"),
    )

    assert_code("XLSX_LIMIT", lambda: extract_xlsx_images_from_bytes(source.read_bytes()))


@pytest.mark.parametrize(
    "limit_name",
    ["MAX_WORKBOOK_IMAGE_PIXELS", "MAX_WORKBOOK_NORMALIZED_IMAGE_BYTES"],
)
def test_xlsx_rejects_excess_unique_normalized_image_totals(
    tmp_path, monkeypatch, limit_name
):
    picture = tmp_path / "picture.png"
    source = tmp_path / f"{limit_name}.xlsx"
    Image.new("RGB", (2, 3), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.save(source)

    monkeypatch.setattr(common_module, limit_name, 1, raising=False)

    assert_code("XLSX_LIMIT", lambda: extract_xlsx_images_from_bytes(source.read_bytes()))


def test_xlsx_pixel_budget_counts_distinct_raw_images_with_same_normalized_png(
    tmp_path, monkeypatch
):
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    source = tmp_path / "same-canonical-image.xlsx"
    picture = Image.new("RGB", (20, 20), "red")
    picture.save(first)
    metadata = PngImagePlugin.PngInfo()
    metadata.add_text("source", "distinct raw bytes")
    picture.save(second, pnginfo=metadata)
    assert first.read_bytes() != second.read_bytes()
    assert common_module._normalize_image(first.read_bytes()).data == common_module._normalize_image(
        second.read_bytes()
    ).data

    workbook = Workbook()
    workbook.active.add_image(XlsxImage(first), "A1")
    workbook.active.add_image(XlsxImage(second), "B2")
    workbook.save(source)

    calls = []
    real_normalize = common_module._normalize_image

    def audited_normalize(raw):
        calls.append(raw)
        return real_normalize(raw)

    monkeypatch.setattr(common_module, "MAX_WORKBOOK_IMAGE_PIXELS", 400, raising=False)
    monkeypatch.setattr(common_module, "_normalize_image", audited_normalize)

    assert_code("XLSX_LIMIT", lambda: extract_xlsx_images_from_bytes(source.read_bytes()))
    assert len(calls) == 1


def test_corrupt_or_unsupported_embedded_images_fail_safely(tmp_path):
    source = tmp_path / "source.xlsx"
    corrupt = tmp_path / "corrupt.xlsx"
    picture = tmp_path / "picture.png"
    Image.new("RGB", (2, 2), "red").save(picture)
    workbook = Workbook()
    workbook.active.add_image(XlsxImage(picture), "A1")
    workbook.save(source)

    def break_image(name, data):
        if name.startswith("xl/media/"):
            return name, b"<svg/>"
        return name, data

    rewrite_zip(source, corrupt, break_image)
    assert_code("IMAGE_UNSAFE", lambda: extract_xlsx_images(corrupt))


def test_source_refs_are_canonical_bounded_and_text_is_neutralized():
    assert source_ref("file-123", "Catalog", "A9") == {
        "file_id": "file-123",
        "sheet_or_page": "Catalog",
        "cell_or_bbox": "A9",
    }
    assert source_ref("file-123", 2, (1, 2.5, 3, 4)) == {
        "file_id": "file-123",
        "sheet_or_page": 2,
        "cell_or_bbox": [1, 2.5, 3, 4],
    }
    for invalid in ("../secret", "x" * 129, "https://secret.invalid"):
        assert_code("SOURCE_REF", lambda invalid=invalid: source_ref(invalid, "Sheet", "A1"))

    assert neutralize_spreadsheet_text(" ordinary Unicode á ") == " ordinary Unicode á "
    assert neutralize_spreadsheet_text("  =1+1") == "'  =1+1"
    assert neutralize_spreadsheet_text("\t@SUM(A1)") == "'\t@SUM(A1)"
    assert neutralize_spreadsheet_text(None) == ""
    assert len(neutralize_spreadsheet_text("x" * 100_000)) <= 32_768


def test_errors_are_redacted_and_source_has_no_cleanup_or_network_calls(tmp_path):
    secret = tmp_path / "SECRET-formula-url.pdf"
    secret.write_bytes(b"%PDF-SECRET https://secret.invalid")
    assert_code("PDF_INVALID", lambda: list(iter_pdf_pages(secret)))

    source = Path("mobiliti_saas/worker/catalog_sync/importers/common.py").read_text(encoding="utf-8")
    assert all(token not in source for token in ("unlink(", "remove(", "rmdir(", "urlopen(", "requests.", "httpx."))
    assert MAX_FILE_BYTES == 67_108_864
    assert MAX_ZIP_ENTRIES > 0
    assert MAX_PDF_PAGES > 0
    assert MAX_IMAGE_BYTES == 8 * 1024 * 1024


def test_public_records_are_frozen_and_contracts_are_typed():
    reference = CellRef("Sheet", "A1")
    with pytest.raises(FrozenInstanceError):
        reference.cell = "B2"

    assert get_type_hints(open_xlsx_data_only)["return"] is Workbook
    assert get_type_hints(open_xlsx_data_only_from_bytes)["return"] is Workbook
    assert get_type_hints(iter_pdf_pages)["return"] == Iterator[PdfPage]
    assert inspect.signature(source_ref).return_annotation is dict
