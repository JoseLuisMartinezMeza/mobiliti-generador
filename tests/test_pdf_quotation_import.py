from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import Workbook
from PIL import Image

try:
    import pymupdf as fitz
except ImportError:  # pragma: no cover
    import fitz  # type: ignore[no-redef]


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mobiliti_saas.quote_engine.images import extract_images  # noqa: E402
from mobiliti_saas.quote_engine.parser import read_items  # noqa: E402
from pdf_quotation_import import convert_pdf_to_quotation  # noqa: E402


def test_convert_pdf_to_engine_quotation_format(tmp_path: Path) -> None:
    source_pdf = tmp_path / "supplier.pdf"
    reference_xlsx = tmp_path / "reference.xlsx"
    output_xlsx = tmp_path / "quotation.xlsx"
    _write_reference_workbook(reference_xlsx)
    _write_supplier_pdf(source_pdf)

    result = convert_pdf_to_quotation(source_pdf, output_xlsx, reference_xlsx)

    assert result.product_count == 4
    assert result.image_count == 4
    assert result.categories == ("PROJECT UNO", "PROJECT DOS")

    items, columns = read_items(output_xlsx)
    products = [item for item in items if item.tipo == "producto"]
    categories = [item for item in items if item.tipo == "categoria"]

    assert columns["unit_price"] == "J"
    assert [category.nombre for category in categories] == ["PROJECT UNO", "PROJECT DOS"]
    assert len(products) == 4
    assert products[0].nombre == "AA1000\nROUND TABLE"
    assert products[0].dimension == "100*100*76CM"
    assert products[0].cantidad == 2
    assert products[0].precio == 10
    assert products[1].nombre == "BB2000 TAIL9\nCHAIR WITH ARMS"
    assert products[1].dimension == "50*60*70CM"
    assert products[2].categoria == "PROJECT DOS"

    image_map, _temp_dir = extract_images(output_xlsx)
    assert sorted(image_map) == [9, 10, 12, 13]


def _write_reference_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    for col, value in enumerate(
        [
            "No.",
            "Item Name",
            "Photo",
            "Description",
            "Dimension",
            "Color",
            "Q'ty",
            "Vol.",
            "Tot.Vol.",
            "Unit Price",
            " Tot.Price",
            "Remark",
        ],
        start=1,
    ):
        sheet.cell(row=7, column=col).value = value
    sheet.cell(row=8, column=1).value = "- SAMPLE"
    sheet.cell(row=9, column=1).value = 1
    workbook.save(path)
    workbook.close()


def _write_supplier_pdf(path: Path) -> None:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    _write_pdf_header(page)

    products = [
        (120, "AA1000", "PROJECT UNO", 2, "$10.00", "$20.00", ["ROUND TABLE 100*100*76CM", "FRAME:SAND TOP:WHITE"]),
        (220, "BB2000", "PROJECT UNO", 3, "$11.00", "$33.00", ["TAIL9|CHAIR WITH ARMS 50*60*70CM", "FINISH:BLACK"]),
        (320, "CC3000", "PROJECT DOS", 1, "$12.50", "$12.50", ["SIDE TABLE 45*45*55CM", "FRAME:OAK"]),
        (420, "DD4000", "PROJECT DOS", 4, "$3.00", "$12.00", ["STOOL 30*30*45CM", "ROPE:SAND"]),
    ]
    for y, code, project, qty, unit_price, total_price, details in products:
        _insert_product_start(page, y, code, project, qty, unit_price, total_price)
        _insert_image(page, y + 22)
        for offset, detail in enumerate(details, start=1):
            if "|" in detail:
                left, right = detail.split("|", 1)
                page.insert_text((22, y + 15 * offset), left, fontsize=9)
                page.insert_text((121, y + 15 * offset), right, fontsize=9)
            else:
                page.insert_text((121, y + 15 * offset), detail, fontsize=9)

    doc.save(path)
    doc.close()


def _write_pdf_header(page) -> None:
    page.insert_text((22, 30), "Proforma Invoice No:", fontsize=9)
    page.insert_text((121, 30), "AL-TEST-0001", fontsize=9)
    page.insert_text((302, 30), "Date:", fontsize=9)
    page.insert_text((398, 30), "10/Jun/26", fontsize=9)
    page.insert_text((22, 75), "Product No.", fontsize=9)
    page.insert_text((121, 75), "Product Name", fontsize=9)
    page.insert_text((331, 75), "Quantity", fontsize=9)
    page.insert_text((398, 75), "uom", fontsize=9)
    page.insert_text((436, 75), "Unit-Price", fontsize=9)
    page.insert_text((526, 75), "Sub-Total", fontsize=9)


def _insert_product_start(page, y: int, code: str, project: str, qty: int, unit_price: str, total_price: str) -> None:
    page.insert_text((22, y), code, fontsize=9)
    page.insert_text((121, y), project, fontsize=9)
    page.insert_text((347, y), str(qty), fontsize=9)
    page.insert_text((398, y), "PCS", fontsize=9)
    page.insert_text((446, y), unit_price, fontsize=9)
    page.insert_text((527, y), total_price, fontsize=9)


def _insert_image(page, y: int) -> None:
    image = Image.new("RGB", (40, 40), (40, 80, 120))
    buffer = BytesIO()
    image.save(buffer, "PNG")
    page.insert_image(fitz.Rect(22, y, 70, y + 48), stream=buffer.getvalue())
