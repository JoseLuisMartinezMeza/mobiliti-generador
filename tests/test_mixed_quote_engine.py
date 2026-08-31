from __future__ import annotations

from copy import deepcopy
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re

import pytest
from openpyxl import Workbook, load_workbook

from mobiliti_saas.quote_engine import generate_quote
from mobiliti_saas.quote_engine import engine
from mobiliti_saas.quote_engine.parser import QuoteItem, read_items


MIXED_HEADERS = {
    1: "No.",
    2: "Item",
    4: "Description",
    5: "Dimension",
    7: "Qty",
    10: "List Price",
    11: "URL",
    12: "Supplier",
    13: "Discount Percent",
    14: "Original Currency",
    15: "Original Unit Price",
    16: "Frozen Exchange Rate",
    17: "Source Reference",
    18: "Price Mode",
    19: "Auto Electrification",
}
ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = engine.OFFICIAL_TEMPLATE_PATH


def _converted_price(original, rate):
    return float(
        (Decimal(str(original)) * Decimal(str(rate))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    )


def _mixed_line(
    *,
    name: str,
    provider: str,
    discount: float,
    mode: str,
    auto: bool,
    price: float,
    original_currency: str,
    original_price: float,
    frozen_rate: float,
    quantity: float = 1,
):
    return {
        "name": name,
        "provider": provider,
        "discount": discount,
        "mode": mode,
        "auto": auto,
        "price": price,
        "original_currency": original_currency,
        "original_price": original_price,
        "frozen_rate": frozen_rate,
        "quantity": quantity,
    }


def _write_mixed_source(path: Path, lines: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    for column, value in MIXED_HEADERS.items():
        ws.cell(7, column).value = value
    ws["A8"] = "- Catalogos mixtos"
    for index, line in enumerate(lines, start=1):
        row = index + 8
        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = line["name"]
        ws.cell(row, 4).value = f"Descripcion {line['name']}"
        ws.cell(row, 5).value = "pieza"
        ws.cell(row, 7).value = line["quantity"]
        ws.cell(row, 10).value = line["price"]
        ws.cell(row, 12).value = line["provider"]
        ws.cell(row, 13).value = line["discount"]
        ws.cell(row, 14).value = line["original_currency"]
        ws.cell(row, 15).value = line["original_price"]
        ws.cell(row, 16).value = line["frozen_rate"]
        ws.cell(row, 17).value = f"source:{line['provider'].lower()}:{index}"
        ws.cell(row, 18).value = line["mode"]
        ws.cell(row, 19).value = line["auto"]
    wb.save(path)
    wb.close()
    return path


def _rate_summary(quote_currency: str, mxn_rate: str, usd_rate: str):
    def rate(catalog, base_currency, exchange_rate):
        identity = base_currency == quote_currency
        return {
            "catalog": catalog,
            "base_currency": base_currency,
            "quote_currency": quote_currency,
            "exchange_rate": exchange_rate,
            "rate_source": "identity" if identity else "saas_exchange_rates",
            "rate_effective_date": "2026-07-15",
            "rate_retrieved_at": "" if identity else "2026-07-15T23:00:00Z",
        }
    return [
        rate("tarkett", "MXN", mxn_rate),
        rate("alma", "USD", usd_rate),
    ]


def _mixed_metadata(
    quote_currency: str = "MXN",
    *,
    mxn_rate: str = "1.000000",
    usd_rate: str = "18.500000",
    auto_rate: bool = False,
):
    summary = _rate_summary(quote_currency, mxn_rate, usd_rate)
    return {
        "catalog_price_mode": "mixed_catalog_converted",
        "quote_currency": quote_currency,
        "rate_summary": summary,
        "auto_electrification_rate": (
            {key: summary[0][key] for key in (
                "base_currency",
                "quote_currency",
                "exchange_rate",
                "rate_source",
                "rate_effective_date",
                "rate_retrieved_at",
            )}
            if auto_rate
            else None
        ),
        "cotizacion": "MIXTA-001",
        "proyecto": "Proyecto mixto",
        "cliente": "Cliente",
    }


def _resolved_simple_reference(workbook, value):
    for _ in range(8):
        match = re.fullmatch(
            r"='?([^']+)'?!([A-Z]+[1-9][0-9]*)",
            value,
        ) if isinstance(value, str) else None
        if match is None:
            return value
        value = workbook[match.group(1)][match.group(2)].value
    raise AssertionError(f"Cadena de referencias demasiado profunda: {value}")


def _row_for_formula(ws, column: int, formula: str) -> int:
    direct = [
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, column).value == formula
    ]
    if direct:
        return direct[0]
    match = re.fullmatch(r"=Quotation!([A-Z]+[1-9][0-9]*)", formula)
    if match is None:
        raise StopIteration(formula)
    expected = ws.parent["Quotation"][match.group(1)].value
    return next(
        row
        for row in range(1, ws.max_row + 1)
        if _resolved_simple_reference(
            ws.parent,
            ws.cell(row, column).value,
        ) == expected
    )


@pytest.fixture(autouse=True)
def _mixed_tests_never_resolve_a_live_legacy_rate(monkeypatch):
    def fail_if_called(_metadata):
        raise AssertionError("mixed mode must not resolve the legacy exchange rate")

    monkeypatch.setattr(engine, "_exchange_rate", fail_if_called)


def test_mobiliti_product_code_and_cost_reference_the_same_quotation_row(tmp_path):
    source = _write_mixed_source(
        tmp_path / "linked-product-codes.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=False,
                price=123.46,
                original_currency="MXN",
                original_price=123.456,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Silla ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=1850,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            ),
            _mixed_line(
                name="Silla importada",
                provider="Proveedor importado",
                discount=40,
                mode="imported",
                auto=False,
                price=1517,
                original_currency="USD",
                original_price=82,
                frozen_rate=18.5,
            ),
        ],
    )
    output = tmp_path / "linked-product-codes-final.xlsx"

    generate_quote(source, output, _mixed_metadata(), TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        mobiliti = wb["Mobiliti"]
        expected_rows = {
            "Tarkett MX": 9,
            "Alma - Exterior": 10,
            "Proveedor Externo": 11,
        }
        for provider, quotation_row in expected_rows.items():
            target_row = next(
                row
                for row in range(1, mobiliti.max_row + 1)
                if mobiliti.cell(row, 6).value == provider
            )
            code_cell = mobiliti.cell(target_row, 4)
            cost_cell = mobiliti.cell(target_row, 10)
            assert code_cell.data_type == "f"
            assert code_cell.value == f"=Quotation!B{quotation_row}"
            assert cost_cell.data_type == "f"
            assert cost_cell.value == f"=Quotation!K{quotation_row}"
    finally:
        wb.close()


def _write_quotation(path: Path, *, mixed: bool) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    headers = MIXED_HEADERS if mixed else {key: MIXED_HEADERS[key] for key in (1, 2, 4, 5, 7, 10, 11)}
    for column, value in headers.items():
        ws.cell(7, column).value = value
    ws["A8"] = "- ALMA"
    ws["A9"] = 1
    ws["B9"] = "Producto ALMA"
    ws["D9"] = "Descripcion"
    ws["E9"] = "pieza"
    ws["G9"] = 2
    ws["J9"] = 1850
    if mixed:
        ws["L9"] = "ALMA"
        ws["M9"] = 0
        ws["N9"] = "USD"
        ws["O9"] = 100
        ws["P9"] = 18.5
        ws["Q9"] = "source:alma:1"
        ws["R9"] = "net"
        ws["S9"] = False
    wb.save(path)
    wb.close()
    return path


def test_parser_reads_mixed_audit_columns_by_header(tmp_path):
    source = _write_quotation(tmp_path / "mixed.xlsx", mixed=True)

    items, columns = read_items(source)

    product = next(item for item in items if item.tipo == "producto")
    assert columns["proveedor"] == "L"
    assert columns["descuento"] == "M"
    assert columns["electrificacion_automatica"] == "S"
    assert product.proveedor == "ALMA"
    assert product.descuento == 0
    assert product.moneda_original == "USD"
    assert product.precio_original == 100
    assert product.tipo_cambio_congelado == 18.5
    assert product.referencia_fuente == "source:alma:1"
    assert product.modo_precio == "net"
    assert product.electrificacion_automatica is False


def test_parser_keeps_legacy_defaults_when_mixed_headers_are_absent(tmp_path):
    source = _write_quotation(tmp_path / "legacy.xlsx", mixed=False)

    product = next(item for item in read_items(source)[0] if item.tipo == "producto")

    assert product.proveedor == ""
    assert product.descuento is None
    assert product.moneda_original == ""
    assert product.precio_original is None
    assert product.tipo_cambio_congelado is None
    assert product.referencia_fuente == ""
    assert product.modo_precio == ""
    assert product.electrificacion_automatica is None


def test_mixed_engine_uses_manual_section_concepts_without_double_numbering(tmp_path):
    source = _write_mixed_source(
        tmp_path / "manual-sections.xlsx",
        [
            _mixed_line(
                name="Silla recepción",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=False,
                price=100,
                original_currency="MXN",
                original_price=100,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Mesa privada",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=1850,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            ),
        ],
    )
    source_wb = load_workbook(source)
    source_ws = source_wb["Quotation"]
    source_ws["A8"] = "- Recepción"
    source_ws.insert_rows(10, 1)
    source_ws["A10"] = "- Privados"
    source_wb.save(source)
    source_wb.close()

    output = tmp_path / "manual-sections-final.xlsx"
    generate_quote(source, output, _mixed_metadata(), TEMPLATE)

    wb = load_workbook(output, data_only=False)
    mobiliti_titles = {
        cell.value
        for row in wb["Mobiliti"].iter_rows()
        for cell in row
        if cell.value in {"Recepción", "Privados"}
    }
    assert "Recepción" in mobiliti_titles
    assert "Privados" in mobiliti_titles
    assert not any(
        "- 1-Recepción" in title or "- 2-Privados" in title
        for title in mobiliti_titles
    )
    cotizacion_categories = {
        cell.value
        for cell in wb["Cotizacion"]["A"]
        if cell.value in {"Recepción", "Privados"}
    }
    assert cotizacion_categories == {"Recepción", "Privados"}
    wb.close()


@pytest.mark.parametrize(
    ("currency", "mxn_rate", "usd_rate", "money_literal"),
    [
        ("MXN", "1.000000", "18.500000", "$"),
        ("USD", "0.054054", "1.000000", "$"),
        ("EUR", "0.048780", "0.902430", "$"),
    ],
)
def test_mixed_engine_converts_once_and_references_one_general_discount(
    tmp_path, currency, mxn_rate, usd_rate, money_literal
):
    source = _write_mixed_source(
        tmp_path / f"mixed-{currency}.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=False,
                price=_converted_price(123.456, mxn_rate),
                original_currency="MXN",
                original_price=123.456,
                frozen_rate=float(mxn_rate),
                quantity=2,
            ),
            _mixed_line(
                name="Silla ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=_converted_price(100, usd_rate),
                original_currency="USD",
                original_price=100,
                frozen_rate=float(usd_rate),
                quantity=3,
            ),
            _mixed_line(
                name="Silla importada",
                provider="Proveedor importado",
                discount=40,
                mode="imported",
                auto=False,
                price=_converted_price(82, usd_rate),
                original_currency="USD",
                original_price=82,
                frozen_rate=float(usd_rate),
                quantity=2,
            ),
        ],
    )
    output = tmp_path / f"final-{currency}.xlsx"

    generate_quote(
        source,
        output,
        _mixed_metadata(currency, mxn_rate=mxn_rate, usd_rate=usd_rate),
        TEMPLATE,
    )

    wb = load_workbook(output, data_only=False)
    try:
        cot = wb["Cotizacion"]
        mobiliti = wb["Mobiliti"]
        tarkett_mob = _row_for_formula(mobiliti, 4, "=Quotation!B9")
        alma_mob = _row_for_formula(mobiliti, 4, "=Quotation!B10")
        imported_mob = _row_for_formula(mobiliti, 4, "=Quotation!B11")
        tarkett_cot = _row_for_formula(cot, 1, "=Quotation!B9")
        alma_cot = _row_for_formula(cot, 1, "=Quotation!B10")
        imported_cot = _row_for_formula(cot, 1, "=Quotation!B11")
        assert mobiliti.cell(tarkett_mob, 6).value == "Tarkett MX"
        assert mobiliti.cell(alma_mob, 6).value == "Alma - Exterior"
        assert mobiliti.cell(imported_mob, 6).value == "Proveedor Externo"
        for mobiliti_row, source_row, expected_price in (
            (tarkett_mob, 9, _converted_price(123.456, mxn_rate)),
            (alma_mob, 10, _converted_price(100, usd_rate)),
            (imported_mob, 11, _converted_price(82, usd_rate)),
        ):
            assert mobiliti.cell(mobiliti_row, 10).value == (
                f"=Quotation!K{source_row}"
            )
            assert wb["Quotation"].cell(source_row, 11).value == expected_price
        assert mobiliti["AD13"].value == 0.4
        for row in (tarkett_cot, alma_cot, imported_cot):
            assert cot.cell(row, 7).value == "=ROUND(Mobiliti!$AD$13,2)"
        for row in (tarkett_mob, alma_mob, imported_mob):
            assert mobiliti[f"AD{row}"].value == f"=IF(H{row}>0,$E$5,0)"
            assert mobiliti[f"AE{row}"].value == f"=IFERROR(AA{row}*AD{row},0)"
        assert cot.cell(tarkett_cot, 6).value == f"=Mobiliti!AA{tarkett_mob}"
        assert cot.cell(alma_cot, 6).value == f"=Mobiliti!AA{alma_mob}"
        assert cot.cell(imported_cot, 6).value == f"=Mobiliti!AA{imported_mob}"
        assert mobiliti["P4"].value is (currency != "MXN")
        for row in (tarkett_cot, alma_cot, imported_cot):
            assert cot.cell(row, 8).value == f"=F{row}*G{row}"
            assert cot.cell(row, 9).value == f"=F{row}-H{row}"
            assert cot.cell(row, 10).value == f"=E{row}*I{row}"
            for column in (6, 8, 9, 10):
                assert money_literal in cot.cell(row, column).number_format
        for row in (tarkett_mob, alma_mob, imported_mob):
            # Columnas monetarias de la plantilla v17; no las posiciones legacy.
            for column in ("J", "M", "O", "R", "T", "U", "W", "X", "Y",
                           "Z", "AA", "AB", "AE", "AF", "AG", "AI", "AJ"):
                assert money_literal in mobiliti[f"{column}{row}"].number_format
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("mode", "unknown", "Modo de precio mixto invalido"),
        ("discount", 40, "Precio neto mixto no admite descuento"),
        ("provider", "ALMA", "Precio de lista mixto solo admite Tarkett u Offiho"),
        ("auto", True, "Electrificacion automatica mixta solo admite Tarkett u Offiho"),
    ],
)
def test_mixed_engine_rejects_invalid_per_line_policy_before_saving(
    tmp_path, field, value, message
):
    line = _mixed_line(
        name="Silla ALMA",
        provider="ALMA",
        discount=0,
        mode="net",
        auto=False,
        price=100,
        original_currency="USD",
        original_price=100,
        frozen_rate=18.5,
    )
    line[field] = value
    if field == "provider":
        line["mode"] = "list"
    source = _write_mixed_source(tmp_path / f"invalid-{field}.xlsx", [line])
    output = tmp_path / f"invalid-{field}-final.xlsx"

    with pytest.raises(ValueError, match=message):
        generate_quote(source, output, _mixed_metadata(), TEMPLATE)

    assert not output.exists()


@pytest.mark.parametrize(
    ("mode", "discount", "expected"),
    (
        ("net", 0, Decimal("0")),
        ("imported", 40, Decimal("0.4")),
    ),
)
def test_mixed_item_discount_policy_accepts_net_zero_and_imported_discount(
    mode, discount, expected
):
    item = QuoteItem(
        tipo="producto",
        row=9,
        proveedor="ALMA" if mode == "net" else "Proveedor importado",
        descuento=discount,
        modo_precio=mode,
    )

    assert engine._mixed_item_discount_fraction(item) == expected


def test_mixed_item_discount_policy_rejects_net_nonzero():
    item = QuoteItem(
        tipo="producto",
        row=9,
        proveedor="ALMA",
        descuento=40,
        modo_precio="net",
    )

    with pytest.raises(ValueError, match="Precio neto mixto no admite descuento"):
        engine._mixed_item_discount_fraction(item)


def test_mixed_metadata_rejects_empty_effective_items_with_empty_rate_summary():
    metadata = _mixed_metadata()
    metadata["rate_summary"] = []

    with pytest.raises(ValueError, match="productos"):
        engine._validate_mixed_catalog_metadata([], metadata)


def test_mixed_metadata_rejects_empty_rate_summary_for_catalog_item():
    metadata = _mixed_metadata()
    metadata["rate_summary"] = []
    item = QuoteItem(
        tipo="producto",
        row=9,
        proveedor="ALMA",
        descuento=0,
        moneda_original="USD",
        precio_original=100,
        precio=1850,
        tipo_cambio_congelado=18.5,
        referencia_fuente="source:alma:1",
        modo_precio="net",
        electrificacion_automatica=False,
    )

    with pytest.raises(ValueError, match="Proveedor mixto sin tasa congelada"):
        engine._validate_mixed_catalog_metadata([item], metadata)


@pytest.mark.parametrize(
    ("mode", "provider", "discount"),
    (
        ("net", "ALMA", 0),
        ("imported", "Proveedor importado", 40),
    ),
)
def test_mixed_metadata_rejects_surplus_rate_summary(mode, provider, discount):
    item = QuoteItem(
        tipo="producto",
        row=9,
        proveedor=provider,
        descuento=discount,
        moneda_original="USD",
        precio_original=100,
        precio=1850,
        tipo_cambio_congelado=18.5,
        referencia_fuente=f"source:{mode}:1",
        modo_precio=mode,
        electrificacion_automatica=False,
    )

    with pytest.raises(ValueError, match="Resumen de tasas mixtas inconsistente"):
        engine._validate_mixed_catalog_metadata([item], _mixed_metadata())


@pytest.mark.parametrize(
    ("currency", "mxn_rate", "usd_rate"),
    [
        ("MXN", "1.000000", "18.500000"),
        ("USD", "0.054054", "1.000000"),
        ("EUR", "0.048780", "0.902430"),
    ],
)
def test_mixed_legacy_flags_add_no_accessories_and_keep_explicit_frozen_cost(
    tmp_path, currency, mxn_rate, usd_rate
):
    source = _write_mixed_source(
        tmp_path / f"accessories-{currency}.xlsx",
        [
            _mixed_line(
                name="Estacion Lido 8PAX",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=True,
                price=_converted_price(1000, mxn_rate),
                original_currency="MXN",
                original_price=1000,
                frozen_rate=float(mxn_rate),
                quantity=1,
            ),
            _mixed_line(
                name="Estacion Lido 8PAX ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=_converted_price(100, usd_rate),
                original_currency="USD",
                original_price=100,
                frozen_rate=float(usd_rate),
                quantity=1,
            ),
            _mixed_line(
                name="LIDO.OP-INT manual",
                provider="Lumbro",
                discount=0,
                mode="net",
                auto=False,
                price=_converted_price(120, mxn_rate),
                original_currency="MXN",
                original_price=120,
                frozen_rate=float(mxn_rate),
                quantity=1,
            ),
        ],
    )
    metadata = _mixed_metadata(
        currency,
        mxn_rate=mxn_rate,
        usd_rate=usd_rate,
        auto_rate=True,
    )
    metadata["rate_summary"].append(
        {
            "catalog": "lumbro",
            "base_currency": "MXN",
            "quote_currency": currency,
            "exchange_rate": mxn_rate,
            "rate_source": "identity" if currency == "MXN" else "saas_exchange_rates",
            "rate_effective_date": "2026-07-15",
            "rate_retrieved_at": "" if currency == "MXN" else "2026-07-15T23:00:00Z",
        }
    )
    output = tmp_path / f"accessories-final-{currency}.xlsx"

    generate_quote(source, output, metadata, TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        cot = wb["Cotizacion"]
        mobiliti = wb["Mobiliti"]
        parent_mob = _row_for_formula(mobiliti, 4, "=Quotation!B9")
        alma_source_row = next(
            row
            for row in range(9, wb["Quotation"].max_row + 1)
            if wb["Quotation"].cell(row, 2).value == "Estacion Lido 8PAX ALMA"
        )
        manual_source_row = next(
            row
            for row in range(9, wb["Quotation"].max_row + 1)
            if wb["Quotation"].cell(row, 2).value == "LIDO.OP-INT manual"
        )
        alma_mob = _row_for_formula(
            mobiliti,
            4,
            f"=Quotation!B{alma_source_row}",
        )
        manual_mob = _row_for_formula(
            mobiliti,
            4,
            f"=Quotation!B{manual_source_row}",
        )
        automatic_rows = [
            row
            for row in range(1, mobiliti.max_row + 1)
            if _resolved_simple_reference(
                wb,
                mobiliti.cell(row, 4).value,
            ) in {"LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"}
        ]
        assert automatic_rows == []
        assert parent_mob < alma_mob < manual_mob
        assert wb["Quotation_Data"].max_row == 4
        assert not any(
            ":lumbro:" in str(wb["Quotation_Data"].cell(row, 1).value)
            for row in range(2, 5)
        )
        parent_cot = _row_for_formula(cot, 1, "=Quotation!B9")
        assert cot.cell(parent_cot + 1, 1).value == f"=Mobiliti!D{alma_mob}"
        assert cot.cell(parent_cot + 2, 1).value == f"=Mobiliti!D{manual_mob}"
        assert mobiliti.cell(manual_mob, 6).value == "Lumbro CH"
        assert mobiliti.cell(manual_mob, 10).value == f"=Quotation!K{manual_source_row}"
        assert wb["Quotation"].cell(manual_source_row, 11).value == _converted_price(
            120, mxn_rate,
        )
        assert wb["Quotation"].cell(manual_source_row, 8).value == 1
    finally:
        wb.close()


def test_mixed_job_without_eligible_auto_lines_needs_no_auto_rate(tmp_path):
    source = _write_mixed_source(
        tmp_path / "alma-only.xlsx",
        [
            _mixed_line(
                name="Estacion Lido 8PAX ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=1850,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            )
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = metadata["rate_summary"][1:]
    output = tmp_path / "alma-only-final.xlsx"

    generate_quote(source, output, metadata, TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        assert not any(
            _resolved_simple_reference(
                wb,
                wb["Mobiliti"].cell(row, 4).value,
            ) in {"LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"}
            for row in range(1, wb["Mobiliti"].max_row + 1)
        )
    finally:
        wb.close()


def test_mixed_headers_are_saved_as_safe_text(tmp_path):
    source = _write_mixed_source(
        tmp_path / "safe-header.xlsx",
        [
            _mixed_line(
                name="Silla ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=1850,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            )
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = metadata["rate_summary"][1:]
    fields = ("cotizacion", "proyecto", "cliente", "correo", "telefono", "direccion", "razon_social")
    dangerous_values = ("=1+1", "+SUM(A1:A2)", "-2+3", "@cmd", "=A1", "+1", "-9")
    metadata.update(dict(zip(fields, dangerous_values, strict=True)))
    output = tmp_path / "safe-headers.xlsx"

    generate_quote(source, output, metadata, TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        cot = wb["Cotizacion"]
        for address, dangerous in zip(
            ("B3", "B7", "B8", "B9", "B10", "B11", "B12"),
            dangerous_values,
            strict=True,
        ):
            assert cot[address].value == "'" + dangerous
            assert cot[address].data_type != "f"
    finally:
        wb.close()


def test_mixed_rate_legend_is_canonical_safe_and_compact(tmp_path):
    source = _write_mixed_source(
        tmp_path / "legend.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=False,
                price=100,
                original_currency="MXN",
                original_price=100,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Silla ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=1850,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            ),
        ],
    )
    output = tmp_path / "legend-final.xlsx"

    metadata = _mixed_metadata()
    expected_legend = (
        "MXN | precios mixtos mas IVA | Tarkett MXN/MXN 1.000000; "
        "ALMA USD/MXN 18.500000 Banco de Mexico / DOF 2026-07-15"
    )
    assert engine._mixed_rate_legend(metadata) == expected_legend

    generate_quote(source, output, metadata, TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb["Cotizacion"]["A4"].value == "=TODAY()"
        assert wb["Cotizacion"]["B4"].value is None
        assert len(expected_legend) <= 1000
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("rate_effective_date", "2026-99-99"),
        ("rate_retrieved_at", "not-a-timestamp"),
    ],
)
def test_mixed_rate_summary_rejects_malformed_dates(field, value):
    metadata = deepcopy(_mixed_metadata())
    metadata["rate_summary"][1][field] = value

    with pytest.raises(ValueError, match="Resumen de tasas mixtas invalido"):
        engine._mixed_rate_summary(metadata)


def test_mixed_auto_line_requires_the_complete_frozen_rate_before_saving(tmp_path):
    source = _write_mixed_source(
        tmp_path / "missing-auto-rate.xlsx",
        [
            _mixed_line(
                name="Estacion Lido 8PAX",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=True,
                price=1000,
                original_currency="MXN",
                original_price=1000,
                frozen_rate=1,
            )
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = metadata["rate_summary"][:1]
    output = tmp_path / "missing-auto-rate-final.xlsx"

    with pytest.raises(ValueError, match="Tasa de electrificacion mixta incompleta"):
        generate_quote(source, output, metadata, TEMPLATE)

    assert not output.exists()


def test_mixed_auto_lines_require_one_identical_snapshot_for_every_eligible_provider(
    tmp_path,
):
    source = _write_mixed_source(
        tmp_path / "two-auto-providers.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=True,
                price=100,
                original_currency="MXN",
                original_price=100,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Escritorio Offiho",
                provider="Offiho",
                discount=40,
                mode="list",
                auto=True,
                price=200,
                original_currency="MXN",
                original_price=200,
                frozen_rate=1,
            ),
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = [
        metadata["rate_summary"][0],
        {
            **metadata["rate_summary"][0],
            "catalog": "offiho",
            "rate_effective_date": "2026-07-14",
        },
    ]
    metadata["auto_electrification_rate"] = {
        key: metadata["rate_summary"][0][key]
        for key in (
            "base_currency",
            "quote_currency",
            "exchange_rate",
            "rate_source",
            "rate_effective_date",
            "rate_retrieved_at",
        )
    }
    output = tmp_path / "missing-parent" / "final.xlsx"
    output.parent.mkdir()

    with pytest.raises(ValueError, match="Tasa de electrificacion mixta inconsistente"):
        generate_quote(source, output, metadata, TEMPLATE)

    assert output.parent.is_dir()
    assert not output.exists()
    assert not output.exists()


def test_mixed_auto_lines_accept_one_matching_snapshot_for_all_eligible_providers(tmp_path):
    source = _write_mixed_source(
        tmp_path / "matching-auto-providers.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett",
                provider="Tarkett",
                discount=40,
                mode="list",
                auto=True,
                price=100,
                original_currency="MXN",
                original_price=100,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Escritorio Offiho",
                provider="Offiho",
                discount=40,
                mode="list",
                auto=True,
                price=200,
                original_currency="MXN",
                original_price=200,
                frozen_rate=1,
            ),
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = [
        metadata["rate_summary"][0],
        {**metadata["rate_summary"][0], "catalog": "offiho"},
    ]
    metadata["auto_electrification_rate"] = {
        key: metadata["rate_summary"][0][key]
        for key in (
            "base_currency",
            "quote_currency",
            "exchange_rate",
            "rate_source",
            "rate_effective_date",
            "rate_retrieved_at",
        )
    }

    items = read_items(source)[0]

    engine._validate_mixed_catalog_metadata(items, metadata)


def test_mixed_converted_price_audit_rejects_arbitrary_final_price_before_paths(
    tmp_path,
):
    source = _write_mixed_source(
        tmp_path / "tampered-converted-price.xlsx",
        [
            _mixed_line(
                name="Silla ALMA",
                provider="ALMA",
                discount=0,
                mode="net",
                auto=False,
                price=999999,
                original_currency="USD",
                original_price=100,
                frozen_rate=18.5,
            )
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = metadata["rate_summary"][1:]
    output = tmp_path / "no-output-dir" / "final.xlsx"
    output.parent.mkdir()

    with pytest.raises(ValueError, match="Precio convertido mixto inconsistente"):
        generate_quote(source, output, metadata, TEMPLATE)

    assert output.parent.is_dir()
    assert not output.exists()
    assert not output.exists()


def test_mixed_discount_precision_and_half_up_price_boundaries_reach_both_sheets(tmp_path):
    source = _write_mixed_source(
        tmp_path / "precision.xlsx",
        [
            _mixed_line(
                name="Piso Tarkett precision alta",
                provider="Tarkett",
                discount=12.345678,
                mode="list",
                auto=False,
                price=2.68,
                original_currency="MXN",
                original_price=2.675,
                frozen_rate=1,
            ),
            _mixed_line(
                name="Piso Tarkett precision minima",
                provider="Tarkett",
                discount=12.345678,
                mode="list",
                auto=False,
                price=0.01,
                original_currency="MXN",
                original_price=0.005,
                frozen_rate=1,
            ),
        ],
    )
    metadata = _mixed_metadata()
    metadata["rate_summary"] = metadata["rate_summary"][:1]
    metadata["descuento"] = 12.345678
    output = tmp_path / "precision-final.xlsx"

    generate_quote(source, output, metadata, TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        cot = wb["Cotizacion"]
        mobiliti = wb["Mobiliti"]
        assert mobiliti["AD13"].value == pytest.approx(0.12345678)
        for source_row in (9, 10):
            cot_row = _row_for_formula(cot, 1, f"=Quotation!B{source_row}")
            mobiliti_row = _row_for_formula(mobiliti, 4, f"=Quotation!B{source_row}")
            # El usuario confirmó conservar ROUND de la plantilla oficial.
            assert cot.cell(cot_row, 7).value == "=ROUND(Mobiliti!$AD$13,2)"
            assert mobiliti.cell(mobiliti_row, 10).value == (
                f"=Quotation!K{source_row}"
            )
            assert wb["Quotation"].cell(source_row, 11).value == (
                2.68 if source_row == 9 else 0.01
            )
            assert mobiliti[f"AE{mobiliti_row}"].value == (
                f"=IFERROR(AA{mobiliti_row}*AD{mobiliti_row},0)"
            )
            assert mobiliti[f"Z{mobiliti_row}"].value.startswith(
                f'=ROUNDUP(IF(OR(F{mobiliti_row}="Offiho",'
            )
            assert mobiliti[f"AA{mobiliti_row}"].value == (
                f"=IF(Z{mobiliti_row}>=Y{mobiliti_row},"
                f"_xlfn.MINIFS($Z$14:$Z$571,$D$14:$D$571,D{mobiliti_row},"
                f"$H$14:$H$571,_xlfn.MAXIFS($H$14:$H$571,"
                f"$D$14:$D$571,D{mobiliti_row})),"
                '"NO SE ESTA RESPETANDO EL MARGEN")'
            )
            assert mobiliti[f"AB{mobiliti_row}"].value == (
                f"=IFERROR(AA{mobiliti_row}*H{mobiliti_row},0)"
            )
    finally:
        wb.close()
