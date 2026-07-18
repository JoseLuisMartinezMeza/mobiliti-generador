import hashlib
import json
from dataclasses import dataclass
from pathlib import Path

import fitz
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image

from mobiliti_saas.quote_engine.supplier_catalog import (
    PUBLIC_ITEM_FIELDS,
    load_supplier_catalog_data,
)
from mobiliti_saas.worker.catalog_sync.importers.common import (
    CellRef,
    SourceSafetyError,
    extract_xlsx_images,
)
from mobiliti_saas.worker.catalog_sync.importers.cr_global import (
    build_cr_global_snapshot,
    build_cr_global_snapshot_with_assets,
)


@dataclass(frozen=True)
class AdapterFile:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path | None


def _sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _png(path, color):
    Image.new("RGB", (24, 32), color).save(path)


def _write_workbook(path, tmp_path):
    workbook = Workbook()
    workbook.active.title = "README"
    sheet = workbook.create_sheet("SPEC-GUIDE-LUMBRO")
    sheet["A6"] = "Linea: CR Global"
    sheet["C6"] = "Sistema: Seating"
    sheet["A7"] = "Famila: Task Chairs"
    for cell, value in zip(
        ("A8", "B8", "C8", "D8", "E8", "F8"),
        ("Cod.", "Imagen.", "Descripcion.", "Medida/Unidad.", "P. Unitario.", "Moneda"),
    ):
        sheet[cell] = value

    rows = (
        (9, None, "ALPHA BLACK", None),
        (10, "CR-100-BLK", "Alpha task chair", "60 x 60 cm"),
        (11, None, "Black polypropylene shell", None),
        (19, None, "ALPHA WHITE", None),
        (20, "CR-100-WHT", "Alpha task chair", "60 x 60 cm"),
        (21, None, "White polypropylene shell", None),
        (29, None, "BETA PANEL", None),
        (30, "CR-200", None, "120 x 40 cm"),
        (39, None, "UNCODED STOOL", None),
        (40, None, "Stool pending supplier code", "45 x 45 cm"),
    )
    for row, code, description, dimensions in rows:
        sheet.cell(row, 1, code)
        sheet.cell(row, 3, description)
        sheet.cell(row, 4, dimensions)

    for row, color in ((10, "black"), (20, "white"), (30, "blue"), (40, "gray")):
        image_path = tmp_path / f"image-{row}.png"
        _png(image_path, color)
        sheet.add_image(WorksheetImage(image_path), f"B{row}")

    workbook.save(path)
    workbook.close()


def _write_pdf(path, pages):
    document = fitz.open()
    for text in pages:
        page = document.new_page()
        page.insert_textbox(fitz.Rect(36, 36, 560, 800), text, fontsize=9)
    document.save(path)
    document.close()


def _price_page(codes, prices, *, currency="MXN", notice=True):
    heading = [
        "LISTA DE PRECIOS DISTRIBUIDORES ABRIL 2026",
        "Precios mas IVA",
        "Tiempos de entrega: 2-3 dias habiles",
        f"Moneda: {currency}",
        "CLAVE",
        "DESCRIPCION",
        "PRECIO",
    ]
    body = list(codes)
    for index, price in enumerate(prices, 1):
        body.extend((f"Producto {index}", *(("$", price) if notice else (price,))))
    if not notice:
        heading = [line for line in heading if not line.startswith("Moneda:")]
    return "\n".join(heading + body)


@pytest.fixture
def source_bundle(tmp_path):
    workbook = tmp_path / "Spec guide-CR Global-2026.xlsx"
    price_list = tmp_path / "CRG_LP_General_Dist_2026-04.pdf"
    catalog = tmp_path / "CRG_FT_General_Dist_2026-04.pdf"
    _write_workbook(workbook, tmp_path)
    _write_pdf(
        price_list,
        [
            _price_page(
                ("CR-100-BLK", "CR-100-WHT", "CR-200"),
                ("1,234.50", "1,345.60", "2,000.00"),
            )
        ],
    )
    _write_pdf(
        catalog,
        [
            "FICHA TECNICA\nModelo: CR-100-BLK\nWrong replacement description",
            "FICHA TECNICA\nAcoustic privacy panel\nModelo: CR-200",
        ],
    )
    return (
        AdapterFile(
            "SPEC GUIDES 2026/CR Global/Spec guide-CR Global-2026.xlsx",
            "spec_guide",
            None,
            _sha256(workbook),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            workbook,
        ),
        AdapterFile(
            "CR GLOBAL/CRG_LP_General_Dist_2026-04.pdf",
            "price_list",
            None,
            _sha256(price_list),
            "application/pdf",
            price_list,
        ),
        AdapterFile(
            "CR GLOBAL/CRG_FT_General_Dist_2026-04.pdf",
            "catalog",
            None,
            _sha256(catalog),
            "application/pdf",
            catalog,
        ),
    )


def _without_generated_at(snapshot):
    return {key: value for key, value in snapshot.items() if key != "generated_at"}


def test_build_snapshot_detects_content_sheet_and_passes_public_contract(source_bundle):
    snapshot = build_cr_global_snapshot(source_bundle)
    loaded = load_supplier_catalog_data(snapshot, expected_supplier="cr-global")

    assert set(snapshot) == {"supplier", "source_hash", "generated_at", "items"}
    assert len(loaded["items"]) == 4
    assert all(set(item) == set(PUBLIC_ITEM_FIELDS) for item in snapshot["items"])
    assert all(item["supplier"] == "cr-global" for item in snapshot["items"])


def test_spec_guide_owns_code_configuration_dimensions_and_image_identity(source_bundle):
    snapshot = build_cr_global_snapshot(source_bundle)
    black = next(item for item in snapshot["items"] if item["sku"] == "CR-100-BLK")
    images = extract_xlsx_images(source_bundle[0].local_path)
    expected = images[CellRef("SPEC-GUIDE-LUMBRO", "B10")]

    assert black["name"] == "ALPHA BLACK"
    assert black["collection"] == "Task Chairs"
    assert black["description"] == "Alpha task chair Black polypropylene shell"
    assert black["attributes"]["configuration"] == "ALPHA BLACK"
    assert black["attributes"]["dimensions"] == "60 x 60 cm"
    assert black["attributes"]["image_sha256"] == expected.sha256
    assert black["attributes"]["image_width"] == expected.width
    assert black["attributes"]["image_height"] == expected.height
    assert black["image_url"] == ""
    assert black["image_kind"] == "placeholder"


def test_asset_build_publishes_only_exact_row_anchored_xlsx_images(source_bundle):
    build = build_cr_global_snapshot_with_assets(source_bundle)

    assert len(build.bindings) == 4
    assert len(build.assets_by_sha256) == 4
    for binding in build.bindings:
        item = next(row for row in build.snapshot["items"] if row["internal_id"] == binding.internal_id)
        asset = build.assets_by_sha256[binding.asset_sha256]
        assert binding.match_status == "exact_xlsx"
        assert binding.source_references[0]["sheet_or_page"] == "SPEC-GUIDE-LUMBRO"
        assert binding.source_references[0]["cell_or_bbox"].startswith("B")
        assert asset.sha256 == binding.asset_sha256
        assert item["image_kind"] == "official"
        assert item["image_url"] == ""
        assert item["attributes"]["approved_asset"]["path"] == binding.object_name
        assert item["attributes"]["image_match"]["status"] == "exact_xlsx"


def test_explicit_spec_price_and_currency_override_pdf_price(source_bundle):
    workbook_path = source_bundle[0].local_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["SPEC-GUIDE-LUMBRO"]
    sheet["E10"] = 9876
    sheet["F10"] = "MXN"
    workbook.save(workbook_path)
    workbook.close()
    files = (
        AdapterFile(**{**source_bundle[0].__dict__, "sha256": _sha256(workbook_path)}),
        *source_bundle[1:],
    )

    item = next(row for row in build_cr_global_snapshot(files)["items"] if row["sku"] == "CR-100-BLK")

    assert item["price_net"] == "9876.000000"
    assert item["base_currency"] == "MXN"
    assert any(reference.get("cell_or_bbox") == "E10" for reference in json.loads(item["source_reference"]))


def test_cost_sheet_sale_price_corrected_code_and_section_metadata_are_authoritative(source_bundle):
    workbook_path = source_bundle[0].local_path
    workbook = load_workbook(workbook_path)
    sheet = workbook.create_sheet("COSTO CR GLOBAL ")
    sheet["A6"] = "Linea: CR Global"
    sheet["C6"] = "Sistema: Sistemas"
    sheet["A7"] = "Famila: Estructuras"
    for cell, value in zip(
        ("A8", "B8", "C8", "D8", "E8", "F8", "G8", "H8"),
        ("Imagen.", "Cod.", "Descripcion.", "Medida/Unidad.", "P. Unitario.", "LAB Cedis", "Moneda", "Precio Venta 50% GP"),
    ):
        sheet[cell] = value
    sheet["B10"] = "CR-CORRECTED"
    sheet["C10"] = "Producto corregido con medidas 60 x 70 cm"
    sheet["D10"] = "60 x 70 cm"
    sheet["E10"] = 500
    sheet["G10"] = "MXN"
    sheet["H10"] = 1000
    sheet["A15"] = "Linea: CR Global"
    sheet["C15"] = "Sistema: Accesorios"
    sheet["A16"] = "Famila: Complementos"
    sheet["B18"] = "CR-SECOND"
    sheet["C18"] = "Segundo producto de 80 cm"
    sheet["E18"] = 400
    sheet["G18"] = "MXN"
    sheet["H18"] = 800
    workbook.save(workbook_path)
    workbook.close()
    files = (
        AdapterFile(**{**source_bundle[0].__dict__, "sha256": _sha256(workbook_path)}),
        *source_bundle[1:],
    )

    items = build_cr_global_snapshot(files)["items"]
    by_sku = {item["sku"]: item for item in items}

    assert set(by_sku) == {"CR-CORRECTED", "CR-SECOND"}
    assert by_sku["CR-CORRECTED"]["price_net"] == "1000.000000"
    assert by_sku["CR-CORRECTED"]["collection"] == "Estructuras"
    assert by_sku["CR-SECOND"]["price_net"] == "800.000000"
    assert by_sku["CR-SECOND"]["collection"] == "Complementos"
    assert by_sku["CR-SECOND"]["attributes"]["system"] == "Accesorios"
    assert "80 cm" in by_sku["CR-SECOND"]["attributes"]["dimensions"]


def test_price_list_is_authoritative_and_technical_pdf_only_fills_missing_description(source_bundle):
    snapshot = build_cr_global_snapshot(source_bundle)
    by_sku = {item["sku"]: item for item in snapshot["items"] if item["sku"]}

    assert by_sku["CR-100-BLK"]["price_net"] == "1234.500000"
    assert by_sku["CR-100-BLK"]["description"] == "Alpha task chair Black polypropylene shell"
    assert "Acoustic privacy panel" in by_sku["CR-200"]["description"]
    assert all(item["base_currency"] == "MXN" for item in snapshot["items"])
    assert all(item["tax_rate"] == "0.160000" for item in snapshot["items"])
    assert all(item["lead_time"] == "2-3 dias habiles" for item in snapshot["items"])


def test_exact_variant_keeps_own_sku_price_and_image(source_bundle):
    snapshot = build_cr_global_snapshot(source_bundle)
    variants = {item["sku"]: item for item in snapshot["items"] if item["sku"].startswith("CR-100")}

    assert set(variants) == {"CR-100-BLK", "CR-100-WHT"}
    assert variants["CR-100-BLK"]["price_net"] == "1234.500000"
    assert variants["CR-100-WHT"]["price_net"] == "1345.600000"
    assert variants["CR-100-BLK"]["attributes"]["image_sha256"] != variants["CR-100-WHT"]["attributes"]["image_sha256"]


def test_base_code_requires_an_explicit_variant_attribute(source_bundle, tmp_path):
    price_list = tmp_path / "base-code-price.pdf"
    _write_pdf(price_list, [_price_page(("CR-100", "BLACK"), ("1,111.00",))])
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(price_list), row.mime_type, price_list)
        if row.kind == "price_list" else row
        for row in source_bundle
    )
    by_sku = {item["sku"]: item for item in build_cr_global_snapshot(files)["items"] if item["sku"]}

    assert by_sku["CR-100-BLK"]["price_net"] == "1111.000000"
    assert by_sku["CR-100-WHT"]["price_net"] == "0.000000"


def test_base_code_does_not_match_short_or_numeric_substrings(source_bundle, tmp_path):
    workbook_file = source_bundle[0].local_path
    workbook = load_workbook(workbook_file)
    workbook["SPEC-GUIDE-LUMBRO"]["C9"] = "ALPHA 1"
    workbook["SPEC-GUIDE-LUMBRO"]["C19"] = "ALPHA 2"
    workbook.save(workbook_file)
    workbook.close()
    price_list = tmp_path / "substring-attribute-price.pdf"
    _write_pdf(price_list, [_price_page(("CR-100", "10"), ("1,111.00",))])
    files = tuple(
        AdapterFile(
            row.path,
            row.kind,
            row.brand,
            _sha256(workbook_file if row.kind == "spec_guide" else price_list),
            row.mime_type,
            workbook_file if row.kind == "spec_guide" else price_list,
        )
        if row.kind in {"spec_guide", "price_list"}
        else row
        for row in source_bundle
    )
    variants = [item for item in build_cr_global_snapshot(files)["items"] if item["sku"].startswith("CR-100")]

    assert variants
    assert all(item["price_net"] == "0.000000" for item in variants)


def test_normalized_name_matches_only_when_unique(source_bundle, tmp_path):
    price_list = tmp_path / "name-price.pdf"
    _write_pdf(
        price_list,
        ["LISTA DE PRECIOS\nPrecios mas IVA\nMoneda: MXN\nBETA   PANEL\n$\n2,222.00"],
    )
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(price_list), row.mime_type, price_list)
        if row.kind == "price_list" else row
        for row in source_bundle
    )

    item = next(item for item in build_cr_global_snapshot(files)["items"] if item["sku"] == "CR-200")
    assert item["price_net"] == "2222.000000"


def test_normalized_name_does_not_cross_an_unrelated_row(source_bundle, tmp_path):
    price_list = tmp_path / "non-adjacent-name-price.pdf"
    _write_pdf(
        price_list,
        ["LISTA DE PRECIOS\nPrecios mas IVA\nMoneda: MXN\nBETA PANEL\nOTRO PRODUCTO\n$\n2,222.00"],
    )
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(price_list), row.mime_type, price_list)
        if row.kind == "price_list" else row
        for row in source_bundle
    )

    item = next(item for item in build_cr_global_snapshot(files)["items"] if item["sku"] == "CR-200")
    assert item["price_net"] == "0.000000"


def test_conflicting_price_requires_review_regardless_of_file_order(source_bundle, tmp_path):
    conflict = tmp_path / "conflicting-price.pdf"
    _write_pdf(
        conflict,
        [
            _price_page(("CR-100-BLK",), ("1,234.50",)),
            _price_page(("CR-100-BLK",), ("9,999.99",)),
        ],
    )
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(conflict), row.mime_type, conflict)
        if row.kind == "price_list"
        else row
        for row in source_bundle
    )

    item = next(item for item in build_cr_global_snapshot(files)["items"] if item["sku"] == "CR-100-BLK")
    assert item["price_net"] == "0.000000"
    assert any("precio" in warning.lower() and "conflic" in warning.lower() for warning in item["warnings"])


def test_missing_code_is_published_for_review_without_sku(source_bundle):
    snapshot = build_cr_global_snapshot(source_bundle)
    item = next(item for item in snapshot["items"] if item["name"] == "UNCODED STOOL")

    assert item["code_status"] == "needs_review"
    assert item["sku"] == ""
    assert item["price_net"] == "0.000000"
    assert any("codigo" in warning.lower() for warning in item["warnings"])


def test_duplicate_source_code_is_not_arbitrarily_resolved(source_bundle):
    workbook_file = source_bundle[0].local_path
    workbook = load_workbook(workbook_file)
    workbook["SPEC-GUIDE-LUMBRO"]["A30"] = "CR-100-BLK"
    workbook.save(workbook_file)
    workbook.close()
    files = (AdapterFile(**{**source_bundle[0].__dict__, "sha256": _sha256(workbook_file)}),) + source_bundle[1:]

    duplicates = [item for item in build_cr_global_snapshot(files)["items"] if "duplicado" in " ".join(item["warnings"]).lower()]
    assert len(duplicates) == 2
    assert all(item["code_status"] == "needs_review" and item["sku"] == "" for item in duplicates)
    assert all(item["price_net"] == "0.000000" for item in duplicates)


@pytest.mark.parametrize(
    "mutation",
    [
        lambda rows: rows[:-1],
        lambda rows: rows + (rows[0],),
        lambda rows: tuple(
            AdapterFile(row.path, row.kind, row.brand, row.sha256, row.mime_type, None)
            if row.kind == "catalog" else row
            for row in rows
        ),
        lambda rows: tuple(
            AdapterFile(row.path + ".pdf", row.kind, row.brand, row.sha256, row.mime_type, row.local_path)
            if row.kind == "spec_guide" else row
            for row in rows
        ),
    ],
)
def test_bundle_shape_fails_closed(source_bundle, mutation):
    with pytest.raises((ValueError, SourceSafetyError)):
        build_cr_global_snapshot(mutation(source_bundle))


def test_declared_hash_mismatch_and_unsafe_content_fail_closed(source_bundle, tmp_path):
    mismatched = (AdapterFile(**{**source_bundle[0].__dict__, "sha256": "0" * 64}),) + source_bundle[1:]
    with pytest.raises((ValueError, SourceSafetyError)):
        build_cr_global_snapshot(mismatched)

    unsafe = tmp_path / "CRG_LP_General_Dist_2026-04.pdf"
    unsafe.write_bytes(b"not a pdf")
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(unsafe), row.mime_type, unsafe)
        if row.kind == "price_list" else row
        for row in source_bundle
    )
    with pytest.raises(SourceSafetyError):
        build_cr_global_snapshot(files)


def test_currency_missing_or_contradictory_blocks_prices(source_bundle, tmp_path):
    for currency, notice in (("USD", True), ("MXN", False)):
        price_list = tmp_path / f"prices-{currency}-{notice}.pdf"
        _write_pdf(price_list, [_price_page(("CR-100-BLK",), ("1,234.50",), currency=currency, notice=notice)])
        files = tuple(
            AdapterFile(row.path, row.kind, row.brand, _sha256(price_list), row.mime_type, price_list)
            if row.kind == "price_list" else row
            for row in source_bundle
        )
        item = next(item for item in build_cr_global_snapshot(files)["items"] if item["sku"] == "CR-100-BLK")
        assert item["price_net"] == "0.000000"
        assert any("moneda" in warning.lower() for warning in item["warnings"])


def test_dollar_symbol_without_explicit_pdf_currency_blocks_price(source_bundle, tmp_path):
    price_list = tmp_path / "missing-explicit-currency.pdf"
    _write_pdf(
        price_list,
        ["LISTA DE PRECIOS\nPrecios mas IVA\nCLAVE\nCR-100-BLK\nDESCRIPCION\n$\n1,234.50"],
    )
    files = tuple(
        AdapterFile(row.path, row.kind, row.brand, _sha256(price_list), row.mime_type, price_list)
        if row.kind == "price_list" else row
        for row in source_bundle
    )

    item = next(item for item in build_cr_global_snapshot(files)["items"] if item["sku"] == "CR-100-BLK")
    assert item["price_net"] == "0.000000"
    assert any("moneda" in warning.lower() for warning in item["warnings"])


def test_source_hash_source_reference_and_output_are_deterministic(source_bundle):
    first = build_cr_global_snapshot(source_bundle)
    second = build_cr_global_snapshot(tuple(reversed(source_bundle)))

    expected_hash = hashlib.sha256(
        "\n".join(f"{row.path}\0{row.sha256}" for row in sorted(source_bundle, key=lambda row: row.path)).encode()
    ).hexdigest()
    assert first["source_hash"] == expected_hash
    assert _without_generated_at(first) == _without_generated_at(second)
    assert first["items"] == sorted(first["items"], key=lambda item: item["internal_id"])
    for item in first["items"]:
        evidence = json.loads(item["source_reference"])
        assert item["source_reference"] == json.dumps(evidence, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
        assert evidence


def test_ignored_real_sources_reconcile_full_cr_global_catalog():
    root = Path(".cache/catalog-sources/cr-global")
    spec = root / "Spec guide-CR Global-2026.xlsx"
    price_list = root / "CRG_LP_General_Dist_2026-04.pdf"
    catalog = root / "CRG_FT_General_Dist_2026-04.pdf"
    if not all(path.exists() for path in (spec, price_list, catalog)):
        pytest.skip("ignored CR Global sources are not available")
    assert _sha256(spec) == "25b2f1984b2666d0fa004527a271f097cba56683f233d6b905e09fcb0716ff9b"
    assert _sha256(price_list) == "01293de9e48b0a5bc41d4a451fa458dbd759b7b0de32efb9ef9872d918f1057e"
    assert _sha256(catalog) == "471b6825958243b6f1bd1e7be32a5099fb3371047a82888efa68eee8d456f17c"
    files = (
        AdapterFile(
            "SPEC GUIDES 2026/CR Global/Spec guide-CR Global-2026.xlsx",
            "spec_guide",
            None,
            _sha256(spec),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            spec,
        ),
        AdapterFile(
            "CR GLOBAL/CRG_LP_General_Dist_2026-04.pdf",
            "price_list",
            None,
            _sha256(price_list),
            "application/pdf",
            price_list,
        ),
        AdapterFile(
            "CR GLOBAL/CRG_FT_General_Dist_2026-04.pdf",
            "catalog",
            None,
            _sha256(catalog),
            "application/pdf",
            catalog,
        ),
    )

    build = build_cr_global_snapshot_with_assets(files)
    items = load_supplier_catalog_data(build.snapshot, expected_supplier="cr-global")["items"]
    metrics = {
        "items": len(items),
        "verified": sum(item["code_status"] == "verified" for item in items),
        "priced": sum(item["price_net"] != "0.000000" for item in items),
        "images": len(build.bindings),
        "assets": len(build.assets_by_sha256),
        "dimensions": sum(bool(item["attributes"].get("dimensions")) for item in items),
        "colors": sum(bool(item["attributes"].get("color")) for item in items),
        "warranties": sum(bool(item["attributes"].get("warranty")) for item in items),
        "links": sum(bool(item["product_url"]) for item in items),
        "collections": len({item["collection"] for item in items}),
        "three_month_lead_time": sum(item["lead_time"] == "3 MESES" for item in items),
        "made_to_order": sum(item["availability_type"] == "made_to_order" for item in items),
    }
    print("CR_GLOBAL_REAL_METRICS=" + json.dumps(metrics, sort_keys=True))
    print("CR_GLOBAL_REAL_WARNINGS=" + json.dumps(sorted({warning for item in items for warning in item["warnings"]})))

    assert metrics["items"] == 56
    assert metrics["verified"] == 56
    assert metrics["priced"] == 56
    assert metrics["images"] == 53
    assert metrics["dimensions"] == 53
    assert metrics["colors"] == 55
    assert metrics["warranties"] == 56
    assert metrics["links"] == 34
    assert metrics["collections"] >= 4
    assert metrics["three_month_lead_time"] == 4
    assert metrics["made_to_order"] == 4
    by_sku = {item["sku"]: item for item in items}
    assert by_sku["CR33-2S2"]["product_url"] == "https://www.crglobal.mx/product-page/estructura-de-un-motor-1"
    assert by_sku["CR33-2S2"]["attributes"]["product_url_match"]["status"] == "exact_code"
    assert by_sku["CR-STVG806"]["product_url"] == ""
    assert by_sku["MONITOR ARM-PS160"]["product_url"] == ""
