from __future__ import annotations

from io import BytesIO
from decimal import Decimal
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import PatternFill
from PIL import Image


MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"


def write_import_fixture(path: Path, *, currency: str | None = None) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    for column, title in {
        1: "No.",
        2: "Item Name",
        3: "Photo",
        4: "Description",
        5: "Dimension",
        7: "Q'ty",
        8: "Vol.",
        10: "Unit Price",
    }.items():
        sheet.cell(7, column, title)
    if currency:
        sheet.cell(7, 14, "Original Currency")

    rows = [
        (8, "category", "SALA DE JUNTAS SECUNDARIO"),
        (9, "product", "DV74 I-Varna II Conference Table"),
        (10, "category", "MUESTRAS"),
        (11, "product", "CAI63SW Alien Task Chair"),
        (12, "product", "CAL61KC Aulenti Task Chair"),
        (13, "product", "CAT60SC Altaes Task Chair"),
        (14, "product", "DL60 Single Seat Workstation"),
        (15, "product", "DL61 Double Seat Workstation"),
        (16, "category", "CONCEJO"),
        (17, "product", "DV74 I-Varna II Conference Table"),
    ]
    product_index = 0
    for row, kind, value in rows:
        if kind == "category":
            sheet.cell(row, 1, f"- {value}")
            continue
        product_index += 1
        sheet.cell(row, 1, product_index)
        sheet.cell(row, 2, value)
        sheet.cell(row, 4, f"Descripción {product_index}")
        sheet.cell(row, 5, f"{600 + product_index} x 600 mm")
        sheet.cell(row, 7, 1 if row != 14 else 2)
        sheet.cell(row, 8, Decimal("0.25"))
        sheet.cell(row, 10, Decimal("80.50") if row == 11 else Decimal("100.00"))
        if currency:
            sheet.cell(row, 14, currency)
        image_path = path.parent / f"fixture-{row}.png"
        Image.new("RGB", (80, 60), (20 * product_index, 80, 120)).save(image_path)
        sheet.add_image(WorkbookImage(str(image_path)), f"C{row}")
    sheet["A1"] = "SUNON TECHNOLOGY CO.,LTD."
    sheet.cell(65536, 14).fill = PatternFill("solid", fgColor="FFFFFF")
    workbook.save(path)
    workbook.close()
    return path


def build_rich_quotation_fixture(
    path: Path,
    *,
    formulas: dict[str, str] | None = None,
    merges: list[str] | None = None,
    image_anchor: str = "B9",
    print_area: str = "A1:N40",
    hidden_rows: list[int] | None = None,
    state: str = "hidden",
) -> Path:
    """Construye un XLSX pequeño con dependencias OOXML difíciles de copiar."""

    formulas = formulas or {"N9": "=G9*J9"}
    merges = merges or ["A1:N1", "B9:C9"]
    hidden_rows = hidden_rows or [12]
    if state not in {"visible", "hidden", "veryHidden"}:
        raise ValueError("Estado de fixture inválido")

    with BytesIO() as image_stream:
        Image.new("RGB", (11, 7), (23, 97, 151)).save(image_stream, format="PNG")
        image_bytes = image_stream.getvalue()

    formula_cells = "".join(
        (
            f'<c r="{escape(cell)}" s="2"><f>{escape(formula.removeprefix("="))}</f>'
            "<v>6</v></c>"
        )
        for cell, formula in sorted(formulas.items())
    )
    hidden = set(hidden_rows)
    row_12_hidden = ' hidden="1"' if 12 in hidden else ""
    merge_xml = "".join(f'<mergeCell ref="{escape(ref)}"/>' for ref in merges)
    anchor_column = _column_index(image_anchor)
    anchor_row = int("".join(character for character in image_anchor if character.isdigit())) - 1
    absolute_print_area = ":".join(
        _absolute_cell(reference) for reference in print_area.split(":")
    )

    workbook_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="{MAIN}" xmlns:r="{OFFICE_REL}">
  <sheets>
    <sheet name="Catalog" sheetId="3" r:id="rId1"/>
    <sheet name="Quotation" sheetId="17" state="{state}" r:id="rId7"/>
  </sheets>
  <definedNames>
    <definedName name="GlobalUnrelated">Catalog!$A$1</definedName>
    <definedName name="_xlnm.Print_Area" localSheetId="1">Quotation!{absolute_print_area}</definedName>
    <definedName name="_xlnm.Print_Titles" localSheetId="1">Quotation!$1:$7</definedName>
    <definedName name="QuoteLocal" hidden="1" localSheetId="1">Quotation!$B$9</definedName>
    <definedName name="CatalogLocal" localSheetId="0">Catalog!$A$1</definedName>
  </definedNames>
</workbook>'''.encode("utf-8")

    worksheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{MAIN}" xmlns:r="{OFFICE_REL}">
  <dimension ref="A1:N40"/>
  <sheetViews><sheetView workbookViewId="0" showGridLines="0"><pane ySplit="7" topLeftCell="A8" state="frozen"/><selection pane="bottomLeft" activeCell="B9" sqref="B9"/></sheetView></sheetViews>
  <sheetFormatPr defaultRowHeight="15"/>
  <cols><col min="2" max="3" width="20" customWidth="1" style="1"/><col min="4" max="4" width="12" hidden="1" customWidth="1" style="2"/></cols>
  <sheetData>
    <row r="1" ht="28" customHeight="1"><c r="A1" s="1" t="s"><v>0</v></c></row>
    <row r="9" s="2" customFormat="1"><c r="B9" s="1" t="s"><v>1</v></c><c r="G9" s="2"><v>2</v></c><c r="J9" s="2"><v>3</v></c>{formula_cells}</row>
    <row r="12"{row_12_hidden}><c r="A12" s="1" t="inlineStr"><is><t>Oculta</t></is></c></row>
  </sheetData>
  <mergeCells count="{len(merges)}">{merge_xml}</mergeCells>
  <conditionalFormatting sqref="N9"><cfRule type="cellIs" dxfId="0" priority="1" operator="greaterThan"><formula>0</formula></cfRule></conditionalFormatting>
  <hyperlinks><hyperlink ref="B9" r:id="rIdHyper" display="Ficha externa"/></hyperlinks>
  <printOptions horizontalCentered="1" headings="1"/>
  <pageMargins left="0.25" right="0.25" top="0.5" bottom="0.5" header="0.2" footer="0.2"/>
  <pageSetup paperSize="9" orientation="landscape" fitToWidth="1" fitToHeight="0" r:id="rIdPrinter"/>
  <legacyDrawing r:id="rIdVml"/>
  <drawing r:id="rIdDrawing"/>
  <tableParts count="1"><tablePart r:id="rIdTable"/></tableParts>
</worksheet>'''.encode("utf-8")

    drawing_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="{OFFICE_REL}">
  <xdr:twoCellAnchor editAs="oneCell">
    <xdr:from><xdr:col>{anchor_column}</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{anchor_row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
    <xdr:to><xdr:col>{anchor_column + 1}</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{anchor_row + 1}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
    <xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="Producto"/><xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill><a:blip r:embed="rIdImage"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>'''.encode("utf-8")

    table_columns = "".join(
        f'<tableColumn id="{index}" name="Columna {index}"'
        + (' dataDxfId="0"' if index == 1 else "")
        + "/>"
        for index in range(1, 15)
    )
    parts = {
        "[Content_Types].xml": _rich_content_types(),
        "_rels/.rels": _rels_xml(
            ("rIdRoot", f"{OFFICE_REL}/officeDocument", "xl/workbook.xml", None),
        ),
        "xl/workbook.xml": workbook_xml,
        "xl/_rels/workbook.xml.rels": _rels_xml(
            ("rId1", f"{OFFICE_REL}/worksheet", "worksheets/catalog.xml", None),
            ("rId7", f"{OFFICE_REL}/worksheet", "worksheets/original-quotation.xml", None),
            ("rIdStyles", f"{OFFICE_REL}/styles", "styles.xml", None),
            ("rIdShared", f"{OFFICE_REL}/sharedStrings", "sharedStrings.xml", None),
            ("rIdTheme", f"{OFFICE_REL}/theme", "theme/theme7.xml", None),
        ),
        "xl/worksheets/catalog.xml": (
            f'<worksheet xmlns="{MAIN}"><sheetData><row r="1"><c r="A1" t="inlineStr"><is><t>Catálogo</t></is></c></row></sheetData></worksheet>'
        ).encode("utf-8"),
        "xl/worksheets/original-quotation.xml": worksheet_xml,
        "xl/worksheets/_rels/original-quotation.xml.rels": _rels_xml(
            ("rIdDrawing", f"{OFFICE_REL}/drawing", "../drawings/drawing7.xml", None),
            ("rIdComments", f"{OFFICE_REL}/comments", "../comments/comment7.xml", None),
            ("rIdVml", f"{OFFICE_REL}/vmlDrawing", "../drawings/vmlDrawing7.vml", None),
            ("rIdTable", f"{OFFICE_REL}/table", "../tables/table7.xml", None),
            ("rIdPrinter", f"{OFFICE_REL}/printerSettings", "../printerSettings/printerSettings7.bin", None),
            ("rIdHyper", f"{OFFICE_REL}/hyperlink", "https://example.com/spec?q=1", "External"),
        ),
        "xl/drawings/drawing7.xml": drawing_xml,
        "xl/drawings/_rels/drawing7.xml.rels": _rels_xml(
            ("rIdImage", f"{OFFICE_REL}/image", "../media/image7.png", None),
        ),
        "xl/media/image7.png": image_bytes,
        "xl/comments/comment7.xml": (
            f'<comments xmlns="{MAIN}"><authors><author>Mobiliti</author></authors><commentList><comment ref="B9" authorId="0"><text><r><rPr><b/></rPr><t>Comentario</t></r></text></comment></commentList></comments>'
        ).encode("utf-8"),
        "xl/drawings/vmlDrawing7.vml": b'<xml xmlns:v="urn:schemas-microsoft-com:vml"><v:shape id="comment-shape"/></xml>',
        "xl/tables/table7.xml": (
            f'<table xmlns="{MAIN}" id="7" name="QuoteTable" displayName="QuoteTable" ref="A8:N12" totalsRowShown="0" headerRowDxfId="0"><autoFilter ref="A8:N12"/><tableColumns count="14">{table_columns}</tableColumns><tableStyleInfo name="CustomQuoteStyle" showFirstColumn="0" showLastColumn="0" showRowStripes="1" showColumnStripes="0"/></table>'
        ).encode("utf-8"),
        "xl/printerSettings/printerSettings7.bin": _valid_printer_settings(),
        "xl/sharedStrings.xml": _rich_shared_strings(),
        "xl/styles.xml": _rich_styles(),
        "xl/theme/theme7.xml": _rich_theme(),
    }
    path = Path(path)
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, content in parts.items():
            archive.writestr(name, content)
    return path


def _column_index(cell: str) -> int:
    letters = "".join(character for character in cell if character.isalpha()).upper()
    value = 0
    for character in letters:
        value = value * 26 + ord(character) - ord("A") + 1
    return value - 1


def _absolute_cell(reference: str) -> str:
    letters = "".join(character for character in reference if character.isalpha()).upper()
    digits = "".join(character for character in reference if character.isdigit())
    if not letters or not digits:
        raise ValueError("Referencia de celda inválida")
    return f"${letters}${digits}"


def _rels_xml(*relationships: tuple[str, str, str, str | None]) -> bytes:
    children = []
    for relationship_id, relationship_type, target, mode in relationships:
        mode_xml = f' TargetMode="{escape(mode)}"' if mode else ""
        children.append(
            f'<Relationship Id="{escape(relationship_id)}" Type="{escape(relationship_type)}" Target="{escape(target)}"{mode_xml}/>'
        )
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{PACKAGE_REL}">{"".join(children)}</Relationships>'
    ).encode("utf-8")


def _rich_content_types() -> bytes:
    defaults = (
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Default Extension="png" ContentType="image/png"/>'
        '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
        '<Default Extension="bin" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings"/>'
    )
    overrides = {
        "/xl/workbook.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "/xl/worksheets/catalog.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
        "/xl/worksheets/original-quotation.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
        "/xl/styles.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml",
        "/xl/sharedStrings.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
        "/xl/theme/theme7.xml": "application/vnd.openxmlformats-officedocument.theme+xml",
        "/xl/drawings/drawing7.xml": "application/vnd.openxmlformats-officedocument.drawing+xml",
        "/xl/comments/comment7.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml",
        "/xl/tables/table7.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml",
    }
    override_xml = "".join(
        f'<Override PartName="{name}" ContentType="{content_type}"/>'
        for name, content_type in overrides.items()
    )
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="{CONTENT_TYPES}">{defaults}{override_xml}</Types>'
    ).encode("utf-8")


def _rich_shared_strings() -> bytes:
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="{MAIN}" count="2" uniqueCount="2">
  <si><r><rPr><b/><color rgb="FF006699"/></rPr><t xml:space="preserve"> Rich </t></r><r><rPr><i/></rPr><t>Text</t></r><rPh sb="0" eb="4"><t>ritchi</t></rPh><phoneticPr fontId="1" type="noConversion"/></si>
  <si><t>Producto Uno</t></si>
</sst>'''.encode("utf-8")


def _valid_printer_settings() -> bytes:
    """DEVMODEW minimo y coherente para validar printerSettings sin payload opaco."""

    payload = bytearray(220)
    device_name = "Mobiliti Printer".encode("utf-16le")
    payload[: len(device_name)] = device_name
    payload[64:66] = (0x0401).to_bytes(2, "little")
    payload[66:68] = (1).to_bytes(2, "little")
    payload[68:70] = (220).to_bytes(2, "little")
    payload[70:72] = (0).to_bytes(2, "little")
    return bytes(payload)


def _rich_styles() -> bytes:
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="{MAIN}">
  <numFmts count="1"><numFmt numFmtId="164" formatCode="&quot;Q-&quot;0.000"/></numFmts>
  <fonts count="2">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/><scheme val="minor"/></font>
    <font><b/><sz val="12"/><color rgb="FF123456"/><name val="FixtureFont"/></font>
  </fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFABCDEF"/><bgColor indexed="64"/></patternFill></fill></fills>
  <borders count="2"><border><left/><right/><top/><bottom/><diagonal/></border><border><left style="thick"><color rgb="FF010203"/></left><right style="thin"><color rgb="FF040506"/></right><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/><xf numFmtId="164" fontId="1" fillId="2" borderId="1" applyNumberFormat="1" applyFont="1" applyFill="1" applyBorder="1"/></cellStyleXfs>
  <cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="164" fontId="1" fillId="2" borderId="1" xfId="1" applyNumberFormat="1" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" wrapText="1"/></xf><xf numFmtId="164" fontId="1" fillId="2" borderId="1" xfId="1" applyNumberFormat="1" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="right"/></xf></cellXfs>
  <cellStyles count="2"><cellStyle name="Normal" xfId="0" builtinId="0"/><cellStyle name="Fixture Style" xfId="1"/></cellStyles>
  <dxfs count="1"><dxf><font><color rgb="FFFF0000"/></font><fill><patternFill patternType="solid"><fgColor rgb="FFFFFF00"/></patternFill></fill><numFmt numFmtId="164" formatCode="&quot;Q-&quot;0.000"/></dxf></dxfs>
  <tableStyles count="1" defaultTableStyle="TableStyleMedium2" defaultPivotStyle="PivotStyleLight16"><tableStyle name="CustomQuoteStyle" pivot="0" table="1" count="1"><tableStyleElement type="firstRowStripe" size="1" dxfId="0"/></tableStyle></tableStyles>
</styleSheet>'''.encode("utf-8")


def _rich_theme() -> bytes:
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/main"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="{drawing}" name="Fixture Theme">
  <a:themeElements>
    <a:clrScheme name="Fixture">
      <a:dk1><a:sysClr val="windowText" lastClr="010203"/></a:dk1>
      <a:lt1><a:sysClr val="window" lastClr="FEFEFE"/></a:lt1>
      <a:dk2><a:srgbClr val="102030"/></a:dk2>
      <a:lt2><a:srgbClr val="E0E1E2"/></a:lt2>
      <a:accent1><a:srgbClr val="123456"/></a:accent1>
      <a:accent2><a:srgbClr val="654321"/></a:accent2>
      <a:accent3><a:srgbClr val="336699"/></a:accent3>
      <a:accent4><a:srgbClr val="993366"/></a:accent4>
      <a:accent5><a:srgbClr val="669933"/></a:accent5>
      <a:accent6><a:srgbClr val="996633"/></a:accent6>
      <a:hlink><a:srgbClr val="0000EE"/></a:hlink>
      <a:folHlink><a:srgbClr val="551A8B"/></a:folHlink>
    </a:clrScheme>
    <a:fontScheme name="Fixture Fonts">
      <a:majorFont><a:latin typeface="Fixture Major"/><a:ea typeface=""/><a:cs typeface=""/></a:majorFont>
      <a:minorFont><a:latin typeface="Fixture Minor"/><a:ea typeface=""/><a:cs typeface=""/></a:minorFont>
    </a:fontScheme>
    <a:fmtScheme name="Fixture Format"/>
  </a:themeElements>
</a:theme>'''.encode("utf-8")
