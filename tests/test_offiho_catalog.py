from decimal import Decimal
from datetime import datetime, timezone
from email.message import Message
from io import BytesIO
from pathlib import Path
from urllib.parse import urlsplit
import hashlib
import json

import pytest
from openpyxl import load_workbook
from PIL import Image

import scripts.build_offiho_catalog as build
from scripts.build_offiho_catalog import extract_offiho_identity
from scripts.build_offiho_catalog import match_official_product
from scripts.build_offiho_catalog import parse_inventory_xls
from scripts.build_offiho_catalog import parse_pdf_price_index


def fake_runtime_catalog(
    *,
    available_quantity: int,
    unit_price: int,
    image_url: str = "",
    description: str = "",
    price_source: str = "inventory",
):
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    item = OffihoCatalogItem(
        inventory_key="OHE-405 NEGRO ALUFSEN",
        code="OHE-405",
        name="ALUFSEN",
        variant="NEGRO",
        unit="PZA",
        pieces_per_box=Decimal("1"),
        available_quantity=Decimal(str(available_quantity)),
        unit_price=Decimal(str(unit_price)),
        price_source=price_source,
        product_url="https://www.offiho.com/productos/alufsen",
        image_url=image_url,
        description=description,
    )
    return {"source_hash": "hash", "items": [item], "by_inventory_key": {item.inventory_key: item}}


def test_offiho_cart_accepts_exhausted_and_overstock_lines():
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    exhausted = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 3}],
        catalog=fake_runtime_catalog(available_quantity=0, unit_price=7999),
    )
    insufficient = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 3}],
        catalog=fake_runtime_catalog(available_quantity=2, unit_price=7999),
    )

    assert exhausted["items"][0]["stock_status"] == "out_of_stock"
    assert insufficient["items"][0]["stock_status"] == "insufficient_stock"
    assert insufficient["items"][0]["unit_price"] == 7999


def test_offiho_cart_uses_catalog_owned_values_and_rejects_non_positive_quantity():
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    payload = build_offiho_cart_payload(
        [
            {
                "inventory_key": "OHE-405 NEGRO ALUFSEN",
                "quantity": "2",
                "unit_price": 1,
                "available_quantity": 999,
                "product_url": "https://example.test/untrusted",
                "image_url": "https://example.test/untrusted.jpg",
            }
        ],
        catalog=fake_runtime_catalog(
            available_quantity=2,
            unit_price=7999,
            image_url="https://www.offiho.com/uploads/alufsen.jpg",
        ),
    )

    assert payload["items"][0]["stock_status"] == "available"
    assert payload["items"][0]["unit_price"] == 7999
    assert payload["items"][0]["available_quantity"] == 2
    assert payload["items"][0]["product_url"] == "https://www.offiho.com/productos/alufsen"
    assert payload["items"][0]["image_url"] == "https://www.offiho.com/uploads/alufsen.jpg"
    with pytest.raises(ValueError, match="Cantidad invalida"):
        build_offiho_cart_payload(
            [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 0}],
            catalog=fake_runtime_catalog(available_quantity=252, unit_price=7999),
        )


def test_offiho_description_reaches_cart_and_quotation_workbook(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )

    description = "Estructura de acero tubular y asiento de polipropileno de alta resistencia."
    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 2}],
        catalog=fake_runtime_catalog(
            available_quantity=5,
            unit_price=7999,
            description=description,
        ),
    )
    output = tmp_path / "offiho-description.xlsx"
    create_offiho_quotation_workbook(payload, output)

    assert payload["items"][0]["description"] == description
    workbook = load_workbook(output, read_only=True)
    assert description in workbook["Quotation"]["D9"].value
    assert "Variante: NEGRO" in workbook["Quotation"]["D9"].value
    workbook.close()

    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    final_output = tmp_path / "offiho-description-final.xlsx"
    generate_online_quote(output, final_output, {"tipo_cambio": "20"})
    final = load_workbook(final_output, read_only=True)
    assert any(
        description in str(cell.value or "")
        for row in final["Cotizacion"].iter_rows()
        for cell in row
    )
    final.close()


def test_exact_offiho_finish_image_reaches_cart_and_quotation_workbook(monkeypatch, tmp_path):
    from PIL import Image

    from mobiliti_saas.quote_engine import catalog_cart
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
        load_offiho_catalog,
    )

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHS-86AL ROJA REVOLUTION", "quantity": 1}],
        catalog=load_offiho_catalog(),
    )
    expected_url = (
        "https://www.offiho.com/operativos/revolution/"
        "OHS-86al/colores/OHS-86alRojo.jpg"
    )
    captured = {}

    def fake_download(image_url, image_dir, code, source_type):
        captured.update(url=image_url, code=code, source_type=source_type)
        image_path = image_dir / "revolution-roja.png"
        Image.new("RGB", (40, 60), "red").save(image_path)
        return image_path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)

    output = create_offiho_quotation_workbook(payload, tmp_path / "revolution-roja.xlsx")
    workbook = load_workbook(output)

    assert payload["items"][0]["image_url"] == expected_url
    assert captured == {
        "url": expected_url,
        "code": "OHS-86AL",
        "source_type": "offiho_cart",
    }
    assert len(workbook["Quotation"]._images) == 1
    workbook.close()


def test_offiho_cart_uses_inventory_key_when_variant_name_is_blank(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        OffihoCatalogItem,
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )
    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    item = OffihoCatalogItem(
        inventory_key="VESPER/05 NEGRA",
        code="VESPER/05",
        name="",
        variant="NEGRA",
        unit="PZA",
        pieces_per_box=Decimal("1"),
        available_quantity=Decimal("312"),
        unit_price=Decimal("2799"),
        description="Estructura de acero tubular y asiento de polipropileno.",
    )
    catalog = {
        "source_hash": "hash",
        "items": [item],
        "by_inventory_key": {item.inventory_key: item},
    }
    payload = build_offiho_cart_payload(
        [{"inventory_key": item.inventory_key, "quantity": 1}],
        catalog=catalog,
    )
    source = create_offiho_quotation_workbook(payload, tmp_path / "vesper-source.xlsx")
    final_path = tmp_path / "vesper-final.xlsx"
    generate_online_quote(source, final_path, {"tipo_cambio": "20"})

    final = load_workbook(final_path, read_only=True)
    assert final["Quotation"]["B9"].value == "VESPER NEGRA"
    assert "modelo VESPER NEGRA" in final["Cotizacion"]["C17"].value
    final.close()
    with pytest.raises(ValueError, match="Cantidad invalida"):
        build_offiho_cart_payload(
            [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": "NaN"}],
            catalog=fake_runtime_catalog(available_quantity=252, unit_price=7999),
        )


@pytest.mark.parametrize("quantity", ["1e5000", "0.0001", "1000000.001"])
def test_offiho_cart_rejects_extreme_or_overprecise_quantity(quantity):
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    with pytest.raises(ValueError, match="Cantidad invalida"):
        build_offiho_cart_payload(
            [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": quantity}],
            catalog=fake_runtime_catalog(available_quantity=0, unit_price=7999),
        )


def test_offiho_cart_accepts_commercial_quantity_limit_and_three_decimals():
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": "1000000.000"}],
        catalog=fake_runtime_catalog(available_quantity=0, unit_price=7999),
    )

    assert payload["items"][0]["quantity"] == 1000000


def test_offiho_cart_accepts_200_unique_lines_and_rejects_201():
    from mobiliti_saas.quote_engine.offiho_catalog import MAX_CART_LINES, OffihoCatalogItem, build_offiho_cart_payload

    items = [
        OffihoCatalogItem(
            inventory_key=f"OHE-{index:03d} NEGRO MODELO {index}",
            code=f"OHE-{index:03d}",
            name=f"MODELO {index}",
            variant="NEGRO",
            unit="PZA",
            pieces_per_box=Decimal("1"),
            available_quantity=Decimal("0"),
            unit_price=Decimal("1"),
        )
        for index in range(MAX_CART_LINES + 1)
    ]
    catalog = {
        "source_hash": "hash",
        "items": items,
        "by_inventory_key": {item.inventory_key: item for item in items},
    }
    raw_items = [{"inventory_key": item.inventory_key, "quantity": 1} for item in items]

    assert len(build_offiho_cart_payload(raw_items[:MAX_CART_LINES], catalog=catalog)["items"]) == 200
    with pytest.raises(ValueError, match="200"):
        build_offiho_cart_payload(raw_items, catalog=catalog)


def test_offiho_cart_rejects_duplicate_inventory_key():
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    line = {"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 1}
    with pytest.raises(ValueError, match="duplicada"):
        build_offiho_cart_payload([line, dict(line)], catalog=fake_runtime_catalog(available_quantity=1, unit_price=7999))


def _runtime_catalog_raw():
    path = Path(__file__).resolve().parents[1] / "mobiliti_saas" / "quote_engine" / "data" / "offiho_catalog.json"
    return json.loads(path.read_text(encoding="utf-8"))


def test_load_offiho_catalog_propagates_inventory_audit():
    from mobiliti_saas.quote_engine.offiho_catalog import load_offiho_catalog

    catalog = load_offiho_catalog()

    assert catalog["source_row_count"] == 1368
    assert catalog["duplicate_row_count"] == 80
    assert catalog["unique_item_count"] == 1288


def test_checked_in_offiho_catalog_keeps_official_media_coverage():
    raw = _runtime_catalog_raw()
    items = raw["items"]
    linked = [item for item in items if item.get("product_url")]
    imaged = [item for item in items if item.get("image_url")]
    official_hosts = {
        "offiho.com",
        "www.offiho.com",
        "offihoblack.com",
        "www.offihoblack.com",
        "colos.it",
        "www.colos.it",
        "web-lemon-one-45.vercel.app",
        "mobiliti11-my.sharepoint.com",
    }

    assert len(items) == 1288
    assert len(linked) >= 850
    # Las fotos oficiales exactas conservan prioridad. Los antiguos faltantes
    # sólo se completan con referencias generadas, explícitamente etiquetadas y
    # con procedencia auditable.
    assert len(imaged) == 1288
    assert sum(item.get("image_kind") == "official" for item in items) == 1073
    generated = [
        item for item in items if item.get("image_kind") == "generated_reference"
    ]
    assert len(generated) == 215
    assert all(item.get("image_label", "").startswith("Imagen generada") for item in generated)
    assert all(len(item.get("image_references", [])) >= 2 for item in generated)
    assert sum(not item.get("image_url") for item in items) == 0
    assert all(item.get("product_url") for item in imaged)
    for item in linked:
        fields = ("product_url", "image_url") if item.get("image_url") else ("product_url",)
        for field in fields:
            parsed = urlsplit(item[field])
            assert parsed.scheme == "https"
            assert parsed.hostname in official_hosts
            if parsed.hostname == "web-lemon-one-45.vercel.app":
                assert parsed.path.startswith("/catalog-assets/offiho/")
        if item.get("image_url"):
            assert item["product_url"] != item["image_url"]


def test_checked_in_revolution_variants_use_their_exact_official_color_image():
    items = [
        item
        for item in _runtime_catalog_raw()["items"]
        if item.get("code") == "OHS-86AL"
    ]
    expected_suffixes = {
        "BLANCO": "OHS-86alBlanco.jpg",
        "ARENA": "OHS-86alArena.jpg",
        "GRIS": "OHS-86alGris.jpg",
        "NEGRO": "OHS-86alNegro.jpg",
        "VERDE": "OHS-86alVerde.jpg",
        "AZUL MARINO": "OHS-86alMarino.jpg",
        "ROJA": "OHS-86alRojo.jpg",
        "NARANJA": "OHS-86alNaranja.jpg",
    }

    assert {item["variant"] for item in items} == set(expected_suffixes)
    for item in items:
        assert item["image_url"].endswith(expected_suffixes[item["variant"]])
    assert len({item["image_url"] for item in items}) == len(expected_suffixes)


def test_checked_in_kyos_keeps_exacts_and_labels_generated_colors():
    catalog = json.loads(build.DEFAULT_OUTPUT.read_text(encoding="utf-8"))
    items = catalog["items"]

    kyos = {
        (item["code"], item["variant"]): item
        for item in items
        if item["code"] in {"OHV-331", "OHV-335CR", "OHV-337"}
    }
    expected_exact = {
        ("OHV-331", "BLANCO"),
        ("OHV-331", "GRIS"),
        ("OHV-335CR", "BLANCO"),
        ("OHV-335CR", "GRIS"),
        ("OHV-335CR", "NEGRO"),
        ("OHV-337", "BLANCO"),
        ("OHV-337", "GRIS"),
    }
    assert all(kyos[key]["image_url"] for key in expected_exact)
    for key in (("OHV-331", "NEGRO"), ("OHV-337", "NEGRO")):
        assert kyos[key]["image_url"]
        assert kyos[key]["image_kind"] == "generated_reference"
        assert kyos[key]["image_label"].startswith("Imagen generada")

    sling = {
        (item["code"], item["variant"]): item
        for item in items
        if item["code"] in {"OHE-94", "OHV-94"}
    }
    assert sling[("OHE-94", "PLUS GRIS")]["image_url"].endswith("/OHE-94plusGris.jpg")
    assert sling[("OHE-94", "PLUS NEGRO")]["image_url"].endswith("/OHE-94plusnegro.jpg")
    assert sling[("OHV-94", "PLUS GRIS")]["image_url"].endswith("/OHV-94plusGris.jpg")
    assert sling[("OHV-94", "PLUS NEGRO")]["image_url"].endswith("/OHV-94plus.jpg")
    plus_cr = next(item for item in items if item["inventory_key"] == "OHV-94 PLUS CR NEGRO SLING *")
    assert plus_cr["image_url"].endswith("/OHV-94plusNegroCr.jpg")


def test_checked_in_official_images_never_belong_to_another_structured_code():
    official_hosts = {
        "offiho.com",
        "www.offiho.com",
        "offihoblack.com",
        "www.offihoblack.com",
    }
    mismatches = []
    for item in _runtime_catalog_raw()["items"]:
        image_url = item.get("image_url", "")
        if urlsplit(image_url).hostname not in official_hosts:
            continue
        if build.CODE_RE.fullmatch(item.get("code", "")) is None:
            continue
        identity = extract_offiho_identity(item["inventory_key"])
        product_url = item.get("product_url", "")
        compact_image = build._compact_variant_value(image_url)
        curated_name_match = (
            identity.code in build.OFFICIAL_NAME_ALIASES
            and build._compact_variant_value(identity.name).rstrip("*") in compact_image
        )
        if not (
            build._image_targets_identity(image_url, identity)
            or build._image_targets_identity(product_url, identity)
            or curated_name_match
        ):
            mismatches.append((item["inventory_key"], image_url))

    assert mismatches == []


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("available_quantity", -1),
        ("available_quantity", "invalid"),
        ("available_quantity", "1000000000.000001"),
        ("unit_price", -1),
        ("unit_price", "NaN"),
        ("unit_price", "1e5000"),
        ("unit_price", "0.0000001"),
        ("pieces_per_box", 0),
        ("pieces_per_box", "invalid"),
        ("pieces_per_box", "1000000.000001"),
    ],
)
def test_load_offiho_catalog_rejects_corrupt_numeric_item_in_1206_index(tmp_path, field, value):
    from mobiliti_saas.quote_engine.offiho_catalog import load_offiho_catalog

    raw = _runtime_catalog_raw()
    raw["items"][0][field] = value
    path = tmp_path / "offiho-corrupt.json"
    path.write_text(json.dumps(raw), encoding="utf-8")

    with pytest.raises(ValueError, match=field):
        load_offiho_catalog(path)


def test_offiho_json_number_rejects_extreme_decimal_before_int_conversion():
    from mobiliti_saas.quote_engine.offiho_catalog import _json_number

    with pytest.raises(ValueError, match="fuera de rango"):
        _json_number(Decimal("1e5000"))


@pytest.mark.parametrize("field", ["inventory_key", "code", "unit", "price_source", "match_status"])
def test_offiho_catalog_item_rejects_blank_required_field(field):
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw[field] = " "

    with pytest.raises(ValueError, match=field):
        OffihoCatalogItem.from_dict(raw)


@pytest.mark.parametrize(
    ("field", "url"),
    [
        ("product_url", "http://www.offiho.com/productos/alufsen"),
        ("product_url", "https://example.com/productos/alufsen"),
        ("image_url", "https://econosillas.com/alufsen.jpg"),
    ],
)
def test_offiho_catalog_item_rejects_non_official_https_url(field, url):
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw[field] = url

    with pytest.raises(ValueError, match=field):
        OffihoCatalogItem.from_dict(raw)


def test_offiho_catalog_item_accepts_official_mexico_image_host():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["image_url"] = "https://offiho.com.mx/e2e-offiho.png"

    item = OffihoCatalogItem.from_dict(raw)

    assert item.image_url == "https://offiho.com.mx/e2e-offiho.png"


def test_offiho_catalog_accepts_only_production_pdf_asset_prefix():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["product_url"] = (
        "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/"
        "lp-black-colos-jul2026.pdf#page=15"
    )
    raw["image_url"] = (
        "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/images/vesper-1-15.jpg"
    )
    assert OffihoCatalogItem.from_dict(raw).image_url.endswith("vesper-1-15.jpg")

    raw["image_url"] = "https://web-lemon-one-45.vercel.app/arbitrary/user-content.jpg"
    with pytest.raises(ValueError, match="image_url"):
        OffihoCatalogItem.from_dict(raw)


def test_offiho_catalog_accepts_official_colos_product_and_image_urls():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["product_url"] = "https://colos.it/en/products/stecca-2"
    raw["image_url"] = "https://colos.it/storage/products/24/S_stecca-2_bluscuro.jpg"

    item = OffihoCatalogItem.from_dict(raw)

    assert item.product_url == raw["product_url"]
    assert item.image_url == raw["image_url"]


def test_offiho_catalog_accepts_canonical_sharepoint_product_url_only():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["product_url"] = (
        "https://mobiliti11-my.sharepoint.com/personal/joel_meza_mobiliti_mx/"
        "Documents/catalogo.pdf#page=42"
    )

    item = OffihoCatalogItem.from_dict(raw)

    assert item.product_url == raw["product_url"]

    raw["image_url"] = raw["product_url"]
    with pytest.raises(ValueError, match="image_url"):
        OffihoCatalogItem.from_dict(raw)


def test_offiho_catalog_rejects_sharepoint_lookalike_product_host():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["product_url"] = "https://mobiliti11-my.sharepoint.com.example.com/catalog.pdf"

    with pytest.raises(ValueError, match="product_url"):
        OffihoCatalogItem.from_dict(raw)


def test_offiho_catalog_rejects_colos_lookalike_host():
    from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem

    raw = _runtime_catalog_raw()["items"][0]
    raw["image_url"] = "https://colos.it.example.com/storage/product.jpg"

    with pytest.raises(ValueError, match="image_url"):
        OffihoCatalogItem.from_dict(raw)


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("source_hash", ""),
        ("source_hash", None),
        ("generated_at", " "),
        ("generated_at", None),
    ],
)
def test_load_offiho_catalog_rejects_blank_root_metadata(tmp_path, field, value):
    from mobiliti_saas.quote_engine.offiho_catalog import load_offiho_catalog

    raw = _runtime_catalog_raw()
    raw[field] = value
    path = tmp_path / "offiho-blank-root.json"
    path.write_text(json.dumps(raw), encoding="utf-8")

    with pytest.raises(ValueError, match=field):
        load_offiho_catalog(path)


class _FakePeerSocket:
    def __init__(self, address):
        self.address = address
        self.calls = 0

    def getpeername(self):
        self.calls += 1
        return (self.address, 443)


def _valid_png_bytes():
    stream = BytesIO()
    Image.new("RGB", (2, 2), "blue").save(stream, format="PNG")
    return stream.getvalue()


_VALID_PNG_BYTES = _valid_png_bytes()


class _FakeImageResponse:
    def __init__(self, *, peer_address="93.184.216.34", include_peer=True, payload=_VALID_PNG_BYTES):
        self.headers = {"content-type": "image/png", "content-length": str(len(payload))}
        self.payload = payload
        self.socket = _FakePeerSocket(peer_address)
        if include_peer:
            raw = type("Raw", (), {"_sock": self.socket})()
            self.fp = type("Fp", (), {"raw": raw})()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def read(self, size):
        return self.payload[:size]


def _install_fake_image_opener(monkeypatch, response):
    import mobiliti_saas.quote_engine.catalog_cart as catalog_cart

    class _FakeOpener:
        def open(self, request, timeout):
            assert request.full_url == "https://www.offiho.com/uploads/alufsen.png"
            assert timeout == 18
            return response

    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    monkeypatch.setattr(catalog_cart.urllib.request, "build_opener", lambda *handlers: _FakeOpener())
    return catalog_cart


def test_catalog_image_write_error_is_omitted(monkeypatch, tmp_path):
    response = _FakeImageResponse()
    catalog_cart = _install_fake_image_opener(monkeypatch, response)
    monkeypatch.setattr(Path, "write_bytes", lambda self, data: (_ for _ in ()).throw(OSError("disk full")))

    result = catalog_cart._download_catalog_image(
        "https://www.offiho.com/uploads/alufsen.png",
        tmp_path,
        "OHE-405",
        "offiho_cart",
    )

    assert result is None
    assert response.socket.calls == 1


def test_catalog_image_rejects_private_connected_peer(monkeypatch, tmp_path):
    response = _FakeImageResponse(peer_address="127.0.0.1")
    catalog_cart = _install_fake_image_opener(monkeypatch, response)

    result = catalog_cart._download_catalog_image(
        "https://www.offiho.com/uploads/alufsen.png",
        tmp_path,
        "OHE-405",
        "offiho_cart",
    )

    assert result is None
    assert response.socket.calls == 1
    assert not list(tmp_path.iterdir())


def test_catalog_image_accepts_public_connected_peer(monkeypatch, tmp_path):
    response = _FakeImageResponse(peer_address="93.184.216.34")
    catalog_cart = _install_fake_image_opener(monkeypatch, response)

    result = catalog_cart._download_catalog_image(
        "https://www.offiho.com/uploads/alufsen.png",
        tmp_path,
        "OHE-405",
        "offiho_cart",
    )

    assert response.socket.calls == 1
    assert result is not None
    assert result.read_bytes() == _VALID_PNG_BYTES


def test_catalog_image_rejects_uninspectable_connected_peer(monkeypatch, tmp_path):
    response = _FakeImageResponse(include_peer=False)
    catalog_cart = _install_fake_image_opener(monkeypatch, response)

    result = catalog_cart._download_catalog_image(
        "https://www.offiho.com/uploads/alufsen.png",
        tmp_path,
        "OHE-405",
        "offiho_cart",
    )

    assert result is None
    assert not list(tmp_path.iterdir())


def test_catalog_image_redirect_validates_private_and_public_peers(monkeypatch):
    import mobiliti_saas.quote_engine.catalog_cart as catalog_cart

    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    handler = catalog_cart._OfficialRedirectHandler(catalog_cart.OFFICIAL_IMAGE_HOSTS["offiho_cart"])
    request = catalog_cart.urllib.request.Request("https://www.offiho.com/start")
    private_response = _FakeImageResponse(peer_address="10.0.0.8")

    with pytest.raises(ValueError, match="publica"):
        handler.redirect_request(
            request,
            private_response,
            302,
            "Found",
            {},
            "https://www.offiho.com/uploads/alufsen.png",
        )

    public_response = _FakeImageResponse(peer_address="93.184.216.34")
    redirected = handler.redirect_request(
        request,
        public_response,
        302,
        "Found",
        {},
        "https://www.offiho.com/uploads/alufsen.png",
    )
    assert redirected.full_url == "https://www.offiho.com/uploads/alufsen.png"
    assert private_response.socket.calls == 1
    assert public_response.socket.calls == 1


def test_catalog_image_disables_environment_proxies(monkeypatch, tmp_path):
    import mobiliti_saas.quote_engine.catalog_cart as catalog_cart

    response = _FakeImageResponse(peer_address="93.184.216.34")
    captured_handlers = []

    class _FakeOpener:
        def open(self, request, timeout):
            return response

    def fake_build_opener(*handlers):
        captured_handlers.extend(handlers)
        return _FakeOpener()

    monkeypatch.setenv("HTTPS_PROXY", "http://127.0.0.1:9999")
    monkeypatch.setenv("HTTP_PROXY", "http://127.0.0.1:9999")
    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    monkeypatch.setattr(catalog_cart.urllib.request, "build_opener", fake_build_opener)

    result = catalog_cart._download_catalog_image(
        "https://www.offiho.com/uploads/alufsen.png",
        tmp_path,
        "OHE-405",
        "offiho_cart",
    )

    proxy_handlers = [
        handler
        for handler in captured_handlers
        if isinstance(handler, catalog_cart.urllib.request.ProxyHandler)
    ]
    assert result is not None
    assert len(proxy_handlers) == 1
    assert proxy_handlers[0].proxies == {}


def test_offiho_workbook_writes_price_and_warning(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 1}],
        catalog=fake_runtime_catalog(available_quantity=0, unit_price=7999),
    )
    output = create_offiho_quotation_workbook(payload, tmp_path / "offiho.xlsx")
    wb = load_workbook(output)
    ws = wb["Quotation"]

    assert ws["J9"].value == 7999
    assert "ADVERTENCIA: PRODUCTO AGOTADO" in ws["D9"].value
    assert ws["D9"].fill.fgColor.rgb.endswith("FFF2CC")
    wb.close()


def test_offiho_missing_price_warns_in_temporary_and_final_workbooks(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )
    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 1}],
        catalog=fake_runtime_catalog(
            available_quantity=2,
            unit_price=0,
            price_source="missing",
        ),
    )
    assert payload["items"][0]["price_source"] == "missing"
    source = create_offiho_quotation_workbook(payload, tmp_path / "offiho-price-pending.xlsx")
    temporary = load_workbook(source)
    assert temporary["Quotation"]["J9"].value == 0
    assert "ADVERTENCIA: PRECIO POR CONFIRMAR" in temporary["Quotation"]["D9"].value
    assert temporary["Quotation"]["D9"].fill.fgColor.rgb.endswith("FFF2CC")
    temporary.close()

    output = tmp_path / "cotizacion-price-pending.xlsx"
    generate_online_quote(source, output, {"tipo_cambio": "20"})
    final = load_workbook(output, data_only=False)
    assert any(
        "ADVERTENCIA: PRECIO POR CONFIRMAR" in str(cell.value or "")
        for row in final["Cotizacion"].iter_rows()
        for cell in row
    )
    assert any(
        cell.value == 0
        for row in final["Quotation"].iter_rows()
        for cell in row
    )
    final.close()


def test_offiho_warning_survives_online_quote_generation(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )
    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 1}],
        catalog=fake_runtime_catalog(available_quantity=0, unit_price=7999),
    )
    source = create_offiho_quotation_workbook(payload, tmp_path / "offiho.xlsx")
    output = tmp_path / "cotizacion.xlsx"

    generate_online_quote(source, output, {"tipo_cambio": "20"})

    wb = load_workbook(output, data_only=False)
    assert {"Cotizacion", "Mobiliti", "Quotation"}.issubset(wb.sheetnames)
    assert any(
        "ADVERTENCIA: PRODUCTO AGOTADO" in str(cell.value or "")
        for row in wb["Cotizacion"].iter_rows()
        for cell in row
    )
    wb.close()


def test_offiho_insufficient_warning_preserves_available_quantity_in_final_quote(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )
    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 3}],
        catalog=fake_runtime_catalog(available_quantity=2, unit_price=7999),
    )
    source = create_offiho_quotation_workbook(payload, tmp_path / "offiho-insufficient.xlsx")
    output = tmp_path / "cotizacion-insufficient.xlsx"

    generate_online_quote(source, output, {"tipo_cambio": "20"})

    wb = load_workbook(output, data_only=False)
    warning_cells = [
        str(cell.value or "")
        for row in wb["Cotizacion"].iter_rows()
        for cell in row
        if "ADVERTENCIA: EXISTENCIA INSUFICIENTE" in str(cell.value or "").upper()
    ]
    assert warning_cells
    assert "DISPONIBLE 2" in warning_cells[0].upper()
    wb.close()


def test_final_cotizacion_fills_only_warning_descriptions_yellow(tmp_path):
    from mobiliti_saas.quote_engine.offiho_catalog import (
        OffihoCatalogItem,
        build_offiho_cart_payload,
        create_offiho_quotation_workbook,
    )
    from mobiliti_saas.worker.online_quote_generator import generate_online_quote

    definitions = [
        ("OHE-501", "AGOTADO", 0, 7999, "inventory"),
        ("OHE-502", "PRECIO PENDIENTE", 2, 0, "missing"),
        ("OHE-503", "AGOTADO SIN PRECIO", 0, 0, "missing"),
        ("OHE-504", "NORMAL", 2, 7999, "inventory"),
    ]
    items = [
        OffihoCatalogItem(
            inventory_key=f"{code} NEGRO {name}",
            code=code,
            name=name,
            variant="NEGRO",
            unit="PZA",
            pieces_per_box=Decimal("1"),
            available_quantity=Decimal(str(stock)),
            unit_price=Decimal(str(price)),
            price_source=price_source,
        )
        for code, name, stock, price, price_source in definitions
    ]
    catalog = {
        "source_hash": "hash",
        "items": items,
        "by_inventory_key": {item.inventory_key: item for item in items},
    }
    payload = build_offiho_cart_payload(
        [{"inventory_key": item.inventory_key, "quantity": 1} for item in items],
        catalog=catalog,
    )
    source = create_offiho_quotation_workbook(payload, tmp_path / "offiho-all-warnings.xlsx")
    output = tmp_path / "cotizacion-all-warnings.xlsx"

    generate_online_quote(source, output, {"tipo_cambio": "20"})

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Cotizacion"]
    warning_cells = [cell for cell in sheet["C"] if "ADVERTENCIA:" in str(cell.value or "").upper()]
    assert len(warning_cells) == 3
    for code in ("OHE-501", "OHE-502", "OHE-503"):
        cell = next(cell for cell in warning_cells if code in str(cell.value or ""))
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb.endswith("FFF2CC")
    both = next(cell for cell in warning_cells if "OHE-503" in str(cell.value or ""))
    assert "ADVERTENCIA: PRODUCTO AGOTADO" in both.value
    assert "ADVERTENCIA: PRECIO POR CONFIRMAR" in both.value

    normal = next(cell for cell in sheet["C"] if "OHE-504" in str(cell.value or ""))
    assert not (
        normal.fill.fill_type == "solid"
        and str(normal.fill.fgColor.rgb or "").endswith("FFF2CC")
    )
    category = next(cell for cell in sheet["A"] if cell.value == "=Quotation!A8")
    assert not (
        category.fill.fill_type == "solid"
        and str(category.fill.fgColor.rgb or "").endswith("FFF2CC")
    )
    workbook.close()


def test_catalog_workbook_closes_when_save_fails(monkeypatch, tmp_path):
    import mobiliti_saas.quote_engine.catalog_cart as catalog_cart
    from mobiliti_saas.quote_engine.offiho_catalog import build_offiho_cart_payload

    payload = build_offiho_cart_payload(
        [{"inventory_key": "OHE-405 NEGRO ALUFSEN", "quantity": 1}],
        catalog=fake_runtime_catalog(available_quantity=2, unit_price=7999),
    )
    workbook = catalog_cart.Workbook()
    original_close = workbook.close
    close_calls = []

    def fail_save(path):
        raise OSError("save failed")

    def record_close():
        close_calls.append(True)
        original_close()

    monkeypatch.setattr(catalog_cart, "Workbook", lambda: workbook)
    monkeypatch.setattr(workbook, "save", fail_save)
    monkeypatch.setattr(workbook, "close", record_close)

    with pytest.raises(OSError, match="save failed"):
        catalog_cart.create_catalog_quotation_workbook(
            payload,
            tmp_path / "save-failure.xlsx",
            source_type="offiho_cart",
            category_label="Offiho",
            image_dir=tmp_path / "images",
        )

    assert close_calls == [True]


class _Sheet:
    nrows = 7

    def cell_value(self, row, column):
        values = {
            (5, 1): "OHE-405 NEGRO ALUFSEN",
            (5, 2): 252,
            (5, 3): 1,
            (5, 4): 7999,
            (6, 1): "OHE-406 GRIS ALUFSEN",
            (6, 2): 0,
            (6, 3): 1,
            (6, 4): 8100,
        }
        return values.get((row, column), "")


class _Workbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _Sheet()


class _DuplicateSheet(_Sheet):
    nrows = 8

    def cell_value(self, row, column):
        if row == 7:
            row = 5
        return super().cell_value(row, column)


class _DuplicateWorkbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _DuplicateSheet()


class _ConflictingSheet(_DuplicateSheet):
    def cell_value(self, row, column):
        if row == 7 and column == 2:
            return 251
        return super().cell_value(row, column)


class _ConflictingWorkbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _ConflictingSheet()


class _CorruptNumericSheet(_Sheet):
    def __init__(self, column, value):
        self.column = column
        self.value = value

    def cell_value(self, row, column):
        if row == 5 and column == self.column:
            return self.value
        return super().cell_value(row, column)


class _CorruptNumericWorkbook:
    def __init__(self, column, value):
        self.column = column
        self.value = value

    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _CorruptNumericSheet(self.column, self.value)


class _SpecialOrderWorkbook:
    def __init__(self, status):
        self.status = status

    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _CorruptNumericSheet(2, self.status)


class _SpecialOrderPiecesWorkbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _CorruptNumericSheet(3, "Sobre Pedido")


class _BlankStockWorkbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _CorruptNumericSheet(2, "")


class _RepeatedHeaderSheet(_Sheet):
    nrows = 8

    def cell_value(self, row, column):
        if row == 7:
            return {
                1: "CODIGO",
                2: "EXISTENCIA",
                3: "Piezas por Caja",
                4: "Precio",
            }.get(column, "")
        return super().cell_value(row, column)


class _RepeatedHeaderWorkbook:
    def sheet_by_name(self, name):
        assert name == "Publicaci\u00f3n"
        return _RepeatedHeaderSheet()


@pytest.fixture
def html_inventory_xls(tmp_path):
    path = tmp_path / "offiho-html.xls"
    path.write_text(
        """<!DOCTYPE html>
<html><head><meta charset="utf-8"></head><body>
<table>
  <tr><th>CODIGO</th><th>Existencia</th><th>Piezas por Caja</th><th>Precio Lista 1</th></tr>
  <tr><td>OHE-405 NEGRO ALUFSEN</td><td>252</td><td>1</td><td>7,999</td></tr>
  <tr><td>OHE-406 GRIS ALUFSEN</td><td>0</td><td>2</td><td></td></tr>
</table>
</body></html>""",
        encoding="utf-8",
    )
    return path


def test_parse_inventory_keeps_available_and_exhausted_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _Workbook())

    rows = parse_inventory_xls(tmp_path / "offiho-small.xls")

    assert rows[0]["inventory_key"] == "OHE-405 NEGRO ALUFSEN"
    assert rows[0]["available_quantity"] == Decimal("252")
    assert rows[0]["unit_price"] == Decimal("7999")
    assert any(row["available_quantity"] == 0 for row in rows)


def test_parse_inventory_removes_materially_identical_duplicate_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _DuplicateWorkbook())

    rows = parse_inventory_xls(tmp_path / "offiho-duplicates.xls")

    assert len(rows) == 2
    assert [row["inventory_key"] for row in rows] == [
        "OHE-405 NEGRO ALUFSEN",
        "OHE-406 GRIS ALUFSEN",
    ]


def test_parse_inventory_rejects_duplicate_key_with_different_data(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _ConflictingWorkbook())

    with pytest.raises(RuntimeError, match="OHE-405 NEGRO ALUFSEN.*datos distintos"):
        parse_inventory_xls(tmp_path / "offiho-conflict.xls")


def test_build_catalog_reports_deduplicated_inventory_audit(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _DuplicateWorkbook())
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {})
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})
    inventory = tmp_path / "offiho-duplicates.xls"
    inventory.write_bytes(b"xls")

    catalog = build.build_catalog(
        inventory,
        [],
        tmp_path / "cache.json",
        tmp_path / "catalog.json",
    )

    assert catalog["source_row_count"] == 3
    assert catalog["duplicate_row_count"] == 1
    assert catalog["unique_item_count"] == 2
    assert catalog["total"] == 2
    assert catalog["out_of_stock"] == 1
    assert catalog["inventory_prices"] == 2


def test_build_catalog_uses_exact_spec_image_when_site_variant_is_missing(monkeypatch, tmp_path):
    inventory = tmp_path / "inventory.xls"
    inventory.write_bytes(b"inventory")
    spec_guide = tmp_path / "spec.xlsx"
    spec_guide.write_bytes(b"spec")
    item = {
        "inventory_key": "CABECERA OHE-112 GRIS",
        "code": "OHE-112",
        "name": "CABECERA",
        "variant": "GRIS",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 2,
        "unit_price": 100,
        "price_source": "inventory",
    }
    monkeypatch.setattr(
        build,
        "_parse_inventory_xls",
        lambda path: ([dict(item)], {"source_row_count": 1, "duplicate_row_count": 0, "unique_item_count": 1}),
    )
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {})
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})
    monkeypatch.setattr(
        build,
        "extract_offiho_spec_images",
        lambda paths, items, **kwargs: {
            item["inventory_key"]: {
                "image_url": "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/spec-images/cabecera.png",
                "description": "Cabecera Crew gris de la guia oficial.",
                "match_status": "spec_guide_exact",
                "image_sha256": "a" * 64,
            }
        },
    )

    catalog = build.build_catalog(
        inventory,
        [],
        tmp_path / "cache.json",
        tmp_path / "catalog.json",
        spec_guide_paths=[spec_guide],
        colos_exact_manifest_path=None,
        offiho_exact_manifest_path=None,
        catalog_exact_crop_manifest_paths=[],
        spec_visual_exact_manifest_paths=[],
    )

    enriched = catalog["items"][0]
    assert enriched["image_url"].endswith("/spec-images/cabecera.png")
    assert enriched["description"] == "Cabecera Crew gris de la guia oficial."
    assert enriched["description_source"] == "spec_guide"
    assert enriched["match_status"] == "spec_guide_exact"
    assert catalog["sources"]["spec_guides"][0]["sha256"]
    assert catalog["sources"]["spec_image_index"]["record_count"] == 1


@pytest.mark.parametrize(
    ("fixture_name", "allowed_hosts", "expected_count"),
    [
        ("offiho_exact_variant_images.json", build.OFFIHO_HOSTS, 50),
        ("colos_exact_images.json", build.COLOS_HOSTS, 175),
    ],
)
def test_checked_in_exact_image_manifests_are_strict_and_unique(
    fixture_name,
    allowed_hosts,
    expected_count,
):
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho") / fixture_name,
        allowed_hosts=allowed_hosts,
        match_status="exact",
    )

    assert len(index) == expected_count
    assert source["record_count"] == expected_count
    assert len(source["sha256"]) == 64
    assert all(item["image_url"] and item["url"] for item in index.values())


def test_checked_in_catalog_crop_manifest_separates_source_and_asset_hosts():
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho/colos_pdf_exact_images.json"),
        allowed_hosts=build.COLOS_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_catalog_exact_crop",
    )

    assert len(index) == 8
    assert source["record_count"] == 8
    assert all(
        item["url"].startswith("https://colos.it/")
        and item["image_url"].startswith(
            "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/catalog-crops/"
        )
        for item in index.values()
    )


def test_checked_in_spec_visual_manifest_keeps_only_independently_reviewed_exacts():
    index, source = build.load_exact_image_manifest(
        Path(
            "catalog_sources/offiho/"
            "offiho_spec_visual_independent_exact_images.json"
        ),
        allowed_hosts=build.SHAREPOINT_CATALOG_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="spec_guide_visual_exact",
    )

    assert len(index) == 78
    assert source["record_count"] == 78
    assert len(source["sha256"]) == 64
    assert all(
        item["url"].startswith("https://mobiliti11-my.sharepoint.com/")
        and "/catalog-assets/offiho/spec-visual-exact/" in item["image_url"]
        for item in index.values()
    )
    assert "OHM-60 TRAVERTINO MARBEL" not in index
    assert "OHM-70 TRAVERTINO MARBEL" not in index
    assert "TA/1Q-600 N NEGRO" not in index
    assert "TA/1Q-700 N NEGRA" not in index


def test_checked_in_spec_auto_audited_manifest_contains_only_visual_exacts():
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho/offiho_spec_auto_audited_exact_images.json"),
        allowed_hosts=build.SHAREPOINT_CATALOG_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="spec_guide_visual_exact",
    )

    assert len(index) == 45
    assert source["record_count"] == 45
    assert all(
        item["url"].startswith("https://mobiliti11-my.sharepoint.com/")
        and "/catalog-assets/offiho/spec-auto-audited-exact/" in item["image_url"]
        for item in index.values()
    )
    assert "OHV-80 ALTO GRIS SHINE *" not in index
    assert "VESPER 3B" not in index


def test_checked_in_official_web_visual_manifest_keeps_shopify_configurations_separate():
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho/offiho_official_web_visual_exact_images.json"),
        allowed_hosts=build.OFFICIAL_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_web_visual_exact",
    )

    assert len(index) == 45
    assert source["record_count"] == 45
    assert index["OHV-128 GRIS FABRIZIA *"]["url"].endswith(
        "/products/fabrizia-ohv-128"
    )
    assert index["OHV-128 LOUNGE GRIS FABRIZIA"]["url"].endswith(
        "/products/fabrizia-ohv-128-lounge"
    )
    assert (
        index["OHV-128 GRIS FABRIZIA *"]["image_url"]
        != index["OHV-128 LOUNGE GRIS FABRIZIA"]["image_url"]
    )


def test_checked_in_live_visual_manifest_contains_only_reviewed_offiho_variants():
    manifest_path = Path(
        "catalog_sources/offiho/offiho_live_visual_exact_images.json"
    )
    index, source = build.load_exact_image_manifest(
        manifest_path,
        allowed_hosts=build.OFFICIAL_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_web_visual_exact",
    )

    assert len(index) == 66
    assert source["record_count"] == 66
    assert all(
        item["url"].startswith("https://www.offiho.com/")
        and "/catalog-assets/offiho/live-visual-exact/" in item["image_url"]
        for item in index.values()
    )
    assert index["OHV-340CR BLANCO KYOS"]["url"].endswith(
        "/visitantes-interior-kyoscollection-modelo-OHT-340cr"
    )
    raw_items = json.loads(manifest_path.read_text(encoding="utf-8"))["items"]
    for item in raw_items:
        filename = Path(urlsplit(item["image_url"]).path).name
        asset = Path(
            "mobiliti_saas/web/public/catalog-assets/offiho/live-visual-exact"
        ) / filename
        assert asset.is_file()
        digest = hashlib.sha256(asset.read_bytes()).hexdigest()
        assert digest == item["source_sha256"] == asset.stem


def test_default_official_web_visual_manifests_are_plural_and_unique():
    combined = {}
    sources = []
    assert len(build.DEFAULT_OFFICIAL_WEB_VISUAL_EXACT_MANIFEST_PATHS) == 4

    for path in build.DEFAULT_OFFICIAL_WEB_VISUAL_EXACT_MANIFEST_PATHS:
        index, source = build.load_exact_image_manifest(
            path,
            allowed_hosts=build.OFFICIAL_HOSTS,
            allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
            match_status="official_web_visual_exact",
        )
        assert not set(combined).intersection(index)
        combined.update(index)
        sources.append(source)

    assert len(combined) == 166
    assert sorted(source["record_count"] for source in sources) == [10, 45, 45, 66]


def test_checked_in_residual_visual_manifest_contains_only_independent_exacts():
    manifest_path = Path(
        "catalog_sources/offiho/offiho_residual_visual_exact_images.json"
    )
    index, source = build.load_exact_image_manifest(
        manifest_path,
        allowed_hosts=build.OFFICIAL_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_web_visual_exact",
    )

    assert len(index) == source["record_count"] == 45
    assert "CABECERA OHE-44 ARENA" not in index
    assert "CABECERA OHE-44 NEGRO" not in index
    assert "STECCA 8 S MOSTAZA" not in index
    raw_items = json.loads(manifest_path.read_text(encoding="utf-8"))["items"]
    for item in raw_items:
        filename = Path(urlsplit(item["image_url"]).path).name
        asset = Path(
            "mobiliti_saas/web/public/catalog-assets/offiho/residual-visual-exact"
        ) / filename
        assert asset.is_file()
        digest = hashlib.sha256(asset.read_bytes()).hexdigest()
        assert digest == item["source_sha256"] == asset.stem


def test_build_catalog_prefers_visual_web_evidence_over_overlapping_spec(
    monkeypatch,
    tmp_path,
):
    inventory = tmp_path / "inventory.xls"
    inventory.write_bytes(b"inventory")
    item = {
        "inventory_key": "OHP-341 GRIS KYOS",
        "code": "OHP-341",
        "name": "KYOS",
        "variant": "GRIS",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 2,
        "unit_price": 100,
        "price_source": "inventory",
    }
    audit = {
        "source_row_count": 1,
        "duplicate_row_count": 0,
        "unique_item_count": 1,
    }
    monkeypatch.setattr(build, "_parse_inventory_xls", lambda path: ([dict(item)], audit))
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(
        build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {}
    )
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})

    catalog = build.build_catalog(
        inventory,
        [],
        tmp_path / "cache.json",
        tmp_path / "catalog.json",
        colos_exact_manifest_path=None,
        offiho_exact_manifest_path=None,
        catalog_exact_crop_manifest_paths=[],
    )

    enriched = catalog["items"][0]
    assert enriched["match_status"] == "official_web_visual_exact"
    assert enriched["product_url"].startswith("https://www.offiho.com/")
    assert "/live-visual-exact/" in enriched["image_url"]
    assert len(catalog["sources"]["official_web_visual_exact_images"]) == 4


def test_checked_in_offiho_catalog_crop_manifest_remains_reproducible():
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho/offiho_catalog_exact_crops.json"),
        allowed_hosts=build.CATALOG_SOURCE_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_catalog_exact_crop",
    )

    assert len(index) == 106
    assert source["record_count"] == 106


def test_checked_in_internet_exact_manifest_contains_verified_rescues():
    index, source = build.load_exact_image_manifest(
        Path("catalog_sources/offiho/offiho_internet_exact_images.json"),
        allowed_hosts=build.CATALOG_SOURCE_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_catalog_exact_crop",
    )

    assert len(index) == 4
    assert source["record_count"] == 4
    assert {
        "PISTÓNCAJERO",
        "G03C PISTON CROMADO REBASADOR",
        "G03C CROMADO REBASADOR",
        "R34 ALUMINIO LUCIDO",
    }.issubset(index)


def test_catalog_crop_manifest_accepts_canonical_sharepoint_source(monkeypatch, tmp_path):
    manifest = tmp_path / "sharepoint-crop.json"
    manifest.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "items": [
                    {
                        "inventory_key": "ARO CROMADO CROMADO",
                        "product_url": (
                            "https://mobiliti11-my.sharepoint.com/personal/joel_meza_mobiliti_mx/"
                            "Documents/catalogo.pdf"
                        ),
                        "image_url": (
                            "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/"
                            "catalog-crops/offiho/aro.png"
                        ),
                        "evidence_as_of": "2026-08-12",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    index, _ = build.load_exact_image_manifest(
        manifest,
        allowed_hosts=build.CATALOG_SOURCE_HOSTS,
        allowed_image_hosts=build.MANAGED_ASSET_HOSTS,
        match_status="official_catalog_exact_crop",
    )

    assert index["ARO CROMADO CROMADO"]["url"].startswith(
        "https://mobiliti11-my.sharepoint.com/"
    )


def test_build_catalog_rejects_audited_wrong_managed_catalog_crop(monkeypatch, tmp_path):
    inventory = tmp_path / "inventory.xls"
    inventory.write_bytes(b"inventory")
    item = {
        "inventory_key": "VESPER/03W BLANCA",
        "code": "VESPER/03W",
        "name": "",
        "variant": "BLANCA",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 2,
        "unit_price": 100,
        "price_source": "inventory",
    }
    audit = {
        "source_row_count": 1,
        "duplicate_row_count": 0,
        "unique_item_count": 1,
    }
    monkeypatch.setattr(build, "_parse_inventory_xls", lambda path: ([dict(item)], audit))
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(
        build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {}
    )
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})

    catalog = build.build_catalog(
        inventory,
        [],
        tmp_path / "cache.json",
        tmp_path / "catalog.json",
        colos_exact_manifest_path=None,
        offiho_exact_manifest_path=None,
        catalog_exact_crop_manifest_paths=[
            Path("catalog_sources/offiho/colos_pdf_exact_images.json")
        ],
        generated_image_manifest_path=None,
    )

    enriched = catalog["items"][0]
    assert enriched["match_status"] == "visual_conflict_rejected"
    assert enriched["image_url"] == ""
    assert catalog["sources"]["catalog_exact_crops"][0]["record_count"] == 8
    assert catalog["sources"]["visual_rejections"]["record_count"] == 7


def test_visual_rejection_clears_only_the_audited_wrong_image():
    wrong_url = (
        "https://www.offiho.com/visitantes-interior/innova/"
        "OHV-3000/zoom/OHV-3000Blanco.jpg"
    )
    corrected_url = (
        "https://www.offiho.com/visitantes-interior/innova/"
        "OHV-3000/zoom/OHV-3000-respaldo-blanco.jpg"
    )
    rejections = {
        "OHV-3000 W BLANCO INNOVA": {
            "image_urls": {wrong_url},
            "reason": "El render cambia el asiento, no el respaldo.",
        }
    }
    items = [
        {
            "inventory_key": "OHV-3000 W BLANCO INNOVA",
            "image_url": wrong_url,
            "match_status": "official_code_match",
        },
        {
            "inventory_key": "OHV-3000 W BLANCO INNOVA",
            "image_url": corrected_url,
            "match_status": "official_web_visual_exact",
        },
    ]

    build.apply_visual_rejections(items, rejections)

    assert items[0]["image_url"] == ""
    assert items[0]["match_status"] == "visual_conflict_rejected"
    assert items[1]["image_url"] == corrected_url
    assert items[1]["match_status"] == "official_web_visual_exact"


@pytest.mark.parametrize(
    ("rejected_host", "catalog_host"),
    [
        ("www.offiho.com", "offiho.com"),
        ("offiho.com", "www.offiho.com"),
        ("www.offihoblack.com", "offihoblack.com"),
        ("offihoblack.com", "www.offihoblack.com"),
        ("www.colos.it", "colos.it"),
        ("colos.it", "www.colos.it"),
    ],
)
def test_visual_rejection_treats_offiho_host_aliases_as_the_same_resource(
    rejected_host, catalog_host
):
    path = "/visitantes-interior/innova/OHV-3000/zoom/OHV-3000Blanco.jpg"
    items = [
        {
            "inventory_key": "OHV-3000 W BLANCO INNOVA",
            "image_url": f"https://{catalog_host}{path}",
            "match_status": "official_code_match",
        }
    ]
    rejections = {
        "OHV-3000 W BLANCO INNOVA": {
            "image_urls": {f"https://{rejected_host}{path}"},
            "reason": "El respaldo visible sigue siendo negro.",
        }
    }

    build.apply_visual_rejections(items, rejections)

    assert items[0]["image_url"] == ""
    assert items[0]["match_status"] == "visual_conflict_rejected"


def test_rejected_higher_priority_candidate_does_not_hide_correct_lower_source():
    inventory_key = "OHV-3000 W BLANCO INNOVA"
    wrong_url = "https://offiho.com/OHV-3000/zoom/OHV-3000Blanco.jpg"
    correct_url = "https://offiho.com/OHV-3000/zoom/OHV-3000RespaldoBlanco.jpg"
    rejections = {
        inventory_key: {
            "image_urls": {wrong_url},
            "reason": "El respaldo visible sigue siendo negro.",
        }
    }

    selected = build._first_non_rejected_product(
        inventory_key,
        [
            {"image_url": wrong_url, "match_status": "official_web_visual_exact"},
            {"image_url": correct_url, "match_status": "spec_guide_visual_exact"},
        ],
        rejections,
    )

    assert selected["image_url"] == correct_url
    assert selected["match_status"] == "spec_guide_visual_exact"


def test_visual_rejection_manifest_is_strict_and_url_specific(tmp_path):
    path = tmp_path / "visual-rejections.json"
    wrong_url = (
        "https://www.offiho.com/visitantes-interior/innova/"
        "OHV-3000/zoom/OHV-3000Blanco.jpg"
    )
    path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "items": [
                    {
                        "inventory_key": "OHV-3000 W BLANCO INNOVA",
                        "rejected_image_url": wrong_url,
                        "evidence_as_of": "2026-08-12",
                        "review": "visual-v46-provisional #0423",
                        "reason": "El respaldo visible sigue siendo negro.",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    index, source = build.load_visual_rejection_manifest(path)

    assert index["OHV-3000 W BLANCO INNOVA"]["image_urls"] == {
        wrong_url.replace("www.offiho.com", "offiho.com")
    }
    assert source["record_count"] == 1
    assert len(source["sha256"]) == 64

    with pytest.raises(RuntimeError, match="inseguro o duplicado"):
        build.load_visual_rejection_manifest(
            path,
            inventory_keys={"OTRA CLAVE"},
        )

    unsafe = json.loads(path.read_text(encoding="utf-8"))
    unsafe["items"][0]["rejected_image_url"] = "https://offiho.com:bad/x.jpg"
    path.write_text(json.dumps(unsafe), encoding="utf-8")
    with pytest.raises(RuntimeError, match="inseguro o duplicado"):
        build.load_visual_rejection_manifest(path)


def test_visual_rejection_filters_wrong_site_candidate_before_ranking():
    inventory_key = "OHE-705 GRIS AIKO"
    wrong_url = "https://offiho.com/ejecutivos/aiko/OHE-705gris/wrong.jpg"
    correct_url = "https://www.offiho.com/ejecutivos/aiko/OHE-705gris/correct.jpg"
    base = {
        "codes": ["OHE-705GRIS"],
        "names": ["AIKO"],
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    candidates = [
        {**base, "url": "https://offiho.com/ejecutivos/aiko/modelo-OHE-705gris", "image_url": wrong_url},
        {**base, "url": "https://www.offiho.com/ejecutivos/aiko/modelo-OHE-705gris", "image_url": correct_url},
    ]
    rejections = {
        inventory_key: {"image_urls": {wrong_url}, "reason": "Imagen equivocada."}
    }
    sanitized = [
        build._without_rejected_candidate_images(inventory_key, row, rejections)
        for row in candidates
    ]

    product = build.match_official_product(
        build.extract_offiho_identity(inventory_key), sanitized
    )

    assert product["image_url"] == correct_url


def test_pdf_support_skips_rejected_image_and_uses_correct_brochure():
    inventory_key = "OHE-705 GRIS AIKO"
    identity = build.extract_offiho_identity(inventory_key)
    wrong_url = "https://offiho.com/catalog/wrong-gris.jpg"
    correct_url = "https://offiho.com/catalog/correct-gris.jpg"
    pdf_product = {
        "matched_title": "OHE-705 GRIS AIKO",
        "image_url": wrong_url,
        "product_url": "https://offiho.com/catalog/lista.pdf#page=1",
    }
    brochure_product = {
        "matched_title": "OHE-705 GRIS AIKO",
        "image_url": correct_url,
        "product_url": "https://offiho.com/catalog/catalogo.pdf#page=2",
    }
    rejections = {
        inventory_key: {"image_urls": {wrong_url}, "reason": "Imagen equivocada."}
    }

    selected = build._support_product_for_identity(
        identity,
        pdf_product,
        brochure_product,
        inventory_key=inventory_key,
        visual_rejections=rejections,
    )

    assert selected == brochure_product


def test_build_catalog_uses_separate_spec_visual_exact_manifest(monkeypatch, tmp_path):
    inventory = tmp_path / "inventory.xls"
    inventory.write_bytes(b"inventory")
    spec_manifest = tmp_path / "spec-visual-exact.json"
    spec_manifest.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "items": [
                    {
                        "inventory_key": "OHE-100 BLANCO WHALE",
                        "product_url": (
                            "https://mobiliti11-my.sharepoint.com/personal/joel_meza_mobiliti_mx/"
                            "Documents/spec-guide.xlsx"
                        ),
                        "image_url": (
                            "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/"
                            "spec-visual-exact/ohe-100-blanco-whale.png"
                        ),
                        "evidence_as_of": "2026-08-12",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    item = {
        "inventory_key": "OHE-100 BLANCO WHALE",
        "code": "OHE-100",
        "name": "WHALE",
        "variant": "BLANCO",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 2,
        "unit_price": 100,
        "price_source": "inventory",
    }
    audit = {
        "source_row_count": 1,
        "duplicate_row_count": 0,
        "unique_item_count": 1,
    }
    monkeypatch.setattr(build, "_parse_inventory_xls", lambda path: ([dict(item)], audit))
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(
        build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {}
    )
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})

    catalog = build.build_catalog(
        inventory,
        [],
        tmp_path / "cache.json",
        tmp_path / "catalog.json",
        colos_exact_manifest_path=None,
        offiho_exact_manifest_path=None,
        catalog_exact_crop_manifest_paths=[],
        spec_visual_exact_manifest_paths=[spec_manifest],
    )

    enriched = catalog["items"][0]
    assert enriched["match_status"] == "spec_guide_visual_exact"
    assert enriched["product_url"].startswith("https://mobiliti11-my.sharepoint.com/")
    assert enriched["image_url"].endswith(
        "/spec-visual-exact/ohe-100-blanco-whale.png"
    )
    assert catalog["sources"]["spec_visual_exact_images"][0]["record_count"] == 1


def test_cross_model_pdf_support_image_is_removed_but_equivalent_identity_is_kept():
    base_url = "https://example.test/catalog-assets/offiho"
    shared_kyos = f"{base_url}/images/negro-kyos.jpg"
    shared_vesper = f"{base_url}/images/vesper-103.jpg"
    items = [
        {"code": "OHV-348", "name": "KYOS", "image_url": shared_kyos},
        {"code": "OHV-318CR", "name": "KYOS", "image_url": shared_kyos},
        {"code": "VESPER", "name": "103", "image_url": shared_vesper},
        {"code": "VESPER/103", "name": "", "image_url": shared_vesper},
        {
            "code": "OHV-2700",
            "name": "INNOVA",
            "image_url": "https://www.offiho.com/OHV-2700/colores/negro.jpg",
        },
    ]

    build._clear_cross_model_support_images(items, base_url)

    assert items[0]["image_url"] == ""
    assert items[1]["image_url"] == ""
    assert items[2]["image_url"] == shared_vesper
    assert items[3]["image_url"] == shared_vesper
    assert items[4]["image_url"].startswith("https://www.offiho.com/")


def test_parse_inventory_supports_html_disguised_as_xls(html_inventory_xls):
    rows = parse_inventory_xls(html_inventory_xls)

    assert len(rows) == 2
    assert rows[0]["inventory_key"] == "OHE-405 NEGRO ALUFSEN"
    assert rows[0]["available_quantity"] == 252
    assert rows[0]["unit_price"] == 7999
    assert rows[1]["available_quantity"] == 0
    assert rows[1]["price_source"] == "missing"


@pytest.mark.parametrize(
    ("column", "column_name", "field"),
    [
        (2, "C", "Existencia"),
        (3, "D", "Piezas por Caja"),
        (4, "E", "Precio Lista 1"),
    ],
)
def test_parse_inventory_rejects_corrupt_numeric_values_with_location(
    monkeypatch, tmp_path, column, column_name, field
):
    monkeypatch.setattr(
        build.xlrd,
        "open_workbook",
        lambda path: _CorruptNumericWorkbook(column, "corrupto"),
    )

    with pytest.raises(RuntimeError, match=rf"Fila 6.*columna {column_name}.*{field}"):
        parse_inventory_xls(tmp_path / "offiho-corrupt.xls")


@pytest.mark.parametrize("status", ["SOBRE PEDIDO", "Consultar Existencias"])
def test_parse_inventory_audits_recognized_nonquantitative_stock(monkeypatch, tmp_path, status):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _SpecialOrderWorkbook(status))

    rows, audit = build._parse_inventory_xls(tmp_path / "offiho-special-order.xls")

    assert [row["inventory_key"] for row in rows] == ["OHE-406 GRIS ALUFSEN"]
    assert audit["excluded_stock_status_count"] == 1


def test_parse_inventory_audits_special_order_pieces_default(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _SpecialOrderPiecesWorkbook())

    rows, audit = build._parse_inventory_xls(tmp_path / "offiho-special-order-pieces.xls")

    assert rows[0]["pieces_per_box"] == 1
    assert audit["defaulted_pieces_status_count"] == 1


def test_parse_inventory_audits_blank_stock_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _BlankStockWorkbook())

    rows, audit = build._parse_inventory_xls(tmp_path / "offiho-blank-stock.xls")

    assert [row["inventory_key"] for row in rows] == ["OHE-406 GRIS ALUFSEN"]
    assert audit["excluded_blank_stock_count"] == 1


def test_parse_inventory_audits_repeated_header_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(build.xlrd, "open_workbook", lambda path: _RepeatedHeaderWorkbook())

    rows, audit = build._parse_inventory_xls(tmp_path / "offiho-repeated-header.xls")

    assert len(rows) == 2
    assert audit["excluded_header_row_count"] == 1


def test_extract_identity_separates_model_name_and_variant():
    identity = extract_offiho_identity("OHE-405 NEGRO ALUFSEN")

    assert identity.code == "OHE-405"
    assert identity.name == "ALUFSEN"
    assert identity.variant == "NEGRO"


@pytest.mark.parametrize(
    "color",
    [
        "AMARILLO",
        "AQUA",
        "AVOCADO",
        "BERENJENA",
        "CAPUCCINO",
        "CAMEL",
        "CELESTE",
        "CREMA",
        "FANGO",
        "GRISVERDE",
        "MARRON",
        "MORADO",
        "MOSTAZA",
        "NARANANJA",
        "OCENAO",
        "PANTIKAN",
        "PERLA",
        "PLATA",
        "TABACO",
    ],
)
def test_extract_identity_recognizes_shopify_color_variants(color):
    identity = extract_offiho_identity(f"OHV-126 {color} FABRIZIA")

    assert identity.name == "FABRIZIA"
    assert identity.variant == color


@pytest.mark.parametrize(
    ("inventory_key", "expected_name", "expected_variant"),
    [
        ("OHV-128 LOUNGE GRIS FABRIZIA", "LOUNGE FABRIZIA", "GRIS"),
        ("OHV-47 ALTA BLANCA QUICK", "ALTA QUICK", "BLANCA"),
        ("OHE-204 ALTO CAFÉ GOETZ", "ALTO GOETZ", "CAFÉ"),
        ("OHE-503 G GRIS NET", "G NET", "GRIS"),
        ("OHV-3000-W CR BLANCO INNOVA", "-W CR INNOVA", "BLANCO"),
        ("OHV-12 C/B NEGRO/NEGRO CIAO", "C/B CIAO", "NEGRO NEGRO"),
        ("OHV-7211B W9 BLANCO SENSILLA", "W9 SENSILLA", "BLANCO"),
        ("NOVAISO CON BRAZOS AZUL", "CON BRAZOS", "AZUL"),
        ("STATO B/T-600 N NEGRA", "B/T-600 N", "NEGRA"),
    ],
)
def test_extract_identity_keeps_configuration_before_color_in_model_name(
    inventory_key,
    expected_name,
    expected_variant,
):
    identity = extract_offiho_identity(inventory_key)

    assert identity.name == expected_name
    assert identity.variant == expected_variant


def test_pdf_price_index_normalizes_compact_variant(monkeypatch, tmp_path):
    monkeypatch.setattr(build, "extract_pdf_pages", lambda paths: ["ALUFSEN OHE-405 negro $ 7,999"])

    prices = parse_pdf_price_index([tmp_path / "prices.pdf"])

    assert prices["OHE-405 NEGRO"] == Decimal("7999")


def test_page_description_extracts_only_official_product_description():
    page_text = """
    FENIX OHE-165 NEGRO
    DESCRIPCIÓN
    base: de nylon con rodajas.
    elevación: pistón neumático.
    mecanismo: reclinable.
    Accesorios opcionales
    Cabecera
    Modelados 3D
    Descargar
    """

    description = build._page_description(page_text)

    assert description == "base: de nylon con rodajas. elevación: pistón neumático. mecanismo: reclinable."


def test_real_black_pdf_index_maps_vesper_family_to_pdf_asset(tmp_path):
    project = Path(__file__).resolve().parents[1]
    pdf_path = project / "LP BLACK®️ & COLOS®️ JUL2026.pdf"
    if not pdf_path.exists():
        pytest.skip("El PDF BLACK/COLOS no esta disponible en este checkout")
    items = [
        {
            "inventory_key": "VESPER 1",
            "code": "VESPER",
            "name": "1",
            "variant": "",
        },
        {
            "inventory_key": "VESPER/05 NEGRA",
            "code": "VESPER/05",
            "name": "VESPER/05 NEGRA",
            "variant": "NEGRA",
        },
    ]

    index = build.parse_pdf_product_index(
        [pdf_path],
        items,
        tmp_path / "assets",
        "https://web-lemon-one-45.vercel.app/catalog-assets/offiho",
    )

    product = index["VESPER/05 NEGRA"]
    assert product["unit_price"] == Decimal("2799")
    assert "acero tubular" in product["description"].casefold()
    assert product["image_url"].startswith(
        "https://web-lemon-one-45.vercel.app/catalog-assets/offiho/images/"
    )
    assert product["product_url"].endswith("lp-black-colos-jul2026.pdf#page=15")
    assert (tmp_path / "assets" / "lp-black-colos-jul2026.pdf").exists()
    assert list((tmp_path / "assets" / "images").glob("vesper-1*"))


def test_pdf_title_match_uses_longest_family_prefix_for_color_variants():
    records = [
        {"title": "PIAZZA 1", "image_name": "piazza-1"},
        {"title": "PIAZZA 3A", "image_name": "piazza-3a"},
        {"title": "VILLA 3", "image_name": "villa-3"},
    ]

    product = build._title_record_for_inventory("PIAZZA 3A N NEGRO", records)

    assert product is not None
    assert product["image_name"] == "piazza-3a"


def test_pdf_title_match_does_not_cross_product_families():
    records = [{"title": "PIAZZA 1", "image_name": "piazza-1"}]

    assert build._title_record_for_inventory("VILLA 1 N NEGRO", records) is None


@pytest.mark.parametrize(
    ("inventory_key", "asset_name", "page"),
    [
        ("ECOGERENCIAL NEGRO", "econosillas-ecogerencial.jpg", 6),
        ("ECONOMALLA BLANCO", "econosillas-economalla.jpg", 11),
        ("ECOVISITA VISITA", "econosillas-ecovisita.jpg", 16),
        ("ISO SIN BRAZOS BRAZOS", "econosillas-iso-sin-brazos.jpg", 17),
        ("ISO CON BRAZOS BRAZOS", "econosillas-iso-con-brazos.jpg", 18),
        ("NOVAISO SIN BRAZOS NEGRO", "econosillas-novaiso-sin-brazos.jpg", 19),
        ("NOVAISO CON BRAZOS AZUL", "econosillas-novaiso-con-brazos.jpg", 20),
        ("OHV-64 BLANCO SAND", "econosillas-sand.jpg", 26),
    ],
)
def test_official_brochure_fallback_maps_verified_product_family(
    tmp_path, inventory_key, asset_name, page
):
    image_dir = tmp_path / "images"
    image_dir.mkdir()
    (image_dir / asset_name).write_bytes(b"verified brochure image")

    product = build.match_official_brochure_product(
        {"inventory_key": inventory_key},
        tmp_path,
        "https://web-lemon-one-45.vercel.app/catalog-assets/offiho",
    )

    assert product["match_status"] == "official_brochure_match"
    assert product["image_url"].endswith(f"/images/{asset_name}")
    assert product["product_url"] == f"https://www.offiho.com/folletoeconosillas.pdf#page={page}"


def test_official_brochure_fallback_rejects_unmapped_or_missing_asset(tmp_path):
    assert build.match_official_brochure_product(
        {"inventory_key": "PRODUCTO SIN MAPA"}, tmp_path, "https://example.test/assets"
    ) == {}
    assert build.match_official_brochure_product(
        {"inventory_key": "OHV-64 BLANCO SAND"}, tmp_path, "https://example.test/assets"
    ) == {}


def test_site_match_requires_expected_model_code():
    product = match_official_product(
        extract_offiho_identity("OHE-405 NEGRO ALUFSEN"),
        [
            {
                "codes": ["OHE-405"],
                "url": "https://www.offiho.com/directivos/alufsen",
                "image_url": "https://www.offiho.com/alufsen.jpg",
            }
        ],
    )

    assert product["url"].endswith("/directivos/alufsen")
    assert product["image_url"] == ""


def test_site_match_uses_exact_name_link_without_borrowing_an_unverified_model_image():
    product = match_official_product(
        extract_offiho_identity("OHE-405 NEGRO ALUFSEN"),
        [
            {
                "codes": ["OHE-999"],
                "names": ["Alufsen"],
                "url": "https://www.offiho.com/directivos/alufsen/",
                "image_url": "https://www.offiho.com/images/alufsen-frente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
                "description": "Respaldo de malla y asiento tapizado.",
            }
        ],
    )

    assert product["url"] == "https://www.offiho.com/directivos/alufsen/"
    assert product["image_url"] == ""
    assert product["match_status"] == "official_name_match"
    assert product["description"] == "Respaldo de malla y asiento tapizado."


def test_site_match_prefers_exact_code_over_name_fallback():
    product = match_official_product(
        extract_offiho_identity("OHE-405 NEGRO ALUFSEN"),
        [
            {
                "codes": ["OHE-999"],
                "names": ["ALUFSEN"],
                "url": "https://www.offiho.com/directivos/alufsen/",
            },
            {
                "codes": ["OHE-405"],
                "names": ["ALUFSEN"],
                "url": "https://www.offiho.com/directivos/alufsen/directivos-alufsen-modelo-OHE-405",
            },
        ],
    )

    assert product["url"].endswith("modelo-OHE-405")
    assert product["match_status"] == "official_code_match"


def test_site_match_accepts_official_code_with_inventory_variant_suffix():
    product = match_official_product(
        extract_offiho_identity("OHE-165 NEGRO FENIX"),
        [
            {
                "codes": ["OHE-165NEGRO"],
                "names": ["FENIX"],
                "url": "https://www.offiho.com/directivos/fenix/directivos-fenix-modelo-OHE-165negro",
                "image_url": "https://www.offiho.com/directivos/fenix/OHE-165/OHE-165negroFrente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["url"].endswith("modelo-OHE-165negro")
    assert product["image_url"].endswith("OHE-165negroFrente.jpg")
    assert product["match_status"] == "official_code_match"


def test_site_match_accepts_page_coded_finish_when_image_names_secondary_upholstery():
    image_url = (
        "https://www.offiho.com/ejecutivos/equa/"
        "OHE-185gris/OHE-185frenteaquaMari.jpg"
    )
    product = match_official_product(
        extract_offiho_identity("OHE-185 GRIS EQUA"),
        [
            {
                "codes": ["OHE-185GRIS"],
                "names": ["EQUA"],
                "url": (
                    "https://www.offiho.com/ejecutivos/equa/"
                    "ejecutivos-equa-modelo-OHE-185gris"
                ),
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
                "variant_images": {
                    "AQUA": {
                        "image_url": image_url,
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            }
        ],
    )

    assert product["image_url"] == image_url
    assert product["match_status"] == "official_code_match"


def test_site_match_accepts_compound_variant_suffix_from_official_code():
    image_url = (
        "https://www.offiho.com/operativos/slingplus/"
        "OHE-94plusgris/colores/OHE-94plusGris.jpg"
    )
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 57907,
    }

    product = build.match_official_product(
        extract_offiho_identity("OHE-94 PLUS GRIS SLING"),
        [
            {
                "url": (
                    "https://www.offiho.com/operativos/slingplus/"
                    "operativos-sling-modelo-OHE-94plusgris"
                ),
                "codes": ["OHE-94PLUSGRIS"],
                "names": ["SLING"],
                "variant_images": {"GRIS": metadata},
                **metadata,
            }
        ],
    )

    assert product["match_status"] == "official_code_match"
    assert product["image_url"] == image_url


def test_site_match_handles_configuration_between_compound_variant_words():
    image_url = "https://www.offiho.com/sling/OHV-94plusNegroCrFrente.jpg"
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    identity = extract_offiho_identity("OHV-94 PLUS CR NEGRO SLING *")
    product = match_official_product(
        identity,
        [
            {
                "codes": ["OHV-94PLUSNEGROCR"],
                "names": ["SLING"],
                "url": "https://www.offiho.com/sling/modelo-OHV-94plusnegrocr",
                "variant_images": {"PLUS NEGRO": metadata},
                **metadata,
            }
        ],
    )

    assert identity.variant == "PLUS NEGRO"
    assert product["image_url"] == image_url


@pytest.mark.parametrize("requested", ["G6", "G8"])
def test_site_match_rejects_conflicting_explicit_configuration_token_in_image(requested):
    image_url = "https://www.offiho.com/naples/OHV-7220B_G7Frente.jpg"
    product = match_official_product(
        extract_offiho_identity(f"OHV-7220B {requested} GRSI CLARO NAPLES"),
        [
            {
                "codes": ["OHV-7220B"],
                "names": ["NAPLES"],
                "url": "https://www.offiho.com/naples/modelo-OHV-7220B",
                "description": "Color G7: gris oscuro.",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_site_match_accepts_matching_explicit_configuration_token_in_image():
    image_url = "https://www.offiho.com/naples/OHV-7220B_G7Frente.jpg"
    product = match_official_product(
        extract_offiho_identity("OHV-7220B G7 GRIS OSCURO NAPLES"),
        [
            {
                "codes": ["OHV-7220B"],
                "names": ["NAPLES"],
                "url": "https://www.offiho.com/naples/modelo-OHV-7220B",
                "description": "Color G7: gris oscuro.",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == image_url


def test_site_match_keeps_generic_documented_alias_image_blank_without_color_evidence():
    image_url = (
        "https://www.offiho.com/visitantes-interior/kyos-collection/"
        "kyos-tapizadas/galeria/OHT-337.jpg"
    )

    product = build.match_official_product(
        extract_offiho_identity("OHV-337 BLANCO KYOS"),
        [
            {
                "url": (
                    "https://www.offiho.com/visitantes-interior/kyos-collection/"
                    "kyos-tapizadas/visitantes-interior-kyoscollection-modelo-OHT-337"
                ),
                "codes": ["OHT-337"],
                "names": ["KYOS"],
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "official_code_match"
    assert product["image_url"] == ""


@pytest.mark.parametrize(
    ("inventory_key", "official_code"),
    [
        ("OHV-338 NEGRO KYOS", "OHT-338"),
        ("OHV-339CR GRIS KYOS", "OHT-339CR"),
        ("OHV-340CR BLANCO KYOS", "OHT-340CR"),
        ("OHR-2800-3P CR CROMADA IVY", "OHR-2800-3PCR"),
        ("OHR-2800-4P CR CROMADA IVY", "OHR-2800-4PCR"),
    ],
)
def test_site_match_keeps_generic_inventory_code_aliases_blank_without_color_evidence(
    inventory_key,
    official_code,
):
    image_url = f"https://www.offiho.com/galeria/{official_code}.jpg"
    product = match_official_product(
        extract_offiho_identity(inventory_key),
        [
            {
                "url": f"https://www.offiho.com/modelo-{official_code}",
                "codes": [official_code],
                "names": ["KYOS", "IVY"],
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "official_code_match"
    assert product["image_url"] == ""


@pytest.mark.parametrize(
    ("inventory_key", "official_name", "image_name"),
    [
        ("OHV-90 GRIS VIOLET", "VIOLET 90", "VioletGris.jpg"),
        ("GAMER-002 MESA DRAGON", "ESCRITORIO DRAGON GAMER002", "Gamer002.jpg"),
        ("SILLA ELEFANTE CAFÉ", "SILLA ELEFANTE", "ElefanteCafe.jpg"),
    ],
)
def test_site_match_uses_curated_official_name_aliases(inventory_key, official_name, image_name):
    image_url = f"https://www.offiho.com/galeria/{image_name}"
    product = match_official_product(
        extract_offiho_identity(inventory_key),
        [
            {
                "url": f"https://www.offiho.com/productos/{image_name.lower()}",
                "codes": [],
                "names": [official_name],
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "official_name_match"
    assert product["image_url"] == image_url


def test_site_match_falls_back_to_exact_name_when_false_code_match_has_no_image():
    product = match_official_product(
        extract_offiho_identity("OHV-90 GRIS VIOLET"),
        [
            {
                "url": "https://www.offihoblack.com/collections/sillas-interior",
                "codes": ["OHV-90"],
                "names": ["SILLAS"],
                "image_url": "https://www.offihoblack.com/images/OHV-127.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            },
            {
                "url": "https://www.offihoblack.com/products/violet-90",
                "codes": ["OHV-128"],
                "names": ["VIOLET 90"],
                "image_url": "https://www.offihoblack.com/images/VioletGris.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            },
        ],
    )

    assert product["match_status"] == "official_name_match"
    assert product["image_url"].endswith("VioletGris.jpg")


def test_site_match_never_borrows_a_variant_image_from_another_model():
    product = match_official_product(
        extract_offiho_identity("OHS-85AL NEGRO REVOLUTION"),
        [
            {
                "codes": ["OHS-85AL"],
                "url": (
                    "https://www.offiho.com/operativos/revolution/"
                    "operativos-revolution-modelo-OHS-85al"
                ),
                "image_url": (
                    "https://www.offiho.com/operativos/revolution/"
                    "OHS-85al/OHS-85alFrente.jpg"
                ),
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            },
            {
                "codes": ["OHS-85AL", "OHV-87CR"],
                "url": (
                    "https://www.offiho.com/visitantes-interior/revolution/"
                    "visitantes-interior-revolution-modelo-OHV-87cr"
                ),
                "variant_images": {
                    "NEGRO": {
                        "image_url": (
                            "https://www.offiho.com/visitantes-interior/revolution/"
                            "OHV-87cr/colores/OHV-87crNegroTurquesa.jpg"
                        ),
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            },
        ],
    )

    assert product["image_url"] == ""


def test_site_match_rejects_same_model_generic_image_when_exact_color_is_missing():
    product = match_official_product(
        extract_offiho_identity("OHE-805 BLANCO QUO"),
        [
            {
                "codes": ["OHE-805"],
                "url": "https://www.offiho.com/ejecutivos/quo/modelo-OHE-805",
                "image_url": "https://www.offiho.com/ejecutivos/quo/OHE-805Frente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
                "variant_images": {
                    "NEGRO": {
                        "image_url": "https://www.offiho.com/ejecutivos/quo/OHE-805negro.jpg",
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            }
        ],
    )

    assert product["image_url"] == ""
    assert product["has_variant_catalog"] is True


def test_site_match_rejects_generic_image_labeled_as_another_color():
    product = match_official_product(
        extract_offiho_identity("OHE-805 BLANCO QUO"),
        [
            {
                "codes": ["OHE-805"],
                "url": "https://www.offiho.com/ejecutivos/quo/modelo-OHE-805",
                "image_url": "https://www.offiho.com/ejecutivos/quo/OHE-805negroFrente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "official_code_match"
    assert product["image_url"] == ""


def test_site_match_does_not_use_full_chair_page_for_headrest_accessory():
    product = match_official_product(
        extract_offiho_identity("CABECERA OHE-112 GRIS"),
        [
            {
                "codes": ["OHE-112"],
                "names": ["CREW"],
                "url": "https://www.offiho.com/ejecutivos/crew/modelo-OHE-112",
                "image_url": "https://www.offiho.com/ejecutivos/crew/OHE-112grisFrente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "unmatched"
    assert product["url"] == ""
    assert product["image_url"] == ""


def test_site_match_accepts_page_that_explicitly_matches_headrest_accessory():
    product = match_official_product(
        extract_offiho_identity("CABECERA OHE-112 GRIS"),
        [
            {
                "codes": ["OHE-112"],
                "names": ["CABECERA CREW"],
                "url": "https://www.offiho.com/accesorios/cabecera-crew-OHE-112",
                "image_url": "https://www.offiho.com/accesorios/OHE-112CabeceraGris.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["url"].endswith("/cabecera-crew-OHE-112")
    assert product["image_url"].endswith("OHE-112CabeceraGris.jpg")


def test_pdf_support_does_not_reintroduce_full_chair_image_for_headrest_accessory():
    identity = extract_offiho_identity("CABECERA OHE-112 GRIS")
    full_chair = {
        "matched_title": "CREW OHE-112",
        "image_url": "https://example.test/crew-chair.jpg",
        "product_url": "https://example.test/crew.pdf#page=1",
    }
    headrest = {
        "matched_title": "CABECERA CREW OHE-112 GRIS",
        "image_url": "https://example.test/crew-headrest-gris.jpg",
        "product_url": "https://example.test/crew.pdf#page=2",
    }

    assert build._support_product_for_identity(identity, full_chair, {}) == {}
    assert build._support_product_for_identity(identity, headrest, {}) == headrest


@pytest.mark.parametrize(
    "inventory_key",
    [
        "OHV-13 ALTA BLANCO STRIKE",
        "OHV-66 CR BLANCO MILK",
        "OHV-11 C/B NEGRO/NEGRO CIAO",
    ],
)
def test_pdf_support_does_not_reintroduce_base_product_for_configuration(inventory_key):
    identity = extract_offiho_identity(inventory_key)
    base_product = {
        "matched_title": f"MODELO {identity.code}",
        "image_url": f"https://example.test/{identity.code}.jpg",
        "product_url": "https://www.offiho.com/lista.pdf#page=1",
    }

    assert build._support_product_for_identity(identity, base_product, {}) == {}


def test_pdf_support_requires_exact_visual_variant_evidence():
    identity = extract_offiho_identity("VESPER/05 NEGRA")
    generic = {
        "matched_title": "VESPER 1",
        "image_url": "https://example.test/vesper-1.jpg",
        "product_url": "https://www.offiho.com/lista.pdf#page=15",
    }
    exact = {
        "matched_title": "VESPER 1 NEGRA",
        "image_url": "https://example.test/vesper-1-negra.jpg",
        "product_url": "https://www.offiho.com/lista.pdf#page=15",
    }

    assert build._support_product_for_identity(identity, generic, {}) == {}
    assert build._support_product_for_identity(identity, exact, {}) == exact


def test_pdf_support_does_not_match_short_color_inside_model_name():
    identity = build.OffihoIdentity(code="RE-AL", name="1", variant="W")
    generic = {
        "matched_title": "RE-AL 1",
        "image_url": "https://example.test/re-al-1.jpg",
        "product_url": "https://www.offiho.com/lista.pdf#page=21",
    }

    assert build._support_product_for_identity(identity, generic, {}) == {}


def test_site_match_keeps_generic_shopify_image_blank_without_color_evidence():
    product = match_official_product(
        extract_offiho_identity("OHE-75 NEGRO VANTO"),
        [
            {
                "codes": ["OHE-75"],
                "names": ["VANTO"],
                "url": "https://www.offihoblack.com/products/vanto-ohe-75",
                "image_url": "https://www.offihoblack.com/cdn/shop/products/VantoE_1400x.jpg?v=1",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["match_status"] == "official_code_match"
    assert product["image_url"] == ""


def test_site_match_prefers_variant_specific_page_generic_image():
    gray_image = "https://www.offiho.com/ejecutivos/aiko/OHE-705gris/OHE-705grisFrente.jpg"
    product = match_official_product(
        extract_offiho_identity("OHE-705 GRIS AIKO"),
        [
            {
                "codes": ["OHE-705"],
                "names": ["AIKO"],
                "url": "https://www.offiho.com/ejecutivos/aiko/modelo-OHE-705",
                "image_url": "https://www.offiho.com/ejecutivos/aiko/OHE-705/OHE-705negroFrente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
                "variant_images": {
                    "NEGRO": {
                        "image_url": "https://www.offiho.com/ejecutivos/aiko/OHE-705/colores/OHE-705.jpg",
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            },
            {
                "codes": ["OHE-705GRIS"],
                "names": ["AIKO"],
                "url": "https://www.offiho.com/ejecutivos/aiko/modelo-OHE-705gris",
                "image_url": gray_image,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            },
        ],
    )

    assert product["image_url"] == gray_image


def test_site_match_uses_model_configuration_to_choose_between_same_code_pages():
    def candidate(url, names, image_prefix):
        return {
            "codes": ["OHV-128"],
            "names": names,
            "url": url,
            "image_url": f"{image_prefix}Default.jpg",
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
            "variant_images": {
                "GRIS": {
                    "image_url": f"{image_prefix}Gris.jpg",
                    "image_verified": True,
                    "image_content_type": "image/jpeg",
                    "image_content_length": 2048,
                }
            },
        }

    candidates = [
        candidate(
            "https://www.offihoblack.com/products/fabrizia-ohv-128",
            ["FABRIZIA"],
            "https://www.offihoblack.com/cdn/shop/products/OHV-128",
        ),
        candidate(
            "https://www.offihoblack.com/products/fabrizia-ohv-128-lounge",
            ["FABRIZIA LOUNGE"],
            "https://www.offihoblack.com/cdn/shop/products/OHV-128Lounge",
        ),
    ]

    standard = match_official_product(
        extract_offiho_identity("OHV-128 GRIS FABRIZIA"),
        candidates,
    )
    lounge = match_official_product(
        extract_offiho_identity("OHV-128 LOUNGE GRIS FABRIZIA"),
        candidates,
    )

    assert standard["url"].endswith("/fabrizia-ohv-128")
    assert standard["image_url"].endswith("OHV-128Gris.jpg")
    assert lounge["url"].endswith("/fabrizia-ohv-128-lounge")
    assert lounge["image_url"].endswith("OHV-128LoungeGris.jpg")


def test_site_match_prefers_official_code_with_requested_high_configuration():
    def candidate(code, suffix):
        image_url = f"https://www.offiho.com/visitantes/strike/{code}/colores/{code}blanco.jpg"
        return {
            "codes": [code],
            "names": ["STRIKE"],
            "url": f"https://www.offiho.com/visitantes/strike/modelo-{code}",
            "image_url": image_url,
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
            "variant_images": {
                "BLANCO": {
                    "image_url": image_url,
                    "image_verified": True,
                    "image_content_type": "image/jpeg",
                    "image_content_length": 2048,
                }
            },
            "source_updated_at": suffix,
        }

    product = match_official_product(
        extract_offiho_identity("OHV-13 ALTA BLANCO STRIKE"),
        [candidate("OHV-13", "standard"), candidate("OHV-13ALTA", "high")],
    )

    assert product["url"].endswith("modelo-OHV-13ALTA")
    assert product["image_url"].endswith("OHV-13ALTAblanco.jpg")


@pytest.mark.parametrize(
    ("inventory_key", "base_code", "configured_code", "variant"),
    [
        ("OHV-66 CR BLANCO MILK", "OHV-66", "OHV-66CR", "BLANCO"),
        ("OHV-11 C/B NEGRO/NEGRO CIAO", "OHV-11", "OHV-11CB", "NEGRO NEGRO"),
    ],
)
def test_site_match_prefers_official_code_with_requested_base_configuration(
    inventory_key,
    base_code,
    configured_code,
    variant,
):
    def candidate(code):
        image_url = f"https://www.offiho.com/modelos/{code}/{code}-{variant.replace(' ', '-')}.jpg"
        return {
            "codes": [code],
            "names": ["MILK", "CIAO"],
            "url": f"https://www.offiho.com/modelos/modelo-{code}",
            "image_url": image_url,
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
            "variant_images": {
                variant: {
                    "image_url": image_url,
                    "image_verified": True,
                    "image_content_type": "image/jpeg",
                    "image_content_length": 2048,
                }
            },
        }

    product = match_official_product(
        extract_offiho_identity(inventory_key),
        [candidate(base_code), candidate(configured_code)],
    )

    assert product["url"].endswith(f"modelo-{configured_code}")
    assert f"/{configured_code}/" in product["image_url"]


@pytest.mark.parametrize(
    "inventory_key",
    [
        "OHV-13 ALTA BLANCO STRIKE",
        "OHV-58 ALTO NEGRO WAY",
        "OHV-66 CR BLANCO MILK",
        "OHV-11 C/B NEGRO/NEGRO CIAO",
    ],
)
def test_site_match_rejects_base_product_without_requested_configuration(inventory_key):
    identity = extract_offiho_identity(inventory_key)
    image_url = f"https://www.offiho.com/modelos/{identity.code}/{identity.code}.jpg"
    product = match_official_product(
        identity,
        [
            {
                "codes": [identity.code],
                "names": ["STRIKE", "WAY", "MILK", "CIAO"],
                "url": f"https://www.offiho.com/modelos/modelo-{identity.code}",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


@pytest.mark.parametrize("misleading_name", ["CREW", "CROMO"])
def test_configuration_guard_does_not_accept_short_suffix_inside_unrelated_word(misleading_name):
    identity = extract_offiho_identity("OHV-66 CR BLANCO MILK")
    candidate = {
        "codes": ["OHV-66"],
        "names": [misleading_name],
        "url": f"https://www.offiho.com/modelos/{misleading_name.casefold()}-OHV-66",
    }

    assert build._candidate_supports_identity_configuration(candidate, identity) is False


def test_site_match_accepts_base_code_when_product_text_explicitly_declares_configuration():
    image_url = "https://www.offihoblack.com/cdn/shop/products/OHE-204AltoCafe.jpg"
    product = match_official_product(
        extract_offiho_identity("OHE-204 ALTO CAFÉ GOETZ"),
        [
            {
                "codes": ["OHE-204"],
                "names": ["GOETZ ALTO"],
                "url": "https://www.offihoblack.com/products/goetz-ohe-204-alto",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == image_url


@pytest.mark.parametrize("feature", ["CABECERA", "CONECTOR", "CUBIERTA"])
def test_site_match_rejects_accessory_page_for_base_product(feature):
    image_url = f"https://www.offiho.com/accesorios/OHE-112-{feature}.jpg"
    product = match_official_product(
        extract_offiho_identity("OHE-112 GRIS CREW"),
        [
            {
                "codes": ["OHE-112"],
                "names": [f"{feature} CREW"],
                "url": f"https://www.offiho.com/accesorios/{feature.casefold()}-OHE-112",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_name_match_rejects_opposite_arm_configuration():
    product = match_official_product(
        extract_offiho_identity("NOVAISO CON BRAZOS AZUL"),
        [
            {
                "codes": [],
                "names": ["NOVAISO"],
                "url": "https://www.offiho.com/econosillas/novaiso-sin-brazos",
                "image_url": "https://www.offiho.com/econosillas/novaiso-sin-brazos.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_site_match_rejects_generic_image_with_conflicting_compound_variant():
    image_url = "https://www.offiho.com/ciao/OHV-12Negro-Perla-frente.jpg"
    product = match_official_product(
        extract_offiho_identity("OHV-12 NEGRO/NEGRO CIAO"),
        [
            {
                "codes": ["OHV-12"],
                "names": ["CIAO"],
                "url": "https://www.offiho.com/ciao/modelo-OHV-12",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_site_match_rejects_generic_image_without_exact_color_evidence():
    image_url = "https://www.offiho.com/kyos/galeria/OHV-332.jpg"
    product = match_official_product(
        extract_offiho_identity("OHV-332 BLANCO KYOS"),
        [
            {
                "codes": ["OHV-332"],
                "names": ["KYOS"],
                "url": "https://www.offiho.com/kyos/modelo-OHV-332",
                "description": "Estructura color: gris o negro.",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_site_match_accepts_generic_image_when_description_proves_only_requested_color():
    image_url = "https://www.offiho.com/directivos/alufsen/OHE-405.jpg"
    product = match_official_product(
        extract_offiho_identity("OHE-405 NEGRO ALUFSEN"),
        [
            {
                "codes": ["OHE-405"],
                "names": ["ALUFSEN"],
                "url": "https://www.offiho.com/directivos/alufsen/modelo-OHE-405",
                "description": "Tapiz en leather de color negro.",
                "image_url": image_url,
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == image_url


def test_site_match_rejects_exact_variant_image_that_names_another_model_code():
    image_url = "https://www.offiho.com/aiko/OHE-703gris/colores/OHE-703gris.jpg"
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    product = match_official_product(
        extract_offiho_identity("OHE-705 GRIS AIKO"),
        [
            {
                "codes": ["OHE-705GRIS"],
                "names": ["AIKO"],
                "url": "https://www.offiho.com/aiko/modelo-OHE-705gris",
                "variant_images": {"GRIS": metadata},
                **metadata,
            }
        ],
    )

    assert product["image_url"] == ""


def test_shopify_exact_product_variant_mapping_overrides_legacy_filename_code():
    image_url = (
        "https://www.offihoblack.com/cdn/shop/products/"
        "OHV-126M.jpg?v=1692902295"
    )
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 70812,
    }
    product = match_official_product(
        extract_offiho_identity("OHV-127 MORADO FABRIZIA *"),
        [
            {
                "codes": ["OHV-127"],
                "names": ["FABRIZIA"],
                "url": "https://www.offihoblack.com/products/fabrizia-ohv-127",
                "variant_images": {"MORADO": metadata},
                **metadata,
            }
        ],
    )

    assert product["image_url"] == image_url


def test_site_match_treats_compound_kit_token_as_required_feature():
    product = match_official_product(
        extract_offiho_identity("OHS-42 KITCR NEGRO PROCHAIR"),
        [
            {
                "codes": ["OHS-42"],
                "names": ["PROCHAIR"],
                "url": "https://www.offiho.com/operativos/prochair/modelo-OHS-42",
                "image_url": "https://www.offiho.com/operativos/prochair/OHS-42frente.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["image_url"] == ""


def test_site_match_prefers_configured_product_page_over_multi_code_collection():
    image_url = "https://www.offiho.com/innova/OHV-3000CRverde.jpg"
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    product = match_official_product(
        extract_offiho_identity("OHV-3000 CR VERDE INNOVA"),
        [
            {
                "codes": ["OHV-3000", "OHV-3000CR"],
                "names": ["INNOVA"],
                "url": "https://www.offiho.com/visitantes-interior/innova/",
            },
            {
                "codes": ["OHV-3000CR"],
                "names": ["INNOVACROMA"],
                "url": "https://www.offiho.com/visitantes-interior/innova/modelo-OHV-3000cr",
                "variant_images": {"VERDE": metadata},
                **metadata,
            },
        ],
    )

    assert product["url"].endswith("modelo-OHV-3000cr")
    assert product["image_url"] == image_url


def test_site_match_does_not_borrow_variant_image_from_other_configuration():
    def candidate(url, names, variant, image_url):
        metadata = {
            "image_url": image_url,
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
        }
        return {
            "codes": ["OHV-128"],
            "names": names,
            "url": url,
            "variant_images": {variant: metadata},
            **metadata,
        }

    product = match_official_product(
        extract_offiho_identity("OHV-128 LOUNGE GRIS FABRIZIA"),
        [
            candidate(
                "https://www.offihoblack.com/products/fabrizia-ohv-128-lounge",
                ["FABRIZIA LOUNGE"],
                "AZUL",
                "https://www.offihoblack.com/cdn/shop/products/OHV-128LoungeAzul.jpg",
            ),
            candidate(
                "https://www.offihoblack.com/products/fabrizia-ohv-128",
                ["FABRIZIA"],
                "GRIS",
                "https://www.offihoblack.com/cdn/shop/products/OHV-128Gris.jpg",
            ),
        ],
    )

    assert product["url"].endswith("fabrizia-ohv-128-lounge")
    assert product["image_url"] == ""


@pytest.mark.parametrize(
    ("url", "expected"),
    [
        ("https://www.offiho.com:443/producto", True),
        ("https://www.offiho.com:444/producto", False),
    ],
)
def test_official_url_rejects_non_https_ports(url, expected):
    assert build.is_official_url(url) is expected


@pytest.mark.parametrize(
    ("inventory_key", "expected_suffix"),
    [
        ("OHE-143 ALUMINIO AZUL ENEAS", "/eneas-ohe-143-aluminio"),
        ("OHE-143 NEGRO GRIS ENEAS", "/eneas-ohe-143-negro"),
    ],
)
def test_site_match_uses_finish_words_to_choose_same_code_product_page(
    inventory_key,
    expected_suffix,
):
    def candidate(finish):
        color = "azul" if finish == "aluminio" else "gris"
        return {
            "codes": ["OHE-143"],
            "names": [f"ENEAS {finish.upper()}"],
            "url": f"https://www.offihoblack.com/products/eneas-ohe-143-{finish}",
            "image_url": f"https://www.offihoblack.com/cdn/shop/products/OHE-143-{finish}-{color}.jpg",
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
        }

    product = match_official_product(
        extract_offiho_identity(inventory_key),
        [candidate("aluminio"), candidate("negro")],
    )

    assert product["url"].endswith(expected_suffix)
    assert expected_suffix.rsplit("-", 1)[-1] in product["image_url"]


@pytest.mark.parametrize("inventory_key", ["RIMINI NEGRO", "FESTINA PERLA"])
def test_site_match_uses_unstructured_inventory_identifier_as_name(inventory_key):
    identity = extract_offiho_identity(inventory_key)
    product = match_official_product(
        identity,
        [
            {
                "codes": [],
                "names": [identity.code],
                "url": f"https://www.offiho.com/econosillas/{identity.code.casefold()}/",
                "image_url": f"https://www.offiho.com/images/{identity.code.casefold()}.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )

    assert product["url"].endswith(f"/{identity.code.casefold()}/")
    assert product["match_status"] == "official_name_match"


def test_image_extraction_never_uses_page_url_for_empty_candidate():
    page_url = "https://www.offiho.com/directivos/alufsen"
    parser = build._PageParser()

    image_url = build._extract_official_image_url(page_url, parser)

    assert image_url == ""
    assert image_url != page_url


def test_image_extraction_accepts_realistic_relative_og_image():
    page_url = "https://www.offiho.com/directivos/alufsen"
    parser = build._PageParser()
    parser.feed(
        '<meta property="og:image" '
        'content="/wp-content/uploads/2026/07/alufsen-negra.jpg?size=large">'
    )

    image_url = build._extract_official_image_url(page_url, parser)

    assert image_url == "https://www.offiho.com/wp-content/uploads/2026/07/alufsen-negra.jpg?size=large"


def test_image_extraction_prefers_ciao_product_front_over_branding_assets():
    page_url = "https://www.offiho.com/econosillas/ciao"
    parser = build._PageParser()
    parser.feed(
        '<img src="/images/logo-econosillas.png">'
        '<img src="/images/logociao.jpg">'
        '<img src="/uploads/precios-lista.jpg">'
        '<img src="/uploads/OHS-12CB/OHS-12B-Negro-frente.jpg">'
        '<img src="/uploads/OHS-12CB/OHS-12B-Negro-lateral.jpg">'
    )

    image_url = build._extract_official_image_url(page_url, parser, codes=["OHS-12B"])

    assert image_url == "https://www.offiho.com/uploads/OHS-12CB/OHS-12B-Negro-frente.jpg"


def test_image_extraction_rejects_logo_menu_social_and_accessory_candidates():
    page_url = "https://www.offiho.com/econosillas/ciao"
    parser = build._PageParser()
    parser.feed(
        '<img src="/images/logo.png">'
        '<img src="/images/menu-icon.png">'
        '<img src="/images/instagram-social.png">'
        '<img src="/images/garantia.jpg">'
        '<img src="/images/caja-accesorio.jpg">'
    )

    assert build._extract_official_image_url(page_url, parser, codes=["OHS-12B"]) == ""


def test_image_extraction_uses_shopify_product_srcset_over_header_logo():
    page_url = "https://www.offihoblack.com/products/vanto-ohe-75"
    parser = build._PageParser()
    parser.feed(
        '<img src="/cdn/shop/files/OffihoBlack_Logo.png">'
        '<source srcset="/cdn/shop/products/VantoEFrente_1800x1800.jpg?v=1 1800w">'
    )

    image_url = build._extract_official_image_url(page_url, parser, codes=["OHE-75"])

    assert image_url == "https://www.offihoblack.com/cdn/shop/products/VantoEFrente_1800x1800.jpg?v=1"


def test_discovery_prioritizes_canonical_shopify_product_pages():
    urls = build._prioritize_product_pages(
        {
            "https://www.offihoblack.com/collections/all",
            "https://www.offihoblack.com/collections/all/products/vanto-ohe-75?variant=123",
            "https://www.offihoblack.com/products/shine-ohv-80",
            "https://www.offiho.com/operativos/ciao/operativos-ciao-modelo-OHS-12CB",
        }
    )

    assert urls[:3] == [
        "https://www.offihoblack.com/products/shine-ohv-80",
        "https://www.offihoblack.com/products/vanto-ohe-75",
        "https://www.offiho.com/operativos/ciao/operativos-ciao-modelo-OHS-12CB",
    ]


def test_canonical_product_url_preserves_offiho_directory_trailing_slash():
    category = "https://www.offiho.com/directivos/"

    assert build._canonical_product_url(category) == category


def test_fetch_page_resolves_relative_links_from_redirected_directory(monkeypatch):
    payload = b'<html><head><title>Directivos - Offiho</title></head><body><a href="alufsen/">ALUFSEN</a></body></html>'

    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = "text/html; charset=utf-8"

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return "https://www.offiho.com/directivos/"

        def read(self, size):
            return payload[:size]

    monkeypatch.setattr(build, "_open_official", lambda request, timeout: _Response())

    record = build._fetch_official_page("https://www.offiho.com/directivos")

    assert record["url"] == "https://www.offiho.com/directivos/"
    assert "https://www.offiho.com/directivos/alufsen/" in record["links"]


def test_fetch_page_normalizes_mixed_case_product_codes(monkeypatch):
    payload = b"<html><body><h1>KYOS</h1><p>Modelo OHP-325cr</p></body></html>"

    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = "text/html"

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return "https://www.offiho.com/escolar/kyos-collection/modelo-OHP-325cr"

        def read(self, size):
            return payload[:size]

    monkeypatch.setattr(build, "_open_official", lambda request, timeout: _Response())

    record = build._fetch_official_page("https://www.offiho.com/escolar/kyos-collection/modelo-OHP-325cr")

    assert record["codes"] == ["OHP-325CR"]


def test_fetch_product_page_ignores_related_model_codes_in_body(monkeypatch):
    payload = (
        b"<html><body><h1>OHS-86al</h1>"
        b"<a href='/modelo-OHV-87cr'>Relacionado OHV-87cr</a></body></html>"
    )

    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = "text/html"

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return (
                "https://www.offiho.com/operativos/revolution/"
                "operativos-revolution-modelo-OHS-86al"
            )

        def read(self, size):
            return payload[:size]

    monkeypatch.setattr(build, "_open_official", lambda request, timeout: _Response())

    record = build._fetch_official_page("https://www.offiho.com/modelo-OHS-86al")

    assert record["codes"] == ["OHS-86AL"]


def test_page_names_do_not_treat_category_card_headings_as_page_name():
    parser = build._PageParser()
    parser.feed("<html><head><title>Visitantes Interior - Offiho</title></head><body><h3>KYOS</h3></body></html>")

    assert "KYOS" not in build._page_names("https://www.offiho.com/visitantes-interior/", parser, [])


def test_page_names_normalize_collection_family_slug():
    parser = build._PageParser()

    assert build._page_names(
        "https://www.offiho.com/visitantes-interior/kyos-collection/", parser, []
    ) == ["KYOS"]


def test_page_parser_excludes_script_and_style_payloads_from_visible_product_text():
    parser = build._PageParser()
    parser.feed(
        "<script>var hidden = 'DESCRIPCIÓN imagen de otro modelo';</script>"
        "<style>.description { color: red; }</style>"
        "<h1>VIOLET OHV-90</h1>"
        "<p>DESCRIPCIÓN Silla visitante tapizada en color a elegir.</p>"
    )

    page_text = " ".join(parser.text)
    assert "otro modelo" not in page_text
    assert "color: red" not in page_text
    assert build._page_description(page_text) == "Silla visitante tapizada en color a elegir."


def test_site_index_keeps_name_only_official_product_pages(monkeypatch):
    page_url = "https://www.offiho.com/directivos/alufsen/"
    record = {
        "url": page_url,
        "links": [],
        "codes": [],
        "names": ["ALUFSEN"],
        "image_url": "https://www.offiho.com/images/alufsen-frente.jpg",
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
        "source_updated_at": "",
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (page_url,))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: record)

    index = build.build_site_product_index({}, now=datetime(2026, 7, 10, tzinfo=timezone.utc))

    assert index["name:ALUFSEN"]["url"] == page_url
    assert index["name:ALUFSEN"]["names"] == ["ALUFSEN"]


def test_site_index_crawls_product_pages_beyond_two_discovery_rounds(monkeypatch):
    seed = "https://www.offiho.com/visitantes-interior/"
    landing = f"{seed}kyos-collection/"
    subdivision = f"{landing}kyos-tapizadas/"
    family = f"{subdivision}familia/"
    product = f"{subdivision}visitantes-modelo-OHV-331"
    records = {
        seed: {"url": seed, "links": [landing], "codes": [], "names": []},
        landing: {"url": landing, "links": [subdivision], "codes": [], "names": ["KYOS"]},
        subdivision: {"url": subdivision, "links": [family], "codes": [], "names": ["KYOS"]},
        family: {"url": family, "links": [product], "codes": [], "names": ["KYOS"]},
        product: {
            "url": product,
            "links": [],
            "codes": ["OHV-331"],
            "names": ["KYOS"],
            "source_updated_at": "",
        },
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (seed,))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: records[url])
    monkeypatch.setattr(build, "_fetch_official_page", lambda url: records[url])

    index = build.build_site_product_index(
        {}, now=datetime(2026, 7, 10, tzinfo=timezone.utc)
    )

    assert index["OHV-331"]["url"] == product


def test_site_index_merges_linked_official_full_color_page_into_product(monkeypatch):
    product = (
        "https://www.offiho.com/visitantes-interior/kyos-collection/"
        "kyos-tapizadas/visitantes-interior-kyoscollection-modelo-OHT-338"
    )
    colors = (
        "https://www.offiho.com/visitantes-interior/kyos-collection/"
        "kyos-tapizadas/OHT-338/colores/"
    )

    def verified(url):
        return {
            "image_url": url,
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
        }

    records = {
        product: {
            "url": product,
            "links": [colors],
            "codes": ["OHT-338"],
            "names": ["KYOS"],
            "variant_images": {"ROJO": verified(f"{product}/OHT-338-rojo.jpg")},
            "source_updated_at": "",
        },
        colors: {
            "url": colors,
            "links": [],
            "codes": ["OHT-338"],
            "names": ["KYOS"],
            "variant_images": {
                "GRIS": verified(f"{colors}OHT-338-gris.jpg"),
                "ROJO": verified(f"{colors}OHT-338-rojo-alterno.jpg"),
            },
            "source_updated_at": "",
        },
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (product,))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: records[url])
    monkeypatch.setattr(build, "_fetch_official_page", lambda url: records[url])

    index = build.build_site_product_index(
        {}, now=datetime(2026, 7, 10, tzinfo=timezone.utc)
    )

    assert index["OHT-338"]["url"] == product
    assert index["OHT-338"]["variant_images"]["GRIS"]["image_url"].endswith(
        "OHT-338-gris.jpg"
    )
    assert index["OHT-338"]["variant_images"]["ROJO"]["image_url"].endswith(
        "OHT-338-rojo.jpg"
    )
    matched = build.match_official_product(
        build.OffihoIdentity("OHT-338", "KYOS", "GRIS"),
        build._site_index_candidates(index),
    )
    assert matched["image_url"].endswith("OHT-338-gris.jpg")
    assert matched["has_variant_catalog"] is True


def test_site_search_enumerates_hidden_official_product_pages(monkeypatch):
    payload = (
        b'<ul><li><a href="https://www.offiho.com/operativos/rainbow/'
        b'operativos-rainbow-modelo-OHE-101">Rainbow</a></li>'
        b'<li><a href="https://evil.example/modelo-OHE-999">Falso</a></li></ul>'
    )

    class _Response:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return "https://www.offiho.com/search.php"

        def read(self, size):
            return payload[:size]

    captured = []

    def fake_open(request, timeout):
        captured.append(request)
        return _Response()

    monkeypatch.setattr(build, "_open_official", fake_open)

    urls = build._fetch_offiho_search_product_urls(["OHE"])

    assert urls == [
        "https://www.offiho.com/operativos/rainbow/operativos-rainbow-modelo-OHE-101"
    ]
    assert captured[0].method == "POST"
    assert captured[0].data == b"keyword=OHE"


def test_site_index_includes_products_enumerated_by_official_search(monkeypatch):
    seed = "https://www.offiho.com/"
    hidden = (
        "https://www.offiho.com/operativos/rainbow/"
        "operativos-rainbow-modelo-OHE-101"
    )
    records = {
        seed: {"url": seed, "links": [], "codes": [], "names": []},
        hidden: {
            "url": hidden,
            "links": [],
            "codes": ["OHE-101"],
            "names": ["RAINBOW"],
            "source_updated_at": "",
        },
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (seed,))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: records[url])
    monkeypatch.setattr(build, "_fetch_official_page", lambda url: records[url])
    monkeypatch.setattr(build, "_fetch_offiho_search_product_urls", lambda: [hidden])

    index = build.build_site_product_index(
        {}, now=datetime(2026, 7, 10, tzinfo=timezone.utc), include_search=True
    )

    assert index["OHE-101"]["url"] == hidden


def test_site_index_preserves_all_product_pages_that_share_a_code(monkeypatch):
    standard_url = "https://www.offihoblack.com/products/fabrizia-ohv-128"
    lounge_url = "https://www.offihoblack.com/products/fabrizia-ohv-128-lounge"
    records = {
        standard_url: {
            "url": standard_url,
            "links": [],
            "codes": ["OHV-128"],
            "names": ["FABRIZIA"],
            "source_updated_at": "",
        },
        lounge_url: {
            "url": lounge_url,
            "links": [],
            "codes": ["OHV-128"],
            "names": ["FABRIZIA LOUNGE"],
            "source_updated_at": "",
        },
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (standard_url, lounge_url))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: records[url])

    index = build.build_site_product_index({}, now=datetime(2026, 7, 10, tzinfo=timezone.utc))

    preserved_urls = {
        candidate["url"]
        for key, candidate in index.items()
        if key.startswith("page:")
    }
    assert preserved_urls == {standard_url, lounge_url}
    candidates = build._site_index_candidates(index)
    assert not any(
        code.startswith("PAGE:")
        for candidate in candidates
        for code in candidate.get("codes", [])
    )


def test_name_index_prefers_family_page_over_different_model_page():
    family = {
        "url": "https://www.offiho.com/directivos/alufsen/",
        "codes": ["OHE-405", "OHV-408"],
        "names": ["ALUFSEN"],
        "image_url": "https://www.offiho.com/images/alufsen.jpg",
    }
    model = {
        "url": "https://www.offiho.com/directivos/alufsen/directivos-alufsen-modelo-OHE-405",
        "codes": ["OHE-405"],
        "names": ["ALUFSEN"],
        "image_url": "https://www.offiho.com/images/ohe-405.jpg",
    }

    assert build._site_candidate_rank("name:ALUFSEN", family) > build._site_candidate_rank(
        "name:ALUFSEN", model
    )


def test_site_seeds_include_each_offiho_catalog_section():
    expected = {
        "directivos",
        "ejecutivos",
        "operativos",
        "industrial",
        "accesorios",
        "visitantes-interior",
        "visitantes-exterior",
        "mesas",
        "bancos",
        "confortables",
        "bancas",
        "escolar",
        "nuevos-productos",
    }
    paths = {urlsplit(url).path.strip("/") for url in build.SITE_SEEDS}

    assert expected <= paths


def test_official_link_normalization_escapes_spaces_before_fetching():
    assert build._normalize_official_link("https://www.offiho.com/3d/WAY OHV-58_DWG.dwg") == (
        "https://www.offiho.com/3d/WAY%20OHV-58_DWG.dwg"
    )


def test_shopify_asset_path_is_not_canonicalized_or_crawled_as_product_page():
    asset = "https://www.offihoblack.com/cdn/shop/products/VantoEFrente_1800x1800.jpg"

    assert build._canonical_product_url(asset) == asset
    assert build._is_official_page_url(asset) is False


def test_image_extraction_ranks_shopify_gallery_link_as_product_image():
    page_url = "https://www.offihoblack.com/products/amelia-ohm-41001"
    parser = build._PageParser()
    parser.feed('<img src="/cdn/shop/files/OffihoBlack_Logo.png">')
    gallery_url = "https://www.offihoblack.com/cdn/shop/files/Amelia-Frente41001_1800x1800.jpg?v=1"

    image_url = build._extract_official_image_url(
        page_url,
        parser,
        codes=["OHM-41001"],
        extra_candidates=[gallery_url],
    )

    assert image_url == gallery_url


def test_image_extraction_rejects_shopify_width_template_before_real_photo():
    page_url = "https://www.offihoblack.com/products/amelia-ohm-41001"
    parser = build._PageParser()
    parser.feed(
        '<meta property="og:image" content="/cdn/shop/files/AMELIA_OHM-41001_{width}x.jpg">'
        '<img src="/cdn/shop/files/Amelia-Frente41001_1800x1800.jpg">'
    )

    image_url = build._extract_official_image_url(page_url, parser, codes=["OHM-41001"])

    assert image_url == "https://www.offihoblack.com/cdn/shop/files/Amelia-Frente41001_1800x1800.jpg"


def test_shopify_product_json_maps_every_exact_color_to_its_featured_image():
    page_url = "https://www.offihoblack.com/products/fabrizia-ohv-126"
    payload = r'''
    <script type="text/javascript">
      var GRFQConfigs = GRFQConfigs || {};
      GRFQConfigs.product = {
        "title":"FABRIZIA OHV-126 /PATAS DE MADERA",
        "handle":"fabrizia-ohv-126",
        "variants":[
          {
            "title":"Verde",
            "option1":"Verde",
            "featured_image":{
              "src":"//www.offihoblack.com/cdn/shop/products/OHV-126-Verde.jpg?v=1"
            }
          },
          {
            "title":"Gris",
            "option1":"Gris",
            "featured_media":{
              "preview_image":{
                "src":"//www.offihoblack.com/cdn/shop/products/OHV-126FrenteGris.jpg?v=1"
              }
            }
          }
        ]
      };
    </script>
    '''

    images = build._extract_shopify_variant_image_urls(
        page_url,
        payload,
        codes=["OHV-126"],
    )

    assert images == {
        "GRIS": "https://www.offihoblack.com/cdn/shop/products/OHV-126FrenteGris.jpg?v=1",
        "VERDE": "https://www.offihoblack.com/cdn/shop/products/OHV-126-Verde.jpg?v=1",
    }


def test_shopify_product_json_ignores_unofficial_or_non_image_variant_sources():
    page_url = "https://www.offihoblack.com/products/fabrizia-ohv-126"
    payload = r'''
    <script>
      GRFQConfigs.product = {
        "title":"FABRIZIA OHV-126",
        "variants":[
          {"option1":"Azul","featured_image":{"src":"https://example.com/OHV-126Azul.jpg"}},
          {"option1":"Morado","featured_image":{"src":"/products/fabrizia-ohv-126"}}
        ]
      };
    </script>
    '''

    assert build._extract_shopify_variant_image_urls(
        page_url,
        payload,
        codes=["OHV-126"],
    ) == {}


def test_shopify_product_json_ignores_default_title_as_a_variant_catalog():
    page_url = "https://www.offihoblack.com/products/yodo-ohe-135plus"
    payload = r'''
    <script>
      GRFQConfigs.product = {
        "title":"YODO OHE-135 PLUS",
        "handle":"yodo-ohe-135plus",
        "variants":[
          {
            "title":"Default Title",
            "option1":"Default Title",
            "featured_image":{"src":"//www.offihoblack.com/cdn/shop/products/YODO.jpg?v=1"}
          }
        ]
      };
    </script>
    '''

    assert build._extract_shopify_variant_image_urls(
        page_url,
        payload,
        codes=["OHE-135"],
    ) == {}


def test_shopify_product_json_keeps_only_unambiguous_multi_option_keys():
    page_url = "https://www.offihoblack.com/products/example-ohv-999"
    payload = r'''
    <script>
      GRFQConfigs.product = {
        "title":"EXAMPLE OHV-999",
        "handle":"example-ohv-999",
        "options":["Color","Base"],
        "variants":[
          {
            "title":"Gris / Madera",
            "option1":"Gris",
            "option2":"Madera",
            "featured_image":{"src":"//www.offihoblack.com/cdn/shop/products/OHV-999GrisMadera.jpg?v=1"}
          },
          {
            "title":"Gris / Cromo",
            "option1":"Gris",
            "option2":"Cromo",
            "featured_image":{"src":"//www.offihoblack.com/cdn/shop/products/OHV-999GrisCromo.jpg?v=1"}
          }
        ]
      };
    </script>
    '''

    images = build._extract_shopify_variant_image_urls(
        page_url,
        payload,
        codes=["OHV-999"],
    )

    assert "GRIS" not in images
    assert images == {
        "GRIS CROMO": "https://www.offihoblack.com/cdn/shop/products/OHV-999GrisCromo.jpg?v=1",
        "GRIS MADERA": "https://www.offihoblack.com/cdn/shop/products/OHV-999GrisMadera.jpg?v=1",
    }


def test_shopify_product_json_exposes_material_prefixed_color_as_exact_color_alias():
    page_url = "https://www.offihoblack.com/products/net-ohe-503-kit-cajero"
    payload = r'''
    <script>
      GRFQConfigs.product = {
        "title":"NET OHE-503 Kit cajero",
        "handle":"net-ohe-503-kit-cajero",
        "variants":[
          {
            "title":"Malla gris",
            "option1":"Malla gris",
            "featured_image":{"src":"//www.offihoblack.com/cdn/shop/products/OHE-503MallaGris.jpg?v=1"}
          }
        ]
      };
    </script>
    '''

    images = build._extract_shopify_variant_image_urls(
        page_url,
        payload,
        codes=["OHE-503"],
    )

    assert images["MALLA GRIS"] == images["GRIS"]


def test_page_product_codes_prefer_visible_product_heading_over_stale_shopify_slug():
    page_url = "https://www.offihoblack.com/products/fabrizia-ohv-128-copia"
    parser = build._PageParser()
    parser.feed(
        '<title>VIOLET OHV-90 – offihoblack</title>'
        '<h1>VIOLET OHV-90</h1>'
        '<a href="/products/fabrizia-ohv-129">Relacionado OHV-129</a>'
    )

    assert build._page_product_codes(page_url, parser) == ["OHV-90"]


def test_site_match_accepts_exact_variant_from_single_code_page_with_stale_slug():
    gray_image = "https://www.offihoblack.com/cdn/shop/files/VIOLETAtrasGris.jpg?v=1"
    product = match_official_product(
        extract_offiho_identity("OHV-90 GRIS VIOLET"),
        [
            {
                "codes": ["OHV-90"],
                "names": ["VIOLET"],
                "url": "https://www.offihoblack.com/products/fabrizia-ohv-128-copia",
                "variant_images": {
                    "GRIS": {
                        "image_url": gray_image,
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            }
        ],
    )

    assert product["url"].endswith("/fabrizia-ohv-128-copia")
    assert product["image_url"] == gray_image


def test_collection_swatch_extraction_maps_large_images_by_exact_finish():
    page_url = (
        "https://www.offiho.com/operativos/revolution/"
        "operativos-revolution-modelo-OHS-86al"
    )
    parser = build._PageParser()
    parser.feed(
        """
        <div class="product-options">
        <img class="cloudzoom-gallery colecciones"
             src="OHS-86al/colores/gris.jpg"
             data-cloudzoom="useZoom: '.cloudzoom', image: 'OHS-86al/colores/OHS-86alGris.jpg', zoomImage: 'OHS-86al/zoom/OHS-86alGris.jpg'">
        <img class="cloudzoom-gallery colecciones"
             src="OHS-86al/colores/azul.jpg"
             data-cloudzoom="useZoom: '.cloudzoom', image: 'OHS-86al/colores/OHS-86alMarino.jpg', zoomImage: 'OHS-86al/zoom/OHS-86alMarino.jpg'">
        </div>
        """
    )

    images = build._extract_variant_image_urls(page_url, parser, codes=["OHS-86AL"])

    assert images == {
        "AZUL": (
            "https://www.offiho.com/operativos/revolution/"
            "OHS-86al/colores/OHS-86alMarino.jpg"
        ),
        "GRIS": (
            "https://www.offiho.com/operativos/revolution/"
            "OHS-86al/colores/OHS-86alGris.jpg"
        ),
        "MARINO": (
            "https://www.offiho.com/operativos/revolution/"
            "OHS-86al/colores/OHS-86alMarino.jpg"
        ),
    }


def test_colos_color_controls_map_visible_finish_to_matching_product_image():
    page_url = "https://colos.it/en/products/stecca-2"
    payload = """
        <a class="ico-colore" data-rel="img-colore1"><span class="info">White</span></a>
        <a class="ico-colore" data-rel="img-colore2"><span class="info">Dark Blue</span></a>
        <a class="ico-colore" data-rel="img-colore3"><span class="info">Aubergine</span></a>
        <div class="img-colore img-colore1">
          <div class="img-wrapper"><img data-src="/storage/products/24/S_stecca-2_bianco.jpg"></div>
        </div>
        <div class="img-colore img-colore2">
          <div class="img-wrapper"><img data-src="/storage/products/24/S_stecca-2_bluscuro.jpg"></div>
        </div>
        <div class="img-colore img-colore3">
          <div class="img-wrapper"><img data-src="/storage/products/24/S_stecca-2_melanzana.jpg"></div>
        </div>
    """

    images = build._extract_colos_variant_image_urls(page_url, payload)

    assert images == {
        "AZUL OSCURO": "https://colos.it/storage/products/24/S_stecca-2_bluscuro.jpg",
        "BERENJENA": "https://colos.it/storage/products/24/S_stecca-2_melanzana.jpg",
        "BLANCO": "https://colos.it/storage/products/24/S_stecca-2_bianco.jpg",
    }


@pytest.mark.parametrize(
    ("website_label", "inventory_finish"),
    [
        ("Forest Green", "VERDE BOSQUE"),
        ("Ice Blue", "AZUL HIELO"),
        ("Mud", "FANGO"),
        ("Grigio Caldo ECO", "GRIS CALIDO"),
        ("Pale blue", "AZUL CLARO"),
        ("Terracotta", "TERRACOTA"),
        ("tobacco", "TABACO"),
    ],
)
def test_colos_translates_official_finish_labels_to_inventory_terms(
    website_label,
    inventory_finish,
):
    assert build._colos_color_key(website_label) == inventory_finish


def test_colos_match_uses_model_and_exact_color_instead_of_family_default():
    image_url = "https://colos.it/storage/products/24/S_stecca-2_bluscuro.jpg"
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    product = build.match_colos_product(
        extract_offiho_identity("STECCA 2 B AZUL OSCURO"),
        [
            {
                "codes": [],
                "names": ["STECCA 2"],
                "url": "https://colos.it/en/products/stecca-2",
                "variant_images": {"AZUL OSCURO": metadata},
                **metadata,
            }
        ],
    )

    assert product["url"] == "https://colos.it/en/products/stecca-2"
    assert product["image_url"] == image_url
    assert product["match_status"] == "official_colos_match"


@pytest.mark.parametrize(
    ("inventory_key", "expected_model"),
    [
        ("TORRE S W BLANCO", "TORRE S"),
        ("STECCA L B VISITANTE AZUL OSCURO", "STECCA L"),
        ("VESPER/01B AZUL CLARO", "VESPER 1"),
    ],
)
def test_colos_model_key_keeps_configuration_but_removes_finish_code(
    inventory_key,
    expected_model,
):
    identity = extract_offiho_identity(inventory_key)

    assert build._colos_identity_model_key(identity) == build._product_name_key(expected_model)


def test_colos_match_trusts_exact_model_scope_for_slash_inventory_codes():
    image_url = "https://colos.it/storage/products/24/S_vesper-1_azzurro.jpg"
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }

    product = build.match_colos_product(
        extract_offiho_identity("VESPER/01B AZUL CLARO"),
        [
            {
                "codes": [],
                "names": ["VESPER 1"],
                "url": "https://colos.it/en/products/vesper-1",
                "variant_images": {"AZUL CLARO": metadata},
                **metadata,
            }
        ],
    )

    assert product["url"] == "https://colos.it/en/products/vesper-1"
    assert product["image_url"] == image_url
    assert product["match_status"] == "official_colos_match"


def test_cloudzoom_gallery_maps_finish_code_to_large_product_image():
    page_url = (
        "https://www.offiho.com/visitantes-exterior/sensilla/"
        "visitantes-exterior-sensilla-modelo-OHV-7211B"
    )
    parser = build._PageParser()
    parser.feed(
        """
        <div class="product-options">
        <img class="cloudzoom-gallery"
             src="OHV-7211B/colores/g6.jpg"
             data-cloudzoom="useZoom: '.cloudzoom', image: 'OHV-7211B/colores/OHV-7211Bcapuccino.jpg', zoomImage: 'OHV-7211B/zoom/OHV-7211Bcapuccino.jpg'">
        </div>
        """
    )

    images = build._extract_variant_image_urls(page_url, parser, codes=["OHV-7211B"])
    large_image = (
        "https://www.offiho.com/visitantes-exterior/sensilla/"
        "OHV-7211B/colores/OHV-7211Bcapuccino.jpg"
    )

    assert images["G6"] == large_image
    assert images["CAPUCCINO"] == large_image


def test_cloudzoom_variant_extraction_ignores_gallery_thumbnails_outside_product_options():
    page_url = (
        "https://www.offiho.com/visitantes-exterior/joyous/"
        "visitantes-exterior-joyous-modelo-OHV-81"
    )
    parser = build._PageParser()
    parser.feed(
        """
        <ul class="slides">
          <li><img class="cloudzoom-gallery"
                   src="OHV-81/OHV-81frenteSalmon.jpg"
                   data-cloudzoom="image: 'OHV-81/OHV-81frenteSalmon.jpg'"></li>
        </ul>
        <div class="product-options">
          <img class="cloudzoom-gallery colecciones"
               src="OHV-81/colores/r5.jpg"
               data-cloudzoom="image: 'OHV-81/colores/OHV-81salmon.jpg'">
        </div>
        """
    )

    images = build._extract_variant_image_urls(page_url, parser, codes=["OHV-81"])

    assert images == {
        "R5": (
            "https://www.offiho.com/visitantes-exterior/joyous/"
            "OHV-81/colores/OHV-81salmon.jpg"
        ),
        "SALMON": (
            "https://www.offiho.com/visitantes-exterior/joyous/"
            "OHV-81/colores/OHV-81salmon.jpg"
        ),
    }


def test_site_match_uses_inventory_finish_code_for_exact_variant_image():
    large_image = (
        "https://www.offiho.com/visitantes-exterior/sensilla/"
        "OHV-7211B/colores/OHV-7211Bcapuccino.jpg"
    )
    product = match_official_product(
        extract_offiho_identity("OHV-7211B G6 GRIS SENSILLA"),
        [
            {
                "codes": ["OHV-7211B"],
                "names": ["SENSILLA"],
                "url": (
                    "https://www.offiho.com/visitantes-exterior/sensilla/"
                    "visitantes-exterior-sensilla-modelo-OHV-7211B"
                ),
                "variant_images": {
                    "G6": {
                        "image_url": large_image,
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            }
        ],
    )

    assert product["image_url"] == large_image


def test_site_match_uses_compound_finish_code_when_inventory_variant_is_blank():
    image_url = "https://www.offiho.com/econosillas/OHV-7067F/colores/OHV-7067Fp5-w9.jpg"
    product = match_official_product(
        extract_offiho_identity("OHV-7067F P5/W9 P5/W9 PENGUIN"),
        [
            {
                "codes": ["OHV-7067F"],
                "names": ["PENGUIN"],
                "url": "https://www.offiho.com/econosillas/penguin-modelo-OHV-7067F",
                "variant_images": {
                    "P5 W9": {
                        "image_url": image_url,
                        "image_verified": True,
                        "image_content_type": "image/jpeg",
                        "image_content_length": 2048,
                    }
                },
            }
        ],
    )

    assert product["image_url"] == image_url


def test_site_match_does_not_reduce_ambiguous_multi_finish_identity_to_one_code():
    identity = extract_offiho_identity("OHV-124 K5 W9/N4 HERON")
    assert build._identity_finish_lookup_keys(identity) == []
    product = match_official_product(
        identity,
        [
            {
                "codes": ["OHV-124"],
                "names": ["HERON"],
                "url": "https://www.offiho.com/econosillas/heron-modelo-OHV-124",
                "image_url": "https://www.offiho.com/econosillas/OHV-124/OHV-124frenteRojo.jpg",
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
            }
        ],
    )
    assert product["image_url"] == ""


def test_collection_gallery_extraction_maps_each_image_to_its_exact_code():
    page_url = (
        "https://www.offiho.com/visitantes-interior/kyos-collection/"
        "kyos-tapizadas/"
    )
    parser = build._PageParser()
    parser.feed(
        '<img src="galeria/OHV-331.jpg" alt="OHV-331">'
        '<img src="galeria/OHV-335cr.jpg" alt="OHV-335cr">'
        '<img src="galeria/OHT-337.jpg" alt="OHT-337">'
    )

    images = build._extract_code_image_urls(
        page_url,
        parser,
        codes=["OHV-331", "OHV-335CR", "OHT-337"],
    )

    assert images == {
        "OHT-337": f"{page_url}galeria/OHT-337.jpg",
        "OHV-331": f"{page_url}galeria/OHV-331.jpg",
        "OHV-335CR": f"{page_url}galeria/OHV-335cr.jpg",
    }


def test_site_index_uses_exact_collection_image_for_each_code(monkeypatch):
    page_url = (
        "https://www.offiho.com/visitantes-interior/kyos-collection/"
        "kyos-tapizadas/"
    )

    def metadata(code):
        return {
            "image_url": f"{page_url}galeria/{code}.jpg",
            "image_verified": True,
            "image_content_type": "image/jpeg",
            "image_content_length": 2048,
        }

    record = {
        "url": page_url,
        "links": [],
        "codes": ["OHV-331", "OHV-335CR"],
        "names": ["KYOS TAPIZADAS"],
        "image_url": f"{page_url}galeria/OHV-331.jpg",
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
        "code_images": {
            "OHV-331": metadata("OHV-331"),
            "OHV-335CR": metadata("OHV-335cr"),
        },
        "source_updated_at": "",
    }
    monkeypatch.setattr(build, "SITE_SEEDS", (page_url,))
    monkeypatch.setattr(build, "_cached_or_fetch_page", lambda url, pages: record)

    index = build.build_site_product_index({}, now=datetime(2026, 7, 10, tzinfo=timezone.utc))

    assert index["OHV-331"]["image_url"].endswith("/OHV-331.jpg")
    assert index["OHV-335CR"]["image_url"].endswith("/OHV-335cr.jpg")
    assert index["OHV-331"]["codes"] == ["OHV-331"]
    assert index["OHV-335CR"]["codes"] == ["OHV-335CR"]


@pytest.mark.parametrize(
    ("variant", "finish"),
    [("GRIS", "Gris"), ("AZUL MARINO", "Marino"), ("ROJA", "Rojo")],
)
def test_site_match_prefers_exact_finish_image_over_generic_gallery(variant, finish):
    image_url = (
        "https://www.offiho.com/operativos/revolution/"
        f"OHS-86al/colores/OHS-86al{finish}.jpg"
    )
    metadata = {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    product = match_official_product(
        extract_offiho_identity(f"OHS-86AL {variant} REVOLUTION"),
        [
            {
                "codes": ["OHS-86AL"],
                "names": ["REVOLUTION"],
                "url": (
                    "https://www.offiho.com/operativos/revolution/"
                    "operativos-revolution-modelo-OHS-86al"
                ),
                "image_url": (
                    "https://www.offiho.com/operativos/revolution/"
                    "galeria/OHV-85cr.jpg"
                ),
                "image_verified": True,
                "image_content_type": "image/jpeg",
                "image_content_length": 2048,
                "variant_images": {finish.upper(): metadata},
            }
        ],
    )

    assert product["image_url"] == image_url


def _mock_image_response(url, content_type, content_length):
    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = content_type
            self.headers["Content-Length"] = str(content_length)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return url

    return _Response()


def test_image_verification_rejects_jpg_served_as_html(monkeypatch):
    image_url = "https://www.offiho.com/uploads/product.jpg"
    monkeypatch.setattr(
        build,
        "_open_official",
        lambda request, timeout: _mock_image_response(image_url, "text/html", 512),
    )

    verification = build._verify_official_image(image_url)

    assert verification["image_verified"] is False
    assert verification["image_url"] == ""


def test_image_verification_accepts_valid_image_content_type_and_size(monkeypatch):
    image_url = "https://www.offiho.com/uploads/product.jpg"
    seen = []

    def fake_open(request, *, timeout):
        seen.append((request.get_method(), request.full_url, timeout))
        return _mock_image_response(image_url, "image/jpeg", 2048)

    monkeypatch.setattr(build, "_open_official", fake_open)

    verification = build._verify_official_image(image_url)

    assert verification == {
        "image_url": image_url,
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    assert seen == [("HEAD", image_url, 10)]


def test_redirect_handler_blocks_external_location_before_request():
    contacted = []
    handler = build._OfficialRedirectHandler()

    class _Parent:
        def open(self, request, timeout=None):
            contacted.append(request.full_url)
            raise AssertionError("external redirect target must not be contacted")

    handler.parent = _Parent()
    headers = Message()
    headers["Location"] = "https://attacker.example/collect"
    request = build.urllib.request.Request("https://www.offiho.com/start")

    with pytest.raises(ValueError, match="redireccion.*host oficial"):
        handler.http_error_302(request, object(), 302, "Found", headers)

    assert contacted == []


def test_download_inventory_uses_validated_mocked_response(monkeypatch, tmp_path):
    payload = b"mock-biff-inventory"
    seen = []

    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = "application/vnd.ms-excel"

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return "https://www.offiho.com/existencias.xls"

        def read(self, size):
            return payload

    def fake_open(request, *, timeout):
        seen.append((request.full_url, timeout))
        return _Response()

    monkeypatch.setattr(build, "_open_official", fake_open)
    output = tmp_path / "inventory.xls"

    result = build.download_inventory("https://www.offiho.com/existencias.xls", output)

    assert result == output
    assert output.read_bytes() == payload
    assert seen == [("https://www.offiho.com/existencias.xls", 30)]


def _mock_download_response(payload, content_type):
    class _Response:
        def __init__(self):
            self.headers = Message()
            self.headers["Content-Type"] = content_type

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def geturl(self):
            return "https://www.offiho.com/existencias.xls"

        def read(self, size):
            return payload if len(payload) <= size else payload[:size]

    return _Response()


def test_download_inventory_accepts_valid_html_table_mime(monkeypatch, tmp_path):
    payload = b"""<!DOCTYPE html><html><body><table>
<tr><th>CODIGO</th><th>Existencia</th><th>Piezas por Caja</th><th>Precio Lista 1</th></tr>
<tr><td>OHE-405 NEGRO ALUFSEN</td><td>252</td><td>1</td><td>7999</td></tr>
</table></body></html>"""
    monkeypatch.setattr(
        build,
        "_open_official",
        lambda request, timeout: _mock_download_response(payload, "text/html; charset=utf-8"),
    )
    output = tmp_path / "inventory-html.xls"

    build.download_inventory("https://www.offiho.com/existencias.xls", output)

    assert output.read_bytes() == payload
    assert parse_inventory_xls(output)[0]["available_quantity"] == 252


def test_download_inventory_rejects_arbitrary_html_landing(monkeypatch, tmp_path):
    payload = b"<html><body><h1>Maintenance</h1></body></html>"
    monkeypatch.setattr(
        build,
        "_open_official",
        lambda request, timeout: _mock_download_response(payload, "text/html"),
    )
    output = tmp_path / "landing.xls"

    with pytest.raises(ValueError, match="HTML.*inventario"):
        build.download_inventory("https://www.offiho.com/existencias.xls", output)

    assert not output.exists()


def test_download_inventory_rejects_payload_over_limit(monkeypatch, tmp_path):
    payload = b"x" * (10 * 1024 * 1024 + 1)
    monkeypatch.setattr(
        build,
        "_open_official",
        lambda request, timeout: _mock_download_response(payload, "application/vnd.ms-excel"),
    )
    output = tmp_path / "oversized.xls"

    with pytest.raises(ValueError, match="excede el limite"):
        build.download_inventory("https://www.offiho.com/existencias.xls", output)

    assert not output.exists()


def test_download_inventory_rejects_colos_even_though_it_is_an_official_image_source(tmp_path):
    with pytest.raises(ValueError, match="host oficial Offiho"):
        build.download_inventory(
            "https://colos.it/en/products/stecca-2",
            tmp_path / "inventory.xls",
        )


def _build_catalog_with_sources(monkeypatch, tmp_path, label, *, pdf_payload, site_index):
    inventory = tmp_path / f"inventory-{label}.xls"
    pdf = tmp_path / f"prices-{label}.pdf"
    inventory.write_bytes(b"same inventory bytes")
    pdf.write_bytes(pdf_payload)
    item = {
        "inventory_key": "OHE-405 NEGRO ALUFSEN",
        "code": "OHE-405",
        "name": "ALUFSEN",
        "variant": "NEGRO",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 252,
        "unit_price": 7999,
        "price_source": "inventory",
    }
    monkeypatch.setattr(
        build,
        "_parse_inventory_xls",
        lambda path: ([dict(item)], {"source_row_count": 1, "duplicate_row_count": 0, "unique_item_count": 1}),
    )
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(build, "parse_pdf_product_index", lambda paths, items, assets, base_url: {})
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: site_index)
    return build.build_catalog(
        inventory,
        [pdf],
        tmp_path / f"cache-{label}.json",
        tmp_path / f"catalog-{label}.json",
    )


def test_source_hash_covers_pdf_and_site_snapshot_with_provenance(monkeypatch, tmp_path):
    site_a = {
        "OHE-405": {
            "url": "https://www.offiho.com/directivos/alufsen",
            "image_url": "https://www.offiho.com/uploads/alufsen.jpg",
            "source_updated_at": "Wed, 08 Jul 2026 12:00:00 GMT",
        }
    }
    site_b = {
        "OHE-405": {
            **site_a["OHE-405"],
            "image_url": "https://www.offiho.com/uploads/alufsen-v2.jpg",
        }
    }

    first = _build_catalog_with_sources(monkeypatch, tmp_path, "first", pdf_payload=b"pdf-a", site_index=site_a)
    same = _build_catalog_with_sources(monkeypatch, tmp_path, "same", pdf_payload=b"pdf-a", site_index=site_a)
    changed_pdf = _build_catalog_with_sources(
        monkeypatch, tmp_path, "changed-pdf", pdf_payload=b"pdf-b", site_index=site_a
    )
    changed_site = _build_catalog_with_sources(
        monkeypatch, tmp_path, "changed-site", pdf_payload=b"pdf-a", site_index=site_b
    )

    assert first["source_hash"] == same["source_hash"]
    assert first["source_hash"] != changed_pdf["source_hash"]
    assert first["source_hash"] != changed_site["source_hash"]
    assert first["sources"]["inventory"]["sha256"]
    assert first["sources"]["pdfs"][0]["sha256"]
    assert first["sources"]["site_index"]["sha256"]
    assert first["sources"]["site_index"]["cache_version"] == build.CACHE_VERSION


def test_build_catalog_is_byte_reproducible_for_identical_sources(monkeypatch, tmp_path):
    inventory = tmp_path / "inventory.xls"
    inventory.write_bytes(b"identical inventory")
    item = {
        "inventory_key": "OHE-405 NEGRO ALUFSEN",
        "code": "OHE-405",
        "name": "ALUFSEN",
        "variant": "NEGRO",
        "unit": "PZA",
        "pieces_per_box": 1,
        "available_quantity": 252,
        "unit_price": 7999,
        "price_source": "inventory",
    }
    audit = {"source_row_count": 1, "duplicate_row_count": 0, "unique_item_count": 1}
    monkeypatch.setattr(build, "_parse_inventory_xls", lambda path: ([dict(item)], dict(audit)))
    monkeypatch.setattr(build, "parse_pdf_price_index", lambda paths: {})
    monkeypatch.setattr(build, "build_site_product_index", lambda cache, **kwargs: {})
    first_path = tmp_path / "first.json"
    second_path = tmp_path / "second.json"

    first = build.build_catalog(inventory, [], tmp_path / "cache.json", first_path)
    second = build.build_catalog(inventory, [], tmp_path / "cache.json", second_path)

    assert first_path.read_bytes() == second_path.read_bytes()
    assert first["generated_at"] == second["generated_at"]
    assert datetime.fromisoformat(first["generated_at"]).tzinfo is not None


def test_no_network_uses_compatible_cache_without_refresh(monkeypatch):
    product = {
        "url": "https://www.offiho.com/directivos/alufsen",
        "image_url": "https://www.offiho.com/uploads/alufsen.jpg",
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
        "source_updated_at": "Wed, 08 Jul 2026 12:00:00 GMT",
    }
    cache = {
        "cache_version": build.CACHE_VERSION,
        "site_index": {"OHE-405": product},
        "site_index_expires_at": "2000-01-01T00:00:00+00:00",
    }
    monkeypatch.setattr(
        build,
        "_fetch_official_page",
        lambda url: (_ for _ in ()).throw(AssertionError("network must remain disabled")),
    )

    index = build.build_site_product_index(cache, no_network=True)

    assert index["OHE-405"] == product


def test_no_network_preserves_verified_description_and_finish_images():
    variant_image = {
        "image_url": "https://www.offiho.com/operativos/revolution/OHS-86alRojo.jpg",
        "image_verified": True,
        "image_content_type": "image/jpeg",
        "image_content_length": 2048,
    }
    cache = {
        "cache_version": build.CACHE_VERSION,
        "site_index": {
            "OHS-86AL": {
                "url": "https://www.offiho.com/operativos/revolution/modelo-OHS-86al",
                "description": "Asiento y respaldo de polipropileno de alta resistencia.",
                "variant_images": {"ROJO": variant_image},
                "source_updated_at": "",
            }
        },
        "site_index_expires_at": "2000-01-01T00:00:00+00:00",
    }

    index = build.build_site_product_index(cache, no_network=True)

    assert index["OHS-86AL"]["description"].startswith("Asiento y respaldo")
    assert index["OHS-86AL"]["variant_images"]["ROJO"] == variant_image


def test_no_network_discards_unverified_cache_image():
    cache = {
        "cache_version": build.CACHE_VERSION,
        "site_index": {
            "OHE-405": {
                "url": "https://www.offiho.com/directivos/alufsen",
                "image_url": "https://www.offiho.com/uploads/alufsen.jpg",
                "source_updated_at": "",
            }
        },
        "site_index_expires_at": "2000-01-01T00:00:00+00:00",
    }

    index = build.build_site_product_index(cache, no_network=True)

    assert index["OHE-405"]["image_url"] == ""
    assert index["OHE-405"]["image_verified"] is False
    assert cache["site_index"]["OHE-405"]["image_verified"] is False
    assert cache["site_index"]["OHE-405"]["image_content_type"] == ""


def test_no_network_deterministically_migrates_legacy_snapshot(monkeypatch):
    page_url = "https://www.offiho.com/directivos/alufsen"
    cache = {
        "site_index": {
            "OHE-405": {
                "url": page_url,
                "image_url": "https://www.offiho.com/uploads/unverified.jpg",
                "source_updated_at": "",
            }
        }
    }
    monkeypatch.setattr(
        build,
        "_fetch_official_page",
        lambda url: (_ for _ in ()).throw(AssertionError("network must remain disabled")),
    )

    index = build.build_site_product_index(cache, no_network=True)

    assert cache["cache_version"] == build.CACHE_VERSION
    assert cache["site_index_created_at"] == build.LEGACY_CACHE_TIMESTAMP
    assert index["OHE-405"]["image_url"] == ""


def test_no_network_rejects_incompatible_cache_version():
    cache = {"cache_version": -1, "site_index": {}}

    with pytest.raises(RuntimeError, match="version.*cache"):
        build.build_site_product_index(cache, no_network=True)
