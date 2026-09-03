import hashlib
import json
import re
import zipfile
from collections import Counter
from io import BytesIO
from dataclasses import dataclass, replace
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from xml.etree import ElementTree

import fitz
import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image

from mobiliti_saas.quote_engine.supplier_catalog import PUBLIC_ITEM_FIELDS
from mobiliti_saas.worker.catalog_sync.importers import parse_lumbro_pdf_prices
from mobiliti_saas.worker.catalog_sync.importers import lumbro as lumbro_importer


FIXTURES = Path("tests/fixtures/catalog_graph/lumbro")
GENERAL_PATH = "LUMBRO/LP/LISTA DE PRECIOS MULTICONTACTOS 2026.pdf"
NEW_PATH = "LUMBRO/LP/LISTA DE PRECIOS NUEVOS PRODUCTOS LUMBRO 2025.pdf"
SPEC_PATH = "SPEC GUIDES 2026/LUMBRO/Spec guide-Lumbro-2026.xlsx"
INTERCONNECTION_PATH = "LUMBRO/LP/Precios Interconexión Sunón act.xlsx"
CATALOG_PATH = "LUMBRO/CATALOGO/CATALOGO LUMBRO 2024 DIGITAL (1).pdf"
INTERCONNECTION_HEADER = "PRECIOS UNITARIOS MENOS EL 10% DESCUENTO MAS IVA"
MULT_LIDO_DESCRIPTION = (
    "Multicontacto Especial LINEA LIDO PARA INTERCONECTAR 4 Puertos AC , "
    "1 Puerto USB DE CARGA DOBLE TIPO A, CON ENTRADAS PARA ARNES DE AMBOS "
    "LADOS, MEDIDAS DE 42 X 16 CM"
)
LIDO_OP_DESCRIPTION = (
    "Multicontacto LIDO para canaleta COLOR GRIS OXFORD con 3 puertos AC No "
    "Regulados y 1 PUERTO USB CARGA DOBLE, para INTERCONECTAR"
)
LIDO_OP_A_C_DESCRIPTION = (
    "MULTICONTACTO LIDO PARA CANALETA COLOR GRIS OXFORD, CON 3 PUERTOS TOMA "
    "CORRIENTE NO REGULADOS Y 1 PUERTO USB DE CARGA A+C PARA INTERCONECTAR"
)
JUMPER_DESCRIPTION = (
    "Cable de interconexión o JUMPER CON SALIDA PARA ARNES POR AMBOS LADOS "
    "de 1.5 metros para Carga No Regula"
)
FUSE_BOX_DESCRIPTION = (
    "Caja de Fusible, PARA CARGA NO REGULADA, con entrada para ARNES de un "
    "costado y del otro cable cal 14 con clavija de 2.5 m de longitud"
)


@dataclass(frozen=True)
class AdapterFile:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path | None


def _write_pdf(path: Path, fixture_name: str) -> None:
    pages = json.loads((FIXTURES / fixture_name).read_text(encoding="utf-8"))
    document = fitz.open()
    for expected_page, fixture in enumerate(pages, 1):
        assert fixture["page"] == expected_page
        page = document.new_page(width=612, height=792)
        y = 45
        for line in fixture["text"].splitlines():
            page.insert_text((45, y), line, fontsize=9)
            y += 15
    document.save(path)
    document.close()


def _adapter_file(logical_path: str, local_path: Path) -> AdapterFile:
    return AdapterFile(
        path=logical_path,
        kind="price_list",
        brand=None,
        sha256=hashlib.sha256(local_path.read_bytes()).hexdigest(),
        mime_type="application/pdf",
        local_path=local_path,
    )


def _spec_file(local_path: Path) -> AdapterFile:
    return AdapterFile(
        path=SPEC_PATH,
        kind="spec_guide",
        brand=None,
        sha256=hashlib.sha256(local_path.read_bytes()).hexdigest(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        local_path=local_path,
    )


def _interconnection_file(local_path: Path) -> AdapterFile:
    return AdapterFile(
        path=INTERCONNECTION_PATH,
        kind="price_list",
        brand=None,
        sha256=hashlib.sha256(local_path.read_bytes()).hexdigest(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        local_path=local_path,
    )


def _catalog_file(local_path: Path) -> AdapterFile:
    return AdapterFile(
        path=CATALOG_PATH,
        kind="catalog",
        brand=None,
        sha256=hashlib.sha256(local_path.read_bytes()).hexdigest(),
        mime_type="application/pdf",
        local_path=local_path,
    )


def _png_bytes(color="#303030") -> bytes:
    output = BytesIO()
    Image.new("RGB", (24, 16), color).save(output, format="PNG")
    return output.getvalue()


@pytest.fixture
def pdf_sources(tmp_path):
    general = tmp_path / "LISTA DE PRECIOS MULTICONTACTOS 2026.pdf"
    new = tmp_path / "LISTA DE PRECIOS NUEVOS PRODUCTOS LUMBRO 2025.pdf"
    _write_pdf(general, "price_general_pages.json")
    _write_pdf(new, "price_new_pages.json")
    return (
        _adapter_file(GENERAL_PATH, general),
        _adapter_file(NEW_PATH, new),
    )


@pytest.fixture
def spec_source(tmp_path):
    path = tmp_path / "Spec guide-Lumbro-2026.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SPEC-GUIDE-LUMBRO"
    sheet.append([])
    for column, value in enumerate(
        ("Cod.", "Imagen.", "Descripcion.", "Medida/Unidad.", "P. Unitario.", "Moneda"),
        1,
    ):
        sheet.cell(8, column, value)

    sheet.merge_cells("C9:F9")
    sheet["C9"] = "Barcelona"
    sheet["A10"] = "BARCELONA"
    sheet["C10"] = "Multicontacto, Modelo Barcelona. Incluye: 3 puertos de corriente"
    sheet["D10"] = "245 x 102 x 60 mm"
    sheet["E10"] = 5648
    sheet["F10"] = "MXN"
    sheet["C11"] = "  y 1 cable de metro con clavija  "
    sheet["C12"] = "Color: Gris Aluminio, Negro"
    sheet["C14"] = "Su montaje requiere orificio de 220 x 68 mm"
    sheet["C15"] = "NOTA: SE PUEDEN MODIFICAR LAS CONEXIONES"

    sheet.merge_cells("C19:F19")
    sheet["C19"] = "Barcelona Carga"
    sheet["A20"] = "BARCELONA"
    sheet["C20"] = "Multicontacto, Modelo Barcelona. Incluye: USB de carga"
    sheet["D20"] = "245 x 102 x 60 mm"
    sheet["E20"] = 5648
    sheet["F20"] = "MXN"
    sheet["C21"] = "y 1 cable de metro con clavija"
    sheet["C22"] = "Color: Blanco"
    sheet["C24"] = "Su montaje requiere orificio de 220 x 68 mm"

    sheet.merge_cells("C29:F29")
    sheet["C29"] = "Lisboa"
    sheet["A30"] = "LISBOA"
    sheet["C30"] = "Multicontacto, Modelo Lisboa"
    sheet["D30"] = "245 x 102 x 60 mm"
    sheet["E30"] = 4000
    sheet["F30"] = "MXN"
    sheet["C31"] = "Color: Negro"

    sheet.merge_cells("C39:F39")
    sheet["C39"] = "Ibiza Carga"
    sheet["A40"] = "IBIZA"
    sheet["C40"] = "Multicontacto, Modelo Ibiza con USB de carga"
    sheet["D40"] = "185 x 75 x 60 mm"
    sheet["E40"] = 1648
    sheet["F40"] = "MXN"
    sheet["C41"] = "Color: Negro"

    sheet.merge_cells("C49:F49")
    sheet["C49"] = "Monaco"
    sheet["A50"] = "MONACO-G"
    sheet["C50"] = "Pasacables, Modelo Monaco G"
    sheet["D50"] = "300 x 111 x 28 mm"
    sheet["E50"] = 1000
    sheet["F50"] = "MXN"
    sheet["C51"] = "Color: Negro"

    sheet.merge_cells("C59:F59")
    sheet["C59"] = "Barcelona box HDMI"
    sheet["A60"] = "BARCELONA BOX IN"
    sheet["C60"] = "Multicontacto Barcelona con HDMI inalambrico"
    sheet["E60"] = 5648
    sheet["F60"] = "MXN"
    sheet["C61"] = "Color: Negro"

    sheet.add_image(XlsxImage(BytesIO(_png_bytes("#303030"))), "B10")
    sheet.add_image(XlsxImage(BytesIO(_png_bytes("#0044AA"))), "B20")
    sheet.add_image(XlsxImage(BytesIO(_png_bytes("#AA0000"))), "A1")
    sheet.add_image(XlsxImage(BytesIO(_png_bytes("#00AA00"))), "B31")
    sheet.add_image(XlsxImage(BytesIO(_png_bytes("#AA00AA"))), "C31")
    workbook.save(path)
    workbook.close()
    return _spec_file(path)


@pytest.fixture
def lumbro_spec(spec_source):
    return lumbro_importer.parse_lumbro_spec_guide(spec_source)


def _write_interconnection_workbook(
    path: Path,
    *,
    active_sheet: str = "2026",
    header: str = INTERCONNECTION_HEADER,
    duplicate_anchor: bool = False,
    corrupt_2025_image: bool = False,
) -> None:
    workbook = Workbook()
    old = workbook.active
    old.title = "2025"
    current = workbook.create_sheet("2026")

    old["H3"] = INTERCONNECTION_HEADER
    old["G4"] = MULT_LIDO_DESCRIPTION
    old["H4"] = 2587.5
    old.add_image(XlsxImage(BytesIO(_png_bytes("#AA0000"))), "C13")

    current["H3"] = header
    current["G4"] = MULT_LIDO_DESCRIPTION
    current["H4"] = 3003
    current["G7"] = LIDO_OP_DESCRIPTION
    current["H7"] = 1394.07
    current["G13"] = JUMPER_DESCRIPTION
    current["H13"] = 350
    current["G19"] = FUSE_BOX_DESCRIPTION
    current["H19"] = 772
    current["G26"] = "Multicontacto sin código oficial verificable"
    current["H26"] = "=3536.6666666667*0.9"
    current["G31"] = "Multicontacto con cable sin código verificable"
    current["H31"] = "=3614.4444444444*0.9"
    current["G36"] = "Amberes de carga sin código verificable"
    current["H36"] = 3208
    current["G51"] = LIDO_OP_A_C_DESCRIPTION
    current["H51"] = 1394.07
    current["G62"] = "Jumper de un metro sin código verificable"
    current["H62"] = 270
    current["G76"] = "UP1 para canaleta sin código verificable"
    current["H76"] = "=1552.6*0.9"
    current["O4"] = "Configuración alterna sin código verificable"
    current["P4"] = 3277
    current["O5"] = "Esta pareja no está autorizada"
    current["P5"] = 1234

    for index, anchor in enumerate(
        (
            "C4",
            "J4",
            "M4",
            "J5",
            "J7",
            "C13",
            "C19",
            "C26",
            "I26",
            "B36",
            "A51",
            "D51",
            "I51",
            "C62",
            "A76",
            "D76",
        )
    ):
        color = f"#{(index + 1) * 1000:06X}"
        current.add_image(XlsxImage(BytesIO(_png_bytes(color))), anchor)
    if duplicate_anchor:
        current.add_image(XlsxImage(BytesIO(_png_bytes("#EFEFEF"))), "C13")

    workbook.active = workbook.sheetnames.index(active_sheet)
    raw_path = path.with_name(f"{path.stem}-openpyxl.xlsx")
    workbook.save(raw_path)
    workbook.close()

    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    formula_caches = {"H26": "3183", "H31": "3253", "H76": "1397.34"}
    with zipfile.ZipFile(raw_path) as source_archive, zipfile.ZipFile(
        path, "w", zipfile.ZIP_DEFLATED
    ) as output_archive:
        for info in source_archive.infolist():
            data = source_archive.read(info)
            if info.filename == "xl/worksheets/sheet2.xml":
                root = ElementTree.fromstring(data)
                for cell in root.iter(f"{{{namespace}}}c"):
                    cached = formula_caches.get(cell.get("r"))
                    if cached is None:
                        continue
                    value = cell.find(f"{{{namespace}}}v")
                    assert value is not None
                    value.text = cached
                data = ElementTree.tostring(
                    root, encoding="utf-8", xml_declaration=True
                )
            if corrupt_2025_image and info.filename == "xl/media/image1.png":
                data = b"not-a-decodable-png"
            output_archive.writestr(info, data)


@pytest.fixture
def interconnection_source(tmp_path):
    path = tmp_path / "Precios Interconexion Sunon act.xlsx"
    _write_interconnection_workbook(path)
    return _interconnection_file(path)


@pytest.fixture
def catalog_source(tmp_path):
    path = tmp_path / "CATALOGO LUMBRO 2024 DIGITAL (1).pdf"
    document = fitz.open()
    page = document.new_page(width=612, height=792)
    page.insert_textbox(
        fitz.Rect(45, 45, 560, 740),
        "Empotrables\nBarcelona\nMedidas\n245 mm\n102 mm\n60 mm\n",
        fontsize=11,
    )
    document.save(path)
    document.close()
    return _catalog_file(path)


@pytest.fixture
def lumbro_sources(pdf_sources, spec_source, interconnection_source, catalog_source):
    return (*pdf_sources, spec_source, interconnection_source, catalog_source)


@pytest.fixture
def lumbro_build(lumbro_sources):
    return lumbro_importer.build_lumbro_snapshot_with_assets(lumbro_sources)


def test_general_pdf_keeps_published_net_price(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    barcelona = next(row for row in rows if row.identity == "barcelona")

    assert barcelona.model == "Barcelona"
    assert barcelona.configuration == ""
    assert barcelona.net_price == Decimal("2824")
    assert barcelona.currency == "MXN"
    assert barcelona.tax_rate == Decimal("0.16")
    assert barcelona.authority_rank == 2
    assert barcelona.parse_status == "parsed"
    assert barcelona.source.path == GENERAL_PATH
    assert barcelona.source.file_id == pdf_sources[0].sha256
    assert barcelona.source.page > 0
    assert not isinstance(barcelona.net_price, float)


def test_general_pdf_parses_explicit_ibiza_carga_a_c_configuration(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    ibiza = next(row for row in rows if row.identity == "ibiza carga a c")

    assert ibiza.model == "Ibiza"
    assert ibiza.configuration == "Carga A+C"
    assert ibiza.net_price == Decimal("824")
    assert ibiza.parse_status == "parsed"


def test_general_pdf_does_not_swallow_currency_separator_before_wrapped_model(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    box = next(row for row in rows if row.identity == "barcelona box hdmi inalambrico")

    assert box.model == "Barcelona"
    assert box.configuration == "Box/HDMI Inalámbrico"
    assert box.net_price == Decimal("2824")


def test_new_pdf_parses_venecia_inalambrico_with_higher_authority(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    venecia = [
        row
        for row in rows
        if row.identity == "venecia inalambrico" and row.source.path == NEW_PATH
    ]

    assert len(venecia) == 1
    assert venecia[0].model == "Venecia"
    assert venecia[0].configuration == "Inalámbrico"
    assert venecia[0].net_price == Decimal("1490")
    assert venecia[0].authority_rank == 3
    assert venecia[0].parse_status == "parsed"


def test_pdf_repeated_headings_are_not_product_rows(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)

    assert all(row.identity not in {"modelo", "precio"} for row in rows)
    assert sum(row.identity == "torre octa" for row in rows) == 2


def test_general_pdf_malformed_currency_is_retained_for_review(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    hamburgo = next(row for row in rows if row.identity == "hamburgo")

    assert hamburgo.net_price is None
    assert hamburgo.parse_status == "needs_review"
    assert "malformed_currency" in hamburgo.warnings
    assert hamburgo.source.path == GENERAL_PATH


def test_new_pdf_conflicting_duplicate_prices_remain_as_review_evidence(pdf_sources):
    rows = parse_lumbro_pdf_prices(pdf_sources)
    duplicates = [row for row in rows if row.identity == "torre octa"]

    assert [row.net_price for row in duplicates] == [Decimal("5480"), Decimal("5580")]
    assert all(row.parse_status == "needs_review" for row in duplicates)
    assert all("conflicting_price" in row.warnings for row in duplicates)
    assert [row.source.page for row in duplicates] == [1, 2]


@pytest.mark.parametrize(
    "mutation",
    [
        lambda rows: rows[:1],
        lambda rows: (rows[0], rows[0]),
        lambda rows: (replace(rows[0], path="LUMBRO/LP/otro.pdf"), rows[1]),
        lambda rows: (replace(rows[0], kind="catalog"), rows[1]),
        lambda rows: (replace(rows[0], mime_type="application/octet-stream"), rows[1]),
        lambda rows: (replace(rows[0], sha256="0" * 64), rows[1]),
    ],
)
def test_pdf_source_set_is_strictly_validated(pdf_sources, mutation):
    with pytest.raises(ValueError, match="LUMBRO_PDF_(?:BUNDLE|HASH)"):
        parse_lumbro_pdf_prices(mutation(pdf_sources))


def test_spec_guide_extracts_identity_dimensions_colors_and_exact_provenance(lumbro_spec):
    variants = [
        record
        for record in lumbro_spec.records
        if record.model == "Barcelona" and record.configuration == ""
    ]

    assert {record.color for record in variants} == {"Gris Aluminio", "Negro"}
    assert {record.code for record in variants} == {"BARCELONA"}
    assert {record.dimensions for record in variants} == {"245 x 102 x 60 mm"}
    assert all(
        record.description
        == "Multicontacto, Modelo Barcelona. Incluye: 3 puertos de corriente y 1 cable de metro con clavija"
        for record in variants
    )
    assert all(record.mounting == "Su montaje requiere orificio de 220 x 68 mm" for record in variants)
    assert all(record.source.row == 10 and record.source.heading_row == 9 for record in variants)
    assert all(record.provenance["code"][0]["cell_or_bbox"] == "A10" for record in variants)
    assert all(
        [reference["cell_or_bbox"] for reference in record.provenance["description"]]
        == ["C10", "C11"]
        for record in variants
    )
    assert all(record.provenance["dimensions"][0]["cell_or_bbox"] == "D10" for record in variants)
    assert all(record.provenance["color"][0]["cell_or_bbox"] == "C12" for record in variants)
    assert all(record.notes == ("NOTA: SE PUEDEN MODIFICAR LAS CONEXIONES",) for record in variants)
    assert all(record.provenance["notes"][0]["cell_or_bbox"] == "C15" for record in variants)
    assert all(record.spec_price_evidence == 5648 for record in variants)
    assert all(record.net_price is None for record in variants)


def test_cost_sheet_column_e_is_authoritative_and_uses_its_image_anchor(tmp_path):
    path = tmp_path / "Spec guide-Lumbro-2026.xlsx"
    workbook = Workbook()
    legacy = workbook.active
    legacy.title = "SPEC-GUIDE-LUMBRO"
    for column, value in enumerate(
        ("Cod.", "Imagen.", "Descripcion.", "Medida/Unidad.", "P. Unitario.", "Moneda"),
        1,
    ):
        legacy.cell(8, column, value)
    legacy["C9"] = "Barcelona"
    legacy["A10"] = "BARCELONA"
    legacy["C10"] = "Multicontacto Barcelona"
    legacy["D10"] = "245 x 102 x 60 mm"
    legacy["E10"] = 5648
    legacy["F10"] = "MXN"
    legacy["C11"] = "Color: Negro"
    legacy.add_image(XlsxImage(BytesIO(_png_bytes("#AA0000"))), "B10")

    cost = workbook.create_sheet("COSTO LUMBRO ")
    for column, value in enumerate(
        (
            "Imagen.",
            "Cod.",
            "Descripcion.",
            "Medida/Unidad.",
            "P. Unitario.",
            "LAB Cedis",
            "Moneda",
            "Precio Venta 50% GP",
        ),
        1,
    ):
        cost.cell(8, column, value)
    cost["C9"] = "Barcelona"
    cost["B10"] = "BARCELONA"
    cost["C10"] = "Multicontacto Barcelona"
    cost["D10"] = "245 x 102 x 60 mm"
    cost["E10"] = 2824
    cost["F10"] = Decimal("0.5")
    cost["G10"] = "MXN"
    cost["H10"] = 5648
    cost["C11"] = "Color: Negro"
    cost.add_image(XlsxImage(BytesIO(_png_bytes("#303030"))), "A10")
    workbook.save(path)
    workbook.close()

    build = lumbro_importer.parse_lumbro_spec_guide(_spec_file(path))
    record = build.records[0]

    assert build.price_authoritative is True
    assert record.net_price == Decimal("2824")
    assert record.spec_price_evidence == 2824
    assert record.source.sheet == "COSTO LUMBRO "
    assert record.provenance["spec_price_evidence"][0]["cell_or_bbox"] == "E10"
    assert build.bindings[0].source_references[0]["cell_or_bbox"] == "A10"


def test_spec_guide_emits_only_explicit_variants_and_safe_family_image_reuse(lumbro_spec):
    variants = {
        (record.model, record.configuration, record.color): record
        for record in lumbro_spec.records
    }

    assert set(variants) == {
        ("Barcelona", "", "Gris Aluminio"),
        ("Barcelona", "", "Negro"),
        ("Barcelona", "Carga", "Blanco"),
        ("Barcelona", "Box/HDMI Inalámbrico", "Negro"),
        ("Ibiza", "Carga", "Negro"),
        ("Lisboa", "", "Negro"),
        ("Monaco", "G", "Negro"),
    }
    bindings = {binding.internal_id: binding for binding in lumbro_spec.bindings}
    for identity, image_cell in (
        (("Barcelona", "", "Gris Aluminio"), "B10"),
        (("Barcelona", "", "Negro"), "B10"),
        (("Barcelona", "Carga", "Blanco"), "B20"),
    ):
        record = variants[identity]
        assert record.image_warning == "El color puede variar"
        assert bindings[record.internal_id].match_status == "family_xlsx"
        assert bindings[record.internal_id].source_references[0]["cell_or_bbox"] == image_cell

    lisboa = variants[("Lisboa", "", "Negro")]
    assert lisboa.image_sha256 is None
    assert lisboa.internal_id not in bindings
    assert len(lumbro_spec.assets_by_sha256) == 2


def test_spec_guide_excludes_non_candidate_anchors_and_ambiguous_family_reuse(lumbro_spec):
    variants = {
        (record.model, record.configuration, record.color): record
        for record in lumbro_spec.records
    }
    bindings = {binding.internal_id: binding for binding in lumbro_spec.bindings}
    box = variants[("Barcelona", "Box/HDMI Inalámbrico", "Negro")]

    assert box.image_sha256 is None
    assert box.internal_id not in bindings
    assert {
        reference["cell_or_bbox"]
        for binding in bindings.values()
        for reference in binding.source_references
    } == {"B10", "B20"}
    assert len(lumbro_spec.assets_by_sha256) == 2


def test_spec_guide_prefers_explicit_configuration_in_heading_or_code(lumbro_spec):
    identities = {
        (record.source.row, record.model, record.configuration)
        for record in lumbro_spec.records
    }

    assert (40, "Ibiza", "Carga") in identities
    assert (50, "Monaco", "G") in identities
    assert (60, "Barcelona", "Box/HDMI Inalámbrico") in identities

    monaco = next(record for record in lumbro_spec.records if record.source.row == 50)
    assert [ref["cell_or_bbox"] for ref in monaco.provenance["model"]] == ["C49", "A50"]
    assert [ref["cell_or_bbox"] for ref in monaco.provenance["configuration"]] == ["A50"]


def test_spec_price_is_never_commercial_authority(lumbro_spec, pdf_sources):
    prices = parse_lumbro_pdf_prices(pdf_sources)
    enriched = lumbro_importer.reconcile_lumbro_spec_prices(lumbro_spec.records, prices)
    barcelona = next(
        record
        for record in enriched
        if record.model == "Barcelona" and record.configuration == "" and record.color == "Negro"
    )

    assert barcelona.net_price == Decimal("2824")
    assert barcelona.spec_price_evidence == 5648
    assert barcelona.price_source is not None
    assert barcelona.price_source.path.endswith("LISTA DE PRECIOS MULTICONTACTOS 2026.pdf")
    assert barcelona.provenance["spec_price_evidence"][0]["cell_or_bbox"] == "E10"


def test_spec_price_collision_for_same_identity_fails_closed(lumbro_spec, pdf_sources):
    price = next(row for row in parse_lumbro_pdf_prices(pdf_sources) if row.identity == "barcelona")
    collision = replace(price, net_price=Decimal("9999"))

    enriched = lumbro_importer.reconcile_lumbro_spec_prices(
        lumbro_spec.records, (price, collision)
    )
    barcelona = next(
        record
        for record in enriched
        if record.model == "Barcelona" and record.configuration == ""
    )

    assert barcelona.net_price is None
    assert barcelona.price_source is None
    assert barcelona.spec_price_evidence == 5648


def test_spec_workbook_is_closed_when_parsing_raises(spec_source, monkeypatch):
    workbook = Workbook()
    workbook.active.title = "SPEC-GUIDE-LUMBRO"

    class TrackingWorkbook:
        closed = False

        @property
        def sheetnames(self):
            return workbook.sheetnames

        def __getitem__(self, name):
            return workbook[name]

        def close(self):
            self.closed = True
            workbook.close()

    tracking = TrackingWorkbook()
    monkeypatch.setattr(lumbro_importer, "open_xlsx_data_only", lambda _path: tracking)

    with pytest.raises(ValueError, match="LUMBRO_SPEC_HEADER"):
        lumbro_importer.parse_lumbro_spec_guide(spec_source)

    assert tracking.closed is True


def test_interconnection_uses_active_2026_cached_net_value(interconnection_source):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    item = next(record for record in build.records if record.code == "MULT-LIDO-INT")

    assert item.net_price == Decimal("3003")
    assert item.net_price != Decimal("2702.7")
    assert item.source.sheet == "2026"
    assert item.source.row == 4
    assert item.source.description_cell == "G4"
    assert item.source.price_cell == "H4"
    assert item.authority_rank == 4
    assert all(record.net_price != Decimal("2587.5") for record in build.records)


def test_interconnection_preserves_validated_xml_formula_caches(interconnection_source):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    by_cell = {record.source.price_cell: record for record in build.records}

    assert by_cell["H26"].net_price == Decimal("3183")
    assert by_cell["H31"].net_price == Decimal("3253")
    assert by_cell["H76"].net_price == Decimal("1397.34")
    assert all(record.parse_status == "needs_review" for record in (
        by_cell["H26"], by_cell["H31"], by_cell["H76"]
    ))


def test_interconnection_maps_only_exact_spec_descriptions(interconnection_source):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    by_description = {record.description: record for record in build.records}

    assert by_description[MULT_LIDO_DESCRIPTION].code == "MULT-LIDO-INT"
    assert by_description[LIDO_OP_DESCRIPTION].code == "LIDO.OP-INT"
    assert by_description[JUMPER_DESCRIPTION].code == "JUMP-1.5M"
    assert by_description[FUSE_BOX_DESCRIPTION].code == "CAJA-FUS"

    unknown = by_description["Multicontacto sin código oficial verificable"]
    assert unknown.code == ""
    assert unknown.parse_status == "needs_review"
    assert "missing_code" in unknown.warnings
    assert all(record.description != "Esta pareja no está autorizada" for record in build.records)


def test_interconnection_real_a_c_near_match_never_receives_lido_op_code(
    interconnection_source,
):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    near_match = next(
        record for record in build.records if record.description == LIDO_OP_A_C_DESCRIPTION
    )

    assert near_match.source.description_cell == "G51"
    assert near_match.net_price == Decimal("1394.07")
    assert near_match.code == ""
    assert near_match.parse_status == "needs_review"
    assert "missing_code" in near_match.warnings


def test_interconnection_verified_codes_have_distinct_reconciliation_identities(
    interconnection_source,
):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    verified = {record.code: record for record in build.records if record.code}

    assert set(verified) == {
        "MULT-LIDO-INT",
        "LIDO.OP-INT",
        "JUMP-1.5M",
        "CAJA-FUS",
    }
    assert len({record.identity for record in verified.values()}) == 4
    assert all(record.model and record.configuration for record in verified.values())


def test_interconnection_rank_four_wins_mixed_real_reconciliation(
    interconnection_source, lumbro_spec, pdf_sources
):
    interconnection = lumbro_importer.parse_lumbro_interconnection(
        interconnection_source
    )
    base_spec = next(record for record in lumbro_spec.records if record.code == "BARCELONA")
    target = replace(
        base_spec,
        code="MULT-LIDO-INT",
        model="LIDO",
        configuration="Multicontacto especial para interconectar",
        price_identity="lido",
        net_price=None,
        price_source=None,
    )
    pdf = next(
        record for record in parse_lumbro_pdf_prices(pdf_sources) if record.net_price
    )
    rank_three = replace(pdf, identity="lido", net_price=Decimal("2824"), authority_rank=3)
    rank_two = replace(pdf, identity="lido", net_price=Decimal("2587.5"), authority_rank=2)

    reconciled = lumbro_importer.reconcile_lumbro_spec_prices(
        (target,), (*interconnection.records, rank_three, rank_two)
    )[0]

    assert reconciled.net_price == Decimal("3003")
    assert reconciled.price_source is not None
    assert reconciled.price_source.sheet == "2026"
    assert reconciled.price_source.price_cell == "H4"


def test_interconnection_parses_only_the_o_p_pair_on_row_four(interconnection_source):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    alternate = next(
        record
        for record in build.records
        if record.description == "Configuración alterna sin código verificable"
    )

    assert alternate.source.description_cell == "O4"
    assert alternate.source.price_cell == "P4"
    assert alternate.net_price == Decimal("3277")
    assert alternate.code == ""
    assert alternate.parse_status == "needs_review"


def test_interconnection_images_are_active_sheet_only_and_fail_closed(
    interconnection_source,
):
    build = lumbro_importer.parse_lumbro_interconnection(interconnection_source)
    bindings = {binding.internal_id: binding for binding in build.bindings}
    records = {record.description: record for record in build.records}

    jumper = records[JUMPER_DESCRIPTION]
    assert bindings[jumper.internal_id].source_references[0]["sheet_or_page"] == "2026"
    assert bindings[jumper.internal_id].source_references[0]["cell_or_bbox"] == "C13"

    fuse_box = records[FUSE_BOX_DESCRIPTION]
    assert bindings[fuse_box.internal_id].source_references[0]["cell_or_bbox"] == "C19"
    assert len(build.assets_by_sha256) == 5
    assert len(build.image_evidence) == 16
    assert {evidence.source.sheet for evidence in build.image_evidence} == {"2026"}
    assert Counter(
        (evidence.status, evidence.reason) for evidence in build.image_evidence
    ) == {
        ("bound", None): 5,
        ("excluded", "ambiguous_images"): 7,
        ("excluded", "ambiguous_product_row"): 3,
        ("excluded", "no_product_row"): 1,
    }


@pytest.mark.filterwarnings("error")
def test_interconnection_never_decodes_invalid_2025_image(tmp_path):
    path = tmp_path / "Precios Interconexion Sunon act.xlsx"
    _write_interconnection_workbook(path, corrupt_2025_image=True)

    build = lumbro_importer.parse_lumbro_interconnection(_interconnection_file(path))

    assert len(build.image_evidence) == 16
    assert {evidence.source.sheet for evidence in build.image_evidence} == {"2026"}


def test_interconnection_duplicate_same_anchor_is_counted_as_ambiguous(tmp_path):
    path = tmp_path / "Precios Interconexion Sunon act.xlsx"
    _write_interconnection_workbook(path, duplicate_anchor=True)

    build = lumbro_importer.parse_lumbro_interconnection(_interconnection_file(path))
    c13 = [evidence for evidence in build.image_evidence if evidence.source.cell == "C13"]
    jumper = next(record for record in build.records if record.code == "JUMP-1.5M")

    assert len(build.image_evidence) == 17
    assert len(c13) == 2
    assert all(evidence.status == "excluded" for evidence in c13)
    assert all(evidence.reason == "ambiguous_images" for evidence in c13)
    assert jumper.internal_id not in {binding.internal_id for binding in build.bindings}


@pytest.mark.parametrize(
    ("active_sheet", "header", "error"),
    [
        ("2025", INTERCONNECTION_HEADER, "LUMBRO_INTERCONNECTION_ACTIVE_SHEET"),
        ("2026", "Precio neto aproximado", "LUMBRO_INTERCONNECTION_HEADER"),
    ],
)
def test_interconnection_rejects_wrong_active_sheet_or_header(
    tmp_path, active_sheet, header, error
):
    path = tmp_path / "Precios Interconexion Sunon act.xlsx"
    _write_interconnection_workbook(path, active_sheet=active_sheet, header=header)

    with pytest.raises(ValueError, match=error):
        lumbro_importer.parse_lumbro_interconnection(_interconnection_file(path))


def _without_generated_at(snapshot):
    return {key: value for key, value in snapshot.items() if key != "generated_at"}


def test_snapshot_uses_approved_new_products_precedence(lumbro_sources, tmp_path):
    replacement = tmp_path / "LISTA DE PRECIOS NUEVOS PRODUCTOS LUMBRO 2025.pdf"
    pages = json.loads((FIXTURES / "price_new_pages.json").read_text(encoding="utf-8"))
    pages[0]["text"] = pages[0]["text"].replace("1,490.00", "1,777.00")
    document = fitz.open()
    for fixture in pages:
        page = document.new_page(width=612, height=792)
        page.insert_textbox(
            fitz.Rect(45, 45, 560, 740), fixture["text"], fontsize=9
        )
    document.save(replacement)
    document.close()
    changed = tuple(
        _adapter_file(NEW_PATH, replacement) if row.path == NEW_PATH else row
        for row in lumbro_sources
    )

    venecia = next(
        item
        for item in lumbro_importer.build_lumbro_snapshot(changed)["items"]
        if item["attributes"]["model"] == "Venecia"
    )

    assert venecia["price_net"] == "1777.000000"
    assert venecia["attributes"]["price_source"]["path"] == NEW_PATH
    assert venecia["attributes"]["price_source"]["authority_rank"] == 3


def test_snapshot_contract_hash_ids_and_source_json_are_deterministic(
    lumbro_sources,
):
    first = lumbro_importer.build_lumbro_snapshot(lumbro_sources)
    second = lumbro_importer.build_lumbro_snapshot(tuple(reversed(lumbro_sources)))

    assert _without_generated_at(first) == _without_generated_at(second)
    descriptors = sorted(
        (
            {
                "path": row.path,
                "kind": row.kind,
                "brand": row.brand,
                "sha256": row.sha256,
                "mime_type": row.mime_type,
            }
            for row in lumbro_sources
        ),
        key=lambda row: row["path"],
    )
    expected_material = {
        "link_manifest_fingerprint": lumbro_importer.load_lumbro_link_index().resource_fingerprint,
        "sources": descriptors,
    }
    expected_hash = hashlib.sha256(
        json.dumps(
            expected_material,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=True,
        ).encode("utf-8")
    ).hexdigest()
    assert first["source_hash"] == expected_hash
    assert len({item["internal_id"] for item in first["items"]}) == len(first["items"])
    for item in first["items"]:
        assert set(item) == set(PUBLIC_ITEM_FIELDS)
        evidence = json.loads(item["source_reference"])
        assert item["source_reference"] == json.dumps(
            evidence, sort_keys=True, separators=(",", ":"), ensure_ascii=True
        )


def test_snapshot_exact_contract_and_non_quoteable_review_variants(lumbro_build):
    items = lumbro_build.snapshot["items"]

    assert items
    assert all(item["supplier"] == "lumbro" for item in items)
    assert all(item["brand"] == "Lumbro" for item in items)
    assert all(item["base_currency"] == "MXN" for item in items)
    assert all(item["tax_rate"] == "0.160000" for item in items)
    assert all(item["unit"] == "PZA" for item in items)
    assert all(item["availability_type"] == "unknown" for item in items)
    assert all(item["stock"] is None for item in items)
    assert all(re.fullmatch(r"[0-9]+\.[0-9]{6}", item["price_net"]) for item in items)
    assert all(
        (item["code_status"] == "verified" and item["sku"])
        or (item["code_status"] == "needs_review" and not item["sku"])
        for item in items
    )
    assert any(
        item["code_status"] == "needs_review"
        and not item["sku"]
        and Decimal(item["price_net"]) > 0
        for item in items
    )


def test_snapshot_uses_only_unique_literal_official_codes_as_sku(lumbro_build):
    items = lumbro_build.snapshot["items"]
    assert all(not item["sku"].startswith("lumbro:") for item in items)

    unique = next(
        item
        for item in items
        if item["attributes"]["source_code"] == "BARCELONA BOX IN"
    )
    assert unique["price_net"] == "2824.000000"
    assert unique["sku"] == "BARCELONA BOX IN"
    assert unique["code_status"] == "verified"

    repeated = [
        item
        for item in items
        if item["attributes"]["source_code"] == "BARCELONA"
    ]
    assert len(repeated) > 1
    assert all(item["sku"] == "" for item in repeated)
    assert all(item["code_status"] == "needs_review" for item in repeated)
    assert all(
        any("no es \u00fanico" in warning for warning in item["warnings"])
        for item in repeated
    )
    assert any(
        item["code_status"] == "needs_review"
        and not item["sku"]
        and item["attributes"]["source_code"]
        and Decimal(item["price_net"]) == 0
        for item in items
    )


def test_snapshot_reconciles_verified_source_code_and_never_uses_spec_price(
    lumbro_build,
):
    items = lumbro_build.snapshot["items"]
    mult_lido = next(
        item
        for item in items
        if item["attributes"]["source_code"] == "MULT-LIDO-INT"
    )

    assert mult_lido["price_net"] == "3003.000000"
    assert mult_lido["attributes"]["price_source"] == {
        "authority_rank": 4,
        "cell": "H4",
        "path": INTERCONNECTION_PATH,
        "sheet": "2026",
    }
    barcelona = next(
        item
        for item in items
        if item["attributes"]["source_code"] == "BARCELONA"
        and item["attributes"]["configuration"] == ""
    )
    assert barcelona["price_net"] == "2824.000000"
    assert barcelona["attributes"]["spec_price_evidence"] == 5648
    assert barcelona["attributes"]["price_source"]["path"] == GENERAL_PATH


def test_catalog_pdf_only_enriches_exact_category_and_measurements(lumbro_build):
    barcelona = next(
        item
        for item in lumbro_build.snapshot["items"]
        if item["attributes"]["source_code"] == "BARCELONA"
        and item["attributes"]["configuration"] == ""
    )

    assert barcelona["collection"] == "Empotrables"
    assert barcelona["attributes"]["catalog_measurements"] == [
        "245 mm",
        "102 mm",
        "60 mm",
    ]
    assert barcelona["price_net"] == "2824.000000"
    catalog_references = [
        reference
        for reference in json.loads(barcelona["source_reference"])
        if reference["file_id"] == next(
            source["sha256"]
            for source in lumbro_build.snapshot["metadata"]["sources"]
            if source["path"] == CATALOG_PATH
        )
    ]
    assert catalog_references[0]["sheet_or_page"] == 1


def test_catalog_elevated_profile_requires_exact_pinned_sha(
    catalog_source, monkeypatch
):
    pinned_hash = "bbd810ebab20336d2a6bdc61123955bd062c5a64d57d4359556fcf6aef57e053"
    validations = []
    iterations = []
    current_hash = {"value": catalog_source.sha256}
    actual_hash = {"value": catalog_source.sha256}

    def validated(_path, _extension, **kwargs):
        validations.append(kwargs)
        return type("Validated", (), {"sha256": current_hash["value"]})()

    def pages(_path, **kwargs):
        iterations.append(kwargs)
        return iter(())

    monkeypatch.setattr(lumbro_importer, "validate_source_file", validated)
    monkeypatch.setattr(lumbro_importer, "iter_pdf_pages", pages)
    monkeypatch.setattr(
        lumbro_importer._common,
        "_read_source",
        lambda *_args: (
            type("RawValidated", (), {"sha256": actual_hash["value"]})(),
            b"%PDF-fixture",
        ),
    )

    assert lumbro_importer._parse_lumbro_catalog(catalog_source) == {}
    assert validations == [{}]
    assert iterations == [{}]

    validations.clear()
    iterations.clear()
    current_hash["value"] = pinned_hash
    pinned = replace(catalog_source, sha256=pinned_hash)
    with pytest.raises(ValueError, match="LUMBRO_CATALOG_HASH"):
        lumbro_importer._parse_lumbro_catalog(pinned)
    assert validations == []
    assert iterations == []

    actual_hash["value"] = pinned_hash
    assert lumbro_importer._parse_lumbro_catalog(pinned) == {}
    expected_profile = {"pdf_profile": "lumbro_catalog_2024"}
    assert validations == [expected_profile]
    assert iterations == [expected_profile]


@pytest.mark.parametrize(
    "change",
    [
        {"path": "LUMBRO/CATALOGO/otro.pdf"},
        {"kind": "price_list"},
        {"brand": "Lumbro"},
        {"mime_type": "application/octet-stream"},
        {"sha256": "a" * 64},
        {"local_path": Path("catalog.txt")},
    ],
)
def test_catalog_named_profile_requires_complete_pinned_descriptor(
    catalog_source, change
):
    pinned = replace(
        catalog_source,
        sha256="bbd810ebab20336d2a6bdc61123955bd062c5a64d57d4359556fcf6aef57e053",
    )

    assert lumbro_importer._catalog_pdf_profile(pinned) == {
        "pdf_profile": "lumbro_catalog_2024"
    }
    assert lumbro_importer._catalog_pdf_profile(replace(pinned, **change)) == {}


def _synthetic_price_record(*, color: str, price: str, page: int):
    return SimpleNamespace(
        identity="modelo carga",
        model="Modelo",
        configuration="Carga",
        color=color,
        net_price=Decimal(price),
        currency="MXN",
        tax_rate=Decimal("0.16"),
        source=lumbro_importer.LumbroPriceSource(GENERAL_PATH, "1" * 64, page),
        authority_rank=2,
        parse_status="parsed",
        warnings=(),
    )


def _synthetic_spec_record(*, color: str, row: int):
    return lumbro_importer.LumbroSpecRecord(
        internal_id=f"lumbro:variant:{color.casefold()}",
        identity=lumbro_importer._identity("Modelo", f"Carga {color}"),
        price_identity=lumbro_importer._identity("Modelo", "Carga"),
        model="Modelo",
        configuration="Carga",
        color=color,
        code="COLOR-1",
        description="Modelo de color explÃ­cito",
        dimensions="",
        mounting="",
        notes=(),
        currency="MXN",
        spec_price_evidence=None,
        source=lumbro_importer.LumbroSpecSource(
            SPEC_PATH, "2" * 64, "SPEC-GUIDE-LUMBRO", row - 1, row
        ),
        provenance={},
    )


def _build_synthetic_color_snapshot(monkeypatch, lumbro_sources, specs, prices):
    monkeypatch.setattr(
        lumbro_importer, "parse_lumbro_pdf_prices", lambda _files: tuple(prices)
    )
    monkeypatch.setattr(
        lumbro_importer,
        "parse_lumbro_spec_guide",
        lambda _source: lumbro_importer.LumbroSpecBuild(tuple(specs), {}, ()),
    )
    monkeypatch.setattr(
        lumbro_importer,
        "parse_lumbro_interconnection",
        lambda _source: lumbro_importer.LumbroInterconnectionBuild((), {}, (), ()),
    )
    monkeypatch.setattr(lumbro_importer, "_parse_lumbro_catalog", lambda _source: {})
    return lumbro_importer.build_lumbro_snapshot(lumbro_sources)


def test_reconciliation_keeps_explicit_colors_in_separate_cache_entries(
    monkeypatch, lumbro_sources
):
    snapshot = _build_synthetic_color_snapshot(
        monkeypatch,
        lumbro_sources,
        (
            _synthetic_spec_record(color="Rojo", row=10),
            _synthetic_spec_record(color="Azul", row=20),
        ),
        (
            _synthetic_price_record(color="Rojo", price="100", page=1),
            _synthetic_price_record(color="Azul", price="200", page=2),
        ),
    )

    by_color = {item["attributes"]["color"]: item for item in snapshot["items"]}
    assert by_color["Rojo"]["price_net"] == "100.000000"
    assert by_color["Azul"]["price_net"] == "200.000000"
    assert snapshot["metadata"]["coverage"]["reconciled_rows"] == 2


def test_reconciliation_rejects_explicit_incompatible_price_color(
    monkeypatch, lumbro_sources
):
    snapshot = _build_synthetic_color_snapshot(
        monkeypatch,
        lumbro_sources,
        (_synthetic_spec_record(color="Azul", row=20),),
        (_synthetic_price_record(color="Rojo", price="100", page=1),),
    )

    blue = next(
        item for item in snapshot["items"] if item["attributes"]["color"] == "Azul"
    )
    assert blue["price_net"] == "0.000000"
    assert blue["code_status"] == "needs_review"
    assert blue["sku"] == ""


def test_link_resolution_preserves_truthful_status_and_label(lumbro_build):
    items = lumbro_build.snapshot["items"]
    barcelona = next(
        item
        for item in items
        if item["attributes"]["source_code"] == "BARCELONA"
    )
    venecia = next(
        item for item in items if item["attributes"]["model"] == "Venecia"
    )
    octa = next(
        item for item in items if item["attributes"]["model"] == "Torre Octa"
    )

    assert barcelona["attributes"]["product_url_match"]["status"] == "collection_index"
    assert barcelona["attributes"]["product_url_match"]["label"] == "Ver catálogo Lumbro"
    assert venecia["attributes"]["product_url_match"]["status"] == "exact_index"
    assert venecia["attributes"]["product_url_match"]["label"] == "Ver producto"
    assert octa["attributes"]["product_url_match"]["status"] == "catalog_fallback"
    assert octa["attributes"]["product_url_match"]["label"] == "Ver catálogo Lumbro"


def test_every_price_row_has_an_audit_disposition(lumbro_build):
    coverage = lumbro_build.snapshot["metadata"]["coverage"]

    assert coverage["parsed_price_rows"] == (
        coverage["imported_rows"]
        + coverage["reconciled_rows"]
        + coverage["excluded_rows"]
    )
    assert len(coverage["exclusions"]) == coverage["excluded_rows"]
    assert all(row["reason"] for row in coverage["exclusions"])


def test_asset_merge_has_one_binding_per_item_and_no_orphans(lumbro_build):
    bindings = lumbro_build.bindings
    item_by_id = {
        item["internal_id"]: item for item in lumbro_build.snapshot["items"]
    }

    assert len({binding.internal_id for binding in bindings}) == len(bindings)
    assert set(lumbro_build.assets_by_sha256) == {
        binding.asset_sha256 for binding in bindings
    }
    assert all(binding.internal_id in item_by_id for binding in bindings)
    assert all(binding.object_name == f"{binding.asset_sha256}.png" for binding in bindings)
    assert all(
        item_by_id[binding.internal_id]["attributes"]["image_match"]["asset_sha256"]
        == binding.asset_sha256
        for binding in bindings
    )
