from decimal import Decimal
from io import BytesIO
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

from mobiliti_saas.quote_engine import engine
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from mobiliti_saas.quote_engine.quotation_sheets import (
    _region,
    inline_source_shared_strings,
)


_XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
_OFFICE_REL = (
    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
)
_PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def _png_bytes(color: str) -> bytes:
    output = BytesIO()
    Image.new("RGB", (32, 24), color).save(output, format="PNG")
    return output.getvalue()


def _source_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet["A1"] = "SUNON TECHNOLOGY CO.,LTD."
    sheet["A7"] = "No."
    sheet["B7"] = "Item Name"
    sheet["C7"] = "Photo"
    sheet["J7"] = "Unit Price"
    sheet["K7"] = "Tot.Price"
    sheet["A8"] = "- RECEPCION"
    sheet["A8"].font = Font(bold=True)
    sheet["A8"].fill = PatternFill("solid", fgColor="4F81BD")
    sheet["A9"] = 1
    sheet["B9"] = "Producto original"
    sheet["E9"] = "100 x 50 cm"
    sheet["G9"] = 1
    sheet["J9"] = 100
    sheet["K9"] = "=G9*J9"
    sheet["A9"].fill = PatternFill("solid", fgColor="FFF2CC")
    image = WorkbookImage(BytesIO(_png_bytes("blue")))
    image.anchor = "C9"
    sheet.add_image(image)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _source_workbook_with_oversized_product_image() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet["A7"] = "No."
    sheet["B7"] = "Item Name"
    sheet["C7"] = "Photo"
    sheet["D7"] = "Description"
    sheet["E7"] = "Dimension"
    sheet["G7"] = "Q'ty"
    sheet["J7"] = "Unit Price"
    sheet["K7"] = "Tot.Price"
    sheet["A8"] = "- RECEPCION"
    sheet["A9"] = 1
    sheet["B9"] = "Producto grande"
    sheet["D9"] = "Descripcion"
    sheet["E9"] = "100 x 50 cm"
    sheet["G9"] = 1
    sheet["J9"] = 100
    sheet["K9"] = "=G9*J9"
    sheet.column_dimensions["C"].width = 20
    sheet.row_dimensions[9].height = 90
    image = WorkbookImage(BytesIO(_png_bytes("blue")))
    image.width = 900
    image.height = 700
    image.anchor = "C9"
    sheet.add_image(image)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _source_workbook_with_unused_ignorable_prefix() -> bytes:
    source = _source_workbook()
    output = BytesIO()
    with ZipFile(BytesIO(source), "r") as source_zip:
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as target_zip:
            for info in source_zip.infolist():
                content = source_zip.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    content = content.replace(
                        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
                        b'spreadsheetml/2006/main">',
                        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
                        b'spreadsheetml/2006/main" '
                        b'xmlns:mc="http://schemas.openxmlformats.org/'
                        b'markup-compatibility/2006" mc:Ignorable="x14ac" '
                        b'xmlns:x14ac="http://schemas.microsoft.com/office/'
                        b'spreadsheetml/2009/9/ac">',
                        1,
                    )
                target_zip.writestr(info, content)
    return output.getvalue()


def _rich_source_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet.merge_cells("A1:L1")
    sheet["A1"] = "SUNON TECHNOLOGY CO.,LTD."
    for coordinate, value in {
        "A7": "No.",
        "B7": "Item Name",
        "C7": "Photo",
        "D7": "Description",
        "E7": "Dimension",
        "G7": "Q'ty",
        "H7": "Vol.",
        "I7": "Tot.Vol.",
        "J7": "Unit Price",
        "K7": "Tot.Price",
        "L7": "Remark",
    }.items():
        sheet[coordinate] = value
    sheet.freeze_panes = "A8"
    sheet.column_dimensions["C"].width = 27.5
    sheet.column_dimensions["D"].width = 45.25
    sheet.column_dimensions["N"].width = 4.5

    sheet.merge_cells("A8:D8")
    sheet["A8"] = "- RECEPCION"
    sheet["A8"].font = Font(bold=True, color="FFFFFF")
    sheet["A8"].fill = PatternFill("solid", fgColor="4F81BD")
    sheet["I8"] = "- RECEPCION Tot.Price :"
    sheet["J8"] = "=SUM(K9:K9)"
    sheet.row_dimensions[8].height = 24

    sheet["A9"] = 1
    sheet["B9"] = "Producto Recepcion"
    sheet["D9"] = "Descripcion Producto Recepcion"
    sheet["E9"] = "100 x 50 cm"
    sheet["G9"] = 2
    sheet["H9"] = 0.5
    sheet["I9"] = "=G9*H9"
    sheet["J9"] = 125
    sheet["K9"] = "=G9*J9"
    sheet["A9"].fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.row_dimensions[9].height = 120
    reception_image = WorkbookImage(BytesIO(_png_bytes("blue")))
    reception_image.width = 37
    reception_image.height = 29
    reception_image.anchor = "C9"
    sheet.add_image(reception_image)

    sheet.merge_cells("A10:D10")
    sheet["A10"] = "- PRIVADOS"
    sheet["A10"].font = Font(bold=True, color="FFFFFF")
    sheet["A10"].fill = PatternFill("solid", fgColor="9BBB59")
    sheet["I10"] = "- PRIVADOS Tot.Price :"
    sheet["J10"] = "=SUM(K11:K11)"
    sheet.row_dimensions[10].height = 30

    sheet["A11"] = 2
    sheet["B11"] = "Producto Privado"
    sheet["D11"] = "Descripcion Producto Privado"
    sheet["E11"] = "100 x 50 cm"
    sheet["G11"] = 3
    sheet["H11"] = 0.75
    sheet["I11"] = "=G11*H11"
    sheet["J11"] = 200
    sheet["K11"] = "=G11*J11"
    sheet["A11"].fill = PatternFill("solid", fgColor="F4CCCC")
    sheet.row_dimensions[11].height = 240
    private_image = WorkbookImage(BytesIO(_png_bytes("red")))
    private_image.width = 71
    private_image.height = 33
    private_image.anchor = "C11"
    sheet.add_image(private_image)

    sheet["I12"] = "=SUM(I9:I11)"
    sheet["J12"] = "TOTAL"
    sheet["K12"] = "=SUM(K9:K11)"
    sheet["I12"].fill = PatternFill("solid", fgColor="D9EAD3")
    sheet.row_dimensions[12].height = 28
    sheet.merge_cells("A14:L14")
    sheet["A14"] = "TERMINOS Y CONDICIONES"
    sheet["A14"].font = Font(bold=True, color="FFFFFF")
    sheet["A14"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.row_dimensions[14].height = 55
    sheet["A15"] = "1. Vigencia de la cotizacion: 30 dias."
    sheet.merge_cells("A16:F16")
    sheet["A16"] = "ELABORO"
    sheet.row_dimensions[16].height = 36
    sheet["N200"].fill = PatternFill("solid", fgColor="FFFFFF")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _structured_source_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    headers = (
        "No.",
        "Item Name",
        "Photo",
        "Description",
        "Dimension",
        "Color",
        "Q'ty",
        "Vol.",
        "Tot.Vol.",
        "Unit Price",
        "Tot.Price",
        "Remark",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(7, column).value = header
    sheet["A8"] = "- RECEPCION"
    sheet["A9"] = 1
    sheet["B9"] = "Producto Recepcion"
    sheet["D9"] = "Descripcion Recepcion"
    sheet["E9"] = "100 x 50 cm"
    sheet["F9"] = "Nogal"
    sheet["G9"] = 2
    sheet["H9"] = 0.5
    sheet["I9"] = "=G9*H9"
    sheet["J9"] = 125
    sheet["K9"] = "=G9*J9"
    sheet["L9"] = "Entrega 8 semanas"
    sheet["A10"] = "- PRIVADOS"
    sheet["A11"] = 2
    sheet["B11"] = "Producto Privado"
    sheet["D11"] = "Descripcion Privado"
    sheet["E11"] = "120 x 60 cm"
    sheet["F11"] = "Negro"
    sheet["G11"] = 3
    sheet["H11"] = 0.75
    sheet["I11"] = "=G11*H11"
    sheet["J11"] = 200
    sheet["K11"] = "=G11*J11"
    sheet["L11"] = "Entrega 10 semanas"

    table = Table(displayName="QuotationTable", ref="A7:L11")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    validation = DataValidation(
        type="list",
        formula1="=$D$9:$D$11",
    )
    sheet.add_data_validation(validation)
    validation.add("D9:D11")
    sheet.conditional_formatting.add(
        "D9:D11",
        FormulaRule(
            formula=['D9<>""'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return engine._normalized_quotation_snapshot(output.getvalue())


def _drawing_anchors(content: bytes) -> dict[int, dict[str, object]]:
    with ZipFile(BytesIO(content), "r") as workbook:
        drawing_part = next(
            name
            for name in workbook.namelist()
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        )
        root = ET.fromstring(workbook.read(drawing_part))
    result: dict[int, dict[str, object]] = {}
    for anchor in list(root):
        marker = anchor.find(f"{_XDR}from")
        picture = anchor.find(f"{_XDR}pic")
        extent = anchor.find(f"{_XDR}ext")
        if marker is None or picture is None or extent is None:
            continue
        row = int(marker.findtext(f"{_XDR}row")) + 1
        non_visual = picture.find(f".//{_XDR}cNvPr")
        assert non_visual is not None
        result[row] = {
            "col": int(marker.findtext(f"{_XDR}col")),
            "colOff": marker.findtext(f"{_XDR}colOff"),
            "rowOff": marker.findtext(f"{_XDR}rowOff"),
            "cx": extent.attrib["cx"],
            "cy": extent.attrib["cy"],
            "name": non_visual.attrib["name"],
        }
    return result


def _sheet_dimension(content: bytes) -> str:
    with ZipFile(BytesIO(content), "r") as workbook:
        root = ET.fromstring(workbook.read("xl/worksheets/sheet1.xml"))
    dimension = root.find(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dimension"
    )
    assert dimension is not None
    return dimension.attrib["ref"]


def _source_workbook_with_linked_image_fallback() -> bytes:
    source = _source_workbook()
    output = BytesIO()
    with ZipFile(BytesIO(source), "r") as source_zip:
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as target_zip:
            for info in source_zip.infolist():
                content = source_zip.read(info.filename)
                if info.filename == "xl/drawings/drawing1.xml":
                    root = ET.fromstring(content)
                    blip = root.find(
                        ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
                    )
                    assert blip is not None
                    blip.attrib[f"{_OFFICE_REL}link"] = "rIdLinkedImage"
                    content = ET.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )
                elif info.filename == "xl/drawings/_rels/drawing1.xml.rels":
                    root = ET.fromstring(content)
                    ET.SubElement(
                        root,
                        f"{_PKG_REL}Relationship",
                        {
                            "Id": "rIdLinkedImage",
                            "Type": (
                                "http://schemas.openxmlformats.org/"
                                "officeDocument/2006/relationships/image"
                            ),
                            "Target": "file:///C:/Temp/vendor.png",
                            "TargetMode": "External",
                        },
                    )
                    content = ET.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    )
                target_zip.writestr(info, content)
    return output.getvalue()


def _line(
    key: str,
    *,
    origin: str,
    source_row: int | None,
    title: str,
    name: str,
    image: bytes | None = None,
    cost: Decimal = Decimal("50"),
    quantity: Decimal = Decimal("1"),
    section_id: str | None = None,
) -> engine._OfficialPresentationLine:
    return engine._OfficialPresentationLine(
        item_key=key,
        section_id=section_id or f"section-{' '.join(title.split()).casefold()}",
        section_title=title,
        item=None,
        name=name,
        description=f"Descripcion {name}",
        dimensions="100 x 50 cm",
        m3=Decimal("0.25"),
        quantity=quantity,
        category="Silla",
        provider="Offiho" if origin != "imported" else "Sunon Inc",
        region="Centro",
        original_currency="USD",
        original_cost=cost,
        frozen_rate=Decimal("1"),
        converted_cost=cost,
        origin=origin,
        source_row=source_row,
        upstream_row_hash="",
        image_content=image,
        image_content_type="image/png" if image is not None else None,
    )


def test_provider_and_region_are_normalized_to_the_official_template_values():
    assert engine._official_mobiliti_provider("SUNON TECHNOLOGY CO.,LTD.") == "Sunon Inc"
    assert engine._official_mobiliti_provider("ALMA") == "Alma - Exterior"
    assert engine._official_mobiliti_provider("CR Global") == (
        "CR Global - Mesas Ajustables"
    )
    assert engine._official_mobiliti_provider("yabo - hoteleria") == "Yabo - Hotelería"
    assert engine._official_mobiliti_provider("texto libre") == "Proveedor Externo"
    assert _region({"catalog": "sunon"}, "sunon") == "Centro"
    assert _region({"region": "Occidente"}, "imported") == "Occidente"


def test_original_quotation_keeps_one_final_project_list_without_stale_imports():
    original = _source_workbook()
    imported = _line(
        "imported-1",
        origin="imported",
        source_row=9,
        title="RECEPCION",
        name="Producto original",
        image=_png_bytes("blue"),
        cost=Decimal("125"),
        quantity=Decimal("2"),
    )
    catalog = _line(
        "catalog-1",
        origin="offiho",
        source_row=None,
        title="RECEPCION",
        name="Producto Offiho",
        image=_png_bytes("green"),
    )

    augmented, quotation_rows, quotation_rates = engine._augment_original_quotation(
        original,
        (imported, catalog),
    )

    workbook = load_workbook(BytesIO(augmented), data_only=False)
    sheet = workbook["Quotation"]
    try:
        assert sheet["A1"].value == "SUNON TECHNOLOGY CO.,LTD."
        assert sheet["A8"].value == "- RECEPCION"
        assert sheet["A8"].font.bold is True
        assert sheet["A8"].fill.fgColor.rgb == "004F81BD"
        assert sheet["B9"].value == "Producto original"
        assert sheet["G9"].value == 2
        assert sheet["J9"].value == 125
        assert sheet["K9"].value == "=G9*J9"
        assert sheet["A9"].fill.fgColor.rgb == "00FFF2CC"
        assert quotation_rows["imported-1"] == 9
        assert quotation_rates["imported-1"] == Decimal("1")
        appended_row = quotation_rows["catalog-1"]
        assert sheet.cell(appended_row, 2).value == "Producto Offiho"
        assert sheet.cell(appended_row, 7).value == 1
        assert sheet.cell(appended_row, 10).value == 50
        assert sheet.cell(appended_row, 11).value == f"=G{appended_row}*J{appended_row}"
        section_rows = [
            row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "- RECEPCION"
        ]
        product_rows = [
            row
            for row in range(1, sheet.max_row + 1)
            if isinstance(sheet.cell(row, 1).value, (int, float))
            and sheet.cell(row, 2).value
        ]
        assert section_rows == [8]
        assert product_rows == [9, 10]
        assert len(sheet._images) == 2
    finally:
        workbook.close()


def test_original_quotation_replaces_preformatted_blank_rows_when_project_expands():
    workbook = load_workbook(BytesIO(_source_workbook()))
    sheet = workbook["Quotation"]
    for row in range(10, 65):
        for column in range(1, 13):
            sheet.cell(row, column).fill = PatternFill(
                "solid",
                fgColor="FFFFFF",
            )
    source = BytesIO()
    workbook.save(source)
    workbook.close()

    imported = _line(
        "imported-1",
        origin="imported",
        source_row=9,
        title="RECEPCION",
        name="Producto original",
        cost=Decimal("125"),
    )
    catalog = _line(
        "catalog-1",
        origin="alma",
        source_row=None,
        title="RECEPCION",
        name="Producto ALMA",
    )

    augmented, _, _ = engine._augment_original_quotation(
        source.getvalue(),
        (imported, catalog),
    )
    package = XlsxPackage.from_bytes(augmented)
    quotation_part = package.sheet_part("Quotation")

    inline_source_shared_strings(
        package.parts[quotation_part],
        package.shared_strings(),
    )


def test_original_quotation_inserts_catalog_line_without_rebuilding_supplier_layout():
    original = _rich_source_workbook()
    original_anchors = _drawing_anchors(original)
    imported_reception = _line(
        "imported-reception",
        origin="imported",
        source_row=9,
        title="RECEPCION",
        name="Producto Recepcion",
        image=_png_bytes("blue"),
        cost=Decimal("125"),
        quantity=Decimal("2"),
    )
    catalog = _line(
        "catalog-reception",
        origin="offiho",
        source_row=None,
        title="RECEPCION",
        name="Producto Offiho",
        image=_png_bytes("green"),
        cost=Decimal("50"),
    )
    imported_private = _line(
        "imported-private",
        origin="imported",
        source_row=11,
        title="PRIVADOS",
        name="Producto Privado",
        image=_png_bytes("red"),
        cost=Decimal("200"),
        quantity=Decimal("3"),
    )

    augmented, rows, rates = engine._augment_original_quotation(
        original,
        (imported_reception, catalog, imported_private),
    )

    assert rows == {
        "imported-reception": 9,
        "catalog-reception": 10,
        "imported-private": 12,
    }
    assert rates == {key: Decimal("1") for key in rows}
    assert _sheet_dimension(augmented) == "A1:N200"
    workbook = load_workbook(BytesIO(augmented), data_only=False)
    sheet = workbook["Quotation"]
    try:
        assert sheet.freeze_panes == "A8"
        assert sheet.column_dimensions["C"].width == 27.5
        assert sheet.column_dimensions["D"].width == 45.25
        assert sheet.column_dimensions["N"].width == 4.5
        assert {str(item) for item in sheet.merged_cells.ranges} == {
            "A1:L1",
            "A8:D8",
            "A11:D11",
            "A15:L15",
            "A17:F17",
        }
        assert sheet["A8"].value == "- RECEPCION"
        assert sheet["A11"].value == "- PRIVADOS"
        assert sheet["B9"].value == "Producto Recepcion"
        assert sheet["B10"].value == "Producto Offiho"
        assert sheet["B12"].value == "Producto Privado"
        assert sheet["A8"].fill.fgColor.rgb == "004F81BD"
        assert sheet["A11"].fill.fgColor.rgb == "009BBB59"
        assert sheet["A9"].fill.fgColor.rgb == "00FFF2CC"
        assert sheet["A12"].fill.fgColor.rgb == "00F4CCCC"
        assert sheet.row_dimensions[9].height == 120
        assert sheet.row_dimensions[12].height == 240
        assert sheet.row_dimensions[13].height == 28
        assert sheet.row_dimensions[15].height == 55
        assert sheet.row_dimensions[17].height == 36
        assert sheet["K9"].value == "=G9*J9"
        assert sheet["K10"].value == "=G10*J10"
        assert sheet["K12"].value == "=G12*J12"
        assert sheet["J8"].value == "=SUM(K9:K10)"
        assert sheet["J11"].value == "=SUM(K12:K12)"
        assert sheet["I13"].value == "=SUM(I9:I12)"
        assert sheet["K13"].value == "=SUM(K9:K12)"
        assert sheet["A15"].value == "TERMINOS Y CONDICIONES"
        assert sheet["A16"].value == "1. Vigencia de la cotizacion: 30 dias."
        assert sheet["A17"].value == "ELABORO"
    finally:
        workbook.close()
    result_anchors = _drawing_anchors(augmented)
    assert set(result_anchors) == {9, 10, 12}
    with ZipFile(BytesIO(augmented), "r") as package:
        quotation_root = ET.fromstring(package.read("xl/worksheets/sheet1.xml"))
    column_width_emu = engine._quotation_column_width_emu(quotation_root, 3)
    for source_row, target_row in ((9, 9), (11, 12)):
        source_anchor = original_anchors[source_row]
        result_anchor = result_anchors[target_row]
        assert {
            key: result_anchor[key]
            for key in ("col", "cx", "cy", "name")
        } == {
            key: source_anchor[key]
            for key in ("col", "cx", "cy", "name")
        }
        row_height_emu = engine._quotation_row_height_emu(
            quotation_root,
            target_row,
        )
        assert int(result_anchor["colOff"]) >= 0
        assert int(result_anchor["rowOff"]) >= 0
        assert (
            int(result_anchor["colOff"]) + int(result_anchor["cx"])
            <= column_width_emu
        )
        assert (
            int(result_anchor["rowOff"]) + int(result_anchor["cy"])
            <= row_height_emu
        )
    assert result_anchors[10]["col"] == 2


def test_original_quotation_contains_oversized_product_image_inside_photo_cell():
    original = _source_workbook_with_oversized_product_image()
    imported = _line(
        "imported-large",
        origin="imported",
        source_row=9,
        title="RECEPCION",
        name="Producto grande",
        image=_png_bytes("blue"),
        cost=Decimal("100"),
    )

    augmented, rows, _rates = engine._augment_original_quotation(
        original,
        (imported,),
    )

    assert rows == {"imported-large": 9}
    anchor = _drawing_anchors(augmented)[9]
    assert anchor["col"] == 2
    with ZipFile(BytesIO(augmented), "r") as package:
        quotation_root = ET.fromstring(package.read("xl/worksheets/sheet1.xml"))
    column_width_emu = engine._quotation_column_width_emu(
        quotation_root,
        3,
    )
    row_height_emu = engine._quotation_row_height_emu(
        quotation_root,
        9,
    )
    assert int(anchor["colOff"]) >= 0
    assert int(anchor["rowOff"]) >= 0
    assert int(anchor["colOff"]) + int(anchor["cx"]) <= column_width_emu
    assert int(anchor["rowOff"]) + int(anchor["cy"]) <= row_height_emu


def test_quotation_normalization_removes_external_link_when_image_is_embedded():
    normalized = engine._normalized_quotation_snapshot(
        _source_workbook_with_linked_image_fallback()
    )

    with ZipFile(BytesIO(normalized), "r") as workbook:
        drawing = ET.fromstring(workbook.read("xl/drawings/drawing1.xml"))
        relationships = ET.fromstring(
            workbook.read("xl/drawings/_rels/drawing1.xml.rels")
        )
    assert all(
        f"{_OFFICE_REL}link" not in blip.attrib
        for blip in drawing.findall(
            ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
        )
    )
    assert all(
        relationship.attrib.get("TargetMode", "").casefold() != "external"
        for relationship in relationships
        if relationship.attrib.get("Type", "").endswith("/image")
    )


def test_quotation_extension_preserves_unused_mc_namespace_declarations():
    original = _source_workbook_with_unused_ignorable_prefix()
    catalog = _line(
        "catalog-1",
        origin="offiho",
        source_row=None,
        title="PRIVADOS",
        name="Producto Offiho",
    )

    snapshot, _, _ = engine._augment_original_quotation(original, (catalog,))

    with ZipFile(BytesIO(snapshot), "r") as workbook:
        worksheet = workbook.read("xl/worksheets/sheet1.xml")
    assert b'mc:Ignorable="x14ac"' in worksheet
    assert b'xmlns:x14ac="http://schemas.microsoft.com/office/' in worksheet


def test_transformation_column_is_inserted_without_damaging_rich_quotation():
    source_workbook = load_workbook(
        BytesIO(_rich_source_workbook()),
        data_only=False,
    )
    source_sheet = source_workbook["Quotation"]
    source_sheet["F7"] = "Color"
    source_sheet["F9"] = "Nogal"
    source_sheet["F11"] = "Negro"
    source_sheet["L9"] = "Entrega 8 semanas"
    source_sheet["L11"] = "Entrega 10 semanas"
    source_sheet["B12"] = "=SUM($K$9:$K$11)"
    source_sheet["B13"] = "=D9"
    for column, width in {
        "E": 18.25,
        "F": 13.5,
        "G": 9.25,
        "H": 11.75,
        "I": 12.5,
        "J": 16.25,
        "K": 17.5,
        "L": 21.25,
    }.items():
        source_sheet.column_dimensions[column].width = width
    source_output = BytesIO()
    source_workbook.save(source_output)
    source_workbook.close()
    source = engine._normalized_quotation_snapshot(source_output.getvalue())
    source_anchors = _drawing_anchors(source)

    transformed = engine._insert_quotation_transformation_column(
        source,
        {
            "reception": 9,
            "private": 11,
        },
        {
            "reception": "Descripcion transformada recepcion",
            "private": "Descripcion transformada privado",
        },
    )

    assert _sheet_dimension(transformed) == "A1:O200"
    assert _drawing_anchors(transformed) == source_anchors
    workbook = load_workbook(BytesIO(transformed), data_only=False)
    sheet = workbook["Quotation"]
    try:
        assert sheet.freeze_panes == "A8"
        assert tuple(sheet.cell(7, column).value for column in range(4, 14)) == (
            "Trasformacion",
            "Description",
            "Dimension",
            "Color",
            "Q'ty",
            "Vol.",
            "Tot.Vol.",
            "Unit Price",
            "Tot.Price",
            "Remark",
        )
        assert tuple(sheet.cell(9, column).value for column in range(4, 14)) == (
            "Descripcion transformada recepcion",
            "Descripcion Producto Recepcion",
            "100 x 50 cm",
            "Nogal",
            2,
            0.5,
            "=H9*I9",
            125,
            "=H9*K9",
            "Entrega 8 semanas",
        )
        assert tuple(sheet.cell(11, column).value for column in range(4, 14)) == (
            "Descripcion transformada privado",
            "Descripcion Producto Privado",
            "100 x 50 cm",
            "Negro",
            3,
            0.75,
            "=H11*I11",
            200,
            "=H11*K11",
            "Entrega 10 semanas",
        )
        assert sheet["J8"].value == "- RECEPCION Tot.Price :"
        assert sheet["K8"].value == "=SUM(L9:L9)"
        assert sheet["J10"].value == "- PRIVADOS Tot.Price :"
        assert sheet["K10"].value == "=SUM(L11:L11)"
        assert sheet["J12"].value == "=SUM(J9:J11)"
        assert sheet["K12"].value == "TOTAL"
        assert sheet["L12"].value == "=SUM(L9:L11)"
        assert sheet["B12"].value == "=SUM($L$9:$L$11)"
        assert sheet["B13"].value == "=E9"
        assert {str(item) for item in sheet.merged_cells.ranges} == {
            "A1:M1",
            "A8:E8",
            "A10:E10",
            "A14:M14",
            "A16:G16",
        }
        assert {
            column: sheet.column_dimensions[column].width
            for column in ("D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "O")
        } == {
            "D": 45.25,
            "E": 45.25,
            "F": 18.25,
            "G": 13.5,
            "H": 9.25,
            "I": 11.75,
            "J": 12.5,
            "K": 16.25,
            "L": 17.5,
            "M": 21.25,
            "O": 4.5,
        }
    finally:
        workbook.close()


def test_transformation_column_expands_quotation_table_contract():
    transformed = engine._insert_quotation_transformation_column(
        _structured_source_workbook(),
        {
            "reception": 9,
            "private": 11,
        },
        {
            "reception": "Descripcion transformada recepcion",
            "private": "Descripcion transformada privado",
        },
    )

    with ZipFile(BytesIO(transformed), "r") as workbook:
        table_part = next(
            name
            for name in workbook.namelist()
            if name.startswith("xl/tables/") and name.endswith(".xml")
        )
        table = ET.fromstring(workbook.read(table_part))
    columns = table.find(f"{_MAIN}tableColumns")
    assert columns is not None
    assert table.attrib["ref"] == "A7:M11"
    auto_filter = table.find(f"{_MAIN}autoFilter")
    assert auto_filter is not None
    assert auto_filter.attrib["ref"] == "A7:M11"
    assert columns.attrib["count"] == "13"
    assert [
        column.attrib["name"]
        for column in columns.findall(f"{_MAIN}tableColumn")
    ] == [
        "No.",
        "Item Name",
        "Photo",
        "Trasformacion",
        "Description",
        "Dimension",
        "Color",
        "Q'ty",
        "Vol.",
        "Tot.Vol.",
        "Unit Price",
        "Tot.Price",
        "Remark",
    ]


def test_transformation_column_shifts_validation_and_conditional_formulas():
    transformed = engine._insert_quotation_transformation_column(
        _structured_source_workbook(),
        {
            "reception": 9,
            "private": 11,
        },
        {
            "reception": "Descripcion transformada recepcion",
            "private": "Descripcion transformada privado",
        },
    )

    with ZipFile(BytesIO(transformed), "r") as workbook:
        worksheet = ET.fromstring(
            workbook.read("xl/worksheets/sheet1.xml")
        )
    validation = worksheet.find(
        f"{_MAIN}dataValidations/{_MAIN}dataValidation"
    )
    assert validation is not None
    assert validation.attrib["sqref"] == "E9:E11"
    assert validation.findtext(f"{_MAIN}formula1") == "=$E$9:$E$11"
    conditional = worksheet.find(f"{_MAIN}conditionalFormatting")
    assert conditional is not None
    assert conditional.attrib["sqref"] == "E9:E11"
    rule = conditional.find(f"{_MAIN}cfRule")
    assert rule is not None
    assert rule.findtext(f"{_MAIN}formula") == 'E9<>""'


def test_transformation_column_insertion_is_idempotent_and_updates_values():
    source = engine._normalized_quotation_snapshot(_rich_source_workbook())
    rows = {
        "reception": 9,
        "private": 11,
    }
    first = engine._insert_quotation_transformation_column(
        source,
        rows,
        {
            "reception": "Primera recepcion",
            "private": "Primera privado",
        },
    )

    second = engine._insert_quotation_transformation_column(
        first,
        rows,
        {
            "reception": "Segunda recepcion",
            "private": "Segunda privado",
        },
    )

    assert _sheet_dimension(second) == "A1:O200"
    workbook = load_workbook(BytesIO(second), data_only=False)
    sheet = workbook["Quotation"]
    try:
        headers = [
            sheet.cell(7, column).value
            for column in range(1, sheet.max_column + 1)
        ]
        assert headers.count("Trasformacion") == 1
        assert sheet["D7"].value == "Trasformacion"
        assert sheet["D9"].value == "Segunda recepcion"
        assert sheet["D11"].value == "Segunda privado"
        assert sheet["E9"].value == "Descripcion Producto Recepcion"
        assert sheet["E11"].value == "Descripcion Producto Privado"
        assert sheet["F9"].value == "100 x 50 cm"
        assert sheet["F11"].value == "100 x 50 cm"
    finally:
        workbook.close()


def test_transformation_column_shifts_comments_and_only_quotation_defined_names():
    workbook = load_workbook(
        BytesIO(_rich_source_workbook()),
        data_only=False,
    )
    sheet = workbook["Quotation"]
    sheet["E9"].comment = Comment("Medida confirmada", "QA")
    workbook.defined_names.add(
        DefinedName(
            "QuotationRange",
            attr_text="Quotation!$D$9:$D$11",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "OtherSheetRange",
            attr_text="NotQuotation!$D$1",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "ExternalQuotationRange",
            attr_text="[Book.xlsx]Quotation!$D$1",
        )
    )
    source = BytesIO()
    workbook.save(source)
    workbook.close()

    transformed = engine._insert_quotation_transformation_column(
        engine._normalized_quotation_snapshot(source.getvalue()),
        {
            "reception": 9,
            "private": 11,
        },
        {
            "reception": "Descripcion transformada recepcion",
            "private": "Descripcion transformada privado",
        },
    )

    result = load_workbook(BytesIO(transformed), data_only=False)
    try:
        sheet = result["Quotation"]
        assert sheet["E9"].comment is None
        assert sheet["F9"].comment is not None
        assert sheet["F9"].comment.text == "Medida confirmada"
        assert (
            result.defined_names["QuotationRange"].attr_text
            == "Quotation!$E$9:$E$11"
        )
        assert (
            result.defined_names["OtherSheetRange"].attr_text
            == "NotQuotation!$D$1"
        )
        assert (
            result.defined_names["ExternalQuotationRange"].attr_text
            == "[Book.xlsx]Quotation!$D$1"
        )
    finally:
        result.close()

    with ZipFile(BytesIO(transformed), "r") as package:
        vml_part = next(
            name
            for name in package.namelist()
            if name.startswith("xl/drawings/") and name.endswith(".vml")
        )
        vml = ET.fromstring(package.read(vml_part))
    comment_columns = [
        element.text
        for element in vml.iter()
        if element.tag.rsplit("}", 1)[-1] == "Column"
    ]
    assert comment_columns == ["5"]


def test_reaugmenting_a_transformed_quotation_keeps_one_column_layout():
    lines = (
        _line(
            "reception",
            origin="imported",
            source_row=9,
            title="RECEPCION",
            name="Producto Recepcion",
            cost=Decimal("125"),
            quantity=Decimal("2"),
        ),
        _line(
            "private",
            origin="imported",
            source_row=11,
            title="PRIVADOS",
            name="Producto Privado",
            cost=Decimal("200"),
            quantity=Decimal("3"),
        ),
    )
    first_snapshot, first_rows, _ = engine._augment_original_quotation(
        _rich_source_workbook(),
        lines,
    )
    first = engine._insert_quotation_transformation_column(
        first_snapshot,
        first_rows,
        {
            "reception": "Primera recepcion",
            "private": "Primera privado",
        },
    )

    second_snapshot, second_rows, _ = engine._augment_original_quotation(
        first,
        lines,
    )
    second = engine._insert_quotation_transformation_column(
        second_snapshot,
        second_rows,
        {
            "reception": "Segunda recepcion",
            "private": "Segunda privado",
        },
    )

    workbook = load_workbook(BytesIO(second), data_only=False)
    sheet = workbook["Quotation"]
    try:
        headers = [
            sheet.cell(7, column).value
            for column in range(1, sheet.max_column + 1)
        ]
        assert headers.count("Trasformacion") == 1
        assert tuple(sheet.cell(9, column).value for column in range(4, 13)) == (
            "Segunda recepcion",
            "Descripcion Producto Recepcion",
            "100 x 50 cm",
            None,
            2,
            0.25,
            "=H9*I9",
            125,
            "=H9*K9",
        )
        assert tuple(sheet.cell(11, column).value for column in range(4, 13)) == (
            "Segunda privado",
            "Descripcion Producto Privado",
            "100 x 50 cm",
            None,
            3,
            0.25,
            "=H11*I11",
            200,
            "=H11*K11",
        )
    finally:
        workbook.close()
