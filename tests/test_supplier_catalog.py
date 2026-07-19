from copy import deepcopy
from datetime import date, datetime, timedelta
from decimal import Decimal, localcontext
import hashlib
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from mobiliti_saas.quote_engine import catalog_cart
from mobiliti_saas.quote_engine import supplier_catalog as supplier_module
from mobiliti_saas.quote_engine.supplier_catalog import (
    build_supplier_cart_payload,
    create_supplier_quotation_workbook,
    load_supplier_catalog_data,
    resolve_conversion_rate,
    safe_excel_text,
)


PUBLIC_FIELDS = {
    "internal_id", "supplier", "product_key", "sku", "code_status",
    "brand", "collection", "name", "description", "unit",
    "availability_type", "stock", "lead_time", "base_price_options",
    "add_on_options", "base_currency", "price_net", "tax_rate",
    "attributes", "image_url", "image_kind", "product_url", "warnings",
    "source_reference",
}
LUMBRO_OFFICIAL_SKUS = (
    "MULT-LIDO-INT",
    "LIDO.OP-INT",
    "JUMP-1.5M",
    "CAJA-FUS",
)
LUMBRO_EXPECTED_NET = Decimal("5519.07")
LUMBRO_EXPECTED_IVA = Decimal("883.05")
LUMBRO_EXPECTED_TOTAL = Decimal("6402.12")


def catalog_payload(*, supplier="alma", items=None):
    item = {
        "internal_id": "alma:kun:kc8611n01rop",
        "supplier": supplier,
        "product_key": "kun-kc8611n01rop",
        "sku": "KC8611N01ROP",
        "code_status": "verified",
        "brand": "KUN",
        "collection": "KUN",
        "name": "Silla KUN",
        "description": "Silla configurable",
        "unit": "pieza",
        "availability_type": "made_to_order",
        "stock": None,
        "lead_time": "Sobre pedido",
        "base_price_options": [
            {"id": "powder-coated-aluminium", "name": "Powder coated aluminium", "price_net": "250.000000", "available": True},
            {"id": "solid-wood", "name": "Solid wood", "price_net": "300.000000", "available": True},
        ],
        "add_on_options": [
            {"id": "cushion:a-plus", "name": "Cushion A+", "family": "cushion", "price_net": "35.100000", "available": True},
            {"id": "cushion:a-plus-plus", "name": "Cushion A++", "family": "cushion", "price_net": "50.000000", "available": True},
            {"id": "ceramic:a", "name": "Ceramic A", "family": "ceramic", "price_net": "22.500000", "available": True, "compatible_base_option_ids": ["powder-coated-aluminium"]},
            {"id": "ceramic:b", "name": "Ceramic B", "family": "ceramic", "price_net": "30.000000", "available": False},
        ],
        "base_currency": "USD",
        "price_net": "199.990000",
        "tax_rate": "0.160000",
        "attributes": {"color": "Black", "dimensions": "65 x 45 x 83 cm"},
        "image_url": "https://example.test/kun.webp",
        "image_kind": "official",
        "product_url": "https://example.test/kun",
        "warnings": ["Imagen ilustrativa"],
        "source_reference": "SPEC Guide-Alma-KUN.xlsx:E9:I9",
    }
    return {
        "supplier": supplier,
        "source_hash": "a" * 64,
        "generated_at": "2026-07-15T00:00:00Z",
        "items": items or [item],
    }


def lumbro_catalog_payload():
    payload = catalog_payload(supplier="lumbro")
    payload["items"][0].update(
        internal_id="lumbro:variant:barcelona",
        product_key="barcelona",
        sku="BARCELONA",
        brand="Lumbro",
        collection="Empotrables",
        name="Barcelona",
        description="Multicontacto empotrable Barcelona",
        unit="PZA",
        base_price_options=[],
        add_on_options=[],
        base_currency="MXN",
        price_net="2824.000000",
        tax_rate="0.160000",
    )
    payload["metadata"] = {
        "coverage": {
            "parsed_price_rows": 1,
            "imported_rows": 1,
            "reconciled_rows": 0,
            "excluded_rows": 0,
            "exclusions": [],
        }
    }
    return payload


@pytest.fixture
def representative_lumbro_catalog():
    base = lumbro_catalog_payload()["items"][0]
    verified = []
    for index, (sku, price) in enumerate(
        zip(
            LUMBRO_OFFICIAL_SKUS,
            ("3003.000000", "1394.070000", "350.000000", "772.000000"),
            strict=True,
        ),
        start=1,
    ):
        row = deepcopy(base)
        row.update(
            internal_id=f"lumbro:variant:official-{index}",
            product_key=f"official-{index}",
            sku=sku,
            code_status="verified",
            name=sku,
            description=f"Accesorio oficial Lumbro {sku}",
            price_net=price,
            image_url="",
            image_kind="placeholder",
            product_url="",
            warnings=[],
        )
        row["attributes"] = {"source_code": sku, "configuration": "Interconexion"}
        verified.append(row)
    review = deepcopy(base)
    review.update(
        internal_id="lumbro:variant:needs-review",
        product_key="needs-review",
        sku="",
        code_status="needs_review",
        name="Multicontacto por revisar",
        price_net="3183.000000",
        image_url="",
        image_kind="placeholder",
        product_url="",
        warnings=["Codigo oficial no verificable"],
    )
    review["attributes"] = {"source_code": "", "configuration": "Por revisar"}
    return {
        "supplier": "lumbro",
        "source_hash": "f" * 64,
        "generated_at": "2026-07-18T12:00:00Z",
        "items": [*verified, review],
        "metadata": {
            "coverage": {
                "items": 5,
                "verified_items": 4,
                "needs_review_items": 1,
                "priced_items": 5,
            }
        },
    }


def rate_rows(*, effective_date=None):
    day = effective_date or (date.today() - timedelta(days=1)).isoformat()
    return [
        {"currency": "USD", "effective_date": day, "mxn_per_unit": "18.500000", "retrieved_at": f"{day}T23:00:00Z"},
        {"currency": "EUR", "effective_date": day, "mxn_per_unit": "21.000000", "retrieved_at": f"{day}T23:05:00Z"},
    ]


def test_load_catalog_validates_and_preserves_public_decimal_strings():
    loaded = load_supplier_catalog_data(catalog_payload(), expected_supplier="alma")
    item = loaded["items"][0]

    assert set(item) == PUBLIC_FIELDS
    assert item["price_net"] == "199.990000"
    assert item["tax_rate"] == "0.160000"
    assert item["base_price_options"][0]["price_net"] == "250.000000"
    assert loaded["by_internal_id"][item["internal_id"]] is item


def test_load_catalog_preserves_a_normalized_copy_of_optional_metadata():
    payload = lumbro_catalog_payload()

    loaded = load_supplier_catalog_data(payload, expected_supplier="lumbro")

    assert loaded["metadata"] == payload["metadata"]
    assert loaded["metadata"] is not payload["metadata"]
    assert loaded["metadata"]["coverage"] is not payload["metadata"]["coverage"]


def test_normalized_catalog_can_be_reloaded_for_cache_api_flow_without_trusting_plain_index():
    raw = lumbro_catalog_payload()
    cached = load_supplier_catalog_data(raw, expected_supplier="lumbro")
    cached["by_internal_id"] = {"attacker": raw["items"][0]}

    api_loaded = load_supplier_catalog_data(cached, expected_supplier="lumbro")

    assert api_loaded["items"] == cached["items"]
    assert api_loaded["metadata"] == cached["metadata"]
    assert api_loaded is not cached
    assert api_loaded["by_internal_id"] is not cached["by_internal_id"]
    assert set(api_loaded["by_internal_id"]) == {"lumbro:variant:barcelona"}
    with pytest.raises(ValueError, match="supplier no coincide"):
        load_supplier_catalog_data(cached, expected_supplier="alma")
    plain_with_index = dict(cached)
    plain_with_index["by_internal_id"] = {"attacker": raw["items"][0]}
    with pytest.raises(ValueError, match="raiz.*by_internal_id"):
        load_supplier_catalog_data(plain_with_index)


def test_normalized_catalog_output_remains_valid_cart_input():
    normalized = load_supplier_catalog_data(lumbro_catalog_payload())

    cart = build_supplier_cart_payload(
        [{"internal_id": "lumbro:variant:barcelona", "quantity": "1", "add_on_option_ids": []}],
        normalized,
        "MXN",
        [],
    )

    assert cart["items"][0]["sku"] == "BARCELONA"


def test_metadata_accepts_only_exact_non_boolean_json_scalar_types():
    payload = lumbro_catalog_payload()
    payload["metadata"] = {
        "object": {"list": [None, "texto", 7, 1.25]},
    }

    loaded = load_supplier_catalog_data(payload)

    assert loaded["metadata"] == payload["metadata"]
    payload["items"][0]["attributes"] = {"historical_boolean": True}
    assert load_supplier_catalog_data(payload)["items"][0]["attributes"] == {
        "historical_boolean": True
    }


def test_metadata_rejects_boolean_non_finite_bytes_non_string_keys_and_subclasses():
    class DictSubclass(dict):
        pass

    class ListSubclass(list):
        pass

    class StringSubclass(str):
        pass

    class IntSubclass(int):
        pass

    class FloatSubclass(float):
        pass

    invalid_values = (
        DictSubclass({"value": "x"}),
        {"value": ListSubclass(["x"])},
        {"value": DictSubclass({"nested": "x"})},
        {"value": StringSubclass("x")},
        {"value": IntSubclass(1)},
        {"value": FloatSubclass(1.5)},
        {"value": True},
        {"value": False},
        {"value": float("nan")},
        {"value": float("inf")},
        {"value": float("-inf")},
        {"value": b"bytes"},
        {"value": "\ud800"},
        {1: "non-string-key"},
        {StringSubclass("key"): "string-subclass-key"},
    )
    for metadata in invalid_values:
        payload = lumbro_catalog_payload()
        payload["metadata"] = metadata
        with pytest.raises(ValueError, match="metadata"):
            load_supplier_catalog_data(payload)


def test_metadata_enforces_incremental_byte_node_depth_and_cycle_boundaries():
    maximum_bytes = supplier_module.MAX_METADATA_JSON_BYTES
    encoded_overhead = len('{"value":""}'.encode("utf-8"))
    payload = lumbro_catalog_payload()
    payload["metadata"] = {"value": "x" * (maximum_bytes - encoded_overhead)}
    assert len(load_supplier_catalog_data(payload)["metadata"]["value"]) == (
        maximum_bytes - encoded_overhead
    )
    payload["metadata"] = {"value": "x" * (maximum_bytes - encoded_overhead + 1)}
    with pytest.raises(ValueError, match="metadata.*bytes"):
        load_supplier_catalog_data(payload)

    payload = lumbro_catalog_payload()
    payload["metadata"] = {"values": [None] * (supplier_module.MAX_METADATA_NODES - 2)}
    load_supplier_catalog_data(payload)
    payload["metadata"] = {"values": [None] * (supplier_module.MAX_METADATA_NODES - 1)}
    with pytest.raises(ValueError, match="metadata.*valores"):
        load_supplier_catalog_data(payload)

    at_limit = "leaf"
    for _ in range(supplier_module.MAX_METADATA_DEPTH):
        at_limit = {"level": at_limit}
    payload = lumbro_catalog_payload()
    payload["metadata"] = at_limit
    load_supplier_catalog_data(payload)
    payload["metadata"] = {"level": at_limit}
    with pytest.raises(ValueError, match="metadata.*profundidad"):
        load_supplier_catalog_data(payload)

    circular = {}
    circular["self"] = circular
    payload = lumbro_catalog_payload()
    payload["metadata"] = circular
    with pytest.raises(ValueError, match="metadata.*circular"):
        load_supplier_catalog_data(payload)

    payload = lumbro_catalog_payload()
    payload.pop("metadata")
    assert "metadata" not in load_supplier_catalog_data(payload)


def test_oversized_metadata_string_fails_before_any_utf8_encoding():
    encode_calls = []

    def tracked_encode(chunk):
        encode_calls.append(len(chunk))
        return chunk.encode("utf-8")

    with pytest.raises(ValueError, match="metadata.*bytes"):
        supplier_module._bounded_metadata_string_utf8_size(
            "x" * (supplier_module.MAX_METADATA_JSON_BYTES + 1),
            encode_chunk=tracked_encode,
        )

    assert encode_calls == []


def test_metadata_string_encoding_is_chunked_and_huge_int_fails_before_decimal_render():
    encoded_chunks = []

    def tracked_encode(chunk):
        encoded_chunks.append(len(chunk))
        return chunk.encode("utf-8")

    assert supplier_module._bounded_metadata_string_utf8_size(
        "a" * (supplier_module._METADATA_UTF8_CHUNK_CHARS + 1),
        encode_chunk=tracked_encode,
    ) == supplier_module._METADATA_UTF8_CHUNK_CHARS + 1
    assert encoded_chunks == [supplier_module._METADATA_UTF8_CHUNK_CHARS, 1]

    render_calls = []

    def tracked_render(value):
        render_calls.append(value)
        return str(value)

    with pytest.raises(ValueError, match="metadata.*bytes"):
        supplier_module._bounded_metadata_integer_text_size(
            1 << 10_000,
            maximum_bytes=64,
            render_decimal=tracked_render,
        )

    assert render_calls == []


@pytest.mark.parametrize(
    "metadata",
    [
        [],
        {"unsafe": object()},
        {"oversized": "x" * 300_000},
        {"nested": {"level": {"level": {"level": {"level": {"level": {"level": {"level": {"level": {"level": "too deep"}}}}}}}}}},
    ],
)
def test_load_catalog_rejects_invalid_or_unbounded_metadata(metadata):
    payload = lumbro_catalog_payload()
    payload["metadata"] = metadata

    with pytest.raises(ValueError, match="metadata"):
        load_supplier_catalog_data(payload)


def test_load_catalog_rejects_unknown_root_key_even_with_valid_metadata():
    payload = lumbro_catalog_payload()
    payload["unexpected"] = True

    with pytest.raises(ValueError, match="raiz|inesperado"):
        load_supplier_catalog_data(payload)


def test_unknown_base_currency_is_allowed_only_for_blocked_zero_prices():
    payload = catalog_payload()
    item = payload["items"][0]
    item["base_currency"] = "XXX"
    item["price_net"] = "0.000000"
    item["base_price_options"] = []
    item["add_on_options"] = []

    loaded = load_supplier_catalog_data(payload)

    assert loaded["items"][0]["base_currency"] == "XXX"

    item["price_net"] = "1.000000"
    with pytest.raises(ValueError, match="moneda base.*verificar"):
        load_supplier_catalog_data(payload)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda payload: payload.update(items=[]), "vacio"),
        (lambda payload: payload["items"][0].pop("name"), "name"),
        (lambda payload: payload["items"][0].update(supplier="sunon"), "supplier"),
        (lambda payload: payload["items"][0].update(price_net=0.1), "price_net"),
        (lambda payload: payload["items"][0].update(price_net="NaN"), "price_net"),
        (lambda payload: payload["items"][0].update(sku="", code_status="verified"), "sku"),
        (lambda payload: payload["items"][0].update(sku="INVENTED", code_status="needs_review"), "sku"),
        (lambda payload: payload.update(source_hash="not-a-hash"), "source_hash"),
        (lambda payload: payload.update(generated_at="not-a-date"), "generated_at"),
    ],
)
def test_load_catalog_rejects_malformed_or_mismatched_data(mutation, message):
    payload = catalog_payload()
    mutation(payload)

    with pytest.raises(ValueError, match=message):
        load_supplier_catalog_data(payload, expected_supplier="alma")


def test_load_catalog_rejects_incompatible_duplicate_ids_and_skus():
    duplicate = deepcopy(catalog_payload()["items"][0])
    duplicate["name"] = "Conflicting name"
    with pytest.raises(ValueError, match="internal_id duplicado"):
        load_supplier_catalog_data(catalog_payload(items=[catalog_payload()["items"][0], duplicate]))


@pytest.mark.parametrize("unknown_fields", [{"unexpected": 1}, {"unexpected": 1, "other": 2}])
def test_load_catalog_rejects_one_or_multiple_unknown_item_fields(unknown_fields):
    payload = catalog_payload()
    payload["items"][0].update(unknown_fields)

    with pytest.raises(ValueError, match="campos inesperados"):
        load_supplier_catalog_data(payload)

    duplicate = deepcopy(catalog_payload()["items"][0])
    duplicate["internal_id"] = "alma:kun:another"
    with pytest.raises(ValueError, match="sku duplicado"):
        load_supplier_catalog_data(catalog_payload(items=[catalog_payload()["items"][0], duplicate]))


def test_build_cart_uses_only_catalog_values_and_central_rounding_example():
    catalog = load_supplier_catalog_data(catalog_payload())
    cart = build_supplier_cart_payload(
        [{
            "internal_id": "alma:kun:kc8611n01rop",
            "quantity": "2",
            "base_option_id": "powder-coated-aluminium",
            "add_on_option_ids": ["cushion:a-plus", "ceramic:a"],
        }],
        catalog=catalog,
        quote_currency="EUR",
        rate_rows=rate_rows(),
    )

    line = cart["items"][0]
    assert cart["source_type"] == "supplier_cart"
    assert cart["supplier"] == "alma"
    assert cart["catalog_source_hash"] == "a" * 64
    assert cart["exchange_rate"] == "0.880952"
    assert line["name"] == "Silla KUN"
    assert line["sku"] == "KC8611N01ROP"
    assert line["image_url"] == "https://example.test/kun.webp"
    assert line["unit_price_base"] == "307.600000"
    assert line["unit_price"] == "270.98"
    assert line["line_total"] == "541.96"
    assert line["quantity"] == "2"
    assert line["configuration"] == "Powder coated aluminium; Cushion A+; Ceramic A"
    assert all(not isinstance(value, float) for value in (cart["exchange_rate"], line["unit_price_base"], line["unit_price"], line["line_total"]))


def test_lumbro_cart_requires_integer_piece_quantity_and_preserves_net_tax_totals():
    catalog = load_supplier_catalog_data(lumbro_catalog_payload())
    cart = build_supplier_cart_payload(
        [{"internal_id": "lumbro:variant:barcelona", "quantity": "2", "add_on_option_ids": []}],
        catalog,
        "MXN",
        [],
    )

    line = cart["items"][0]
    net = Decimal(line["line_total"])
    tax = (net * Decimal(line["tax_rate"])).quantize(Decimal("0.01"))
    total = net + tax
    assert cart["supplier"] == "lumbro"
    assert supplier_module.SUPPLIER_LABELS["lumbro"] == "Lumbro"
    assert cart["base_currency"] == "MXN"
    assert line["quantity"] == "2"
    assert line["base_price"] == "2824.000000"
    assert line["unit_price"] == "2824.00"
    assert line["line_total"] == "5648.00"
    assert line["tax_rate"] == "0.160000"
    assert tax == Decimal("903.68")
    assert total == Decimal("6551.68")
    with pytest.raises(ValueError, match="entera"):
        build_supplier_cart_payload(
            [{"internal_id": "lumbro:variant:barcelona", "quantity": "2.5", "add_on_option_ids": []}],
            catalog,
            "MXN",
            [],
        )


def test_representative_lumbro_snapshot_quotes_official_codes_and_warns_on_review_item(
    representative_lumbro_catalog,
    tmp_path,
):
    loaded = load_supplier_catalog_data(
        representative_lumbro_catalog,
        expected_supplier="lumbro",
    )
    verified = [row for row in loaded["items"] if row["code_status"] == "verified"]
    review = next(row for row in loaded["items"] if row["code_status"] == "needs_review")
    assert len(verified) == 4
    assert loaded["metadata"]["coverage"]["verified_items"] == 4
    assert tuple(row["sku"] for row in verified) == LUMBRO_OFFICIAL_SKUS
    assert all(not row["sku"].startswith("lumbro:") for row in verified)
    assert review["sku"] == ""

    cart = build_supplier_cart_payload(
        [
            {"internal_id": row["internal_id"], "quantity": "1", "add_on_option_ids": []}
            for row in verified
        ],
        loaded,
        "MXN",
        [],
    )
    assert len(cart["items"]) == 4
    assert tuple(row["sku"] for row in cart["items"]) == LUMBRO_OFFICIAL_SKUS
    assert tuple(row["line_total"] for row in cart["items"]) == (
        "3003.00",
        "1394.07",
        "350.00",
        "772.00",
    )
    assert {row["tax_rate"] for row in cart["items"]} == {"0.160000"}
    actual_net = sum((Decimal(row["line_total"]) for row in cart["items"]), Decimal(0))
    actual_iva = (actual_net * Decimal("0.160000")).quantize(Decimal("0.01"))
    assert actual_net == LUMBRO_EXPECTED_NET
    assert actual_iva == LUMBRO_EXPECTED_IVA
    assert actual_net + actual_iva == LUMBRO_EXPECTED_TOTAL

    review_cart = build_supplier_cart_payload(
        [{"internal_id": review["internal_id"], "quantity": "1", "add_on_option_ids": []}],
        loaded,
        "MXN",
        [],
    )
    review_line = review_cart["items"][0]
    assert review_line["code_status"] == "needs_review"
    assert review_line["sku"] == ""
    assert review_line["warnings"]
    description, warning = catalog_cart._description_for_item(
        review_line,
        review_line["sku"],
        review_line["product_url"],
        Decimal(review_line["quantity"]),
    )
    assert "Codigo por verificar" in description
    assert "Codigo por verificar" in warning

    with pytest.raises(ValueError, match="entera"):
        build_supplier_cart_payload(
            [{"internal_id": verified[0]["internal_id"], "quantity": "2.5", "add_on_option_ids": []}],
            loaded,
            "MXN",
            [],
        )

    output = create_supplier_quotation_workbook(cart, tmp_path / "lumbro-representative.xlsx")
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Quotation"]
        assert sheet["A8"].value == "- Lumbro"
        assert tuple(sheet.cell(row, 10).value for row in range(9, 13)) == (
            3003,
            1394.07,
            350,
            772,
        )
    finally:
        workbook.close()


def test_repeated_kun_source_code_survives_cart_and_embeds_official_image_in_xlsx(
    monkeypatch, tmp_path
):
    first = deepcopy(catalog_payload()["items"][0])
    second = deepcopy(first)
    first["attributes"]["source_code"] = "KUN-REPETIDO"
    second.update(
        internal_id="alma:kun:variant:second",
        product_key="kun-second",
        sku="kun-internal-second",
        name="Silla KUN segunda",
        description="Descripcion oficial segunda",
        product_url="https://www.kundesign.com/s/1/product.html",
        image_url="https://project-ref.supabase.co/storage/v1/object/public/catalog-assets/"
        + "b" * 64
        + ".png",
    )
    second["attributes"]["source_code"] = "KUN-REPETIDO"
    catalog = load_supplier_catalog_data(catalog_payload(items=[first, second]))
    cart = build_supplier_cart_payload(
        [
            {
                "internal_id": second["internal_id"],
                "quantity": "2",
                "base_option_id": "powder-coated-aluminium",
                "add_on_option_ids": ["cushion:a-plus"],
            }
        ],
        catalog,
        "USD",
        [],
    )
    line = cart["items"][0]
    assert line["sku"] == "KUN-REPETIDO"
    assert line["description"] == second["description"]
    assert line["product_url"] == second["product_url"]
    assert line["image_url"] == second["image_url"]
    assert line["unit_price"] == "285.10"

    encoded = io.BytesIO()
    Image.new("RGB", (20, 20), "white").save(encoded, format="PNG")
    image_bytes = encoded.getvalue()

    class PeerSocket:
        def getpeername(self):
            return ("93.184.216.34", 443)

    class Response:
        headers = {
            "content-type": "image/png",
            "content-length": str(len(image_bytes)),
        }

        def __init__(self):
            self.fp = type("Fp", (), {"raw": type("Raw", (), {"_sock": PeerSocket()})()})()

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def read(self, size):
            return image_bytes[:size]

    class Opener:
        def open(self, request, timeout):
            assert request.full_url == second["image_url"]
            assert timeout == 18
            return Response()

    monkeypatch.setenv("SUPABASE_URL", "https://project-ref.supabase.co")
    monkeypatch.delenv("CATALOG_ASSET_PUBLIC_BASE_URL", raising=False)
    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    monkeypatch.setattr(catalog_cart.urllib.request, "build_opener", lambda *handlers: Opener())
    output = create_supplier_quotation_workbook(cart, tmp_path / "kun-cart.xlsx")
    workbook = load_workbook(output)
    sheet = workbook["Quotation"]
    assert sheet["B9"].value == second["name"]
    assert "Descripcion oficial segunda" in sheet["D9"].value
    assert "KUN-REPETIDO" in sheet["D9"].value
    assert "Powder coated aluminium; Cushion A+" in sheet["D9"].value
    assert sheet["E9"].value == "65 x 45 x 83 cm"
    assert sheet["G9"].value == 2
    assert isinstance(sheet["G9"].value, int)
    assert sheet["J9"].value == 285.1
    assert sheet["K9"].value == second["product_url"]
    assert len(sheet._images) == 1
    workbook.close()


def test_supplier_xlsx_description_keeps_exact_link_color_and_availability_buckets(tmp_path):
    payload = catalog_payload(supplier="sunon")
    item = payload["items"][0]
    item.update(
        internal_id="sunon:erp:9001",
        product_key="fast10",
        sku="9001",
        brand="Sunon",
        collection="",
        name="Fast Chair",
        description="Silla operativa",
        unit="PZA",
        availability_type="stocked",
        stock="19",
        lead_time="1-2 semanas",
        base_price_options=[],
        add_on_options=[],
        price_net="200.000000",
        product_url="https://www.sunonglobal.com/product/fast-chair/",
        image_url="",
        image_kind="placeholder",
    )
    item["attributes"] = {
        "source_erp_code": "9001",
        "source_model_code": "FAST10",
        "dimensions": "700 x 700 mm",
        "color": "Black frame",
        "availability_buckets": [
            {"lead_time": "1-2 semanas", "quantity": "12", "source_refs": []},
            {"lead_time": "1-2 semanas", "quantity": "3", "source_refs": []},
            {"lead_time": "4-6 semanas", "quantity": "7", "source_refs": []},
        ],
        "product_url_match": {
            "status": "exact_code",
            "matched_code": "FAST10",
            "lookup_code": "FAST10",
        },
    }
    catalog = load_supplier_catalog_data(payload, expected_supplier="sunon")
    cart = build_supplier_cart_payload(
        [{"internal_id": item["internal_id"], "quantity": "2"}],
        catalog,
        "USD",
        [],
    )

    output = create_supplier_quotation_workbook(cart, tmp_path / "sunon-cart.xlsx")
    workbook = load_workbook(output)
    sheet = workbook["Quotation"]
    description = sheet["D9"].value

    assert "Silla operativa" in description
    assert "SKU: 9001" in description
    assert "Color: Black frame" in description
    assert "Entrega: 1-2 semanas" in description
    assert "Disponibilidad: 15 PZA (1-2 semanas); 7 PZA (4-6 semanas)" in description
    assert "URL: https://www.sunonglobal.com/product/fast-chair/" in description
    assert sheet["E9"].value == "700 x 700 mm"
    assert sheet["G9"].value == 2
    assert isinstance(sheet["G9"].value, int)
    assert sheet["K9"].value == "https://www.sunonglobal.com/product/fast-chair/"
    workbook.close()


def test_supplier_xlsx_discloses_unknown_availability():
    description, warning = catalog_cart._description_for_item(
        {
            "description": "Producto sin existencia publicada",
            "availability_type": "unknown",
            "warnings": [],
        },
        "",
        "",
        Decimal("1"),
    )

    assert "Disponibilidad: por confirmar" in description
    assert warning == ""


def test_supplier_xlsx_preserves_warranty_and_product_notes():
    description, warning = catalog_cart._description_for_item(
        {
            "description": "Producto CR Global",
            "availability_type": "unknown",
            "attributes": {
                "warranty": "GARANTIA DE 5 ANOS",
                "product_notes": ["NO INCLUYE CUBIERTA"],
            },
            "warnings": [],
        },
        "CR-1",
        "",
        Decimal("1"),
    )

    assert "Garantia: GARANTIA DE 5 ANOS" in description
    assert "Notas: NO INCLUYE CUBIERTA" in description
    assert warning == ""


def test_supplier_cart_rejects_verified_item_with_zero_price():
    payload = catalog_payload()
    payload["items"][0].update(
        code_status="verified",
        sku="ALMA-ZERO",
        base_currency="USD",
        price_net="0.000000",
        base_price_options=[],
        add_on_options=[],
    )

    with pytest.raises(ValueError, match="precio por confirmar"):
        build_supplier_cart_payload(
            [{"internal_id": payload["items"][0]["internal_id"], "quantity": "1"}],
            payload,
            "MXN",
            [],
        )


@pytest.mark.parametrize(
    "host",
    ("kundesign.com", "www.kundesign.com", "assets.kundesign.com"),
)
def test_supplier_cart_never_allows_kundesign_as_image_storage(monkeypatch, tmp_path, host):
    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", f"https://{host}/assets")
    monkeypatch.delenv("SUPABASE_URL", raising=False)
    monkeypatch.setattr(
        catalog_cart.urllib.request,
        "build_opener",
        lambda *handlers: pytest.fail("Kundesign no debe alcanzar el transporte"),
    )

    assert catalog_cart._allowed_image_hosts("supplier_cart") == frozenset()
    assert catalog_cart._download_catalog_image(
        f"https://{host}/image.png", tmp_path, "KUN", "supplier_cart"
    ) is None


def test_build_cart_rejects_client_catalog_fields():
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": "1",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
        "unit_price": "0.01",
    }
    with pytest.raises(ValueError, match="no permitido"):
        build_supplier_cart_payload([raw], load_supplier_catalog_data(catalog_payload()), "USD", [])


def test_build_cart_rejects_review_codes():
    review_payload = catalog_payload()
    review_item = review_payload["items"][0]
    review_item.update({"code_status": "needs_review", "sku": ""})
    raw = {
        "internal_id": review_item["internal_id"],
        "quantity": "1",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
    }
    with pytest.raises(ValueError, match="verificar"):
        build_supplier_cart_payload([raw], review_payload, "USD", [])

    cr_global_payload = catalog_payload(supplier="cr-global")
    cr_global_item = cr_global_payload["items"][0]
    cr_global_item.update({"code_status": "needs_review", "sku": "", "base_currency": "MXN"})
    with pytest.raises(ValueError, match="codigo por verificar"):
        build_supplier_cart_payload(
            [{"internal_id": cr_global_item["internal_id"], "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}],
            cr_global_payload,
            "MXN",
            [],
        )


def test_sonara_cart_accepts_review_item_with_positive_mxn_price_and_warning():
    payload = catalog_payload(supplier="sonara")
    item = payload["items"][0]
    item.update(
        internal_id="sonara:review-panel",
        supplier="sonara",
        sku="",
        code_status="needs_review",
        base_currency="MXN",
        price_net="77.000000",
        tax_rate="0.160000",
        base_price_options=[],
        add_on_options=[],
        warnings=["Codigo por verificar"],
    )
    cart = build_supplier_cart_payload(
        [{"internal_id": item["internal_id"], "quantity": "2", "add_on_option_ids": []}],
        payload,
        "MXN",
        [],
    )
    line = cart["items"][0]
    assert line["code_status"] == "needs_review"
    assert line["sku"] == ""
    assert line["unit_price"] == "77.00"
    assert line["tax_rate"] == "0.160000"
    assert "Codigo por verificar" in line["warnings"]


@pytest.mark.parametrize(
    ("field", "value", "message"),
    (
        ("price_net", "0.000000", "precio por confirmar"),
        ("base_currency", "XXX", "moneda base por verificar"),
        ("base_currency", "USD", "moneda base por verificar"),
        ("tax_rate", "0.080000", "IVA 16"),
    ),
)
def test_sonara_review_item_fails_closed_without_valid_commercial_data(field, value, message):
    payload = catalog_payload(supplier="sonara")
    item = payload["items"][0]
    item.update(
        internal_id="sonara:review-panel",
        supplier="sonara",
        sku="",
        code_status="needs_review",
        base_currency="MXN",
        price_net="77.000000",
        tax_rate="0.160000",
        base_price_options=[],
        add_on_options=[],
    )
    item[field] = value
    with pytest.raises(ValueError, match=message):
        build_supplier_cart_payload(
            [{"internal_id": item["internal_id"], "quantity": "1", "add_on_option_ids": []}],
            payload,
            "MXN",
            [],
        )


def test_sonara_verified_item_with_usd_base_currency_fails_closed():
    payload = catalog_payload(supplier="sonara")
    item = payload["items"][0]
    item.update(
        internal_id="sonara:verified-panel",
        supplier="sonara",
        base_currency="USD",
        price_net="77.000000",
        base_price_options=[],
        add_on_options=[],
    )
    with pytest.raises(ValueError, match="moneda base por verificar"):
        build_supplier_cart_payload(
            [{"internal_id": item["internal_id"], "quantity": "1", "add_on_option_ids": []}],
            payload,
            "MXN",
            [],
        )


def test_review_warning_is_canonicalized_and_verified_warnings_are_preserved():
    payload = catalog_payload(supplier="sonara")
    item = payload["items"][0]
    item.update(
        internal_id="sonara:review-panel",
        supplier="sonara",
        sku="",
        code_status="needs_review",
        base_currency="MXN",
        price_net="77.000000",
        tax_rate="0.160000",
        base_price_options=[],
        add_on_options=[],
        warnings=["C\u00f3digo por verificar", " codigo por VERIFICAR ", "Evidencia detallada"],
    )
    review_line = build_supplier_cart_payload(
        [{"internal_id": item["internal_id"], "quantity": "1", "add_on_option_ids": []}], payload, "MXN", []
    )["items"][0]
    assert review_line["warnings"] == ["Evidencia detallada", "Codigo por verificar"]

    payload = catalog_payload(supplier="sonara")
    payload["items"][0]["base_currency"] = "MXN"
    payload["items"][0]["warnings"] = ["Imagen ilustrativa", "Otra advertencia"]
    verified_line = build_supplier_cart_payload(
        [{"internal_id": payload["items"][0]["internal_id"], "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}], payload, "MXN", []
    )["items"][0]
    assert verified_line["warnings"] == ["Imagen ilustrativa", "Otra advertencia"]


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("internal_id", 1), ("internal_id", True), ("internal_id", 1.5),
        ("internal_id", []), ("internal_id", {}),
        ("base_option_id", 1), ("base_option_id", True), ("base_option_id", 1.5),
        ("base_option_id", []), ("base_option_id", {}),
    ],
)
def test_build_cart_rejects_non_string_browser_ids(field, value):
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": "1",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
    }
    raw[field] = value

    with pytest.raises(ValueError, match=field):
        build_supplier_cart_payload([raw], load_supplier_catalog_data(catalog_payload()), "USD", [])


@pytest.mark.parametrize("value", [1, True, 1.5, [], {}])
def test_build_cart_rejects_non_string_add_on_ids(value):
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": "1",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [value],
    }
    with pytest.raises(ValueError, match="add_on_option_id"):
        build_supplier_cart_payload([raw], load_supplier_catalog_data(catalog_payload()), "USD", [])


@pytest.mark.parametrize("quantity", [1, True, 1.5])
def test_build_cart_rejects_non_decimal_string_browser_quantities(quantity):
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": quantity,
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
    }
    with pytest.raises(ValueError, match="cantidad"):
        build_supplier_cart_payload([raw], load_supplier_catalog_data(catalog_payload()), "USD", [])


@pytest.mark.parametrize("quantity", ["0.5", "1.25", Decimal("2.5")])
def test_build_cart_rejects_fractional_piece_quantities(quantity):
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": quantity,
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
    }
    with pytest.raises(ValueError, match="entera"):
        build_supplier_cart_payload(
            [raw], load_supplier_catalog_data(catalog_payload()), "USD", []
        )


@pytest.mark.parametrize("unit", ["M2", "m²"])
def test_build_cart_accepts_fractional_square_meter_quantities(unit):
    payload = catalog_payload()
    payload["items"][0]["unit"] = unit
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": "1.25",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": [],
    }
    cart = build_supplier_cart_payload([raw], payload, "USD", [])
    assert cart["items"][0]["quantity"] == "1.25"


def test_decimal_results_do_not_depend_on_ambient_context_precision():
    raw = {
        "internal_id": "alma:kun:kc8611n01rop",
        "quantity": "2",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": ["cushion:a-plus", "ceramic:a"],
    }
    expected = build_supplier_cart_payload([raw], catalog_payload(), "EUR", rate_rows())

    with localcontext() as context:
        context.prec = 2
        actual = build_supplier_cart_payload([raw], catalog_payload(), "EUR", rate_rows())

    assert actual["exchange_rate"] == expected["exchange_rate"] == "0.880952"
    assert actual["items"][0]["unit_price_base"] == expected["items"][0]["unit_price_base"] == "307.600000"
    assert actual["items"][0]["unit_price"] == expected["items"][0]["unit_price"] == "270.98"
    assert actual["items"][0]["line_total"] == expected["items"][0]["line_total"] == "541.96"


def test_derived_line_boundary_is_accepted_under_low_ambient_precision():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = []
    payload["items"][0]["add_on_options"] = []
    payload["items"][0]["price_net"] = "1000000000.000000"
    raw = {"internal_id": "alma:kun:kc8611n01rop", "quantity": "1000000", "add_on_option_ids": []}

    with localcontext() as context:
        context.prec = 2
        cart = build_supplier_cart_payload([raw], payload, "USD", [])

    assert cart["items"][0]["line_total"] == "1000000000000000.00"


def test_extreme_derived_ratio_and_line_total_are_rejected_cleanly():
    today = date(2026, 7, 15)
    extreme_rates = [
        {"currency": "USD", "effective_date": "2026-07-15", "mxn_per_unit": "1000000000.000000", "retrieved_at": "2026-07-15T20:00:00Z"},
        {"currency": "EUR", "effective_date": "2026-07-15", "mxn_per_unit": "0.000001", "retrieved_at": "2026-07-15T20:00:00Z"},
    ]
    with pytest.raises(ValueError, match="exchange_rate fuera de rango"):
        resolve_conversion_rate("USD", "EUR", extreme_rates, today)

    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = [
        {"id": "expensive", "name": "Expensive", "price_net": "1000000000.000000", "available": True}
    ]
    payload["items"][0]["add_on_options"] = [
        {"id": "extra", "name": "Extra", "family": "extra", "price_net": "1000000000.000000", "available": True}
    ]
    raw = {"internal_id": "alma:kun:kc8611n01rop", "quantity": "1000000", "base_option_id": "expensive", "add_on_option_ids": ["extra"]}
    with pytest.raises(ValueError, match="line_total fuera de rango"):
        build_supplier_cart_payload([raw], payload, "USD", [])


def test_build_cart_uses_price_net_when_there_are_no_options():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = []
    payload["items"][0]["add_on_options"] = []
    payload["items"][0]["unit"] = "M2"
    catalog = load_supplier_catalog_data(payload)

    cart = build_supplier_cart_payload(
        [{"internal_id": "alma:kun:kc8611n01rop", "quantity": "1.25", "add_on_option_ids": []}],
        catalog=catalog,
        quote_currency="USD",
        rate_rows=[],
    )

    assert cart["exchange_rate"] == "1.000000"
    assert cart["items"][0]["unit_price_base"] == "199.990000"
    assert cart["items"][0]["unit_price"] == "199.99"
    assert cart["items"][0]["line_total"] == "249.99"


def test_build_cart_line_total_uses_the_rounded_display_unit_price():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = []
    payload["items"][0]["add_on_options"] = []
    payload["items"][0]["price_net"] = "0.335000"

    cart = build_supplier_cart_payload(
        [{"internal_id": "alma:kun:kc8611n01rop", "quantity": "3", "add_on_option_ids": []}],
        payload,
        "USD",
        [],
    )

    assert cart["items"][0]["unit_price"] == "0.34"
    assert cart["items"][0]["line_total"] == "1.02"


def test_build_cart_rejects_zero_product_price():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = []
    payload["items"][0]["add_on_options"] = []
    payload["items"][0]["price_net"] = "0.000000"

    with pytest.raises(ValueError, match="precio por confirmar"):
        build_supplier_cart_payload(
            [{"internal_id": "alma:kun:kc8611n01rop", "quantity": "2", "add_on_option_ids": []}],
            payload,
            "USD",
            [],
        )


def test_build_cart_rejects_unknown_base_currency_until_verified():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"] = []
    payload["items"][0]["add_on_options"] = []
    payload["items"][0]["base_currency"] = "XXX"
    payload["items"][0]["price_net"] = "0.000000"

    with pytest.raises(ValueError, match="moneda base.*verificar"):
        build_supplier_cart_payload(
            [{"internal_id": "alma:kun:kc8611n01rop", "quantity": "2", "add_on_option_ids": []}],
            payload,
            "USD",
            [],
        )


def test_build_cart_rejects_zero_base_and_add_on_option_prices():
    payload = catalog_payload()
    payload["items"][0]["base_price_options"][0]["price_net"] = "0.000000"
    payload["items"][0]["add_on_options"][0]["price_net"] = "0.000000"

    with pytest.raises(ValueError, match="precio por confirmar"):
        build_supplier_cart_payload(
            [{
                "internal_id": "alma:kun:kc8611n01rop",
                "quantity": "2",
                "base_option_id": "powder-coated-aluminium",
                "add_on_option_ids": ["cushion:a-plus"],
            }],
            payload,
            "USD",
            [],
        )


@pytest.mark.parametrize(
    ("raw", "message"),
    [
        ({"internal_id": "missing", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}, "no encontrado"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "0", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}, "cantidad"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1000000.000001", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}, "cantidad"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1.0000001", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}, "cantidad"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": None, "add_on_option_ids": []}, "base"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "unknown", "add_on_option_ids": []}, "base"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": ["ceramic:b"]}, "disponible"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "solid-wood", "add_on_option_ids": ["ceramic:a"]}, "incompatible"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": ["cushion:a-plus", "cushion:a-plus-plus"]}, "familia"),
        ({"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": ["ceramic:a", "ceramic:a"]}, "duplicado"),
    ],
)
def test_build_cart_rejects_invalid_client_configurations(raw, message):
    with pytest.raises(ValueError, match=message):
        build_supplier_cart_payload([raw], load_supplier_catalog_data(catalog_payload()), "USD", [])


def test_build_cart_rejects_duplicate_configuration_rows():
    raw = {"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": ["ceramic:a"]}
    with pytest.raises(ValueError, match="configuracion duplicada"):
        build_supplier_cart_payload([raw, deepcopy(raw)], load_supplier_catalog_data(catalog_payload()), "USD", [])


def test_resolve_conversion_rate_uses_common_date_and_decimal_cross_rate():
    rows = rate_rows(effective_date="2026-07-14") + [
        {"currency": "USD", "effective_date": "2026-07-15", "mxn_per_unit": "19.000000", "retrieved_at": "2026-07-15T20:00:00Z"}
    ]
    snapshot = resolve_conversion_rate("USD", "EUR", rows, date(2026, 7, 15))

    assert snapshot.exchange_rate == Decimal("0.880952")
    assert snapshot.rate_effective_date == date(2026, 7, 14)
    assert snapshot.rate_retrieved_at == "2026-07-14T23:05:00Z"


def test_supplier_cart_uses_explicit_effective_date_for_rate_selection():
    catalog = load_supplier_catalog_data(catalog_payload())
    rows = [
        {"currency": "USD", "effective_date": "2026-07-17", "mxn_per_unit": "17.000000", "retrieved_at": "2026-07-17T12:00:00Z"},
        {"currency": "USD", "effective_date": "2026-07-18", "mxn_per_unit": "18.000000", "retrieved_at": "2026-07-18T12:00:00Z"},
        {"currency": "USD", "effective_date": "2026-07-19", "mxn_per_unit": "19.000000", "retrieved_at": "2026-07-19T12:00:00Z"},
    ]

    payload = build_supplier_cart_payload(
        [{"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}],
        catalog, "MXN", rows, today=date(2026, 7, 18),
    )

    assert payload["rate_effective_date"] == "2026-07-18"
    assert payload["exchange_rate"] == "18.000000"


def test_resolve_conversion_rate_supports_mxn_and_rejects_stale_or_invalid_rows():
    today = date(2026, 7, 15)
    assert resolve_conversion_rate("USD", "MXN", rate_rows(effective_date="2026-07-15"), today).exchange_rate == Decimal("18.500000")
    assert resolve_conversion_rate("MXN", "USD", rate_rows(effective_date="2026-07-15"), today).exchange_rate == Decimal("0.054054")
    assert resolve_conversion_rate("USD", "USD", [], today).exchange_rate == Decimal("1.000000")

    with pytest.raises(ValueError, match="vencida"):
        resolve_conversion_rate("USD", "MXN", rate_rows(effective_date="2026-07-09"), today)
    bad_rows = rate_rows(effective_date="2026-07-15")
    bad_rows[0]["mxn_per_unit"] = "Infinity"
    with pytest.raises(ValueError, match="tasa"):
        resolve_conversion_rate("USD", "MXN", bad_rows, today)


def test_identity_rate_validates_supplied_rows_without_requiring_any():
    today = date(2026, 7, 15)
    assert resolve_conversion_rate("USD", "USD", [], today).exchange_rate == Decimal("1.000000")
    assert resolve_conversion_rate("USD", "USD", rate_rows(effective_date="2026-07-15"), today).exchange_rate == Decimal("1.000000")

    malformed = rate_rows(effective_date="2026-07-15")
    malformed[0]["mxn_per_unit"] = 18.5
    with pytest.raises(ValueError, match="tasa"):
        resolve_conversion_rate("USD", "USD", malformed, today)

    duplicates = rate_rows(effective_date="2026-07-15")
    duplicates.append({**duplicates[0], "mxn_per_unit": "19.000000"})
    with pytest.raises(ValueError, match="duplicada"):
        resolve_conversion_rate("USD", "USD", duplicates, today)

    future = rate_rows(effective_date="2026-07-16")
    assert resolve_conversion_rate("USD", "USD", future, today).exchange_rate == Decimal("1.000000")


def test_resolve_conversion_rate_rejects_datetime_for_today():
    with pytest.raises(ValueError, match="Fecha de conversion invalida"):
        resolve_conversion_rate("USD", "USD", [], datetime(2026, 7, 15, 12, 0))


@pytest.mark.parametrize(
    "url",
    [
        "javascript:alert(1)", "data:text/plain,bad", "file:///tmp/bad",
        "//cdn.example.test/image.webp", "https://user:pass@example.test/image.webp",
        "https://example.test/bad\nheader", "x" * 2049,
        "https://exa mple.com/x", "https://./x", "https://example..com/x",
        "https://-example.com/x", "https://example-.com/x",
        "https://exa%mple.com/x", "https://example.com\\evil/x",
        f"https://{'a' * 64}.com/x",
        f"https://{'.'.join(['a' * 63] * 4)}/x",
        "https://256.1.1.1/x", "https://999.999.999.999/x",
        "https://127.1/x", "https://xn--a.example/x",
    ],
)
def test_load_catalog_rejects_unsafe_or_unbounded_urls(url):
    payload = catalog_payload()
    payload["items"][0]["image_url"] = url
    with pytest.raises(ValueError, match="image_url"):
        load_supplier_catalog_data(payload)


def test_load_catalog_accepts_empty_and_absolute_http_storage_urls():
    payload = catalog_payload()
    payload["items"][0]["image_url"] = "http://assets.example.test/catalog/item.jpg"
    payload["items"][0]["product_url"] = "https://storage.example.test/catalog/item"
    loaded = load_supplier_catalog_data(payload)
    assert loaded["items"][0]["image_url"].startswith("http://")

    payload["items"][0]["image_url"] = ""
    payload["items"][0]["product_url"] = ""
    load_supplier_catalog_data(payload)


@pytest.mark.parametrize(
    "url",
    [
        "https://catalog.example.com/x",
        "https://192.0.2.10/x",
        "https://127.0.0.1/x",
        "https://[2001:db8::10]/x",
        "https://münich.example/x",
        "https://xn--mnich-kva.example/x",
    ],
)
def test_load_catalog_accepts_valid_dns_ip_and_idna_urls(url):
    payload = catalog_payload()
    payload["items"][0]["image_url"] = url
    assert load_supplier_catalog_data(payload)["items"][0]["image_url"] == url


def test_catalog_item_and_cart_row_limits_are_enforced_before_iteration():
    payload = catalog_payload()
    payload["items"] = [payload["items"][0]] * (supplier_module.MAX_CATALOG_ITEMS + 1)
    with pytest.raises(ValueError, match="limite de items"):
        load_supplier_catalog_data(payload)

    raw = {"internal_id": "alma:kun:kc8611n01rop", "quantity": "1", "base_option_id": "powder-coated-aluminium", "add_on_option_ids": []}
    with pytest.raises(ValueError, match="limite de filas"):
        build_supplier_cart_payload([raw] * (supplier_module.MAX_CART_ROWS + 1), catalog_payload(), "USD", [])


def test_options_warnings_and_compatibility_limits_are_enforced():
    payload = catalog_payload()
    option = payload["items"][0]["base_price_options"][0]
    payload["items"][0]["base_price_options"] = [{**option, "id": f"base-{index}"} for index in range(supplier_module.MAX_OPTIONS_PER_ITEM + 1)]
    with pytest.raises(ValueError, match="limite de options"):
        load_supplier_catalog_data(payload)

    payload = catalog_payload()
    payload["items"][0]["warnings"] = ["warning"] * (supplier_module.MAX_WARNINGS_PER_ITEM + 1)
    with pytest.raises(ValueError, match="limite de warnings"):
        load_supplier_catalog_data(payload)

    payload = catalog_payload()
    payload["items"][0]["add_on_options"][0]["compatible_base_option_ids"] = [
        "powder-coated-aluminium"
    ] * (supplier_module.MAX_COMPATIBLE_OPTION_IDS + 1)
    with pytest.raises(ValueError, match="limite de compatible"):
        load_supplier_catalog_data(payload)


@pytest.mark.parametrize(
    ("option_kind", "mutation", "message"),
    [
        ("base_price_options", lambda option: option.update(unexpected="value"), "campos inesperados"),
        ("base_price_options", lambda option: option.update(compatible_base_option_ids=[]), "campos inesperados"),
        ("add_on_options", lambda option: option.update(unexpected="value"), "campos inesperados"),
        ("base_price_options", lambda option: option.pop("name"), "campos faltantes"),
        ("add_on_options", lambda option: option.pop("family"), "campos faltantes"),
    ],
)
def test_options_require_strict_fields_by_option_kind(option_kind, mutation, message):
    payload = catalog_payload()
    mutation(payload["items"][0][option_kind][0])

    with pytest.raises(ValueError, match=message):
        load_supplier_catalog_data(payload)


def test_add_on_allows_only_documented_optional_compatibility_field():
    payload = catalog_payload()
    loaded = load_supplier_catalog_data(payload)

    add_ons = loaded["items"][0]["add_on_options"]
    assert add_ons[0].keys() == {"id", "name", "family", "price_net", "available"}
    assert add_ons[2]["compatible_base_option_ids"] == ["powder-coated-aluminium"]


@pytest.mark.parametrize("field", ["name", "description", "source_reference"])
def test_catalog_text_limits_are_enforced_without_truncation(field):
    payload = catalog_payload()
    limit = {
        "name": supplier_module.MAX_TEXT_LENGTH,
        "description": supplier_module.MAX_DESCRIPTION_LENGTH,
        "source_reference": supplier_module.MAX_SOURCE_REFERENCE_LENGTH,
    }[field]
    payload["items"][0][field] = "x" * (limit + 1)
    with pytest.raises(ValueError, match=field):
        load_supplier_catalog_data(payload)


@pytest.mark.parametrize("attributes", [{"bad": object()}, {"bad": {1, 2}}, {"bad": Decimal("1")}])
def test_attributes_reject_non_json_values(attributes):
    payload = catalog_payload()
    payload["items"][0]["attributes"] = attributes
    with pytest.raises(ValueError, match="attributes"):
        load_supplier_catalog_data(payload)


def test_attributes_reject_excessive_size_and_depth():
    payload = catalog_payload()
    payload["items"][0]["attributes"] = {"value": "x" * supplier_module.MAX_ATTRIBUTES_JSON_BYTES}
    with pytest.raises(ValueError, match="attributes"):
        load_supplier_catalog_data(payload)

    nested = "leaf"
    for _ in range(supplier_module.MAX_ATTRIBUTES_DEPTH + 1):
        nested = {"nested": nested}
    payload = catalog_payload()
    payload["items"][0]["attributes"] = nested
    with pytest.raises(ValueError, match="attributes"):
        load_supplier_catalog_data(payload)


def test_rate_row_limit_applies_to_identity_and_conversion():
    rows = rate_rows() * (supplier_module.MAX_RATE_ROWS // 2 + 1)
    with pytest.raises(ValueError, match="limite de tasas"):
        resolve_conversion_rate("USD", "USD", rows, date.today())
    with pytest.raises(ValueError, match="limite de tasas"):
        resolve_conversion_rate("USD", "EUR", rows, date.today())


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("ordinary text", "ordinary text"),
        ("  =1+1", "'  =1+1"),
        ("\t+SUM(A1:A2)", "'\t+SUM(A1:A2)"),
        ("-2", "'-2"),
        ("@command", "'@command"),
        (None, ""),
        ("line\r\ntext", "line\ntext"),
    ],
)
def test_safe_excel_text_neutralizes_formula_prefixes(value, expected):
    assert safe_excel_text(value) == expected


def test_workbook_boundary_rejects_invalid_supplier_payload(tmp_path):
    with pytest.raises(ValueError, match="Proveedor no soportado"):
        create_supplier_quotation_workbook({}, tmp_path / "supplier.xlsx")


def test_supplier_images_only_allow_configured_public_asset_hosts(monkeypatch):
    monkeypatch.delenv("CATALOG_ASSET_PUBLIC_BASE_URL", raising=False)
    monkeypatch.delenv("SUPABASE_URL", raising=False)
    assert catalog_cart._allowed_image_hosts("supplier_cart") == frozenset()

    monkeypatch.setenv("SUPABASE_URL", "https://project-ref.supabase.co")
    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", "https://catalog-assets.example.test/public")
    assert catalog_cart._allowed_image_hosts("supplier_cart") == frozenset(
        {"project-ref.supabase.co", "catalog-assets.example.test"}
    )
    assert "www.offiho.com" not in catalog_cart._allowed_image_hosts("supplier_cart")


def test_supplier_image_allowlist_ignores_insecure_or_credentialed_bases(monkeypatch):
    monkeypatch.setenv("SUPABASE_URL", "https://project-ref.supabase.co:8443")
    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", "https://user:pass@assets.example.test")
    assert catalog_cart._allowed_image_hosts("supplier_cart") == frozenset()


def test_supplier_module_copies_are_byte_identical():
    root = Path(__file__).resolve().parents[1]
    for module_name in ("supplier_catalog.py", "catalog_cart.py"):
        source = (root / "mobiliti_saas/quote_engine" / module_name).read_bytes()
        mirror = (
            root / "mobiliti_saas/web/mobiliti_saas/quote_engine" / module_name
        ).read_bytes()
        assert source == mirror
        assert hashlib.sha256(source).hexdigest() == hashlib.sha256(mirror).hexdigest()
