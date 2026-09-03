from copy import deepcopy
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mobiliti_saas.quote_engine import generate_quote
from mobiliti_saas.quote_engine import catalog_cart
from mobiliti_saas.quote_engine.catalog_search import search_catalog_products
from mobiliti_saas.quote_engine.engine import MIXED_CATALOG_ORDER as ENGINE_MIXED_CATALOG_ORDER
from mobiliti_saas.quote_engine.mixed_catalog import (
    MIXED_CATALOG_ORDER,
    build_mixed_catalog_cart_payload,
)
from mobiliti_saas.quote_engine.project_model import normalize_project_payload
from mobiliti_saas.quote_engine.quotation_sheets import quotation_data_rows
from mobiliti_saas.quote_engine.supplier_catalog import (
    EXPECTED_SUPPLIER_BASE_CURRENCY,
    SUPPLIER_LABELS,
    build_supplier_cart_payload,
    load_supplier_catalog_data,
)
from mobiliti_saas.worker.quote_worker import _convert_mixed_catalog_cart_to_quotation


TODAY = date(2026, 8, 3)
RATE_ROWS = [{
    "currency": "USD",
    "effective_date": TODAY.isoformat(),
    "mxn_per_unit": "20.000000",
    "retrieved_at": "2026-08-03T12:00:00+00:00",
}]


def _item(
    supplier: str,
    *,
    internal_id: str,
    price: str | None,
    warnings: list[str] | None = None,
    code_status: str = "verified",
    base_price_options: list[dict] | None = None,
    add_on_options: list[dict] | None = None,
) -> dict:
    return {
        "internal_id": internal_id,
        "supplier": supplier,
        "product_key": internal_id,
        "sku": f"{supplier.upper()}-1" if code_status == "verified" else "",
        "code_status": code_status,
        "brand": SUPPLIER_LABELS[supplier],
        "collection": "Integracion",
        "name": f"Silla {SUPPLIER_LABELS[supplier]}",
        "description": "Producto de integracion local.",
        "unit": "PZA",
        "availability_type": "stocked",
        "stock": "5.000000",
        "lead_time": "Entrega inmediata",
        "base_price_options": base_price_options or [],
        "add_on_options": add_on_options or [],
        "base_currency": "MXN",
        "price_net": price,
        "tax_rate": "0.160000",
        "attributes": {
            "reference_price_mxn": "999.000000",
            "price_status": "precio_por_confirmar" if price is None else "confirmado",
            "quotable": True,
        },
        "image_url": "",
        "image_kind": "placeholder",
        "product_url": "",
        "warnings": warnings or [],
        "source_reference": f"{supplier}:integration",
    }


def _catalog(supplier: str, items: list[dict]) -> dict:
    return load_supplier_catalog_data({
        "supplier": supplier,
        "source_hash": ("a" if supplier == "idelika" else "b") * 64,
        "generated_at": "2026-08-03T00:00:00+00:00",
        "items": items,
    })


def _project_line(catalog: str, internal_id: str, *, base_option_id: str = "", add_on_option_ids: list[str] | None = None) -> dict:
    return {
        "line_id": "11111111-1111-4111-8111-111111111111",
        "role": "principal",
        "section_id": "principal",
        "parent_line_id": None,
        "position": 0,
        "quantity": "1",
        "source": "catalog",
        "official_code": f"{catalog.upper()}-1",
        "display_cache": {"name": "Silla", "code": f"{catalog.upper()}-1", "image_url": ""},
        "catalog": catalog,
        "identity": {
            "internal_id": internal_id,
            "base_option_id": base_option_id,
            "add_on_option_ids": add_on_option_ids or [],
        },
    }


def test_precio_nulo_falla_cerrado_sin_contrato_explicito_completo():
    raw = _item(
        "idelika",
        internal_id="idelika:pending-invalid",
        price=None,
        warnings=["price_pending"],
        code_status="needs_review",
    )
    for mutation in ("warning", "quotable", "price_status"):
        invalid = deepcopy(raw)
        if mutation == "warning":
            invalid["warnings"] = ["missing_code"]
        elif mutation == "quotable":
            invalid["attributes"]["quotable"] = False
        else:
            invalid["attributes"]["price_status"] = "confirmado"
        with pytest.raises(ValueError):
            _catalog("idelika", [invalid])


def test_idelika_y_conceptos_participan_en_busqueda_proyecto_y_cotizacion_mixta():
    idelika = _catalog("idelika", [
        _item("idelika", internal_id="idelika:priced", price="100.000000"),
        _item(
            "idelika",
            internal_id="idelika:pending",
            price=None,
            warnings=["price_pending", "missing_code"],
            code_status="needs_review",
        ),
    ])
    conceptos = _catalog("conceptos", [
        _item(
            "conceptos",
            internal_id="conceptos:configurable",
            price="0.000000",
            base_price_options=[{
                "id": "conceptos:base", "name": "Tapiz base", "price_net": "200.000000", "available": True,
            }],
            add_on_options=[{
                "id": "conceptos:cojin", "name": "Cojin", "family": "cojin", "price_net": "15.000000",
                "available": True, "compatible_base_option_ids": ["conceptos:base"],
            }],
        ),
    ])
    catalogs = {"idelika": idelika, "conceptos": conceptos}

    assert {supplier: SUPPLIER_LABELS[supplier] for supplier in catalogs} == {
        "idelika": "IDÉLIKA", "conceptos": "Conceptos",
    }
    assert {supplier: EXPECTED_SUPPLIER_BASE_CURRENCY[supplier] for supplier in catalogs} == {
        "idelika": "MXN", "conceptos": "MXN",
    }
    assert MIXED_CATALOG_ORDER[-4:-2] == ("idelika", "conceptos")
    assert ENGINE_MIXED_CATALOG_ORDER[-4:-2] == ("idelika", "conceptos")

    search_items = [
        item
        for supplier in ("idelika", "conceptos")
        for item in search_catalog_products(
            catalogs, query="silla", supplier=supplier, offset=0, limit=20
        )["items"]
    ]
    assert [item["catalog"] for item in search_items] == ["idelika", "idelika", "conceptos"]
    assert search_items[0]["base_currency"] == "MXN"
    pending_search = next(
        item for item in search_items
        if item["identity"].get("internal_id") == "idelika:pending"
    )
    assert pending_search["official_code"] == ""
    assert pending_search["display_key"] == "idelika:pending"
    assert pending_search["snapshot"]["code"] == ""
    assert pending_search["snapshot"]["warnings"] == [
        "Código por verificar",
        "Precio por confirmar",
    ]
    project = normalize_project_payload({
        "schema_version": 1,
        "quote_fields": {
            "proyecto": "", "cliente": "", "correo": "", "telefono": "",
            "direccion": "", "razon_social": "", "quote_currency": "MXN", "descuento": "40",
            "template": "official_2026_gdl", "description_language": "es",
        },
        "sections": [{"section_id": "principal", "concept": "Principal", "position": 0}],
        "lines": [_project_line(
            "conceptos", "conceptos:configurable",
            base_option_id="conceptos:base", add_on_option_ids=["conceptos:cojin"],
        )],
    })
    assert project["lines"][0]["catalog"] == "conceptos"
    assert project["lines"][0]["identity"] == {
        "internal_id": "conceptos:configurable",
        "base_option_id": "conceptos:base",
        "add_on_option_ids": ["conceptos:cojin"],
    }

    rows = [
        {"catalog": "idelika", "internal_id": "idelika:priced", "quantity": "1"},
        {"catalog": "idelika", "internal_id": "idelika:pending", "quantity": "1"},
        {
            "catalog": "conceptos", "internal_id": "conceptos:configurable", "quantity": "1",
            "base_option_id": "conceptos:base", "add_on_option_ids": ["conceptos:cojin"],
        },
    ]
    mxn_payload = build_mixed_catalog_cart_payload(
        rows, catalogs=catalogs, rate_rows=RATE_ROWS, quote_currency="MXN",
        commercial_discount_percent="40", today=TODAY,
    )
    mxn_lines = {line["canonical_key"]: line for group in mxn_payload["groups"] for line in group["items"]}
    assert mxn_lines['idelika:["idelika:priced","",[]]']["original_unit_price"] == "100.000000"
    assert mxn_lines['idelika:["idelika:priced","",[]]']["unit_price"] == "100.00"
    assert mxn_lines['idelika:["idelika:pending","",[]]']["original_unit_price"] is None
    assert mxn_lines['idelika:["idelika:pending","",[]]']["unit_price"] is None
    assert mxn_lines['idelika:["idelika:pending","",[]]']["frozen_exchange_rate"] is None
    assert "Precio por confirmar" in mxn_lines['idelika:["idelika:pending","",[]]']["warnings"]
    assert mxn_lines['conceptos:["conceptos:configurable","conceptos:base",["conceptos:cojin"]]']["original_unit_price"] == "215.000000"
    assert mxn_lines['conceptos:["conceptos:configurable","conceptos:base",["conceptos:cojin"]]']["unit_price"] == "215.00"
    canonical_pending = next(
        row for row in quotation_data_rows(mxn_payload)
        if row.original_cost is None
    )
    assert canonical_pending.original_cost is None
    assert canonical_pending.frozen_rate is None
    assert canonical_pending.converted_cost is None

    pending_payload = build_supplier_cart_payload(
        [{"internal_id": "idelika:pending", "quantity": "1"}],
        idelika, "MXN", RATE_ROWS, today=TODAY,
    )
    pending_line = pending_payload["items"][0]
    assert pending_line["base_price"] is None
    assert pending_line["unit_price_base"] is None
    assert pending_line["unit_price"] is None
    assert pending_line["line_total"] is None
    assert pending_line["attributes"]["reference_price_mxn"] == "999.000000"
    assert pending_line["unit_price_base"] != pending_line["attributes"]["reference_price_mxn"]

    usd_payload = build_mixed_catalog_cart_payload(
        rows[:1], catalogs=catalogs, rate_rows=RATE_ROWS, quote_currency="USD",
        commercial_discount_percent="40", today=TODAY,
    )
    usd_line = usd_payload["groups"][0]["items"][0]
    assert usd_line["original_unit_price"] == "100.000000"
    assert usd_line["frozen_exchange_rate"] == "0.050000"
    assert usd_line["unit_price"] == "5.00"


def test_worker_y_compositor_xlsx_muestran_pendiente_sin_ceros(
    monkeypatch,
    tmp_path: Path,
):
    catalog = _catalog("idelika", [
        _item("idelika", internal_id="idelika:priced-xlsx", price="100.000000"),
        _item(
            "idelika",
            internal_id="idelika:pending-xlsx",
            price=None,
            warnings=["price_pending", "missing_code"],
            code_status="needs_review",
        ),
    ])
    payload = build_mixed_catalog_cart_payload(
        [
            {"catalog": "idelika", "internal_id": "idelika:priced-xlsx", "quantity": "2"},
            {"catalog": "idelika", "internal_id": "idelika:pending-xlsx", "quantity": "1"},
        ],
        catalogs={"idelika": catalog},
        rate_rows=RATE_ROWS,
        quote_currency="MXN",
        commercial_discount_percent="40",
        today=TODAY,
    )
    monkeypatch.setattr(catalog_cart, "_download_catalog_image", lambda *_args, **_kwargs: None)
    intermediate = tmp_path / "idelika-pending-worker.xlsx"
    _convert_mixed_catalog_cart_to_quotation(
        tmp_path / "idelika-pending.json",
        intermediate,
        payload,
    )
    workbook = load_workbook(intermediate, data_only=False)
    try:
        quotation = workbook["Quotation"]
        pending_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == "Silla IDÉLIKA"
            and quotation.cell(row, 10).value == "Por confirmar"
        )
        assert quotation.cell(pending_row, 15).value is None
        assert quotation.cell(pending_row, 16).value is None
    finally:
        workbook.close()

    template = Path("mobiliti_saas/worker/templates/Formato Cotizacion 2026 Oficial.xlsx")
    output = tmp_path / "idelika-pending-official.xlsx"
    generate_quote(
        intermediate,
        output,
        {
            "catalog_price_mode": "mixed_catalog_converted",
            "catalog_source_hashes": {"idelika": catalog["source_hash"]},
            "quote_currency": "MXN",
            "rate_summary": deepcopy(payload["rate_summary"]),
            "auto_electrification_rate": None,
            "descuento": 40,
            "cotizacion": "IDELIKA-PENDING",
            "proyecto": "Pendientes",
            "cliente": "Cliente",
        },
        template,
        original_quotation_path=None,
        quotation_data_rows=quotation_data_rows(payload),
    )
    result = load_workbook(output, data_only=False)
    try:
        quotation = result["Quotation"]
        pending_row = next(
            row for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 2).value == "Silla IDÉLIKA"
            and quotation.cell(row, 11).value == "Por confirmar"
        )
        assert quotation.cell(pending_row, 12).value in {None, ""}
        mobiliti = result["Mobiliti"]
        mobiliti_row = next(
            row for row in range(1, mobiliti.max_row + 1)
            if mobiliti.cell(row, 4).value == f"=Quotation!B{pending_row}"
        )
        assert mobiliti.cell(mobiliti_row, 10).value == "Por confirmar"
        cotizacion = result["Cotizacion"]
        cotizacion_row = next(
            row for row in range(1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 1).value == f"=Mobiliti!D{mobiliti_row}"
        )
        assert cotizacion.cell(cotizacion_row, 6).value == "Por confirmar"
        assert cotizacion.cell(cotizacion_row, 9).value == "Por confirmar"
        assert cotizacion.cell(cotizacion_row, 10).value in {None, ""}
        audit = result["Quotation_Data"]
        pending_audit_row = next(
            row for row in range(2, audit.max_row + 1)
            if audit.cell(row, 8).value is None
        )
        assert audit.cell(pending_audit_row, 9).value is None
        assert audit.cell(pending_audit_row, 10).value is None
    finally:
        result.close()
