from datetime import date
from decimal import Decimal
from copy import deepcopy
import base64
from io import BytesIO
import hashlib
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorkbookImage
from PIL import Image
import pytest

from mobiliti_saas.quote_engine import catalog_cart
from mobiliti_saas.quote_engine.catalog_cart import (
    _catalog_warning_key,
    _description_for_item,
    create_catalog_quotation_workbook,
)
from mobiliti_saas.quote_engine.mixed_catalog import (
    build_mixed_catalog_cart_payload,
    create_mixed_catalog_quotation_workbook,
    validate_mixed_catalog_payload,
)
from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem
from mobiliti_saas.quote_engine.quotation_import import build_import_manifest
from mobiliti_saas.quote_engine.tarkett_catalog import TarkettCatalogItem
from quotation_import_fixtures import write_import_fixture


def _supplier_item(catalog, *, currency, internal_id):
    return {
        "internal_id": internal_id,
        "supplier": catalog,
        "product_key": internal_id,
        "sku": internal_id.upper(),
        "code_status": "verified",
        "brand": catalog,
        "collection": catalog,
        "name": f"Producto {catalog}",
        "description": "",
        "unit": "pieza",
        "availability_type": "made_to_order",
        "stock": None,
        "lead_time": "Sobre pedido",
        "base_price_options": [],
        "add_on_options": [],
        "base_currency": currency,
        "price_net": "100.000000",
        "tax_rate": "0.160000",
        "attributes": {},
        "image_url": "",
        "image_kind": "placeholder",
        "product_url": "",
        "warnings": [],
        "source_reference": f"{catalog}:source",
    }


def frozen_payload():
    tarkett = TarkettCatalogItem(
        "25731726",
        "Piso Tarkett",
        "m2",
        Decimal("10"),
        unit_price=Decimal("100"),
        price_source="catalog",
    )
    catalogs = {
        "tarkett": {
            "source_hash": "a" * 64,
            "items": [tarkett],
            "by_code": {tarkett.code: tarkett},
        },
        "sonara": {
            "supplier": "sonara",
            "source_hash": "b" * 64,
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [_supplier_item("sonara", currency="MXN", internal_id="sonara:desk-1")],
        },
        "alma": {
            "supplier": "alma",
            "source_hash": "c" * 64,
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [_supplier_item("alma", currency="USD", internal_id="alma:desk-1")],
        },
    }
    rows = [
        {"catalog": "tarkett", "code": tarkett.code, "quantity": "1"},
        {"catalog": "sonara", "internal_id": "sonara:desk-1", "quantity": "1"},
        {"catalog": "alma", "internal_id": "alma:desk-1", "quantity": "1"},
    ]
    rates = [
        {
            "currency": "USD",
            "effective_date": "2026-07-19",
            "mxn_per_unit": "18.500000",
            "retrieved_at": "2026-07-19T12:00:00+00:00",
        }
    ]
    return build_mixed_catalog_cart_payload(
        rows,
        catalogs=catalogs,
        rate_rows=rates,
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[
            {"id": "section-1", "title": "Tarkett", "item_keys": ["tarkett:25731726"]},
            {"id": "section-2", "title": "Sonara", "item_keys": ["sonara:[\"sonara:desk-1\",\"\",[]]"]},
            {"id": "section-3", "title": "ALMA", "item_keys": ["alma:[\"alma:desk-1\",\"\",[]]"]},
        ],
        today=date(2026, 7, 19),
    )


def _png_bytes(color):
    stream = BytesIO()
    Image.new("RGB", (2, 2), color).save(stream, format="PNG")
    return stream.getvalue()


def _write_two_product_import_fixture(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    for column, title in {
        1: "No.",
        2: "Item Name",
        3: "Photo",
        4: "Description",
        5: "Dimension",
        7: "Q'ty",
        10: "Unit Price",
    }.items():
        sheet.cell(7, column, title)
    sheet["A1"] = "Proveedor prueba"
    image_streams = []
    for row, index, category, name, color in (
        (9, 1, "SECCION A", "Producto A", "navy"),
        (11, 2, "SECCION B", "Producto B", "orange"),
    ):
        sheet.cell(row - 1, 1, f"- {category}")
        sheet.cell(row, 1, index)
        sheet.cell(row, 2, name)
        sheet.cell(row, 4, f"Descripcion {name}")
        sheet.cell(row, 5, f"Dimension {name}")
        sheet.cell(row, 7, 1)
        sheet.cell(row, 10, 100 + index)
        stream = BytesIO(_png_bytes(color))
        image_streams.append(stream)
        sheet.add_image(WorkbookImage(stream), f"C{row}")
    workbook.save(path)
    workbook.close()
    return path


def _stub_catalog_image_transport(monkeypatch, responses):
    class Response:
        def __init__(self, data, content_type):
            self.data = data
            self.headers = {
                "content-type": content_type,
                "content-length": str(len(data)),
            }

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def read(self, size):
            return self.data[:size]

    class Opener:
        def open(self, request, timeout):
            assert timeout == 18
            data, content_type = responses[request.full_url]
            return Response(data, content_type)

    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    monkeypatch.setattr(catalog_cart, "_validate_connected_peer", lambda response: None)
    monkeypatch.setattr(
        catalog_cart.urllib.request,
        "build_opener",
        lambda *handlers: Opener(),
    )


def _three_image_payload():
    tarkett = TarkettCatalogItem(
        "25731726",
        "Piso Tarkett",
        "m2",
        Decimal("10"),
        image_url="https://media.tarkett-image.com/tarkett.png",
        unit_price=Decimal("100"),
        price_source="catalog",
    )
    offiho = OffihoCatalogItem(
        "offiho:desk-1",
        "OHE-1",
        "Escritorio",
        "Negro",
        "PZA",
        Decimal("1"),
        Decimal("8"),
        Decimal("200"),
        price_source="catalog",
        image_url="https://www.offiho.com/offiho.png",
    )
    alma = _supplier_item("alma", currency="USD", internal_id="alma:desk-1")
    alma["image_url"] = "https://assets.example.test/alma.png"
    alma["product_url"] = "https://products.example.test/no-get"
    catalogs = {
        "tarkett": {
            "source_hash": "a" * 64,
            "items": [tarkett],
            "by_code": {tarkett.code: tarkett},
        },
        "offiho": {
            "source_hash": "b" * 64,
            "items": [offiho],
            "by_inventory_key": {offiho.inventory_key: offiho},
        },
        "alma": {
            "supplier": "alma",
            "source_hash": "c" * 64,
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [alma],
        },
    }
    rows = [
        {"catalog": "tarkett", "code": tarkett.code, "quantity": "1"},
        {"catalog": "offiho", "inventory_key": offiho.inventory_key, "quantity": "1"},
        {"catalog": "alma", "internal_id": alma["internal_id"], "quantity": "1"},
    ]
    return build_mixed_catalog_cart_payload(
        rows,
        catalogs=catalogs,
        rate_rows=[
            {
                "currency": "USD",
                "effective_date": "2026-07-19",
                "mxn_per_unit": "18.500000",
                "retrieved_at": "2026-07-19T12:00:00+00:00",
            }
        ],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[
            {"id": f"section-{index}", "title": title, "item_keys": [key]}
            for index, (title, key) in enumerate(
                (
                    ("Tarkett", "tarkett:25731726"),
                    ("Offiho", "offiho:offiho:desk-1"),
                    ("ALMA", "alma:[\"alma:desk-1\",\"\",[]]"),
                ),
                1,
            )
        ],
        today=date(2026, 7, 19),
    )


def test_mixed_workbook_has_one_quotation_with_provider_sections_and_audit_columns(tmp_path):
    output = tmp_path / "mixed.xlsx"
    create_mixed_catalog_quotation_workbook(
        frozen_payload(), output, image_dir=tmp_path / "images"
    )
    wb = load_workbook(output, data_only=False)
    assert wb.sheetnames == ["Quotation"]
    ws = wb["Quotation"]
    assert [ws.cell(7, column).value for column in range(12, 20)] == [
        "Supplier",
        "Discount Percent",
        "Original Currency",
        "Original Unit Price",
        "Frozen Exchange Rate",
        "Source Reference",
        "Price Mode",
        "Auto Electrification",
    ]
    assert [ws.cell(row, 1).value for row in (8, 10, 12)] == [
        "- Tarkett",
        "- Sonara",
        "- ALMA",
    ]
    assert [ws.cell(row, 1).value for row in (9, 11, 13)] == [1, 2, 3]
    assert [ws.cell(row, 12).value for row in (9, 11, 13)] == [
        "Tarkett",
        "Sonara",
        "ALMA",
    ]
    assert [ws.cell(row, 13).value for row in (9, 11, 13)] == [40, 0, 0]
    assert ws["M9"].number_format == "0.000000"
    assert ws["N13"].value == "USD"
    assert ws["O13"].value == 100
    assert ws["P13"].value == 18.5
    assert ws["J13"].value == 1850
    assert ws["R13"].value == "net"
    assert ws["S9"].value is True
    assert ws["S11"].value is False
    assert ws["S13"].value is False
    wb.close()


def test_mixed_workbook_uses_manual_sections_and_interleaved_supplier_order(tmp_path):
    payload = frozen_payload()
    tarkett_key = payload["groups"][0]["items"][0]["canonical_key"]
    sonara_key = payload["groups"][1]["items"][0]["canonical_key"]
    alma_key = payload["groups"][2]["items"][0]["canonical_key"]
    payload["sections"] = [
        {"id": "section-1", "title": "Recepción", "item_keys": [alma_key, tarkett_key]},
        {"id": "section-2", "title": "Privados", "item_keys": [sonara_key]},
    ]

    output = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "manual-sections.xlsx",
        image_dir=tmp_path / "images",
    )
    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]
    assert [ws.cell(row, 1).value for row in (8, 11)] == [
        "- Recepción",
        "- Privados",
    ]
    assert [ws.cell(row, 12).value for row in (9, 10, 12)] == [
        "ALMA",
        "Tarkett",
        "Sonara",
    ]
    assert [ws.cell(row, 1).value for row in (9, 10, 12)] == [1, 2, 3]
    wb.close()


def test_mixed_workbook_interleaves_catalog_and_imported_rows_with_original_image(
    monkeypatch, tmp_path
):
    source = write_import_fixture(tmp_path / "imported-source.xlsx")
    source_hash_before = hashlib.sha256(source.read_bytes()).hexdigest()
    manifest, image_map = build_import_manifest(
        source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=source.name,
    )
    imported = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[{
            "currency": "USD",
            "effective_date": "2026-07-21",
            "mxn_per_unit": "18.500000",
            "retrieved_at": "2026-07-21T00:00:00Z",
        }],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Recepción",
            "item_keys": [f"import:{manifest['import_id']}:11"],
        }],
        imported_source={
            "manifest": manifest,
            "items": [{
                "kind": "imported",
                "import_id": manifest["import_id"],
                "source_row": 11,
                "source_currency": "USD",
                "quantity": "2",
                "overrides": {
                    "name": "Alien Task Chair revisada",
                    "description": "Silla operativa revisada",
                    "dimension": "630 x 565 x 1000 mm",
                    "unit_price": "82.00",
                    "provider": "Sunon",
                },
            }],
            "source_currency": "USD",
        },
        today=date(2026, 7, 21),
    )
    payload = _three_image_payload()
    catalog_key = payload["groups"][0]["items"][0]["canonical_key"]
    offiho_key = payload["groups"][1]["items"][0]["canonical_key"]
    alma_key = payload["groups"][2]["items"][0]["canonical_key"]
    for group in payload["groups"][1:]:
        group["items"][0]["image_url"] = ""
        group["items"][0]["image_kind"] = "placeholder"
    imported_key = imported["imported_source"]["items"][0]["canonical_key"]
    payload["imported_source"] = imported["imported_source"]
    payload["sections"] = [
        {
            "id": "section-1",
            "title": "Recepción",
            "item_keys": [catalog_key, imported_key],
        },
        {"id": "section-2", "title": "Operación", "item_keys": [offiho_key]},
        {"id": "section-3", "title": "Dirección", "item_keys": [alma_key]},
    ]
    payload["item_count"] += 1

    catalog_image = _png_bytes("navy")

    def fake_download(url, image_dir, code, source_type, destination_key=None):
        path = image_dir / "catalog.png"
        path.write_bytes(catalog_image)
        return path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)
    output = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "mixed-imported.xlsx",
        image_dir=tmp_path / "images",
        imported_source_path=source,
    )

    workbook = load_workbook(output)
    quotation = workbook["Quotation"]
    assert quotation["A8"].value == "- Recepción"
    assert quotation["B9"].value == "Piso Tarkett"
    assert quotation["B10"].value == "Alien Task Chair revisada"
    assert quotation["D10"].value.startswith("Silla operativa revisada")
    assert quotation["E10"].value == "630 x 565 x 1000 mm"
    assert quotation["G10"].value == 2
    assert quotation["J10"].value == 1517
    assert quotation["K10"].value is None
    assert quotation["L10"].value == "Sunon"
    assert quotation["N10"].value == "USD"
    assert quotation["O10"].value == 82
    assert quotation["P10"].value == 18.5
    assert quotation["Q10"].value == f"{source.name}#Quotation!11"
    assert quotation["R10"].value == "imported"
    assert quotation["S10"].value is False
    anchored_hashes = {
        image.anchor._from.row + 1: hashlib.sha256(image._data()).hexdigest()
        for image in quotation._images
    }
    assert anchored_hashes == {
        9: hashlib.sha256(catalog_image).hexdigest(),
        10: hashlib.sha256(image_map[11][0]).hexdigest(),
    }
    workbook.close()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash_before


def test_imported_only_workbook_keeps_duplicate_names_bound_to_distinct_source_rows(
    tmp_path
):
    source = write_import_fixture(tmp_path / "duplicate-names.xlsx")
    source_workbook = load_workbook(source)
    source_sheet = source_workbook["Quotation"]
    source_sheet._images = [
        image
        for image in source_sheet._images
        if image.anchor._from.row + 1 != 9
    ]
    source_workbook.save(source)
    source_workbook.close()
    source_hash_before = hashlib.sha256(source.read_bytes()).hexdigest()
    manifest, image_map = build_import_manifest(
        source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=source.name,
    )
    imported_keys = [
        f"import:{manifest['import_id']}:9",
        f"import:{manifest['import_id']}:17",
    ]
    items = [
        {
            "kind": "imported",
            "import_id": manifest["import_id"],
            "source_row": source_row,
            "source_currency": "MXN",
            "quantity": "1",
            "overrides": {
                "name": "Mesa DV74 repetida",
                "description": f"Fila original {source_row}",
                "dimension": f"Dimension {source_row}",
                "unit_price": "100.00",
                "provider": "Sunon",
            },
        }
        for source_row in (9, 17)
    ]
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Mesas",
            "item_keys": imported_keys,
        }],
        imported_source={
            "manifest": manifest,
            "items": items,
            "source_currency": "MXN",
        },
        today=date(2026, 7, 21),
    )

    output = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "imported-only.xlsx",
        imported_source_path=source,
    )
    workbook = load_workbook(output)
    quotation = workbook["Quotation"]
    assert payload["groups"] == []
    assert [quotation.cell(row, 1).value for row in (9, 10)] == [1, 2]
    assert [quotation.cell(row, 2).value for row in (9, 10)] == [
        "Mesa DV74 repetida",
        "Mesa DV74 repetida",
    ]
    assert [quotation.cell(row, 5).value for row in (9, 10)] == [
        "Dimension 9",
        "Dimension 17",
    ]
    assert [quotation.cell(row, 17).value for row in (9, 10)] == [
        f"{source.name}#Quotation!9",
        f"{source.name}#Quotation!17",
    ]
    assert {
        image.anchor._from.row + 1: hashlib.sha256(image._data()).hexdigest()
        for image in quotation._images
    } == {10: hashlib.sha256(image_map[17][0]).hexdigest()}
    workbook.close()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash_before


def test_mixed_workbook_rejects_changed_import_source_before_creating_output_paths(
    tmp_path
):
    source = write_import_fixture(tmp_path / "changed-source.xlsx")
    manifest, _images = build_import_manifest(
        source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=source.name,
    )
    imported_key = f"import:{manifest['import_id']}:11"
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Recepción",
            "item_keys": [imported_key],
        }],
        imported_source={
            "manifest": manifest,
            "items": [{
                "kind": "imported",
                "import_id": manifest["import_id"],
                "source_row": 11,
                "source_currency": "MXN",
                "quantity": "1",
                "overrides": {
                    "name": "Alien Task Chair",
                    "description": "Silla operativa",
                    "dimension": "630 x 565 x 1000 mm",
                    "unit_price": "82.00",
                    "provider": "Sunon",
                },
            }],
            "source_currency": "MXN",
        },
        today=date(2026, 7, 21),
    )
    altered = source.read_bytes() + b"changed"
    output = tmp_path / "output" / "changed.xlsx"
    images = tmp_path / "images"

    with pytest.raises(ValueError, match="fuente importada cambio"):
        create_mixed_catalog_quotation_workbook(
            payload,
            output,
            image_dir=images,
            imported_source_path=altered,
        )

    assert not output.parent.exists()
    assert not images.exists()


def test_mixed_workbook_rejects_self_consistent_imported_row_remap_before_output(
    tmp_path,
):
    source = _write_two_product_import_fixture(tmp_path / "two-products.xlsx")
    manifest, _image_map = build_import_manifest(
        source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=source.name,
    )
    original_key = f"import:{manifest['import_id']}:11"
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Presentacion editable",
            "item_keys": [original_key],
        }],
        imported_source={
            "manifest": manifest,
            "items": [{
                "kind": "imported",
                "import_id": manifest["import_id"],
                "source_row": 11,
                "source_currency": "MXN",
                "quantity": "1",
                "overrides": {
                    "name": "Producto B editado",
                    "description": "Descripcion comercial permitida",
                    "dimension": "Dimension editada",
                    "unit_price": "222.00",
                    "provider": "Proveedor editado",
                },
            }],
            "source_currency": "MXN",
        },
        today=date(2026, 7, 21),
    )
    authoritative_row_9 = next(
        item for item in manifest["items"] if item["source_row"] == 9
    )
    remapped_key = f"import:{manifest['import_id']}:9"
    imported_line = payload["imported_source"]["items"][0]
    imported_line.update(
        source_row=9,
        canonical_key=remapped_key,
        source_reference=f"{source.name}#Quotation!9",
        row_hash=authoritative_row_9["row_hash"],
    )
    payload["sections"][0]["item_keys"] = [remapped_key]

    # El validador estructural no puede ligar la fila al XLSX; el builder si debe.
    assert validate_mixed_catalog_payload(payload) is payload
    output = tmp_path / "output" / "remapped.xlsx"
    images = tmp_path / "images"
    with pytest.raises(ValueError, match="fila importada"):
        create_mixed_catalog_quotation_workbook(
            payload,
            output,
            image_dir=images,
            imported_source_path=source,
        )

    assert not output.parent.exists()
    assert not images.exists()


def test_local_image_data_enforces_byte_limit_before_decode_and_anchor(monkeypatch):
    workbook = Workbook()
    sheet = workbook.active
    calls = []
    monkeypatch.setattr(catalog_cart, "MAX_IMAGE_BYTES", 4)
    monkeypatch.setattr(
        catalog_cart,
        "_validated_catalog_image_suffix",
        lambda data, content_type: calls.append(("validate", data, content_type)),
    )
    monkeypatch.setattr(
        catalog_cart,
        "_anchor_catalog_image",
        lambda ws, row, source: calls.append(("anchor", row, source.read())),
    )

    catalog_cart._add_local_catalog_image(sheet, 9, (b"1234", "image/png"))
    assert calls == [
        ("validate", b"1234", "image/png"),
        ("anchor", 9, b"1234"),
    ]

    catalog_cart._add_local_catalog_image(sheet, 10, (b"12345", "image/png"))
    assert calls == [
        ("validate", b"1234", "image/png"),
        ("anchor", 9, b"1234"),
    ]
    workbook.close()


def test_mixed_workbook_preserves_configuration_review_warning_and_safe_text(tmp_path):
    payload = frozen_payload()
    original = deepcopy(payload)
    sonara = payload["groups"][1]["items"][0]
    sonara["name"] = '=HYPERLINK("bad")'
    sonara["code"] = ""
    sonara["code_status"] = "needs_review"
    sonara["warnings"] = ["Codigo por verificar"]
    sonara["source_reference"] = "sonara:catalogo-2026:pagina-4"
    alma = payload["groups"][2]["items"][0]
    alma["configuration"] = "Cubierta nogal; electrificacion A+C"
    submitted = deepcopy(payload)

    output = create_mixed_catalog_quotation_workbook(payload, tmp_path / "safe.xlsx")
    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]
    assert str(ws["B11"].value).startswith("'")
    assert "Codigo por verificar" in ws["D11"].value
    assert "Fuente: sonara:catalogo-2026:pagina-4" in ws["D11"].value
    assert ws["D11"].fill.fgColor.rgb.endswith("FFF2CC")
    assert "Cubierta nogal; electrificacion A+C" in ws["D13"].value
    assert ws["Q11"].value == "sonara:catalogo-2026:pagina-4"
    wb.close()

    assert original["groups"][0] == payload["groups"][0]
    assert payload == submitted


def test_mixed_workbook_validates_before_creating_output_or_image_directories(tmp_path):
    output = tmp_path / "output" / "mixed.xlsx"
    images = tmp_path / "images"
    with pytest.raises(ValueError, match="Grupos mixtos invalidos"):
        create_mixed_catalog_quotation_workbook({}, output, image_dir=images)
    assert not output.parent.exists()
    assert not images.exists()


def test_catalog_warning_merge_keeps_each_semantic_warning_once():
    description, warning = _description_for_item(
        {
            "description": "Descripcion original",
            "code_status": "needs_review",
            "image_kind": "generated_reference",
            "availability_type": "stocked",
            "stock_status": "insufficient_stock",
            "available_after_reservations": "2",
            "price_source": "missing",
            "warnings": [
                "precio por confirmar",
                "IMAGEN DE REFERENCIA",
                "Código por verificar",
                "existencia insuficiente; verificar disponibilidad",
                "Validar acabado manualmente",
            ],
        },
        "",
        "",
        Decimal("3"),
    )
    semantic_keys = [_catalog_warning_key(value) for value in warning.split(" | ")]
    for key in (
        "precio por confirmar",
        "imagen de referencia",
        "codigo por verificar",
        "existencia insuficiente",
    ):
        assert semantic_keys.count(key) == 1
    assert "SOLICITADO 3 - DISPONIBLE 2" in description
    assert description.count("Validar acabado manualmente") == 1


def test_legacy_supplier_sku_fallback_keeps_a_to_k_and_image_anchor(monkeypatch, tmp_path):
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
    )
    image_path = tmp_path / "SKU-ONLY.png"
    image_path.write_bytes(png)
    calls = []

    def fake_download(url, image_dir, code, source_type):
        calls.append((url, image_dir, code, source_type))
        return image_path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)
    payload = {
        "source_type": "supplier_cart",
        "items": [
            {
                "sku": "SKU-ONLY",
                "name": "Silla legacy",
                "description": "Descripcion legacy",
                "unit": "pieza",
                "quantity": "2",
                "unit_price": "125.5",
                "product_url": "https://example.test/producto",
                "image_url": "https://assets.example.test/sku-only.png",
                "attributes": {"dimensions": "60 x 60 cm"},
                "warnings": [],
            }
        ],
    }
    output = create_catalog_quotation_workbook(
        payload,
        tmp_path / "legacy.xlsx",
        source_type="supplier_cart",
        category_label="Legacy",
        image_dir=tmp_path / "images",
    )
    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]
    assert [ws.cell(9, column).value for column in range(1, 12)] == [
        1,
        "Silla legacy",
        None,
        "Descripcion legacy | SKU: SKU-ONLY | URL: https://example.test/producto",
        "60 x 60 cm",
        None,
        2,
        None,
        None,
        125.5,
        "https://example.test/producto",
    ]
    assert calls == [
        (
            "https://assets.example.test/sku-only.png",
            tmp_path / "images",
            "SKU-ONLY",
            "supplier_cart",
        )
    ]
    assert [image.anchor._from.row + 1 for image in ws._images] == [9]
    wb.close()


def test_mixed_workbook_uses_only_frozen_image_urls_with_source_policy(monkeypatch, tmp_path):
    payload = _three_image_payload()
    calls = []

    def fake_download(url, image_dir, code, source_type, destination_key=None):
        row = int(destination_key.split("-", 2)[1])
        calls.append((source_type, url, row))
        path = image_dir / f"fixture-{row}.png"
        path.write_bytes(_png_bytes("blue"))
        return path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)
    output = create_mixed_catalog_quotation_workbook(
        payload, tmp_path / "images.xlsx", image_dir=tmp_path / "images"
    )
    wb = load_workbook(output)
    ws = wb["Quotation"]
    assert calls == [
        ("tarkett_cart", "https://media.tarkett-image.com/tarkett.png", 9),
        ("offiho_cart", "https://www.offiho.com/offiho.png", 11),
        ("supplier_cart", "https://assets.example.test/alma.png", 13),
    ]
    assert sorted(image.anchor._from.row + 1 for image in ws._images) == [9, 11, 13]
    assert all("products.example.test" not in url for _, url, _ in calls)
    wb.close()


def test_mixed_workbook_keeps_structured_visual_semantics(tmp_path):
    offiho = OffihoCatalogItem(
        "offiho:desk-1",
        "OHE-1",
        "Escritorio",
        "Negro",
        "PZA",
        Decimal("1"),
        Decimal("1"),
        Decimal("0"),
        price_source="missing",
    )
    sonara = _supplier_item("sonara", currency="MXN", internal_id="sonara:desk-1")
    sonara.update(
        image_kind="generated_reference",
        warnings=["Imagen de referencia"],
    )
    alma = _supplier_item("alma", currency="USD", internal_id="alma:desk-1")
    alma["lead_time"] = "6 semanas"
    catalogs = {
        "offiho": {
            "source_hash": "a" * 64,
            "items": [offiho],
            "by_inventory_key": {offiho.inventory_key: offiho},
        },
        "sonara": {
            "supplier": "sonara",
            "source_hash": "b" * 64,
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [sonara],
        },
        "alma": {
            "supplier": "alma",
            "source_hash": "c" * 64,
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [alma],
        },
    }
    payload = build_mixed_catalog_cart_payload(
        [
            {"catalog": "offiho", "inventory_key": offiho.inventory_key, "quantity": "2"},
            {"catalog": "sonara", "internal_id": sonara["internal_id"], "quantity": "1"},
            {"catalog": "alma", "internal_id": alma["internal_id"], "quantity": "1"},
        ],
        catalogs=catalogs,
        rate_rows=[
            {
                "currency": "USD",
                "effective_date": "2026-07-19",
                "mxn_per_unit": "18.500000",
                "retrieved_at": "2026-07-19T12:00:00+00:00",
            }
        ],
        quote_currency="MXN",
        commercial_discount_percent="40",
        today=date(2026, 7, 19),
    )
    offiho_line = payload["groups"][0]["items"][0]
    offiho_line.update(
        reserved_quantity="0.000000",
        available_after_reservations="1.000000",
        reserved_by_others=False,
    )
    offiho_line["warnings"].append("Existencia insuficiente; verificar disponibilidad.")
    output = create_mixed_catalog_quotation_workbook(payload, tmp_path / "visual.xlsx")
    wb = load_workbook(output)
    ws = wb["Quotation"]
    descriptions = {9: ws["D9"].value, 10: ws["D10"].value, 11: ws["D11"].value}
    for expected in ("Variante: Negro", "PRECIO POR CONFIRMAR", "EXISTENCIA INSUFICIENTE"):
        assert descriptions[9].count(expected) == 1
    assert descriptions[10].count("Imagen de referencia") == 1
    assert descriptions[11].count("Entrega: 6 semanas") == 1
    wb.close()


def test_mixed_image_destination_keys_prevent_same_code_collisions(monkeypatch, tmp_path):
    first = _supplier_item("sonara", currency="MXN", internal_id="sonara:first")
    second = _supplier_item("sonara", currency="MXN", internal_id="sonara:second")
    first.update(
        sku="",
        code_status="needs_review",
        image_url="https://assets.example.test/first.png",
        warnings=["Codigo por verificar"],
    )
    second.update(
        sku="",
        code_status="needs_review",
        image_url="https://assets.example.test/second.png",
        warnings=["Codigo por verificar"],
    )
    catalog = {
        "supplier": "sonara",
        "source_hash": "d" * 64,
        "generated_at": "2026-07-19T00:00:00+00:00",
        "items": [first, second],
    }
    payload = build_mixed_catalog_cart_payload(
        [
            {"catalog": "sonara", "internal_id": "sonara:first", "quantity": "1"},
            {"catalog": "sonara", "internal_id": "sonara:second", "quantity": "1"},
        ],
        catalogs={"sonara": catalog},
        rate_rows=[],
        quote_currency="MXN",
        commercial_discount_percent="40",
        today=date(2026, 7, 19),
    )
    bodies = {
        first["image_url"]: _png_bytes("red"),
        second["image_url"]: _png_bytes("green"),
    }

    class Response:
        headers = {"content-type": "image/png"}

        def __init__(self, data):
            self.data = data

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def read(self, size):
            return self.data[:size]

    class Opener:
        def open(self, request, timeout):
            assert timeout == 18
            return Response(bodies[request.full_url])

    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", "https://assets.example.test")
    monkeypatch.setattr(catalog_cart, "_resolve_public_host", lambda host: None)
    monkeypatch.setattr(catalog_cart, "_validate_connected_peer", lambda response: None)
    monkeypatch.setattr(catalog_cart.urllib.request, "build_opener", lambda *handlers: Opener())
    output = create_mixed_catalog_quotation_workbook(
        payload, tmp_path / "collision.xlsx", image_dir=tmp_path / "images"
    )
    downloaded = sorted(path.name for path in (tmp_path / "images").iterdir())
    assert any(name.startswith("sonara-9-") for name in downloaded)
    assert any(name.startswith("sonara-10-") for name in downloaded)
    with ZipFile(output) as archive:
        media = sorted(name for name in archive.namelist() if name.startswith("xl/media/"))
        hashes = {hashlib.sha256(archive.read(name)).hexdigest() for name in media}
    assert len(media) == 2
    assert hashes == {hashlib.sha256(data).hexdigest() for data in bodies.values()}
    wb = load_workbook(output)
    anchored_hashes = {
        image.anchor._from.row + 1: hashlib.sha256(image._data()).hexdigest()
        for image in wb["Quotation"]._images
    }
    assert anchored_hashes == {
        9: hashlib.sha256(bodies[first["image_url"]]).hexdigest(),
        10: hashlib.sha256(bodies[second["image_url"]]).hexdigest(),
    }
    wb.close()


def test_mixed_workbook_downloads_each_unique_image_url_once(monkeypatch, tmp_path):
    payload = frozen_payload()
    alma_group = payload["groups"][2]
    first = alma_group["items"][0]
    first.update(
        image_url="https://assets.example.test/shared.png",
        image_kind="official",
    )
    second = deepcopy(first)
    second.update(
        canonical_key='alma:["alma:desk-2","",[]]',
        name="Segundo producto ALMA",
        source_reference="alma:source:second",
    )
    alma_group["items"].append(second)
    payload["sections"][2]["item_keys"].append(second["canonical_key"])
    payload["item_count"] += 1
    calls = []
    image_path = tmp_path / "shared.png"
    image_path.write_bytes(_png_bytes("purple"))

    def fake_download(url, image_dir, code, source_type, destination_key=None):
        calls.append((url, source_type, destination_key))
        return image_path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)
    output = create_mixed_catalog_quotation_workbook(payload, tmp_path / "cached.xlsx")
    wb = load_workbook(output)
    assert len(calls) == 1
    assert calls[0][:2] == (
        "https://assets.example.test/shared.png",
        "supplier_cart",
    )
    assert sorted(image.anchor._from.row + 1 for image in wb["Quotation"]._images) == [13, 14]
    wb.close()


def test_catalog_downloader_validates_real_png_before_writing(monkeypatch, tmp_path):
    url = "https://www.offiho.com/valid.png"
    body = _png_bytes("blue")
    _stub_catalog_image_transport(monkeypatch, {url: (body, "image/png")})

    result = catalog_cart._download_catalog_image(
        url, tmp_path, "VALID", "offiho_cart"
    )

    assert result == tmp_path / "VALID.png"
    assert result.read_bytes() == body


def test_catalog_downloader_reads_exact_published_dev_asset_without_network(
    monkeypatch, tmp_path
):
    body = _png_bytes("blue")
    object_name = f"{hashlib.sha256(body).hexdigest()}.png"
    asset_dir = tmp_path / "catalog-assets"
    asset_dir.mkdir()
    (asset_dir / object_name).write_bytes(body)
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    monkeypatch.setenv("MOBILITI_DEV_MODE", "1")
    monkeypatch.setenv("MOBILITI_DEV_PUBLIC_BASE_URL", "http://127.0.0.1:8000")
    monkeypatch.setenv("MOBILITI_DEV_STORE_DIR", str(tmp_path))
    monkeypatch.setattr(
        catalog_cart.urllib.request,
        "build_opener",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("network used")),
    )

    result = catalog_cart._download_catalog_image(
        f"http://127.0.0.1:8000/dev/catalog-assets/{object_name}",
        output_dir,
        "LOCAL",
        "supplier_cart",
    )

    assert result == output_dir / "LOCAL.png"
    assert result.read_bytes() == body


def test_catalog_downloader_never_writes_corrupt_image_body(monkeypatch, tmp_path):
    url = "https://www.offiho.com/corrupt.png"
    _stub_catalog_image_transport(
        monkeypatch, {url: (b"not-an-image", "image/png")}
    )

    result = catalog_cart._download_catalog_image(
        url, tmp_path, "CORRUPT", "offiho_cart"
    )

    assert result is None
    assert list(tmp_path.iterdir()) == []


def test_catalog_downloader_rejects_image_dimensions_over_catalog_limit(monkeypatch, tmp_path):
    url = "https://www.offiho.com/oversized.png"
    _stub_catalog_image_transport(
        monkeypatch, {url: (_png_bytes("orange"), "image/png")}
    )
    monkeypatch.setattr(catalog_cart, "MAX_CATALOG_IMAGE_PIXELS", 3)

    result = catalog_cart._download_catalog_image(
        url, tmp_path, "OVERSIZED", "offiho_cart"
    )

    assert result is None
    assert list(tmp_path.iterdir()) == []


def test_catalog_image_pixel_limit_is_exactly_forty_million():
    assert catalog_cart.MAX_CATALOG_IMAGE_PIXELS == 40_000_000


def test_catalog_downloader_rejects_mime_decoded_format_mismatch(monkeypatch, tmp_path):
    url = "https://www.offiho.com/mismatch.jpg"
    _stub_catalog_image_transport(
        monkeypatch, {url: (_png_bytes("green"), "image/jpeg")}
    )

    result = catalog_cart._download_catalog_image(
        url, tmp_path, "MISMATCH", "offiho_cart"
    )

    assert result is None
    assert list(tmp_path.iterdir()) == []


def test_mixed_workbook_persists_exact_reference_and_source_hash_without_mutation(tmp_path):
    payload = frozen_payload()
    submitted = deepcopy(payload)
    output = create_mixed_catalog_quotation_workbook(payload, tmp_path / "source-audit.xlsx")
    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]

    assert ws["Q9"].value == payload["groups"][0]["items"][0]["source_reference"]
    assert f"Hash fuente: {'a' * 64}" in ws["D9"].value
    assert ws["Q11"].value == payload["groups"][1]["items"][0]["source_reference"]
    assert f"Hash fuente: {'b' * 64}" in ws["D11"].value
    assert payload == submitted
    wb.close()


def test_legacy_same_url_uses_each_code_filename_and_download(monkeypatch, tmp_path):
    body = _png_bytes("teal")
    calls = []

    def fake_download(url, image_dir, code, source_type):
        calls.append((url, code, source_type))
        path = image_dir / f"{code}.png"
        path.write_bytes(body)
        return path

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", fake_download)
    shared_url = "https://www.offiho.com/shared.png"
    payload = {
        "source_type": "offiho_cart",
        "items": [
            {
                "code": code,
                "name": f"Producto {code}",
                "unit": "PZA",
                "quantity": "1",
                "unit_price": "10",
                "image_url": shared_url,
                "warnings": [],
            }
            for code in ("CODE-A", "CODE-B")
        ],
    }
    output = create_catalog_quotation_workbook(
        payload,
        tmp_path / "legacy-shared.xlsx",
        source_type="offiho_cart",
        category_label="Offiho",
        image_dir=tmp_path / "images",
    )
    wb = load_workbook(output)
    ws = wb["Quotation"]

    assert calls == [
        (shared_url, "CODE-A", "offiho_cart"),
        (shared_url, "CODE-B", "offiho_cart"),
    ]
    assert sorted(path.name for path in (tmp_path / "images").iterdir()) == [
        "CODE-A.png",
        "CODE-B.png",
    ]
    assert sorted(image.anchor._from.row + 1 for image in ws._images) == [9, 10]
    assert [ws.cell(9, column).value for column in range(1, 12)] == [
        1,
        "Producto CODE-A",
        None,
        "Clave: CODE-A",
        "PZA",
        None,
        1,
        None,
        None,
        10,
        None,
    ]
    assert [ws.cell(10, column).value for column in range(1, 12)] == [
        2,
        "Producto CODE-B",
        None,
        "Clave: CODE-B",
        "PZA",
        None,
        1,
        None,
        None,
        10,
        None,
    ]
    wb.close()


def test_malformed_supplier_allowlist_is_isolated_before_valid_family_image(monkeypatch, tmp_path):
    valid_url = "https://www.offiho.com/valid-after-invalid.png"
    _stub_catalog_image_transport(
        monkeypatch, {valid_url: (_png_bytes("purple"), "image/png")}
    )
    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", "https://[")
    monkeypatch.delenv("SUPABASE_URL", raising=False)
    wb = Workbook()
    ws = wb.active
    image_dir = tmp_path / "images"
    image_dir.mkdir()

    assert catalog_cart._download_catalog_image(
        "https://assets.example.test/invalid.png",
        image_dir,
        "INVALID",
        "supplier_cart",
        destination_key="supplier-invalid",
    ) is None
    catalog_cart._add_catalog_image(
        ws,
        10,
        valid_url,
        image_dir,
        "VALID",
        "offiho_cart",
        destination_key="offiho-valid",
    )
    output = tmp_path / "isolated.xlsx"
    wb.save(output)
    wb.close()
    saved = load_workbook(output)

    assert [image.anchor._from.row + 1 for image in saved.active._images] == [10]
    assert not (image_dir / "supplier-invalid.png").exists()
    assert (image_dir / "offiho-valid.png").exists()
    saved.close()


@pytest.mark.parametrize(
    "configured_url",
    (
        "https://[",
        "https://bad host.example/path",
        "https://assets.example.test:not-a-port/path",
    ),
)
def test_supplier_allowlist_fails_closed_for_malformed_config(monkeypatch, configured_url):
    monkeypatch.setenv("CATALOG_ASSET_PUBLIC_BASE_URL", configured_url)
    monkeypatch.delenv("SUPABASE_URL", raising=False)
    assert catalog_cart._allowed_image_hosts("supplier_cart") == frozenset()


def test_corrupt_image_isolated_before_valid_image(monkeypatch, tmp_path):
    corrupt_url = "https://www.offiho.com/first-corrupt.png"
    valid_url = "https://www.offiho.com/second-valid.png"
    _stub_catalog_image_transport(
        monkeypatch,
        {
            corrupt_url: (b"corrupt", "image/png"),
            valid_url: (_png_bytes("yellow"), "image/png"),
        },
    )
    wb = Workbook()
    ws = wb.active
    image_dir = tmp_path / "images"
    image_dir.mkdir()

    catalog_cart._add_catalog_image(
        ws, 9, corrupt_url, image_dir, "BAD", "offiho_cart", destination_key="bad"
    )
    catalog_cart._add_catalog_image(
        ws, 10, valid_url, image_dir, "GOOD", "offiho_cart", destination_key="good"
    )
    output = tmp_path / "corrupt-then-valid.xlsx"
    wb.save(output)
    wb.close()
    saved = load_workbook(output)

    assert [image.anchor._from.row + 1 for image in saved.active._images] == [10]
    assert not (image_dir / "bad.png").exists()
    assert (image_dir / "good.png").exists()
    saved.close()


def test_mixed_workbook_rejects_final_description_over_excel_limit_before_paths(tmp_path):
    payload = frozen_payload()
    payload["groups"][2]["items"][0]["warnings"] = [
        f"Aviso unico {index}: " + (chr(65 + index % 26) * 700)
        for index in range(50)
    ]
    submitted = deepcopy(payload)
    output = tmp_path / "output" / "oversized-description.xlsx"
    image_dir = tmp_path / "images"

    with pytest.raises(ValueError, match="limite de Excel"):
        create_mixed_catalog_quotation_workbook(
            payload, output, image_dir=image_dir
        )

    assert not output.parent.exists()
    assert not image_dir.exists()
    assert payload == submitted


@pytest.mark.parametrize(
    "attributes",
    (
        {"color": "Rojo\x00oculto"},
        {"dimensions": "10\x0b20 cm"},
        {"product_notes": ["Visible", "Oculto\x1f"]},
        {"nested": {"rows": [{"value": "Control\x0c"}]}},
        {"clave\x07ilegal": "valor"},
    ),
)
def test_mixed_workbook_rejects_recursive_illegal_attribute_controls_before_paths(
    tmp_path, attributes
):
    payload = frozen_payload()
    payload["groups"][2]["items"][0]["attributes"] = attributes
    submitted = deepcopy(payload)
    output = tmp_path / "output" / "illegal-attributes.xlsx"
    image_dir = tmp_path / "images"

    with pytest.raises(ValueError, match="caracteres de control"):
        create_mixed_catalog_quotation_workbook(
            payload, output, image_dir=image_dir
        )

    assert not output.parent.exists()
    assert not image_dir.exists()
    assert payload == submitted
