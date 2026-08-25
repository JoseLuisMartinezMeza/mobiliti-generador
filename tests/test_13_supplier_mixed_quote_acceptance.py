from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import importlib.util
import json
import os
from pathlib import Path
import sys
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

import test_mixed_catalog_quote_e2e as mixed_e2e
from mobiliti_saas.quote_engine import catalog_cart
from mobiliti_saas.quote_engine.mixed_catalog import MIXED_CATALOG_ORDER
from mobiliti_saas.quote_engine.quotation_sheets import (
    QUOTATION_DATA_HEADERS,
    quotation_data_rows,
)
from mobiliti_saas.worker.catalog_sync.importers.labenze import (
    build_labenze_snapshot_with_assets,
)
from mobiliti_saas.worker.catalog_sync.importers.requiez import (
    build_requiez_snapshot_with_assets,
)


ROOT = Path(__file__).resolve().parents[1]
WORKER_DIR = ROOT / "mobiliti_saas" / "worker"
WORKER_TEMPLATE = (
    WORKER_DIR / "templates" / "Formato Cotizacion 2026 Oficial.xlsx"
)
GENERIC_SUPPLIERS = MIXED_CATALOG_ORDER[2:]
PDF_ENVIRONMENTS = {
    "labenze": "MOBILITI_E2E_LABENZE_PDF",
    "requiez": "MOBILITI_E2E_REQUIEZ_PDF",
}
OFFICIAL_PDF_SHA256 = {
    "labenze": "c4fc2d2152b5e854f7c36c9106c71cd21853abb50efcde96ba2566cb72f1d6f3",
    "requiez": (
        "7f3281d1965c67a234bac551128000670"
        "19ad471f835de59ff758e759eca56ba"
    ),
}
EXPECTED_TEMPLATE_REF_DEBT = (
    "Cotizacion!P21: =#REF!-#REF!",
    "Cotizacion!L121: =IF(ROUND(#REF!,2)=ROUND('Estrategia Comercial '!B70,2),"
    "\"ESTRATEGIA CORRECTA\",\"REVISA TU ESTRATEGIA\")",
    "SPEC-GUIDE ESTRUCTURAS!E474: ='[4]COSTO CDMX,QRO,GDL '!#REF!",
)
SUPPLIER_LABELS = {
    "tarkett": "Tarkett",
    "offiho": "Offiho",
    "cr-global": "CR Global",
    "sonara": "Sonara",
    "sunon": "Sunon",
    "alma": "ALMA",
    "lumbro": "Lumbro",
    "jome": "JOME",
    "lauco": "Lauco",
    "idelika": "IDÉLIKA",
    "conceptos": "Conceptos",
    "labenze": "Labenze",
    "requiez": "Requiez",
}
SYNTHETIC_CODES = {
    "cr-global": "CR-13-E2E",
    "sonara": "SON-13-E2E",
    "sunon": "SUN-13-E2E",
    "alma": "ALMA-13-E2E",
    "lumbro": "LUMBRO-13-E2E",
    "jome": "JOME-13-E2E",
    "lauco": "LAUCO-13-E2E",
    "idelika": "IDELIKA-13-E2E",
    "conceptos": "CONCEPTOS-13-E2E",
}


@dataclass(frozen=True)
class SourceDocument:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path


@dataclass(frozen=True)
class OfficialSelection:
    supplier: str
    snapshot: dict
    item: dict
    asset_data: bytes
    asset_sha256: str
    object_name: str
    match_status: str


@pytest.fixture()
def isolated_13_supplier_runtime(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SUPABASE_URL", "https://supabase.example.test")
    monkeypatch.setenv("SUPABASE_SERVICE_KEY", "test-key")
    monkeypatch.setenv("JWT_SECRET_KEY", "test-secret-32-chars-long!!!!!")
    monkeypatch.setenv("CATALOG_ENABLED_SUPPLIERS", ",".join(GENERIC_SUPPLIERS))
    monkeypatch.syspath_prepend(str(WORKER_DIR))
    modules_before = set(sys.modules)
    suffix = uuid.uuid4().hex

    def load_module(name: str, path: Path):
        spec = importlib.util.spec_from_file_location(name, path)
        assert spec is not None and spec.loader is not None
        module = importlib.util.module_from_spec(spec)
        sys.modules[name] = module
        spec.loader.exec_module(module)
        return module

    api_index = load_module(
        f"mixed_quote_13_api_{suffix}",
        ROOT / "mobiliti_saas" / "api" / "index.py",
    )
    quote_worker = load_module(
        f"mixed_quote_13_worker_{suffix}",
        WORKER_DIR / "quote_worker.py",
    )
    assert api_index._enabled_catalog_suppliers() == GENERIC_SUPPLIERS
    try:
        yield api_index, quote_worker
    finally:
        for name in set(sys.modules) - modules_before:
            module_path = getattr(sys.modules.get(name), "__file__", None)
            if name.endswith(suffix) or (
                module_path
                and Path(module_path).resolve().is_relative_to(WORKER_DIR.resolve())
            ):
                sys.modules.pop(name, None)


def _required_pdf(supplier: str) -> Path:
    environment = PDF_ENVIRONMENTS[supplier]
    configured = str(os.environ.get(environment) or "").strip()
    if not configured:
        pytest.skip(f"Configura {environment} con el PDF oficial {supplier}")
    path = Path(configured).expanduser().resolve()
    if not path.is_file():
        pytest.skip(f"PDF oficial {supplier} no disponible: {path}")
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    assert digest == OFFICIAL_PDF_SHA256[supplier]
    return path


def _source_document(supplier: str, path: Path) -> SourceDocument:
    logical_path = {
        "labenze": "LABENZE/LP Labenze B26.pdf",
        "requiez": "REQUIEZ/Lista de precios A-26.pdf",
    }[supplier]
    return SourceDocument(
        path=logical_path,
        kind="price_list",
        brand=None,
        sha256=OFFICIAL_PDF_SHA256[supplier],
        mime_type="application/pdf",
        local_path=path,
    )


def _official_selection(supplier: str, path: Path) -> OfficialSelection:
    builders = {
        "labenze": build_labenze_snapshot_with_assets,
        "requiez": build_requiez_snapshot_with_assets,
    }
    build = builders[supplier](
        (_source_document(supplier, path),),
        synced_at=datetime(2026, 8, 18, 12, tzinfo=timezone.utc),
    )
    items = {item["internal_id"]: item for item in build.snapshot["items"]}
    preferred_skus = {
        "labenze": ("155-20400", "160-S1170"),
        "requiez": ("RP-1400/GC", "RE-1450/GC"),
    }[supplier]
    candidates = []
    for binding in build.bindings:
        item = items[binding.internal_id]
        folded_name = str(item["name"]).casefold()
        if (
            binding.match_status == "exact_pdf"
            and item["code_status"] == "verified"
            and item["sku"]
            and item["price_net"]
            and binding.asset_sha256 in build.assets_by_sha256
            and not any(
                token in folded_name
                for token in ("accesorio", "cabecera", "kit de base", "base tapiz")
            )
        ):
            preference = (
                preferred_skus.index(item["sku"])
                if item["sku"] in preferred_skus
                else len(preferred_skus)
            )
            candidates.append(
                (
                    preference,
                    bool(item["base_price_options"]),
                    binding.internal_id,
                    binding,
                )
            )
    assert candidates, f"{supplier} debe ofrecer al menos una unión exact_pdf verificable"
    _preference, _has_options, internal_id, binding = min(candidates)
    item = deepcopy(items[internal_id])
    asset = build.assets_by_sha256[binding.asset_sha256]
    assert asset.sha256 == hashlib.sha256(asset.data).hexdigest()
    assert asset.media_type == "image/png"
    assert item["image_kind"] == "official"
    assert item["attributes"]["image_match"]["status"] == "exact_pdf"
    assert binding.source_references
    assert all(
        reference["file_id"] == OFFICIAL_PDF_SHA256[supplier]
        and isinstance(reference["sheet_or_page"], int)
        and len(reference["cell_or_bbox"]) == 4
        for reference in binding.source_references
    )
    return OfficialSelection(
        supplier=supplier,
        snapshot={**deepcopy(build.snapshot), "items": [item]},
        item=item,
        asset_data=asset.data,
        asset_sha256=asset.sha256,
        object_name=binding.object_name,
        match_status=binding.match_status,
    )


def _synthetic_item(supplier: str, index: int) -> dict:
    code = SYNTHETIC_CODES[supplier]
    source_hash = hashlib.sha256(
        f"acceptance-13:{supplier}".encode("utf-8")
    ).hexdigest()
    source_reference = json.dumps(
        [
            {
                "file_id": source_hash,
                "sheet_or_page": "Aceptacion sintetica",
                "cell_or_bbox": f"A{index + 1}",
            }
        ],
        ensure_ascii=True,
        separators=(",", ":"),
        sort_keys=True,
    )
    return {
        "internal_id": f"{supplier}:acceptance-13",
        "supplier": supplier,
        "product_key": "acceptance-13",
        "sku": code,
        "code_status": "verified",
        "brand": SUPPLIER_LABELS[supplier],
        "collection": "Aceptación E2E auditable",
        "name": f"Producto aceptación {SUPPLIER_LABELS[supplier]}",
        "description": f"Snapshot sintético auditable de {SUPPLIER_LABELS[supplier]}",
        "unit": "PZA",
        "availability_type": "stocked",
        "stock": "20.000000",
        "lead_time": "Entrega inmediata",
        "base_price_options": [],
        "add_on_options": [],
        "base_currency": "USD" if supplier in {"sunon", "alma"} else "MXN",
        "price_net": f"{700 + index * 25}.000000",
        "tax_rate": "0.160000",
        "attributes": {"acceptance_fixture": True},
        "image_url": "",
        "image_kind": "placeholder",
        "product_url": "",
        "warnings": [],
        "source_reference": source_reference,
    }


def _catalogs_and_official_assets(api_index, selections: dict[str, OfficialSelection]):
    base_catalogs = mixed_e2e.authoritative_catalogs()
    catalogs = {
        "tarkett": base_catalogs["tarkett"],
        "offiho": base_catalogs["offiho"],
    }
    for index, supplier in enumerate(SYNTHETIC_CODES, start=1):
        item = _synthetic_item(supplier, index)
        catalogs[supplier] = {
            "supplier": supplier,
            "source_hash": hashlib.sha256(
                f"acceptance-13:{supplier}".encode("utf-8")
            ).hexdigest(),
            "generated_at": "2026-08-18T12:00:00+00:00",
            "items": [item],
        }
    official_urls = {}
    for supplier, selection in selections.items():
        hydrated = api_index._hydrate_catalog_asset_urls(selection.snapshot)
        catalogs[supplier] = hydrated
        official_urls[supplier] = hydrated["items"][0]["image_url"]
        assert hydrated["items"][0]["image_kind"] == "official"
        assert hydrated["source_hash"] == selection.snapshot["source_hash"]
    assert tuple(catalogs) == MIXED_CATALOG_ORDER
    return catalogs, official_urls


def _browser_rows(catalogs: dict) -> list[dict]:
    rows = [
        {"line_id": "line-tarkett", "catalog": "tarkett", "code": "TARK-E2E", "quantity": "1"},
        {
            "line_id": "line-offiho",
            "catalog": "offiho",
            "inventory_key": "OFF-E2E NEGRO",
            "quantity": "1",
        },
    ]
    for supplier in GENERIC_SUPPLIERS:
        item = catalogs[supplier]["items"][0]
        row = {
            "line_id": f"line-{supplier}",
            "catalog": supplier,
            "internal_id": item["internal_id"],
            "quantity": "1",
        }
        if item["base_price_options"]:
            row["base_option_id"] = item["base_price_options"][0]["id"]
        rows.append(row)
    assert [row["catalog"] for row in rows] == list(MIXED_CATALOG_ORDER)
    return rows


def _formula_text(value: object) -> str:
    return str(getattr(value, "text", value) or "")


def _broken_ref_formulas(workbook) -> list[str]:
    formulas = 0
    broken = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = _formula_text(cell.value)
                if cell.data_type == "f" or value.startswith("="):
                    formulas += 1
                    if "#REF!" in value.upper():
                        broken.append(f"{worksheet.title}!{cell.coordinate}: {value}")
    assert formulas > 0
    return broken


def _assert_operational_formulas_without_ref(workbook) -> None:
    """El legado fuera del rango operativo se audita aparte, no se propaga."""

    for sheet_name in ("Quotation", "Quotation_Data", "Mobiliti"):
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                value = _formula_text(cell.value)
                if cell.data_type == "f" or value.startswith("="):
                    assert "#REF!" not in value.upper(), (
                        f"Referencia rota operativa en {sheet_name}!{cell.coordinate}"
                    )
    quotation = workbook["Cotizacion"]
    for row in quotation.iter_rows(min_col=1, max_col=10):
        for cell in row:
            value = _formula_text(cell.value)
            if cell.data_type == "f" or value.startswith("="):
                assert "#REF!" not in value.upper(), (
                    f"Referencia rota en producto/total Cotizacion!{cell.coordinate}"
                )


def _image_hashes(worksheet) -> set[str]:
    return {
        hashlib.sha256(image._data()).hexdigest()
        for image in worksheet._images
    }


def _header_columns(worksheet) -> dict[str, int]:
    return {
        str(worksheet.cell(7, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(7, column).value is not None
    }


def test_post_mixed_quote_runs_python_worker_for_all_13_suppliers(
    isolated_13_supplier_runtime,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    """Aceptación local completa; los dos PDF y sus recortes nunca salen de tmp."""

    assert WORKER_TEMPLATE.is_file()
    template_workbook = load_workbook(WORKER_TEMPLATE, data_only=False)
    try:
        assert _broken_ref_formulas(template_workbook) == list(
            EXPECTED_TEMPLATE_REF_DEBT
        )
    finally:
        template_workbook.close()
    pdf_paths = {supplier: _required_pdf(supplier) for supplier in PDF_ENVIRONMENTS}
    selections = {
        supplier: _official_selection(supplier, path)
        for supplier, path in pdf_paths.items()
    }
    assert all(
        not path.is_relative_to(ROOT)
        for path in pdf_paths.values()
    )

    api_index, quote_worker = isolated_13_supplier_runtime
    catalogs, official_urls = _catalogs_and_official_assets(api_index, selections)
    rows = _browser_rows(catalogs)
    api_state = mixed_e2e.install_api_boundary(
        monkeypatch,
        api_index,
        catalogs,
    )

    assets_dir = tmp_path / "official-pdf-assets"
    assets_dir.mkdir()
    image_paths = {
        "https://media.tarkett-image.com/e2e-tarkett.png":
            mixed_e2e._make_png(tmp_path / "tarkett-synthetic.png", (20, 70, 120)),
        "https://offiho.com.mx/e2e-offiho.png":
            mixed_e2e._make_png(tmp_path / "offiho-synthetic.png", (120, 70, 20)),
    }
    for supplier, selection in selections.items():
        path = assets_dir / selection.object_name
        path.write_bytes(selection.asset_data)
        assert path.is_relative_to(tmp_path)
        assert hashlib.sha256(path.read_bytes()).hexdigest() == selection.asset_sha256
        image_paths[official_urls[supplier]] = path

    image_calls = []

    def local_catalog_image(url, image_dir, code, source_type, destination_key=None):
        image_calls.append((source_type, str(url), destination_key))
        return image_paths[str(url)]

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", local_catalog_image)

    body = {
        "items": rows,
        "quote_currency": "MXN",
        "descuento": "40",
        "proyecto": "Aceptación mixta 13 proveedores",
        "cliente": "Cliente aceptación",
        "correo": "aceptacion@example.test",
        "telefono": "3330000000",
        "direccion": "Guadalajara",
        "razon_social": "Cliente Aceptación SA de CV",
        "image_provider": "pillow",
        "template": WORKER_TEMPLATE.name,
    }
    request_bytes = json.dumps(
        body,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    assert len(request_bytes) < int(3.5 * 1024 * 1024)

    with TestClient(api_index.app) as api_client:
        response = api_client.post(
            "/catalogs/mixed-quote",
            headers=mixed_e2e.auth_headers(api_index),
            json=body,
        )
    assert response.status_code == 200, response.text
    queued_job = response.json()["job"]
    assert api_state["events"] == [
        "create_job", "reserve_mixed", "upload", "queue", "wake"
    ]
    assert len(api_state["uploads"]) == 1
    input_path, input_bytes, content_type = api_state["uploads"][0]
    assert input_path == queued_job["input_path"]
    assert content_type == "application/json"
    assert len(input_bytes) < int(3.5 * 1024 * 1024)

    payload = json.loads(input_bytes)
    assert payload["source_type"] == "mixed_catalog_cart"
    assert payload["item_count"] == 13
    assert [group["catalog"] for group in payload["groups"]] == list(
        MIXED_CATALOG_ORDER
    )
    assert all(len(group["items"]) == 1 for group in payload["groups"])
    assert [group["items"][0]["supplier"] for group in payload["groups"]] == [
        SUPPLIER_LABELS[supplier] for supplier in MIXED_CATALOG_ORDER
    ]
    assert [group["catalog_source_hash"] for group in payload["groups"][-2:]] == [
        selections["labenze"].snapshot["source_hash"],
        selections["requiez"].snapshot["source_hash"],
    ]

    payload_lines = {
        group["catalog"]: group["items"][0]
        for group in payload["groups"]
    }
    expected_codes = {
        "tarkett": "TARK-E2E",
        "offiho": "OFF-E2E",
        **SYNTHETIC_CODES,
        **{
            supplier: selection.item["sku"]
            for supplier, selection in selections.items()
        },
    }
    assert {supplier: line["code"] for supplier, line in payload_lines.items()} == (
        expected_codes
    )
    for supplier, selection in selections.items():
        line = payload_lines[supplier]
        assert line["source_reference"] == selection.item["source_reference"]
        assert line["image_url"] == official_urls[supplier]
        assert line["image_kind"] == "official"
        assert line["product_url"].startswith("https://")
        assert "#page=" in line["product_url"]

    worker_client = mixed_e2e.EndToEndWorkerClient(
        queued_job,
        {input_path: input_bytes},
    )
    expected_handoff = quotation_data_rows(payload)
    real_run_generator = quote_worker._run_generator

    def audited_run_generator(job, generator_input, local_output):
        assert isinstance(generator_input, quote_worker.PreparedGeneratorInput)
        assert generator_input.parser_source.name == "quotation_from_mixed_catalog.xlsx"
        assert generator_input.parser_source.is_file()
        assert generator_input.original_quotation is None
        assert generator_input.quotation_data == expected_handoff
        converted = load_workbook(generator_input.parser_source, data_only=False)
        try:
            assert converted.sheetnames == ["Quotation"]
            official_hashes = {
                selection.asset_sha256 for selection in selections.values()
            }
            assert official_hashes <= _image_hashes(converted["Quotation"])
        finally:
            converted.close()
        worker_client.record_event("converter")
        worker_client.record_event("generator")
        return real_run_generator(job, generator_input, local_output)

    monkeypatch.setattr(quote_worker, "_run_generator", audited_run_generator)
    monkeypatch.setattr(quote_worker, "_template_path", lambda: str(WORKER_TEMPLATE))
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")

    completed_rows = quote_worker.process_job(worker_client, queued_job)
    assert completed_rows and completed_rows[0]["status"] == "completed"
    assert worker_client.events == list(worker_client.EXPECTED_EVENTS)
    assert worker_client.failed_updates == []
    assert len(worker_client.uploads) == 1
    _uploaded_path, output_bytes = worker_client.uploads[0]
    assert output_bytes.startswith(b"PK")
    assert len(output_bytes) <= 60 * 1024 * 1024
    assert len(output_bytes) < 150 * 1024 * 1024

    output_path = tmp_path / "cotizacion_mixta_13_proveedores.xlsx"
    output_path.write_bytes(output_bytes)
    output_hash = hashlib.sha256(output_bytes).hexdigest()
    assert output_path.is_file() and output_path.is_relative_to(tmp_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert {"Cotizacion", "Mobiliti", "Quotation", "Quotation_Data"} <= set(
            workbook.sheetnames
        )
        broken_ref_formulas = _broken_ref_formulas(workbook)
        assert broken_ref_formulas == list(EXPECTED_TEMPLATE_REF_DEBT)
        _assert_operational_formulas_without_ref(workbook)
        quotation = workbook["Quotation"]
        audit = workbook["Quotation_Data"]
        headers = _header_columns(quotation)
        assert {
            "Description", "Source Reference", "Canonical Key"
        } <= set(headers)
        product_rows = [
            row
            for row in range(8, quotation.max_row + 1)
            if isinstance(quotation.cell(row, 1).value, (int, float))
        ]
        descriptions = {
            row: str(quotation.cell(row, headers["Description"]).value or "")
            for row in product_rows
        }
        main_rows_by_supplier = {}
        for supplier, code in expected_codes.items():
            matches = [row for row, description in descriptions.items() if code in description]
            assert len(matches) == 1, f"Código principal {supplier}/{code} ambiguo"
            row = matches[0]
            main_rows_by_supplier[supplier] = row
            assert payload_lines[supplier]["source_reference"] in descriptions[row]
        assert len(set(main_rows_by_supplier.values())) == 13
        derived_rows = [
            row for row in product_rows if row not in set(main_rows_by_supplier.values())
        ]
        assert len(derived_rows) <= 3
        assert {
            quotation.cell(row, 2).value for row in derived_rows
        } <= {"LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"}

        frozen_sources = [
            quotation.cell(row, headers["Source Reference"]).value
            for row in product_rows
            if quotation.cell(row, headers["Source Reference"]).value
        ]
        assert frozen_sources == [
            payload_lines[supplier]["source_reference"]
            for supplier in MIXED_CATALOG_ORDER
        ]

        assert audit.sheet_state == "veryHidden"
        assert tuple(
            audit.cell(1, column).value
            for column in range(1, len(QUOTATION_DATA_HEADERS) + 1)
        ) == QUOTATION_DATA_HEADERS
        audit_records = [
            {
                field: audit.cell(row, column).value
                for column, field in enumerate(QUOTATION_DATA_HEADERS, start=1)
            }
            for row in range(2, audit.max_row + 1)
        ]
        main_line_ids = {line["line_id"] for line in payload_lines.values()}
        assert sum(record["item_key"] in main_line_ids for record in audit_records) == 13
        expected_audit = {row.item_key: row for row in expected_handoff}
        assert {
            record["item_key"]: record["provider"]
            for record in audit_records
            if record["item_key"] in main_line_ids
        } == {
            line_id: expected_audit[line_id].provider
            for line_id in main_line_ids
        }
        derived = [
            record for record in audit_records if record["item_key"] not in main_line_ids
        ]
        assert len(derived) <= 3
        assert all(record["origin"] == "lumbro" for record in derived)

        official_hashes = {
            selection.asset_sha256 for selection in selections.values()
        }
        quotation_images = [
            (image.anchor._from.row + 1, image._data())
            for image in quotation._images
        ]
        assert official_hashes <= {
            hashlib.sha256(data).hexdigest() for _row, data in quotation_images
        }
        for supplier, selection in selections.items():
            source_row = main_rows_by_supplier[supplier]
            assert selection.asset_sha256 in {
                hashlib.sha256(data).hexdigest()
                for row, data in quotation_images
                if row == source_row
            }

        cotizacion = workbook["Cotizacion"]
        cotizacion_image_rows = {
            image.anchor._from.row + 1
            for image in cotizacion._images
            if image.anchor._from.col + 1 == 2
        }
        for supplier in selections:
            source_row = main_rows_by_supplier[supplier]
            target_rows = [
                row
                for row in range(1, cotizacion.max_row + 1)
                if cotizacion.cell(row, 3).value == f"=Quotation!D{source_row}"
            ]
            assert len(target_rows) == 1
            assert target_rows[0] in cotizacion_image_rows
    finally:
        workbook.close()

    called_urls = [url for _source_type, url, _key in image_calls]
    assert all(called_urls.count(url) == 1 for url in official_urls.values())
    assert all(key for _source, _url, key in image_calls)
    print(
        "ACCEPTANCE_13_SUPPLIERS "
        f"path={output_path} input_bytes={len(input_bytes)} "
        f"output_bytes={len(output_bytes)} sha256={output_hash} "
        f"labenze_sku={selections['labenze'].item['sku']} "
        f"requiez_sku={selections['requiez'].item['sku']} "
        f"template_ref_debt={len(EXPECTED_TEMPLATE_REF_DEBT)}"
    )
