from __future__ import annotations

import hashlib
import importlib
from pathlib import Path
from zipfile import ZipFile

from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils.units import pixels_to_EMU


def _png(path: Path, color: tuple[int, int, int] = (31, 96, 145)) -> None:
    Image.new("RGB", (24, 16), color).save(path, "PNG")


def _workbook(
    path: Path,
    rows: list[tuple[int, str, str]],
    pictures: list[tuple[Path, object]],
    *,
    row_heights: dict[int, float] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SPEC Offiho"
    sheet["A8"] = "Cod."
    sheet["B8"] = "Imagen."
    sheet["C8"] = "Descripcion."
    for row, code, description in rows:
        sheet.cell(row, 1).value = code
        sheet.cell(row, 3).value = description
    for row, height in (row_heights or {}).items():
        sheet.row_dimensions[row].height = height
    for image_path, anchor in pictures:
        sheet.add_image(WorkbookImage(str(image_path)), anchor)
    workbook.save(path)
    workbook.close()


def _embedded_png(path: Path) -> bytes:
    with ZipFile(path) as package:
        media = next(name for name in package.namelist() if name.startswith("xl/media/"))
        return package.read(media)


def _duplicated_spec_cost_workbook(path: Path, image_path: Path) -> None:
    workbook = Workbook()
    spec_sheet = workbook.active
    spec_sheet.title = "SPEC Sillas"
    cost_sheet = workbook.create_sheet("Cost Sillas")
    for sheet, description in (
        (spec_sheet, "Descripcion tecnica de SPEC."),
        (cost_sheet, "Descripcion abreviada de Cost."),
    ):
        sheet["A8"] = "Cod."
        sheet["B8"] = "Imagen."
        sheet["C8"] = "Descripcion."
        sheet["A9"] = "OHE-705gris"
        sheet["C9"] = description
        sheet.add_image(WorkbookImage(str(image_path)), "B9")
    workbook.save(path)
    workbook.close()


def test_extracts_exact_code_variant_and_materializes_source_image(tmp_path):
    source_image = tmp_path / "aiko.png"
    _png(source_image)
    workbook = tmp_path / "offiho-aiko.xlsx"
    _workbook(
        workbook,
        [(9, "OHE-705gris", "Silla ejecutiva Aiko, acabado gris.")],
        [(source_image, "B9")],
    )
    embedded = _embedded_png(workbook)
    expected_hash = hashlib.sha256(embedded).hexdigest()
    assets_dir = tmp_path / "assets"
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "OHE-705 GRIS AIKO",
                "code": "OHE-705",
                "name": "AIKO",
                "variant": "GRIS",
            }
        ],
        assets_dir=assets_dir,
        base_url="/catalog-assets/offiho/spec",
        source_urls={workbook.name: "https://sharepoint.example/offiho/aiko.xlsx"},
    )

    match = matches["OHE-705 GRIS AIKO"]
    assert match["product_url"] == "https://sharepoint.example/offiho/aiko.xlsx"
    assert match["description"] == "Silla ejecutiva Aiko, acabado gris."
    assert match["match_status"] == "spec_guide_exact"
    assert match["image_sha256"] == expected_hash
    assert match["image_content_type"] == "image/png"
    assert match["image_content_length"] == len(embedded)
    assert match["image_width"] == 24
    assert match["image_height"] == 16
    assert match["source_workbook"] == workbook.name
    assert match["source_sheet"] == "SPEC Offiho"
    assert match["source_row"] == 9
    assert match["source_code"] == "OHE-705gris"
    assert match["image_url"].startswith("/catalog-assets/offiho/spec/")
    assert match["image_url"].endswith(f"-{expected_hash[:16]}.png")
    materialized = assets_dir / match["image_url"].rsplit("/", 1)[-1]
    assert materialized.read_bytes() == embedded


def test_uses_vertical_overlap_when_anchor_starts_in_previous_row(tmp_path):
    source_image = tmp_path / "fenix.png"
    _png(source_image, (18, 80, 45))
    workbook = tmp_path / "offiho-fenix.xlsx"
    shifted_anchor = TwoCellAnchor(
        _from=AnchorMarker(col=1, colOff=0, row=8, rowOff=pixels_to_EMU(19)),
        to=AnchorMarker(col=1, colOff=pixels_to_EMU(20), row=9, rowOff=pixels_to_EMU(10)),
    )
    _workbook(
        workbook,
        [
            (9, "OHE-999", "Producto de la fila anterior."),
            (10, "OHE-165negro", "Silla ejecutiva Fenix negra."),
        ],
        [(source_image, shifted_anchor)],
        row_heights={9: 15, 10: 30},
    )
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "OHE-165 NEGRO FENIX",
                "code": "OHE-165",
                "name": "FENIX",
                "variant": "NEGRO",
            }
        ],
        assets_dir=tmp_path / "assets",
        base_url="/assets/spec",
    )

    assert matches["OHE-165 NEGRO FENIX"]["source_row"] == 10
    assert matches["OHE-165 NEGRO FENIX"]["source_code"] == "OHE-165negro"
    assert matches["OHE-165 NEGRO FENIX"]["anchor_row_delta"] == 1


def test_treats_cr_code_suffix_as_exact_cromada_variant(tmp_path):
    source_image = tmp_path / "ivy.png"
    _png(source_image, (120, 120, 120))
    workbook = tmp_path / "offiho-ivy.xlsx"
    _workbook(
        workbook,
        [(9, "OHR-2800-4PCR", "Banca Ivy de cuatro plazas con base cromada.")],
        [(source_image, "B9")],
    )
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "OHR-2800-4P CR CROMADA IVY",
                "code": "OHR-2800-4P",
                "name": "CR IVY",
                "variant": "CROMADA",
            }
        ],
        assets_dir=tmp_path / "assets",
        base_url="/assets/spec",
    )

    assert matches["OHR-2800-4P CR CROMADA IVY"]["source_code"] == "OHR-2800-4PCR"


def test_required_feature_selects_cabecera_image_not_base_model(tmp_path):
    base_image = tmp_path / "crew-base.png"
    cabecera_image = tmp_path / "crew-cabecera.png"
    _png(base_image, (180, 180, 180))
    _png(cabecera_image, (22, 58, 100))
    workbook = tmp_path / "offiho-crew.xlsx"
    _workbook(
        workbook,
        [
            (9, "OHE-112gris", "Silla Crew gris."),
            (10, "OHE-112griscab", "Silla Crew gris con cabecera."),
        ],
        [(base_image, "B9"), (cabecera_image, "B10")],
    )
    with ZipFile(workbook) as package:
        media = sorted(name for name in package.namelist() if name.startswith("xl/media/"))
        expected_hash = hashlib.sha256(package.read(media[1])).hexdigest()
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "CABECERA OHE-112 GRIS",
                "code": "OHE-112",
                "name": "CABECERA",
                "variant": "GRIS",
            }
        ],
        assets_dir=tmp_path / "assets",
        base_url="/assets/spec",
    )

    match = matches["CABECERA OHE-112 GRIS"]
    assert match["source_code"] == "OHE-112griscab"
    assert match["source_row"] == 10
    assert match["image_sha256"] == expected_hash


def test_no_variant_requires_full_identity_and_rejects_generic_base_rows(tmp_path):
    exact_image = tmp_path / "ta.png"
    configured_image = tmp_path / "heron.png"
    generic_image = tmp_path / "aiko-generic.png"
    _png(exact_image, (80, 54, 32))
    _png(configured_image, (42, 76, 110))
    _png(generic_image, (155, 155, 155))
    workbook = tmp_path / "offiho-identities.xlsx"
    _workbook(
        workbook,
        [
            (9, "TA 1/Q-600", "Mesa baja TA de 600 mm."),
            (10, "OHV-124", "Silla visitante Heron con configuraciones disponibles."),
            (11, "OHE-705", "Silla Aiko disponible en gris o negro."),
        ],
        [
            (exact_image, "B9"),
            (configured_image, "B10"),
            (generic_image, "B11"),
        ],
    )
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "TA/1Q-600",
                "code": "TA/1Q-600",
                "name": "",
                "variant": "",
            },
            {
                "inventory_key": "OHV-124 K5 W9/N4 HERON",
                "code": "OHV-124",
                "name": "HERON",
                "variant": "",
            },
            {
                "inventory_key": "OHE-705 GRIS AIKO",
                "code": "OHE-705",
                "name": "AIKO",
                "variant": "GRIS",
            },
            {
                "inventory_key": "OHE-705 NEGRO AIKO",
                "code": "OHE-705",
                "name": "AIKO",
                "variant": "NEGRO",
            },
        ],
        assets_dir=tmp_path / "assets",
        base_url="/assets/spec",
    )

    assert matches["TA/1Q-600"]["source_code"] == "TA 1/Q-600"
    assert "OHV-124 K5 W9/N4 HERON" not in matches
    assert "OHE-705 GRIS AIKO" not in matches
    assert "OHE-705 NEGRO AIKO" not in matches


def test_deduplicates_same_image_across_spec_and_cost_and_preserves_sources(tmp_path):
    source_image = tmp_path / "aiko.png"
    _png(source_image, (19, 71, 111))
    workbook = tmp_path / "offiho-duplicated.xlsx"
    _duplicated_spec_cost_workbook(workbook, source_image)
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")
    assets_dir = tmp_path / "assets"

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "OHE-705 GRIS AIKO",
                "code": "OHE-705",
                "name": "AIKO",
                "variant": "GRIS",
            }
        ],
        assets_dir=assets_dir,
        base_url="/assets/spec",
    )

    match = matches["OHE-705 GRIS AIKO"]
    assert match["source_sheet"] == "SPEC Sillas"
    assert match["description"] == "Descripcion tecnica de SPEC."
    assert match["source_reference_count"] == 2
    assert {(ref["sheet"], ref["row"]) for ref in match["source_references"]} == {
        ("SPEC Sillas", 9),
        ("Cost Sillas", 9),
    }
    assert match["exact_image_candidate_count"] == 1
    assert len(list(assets_dir.iterdir())) == 1


def test_handles_attached_feature_and_composite_variant_tokens(tmp_path):
    kit_image = tmp_path / "kit.png"
    composite_image = tmp_path / "composite.png"
    _png(kit_image, (130, 130, 130))
    _png(composite_image, (40, 130, 85))
    workbook = tmp_path / "offiho-compact-codes.xlsx"
    _workbook(
        workbook,
        [
            (9, "KITCR", "Accesorio con acabado cromado."),
            (10, "OHX-10verdegris", "Modelo con acabado verde y gris."),
        ],
        [(kit_image, "B9"), (composite_image, "B10")],
    )
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "KIT CR CROMADO",
                "code": "KIT",
                "name": "KIT",
                "variant": "CROMADO",
            },
            {
                "inventory_key": "OHX-10 GRIS/VERDE MODELO",
                "code": "OHX-10",
                "name": "MODELO",
                "variant": "GRIS/VERDE",
            },
        ],
        assets_dir=tmp_path / "assets",
        base_url="/assets/spec",
    )

    assert matches["KIT CR CROMADO"]["source_code"] == "KITCR"
    assert matches["OHX-10 GRIS/VERDE MODELO"]["source_code"] == "OHX-10verdegris"


def test_ignores_zero_sized_hidden_anchor(tmp_path):
    source_image = tmp_path / "hidden.png"
    _png(source_image)
    workbook = tmp_path / "offiho-hidden.xlsx"
    zero_anchor = TwoCellAnchor(
        _from=AnchorMarker(col=1, colOff=0, row=8, rowOff=0),
        to=AnchorMarker(col=1, colOff=0, row=8, rowOff=0),
    )
    _workbook(
        workbook,
        [(9, "OHE-705gris", "Silla Aiko gris.")],
        [(source_image, zero_anchor)],
    )
    module = importlib.import_module("mobiliti_saas.quote_engine.offiho_spec_images")
    assets_dir = tmp_path / "assets"

    matches = module.extract_offiho_spec_images(
        [workbook],
        [
            {
                "inventory_key": "OHE-705 GRIS AIKO",
                "code": "OHE-705",
                "name": "AIKO",
                "variant": "GRIS",
            }
        ],
        assets_dir=assets_dir,
        base_url="/assets/spec",
    )

    assert matches == {}
    assert list(assets_dir.iterdir()) == []
