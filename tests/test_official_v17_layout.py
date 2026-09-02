"""Contrato de regresión para la plantilla oficial Mobiliti versión 17."""

from hashlib import sha256
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
import pytest

from mobiliti_saas.quote_engine import generate_quote
from mobiliti_saas.quote_engine.mobiliti_layout import SectionNeed, plan_mobiliti_layout
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from mobiliti_saas.quote_engine.ooxml_worksheet import (
    MobilitiCellWrite,
    build_mobiliti_sheet,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 Oficial.xlsx"
)
EXPECTED_SHA256 = "39f5cebd3cbe3e7356f4d4174161e8599bf7158e7b495a789c9fc04850928ee4"
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XM = "http://schemas.microsoft.com/office/excel/2006/main"


def _formula(root: ET.Element, coordinate: str) -> str:
    cell = root.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
    assert cell is not None, coordinate
    formula = cell.find(f"{{{MAIN}}}f")
    assert formula is not None and formula.text, coordinate
    return formula.text


def test_active_official_asset_is_the_signed_v17_workbook() -> None:
    package = XlsxPackage.read(TEMPLATE)
    states = {
        name: state
        for name, state, _index, _part in package._sheet_rows()
    }

    assert sha256(TEMPLATE.read_bytes()).hexdigest() == EXPECTED_SHA256
    assert states["Control Administrativo"] == "visible"
    assert states["Fletes"] == "hidden"

    cotizacion = ET.fromstring(package.parts[package.sheet_part("Cotizacion")])
    mobiliti = ET.fromstring(package.parts[package.sheet_part("Mobiliti")])
    control = ET.fromstring(
        package.parts[package.sheet_part("Control Administrativo")]
    )

    assert cotizacion.find(f"{{{MAIN}}}dimension").attrib["ref"] == "A3:BA184"
    assert mobiliti.find(f"{{{MAIN}}}dimension").attrib["ref"] == "A1:AZ614"
    assert _formula(cotizacion, "G17") == "ROUND(Mobiliti!$AD$14,2)"
    assert _formula(cotizacion, "H38") == "H37*$N$39"
    assert _formula(mobiliti, "P6") == 'IF(P4=TRUE,_FV(J6,"Price"),0)'
    assert _formula(mobiliti, "AD14") == "E6"
    assert _formula(mobiliti, "AD15") == "IF(H15>0,$E$5,0)"
    assert _formula(control, "E4") == "Cotizacion!$H$41"


def test_v17_dynamic_rows_keep_the_uniform_price_from_the_largest_quantity() -> None:
    package = XlsxPackage.read(TEMPLATE)
    needs = (
        SectionNeed("large", "SILLAS 60", 1),
        SectionNeed("small", "SILLAS 12", 1),
    )
    row_map = plan_mobiliti_layout(
        needs,
        first_section_row=14,
        canonical_auxiliary_row_count=40,
    )
    large_row, small_row = row_map.item_rows
    mutation = build_mobiliti_sheet(
        package.parts[package.sheet_part("Mobiliti")],
        needs,
        (
            MobilitiCellWrite(f"D{large_row}", "text", "CHT85SW"),
            MobilitiCellWrite(f"H{large_row}", "number", Decimal("60")),
            MobilitiCellWrite(f"D{small_row}", "text", "CHT85SW"),
            MobilitiCellWrite(f"H{small_row}", "number", Decimal("12")),
        ),
    )
    output = ET.fromstring(mutation.xml)
    last_row = row_map.last_product_row

    for row in (large_row, small_row):
        assert _formula(output, f"AA{row}") == (
            f"IF(Z{row}>=Y{row},"
            f"_xlfn.MINIFS($Z$15:$Z${last_row},"
            f"$D$15:$D${last_row},D{row},"
            f"$H$15:$H${last_row},"
            f"_xlfn.MAXIFS($H$15:$H${last_row},"
            f"$D$15:$D${last_row},D{row})),"
            '"NO SE ESTA RESPETANDO EL MARGEN")'
        )
        assert _formula(output, f"AB{row}") == f"IFERROR(AA{row}*H{row},0)"
        assert _formula(output, f"AF{row}") == (
            f'IF($E$5>$E$6,"ERROR",AA{row}-AE{row})'
        )
        assert _formula(output, f"AG{row}") == f"AF{row}*H{row}"


def test_v17_end_to_end_links_new_financial_and_control_surfaces(
    tmp_path: Path,
) -> None:
    source = tmp_path / "quotation-v17-e2e.xlsx"
    workbook = Workbook()
    quotation = workbook.active
    quotation.title = "Quotation"
    for column, header in enumerate(
        (
            "No.",
            "Item Name",
            "Photo",
            "Description",
            "Dimension",
            "Color",
            "Q'ty",
            "Vol.",
            "Tot.Vol.",
            "Unit Price",
            "Tot.Price",
            "Remark",
        ),
        start=1,
    ):
        quotation.cell(7, column).value = header
    quotation["A8"] = "- Sillas"
    for row, quantity, price in ((9, 12, 1000), (10, 60, 900)):
        quotation.cell(row, 1).value = row - 8
        quotation.cell(row, 2).value = "CHT85SW"
        quotation.cell(row, 4).value = f"Silla ejecutiva {quantity} piezas"
        quotation.cell(row, 5).value = "60 x 60 x 90 cm"
        quotation.cell(row, 7).value = quantity
        quotation.cell(row, 8).value = Decimal("0.12")
        quotation.cell(row, 9).value = f"=G{row}*H{row}"
        quotation.cell(row, 10).value = price
        quotation.cell(row, 11).value = f"=G{row}*J{row}"
    workbook.save(source)
    workbook.close()

    output = tmp_path / "cotizacion-v17-e2e.xlsx"
    generate_quote(
        source,
        output,
        {
            "cotizacion": "V17-E2E",
            "proyecto": "Migración de plantilla",
            "cliente": "Cliente QA",
            "lugar_entrega": "Monterrey",
            "quote_currency": "USD",
            "descuento": 35,
        },
        TEMPLATE,
    )

    result = XlsxPackage.read(output)
    assert result.sheet_state("Fletes") == "hidden"
    assert result.sheet_state("Control Administrativo") == "visible"
    cotizacion = ET.fromstring(result.parts[result.sheet_part("Cotizacion")])
    mobiliti = ET.fromstring(result.parts[result.sheet_part("Mobiliti")])
    control = ET.fromstring(
        result.parts[result.sheet_part("Control Administrativo")]
    )
    estrategia = ET.fromstring(
        result.parts[result.sheet_part("Estrategia Comercial ")]
    )

    assert _formula(cotizacion, "F17") == "Mobiliti!AA15"
    assert _formula(cotizacion, "G17") == "ROUND(Mobiliti!$AD$14,2)"
    total_row = int(_formula(control, "E4").removeprefix("Cotizacion!$H$"))
    assert total_row == 23
    delta = total_row - 41
    assert _formula(cotizacion, f"M{37 + delta}") == (
        f"VLOOKUP($D${64 + delta},Tabla_Regiones,2,0)"
    )
    assert _formula(mobiliti, "P8") == "Cotizacion!$D$46"
    delivery_validation = cotizacion.find(f".//{{{X14}}}dataValidation")
    assert delivery_validation is not None
    assert delivery_validation.findtext(
        f"{{{X14}}}formula1/{{{XM}}}f"
    ) == "Fletes!$A$46:$A$55"
    assert delivery_validation.findtext(f"{{{XM}}}sqref") == "D46"
    assert _formula(control, "E3") == f"Cotizacion!H{total_row - 2}"
    assert _formula(estrategia, "B70") == f"Cotizacion!H{total_row}"

    rendered = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert rendered["Mobiliti"]["P15"].value == "=Quotation!I9"
        assert rendered["Mobiliti"]["S15"].value == "Centro"
        assert rendered["Mobiliti"]["AD14"].value == pytest.approx(0.35)
        assert rendered["Cotizacion"]["D46"].value == "Nuevo León"
    finally:
        rendered.close()
