from __future__ import annotations

from decimal import Decimal
from io import BytesIO
import importlib.util
import json
from pathlib import Path
import subprocess
import sys
import threading
import unicodedata
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
import pytest


ROOT = Path(__file__).resolve().parents[1]
WORKER_DIR = ROOT / "mobiliti_saas" / "worker"

from mobiliti_saas.quote_engine import catalog_cart
from mobiliti_saas.quote_engine.supplier_catalog import (
    build_supplier_cart_payload,
    load_supplier_catalog_data,
)


VERIFIED_CODES = (
    "MULT-LIDO-INT",
    "LIDO.OP-INT",
    "JUMP-1.5M",
    "CAJA-FUS",
)
INTERCONNECTION_SOURCE = (
    "LUMBRO/LP/Precios Interconexión Sunón act.xlsx:2026!G4:H4"
)
GENERAL_PRICE_SOURCE = "LUMBRO/LP/LISTA DE PRECIOS MULTICONTACTOS 2026.pdf:5"
SELECTED_INTERNAL_ID = "lumbro:variant:mult-lido-int"
BARCELONA_REVIEW_ID = "lumbro:variant:barcelona-gris"


def test_module_import_does_not_mutate_environment_or_sys_path():
    script = f"""
import importlib.util
import os
from pathlib import Path
import sys

for key in ("SUPABASE_URL", "SUPABASE_SERVICE_KEY", "JWT_SECRET_KEY"):
    os.environ.pop(key, None)
before_environment = dict(os.environ)
before_path = list(sys.path)
module_path = Path({str(Path(__file__).resolve())!r})
spec = importlib.util.spec_from_file_location("lumbro_e2e_isolation_probe", module_path)
module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(module)
assert dict(os.environ) == before_environment
assert sys.path == before_path
"""
    result = subprocess.run(
        [sys.executable, "-c", script],
        cwd=ROOT,
        capture_output=True,
        text=True,
        timeout=30,
        check=False,
    )

    assert result.returncode == 0, result.stderr


@pytest.fixture
def isolated_quote_runtime(monkeypatch):
    monkeypatch.setenv("SUPABASE_URL", "http://localhost:54321")
    monkeypatch.setenv("SUPABASE_SERVICE_KEY", "test-key")
    monkeypatch.setenv("JWT_SECRET_KEY", "test-secret-32-chars-long!!!!!")
    monkeypatch.syspath_prepend(str(WORKER_DIR))
    modules_before = set(sys.modules)
    suffix = uuid.uuid4().hex

    def load_module(name: str, path: Path):
        spec = importlib.util.spec_from_file_location(name, path)
        assert spec is not None and spec.loader is not None
        module = importlib.util.module_from_spec(spec)
        sys.modules[name] = module
        spec.loader.exec_module(module)
        return module

    api_index = load_module(
        f"lumbro_e2e_api_{suffix}",
        ROOT / "mobiliti_saas" / "api" / "index.py",
    )
    quote_worker = load_module(
        f"lumbro_e2e_worker_{suffix}",
        WORKER_DIR / "quote_worker.py",
    )
    try:
        yield api_index, quote_worker
    finally:
        for name in set(sys.modules) - modules_before:
            module_path = getattr(sys.modules.get(name), "__file__", None)
            if name.endswith(suffix) or (
                module_path
                and Path(module_path).resolve().is_relative_to(WORKER_DIR.resolve())
            ):
                sys.modules.pop(name, None)


def _item(**overrides) -> dict:
    item = {
        "internal_id": SELECTED_INTERNAL_ID,
        "supplier": "lumbro",
        "product_key": "mult-lido-int",
        "sku": "MULT-LIDO-INT",
        "code_status": "verified",
        "brand": "Lumbro",
        "collection": "Interconexión",
        "name": "=Multicontacto LIDO para interconectar",
        "description": "+Multicontacto especial de 4 puertos AC y USB doble",
        "unit": "PZA",
        "availability_type": "unknown",
        "stock": None,
        "lead_time": "",
        "base_price_options": [],
        "add_on_options": [],
        "base_currency": "MXN",
        "price_net": "3003.000000",
        "tax_rate": "0.160000",
        "attributes": {
            "source_code": "MULT-LIDO-INT",
            "configuration": "Interconexión",
            "dimensions": "420 x 160 mm",
        },
        "image_url": (
            "https://project-ref.supabase.co/storage/v1/object/public/"
            "catalog-assets/lumbro/mult-lido-int.png"
        ),
        "image_kind": "generated_reference",
        "product_url": "https://www.lumbromx.com/productos-1",
        "warnings": [],
        "source_reference": INTERCONNECTION_SOURCE,
    }
    item.update(overrides)
    return item


@pytest.fixture
def representative_lumbro_snapshot() -> dict:
    verified = [_item()]
    for index, (code, price) in enumerate(
        zip(
            VERIFIED_CODES[1:],
            ("1394.070000", "350.000000", "772.000000"),
            strict=True,
        ),
        start=2,
    ):
        verified.append(
            _item(
                internal_id=f"lumbro:variant:official-{index}",
                product_key=f"official-{index}",
                sku=code,
                name=code,
                description=f"Accesorio de interconexión {code}",
                price_net=price,
                attributes={
                    "source_code": code,
                    "configuration": "Interconexión",
                    "dimensions": "",
                },
                image_url="",
                image_kind="placeholder",
                product_url="",
            )
        )
    barcelona = _item(
        internal_id=BARCELONA_REVIEW_ID,
        product_key="barcelona-gris",
        sku="",
        code_status="needs_review",
        collection="Empotrables",
        name="Barcelona",
        description="Multicontacto empotrable Barcelona",
        price_net="2824.000000",
        attributes={
            "source_code": "BARCELONA",
            "configuration": "",
            "dimensions": "245 x 102 x 60 mm",
        },
        image_url="",
        image_kind="placeholder",
        product_url="https://www.lumbromx.com/empotrados",
        warnings=["Código oficial repetido; requiere revisión"],
        source_reference=GENERAL_PRICE_SOURCE,
    )
    return {
        "supplier": "lumbro",
        "source_hash": "c" * 64,
        "generated_at": "2026-07-18T12:00:00Z",
        "items": [*verified, barcelona],
        "metadata": {
            "coverage": {
                "items": 5,
                "verified_items": 4,
                "needs_review_items": 1,
                "priced_items": 5,
            }
        },
    }


def test_local_lumbro_review_item_crosses_supplier_cart_with_one_canonical_warning(
    representative_lumbro_snapshot,
):
    review = next(
        item for item in representative_lumbro_snapshot["items"]
        if item["code_status"] == "needs_review"
    )
    line = build_supplier_cart_payload(
        [{"internal_id": review["internal_id"], "quantity": "1", "add_on_option_ids": []}],
        representative_lumbro_snapshot,
        "MXN",
        [],
    )["items"][0]

    def normalized(warning):
        return " ".join(
            "".join(
                character for character in unicodedata.normalize("NFKD", warning.casefold())
                if not unicodedata.combining(character)
            ).split()
        )

    assert any("repetido" in warning.lower() for warning in line["warnings"])
    assert [warning for warning in line["warnings"] if normalized(warning) == "codigo por verificar"] == ["Codigo por verificar"]


def _isolated_barcelona_arithmetic_catalog() -> dict:
    return {
        "supplier": "lumbro",
        "source_hash": "a" * 64,
        "generated_at": "2026-07-18T12:00:00Z",
        "items": [
            _item(
                internal_id="lumbro:contract:barcelona-arithmetic",
                product_key="barcelona-arithmetic",
                sku="BARCELONA",
                name="Barcelona",
                description="Fixture contractual aislado de aritmética",
                collection="Empotrables",
                price_net="2824.000000",
                attributes={
                    "source_code": "BARCELONA",
                    "configuration": "",
                    "dimensions": "245 x 102 x 60 mm",
                },
                image_url="",
                image_kind="placeholder",
                product_url="",
                source_reference=GENERAL_PRICE_SOURCE,
            )
        ],
    }


def test_generic_net_price_arithmetic_for_isolated_barcelona_contract():
    """Prueba aritmética aislada; no afirma que Barcelona sea cotizable en Task 6."""
    catalog = load_supplier_catalog_data(
        _isolated_barcelona_arithmetic_catalog(),
        expected_supplier="lumbro",
    )

    cart = build_supplier_cart_payload(
        [
            {
                "internal_id": "lumbro:contract:barcelona-arithmetic",
                "quantity": "2",
                "add_on_option_ids": [],
            }
        ],
        catalog,
        "MXN",
        [],
    )

    line = cart["items"][0]
    net = Decimal(line["line_total"])
    iva = (net * Decimal(line["tax_rate"])).quantize(Decimal("0.01"))
    assert line["sku"] == "BARCELONA"
    assert line["unit_price"] == "2824.00"
    assert line["quantity"] == "2"
    assert net == Decimal("5648.00")
    assert iva == Decimal("903.68")


def _install_api_boundary(monkeypatch, catalog: dict, api_index) -> dict:
    state = {"created": [], "uploaded": [], "queued": []}
    monkeypatch.setattr(
        api_index,
        "db_get_usuario_by_id",
        lambda user_id: {
            "id": user_id,
            "email": "cliente@example.com",
            "nombre": "Cliente",
            "empresa": "Mobiliti",
            "es_admin": False,
            "activo": True,
        },
    )
    monkeypatch.setattr(
        api_index,
        "db_get_suscripcion_by_usuario",
        lambda user_id: {
            "id": 1,
            "usuario_id": user_id,
            "estado": "activa",
            "plan": "mensual",
            "fecha_fin": "2099-01-01T00:00:00+00:00",
        },
    )
    monkeypatch.setattr(api_index, "CATALOG_ENABLED_SUPPLIERS", ("lumbro",))
    monkeypatch.setattr(api_index, "_load_supplier_catalog_cached", lambda supplier: catalog)
    monkeypatch.setattr(api_index, "db_list_exchange_rates", lambda: [])
    monkeypatch.setattr(api_index, "_next_quote_number_for_user", lambda user: None)
    monkeypatch.setattr(api_index, "_enforce_active_quote_limit", lambda user_id: None)
    monkeypatch.setattr(api_index, "_enforce_quote_history_limit", lambda user_id: [])
    monkeypatch.setattr(api_index, "_storage_provider_name", lambda: "supabase")
    monkeypatch.setattr(api_index, "_wake_worker", lambda: None)

    def create_job(user_id, template, metadata, input_path, job_id=None):
        row = {
            "id": job_id,
            "usuario_id": user_id,
            "status": "draft",
            "template": template,
            "metadata": dict(metadata),
            "input_path": input_path,
        }
        state["created"].append(row)
        return row

    def upload(path, content, content_type="application/octet-stream"):
        state["uploaded"].append(
            {"path": path, "content": content, "content_type": content_type}
        )

    def update_job(job_id, updates):
        row = {"id": job_id, **updates}
        state["queued"].append(row)
        return row

    monkeypatch.setattr(api_index, "db_create_quote_job", create_job)
    monkeypatch.setattr(api_index, "_storage_upload_bytes", upload)
    monkeypatch.setattr(api_index, "db_update_quote_job", update_job)
    return state


def _auth_headers(api_index) -> dict[str, str]:
    token = api_index.create_access_token(
        {"sub": "7", "email": "cliente@example.com"}
    )
    return {"Authorization": f"Bearer {token}"}


def _quote_body(internal_id: str) -> dict:
    return {
        "proyecto": "=Proyecto Lumbro E2E",
        "cliente": "Cliente",
        "correo": "cliente@example.com",
        "telefono": "5551234567",
        "direccion": "Guadalajara",
        "razon_social": "Cliente SA de CV",
        "descuento": 40,
        "quote_currency": "MXN",
        "items": [
            {
                "internal_id": internal_id,
                "quantity": "2",
                "add_on_option_ids": [],
            }
        ],
    }


class _MemoryWorkerClient:
    def __init__(self, job: dict, input_content: bytes):
        assert job["status"] == "queued"
        self.job = dict(job)
        self.input_content = input_content
        self.output_content = b""
        self.completed = None
        self.uploaded_path = None
        self.progress_updates = []
        self.input_deleted = False
        self.input_cleared = False
        self.lock = threading.RLock()

    def _update_current_row(self, data: dict) -> list[dict]:
        self.job.update(data)
        return [dict(self.job)]

    def rest(self, method, path, params=None, data=None):
        assert params is None
        assert method == "PATCH"
        assert isinstance(data, dict)
        with self.lock:
            job_id = self.job["id"]
            claim_path = f"/saas_quote_jobs?id=eq.{job_id}&status=eq.queued"
            if path == claim_path:
                assert self.job["status"] == "queued"
                assert set(data) == {
                    "status", "attempt_token", "lease_expires_at", "updated_at",
                }
                assert data["status"] == "processing"
                assert data["attempt_token"]
                assert data["lease_expires_at"]
                return self._update_current_row(data)

            attempt_token = self.job.get("attempt_token")
            processing_path = (
                f"/saas_quote_jobs?id=eq.{job_id}&status=eq.processing"
                f"&attempt_token=eq.{attempt_token}"
            )
            if path == processing_path:
                assert self.job["status"] == "processing"
                if data.get("status") == "completed":
                    assert set(data) == {
                        "status", "output_path", "metadata", "error_message",
                        "lease_expires_at", "updated_at", "completed_at",
                    }
                    assert self.uploaded_path == data["output_path"]
                    assert data["metadata"]["progress_percent"] == 100
                    row = self._update_current_row(data)[0]
                    self.completed = dict(row)
                    return [row]

                assert set(data) in (
                    {"metadata", "lease_expires_at", "updated_at"},
                    {"lease_expires_at", "updated_at"},
                )
                if "metadata" in data:
                    progress = data["metadata"]["progress_percent"]
                    assert progress in {45, 55, 90}
                    self.progress_updates.append(progress)
                return self._update_current_row(data)

            completed_path = (
                f"/saas_quote_jobs?id=eq.{job_id}&status=eq.completed"
                f"&attempt_token=eq.{attempt_token}"
            )
            if path == completed_path:
                assert self.job["status"] == "completed"
                assert self.completed is not None
                assert self.input_deleted is True
                assert set(data) == {"input_path", "updated_at"}
                assert data["input_path"] is None
                self.input_cleared = True
                return self._update_current_row(data)

        raise AssertionError(f"Ruta REST inesperada: {method} {path}")

    def storage_download(self, object_path, destination):
        assert self.job["status"] == "processing"
        assert object_path == self.job["input_path"]
        Path(destination).write_bytes(self.input_content)

    def storage_upload(self, object_path, source):
        expected_path = (
            f"users/{self.job['usuario_id']}/jobs/{self.job['id']}/attempts/"
            f"{self.job['attempt_token']}/output.xlsx"
        )
        assert self.job["status"] == "processing"
        assert object_path == expected_path
        self.uploaded_path = object_path
        self.output_content = Path(source).read_bytes()

    def storage_delete(self, object_path):
        assert self.completed is not None
        assert self.job["status"] == "completed"
        assert self.job["output_path"] == self.uploaded_path
        assert object_path == self.job["input_path"]
        self.input_deleted = True


def test_real_verified_lumbro_item_crosses_api_worker_and_xlsx_without_second_discount(
    representative_lumbro_snapshot,
    isolated_quote_runtime,
    monkeypatch,
    tmp_path,
):
    api_index, quote_worker = isolated_quote_runtime
    selected = next(
        item
        for item in representative_lumbro_snapshot["items"]
        if item["internal_id"] == SELECTED_INTERNAL_ID
    )
    selected["attributes"]["product_notes"] = [
        "NOTA: SE PUEDEN MODIFICAR LAS CONEXIONES CON PRECIO ESPECIAL"
    ]
    loaded = load_supplier_catalog_data(
        representative_lumbro_snapshot,
        expected_supplier="lumbro",
    )
    verified = [item for item in loaded["items"] if item["code_status"] == "verified"]
    review = next(item for item in loaded["items"] if item["code_status"] == "needs_review")
    assert tuple(item["sku"] for item in verified) == VERIFIED_CODES
    assert len(verified) == 4
    assert review["internal_id"] == BARCELONA_REVIEW_ID
    assert review["name"] == "Barcelona"
    assert review["sku"] == ""

    state = _install_api_boundary(
        monkeypatch,
        representative_lumbro_snapshot,
        api_index,
    )
    client = TestClient(api_index.app)
    body = _quote_body(SELECTED_INTERNAL_ID)
    body["image_provider"] = "pillow"
    accepted = client.post(
        "/catalogs/lumbro/quote",
        headers=_auth_headers(api_index),
        json=body,
    )
    assert accepted.status_code == 200
    assert len(state["created"]) == len(state["uploaded"]) == len(state["queued"]) == 1
    payload = json.loads(state["uploaded"][0]["content"].decode("utf-8"))
    line = payload["items"][0]
    assert line["sku"] == "MULT-LIDO-INT"
    assert line["name"] == "=Multicontacto LIDO para interconectar"
    assert line["description"] == "+Multicontacto especial de 4 puertos AC y USB doble"
    assert line["attributes"]["dimensions"] == "420 x 160 mm"
    assert line["unit"] == "PZA"
    assert line["quantity"] == "2"
    assert line["unit_price"] == "3003.00"
    assert line["line_total"] == "6006.00"
    assert line["tax_rate"] == "0.160000"
    assert line["source_reference"] == INTERCONNECTION_SOURCE
    assert line["image_kind"] == "generated_reference"
    assert line["attributes"]["product_notes"] == [
        "NOTA: SE PUEDEN MODIFICAR LAS CONEXIONES CON PRECIO ESPECIAL"
    ]
    assert Decimal(line["line_total"]) * Decimal(line["tax_rate"]) == Decimal("960.96000000")
    assert all(item["internal_id"] != BARCELONA_REVIEW_ID for item in payload["items"])

    local_image = tmp_path / "lumbro-safe-fixture.png"
    Image.new("RGB", (24, 16), (35, 90, 140)).save(local_image, format="PNG")

    def local_catalog_image(url, image_dir, code, source_type):
        assert url == line["image_url"]
        assert code == "MULT-LIDO-INT"
        assert source_type == "supplier_cart"
        return local_image

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", local_catalog_image)
    created_job = {**state["created"][0], **state["queued"][0]}
    worker_client = _MemoryWorkerClient(created_job, state["uploaded"][0]["content"])
    quote_worker.process_job(worker_client, created_job)

    assert worker_client.completed is not None
    assert worker_client.completed["status"] == "completed"
    assert worker_client.completed["metadata"]["catalog_supplier"] == "lumbro"
    assert worker_client.completed["metadata"]["descuento"] == 0
    assert worker_client.progress_updates == [45, 55, 90]
    assert worker_client.input_deleted is True
    assert worker_client.input_cleared is True
    assert worker_client.uploaded_path == (
        f"users/{created_job['usuario_id']}/jobs/{created_job['id']}/attempts/"
        f"{worker_client.completed['attempt_token']}/output.xlsx"
    )
    assert worker_client.output_content.startswith(b"PK")

    workbook = load_workbook(BytesIO(worker_client.output_content), data_only=False)
    try:
        assert {"Cotizacion", "Mobiliti", "Quotation"} <= set(workbook.sheetnames)
        quotation = workbook["Quotation"]
        assert quotation["A8"].value == "- Lumbro"
        assert quotation["B9"].value == "'=Multicontacto LIDO para interconectar"
        assert quotation["D9"].value.startswith(
            "Multicontacto modelo '=Multicontacto LIDO para interconectar."
        )
        assert "'+Multicontacto especial" in quotation["D9"].value
        assert f"Fuente: {INTERCONNECTION_SOURCE}" in quotation["E9"].value
        assert "SKU: MULT-LIDO-INT" in quotation["E9"].value
        assert "Unidad: PZA" in quotation["E9"].value
        assert "Imagen de referencia" in quotation["E9"].value
        assert "Notas:" not in quotation["D9"].value
        assert "Disponibilidad: por confirmar" not in quotation["D9"].value
        assert quotation["F9"].value == "420 x 160 mm"
        assert quotation["H9"].value == 2
        assert isinstance(quotation["H9"].value, int)
        assert quotation["K9"].value == 3003
        assert "URL: https://www.lumbromx.com/productos-1" in quotation["E9"].value
        assert len(quotation._images) == 1

        cotizacion = workbook["Cotizacion"]
        product_row = next(
            row
            for row in range(1, cotizacion.max_row + 1)
            if str(cotizacion.cell(row, 1).value or "").startswith("=Mobiliti!D")
        )
        assert cotizacion["B7"].value == "'=Proyecto Lumbro E2E"
        assert cotizacion.cell(product_row, 7).value == 0
        assert cotizacion.cell(product_row, 10).value == f"=E{product_row}*I{product_row}"
        product_images = [
            image
            for image in cotizacion._images
            if image.anchor._from.row + 1 == product_row
        ]
        assert len(product_images) == 1
        with Image.open(BytesIO(product_images[0]._data())) as generated_image:
            with Image.open(local_image) as native_image:
                assert generated_image.size == native_image.size
                assert generated_image.convert("RGBA").tobytes() == (
                    native_image.convert("RGBA").tobytes()
                )
        iva_rows = [
            row
            for row in range(product_row + 1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 4).value == "IVA:"
        ]
        assert len(iva_rows) == 1
        iva_row = iva_rows[0]
        assert cotizacion.cell(iva_row, 8).value == f"=H{iva_row - 1}*16%"
        total_row = next(
            row
            for row in range(iva_row + 1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 4).value == "TOTAL:"
        )
        assert total_row == iva_row + 1
        assert cotizacion.cell(total_row, 8).value == (
            f"=H{iva_row - 1}+H{iva_row}"
        )
        output_product_rows = [
            row
            for row in range(9, quotation.max_row + 1)
            if isinstance(quotation.cell(row, 1).value, (int, float))
        ]
        assert output_product_rows == [9]
        assert "Barcelona" not in str(quotation["B9"].value)
        assert "Barcelona" not in str(quotation["D9"].value)
    finally:
        workbook.close()
