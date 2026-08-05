from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import struct

import pytest
from openpyxl import Workbook, load_workbook

import mobiliti_saas.quote_engine.quotation_import as quotation_import
from mobiliti_saas.quote_engine.quotation_import import (
    build_import_manifest,
    extract_images_from_bytes,
    normalize_imported_items,
    read_items_from_bytes,
    validate_import_manifest,
)
from quotation_import_fixtures import write_import_fixture


IMPORT_ID = "7b1d6d42-236a-4bc1-9aa8-8d9db793c30b"


@pytest.fixture
def import_manifest(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")
    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )
    return manifest


def test_build_import_manifest_preserves_sections_rows_images_and_requires_currency(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")

    manifest, image_map = build_import_manifest(
        source.read_bytes(),
        import_id=IMPORT_ID,
        original_filename=source.name,
    )

    assert manifest["source_currency"] is None
    assert manifest["currency_status"] == "required"
    assert manifest["provider"] == "SUNON TECHNOLOGY CO.,LTD."
    assert [row["title"] for row in manifest["sections"]] == [
        "SALA DE JUNTAS SECUNDARIO",
        "MUESTRAS",
        "CONCEJO",
    ]
    assert len(manifest["items"]) == 7
    assert manifest["items"][0]["key"].endswith(":9")
    assert sorted(image_map) == [9, 11, 12, 13, 14, 15, 17]
    assert all(image and media_type.startswith("image/") for image, media_type in image_map.values())


def test_build_import_manifest_detects_explicit_currency(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx", currency="USD")

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert manifest["source_currency"] == "USD"
    assert manifest["currency_status"] == "detected"
    assert {item["source_currency"] for item in manifest["items"]} == {"USD"}
    assert len({item["row_hash"] for item in manifest["items"]}) == 7


def test_build_import_manifest_rounds_excel_float_noise_to_six_places(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    workbook["Quotation"]["J9"] = 92.39999999999999
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert manifest["items"][0]["unit_price"] == "92.4"


def test_build_import_manifest_rounds_excel_quantity_float_noise_to_six_places(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    workbook["Quotation"]["G9"] = 3.600000000000001
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert manifest["items"][0]["quantity"] == "3.6"


def test_build_import_manifest_normalizes_multiline_workbook_descriptions(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    workbook["Quotation"]["B9"] = "Mesa\nprincipal"
    workbook["Quotation"]["D9"] = "Linea uno\nLinea dos\r\nLinea tres\tfin"
    workbook["Quotation"]["E9"] = None
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert manifest["items"][0]["name"] == "Mesa principal"
    assert manifest["items"][0]["description"] == "Linea uno Linea dos Linea tres fin"
    assert manifest["items"][0]["dimension"] == ""


@pytest.mark.parametrize(
    ("provider", "expected_provider"),
    [
        (None, ""),
        ("", ""),
        ("SUNON\nTECHNOLOGY\rCO.\tLTD.", "SUNON TECHNOLOGY CO. LTD."),
    ],
)
def test_build_import_manifest_normalizes_section_and_optional_provider_workbook_text(
    tmp_path, provider, expected_provider
):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    sheet = workbook["Quotation"]
    sheet["A1"] = provider
    sheet["A8"] = "- SALA\nDE\rJUNTAS\tNORTE"
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert manifest["provider"] == expected_provider
    assert manifest["sections"][0]["title"] == "SALA DE JUNTAS NORTE"


def test_build_import_manifest_uses_default_for_blank_section_title(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    workbook["Quotation"]["A8"] = "-  "
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(),
        import_id=IMPORT_ID,
        original_filename=source.name,
    )

    assert manifest["sections"][0]["title"] == "Sin categoria"
    assert manifest["items"][0]["category"] == "Sin categoria"


@pytest.mark.parametrize(
    ("cell", "value"),
    [
        ("A1", "=SUM(A2:A3)"),
        ("A8", "- =SUM(A2:A3)"),
    ],
)
def test_build_import_manifest_rejects_formulas_in_provider_and_section_titles(
    tmp_path, cell, value
):
    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    workbook["Quotation"][cell] = value
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError):
        build_import_manifest(
            source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
        )


def test_workbook_text_only_accepts_none_for_explicit_optional_fields():
    with pytest.raises(ValueError, match="Descripcion"):
        quotation_import._workbook_text(None, "Descripcion", allow_empty=True)
    with pytest.raises(ValueError, match="Nombre"):
        quotation_import._workbook_text(None, "Nombre")
    with pytest.raises(ValueError, match="Categoria"):
        quotation_import._workbook_text(None, "Categoria")

    assert quotation_import._workbook_text(
        None, "Dimension", allow_empty=True, allow_none=True
    ) == ""
    assert quotation_import._workbook_text(
        None, "Proveedor", allow_empty=True, allow_none=True
    ) == ""


@pytest.mark.parametrize("value", ["=SUM(A1:A2)", "+cmd", "texto\x0bcontrol"])
def test_workbook_text_normalization_keeps_formula_and_control_rejection(value):
    with pytest.raises(ValueError, match="Campo"):
        quotation_import._workbook_text(value, "Campo", allow_empty=True)


def test_quotation_import_module_copies_are_byte_identical():
    paths = [
        Path("mobiliti_saas/quote_engine/quotation_import.py"),
        Path("mobiliti_saas/web/mobiliti_saas/quote_engine/quotation_import.py"),
    ]

    assert len({path.read_bytes() for path in paths}) == 1


def test_normalize_imported_items_uses_selected_currency_and_allowed_overrides(import_manifest):
    rows = normalize_imported_items(
        [
            {
                "kind": "imported",
                "import_id": import_manifest["import_id"],
                "source_row": 11,
                "source_currency": "USD",
                "quantity": "2",
                "overrides": {
                    "name": "Alien Task Chair revisada",
                    "description": "Silla operativa revisada",
                    "dimension": "630 x 565 x 1000 mm",
                    "unit_price": "82.00",
                    "provider": "Sunon",
                },
            }
        ],
        import_manifest,
        source_currency="USD",
        quote_currency="MXN",
        rate_rows=[
            {
                "currency": "USD",
                "mxn_per_unit": "18.50",
                "effective_date": "2026-07-21",
                "retrieved_at": "2026-07-21T00:00:00Z",
            }
        ],
        discount_percent="40",
    )

    assert rows[0]["original_unit_price"] == "82.000000"
    assert rows[0]["unit_price"] == "1517.00"
    assert rows[0]["frozen_exchange_rate"] == "18.500000"
    assert rows[0]["source_reference"].endswith("#Quotation!11")


@pytest.mark.parametrize("field", ("quantity", "unit_price"))
@pytest.mark.parametrize(
    "value",
    (
        True,
        1,
        1.0,
        float("nan"),
        float("inf"),
        "NaN",
        "Infinity",
        "1e0",
        "1.0000001",
        "+1",
        "01",
        " 1",
        "1 ",
        "1\t",
        "1١",
        "1.١",
    ),
)
def test_normalize_imported_items_rejects_noncanonical_wire_decimals(
    import_manifest, field, value
):
    item = {
        "kind": "imported",
        "import_id": import_manifest["import_id"],
        "source_row": 11,
        "source_currency": "MXN",
        "quantity": "1",
        "overrides": {
            "name": "Alien Task Chair",
            "description": "Silla operativa",
            "dimension": "630 x 565 x 1000 mm",
            "unit_price": "82.00",
            "provider": "Sunon",
        },
    }
    if field == "quantity":
        item["quantity"] = value
    else:
        item["overrides"]["unit_price"] = value

    with pytest.raises(ValueError):
        normalize_imported_items(
            [item],
            import_manifest,
            source_currency="MXN",
            quote_currency="MXN",
            rate_rows=[],
            discount_percent="40",
        )


def test_normalize_imported_items_rejects_payload_currency_that_replaces_explicit_row_currency(tmp_path):
    source = write_import_fixture(tmp_path / "source.xlsx", currency="USD")
    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )
    item = {
        "kind": "imported",
        "import_id": manifest["import_id"],
        "source_row": 11,
        "source_currency": "MXN",
        "quantity": "1",
        "overrides": {
            "name": "Alien Task Chair",
            "description": "Silla operativa",
            "dimension": "630 x 565 x 1000 mm",
            "unit_price": "82.00",
            "provider": "Sunon",
        },
    }

    with pytest.raises(ValueError, match="Moneda de origen explicita"):
        normalize_imported_items(
            [item],
            manifest,
            source_currency="MXN",
            quote_currency="MXN",
            rate_rows=[{"currency": "USD", "mxn_per_unit": "18.50", "effective_date": "2026-07-21"}],
            discount_percent="0",
        )


def test_normalize_imported_items_allows_payload_currency_when_original_row_has_none(import_manifest):
    item = {
        "kind": "imported",
        "import_id": import_manifest["import_id"],
        "source_row": 11,
        "source_currency": "USD",
        "quantity": "1",
        "overrides": {
            "name": "Alien Task Chair",
            "description": "Silla operativa",
            "dimension": "630 x 565 x 1000 mm",
            "unit_price": "82.00",
            "provider": "Sunon",
        },
    }

    rows = normalize_imported_items(
        [item],
        import_manifest,
        source_currency="MXN",
        quote_currency="MXN",
        rate_rows=[{"currency": "USD", "mxn_per_unit": "18.50", "effective_date": "2026-07-21"}],
        discount_percent="0",
    )

    assert rows[0]["original_currency"] == "USD"
    assert rows[0]["frozen_exchange_rate"] == "18.500000"


def _zip_with_declared_expansion(
    uncompressed_size: int, *, name: bytes = b"xl/worksheets/sheet1.xml"
) -> bytes:
    """Crea un ZIP minimo cuyo directorio central declara una expansion excesiva."""
    compressed = b"x"
    local = struct.pack(
        "<IHHHHHIIIHH",
        0x04034B50, 20, 0, 0, 0, 0, 0, len(compressed), uncompressed_size, len(name), 0,
    ) + name + compressed
    central = struct.pack(
        "<IHHHHHHIIIHHHHHII",
        0x02014B50, 20, 20, 0, 0, 0, 0, 0, len(compressed), uncompressed_size,
        len(name), 0, 0, 0, 0, 0, 0,
    ) + name
    end = struct.pack("<IHHHHIIH", 0x06054B50, 0, 0, 1, 1, len(central), len(local), 0)
    return local + central + end


def test_xlsx_preflight_rejects_declared_expansion_before_openpyxl_or_image_reads(monkeypatch):
    source = _zip_with_declared_expansion(quotation_import.MAX_ZIP_MEMBER_UNCOMPRESSED + 1)
    load_workbook_calls = []

    def unexpected_load_workbook(*args, **kwargs):
        load_workbook_calls.append((args, kwargs))
        raise AssertionError("load_workbook no debe ejecutarse despues de un preflight rechazado")

    monkeypatch.setattr(quotation_import, "load_workbook", unexpected_load_workbook)

    with pytest.raises(ValueError, match="archivo .xlsx inseguro"):
        read_items_from_bytes(source)
    with pytest.raises(ValueError, match="archivo .xlsx inseguro"):
        extract_images_from_bytes(source)

    assert load_workbook_calls == []


def test_xlsx_preflight_rejects_anomalous_member_names():
    source = _zip_with_declared_expansion(1, name=b"xl/../escape.xml")

    with pytest.raises(ValueError, match="nombres de miembros"):
        extract_images_from_bytes(source)


@pytest.mark.parametrize(
    "mutation",
    [
        lambda item: item["overrides"].update(name="=SUM(A1:A2)"),
        lambda item: item.update(quantity="0"),
        lambda item: item["overrides"].update(unit_price="-1"),
        lambda item: item.update(source_currency=""),
        lambda item: item["overrides"].update(unexpected="no permitido"),
    ],
)
def test_normalize_imported_items_rejects_unsafe_or_invalid_values(import_manifest, mutation):
    item = {
        "kind": "imported",
        "import_id": import_manifest["import_id"],
        "source_row": 11,
        "source_currency": "USD",
        "quantity": "1",
        "overrides": {
            "name": "Alien Task Chair",
            "description": "Silla operativa",
            "dimension": "630 x 565 x 1000 mm",
            "unit_price": "82.00",
            "provider": "Sunon",
        },
    }
    mutation(item)

    with pytest.raises(ValueError):
        normalize_imported_items(
            [item],
            import_manifest,
            source_currency=None,
            quote_currency="MXN",
            rate_rows=[],
            discount_percent="40",
        )


def test_normalize_imported_items_rejects_duplicate_rows(import_manifest):
    item = {
        "kind": "imported",
        "import_id": import_manifest["import_id"],
        "source_row": 11,
        "source_currency": "MXN",
        "quantity": "1",
        "overrides": {
            "name": "Alien Task Chair",
            "description": "Silla operativa",
            "dimension": "630 x 565 x 1000 mm",
            "unit_price": "82.00",
            "provider": "Sunon",
        },
    }

    with pytest.raises(ValueError, match="Fila importada invalida"):
        normalize_imported_items(
            [item, deepcopy(item)],
            import_manifest,
            source_currency="MXN",
            quote_currency="MXN",
            rate_rows=[],
            discount_percent="40",
        )


def test_imported_line_limit_is_derived_from_physical_xlsx_capacity():
    assert quotation_import.MAX_IMPORTED_LINES > 500
    assert quotation_import.MAX_IMPORTED_LINES == (
        quotation_import.XLSX_MAX_ROWS
        - quotation_import.MOBILITI_RESERVED_ROWS_AFTER_TOTAL
    )


def test_build_import_manifest_accepts_1000_rows_above_old_limit(tmp_path):
    source = tmp_path / "large.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet["A1"] = "Proveedor grande"
    for column, title in {
        1: "No.",
        2: "Item Name",
        4: "Description",
        5: "Dimension",
        7: "Q'ty",
        10: "Unit Price",
        14: "Original Currency",
    }.items():
        sheet.cell(7, column, title)
    for index in range(1_000):
        row = index + 8
        sheet.cell(row, 1, index + 1)
        sheet.cell(row, 2, f"Producto {index + 1}")
        sheet.cell(row, 4, "Descripcion compacta")
        sheet.cell(row, 5, "600 x 600 mm")
        sheet.cell(row, 7, 1)
        sheet.cell(row, 10, 100)
        sheet.cell(row, 14, "MXN")
    workbook.save(source)
    workbook.close()

    manifest, image_map = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    assert len(manifest["items"]) == 1_000
    assert manifest["sections"][0]["title"] == "Sin categoria"
    assert image_map == {}


def test_validate_quote_size_reports_required_physical_row():
    with pytest.raises(ValueError, match=r"fila .*1048576"):
        quotation_import.validate_quote_size(
            section_counts=[quotation_import.XLSX_MAX_ROWS],
            encoded_bytes=0,
        )


def test_validate_quote_size_rejects_non_integer_section_counts():
    with pytest.raises(ValueError, match="seccion"):
        quotation_import.validate_quote_size(
            section_counts=["1"],
            encoded_bytes=0,
        )


def test_manifest_item_preserves_provider_and_official_code_from_explicit_columns(tmp_path):
    from openpyxl import load_workbook

    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    sheet = workbook["Quotation"]
    sheet["K7"] = "Official Code"
    sheet["L7"] = "Provider"
    sheet["K11"] = "ALIEN-001"
    sheet["L11"] = "Sunon\nMexico"
    workbook.save(source)
    workbook.close()

    manifest, _ = build_import_manifest(
        source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
    )

    item = next(row for row in manifest["items"] if row["source_row"] == 11)
    assert item["official_code"] == "ALIEN-001"
    assert item["provider"] == "Sunon Mexico"
    assert item["official_code"] != item["name"]


@pytest.mark.parametrize("cell", ("K11", "L11"))
def test_manifest_item_rejects_formula_prefixes_in_structured_columns(tmp_path, cell):
    from openpyxl import load_workbook

    source = write_import_fixture(tmp_path / "source.xlsx")
    workbook = load_workbook(source)
    sheet = workbook["Quotation"]
    sheet["K7"] = "Official Code"
    sheet["L7"] = "Provider"
    sheet[cell] = "=SUM(A1:A2)"
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError):
        build_import_manifest(
            source.read_bytes(), import_id=IMPORT_ID, original_filename=source.name
        )
