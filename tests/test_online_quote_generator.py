import os
import sys
from pathlib import Path

from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorkbookImage

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "mobiliti_saas", "worker"))

from online_quote_generator import generate_online_quote, read_items


def _sample_quotation(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    headers = {
        1: "No.",
        2: "Item",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
    }
    for col, value in headers.items():
        ws.cell(row=7, column=col, value=value)
    ws.cell(row=8, column=1, value="- Escritorios")
    ws.cell(row=9, column=1, value=1)
    ws.cell(row=9, column=2, value="Mesa Uno")
    ws.cell(row=9, column=4, value="Mesa operativa")
    ws.cell(row=9, column=5, value="1200 x 600")
    ws.cell(row=9, column=7, value=2)
    ws.cell(row=9, column=10, value=1000)
    wb.save(path)


def _sample_quotation_with_image(path: Path, image_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    headers = {
        1: "No.",
        2: "Item",
        4: "Description",
        5: "Dimension",
        7: "Qty",
        10: "List Price",
    }
    for col, value in headers.items():
        ws.cell(row=7, column=col, value=value)
    ws.cell(row=8, column=1, value="- Escritorios")
    ws.cell(row=9, column=1, value=1)
    ws.cell(row=9, column=2, value="Mesa Uno")
    ws.cell(row=9, column=4, value="Mesa operativa")
    ws.cell(row=9, column=5, value="1200 x 600")
    ws.cell(row=9, column=7, value=2)
    ws.cell(row=9, column=10, value=1000)
    ws.add_image(WorkbookImage(str(image_path)), "B9")
    wb.save(path)


def test_read_items_from_quotation(tmp_path):
    source = tmp_path / "quotation.xlsx"
    _sample_quotation(source)

    items, column_map = read_items(source)

    assert column_map["cantidad"] == "G"
    assert column_map["list_price"] == "J"
    assert [item.tipo for item in items] == ["categoria", "producto"]
    assert items[1].nombre == "Mesa Uno"


def test_generate_online_quote_creates_valid_xlsx(tmp_path):
    source = tmp_path / "quotation.xlsx"
    output = tmp_path / "cotizacion.xlsx"
    _sample_quotation(source)

    generate_online_quote(
        source,
        output,
        {
            "cotizacion": "COT-001",
            "proyecto": "Demo",
            "cliente": "Cliente Test",
            "descuento": "30",
            "tipo_cambio": "20",
        },
    )

    wb = load_workbook(output, data_only=False)
    assert "Cotizacion" in wb.sheetnames
    assert "Mobiliti" in wb.sheetnames
    assert "Quotation" in wb.sheetnames
    ws = wb["Cotizacion"]
    mob = wb["Mobiliti"]
    assert ws["B3"].value == "COT-001"
    assert ws["A16"].value == "Escritorios"
    assert ws["A17"].value == "=Mobiliti!D14"
    assert ws["C17"].value == "=Quotation!D9"
    assert ws["D17"].value == "=Quotation!F9"
    assert ws["E17"].value == "=Mobiliti!H14"
    assert ws["G17"].value == 0.3
    assert ws["H17"].value == "=F17*G17"
    assert ws["I17"].value == "=F17-H17"
    assert ws["J17"].value == "=E17*I17"
    assert mob["D14"].value == "=Quotation!B9"
    assert mob["H14"].value == "=Quotation!H9"
    assert mob["J14"].value == "=Quotation!K9"
    assert mob["K14"].value == "=Quotation!I9"
    assert wb["Quotation_Data"].max_column == 16
    wb.close()


def test_generate_online_quote_preserves_and_places_source_image(tmp_path):
    source = tmp_path / "quotation.xlsx"
    original = tmp_path / "original.png"
    output = tmp_path / "cotizacion.xlsx"
    Image.new("RGB", (80, 60), (20, 20, 20)).save(original)
    _sample_quotation_with_image(source, original)

    generate_online_quote(
        source,
        output,
        {
            "cotizacion": "COT-IA",
            "proyecto": "Demo IA",
            "cliente": "Cliente Test",
            "image_background": "white",
            "image_prompt": "Mejora la calidad de imagen y que este en fondo blanco",
            "tipo_cambio": "20",
        },
    )

    wb = load_workbook(output, data_only=False)
    try:
        cotizacion_images = [
            img
            for img in wb["Cotizacion"]._images
            if int(getattr(img.anchor, "_from").row) == 16 and int(getattr(img.anchor, "_from").col) == 1
        ]
        assert cotizacion_images
        anchor = getattr(cotizacion_images[0].anchor, "_from")
        assert int(anchor.rowOff) > 0
        assert int(anchor.colOff) > 0
        with Image.open(cotizacion_images[0].ref) as embedded:
            pixel = embedded.convert("RGB").getpixel((embedded.width // 2, embedded.height // 2))
        assert pixel == (20, 20, 20)
    finally:
        wb.close()


def test_generate_online_quote_uses_vol_for_mobiliti_m3(tmp_path):
    source = tmp_path / "quotation.xlsx"
    output = tmp_path / "cotizacion.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    for col, value in {1: "No.", 2: "Item", 4: "Description", 5: "Dimension", 7: "Qty", 8: "Vol", 10: "List Price"}.items():
        ws.cell(row=7, column=col, value=value)
    ws.cell(row=8, column=1, value="- Sillas")
    ws.cell(row=9, column=1, value=1)
    ws.cell(row=9, column=2, value="Silla Uno")
    ws.cell(row=9, column=4, value="Silla operativa")
    ws.cell(row=9, column=5, value="600 x 600")
    ws.cell(row=9, column=7, value=3)
    ws.cell(row=9, column=8, value=0.45)
    ws.cell(row=9, column=10, value=500)
    wb.save(source)

    generate_online_quote(source, output, {"tipo_cambio": "20"})

    out = load_workbook(output, data_only=False)
    assert out["Cotizacion"]["D17"].value == "=Quotation!F9"
    assert out["Mobiliti"]["H14"].value == "=Quotation!H9"
    assert out["Mobiliti"]["J14"].value == "=Quotation!K9"
    assert out["Mobiliti"]["K14"].value == "=Quotation!I9"
    out.close()
