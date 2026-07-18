import hashlib
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import Workbook

from mobiliti_saas.quote_engine.supplier_catalog import (
    build_supplier_cart_payload,
    load_supplier_catalog_data,
)
from mobiliti_saas.worker.catalog_sync import load_source_config
from mobiliti_saas.worker.catalog_sync.importers import alma


MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ROOT_PATH = "PROYECTOS CET - 2026/LISTAS DE PRECIOS PROVEEDORES"
IDENTITY_PATH = "SPEC Guide-Alma-KUN.xlsx"
PRICE_PATH = "SPEC GUIDES 2026/ALMA/Spec guide-Alma-KUN Design.xlsx"
MONDECASA_PATH = "SPEC Guide-Alma-Mondecasa.xlsx"


@dataclass(frozen=True)
class AdapterFile:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path


def _save_identity(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KUN DESIGN"
    headers = (
        "Código", "Imagen", "Descripción", "Dimensiones",
        "Sin cojín Aluminio: Recubrimiento en polvo",
        "Sin cojín Aluminio: Aspecto de teca",
        "Solo Cojín Calidad: Tela A Espuma Normal Ceramica A",
        "Solo Cojín Calidad: Tela A+ Espuma Normal Ceramica A+",
        "Solo Cojín Calidad: Tela A++ Espuma Normal Ceramica A++",
    )
    for column, value in enumerate(headers, 1):
        sheet.cell(6, column, value)
    sheet.merge_cells("A7:I7")
    sheet["A7"] = "COLECCION PRUEBA"
    rows = (
        ("DUP-1", "Silla corta", "60*60cm", 100),
        ("DUP-1", "Silla larga", "80*60cm", 120),
        ("GOOD-1", "Producto que deriva", "90*70cm", 150),
    )
    for row, values in enumerate(rows, 8):
        code, description, dimensions, cost = values
        sheet.cell(row, 1, code)
        sheet.cell(row, 3, description)
        sheet.cell(row, 4, dimensions)
        sheet.cell(row, 5, cost)

    pavilion = workbook.create_sheet("PAVILION ")
    for column, value in enumerate(("Picture", "Item no.", "Description", "Dimensions", "FOB Price"), 1):
        pavilion.cell(1, column, value)
    pavilion.merge_cells("A2:O2")
    pavilion["A2"] = "PAVILION"
    for row, price in enumerate((300, 400, 500), 3):
        pavilion.cell(row, 2, "AP6500")
        pavilion.cell(row, 3, f"Pavilion {row - 2}")
        pavilion.cell(row, 4, f"{row}00*300cm")
        pavilion.cell(row, 5, price)
    workbook.save(path)
    workbook.close()


def _save_prices(path: Path) -> None:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "SPEC Alma"
    costs = workbook.create_sheet("Costo Alma")
    for sheet, price_header in ((sales, "Precio Venta"), (costs, "P. Unitario.")):
        headers = ("Imagen.", "Cod.", "Descripcion.", *(price_header for _ in range(5)))
        for column, value in enumerate(headers, 1):
            sheet.cell(8, column, value)
    sales["I8"] = "Moneda."
    costs["I8"] = "Tipo de Cambio"
    costs["J8"] = "LAB Cedis"
    rows = (
        ("DUP-1", "Silla corta", 100, 666.666667),
        ("DUP-1", "Silla larga", 120, 800),
        ("BROKEN-0.3", "Identidad corrupta", 999, 6660),
    )
    for row, (code, description, cost, sale) in enumerate(rows, 9):
        sales.cell(row, 2, code)
        sales.cell(row, 3, description)
        sales.cell(row, 4, sale)
        sales.cell(row, 9, "USD")
        costs.cell(row, 2, code)
        costs.cell(row, 3, description)
        costs.cell(row, 4, cost)
        costs.cell(row, 9, 0.3)
        costs.cell(row, 10, 0.5)
    workbook.save(path)
    workbook.close()


def _save_mondecasa(path: Path) -> None:
    workbook = Workbook()
    products = workbook.active
    products.title = "MONDECASA"
    for column, value in enumerate(
        ("Cod.", "TEST", "Descripcion.", "Dimensiones", "Costo SIN Cojín"), 1
    ):
        products.cell(8, column, value)
    products["A9"] = "MON-1"
    products["C9"] = "Mesa"
    products["D9"] = "50*50cm"
    products["E9"] = 50
    pavilions = workbook.create_sheet("PAVILIONS")
    for row in (5, 10, 17):
        pavilions.cell(row, 1, "Item no.")
        pavilions.cell(row, 3, "Description")
    workbook.save(path)
    workbook.close()


def _file(path: str, brand: str, local_path: Path) -> AdapterFile:
    digest = hashlib.sha256(local_path.read_bytes()).hexdigest()
    return AdapterFile(path, "spec_guide", brand, digest, MIME, local_path)


@pytest.fixture
def source_bundle(tmp_path, monkeypatch):
    monkeypatch.setattr(alma, "_KUN_DESIGN_COUNT", 3, raising=False)
    monkeypatch.setattr(alma, "_KUN_PAVILION_COUNT", 3, raising=False)
    monkeypatch.setattr(alma, "_KUN_MIN_IDENTITY_MATCHES", 2, raising=False)
    identity = tmp_path / "identity.xlsx"
    prices = tmp_path / "prices.xlsx"
    mondecasa = tmp_path / "mondecasa.xlsx"
    _save_identity(identity)
    _save_prices(prices)
    _save_mondecasa(mondecasa)
    return (
        _file(IDENTITY_PATH, "KUN", identity),
        _file(PRICE_PATH, "KUN", prices),
        _file(MONDECASA_PATH, "Mondecasa", mondecasa),
    )


def _counted_bundle(tmp_path, design_count, pavilion_count, *, mismatched_first=False):
    identity_path = tmp_path / "identity-counted.xlsx"
    identity = Workbook()
    sheet = identity.active
    sheet.title = "KUN DESIGN"
    headers = (
        "Código", "Imagen", "Descripción", "Dimensiones",
        "Sin cojín Aluminio: Recubrimiento en polvo",
        "Sin cojín Aluminio: Aspecto de teca",
        "Solo Cojín Calidad: Tela A Espuma Normal Ceramica A",
        "Solo Cojín Calidad: Tela A+ Espuma Normal Ceramica A+",
        "Solo Cojín Calidad: Tela A++ Espuma Normal Ceramica A++",
    )
    for column, value in enumerate(headers, 1):
        sheet.cell(6, column, value)
    sheet.merge_cells("A7:I7")
    sheet["A7"] = "COLECCION CONTADA"
    row = 8
    for index in range(design_count):
        code = f"COUNT-{index:03d}"
        description = f"Producto contado {index:03d}"
        dimensions = f"{50 + index}*50cm"
        if index == 0 and mismatched_first:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
            sheet.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)
            sheet.merge_cells(start_row=row, start_column=4, end_row=row + 1, end_column=4)
            sheet.cell(row + 1, 7, 5)
            span = 2
        else:
            span = 1
        sheet.cell(row, 1, code)
        sheet.cell(row, 3, description)
        sheet.cell(row, 4, dimensions)
        sheet.cell(row, 5, 100 + index)
        row += span
    pavilion = identity.create_sheet("PAVILION ")
    for column, value in enumerate(("Picture", "Item no.", "Description", "Dimensions", "FOB Price"), 1):
        pavilion.cell(1, column, value)
    pavilion.merge_cells("A2:O2")
    pavilion["A2"] = "PAVILION"
    for index in range(pavilion_count):
        pavilion.cell(index + 3, 2, f"PAV-{index}")
        pavilion.cell(index + 3, 3, f"Pavilion contado {index}")
        pavilion.cell(index + 3, 4, f"{300 + index}*300cm")
        pavilion.cell(index + 3, 5, 300 + index)
    identity.save(identity_path)
    identity.close()

    price_path = tmp_path / "prices-counted.xlsx"
    prices = Workbook()
    sales = prices.active
    sales.title = "SPEC Alma"
    costs = prices.create_sheet("Costo Alma")
    for target, price_header in ((sales, "Precio Venta"), (costs, "P. Unitario.")):
        headers = ("Imagen.", "Cod.", "Descripcion.", *(price_header for _ in range(5)))
        for column, value in enumerate(headers, 1):
            target.cell(8, column, value)
    sales["I8"] = "Moneda."
    costs["I8"] = "Tipo de Cambio"
    costs["J8"] = "LAB Cedis"
    for index in range(design_count):
        row = index + 9
        cost = 100 + index
        for target in (sales, costs):
            target.cell(row, 2, f"COUNT-{index:03d}")
            target.cell(row, 3, f"Producto contado {index:03d}")
        sales.cell(row, 4, cost / 0.3 / 0.5)
        sales.cell(row, 9, "USD")
        costs.cell(row, 4, cost)
        costs.cell(row, 9, 0.3)
        costs.cell(row, 10, 0.5)
    prices.save(price_path)
    prices.close()

    mondecasa_path = tmp_path / "mondecasa-counted.xlsx"
    _save_mondecasa(mondecasa_path)
    return (
        _file(IDENTITY_PATH, "KUN", identity_path),
        _file(PRICE_PATH, "KUN", price_path),
        _file(MONDECASA_PATH, "Mondecasa", mondecasa_path),
    )


def _rehash(bundle, logical_path):
    row = next(item for item in bundle if item.path == logical_path)
    changed = AdapterFile(
        row.path,
        row.kind,
        row.brand,
        hashlib.sha256(row.local_path.read_bytes()).hexdigest(),
        row.mime_type,
        row.local_path,
    )
    return tuple(changed if item is row else item for item in bundle)


def test_config_uses_exact_2026_root_and_three_alma_sources():
    sources = load_source_config(Path("mobiliti_saas/worker/catalog_sync/sources.json"))
    assert {source.root_path for source in sources} == {ROOT_PATH}
    alma_source = next(source for source in sources if source.supplier == "alma")
    assert [(row.path, row.kind, row.brand) for row in alma_source.files] == [
        (IDENTITY_PATH, "spec_guide", "KUN"),
        (PRICE_PATH, "spec_guide", "KUN"),
        (MONDECASA_PATH, "spec_guide", "Mondecasa"),
    ]


@pytest.mark.parametrize(
    ("design_count", "pavilion_count"),
    ((306, 3), (308, 3), (307, 4)),
)
def test_production_contract_rejects_wrong_kun_counts(tmp_path, design_count, pavilion_count):
    with pytest.raises(ValueError, match="ALMA_KUN_COUNT"):
        alma.build_alma_snapshot(_counted_bundle(tmp_path, design_count, pavilion_count))


def test_group_row_length_mismatch_fails_closed(tmp_path):
    with pytest.raises(ValueError, match="ALMA_KUN_ALIGNMENT"):
        alma.build_alma_snapshot(
            _counted_bundle(tmp_path, 307, 3, mismatched_first=True)
        )


def test_three_derived_groups_fail_closed(source_bundle):
    price = next(row for row in source_bundle if row.path == PRICE_PATH)
    workbook = __import__("openpyxl").load_workbook(price.local_path)
    for row in range(9, 12):
        for sheet_name in ("SPEC Alma", "Costo Alma"):
            sheet = workbook[sheet_name]
            sheet.cell(row, 2, f"CORRUPTO-{row}")
            sheet.cell(row, 3, f"Identidad corrupta {row}")
        cost = 900 + row
        workbook["Costo Alma"].cell(row, 4, cost)
        workbook["SPEC Alma"].cell(row, 4, cost / 0.3 / 0.5)
    workbook.save(price.local_path)
    workbook.close()
    changed = _rehash(source_bundle, PRICE_PATH)

    with pytest.raises(ValueError, match="ALMA_KUN_RECONCILIATION"):
        alma.build_alma_snapshot(changed)


def test_wide_price_group_permutation_fails_closed(tmp_path, monkeypatch):
    monkeypatch.setattr(alma, "_KUN_DESIGN_COUNT", 9, raising=False)
    monkeypatch.setattr(alma, "_KUN_PAVILION_COUNT", 3, raising=False)
    monkeypatch.setattr(alma, "_KUN_MIN_IDENTITY_MATCHES", 0, raising=False)
    bundle = _counted_bundle(tmp_path, 9, 3)
    price = next(row for row in bundle if row.path == PRICE_PATH)
    workbook = __import__("openpyxl").load_workbook(price.local_path)
    for sheet_name in ("SPEC Alma", "Costo Alma"):
        sheet = workbook[sheet_name]
        values = [
            [sheet.cell(row, column).value for column in range(2, 11)]
            for row in range(9, 18)
        ]
        values = values[1:] + values[:1]
        for row, row_values in zip(range(9, 18), values):
            for column, value in zip(range(2, 11), row_values):
                sheet.cell(row, column, value)
    workbook.save(price.local_path)
    workbook.close()
    changed = _rehash(bundle, PRICE_PATH)

    with pytest.raises(ValueError, match="ALMA_KUN_RECONCILIATION"):
        alma.build_alma_snapshot(changed)


def test_variant_ids_do_not_change_when_product_rows_shift(source_bundle):
    before = alma.build_alma_snapshot(source_bundle)
    shifted = source_bundle
    for logical_path, sheet_names, insert_at in (
        (IDENTITY_PATH, ("KUN DESIGN",), 8),
        (PRICE_PATH, ("SPEC Alma", "Costo Alma"), 9),
    ):
        row = next(item for item in shifted if item.path == logical_path)
        workbook = __import__("openpyxl").load_workbook(row.local_path)
        for sheet_name in sheet_names:
            workbook[sheet_name].insert_rows(insert_at, 5)
        workbook.save(row.local_path)
        workbook.close()
        shifted = _rehash(shifted, logical_path)
    after = alma.build_alma_snapshot(shifted)

    def commerce_ids(snapshot):
        return {
            (item["brand"], item["attributes"]["source_code"], item["description"]):
            (item["internal_id"], item["sku"])
            for item in snapshot["items"]
            if item["brand"] == "KUN"
        }

    assert commerce_ids(after) == commerce_ids(before)


def test_kun_variants_are_quotable_and_public_code_survives_cart(source_bundle):
    snapshot = alma.build_alma_snapshot(source_bundle)
    items = load_supplier_catalog_data(snapshot, expected_supplier="alma")["items"]
    kun = [item for item in items if item["brand"] == "KUN"]
    variants = [item for item in kun if item["attributes"]["source_code"] == "DUP-1"]

    assert len(kun) == 6
    assert len(variants) == 2
    assert len({item["sku"] for item in kun}) == len(kun)
    assert len({item["internal_id"] for item in kun}) == len(kun)
    assert all(item["code_status"] == "verified" for item in variants)

    line = build_supplier_cart_payload(
        [{
            "internal_id": variants[0]["internal_id"],
            "quantity": "1",
            "base_option_id": variants[0]["base_price_options"][0]["id"],
            "add_on_option_ids": [],
        }],
        snapshot,
        "USD",
        [],
    )["items"][0]
    assert line["sku"] == "DUP-1"


def test_corrupt_current_identity_is_never_public_and_inequivalent_price_is_derived(source_bundle):
    snapshot = alma.build_alma_snapshot(source_bundle)
    item = next(
        row for row in snapshot["items"]
        if row["attributes"].get("source_code") == "GOOD-1"
    )

    assert item["name"] == "Producto que deriva"
    assert item["base_price_options"][0]["price_net"] == "1000.000000"
    assert item["attributes"]["price_reconciliation"]["method"] == "derived_from_identity_cost"
    assert any("deriv" in warning.casefold() for warning in item["warnings"])
    assert "BROKEN-0.3" not in str(snapshot)


def test_pavilion_rows_keep_direct_identity_prices_without_factor(source_bundle):
    snapshot = alma.build_alma_snapshot(source_bundle)
    pavilions = [
        row for row in snapshot["items"]
        if row["brand"] == "KUN" and row["collection"] == "PAVILION"
    ]

    assert sorted(row["price_net"] for row in pavilions) == ["300.000000", "400.000000", "500.000000"]
    assert all(
        row["attributes"]["price_reconciliation"]["method"] == "identity_direct"
        for row in pavilions
    )


def test_real_2026_cache_has_310_kun_rows_with_ambiguous_addons_fail_closed():
    root = Path(".cache/catalog_sources/alma/sharepoint_2026-07-17")
    paths = (
        (IDENTITY_PATH, "KUN", root / "SPEC Guide-Alma-KUN.root.xlsx"),
        (PRICE_PATH, "KUN", root / "Spec guide-Alma-KUN Design.current.xlsx"),
        (MONDECASA_PATH, "Mondecasa", root / "SPEC Guide-Alma-Mondecasa.current.xlsx"),
    )
    if any(not local_path.exists() for _, _, local_path in paths):
        pytest.skip("Fuentes ALMA 2026 ignoradas no disponibles")
    snapshot = alma.build_alma_snapshot(
        tuple(_file(path, brand, local_path) for path, brand, local_path in paths)
    )
    items = load_supplier_catalog_data(snapshot, expected_supplier="alma")["items"]
    kun = [item for item in items if item["brand"] == "KUN"]
    nonempty_skus = [item["sku"] for item in items if item["sku"]]

    assert len(kun) == 310
    assert sum(item["collection"] == "PAVILION" for item in kun) == 3
    assert sum(item["code_status"] == "verified" and bool(item["sku"]) for item in kun) == 262
    assert sum(item["code_status"] == "needs_review" and not item["sku"] for item in kun) == 48
    assert len(set(nonempty_skus)) == len(nonempty_skus)
    assert len({item["internal_id"] for item in items}) == len(items)
    assert sum(
        item["attributes"]["price_reconciliation"]["method"]
        == "derived_from_identity_cost"
        for item in kun
    ) == 2
