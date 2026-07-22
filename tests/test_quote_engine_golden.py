from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import os
import sys

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mobiliti_saas.quote_engine import generate_quote  # noqa: E402
from mobiliti_saas.quote_engine.engine import SECTION_PROD_STARTS, _copy_source_sheet, _write_estrategia_comercial  # noqa: E402
from mobiliti_saas.quote_engine.mixed_catalog import (  # noqa: E402
    build_mixed_catalog_cart_payload,
    create_mixed_catalog_quotation_workbook,
)
from mobiliti_saas.quote_engine.supplier_catalog import create_supplier_quotation_workbook  # noqa: E402
from mobiliti_saas.quote_engine.tarkett_catalog import TarkettCatalogItem  # noqa: E402


DOWNLOADS = Path(r"C:\Users\pepem\Downloads")
TEMPLATE_DIR = ROOT / "versiones historial" / "HISTORIAL DE VERSIONES" / "Mobiliti_Generador_Windows"
TEMPLATE = next(TEMPLATE_DIR.glob("Formato*.xlsx"), TEMPLATE_DIR / "Formato Cotizacion 2026 GDL (1).xlsx")
WORKER_TEMPLATE = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 Oficial.xlsx"
)
GOLDENS = [
    DOWNLOADS / "IZA REFORMA-Quotation Sheet - V1.xlsx",
    ROOT / "versiones historial" / "KIVO BRAVANTE-Quotation Sheet - V1.xlsx",
]
TARGET_VISUAL = next(
    (path for path in DOWNLOADS.glob("*.xlsx") if "ESSENTIA EVIDIKA" in path.name and not path.name.startswith("~$")),
    DOWNLOADS / "Cotizacion ESSENTIA EVIDIKA.xlsx",
)
IZA_CORRECTA = DOWNLOADS / "Cotizacion_IZA_REFORMA.xlsx"
LEAC_QUOTATION = DOWNLOADS / "LEAC" / "LEAC- GOTAPP - MORGINS-Quotation Sheet - V1.xlsx"
EXPECTED_TEMPLATE_SHEETS = {
    "Cotizacion",
    "Mobiliti",
    "Estrategia Comercial ",
    "Fletes",
    "Proveedores",
    "SPEC LAMINADO JOME",
    "SPEC-GUIDE-LUMBRO",
    "SPEC-GUIDE ESTRUCTURAS",
    "Spec Guide Estructura ",
    "SPEC-GUIDE-CR GLOBAL",
    "Meses Sin Intereses Tarjetas",
    "Quotation",
}
MONEY = Decimal("0.01")


def money(value: Decimal) -> Decimal:
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def reference_totals(rows: list[tuple[Decimal, Decimal, Decimal]]):
    subtotal = Decimal("0")
    for price, quantity, discount in rows:
        unit = money(price)
        discount_amount = money(unit * discount)
        net_unit = money(unit - discount_amount)
        subtotal += money(quantity * net_unit)
    subtotal = money(subtotal)
    freight = money(subtotal * Decimal("0.12"))
    before_tax = money(subtotal + freight)
    tax = money(before_tax * Decimal("0.16"))
    return subtotal, freight, before_tax, tax, money(before_tax + tax)


def _supplier_line(**overrides):
    line = {
        "internal_id": "alma:kun:kc8611n01rop",
        "supplier": "alma",
        "product_key": "kun-kc8611n01rop",
        "sku": "KC8611N01ROP",
        "code_status": "verified",
        "brand": "KUN",
        "collection": "CLOGS",
        "name": "Silla KUN",
        "description": "Silla configurable",
        "unit": "pieza",
        "availability_type": "made_to_order",
        "stock": None,
        "lead_time": "Sobre pedido",
        "quantity": "1",
        "base_option_id": "powder-coated-aluminium",
        "add_on_option_ids": ["cushion:a-plus"],
        "configuration": "Aluminio; Cojin A+",
        "base_currency": "USD",
        "base_price": "250.000000",
        "unit_price_base": "285.100000",
        "unit_price": "5274.35",
        "line_total": "5274.35",
        "tax_rate": "0.160000",
        "attributes": {"color": "verde"},
        "image_url": "",
        "image_kind": "official",
        "product_url": "https://example.test/kun",
        "warnings": [],
        "source_reference": "SPEC Guide-Alma-KUN.xlsx:E9:I9",
    }
    line.update(overrides)
    return line


def _supplier_payload(items):
    return {
        "source_type": "supplier_cart",
        "supplier": "alma",
        "catalog_source_hash": "a" * 64,
        "base_currency": "USD",
        "quote_currency": "MXN",
        "exchange_rate": "18.500000",
        "rate_source": "saas_exchange_rates",
        "rate_effective_date": "2026-07-15",
        "rate_retrieved_at": "2026-07-15T23:00:00Z",
        "items": items,
    }


def _write_minimal_quotation(path: Path, *, unit_price=123.456) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    for column, value in {1: "No.", 2: "Item", 4: "Description", 5: "Dimension", 7: "Qty", 10: "List Price", 11: "URL"}.items():
        ws.cell(7, column).value = value
    ws["A8"] = "- Catalogo proveedor"
    ws["A9"] = 1
    ws["B9"] = "Producto de prueba"
    ws["D9"] = "Descripcion configurable | SKU: TEST-001"
    ws["E9"] = "pieza"
    ws["G9"] = 2
    ws["J9"] = unit_price
    ws["K9"] = "https://example.test/producto"
    wb.save(path)
    wb.close()


def _write_mixed_totals_quotation(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    headers = {
        1: "No.", 2: "Item", 4: "Description", 5: "Dimension", 7: "Qty",
        10: "List Price", 11: "URL", 12: "Supplier", 13: "Discount Percent",
        14: "Original Currency", 15: "Original Unit Price", 16: "Frozen Exchange Rate",
        17: "Source Reference", 18: "Price Mode", 19: "Auto Electrification",
    }
    for column, value in headers.items():
        ws.cell(7, column).value = value
    ws["A8"] = "- Catalogos mixtos"
    rows = [
        (9, "Piso Tarkett", 2, 123.46, "Tarkett", 40, "MXN", 123.456, 1, "list"),
        (10, "Silla ALMA", 3, 200.13, "ALMA", 0, "USD", 10.817568, 18.5, "net"),
    ]
    for row, name, quantity, price, provider, discount, original_currency, original_price, rate, mode in rows:
        ws.cell(row, 1).value = row - 8
        ws.cell(row, 2).value = name
        ws.cell(row, 4).value = f"Descripcion {name}"
        ws.cell(row, 5).value = "pieza"
        ws.cell(row, 7).value = quantity
        ws.cell(row, 10).value = price
        ws.cell(row, 12).value = provider
        ws.cell(row, 13).value = discount
        ws.cell(row, 14).value = original_currency
        ws.cell(row, 15).value = original_price
        ws.cell(row, 16).value = rate
        ws.cell(row, 17).value = f"source:{provider.lower()}"
        ws.cell(row, 18).value = mode
        ws.cell(row, 19).value = False
    wb.save(path)
    wb.close()


def _mixed_totals_metadata():
    return {
        "catalog_price_mode": "mixed_catalog_converted",
        "quote_currency": "MXN",
        "auto_electrification_rate": None,
        "rate_summary": [
            {
                "catalog": "tarkett", "base_currency": "MXN", "quote_currency": "MXN",
                "exchange_rate": "1.000000", "rate_source": "identity",
                "rate_effective_date": "2026-07-15", "rate_retrieved_at": "",
            },
            {
                "catalog": "alma", "base_currency": "USD", "quote_currency": "MXN",
                "exchange_rate": "18.500000", "rate_source": "saas_exchange_rates",
                "rate_effective_date": "2026-07-15",
                "rate_retrieved_at": "2026-07-15T23:00:00Z",
            },
        ],
        "cotizacion": "MIXTA-GOLDEN",
    }


def _build_real_mixed_task5_payload():
    tarkett = TarkettCatalogItem(
        "25731726",
        "Piso Tarkett",
        "m2",
        Decimal("10"),
        unit_price=Decimal("123.456"),
        price_source="catalog",
    )
    alma = {
        "internal_id": "alma:kun:configured",
        "supplier": "alma",
        "product_key": "kun-configured",
        "sku": "KC8611N01ROP",
        "code_status": "verified",
        "brand": "KUN",
        "collection": "KUN",
        "name": "Silla KUN configurable",
        "description": "Silla ejecutiva configurable",
        "unit": "pieza",
        "availability_type": "made_to_order",
        "stock": None,
        "lead_time": "Sobre pedido",
        "base_price_options": [
            {
                "id": "powder-coated-aluminium",
                "name": "Powder coated aluminium",
                "price_net": "250.000000",
                "available": True,
            }
        ],
        "add_on_options": [
            {
                "id": "cushion:a-plus",
                "name": "Cushion A+",
                "family": "cushion",
                "price_net": "35.100000",
                "available": True,
            }
        ],
        "base_currency": "USD",
        "price_net": "199.990000",
        "tax_rate": "0.160000",
        "attributes": {"color": "Black", "dimensions": "65 x 45 x 83 cm"},
        "image_url": "",
        "image_kind": "generated_reference",
        "product_url": "",
        "warnings": ["Imagen ilustrativa"],
        "source_reference": "SPEC Guide-Alma-KUN.xlsx:E9:I9",
    }
    return build_mixed_catalog_cart_payload(
        [
            {"catalog": "tarkett", "code": tarkett.code, "quantity": "2"},
            {
                "catalog": "alma",
                "internal_id": alma["internal_id"],
                "quantity": "1",
                "base_option_id": "powder-coated-aluminium",
                "add_on_option_ids": ["cushion:a-plus"],
            },
        ],
        catalogs={
            "tarkett": {
                "source_hash": "a" * 64,
                "items": [tarkett],
                "by_code": {tarkett.code: tarkett},
            },
            "alma": {
                "supplier": "alma",
                "source_hash": "b" * 64,
                "generated_at": "2026-07-19T00:00:00+00:00",
                "items": [alma],
            },
        },
        rate_rows=[
            {
                "currency": "USD",
                "effective_date": "2026-07-19",
                "mxn_per_unit": "18.500000",
                "retrieved_at": "2026-07-19T12:00:00+00:00",
            }
        ],
        quote_currency="MXN",
        commercial_discount_percent="40",
        today=date(2026, 7, 19),
    )


def _formula_uses_round_2(value):
    text = str(value or "").upper().replace(" ", "")
    return text.startswith("=") and "ROUND(" in text and ",2)" in text


def test_supplier_intermediate_workbook_contains_configuration_and_stock_warnings(tmp_path):
    output = tmp_path / "supplier-quotation.xlsx"
    payload = _supplier_payload(
        [
            _supplier_line(
                code_status="needs_review",
                sku="",
                image_kind="generated_reference",
                quantity="2",
            ),
            _supplier_line(
                internal_id="cr-global:out",
                supplier="cr-global",
                product_key="crg-out",
                sku="CRG-OUT",
                name="Producto agotado",
                availability_type="stocked",
                stock="0.000000",
                lead_time="Entrega inmediata",
                configuration="Standard",
                quantity="1",
            ),
            _supplier_line(
                internal_id="cr-global:low",
                supplier="cr-global",
                product_key="crg-low",
                sku="CRG-LOW",
                name="Producto insuficiente",
                availability_type="stocked",
                stock="1.000000",
                available_after_reservations="0.500000",
                lead_time="Entrega inmediata",
                configuration="Standard",
                quantity="2",
            ),
        ]
    )

    create_supplier_quotation_workbook(payload, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["Quotation"]
    first_description = str(ws["D9"].value)
    assert ws["B9"].value == "Silla KUN"
    assert "Silla configurable" in first_description
    assert "Aluminio; Cojin A+" in first_description
    assert "Codigo por verificar" in first_description
    assert "Imagen de referencia" in first_description
    assert "Sobre pedido" in first_description
    assert "AGOTADO" not in first_description.upper()
    assert ws["J9"].value == 5274.35
    assert ws["G9"].value == 2
    assert ws["K9"].value == "https://example.test/kun"
    assert ws["D9"].fill.fgColor.rgb.endswith("FFF2CC")
    assert "SKU: CRG-OUT" in str(ws["D10"].value)
    assert "AGOTADO" in str(ws["D10"].value).upper()
    assert ws["D10"].fill.fgColor.rgb.endswith("FFF2CC")
    assert "INSUFICIENTE" in str(ws["D11"].value).upper()
    assert "DISPONIBLE 0.5" in str(ws["D11"].value).upper()
    assert ws["D11"].fill.fgColor.rgb.endswith("FFF2CC")
    wb.close()


def test_supplier_final_workbook_uses_official_template_and_frozen_price_contract(tmp_path):
    source = tmp_path / "supplier-source.xlsx"
    output = tmp_path / "supplier-final.xlsx"
    _write_minimal_quotation(source)
    metadata = {
        "cotizacion": "SUP-001",
        "proyecto": "Catalogo ALMA",
        "cliente": "Cliente",
        "descuento": 40,
        "catalog_supplier": "alma",
        "catalog_supplier_label": "ALMA",
        "catalog_price_mode": "list_price_net",
        "base_currency": "USD",
        "quote_currency": "MXN",
        "exchange_rate": "18.500000",
        "rate_source": "saas_exchange_rates",
        "rate_effective_date": "2026-07-15",
        "rate_retrieved_at": "2026-07-15T23:00:00Z",
    }

    generate_quote(source, output, metadata, WORKER_TEMPLATE)

    wb = load_workbook(output, data_only=False)
    cot = wb["Cotizacion"]
    mobiliti = wb["Mobiliti"]
    # El paquete oficial, no el writer legacy, gobierna B4, K4/K6 y las
    # fórmulas visibles; J contiene el único costo congelado por producto.
    assert cot["B4"].value is None
    mobiliti_row = next(
        row
        for row in range(1, mobiliti.max_row + 1)
        if mobiliti.cell(row, 4).value == "Producto de prueba"
    )
    product_row = next(
        row
        for row in range(1, cot.max_row + 1)
        if cot.cell(row, 1).value == "Producto de prueba"
    )
    assert mobiliti.cell(mobiliti_row, 6).value == "ALMA"
    assert mobiliti.cell(mobiliti_row, 10).value == pytest.approx(123.46)
    assert mobiliti["K4"].value is False
    assert getattr(mobiliti["K6"].value, "text", None) == '=_FV(J6,"High")'
    assert cot.cell(product_row, 7).value == 0.4
    assert cot.cell(product_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
    assert cot.cell(product_row, 8).value == f"=F{product_row}*G{product_row}"
    assert cot.cell(product_row, 9).value == f"=F{product_row}-H{product_row}"
    assert cot.cell(product_row, 10).value == f"=E{product_row}*I{product_row}"
    total_row = next(
        row
        for row in range(product_row + 1, cot.max_row + 1)
        if str(cot.cell(row, 4).value or "").strip() == "TOTAL:"
    )
    assert "16%" in str(cot.cell(total_row - 1, 8).value)
    assert wb["Quotation_Data"].sheet_state == "veryHidden"
    wb.close()


def test_mixed_final_workbook_uses_official_formulas_and_one_frozen_cost_per_item(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine._exchange_rate",
        lambda _metadata: (_ for _ in ()).throw(
            AssertionError("mixed mode must not use the legacy exchange rate")
        ),
    )
    source = tmp_path / "mixed-totals.xlsx"
    output = tmp_path / "mixed-totals-final.xlsx"
    _write_mixed_totals_quotation(source)

    generate_quote(source, output, _mixed_totals_metadata(), WORKER_TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        assert {"Cotizacion", "Mobiliti", "Quotation"} <= set(wb.sheetnames)
        assert wb.sheetnames.count("Cotizacion") == 1
        assert wb.sheetnames.count("Mobiliti") == 1
        assert wb.sheetnames.count("Quotation") == 1
        cot = wb["Cotizacion"]
        mobiliti = wb["Mobiliti"]
        # Las identidades se materializan como texto y Cotizacion referencia X;
        # ya no se prueban fórmulas ROUND ni enlaces directos a Quotation.
        product_row_map = []
        for source_row in (9, 10):
            product_name = wb["Quotation"].cell(source_row, 2).value
            cot_row = next(
                row for row in range(1, cot.max_row + 1)
                if cot.cell(row, 1).value == product_name
            )
            mobiliti_row = next(
                row for row in range(1, mobiliti.max_row + 1)
                if mobiliti.cell(row, 4).value == product_name
            )
            product_row_map.append((cot_row, mobiliti_row))

        first_product = product_row_map[0][0]
        for cot_row, mobiliti_row in product_row_map:
            assert cot.cell(cot_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
            assert cot.cell(cot_row, 8).value == f"=F{cot_row}*G{cot_row}"
            assert cot.cell(cot_row, 9).value == f"=F{cot_row}-H{cot_row}"
            assert cot.cell(cot_row, 10).value == f"=E{cot_row}*I{cot_row}"
        assert cot.cell(first_product, 7).value == 0.4
        assert cot.cell(product_row_map[1][0], 7).value == f"=$G${first_product}"
        assert mobiliti.cell(product_row_map[0][1], 10).value == pytest.approx(123.46)
        assert mobiliti.cell(product_row_map[1][1], 10).value == pytest.approx(200.13)
        assert wb["Quotation_Data"].sheet_state == "veryHidden"
        assert reference_totals(
            [
                (Decimal("123.456"), Decimal("2"), Decimal("0.4")),
                (Decimal("200.13"), Decimal("3"), Decimal("0")),
            ]
        ) == (
            Decimal("748.55"),
            Decimal("89.83"),
            Decimal("838.38"),
            Decimal("134.14"),
            Decimal("972.52"),
        )
    finally:
        wb.close()


def test_real_task5_mixed_workbook_preserves_structured_description_and_identity_price(
    tmp_path, monkeypatch
):
    payload = _build_real_mixed_task5_payload()
    source = tmp_path / "real-task5-mixed.xlsx"
    output = tmp_path / "real-task6-final.xlsx"
    create_mixed_catalog_quotation_workbook(
        payload,
        source,
        image_dir=tmp_path / "mixed-images",
    )
    metadata = {
        "catalog_price_mode": "mixed_catalog_converted",
        "quote_currency": payload["quote_currency"],
        "rate_summary": payload["rate_summary"],
        "auto_electrification_rate": payload["auto_electrification_rate"],
        "cotizacion": "MIXTA-REAL",
    }
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine._exchange_rate",
        lambda _metadata: (_ for _ in ()).throw(
            AssertionError("mixed mode must not use the legacy exchange rate")
        ),
    )

    generate_quote(source, output, metadata, WORKER_TEMPLATE)

    wb = load_workbook(output, data_only=False)
    try:
        cot = wb["Cotizacion"]
        mobiliti = wb["Mobiliti"]
        quotation = wb["Quotation"]
        alma_source_row = next(
            row
            for row in range(8, quotation.max_row + 1)
            if quotation.cell(row, 12).value == "ALMA"
        )
        alma_name = quotation.cell(alma_source_row, 2).value
        cot_row = next(
            row
            for row in range(1, cot.max_row + 1)
            if cot.cell(row, 1).value == alma_name
        )
        mobiliti_row = next(
            row
            for row in range(1, mobiliti.max_row + 1)
            if mobiliti.cell(row, 4).value == alma_name
        )
        description = str(cot.cell(cot_row, 3).value)
        for expected in (
            "Silla ejecutiva configurable",
            "Powder coated aluminium",
            "Cushion A+",
            "Sobre pedido",
            "Imagen de referencia",
        ):
            assert expected in description
        # El contrato Task 7 reemplaza W/X legacy por fórmula oficial y deja J
        # como valor numérico congelado, incluso para un renglón importado.
        assert mobiliti.cell(mobiliti_row, 10).value == pytest.approx(
            float(quotation.cell(alma_source_row, 10).value)
        )
        assert str(mobiliti.cell(mobiliti_row, 23).value).startswith("=IF(")
        assert str(mobiliti.cell(mobiliti_row, 24).value).startswith("=_xlfn.MINIFS(")
        assert cot.cell(cot_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
        assert wb["Quotation_Data"].sheet_state == "veryHidden"
    finally:
        wb.close()


def test_standard_workbook_keeps_official_provider_header_and_formulas(tmp_path):
    source = tmp_path / "legacy-source.xlsx"
    output = tmp_path / "legacy-final.xlsx"
    _write_minimal_quotation(source, unit_price=100)

    generate_quote(source, output, {"cotizacion": "LEGACY", "tipo_cambio": 20, "descuento": 40}, WORKER_TEMPLATE)

    wb = load_workbook(output, data_only=False)
    cot = wb["Cotizacion"]
    mobiliti = wb["Mobiliti"]
    mobiliti_row = next(
        row
        for row in range(1, mobiliti.max_row + 1)
        if mobiliti.cell(row, 4).value == "Producto de prueba"
    )
    product_row = next(
        row
        for row in range(1, cot.max_row + 1)
        if cot.cell(row, 1).value == "Producto de prueba"
    )
    # Las filas vacías conservan las fórmulas nativas del template oficial; no
    # se reinstalan los guards sintéticos producidos por el writer anterior.
    assert cot["B4"].value is None
    assert mobiliti.cell(mobiliti_row, 6).value == "Sunon Inc"
    assert cot.cell(product_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
    assert cot.cell(product_row, 8).value == f"=F{product_row}*G{product_row}"
    assert cot.cell(product_row, 9).value == f"=F{product_row}-H{product_row}"
    assert cot.cell(product_row, 10).value == f"=E{product_row}*I{product_row}"
    unused_row = mobiliti_row + 1
    assert all(
        mobiliti.cell(unused_row, column).value is None
        for column in (4, 5, 6, 8, 10, 11, 16)
    )
    assert str(mobiliti.cell(unused_row, 23).value).startswith("=IF(")
    assert str(mobiliti.cell(unused_row, 24).value).startswith("=_xlfn.MINIFS(")
    assert mobiliti["K4"].value is False
    assert getattr(mobiliti["K6"].value, "text", None) == '=_FV(J6,"High")'
    assert wb["Quotation_Data"].sheet_state == "veryHidden"
    wb.close()


@pytest.mark.parametrize("source", GOLDENS)
def test_python_engine_generates_golden_structure(source, tmp_path):
    if not source.exists() or not TEMPLATE.exists():
        pytest.skip("Golden input/template not available on this machine")

    output = tmp_path / f"{source.stem}_python.xlsx"
    generate_quote(
        source,
        output,
        {"cotizacion": "GOLDEN", "proyecto": "Golden", "cliente": "Cliente", "tipo_cambio": 20},
        TEMPLATE,
    )

    wb = load_workbook(output, data_only=False)
    assert "Cotizacion" in wb.sheetnames
    assert "Mobiliti" in wb.sheetnames
    assert "Quotation" in wb.sheetnames
    assert wb["Cotizacion"]["D17"].value == "=Quotation!E9"
    assert str(wb["Mobiliti"]["K14"].value).startswith("=Quotation!")
    assert wb["Cotizacion"].print_area
    assert len(wb["Cotizacion"]._images) > 0
    wb.close()


def test_copy_source_sheet_handles_leac_external_styles(tmp_path):
    if os.environ.get("RUN_SLOW_QUOTE_TESTS") != "1":
        pytest.skip("LEAC copy is a slow local regression test")
    if not LEAC_QUOTATION.exists():
        pytest.skip("LEAC quotation fixture not available on this machine")

    workbook = Workbook()
    _copy_source_sheet(LEAC_QUOTATION, workbook)
    output = tmp_path / "leac-copy.xlsx"
    workbook.save(output)
    workbook.close()

    copied = load_workbook(output)
    assert "Quotation" in copied.sheetnames
    assert len(copied["Quotation"].merged_cells.ranges) > 0
    copied.close()


def test_visual_golden_references_are_intact():
    if not TARGET_VISUAL.exists() or not IZA_CORRECTA.exists():
        pytest.skip("Visual golden files not available on this machine")

    for path in [TARGET_VISUAL, IZA_CORRECTA]:
        wb = load_workbook(path, data_only=False)
        assert EXPECTED_TEMPLATE_SHEETS.issubset(set(wb.sheetnames))
        cot = wb["Cotizacion"]
        assert cot.max_column >= 53
        assert cot.print_area
        assert len(cot._images) >= 30
        assert len(cot.merged_cells.ranges) >= 65
        assert cot["B4"].value is None
        assert cot["D17"].value == "=Quotation!E9"
        assert str(cot["F17"].value).startswith("=Mobiliti!W")
        assert (cot.row_dimensions[17].height or 0) >= 300
        wb.close()


def test_python_engine_does_not_fallback_to_blank_workbook(tmp_path):
    source = DOWNLOADS / "IZA REFORMA-Quotation Sheet - V1.xlsx"
    if not source.exists():
        pytest.skip("IZA input not available on this machine")

    with pytest.raises(FileNotFoundError):
        generate_quote(source, tmp_path / "out.xlsx", {}, tmp_path / "missing-template.xlsx")


def test_python_engine_passes_image_provider_metadata(monkeypatch, tmp_path):
    source = DOWNLOADS / "IZA REFORMA-Quotation Sheet - V1.xlsx"
    if not source.exists() or not TEMPLATE.exists():
        pytest.skip("Golden input/template not available on this machine")

    captured = {}

    def fake_improve_image_map(image_map, temp_dir, **kwargs):
        captured.update(kwargs)
        return {}

    monkeypatch.setattr("mobiliti_saas.quote_engine.engine.improve_image_map", fake_improve_image_map)
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine._generate_missing_dezgo_images",
        lambda image_map, items, temp_dir, metadata, stats=None: image_map,
    )

    generate_quote(
        source,
        tmp_path / "out.xlsx",
        {"cotizacion": "GOLDEN", "image_provider": "dezgo", "tipo_cambio": 20},
        TEMPLATE,
    )

    assert captured["image_provider"] == "dezgo"


def test_estrategia_comercial_references_generated_total_row():
    if not TEMPLATE.exists():
        pytest.skip("Template not available on this machine")

    wb = load_workbook(TEMPLATE, data_only=False)
    if "Estrategia Comercial " not in wb.sheetnames:
        wb.close()
        pytest.skip("Template with Estrategia Comercial not available on this machine")

    ws = wb["Estrategia Comercial "]
    _write_estrategia_comercial(ws, total_row=74)

    assert ws["D59"].value == "=Cotizacion!H74"
    wb.close()


def test_python_engine_preserves_workbook_contract(tmp_path):
    source = DOWNLOADS / "IZA REFORMA-Quotation Sheet - V1.xlsx"
    if not source.exists() or not TEMPLATE.exists():
        pytest.skip("Golden input/template not available on this machine")

    output = tmp_path / "iza_python_contract.xlsx"
    generate_quote(
        source,
        output,
        {"cotizacion": "GOLDEN", "proyecto": "Golden", "cliente": "Cliente", "tipo_cambio": 20},
        TEMPLATE,
    )

    wb = load_workbook(output, data_only=False)
    assert EXPECTED_TEMPLATE_SHEETS.issubset(set(wb.sheetnames))
    cot = wb["Cotizacion"]
    assert cot.max_column >= 53
    assert cot.print_area
    assert len(cot._images) >= 30
    assert len(cot.merged_cells.ranges) >= 50
    assert cot["B4"].value is None
    assert cot["A16"].value == "=Quotation!A8"
    assert cot["D17"].value == "=Quotation!E9"
    assert str(cot["F17"].value).startswith(("=Mobiliti!X", "=IFERROR((Mobiliti!X"))
    assert (cot.row_dimensions[19].height or 0) >= 300
    merged_ranges = {str(rng) for rng in cot.merged_cells.ranges}
    assert "D58:G58" in merged_ranges
    assert "H58:J58" in merged_ranges
    for row in range(58, 63):
        assert cot[f"H{row}"].alignment.horizontal == "right"
    assert (cot.row_dimensions[58].height or 0) < 50
    assert (cot.row_dimensions[59].height or 0) < 50
    assert cot["C20"].alignment.vertical == "top"
    assert cot["C32"].alignment.vertical == "top"
    assert any(
        isinstance(cot.cell(row, 6).value, str) and "+Mobiliti!Y" in cot.cell(row, 6).value
        for row in range(1, cot.max_row + 1)
    )
    assert any(
        getattr(getattr(img, "anchor", None), "_from", None) is not None
        and int(img.anchor._from.col) == 2
        for img in cot._images
    )
    assert cot["F32"].value == (
        "=IFERROR((Mobiliti!X155*Mobiliti!H155+Mobiliti!Y156+Mobiliti!Y157+Mobiliti!Y158)/Mobiliti!H155,0)"
    )
    product_rows = [
        row
        for row in range(1, cot.max_row + 1)
        if str(cot.cell(row, 4).value or "").startswith("=Quotation!E")
    ]
    assert product_rows
    first_product_row = product_rows[0]
    assert cot.cell(first_product_row, 7).value == 0.4
    assert cot.cell(first_product_row, 8).value == f"=F{first_product_row}*G{first_product_row}"
    assert cot.cell(first_product_row, 9).value == f"=F{first_product_row}-H{first_product_row}"
    assert cot.cell(first_product_row, 10).value == f"=E{first_product_row}*I{first_product_row}"
    for row in product_rows[1:]:
        assert cot.cell(row, 7).value == f"=G${first_product_row}"
        assert cot.cell(row, 8).value == f"=F{row}*G{row}"
        assert cot.cell(row, 9).value == f"=F{row}-H{row}"
        assert cot.cell(row, 10).value == f"=E{row}*I{row}"
    assert not any(
        isinstance(cot.cell(row, 10).value, str) and "Mobiliti!AD" in cot.cell(row, 10).value
        for row in product_rows
    )
    mob = wb["Mobiliti"]
    assert mob["J6"].value == "USD/MXN"
    assert mob["K6"].value == 20
    assert mob["K8"].value == "Guadalajara"
    mob_merged_ranges = {str(rng) for rng in mob.merged_cells.ranges}
    assert "D13:J13" in mob_merged_ranges
    assert "A47:F47" in mob_merged_ranges
    assert "A48:J48" in mob_merged_ranges
    assert str(mob["D13"].value).startswith("Sección 1 - ")
    assert str(mob["A48"].value).startswith("Sección 2 - ")
    estrategia = wb["Estrategia Comercial "]
    total_rows = [
        row
        for row in range(1, cot.max_row + 1)
        if str(cot.cell(row, 4).value or "").strip().upper() == "TOTAL:"
    ]
    assert total_rows
    assert estrategia["D59"].value == f"=Cotizacion!H{total_rows[-1]}"
    fletes = wb["Fletes"]
    assert fletes["A5"].value == "Guadalajara"
    assert fletes["C5"].value == "Monterrey"
    assert fletes["A17"].value == "Guadalajara"
    assert fletes["C17"].value == "Guadalajara"
    assert fletes["I8"].value == "Escritorios-WorkStation"
    assert fletes["M8"].value == "Escritorios-WorkStation"
    template_wb = load_workbook(TEMPLATE, data_only=False, keep_links=False)
    template_fletes = template_wb["Fletes"]
    for cell in ["I16", "M16", "I17", "M17", "I18", "M18"]:
        assert fletes[cell].value == template_fletes[cell].value
    template_wb.close()
    provider_validations = [
        dv
        for dv in mob.data_validations.dataValidation
        if dv.type == "list" and dv.formula1 == "=Proveedores!$A$2:$A$32"
    ]
    assert provider_validations
    assert "F14:F45" in str(provider_validations[0].sqref)
    region_validations = [
        dv
        for dv in mob.data_validations.dataValidation
        if dv.type == "list" and dv.formula1 == "Taba_Region"
    ]
    assert region_validations
    assert "P14:P45" in str(region_validations[0].sqref)
    populated_rows = [
        row
        for start in [14, 49, 84, 119, 154, 189, 224, 259, 294, 329, 364, 398, 432]
        for row in range(start, start + 32)
        if any(mob.cell(row, col).value for col in (4, 5, 6))
    ]
    assert populated_rows
    for row in populated_rows:
        assert mob.cell(row, 16).value == "Centro"
        assert mob.cell(row, 27).value == f"=MIN(0.4,Z{row})"
        assert "ERROR" not in str(mob.cell(row, 29).value or "")
        assert str(mob.cell(row, 30).value or "") == f"=AC{row}*H{row}"
    blank_formula_errors = []
    for start in [14, 49, 84, 119, 154, 189, 224, 259, 294, 329, 364, 398, 432]:
        for row in range(start, start + 32):
            if any(mob.cell(row, col).value for col in (4, 5, 6)):
                continue
            values = [str(mob.cell(row, col).value or "") for col in range(11, 36)]
            if any("#REF!" in value or "ERROR" in value for value in values):
                blank_formula_errors.append(row)
    assert blank_formula_errors == []
    assert wb["Cotizacion"]["G17"].value == 0.4
    assert wb["Cotizacion"]["H17"].value == "=F17*G17"
    assert wb["Cotizacion"]["I17"].value == "=F17-H17"
    assert wb["Cotizacion"]["J17"].value == "=E17*I17"
    quotation = wb["Quotation"]
    assert len(quotation._images) >= 30
    assert quotation.freeze_panes is None
    assert quotation.sheet_view.zoomScale == 110
    assert quotation.page_setup.orientation == "landscape"
    assert quotation.page_setup.paperSize == 9
    assert quotation.sheet_properties.pageSetUpPr.fitToPage is True
    assert quotation.print_title_rows == "$7:$7"
    assert quotation.page_margins.left < 0.3
    assert quotation.page_margins.right < 0.3
    wb.close()


def test_mobiliti_product_starts_are_not_merged_rows():
    if not TEMPLATE.exists():
        pytest.skip("Template not available on this machine")

    wb = load_workbook(TEMPLATE, data_only=False, keep_links=False)
    ws = wb["Mobiliti"]
    try:
        merged_rows = {
            row
            for merged in ws.merged_cells.ranges
            if merged.min_col <= 29 and merged.max_col >= 4
            for row in range(merged.min_row, merged.max_row + 1)
        }
        assert not (set(SECTION_PROD_STARTS) & merged_rows)
        assert 398 not in SECTION_PROD_STARTS
        assert 399 in SECTION_PROD_STARTS
    finally:
        wb.close()


def test_python_engine_rejects_cummins_external_image_relationships_fail_closed(tmp_path):
    source = DOWNLOADS / "CUMMINS-Quotation Sheet - V1.xlsx"
    template = WORKER_TEMPLATE
    if not source.exists() or not template.exists():
        pytest.skip("CUMMINS input/template not available on this machine")

    output = tmp_path / "cummins_python.xlsx"
    # El quotation real contiene una relación de imagen externa. Task 6 exige
    # rechazarla: no se degrada silenciosamente ni se produce un XLSX parcial.
    with pytest.raises(ValueError, match="TargetMode externo no permitido para image"):
        generate_quote(
            source,
            output,
            {
                "cotizacion": "CUMMINS-TEST",
                "proyecto": "Cummins",
                "cliente": "Cummins",
                "image_provider": "pillow",
                "tipo_cambio": 20,
            },
            template,
        )

    assert not output.exists()
