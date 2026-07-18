from pathlib import Path
import shutil
import sys

import pytest
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.utils.units import pixels_to_EMU


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mobiliti_saas.quote_engine.images import (  # noqa: E402
    center_image_in_cell,
    extract_images,
    fit_image_to_cell,
    image_scale_for_category,
)
from mobiliti_saas.quote_engine.engine import _align_image_map_to_product_rows  # noqa: E402
from mobiliti_saas.quote_engine.engine import _generate_missing_dezgo_images  # noqa: E402
from mobiliti_saas.quote_engine.engine import _resolve_sunon_catalog_images  # noqa: E402
from mobiliti_saas.quote_engine.engine import _resolve_sunon_web_images  # noqa: E402
from mobiliti_saas.quote_engine.parser import QuoteItem  # noqa: E402


def _image(path: Path, size: tuple[int, int]) -> None:
    Image.new("RGB", size, (20, 20, 20)).save(path, "PNG")


def test_image_scale_for_category_contract():
    assert image_scale_for_category("Silla") == 0.8
    assert image_scale_for_category("Mesas de Juntas") == 0.9
    assert image_scale_for_category("Escritorios-WorkStation") == 0.9
    assert image_scale_for_category("Sofas") == 0.8
    assert image_scale_for_category("Sillones") == 0.8
    assert image_scale_for_category("Mesas de Apoyo") == 0.8
    assert image_scale_for_category("Banco") == 0.8
    assert image_scale_for_category("Categoria rara") == 0.7


def test_fit_image_to_cell_applies_category_scale(tmp_path):
    source = tmp_path / "product.png"
    _image(source, (200, 100))

    img = fit_image_to_cell(str(source), max_width=100, max_height=100, scale=0.8)

    assert img.width == 80
    assert img.height == 40


def test_center_image_in_cell_sets_middle_offsets(tmp_path):
    source = tmp_path / "product.png"
    _image(source, (100, 50))
    img = fit_image_to_cell(str(source), max_width=40, max_height=80)

    center_image_in_cell(img, row=7, column=2, cell_width=100, cell_height=80)

    assert img.anchor._from.col == 1
    assert img.anchor._from.row == 6
    assert img.anchor._from.colOff == pixels_to_EMU(30)
    assert img.anchor._from.rowOff == pixels_to_EMU(30)
    assert img.anchor.ext.cx == pixels_to_EMU(40)
    assert img.anchor.ext.cy == pixels_to_EMU(20)


def test_extract_images_reads_quotation_sheet_even_when_not_first(tmp_path):
    product = tmp_path / "product.png"
    _image(product, (60, 40))

    wb = Workbook()
    intro = wb.active
    intro.title = "Intro"
    ws = wb.create_sheet("Quotation")
    ws["A7"] = "No"
    ws["B7"] = "Item"
    ws["A9"] = 1
    ws["B9"] = "Desk"
    ws.add_image(WorkbookImage(str(product)), "B9")
    source = tmp_path / "quotation_not_first.xlsx"
    wb.save(source)
    wb.close()

    image_map, temp_dir = extract_images(source)
    try:
        assert 9 in image_map
        assert Path(image_map[9]).exists()
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def test_align_image_map_to_product_rows_accepts_nearby_excel_anchor():
    items = [
        QuoteItem(tipo="categoria", row=8, nombre="Workstations"),
        QuoteItem(tipo="producto", row=9, nombre="Desk"),
        QuoteItem(tipo="producto", row=14, nombre="Chair"),
    ]

    aligned = _align_image_map_to_product_rows({8: "desk.png", 15: "chair.png"}, items)

    assert aligned[9] == "desk.png"
    assert aligned[14] == "chair.png"


def test_generate_missing_dezgo_images_only_fills_missing_rows(monkeypatch, tmp_path):
    items = [
        QuoteItem(tipo="producto", row=9, nombre="Double Seat Workstation", descripcion="office workstation"),
        QuoteItem(tipo="producto", row=13, nombre="Alien Task Chair", descripcion="mesh chair"),
    ]
    existing = tmp_path / "existing.png"
    _image(existing, (20, 20))
    generated = []

    monkeypatch.setenv("DEZGO_API_KEY", "fake-key")

    def fake_generate(prompt, output_path, config):
        generated.append((prompt, Path(output_path)))
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        _image(Path(output_path), (32, 32))
        return Path(output_path)

    monkeypatch.setattr("mobiliti_saas.quote_engine.engine.generate_with_dezgo", fake_generate)

    image_map = _generate_missing_dezgo_images(
        {13: str(existing)},
        items,
        tmp_path,
        {"image_provider": "dezgo", "image_prompt": "Mejora la calidad de imagen y que este en fondo blanco"},
    )

    assert 9 in image_map
    assert image_map[13] == str(existing)
    assert len(generated) == 1
    assert "Mejora la calidad de imagen y que este en fondo blanco" in generated[0][0]
    assert "Double Seat Workstation" in generated[0][0]
    assert Path(image_map[9]).exists()


def test_sunon_web_provider_replaces_local_image_by_product_code(monkeypatch, tmp_path):
    local = tmp_path / "local.png"
    sunon = tmp_path / "sunon.png"
    _image(local, (20, 20))
    _image(sunon, (40, 40))
    items = [QuoteItem(tipo="producto", row=9, nombre="CHJ80SW H7 Task Chair")]

    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine.fetch_sunon_product_image",
        lambda _name, _output_dir, **_kwargs: sunon,
    )

    stats = {}
    result = _resolve_sunon_web_images(
        {9: str(local)},
        items,
        tmp_path,
        {"image_provider": "sunon_web"},
        stats=stats,
    )

    assert result[9] == str(sunon)
    assert stats["image_sunon_attempted_count"] == 1
    assert stats["image_sunon_found_count"] == 1


def test_sunon_web_provider_deduplicates_codes(monkeypatch, tmp_path):
    sunon = tmp_path / "sunon.png"
    _image(sunon, (40, 40))
    items = [
        QuoteItem(tipo="producto", row=9, nombre="CHJ80SW H7 Task Chair"),
        QuoteItem(tipo="producto", row=13, nombre="CHJ80SW H7 Task Chair duplicate"),
    ]
    calls = []

    def fake_fetch(name, _output_dir, **_kwargs):
        calls.append(name)
        return sunon

    monkeypatch.setattr("mobiliti_saas.quote_engine.engine.fetch_sunon_product_image", fake_fetch)
    stats = {}

    result = _resolve_sunon_web_images({}, items, tmp_path, {"image_provider": "sunon_web"}, stats=stats)

    assert len(calls) == 1
    assert result == {9: str(sunon), 13: str(sunon)}
    assert stats["image_sunon_attempted_count"] == 1
    assert stats["image_sunon_cache_hit_count"] == 1


def test_sunon_web_provider_caps_remote_lookups(monkeypatch, tmp_path):
    monkeypatch.setenv("SUNON_MAX_LOOKUPS_PER_JOB", "2")
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine.fetch_sunon_product_image",
        lambda *_args, **_kwargs: None,
    )
    items = [
        QuoteItem(tipo="producto", row=9, nombre="AAA1 Product"),
        QuoteItem(tipo="producto", row=13, nombre="BBB2 Product"),
        QuoteItem(tipo="producto", row=17, nombre="CCC3 Product"),
    ]
    stats = {}

    _resolve_sunon_web_images({}, items, tmp_path, {"image_provider": "sunon_web"}, stats=stats)

    assert stats["image_sunon_attempted_count"] == 2
    assert stats["image_sunon_skipped_limit_count"] == 1


def test_sunon_catalog_provider_obeys_job_time_budget(monkeypatch, tmp_path):
    monkeypatch.setenv("SUNON_LOOKUP_BUDGET_SECONDS", "30")
    clock = iter([0.0, 0.0, 31.0])
    monkeypatch.setattr("mobiliti_saas.quote_engine.engine.time.monotonic", lambda: next(clock))
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine.find_sunon_catalog_match",
        lambda code: ({"code": code}, code, "exact_code"),
    )
    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine.fetch_sunon_catalog_product_image",
        lambda *_args, **_kwargs: None,
    )
    items = [
        QuoteItem(tipo="producto", row=9, nombre="AAA1 Product"),
        QuoteItem(tipo="producto", row=13, nombre="BBB2 Product"),
    ]
    stats = {}

    _resolve_sunon_catalog_images({}, items, tmp_path, {"image_provider": "sunon_catalog"}, stats=stats)

    assert stats["image_sunon_catalog_attempted_count"] == 1
    assert stats["image_sunon_catalog_skipped_limit_count"] == 1


def test_sunon_catalog_provider_replaces_only_exact_catalog_matches(monkeypatch, tmp_path):
    local = tmp_path / "local.png"
    sunon = tmp_path / "sunon.png"
    _image(local, (20, 20))
    _image(sunon, (40, 40))
    items = [
        QuoteItem(tipo="producto", row=9, nombre="CHJ80SW H7 Task Chair"),
        QuoteItem(tipo="producto", row=13, nombre="SIN CODIGO"),
    ]

    monkeypatch.setattr(
        "mobiliti_saas.quote_engine.engine.fetch_sunon_catalog_product_image",
        lambda name, _output_dir, **_kwargs: sunon if "CHJ80SW" in name else None,
    )

    stats = {}
    result = _resolve_sunon_catalog_images(
        {9: str(local), 13: str(local)},
        items,
        tmp_path,
        {"image_provider": "sunon_catalog"},
        stats=stats,
    )

    assert result[9] == str(sunon)
    assert result[13] == str(local)
    assert stats["image_sunon_catalog_attempted_count"] == 1
    assert stats["image_sunon_catalog_exact_code_count"] == 1


def test_generate_missing_dezgo_images_raises_when_explicit_provider_fails(monkeypatch, tmp_path):
    items = [QuoteItem(tipo="producto", row=9, nombre="Double Seat Workstation", descripcion="office workstation")]
    monkeypatch.setenv("DEZGO_API_KEY", "fake-key")
    monkeypatch.setenv("IMAGE_PROVIDER_STRICT", "true")

    def fake_generate(_prompt, _output_path, _config):
        raise RuntimeError("dezgo generation failed")

    monkeypatch.setattr("mobiliti_saas.quote_engine.engine.generate_with_dezgo", fake_generate)

    with pytest.raises(RuntimeError, match="dezgo generation failed"):
        _generate_missing_dezgo_images(
            {},
            items,
            tmp_path,
            {"image_provider": "dezgo", "image_prompt": "Mejora la calidad de imagen y que este en fondo blanco"},
        )
