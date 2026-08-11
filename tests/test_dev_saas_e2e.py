import os
import sys
import hashlib
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "vercel_deploy", "api"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "mobiliti_saas", "worker"))

os.environ.setdefault("JWT_SECRET_KEY", "test-secret-32-chars-long!!!!!")

import index
import quote_worker
from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from quotation_import_fixtures import write_import_fixture


def _sample_quotation(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    for col, value in {1: "No.", 2: "Item", 4: "Description", 5: "Dimension", 7: "Qty", 10: "List Price"}.items():
        ws.cell(row=7, column=col, value=value)
    ws.cell(row=8, column=1, value="- Escritorios")
    ws.cell(row=9, column=1, value=1)
    ws.cell(row=9, column=2, value="Mesa Uno")
    ws.cell(row=9, column=4, value="Mesa operativa")
    ws.cell(row=9, column=5, value="1200 x 600")
    ws.cell(row=9, column=7, value=2)
    ws.cell(row=9, column=10, value=1000)
    wb.save(path)


def test_dev_mode_full_quote_flow(tmp_path, monkeypatch):
    store_dir = tmp_path / "dev-store"
    monkeypatch.setattr(index, "DEV_MODE", True)
    monkeypatch.setattr(index, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(index, "DEV_PUBLIC_BASE_URL", "http://testserver")
    monkeypatch.setattr(quote_worker, "DEV_MODE", True)
    monkeypatch.setattr(quote_worker, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")
    monkeypatch.setattr(quote_worker, "_delete_job_input", lambda *_args, **_kwargs: None)
    index._RATE_LIMIT_STORE.clear()

    client = TestClient(index.app)
    login = client.post("/login", json={"email": "dev@mobiliti.local", "password": "dev12345"})
    assert login.status_code == 200
    token = login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    source = tmp_path / "quotation.xlsx"
    _sample_quotation(source)

    init = client.post(
        "/cotizaciones/init-upload",
        headers=headers,
        json={"filename": "quotation.xlsx", "size": source.stat().st_size, "template": "official_2026_gdl"},
    )
    assert init.status_code == 200
    init_data = init.json()
    assert init_data["upload_url"].endswith("/dev-upload")

    upload = client.post(
        init_data["upload_url"],
        headers=headers,
        files={"file": ("quotation.xlsx", source.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200

    submit = client.post(
        f"/cotizaciones/{init_data['job_id']}/submit",
        headers=headers,
        json={
            "cotizacion": "COT-DEV",
            "proyecto": "Proyecto Dev",
            "cliente": "Cliente Dev",
            "correo": "cliente@example.com",
            "telefono": "555",
            "direccion": "Direccion",
            "razon_social": "Empresa Dev",
            "template": "official_2026_gdl",
            "image_provider": "pillow",
        },
    )
    assert submit.status_code == 200
    assert submit.json()["job"]["status"] == "queued"

    assert quote_worker.run_once() is True

    status = client.get(f"/cotizaciones/{init_data['job_id']}", headers=headers)
    assert status.status_code == 200
    job = status.json()["job"]
    assert job["status"] == "completed"
    assert job["output_path"].endswith("/output.xlsx")

    download = client.get(f"/cotizaciones/{init_data['job_id']}/download", headers=headers)
    assert download.status_code == 200
    assert "/dev/storage/" in download.json()["download_url"]

    encoded = quote(job["output_path"], safe="")
    file_resp = client.get(f"/dev/storage/{encoded}")
    assert file_resp.status_code == 200
    assert file_resp.content.startswith(b"PK")
    package = XlsxPackage.from_bytes(file_resp.content)
    assert package.sheet_state("Quotation_Data") == "veryHidden"
    assert package.sheet_part("Quotation")


def test_dev_mode_import_preview_mixed_checkout_worker_and_xlsx(tmp_path, monkeypatch):
    store_dir = tmp_path / "dev-store-mixed"
    monkeypatch.setattr(index, "DEV_MODE", True)
    monkeypatch.setattr(index, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(index, "DEV_PUBLIC_BASE_URL", "http://testserver")
    monkeypatch.setattr(quote_worker, "DEV_MODE", True)
    monkeypatch.setattr(quote_worker, "DEV_STORE_DIR", store_dir)
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")
    monkeypatch.setattr(quote_worker, "_delete_job_input", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(index, "_wake_worker", lambda: None)
    index._RATE_LIMIT_STORE.clear()

    offiho = OffihoCatalogItem(
        inventory_key="OHE-405 NEGRO ALUFSEN",
        code="OHE-405",
        name="ALUFSEN",
        variant="NEGRO",
        unit="PZA",
        pieces_per_box=Decimal("1"),
        available_quantity=Decimal("8"),
        unit_price=Decimal("7999"),
    )
    monkeypatch.setattr(index, "_load_offiho_catalog_cached", lambda: {
        "source_hash": hashlib.sha256(b"offiho-dev-e2e").hexdigest(),
        "generated_at": "2026-07-21T00:00:00Z",
        "items": [offiho],
        "by_inventory_key": {offiho.inventory_key: offiho},
    })
    exchange_date = date.today().isoformat()
    exchange_retrieved_at = datetime.now(timezone.utc).isoformat()
    monkeypatch.setattr(index, "db_list_exchange_rates", lambda: [{
        "currency": "USD",
        "effective_date": exchange_date,
        "mxn_per_unit": "18.500000",
        "retrieved_at": exchange_retrieved_at,
    }])

    client = TestClient(index.app)
    login = client.post(
        "/login", json={"email": "dev@mobiliti.local", "password": "dev12345"}
    )
    assert login.status_code == 200
    headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    source = write_import_fixture(tmp_path / "quotation-import.xlsx")
    source_hash_before = hashlib.sha256(source.read_bytes()).hexdigest()
    init = client.post(
        "/cotizaciones/init-upload",
        headers=headers,
        json={"filename": source.name, "size": source.stat().st_size, "template": "official_2026_gdl"},
    )
    assert init.status_code == 200, init.json()
    job_id = init.json()["job_id"]
    upload = client.post(
        init.json()["upload_url"],
        headers=headers,
        files={"file": (
            source.name,
            source.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )
    assert upload.status_code == 200, upload.json()

    preview_response = client.post(
        f"/cotizaciones/{job_id}/import-preview", headers=headers
    )
    assert preview_response.status_code == 200, preview_response.json()
    preview = preview_response.json()
    assert len(preview["sections"]) == 3
    assert len(preview["items"]) == 7
    assert sum(bool(item["image_url"]) for item in preview["items"]) == 7
    assert preview["currency_status"] == "required"
    assert preview["source_currency"] is None
    assert preview["provider"] == "SUNON TECHNOLOGY CO.,LTD."

    imported_items = []
    for item in preview["items"]:
        edited = item["source_row"] == 11
        imported_items.append({
            "kind": "imported",
            "import_id": preview["import_id"],
            "source_row": item["source_row"],
            "source_currency": "USD",
            "quantity": "2" if edited else item["quantity"],
            "overrides": {
                "name": item["name"],
                "description": "Descripcion revisada dev" if edited else item["description"],
                "dimension": item["dimension"],
                "unit_price": "82.00" if edited else item["unit_price"],
                "provider": preview["provider"],
            },
        })

    offiho_item = {
        "catalog": "offiho",
        "inventory_key": offiho.inventory_key,
        "quantity": "1",
    }
    sections = [
        {
            "id": f"section-{section_index}",
            "title": section["title"],
            "item_keys": list(section["item_keys"]),
        }
        for section_index, section in enumerate(preview["sections"], start=1)
    ]
    sections[1]["item_keys"].insert(2, f"offiho:{offiho.inventory_key}")
    checkout = client.post(
        "/catalogs/mixed-quote",
        headers=headers,
        json={
            "proyecto": "Proyecto Dev Mixto",
            "cliente": "Cliente Dev",
            "correo": "cliente@example.com",
            "telefono": "555",
            "direccion": "Direccion",
            "razon_social": "Empresa Dev",
            "descuento": "40",
            "quote_currency": "MXN",
            "template": "official_2026_gdl",
            "image_provider": "pillow",
            "items": [*imported_items, offiho_item],
            "sections": sections,
        },
    )
    assert checkout.status_code == 200, checkout.json()
    final_job_id = checkout.json()["job"]["id"]
    assert quote_worker.run_once() is True

    status = client.get(f"/cotizaciones/{final_job_id}", headers=headers)
    assert status.status_code == 200
    job = status.json()["job"]
    assert job["status"] == "completed", job
    assert job["metadata"]["source_type"] == "mixed_catalog_cart"
    encoded = quote(job["output_path"], safe="")
    output_response = client.get(f"/dev/storage/{encoded}")
    assert output_response.status_code == 200
    assert output_response.content.startswith(b"PK")
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash_before

    workbook = load_workbook(BytesIO(output_response.content), data_only=False)
    try:
        quotation = workbook["Quotation"]
        source_rows = [
            row for row in range(8, quotation.max_row + 1)
            if isinstance(quotation.cell(row, 1).value, (int, float))
        ]
        names = [quotation.cell(row, 2).value for row in source_rows]
        assert len(source_rows) == 8
        assert names[3:5] == ["ALUFSEN", "CAT60SC Altaes Task Chair"]
        edited_row = source_rows[names.index("CAI63SW Alien Task Chair")]
        assert quotation.cell(edited_row, 4).value == (
            "Silla modelo Alien Task Chair.\nDescripcion revisada dev."
        )
        assert quotation.cell(edited_row, 5).value == "Descripción 2"
        assert quotation.cell(edited_row, 6).value == "602 x 600 mm"
        assert quotation.cell(edited_row, 8).value == 2
        assert quotation.cell(edited_row, 11).value == 1517
        assert len(quotation._images) >= 7

        cotizacion = workbook["Cotizacion"]
        product_rows = [
            row for row in range(1, cotizacion.max_row + 1)
            if str(cotizacion.cell(row, 1).value or "").startswith("=Mobiliti!D")
        ]
        assert len(product_rows) == 8
        assert cotizacion.cell(product_rows[0], 7).value == 0.4
        assert [cotizacion.cell(row, 7).value for row in product_rows[1:]] == [
            f"=$G${product_rows[0]}"
        ] * 7
        mobiliti = workbook["Mobiliti"]
        official_blank_rows = [
            row for row in range(1, mobiliti.max_row + 1)
            if mobiliti.cell(row, 4).value is None
            and str(mobiliti.cell(row, 23).value or "").startswith("=IF(F")
            and str(mobiliti.cell(row, 24).value or "").startswith("=_xlfn.MINIFS(")
        ]
        assert official_blank_rows
        blank_row = official_blank_rows[0]
        assert all(
            mobiliti.cell(blank_row, column).value is None
            for column in (4, 5, 6, 8, 10, 11, 16)
        )
        assert str(mobiliti.cell(blank_row, 23).value).startswith(
            f'=IF(F{blank_row}="Offiho",J{blank_row},'
        )
        assert str(mobiliti.cell(blank_row, 24).value).startswith(
            "=_xlfn.MINIFS("
        )
        assert mobiliti.cell(blank_row, 35).value == (
            f'=IF(AH{blank_row}<30%,"ERROR","OK")'
        )
        assert all(
            "$K$6" not in str(cell.value or "")
            for sheet in (quotation, mobiliti, cotizacion)
            for cell in sheet._cells.values()
        )
    finally:
        workbook.close()

    artifact_dir = os.environ.get("MOBILITI_TASK8_ARTIFACT_DIR")
    if artifact_dir:
        destination = Path(artifact_dir)
        destination.mkdir(parents=True, exist_ok=True)
        (destination / "mixed-dev-final-mxn.xlsx").write_bytes(output_response.content)
