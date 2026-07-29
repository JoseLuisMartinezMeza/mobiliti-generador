from __future__ import annotations

from copy import deepcopy
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw

from mobiliti_saas.quote_engine import engine, generate_quote
from mobiliti_saas.quote_engine.mixed_catalog import (
    build_mixed_catalog_cart_payload,
    create_mixed_catalog_quotation_workbook,
)
from mobiliti_saas.quote_engine.mobiliti_layout import SectionNeed, plan_mobiliti_layout
from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage
from mobiliti_saas.quote_engine.ooxml_worksheet import MobilitiSheetMutation
from mobiliti_saas.quote_engine.official_composer import (
    CotizacionMetadata,
    CotizacionPriceTerm,
    CotizacionProduct,
    CotizacionSection,
    CotizacionSheetEditor,
    _validate_exact_cotizacion_surface,
)
from mobiliti_saas.quote_engine.project_quote import project_context
from mobiliti_saas.quote_engine.quotation_sheets import quotation_data_rows


PRINCIPAL_ID = "11111111-1111-4111-8111-111111111111"
PER_UNIT_ID = "22222222-2222-4222-8222-222222222222"
FIXED_ID = "33333333-3333-4333-8333-333333333333"


def _line(
    line_id: str,
    *,
    description: str,
    quantity: str,
    section_id: str = "section-1",
    parent_item_key: str | None = None,
    origin: str = "sunon",
    image_content: bytes | None = None,
    image_content_type: str | None = None,
) -> engine._OfficialPresentationLine:
    return engine._OfficialPresentationLine(
        item_key=line_id,
        section_id=section_id,
        section_title="Recepción",
        item=None,
        name={
            PRINCIPAL_ID: "MAIN-1",
            PER_UNIT_ID: "PER-1",
            FIXED_ID: "FIXED-1",
        }.get(line_id, "LIDO.OP-INT"),
        description=description,
        dimensions="600 x 600 mm",
        m3=Decimal("0"),
        quantity=Decimal(quantity),
        category="Terminados",
        provider="Sunon",
        region="Centro",
        original_currency="USD",
        original_cost=Decimal("100"),
        frozen_rate=Decimal("18.5"),
        converted_cost=Decimal("1850"),
        origin=origin,
        source_row=None,
        upstream_row_hash="",
        parent_item_key=parent_item_key,
        image_content=image_content,
        image_content_type=(
            image_content_type
            if image_content is not None
            else None
        )
        or ("image/png" if image_content is not None else None),
    )


def _mobiliti(lines):
    row_map = plan_mobiliti_layout(
        (SectionNeed("section-1", "Recepción", len(lines)),)
    )
    return MobilitiSheetMutation(b"<worksheet/>", row_map)


def _quotation_data_row_map(lines):
    return {
        line.item_key: position + 2
        for position, line in enumerate(lines)
    }


def _context():
    return {
        "project_id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
        "project_revision": 3,
        "project_payload_hash": "a" * 64,
        "normalized_project_payload": {
            "sections": [
                {"section_id": "section-1", "concept": "Recepción", "position": 0}
            ],
            "lines": [
                {
                    "line_id": PRINCIPAL_ID,
                    "role": "principal",
                    "section_id": "section-1",
                    "parent_line_id": None,
                },
                {
                    "line_id": PER_UNIT_ID,
                    "role": "complement",
                    "section_id": None,
                    "parent_line_id": PRINCIPAL_ID,
                },
                {
                    "line_id": FIXED_ID,
                    "role": "complement",
                    "section_id": None,
                    "parent_line_id": PRINCIPAL_ID,
                },
            ],
        },
        "compositions": [
            {
                "principal_line_id": PRINCIPAL_ID,
                "section_id": "section-1",
                "component_line_ids": [PRINCIPAL_ID, PER_UNIT_ID, FIXED_ID],
                "price_terms": [
                    {
                        "line_id": PRINCIPAL_ID,
                        "numerator": "1",
                        "denominator": "1",
                    },
                    {
                        "line_id": PER_UNIT_ID,
                        "numerator": "2",
                        "denominator": "1",
                    },
                    {
                        "line_id": FIXED_ID,
                        "numerator": "3",
                        "denominator": "10",
                    },
                ],
            }
        ],
    }


def _project_lines():
    return (
        _line(PRINCIPAL_ID, description="Principal", quantity="10"),
        _line(PER_UNIT_ID, description="Complemento por unidad", quantity="20"),
        _line(FIXED_ID, description="Complemento fijo", quantity="3"),
    )


def _supplier_item(internal_id: str, name: str, description: str):
    return {
        "internal_id": internal_id,
        "supplier": "sunon",
        "product_key": internal_id,
        "sku": internal_id.upper(),
        "code_status": "verified",
        "brand": "sunon",
        "collection": "project",
        "name": name,
        "description": description,
        "unit": "pieza",
        "availability_type": "made_to_order",
        "stock": None,
        "lead_time": "Sobre pedido",
        "base_price_options": [],
        "add_on_options": [],
        "base_currency": "USD",
        "price_net": "100.000000",
        "tax_rate": "0.160000",
        "attributes": {"dimensions": "600 x 600 mm"},
        "image_url": "",
        "image_kind": "placeholder",
        "product_url": "",
        "warnings": [],
        "source_reference": f"sunon:{internal_id}",
    }


def _project_payload():
    common = {
        "source": "catalog",
        "catalog": "sunon",
        "official_code": "",
        "display_cache": {"name": "", "code": "", "image_url": ""},
    }
    return {
        "schema_version": 1,
        "quote_fields": {
            "proyecto": "Proyecto compuesto",
            "cliente": "Cliente",
            "correo": "cliente@example.com",
            "telefono": "33",
            "direccion": "Guadalajara",
            "razon_social": "Cliente SA",
            "quote_currency": "MXN",
            "descuento": "40",
        },
        "sections": [
            {"section_id": "section-1", "concept": "Recepción", "position": 0}
        ],
        "lines": [
            {
                **common,
                "line_id": PRINCIPAL_ID,
                "role": "principal",
                "section_id": "section-1",
                "parent_line_id": None,
                "position": 0,
                "quantity": "10",
                "official_code": "MAIN-1",
                "identity": {
                    "internal_id": "sunon:main-1",
                    "base_option_id": "",
                    "add_on_option_ids": [],
                },
                "display_cache": {
                    "name": "MAIN-1",
                    "code": "MAIN-1",
                    "image_url": "",
                },
            },
            {
                **common,
                "line_id": PER_UNIT_ID,
                "role": "complement",
                "section_id": None,
                "parent_line_id": PRINCIPAL_ID,
                "position": 0,
                "quantity": "2",
                "quantity_mode": "per_parent_unit",
                "official_code": "PER-1",
                "identity": {
                    "internal_id": "sunon:per-1",
                    "base_option_id": "",
                    "add_on_option_ids": [],
                },
                "display_cache": {
                    "name": "PER-1",
                    "code": "PER-1",
                    "image_url": "",
                },
            },
            {
                **common,
                "line_id": FIXED_ID,
                "role": "complement",
                "section_id": None,
                "parent_line_id": PRINCIPAL_ID,
                "position": 1,
                "quantity": "3",
                "quantity_mode": "fixed_project",
                "official_code": "FIXED-1",
                "identity": {
                    "internal_id": "sunon:fixed-1",
                    "base_option_id": "",
                    "add_on_option_ids": [],
                },
                "display_cache": {
                    "name": "FIXED-1",
                    "code": "FIXED-1",
                    "image_url": "",
                },
            },
        ],
    }


def _project_quote_input(tmp_path):
    catalog = {
        "supplier": "sunon",
        "source_hash": "c" * 64,
        "generated_at": "2026-07-23T00:00:00+00:00",
        "items": [
            _supplier_item("sunon:main-1", "MAIN-1", "Principal"),
            _supplier_item(
                "sunon:per-1",
                "PER-1",
                "Complemento por unidad",
            ),
            _supplier_item(
                "sunon:fixed-1",
                "FIXED-1",
                "Complemento fijo",
            ),
        ],
    }
    project = _project_payload()
    context = project_context(
        project,
        "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa",
        3,
    )
    rows = [
        {
            "line_id": PRINCIPAL_ID,
            "catalog": "sunon",
            "internal_id": "sunon:main-1",
            "quantity": "10",
        },
        {
            "line_id": PER_UNIT_ID,
            "catalog": "sunon",
            "internal_id": "sunon:per-1",
            "quantity": "20",
        },
        {
            "line_id": FIXED_ID,
            "catalog": "sunon",
            "internal_id": "sunon:fixed-1",
            "quantity": "3",
        },
    ]
    payload = build_mixed_catalog_cart_payload(
        rows,
        catalogs={"sunon": catalog},
        rate_rows=[
            {
                "currency": "USD",
                "effective_date": "2026-07-23",
                "mxn_per_unit": "18.500000",
                "retrieved_at": "2026-07-23T00:00:00+00:00",
            }
        ],
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[
            {
                "id": "section-1",
                "title": "Recepción",
                "line_ids": [PRINCIPAL_ID, PER_UNIT_ID, FIXED_ID],
            }
        ],
        project_context=context,
        today=date(2026, 7, 23),
    )
    source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "project-source.xlsx",
        image_dir=tmp_path / "images",
    )
    metadata = {
        **project["quote_fields"],
        "cotizacion": "PROJECT-001",
        "catalog_price_mode": "mixed_catalog_converted",
        "base_currency": "MXN",
        "quote_currency": "MXN",
        "exchange_rate": "1.000000",
        "rate_summary": deepcopy(payload["rate_summary"]),
        "auto_electrification_rate": None,
        "catalog_source_hashes": {"sunon": "c" * 64},
        "project_context": deepcopy(payload["project_context"]),
    }
    return source, payload, metadata


def test_official_engine_separates_mobiliti_and_composes_cotizacion(tmp_path):
    source, payload, metadata = _project_quote_input(tmp_path)
    output = tmp_path / "project.xlsx"

    generate_quote(
        source,
        output,
        metadata,
        engine.OFFICIAL_TEMPLATE_PATH,
        original_quotation_path=None,
        quotation_data_rows=quotation_data_rows(payload),
    )

    workbook = load_workbook(output, data_only=False)
    try:
        mobiliti = workbook["Mobiliti"]
        assert [mobiliti.cell(row, 4).value for row in (14, 15, 16)] == [
            "=Quotation!B9",
            "=Quotation!B10",
            "=Quotation!B11",
        ]
        assert [mobiliti.cell(row, 8).value for row in (14, 15, 16)] == [
            "=Quotation!H9",
            "=Quotation!H10",
            "=Quotation!H11",
        ]
        assert [mobiliti.cell(row, 10).value for row in (14, 15, 16)] == [
            "=Quotation!K9",
            "=Quotation!K10",
            "=Quotation!K11",
        ]
        quotation = workbook["Quotation"]
        assert [quotation.cell(row, 8).value for row in (9, 10, 11)] == [
            10,
            20,
            3,
        ]
        cotizacion = workbook["Cotizacion"]
        product_rows = [
            row
            for row in range(1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 1).value == "=Mobiliti!D14"
        ]
        assert product_rows == [17]
        assert all(
            cotizacion.cell(row, 1).value not in {
                "=Mobiliti!D15",
                "=Mobiliti!D16",
            }
            for row in range(1, cotizacion.max_row + 1)
        )
        assert cotizacion["F17"].value == (
            "=Mobiliti!X14+Mobiliti!X15*2+Mobiliti!X16*3/10"
        )
        assert cotizacion["A17"].value == "=Mobiliti!D14"
        assert cotizacion["C17"].value == "=Quotation!D9"
        assert cotizacion["D17"].value == "=Quotation!F9"
        assert cotizacion["E17"].value == "=Mobiliti!H14"
        processed_description = quotation["D9"].value
        assert "\n+ " in processed_description
        assert processed_description.count("\n+ ") == 2
    finally:
        workbook.close()


def test_project_projection_keeps_physical_mobiliti_rows_and_groups_cotizacion():
    lines = _project_lines()

    sections = engine._project_cotizacion_sections(
        lines,
        _mobiliti(lines),
        {"project_context": _context(), "descuento": "40"},
        _quotation_data_row_map(lines),
    )

    assert len(sections) == 1
    assert sections[0].title == "Recepción"
    assert len(sections[0].products) == 1
    product = sections[0].products[0]
    assert product.item_key == PRINCIPAL_ID
    assert product.name == "MAIN-1"
    assert product.description == (
        "Principal\n+ Complemento por unidad\n+ Complemento fijo"
    )
    assert product.quantity == Decimal("10")
    assert product.discount == Decimal("0.4")
    assert [
        (term.mobiliti_row, term.numerator, term.denominator)
        for term in product.price_terms
    ] == [
        (14, Decimal("1"), Decimal("1")),
        (15, Decimal("2"), Decimal("1")),
        (16, Decimal("3"), Decimal("10")),
    ]


def test_project_projection_keeps_ordered_images_separate():
    lines = (
        _line(
            PRINCIPAL_ID,
            description="Principal",
            quantity="10",
            image_content=b"principal",
        ),
        _line(
            PER_UNIT_ID,
            description="Complemento por unidad",
            quantity="20",
            image_content=b"per-unit",
        ),
        _line(
            FIXED_ID,
            description="Complemento fijo",
            quantity="3",
            image_content=b"fixed",
        ),
    )
    product = engine._project_cotizacion_sections(
        lines,
        _mobiliti(lines),
        {"project_context": _context()},
        _quotation_data_row_map(lines),
    )[0].products[0]

    assert product.image_content == b"principal"
    assert product.image_content_type == "image/png"
    assert [
        (image.content, image.content_type)
        for image in product.complement_images
    ] == [
        (b"per-unit", "image/png"),
        (b"fixed", "image/png"),
    ]


def test_official_image_improvement_changes_only_cotizacion_projection():
    source = BytesIO()
    source_image = Image.new("RGB", (80, 60), (232, 232, 232))
    ImageDraw.Draw(source_image).rectangle(
        (24, 12, 56, 52),
        fill=(35, 35, 35),
    )
    source_image.save(source, format="JPEG")
    line = _line(
        PRINCIPAL_ID,
        description="Principal",
        quantity="1",
        image_content=source.getvalue(),
        image_content_type="image/jpeg",
    )

    improved = engine._improve_official_cotizacion_images(
        (line,),
        {
            "image_provider": "pillow",
            "image_background": "white",
            "image_cleanup_strength": "balanced",
        },
    )

    assert line.image_content == source.getvalue()
    assert line.image_content_type == "image/jpeg"
    assert improved[0].image_content_type == "image/png"
    assert improved[0].image_content.startswith(b"\x89PNG\r\n\x1a\n")
    with Image.open(BytesIO(improved[0].image_content)) as result:
        assert result.format == "PNG"
        assert result.mode == "RGB"
        assert result.getpixel((0, 0)) == (255, 255, 255)


def test_official_image_improvement_removes_shadow_only_for_imported_lines(monkeypatch):
    calls = []

    def fake_improve(content, content_type, **options):
        calls.append((content, content_type, options["remove_shadow"]))
        return content, content_type

    monkeypatch.setattr(engine, "improve_product_image_bytes", fake_improve)
    imported = _line(
        PRINCIPAL_ID,
        description="Importado",
        quantity="1",
        origin="imported",
        image_content=b"imported",
    )
    catalog = _line(
        PER_UNIT_ID,
        description="Catálogo",
        quantity="1",
        origin="sunon",
        image_content=b"catalog",
    )

    engine._improve_official_cotizacion_images(
        (imported, catalog),
        {"image_provider": "pillow", "image_background": "white"},
    )

    assert calls == [
        (b"imported", "image/png", True),
        (b"catalog", "image/png", False),
    ]


def test_official_image_improvement_falls_back_to_original_invalid_bytes():
    line = _line(
        PRINCIPAL_ID,
        description="Principal",
        quantity="1",
        image_content=b"invalid",
        image_content_type="image/png",
    )

    assert engine._improve_official_cotizacion_images(
        (line,),
        {"image_provider": "pillow", "image_background": "white"},
    ) == (line,)


def test_project_projection_keeps_generated_lumbro_as_independent_line():
    principal, per_unit, fixed = _project_lines()
    accessory = _line(
        "lumbro-auto-1",
        description="Accesorio de electrificación Lumbro",
        quantity="1",
        parent_item_key=PRINCIPAL_ID,
        origin="lumbro",
    )
    lines = (principal, accessory, per_unit, fixed)

    products = engine._project_cotizacion_sections(
        lines,
        _mobiliti(lines),
        {"project_context": _context()},
        _quotation_data_row_map(lines),
    )[0].products

    assert [product.item_key for product in products] == [
        PRINCIPAL_ID,
        "lumbro-auto-1",
    ]
    assert [term.mobiliti_row for term in products[0].price_terms] == [14, 16, 17]
    assert [term.mobiliti_row for term in products[1].price_terms] == [15]


def test_project_projection_distinguishes_catalog_lumbro_from_auto_accessory():
    principal = _line(
        PRINCIPAL_ID,
        description="Principal Lumbro",
        quantity="10",
        origin="lumbro",
    )
    complement = _line(
        PER_UNIT_ID,
        description="Complemento Lumbro",
        quantity="20",
        origin="lumbro",
    )
    fixed = _line(FIXED_ID, description="Complemento fijo", quantity="3")
    accessory = _line(
        "lumbro-auto-1",
        description="Accesorio de electrificación Lumbro",
        quantity="1",
        parent_item_key=PRINCIPAL_ID,
        origin="lumbro",
    )
    lines = (principal, accessory, complement, fixed)

    products = engine._project_cotizacion_sections(
        lines,
        _mobiliti(lines),
        {"project_context": _context()},
        _quotation_data_row_map(lines),
    )[0].products

    assert [product.item_key for product in products] == [
        PRINCIPAL_ID,
        "lumbro-auto-1",
    ]
    assert [term.mobiliti_row for term in products[0].price_terms] == [14, 16, 17]
    assert [term.mobiliti_row for term in products[1].price_terms] == [15]


def test_exact_auditor_requires_explicit_project_signal_for_grouped_rows():
    base = XlsxPackage.read(engine.OFFICIAL_TEMPLATE_PATH)
    row_map = plan_mobiliti_layout(
        (SectionNeed("section-1", "Recepción", 2),)
    )
    grouped = CotizacionSheetEditor.from_xml(
        base.parts[base.sheet_part("Cotizacion")]
    ).compose(
        metadata=CotizacionMetadata(
            quotation_number="PROJECT-AUDIT",
            project="Proyecto",
            client="Cliente",
            email="cliente@example.com",
            phone="33",
            address="Guadalajara",
            business_name="Cliente SA",
        ),
        sections=(
            CotizacionSection(
                title="Recepción",
                products=(
                    CotizacionProduct(
                        item_key=PRINCIPAL_ID,
                        name="MAIN-1",
                        description="Principal\n+ Complemento",
                        dimensions="600 x 600 mm",
                        quantity=Decimal("10"),
                        mobiliti_row=14,
                        discount=Decimal("0.4"),
                        price_terms=(
                            CotizacionPriceTerm(14),
                            CotizacionPriceTerm(15),
                        ),
                    ),
                ),
            ),
        ),
    )

    with pytest.raises(ValueError, match="filas de producto"):
        _validate_exact_cotizacion_surface(base, grouped, row_map)

    _validate_exact_cotizacion_surface(
        base,
        grouped,
        row_map,
        project_composition=True,
    )


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (
            lambda context: context["compositions"][0][
                "component_line_ids"
            ].__setitem__(1, "44444444-4444-4444-8444-444444444444"),
            "desconocida",
        ),
        (
            lambda context: context["compositions"][0][
                "component_line_ids"
            ].__setitem__(2, PER_UNIT_ID),
            "repetida",
        ),
        (
            lambda context: context["compositions"][0]["price_terms"].pop(),
            "incompletos",
        ),
        (
            lambda context: context["compositions"][0].__setitem__(
                "section_id", "section-2"
            ),
            "sección",
        ),
        (
            lambda context: (
                context["compositions"][0]["component_line_ids"].pop(),
                context["compositions"][0]["price_terms"].pop(),
            ),
            "sin consumir",
        ),
    ],
)
def test_project_projection_rejects_invalid_component_consumption(mutation, message):
    context = deepcopy(_context())
    mutation(context)

    with pytest.raises(ValueError, match=message):
        engine._project_cotizacion_sections(
            _project_lines(),
            _mobiliti(_project_lines()),
            {"project_context": context},
            _quotation_data_row_map(_project_lines()),
        )


def test_project_projection_without_context_is_identical_to_legacy():
    lines = _project_lines()
    mobiliti = _mobiliti(lines)
    metadata = {"descuento": "40"}

    quotation_rows = _quotation_data_row_map(lines)
    assert engine._project_cotizacion_sections(
        lines,
        mobiliti,
        metadata,
        quotation_rows,
    ) == (
        engine._legacy_cotizacion_sections(
            lines,
            mobiliti,
            metadata,
            quotation_rows,
        )
    )
