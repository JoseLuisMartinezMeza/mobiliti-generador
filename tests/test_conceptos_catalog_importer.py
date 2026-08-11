from __future__ import annotations

import hashlib
import importlib
import io
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image

from mobiliti_saas.worker.catalog_sync.importers.common import CatalogSnapshotBuild


MODULE = "mobiliti_saas.worker.catalog_sync.importers.conceptos"
MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SPEC_SHEET = "Spec sofas - Cdmx-Gdl-Qro"
COST_SHEET = "Costo Sofas - Cdmx-Gdl-Qro"
ROOT = Path(__file__).resolve().parents[1]
OFFICIAL_SOURCE = (
    ROOT / "outputs" / "019f7907-1ecc-7001-b3f3-8eb209086fa8" / "sources"
    / "Spec guide - Conceptos - Sofas - CdMx - Gdl - Qro - 2021.xlsx"
)


@dataclass(frozen=True)
class SourceDocument:
    path: str
    kind: str
    sha256: str
    mime_type: str
    local_path: Path


def _conceptos_module():
    try:
        return importlib.import_module(MODULE)
    except ModuleNotFoundError as error:
        pytest.fail(f"Falta el importador Conceptos: {error.name}")


def _image(color: str) -> WorksheetImage:
    data = io.BytesIO()
    Image.new("RGB", (8, 6), color).save(data, "PNG")
    data.seek(0)
    return WorksheetImage(data)


def _headers(sheet, values: tuple[str, ...]) -> None:
    for index, value in enumerate(values, 1):
        sheet.cell(8, index, value)


def _row(sheet, number: int, values: tuple[object, ...]) -> None:
    for index, value in enumerate(values, 1):
        sheet.cell(number, index, value)


@pytest.fixture
def conceptos_files(tmp_path):
    path = tmp_path / "Spec guide - Conceptos - Sofas.xlsx"
    workbook = Workbook()
    spec = workbook.active
    spec.title = SPEC_SHEET
    cost = workbook.create_sheet(COST_SHEET)
    _headers(spec, ("Imagen", "Codigo", "Descripcion", "Material", "Medidas", "Unidad"))
    _headers(cost, ("Imagen", "Codigo", "Descripcion", "Material", "Costo", "Unidad", "Referencia", "Medidas"))

    # El bloque probado permite heredar solamente A/B entre sus dos variantes.
    _row(spec, 9, (None, "SOF-100", "Sofa Lirio", "Tela Oslo", "180 x 80 cm", "PZA"))
    _row(spec, 10, (None, None, "Sofa Lirio Grande", "Piel", "200 x 80 cm", "JGO"))
    _row(spec, 11, (None, None, "FILA DECORATIVA", "NO IMPORTAR", "SIN MEDIDA", "NO-UNIDAD"))
    _row(spec, 12, (None, "SOF-200", "Sofa Nube", "Tela", "160 x 75 cm", "PZA"))
    spec.merge_cells("A9:A10")
    spec.merge_cells("B9:B10")
    # El dibujo oficial puede iniciar dentro del bloque combinado, no solo en
    # su primera fila. La interseccion del ancla con A9:A10 prueba pertenencia.
    spec.add_image(_image("red"), "A10")

    _row(cost, 9, (None, "SOF-100", "Texto que no se usa para unir", "Tela Oslo", 12000, "PZA", 18000, "180 x 80 cm"))
    _row(cost, 10, (None, None, "Otra descripcion", "Piel", 15500, "PZA", 22000, "200 x 80 cm"))
    # Tiene importe, pero no pertenece a ningun bloque probado y debe ignorarse.
    _row(cost, 11, (None, None, "Separador decorativo", "", 999999, "PZA", 999999, ""))
    _row(cost, 12, (None, "SOF-200", "Sofa Nube", "Tela", 9000, "PZA", 13000, "160 x 75 cm"))
    cost.merge_cells("A9:A10")
    cost.merge_cells("B9:B10")
    workbook.save(path)
    workbook.close()
    return (
        SourceDocument(path.name, "spec_guide", hashlib.sha256(path.read_bytes()).hexdigest(), MIME, path),
    )


def test_parsea_las_dos_hojas_por_bloque_y_usa_e_no_g(conceptos_files):
    rows = _conceptos_module().parse_conceptos_rows(conceptos_files)

    assert len(rows) == 3
    tela, piel, nube = rows
    assert tela["code"] == piel["code"] == "SOF-100"
    assert tela["name"] == "Sofa Lirio"
    assert piel["name"] == "Sofa Lirio Grande"
    assert tela["dimensions"] == "180 x 80 cm"
    assert piel["dimensions"] == "200 x 80 cm"
    assert tela["material"] == "Tela Oslo"
    assert piel["material"] == "Piel"
    assert tela["unit"] == "PZA"
    assert piel["unit"] == "JGO"
    assert tela["raw_cost"] == Decimal("12000")
    assert piel["raw_cost"] == Decimal("15500")
    assert tela["reference_price_mxn"] == Decimal("18000")
    assert piel["reference_price_mxn"] == Decimal("22000")
    assert all(row["base_currency"] == "MXN" for row in rows)
    assert all(row["raw_cost"] != Decimal("999999") for row in rows)
    assert all(row["name"] != "FILA DECORATIVA" for row in rows)
    assert all(row["dimensions"] != "SIN MEDIDA" for row in rows)
    assert all(row["unit"] != "NO-UNIDAD" for row in rows)


def test_imagen_y_codigo_solo_se_arrastran_dentro_del_bloque_combinado(conceptos_files):
    rows = _conceptos_module().parse_conceptos_rows(conceptos_files)

    tela, piel, nube = rows
    assert tela["image"]["source_reference"]["cell_or_bbox"] == "A10"
    assert piel["image"]["sha256"] == tela["image"]["sha256"]
    assert nube.get("image") is None
    assert piel["provenance"]["spec_row"] == 10
    assert piel["provenance"]["cost_row"] == 10


def test_rechaza_encabezados_oficiales_alterados_en_fila_ocho(conceptos_files):
    path = conceptos_files[0].local_path
    workbook = Workbook()
    spec = workbook.active
    spec.title = SPEC_SHEET
    cost = workbook.create_sheet(COST_SHEET)
    _headers(spec, ("Imagen", "Codigo", "Descripcion", "Material", "Unidad", "Medidas"))
    _headers(cost, ("Imagen", "Codigo", "Descripcion", "Material", "Costo", "Unidad", "Referencia", "Medidas"))
    _row(spec, 9, (None, "SOF-1", "Sofa", "Tela", "180 cm", "PZA"))
    _row(cost, 9, (None, "SOF-1", "Sofa", "Tela", 100, "PZA", 150, "180 cm"))
    workbook.save(path)
    workbook.close()
    file = SourceDocument(path.name, "spec_guide", hashlib.sha256(path.read_bytes()).hexdigest(), MIME, path)

    with pytest.raises(ValueError, match="CONCEPTOS_HEADERS"):
        _conceptos_module().parse_conceptos_rows((file,))


def test_snapshot_publica_variantes_seleccionables_ids_deterministas_y_assets(conceptos_files):
    adapter = _conceptos_module()
    first = adapter.build_conceptos_snapshot_with_assets(conceptos_files)
    second = adapter.build_conceptos_snapshot_with_assets(conceptos_files)

    assert isinstance(first, CatalogSnapshotBuild)
    assert first.snapshot["supplier"] == "conceptos"
    lirio = next(item for item in first.snapshot["items"] if item["sku"] == "SOF-100")
    assert lirio["base_currency"] == "MXN"
    assert lirio["price_net"] == "12000.000000"
    assert [option["price_net"] for option in lirio["base_price_options"]] == [
        "12000.000000", "15500.000000",
    ]
    assert [option["id"] for option in lirio["base_price_options"]] == [
        option["id"]
        for option in next(item for item in second.snapshot["items"] if item["sku"] == "SOF-100")["base_price_options"]
    ]
    assert {variant["material"] for variant in lirio["attributes"]["variants"]} == {"Tela Oslo", "Piel"}
    assert first.assets_by_sha256
    assert first.bindings[0].image_kind == "official"


def test_builder_conceptos_se_expone_desde_el_paquete_de_importadores(conceptos_files):
    from mobiliti_saas.worker.catalog_sync.importers import build_conceptos_snapshot_with_assets

    build = build_conceptos_snapshot_with_assets(conceptos_files)
    assert isinstance(build, CatalogSnapshotBuild)


def test_rechaza_costos_conflictivos_para_la_misma_variante(conceptos_files):
    path = conceptos_files[0].local_path
    workbook = Workbook()
    spec = workbook.active
    spec.title = SPEC_SHEET
    cost = workbook.create_sheet(COST_SHEET)
    _headers(spec, ("Imagen", "Codigo", "Descripcion", "Material", "Medidas", "Unidad"))
    _headers(cost, ("Imagen", "Codigo", "Descripcion", "Material", "Costo", "Unidad", "Referencia", "Medidas"))
    _row(spec, 9, (None, "DUP-1", "Sofa", "Tela", "180 cm", "PZA"))
    _row(spec, 10, (None, "DUP-1", "Sofa", "Tela", "180 cm", "PZA"))
    _row(cost, 9, (None, "DUP-1", "Uno", "Tela", 100, "PZA", 200, "180 cm"))
    _row(cost, 10, (None, "DUP-1", "Dos", "Tela", 120, "PZA", 200, "180 cm"))
    spec.merge_cells("B9:B10")
    cost.merge_cells("B9:B10")
    workbook.save(path)
    workbook.close()
    file = SourceDocument(path.name, "spec_guide", hashlib.sha256(path.read_bytes()).hexdigest(), MIME, path)

    with pytest.raises(ValueError, match="CONCEPTOS_CONFLICTING_COST"):
        _conceptos_module().parse_conceptos_rows((file,))


def test_archivo_oficial_acepta_solo_sus_encabezados_publicados_y_conserva_e_y_g():
    """El layout local publicado no es el fixture detallado, pero sigue siendo cerrado."""
    workbook = load_workbook(OFFICIAL_SOURCE, data_only=True)
    try:
        assert tuple(workbook[SPEC_SHEET].cell(8, column).value for column in range(1, 7)) == (
            "Imagen", "Codigo ", "Descripción ", "Unidad ", "Precio Venta", "Moneda",
        )
        assert tuple(workbook[COST_SHEET].cell(8, column).value for column in range(1, 9)) == (
            "Imagen", "Codigo ", "Descripción ", "Unidad ", "Precio Unitario ",
            "Utilidad (50%)", "Precio Venta", "Moneda",
        )
    finally:
        workbook.close()

    document = SourceDocument(
        OFFICIAL_SOURCE.name,
        "spec_guide",
        hashlib.sha256(OFFICIAL_SOURCE.read_bytes()).hexdigest(),
        MIME,
        OFFICIAL_SOURCE,
    )
    rows = _conceptos_module().parse_conceptos_rows((document,))

    assert len(rows) == 171
    assert all(row["raw_cost"] > 0 for row in rows)
    assert all(row["provenance"]["cost_cell"].startswith("E") for row in rows)
    assert all(row["provenance"]["reference_cell"].startswith("G") for row in rows)


def test_archivo_oficial_publica_una_imagen_por_cada_bloque_de_producto():
    document = SourceDocument(
        OFFICIAL_SOURCE.name,
        "spec_guide",
        hashlib.sha256(OFFICIAL_SOURCE.read_bytes()).hexdigest(),
        MIME,
        OFFICIAL_SOURCE,
    )

    build = _conceptos_module().build_conceptos_snapshot_with_assets((document,))

    items = build.snapshot["items"]
    assert len(items) == 40
    # Los dos bloques de cushion comparten el mismo PNG oficial.
    assert len(build.assets_by_sha256) == 39
    assert len(build.bindings) == 40
    assert {binding.internal_id for binding in build.bindings} == {
        item["internal_id"] for item in items
    }
    assert all(binding.match_status == "merged_xlsx" for binding in build.bindings)
    assert all(item["image_kind"] == "official" for item in items)
