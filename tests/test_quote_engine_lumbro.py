from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from mobiliti_saas.quote_engine import engine  # noqa: E402
from mobiliti_saas.quote_engine.engine import _load_lumbro_prices, _lumbro_accessories_for_item, _write_mobiliti  # noqa: E402
from mobiliti_saas.quote_engine.parser import QuoteItem  # noqa: E402
from mobiliti_saas.quote_engine.parser import read_items  # noqa: E402


DOWNLOADS = Path(r"C:\Users\pepem\Downloads")
TEMPLATE_DIR = ROOT / "versiones historial" / "HISTORIAL DE VERSIONES" / "Mobiliti_Generador_Windows"
TEMPLATE = next(TEMPLATE_DIR.glob("Formato*.xlsx"), TEMPLATE_DIR / "Formato Cotizacion 2026 GDL (1).xlsx")


def test_lumbro_accessories_for_workstation_pax_multiplies_quantity():
    item = QuoteItem(tipo="producto", row=9, nombre="Estacion Lido 8PAX", cantidad=2)

    accessories = _lumbro_accessories_for_item(item, "Escritorios-WorkStation")

    assert accessories == [
        ("LIDO.OP-INT", 16),
        ("JUMP-1.5M", 16),
        ("CAJA-FUS", 4),
    ]


def test_lumbro_accessories_for_meeting_table_adds_jump_plus_one():
    item = QuoteItem(tipo="producto", row=14, nombre="Sala de juntas para 8 pax", cantidad=1)

    accessories = _lumbro_accessories_for_item(item, "Mesas de Juntas")

    assert accessories == [
        ("MULT-LIDO-INT", 2),
        ("JUMP-1.5M", 3),
    ]


def test_lumbro_accessories_for_workstation_without_pax_adds_default_multicontact():
    item = QuoteItem(tipo="producto", row=21, nombre="DU688-Lido Ejecutivo", cantidad=1)

    accessories = _lumbro_accessories_for_item(item, "Escritorios-WorkStation")

    assert accessories == [("MULT-LIDO-INT", 1)]


def test_lumbro_accessories_for_workstation_without_pax_respects_quantity():
    item = QuoteItem(tipo="producto", row=21, nombre="UP 1 IND Escritorio", cantidad=4)

    accessories = _lumbro_accessories_for_item(item, "Escritorios-WorkStation")

    assert accessories == [("MULT-LIDO-INT", 4)]


def test_mobiliti_lumbro_rows_use_safe_discount_and_region_formulas():
    source = DOWNLOADS / "IZA REFORMA-Quotation Sheet - V1.xlsx"
    if not source.exists() or not TEMPLATE.exists():
        pytest.skip("IZA input/template not available on this machine")

    items, column_map = read_items(source)
    wb = load_workbook(TEMPLATE, data_only=False)
    try:
        ws = wb["Mobiliti"]
        _, lumbro_row_map = _write_mobiliti(ws, items, column_map, _load_lumbro_prices(TEMPLATE))
        lumbro_rows = [row for rows in lumbro_row_map.values() for row in rows]
        assert lumbro_rows

        for row in lumbro_rows:
            assert ws.cell(row, 16).value == "Centro"
            assert str(ws.cell(row, 10).value).startswith("='SPEC-GUIDE-LUMBRO'!E")
            assert str(ws.cell(row, 10).value).endswith("/$K$6")
            assert ws.cell(row, 27).value == f"=MIN(0.4,Z{row})"
            assert ws.cell(row, 28).value == f"=X{row}*AA{row}"
            assert ws.cell(row, 29).value == f'=IF(AA{row}>Z{row},"ERROR",(X{row}-AB{row}))'
            assert ws.cell(row, 30).value == f"=AC{row}*H{row}"
    finally:
        wb.close()


def test_load_lumbro_prices_preserves_source_row_reference():
    if not TEMPLATE.exists():
        pytest.skip("template not available on this machine")

    prices = _load_lumbro_prices(TEMPLATE)

    assert prices["LIDO.OP-INT"].row == 380
    assert prices["LIDO.OP-INT"].price_mxn > 0


def test_exchange_rate_prefers_explicit_metadata(monkeypatch):
    monkeypatch.setattr(engine, "_fetch_usd_mxn_exchange_rate", lambda: 17.25)

    assert engine._exchange_rate({"tipo_cambio": "20"}) == 20


def test_exchange_rate_fetches_live_rate_when_no_metadata(monkeypatch, tmp_path):
    monkeypatch.setattr(engine, "EXCHANGE_RATE_CACHE_PATH", tmp_path / "rate.json")
    monkeypatch.setattr(engine, "_fetch_usd_mxn_exchange_rate", lambda: 17.25)

    assert engine._exchange_rate({}) == 17.25
