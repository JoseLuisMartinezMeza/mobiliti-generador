from __future__ import annotations

import hashlib
from copy import copy
from pathlib import Path
import re
from xml.etree import ElementTree as ET

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

from mobiliti_saas.quote_engine.ooxml_package import XlsxPackage


ROOT = Path(__file__).resolve().parents[1]
SOURCE = (
    Path.home() / "Downloads" / "Formato-Cotizacion-Unico - Sunon-Cdmx-V1C.xlsx"
)
CDMX = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion Sunon CDMX V1C.xlsx"
)


def _formula(cell) -> str:
    value = cell.value
    if isinstance(value, ArrayFormula):
        return value.text
    assert isinstance(value, str) and value.startswith(
        "="
    ), f"{cell.coordinate} debe contener una fórmula"
    return value


def _image_hashes(sheet) -> set[str]:
    return {
        hashlib.sha256(image._data()).hexdigest()
        for image in sheet._images
    }


def _row_style_signature(sheet, row: int) -> tuple[int, ...]:
    return tuple(
        sheet.cell(row=row, column=column).style_id
        for column in range(8, 17)
    )


def _effective_style_owner(sheet, row: int, column: int):
    """Resuelve la herencia OOXML celda -> fila -> columna para celdas vacías."""

    cell = sheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return cell
    if cell.has_style:
        return cell
    row_dimension = sheet.row_dimensions.get(row)
    if row_dimension is not None and row_dimension.has_style:
        return row_dimension
    for column_dimension in sheet.column_dimensions.values():
        if (
            column_dimension.has_style
            and column_dimension.min <= column <= column_dimension.max
        ):
            return column_dimension
    return cell


def _merged_ranges_relative_to(
    sheet,
    *,
    first_row: int,
    last_row: int,
) -> set[tuple[int, int, int, int]]:
    return {
        (
            cell_range.min_col,
            cell_range.min_row - first_row,
            cell_range.max_col,
            cell_range.max_row - first_row,
        )
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row >= first_row
        and cell_range.max_row <= last_row
    }


def test_cotizacion_preserves_official_cdmx_terms_and_conditions() -> None:
    source_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    cdmx_workbook = load_workbook(CDMX, data_only=False, read_only=False)
    source = source_workbook["Cotizacion"]
    cdmx = cdmx_workbook["Cotizacion"]

    source_values = tuple(
        tuple(source.cell(row=row, column=column).value for column in range(1, 11))
        for row in range(35, 84)
    )
    cdmx_values = tuple(
        tuple(cdmx.cell(row=row, column=column).value for column in range(1, 11))
        for row in range(28, 77)
    )
    assert cdmx_values == source_values

    legal_text = "\n".join(
        str(value)
        for row in cdmx_values
        for value in row
        if value not in (None, "")
    )
    assert "COMERCIALIZADORA VICARJOFRAA DE OCCIDENTE" in legal_text
    assert "Pago 70% Anticipo + 20% Contra Aviso de Embarque + 10% Contra Entrega" in legal_text
    assert "10-12 SEMANAS" in legal_text
    assert "Ciudad de México, CDMX" in legal_text
    assert "Pago 60% Anticipo" not in legal_text


def test_cotizacion_preserves_official_cdmx_terms_presentation() -> None:
    source_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    cdmx_workbook = load_workbook(CDMX, data_only=False, read_only=False)
    source = source_workbook["Cotizacion"]
    cdmx = cdmx_workbook["Cotizacion"]
    assert (36, 1) not in source._cells
    assert source.column_dimensions["A"].has_style

    for source_row, cdmx_row in zip(range(35, 84), range(28, 77)):
        # Excel cuantiza ciertas alturas al cuarto de punto al guardar. Una
        # diferencia máxima de 0.10 puntos representa el mismo alto visual.
        source_height = source.row_dimensions[source_row].height or 15
        cdmx_height = cdmx.row_dimensions[cdmx_row].height or 15
        assert abs(source_height - cdmx_height) <= 0.11

        for column in range(1, 11):
            source_cell = source.cell(source_row, column)
            cdmx_cell = cdmx.cell(cdmx_row, column)
            source_style = _effective_style_owner(source, source_row, column)
            cdmx_style = _effective_style_owner(cdmx, cdmx_row, column)
            for style_component in (
                "font",
                "fill",
                "border",
                "alignment",
                "protection",
            ):
                assert copy(getattr(cdmx_style, style_component)) == copy(
                    getattr(source_style, style_component)
                )
            assert cdmx_style.number_format == source_style.number_format
            source_link = (
                source_cell.hyperlink.target
                if source_cell.hyperlink
                else None
            )
            cdmx_link = (
                cdmx_cell.hyperlink.target
                if cdmx_cell.hyperlink
                else None
            )
            assert cdmx_link == source_link

    assert _merged_ranges_relative_to(
        cdmx,
        first_row=28,
        last_row=76,
    ) == _merged_ranges_relative_to(
        source,
        first_row=35,
        last_row=83,
    )


def test_cotizacion_has_no_visible_payload_or_merges_after_print_layout() -> None:
    package = XlsxPackage.read(CDMX)
    root = ET.fromstring(package.parts[package.sheet_part("Cotizacion")])
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sidecar_after_print_end = False

    for cell in root.findall(f".//{{{main}}}c"):
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", cell.attrib["r"])
        assert match is not None
        column = column_index_from_string(match.group(1))
        row = int(match.group(2))
        if row <= 76:
            continue
        if column > 10:
            sidecar_after_print_end = True
            continue
        assert cell.find(f"{{{main}}}f") is None
        assert cell.find(f"{{{main}}}v") is None
        assert cell.find(f"{{{main}}}is") is None

    merge_refs = [
        node.attrib["ref"]
        for node in root.findall(f"{{{main}}}mergeCells/{{{main}}}mergeCell")
    ]
    assert len(merge_refs) == len(set(merge_refs))
    assert not any(
        min_column <= 10 and max_row > 76
        for min_column, _min_row, _max_column, max_row in map(
            range_boundaries,
            merge_refs,
        )
    )
    assert sidecar_after_print_end


def test_cotizacion_has_no_official_payload_in_cdmx_residue_rows() -> None:
    package = XlsxPackage.read(CDMX)
    root = ET.fromstring(package.parts[package.sheet_part("Cotizacion")])
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    for cell in root.findall(f".//{{{main}}}c"):
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", cell.attrib["r"])
        assert match is not None
        column = column_index_from_string(match.group(1))
        row = int(match.group(2))
        if column > 10 or not 19 <= row <= 27:
            continue
        assert cell.find(f"{{{main}}}f") is None
        assert cell.find(f"{{{main}}}v") is None
        assert cell.find(f"{{{main}}}is") is None

    prototype = root.find(f".//{{{main}}}c[@r='J18']/{{{main}}}f")
    assert prototype is not None and prototype.text == "SUM(J13:J15)"


def test_cotizacion_exposes_cdmx_section_subtotal_prototype() -> None:
    source_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    cdmx_workbook = load_workbook(CDMX, data_only=False, read_only=False)
    source = source_workbook["Cotizacion"]
    cdmx = cdmx_workbook["Cotizacion"]

    assert cdmx["I18"].value == source["I16"].value == "SUBTOTAL AREA"
    assert cdmx["J18"].value == source["J16"].value == "=SUM(J13:J15)"
    assert cdmx.row_dimensions[18].height == source.row_dimensions[16].height


def test_cotizacion_adopts_cdmx_logo() -> None:
    source_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    cdmx_workbook = load_workbook(CDMX, data_only=False, read_only=False)
    source = source_workbook["Cotizacion"]
    cdmx = cdmx_workbook["Cotizacion"]

    assert _image_hashes(cdmx) == _image_hashes(source)


def test_cotizacion_adopts_cdmx_view_in_canonical_print_area() -> None:
    source_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    cdmx_workbook = load_workbook(CDMX, data_only=False, read_only=False)
    source = source_workbook["Cotizacion"]
    cdmx = cdmx_workbook["Cotizacion"]

    assert cdmx.sheet_view.zoomScale == source.sheet_view.zoomScale == 40
    assert str(cdmx.print_area) == "'Cotizacion'!$A$1:$J$76"


def test_cotizacion_maps_cdmx_merges_without_blocking_functional_anchors() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    cdmx = workbook["Cotizacion"]

    assert {
        "A14:J14",
        "A16:J16",
        "B11:G11",
        "D20:G20",
        "H20:J20",
        "D28:J28",
    }.issubset({str(cell_range) for cell_range in cdmx.merged_cells.ranges})

    for coordinate in ("B3", "B7", "B8", "B9", "B10", "B11", "B12"):
        assert not isinstance(cdmx[coordinate], MergedCell), (
            f"{coordinate} debe seguir siendo el ancla escribible del motor"
        )


def test_cantidades_lumbro_uses_only_live_mobiliti_references() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    sheet = workbook["Cantidades Lumbro "]

    code = _formula(sheet["H4"])
    quantity = _formula(sheet["I4"])
    unit_cost = _formula(sheet["L4"])

    for formula in (code, quantity, unit_cost):
        assert "Mobiliti!$F$14:$F$5000" in formula
        assert 'SEARCH("Lumbro"' in formula
        assert "_xlws.FILTER" in formula
    assert "Mobiliti!$D$14:$D$5000" in code
    assert "Mobiliti!$H$14:$H$5000" in quantity
    assert "Mobiliti!$J$14:$J$5000" in unit_cost


def test_cantidades_lumbro_has_no_hardcoded_product_data_or_exchange_rate() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    sheet = workbook["Cantidades Lumbro "]

    for coordinate in ("H4", "I4", "J4", "K4", "L4", "M4", "N4", "P4"):
        normalized = _formula(sheet[coordinate]).replace(" ", "")
        assert "/18" not in normalized
        assert "/18.5" not in normalized
        assert "COSTOLUMBRO" not in normalized.replace("'", "").replace("!", "")


def test_cantidades_lumbro_has_dynamic_capacity_and_live_totals() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    sheet = workbook["Cantidades Lumbro "]

    assert sheet["M2"].value == 0.40
    assert sheet["I3"].value == "# ESTACIONES"
    assert sheet["L3"].value == "Costo unitario"
    assert sheet["P3"].value == "TOTALES VENTA"
    assert "ANCHORARRAY(H4)" in _formula(sheet["J4"])
    assert "ANCHORARRAY(I4)" in _formula(sheet["K4"])
    assert "ANCHORARRAY(L4)" in _formula(sheet["M4"])
    assert "ANCHORARRAY(N4)" in _formula(sheet["P4"])
    assert sheet.column_dimensions["O"].hidden is True


def test_cantidades_lumbro_has_no_sample_images() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    sheet = workbook["Cantidades Lumbro "]

    assert sheet._images == []


def test_cantidades_lumbro_offers_compact_capacity_for_25_rows() -> None:
    workbook = load_workbook(CDMX, data_only=False, read_only=False)
    sheet = workbook["Cantidades Lumbro "]

    assert 28 <= sheet.max_row < 100
    assert str(sheet.print_area) == "'Cantidades Lumbro '!$H$1:$P$28"

    product_style = _row_style_signature(sheet, 4)
    assert product_style != (0,) * 9
    for row in range(4, 29):
        assert _row_style_signature(sheet, row) == product_style
        effective_height = (
            sheet.row_dimensions[row].height
            or sheet.sheet_format.defaultRowHeight
            or 15
        )
        assert effective_height <= 30, (
            f"Fila {row} no es compacta: {effective_height}"
        )
