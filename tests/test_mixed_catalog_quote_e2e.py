from __future__ import annotations

from copy import deepcopy
from dataclasses import replace
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import hashlib
import importlib.util
import json
from pathlib import Path
import re
import sys
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from PIL import Image
import pytest

from mobiliti_saas.quote_engine import catalog_cart, generate_quote, quotation_sheets
from mobiliti_saas.quote_engine import engine as quote_engine
from mobiliti_saas.quote_engine.mixed_catalog import (
    build_mixed_catalog_cart_payload,
    create_mixed_catalog_quotation_workbook,
    mixed_cart_key,
)
from mobiliti_saas.quote_engine.offiho_catalog import OffihoCatalogItem
from mobiliti_saas.quote_engine.quotation_import import build_import_manifest
from mobiliti_saas.quote_engine.quotation_sheets import (
    QUOTATION_DATA_HEADERS,
    quotation_data_rows,
)
from mobiliti_saas.quote_engine.tarkett_catalog import TarkettCatalogItem
from quotation_import_fixtures import write_import_fixture


ROOT = Path(__file__).resolve().parents[1]
WORKER_DIR = ROOT / "mobiliti_saas" / "worker"
WORKER_TEMPLATE = (
    WORKER_DIR / "templates" / "Formato Cotizacion 2026 Oficial.xlsx"
)
MONEY = Decimal("0.01")
CATALOGS = (
    "tarkett",
    "offiho",
    "cr-global",
    "sonara",
    "sunon",
    "alma",
    "lumbro",
)
CATALOG_LABELS = {
    "tarkett": "Tarkett",
    "offiho": "Offiho",
    "cr-global": "CR Global",
    "sonara": "Sonara",
    "sunon": "Sunon",
    "alma": "ALMA",
    "lumbro": "Lumbro",
}
SOURCE_HASHES = {
    catalog: hashlib.sha256(f"mixed-e2e:{catalog}".encode("utf-8")).hexdigest()
    for catalog in CATALOGS
}

@pytest.fixture
def isolated_quote_runtime(monkeypatch):
    monkeypatch.setenv("SUPABASE_URL", "http://localhost:54321")
    monkeypatch.setenv("SUPABASE_SERVICE_KEY", "test-key")
    monkeypatch.setenv("JWT_SECRET_KEY", "test-secret-32-chars-long!!!!!")
    monkeypatch.setenv(
        "CATALOG_ENABLED_SUPPLIERS", "cr-global,sonara,sunon,alma,lumbro"
    )
    monkeypatch.syspath_prepend(str(WORKER_DIR))
    modules_before = set(sys.modules)
    suffix = uuid.uuid4().hex

    def load_module(name: str, path: Path):
        spec = importlib.util.spec_from_file_location(name, path)
        assert spec is not None and spec.loader is not None
        module = importlib.util.module_from_spec(spec)
        sys.modules[name] = module
        spec.loader.exec_module(module)
        return module

    api_index = load_module(
        f"mixed_quote_e2e_api_{suffix}",
        ROOT / "mobiliti_saas" / "api" / "index.py",
    )
    quote_worker = load_module(
        f"mixed_quote_e2e_worker_{suffix}",
        WORKER_DIR / "quote_worker.py",
    )
    assert api_index._enabled_catalog_suppliers() == CATALOGS[2:]
    try:
        yield api_index, quote_worker
    finally:
        for name in set(sys.modules) - modules_before:
            module_path = getattr(sys.modules.get(name), "__file__", None)
            if name.endswith(suffix) or (
                module_path
                and Path(module_path).resolve().is_relative_to(WORKER_DIR.resolve())
            ):
                sys.modules.pop(name, None)


def _supplier_item(
    supplier: str,
    *,
    internal_id: str,
    name: str,
    price: str,
    currency: str,
    sku: str | None = None,
    code_status: str = "verified",
    image_url: str = "",
    image_kind: str = "placeholder",
    warnings: list[str] | None = None,
    base_price_options: list[dict] | None = None,
    add_on_options: list[dict] | None = None,
) -> dict:
    return {
        "internal_id": internal_id,
        "supplier": supplier,
        "product_key": internal_id,
        "sku": (sku or internal_id.upper()) if code_status == "verified" else "",
        "code_status": code_status,
        "brand": CATALOG_LABELS[supplier],
        "collection": "E2E",
        "name": name,
        "description": f"Descripcion {name}",
        "unit": "PZA",
        "availability_type": "stocked",
        "stock": "20.000000",
        "lead_time": "Entrega inmediata",
        "base_price_options": base_price_options or [],
        "add_on_options": add_on_options or [],
        "base_currency": currency,
        "price_net": price,
        "tax_rate": "0.160000",
        "attributes": {},
        "image_url": image_url,
        "image_kind": image_kind,
        "product_url": "",
        "warnings": warnings or [],
        "source_reference": f"{supplier}:e2e:{internal_id}",
    }


def authoritative_catalogs() -> dict[str, dict]:
    tarkett = TarkettCatalogItem(
        code="TARK-E2E",
        name="Estacion Lido 8PAX Tarkett",
        unit="PZA",
        available_quantity=Decimal("20"),
        image_url="https://media.tarkett-image.com/e2e-tarkett.png",
        unit_price=Decimal("1000"),
        price_source="catalog",
    )
    offiho = OffihoCatalogItem(
        inventory_key="OFF-E2E NEGRO",
        code="OFF-E2E",
        name="Silla Offiho",
        variant="Negro",
        unit="PZA",
        pieces_per_box=Decimal("1"),
        available_quantity=Decimal("20"),
        unit_price=Decimal("500"),
        image_url="https://offiho.com.mx/e2e-offiho.png",
        price_source="catalog",
    )
    alma_options = [
        {
            "id": "alma-base",
            "name": "Base operativa",
            "price_net": "100.000000",
            "available": True,
        }
    ]
    alma_add_ons = [
        {
            "id": "alma-electrificacion-a",
            "name": "Electrificacion A",
            "family": "electrificacion",
            "price_net": "10.000000",
            "available": True,
            "compatible_base_option_ids": ["alma-base"],
        },
        {
            "id": "alma-pasacables-b",
            "name": "Pasacables B",
            "family": "pasacables",
            "price_net": "20.000000",
            "available": True,
            "compatible_base_option_ids": ["alma-base"],
        },
    ]
    supplier_items = {
        "cr-global": _supplier_item(
            "cr-global", internal_id="cr:e2e", name="Silla CR Global",
            price="300.000000", currency="MXN", sku="CR-E2E",
        ),
        "sonara": _supplier_item(
            "sonara", internal_id="sonara:e2e-review",
            name="Panel Sonara por verificar", price="400.000000",
            currency="MXN", code_status="needs_review",
            warnings=["Revision documental local"],
        ),
        "sunon": _supplier_item(
            "sunon", internal_id="sunon:e2e", name="Silla Sunon",
            price="50.000000", currency="USD", sku="SUN-E2E",
        ),
        "alma": _supplier_item(
            "alma", internal_id="alma:e2e-configurable", name="Mesa ALMA",
            price="0.000000", currency="USD", sku="ALMA-E2E",
            image_url="https://alma.example.test/e2e-alma.png",
            image_kind="official", base_price_options=alma_options,
            add_on_options=alma_add_ons,
        ),
        "lumbro": _supplier_item(
            "lumbro", internal_id="lumbro:e2e-manual",
            name="LIDO.OP-INT manual", price="250.000000",
            currency="MXN", sku="LUMBRO-MANUAL",
        ),
    }
    catalogs = {
        "tarkett": {
            "source_hash": SOURCE_HASHES["tarkett"],
            "items": [tarkett],
            "by_code": {tarkett.code: tarkett},
        },
        "offiho": {
            "source_hash": SOURCE_HASHES["offiho"],
            "items": [offiho],
            "by_inventory_key": {offiho.inventory_key: offiho},
        },
    }
    for supplier, item in supplier_items.items():
        catalogs[supplier] = {
            "supplier": supplier,
            "source_hash": SOURCE_HASHES[supplier],
            "generated_at": "2026-07-19T00:00:00+00:00",
            "items": [item],
        }
    return catalogs


def browser_rows_for_all_catalogs_and_two_alma_configs() -> list[dict]:
    return [
        {"catalog": "tarkett", "code": "TARK-E2E", "quantity": "1"},
        {
            "catalog": "offiho",
            "inventory_key": "OFF-E2E NEGRO",
            "quantity": "2",
        },
        {"catalog": "cr-global", "internal_id": "cr:e2e", "quantity": "1"},
        {
            "catalog": "sonara",
            "internal_id": "sonara:e2e-review",
            "quantity": "1",
        },
        {"catalog": "sunon", "internal_id": "sunon:e2e", "quantity": "1"},
        {
            "catalog": "alma",
            "internal_id": "alma:e2e-configurable",
            "quantity": "1",
            "base_option_id": "alma-base",
            "add_on_option_ids": ["alma-electrificacion-a"],
        },
        {
            "catalog": "alma",
            "internal_id": "alma:e2e-configurable",
            "quantity": "1",
            "base_option_id": "alma-base",
            "add_on_option_ids": ["alma-pasacables-b"],
        },
        {
            "catalog": "lumbro",
            "internal_id": "lumbro:e2e-manual",
            "quantity": "1",
        },
    ]


def rate_rows() -> list[dict]:
    today = date.today().isoformat()
    return [
        {
            "currency": "USD",
            "effective_date": today,
            "mxn_per_unit": "18.500000",
            "retrieved_at": f"{today}T12:00:00+00:00",
        },
        {
            "currency": "EUR",
            "effective_date": today,
            "mxn_per_unit": "20.500000",
            "retrieved_at": f"{today}T12:00:00+00:00",
        },
    ]


def install_api_boundary(monkeypatch, api_index, catalogs: dict) -> dict:
    state = {
        "events": [],
        "jobs": [],
        "uploads": [],
        "released": [],
        "deleted_jobs": [],
        "deleted_inputs": [],
    }
    monkeypatch.setattr(
        api_index,
        "db_get_usuario_by_id",
        lambda user_id: {
            "id": user_id,
            "email": "cliente@example.test",
            "nombre": "Cliente",
            "empresa": "Mobiliti",
            "es_admin": False,
            "activo": True,
        },
    )
    monkeypatch.setattr(
        api_index,
        "db_get_suscripcion_by_usuario",
        lambda user_id: {
            "id": 1,
            "usuario_id": user_id,
            "estado": "activa",
            "plan": "mensual",
            "fecha_fin": "2099-01-01T00:00:00+00:00",
        },
    )
    monkeypatch.setattr(
        api_index,
        "_enforce_active_quote_limit",
        lambda _user, **_kwargs: None,
    )
    monkeypatch.setattr(api_index, "_next_quote_number_for_user", lambda _user: None)
    monkeypatch.setattr(api_index, "_storage_provider_name", lambda: "supabase")
    monkeypatch.setattr(
        api_index, "_load_tarkett_catalog_cached", lambda: catalogs["tarkett"]
    )
    monkeypatch.setattr(
        api_index, "_load_offiho_catalog_cached", lambda: catalogs["offiho"]
    )
    monkeypatch.setattr(
        api_index,
        "_load_supplier_catalog_cached",
        lambda supplier: catalogs[supplier],
    )
    monkeypatch.setattr(api_index, "db_list_exchange_rates", rate_rows)

    def create_job(user_id, template, metadata, input_path, job_id=None):
        state["events"].append("create_job")
        job = {
            "id": job_id,
            "usuario_id": user_id,
            "status": "draft",
            "template": template,
            "metadata": deepcopy(metadata),
            "input_path": input_path,
        }
        state["jobs"].append(job)
        return deepcopy(job)

    def reserve_mixed(_user_id, _job_id, groups):
        state["events"].append("reserve_mixed")
        return [
            {
                "catalog": group["catalog"],
                "identity": item["identity"],
                "reserved_before": "0.000000",
                "available_before": item["stock"],
                "insufficient": False,
                "reserved_by_others": False,
            }
            for group in groups
            for item in group["items"]
        ]

    def upload(path, content, content_type="application/octet-stream"):
        state["events"].append("upload")
        state["uploads"].append((path, bytes(content), content_type))

    def queue(job_id, metadata):
        state["events"].append("queue")
        job = next(job for job in state["jobs"] if job["id"] == job_id)
        job.update(status="queued", metadata=deepcopy(metadata))
        return deepcopy(job)

    monkeypatch.setattr(api_index, "db_create_quote_job", create_job)
    monkeypatch.setattr(api_index, "db_reserve_mixed_cart", reserve_mixed)
    monkeypatch.setattr(api_index, "_storage_upload_bytes", upload)
    monkeypatch.setattr(api_index, "db_queue_mixed_quote_job", queue)
    monkeypatch.setattr(
        api_index,
        "db_release_mixed_cart",
        lambda job_id: state["released"].append(job_id),
    )
    monkeypatch.setattr(
        api_index,
        "db_delete_quote_job",
        lambda job_id: state["deleted_jobs"].append(job_id),
    )
    monkeypatch.setattr(
        api_index,
        "_delete_storage_paths",
        lambda paths: state["deleted_inputs"].extend(paths),
    )
    monkeypatch.setattr(api_index, "_wake_worker", lambda: state["events"].append("wake"))
    return state


class EndToEndWorkerClient:
    EXPECTED_EVENTS = (
        "claim",
        "progress:45",
        "download",
        "progress:55",
        "converter",
        "generator",
        "progress:90",
        "upload",
        "completed",
        "delete",
        "clear_input",
    )

    def __init__(self, job: dict, objects: dict[str, bytes]):
        self.job = deepcopy(job)
        self.objects = dict(objects)
        self.events: list[str] = []
        self.downloads: list[str] = []
        self.uploads: list[tuple[str, bytes]] = []
        self.deleted_inputs: list[str] = []
        self.completed_updates: list[dict] = []
        self.failed_updates: list[dict] = []

    def record_event(self, event: str) -> None:
        assert len(self.events) < len(self.EXPECTED_EVENTS)
        assert event == self.EXPECTED_EVENTS[len(self.events)]
        self.events.append(event)

    def rest(self, method, path, params=None, data=None):
        assert method == "PATCH"
        assert params is None
        data = deepcopy(data or {})
        job_path = f"/saas_quote_jobs?id=eq.{self.job['id']}"
        if path == job_path + "&status=eq.queued":
            assert set(data) == {
                "status", "attempt_token", "lease_expires_at", "updated_at",
            }
            assert data["status"] == "processing"
            assert data["attempt_token"]
            assert data["lease_expires_at"]
            assert isinstance(data["updated_at"], str) and data["updated_at"]
            self.record_event("claim")
            claimed = {**self.job, **data}
            self.job = claimed
            return [deepcopy(claimed)]
        fenced_processing = (
            job_path + "&status=eq.processing"
            f"&attempt_token=eq.{self.job.get('attempt_token')}"
        )
        fenced_completed = (
            job_path + "&status=eq.completed"
            f"&attempt_token=eq.{self.job.get('attempt_token')}"
        )
        if path == fenced_processing and set(data) == {
            "lease_expires_at", "updated_at",
        }:
            self.job = {**self.job, **data}
            return [deepcopy(self.job)]
        if path == fenced_processing and set(data) == {
            "metadata", "lease_expires_at", "updated_at",
        }:
            metadata = data["metadata"]
            assert isinstance(metadata, dict)
            progress = metadata.get("progress_percent")
            assert progress in {45, 55, 90}
            self.record_event(f"progress:{progress}")
            self.job = {**self.job, **data}
            return [deepcopy(self.job)]
        if path == fenced_processing and data.get("status") == "completed":
            assert set(data) == {
                "status", "output_path", "metadata", "error_message",
                "lease_expires_at", "updated_at", "completed_at",
            }
            assert data["output_path"].startswith(
                f"users/{self.job['usuario_id']}/jobs/{self.job['id']}/attempts/"
            )
            assert data["output_path"].endswith(
                f"/{self.job['attempt_token']}/output.xlsx"
            )
            assert data["metadata"]["progress_percent"] == 100
            assert data["error_message"] is None
            assert data["lease_expires_at"] is None
            self.record_event("completed")
            completed = {**self.job, **data}
            self.job = completed
            self.completed_updates.append(deepcopy(completed))
            return [deepcopy(completed)]
        if path == fenced_completed and set(data) == {"input_path", "updated_at"}:
            assert data["input_path"] is None
            self.record_event("clear_input")
            self.job = {**self.job, **data}
            self.completed_updates[-1] = deepcopy(self.job)
            return [deepcopy(self.job)]
        if path == fenced_processing and data.get("status") == "failed":
            assert set(data) == {
                "status", "metadata", "error_message", "lease_expires_at", "updated_at"
            }
            failed = {**self.job, **data}
            self.job = failed
            self.failed_updates.append(deepcopy(failed))
            return [deepcopy(failed)]
        raise AssertionError(f"Contrato REST worker inesperado: {method} {path} {data}")

    def storage_download(self, object_path, destination):
        self.record_event("download")
        assert object_path == self.job["input_path"]
        assert object_path in self.objects
        self.downloads.append(object_path)
        Path(destination).write_bytes(self.objects[object_path])

    def storage_upload(self, object_path, source):
        self.record_event("upload")
        assert object_path.startswith(
            f"users/{self.job['usuario_id']}/jobs/{self.job['id']}/attempts/"
        )
        assert object_path.endswith(f"/{self.job['attempt_token']}/output.xlsx")
        assert Path(source).is_file()
        self.uploads.append((object_path, Path(source).read_bytes()))

    def storage_delete(self, object_path):
        self.record_event("delete")
        assert object_path == self.job["input_path"]
        assert object_path in self.objects
        self.deleted_inputs.append(object_path)
        self.objects.pop(object_path)


def auth_headers(api_index) -> dict[str, str]:
    token = api_index.create_access_token(
        {"sub": "7", "email": "cliente@example.test"}
    )
    return {"Authorization": f"Bearer {token}"}


def money(value: Decimal) -> Decimal:
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def expected_mixed_totals(
    payload: dict,
    accessories_by_parent: dict[str, list[tuple[Decimal, Decimal]]],
) -> tuple[Decimal, Decimal, Decimal, Decimal, Decimal]:
    net = Decimal("0")
    auto_rate = Decimal(payload["auto_electrification_rate"]["exchange_rate"])
    for group in payload["groups"]:
        for item in group["items"]:
            price = Decimal(item["unit_price"])
            discount = Decimal("0.40")
            quantity = Decimal(item["quantity"])
            accessory_total = sum(
                (
                    money(unit_price_mxn * auto_rate) * accessory_quantity
                    for unit_price_mxn, accessory_quantity
                    in accessories_by_parent.get(item["canonical_key"], [])
                ),
                Decimal("0"),
            )
            combined_unit = money(((price * quantity) + accessory_total) / quantity)
            discount_amount = money(combined_unit * discount)
            net_unit = money(combined_unit - discount_amount)
            net += money(net_unit * quantity)
    net = money(net)
    freight = money(net * Decimal("0.12"))
    before_tax = money(net + freight)
    tax = money(before_tax * Decimal("0.16"))
    return net, freight, before_tax, tax, money(before_tax + tax)


def _row_for_formula(ws, column: int, formula: str) -> int:
    return next(
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, column).value == formula
    )


def _rows_with_value(ws, column: int, value) -> list[int]:
    return [
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, column).value == value
    ]


def _ordered_rows_for_values(ws, column: int, values) -> list[int]:
    rows = []
    cursor = 0
    for value in values:
        cursor = next(
            row
            for row in range(cursor + 1, ws.max_row + 1)
            if ws.cell(row, column).value == value
        )
        rows.append(cursor)
    return rows


def _assert_exact_quotation_data(audit, expected_rows) -> None:
    assert audit.sheet_state == "veryHidden"
    assert tuple(audit.cell(1, column).value for column in range(1, 17)) == (
        QUOTATION_DATA_HEADERS
    )
    assert audit.max_row == len(expected_rows) + 1
    decimal_fields = {
        "original_cost",
        "frozen_rate",
        "converted_cost",
        "quantity",
    }
    for row_number, expected in enumerate(expected_rows, start=2):
        for column, field in enumerate(QUOTATION_DATA_HEADERS, start=1):
            actual = audit.cell(row_number, column).value
            wanted = getattr(expected, field)
            if field in decimal_fields:
                assert Decimal(str(actual)) == wanted
            else:
                assert actual == wanted


def _mixed_import_metadata(payload: dict, quote_currency: str) -> dict:
    return {
        "catalog_price_mode": "mixed_catalog_converted",
        "catalog_source_hashes": {
            group["catalog"]: group["catalog_source_hash"]
            for group in payload["groups"]
        },
        "quote_currency": quote_currency,
        "rate_summary": deepcopy(payload["rate_summary"]),
        "auto_electrification_rate": deepcopy(
            payload["auto_electrification_rate"]
        ),
        "descuento": 40,
        "cotizacion": f"IMPORT-{quote_currency}",
        "proyecto": "Importado y catalogo",
        "cliente": "Cliente E2E",
    }


def _assert_blank_mobiliti_formulas(ws, row: int) -> None:
    input_columns = (4, 5, 6, 8, 10, 11, 16)
    guard = (
        f'=IF(COUNTA($D{row},$E{row},$F{row},$H{row},$J{row},$K{row})=0,"",'
    )
    assert all(ws.cell(row, column).value is None for column in input_columns)
    assert str(ws.cell(row, 23).value).startswith(guard)
    assert str(ws.cell(row, 24).value).startswith(guard)
    assert ws.cell(row, 35).value == f'{guard}IF(AH{row}<30%,"ERROR","OK"))'


def _one_discount_totals(payload: dict) -> tuple[Decimal, ...]:
    items = {
        item["line_id"]: item
        for item in [
            *(item for group in payload["groups"] for item in group["items"]),
            *payload["imported_source"]["items"],
        ]
    }
    subtotal = Decimal("0")
    for section in payload["sections"]:
        for key in section["line_ids"]:
            item = items[key]
            price = Decimal(item["unit_price"])
            quantity = Decimal(item["quantity"])
            discount = (price * Decimal("0.40")).quantize(
                MONEY, rounding=ROUND_HALF_UP
            )
            net_unit = (price - discount).quantize(MONEY, rounding=ROUND_HALF_UP)
            subtotal += (quantity * net_unit).quantize(MONEY, rounding=ROUND_HALF_UP)
    subtotal = subtotal.quantize(MONEY, rounding=ROUND_HALF_UP)
    freight = (subtotal * Decimal("0.12")).quantize(MONEY, rounding=ROUND_HALF_UP)
    before_tax = (subtotal + freight).quantize(MONEY, rounding=ROUND_HALF_UP)
    tax = (before_tax * Decimal("0.16")).quantize(MONEY, rounding=ROUND_HALF_UP)
    return subtotal, freight, before_tax, tax, (before_tax + tax).quantize(MONEY)


@pytest.mark.parametrize(
    ("mismatch", "message"),
    (("count", "cantidad canónica"), ("value", "quantity")),
)
def test_authoritative_base_mismatch_fails_before_output(
    mismatch,
    message,
    monkeypatch,
    tmp_path,
):
    rows = [
        {
            "catalog": "offiho",
            "inventory_key": "OFF-E2E NEGRO",
            "quantity": "1",
        },
        {"catalog": "cr-global", "internal_id": "cr:e2e", "quantity": "1"},
    ]
    payload = build_mixed_catalog_cart_payload(
        rows,
        catalogs=authoritative_catalogs(),
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Validacion",
            "item_keys": [mixed_cart_key(row) for row in rows],
        }],
        today=date.today(),
    )
    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / f"mismatch-{mismatch}-source.xlsx",
        image_dir=tmp_path / f"mismatch-{mismatch}-images",
    )
    canonical_rows = quotation_data_rows(payload)
    if mismatch == "count":
        handed_off = canonical_rows[:-1]
    else:
        changed = deepcopy(payload)
        changed["groups"][0]["items"][0]["quantity"] = "2.000000"
        handed_off = quotation_data_rows(changed)
    output = tmp_path / f"mismatch-{mismatch}-must-not-exist.xlsx"

    with pytest.raises(ValueError, match=message):
        generate_quote(
            parser_source,
            output,
            _mixed_import_metadata(payload, "MXN"),
            WORKER_TEMPLATE,
            original_quotation_path=None,
            quotation_data_rows=handed_off,
        )

    assert not output.exists()


def test_authoritative_catalog_origin_uses_canonical_key_not_source_reference(
    monkeypatch,
    tmp_path,
):
    catalogs = authoritative_catalogs()
    catalogs["cr-global"]["items"][0]["source_reference"] = (
        '[{"cell_or_bbox":"B565","sheet_or_page":"COSTO CR GLOBAL"}]'
    )
    row = {"catalog": "cr-global", "internal_id": "cr:e2e", "quantity": "1"}
    payload = build_mixed_catalog_cart_payload(
        [row],
        catalogs=catalogs,
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Validacion",
            "item_keys": [mixed_cart_key(row)],
        }],
        today=date.today(),
    )
    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "json-source-reference.xlsx",
        image_dir=tmp_path / "json-source-reference-images",
    )
    output = tmp_path / "json-source-reference-output.xlsx"

    generate_quote(
        parser_source,
        output,
        _mixed_import_metadata(payload, "MXN"),
        WORKER_TEMPLATE,
        original_quotation_path=None,
        quotation_data_rows=quotation_data_rows(payload),
    )

    assert output.is_file()


def test_authoritative_catalog_origin_rejects_another_catalog_with_valid_row_hash(
    monkeypatch,
    tmp_path,
):
    row = {"catalog": "cr-global", "internal_id": "cr:e2e", "quantity": "1"}
    payload = build_mixed_catalog_cart_payload(
        [row],
        catalogs=authoritative_catalogs(),
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Validacion",
            "item_keys": [mixed_cart_key(row)],
        }],
        today=date.today(),
    )
    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "forged-origin-source.xlsx",
        image_dir=tmp_path / "forged-origin-images",
    )
    canonical = quotation_data_rows(payload)[0]
    forged = quotation_sheets._with_canonical_hash(
        replace(canonical, origin="sunon", row_hash="")
    )
    output = tmp_path / "forged-origin-must-not-exist.xlsx"

    with pytest.raises(ValueError, match="origin"):
        generate_quote(
            parser_source,
            output,
            _mixed_import_metadata(payload, "MXN"),
            WORKER_TEMPLATE,
            original_quotation_path=None,
            quotation_data_rows=(forged,),
        )

    assert not output.exists()


def _identical_offiho_payload_with_lumbro_parents() -> dict:
    catalogs = authoritative_catalogs()
    identical_items = [
        OffihoCatalogItem(
            inventory_key=inventory_key,
            code="OFF-LIDO",
            name="Estacion Lido 8PAX",
            variant="Negro",
            unit="PZA",
            pieces_per_box=Decimal("1"),
            available_quantity=Decimal("20"),
            unit_price=Decimal("500"),
            price_source="catalog",
        )
        for inventory_key in ("OFF-LIDO-A", "OFF-LIDO-B")
    ]
    catalogs["offiho"] = {
        "source_hash": SOURCE_HASHES["offiho"],
        "items": identical_items,
        "by_inventory_key": {
            item.inventory_key: item for item in identical_items
        },
    }
    rows = [
        {
            "catalog": "offiho",
            "inventory_key": item.inventory_key,
            "quantity": "1",
        }
        for item in identical_items
    ]
    return build_mixed_catalog_cart_payload(
        rows,
        catalogs=catalogs,
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Identidad",
            "item_keys": [mixed_cart_key(row) for row in rows],
        }],
        today=date.today(),
    )


def test_authoritative_identity_rejects_inverted_identical_lines_and_keeps_lumbro_parents(
    monkeypatch,
    tmp_path,
):
    payload = _identical_offiho_payload_with_lumbro_parents()
    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "identical-offiho-source.xlsx",
        image_dir=tmp_path / "identical-offiho-images",
    )
    canonical_rows = quotation_data_rows(payload)
    correct_output = tmp_path / "identical-offiho-correct.xlsx"

    generate_quote(
        parser_source,
        correct_output,
        _mixed_import_metadata(payload, "MXN"),
        WORKER_TEMPLATE,
        original_quotation_path=None,
        quotation_data_rows=canonical_rows,
    )

    correct = load_workbook(correct_output, data_only=False)
    try:
        audit = correct["Quotation_Data"]
        keys = [audit.cell(row, 1).value for row in range(2, audit.max_row + 1)]
        first, second = (row.item_key for row in canonical_rows)
        assert keys == [
            first,
            f"{first}:lumbro:1:LIDO.OP-INT",
            f"{first}:lumbro:2:JUMP-1.5M",
            f"{first}:lumbro:3:CAJA-FUS",
            second,
            f"{second}:lumbro:1:LIDO.OP-INT",
            f"{second}:lumbro:2:JUMP-1.5M",
            f"{second}:lumbro:3:CAJA-FUS",
        ]
    finally:
        correct.close()

    reordered = deepcopy(payload)
    reordered["sections"][0]["line_ids"].reverse()
    inverted_rows = quotation_data_rows(reordered)
    rejected_output = tmp_path / "identical-offiho-inverted-must-not-exist.xlsx"

    with pytest.raises(ValueError, match="item_key"):
        generate_quote(
            parser_source,
            rejected_output,
            _mixed_import_metadata(payload, "MXN"),
            WORKER_TEMPLATE,
            original_quotation_path=None,
            quotation_data_rows=inverted_rows,
        )

    assert not rejected_output.exists()
    parser_book = load_workbook(parser_source, data_only=False)
    try:
        quotation = parser_book["Quotation"]
        assert tuple(quotation.cell(7, column).value for column in range(20, 24)) == (
            "Canonical Key",
            "Source Hash",
            "Original Source Row",
            "Upstream Row Hash",
        )
        assert all(
            quotation.column_dimensions[column].hidden
            for column in ("T", "U", "V", "W")
        )
    finally:
        parser_book.close()


@pytest.mark.parametrize(
    ("technical_column", "replacement", "message"),
    (
        (20, None, "canonical_key"),
        (21, "f" * 64, "source_hash"),
    ),
)
def test_authoritative_handoff_rejects_missing_or_mismatched_parser_identity_before_output(
    technical_column,
    replacement,
    message,
    monkeypatch,
    tmp_path,
):
    payload = _identical_offiho_payload_with_lumbro_parents()
    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: None,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "missing-identity-source.xlsx",
        image_dir=tmp_path / "missing-identity-images",
    )
    parser_book = load_workbook(parser_source, data_only=False)
    try:
        quotation = parser_book["Quotation"]
        first_product_row = next(
            row
            for row in range(8, quotation.max_row + 1)
            if isinstance(quotation.cell(row, 1).value, (int, float))
        )
        quotation.cell(first_product_row, technical_column).value = replacement
        parser_book.save(parser_source)
    finally:
        parser_book.close()
    output = tmp_path / "missing-identity-must-not-exist.xlsx"

    with pytest.raises(ValueError, match=message):
        generate_quote(
            parser_source,
            output,
            _mixed_import_metadata(payload, "MXN"),
            WORKER_TEMPLATE,
            original_quotation_path=None,
            quotation_data_rows=quotation_data_rows(payload),
        )

    assert not output.exists()


@pytest.mark.parametrize(
    ("technical_column", "replacement", "message"),
    (
        (22, 9, "source_row"),
        (23, "f" * 64, "upstream_row_hash"),
    ),
)
def test_authoritative_imported_identity_mismatch_fails_before_output(
    technical_column,
    replacement,
    message,
    monkeypatch,
    tmp_path,
):
    imported_source = write_import_fixture(tmp_path / "identity-import-source.xlsx")
    manifest, _images = build_import_manifest(
        imported_source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=imported_source.name,
    )
    imported_key = f"import:{manifest['import_id']}:11"
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Importados",
            "item_keys": [imported_key],
        }],
        imported_source={
            "manifest": manifest,
            "items": [{
                "kind": "imported",
                "import_id": manifest["import_id"],
                "source_row": 11,
                "source_currency": "USD",
                "quantity": "2",
                "overrides": {
                    "name": "Alien Task Chair identidad",
                    "description": "Silla operativa importada",
                    "dimension": "630 x 565 x 1000 mm",
                    "unit_price": "82.00",
                    "provider": "Sunon importado",
                },
            }],
            "source_currency": "USD",
        },
        today=date.today(),
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / f"imported-identity-{technical_column}.xlsx",
        image_dir=tmp_path / f"imported-identity-{technical_column}-images",
        imported_source_path=imported_source,
    )
    parser_book = load_workbook(parser_source, data_only=False)
    try:
        quotation = parser_book["Quotation"]
        product_row = next(
            row
            for row in range(8, quotation.max_row + 1)
            if isinstance(quotation.cell(row, 1).value, (int, float))
        )
        quotation.cell(product_row, technical_column).value = replacement
        parser_book.save(parser_source)
    finally:
        parser_book.close()
    output = tmp_path / f"imported-identity-{technical_column}-must-not-exist.xlsx"

    with pytest.raises(ValueError, match=message):
        generate_quote(
            parser_source,
            output,
            _mixed_import_metadata(payload, "MXN"),
            WORKER_TEMPLATE,
            original_quotation_path=imported_source,
            quotation_data_rows=quotation_data_rows(payload),
        )

    assert not output.exists()


def test_imported_only_cart_builds_workbook_and_generates_quote_without_rate_summary(
    monkeypatch,
    tmp_path,
):
    assert WORKER_TEMPLATE.is_file()
    imported_source = write_import_fixture(tmp_path / "imported-only-source.xlsx")
    source_hash_before = hashlib.sha256(imported_source.read_bytes()).hexdigest()
    manifest, imported_images = build_import_manifest(
        imported_source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=imported_source.name,
    )
    imported_key = f"import:{manifest['import_id']}:11"
    payload = build_mixed_catalog_cart_payload(
        [],
        catalogs={},
        rate_rows=[],
        quote_currency="USD",
        commercial_discount_percent="40",
        presentation_sections=[
            {
                "id": "section-1",
                "title": "Importados",
                "item_keys": [imported_key],
            }
        ],
        imported_source={
            "manifest": manifest,
            "items": [
                {
                    "kind": "imported",
                    "import_id": manifest["import_id"],
                    "source_row": 11,
                    "source_currency": "USD",
                    "quantity": "2",
                    "overrides": {
                        "name": "Alien Task Chair imported-only",
                        "description": "Silla operativa importada",
                        "dimension": "630 x 565 x 1000 mm",
                        "unit_price": "82.00",
                        "provider": "Sunon importado",
                    },
                }
            ],
            "source_currency": "USD",
        },
        today=date.today(),
    )
    assert payload["groups"] == []
    assert payload["rate_summary"] == []
    assert payload["item_count"] == 1
    canonical_rows = quotation_data_rows(payload)

    quotation_input = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "imported-only-quotation.xlsx",
        image_dir=tmp_path / "imported-only-images",
        imported_source_path=imported_source,
    )
    output = tmp_path / "imported-only-final.xlsx"

    generate_quote(
        quotation_input,
        output,
        _mixed_import_metadata(payload, "USD"),
        WORKER_TEMPLATE,
        original_quotation_path=imported_source,
        quotation_data_rows=canonical_rows,
    )

    assert hashlib.sha256(imported_source.read_bytes()).hexdigest() == source_hash_before
    wb = load_workbook(output, data_only=False)
    try:
        quotation = wb["Quotation"]
        audit = wb["Quotation_Data"]
        mobiliti = wb["Mobiliti"]
        cotizacion = wb["Cotizacion"]
        assert not _rows_with_value(
            quotation,
            2,
            "CAI63SW Alien Task Chair",
        )
        quotation_hashes = {
            hashlib.sha256(image._data()).hexdigest()
            for image in quotation._images
        }
        selected_imported_hash = hashlib.sha256(
            imported_images[11][0]
        ).hexdigest()
        unused_imported_hashes = {
            hashlib.sha256(image_bytes).hexdigest()
            for source_row, (image_bytes, _extension) in imported_images.items()
            if source_row != 11
        }
        assert selected_imported_hash in quotation_hashes
        assert quotation_hashes.isdisjoint(unused_imported_hashes)
        _assert_exact_quotation_data(audit, canonical_rows)
        assert audit["A2"].value == payload["imported_source"]["items"][0]["line_id"]
        assert audit["F2"].value == 11
        assert audit["O2"].value == manifest["items"][1]["row_hash"]

        mobiliti_row = _rows_with_value(
            mobiliti, 4, "Alien Task Chair imported-only"
        )[0]
        cotizacion_row = _rows_with_value(
            cotizacion, 1, f"=Mobiliti!D{mobiliti_row}"
        )[0]
        cost_formula = str(mobiliti.cell(mobiliti_row, 10).value)
        assert cost_formula.startswith("=Quotation!K")
        edited_row = int(cost_formula.removeprefix("=Quotation!K"))
        assert edited_row != 11
        assert quotation.cell(edited_row, 2).value == "Alien Task Chair imported-only"
        assert quotation.cell(edited_row, 11).value == 82
        assert mobiliti.cell(mobiliti_row, 16).value == "Centro"
        assert mobiliti.cell(mobiliti_row, 6).value == "Sunon Inc"
        assert cotizacion.cell(cotizacion_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
        assert cotizacion.cell(cotizacion_row, 7).value == 0.4
        assert cotizacion.cell(cotizacion_row, 8).value == (
            f"=F{cotizacion_row}*G{cotizacion_row}"
        )
        assert cotizacion.cell(cotizacion_row, 9).value == (
            f"=F{cotizacion_row}-H{cotizacion_row}"
        )
        assert cotizacion.cell(cotizacion_row, 10).value == (
            f"=E{cotizacion_row}*I{cotizacion_row}"
        )
    finally:
        wb.close()


def test_explicit_original_does_not_cross_wire_same_row_parser_image(
    monkeypatch,
    tmp_path,
):
    imported_source = write_import_fixture(tmp_path / "visible-original.xlsx")
    _manifest, imported_images = build_import_manifest(
        imported_source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=imported_source.name,
    )
    catalog_row = {
        "catalog": "offiho",
        "inventory_key": "OFF-E2E NEGRO",
        "quantity": "1",
    }
    payload = build_mixed_catalog_cart_payload(
        [catalog_row],
        catalogs=authoritative_catalogs(),
        rate_rows=rate_rows(),
        quote_currency="MXN",
        commercial_discount_percent="40",
        presentation_sections=[{
            "id": "section-1",
            "title": "Catalogo",
            "item_keys": [mixed_cart_key(catalog_row)],
        }],
        today=date.today(),
    )
    catalog_image = _make_png(tmp_path / "parser-offiho.png", (8, 180, 240))

    monkeypatch.setattr(
        catalog_cart,
        "_download_catalog_image",
        lambda *_args, **_kwargs: catalog_image,
    )
    parser_source = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / "parser-source.xlsx",
        image_dir=tmp_path / "parser-images",
    )
    output = tmp_path / "same-row-image-output.xlsx"

    generate_quote(
        parser_source,
        output,
        _mixed_import_metadata(payload, "MXN"),
        WORKER_TEMPLATE,
        original_quotation_path=imported_source,
        quotation_data_rows=quotation_data_rows(payload),
    )

    workbook = load_workbook(output, data_only=False)
    try:
        original = load_workbook(imported_source, data_only=False)
        try:
            assert workbook["Quotation"]["B9"].value == original["Quotation"]["B9"].value
        finally:
            original.close()
        cotizacion = workbook["Cotizacion"]
        product_row = next(
            row
            for row in range(16, cotizacion.max_row + 1)
            if cotizacion.cell(row, 1).value == "Silla Offiho"
        )
        product_images = [
            image
            for image in cotizacion._images
            if image.anchor._from.row + 1 == product_row
        ]
        assert len(product_images) == 1
        assert hashlib.sha256(product_images[0]._data()).hexdigest() == hashlib.sha256(
            catalog_image.read_bytes()
        ).hexdigest()
        assert hashlib.sha256(catalog_image.read_bytes()).hexdigest() != hashlib.sha256(
            imported_images[9][0]
        ).hexdigest()
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("quote_currency", "expected_import_price", "expected_import_rate"),
    (
        ("MXN", Decimal("1517.00"), Decimal("18.500000")),
        ("USD", Decimal("82.00"), Decimal("1.000000")),
        ("EUR", Decimal("75.44"), Decimal("0.920000")),
    ),
)
def test_imported_and_catalog_items_generate_one_quote_with_single_conversion(
    quote_currency,
    expected_import_price,
    expected_import_rate,
    monkeypatch,
    tmp_path,
):
    assert WORKER_TEMPLATE.is_file()
    imported_source = write_import_fixture(tmp_path / "imported-source.xlsx")
    source_hash_before = hashlib.sha256(imported_source.read_bytes()).hexdigest()
    manifest, imported_images = build_import_manifest(
        imported_source.read_bytes(),
        import_id="7b1d6d42-236a-4bc1-9aa8-8d9db793c30b",
        original_filename=imported_source.name,
    )
    catalog_rows = [
        {
            "catalog": "offiho",
            "inventory_key": "OFF-E2E NEGRO",
            "quantity": "1",
        },
        {"catalog": "cr-global", "internal_id": "cr:e2e", "quantity": "1"},
        {
            "catalog": "sonara",
            "internal_id": "sonara:e2e-review",
            "quantity": "1",
        },
    ]
    imported_key = f"import:{manifest['import_id']}:11"
    catalog_keys = [mixed_cart_key(row) for row in catalog_rows]
    sections = [
        {
            "id": "section-1",
            "title": "Recepcion",
            "item_keys": [catalog_keys[0], imported_key, catalog_keys[1]],
        },
        {
            "id": "section-2",
            "title": "Privados",
            "item_keys": [catalog_keys[2]],
        },
    ]
    today = date.today()
    rates = [
        {
            "currency": "USD",
            "effective_date": today.isoformat(),
            "mxn_per_unit": "18.500000",
            "retrieved_at": f"{today.isoformat()}T12:00:00+00:00",
        },
        {
            "currency": "EUR",
            "effective_date": today.isoformat(),
            "mxn_per_unit": "20.108696",
            "retrieved_at": f"{today.isoformat()}T12:00:00+00:00",
        },
    ]
    catalogs = authoritative_catalogs()
    payload = build_mixed_catalog_cart_payload(
        catalog_rows,
        catalogs=catalogs,
        rate_rows=rates,
        quote_currency=quote_currency,
        commercial_discount_percent="40",
        presentation_sections=sections,
        imported_source={
            "manifest": manifest,
            "items": [
                {
                    "kind": "imported",
                    "import_id": manifest["import_id"],
                    "source_row": 11,
                    "source_currency": "USD",
                    "quantity": "2",
                    "overrides": {
                        "name": "Alien Task Chair revisada",
                        "description": "Silla operativa revisada",
                        "dimension": "630 x 565 x 1000 mm",
                        "unit_price": "82.00",
                        "provider": "Sunon importado",
                    },
                }
            ],
            "source_currency": "USD",
        },
        today=today,
    )
    assert payload["sections"] == [
        {
            "id": "section-1",
            "title": "Recepcion",
            "line_ids": ["legacy-1", "legacy-import-1", "legacy-2"],
        },
        {
            "id": "section-2",
            "title": "Privados",
            "line_ids": ["legacy-3"],
        },
    ]
    assert payload["groups"][0]["items"][0]["name"] == (
        catalogs["offiho"]["items"][0].name
    )
    imported_line = payload["imported_source"]["items"][0]
    assert Decimal(imported_line["unit_price"]) == expected_import_price
    assert Decimal(imported_line["frozen_exchange_rate"]) == expected_import_rate

    catalog_image = _make_png(tmp_path / "catalog-offiho.png", (120, 70, 20))

    def local_catalog_image(url, image_dir, code, source_type, destination_key=None):
        assert source_type == "offiho_cart"
        assert url == "https://offiho.com.mx/e2e-offiho.png"
        return catalog_image

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", local_catalog_image)
    quotation_input = create_mixed_catalog_quotation_workbook(
        payload,
        tmp_path / f"mixed-import-{quote_currency}.xlsx",
        image_dir=tmp_path / f"images-{quote_currency}",
        imported_source_path=imported_source,
    )
    output = tmp_path / f"final-import-{quote_currency}.xlsx"
    canonical_rows = quotation_data_rows(payload)
    generate_quote(
        quotation_input,
        output,
        _mixed_import_metadata(payload, quote_currency),
        WORKER_TEMPLATE,
        original_quotation_path=imported_source,
        quotation_data_rows=canonical_rows,
    )

    assert hashlib.sha256(imported_source.read_bytes()).hexdigest() == source_hash_before
    template_workbook = load_workbook(WORKER_TEMPLATE, data_only=False)
    try:
        template_x_formula = template_workbook["Mobiliti"]["X14"].value
        template_freight_formula = template_workbook["Cotizacion"]["H21"].value
        template_before_tax_formula = template_workbook["Cotizacion"]["H22"].value
        template_tax_formula = template_workbook["Cotizacion"]["H23"].value
        template_total_formula = template_workbook["Cotizacion"]["H24"].value
    finally:
        template_workbook.close()
    assert isinstance(template_x_formula, str) and template_x_formula.startswith("=")
    assert isinstance(template_freight_formula, str)
    assert isinstance(template_before_tax_formula, str)
    assert isinstance(template_tax_formula, str)
    assert isinstance(template_total_formula, str)
    wb = load_workbook(output, data_only=False)
    try:
        quotation = wb["Quotation"]
        audit = wb["Quotation_Data"]
        mobiliti = wb["Mobiliti"]
        cotizacion = wb["Cotizacion"]
        edited_name_rows = _rows_with_value(
            quotation,
            2,
            "Alien Task Chair revisada",
        )
        assert len(edited_name_rows) == 1
        assert not _rows_with_value(
            quotation,
            2,
            "CAI63SW Alien Task Chair",
        )
        quotation_hashes = {
            hashlib.sha256(image._data()).hexdigest()
            for image in quotation._images
        }
        selected_imported_hash = hashlib.sha256(
            imported_images[11][0]
        ).hexdigest()
        unused_imported_hashes = {
            hashlib.sha256(image_bytes).hexdigest()
            for source_row, (image_bytes, _extension) in imported_images.items()
            if source_row != 11
        }
        assert selected_imported_hash in quotation_hashes
        assert quotation_hashes.isdisjoint(unused_imported_hashes)
        assert hashlib.sha256(catalog_image.read_bytes()).hexdigest() in quotation_hashes

        _assert_exact_quotation_data(audit, canonical_rows)
        imported_audit_row = canonical_rows.index(
            next(row for row in canonical_rows if row.origin == "imported")
        ) + 2
        assert audit.cell(imported_audit_row, 1).value == imported_line["line_id"]
        assert audit.cell(imported_audit_row, 6).value == 11
        assert audit.cell(imported_audit_row, 15).value == imported_line["row_hash"]

        items_by_key = {
            item["line_id"]: item
            for item in [
                *(item for group in payload["groups"] for item in group["items"]),
                *payload["imported_source"]["items"],
            ]
        }
        ordered_items = [
            items_by_key[key]
            for section in payload["sections"]
            for key in section["line_ids"]
        ]
        expected_names = [item["name"] for item in ordered_items]
        mobiliti_rows = [
            _rows_with_value(mobiliti, 4, name)[0]
            for name in expected_names
        ]
        cotizacion_rows = [
            _rows_with_value(cotizacion, 1, f"=Mobiliti!D{row}")[0]
            for row in mobiliti_rows
        ]
        assert mobiliti_rows == sorted(mobiliti_rows)
        assert cotizacion_rows == sorted(cotizacion_rows)
        first_product_row = cotizacion_rows[0]
        assert [cotizacion.cell(row, 7).value for row in cotizacion_rows] == [
            0.4,
            *(f"=$G${first_product_row}" for _ in cotizacion_rows[1:]),
        ]
        for canonical, mobiliti_row, cotizacion_row in zip(
            canonical_rows,
            mobiliti_rows,
            cotizacion_rows,
            strict=True,
        ):
            cost_formula = str(mobiliti.cell(mobiliti_row, 10).value)
            quotation_match = re.search(r"Quotation!K([1-9][0-9]*)", cost_formula)
            assert quotation_match is not None
            quotation_row = int(quotation_match.group(1))
            if canonical.origin == "imported":
                if cost_formula.startswith("=ROUND("):
                    assert Decimal(str(quotation.cell(quotation_row, 11).value)) == (
                        canonical.original_cost
                    )
                    expected_formula = (
                        f"=ROUND(Quotation!K{quotation_row}"
                        f"*{canonical.frozen_rate:f},2)"
                    )
                else:
                    assert Decimal(str(quotation.cell(quotation_row, 11).value)) == (
                        canonical.converted_cost
                    )
                    expected_formula = f"=Quotation!K{quotation_row}"
                assert cost_formula == expected_formula
            else:
                assert Decimal(str(quotation.cell(quotation_row, 11).value)) == (
                    canonical.converted_cost
                )
                assert cost_formula == f"=Quotation!K{quotation_row}"
            assert mobiliti.cell(mobiliti_row, 16).value == canonical.region
            assert str(mobiliti.cell(mobiliti_row, 23).value).startswith("=IF(")
            assert mobiliti.cell(mobiliti_row, 24).value == (
                Translator(
                    template_x_formula,
                    origin="X14",
                ).translate_formula(f"X{mobiliti_row}")
            )
            assert cotizacion.cell(cotizacion_row, 6).value == f"=Mobiliti!X{mobiliti_row}"
            assert cotizacion.cell(cotizacion_row, 8).value == (
                f"=F{cotizacion_row}*G{cotizacion_row}"
            )
            assert cotizacion.cell(cotizacion_row, 9).value == (
                f"=F{cotizacion_row}-H{cotizacion_row}"
            )
            assert cotizacion.cell(cotizacion_row, 10).value == (
                f"=E{cotizacion_row}*I{cotizacion_row}"
            )

        assert _rows_with_value(cotizacion, 1, "Recepcion")
        assert _rows_with_value(cotizacion, 1, "Privados")
        mobiliti_values = {
            cell.value
            for row in mobiliti.iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert {"Recepcion", "Privados"} <= mobiliti_values
        assert not any(
            str(cotizacion.cell(row, 1).value or "").startswith("=Quotation!")
            for row in range(16, cotizacion.max_row + 1)
        )

        total_rows = [
            row
            for row in range(1, cotizacion.max_row + 1)
            if cotizacion.cell(row, 4).value
            in {"SUBTOTAL:", "IVA:", "TOTAL:"}
        ]
        assert len(total_rows) == 4
        subtotal, before_tax, tax, total = total_rows
        subtotal_formula = cotizacion.cell(subtotal, 8).value
        assert getattr(subtotal_formula, "text", None) == (
            f"=SUM(IFERROR(J{first_product_row}:J{cotizacion_rows[-1]},0))"
        )
        assert cotizacion.cell(subtotal + 1, 8).value == Translator(
            template_freight_formula,
            origin="H21",
        ).translate_formula(f"H{subtotal + 1}")
        assert cotizacion.cell(before_tax, 8).value == Translator(
            template_before_tax_formula,
            origin="H22",
        ).translate_formula(f"H{before_tax}")
        assert cotizacion.cell(tax, 8).value == Translator(
            template_tax_formula,
            origin="H23",
        ).translate_formula(f"H{tax}")
        assert cotizacion.cell(total, 8).value == Translator(
            template_total_formula,
            origin="H24",
        ).translate_formula(f"H{total}")

        cotizacion_hashes = {
            hashlib.sha256(image._data()).hexdigest()
            for image in cotizacion._images
        }
        assert hashlib.sha256(catalog_image.read_bytes()).hexdigest() in (
            cotizacion_hashes
        )
        assert hashlib.sha256(imported_images[11][0]).hexdigest() in (
            cotizacion_hashes
        )
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("quote_currency", "automatic_rate"),
    (("MXN", "1.000000"), ("USD", "0.054054"), ("EUR", "0.048780")),
)
def test_mixed_api_worker_produces_one_auditable_workbook(
    quote_currency,
    automatic_rate,
    isolated_quote_runtime,
    monkeypatch,
    tmp_path,
):
    assert WORKER_TEMPLATE.is_file()
    api_index, quote_worker = isolated_quote_runtime
    api_state = install_api_boundary(monkeypatch, api_index, authoritative_catalogs())

    local_images = {
        ("tarkett_cart", "https://media.tarkett-image.com/e2e-tarkett.png"):
            _make_png(tmp_path / "tarkett.png", (20, 70, 120)),
        ("offiho_cart", "https://offiho.com.mx/e2e-offiho.png"):
            _make_png(tmp_path / "offiho.png", (120, 70, 20)),
        ("supplier_cart", "https://alma.example.test/e2e-alma.png"):
            _make_png(tmp_path / "alma.png", (40, 120, 70)),
    }
    image_calls = []

    def local_catalog_image(url, image_dir, code, source_type, destination_key=None):
        image_calls.append((source_type, url, destination_key))
        return local_images[(source_type, url)]

    monkeypatch.setattr(catalog_cart, "_download_catalog_image", local_catalog_image)

    body = {
        "items": browser_rows_for_all_catalogs_and_two_alma_configs(),
        "quote_currency": quote_currency,
        "descuento": "40",
        "proyecto": "Proyecto mixto",
        "cliente": "Cliente prueba",
        "correo": "cliente@example.test",
        "telefono": "3330000000",
        "direccion": "Guadalajara",
        "razon_social": "Cliente SA de CV",
        "image_provider": "pillow",
        "template": WORKER_TEMPLATE.name,
    }
    with TestClient(api_index.app) as api_client:
        response = api_client.post(
            "/catalogs/mixed-quote",
            headers=auth_headers(api_index),
            json=body,
        )
    assert response.status_code == 200, response.json()
    queued_job = response.json()["job"]
    assert api_state["events"] == [
        "create_job", "reserve_mixed", "upload", "queue", "wake"
    ]
    assert len(api_state["jobs"]) == 1
    assert api_state["released"] == []
    assert api_state["deleted_jobs"] == []
    assert api_state["deleted_inputs"] == []
    assert len(api_state["uploads"]) == 1
    input_path, input_bytes, content_type = api_state["uploads"][0]
    assert input_path == queued_job["input_path"]
    assert content_type == "application/json"
    payload = json.loads(input_bytes)
    frozen_payload = deepcopy(payload)
    assert payload["source_type"] == "mixed_catalog_cart"
    assert payload["item_count"] == 8
    assert [group["catalog"] for group in payload["groups"]] == list(CATALOGS)
    assert {
        group["catalog"]: group["catalog_source_hash"] for group in payload["groups"]
    } == SOURCE_HASHES
    alma_lines = next(group for group in payload["groups"] if group["catalog"] == "alma")["items"]
    assert len({line["canonical_key"] for line in alma_lines}) == 2
    assert {line["configuration"] for line in alma_lines} == {
        "Base operativa; Electrificacion A",
        "Base operativa; Pasacables B",
    }
    assert payload["auto_electrification_rate"]["exchange_rate"] == automatic_rate
    for line in (item for group in payload["groups"] for item in group["items"]):
        assert line["reservation"] is not None
        assert line["reserved_quantity"] == "0.000000"
        assert line["available_after_reservations"] == "20.000000"
        assert line["reserved_by_others"] is False

    worker_client = EndToEndWorkerClient(
        queued_job,
        {input_path: input_bytes},
    )
    real_run_generator = quote_worker._run_generator
    generator_calls = []
    expected_handoff_rows = quotation_data_rows(payload)

    def counted_run_generator(job, generator_input, local_output):
        assert isinstance(generator_input, quote_worker.PreparedGeneratorInput)
        assert generator_input.parser_source.name == (
            "quotation_from_mixed_catalog.xlsx"
        )
        assert generator_input.parser_source.is_file()
        assert generator_input.original_quotation is None
        assert generator_input.quotation_data == expected_handoff_rows
        converted = load_workbook(
            generator_input.parser_source,
            data_only=False,
            read_only=True,
        )
        try:
            assert converted.sheetnames == ["Quotation"]
        finally:
            converted.close()
        worker_client.record_event("converter")
        worker_client.record_event("generator")
        generator_calls.append(generator_input.parser_source.name)
        result = real_run_generator(job, generator_input, local_output)
        assert generator_input.quotation_data == expected_handoff_rows
        return result

    monkeypatch.setattr(quote_worker, "_run_generator", counted_run_generator)
    monkeypatch.setattr(quote_worker, "_template_path", lambda: str(WORKER_TEMPLATE))
    monkeypatch.setattr(quote_worker, "QUOTE_ENGINE", "python")

    completed = quote_worker.process_job(worker_client, queued_job)
    assert completed
    assert worker_client.events == [
        "claim", "progress:45", "download", "progress:55", "converter",
        "generator", "progress:90", "upload", "completed", "delete", "clear_input",
    ]
    assert generator_calls == ["quotation_from_mixed_catalog.xlsx"]
    assert worker_client.downloads == [input_path]
    assert len(worker_client.uploads) == 1
    uploaded_path, output_bytes = worker_client.uploads[0]
    assert uploaded_path.startswith(f"users/7/jobs/{queued_job['id']}/attempts/")
    assert uploaded_path.endswith("/output.xlsx")
    assert output_bytes.startswith(b"PK")
    assert worker_client.deleted_inputs == [input_path]
    assert input_path not in worker_client.objects
    assert len(worker_client.completed_updates) == 1
    assert worker_client.failed_updates == []
    completed_job = worker_client.completed_updates[0]
    completed_metadata = completed_job["metadata"]
    assert completed_metadata["mixed_catalog_converted"] is True
    assert completed_metadata["catalog_price_mode"] == "mixed_catalog_converted"
    assert completed_metadata["base_currency"] == quote_currency
    assert completed_metadata["quote_currency"] == quote_currency
    assert completed_metadata["exchange_rate"] == "1.000000"
    assert completed_metadata["rate_summary"] == frozen_payload["rate_summary"]
    assert completed_metadata["auto_electrification_rate"] == frozen_payload[
        "auto_electrification_rate"
    ]
    assert completed_metadata["catalog_source_hashes"] == SOURCE_HASHES
    assert completed_metadata["descuento"] == 40
    assert completed_job["input_path"] is None
    assert completed_job["output_path"] == uploaded_path
    assert payload == frozen_payload

    assert [(source, url) for source, url, _key in image_calls] == list(
        local_images
    )
    image_keys = [key for _source, _url, key in image_calls]
    assert all(image_keys)
    assert len(set(image_keys)) == 3

    output_xlsx = tmp_path / f"cotizacion_mixta_final_{quote_currency}.xlsx"
    output_xlsx.write_bytes(output_bytes)
    assert output_xlsx.is_file()
    wb = load_workbook(output_xlsx, data_only=False)
    try:
        _assert_task9_final_workbook(
            wb,
            payload,
            quote_currency=quote_currency,
            automatic_rate=automatic_rate,
            expected_image_hashes={
                hashlib.sha256(path.read_bytes()).hexdigest()
                for path in local_images.values()
            },
        )
    finally:
        wb.close()


def _make_png(path: Path, color: tuple[int, int, int]) -> Path:
    Image.new("RGB", (96, 72), color).save(path, format="PNG")
    return path


def _assert_task9_final_workbook(
    wb,
    payload: dict,
    *,
    quote_currency: str,
    automatic_rate: str,
    expected_image_hashes: set[str],
) -> None:
    assert wb.sheetnames.count("Cotizacion") == 1
    assert wb.sheetnames.count("Mobiliti") == 1
    assert wb.sheetnames.count("Quotation") == 1
    assert wb.sheetnames.count("Quotation_Data") == 1
    quotation = wb["Quotation"]
    mobiliti = wb["Mobiliti"]
    cotizacion = wb["Cotizacion"]
    audit = wb["Quotation_Data"]
    assert audit.sheet_state == "veryHidden"
    assert tuple(audit.cell(1, column).value for column in range(1, 17)) == (
        QUOTATION_DATA_HEADERS
    )

    base_rows = quotation_data_rows(payload)
    base_keys = [row.item_key for row in base_rows]
    assert base_keys == [
        item_key
        for section in payload["sections"]
        for item_key in section["line_ids"]
    ]
    records = [
        {
            field: audit.cell(row_number, column).value
            for column, field in enumerate(QUOTATION_DATA_HEADERS, start=1)
        }
        for row_number in range(2, audit.max_row + 1)
    ]
    assert [record["position"] for record in records] == list(
        range(1, len(records) + 1)
    )
    assert len(records) == len(base_rows) + 3
    assert [
        record["item_key"]
        for record in records
        if record["item_key"] in set(base_keys)
    ] == base_keys

    records_by_key = {record["item_key"]: record for record in records}
    assert len(records_by_key) == len(records)
    decimal_fields = (
        "original_cost",
        "frozen_rate",
        "converted_cost",
        "quantity",
    )
    preserved_fields = (
        "item_key",
        "section_id",
        "section_title",
        "origin",
        "source_row",
        "original_currency",
        "provider",
        "region",
        "source_hash",
        "upstream_row_hash",
    )
    for expected in base_rows:
        actual = records_by_key[expected.item_key]
        assert all(actual[field] == getattr(expected, field) for field in preserved_fields)
        assert all(
            Decimal(str(actual[field])) == getattr(expected, field)
            for field in decimal_fields
        )
        assert len(actual["row_hash"]) == 64

    parent_key = base_keys[0]
    parent_record = records_by_key[parent_key]
    derived = [record for record in records if record["item_key"] not in set(base_keys)]
    assert [record["position"] for record in derived] == [2, 3, 4]
    assert [record["item_key"].rsplit(":", 1)[-1] for record in derived] == [
        "LIDO.OP-INT",
        "JUMP-1.5M",
        "CAJA-FUS",
    ]
    assert all(
        record["item_key"].startswith(f"{parent_key}:lumbro:")
        and record["origin"] == "lumbro"
        and record["source_row"] is None
        and record["section_id"] == parent_record["section_id"]
        and record["section_title"] == parent_record["section_title"]
        and record["upstream_row_hash"] == ""
        and len(record["source_hash"]) == 64
        and len(record["row_hash"]) == 64
        for record in derived
    )
    assert len({record["source_hash"] for record in derived}) == 1
    assert [Decimal(str(record["frozen_rate"])) for record in derived] == [
        Decimal(automatic_rate),
    ] * 3

    template_values = load_workbook(WORKER_TEMPLATE, data_only=True, read_only=True)
    try:
        guide = template_values["SPEC-GUIDE-LUMBRO"]
        expected_original_costs = [
            Decimal(str(guide[f"E{row}"].value)).quantize(
                Decimal("0.000001"),
                rounding=ROUND_HALF_UP,
            )
            for row in (380, 396, 406)
        ]
    finally:
        template_values.close()
    assert [Decimal(str(record["original_cost"])) for record in derived] == (
        expected_original_costs
    )
    assert [Decimal(str(record["converted_cost"])) for record in derived] == [
        (cost * Decimal(automatic_rate)).quantize(
            MONEY,
            rounding=ROUND_HALF_UP,
        )
        for cost in expected_original_costs
    ]
    assert [Decimal(str(record["quantity"])) for record in derived] == [
        Decimal("8"),
        Decimal("8"),
        Decimal("2"),
    ]

    payload_items = {
        item["line_id"]: item
        for group in payload["groups"]
        for item in group["items"]
    }
    base_names = [payload_items[key]["name"] for key in base_keys]
    derived_names = ["LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"]
    final_names = [base_names[0], *derived_names, *base_names[1:]]
    mobiliti_rows = _ordered_rows_for_values(mobiliti, 4, final_names)
    cotizacion_rows = _ordered_rows_for_values(
        cotizacion,
        1,
        [f"=Mobiliti!D{row}" for row in mobiliti_rows],
    )
    assert mobiliti_rows == list(range(14, 25))
    assert cotizacion_rows == list(range(17, 28))

    processed_by_cotizacion_row = {}
    for record, mobiliti_row, cotizacion_row in zip(
        records,
        mobiliti_rows,
        cotizacion_rows,
        strict=True,
    ):
        cost_formula = str(mobiliti.cell(mobiliti_row, 10).value)
        match = re.fullmatch(r"=Quotation!K([1-9][0-9]*)", cost_formula)
        assert match is not None
        assert Decimal(str(quotation.cell(int(match.group(1)), 11).value)) == (
            Decimal(str(record["converted_cost"]))
        )
        assert mobiliti.cell(mobiliti_row, 16).value == record["region"]
        assert str(mobiliti.cell(mobiliti_row, 23).value).startswith("=IF(")
        assert mobiliti.cell(mobiliti_row, 24).value == (
            f"=_xlfn.MINIFS($W$14:$W$571,$D$14:$D$571,D{mobiliti_row})"
        )
        quantity_match = re.fullmatch(
            r"=Quotation!H([1-9][0-9]*)",
            str(mobiliti.cell(mobiliti_row, 8).value),
        )
        assert quantity_match is not None
        quotation_row = int(quantity_match.group(1))
        assert Decimal(str(quotation.cell(quotation_row, 8).value)) == Decimal(
            str(record["quantity"])
        )
        assert mobiliti.cell(mobiliti_row, 11).value == (
            f"=Quotation!I{quotation_row}"
        )
        assert cotizacion.cell(cotizacion_row, 1).value == (
            f"=Mobiliti!D{mobiliti_row}"
        )
        assert cotizacion.cell(cotizacion_row, 3).value == (
            f"=Quotation!D{quotation_row}"
        )
        assert cotizacion.cell(cotizacion_row, 4).value == (
            f"=Quotation!F{quotation_row}"
        )
        assert cotizacion.cell(cotizacion_row, 5).value == (
            f"=Mobiliti!H{mobiliti_row}"
        )
        processed_by_cotizacion_row[cotizacion_row] = quotation.cell(
            quotation_row,
            4,
        ).value
        assert processed_by_cotizacion_row[cotizacion_row]
        assert cotizacion.cell(cotizacion_row, 6).value == (
            f"=Mobiliti!X{mobiliti_row}"
        )
        assert cotizacion.cell(cotizacion_row, 8).value == (
            f"=F{cotizacion_row}*G{cotizacion_row}"
        )
        assert cotizacion.cell(cotizacion_row, 9).value == (
            f"=F{cotizacion_row}-H{cotizacion_row}"
        )
        assert cotizacion.cell(cotizacion_row, 10).value == (
            f"=E{cotizacion_row}*I{cotizacion_row}"
        )
    first_product = cotizacion_rows[0]
    assert [cotizacion.cell(row, 7).value for row in cotizacion_rows] == [
        0.4,
        *(f"=$G${first_product}" for _ in cotizacion_rows[1:]),
    ]
    assert not any(
        str(cotizacion.cell(row, 1).value or "").startswith("=Quotation!")
        for row in range(16, cotizacion.max_row + 1)
    )

    sonara_row = cotizacion_rows[final_names.index("Panel Sonara por verificar")]
    assert "Revision documental local" in str(
        processed_by_cotizacion_row[sonara_row]
    )
    alma_rows = [
        cotizacion_rows[index]
        for index, name in enumerate(final_names)
        if name == "Mesa ALMA"
    ]
    assert "electrificacion a" in str(
        processed_by_cotizacion_row[alma_rows[0]]
    ).casefold()
    assert "pasacables b" in str(
        processed_by_cotizacion_row[alma_rows[1]]
    ).casefold()

    subtotal_rows = _rows_with_value(cotizacion, 4, "SUBTOTAL:")
    tax_row = _rows_with_value(cotizacion, 4, "IVA:")[0]
    total_row = _rows_with_value(cotizacion, 4, "TOTAL:")[0]
    assert len(subtotal_rows) == 2
    subtotal, before_tax = subtotal_rows
    subtotal_formula = cotizacion.cell(subtotal, 8).value
    assert getattr(subtotal_formula, "text", None) == (
        f"=SUM(IFERROR(J{first_product}:J{cotizacion_rows[-1]},0))"
    )
    assert cotizacion.cell(subtotal + 1, 8).value == f"=H{subtotal}*8%"
    assert cotizacion.cell(before_tax, 8).value == (
        f"=H{subtotal}+H{subtotal + 1}"
    )
    assert cotizacion.cell(tax_row, 8).value == f"=H{before_tax}*16%"
    assert cotizacion.cell(total_row, 8).value == f"=H{before_tax}+H{tax_row}"
    assert mobiliti["K4"].value is (quote_currency != "MXN")

    cotizacion_hashes = {
        hashlib.sha256(image._data()).hexdigest()
        for image in cotizacion._images
    }
    assert expected_image_hashes <= cotizacion_hashes
    quotation_hashes = {
        hashlib.sha256(image._data()).hexdigest()
        for image in quotation._images
    }
    assert expected_image_hashes <= quotation_hashes


def _assert_final_workbook(
    wb,
    payload: dict,
    *,
    quote_currency: str,
    automatic_rate: str,
) -> None:
    assert wb.sheetnames.count("Cotizacion") == 1
    assert wb.sheetnames.count("Mobiliti") == 1
    assert wb.sheetnames.count("Quotation") == 1
    quotation = wb["Quotation"]
    mobiliti = wb["Mobiliti"]
    cot = wb["Cotizacion"]
    # Las solicitudes heredadas sin `sections` conservan el orden del carrito en
    # una sola sección de presentación. El proveedor sigue auditado por línea,
    # pero ya no controla las bandas visuales del Excel.
    assert [
        quotation.cell(row, 1).value
        for row in range(8, quotation.max_row + 1)
        if isinstance(quotation.cell(row, 1).value, str)
    ] == ["- Recepción"]
    source_rows = [
        row
        for row in range(8, quotation.max_row + 1)
        if isinstance(quotation.cell(row, 1).value, (int, float))
    ]
    assert [quotation.cell(row, 1).value for row in source_rows] == list(range(1, 9))
    assert [quotation.cell(row, 13).value for row in source_rows] == [
        "Tarkett", "Offiho", "CR Global", "Sonara", "Sunon", "ALMA",
        "ALMA", "Lumbro",
    ]
    assert [quotation.cell(row, 14).value for row in source_rows] == [
        40, 40, 0, 0, 0, 0, 0, 0,
    ]
    assert [quotation.cell(row, 20).value for row in source_rows] == [
        True, True, False, False, False, False, False, False,
    ]
    sonara_description = str(quotation.cell(source_rows[3], 4).value)
    assert "Revision documental local" in sonara_description
    assert "Codigo por verificar" in sonara_description
    alma_a_description = str(quotation.cell(source_rows[5], 4).value)
    alma_b_description = str(quotation.cell(source_rows[6], 4).value)
    assert "electrificacion a" in alma_a_description.casefold()
    assert "pasacables b" in alma_b_description.casefold()
    assert alma_a_description != alma_b_description
    assert "LIDO.OP-INT manual" in str(quotation.cell(source_rows[7], 2).value)
    assert len(quotation._images) >= 3
    assert len(cot._images) >= 3

    row_maps = []
    for source_row in source_rows:
        mobiliti_row = _row_for_formula(mobiliti, 4, f"=Quotation!B{source_row}")
        cot_row = _row_for_formula(cot, 1, f"=Mobiliti!D{mobiliti_row}")
        row_maps.append((source_row, mobiliti_row, cot_row))
    assert [
        (row, cot.cell(row, 1).value)
        for row in range(1, cot.max_row + 1)
        if str(cot.cell(row, 1).value or "").startswith("=Mobiliti!D")
    ] == [
        (cot_row, f"=Mobiliti!D{mobiliti_row}")
        for _source_row, mobiliti_row, cot_row in row_maps
    ]
    assert [mobiliti.cell(row, 6).value for _source, row, _cot in row_maps] == [
        "Tarkett", "Offiho", "CR Global", "Sonara", "Sunon", "ALMA",
        "ALMA", "Lumbro CH",
    ]
    first_discount_row = row_maps[0][2]
    assert [cot.cell(row, 7).value for _source, _mob, row in row_maps] == [
        0.4,
        *(f"=G${first_discount_row}" for _ in row_maps[1:]),
    ]
    assert mobiliti["J6"].value == f"{quote_currency}/{quote_currency}"
    assert mobiliti["K6"].value == 1
    frozen_lines = [
        item for group in payload["groups"] for item in group["items"]
    ]
    assert [quotation.cell(row, 11).value for row in source_rows] == [
        float(Decimal(item["unit_price"])) for item in frozen_lines
    ]
    for index, (_source_row, mobiliti_row, cot_row) in enumerate(row_maps):
        assert mobiliti.cell(mobiliti_row, 10).value == f"=Quotation!K{_source_row}"
        assert "$K$6" not in str(mobiliti.cell(mobiliti_row, 10).value)
        assert "$K$6" not in str(cot.cell(cot_row, 6).value)
        if index:
            assert cot.cell(cot_row, 6).value == f"=ROUND(Mobiliti!X{mobiliti_row},2)"
        assert cot.cell(cot_row, 8).value == f"=ROUND(F{cot_row}*G{cot_row},2)"
        assert cot.cell(cot_row, 9).value == f"=ROUND(F{cot_row}-H{cot_row},2)"
        assert cot.cell(cot_row, 10).value == f"=ROUND(E{cot_row}*I{cot_row},2)"

    automatic_rows = [
        row
        for row in range(1, mobiliti.max_row + 1)
        if mobiliti.cell(row, 4).value in {"LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"}
    ]
    assert len(automatic_rows) == 3
    automatic_by_code = {
        mobiliti.cell(row, 4).value: row for row in automatic_rows
    }
    assert list(automatic_by_code) == ["LIDO.OP-INT", "JUMP-1.5M", "CAJA-FUS"]
    first_unused_row = max(
        *(mobiliti_row for _source, mobiliti_row, _cot in row_maps),
        *automatic_rows,
    ) + 1
    assert first_unused_row == 25
    guard = (
        f'=IF(COUNTA($D{first_unused_row},$E{first_unused_row},$F{first_unused_row},'
        f'$H{first_unused_row},$J{first_unused_row},$K{first_unused_row})=0,"",'
    )
    assert all(
        mobiliti.cell(first_unused_row, column).value is None
        for column in (4, 5, 6, 8, 10, 11, 16)
    )
    assert str(mobiliti.cell(first_unused_row, 23).value).startswith(guard)
    assert str(mobiliti.cell(first_unused_row, 24).value).startswith(guard)
    assert mobiliti.cell(first_unused_row, 35).value == (
        f'{guard}IF(AH{first_unused_row}<30%,"ERROR","OK"))'
    )
    expected_template_rows = {
        "LIDO.OP-INT": 380,
        "JUMP-1.5M": 396,
        "CAJA-FUS": 406,
    }
    assert {
        code: mobiliti.cell(row, 8).value
        for code, row in automatic_by_code.items()
    } == {"LIDO.OP-INT": 8, "JUMP-1.5M": 8, "CAJA-FUS": 2}
    expected_rate_literal = str(float(automatic_rate)).rstrip("0").rstrip(".") or "0"
    for code, template_row in expected_template_rows.items():
        row = automatic_by_code[code]
        assert mobiliti.cell(row, 10).value == (
            f"=ROUND('SPEC-GUIDE-LUMBRO'!E{template_row}*"
            f"{expected_rate_literal},2)"
        )
    parent_cot_row = row_maps[0][2]
    parent_mobiliti_row = row_maps[0][1]
    parent_formula = str(cot.cell(parent_cot_row, 6).value)
    expected_terms = [
        f"Mobiliti!X{parent_mobiliti_row}*Mobiliti!H{parent_mobiliti_row}",
        *(
            f"Mobiliti!X{automatic_by_code[code]}*"
            f"Mobiliti!H{automatic_by_code[code]}"
            for code in expected_template_rows
        ),
    ]
    assert parent_formula == (
        f"=ROUND(IFERROR(({'+'.join(expected_terms)})/"
        f"Mobiliti!H{parent_mobiliti_row},0),2)"
    )
    assert "Mobiliti!Y" not in parent_formula
    assert "$K$6" not in parent_formula
    for _source, _mob, cot_row in row_maps[1:]:
        formula = str(cot.cell(cot_row, 6).value)
        assert all(
            f"Mobiliti!X{row}*Mobiliti!H{row}" not in formula
            for row in automatic_rows
        )

    total_rows = [
        row
        for row in range(1, cot.max_row + 1)
        if cot.cell(row, 4).value
        in {"SUBTOTAL:", "COSTO DE FLETE:", "IVA:", "TOTAL:"}
    ]
    assert len(total_rows) == 5
    assert [cot.cell(row, 4).value for row in total_rows] == [
        "SUBTOTAL:", "COSTO DE FLETE:", "SUBTOTAL:", "IVA:", "TOTAL:",
    ]
    first_product = row_maps[0][2]
    last_product = row_maps[-1][2]
    subtotal, freight, before_tax, tax, total = total_rows
    assert cot.cell(subtotal, 8).value == f"=ROUND(SUM(J{first_product}:J{last_product}),2)"
    assert cot.cell(freight, 8).value == f"=ROUND(H{subtotal}*12%,2)"
    assert cot.cell(before_tax, 8).value == f"=ROUND(H{subtotal}+H{freight},2)"
    assert cot.cell(tax, 8).value == f"=ROUND(H{before_tax}*16%,2)"
    assert cot.cell(total, 8).value == f"=ROUND(H{before_tax}+H{tax},2)"

    template_values = load_workbook(WORKER_TEMPLATE, data_only=True, read_only=True)
    try:
        lumbro = template_values["SPEC-GUIDE-LUMBRO"]
        unit_prices = {
            code: Decimal(str(lumbro[f"E{template_row}"].value))
            for code, template_row in expected_template_rows.items()
        }
    finally:
        template_values.close()
    tarkett_key = payload["groups"][0]["items"][0]["canonical_key"]
    accessories = {
        tarkett_key: [
            (
                unit_prices[code],
                Decimal(str(mobiliti.cell(automatic_by_code[code], 8).value)),
            )
            for code in expected_template_rows
        ]
    }
    totals = expected_mixed_totals(payload, accessories)
    assert totals == {
        "MXN": (
            Decimal("17610.12"), Decimal("2113.21"), Decimal("19723.33"),
            Decimal("3155.73"), Decimal("22879.06"),
        ),
        "USD": (
            Decimal("951.90"), Decimal("114.23"), Decimal("1066.13"),
            Decimal("170.58"), Decimal("1236.71"),
        ),
        "EUR": (
            Decimal("859.00"), Decimal("103.08"), Decimal("962.08"),
            Decimal("153.93"), Decimal("1116.01"),
        ),
    }[quote_currency]
    assert totals[2] == totals[0] + totals[1]
    assert totals[4] == totals[2] + totals[3]
