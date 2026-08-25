from pathlib import Path
import posixpath
from decimal import Decimal
import re
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import pytest
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.formula.tokenizer import Tokenizer

from mobiliti_saas.quote_engine.mobiliti_layout import SectionNeed, plan_mobiliti_layout
from mobiliti_saas.quote_engine.ooxml_formula import translate_formula
from mobiliti_saas.quote_engine.ooxml_package import (
    PackageMutation,
    XlsxPackage,
    assert_package_preserved,
)
from mobiliti_saas.quote_engine.ooxml_worksheet import (
    MobilitiCellWrite,
    WorksheetEditor,
    apply_mobiliti_cell_writes,
    build_mobiliti_sheet,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 GDL.xlsx"
)
OFFICIAL_TEMPLATE = (
    ROOT
    / "mobiliti_saas"
    / "worker"
    / "templates"
    / "Formato Cotizacion 2026 Oficial.xlsx"
)
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
XM = "http://schemas.microsoft.com/office/excel/2006/main"
TEST_CELL_REFERENCE = re.compile(
    r"(?P<column>\$?[A-Z]{1,3})(?P<row_abs>\$?)(?P<row>[1-9][0-9]*)$"
)
TEST_RANGE_REFERENCE = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[^'!]+)!)?"
    r"(?P<first>\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
    r"(?::(?P<last>\$?[A-Z]{1,3}\$?[1-9][0-9]*))?$"
)

pytestmark = [
    pytest.mark.filterwarnings("ignore:Data Validation extension is not supported"),
    pytest.mark.filterwarnings("ignore:wmf image format is not supported"),
]


def _official_part(path: Path, sheet_name: str) -> str:
    with ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships.findall(f"{{{PACKAGE_REL}}}Relationship")
    }
    sheet = next(
        node
        for node in workbook.findall(f".//{{{MAIN}}}sheet")
        if node.attrib["name"] == sheet_name
    )
    target = targets[sheet.attrib[f"{{{OFFICE_REL}}}id"]]
    return posixpath.normpath(posixpath.join("xl", target.lstrip("/")))


def _part_bytes(path: Path, part: str) -> bytes:
    with ZipFile(path) as archive:
        return archive.read(part)


def _official_root() -> ET.Element:
    part = _official_part(TEMPLATE, "Mobiliti")
    return ET.fromstring(_part_bytes(TEMPLATE, part))


def _render(tmp_path: Path, count: int):
    part = _official_part(TEMPLATE, "Mobiliti")
    writes = [
        MobilitiCellWrite(f"D{14 + index}", "text", f"SKU-{index:03d}")
        for index in range(count)
    ]
    mutation = build_mobiliti_sheet(
        _part_bytes(TEMPLATE, part),
        [SectionNeed("chairs", "SILLAS", count)],
        writes,
    )
    output = tmp_path / f"mobiliti-{count}.xlsx"
    XlsxPackage.read(TEMPLATE).write_new(
        output,
        PackageMutation(replacements={part: mutation.xml}),
    )
    assert assert_package_preserved(TEMPLATE, output, {part}).changed_parts == {part}
    workbook = load_workbook(output, data_only=False, keep_links=True)
    return mutation.row_map, workbook, workbook["Mobiliti"]


def _render_sections(tmp_path: Path, section_count: int):
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(section_count)
    ]
    planned = plan_mobiliti_layout(needs)
    writes = [
        MobilitiCellWrite(f"D{section.product_start}", "text", f"SKU-{index:03d}")
        for index, section in enumerate(planned.sections[:section_count])
    ]
    part = _official_part(TEMPLATE, "Mobiliti")
    mutation = build_mobiliti_sheet(_part_bytes(TEMPLATE, part), needs, writes)
    output = tmp_path / f"mobiliti-sections-{section_count}.xlsx"
    XlsxPackage.read(TEMPLATE).write_new(
        output,
        PackageMutation(replacements={part: mutation.xml}),
    )
    workbook = load_workbook(output, data_only=False, keep_links=True)
    return mutation, workbook, workbook["Mobiliti"]


def _official_formula(coordinate: str) -> str:
    part = _official_part(TEMPLATE, "Mobiliti")
    root = ET.fromstring(_part_bytes(TEMPLATE, part))
    formula = root.find(f".//{{{MAIN}}}c[@r='{coordinate}']/{{{MAIN}}}f")
    assert formula is not None and formula.text
    return "=" + formula.text


def _formula_rows(formula: str, column: str) -> list[int]:
    return [int(row) for row in __import__("re").findall(fr"\b{column}(\d+)\b", formula)]


def _reference_with_test_rows(reference: str, rows: tuple[int, ...]) -> str:
    match = TEST_RANGE_REFERENCE.fullmatch(reference)
    assert match is not None
    endpoints = [match.group("first")]
    if match.group("last"):
        endpoints.append(match.group("last"))
    rewritten = []
    for endpoint, row in zip(endpoints, rows, strict=True):
        cell = TEST_CELL_REFERENCE.fullmatch(endpoint)
        assert cell is not None
        rewritten.append(f"{cell.group('column')}{cell.group('row_abs')}{row}")
    prefix = f"{match.group('sheet')}!" if match.group("sheet") else ""
    return prefix + ":".join(rewritten)


def _is_test_local_mobiliti_sheet(sheet: str | None) -> bool:
    if sheet is None:
        return True
    normalized = sheet[1:-1].replace("''", "'") if sheet.startswith("'") else sheet
    return normalized.casefold() == "mobiliti"


def _derive_official_formula(
    coordinate: str,
    target: str,
    row_map,
    *,
    range_overrides: dict[str, str] | None = None,
) -> str:
    formula = _official_formula(coordinate)
    result = ["="]
    for token in Tokenizer(formula).items:
        value = token.value
        match = (
            TEST_RANGE_REFERENCE.fullmatch(value)
            if token.type == "OPERAND" and token.subtype == "RANGE"
            else None
        )
        if match is not None:
            if value in (range_overrides or {}):
                value = range_overrides[value]
            else:
                value = translate_formula(
                    "=" + value,
                    origin=coordinate,
                    target=target,
                )[1:]
                endpoints = [match.group("first")]
                if match.group("last"):
                    endpoints.append(match.group("last"))
                parsed = [TEST_CELL_REFERENCE.fullmatch(item) for item in endpoints]
                rows = [int(item.group("row")) for item in parsed if item is not None]
                if (
                    len(rows) == len(endpoints)
                    and _is_test_local_mobiliti_sheet(match.group("sheet"))
                    and all(573 <= row <= 610 for row in rows)
                ):
                    value = _reference_with_test_rows(
                        token.value,
                        tuple(row_map.total_row + row - 573 for row in rows),
                    )
        result.append(value)
    return "".join(result)


def _expected_total_subtotal_rows(row_map, coordinate: str = "H573") -> list[int]:
    canonical = [
        int(token.value[1:])
        for token in Tokenizer(_official_formula(coordinate)).items
        if token.type == "OPERAND"
        and token.subtype == "RANGE"
        and token.value.startswith("H")
        and token.value[1:].isdigit()
        and int(token.value[1:]) in range(47, 573, 35)
    ]
    assert len(canonical) == 16
    return (
        list(reversed(row_map.subtotal_rows))
        if canonical[0] > canonical[-1]
        else list(row_map.subtotal_rows)
    )


def _formula_token_signature(formula: str) -> list[tuple[str, str, str]]:
    return [(token.type, token.subtype, token.value) for token in Tokenizer(formula).items]


@pytest.mark.parametrize("count", [34, 100])
def test_one_section_keeps_every_product_and_official_formulas(tmp_path, count):
    row_map, workbook, worksheet = _render(tmp_path, count)
    try:
        rows = list(row_map.item_rows)
        assert len(rows) == count
        assert rows == list(range(14, 14 + count))
        assert worksheet.cell(14 + count, 1).value == "Subtotales Sección 1"
        assert all(worksheet.cell(row, 4).value == f"SKU-{row - 14:03d}" for row in rows)
        assert "#REF!" not in "\n".join(
            value
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance((value := cell.value), str) and value.startswith("=")
        )
    finally:
        workbook.close()


@pytest.mark.parametrize("section_count", [17, 20])
def test_more_than_sixteen_sections_clone_official_blocks(tmp_path, section_count):
    mutation, workbook, worksheet = _render_sections(tmp_path, section_count)
    try:
        row_map = mutation.row_map
        assert len(row_map.sections) == section_count
        assert all(
            worksheet.cell(section.section_row, 4 if index == 0 else 1).value
            == f"SECCION {index + 1}"
            for index, section in enumerate(row_map.sections)
        )
        last = row_map.sections[-1]
        assert worksheet.cell(last.product_start, 4).value == f"SKU-{section_count - 1:03d}"
        assert worksheet.cell(last.product_start + 1, 4).value is None
        assert worksheet.cell(last.product_start + 1, 23).value == translate_formula(
            _official_formula("W49"),
            origin="W49",
            target=f"W{last.product_start + 1}",
        )
        subtotal_formula = worksheet.cell(last.subtotal_row, 8).value
        assert subtotal_formula.ref == f"H{last.subtotal_row}"
        assert subtotal_formula.text == _derive_official_formula(
            "H82",
            f"H{last.subtotal_row}",
            row_map,
            range_overrides={
                "H49:H81": f"H{last.product_start}:H{last.product_start}"
            },
        )
        assert _formula_rows(
            worksheet.cell(row_map.total_row, 8).value, "H"
        ) == _expected_total_subtotal_rows(row_map)
    finally:
        workbook.close()


def test_seventeenth_section_extends_official_x_global_product_ranges():
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(17)
    ]
    part = _official_part(OFFICIAL_TEMPLATE, "Mobiliti")
    mutation = build_mobiliti_sheet(
        _part_bytes(OFFICIAL_TEMPLATE, part),
        needs,
        [],
    )
    root = ET.fromstring(mutation.xml)
    last_row = mutation.row_map.sections[-1].product_start
    formula = root.find(f".//{{{MAIN}}}c[@r='X{last_row}']/{{{MAIN}}}f")

    assert formula is not None and formula.text
    assert (
        f"$W$14:$W${mutation.row_map.last_product_row}"
        in formula.text
    )
    assert (
        f"$D$14:$D${mutation.row_map.last_product_row}"
        in formula.text
    )
    assert (
        f"$H$14:$H${mutation.row_map.last_product_row}"
        in formula.text
    )
    assert "_xlfn.MAXIFS(" in formula.text


def test_validations_and_conditional_formatting_reach_twentieth_section(tmp_path):
    mutation, workbook, _worksheet = _render_sections(tmp_path, 20)
    workbook.close()
    root = ET.fromstring(mutation.xml)
    last = mutation.row_map.sections[-1]
    last_range = f"{last.product_start}:{last.product_start + last.capacity - 1}"

    validation_sqrefs = " ".join(
        validation.attrib.get("sqref", "")
        for validation in root.findall(f".//{{{MAIN}}}dataValidation")
    )
    extension_sqrefs = " ".join(
        node.text or "" for node in root.findall(f".//{{{XM}}}sqref")
    )
    conditional_sqrefs = " ".join(
        node.attrib.get("sqref", "")
        for node in root.findall(f"{{{MAIN}}}conditionalFormatting")
    )

    assert f"E{last_range.replace(':', ':E')}" in validation_sqrefs
    assert f"F{last_range.replace(':', ':F')}" in extension_sqrefs
    assert f"AH{last_range.replace(':', ':AH')}" in conditional_sqrefs


def test_static_sidecar_formulas_follow_relocated_total_and_auxiliary_rows(tmp_path):
    mutation, workbook, worksheet = _render_sections(tmp_path, 20)
    official = load_workbook(TEMPLATE, data_only=False, keep_links=True)
    try:
        assert worksheet["P9"].value == f"=P8/H{mutation.row_map.total_row}"
        assert worksheet["AS15"].value == f"=AC{mutation.row_map.total_row}"
        assert worksheet.cell(mutation.row_map.total_row + 37, 36).style_id == official["Mobiliti"]["AJ610"].style_id
        assert worksheet.row_dimensions[mutation.row_map.total_row + 37].height == (
            official["Mobiliti"].row_dimensions[610].height
        )
    finally:
        official.close()
        workbook.close()


def test_unknown_formula_metadata_and_extension_nodes_survive_cloning():
    part = _official_part(TEMPLATE, "Mobiliti")
    root = ET.fromstring(_part_bytes(TEMPLATE, part))
    custom_namespace = "urn:mobiliti:test:unknown"
    formula = root.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")
    assert formula is not None
    formula.set(f"{{{custom_namespace}}}audit", "keep")
    marker = ET.SubElement(root, f"{{{custom_namespace}}}marker", {"keep": "yes"})
    marker.text = "opaque"

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [SectionNeed("chairs", "SILLAS", 1)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    cloned_formula = output.find(f".//{{{MAIN}}}c[@r='W15']/{{{MAIN}}}f")

    assert cloned_formula is not None
    assert cloned_formula.attrib[f"{{{custom_namespace}}}audit"] == "keep"
    assert not ({"t", "ref", "si"} & set(cloned_formula.attrib))
    assert output.find(f"{{{custom_namespace}}}marker").text == "opaque"
    assert output.find(f"{{{MAIN}}}extLst") is not None


def test_shared_formula_follower_is_materialized_from_its_master_before_cloning():
    root = _official_root()
    follower = root.find(f".//{{{MAIN}}}c[@r='G15']/{{{MAIN}}}f")
    assert follower is not None
    assert follower.attrib == {"t": "shared", "si": "1"}
    assert follower.text is None

    editor = WorksheetEditor(root)
    row_map = plan_mobiliti_layout([SectionNeed("one", "UNO", 1)])
    editor.replace_table_row(
        100,
        editor.require_row(15),
        15,
        row_map,
    )
    cloned = editor.root.find(f".//{{{MAIN}}}c[@r='G100']/{{{MAIN}}}f")

    assert cloned is not None
    assert cloned.text == 'IFERROR(VLOOKUP(F100,Tabla_Proveedores_1,2,0)," ")'
    assert not ({"t", "ref", "si"} & set(cloned.attrib))


@pytest.mark.parametrize("formula_type", ["array", "dataTable"])
def test_special_formula_in_a_cloned_row_fails_closed_before_mutation(formula_type):
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")
    assert formula is not None
    formula.set("t", formula_type)
    formula.set("ref", "W14:W46")
    payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    with pytest.raises(ValueError, match=fr"{formula_type}.*W14.*clonar"):
        build_mobiliti_sheet(
            payload,
            [SectionNeed("chairs", "SILLAS", 1)],
            [],
        )

    assert ET.fromstring(payload).find(
        f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f"
    ).attrib["t"] == formula_type


def test_unmoved_array_formula_is_preserved_with_cache_and_attributes():
    root = _official_root()
    cell = root.find(f".//{{{MAIN}}}c[@r='P9']")
    assert cell is not None
    original_formula = cell.find(f"{{{MAIN}}}f")
    assert original_formula is not None
    original_formula.set("t", "array")
    original_formula.set("ref", "P9")
    original_formula.set("keep", "opaque")
    cached = cell.find(f"{{{MAIN}}}v")
    if cached is None:
        cached = ET.SubElement(cell, f"{{{MAIN}}}v")
    cached.text = "123"

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [],
        [],
    )
    output = ET.fromstring(mutation.xml)
    result_cell = output.find(f".//{{{MAIN}}}c[@r='P9']")
    result_formula = result_cell.find(f"{{{MAIN}}}f")

    assert result_formula.text == original_formula.text
    assert result_formula.attrib == original_formula.attrib
    assert result_cell.find(f"{{{MAIN}}}v").text == "123"


@pytest.mark.parametrize(
    "needs",
    [
        [SectionNeed("large", "GRANDE", 34)],
        [
            SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
            for index in range(20)
        ],
    ],
    ids=["34-products", "20-sections"],
)
@pytest.mark.parametrize(
    ("coordinate", "formula_type"),
    [("P9", "array"), ("AS15", "dataTable")],
)
def test_static_special_formulas_follow_relocated_total_preserving_metadata(
    needs, coordinate, formula_type
):
    root = _official_root()
    cell = root.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
    formula = cell.find(f"{{{MAIN}}}f")
    assert formula is not None
    formula.attrib.clear()
    formula.set("t", formula_type)
    formula.set("ref", coordinate)
    formula.set("opaque", "keep")
    if formula_type == "array":
        formula.text = "P8/H573"
    else:
        formula.text = None
        formula.set("r1", "AC573")
        formula.set("r2", "$H$574")
        formula.set("inputCell", "Mobiliti!$AC$573")

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    result_cell = output.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
    result = result_cell.find(f"{{{MAIN}}}f")

    assert result.attrib["t"] == formula_type
    assert result.attrib["ref"] == coordinate
    assert result.attrib["opaque"] == "keep"
    if formula_type == "array":
        assert result.text == f"P8/H{mutation.row_map.total_row}"
    else:
        assert result.text is None
        assert result.attrib["r1"] == f"AC{mutation.row_map.total_row}"
        assert result.attrib["r2"] == f"$H${mutation.row_map.total_row + 1}"
        assert result.attrib["inputCell"] == f"Mobiliti!$AC${mutation.row_map.total_row}"
    assert result_cell.find(f"{{{MAIN}}}v") is None


@pytest.mark.parametrize("formula_type", ["array", "dataTable"])
def test_static_multicell_special_that_references_moved_total_fails_preflight(
    formula_type
):
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='P9']/{{{MAIN}}}f")
    assert formula is not None
    formula.attrib.clear()
    formula.set("t", formula_type)
    formula.set("ref", "P9:Q9")
    formula.set("r1", "H573")
    formula.text = "P8/H573" if formula_type == "array" else None

    with pytest.raises(ValueError, match=fr"{formula_type}.*P9.*preflight"):
        build_mobiliti_sheet(
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
            [SectionNeed("large", "GRANDE", 34)],
            [],
        )


def test_single_cell_datatable_without_text_translates_all_a1_attributes_when_cloned():
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")
    assert formula is not None
    formula.attrib.clear()
    formula.set("t", "dataTable")
    formula.set("ref", "W14")
    formula.set("r1", "H573")
    formula.set("r2", "$AC$574")
    formula.set("inputCell", "Mobiliti!$H$573")
    formula.set("opaque", "keep")
    formula.text = None

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [SectionNeed("large", "GRANDE", 34)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    for row in (14, 15):
        cell = output.find(f".//{{{MAIN}}}c[@r='W{row}']")
        result = cell.find(f"{{{MAIN}}}f")
        assert result.text is None
        assert result.attrib == {
            "t": "dataTable",
            "ref": f"W{row}",
            "r1": f"H{mutation.row_map.total_row}",
            "r2": f"$AC${mutation.row_map.total_row + 1}",
            "inputCell": f"Mobiliti!$H${mutation.row_map.total_row}",
            "opaque": "keep",
        }
        assert cell.find(f"{{{MAIN}}}v") is None


def test_multicell_datatable_without_text_fails_closed_before_mutation():
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")
    assert formula is not None
    formula.attrib.clear()
    formula.set("t", "dataTable")
    formula.set("ref", "W14:W46")
    formula.set("r1", "H573")
    formula.text = None
    payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    with pytest.raises(ValueError, match=r"dataTable.*W14.*clonar"):
        build_mobiliti_sheet(
            payload,
            [SectionNeed("large", "GRANDE", 34)],
            [],
        )

    result = ET.fromstring(payload).find(
        f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f"
    )
    assert result.attrib["ref"] == "W14:W46"
    assert result.text is None


def test_total_shared_follower_without_text_expands_exact_subtotal_sequence():
    root = _official_root()
    total_row = root.find(f".//{{{MAIN}}}row[@r='573']")
    follower = total_row.find(f"{{{MAIN}}}c[@r='H573']/{{{MAIN}}}f")
    assert follower is not None and follower.text
    original = "=" + follower.text
    master_cell = ET.Element(f"{{{MAIN}}}c", {"r": "AK573"})
    master = ET.SubElement(
        master_cell,
        f"{{{MAIN}}}f",
        {"t": "shared", "si": "900", "ref": "H573:AK573"},
    )
    master.text = translate_formula(
        original,
        origin="H573",
        target="AK573",
    )[1:]
    total_row.append(master_cell)
    follower.attrib.clear()
    follower.set("t", "shared")
    follower.set("si", "900")
    follower.text = None
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(20)
    ]

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    result = output.find(
        f".//{{{MAIN}}}c[@r='H{mutation.row_map.total_row}']/{{{MAIN}}}f"
    )
    local_rows = [
        int(token.value[1:])
        for token in Tokenizer("=" + result.text).items
        if token.type == "OPERAND"
        and token.subtype == "RANGE"
        and token.value.startswith("H")
        and token.value[1:].isdigit()
    ]

    assert local_rows == _expected_total_subtotal_rows(mutation.row_map)
    assert len(local_rows) == 20
    assert not ({"t", "ref", "si"} & set(result.attrib))


@pytest.mark.parametrize(
    "sheet_data",
    [
        f'<row xmlns="{MAIN}" r="0"/>',
        f'<row xmlns="{MAIN}" r="2"/><row xmlns="{MAIN}" r="1"/>',
        f'<row xmlns="{MAIN}" r="1"/><row xmlns="{MAIN}" r="1"/>',
        f'<row xmlns="{MAIN}" r="1"><c r="A2"/></row>',
        f'<row xmlns="{MAIN}" r="1"><c r="A1"/><c r="A1"/></row>',
        f'<row xmlns="{MAIN}" r="1"><c r="XFE1"/></row>',
        f'<row xmlns="{MAIN}" r="1"><c r="B1"/><c r="A1"/></row>',
    ],
)
def test_worksheet_editor_rejects_invalid_row_and_cell_structure(sheet_data):
    payload = (
        f'<worksheet xmlns="{MAIN}"><sheetData>{sheet_data}</sheetData></worksheet>'
    ).encode()

    with pytest.raises(ValueError, match="estructura.*Mobiliti"):
        WorksheetEditor.from_xml(payload)


def _official_editor_and_row_map(item_count: int = 1):
    editor = WorksheetEditor.from_xml(
        ET.tostring(_official_root(), encoding="utf-8", xml_declaration=True)
    )
    row_map = plan_mobiliti_layout(
        [SectionNeed("chairs", "SILLAS", item_count)]
    )
    return editor, row_map


def test_cell_writes_reject_an_unused_capacity_row():
    editor, row_map = _official_editor_and_row_map(item_count=1)

    with pytest.raises(ValueError, match="fuera de inputs Mobiliti"):
        apply_mobiliti_cell_writes(
            editor,
            [MobilitiCellWrite("D15", "text", "NO DEBE ESCRIBIRSE")],
            row_map,
        )


@pytest.mark.parametrize("value", [Decimal("NaN"), Decimal("Infinity"), Decimal("-Infinity")])
def test_cell_writes_reject_non_finite_decimals(value):
    editor, row_map = _official_editor_and_row_map()

    with pytest.raises(ValueError, match="finito"):
        apply_mobiliti_cell_writes(
            editor,
            [MobilitiCellWrite("H14", "number", value)],
            row_map,
        )


@pytest.mark.parametrize("value", ["control\x01", "surrogate\ud800"])
def test_cell_writes_reject_text_that_is_not_xml_10(value):
    editor, row_map = _official_editor_and_row_map()

    with pytest.raises(ValueError, match="XML 1.0"):
        apply_mobiliti_cell_writes(
            editor,
            [MobilitiCellWrite("D14", "text", value)],
            row_map,
        )


def test_cell_write_preserves_significant_whitespace_and_escapes_xml():
    editor, row_map = _official_editor_and_row_map()
    value = " <SKU & modelo> "

    apply_mobiliti_cell_writes(
        editor,
        [MobilitiCellWrite("D14", "text", value)],
        row_map,
    )
    payload = editor.to_xml()
    output = ET.fromstring(payload)
    text = output.find(f".//{{{MAIN}}}c[@r='D14']/{{{MAIN}}}is/{{{MAIN}}}t")

    assert text.text == value
    assert text.attrib["{http://www.w3.org/XML/1998/namespace}space"] == "preserve"
    assert b"&lt;SKU &amp; modelo&gt;" in payload


def test_typed_editor_setters_are_strict_and_idempotent():
    editor, _row_map = _official_editor_and_row_map()

    editor.set_boolean("K4", True)
    editor.set_inline_string("K8", " Guadalajara & Norte ")
    once = editor.to_xml()
    editor.set_boolean("K4", True)
    editor.set_inline_string("K8", " Guadalajara & Norte ")

    assert editor.to_xml() == once
    output = ET.fromstring(once)
    assert output.find(f".//{{{MAIN}}}c[@r='K4']/{{{MAIN}}}v").text == "1"
    text = output.find(f".//{{{MAIN}}}c[@r='K8']/{{{MAIN}}}is/{{{MAIN}}}t")
    assert text.text == " Guadalajara & Norte "
    assert text.attrib["{http://www.w3.org/XML/1998/namespace}space"] == "preserve"


def test_typed_editor_batch_is_atomic_when_a_destination_cell_is_absent():
    editor, _row_map = _official_editor_and_row_map()
    row = editor.require_row(8)
    k8 = row.find(f"{{{MAIN}}}c[@r='K8']")
    assert k8 is not None
    row.remove(k8)
    before = editor.to_xml()

    with pytest.raises(ValueError, match="K8"):
        editor.set_typed_values(
            (
                MobilitiCellWrite("K4", "boolean", False),
                MobilitiCellWrite("K8", "text", "Guadalajara"),
            )
        )

    assert editor.to_xml() == before


@pytest.mark.parametrize(
    ("method", "coordinate", "value", "message"),
    [
        ("set_boolean", "K0", True, "Coordenada"),
        ("set_boolean", "K4", 1, "bool"),
        ("set_inline_string", "K999999", "texto", "fila"),
        ("set_inline_string", "K8", "control\x01", "XML 1.0"),
        pytest.param(
            "set_inline_string",
            "K8",
            "x" * 32_768,
            "32767",
            id="oversized-inline-string",
        ),
    ],
)
def test_typed_editor_setters_reject_invalid_coordinates_and_values(
    method, coordinate, value, message
):
    editor, _row_map = _official_editor_and_row_map()

    with pytest.raises((TypeError, ValueError), match=message):
        getattr(editor, method)(coordinate, value)


def test_multirange_sqref_translates_mixed_fixed_dynamic_and_x14_tokens():
    root = _official_root()
    data_validations = root.find(f"{{{MAIN}}}dataValidations")
    assert data_validations is not None
    ET.SubElement(
        data_validations,
        f"{{{MAIN}}}dataValidation",
        {
            "type": "list",
            "marker": "mixed-dv",
            "sqref": "A1 E49:E81 $B$2 E49:E81",
        },
    )
    conditional = ET.Element(
        f"{{{MAIN}}}conditionalFormatting",
        {"marker": "mixed-cf", "sqref": "A1 W49:W81 $B$2 W49:W81"},
    )
    rule = ET.SubElement(conditional, f"{{{MAIN}}}cfRule", {"type": "expression"})
    ET.SubElement(rule, f"{{{MAIN}}}formula").text = "W49>0"
    root.insert(list(root).index(data_validations), conditional)
    ext_lst = root.find(f"{{{MAIN}}}extLst")
    assert ext_lst is not None
    extension = ET.SubElement(ext_lst, "{urn:mobiliti:test:x14}validation")
    extension.set("marker", "mixed-x14")
    ET.SubElement(extension, f"{{{XM}}}sqref").text = (
        "A1 F49:F81 $C$3 F49:F81"
    )
    needs = [SectionNeed("first", "FIRST", 34)] + [
        SectionNeed(f"s-{index}", f"S{index}", 1) for index in range(2, 21)
    ]

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    second = mutation.row_map.sections[1]
    extra = mutation.row_map.sections[16:]
    dynamic_e = [
        f"E{second.product_start}:E{second.product_start + second.capacity - 1}",
        *[
            f"E{section.product_start}:E{section.product_start + section.capacity - 1}"
            for section in extra
        ],
    ]
    dynamic_f = [value.replace("E", "F") for value in dynamic_e]
    dynamic_w = [value.replace("E", "W") for value in dynamic_e]

    validation = output.find(
        f".//{{{MAIN}}}dataValidation[@marker='mixed-dv']"
    )
    assert validation.attrib["sqref"].split() == ["A1", dynamic_e[0], "$B$2", *dynamic_e[1:]]
    x14_sqref = output.find(".//*[@marker='mixed-x14']/{%s}sqref" % XM)
    assert x14_sqref.text.split() == ["A1", dynamic_f[0], "$C$3", *dynamic_f[1:]]

    conditional_nodes = output.findall(
        f"{{{MAIN}}}conditionalFormatting[@marker='mixed-cf']"
    )
    assert len(conditional_nodes) == len(dynamic_w)
    assert [
        token
        for node in conditional_nodes
        for token in node.attrib["sqref"].split()
        if token.startswith("W")
    ] == dynamic_w
    fixed_tokens = [
        token
        for node in conditional_nodes
        for token in node.attrib["sqref"].split()
        if token in {"A1", "$B$2"}
    ]
    assert fixed_tokens == ["A1", "$B$2"]
    assert all(len(tokens := node.attrib["sqref"].split()) == len(set(tokens)) for node in conditional_nodes)


def test_dv_and_x14_extra_sections_preserve_full_absolute_template_tokens():
    root = _official_root()
    data_validations = root.find(f"{{{MAIN}}}dataValidations")
    assert data_validations is not None
    ET.SubElement(
        data_validations,
        f"{{{MAIN}}}dataValidation",
        {
            "type": "list",
            "marker": "absolute-dv",
            "sqref": "'Mobiliti'!$E$49:$E$81 $B$2",
        },
    )
    ext_lst = root.find(f"{{{MAIN}}}extLst")
    extension = ET.SubElement(ext_lst, "{urn:mobiliti:test:x14}validation")
    extension.set("marker", "absolute-x14")
    ET.SubElement(extension, f"{{{XM}}}sqref").text = "$F$49:$F$81 $C$3"
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(20)
    ]

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    target_sections = [mutation.row_map.sections[1], *mutation.row_map.sections[16:]]
    expected_dv = [
        (
            f"'Mobiliti'!$E${section.product_start}:"
            f"$E${section.product_start + section.capacity - 1}"
        )
        for section in target_sections
    ]
    expected_x14 = [
        (
            f"$F${section.product_start}:"
            f"$F${section.product_start + section.capacity - 1}"
        )
        for section in target_sections
    ]
    validation = output.find(
        f".//{{{MAIN}}}dataValidation[@marker='absolute-dv']"
    )
    x14_sqref = output.find(".//*[@marker='absolute-x14']/{%s}sqref" % XM)

    assert validation.attrib["sqref"].split() == [
        expected_dv[0], "$B$2", *expected_dv[1:]
    ]
    assert x14_sqref.text.split() == [expected_x14[0], "$C$3", *expected_x14[1:]]
    assert len(validation.attrib["sqref"].split()) == len(
        set(validation.attrib["sqref"].split())
    )
    assert len(x14_sqref.text.split()) == len(set(x14_sqref.text.split()))


@pytest.mark.parametrize(
    "needs",
    [
        [SectionNeed("first", "FIRST", 34)],
        [SectionNeed("first", "FIRST", 100)],
        [SectionNeed(f"s-{index}", f"S{index}", 1) for index in range(20)],
    ],
    ids=["34-products", "100-products", "20-sections"],
)
def test_merges_clone_product_and_total_rows_preserving_cross_boundary(needs):
    root = _official_root()
    merges = root.find(f"{{{MAIN}}}mergeCells")
    assert merges is not None
    for reference in ("D14:E14", "F49:G49", "A573:B573", "AH14:AI14"):
        ET.SubElement(merges, f"{{{MAIN}}}mergeCell", {"ref": reference})
    merges.set("count", str(len(merges)))

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    references = [
        node.attrib["ref"]
        for node in output.findall(f"{{{MAIN}}}mergeCells/{{{MAIN}}}mergeCell")
    ]
    expected_product_merges = [
        f"{('D' if index == 0 else 'F')}{row}:{('E' if index == 0 else 'G')}{row}"
        for index, section in enumerate(mutation.row_map.sections)
        for row in range(section.product_start, section.product_start + section.capacity)
    ]
    expected_total = f"A{mutation.row_map.total_row}:B{mutation.row_map.total_row}"

    assert all(reference in references for reference in expected_product_merges)
    assert expected_total in references
    assert references.count("AH14:AI14") == 1
    assert references.count("AR14:AT14") == 1
    assert len(references) == len(set(references))
    assert references == sorted(references, key=_merge_sort_key)


def test_moved_total_removes_stale_ai_aj_cells_and_merges_but_keeps_ak_sidecar():
    root = _official_root()
    total_row = root.find(f".//{{{MAIN}}}row[@r='573']")
    assert total_row is not None
    for coordinate, value in (
        ("AI573", "TOTAL-AI"),
        ("AJ573", "TOTAL-AJ"),
        ("AK573", "SIDECAR-AK"),
    ):
        existing = total_row.find(f"{{{MAIN}}}c[@r='{coordinate}']")
        if existing is not None:
            total_row.remove(existing)
        cell = ET.SubElement(total_row, f"{{{MAIN}}}c", {"r": coordinate, "t": "inlineStr"})
        inline = ET.SubElement(cell, f"{{{MAIN}}}is")
        ET.SubElement(inline, f"{{{MAIN}}}t").text = value
    total_row[:] = sorted(
        total_row,
        key=lambda cell: __import__("openpyxl").utils.column_index_from_string(
            "".join(filter(str.isalpha, cell.attrib["r"]))
        ),
    )
    merges = root.find(f"{{{MAIN}}}mergeCells")
    assert merges is not None
    for reference in ("AH573:AI573", "AI573:AJ573"):
        ET.SubElement(merges, f"{{{MAIN}}}mergeCell", {"ref": reference})
    merges.set("count", str(len(merges)))
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(20)
    ]

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    stale_row = output.find(f".//{{{MAIN}}}row[@r='573']")
    moved_row = output.find(
        f".//{{{MAIN}}}row[@r='{mutation.row_map.total_row}']"
    )
    merge_refs = [
        node.attrib["ref"]
        for node in output.findall(f"{{{MAIN}}}mergeCells/{{{MAIN}}}mergeCell")
    ]

    assert stale_row.find(f"{{{MAIN}}}c[@r='AI573']") is None
    assert stale_row.find(f"{{{MAIN}}}c[@r='AJ573']") is None
    assert stale_row.find(f"{{{MAIN}}}c[@r='AK573']") is not None
    assert moved_row.find(
        f"{{{MAIN}}}c[@r='AI{mutation.row_map.total_row}']"
    ) is not None
    assert moved_row.find(
        f"{{{MAIN}}}c[@r='AJ{mutation.row_map.total_row}']"
    ) is not None
    expected_merges = {
        f"AH{mutation.row_map.total_row}:AI{mutation.row_map.total_row}",
        f"AI{mutation.row_map.total_row}:AJ{mutation.row_map.total_row}",
    }
    assert "AH573:AI573" not in merge_refs
    assert "AI573:AJ573" not in merge_refs
    assert expected_merges <= set(merge_refs)
    assert all(merge_refs.count(reference) == 1 for reference in expected_merges)


def _merge_sort_key(reference: str):
    first, last = reference.split(":")
    first_column = "".join(filter(str.isalpha, first))
    last_column = "".join(filter(str.isalpha, last))
    first_row = int("".join(filter(str.isdigit, first)))
    last_row = int("".join(filter(str.isdigit, last)))
    return (
        first_row,
        __import__("openpyxl").utils.column_index_from_string(first_column),
        last_row,
        __import__("openpyxl").utils.column_index_from_string(last_column),
    )


def test_structural_overrides_do_not_rewrite_another_sheet():
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")
    assert formula is not None
    formula.text = "Other!H573+H573+Mobiliti!H573"

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 34)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    result = output.find(f".//{{{MAIN}}}c[@r='W14']/{{{MAIN}}}f")

    assert result.text == f"Other!H573+H{mutation.row_map.total_row}+Mobiliti!H{mutation.row_map.total_row}"


def test_total_translates_whole_formula_and_replaces_only_local_canonical_operands():
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='H573']/{{{MAIN}}}f")
    assert formula is not None and formula.text
    formula.text = f"Other!H47+($Z$1)+({formula.text})+Other!H82"

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 34)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    result = output.find(
        f".//{{{MAIN}}}c[@r='H{mutation.row_map.total_row}']/{{{MAIN}}}f"
    ).text
    expected_external = translate_formula(
        "=Other!H47+($Z$1)+Other!H82",
        origin="H573",
        target=f"H{mutation.row_map.total_row}",
    )[1:]
    local_rows = [
        int(token.value[1:])
        for token in Tokenizer("=" + result).items
        if token.type == "OPERAND"
        and token.subtype == "RANGE"
        and token.value.startswith("H")
        and token.value[1:].isdigit()
    ]

    assert local_rows == _expected_total_subtotal_rows(mutation.row_map)
    assert result.startswith(expected_external.split("+Other!", 1)[0] + "+(")
    assert result.endswith("+" + expected_external.rsplit("+", 1)[-1])


def test_total_translates_structural_prefix_suffix_before_expanding_operands():
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='H573']/{{{MAIN}}}f")
    assert formula is not None and formula.text
    formula.text = (
        f"$AJ$574+Other!$AJ$574+({formula.text})+Mobiliti!$AJ$610"
    )
    needs = [
        SectionNeed(f"section-{index}", f"SECCION {index + 1}", 1)
        for index in range(20)
    ]

    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True), needs, []
    )
    output = ET.fromstring(mutation.xml)
    result = output.find(
        f".//{{{MAIN}}}c[@r='H{mutation.row_map.total_row}']/{{{MAIN}}}f"
    ).text
    local_rows = [
        int(token.value[1:])
        for token in Tokenizer("=" + result).items
        if token.type == "OPERAND"
        and token.subtype == "RANGE"
        and token.value.startswith("H")
        and token.value[1:].isdigit()
    ]

    assert result.startswith(
        f"$AJ${mutation.row_map.total_row + 1}+Other!$AJ$574+("
    )
    assert result.endswith(f")+Mobiliti!$AJ${mutation.row_map.total_row + 37}")
    assert local_rows == _expected_total_subtotal_rows(mutation.row_map)
    assert len(local_rows) == 20


@pytest.mark.parametrize("corruption", ["duplicate", "missing", "reordered"])
def test_total_rejects_non_unique_missing_or_reordered_canonical_operands(corruption):
    root = _official_root()
    formula = root.find(f".//{{{MAIN}}}c[@r='H573']/{{{MAIN}}}f")
    assert formula is not None and formula.text
    if corruption == "duplicate":
        formula.text += "+H47"
    elif corruption == "missing":
        formula.text = formula.text.replace("+H82", "")
    else:
        formula.text = formula.text.replace("H572+H537", "H537+H572")

    with pytest.raises(ValueError, match="16 subtotales.*orden"):
        build_mobiliti_sheet(
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
            [SectionNeed("first", "FIRST", 34)],
            [],
        )


def test_changed_static_formulas_drop_stale_cached_values():
    mutation = build_mobiliti_sheet(
        ET.tostring(_official_root(), encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 34)],
        [],
    )
    output = ET.fromstring(mutation.xml)

    for coordinate in ("P9", "AS15"):
        cell = output.find(f".//{{{MAIN}}}c[@r='{coordinate}']")
        assert cell is not None
        assert cell.find(f"{{{MAIN}}}f") is not None
        assert cell.find(f"{{{MAIN}}}v") is None


def test_empty_section_subtotal_is_official_but_has_no_phantom_product_operand():
    mutation = build_mobiliti_sheet(
        ET.tostring(_official_root(), encoding="utf-8", xml_declaration=True),
        [SectionNeed("empty", "VACIA", 0)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    section = mutation.row_map.sections[0]
    formula = output.find(
        f".//{{{MAIN}}}c[@r='H{section.subtotal_row}']/{{{MAIN}}}f"
    )
    official_tokens = Tokenizer(_official_formula("H47")).items
    expected = "=" + "".join(
        "0" if token.value == "H14:H46" else token.value
        for token in official_tokens
    )

    assert mutation.row_map.item_rows == ()
    assert formula is not None
    assert formula.text == expected[1:]
    assert formula.attrib["t"] == "array"
    assert formula.attrib["ref"] == f"H{section.subtotal_row}"
    assert not any(
        token.type == "OPERAND"
        and token.subtype == "RANGE"
        and token.value in {f"H{row}" for row in range(section.product_start, section.subtotal_row)}
        for token in Tokenizer("=" + formula.text).items
    )


def test_all_product_rows_keep_the_official_currency_reference():
    mutation = build_mobiliti_sheet(
        ET.tostring(_official_root(), encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 1), SectionNeed("second", "SECOND", 1)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    product_rows = [
        row
        for section in mutation.row_map.sections
        for row in range(section.product_start, section.subtotal_row)
    ]

    assert output.find(f".//{{{MAIN}}}c[@r='C{product_rows[0]}']/{{{MAIN}}}f").text == "C13"
    assert {
        output.find(f".//{{{MAIN}}}c[@r='C{row}']/{{{MAIN}}}f").text
        for row in product_rows[1:]
    } == {"$C$13"}


def test_provider_cells_keep_the_yellow_product_row_style_in_every_section():
    mutation = build_mobiliti_sheet(
        ET.tostring(_official_root(), encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 1), SectionNeed("second", "SECOND", 1)],
        [],
    )
    output = ET.fromstring(mutation.xml)

    for section in mutation.row_map.sections:
        for row in range(section.product_start, section.subtotal_row):
            provider = output.find(f".//{{{MAIN}}}c[@r='F{row}']")
            yellow_reference = output.find(f".//{{{MAIN}}}c[@r='H{row}']")
            assert provider is not None
            assert yellow_reference is not None
            assert provider.attrib.get("s") == yellow_reference.attrib.get("s")


def test_all_product_rows_keep_styles_and_used_unused_rows_match_official_formulas():
    root = _official_root()
    mutation = build_mobiliti_sheet(
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        [SectionNeed("first", "FIRST", 34)],
        [],
    )
    output = ET.fromstring(mutation.xml)
    input_columns = {4, 5, 6, 8, 10, 11, 16}
    used_row = mutation.row_map.item_rows[0]
    used_rows = set(mutation.row_map.item_rows)
    unused_row = next(
        row
        for section in mutation.row_map.sections
        for row in range(section.product_start, section.subtotal_row)
        if row not in used_rows
    )
    formula_check_rows = {used_row, unused_row}

    for section in mutation.row_map.sections:
        # La fila 14 es la superficie canónica completa. Las filas posteriores
        # aportan presentación, pero no un contrato alterno de fórmulas.
        source_row_number = 14
        source_row = root.find(f".//{{{MAIN}}}row[@r='{source_row_number}']")
        assert source_row is not None
        source_cells = {
            __import__("openpyxl").utils.column_index_from_string(
                "".join(filter(str.isalpha, cell.attrib["r"]))
            ): cell
            for cell in source_row.findall(f"{{{MAIN}}}c")
            if __import__("openpyxl").utils.column_index_from_string(
                "".join(filter(str.isalpha, cell.attrib["r"]))
            ) <= 34
        }
        formula_columns = {
            column
            for column, cell in source_cells.items()
            if column not in input_columns and cell.find(f"{{{MAIN}}}f") is not None
        }
        for row_number in range(section.product_start, section.subtotal_row):
            target_row = output.find(f".//{{{MAIN}}}row[@r='{row_number}']")
            assert target_row is not None
            target_cells = {
                __import__("openpyxl").utils.column_index_from_string(
                    "".join(filter(str.isalpha, cell.attrib["r"]))
                ): cell
                for cell in target_row.findall(f"{{{MAIN}}}c")
                if __import__("openpyxl").utils.column_index_from_string(
                    "".join(filter(str.isalpha, cell.attrib["r"]))
                ) <= 34
            }
            expected_styles = {
                column: cell.attrib.get("s") for column, cell in source_cells.items()
            }
            expected_styles[6] = source_cells[8].attrib.get("s")
            assert {
                column: cell.attrib.get("s") for column, cell in target_cells.items()
            } == expected_styles
            assert all(
                target_cells[column].find(f"{{{MAIN}}}f") is not None
                for column in formula_columns
            )
            if row_number in formula_check_rows:
                for column in formula_columns:
                    formula_source_row = (
                        49
                        if column == 3 and row_number != mutation.row_map.sections[0].product_start
                        else source_row_number
                    )
                    source_coordinate = (
                        f"{__import__('openpyxl').utils.get_column_letter(column)}"
                        f"{formula_source_row}"
                    )
                    target_coordinate = (
                        f"{__import__('openpyxl').utils.get_column_letter(column)}"
                        f"{row_number}"
                    )
                    actual = target_cells[column].find(f"{{{MAIN}}}f").text
                    if column == 24:
                        first_product_row = mutation.row_map.sections[0].product_start
                        last_product_row = mutation.row_map.last_product_row
                        expected = (
                            f"_xlfn.MINIFS($W${first_product_row}:$W${last_product_row},"
                            f"$D${first_product_row}:$D${last_product_row},D{row_number},"
                            f"$H${first_product_row}:$H${last_product_row},"
                            f"_xlfn.MAXIFS($H${first_product_row}:$H${last_product_row},"
                            f"$D${first_product_row}:$D${last_product_row},D{row_number}))"
                        )
                    elif column == 25:
                        expected = f"(X{row_number}*H{row_number})"
                    elif column == 28:
                        expected = f"X{row_number}*AA{row_number}"
                    elif column == 29:
                        expected = (
                            f'IF(AA{row_number}>Z{row_number},"ERROR",'
                            f"(X{row_number}-AB{row_number}))"
                        )
                    elif column == 31:
                        expected = (
                            f"IF(A{row_number + 1}=TRUE,MAX(0,"
                            f'1-(AF{row_number}/X{row_number})),"NA")'
                        )
                    else:
                        expected = _derive_official_formula(
                            source_coordinate,
                            target_coordinate,
                            mutation.row_map,
                        )[1:]
                    assert _formula_token_signature("=" + actual) == (
                        _formula_token_signature("=" + expected)
                    )
            assert all(
                target_cells[column].find(f"{{{MAIN}}}f").attrib.get("t") != "shared"
                for column in formula_columns
            )
