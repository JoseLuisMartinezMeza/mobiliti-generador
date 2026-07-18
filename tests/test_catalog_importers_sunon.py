import hashlib
import json
from dataclasses import dataclass, replace
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image

from mobiliti_saas.quote_engine.supplier_catalog import (
    PUBLIC_ITEM_FIELDS,
    load_supplier_catalog_data,
)
from mobiliti_saas.worker.catalog_sync.importers.common import SourceSafetyError
import mobiliti_saas.worker.catalog_sync.importers.sunon as sunon_module
from mobiliti_saas.worker.catalog_sync.importers.sunon import (
    SunonSnapshotBuild,
    build_sunon_snapshot,
    build_sunon_snapshot_with_assets,
)


MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SPEC_PATH = "SPEC GUIDES 2026/SUNON MTY/Spec guide-Sunon MTY-2026.xlsx"
CHAIRS_PATH = "SUNON MTY/2026 updated price-Chairs _ Mexico Stock Reserves \uff084-6 weeks).xlsx"
FAST_PATH = "SUNON MTY/2026 updated price-Fast inventory(1-2 Weeks) 02-09.xlsx"
RAW_PATH = (
    "SUNON MTY/2026 updated price-Raw material preparation \u2605 Mexican inventory list "
    "\uff084-6 weeks).xlsx"
)
MALL_PATH = "SUNON MTY/INVENTORY MALL 1 \uff084-6weeks).xlsx"


@dataclass(frozen=True)
class AdapterFile:
    path: str
    kind: str
    brand: str | None
    sha256: str
    mime_type: str
    local_path: Path | None


def _sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _image(color):
    buffer = BytesIO()
    Image.new("RGB", (12, 8), color).save(buffer, "PNG")
    buffer.seek(0)
    return WorksheetImage(buffer)


def _inventory_sheet(workbook, title, rows=(), *, fast=False):
    sheet = workbook.create_sheet(title)
    header_row = 5 if fast else 3
    headings = (
        ("No.", "ERP CODE", "Item Name", "Photo", "Description", "Dimension", "Color", "Q'ty")
        if fast
        else ("No.", "Item Name", "Photo", "Description", "Dimension", "Color", "Q'ty")
    )
    for column, heading in enumerate(headings, 1):
        sheet.cell(header_row, column, heading)
    price_column = 11 if fast else 10
    sheet.cell(header_row, price_column, "Unit Price in CET Usa Dollar")
    sheet.cell(header_row, price_column + 1, "EXW Monterrey Listing Price MXN (Without VAT)")
    sheet.cell(header_row, price_column + 2, "EXW Monterrey Listing Price MXN (With 16%VAT)")

    for offset, row in enumerate(rows, header_row + 1):
        sheet.cell(offset, 1, row.get("serial", offset - header_row))
        if fast:
            sheet.cell(offset, 2, row.get("erp"))
            model_column, image_column, description_column = 3, 4, 5
            dimensions_column, color_column, quantity_column = 6, 7, 8
        else:
            model_column, image_column, description_column = 2, 3, 4
            dimensions_column, color_column, quantity_column = 5, 6, 7
        sheet.cell(offset, model_column, row["model"])
        sheet.cell(offset, description_column, row.get("description", "Inventory description"))
        sheet.cell(offset, dimensions_column, row.get("dimensions", "600 x 600 mm"))
        sheet.cell(offset, color_column, row.get("color", "Black"))
        sheet.cell(offset, quantity_column, row.get("quantity", 0))
        sheet.cell(offset, price_column, row.get("price"))
        sheet.cell(offset, price_column + 1, f"={sheet.cell(offset, price_column).coordinate}*20")
        sheet.cell(offset, price_column + 2, f"={sheet.cell(offset, price_column + 1).coordinate}*1.16")
        if row.get("image"):
            sheet.add_image(_image(row["image"]), f"{sheet.cell(offset, image_column).column_letter}{offset}")
    return sheet


def _write_spec(path):
    workbook = Workbook()
    cost = workbook.active
    cost.title = "Costo Sunon Mty"
    cost["B8"] = "Cod."
    cost["D9"] = "=100*20"
    spec = workbook.create_sheet("SPEC Sunon Mty")
    for cell, value in {
        "A8": "Imagen.",
        "B8": "Cod.",
        "C8": "Descripcion.",
        "E8": "Color",
        "F8": "Precio Venta",
        "G8": "Moneda.",
    }.items():
        spec[cell] = value
    rows = (
        (9, "UNIQUE10\nUnique Chair", "Spec description unique", "Spec black", "red"),
        (10, "FAST10\nFast Chair", "Spec description fast", "Spec black", None),
        (11, "MULTI10\nMulti Chair", "Spec description multi", "Model color", "blue"),
    )
    for row, code, description, color, image_color in rows:
        spec.cell(row, 2, code)
        spec.cell(row, 3, description)
        spec.cell(row, 5, color)
        spec.cell(row, 6, 999999)
        spec.cell(row, 7, "MXN")
        if image_color:
            spec.add_image(_image(image_color), f"A{row}")
    workbook.active = 0
    workbook.save(path)
    workbook.close()


def _write_chairs(path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _inventory_sheet(
        workbook,
        "Raw material inventory for Chai",
        ({
            "model": "UNIQUE10\nUnique Chair",
            "description": "Inventory description unique",
            "quantity": 2,
            "price": 100,
            "image": "green",
        },),
    )
    workbook.save(path)
    workbook.close()


def _write_fast(path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    shared = {
        "erp": "9001",
        "model": "FAST10\nFast Chair",
        "description": "Inventory description fast",
        "dimensions": "700 x 700 mm",
        "color": "Black frame",
        "price": 200,
    }
    _inventory_sheet(
        workbook,
        "The 1st and 2nd batch inventory",
        ({**shared, "quantity": 3}, {**shared, "quantity": 3}),
        fast=True,
    )
    _inventory_sheet(
        workbook,
        "The 3rd batch",
        ({**shared, "quantity": 4},),
        fast=True,
    )
    _inventory_sheet(
        workbook,
        "The 4th batch",
        ({**shared, "quantity": 5},),
        fast=True,
    )
    workbook.active = 2
    workbook.save(path)
    workbook.close()


def _write_raw(path):
    workbook = Workbook()
    ignored = workbook.active
    ignored.title = "Available Color Option"
    ignored["A1"] = "ignored"
    _inventory_sheet(
        workbook,
        "Mall",
        (
            {
                "model": "MULTI10\nMulti Chair",
                "dimensions": "600 x 600 mm",
                "color": "Black",
                "quantity": 7,
                "price": 300,
            },
            {
                "model": "MULTI10\nMulti Chair",
                "dimensions": "700 x 700 mm",
                "color": "White",
                "quantity": 8,
                "price": 310,
            },
        ),
    )
    for title in ("Mandis", "Universal Table", "M Cabinet"):
        _inventory_sheet(workbook, title)
    total = workbook.create_sheet("Total")
    total["A1"] = "ignored"
    workbook.active = 0
    workbook.save(path)
    workbook.close()


def _write_mall(path):
    workbook = Workbook()
    planning = workbook.active
    planning.title = "Quotation"
    planning["A1"] = "planning only"
    _inventory_sheet(
        workbook,
        "Quotation (2)",
        ({
            "model": "FAST10\nFast Chair",
            "description": "Inventory description fast",
            "dimensions": "700 x 700 mm",
            "color": "Black frame",
            "quantity": 7,
            "price": 200,
        },),
    )
    workbook.active = 0
    workbook.save(path)
    workbook.close()


@pytest.fixture
def source_bundle(tmp_path):
    files = {
        SPEC_PATH: ("spec.xlsx", "spec_guide", _write_spec),
        CHAIRS_PATH: ("chairs.xlsx", "inventory", _write_chairs),
        FAST_PATH: ("fast.xlsx", "inventory", _write_fast),
        RAW_PATH: ("raw.xlsx", "inventory", _write_raw),
        MALL_PATH: ("mall.xlsx", "inventory", _write_mall),
    }
    rows = []
    for logical_path, (name, kind, writer) in files.items():
        local_path = tmp_path / name
        writer(local_path)
        rows.append(AdapterFile(logical_path, kind, None, _sha256(local_path), MIME, local_path))
    return tuple(rows)


def _without_generated_at(snapshot):
    return {key: value for key, value in snapshot.items() if key != "generated_at"}


def _expected_source_hash(bundle):
    source_material = "\n".join(
        f"{row.path}\0{row.kind}\0{row.sha256}"
        for row in sorted(bundle, key=lambda value: (value.path, value.kind, value.sha256))
    )
    catalog_digest = hashlib.sha256(sunon_module.SUNON_CATALOG_PATH.read_bytes()).hexdigest()
    return hashlib.sha256(
        f"{source_material}\nsunon_catalog\0{catalog_digest}".encode()
    ).hexdigest()


def _edit(bundle, logical_path, mutation):
    row = next(candidate for candidate in bundle if candidate.path == logical_path)
    workbook = load_workbook(row.local_path)
    mutation(workbook)
    workbook.save(row.local_path)
    workbook.close()
    updated = replace(row, sha256=_sha256(row.local_path))
    return tuple(updated if candidate is row else candidate for candidate in bundle)


def test_snapshot_contract_source_hash_and_order_are_deterministic(source_bundle):
    first = build_sunon_snapshot(source_bundle)
    second = build_sunon_snapshot(tuple(reversed(source_bundle)))
    loaded = load_supplier_catalog_data(first, expected_supplier="sunon")
    assert set(first) == {"supplier", "source_hash", "generated_at", "items"}
    assert first["source_hash"] == _expected_source_hash(source_bundle)
    assert _without_generated_at(first) == _without_generated_at(second)
    assert len(loaded["items"]) == 4
    assert all(set(item) == set(PUBLIC_ITEM_FIELDS) for item in first["items"])
    assert all(
        item["supplier"] == "sunon"
        and item["brand"] == "Sunon"
        and item["unit"] == "PZA"
        and item["base_price_options"] == []
        and item["add_on_options"] == []
        for item in first["items"]
    )


def test_source_hash_changes_with_bundled_catalog_evidence_without_network(
    source_bundle, monkeypatch, tmp_path,
):
    first = build_sunon_snapshot(source_bundle)
    alternate_catalog = tmp_path / "sunon_catalog.json"
    alternate_catalog.write_text(
        json.dumps({"entries": [{"code": "TEST10", "product_url": "https://www.sunonglobal.com/product/test/", "image_url": ""}]}),
        encoding="utf-8",
    )
    monkeypatch.setattr(sunon_module, "SUNON_CATALOG_PATH", alternate_catalog)

    second = build_sunon_snapshot(tuple(reversed(source_bundle)))
    third = build_sunon_snapshot(source_bundle)

    assert first["source_hash"] != second["source_hash"]
    assert second["source_hash"] == third["source_hash"]


def test_spec_description_and_direct_usd_price_are_authoritative(source_bundle):
    item = next(item for item in build_sunon_snapshot(source_bundle)["items"] if item["sku"] == "UNIQUE10")

    assert item["name"] == "Unique Chair"
    assert item["description"] == "Spec description unique"
    assert item["price_net"] == "100.000000"
    assert item["base_currency"] == "USD"
    assert item["tax_rate"] == "0.160000"
    assert item["stock"] == "2"


def test_model_code_requires_spec_guide_authority(source_bundle):
    changed = _edit(
        source_bundle,
        SPEC_PATH,
        lambda workbook: setattr(workbook["SPEC Sunon Mty"]["B9"], "value", "OTHER10\nOther Chair"),
    )
    item = next(item for item in build_sunon_snapshot(changed)["items"] if item["product_key"] == "unique10")

    assert item["code_status"] == "needs_review"
    assert item["sku"] == ""
    assert any("spec guide" in warning.lower() for warning in item["warnings"])


def test_independent_buckets_sum_and_duplicate_row_is_counted_once(source_bundle):
    item = next(item for item in build_sunon_snapshot(source_bundle)["items"] if item["sku"] == "9001")
    buckets = item["attributes"]["availability_buckets"]

    assert item["stock"] == "19"
    assert item["lead_time"] == "1-2 semanas"
    assert sorted(bucket["quantity"] for bucket in buckets) == ["3", "4", "5", "7"]
    assert all(bucket["source_refs"] for bucket in buckets)
    assert len(json.loads(item["source_reference"])) >= 4


def test_fastest_nonzero_bucket_wins_and_zero_total_keeps_source_lead(source_bundle):
    def zero_fast(workbook):
        for sheet_name in (
            "The 1st and 2nd batch inventory",
            "The 3rd batch",
            "The 4th batch",
        ):
            sheet = workbook[sheet_name]
            for row in range(6, sheet.max_row + 1):
                sheet.cell(row, 8, 0)

    slow_only = _edit(source_bundle, FAST_PATH, zero_fast)
    slow_item = next(item for item in build_sunon_snapshot(slow_only)["items"] if item["sku"] == "9001")
    all_zero = _edit(slow_only, MALL_PATH, lambda workbook: setattr(workbook["Quotation (2)"]["G4"], "value", 0))
    zero_item = next(item for item in build_sunon_snapshot(all_zero)["items"] if item["sku"] == "9001")

    assert slow_item["stock"] == "7"
    assert slow_item["lead_time"] == "4-6 semanas"
    assert zero_item["stock"] == "0"
    assert zero_item["availability_type"] == "stocked"
    assert zero_item["lead_time"] == "1-2 semanas"


def test_model_variants_without_erp_need_review_without_invented_sku(source_bundle):
    first = build_sunon_snapshot(source_bundle)
    second = build_sunon_snapshot(tuple(reversed(source_bundle)))
    variants = [item for item in first["items"] if item["product_key"] == "multi10"]

    assert len(variants) == 2
    assert all(item["code_status"] == "needs_review" and item["sku"] == "" for item in variants)
    assert {item["attributes"]["color"] for item in variants} == {"Black", "White"}
    assert {item["internal_id"] for item in variants} == {
        item["internal_id"] for item in second["items"] if item["product_key"] == "multi10"
    }


def test_configuration_name_prevents_uncoded_variant_merge(source_bundle):
    def add_configurations(workbook):
        sheet = workbook["M Cabinet"]
        for row, name in ((4, "Laminate lower door"), (5, "Laminate upper door")):
            sheet.cell(row, 1, row - 3)
            sheet.cell(row, 2, f"DGKM-2\n{name}")
            sheet.cell(row, 4, "M Cabinet door")
            sheet.cell(row, 5, "900 x 450 mm")
            sheet.cell(row, 6, "White")
            sheet.cell(row, 7, 20)
            sheet.cell(row, 10, 50)

    changed = _edit(source_bundle, RAW_PATH, add_configurations)
    variants = [
        item for item in build_sunon_snapshot(changed)["items"] if item["product_key"] == "dgkm2"
    ]

    assert len(variants) == 2
    assert {item["name"] for item in variants} == {"Laminate lower door", "Laminate upper door"}
    assert {item["stock"] for item in variants} == {"20"}
    assert len({item["internal_id"] for item in variants}) == 2
    assert all(item["code_status"] == "needs_review" and item["sku"] == "" for item in variants)


def test_conflicting_attributes_for_one_erp_split_stock_and_block_code_and_price(source_bundle):
    changed = _edit(source_bundle, FAST_PATH, lambda workbook: setattr(workbook["The 3rd batch"]["G6"], "value", "Blue frame"))
    items = [
        item
        for item in build_sunon_snapshot(changed)["items"]
        if item["attributes"].get("source_erp_code") == "9001"
    ]

    assert len(items) == 2
    assert {item["stock"] for item in items} == {"4", "15"}
    assert {item["attributes"]["color"] for item in items} == {"Black frame", "Blue frame"}
    assert all(item["sku"] == "" and item["code_status"] == "needs_review" for item in items)
    assert all(item["price_net"] == "0.000000" for item in items)
    assert all(any("duplicado" in warning.lower() for warning in item["warnings"]) for item in items)


def test_conflicting_direct_usd_prices_block_price(source_bundle):
    changed = _edit(source_bundle, FAST_PATH, lambda workbook: setattr(workbook["The 3rd batch"]["K6"], "value", 201))
    item = next(item for item in build_sunon_snapshot(changed)["items"] if item["sku"] == "9001")

    assert item["price_net"] == "0.000000"
    assert any("precio" in warning.lower() and "conflict" in warning.lower() for warning in item["warnings"])


@pytest.mark.parametrize(
    "mutation",
    (
        lambda workbook: setattr(workbook["The 4th batch"]["K5"], "value", "Unit Price in CET MXN"),
        lambda workbook: setattr(workbook["The 4th batch"]["K6"], "value", None),
    ),
)
def test_missing_or_contradictory_direct_price_blocks_price(source_bundle, mutation):
    changed = _edit(source_bundle, FAST_PATH, mutation)
    item = next(item for item in build_sunon_snapshot(changed)["items"] if item["sku"] == "9001")

    assert item["price_net"] == "0.000000"
    assert any("precio" in warning.lower() for warning in item["warnings"])


def test_inventory_image_wins_and_spec_model_image_stays_reference(source_bundle):
    first = build_sunon_snapshot(source_bundle)
    second = build_sunon_snapshot(tuple(reversed(source_bundle)))
    unique = next(item for item in first["items"] if item["sku"] == "UNIQUE10")
    multi = next(item for item in first["items"] if item["product_key"] == "multi10")

    assert len(unique["attributes"]["embedded_image_sha256"]) == 64
    assert unique["attributes"]["embedded_image_origin"] == "inventory"
    assert "reference_image_sha256" not in unique["attributes"]
    assert len(multi["attributes"]["reference_image_sha256"]) == 64
    assert any("referencia" in warning.lower() for warning in multi["warnings"])
    assert _without_generated_at(first) == _without_generated_at(second)


def test_inventory_image_becomes_exact_xlsx_approved_asset_and_preserves_official_link(
    source_bundle, monkeypatch,
):
    monkeypatch.setattr(
        sunon_module,
        "find_sunon_catalog_match",
        lambda code: (
            {
                "code": "UNIQUE10",
                "image_url": "https://images.example/unique.png",
                "product_url": "https://products.example/unique",
                "confidence": "exact_code",
            },
            "UNIQUE10",
            "exact_code",
        )
        if str(code).replace("-", "").upper() == "UNIQUE10"
        else (None, None, None),
    )

    build = build_sunon_snapshot_with_assets(source_bundle)
    assert isinstance(build, SunonSnapshotBuild)
    item = next(row for row in build.snapshot["items"] if row["sku"] == "UNIQUE10")
    binding = next(row for row in build.bindings if row.internal_id == item["internal_id"])

    assert item["image_url"] == ""
    assert item["image_kind"] == "official"
    assert item["product_url"] == "https://products.example/unique"
    assert item["attributes"]["image_match"]["status"] == "exact_xlsx"
    assert item["attributes"]["image_match"]["selection_reason"].endswith(
        "Raw material inventory for Chai:C4"
    )
    assert binding.source_references == tuple(
        item["attributes"]["image_match"]["source_references"]
    )
    assert binding.source_references[0]["cell_or_bbox"] == "C4"
    assert item["attributes"]["approved_asset"] == {
        "bucket": "catalog-assets",
        "path": binding.object_name,
        "image_kind": "official",
        "label": "Imagen oficial del XLSX SUNON",
        "approved": True,
    }
    assert binding.object_name == f"{binding.asset_sha256}.png"
    assert build.assets_by_sha256[binding.asset_sha256].sha256 == binding.asset_sha256


def test_multiple_inventory_images_bind_a_deterministic_merged_asset(source_bundle):
    def add_second_inventory_image(workbook):
        sheet = workbook["Raw material inventory for Chai"]
        for column, value in {
            1: 2,
            2: "UNIQUE10\nUnique Chair",
            4: "Inventory description unique",
            5: "600 x 600 mm",
            6: "Black",
            7: 1,
            10: 100,
        }.items():
            sheet.cell(5, column, value)
        sheet.add_image(_image("blue"), "C5")

    changed = _edit(source_bundle, CHAIRS_PATH, add_second_inventory_image)

    build = build_sunon_snapshot_with_assets(changed)
    item = next(row for row in build.snapshot["items"] if row["sku"] == "UNIQUE10")
    binding = next(row for row in build.bindings if row.internal_id == item["internal_id"])

    assert item["attributes"]["image_match"]["status"] == "merged_xlsx"
    assert binding.source_references[0]["cell_or_bbox"] == "C4"
    assert item["attributes"]["image_match"]["selection_reason"].endswith(
        "Raw material inventory for Chai:C4"
    )
    selected = next(
        image
        for image in item["attributes"]["embedded_images"]
        if image["sha256"] == binding.asset_sha256
    )
    assert selected["selected_source_reference"] == binding.source_references[0]
    assert any("varias imagenes embebidas exactas" in warning.lower() for warning in item["warnings"])


def test_only_exact_verified_index_match_becomes_public_image(source_bundle, monkeypatch):
    def match(code):
        normalized = str(code).replace("-", "").upper()
        if normalized == "UNIQUE10":
            return ({
                "code": "UNIQUE10",
                "image_url": "https://images.example/unique.png",
                "product_url": "https://products.example/unique",
                "confidence": "exact_code",
            }, "UNIQUE10", "exact_code")
        if normalized in {"FAST10", "MULTI10"}:
            return ({
                "code": normalized,
                "image_url": f"https://images.example/{normalized.lower()}.png",
                "product_url": f"https://products.example/{normalized.lower()}",
                "confidence": "exact_code",
            }, "FAST" if normalized == "FAST10" else normalized, "base_code" if normalized == "FAST10" else "exact_code")
        return None, None, None

    monkeypatch.setattr(sunon_module, "find_sunon_catalog_match", match)
    snapshot = build_sunon_snapshot(source_bundle)
    unique = next(item for item in snapshot["items"] if item["sku"] == "UNIQUE10")
    others = [item for item in snapshot["items"] if item is not unique]

    assert unique["image_url"] == "https://images.example/unique.png"
    assert unique["image_kind"] == "official"
    assert unique["product_url"] == "https://products.example/unique"
    assert all(item["image_url"] == "" and item["image_kind"] == "placeholder" for item in others)
    assert all("catalog_image_reference" in item["attributes"] for item in others)


def test_exact_model_match_preserves_link_but_not_image_for_distinct_erp_sku(
    source_bundle, monkeypatch,
):
    def match(code):
        if str(code).replace("-", "").upper() == "FAST10":
            return (
                {
                    "code": "FAST10",
                    "image_url": "https://images.example/fast.png",
                    "product_url": "https://www.sunonglobal.com/product/fast-chair/",
                    "confidence": "exact_code",
                },
                "FAST10",
                "exact_code",
            )
        return None, None, None

    monkeypatch.setattr(sunon_module, "find_sunon_catalog_match", match)
    item = next(
        row for row in build_sunon_snapshot(source_bundle)["items"] if row["sku"] == "9001"
    )

    assert item["attributes"]["source_model_code"] == "FAST10"
    assert item["product_url"] == "https://www.sunonglobal.com/product/fast-chair/"
    assert item["attributes"]["product_url_match"] == {
        "status": "exact_code",
        "matched_code": "FAST10",
        "lookup_code": "FAST10",
    }
    assert item["image_url"] == ""
    assert item["image_kind"] == "placeholder"


def test_exact_lookup_with_base_only_confidence_stays_reference(source_bundle, monkeypatch):
    monkeypatch.setattr(
        sunon_module,
        "find_sunon_catalog_match",
        lambda code: (
            {
                "code": "UNIQUE10",
                "image_url": "https://images.example/unique.png",
                "product_url": "https://products.example/unique",
                "confidence": "exact_base_code",
            },
            "UNIQUE10",
            "exact_code",
        )
        if str(code).replace("-", "").upper() == "UNIQUE10"
        else (None, None, None),
    )

    item = next(
        item for item in build_sunon_snapshot(source_bundle)["items"] if item["sku"] == "UNIQUE10"
    )

    assert item["image_url"] == ""
    assert item["image_kind"] == "placeholder"
    assert item["attributes"]["catalog_image_reference"]["match_type"] == "exact_code"


def test_catalog_lookup_preserves_raw_model_code_for_base_reference(source_bundle, monkeypatch):
    changed = _edit(
        source_bundle,
        SPEC_PATH,
        lambda workbook: setattr(
            workbook["SPEC Sunon Mty"]["B9"], "value", "DV72-2\nI-Varna II Conference Table"
        ),
    )
    changed = _edit(
        changed,
        CHAIRS_PATH,
        lambda workbook: setattr(
            workbook["Raw material inventory for Chai"]["B4"],
            "value",
            "DV72-2\nI-Varna II Conference Table",
        ),
    )
    calls = []

    def match(code):
        calls.append(code)
        if code == "DV72-2":
            return (
                {
                    "code": "DV72",
                    "image_url": "https://images.example/dv72.png",
                    "product_url": "https://products.example/dv72",
                    "confidence": "exact_base_code",
                },
                "DV72",
                "base_code",
            )
        return None, None, None

    monkeypatch.setattr(sunon_module, "find_sunon_catalog_match", match)
    item = next(
        item for item in build_sunon_snapshot(changed)["items"] if item["product_key"] == "dv722"
    )

    assert "DV72-2" in calls
    assert item["image_url"] == ""
    assert item["image_kind"] == "placeholder"
    assert item["attributes"]["catalog_image_reference"]["matched_code"] == "DV72"


@pytest.mark.parametrize(
    "mutation",
    (
        lambda rows: rows[:-1],
        lambda rows: rows + (rows[0],),
        lambda rows: tuple(replace(row, path="unexpected.xlsx") if row.path == MALL_PATH else row for row in rows),
        lambda rows: tuple(replace(row, kind="catalog") if row.path == MALL_PATH else row for row in rows),
        lambda rows: tuple(replace(row, mime_type="application/octet-stream") if row.path == MALL_PATH else row for row in rows),
        lambda rows: tuple(replace(row, local_path=None) if row.path == MALL_PATH else row for row in rows),
        lambda rows: tuple(replace(row, path=CHAIRS_PATH) if row.path == MALL_PATH else row for row in rows),
    ),
)
def test_bundle_shape_paths_kinds_mime_and_duplicate_paths_fail_closed(source_bundle, mutation):
    with pytest.raises((ValueError, SourceSafetyError)):
        build_sunon_snapshot(mutation(source_bundle))


def test_hash_mismatch_and_corrupt_source_fail_closed(source_bundle):
    mismatched = tuple(replace(row, sha256="0" * 64) if row.path == SPEC_PATH else row for row in source_bundle)
    with pytest.raises((ValueError, SourceSafetyError)):
        build_sunon_snapshot(mismatched)

    row = next(candidate for candidate in source_bundle if candidate.path == MALL_PATH)
    row.local_path.write_bytes(b"not an xlsx")
    corrupt = tuple(replace(candidate, sha256=_sha256(row.local_path)) if candidate is row else candidate for candidate in source_bundle)
    with pytest.raises(SourceSafetyError):
        build_sunon_snapshot(corrupt)


def test_unsafe_formula_fails_closed(source_bundle):
    unsafe = _edit(
        source_bundle,
        SPEC_PATH,
        lambda workbook: setattr(workbook["SPEC Sunon Mty"]["C9"], "value", "='[outside.xlsx]Sheet1'!A1"),
    )
    with pytest.raises(SourceSafetyError):
        build_sunon_snapshot(unsafe)


def test_unsupported_sheet_fails_closed(source_bundle):
    unsupported = _edit(source_bundle, CHAIRS_PATH, lambda workbook: workbook.create_sheet("Unexpected"))
    with pytest.raises(ValueError, match="SUNON_SHEETS"):
        build_sunon_snapshot(unsupported)


def test_parses_the_exact_bytes_returned_by_validation(source_bundle, monkeypatch):
    original = sunon_module.read_validated_source

    def validate_then_change(path, extension):
        validated, data = original(path, extension)
        path.write_bytes(b"changed after validation")
        return validated, data

    monkeypatch.setattr(sunon_module, "read_validated_source", validate_then_change)

    assert len(build_sunon_snapshot(source_bundle)["items"]) == 4


def test_ignored_real_sources_preserve_source_metrics():
    root = Path(".cache/catalog-sources/sunon")
    sources = (
        (SPEC_PATH, "Spec guide-Sunon MTY-2026.xlsx", "spec_guide", "79a12811c9921447f58cf0ea38272b6c458c6982f17f25d5fdc212b30915de6b"),
        (CHAIRS_PATH, "2026 updated price-Chairs _ Mexico Stock Reserves （4-6 weeks).xlsx", "inventory", "42958cc5d245b10faa4ec6d877196e2be6d98c503b7408958d15b805e3926b44"),
        (FAST_PATH, "2026 updated price-Fast inventory(1-2 Weeks) 02-09.xlsx", "inventory", "96668bc5ef1a7a08b9e4ee2f7d935cf80f3c9c874d0ffbce0fdb5cfaf3de656d"),
        (RAW_PATH, "2026 updated price-Raw material preparation ★ Mexican inventory list （4-6 weeks).xlsx", "inventory", "1338afa15dc843169a75c7504dbd964b20241eb1c3fd7b81d87c98b81b49ff80"),
        (MALL_PATH, "INVENTORY MALL 1 （4-6weeks).xlsx", "inventory", "395233272f9c8f9075cfd6c45f4b49d247d5abcdd2712a4106beac617f3f87bf"),
    )
    if any(not (root / filename).exists() for _, filename, _, _ in sources):
        pytest.skip("ignored Sunon source workbooks are not available")
    bundle = tuple(
        AdapterFile(logical, kind, None, expected_hash, MIME, root / filename)
        for logical, filename, kind, expected_hash in sources
    )
    assert all(_sha256(row.local_path) == row.sha256 for row in bundle)

    snapshot = build_sunon_snapshot(bundle)
    items = load_supplier_catalog_data(snapshot, expected_supplier="sunon")["items"]
    bucket_totals = {}
    for item in items:
        for bucket in item["attributes"]["availability_buckets"]:
            bucket_totals[bucket["lead_time"]] = bucket_totals.get(
                bucket["lead_time"], 0
            ) + int(Decimal(bucket["quantity"]))
    conflicts = [
        item
        for item in items
        if item["attributes"].get("source_erp_code") == "C13912815"
    ]
    cabinet_variants = [item for item in items if item["product_key"] == "dgkm2"]
    metrics = {
        "items": len(items),
        "verified": sum(item["code_status"] == "verified" for item in items),
        "needs_review": sum(item["code_status"] == "needs_review" for item in items),
        "stock": sum(Decimal(item["stock"]) for item in items),
        "embedded_images": sum(
            "embedded_image_sha256" in item["attributes"]
            or "embedded_images" in item["attributes"]
            for item in items
        ),
        "blocked_prices": sum(item["price_net"] == "0.000000" for item in items),
    }
    print("SUNON_REAL_METRICS=" + json.dumps(metrics, sort_keys=True, default=str))

    assert snapshot["source_hash"] == _expected_source_hash(bundle)
    assert metrics == {
        "items": 262,
        "verified": 145,
        "needs_review": 117,
        "stock": Decimal("14085"),
        "embedded_images": 245,
        "blocked_prices": 2,
    }
    assert bucket_totals == {"1-2 semanas": 1566, "4-6 semanas": 12519}
    assert len(conflicts) == 2
    assert {item["stock"] for item in conflicts} == {"1", "2"}
    assert all(
        item["sku"] == ""
        and item["code_status"] == "needs_review"
        and item["price_net"] == "0.000000"
        and item["image_kind"] == "placeholder"
        for item in conflicts
    )
    assert len(cabinet_variants) == 5
    assert {item["name"] for item in cabinet_variants} == {
        "Laminate lower door",
        "Laminate upper door",
    }
    assert {item["stock"] for item in cabinet_variants} == {"20"}
    assert len({item["internal_id"] for item in cabinet_variants}) == 5
