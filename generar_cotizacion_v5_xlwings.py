#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti - v5 Pixel-Perfect (xlwings)
===============================================================
Usa xlwings para automatizar Excel nativamente.
Maxima fidelidad visual porque Excel mismo aplica todo el formato.

Uso:
    python generar_cotizacion_v5_xlwings.py \
        --source "Quotation.xlsx" \
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \
        --output "Cotizacion.xlsx" \
        --cotizacion "100-99999" \
        --proyecto "Proyecto" \
        --cliente "Cliente"
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import xlwings as xw
import clasificador
import insertar_imagenes
import mejorador_imagenes

Q_HEADER_ROW = 7


def limpiar_excel():
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes(source_path):
    """Extrae imagenes del source parseando el drawing XML directamente.
    
    openpyxl img.path no es confiable (reporta image1.png para todas).
    Parseamos el drawing XML + relaciones para obtener el mapeo real.
    """
    import xml.etree.ElementTree as ET
    
    temp_dir = tempfile.mkdtemp(prefix="excel_images_")
    image_map = {}
    
    ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
          'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
          'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
          'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        # 1. Extraer TODAS las imagenes de media
        media_files = [f for f in z.namelist() if 'xl/media/' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
        
        # 2. Leer relaciones del worksheet para encontrar el drawing
        sheet_rels = {}
        for rels_path in z.namelist():
            if 'worksheets/_rels/sheet1.xml.rels' in rels_path:
                rels_content = z.read(rels_path).decode('utf-8')
                rels_root = ET.fromstring(rels_content)
                for rel in rels_root.findall('.//rels:Relationship', ns):
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    sheet_rels[rid] = target
                break
        
        # 3. Encontrar el drawing (podria ser drawing1.xml o vmlDrawing)
        drawing_rels_path = None
        drawing_path = None
        for rid, target in sheet_rels.items():
            if 'drawing' in target.lower() and 'vml' not in target.lower():
                drawing_path = target.replace('../', 'xl/')
                drawing_rels_path = 'xl/drawings/_rels/' + os.path.basename(drawing_path) + '.rels'
                break
        
        if not drawing_path:
            # Fallback: usar openpyxl (aunque no sea perfecto)
            wb = openpyxl.load_workbook(str(source_path), data_only=False)
            ws = wb['Quotation']
            for img in ws._images:
                row = img.anchor._from.row + 1 if hasattr(img.anchor, '_from') else img.anchor.row + 1
                # Intentar obtener el path real del objeto interno
                if hasattr(img, '_path'):
                    path = img._path
                else:
                    path = img.path
                filename = os.path.basename(path)
                filepath = os.path.join(temp_dir, filename)
                if os.path.exists(filepath):
                    image_map[row] = filepath
            wb.close()
            return image_map, temp_dir
        
        # 4. Leer relaciones del drawing (rId -> archivo)
        rId_to_file = {}
        if drawing_rels_path in z.namelist():
            rels_content = z.read(drawing_rels_path).decode('utf-8')
            rels_root = ET.fromstring(rels_content)
            for rel in rels_root.findall('.//rels:Relationship', ns):
                rid = rel.get('Id')
                target = rel.get('Target')
                target_mode = rel.get('TargetMode', 'Internal')
                if target_mode == 'External':
                    continue  # Imagenes externas no estan en el archivo
                # Convertir ../media/imageX.png -> imageX.png
                filename = os.path.basename(target)
                rId_to_file[rid] = filename
        
        # 5. Leer drawing XML para obtener posiciones
        if drawing_path in z.namelist():
            drawing_content = z.read(drawing_path).decode('utf-8')
            drawing_root = ET.fromstring(drawing_content)
            
            for anchor in drawing_root.findall('.//xdr:twoCellAnchor', ns):
                pic = anchor.find('.//xdr:pic', ns)
                if pic is not None:
                    blip = pic.find('.//a:blip', ns)
                    if blip is not None:
                        rId = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        from_row_elem = anchor.find('xdr:from/xdr:row', ns)
                        if from_row_elem is not None:
                            row = int(from_row_elem.text) + 1
                            filename = rId_to_file.get(rId)
                            if filename:
                                filepath = os.path.join(temp_dir, filename)
                                if os.path.exists(filepath):
                                    image_map[row] = filepath
    
    return image_map, temp_dir


def _normalizar_header(texto):
    """Normaliza un header para comparacion: minusculas, sin acentos, sin espacios extra."""
    import unicodedata
    texto = str(texto).lower().strip()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                    if unicodedata.category(c) != 'Mn')
    texto = texto.replace("'", "").replace(".", "").replace(",", "")
    return ' '.join(texto.split())


def _detectar_columnas(ws):
    """Detecta las columnas clave escaneando la fila de encabezados (Q_HEADER_ROW)."""
    keywords = {
        'cantidad': ['qty', 'quantity', 'cantidad'],
        'unit_price': ['unit price', 'unitprice', 'price unit', 'precio unitario'],
        'total_price': ['tot price', 'total price', 'totprice', 'total', 'amount', 'precio total'],
        'list_price': ['list price', 'listprice', 'price list'],
        'descripcion': ['description', 'desc', 'descripcion'],
        'dimension': ['dimension', 'dimensions', 'size', 'medida'],
    }
    column_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=Q_HEADER_ROW, column=col).value
        if not val:
            continue
        norm = _normalizar_header(val)
        for key, terms in keywords.items():
            if key in column_map:
                continue
            for term in terms:
                if term in norm or norm in term:
                    column_map[key] = openpyxl.utils.get_column_letter(col)
                    break
    # Fallbacks: si no se detecta unit_price, intentar con 'price' si no hay list_price
    if 'unit_price' not in column_map and 'list_price' not in column_map:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=Q_HEADER_ROW, column=col).value
            if val and 'price' in _normalizar_header(val):
                column_map['unit_price'] = openpyxl.utils.get_column_letter(col)
                break
    return column_map


def leer_items(source_path):
    wb = openpyxl.load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    column_map = _detectar_columnas(ws)
    
    last_row = ws.max_row
    for row in range(last_row, 0, -1):
        if ws.cell(row=row, column=1).value is not None:
            last_row = row
            break
    
    items = []
    for row in range(Q_HEADER_ROW + 1, last_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        # Categoria: no_val empieza con '-' (puede tener item_name vacio)
        if isinstance(no_val, str) and no_val.startswith('-'):
            nombre = no_val.strip('- ').strip()
            items.append({
                'tipo': 'categoria',
                'row': row,
                'nombre': nombre
            })
            continue
        
        # Saltar filas vacias
        if (item_name is None or item_name == "") and (no_val is None or no_val == ""):
            continue
        
        # Producto: no_val es numerico
        if isinstance(no_val, (int, float)):
            items.append({
                'tipo': 'producto',
                'row': row,
                'nombre': item_name
            })
        # Categoria sin guion pero con texto
        elif no_val is None or no_val == "":
            if item_name:
                items.append({
                    'tipo': 'categoria',
                    'row': row,
                    'nombre': item_name.strip()
                })
    
    wb.close()
    return items, column_map


def crear_template_desprotegido(template_original_path):
    """Crea copia del template sin proteccion de hojas usando zipfile."""
    import uuid
    output = Path(tempfile.gettempdir()) / f"template_desprotegido_{uuid.uuid4().hex[:8]}.xlsx"
    with zipfile.ZipFile(template_original_path, 'r') as zin:
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if 'worksheets/sheet' in item and item.endswith('.xml'):
                    content = data.decode('utf-8')
                    if 'sheetProtection' in content:
                        original = content
                        content = re.sub(r'<sheetProtection[^>]*/>', '', content)
                        if content != original:
                            data = content.encode('utf-8')
                zout.writestr(item, data)
    return output


def agregar_electrificacion_lumbro(ws_mob, items, diccionario, mobiliti_row_map):
    """Agrega filas de electrificacion Lumbro para productos Escritorio con terminos especificos."""
    ELECTRIFICACION_TERMS = ['estacion', 'beach', 'pax']
    LUMBRO_CODIGO_REF = "='SPEC-GUIDE-LUMBRO'!B380"
    LUMBRO_PRECIO_REF = "='SPEC-GUIDE-LUMBRO'!E380"
    fmt_contable = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
    
    filas_agregadas = 0
    for item in items:
        if item['tipo'] != 'producto':
            continue
        
        nombre = str(item['nombre'] or '').lower()
        categoria = clasificador.clasificar_producto(item['nombre'], diccionario)
        
        # Solo productos de categoria Escritorios con terminos de electrificacion
        if categoria != 'Escritorios-WorkStation':
            continue
        
        tiene_term = any(term in nombre for term in ELECTRIFICACION_TERMS)
        if not tiene_term:
            continue
        
        q_row = item['row']
        prod_row = mobiliti_row_map.get(q_row)
        if not prod_row:
            continue
        
        # Buscar fila destino (debajo o siguiente vacia)
        dest_row = prod_row + 1
        max_busqueda = prod_row + 50
        while dest_row < max_busqueda:
            val_d = ws_mob.range(f'D{dest_row}').value
            if val_d is None or val_d == '':
                break
            dest_row += 1
        
        # Escribir datos de electrificacion
        ws_mob.range(f'D{dest_row}').value = LUMBRO_CODIGO_REF
        ws_mob.range(f'F{dest_row}').value = "Lumbro"
        ws_mob.range(f'H{dest_row}').value = f"=Mobiliti!H{prod_row}"
        ws_mob.range(f'J{dest_row}').value = LUMBRO_PRECIO_REF
        ws_mob.range(f'J{dest_row}').api.NumberFormat = fmt_contable
        
        filas_agregadas += 1
    
    if filas_agregadas > 0:
        print(f"      [OK] Electrificacion Lumbro agregada ({filas_agregadas} filas)")


def generar_cotizacion(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI - v5 Pixel-Perfect")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source).resolve()
    template_original = Path(args.template).resolve()
    
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx").resolve()
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: no se pudo borrar archivo previo: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_original.exists():
        print(f"ERROR: Template no encontrado: {template_original}")
        sys.exit(1)
    
    # Cargar diccionario de clasificacion
    dic_path = Path(__file__).parent / "diccionario_categorias.json"
    if not dic_path.exists():
        dic_path = Path.cwd() / "diccionario_categorias.json"
    try:
        diccionario = clasificador.cargar_diccionario(dic_path)
        print(f"      [OK] Diccionario de clasificacion cargado ({len(diccionario.get('categorias', {}))} categorias)")
    except Exception as e:
        print(f"      [ADVERTENCIA] No se pudo cargar diccionario: {e}")
        diccionario = None
    
    temp_dir = None
    app = None
    
    try:
        # PASO 0: Crear template desprotegido
        print("[0/8] Preparando template desprotegido...")
        template_path = crear_template_desprotegido(str(template_original))
        print("      [OK] Template desprotegido creado")
        
        # PASO 1: Extraer imagenes
        print("[1/8] Extrayendo imagenes del Quotation...")
        image_map, temp_dir = extraer_imagenes(str(source_path))
        print(f"      [OK] {len(image_map)} imagenes extraidas")
        
        # PASO 1b: Mejorar imagenes con Pillow
        print("[1b/8] Mejorando imagenes...")
        cache_dir = os.path.join(
            os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
            'Mobiliti', 'Cache', 'Imagenes'
        )
        image_map = mejorador_imagenes.mejorar_image_map(image_map, cache_dir)
        print(f"      [OK] Imagenes mejoradas")
        
        # PASO 2: Leer items
        print("[2/8] Leyendo items del Quotation...")
        items, column_map = leer_items(str(source_path))
        num_cats = sum(1 for i in items if i['tipo'] == 'categoria')
        num_prods = sum(1 for i in items if i['tipo'] == 'producto')
        print(f"      [OK] {len(items)} items ({num_cats} categorias, {num_prods} productos)")
        print(f"      [OK] Columnas detectadas: {column_map}")
        
        # PASO 3: Iniciar Excel via xlwings
        print("[3/8] Iniciando Excel...")
        app = xw.App(visible=True)
        app.display_alerts = False
        app.api.Visible = False  # Ocultar despues de crear (evita bug xlwings)
        print("      [OK] Excel iniciado")
        
        # PASO 4: Abrir workbooks
        print("[4/8] Abriendo archivos...")
        import time
        wb_template = app.books.open(str(template_path))
        time.sleep(1)
        wb_source = app.books.open(str(source_path))
        time.sleep(1)
        
        # Agregar categorias nuevas a la lista Mobiliario en Fletes
        try:
            ws_fletes = wb_template.sheets['Fletes']
            ws_fletes.range('I16').value = 'Bancos'
            ws_fletes.range('I17').value = 'Cocineta'
            ws_fletes.range('I18').value = 'Pizarrones'
            
            # Sincronizar nombre de categoria Escritorios en Fletes (I y M)
            # con el nombre exacto del diccionario
            if diccionario:
                nombre_escritorios = None
                for cat in diccionario.get('categorias', {}).keys():
                    if 'escritorio' in cat.lower():
                        nombre_escritorios = cat
                        break
                if nombre_escritorios:
                    for col in ['I', 'M']:
                        for row in range(6, 19):
                            val = ws_fletes.range(f'{col}{row}').value
                            if val and isinstance(val, str) and 'escritorio' in val.lower():
                                ws_fletes.range(f'{col}{row}').value = nombre_escritorios
            
            print("      [OK] Categorias nuevas agregadas a Mobiliario (Bancos, Cocineta, Pizarrones)")
        except Exception as e:
            print(f"      [ADVERTENCIA] No se pudieron agregar categorias: {e}")
        
        print("      [OK] Archivos abiertos")
        
        # PASO 5: Copiar Quotation del source al template
        print("[5/8] Copiando hoja Quotation...")
        ws_source_q = wb_source.sheets['Quotation']
        ws_source_q.api.Copy(After=wb_template.sheets[-1].api)
        
        q_sheet_name = None
        for sh in wb_template.sheets:
            if 'Quotation' in sh.name:
                q_sheet_name = sh.name
                break
        if not q_sheet_name:
            q_sheet_name = wb_template.sheets[-1].name
        
        ws_quotation = wb_template.sheets[q_sheet_name]
        print(f"      [OK] Quotation copiada como '{q_sheet_name}'")
        
        # PASO 5.5: Agregar introduccion en columna D de la hoja Quotation
        if diccionario:
            for item in items:
                if item['tipo'] == 'producto':
                    q_row = item['row']
                    nombre_b = item['nombre']
                    if nombre_b:
                        # Extraer nombre propio (todo despues del primer salto de linea)
                        nombre_partes = str(nombre_b).split('\n', 1)
                        if len(nombre_partes) > 1:
                            nombre_propio = nombre_partes[1].strip()
                        else:
                            nombre_propio = nombre_partes[0].strip()
                        
                        categoria = clasificador.clasificar_producto(nombre_b, diccionario)
                        introduccion = f"{categoria} modelo {nombre_propio}"
                        texto_actual = ws_quotation.range(f'D{q_row}').value
                        if texto_actual:
                            ws_quotation.range(f'D{q_row}').value = introduccion + str(texto_actual)
                        else:
                            ws_quotation.range(f'D{q_row}').value = introduccion
            print("      [OK] Introducciones agregadas en columna D de Quotation")
        
        # PASO 6: Llenar encabezado
        print("[6/8] Llenando encabezado de Cotizacion...")
        ws_cot = wb_template.sheets['Cotizacion']
        
        # Desmergear celdas del encabezado
        for row in range(3, 13):
            try:
                ws_cot.api.Range(f"A{row}:J{row}").UnMerge()
            except Exception:
                pass
        
        ws_cot.range('B3').value = args.cotizacion
        ws_cot.range('B4').value = datetime.now()
        ws_cot.range('B7').value = args.proyecto
        ws_cot.range('B8').value = args.cliente
        ws_cot.range('B9').value = args.correo
        ws_cot.range('B10').value = args.telefono
        ws_cot.range('B11').value = args.direccion
        ws_cot.range('B12').value = args.razon_social
        
        # Insertar logo si existe
        logo_path = Path(__file__).parent / "LOGO.png"
        if not logo_path.exists():
            logo_path = Path.cwd() / "LOGO.png"
        
        if logo_path.exists():
            # Eliminar imagenes previas en el encabezado (top < 500)
            for pic in ws_cot.pictures:
                if pic.top < 500:
                    try:
                        pic.delete()
                    except Exception:
                        pass
            
            # Insertar logo en la posicion del template
            # Posicion original del template: top=2.9, left=1398.4, width=318.6, height=317.1
            pic = ws_cot.pictures.add(str(logo_path), left=1398, top=3, width=319, height=317)
            print("      [OK] Logo insertado")
        
        print("      [OK] Encabezado llenado")
        
        # PASO 7: Generar Mobiliti
        print("[7/8] Generando Mobiliti...")
        ws_mob = wb_template.sheets['Mobiliti']
        
        # Desproteger Mobiliti temporalmente para preparar secciones
        try:
            ws_mob.api.Unprotect("M0b1l1t$")
        except Exception:
            pass
        
        # Estructura de secciones del template (categoria + 32 productos + subtotales)
        # Seccion 1: cat=13, prods=14-45, subtotales=47
        # Seccion 2: cat=48, prods=49-80, subtotales=82
        # ...
        # Seccion 10: cat=328, prods=329-360, subtotales=362
        # Seccion 11: cat=363, prods=364-395, subtotales=396
        # Seccion 12: cat=397, prods=398-429, subtotales=430
        # Seccion 13: cat=431, prods=432-463, subtotales=464
        section_cats = [13, 48, 83, 118, 153, 188, 223, 258, 293, 328, 363, 397, 431]
        section_prod_starts = [14, 49, 84, 119, 154, 189, 224, 259, 294, 329, 364, 398, 432]
        section_subtotals = [47, 82, 117, 152, 187, 222, 257, 292, 327, 362, 396, 430, 464]
        max_prod_per_section = 32
        
        # Preparar secciones 11-13 si no tienen fórmulas
        # Verificar si la fila 364 tiene fórmulas; si no, copiar seccion 1
        print("      Preparando secciones adicionales...")
        try:
            tiene_formulas = ws_mob.range('H364').formula and ws_mob.range('H364').formula.startswith('=')
        except Exception:
            tiene_formulas = False
        
        if not tiene_formulas:
            print("        Copiando seccion 1 a secciones 11-13...")
            # Copiar seccion 1 completa (A13:AV47) a cada nueva seccion
            seccion_origen = ws_mob.range('A13:AV47')
            for sec_num, cat_row in enumerate(section_cats[10:], start=11):
                subtotal_row = section_subtotals[sec_num - 1]
                destino = ws_mob.range(f'A{cat_row}')
                print(f"          Seccion {sec_num} (filas {cat_row}-{subtotal_row})...")
                seccion_origen.api.Copy(destino.api)
            
            # Corregir referencias C en filas de producto para que apunten a C13
            print("        Corrigiendo referencias de moneda...")
            for sec_idx in range(10, len(section_cats)):
                cat_row = section_cats[sec_idx]
                prod_start = section_prod_starts[sec_idx]
                # Limpiar nombre de categoria copiado
                try:
                    ws_mob.range(f'D{cat_row}').value = f'Sección {sec_idx + 1} - NOMBRE'
                except Exception:
                    pass
                for r in range(prod_start, prod_start + max_prod_per_section):
                    try:
                        cell_c = ws_mob.range(f'C{r}')
                        if cell_c.formula:
                            new_c = cell_c.formula
                            # Reemplazar cualquier referencia a la fila de categoria copiada con $C$13
                            for ref_cat in [48, 83, 118, 153, 188, 223, 258, 293, 328, 363, 397, 431]:
                                new_c = new_c.replace(f'$C${ref_cat}', '$C$13')
                                new_c = new_c.replace(f'C${ref_cat}', 'C$13')
                            # Tambien corregir referencias relativas como C363 (sin $)
                            new_c = new_c.replace(f'=C{cat_row}', '=C$13')
                            if new_c != cell_c.formula:
                                cell_c.formula = new_c
                    except Exception:
                        pass
        else:
            print("        Secciones 11-13 ya tienen formulas (template ya preparado)")
        
        mobiliti_row_map = {}
        section_idx = 0
        prod_in_section = 0
        
        # Limpiar SOLO las celdas que modificamos (D, F, H, J, K, P) en todas las filas de producto
        print("      Limpiando celdas de referencia...")
        for sec_idx in range(len(section_cats)):
            cat_row = section_cats[sec_idx]
            prod_start = section_prod_starts[sec_idx]
            for r in range(prod_start, prod_start + max_prod_per_section):
                try:
                    ws_mob.range(f'D{r}').value = None
                    ws_mob.range(f'F{r}').value = None
                    ws_mob.range(f'H{r}').value = None
                    ws_mob.range(f'J{r}').value = None
                    ws_mob.range(f'K{r}').value = None
                    ws_mob.range(f'P{r}').value = None
                except Exception:
                    pass
        
        for item in items:
            if item['tipo'] == 'categoria':
                # Si ya habia productos en la seccion anterior, avanzar a la siguiente
                if prod_in_section > 0:
                    section_idx += 1
                    prod_in_section = 0
                
                # Si nos quedamos sin secciones, salir
                if section_idx >= len(section_cats):
                    print(f"      [ADVERTENCIA] Mas categorias que secciones disponibles")
                    break
                
                cat_row = section_cats[section_idx]
                # Limpiar D en categoria (texto del template)
                ws_mob.range(f'D{cat_row}').value = None
                
            else:
                q_row = item['row']
                
                # Si la seccion actual esta llena, avanzar a la siguiente
                if prod_in_section >= max_prod_per_section:
                    section_idx += 1
                    prod_in_section = 0
                
                # Si nos quedamos sin secciones, salir
                if section_idx >= len(section_cats):
                    print(f"      [ADVERTENCIA] Mas productos que espacio disponible")
                    break
                
                cat_row = section_cats[section_idx]
                prod_row = section_prod_starts[section_idx] + prod_in_section
                
                # Ajustar referencia de moneda en C (=C13 → =C${cat_row} o =$C$13 → =$C${cat_row})
                cell_c = ws_mob.range(f'C{prod_row}')
                if cell_c.formula:
                    new_c_formula = cell_c.formula
                    new_c_formula = new_c_formula.replace('$C$13', f'$C${cat_row}')
                    new_c_formula = new_c_formula.replace('C13', f'C${cat_row}')
                    if new_c_formula != cell_c.formula:
                        cell_c.formula = new_c_formula
                
                # Clasificar producto en columna E
                if diccionario:
                    nombre_producto = item['nombre']
                    categoria = clasificador.clasificar_producto(nombre_producto, diccionario)
                    ws_mob.range(f'E{prod_row}').value = categoria
                
                # Poner referencias a Quotation en D, F, H, J, K, P
                col_cant = column_map.get('cantidad', 'G')
                col_unit = column_map.get('unit_price', 'K')
                col_tot = column_map.get('total_price', 'H')
                
                ws_mob.range(f'D{prod_row}').value = f"={q_sheet_name}!B{q_row}"   # Código de producto
                ws_mob.range(f'F{prod_row}').value = "Sunon Inc"                    # Proveedor
                ws_mob.range(f'H{prod_row}').value = f"={q_sheet_name}!{col_cant}{q_row}"   # Cantidad
                ws_mob.range(f'J{prod_row}').value = f"={q_sheet_name}!{col_unit}{q_row}"   # Precio unitario
                ws_mob.range(f'K{prod_row}').value = f"={q_sheet_name}!{col_tot}{q_row}"   # Precio total
                ws_mob.range(f'P{prod_row}').value = "Centro"                      # Centro
                
                mobiliti_row_map[q_row] = prod_row
                prod_in_section += 1
        
        # Restaurar C en todas las filas de producto para que apunte a C13 (moneda)
        # El template original tiene C apuntando a C13; Excel a veces lo cambia al guardar
        for sec_idx in range(section_idx + 1):
            prod_start = section_prod_starts[sec_idx]
            for r in range(prod_start, prod_start + max_prod_per_section + 1):
                cell_c = ws_mob.range(f'C{r}')
                if cell_c.formula:
                    # Restaurar referencias a C13
                    new_c = cell_c.formula
                    new_c = new_c.replace('$C$48', '$C$13').replace('C$48', 'C$13')
                    new_c = new_c.replace('$C$83', '$C$13').replace('C$83', 'C$13')
                    new_c = new_c.replace('$C$118', '$C$13').replace('C$118', 'C$13')
                    new_c = new_c.replace('$C$153', '$C$13').replace('C$153', 'C$13')
                    new_c = new_c.replace('$C$188', '$C$13').replace('C$188', 'C$13')
                    new_c = new_c.replace('$C$223', '$C$13').replace('C$223', 'C$13')
                    new_c = new_c.replace('$C$258', '$C$13').replace('C$258', 'C$13')
                    new_c = new_c.replace('$C$293', '$C$13').replace('C$293', 'C$13')
                    new_c = new_c.replace('$C$363', '$C$13').replace('C$363', 'C$13')
                    new_c = new_c.replace('$C$397', '$C$13').replace('C$397', 'C$13')
                    new_c = new_c.replace('$C$431', '$C$13').replace('C$431', 'C$13')
                    if new_c != cell_c.formula:
                        cell_c.formula = new_c
        
        print(f"      [OK] Mobiliti generada ({section_idx + 1} secciones usadas)")
        
        # PASO 7.5: Agregar electrificacion Lumbro para Escritorios elegibles
        if diccionario:
            agregar_electrificacion_lumbro(ws_mob, items, diccionario, mobiliti_row_map)
        
        # PASO 8: Generar Cotizacion
        print("[8/8] Generando Cotizacion...")
        
        # Encontrar fila de terminos
        terminos_start_row = None
        for row in range(16, 100):
            val = ws_cot.range(f'A{row}').value
            if val and isinstance(val, str) and 'CONDICIONES' in val:
                terminos_start_row = row
                break
        if not terminos_start_row:
            terminos_start_row = 32
        
        # Encontrar la ultima fila usada del template (bloque de terminos completo)
        last_template_row = ws_cot.api.UsedRange.Rows(ws_cot.api.UsedRange.Rows.Count).Row
        if last_template_row < terminos_start_row:
            last_template_row = terminos_start_row + 50
        
        # 1. GUARDAR bloque de terminos y totales en hoja temporal
        terminos_row_count = last_template_row - terminos_start_row + 1
        print(f"      Guardando bloque de terminos ({terminos_row_count} filas) en hoja temporal...")
        
        temp_sheet = wb_template.sheets.add(name="__TempTerminos")
        
        # Guardar bloque de terminos
        ws_cot.api.Range(f"A{terminos_start_row}:J{last_template_row}").Copy(
            Destination=temp_sheet.api.Range("A1")
        )
        
        # Guardar filas de totales (21-25) del template para copiar formato nativo
        total_template_start = 21
        total_template_end = 25
        ws_cot.api.Range(f"A{total_template_start}:J{total_template_end}").Copy(
            Destination=temp_sheet.api.Range(f"A{terminos_row_count + 2}")
        )
        print(f"      Tambien guardadas filas de totales ({total_template_start}-{total_template_end})")
        
        # 2. Desmergear SOLO las filas de ejemplo (16 hasta terminos_start_row - 1)
        for row in range(16, terminos_start_row):
            try:
                ws_cot.api.Range(f"A{row}:J{row}").UnMerge()
            except Exception:
                pass
        
        # 3. Limpiar SOLO las filas de ejemplo
        for row in range(16, terminos_start_row):
            ws_cot.api.Range(f"A{row}:J{row}").ClearContents()
        
        # 4. Insertar filas si es necesario (antes de los terminos)
        filas_ejemplo = terminos_start_row - 16
        if len(items) > filas_ejemplo:
            filas_extra = len(items) - filas_ejemplo
            ws_cot.api.Rows(f"{terminos_start_row}:{terminos_start_row + filas_extra - 1}").Insert(
                Shift=-4121,
                CopyOrigin=0
            )
            terminos_start_row += filas_extra
        
        # 5. ELIMINAR terminos originales que quedaron empujados
        # (ya los tenemos guardados en la hoja temporal, asi que eliminamos los duplicados)
        print("      Eliminando terminos originales empujados...")
        ws_cot.api.Rows(f"{terminos_start_row}:{terminos_start_row + terminos_row_count - 1}").Delete(
            Shift=-4162  # xlUp
        )
        print(f"      [OK] Eliminadas {terminos_row_count} filas de terminos duplicados")
        
        # Escribir items
        current_row = 16
        first_data_row = None
        last_data_row = None
        descuento_row = None
        
        # Filas template de Cotizacion
        cot_cat_row = 16   # Fila con formato de categoria
        cot_prod_row = 18  # Fila con formato de producto
        cot_cat_height = float(ws_cot.api.Rows(cot_cat_row).RowHeight)  # Altura de categoria
        cot_prod_height = float(ws_cot.api.Rows(cot_prod_row).RowHeight)  # Altura de producto
        
        for item in items:
            if item['tipo'] == 'categoria':
                # 1. Copiar formato de categoria desde fila template (rango A:J)
                src = ws_cot.api.Range(f"A{cot_cat_row}:J{cot_cat_row}")
                dst = ws_cot.api.Range(f"A{current_row}:J{current_row}")
                src.Copy()
                dst.PasteSpecial(-4104)  # xlPasteAll
                app.api.CutCopyMode = False
                
                # 2. Escribir valores (sobreescriben los del template)
                ws_cot.range(f'A{current_row}').value = f"={q_sheet_name}!A{item['row']}"
                
                # 3. Aplicar merge manualmente (PasteSpecial no siempre copia merge)
                ws_cot.api.Range(f"A{current_row}:J{current_row}").Merge()
                
                # 4. Ajustar altura de fila (PasteSpecial con rango no copia altura)
                ws_cot.api.Rows(current_row).RowHeight = cot_cat_height
                
                current_row += 1
            else:
                if first_data_row is None:
                    first_data_row = current_row
                    descuento_row = current_row  # El descuento va en el primer producto
                
                q_row = item['row']
                mob_row = mobiliti_row_map.get(q_row)
                
                # 1. Copiar formato de producto desde fila template (rango A:J)
                src = ws_cot.api.Range(f"A{cot_prod_row}:J{cot_prod_row}")
                dst = ws_cot.api.Range(f"A{current_row}:J{current_row}")
                src.Copy()
                dst.PasteSpecial(-4104)  # xlPasteAll
                app.api.CutCopyMode = False
                
                # 2. Escribir valores (sobreescriben los del template)
                col_desc = column_map.get('descripcion', 'D')
                col_dim = column_map.get('dimension', 'E')
                col_cant = column_map.get('cantidad', 'G')
                col_list = column_map.get('list_price', column_map.get('unit_price', 'J'))
                
                ws_cot.range(f'A{current_row}').value = f"={q_sheet_name}!B{q_row}"
                ws_cot.range(f'C{current_row}').value = f"={q_sheet_name}!{col_desc}{q_row}"
                ws_cot.range(f'D{current_row}').value = f"={q_sheet_name}!{col_dim}{q_row}"
                
                if mob_row:
                    ws_cot.range(f'E{current_row}').value = f"=Mobiliti!H{mob_row}"
                    ws_cot.range(f'F{current_row}').value = f"=Mobiliti!W{mob_row}"
                else:
                    ws_cot.range(f'E{current_row}').value = f"={q_sheet_name}!{col_cant}{q_row}"
                    ws_cot.range(f'F{current_row}').value = f"={q_sheet_name}!{col_list}{q_row}"
                
                # Calcular factor de descuento (ej: 30% -> 0.7)
                try:
                    descuento_pct = float(getattr(args, 'descuento', '30').replace('%', '').strip())
                except (ValueError, AttributeError):
                    descuento_pct = 30.0
                descuento_factor = max(0.0, min(1.0, 1.0 - (descuento_pct / 100.0)))
                
                if current_row == descuento_row:
                    ws_cot.range(f'G{current_row}').value = descuento_factor
                else:
                    ws_cot.range(f'G{current_row}').value = f"=G${descuento_row}"
                ws_cot.range(f'H{current_row}').value = f"=F{current_row}*G{current_row}"
                ws_cot.range(f'I{current_row}').value = f"=F{current_row}-H{current_row}"
                ws_cot.range(f'J{current_row}').value = f"=I{current_row}*E{current_row}"
                
                # Aplicar formato contable ($) a columnas monetarias
                fmt_contable = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
                for col in ['F', 'H', 'I', 'J']:
                    ws_cot.range(f'{col}{current_row}').api.NumberFormat = fmt_contable
                
                # 3. Ajustar altura de fila (PasteSpecial con rango no copia altura)
                # Aumentar altura para que las imagenes se vean mas grandes (max 409 puntos de Excel)
                ws_cot.api.Rows(current_row).RowHeight = min(cot_prod_height * 1.3, 409)
                
                current_row += 1
        
        last_data_row = current_row - 1
        
        # Descuento base ya fue establecido en el primer producto (G{descuento_row})
        
        # Insertar imagenes usando modulo separado
        img_count = insertar_imagenes.insertar_imagenes_cotizacion(ws_cot, items, image_map)
        
        # Totales - copiar formato nativo del template (filas 21-25)
        print("      Agregando totales con formato nativo del template...")
        
        # Los totales van justo despues del ultimo item (sin fila vacia)
        row_sub = last_data_row + 1
        
        # Filas template de totales guardadas en hoja temporal
        tpl_subtotal_row = terminos_row_count + 2      # template fila 21
        tpl_flete_row = terminos_row_count + 3         # template fila 22
        tpl_subtotal2_row = terminos_row_count + 4     # template fila 23
        tpl_iva_row = terminos_row_count + 5           # template fila 24
        tpl_total_row = terminos_row_count + 6         # template fila 25
        
        # Funcion para copiar formato de fila template a fila destino
        def copiar_formato_total(temp_row, dest_row):
            temp_sheet.api.Range(f"A{temp_row}:J{temp_row}").Copy()
            ws_cot.api.Range(f"A{dest_row}:J{dest_row}").PasteSpecial(-4104)
            app.api.CutCopyMode = False
        
        # Fila 1: SUBTOTAL
        copiar_formato_total(tpl_subtotal_row, row_sub)
        ws_cot.range(f'D{row_sub}').value = "SUBTOTAL:"
        ws_cot.range(f'H{row_sub}').value = f"=SUM(J{first_data_row}:J{last_data_row})"
        
        # Fila 2: COSTO DE FLETE
        row_flete = row_sub + 1
        copiar_formato_total(tpl_flete_row, row_flete)
        ws_cot.range(f'D{row_flete}').value = "COSTO DE FLETE:"
        ws_cot.range(f'H{row_flete}').value = f"=H{row_sub}*12%"
        
        # Fila 3: SUBTOTAL (con flete)
        row_sub2 = row_flete + 1
        copiar_formato_total(tpl_subtotal2_row, row_sub2)
        ws_cot.range(f'D{row_sub2}').value = "SUBTOTAL:"
        ws_cot.range(f'H{row_sub2}').value = f"=H{row_sub}+H{row_flete}"
        
        # Fila 4: IVA
        row_iva = row_sub2 + 1
        copiar_formato_total(tpl_iva_row, row_iva)
        ws_cot.range(f'D{row_iva}').value = "IVA:"
        ws_cot.range(f'H{row_iva}').value = f"=H{row_sub2}*16%"
        
        # Fila 5: TOTAL
        row_total = row_iva + 1
        copiar_formato_total(tpl_total_row, row_total)
        ws_cot.range(f'D{row_total}').value = "TOTAL:"
        ws_cot.range(f'H{row_total}').value = f"=H{row_sub2}+H{row_iva}"
        
        print(f"      [OK] Totales en filas {row_sub}-{row_total}")
        
        # INSERTAR 2 filas vacias entre totales y terminos
        print("      Insertando 2 filas vacias entre totales y terminos...")
        ws_cot.api.Rows(f"{row_total + 1}:{row_total + 2}").Insert(
            Shift=-4121,
            CopyOrigin=0  # Copia formato de arriba, luego limpiamos
        )
        # Limpiar formato de las 2 filas insertadas (deben quedar blancas)
        for r in range(row_total + 1, row_total + 3):
            ws_cot.api.Range(f"A{r}:J{r}").UnMerge()
            ws_cot.api.Range(f"A{r}:J{r}").Interior.Color = 0xFFFFFF
            ws_cot.api.Range(f"A{r}:J{r}").Font.Color = 0x000000
            ws_cot.api.Range(f"A{r}:J{r}").Font.Bold = False
            ws_cot.api.Range(f"A{r}:J{r}").Font.Name = "Roboto"
            ws_cot.api.Range(f"A{r}:J{r}").Font.Size = 9
        new_terminos_start = row_total + 3
        
        # RESTAURAR bloque de terminos en su nueva posicion
        print("      Restaurando bloque de terminos...")
        
        # Limpiar cualquier contenido existente desde la nueva posicion hacia abajo
        ws_cot.api.Range(f"A{new_terminos_start}:J{new_terminos_start + terminos_row_count + 10}").ClearContents()
        
        # Copiar bloque de terminos desde hoja temporal
        temp_sheet.api.Range(f"A1:J{terminos_row_count}").Copy(
            Destination=ws_cot.api.Range(f"A{new_terminos_start}")
        )
        
        # Eliminar hoja temporal
        wb_template.sheets["__TempTerminos"].delete()
        
        print(f"      [OK] Terminos restaurados en fila {new_terminos_start}")
        
        # Configurar area de impresion para incluir TODO desde fila 1 hasta el final
        print("      Configurando area de impresion...")
        
        # Cotizacion: desde A1 hasta la ultima fila usada
        last_row_cot = ws_cot.api.UsedRange.Rows.Count + ws_cot.api.UsedRange.Row - 1
        ws_cot.api.PageSetup.PrintArea = f"$A$1:$J${last_row_cot}"
        print(f"      [OK] PrintArea Cotizacion: A1:J{last_row_cot}")
        
        # Mobiliti: desde A1 hasta la ultima fila usada
        last_row_mob = ws_mob.api.UsedRange.Rows.Count + ws_mob.api.UsedRange.Row - 1
        ws_mob.api.PageSetup.PrintArea = f"$A$1:$J${last_row_mob}"
        print(f"      [OK] PrintArea Mobiliti: A1:J{last_row_mob}")
        
        # Volver a proteger la hoja Mobiliti (solo esta hoja, no Cotizacion)
        print("      Protegiendo hoja Mobiliti...")
        try:
            ws_mob.api.Protect(Password="M0b1l1t$", UserInterfaceOnly=False)
            print("      [OK] Hoja Mobiliti protegida")
        except Exception as e:
            print(f"      [ADVERTENCIA] No se pudo proteger Mobiliti: {e}")
        
        # Guardar
        print("      Guardando...")
        wb_template.save(str(output_path))
        
        print()
        print("=" * 60)
        print("[OK] COTIZACION GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print()
        
        # Cerrar workbooks (ignorar errores OLE no criticos)
        try:
            wb_template.close()
        except Exception as e:
            print(f"      [ADVERTENCIA] Al cerrar template: {e}")
        try:
            wb_source.close()
        except Exception as e:
            print(f"      [ADVERTENCIA] Al cerrar source: {e}")
        try:
            app.quit()
        except Exception as e:
            print(f"      [ADVERTENCIA] Al cerrar Excel: {e}")
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass
        if app:
            try:
                app.quit()
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti con Excel nativo (xlwings)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_v5_xlwings.py \\
    --source "Quotation.xlsx" \\
    --template "Formato Cotizacion 2026 GDL (1).xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotización 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    parser.add_argument('--descuento', default='30', help='Porcentaje de descuento (ej: 30 = 30%% descuento)')
    
    args = parser.parse_args()
    generar_cotizacion(args)


if __name__ == '__main__':
    main()
