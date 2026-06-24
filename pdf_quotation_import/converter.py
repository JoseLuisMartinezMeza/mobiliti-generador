from __future__ import annotations

import argparse
from collections import Counter
from copy import copy
from dataclasses import dataclass, field
import re
import tempfile
from pathlib import Path
from typing import Iterable

try:
    import pymupdf as fitz
except ImportError:  # pragma: no cover - older PyMuPDF exposes only fitz
    import fitz  # type: ignore[no-redef]

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_ROW = 7
HEADERS = [
    "No.",
    "Item Name",
    "Photo",
    "Description",
    "Dimension",
    "Color",
    "Q'ty",
    "Vol.",
    "Tot.Vol.",
    "Unit Price",
    " Tot.Price",
    "Remark",
]

MONEY_RE = re.compile(r"^\$?[0-9][0-9,]*\.\d{2}$")
INTEGER_RE = re.compile(r"^\d+$")
CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9./+-]{4,}$")
DIMENSION_RE = re.compile(
    r"(?i)(?:[ΦФØø]\s*)?\d+(?:[.,]\d+)?(?:\s*\*\s*\d+(?:[.,]\d+)?){1,3}\s*(?:cm|mm|m)?"
)
SPEC_MARKERS = (
    "FRAME:",
    "FINISH:",
    "FABRIC:",
    "ROPE:",
    "SLING:",
    "MATERIAL:",
    "TOP:",
    "REMARK:",
    "COLOR:",
)
SUMMARY_MARKERS = (
    "TOTAL QUANTITY",
    "TOTAL VALUE",
    "DISCOUNT",
    "FINAL VALUE",
    "DEPOIST",
    "DEPOSIT",
    "BANK INFORMATION",
    "BENEFICIARY",
    "SWIFT CODE",
    "A/C NO",
)


@dataclass(frozen=True)
class PdfWord:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class PdfLine:
    page_index: int
    page_height: float
    y: float
    words: list[PdfWord]

    @property
    def global_y(self) -> float:
        return self.page_index * self.page_height + self.y

    @property
    def text(self) -> str:
        return " ".join(word.text for word in sorted(self.words, key=lambda word: word.x0))

    def text_between(self, x_min: float, x_max: float) -> str:
        words = [word.text for word in sorted(self.words, key=lambda word: word.x0) if x_min <= word.x0 < x_max]
        return clean_text(" ".join(words))

    def texts(self) -> list[str]:
        return [word.text for word in sorted(self.words, key=lambda word: word.x0)]


@dataclass
class ProductStart:
    line: PdfLine
    code: str
    raw_field: str
    quantity: int
    unit_price: float
    total_price: float


@dataclass
class ParsedProduct:
    number: int
    code: str
    category: str
    quantity: int
    unit_price: float
    total_price: float
    detail_lines: list[str]
    image_path: Path | None = None
    name: str = ""
    description: str = ""
    dimension: str = ""
    color: str = ""


@dataclass
class PdfImageBlock:
    global_y: float
    bbox_area: float
    data_size: int
    path: Path


@dataclass(frozen=True)
class ConversionResult:
    output_path: Path
    product_count: int
    image_count: int
    categories: tuple[str, ...]


@dataclass
class StyleSource:
    reference_path: Path | None = None
    header_rows: dict[int, list] = field(default_factory=dict)
    row_heights: dict[int, float | None] = field(default_factory=dict)
    column_widths: dict[int, float | None] = field(default_factory=dict)
    category_row: list | None = None
    product_row: list | None = None
    total_row: list | None = None


def convert_pdf_to_quotation(
    source_pdf: str | Path,
    output_xlsx: str | Path,
    reference_xlsx: str | Path | None = None,
) -> ConversionResult:
    source = Path(source_pdf).expanduser().resolve()
    output = Path(output_xlsx).expanduser().resolve()
    reference = Path(reference_xlsx).expanduser().resolve() if reference_xlsx else None

    if not source.exists():
        raise FileNotFoundError(f"No existe el PDF: {source}")
    if reference and not reference.exists():
        raise FileNotFoundError(f"No existe el XLSX de referencia: {reference}")

    output.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="pdf_quote_import_") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        with fitz.open(source) as doc:
            lines = extract_lines(doc)
            starts = extract_product_starts(lines)
            if not starts:
                raise ValueError("No se detectaron productos en el PDF")

            category_labels = detect_category_labels(starts)
            products = build_products(lines, starts, category_labels)
            attach_start_positions(products, starts)
            image_blocks = extract_image_blocks(doc, temp_dir)
            assign_images(products, image_blocks)
            metadata = extract_metadata(lines, source.stem, len(doc))

        style_source = load_style_source(reference)
        write_workbook(products, metadata, output, style_source)

    categories = tuple(dict.fromkeys(product.category for product in products))
    return ConversionResult(
        output_path=output,
        product_count=len(products),
        image_count=sum(1 for product in products if product.image_path),
        categories=categories,
    )


def extract_lines(doc) -> list[PdfLine]:
    lines: list[PdfLine] = []
    for page_index, page in enumerate(doc):
        page_lines: list[PdfLine] = []
        for raw_word in page.get_text("words", sort=True):
            x0, y0, x1, y1, text = raw_word[:5]
            word = PdfWord(float(x0), float(y0), float(x1), float(y1), str(text))
            target = next((line for line in page_lines if abs(line.y - word.y0) <= 2.4), None)
            if target:
                target.words.append(word)
                target.y = (target.y + word.y0) / 2
            else:
                page_lines.append(PdfLine(page_index, float(page.rect.height), word.y0, [word]))
        lines.extend(sorted(page_lines, key=lambda line: line.y))
    return sorted(lines, key=lambda line: line.global_y)


def extract_product_starts(lines: Iterable[PdfLine]) -> list[ProductStart]:
    starts: list[ProductStart] = []
    for line in lines:
        words = sorted(line.words, key=lambda word: word.x0)
        texts = [word.text for word in words]
        money_words = [word.text for word in words if MONEY_RE.match(word.text)]
        qty_words = [
            word.text
            for word in words
            if INTEGER_RE.match(word.text) and 320 <= word.x0 <= 375
        ]
        left_code = " ".join(word.text for word in words if word.x0 < 115 and CODE_RE.match(word.text))
        has_pcs = any(text.upper() == "PCS" for text in texts)
        if len(money_words) < 2 or not qty_words or not has_pcs or not left_code:
            continue

        starts.append(
            ProductStart(
                line=line,
                code=left_code,
                raw_field=project_or_name_field(line),
                quantity=int(qty_words[-1]),
                unit_price=parse_money(money_words[0]),
                total_price=parse_money(money_words[-1]),
            )
        )
    return starts


def detect_category_labels(starts: Iterable[ProductStart]) -> set[str]:
    counts = Counter(start.raw_field for start in starts if start.raw_field)
    return {label for label, count in counts.items() if count >= 2}


def build_products(
    lines: list[PdfLine],
    starts: list[ProductStart],
    category_labels: set[str],
) -> list[ParsedProduct]:
    products: list[ParsedProduct] = []
    current_category = "PRODUCTOS"

    for index, start in enumerate(starts):
        next_y = starts[index + 1].line.global_y if index + 1 < len(starts) else float("inf")
        if start.raw_field in category_labels:
            current_category = start.raw_field
        elif current_category == "PRODUCTOS" and start.raw_field:
            current_category = start.raw_field

        code = start.code
        details: list[str] = []
        if start.raw_field and start.raw_field not in category_labels:
            details.append(start.raw_field)

        for line in lines:
            if not (start.line.global_y + 1 < line.global_y < next_y - 1):
                continue
            normalized = normalize_marker_text(line.text)
            if is_repeated_page_header(normalized):
                continue
            if any(marker in normalized for marker in SUMMARY_MARKERS):
                break
            if is_single_buyer_number(line):
                continue
            if looks_like_product_start(line):
                continue

            left_text = line.text_between(0, 115)
            right_text = line.text_between(115, 570)
            if left_text and CODE_RE.match(left_text.replace(" ", "")) and not contains_money_or_uom(line):
                code = clean_text(f"{code} {left_text}")
                if right_text:
                    details.append(right_text)
                continue
            if right_text:
                details.append(right_text)

        product = ParsedProduct(
            number=index + 1,
            code=code,
            category=current_category,
            quantity=start.quantity,
            unit_price=start.unit_price,
            total_price=start.total_price,
            detail_lines=dedupe_adjacent(details),
        )
        enrich_product(product)
        products.append(product)

    return products


def extract_image_blocks(doc, temp_dir: Path) -> list[PdfImageBlock]:
    blocks: list[PdfImageBlock] = []
    for page_index, page in enumerate(doc):
        page_height = float(page.rect.height)
        for block_index, block in enumerate(page.get_text("dict").get("blocks", [])):
            if block.get("type") != 1:
                continue
            bbox = tuple(float(value) for value in block.get("bbox", (0, 0, 0, 0)))
            if len(bbox) != 4:
                continue
            x0, y0, x1, y1 = bbox
            width = max(0.0, x1 - x0)
            height = max(0.0, y1 - y0)
            image_bytes = block.get("image") or b""
            if x0 > 115 or height < 8 or not image_bytes:
                continue

            ext = str(block.get("ext") or "png").lower().lstrip(".")
            if ext not in {"png", "jpg", "jpeg", "bmp", "gif"}:
                ext = "png"
            path = temp_dir / f"page_{page_index + 1:02d}_image_{block_index + 1:02d}.{ext}"
            path.write_bytes(image_bytes)
            blocks.append(
                PdfImageBlock(
                    global_y=page_index * page_height + y0,
                    bbox_area=width * height,
                    data_size=len(image_bytes),
                    path=path,
                )
            )
    return sorted(blocks, key=lambda block: block.global_y)


def assign_images(products: list[ParsedProduct], images: list[PdfImageBlock]) -> None:
    used: set[Path] = set()
    for index, product in enumerate(products):
        start_y = product_start_global_y(product)
        end_y = product_start_global_y(products[index + 1]) if index + 1 < len(products) else float("inf")
        candidates = [
            image
            for image in images
            if image.path not in used and start_y < image.global_y < end_y
        ]
        if not candidates:
            continue
        chosen = max(candidates, key=lambda image: (image.bbox_area, image.data_size))
        product.image_path = chosen.path
        used.add(chosen.path)


def product_start_global_y(product: ParsedProduct) -> float:
    # Filled after product parsing by positional lookup. This keeps ParsedProduct free of layout-only fields.
    return getattr(product, "_start_global_y")


def enrich_product(product: ParsedProduct) -> None:
    full_text = "\n".join(product.detail_lines)
    product.dimension = extract_dimension(full_text)
    product.name = extract_name(product.detail_lines, product.dimension) or product.code
    product.description = "\nProduct Specification:\n" + full_text.strip() + "\n"
    product.color = extract_color(product.detail_lines)


def extract_dimension(text: str) -> str:
    normalized = text.replace("Φ ", "Φ").replace("Ф ", "Ф")
    match = DIMENSION_RE.search(normalized)
    if not match:
        return ""
    value = clean_text(match.group(0))
    return value.upper().replace(" ", "")


def extract_name(lines: list[str], dimension: str) -> str:
    name_parts: list[str] = []
    for line in lines:
        if has_spec_marker(line):
            break
        cleaned = remove_dimensions(line)
        cleaned = clean_text(cleaned)
        if cleaned:
            name_parts.append(cleaned)
    if not name_parts and lines:
        name_parts.append(clean_text(remove_dimensions(lines[0])))
    return clean_text(" ".join(name_parts))


def extract_color(lines: list[str]) -> str:
    color_lines = [line for line in lines if has_spec_marker(line)]
    return "\n".join(color_lines[:4])


def remove_dimensions(text: str) -> str:
    return clean_text(DIMENSION_RE.sub("", text.replace("Φ ", "Φ").replace("Ф ", "Ф")))


def has_spec_marker(text: str) -> bool:
    upper = normalize_marker_text(text)
    return any(marker in upper for marker in SPEC_MARKERS)


def project_or_name_field(line: PdfLine) -> str:
    tokens = []
    for word in sorted(line.words, key=lambda word: word.x0):
        if not (115 <= word.x0 < 320):
            continue
        if word.text in {"+", "/", "-", "|"}:
            continue
        if MONEY_RE.match(word.text) or word.text.upper() == "PCS" or INTEGER_RE.match(word.text):
            continue
        tokens.append(word.text)
    return clean_text(" ".join(tokens))


def looks_like_product_start(line: PdfLine) -> bool:
    money_count = sum(1 for word in line.words if MONEY_RE.match(word.text))
    has_pcs = any(word.text.upper() == "PCS" for word in line.words)
    has_qty = any(INTEGER_RE.match(word.text) and 320 <= word.x0 <= 375 for word in line.words)
    has_code = any(word.x0 < 115 and CODE_RE.match(word.text) for word in line.words)
    return money_count >= 2 and has_pcs and has_qty and has_code


def contains_money_or_uom(line: PdfLine) -> bool:
    return any(MONEY_RE.match(word.text) or word.text.upper() == "PCS" for word in line.words)


def is_single_buyer_number(line: PdfLine) -> bool:
    words = sorted(line.words, key=lambda word: word.x0)
    return len(words) == 1 and INTEGER_RE.match(words[0].text) and 115 <= words[0].x0 <= 140


def is_repeated_page_header(normalized_text: str) -> bool:
    if re.fullmatch(r"PAGE \d+ OF \d+", normalized_text):
        return True
    if re.search(r"\bPAGE \d+ OF \d+\b", normalized_text):
        return True
    if normalized_text.startswith("PROFORMA INVOICE NO"):
        return True
    if "BUYER" in normalized_text and "PRODUCT SPECIFICATION" in normalized_text:
        return True
    header_markers = (
        "PRODUCT NO",
        "BUYER'S NO",
        "PRODUCT NAME",
        "PRODUCT SPECIFICATION",
        "QUANTITY",
        "UNIT-PRICE",
        "SUB-TOTAL",
    )
    return normalized_text in header_markers or all(marker in normalized_text for marker in ("PRODUCT", "QUANTITY", "UNIT"))


def parse_money(value: str) -> float:
    return float(value.replace("$", "").replace(",", ""))


def clean_text(value: str) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").split())


def normalize_marker_text(value: str) -> str:
    return clean_text(value).upper()


def dedupe_adjacent(lines: list[str]) -> list[str]:
    deduped: list[str] = []
    for line in lines:
        cleaned = clean_text(line)
        if cleaned and (not deduped or cleaned != deduped[-1]):
            deduped.append(cleaned)
    return deduped


def extract_metadata(lines: list[PdfLine], fallback_name: str, page_count: int) -> dict[str, str]:
    text = "\n".join(line.text for line in lines[:60])
    invoice = next(iter(re.findall(r"\b[A-Z]{2}-[A-Z0-9]+-\d+\b", text)), fallback_name)
    date = next(iter(re.findall(r"\b\d{1,2}/[A-Za-z]{3}/\d{2}\b", text)), "")
    buyer_lines = [line.text_between(0, 295) for line in lines if line.page_index == 0 and 92 <= line.y <= 132]
    seller_lines = [line.text_between(295, 570) for line in lines if line.page_index == 0 and 92 <= line.y <= 132]
    return {
        "invoice": invoice,
        "date": date,
        "buyer": clean_text(" ".join(buyer_lines)),
        "seller": clean_text(" ".join(seller_lines)) or "ALMA MANUFACTURING COMPANY LIMITED",
        "project": invoice,
        "pages": str(page_count),
    }


def load_style_source(reference_path: Path | None) -> StyleSource:
    source = StyleSource(reference_path=reference_path)
    if not reference_path:
        return source

    workbook = load_workbook(reference_path, data_only=False)
    try:
        sheet = workbook["Quotation"] if "Quotation" in workbook.sheetnames else workbook.active
        for col in range(1, 13):
            letter = sheet.cell(row=HEADER_ROW, column=col).column_letter
            source.column_widths[col] = sheet.column_dimensions[letter].width
        for row in range(1, HEADER_ROW + 1):
            source.row_heights[row] = sheet.row_dimensions[row].height
            source.header_rows[row] = [copy_cell_style(sheet.cell(row=row, column=col)) for col in range(1, 13)]
        source.category_row = [copy_cell_style(sheet.cell(row=8, column=col)) for col in range(1, 13)]
        source.product_row = [copy_cell_style(sheet.cell(row=9, column=col)) for col in range(1, 13)]
        source.total_row = [copy_cell_style(sheet.cell(row=18, column=col)) for col in range(1, 13)]
    finally:
        workbook.close()
    return source


def copy_cell_style(cell) -> dict:
    return {
        "value": cell.value,
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def apply_cell_style(cell, style: dict, include_value: bool = False) -> None:
    if include_value:
        cell.value = style.get("value")
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def write_workbook(
    products: list[ParsedProduct],
    metadata: dict[str, str],
    output: Path,
    style_source: StyleSource,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"

    apply_base_layout(sheet, style_source)
    write_header(sheet, metadata)
    write_products(sheet, products, style_source)

    workbook.save(output)
    workbook.close()


def apply_base_layout(sheet, style_source: StyleSource) -> None:
    for col in range(1, 13):
        letter = sheet.cell(row=HEADER_ROW, column=col).column_letter
        sheet.column_dimensions[letter].width = style_source.column_widths.get(col) or default_width(col)

    for row in range(1, HEADER_ROW + 1):
        sheet.row_dimensions[row].height = style_source.row_heights.get(row)
        for col in range(1, 13):
            cell = sheet.cell(row=row, column=col)
            style = style_source.header_rows.get(row, [None] * 12)[col - 1] if style_source.header_rows else None
            if style:
                apply_cell_style(cell, style, include_value=True)
    if not style_source.header_rows:
        apply_default_header_style(sheet)

    for col, header in enumerate(HEADERS, start=1):
        sheet.cell(row=HEADER_ROW, column=col).value = header


def write_header(sheet, metadata: dict[str, str]) -> None:
    sheet["A1"] = metadata.get("seller") or "ALMA MANUFACTURING COMPANY LIMITED"
    sheet["A4"] = f"Project: {metadata.get('project') or ''}"
    sheet["A5"] = f"To: {metadata.get('buyer') or ''}"
    sheet["J4"] = "REF:"
    sheet["K4"] = metadata.get("invoice") or ""
    sheet["J5"] = "DATE:"
    sheet["K5"] = metadata.get("date") or ""
    sheet["J6"] = "PAGE:"
    sheet["K6"] = metadata.get("pages") or ""


def write_products(sheet, products: list[ParsedProduct], style_source: StyleSource) -> None:
    row = HEADER_ROW + 1
    grouped: dict[str, list[ParsedProduct]] = {}
    for product in products:
        grouped.setdefault(product.category, []).append(product)

    product_number = 1
    for category, category_products in grouped.items():
        category_row = row
        apply_template_row(sheet, category_row, style_source.category_row)
        sheet.cell(row=category_row, column=1).value = f"- {category}"
        sheet.cell(row=category_row, column=9).value = f"- {category}  Tot.Price  :"
        row += 1

        first_product_row = row
        for product in category_products:
            apply_template_row(sheet, row, style_source.product_row)
            write_product_row(sheet, row, product, product_number)
            row += 1
            product_number += 1
        last_product_row = row - 1
        sheet.cell(row=category_row, column=11).value = f"=SUM(K{first_product_row}:K{last_product_row})"

    apply_template_row(sheet, row, style_source.total_row)
    sheet.cell(row=row, column=1).value = " "
    sheet.cell(row=row, column=9).value = f"=SUM(I{HEADER_ROW + 1}:I{row - 1})"
    sheet.cell(row=row, column=10).value = "  Tot.Price "
    sheet.cell(row=row, column=11).value = f"=SUM(K{HEADER_ROW + 1}:K{row - 1})"


def write_product_row(sheet, row: int, product: ParsedProduct, number: int) -> None:
    height = min(320, max(145, 80 + len(product.description) / 5))
    sheet.row_dimensions[row].height = height
    sheet.cell(row=row, column=1).value = number
    sheet.cell(row=row, column=2).value = f"{product.code}\n{product.name}".strip()
    sheet.cell(row=row, column=4).value = product.description
    sheet.cell(row=row, column=5).value = product.dimension
    sheet.cell(row=row, column=6).value = product.color
    sheet.cell(row=row, column=7).value = product.quantity
    sheet.cell(row=row, column=8).value = 0
    sheet.cell(row=row, column=9).value = f"=G{row}*H{row}"
    sheet.cell(row=row, column=10).value = product.unit_price
    sheet.cell(row=row, column=11).value = f"=G{row}*J{row}"

    for col in (2, 4, 5, 6):
        sheet.cell(row=row, column=col).alignment = copy(sheet.cell(row=row, column=col).alignment)
        sheet.cell(row=row, column=col).alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal=sheet.cell(row=row, column=col).alignment.horizontal,
        )

    if product.image_path and product.image_path.exists():
        add_product_image(sheet, row, product.image_path, height)


def add_product_image(sheet, row: int, path: Path, row_height_points: float) -> None:
    image = XlsxImage(str(path))
    max_width = 135
    max_height = max(90, int(row_height_points * 1.1))
    if image.width and image.height:
        scale = min(max_width / image.width, max_height / image.height)
        image.width = max(1, int(image.width * scale))
        image.height = max(1, int(image.height * scale))
    sheet.add_image(image, f"C{row}")


def apply_template_row(sheet, row: int, template: list | None) -> None:
    if template:
        for col, style in enumerate(template, start=1):
            apply_cell_style(sheet.cell(row=row, column=col), style, include_value=False)
        return

    thin = Side(style="thin", color="B7B7B7")
    for col in range(1, 13):
        cell = sheet.cell(row=row, column=col)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_default_header_style(sheet) -> None:
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 13):
        cell = sheet.cell(row=HEADER_ROW, column=col)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def default_width(col: int) -> float:
    widths = {
        1: 5.63,
        2: 12.18,
        3: 21.73,
        4: 32.27,
        5: 12.18,
        6: 13.0,
        7: 5.0,
        8: 8.54,
        9: 9.27,
        10: 10.73,
        11: 13.63,
        12: 7.82,
    }
    return widths.get(col, 10.0)


def attach_start_positions(products: list[ParsedProduct], starts: list[ProductStart]) -> None:
    for product, start in zip(products, starts, strict=True):
        setattr(product, "_start_global_y", start.line.global_y)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convierte una proforma PDF a hoja Quotation XLSX.")
    parser.add_argument("--source", "-s", required=True, help="PDF de proforma del proveedor")
    parser.add_argument("--reference-xlsx", "-r", help="XLSX de ejemplo para copiar formato base")
    parser.add_argument("--output", "-o", help="Ruta del XLSX Quotation generado")
    args = parser.parse_args(argv)

    source = Path(args.source)
    output = Path(args.output) if args.output else source.with_name(f"{source.stem}-Quotation-from-pdf.xlsx")
    result = convert_pdf_to_quotation(source, output, args.reference_xlsx)

    print(f"[OK] Quotation generado: {result.output_path}")
    print(f"[OK] Productos: {result.product_count}")
    print(f"[OK] Imagenes asociadas: {result.image_count}")
    print(f"[OK] Categorias: {', '.join(result.categories)}")
    return 0
