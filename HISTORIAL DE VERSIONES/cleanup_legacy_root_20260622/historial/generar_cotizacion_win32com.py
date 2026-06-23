#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti - win32com Edition
======================================================
Usa Excel nativo via COM para maxima fidelidad visual.
El output es identico al ejemplo objetivo porque Excel mismo
hace todo el trabajo de formato, insercion de filas e imagenes.

Uso:
    python generar_cotizacion_win32com.py \
        --source "Quotation.xlsx" \
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \
        --output "Cotizacion.xlsx" \
        --cotizacion "100-99999" \
        --proyecto "Proyecto" \
        --cliente "Cliente"
"""

import argparse
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import win32com.client as win32


Q_HEADER_ROW = 7

# Constantes Excel
XL_DOWN = -4121
XL_FORMAT_FROM_LEFT_OR_ABOVE = 0
XL_UP = -4162
XL_PASTE_FORMATS = -4122


def rgb(r, g, b):
    """Convierte RGB a color OLE (BGR) para Excel COM."""
    return r | (g << 8) | (b << 16)


def limpiar_excel():
    """Cierra procesos Excel colgados."""
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes(source_path):
    """Extrae imagenes del source y crea mapeo row->imagen_path."""
    temp_dir = tempfile.mkdtemp(prefix="excel_images_")
    image_map = {}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        media_files = [f for f in z.namelist() if 'media' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
    
    wb = openpyxl.load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    for img in ws._images:
        path = img.path
        filename = os.path.basename(path)
        filepath = os.path.join(temp_dir, filename)
        
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1
        else:
            row = img.anchor.row + 1
        
        if os.path.exists(filepath):
            image_map[row] = filepath
    
    wb.close()
    return image_map, temp_dir


def leer_items(source_path):
    """Lee items de Quotation del proveedor."""
    wb = openpyxl.load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    last_row = ws.max_row
    for row in range(last_row, 0, -1):
        if ws.cell(row=row, column=1).value is not None:
            last_row = row
            break
    
    items = []
    for row in range(Q_HEADER_ROW + 1, last_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        if item_name is None or item_name == "":
            continue
        
        if isinstance(no_val, str) and no_val.startswith('-'):
            items.append({
                'tipo': 'categoria',
                'row': row,
                'nombre': no_val.strip('- ').strip()
            })
        elif isinstance(no_val, (int, float)):
            items.append({
                'tipo': 'producto',
                'row': row,
                'nombre': item_name
            })
        elif no_val is None or no_val == "":
            items.append({
                'tipo': 'categoria',
                'row': row,
                'nombre': item_name.strip()
            })
    
    wb.close()
    return items


def generar_cotizacion(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI - win32com Edition")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source).resolve()
    template_path = Path(args.template).resolve()
    
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx").resolve()
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: no se pudo borrar archivo previo: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template no encontrado: {template_path}")
        sys.exit(1)
    
    temp_dir = None
    excel = None
    
    try:
        # PASO 1: Extraer imagenes
        print("[1/8] Extrayendo imagenes del Quotation...")
        image_map, temp_dir = extraer_imagenes(str(source_path))
        print(f"      [OK] {len(image_map)} imagenes extraidas")
        
        # PASO 2: Leer items
        print("[2/8] Leyendo items del Quotation...")
        items = leer_items(str(source_path))
        num_cats = sum(1 for i in items if i['tipo'] == 'categoria')
        num_prods = sum(1 for i in items if i['tipo'] == 'producto')
        print(f"      [OK] {len(items)} items ({num_cats} categorias, {num_prods} productos)")
        
        # PASO 3: Iniciar Excel
        print("[3/8] Iniciando Excel...")
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.AskToUpdateLinks = False
        import time
        time.sleep(2)
        print("      [OK] Excel iniciado")
        
        # PASO 4: Abrir workbooks
        print("[4/8] Abriendo archivos...")
        import time
        max_retries = 3
        for attempt in range(max_retries):
            try:
                wb_template = excel.Workbooks.Open(str(template_path))
                break
            except Exception as e:
                if attempt == max_retries - 1:
                    raise
                time.sleep(2)
        for attempt in range(max_retries):
            try:
                wb_source = excel.Workbooks.Open(str(source_path))
                break
            except Exception as e:
                if attempt == max_retries - 1:
                    raise
                time.sleep(2)
        print("      [OK] Archivos abiertos")
        
        # PASO 5: Copiar Quotation del source al template
        print("[5/8] Copiando hoja Quotation...")
        sheets_before = [sh.Name for sh in wb_template.Sheets]
        ws_source_q = wb_source.Sheets("Quotation")
        ws_source_q.Copy(After=wb_template.Sheets(wb_template.Sheets.Count))
        sheets_after = [sh.Name for sh in wb_template.Sheets]
        new_sheets = [n for n in sheets_after if n not in sheets_before]
        if new_sheets:
            q_sheet_name = new_sheets[0]
            ws_quotation = wb_template.Sheets(q_sheet_name)
        else:
            q_sheet_name = "Quotation"
            ws_quotation = wb_template.Sheets(q_sheet_name)
        print(f"      [OK] Quotation copiada como '{q_sheet_name}'")
        
        # PASO 6: Llenar encabezado
        print("[6/8] Llenando encabezado de Cotizacion...")
        ws_cot = wb_template.Sheets("Cotizacion")
        
        # Desproteger
        try:
            ws_cot.Unprotect()
        except Exception:
            pass
        
        # Desmergear celdas del encabezado
        for row in range(3, 13):
            try:
                ws_cot.Range(f"A{row}:J{row}").UnMerge()
            except Exception:
                pass
        
        ws_cot.Range("B3").Value = args.cotizacion
        ws_cot.Range("B4").Value = datetime.now()
        ws_cot.Range("B7").Value = args.proyecto
        ws_cot.Range("B8").Value = args.cliente
        ws_cot.Range("B9").Value = args.correo
        ws_cot.Range("B10").Value = args.telefono
        ws_cot.Range("B11").Value = args.direccion
        ws_cot.Range("B12").Value = args.razon_social
        print("      [OK] Encabezado llenado")
        
        # PASO 7: Generar Mobiliti
        print("[7/8] Generando Mobiliti...")
        ws_mob = wb_template.Sheets("Mobiliti")
        
        # Desproteger hoja si esta protegida
        try:
            ws_mob.Unprotect()
        except Exception:
            pass
        
        # Limpiar area de datos (filas 17 en adelante)
        last_mob_row = ws_mob.Cells(ws_mob.Rows.Count, 1).End(XL_UP).Row
        if last_mob_row >= 17:
            ws_mob.Range(f"A17:G{last_mob_row}").ClearContents()
        
        # Desmergear
        for row in range(17, 200):
            try:
                ws_mob.Range(f"A{row}:G{row}").UnMerge()
            except Exception:
                pass
        
        current_row = 17
        section_num = 1
        section_start_row = current_row
        mobiliti_row_map = {}
        
        for item in items:
            if item['tipo'] == 'categoria':
                if current_row > section_start_row:
                    ws_mob.Cells(current_row, 1).Value = f"Subtotales Seccion {section_num}"
                    ws_mob.Cells(current_row, 7).Value = f"=SUM(G{section_start_row}:G{current_row-1})"
                    ws_mob.Cells(current_row, 1).Font.Bold = True
                    ws_mob.Cells(current_row, 7).Font.Bold = True
                    current_row += 1
                    section_num += 1
                
                ws_mob.Cells(current_row, 1).Value = f"={q_sheet_name}!A{item['row']}"
                # Formato categoria: Calibri 20 bold, fill marron
                r = ws_mob.Range(f"A{current_row}:G{current_row}")
                r.Interior.Color = rgb(62, 37, 0)
                r.Font.Name = "Calibri"
                r.Font.Size = 20
                r.Font.Bold = True
                r.Font.Color = rgb(255, 255, 255)
                
                current_row += 1
                section_start_row = current_row
            else:
                q_row = item['row']
                ws_mob.Cells(current_row, 1).Value = f"={q_sheet_name}!B{q_row}"
                ws_mob.Cells(current_row, 2).Value = f"={q_sheet_name}!D{q_row}"
                ws_mob.Cells(current_row, 3).Value = "Sunon Inc"
                ws_mob.Cells(current_row, 4).Value = f'=IFERROR(VLOOKUP(C{current_row},Proveedores!A:B,2,0)," ")'
                ws_mob.Cells(current_row, 5).Value = f"={q_sheet_name}!G{q_row}"
                ws_mob.Cells(current_row, 6).Value = f"={q_sheet_name}!J{q_row}"
                ws_mob.Cells(current_row, 7).Value = f"=E{current_row}*F{current_row}"
                
                # Formato producto: Century Gothic 11, fill amarillo
                r = ws_mob.Range(f"A{current_row}:G{current_row}")
                r.Interior.Color = rgb(255, 192, 0)
                r.Font.Name = "Century Gothic"
                r.Font.Size = 11
                ws_mob.Cells(current_row, 1).Font.Bold = True
                ws_mob.Cells(current_row, 2).Font.Bold = True
                ws_mob.Cells(current_row, 3).Font.Bold = False
                ws_mob.Cells(current_row, 4).Font.Bold = True
                ws_mob.Cells(current_row, 5).Font.Bold = True
                ws_mob.Cells(current_row, 6).Font.Bold = False
                ws_mob.Cells(current_row, 7).Font.Bold = False
                
                mobiliti_row_map[q_row] = current_row
                current_row += 1
        
        if current_row > section_start_row:
            ws_mob.Cells(current_row, 1).Value = f"Subtotales Seccion {section_num}"
            ws_mob.Cells(current_row, 7).Value = f"=SUM(G{section_start_row}:G{current_row-1})"
            ws_mob.Cells(current_row, 1).Font.Bold = True
            ws_mob.Cells(current_row, 7).Font.Bold = True
        
        print(f"      [OK] Mobiliti generada ({section_num} secciones)")
        
        # PASO 8: Generar Cotizacion
        print("[8/8] Generando Cotizacion...")
        
        # Encontrar fila de terminos
        terminos_start_row = None
        for row in range(16, 100):
            val = ws_cot.Cells(row, 1).Value
            if val and isinstance(val, str) and 'CONDICIONES' in val:
                terminos_start_row = row
                break
        if not terminos_start_row:
            terminos_start_row = 32
        
        # Desmergear TODO el area de datos + terminos para evitar conflictos
        for row in range(16, terminos_start_row + 30):
            try:
                ws_cot.Range(f"A{row}:J{row}").UnMerge()
            except Exception:
                pass
        
        # Limpiar filas de ejemplo
        for row in range(16, terminos_start_row):
            ws_cot.Range(f"A{row}:J{row}").ClearContents()
        
        # Insertar filas adicionales si es necesario
        filas_ejemplo = terminos_start_row - 16
        if len(items) > filas_ejemplo:
            filas_extra = len(items) - filas_ejemplo
            ws_cot.Rows(f"{terminos_start_row}:{terminos_start_row + filas_extra - 1}").Insert(
                Shift=XL_DOWN,
                CopyOrigin=XL_FORMAT_FROM_LEFT_OR_ABOVE
            )
            terminos_start_row += filas_extra
        
        # Escribir items
        current_row = 16
        first_data_row = None
        last_data_row = None
        descuento_row = None
        cat_rows = set()
        
        for item in items:
            if item['tipo'] == 'categoria':
                ws_cot.Cells(current_row, 1).Value = f"={q_sheet_name}!A{item['row']}"
                
                # Formato categoria: Roboto 16 bold, fill azul
                r = ws_cot.Range(f"A{current_row}:J{current_row}")
                r.Interior.Color = rgb(115, 169, 219)
                r.Font.Name = "Roboto"
                r.Font.Size = 16
                r.Font.Bold = True
                r.Font.Color = rgb(0, 0, 0)
                
                # Mergear fila de categoria
                r.Merge()
                cat_rows.add(current_row)
                
                current_row += 1
            else:
                if first_data_row is None:
                    first_data_row = current_row
                    descuento_row = current_row + 2
                
                q_row = item['row']
                mob_row = mobiliti_row_map.get(q_row)
                
                ws_cot.Cells(current_row, 1).Value = f"={q_sheet_name}!B{q_row}"
                ws_cot.Cells(current_row, 3).Value = f"={q_sheet_name}!D{q_row}"
                ws_cot.Cells(current_row, 4).Value = f"={q_sheet_name}!E{q_row}"
                ws_cot.Cells(current_row, 5).Value = f"={q_sheet_name}!G{q_row}"
                
                if mob_row:
                    ws_cot.Cells(current_row, 6).Value = f"=Mobiliti!F{mob_row}"
                else:
                    ws_cot.Cells(current_row, 6).Value = f"={q_sheet_name}!J{q_row}"
                
                ws_cot.Cells(current_row, 7).Value = f"=G${descuento_row}"
                ws_cot.Cells(current_row, 8).Value = f"=F{current_row}*G{current_row}"
                ws_cot.Cells(current_row, 9).Value = f"=F{current_row}-H{current_row}"
                ws_cot.Cells(current_row, 10).Value = f"=I{current_row}*E{current_row}"
                
                # Formato producto
                r = ws_cot.Range(f"A{current_row}:J{current_row}")
                r.Interior.ColorIndex = 0  # Sin fill
                ws_cot.Cells(current_row, 1).Font.Name = "Roboto"
                ws_cot.Cells(current_row, 1).Font.Size = 16
                ws_cot.Cells(current_row, 1).Font.Bold = True
                ws_cot.Cells(current_row, 2).Font.Name = "Roboto"
                ws_cot.Cells(current_row, 2).Font.Size = 16
                ws_cot.Cells(current_row, 2).Font.Bold = False
                ws_cot.Cells(current_row, 3).Font.Name = "Roboto"
                ws_cot.Cells(current_row, 3).Font.Size = 16
                ws_cot.Cells(current_row, 3).Font.Bold = False
                for col in range(4, 11):
                    ws_cot.Cells(current_row, col).Font.Name = "Roboto"
                    ws_cot.Cells(current_row, col).Font.Size = 16
                    ws_cot.Cells(current_row, col).Font.Bold = True
                
                current_row += 1
        
        last_data_row = current_row - 1
        
        # Descuento base
        if descuento_row and descuento_row <= last_data_row:
            ws_cot.Cells(descuento_row, 7).Value = 0.7
        
        # Insertar imagenes
        print("      Insertando imagenes...")
        img_count = 0
        current_row = 16
        for item in items:
            if item['tipo'] == 'categoria':
                current_row += 1
                continue
            
            q_row = item['row']
            if q_row in image_map:
                try:
                    img_path = image_map[q_row]
                    left = ws_cot.Range(f"B{current_row}").Left + 2
                    top = ws_cot.Range(f"B{current_row}").Top + 2
                    ws_cot.Shapes.AddPicture(
                        img_path,
                        LinkToFile=False,
                        SaveWithDocument=True,
                        Left=left,
                        Top=top,
                        Width=120,
                        Height=90
                    )
                    img_count += 1
                except Exception as e:
                    pass
            
            current_row += 1
        
        print(f"      [OK] {img_count} imagenes insertadas")
        
        # Totales
        print("      Agregando totales...")
        row_sub = terminos_start_row - 5
        if row_sub <= last_data_row + 1:
            ws_cot.Rows(f"{last_data_row + 2}:{last_data_row + 6}").Insert(
                Shift=XL_DOWN,
                CopyOrigin=XL_FORMAT_FROM_LEFT_OR_ABOVE
            )
            row_sub = last_data_row + 2
            terminos_start_row = row_sub + 5
        
        # Desmergear filas de totales
        for row in range(row_sub, row_sub + 5):
            try:
                ws_cot.Range(f"A{row}:J{row}").UnMerge()
            except Exception:
                pass
        
        ws_cot.Cells(row_sub, 4).Value = "SUBTOTAL:"
        ws_cot.Cells(row_sub, 7).Value = f"=SUM(J{first_data_row}:J{last_data_row})"
        ws_cot.Cells(row_sub, 4).Font.Bold = True
        ws_cot.Cells(row_sub, 7).Font.Bold = True
        
        row_flete = row_sub + 1
        ws_cot.Cells(row_flete, 4).Value = "COSTO DE FLETE E INSTALACION:"
        ws_cot.Cells(row_flete, 7).Value = f"=G{row_sub}*12%"
        ws_cot.Cells(row_flete, 4).Font.Bold = True
        ws_cot.Cells(row_flete, 7).Font.Bold = True
        
        row_sub2 = row_flete + 1
        ws_cot.Cells(row_sub2, 4).Value = "SUBTOTAL:"
        ws_cot.Cells(row_sub2, 7).Value = f"=G{row_sub}+G{row_flete}"
        ws_cot.Cells(row_sub2, 4).Font.Bold = True
        ws_cot.Cells(row_sub2, 7).Font.Bold = True
        
        row_iva = row_sub2 + 1
        ws_cot.Cells(row_iva, 4).Value = "IVA:"
        ws_cot.Cells(row_iva, 7).Value = f"=G{row_sub2}*16%"
        ws_cot.Cells(row_iva, 4).Font.Bold = True
        ws_cot.Cells(row_iva, 7).Font.Bold = True
        
        row_total = row_iva + 1
        ws_cot.Cells(row_total, 4).Value = "TOTAL:"
        ws_cot.Cells(row_total, 7).Value = f"=G{row_sub2}+G{row_iva}"
        ws_cot.Cells(row_total, 4).Font.Bold = True
        ws_cot.Cells(row_total, 7).Font.Bold = True
        
        print(f"      [OK] Totales en filas {row_sub}-{row_total}")
        
        # Guardar
        print("      Guardando...")
        wb_template.SaveAs(str(output_path))
        wb_template.Close()
        wb_source.Close()
        excel.Quit()
        
        print()
        print("=" * 60)
        print("[OK] COTIZACION GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti con Excel nativo (win32com)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_win32com.py \\
    --source "Quotation.xlsx" \\
    --template "Formato Cotizacion 2026 GDL (1).xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotización 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    
    args = parser.parse_args()
    generar_cotizacion(args)


if __name__ == '__main__':
    main()
