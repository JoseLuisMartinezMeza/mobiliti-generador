#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti - xlwings Edition
====================================================
Usa Excel nativo para insertar filas dinamicamente y preservar formato.

Flujo:
1. Abre template en Excel
2. Inserta filas en Cotizacion segun cantidad de items
3. Copia hoja Quotation completa del source
4. Llena datos con formulas que referencian Quotation
5. Guarda resultado

Uso:
    python generar_cotizacion_xlwings.py \
        --source "Quotation.xlsx" \
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \
        --output "Cotizacion.xlsx" \
        --cotizacion "100-99999" \
        --proyecto "Proyecto" \
        --cliente "Cliente"
"""

import argparse
import os
import shutil
import subprocess
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import xlwings as xw

Q_HEADER_ROW = 7


def limpiar_excel():
    """Cierra procesos Excel colgados."""
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes_source(source_path):
    """Extrae imagenes del source y crea mapeo row->imagen_path."""
    temp_dir = os.path.join(os.path.dirname(str(source_path)), "_temp_images")
    os.makedirs(temp_dir, exist_ok=True)
    image_map = {}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        media_files = [f for f in z.namelist() if 'media' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
    
    # Mapear usando openpyxl (solo para obtener posiciones)
    from openpyxl import load_workbook
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    for img in ws._images:
        path = img.path
        filename = os.path.basename(path)
        filepath = os.path.join(temp_dir, filename)
        
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1  # 1-indexed
        else:
            row = img.anchor.row + 1
        
        if os.path.exists(filepath):
            image_map[row] = filepath
    
    wb.close()
    return image_map, temp_dir


def leer_datos_quotation(source_path):
    """Lee items de Quotation."""
    from openpyxl import load_workbook
    print("[1/9] Leyendo datos del archivo fuente...")
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    items = []
    for row in range(Q_HEADER_ROW + 1, ws.max_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        if no_val is None and item_name is None:
            continue
        
        if isinstance(no_val, str) and no_val.startswith('-'):
            items.append({
                'tipo': 'categoria',
                'nombre': no_val.strip('- ').strip(),
                'row': row
            })
        elif isinstance(no_val, (int, float)) and item_name:
            desc = ws.cell(row=row, column=4).value
            # Truncar descripcion a primera linea
            if desc and isinstance(desc, str):
                lines = [l.strip() for l in desc.split('\n') if l.strip()]
                if lines:
                    desc = lines[0][:100]
                else:
                    desc = desc[:100]
            
            items.append({
                'tipo': 'producto',
                'no': int(no_val),
                'item_name': item_name,
                'description': desc,
                'dimension': ws.cell(row=row, column=5).value,
                'color': ws.cell(row=row, column=6).value,
                'qty': ws.cell(row=row, column=7).value,
                'unit_price': ws.cell(row=row, column=10).value,
                'row': row
            })
        elif item_name and isinstance(item_name, str) and not no_val:
            items.append({
                'tipo': 'categoria',
                'nombre': item_name.strip(),
                'row': row
            })
    
    print(f"      [OK] {len(items)} items leidos")
    wb.close()
    return items


def copiar_hoja_quotation(wb_dest, source_path):
    """Copia la hoja Quotation completa del source al destino usando xlwings."""
    print("[3/9] Copiando hoja Quotation original...")
    
    # Abrir source en Excel
    app = wb_dest.app
    
    wb_src = app.books.open(str(source_path))
    ws_src = wb_src.sheets['Quotation']
    
    # Verificar si ya existe Quotation en destino
    sheet_names = [s.name for s in wb_dest.sheets]
    if 'Quotation' in sheet_names:
        # Eliminar hoja existente
        wb_dest.sheets['Quotation'].delete()
    
    # Copiar la hoja al destino
    ws_src.api.Copy(After=wb_dest.sheets[-1].api)
    
    # Renombrar si es necesario
    new_sheet = wb_dest.sheets[-1]
    if new_sheet.name != 'Quotation':
        new_sheet.name = 'Quotation'
    
    wb_src.close()
    print("      [OK] Quotation copiada")


def llenar_encabezado_cliente(ws_cot, args):
    print("[4/9] Llenando datos del cliente...")
    
    campos = {
        'B3': args.cotizacion,
        'B4': datetime.now(),
        'B7': args.proyecto,
        'B8': args.cliente,
        'B9': args.correo,
        'B10': args.telefono,
        'B11': args.direccion,
        'B12': args.razon_social,
    }
    
    for cell, valor in campos.items():
        if valor:
            ws_cot.range(cell).value = valor
    
    print(f"      [OK] Cliente: {args.cliente}")


def contar_filas_productos_template(ws_cot):
    """Cuenta cuantas filas de productos tiene el template."""
    # El template tiene productos de ejemplo en filas 17-23
    # Detectamos donde empiezan los terminos (fila 25 es TOTAL del template)
    count = 0
    for row in range(16, 30):
        val = ws_cot.range(f'A{row}').value
        if val and (str(val).startswith('=') or str(val).startswith('Item')):
            count += 1
    return count if count > 0 else 7  # Default 7 filas de ejemplo


def insertar_filas_productos(ws_cot, num_productos_reales, num_productos_template=7):
    """Inserta filas adicionales si hay mas productos que el template."""
    filas_necesarias = num_productos_reales
    filas_existentes = num_productos_template
    
    if filas_necesarias > filas_existentes:
        filas_a_insertar = filas_necesarias - filas_existentes
        print(f"      Insertando {filas_a_insertar} filas adicionales...")
        
        # Insertar filas antes de los terminos (fila 24)
        # Las filas 16-23 son productos del template
        # La fila 24 es vacia, fila 25 es TOTAL del template
        fila_insercion = 23  # Insertar despues del ultimo producto del template
        ws_cot.api.Rows(f"{fila_insercion}:{fila_insercion + filas_a_insertar - 1}").Insert()
        
        # Copiar formato de la fila anterior a las nuevas filas
        for i in range(filas_a_insertar):
            dest_row = fila_insercion + i
            ws_cot.api.Rows(dest_row).ClearContents()
        
        print(f"      [OK] {filas_a_insertar} filas insertadas")
        return True
    return False


def llenar_productos(ws_cot, items, image_map=None):
    """Llena las filas de productos con formulas y opcionalmente imagenes."""
    print("[6/9] Llenando productos en Cotizacion...")
    
    current_row = 16
    first_data_row = None
    
    # Establecer descuento base
    ws_cot.range('G19').value = 0.7
    
    for item in items:
        if item['tipo'] == 'categoria':
            ws_cot.range(f'A{current_row}').value = item['nombre']
            current_row += 1
        else:
            if first_data_row is None:
                first_data_row = current_row
            
            q_row = item['row']
            ws_cot.range(f'A{current_row}').value = f"=Quotation!B{q_row}"
            ws_cot.range(f'C{current_row}').value = f"=Quotation!D{q_row}"
            ws_cot.range(f'D{current_row}').value = f"=Quotation!E{q_row}"
            ws_cot.range(f'E{current_row}').value = f"=Quotation!G{q_row}"
            ws_cot.range(f'F{current_row}').value = f"=Quotation!J{q_row}"
            ws_cot.range(f'G{current_row}').value = "=G$19"
            ws_cot.range(f'H{current_row}').value = f"=F{current_row}*G{current_row}"
            ws_cot.range(f'I{current_row}').value = f"=F{current_row}-H{current_row}"
            ws_cot.range(f'J{current_row}').value = f"=I{current_row}*E{current_row}"
            
            # Insertar imagen si existe
            if image_map and q_row in image_map:
                try:
                    img_path = image_map[q_row]
                    # Insertar imagen con tamaño proporcional
                    left = ws_cot.range(f'B{current_row}').left
                    top = ws_cot.range(f'B{current_row}').top
                    pic = ws_cot.pictures.add(img_path, 
                                               left=left + 2,
                                               top=top + 2,
                                               width=120,
                                               height=90)
                except Exception as e:
                    print(f"      Advertencia: Imagen fila {current_row}: {e}")
            
            current_row += 1
    
    last_data_row = current_row - 1
    print(f"      [OK] {len(items)} items en filas {first_data_row}-{last_data_row}")
    return first_data_row, last_data_row


def agregar_totales(ws_cot, row_inicio, row_fin):
    print("[7/9] Agregando totales...")
    
    # Encontrar fila de TOTAL del template
    # Normalmente esta en fila 25 o despues de los productos
    row_total_template = None
    for row in range(row_fin + 1, row_fin + 10):
        val = ws_cot.range(f'D{row}').value
        if val and 'TOTAL' in str(val).upper():
            row_total_template = row
            break
    
    if not row_total_template:
        row_total_template = row_fin + 2
    
    row_sub = row_total_template
    ws_cot.range(f'D{row_sub}').value = "SUBTOTAL:"
    ws_cot.range(f'G{row_sub}').value = f"=SUM(J{row_inicio}:J{row_fin})"
    
    row_flete = row_sub + 1
    ws_cot.range(f'D{row_flete}').value = "COSTO DE FLETE E INSTALACION:"
    ws_cot.range(f'G{row_flete}').value = f"=G{row_sub}*12%"
    
    row_sub2 = row_flete + 1
    ws_cot.range(f'D{row_sub2}').value = "SUBTOTAL:"
    ws_cot.range(f'G{row_sub2}').value = f"=G{row_sub}+G{row_flete}"
    
    row_iva = row_sub2 + 1
    ws_cot.range(f'D{row_iva}').value = "IVA:"
    ws_cot.range(f'G{row_iva}').value = f"=G{row_sub2}*16%"
    
    row_total = row_iva + 1
    ws_cot.range(f'D{row_total}').value = "TOTAL:"
    ws_cot.range(f'G{row_total}').value = f"=G{row_sub2}+G{row_iva}"
    
    print(f"      [OK] Totales en filas {row_sub}-{row_total}")
    return row_total


def llenar_mobiliti(wb_dest, items):
    """Llena la hoja Mobiliti con datos de Quotation."""
    print("[5/9] Llenando hoja Mobiliti...")
    
    if 'Mobiliti' not in [s.name for s in wb_dest.sheets]:
        print("      [SKIP] Hoja Mobiliti no existe")
        return
    
    ws_m = wb_dest.sheets['Mobiliti']
    
    # Limpiar area de datos (filas 14-100)
    ws_m.range('A14:G100').clear_contents()
    
    # Escribir productos
    current_row = 14
    section_num = 1
    section_start_row = current_row
    
    for item in items:
        if item['tipo'] == 'categoria':
            if current_row > section_start_row:
                ws_m.range(f'A{current_row}').value = f"Subtotales Seccion {section_num}"
                ws_m.range(f'G{current_row}').value = f"=SUM(G{section_start_row}:G{current_row-1})"
                current_row += 1
                section_num += 1
            
            ws_m.range(f'A{current_row}').value = f"=Quotation!A{item['row']}"
            current_row += 1
            section_start_row = current_row
        else:
            q_row = item['row']
            ws_m.range(f'A{current_row}').value = f"=Quotation!B{q_row}"
            ws_m.range(f'B{current_row}').value = f"=Quotation!D{q_row}"
            ws_m.range(f'C{current_row}').value = "Sunon Inc"
            ws_m.range(f'D{current_row}').value = f'=IFERROR(VLOOKUP(C{current_row},Proveedores!A:B,2,0)," ")'
            ws_m.range(f'E{current_row}').value = f"=Quotation!G{q_row}"
            ws_m.range(f'F{current_row}').value = f"=Quotation!J{q_row}"
            ws_m.range(f'G{current_row}').value = f"=E{current_row}*F{current_row}"
            current_row += 1
    
    if current_row > section_start_row:
        ws_m.range(f'A{current_row}').value = f"Subtotales Seccion {section_num}"
        ws_m.range(f'G{current_row}').value = f"=SUM(G{section_start_row}:G{current_row-1})"
    
    print(f"      [OK] Mobiliti llenada ({section_num} secciones)")


def generar_cotizacion(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI - xlwings")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source)
    template_path = Path(args.template)
    
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx")
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template no encontrado: {template_path}")
        sys.exit(1)
    
    temp_dir = None
    app = None
    
    try:
        # PASO 1: Extraer imagenes
        image_map, temp_dir = extraer_imagenes_source(source_path)
        
        # PASO 2: Leer datos
        items = leer_datos_quotation(source_path)
        
        # Contar productos (no categorias)
        num_productos = sum(1 for item in items if item['tipo'] == 'producto')
        num_categorias = sum(1 for item in items if item['tipo'] == 'categoria')
        total_filas = len(items)
        
        # PASO 3: Abrir Excel y template
        print("[2/9] Abriendo Excel...")
        app = xw.App(visible=False)
        app.display_alerts = False
        
        print(f"      Abriendo template: {template_path}")
        wb_dest = app.books.open(str(template_path))
        
        # PASO 4: Copiar Quotation del source
        copiar_hoja_quotation(wb_dest, source_path)
        
        # PASO 5: Llenar Mobiliti
        llenar_mobiliti(wb_dest, items)
        
        # PASO 6: Preparar Cotizacion
        ws_cot = wb_dest.sheets['Cotizacion']
        
        # Llenar encabezado
        llenar_encabezado_cliente(ws_cot, args)
        
        # Limpiar filas de productos del template
        print("[5/9] Preparando filas de productos...")
        ws_cot.range('A16:J50').clear_contents()
        
        # Eliminar imagenes antiguas de productos
        for pic in ws_cot.pictures:
            try:
                # Solo eliminar imagenes que no sean del header (top < fila 15)
                if pic.top > ws_cot.range('A15').top:
                    pic.delete()
            except Exception:
                pass
        
        # PASO 7: Insertar filas si necesario
        # El template tiene ~7 filas de ejemplo (16-22)
        # Necesitamos espacio para todos los items
        filas_template = 7  # Filas 16-22
        if total_filas > filas_template:
            print(f"      {total_filas} items > {filas_template} filas del template")
            print("      Insertando filas adicionales...")
            
            # Calcular cuantas filas insertar
            filas_extra = total_filas - filas_template
            
            # Insertar despues de la fila 22 (ultimo producto del template)
            # Esto empuja los terminos hacia abajo
            fila_insertar = 22
            for _ in range(filas_extra):
                ws_cot.api.Rows(fila_insertar + 1).Insert()
            
            print(f"      [OK] {filas_extra} filas insertadas")
        
        # PASO 8: Llenar productos
        first_row, last_row = llenar_productos(ws_cot, items, image_map)
        
        # PASO 9: Totales
        if first_row and last_row:
            row_total = agregar_totales(ws_cot, first_row, last_row)
        
        # Guardar
        print("[8/9] Guardando archivo...")
        wb_dest.save(str(output_path))
        wb_dest.close()
        
        print()
        print("=" * 60)
        print("[OK] COTIZACION GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print(f"Items: {len(items)} ({num_productos} productos, {num_categorias} categorias)")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if app:
            try:
                app.quit()
            except Exception:
                pass
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti con xlwings',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_xlwings.py \\
    --source "Quotation.xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotizacion 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    
    args = parser.parse_args()
    generar_cotizacion(args)


if __name__ == '__main__':
    main()
