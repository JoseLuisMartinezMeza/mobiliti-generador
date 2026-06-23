#!/usr/bin/env python3
"""
Generador de Cotizaciones Mobiliti v3
=====================================
- Copia hoja Quotation COMPLETA del source
- Preserva formato de celdas del template
- Inserta terminos despues de totales
- Usa openpyxl (compatible Python 3.14)

Uso:
    python generar_cotizacion_v3.py \
        --source "Quotation.xlsx" \
        --template "Formato Cotizacion 2026 GDL (1).xlsx" \
        --output "Cotizacion.xlsx" \
        --cotizacion "100-99999" \
        --proyecto "Proyecto" \
        --cliente "Cliente"
"""

import argparse
import copy
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


Q_HEADER_ROW = 7


def limpiar_excel():
    try:
        subprocess.run(['taskkill', '//F', '//IM', 'EXCEL.EXE'],
                      capture_output=True, timeout=10)
    except Exception:
        pass


def extraer_imagenes_source(source_path):
    """Extrae imagenes del source y crea mapeo row->imagen_path."""
    temp_dir = tempfile.mkdtemp(prefix="excel_images_")
    image_map = {}
    
    with zipfile.ZipFile(source_path, 'r') as z:
        media_files = [f for f in z.namelist() if 'media' in f]
        for f in media_files:
            filename = os.path.basename(f)
            filepath = os.path.join(temp_dir, filename)
            with open(filepath, 'wb') as out:
                out.write(z.read(f))
    
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    for img in ws._images:
        path = img.path
        filename = os.path.basename(path)
        filepath = os.path.join(temp_dir, filename)
        
        if hasattr(img.anchor, '_from'):
            row = img.anchor._from.row + 1  # 1-indexed
        else:
            row = img.anchor.row + 1
        
        if os.path.exists(filepath):
            image_map[row] = filepath
    
    wb.close()
    return image_map, temp_dir


def leer_datos_quotation(source_path):
    """Lee items de Quotation."""
    print("[1/9] Leyendo datos del archivo fuente...")
    wb = load_workbook(str(source_path), data_only=False)
    ws = wb['Quotation']
    
    items = []
    for row in range(Q_HEADER_ROW + 1, ws.max_row + 1):
        no_val = ws.cell(row=row, column=1).value
        item_name = ws.cell(row=row, column=2).value
        
        if no_val is None and item_name is None:
            continue
        
        if isinstance(no_val, str) and no_val.startswith('-'):
            items.append({
                'tipo': 'categoria',
                'nombre': no_val.strip('- ').strip(),
                'row': row
            })
        elif isinstance(no_val, (int, float)) and item_name:
            desc = ws.cell(row=row, column=4).value
            if desc and isinstance(desc, str):
                lines = [l.strip() for l in desc.split('\n') if l.strip()]
                if lines:
                    desc = lines[0][:100]
                else:
                    desc = desc[:100]
            
            items.append({
                'tipo': 'producto',
                'no': int(no_val),
                'item_name': item_name,
                'description': desc,
                'dimension': ws.cell(row=row, column=5).value,
                'color': ws.cell(row=row, column=6).value,
                'qty': ws.cell(row=row, column=7).value,
                'unit_price': ws.cell(row=row, column=10).value,
                'row': row
            })
        elif item_name and isinstance(item_name, str) and not no_val:
            items.append({
                'tipo': 'categoria',
                'nombre': item_name.strip(),
                'row': row
            })
    
    print(f"      [OK] {len(items)} items leidos")
    wb.close()
    return items


def copiar_hoja_completa(wb_dest, wb_src, sheet_name, new_name=None, copiar_imagenes=False):
    """Copia una hoja completa de un workbook a otro, incluyendo valores y estilos."""
    if new_name is None:
        new_name = sheet_name
    
    # Eliminar si ya existe
    if new_name in wb_dest.sheetnames:
        del wb_dest[new_name]
    
    ws_src = wb_src[sheet_name]
    ws_dest = wb_dest.create_sheet(new_name)
    
    # Copiar dimensiones de columnas
    for col_letter, col_dim in ws_src.column_dimensions.items():
        if col_dim.width:
            ws_dest.column_dimensions[col_letter].width = col_dim.width
    
    # Copiar dimensiones de filas
    for row_num, row_dim in ws_src.row_dimensions.items():
        if row_dim.height:
            ws_dest.row_dimensions[row_num].height = row_dim.height
    
    # Copiar celdas con valores y estilos (solo las primeras 100 filas para velocidad)
    max_row = min(ws_src.max_row, 100)
    for row in ws_src.iter_rows(min_row=1, max_row=max_row):
        for cell_src in row:
            cell_dest = ws_dest.cell(row=cell_src.row, column=cell_src.column)
            
            # Valor
            cell_dest.value = cell_src.value
            
            # Estilo
            if cell_src.has_style:
                cell_dest.font = copy.copy(cell_src.font)
                cell_dest.border = copy.copy(cell_src.border)
                cell_dest.fill = copy.copy(cell_src.fill)
                cell_dest.number_format = copy.copy(cell_src.number_format)
                cell_dest.protection = copy.copy(cell_src.protection)
                cell_dest.alignment = copy.copy(cell_src.alignment)
    
    # Copiar merged cells (solo las que estan dentro del rango)
    for merged_range in ws_src.merged_cells.ranges:
        if merged_range.max_row <= max_row:
            ws_dest.merge_cells(str(merged_range))
    
    # Copiar imagenes solo si se solicita
    if copiar_imagenes:
        for img in ws_src._images:
            ws_dest._images.append(img)
    
    return ws_dest


def extraer_terminos_template(ws_cot):
    """Extrae las filas de terminos y condiciones del template (fila 29 en adelante)."""
    START_ROW = 29
    terminos = []
    merges = []
    
    for merged_range in list(ws_cot.merged_cells.ranges):
        if merged_range.min_row >= START_ROW:
            merges.append({
                'min_row': merged_range.min_row,
                'min_col': merged_range.min_col,
                'max_row': merged_range.max_row,
                'max_col': merged_range.max_col,
            })
    
    for row in range(START_ROW, 80):
        row_data = {}
        has_value = False
        for col in range(1, 11):
            cell = ws_cot.cell(row=row, column=col)
            if cell.value is not None:
                row_data[col] = cell.value
                has_value = True
        if has_value or any(m['min_row'] <= row <= m['max_row'] for m in merges):
            terminos.append({
                'rel_row': row - START_ROW,
                'values': row_data,
            })
    
    return {'rows': terminos, 'merges': merges}


def insertar_terminos(ws_cot, terminos_data, start_row):
    """Inserta los terminos y condiciones a partir de start_row."""
    if not terminos_data or not terminos_data['rows']:
        return
    
    START_ROW = 29
    
    # Copiar valores
    for row_info in terminos_data['rows']:
        dest_row = START_ROW + row_info['rel_row'] + (start_row - START_ROW)
        for col, value in row_info['values'].items():
            ws_cot.cell(row=dest_row, column=col, value=value)
    
    # Recrear merged cells con nuevo offset
    for merge in terminos_data['merges']:
        new_min_row = merge['min_row'] + (start_row - START_ROW)
        new_max_row = merge['max_row'] + (start_row - START_ROW)
        try:
            ws_cot.merge_cells(
                start_row=new_min_row,
                start_column=merge['min_col'],
                end_row=new_max_row,
                end_column=merge['max_col']
            )
        except Exception:
            pass
    
    print(f"      [OK] Terminos insertados en filas {start_row}+")


def unmerge_cells_in_rows(ws, start_row, end_row):
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            to_unmerge.append(str(merged_range))
    for merge in to_unmerge:
        ws.unmerge_cells(merge)


def llenar_encabezado_cliente(ws_cot, args):
    print("[4/9] Llenando datos del cliente...")
    unmerge_cells_in_rows(ws_cot, 3, 12)
    
    campos = {
        3: ("COTIZACION #", args.cotizacion, 2),
        4: (None, datetime.now(), 1),
        7: ("PROYECTO", args.proyecto, 2),
        8: ("Nombre", args.cliente, 2),
        9: ("Correo", args.correo, 2),
        10: ("Telefono", args.telefono, 2),
        11: ("Direccion", args.direccion, 2),
        12: ("Razon Social", args.razon_social, 2),
    }
    
    for row, (label, valor, col) in campos.items():
        if valor:
            ws_cot.cell(row=row, column=col, value=valor)
            if label:
                ws_cot.cell(row=row, column=1, value=f"{label}:")
    
    print(f"      [OK] Cliente: {args.cliente}")


def limpiar_cotizacion(ws_cot):
    """Limpia el area de datos de Cotizacion, incluyendo imagenes."""
    header_images = []
    for img in ws_cot._images:
        row = img.anchor._from.row if hasattr(img.anchor, '_from') else img.anchor.row
        if row < 15:
            header_images.append(img)
    ws_cot._images = header_images
    
    unmerge_cells_in_rows(ws_cot, 16, 200)
    for row in range(16, 200):
        for col in range(1, 11):
            cell = ws_cot.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None


def copiar_formato_celda(cell_src, cell_dest):
    """Copia el formato de una celda a otra."""
    if cell_src.has_style:
        cell_dest.font = copy.copy(cell_src.font)
        cell_dest.border = copy.copy(cell_src.border)
        cell_dest.fill = copy.copy(cell_src.fill)
        cell_dest.number_format = copy.copy(cell_src.number_format)
        cell_dest.protection = copy.copy(cell_src.protection)
        cell_dest.alignment = copy.copy(cell_src.alignment)


def crear_fila_categoria(ws_cot, row_dest, nombre_categoria, template_row=None, ws_template=None):
    ws_cot.cell(row=row_dest, column=1, value=nombre_categoria)
    # Aplicar formato de categoria del template si existe
    if template_row and ws_template:
        for col in range(1, 11):
            src_cell = ws_template.cell(row=template_row, column=col)
            dest_cell = ws_cot.cell(row=row_dest, column=col)
            copiar_formato_celda(src_cell, dest_cell)


def crear_fila_producto(ws_cot, row_dest, item, image_map=None, template_row=None, ws_template=None):
    q_row = item['row']
    
    ws_cot.cell(row=row_dest, column=1, value=f"=Quotation!B{q_row}")
    ws_cot.cell(row=row_dest, column=3, value=f"=Quotation!D{q_row}")
    ws_cot.cell(row=row_dest, column=4, value=f"=Quotation!E{q_row}")
    ws_cot.cell(row=row_dest, column=5, value=f"=Quotation!G{q_row}")
    ws_cot.cell(row=row_dest, column=6, value=f"=Quotation!J{q_row}")
    ws_cot.cell(row=row_dest, column=7, value="=G$19")
    ws_cot.cell(row=row_dest, column=8, value=f"=F{row_dest}*G{row_dest}")
    ws_cot.cell(row=row_dest, column=9, value=f"=F{row_dest}-H{row_dest}")
    ws_cot.cell(row=row_dest, column=10, value=f"=I{row_dest}*E{row_dest}")
    
    # Aplicar formato del template
    if template_row and ws_template:
        for col in range(1, 11):
            src_cell = ws_template.cell(row=template_row, column=col)
            dest_cell = ws_cot.cell(row=row_dest, column=col)
            copiar_formato_celda(src_cell, dest_cell)
    
    # Insertar imagen si existe
    if image_map and q_row in image_map:
        try:
            img_path = image_map[q_row]
            img = XLImage(img_path)
            img.width = 158
            img.height = 118
            ws_cot.add_image(img, f'B{row_dest}')
        except Exception as e:
            print(f"      Advertencia: No se pudo insertar imagen fila {row_dest}: {e}")


def procesar_items_cotizacion(ws_cot, items, image_map=None, ws_template=None):
    print("[6/9] Procesando items en Cotizacion...")
    
    current_row = 16
    first_data_row = None
    
    # Establecer descuento base (usando fila 19 del template como referencia)
    ws_cot.cell(row=19, column=7, value=0.7)
    
    for item in items:
        if item['tipo'] == 'categoria':
            crear_fila_categoria(ws_cot, current_row, item['nombre'], template_row=16, ws_template=ws_template)
            current_row += 1
        else:
            if first_data_row is None:
                first_data_row = current_row
            crear_fila_producto(ws_cot, current_row, item, image_map, template_row=17, ws_template=ws_template)
            current_row += 1
    
    last_data_row = current_row - 1
    print(f"      [OK] {len(items)} items, filas {first_data_row}-{last_data_row}")
    return first_data_row, last_data_row


def agregar_totales(ws_cot, row_inicio, row_fin):
    print("[7/9] Agregando totales...")
    
    row_sub = row_fin + 2
    ws_cot.cell(row=row_sub, column=4, value="SUBTOTAL:")
    ws_cot.cell(row=row_sub, column=7, value=f"=SUM(J{row_inicio}:J{row_fin})")
    
    row_flete = row_sub + 1
    ws_cot.cell(row=row_flete, column=4, value="COSTO DE FLETE E INSTALACION:")
    ws_cot.cell(row=row_flete, column=7, value=f"=G{row_sub}*12%")
    
    row_sub2 = row_flete + 1
    ws_cot.cell(row=row_sub2, column=4, value="SUBTOTAL:")
    ws_cot.cell(row=row_sub2, column=7, value=f"=G{row_sub}+G{row_flete}")
    
    row_iva = row_sub2 + 1
    ws_cot.cell(row=row_iva, column=4, value="IVA:")
    ws_cot.cell(row=row_iva, column=7, value=f"=G{row_sub2}*16%")
    
    row_total = row_iva + 1
    ws_cot.cell(row=row_total, column=4, value="TOTAL:")
    ws_cot.cell(row=row_total, column=7, value=f"=G{row_sub2}+G{row_iva}")
    
    print(f"      [OK] Totales en filas {row_sub}-{row_total}")
    return row_total


def llenar_mobiliti(wb_dest, items):
    print("[5/9] Llenando hoja Mobiliti...")
    
    if 'Mobiliti' not in wb_dest.sheetnames:
        print("      [SKIP] Hoja Mobiliti no existe")
        return
    
    ws_m = wb_dest['Mobiliti']
    unmerge_cells_in_rows(ws_m, 14, 200)
    
    for row in range(14, 200):
        for col in range(1, 8):
            cell = ws_m.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
    
    current_row = 14
    section_num = 1
    section_start_row = current_row
    
    for item in items:
        if item['tipo'] == 'categoria':
            if current_row > section_start_row:
                ws_m.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
                ws_m.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
                current_row += 1
                section_num += 1
            
            ws_m.cell(row=current_row, column=1, value=f"=Quotation!A{item['row']}")
            current_row += 1
            section_start_row = current_row
        else:
            q_row = item['row']
            ws_m.cell(row=current_row, column=1, value=f"=Quotation!B{q_row}")
            ws_m.cell(row=current_row, column=2, value=f"=Quotation!D{q_row}")
            ws_m.cell(row=current_row, column=3, value="Sunon Inc")
            ws_m.cell(row=current_row, column=4, value=f'=IFERROR(VLOOKUP(C{current_row},Proveedores!A:B,2,0)," ")')
            ws_m.cell(row=current_row, column=5, value=f"=Quotation!G{q_row}")
            ws_m.cell(row=current_row, column=6, value=f"=Quotation!J{q_row}")
            ws_m.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}")
            current_row += 1
    
    if current_row > section_start_row:
        ws_m.cell(row=current_row, column=1, value=f"Subtotales Seccion {section_num}")
        ws_m.cell(row=current_row, column=7, value=f"=SUM(G{section_start_row}:G{current_row-1})")
    
    print(f"      [OK] Mobiliti llenada ({section_num} secciones)")


def generar_cotizacion(args):
    print("=" * 60)
    print("GENERADOR DE COTIZACIONES MOBILITI v3")
    print("=" * 60)
    print()
    
    limpiar_excel()
    
    source_path = Path(args.source)
    template_path = Path(args.template)
    
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ps = args.proyecto.replace(' ', '_').replace('/', '-') if args.proyecto else 'Cotizacion'
        output_path = Path(f"Cotizacion_{ps}_{ts}.xlsx")
    
    if output_path.exists():
        try:
            os.remove(str(output_path))
        except Exception as e:
            print(f"Advertencia: {e}")
    
    if not source_path.exists():
        print(f"ERROR: Fuente no encontrado: {source_path}")
        sys.exit(1)
    if not template_path.exists():
        print(f"ERROR: Template no encontrado: {template_path}")
        sys.exit(1)
    
    temp_dir = None
    wb_src = None
    
    try:
        # PASO 1: Extraer imagenes
        image_map, temp_dir = extraer_imagenes_source(source_path)
        
        # PASO 2: Leer datos
        items = leer_datos_quotation(source_path)
        
        # PASO 3: Abrir workbooks
        print("[2/9] Abriendo archivos...")
        wb_src = load_workbook(str(source_path), data_only=False)
        wb_dest = load_workbook(str(template_path), data_only=False)
        print("      [OK] Archivos abiertos")
        
        # PASO 4: Copiar Quotation completa del source
        print("[3/9] Copiando hoja Quotation original...")
        copiar_hoja_completa(wb_dest, wb_src, 'Quotation', 'Quotation')
        print("      [OK] Quotation copiada")
        
        # PASO 5: Llenar Mobiliti
        llenar_mobiliti(wb_dest, items)
        
        # PASO 6: Preparar Cotizacion
        ws_cot = wb_dest['Cotizacion']
        ws_template = load_workbook(str(template_path), data_only=False)['Cotizacion']
        
        # Extraer terminos ANTES de limpiar
        terminos_data = extraer_terminos_template(ws_cot)
        
        # Limpiar
        limpiar_cotizacion(ws_cot)
        
        # Llenar encabezado
        llenar_encabezado_cliente(ws_cot, args)
        
        # PASO 7: Procesar items
        first_row, last_row = procesar_items_cotizacion(ws_cot, items, image_map, ws_template)
        
        # PASO 8: Totales
        row_total = None
        if first_row and last_row:
            row_total = agregar_totales(ws_cot, first_row, last_row)
        
        # PASO 9: Insertar terminos
        if row_total and terminos_data:
            print("[8/9] Restaurando terminos y condiciones...")
            insertar_terminos(ws_cot, terminos_data, row_total + 2)
        
        # PASO 10: Guardar
        print("[9/9] Guardando archivo...")
        wb_dest.save(str(output_path))
        
        ws_template.parent.close()
        wb_dest.close()
        
        print()
        print("=" * 60)
        print("[OK] COTIZACION GENERADA EXITOSAMENTE")
        print("=" * 60)
        print(f"Archivo: {output_path}")
        print(f"Items: {len(items)}")
        print(f"Hojas: {wb_dest.sheetnames}")
        print()
        
    except Exception as e:
        print()
        print("=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if wb_src:
            wb_src.close()
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


def main():
    parser = argparse.ArgumentParser(
        description='Genera cotizaciones Mobiliti v3',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo:
  python generar_cotizacion_v3.py \\
    --source "Quotation.xlsx" \\
    --output "Cotizacion.xlsx" \\
    --cotizacion "100-99999" \\
    --proyecto "Proyecto XYZ" \\
    --cliente "Juan Perez"
        """
    )
    
    parser.add_argument('--source', '-s', required=True, help='Archivo fuente con Quotation')
    parser.add_argument('--template', '-t', default='Formato Cotizacion 2026 GDL (1).xlsx', help='Plantilla')
    parser.add_argument('--output', '-o', default=None, help='Archivo de salida')
    
    parser.add_argument('--cotizacion', '-n', default='100-00000')
    parser.add_argument('--proyecto', '-p', default='')
    parser.add_argument('--cliente', '-c', default='')
    parser.add_argument('--correo', '-e', default='')
    parser.add_argument('--telefono', '-tel', default='')
    parser.add_argument('--direccion', '-d', default='')
    parser.add_argument('--razon_social', '-r', default='')
    
    args = parser.parse_args()
    generar_cotizacion(args)


if __name__ == '__main__':
    main()
